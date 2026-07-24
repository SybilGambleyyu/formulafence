"""Generate a safe demonstration pair; run with `python generate_workbooks.py`."""

from pathlib import Path

from openpyxl import Workbook

OUTPUT = Path(__file__).parent


def build(path: Path, candidate: bool = False) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Revenue input"
    inputs["B2"] = 100

    model = workbook.create_sheet("Model")
    model["A1"] = "Calculated revenue"
    model["B2"] = 250 if candidate else "=Inputs!B2*2"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Headline output"
    dashboard["B12"] = "=Model!B2"
    workbook.save(path)


if __name__ == "__main__":
    build(OUTPUT / "baseline.xlsx")
    build(OUTPUT / "candidate-risky.xlsx", candidate=True)
    print("Wrote baseline.xlsx and candidate-risky.xlsx")
