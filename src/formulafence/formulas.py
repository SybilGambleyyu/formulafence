"""Small, safe helpers for inspecting—not evaluating—Excel formulas."""

from __future__ import annotations

import re
import unicodedata
from collections.abc import Mapping, Sequence
from dataclasses import dataclass

from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_to_tuple,
    range_boundaries,
)

from formulafence.models import (
    ExternalWorkbookDefinedNameReference,
    ExternalWorkbookReference,
    ExternalWorkbookStructuredReference,
    ExternalWorkbookThreeDReference,
)

MAX_EXCEL_ROW = 1_048_576
MAX_EXCEL_COLUMN = 16_384
_DYNAMIC_REFERENCE_FUNCTIONS = {"INDIRECT", "OFFSET"}
# These functions can cause a workbook to navigate a reviewer, request content,
# or bind to a host-side data provider.  HYPERLINK also supports in-workbook
# destinations, but its link location may be dynamically computed, so retain
# every call as a reviewable action surface rather than evaluating arguments.
#
# ``STOCKHISTORY`` retrieves provider-backed market history.  The documented
# Cube family can bind a stored formula to a workbook connection, an OLAP server,
# or an offline cube.  Keep all seven Cube functions together: even helpers
# such as ``CUBESETCOUNT`` carry a value whose semantics depend on a cube-set
# chain, and splitting that family would create a misleading partial boundary.
_EXTERNAL_ACTION_CUBE_FUNCTIONS = frozenset(
    {
        "CUBEKPIMEMBER",
        "CUBEMEMBER",
        "CUBEMEMBERPROPERTY",
        "CUBERANKEDMEMBER",
        "CUBESET",
        "CUBESETCOUNT",
        "CUBEVALUE",
    }
)
_EXTERNAL_ACTION_FUNCTIONS = frozenset(
    {
        "HYPERLINK",
        "WEBSERVICE",
        "IMAGE",
        "RTD",
        "STOCKHISTORY",
        *_EXTERNAL_ACTION_CUBE_FUNCTIONS,
    }
)
# ``PY`` is intentionally separate from the remote-content/action ledger. Excel
# stores its executable code in a workbook-level Python part, so callers need a
# dedicated code boundary that can fingerprint both the formula and that part.
_PYTHON_FUNCTIONS = {"PY"}
# ``REGISTER.ID`` is the worksheet-capable member of Excel's legacy
# DLL/code-resource registration family.  Microsoft documents that it
# registers the resource when needed, then returns its registration ID.  Keep
# it separate from generic external-action calls: the module, procedure, type
# string, and arguments are sensitive implementation material, and the
# surrounding XLM-only ``CALL`` / ``REGISTER`` forms are covered by the raw
# macro-sheet boundary instead.
_WORKSHEET_CODE_RESOURCE_REGISTRATION_FUNCTIONS = {"REGISTER.ID"}
# ``REGISTER`` is an XLM macro function rather than a normal worksheet
# function. Microsoft documents that its registration modes can be called from
# a defined-name definition, so FormulaFence inventories it only while it is
# inspecting formula-defined names and named LAMBDAs. Direct worksheet calls
# remain outside this narrow boundary; raw XLM macro sheets have their own
# package-level scanner.
_FORMULA_DEFINED_XLM_REGISTRATION_FUNCTIONS = {"REGISTER"}
# `EVALUATE` is an XLM function that parses a supplied text expression at
# calculation time. Keep it in its own stored-definition boundary rather than
# treating the text as a statically visible formula. Direct worksheet calls
# remain outside this narrow boundary; the caller enables inspection only for
# formula-defined names and named LAMBDAs.
_FORMULA_DEFINED_XLM_EVALUATION_FUNCTIONS = {"EVALUATE"}
# These selected XLM macro-sheet functions can dispatch another macro, launch
# a program, invoke a DLL entry point, send a DDE command, or install an event
# handler that runs a macro later. They are deliberately inspected only inside
# formula-defined names and named LAMBDAs: raw XLM macro-sheet programs remain
# under the separate package-level boundary, and FormulaFence never executes or
# resolves any stored action argument.
_FORMULA_DEFINED_XLM_ACTION_FUNCTIONS = {
    "CALL",
    "EXEC",
    "EXECUTE",
    "ON.DATA",
    "ON.DOUBLECLICK",
    "ON.ENTRY",
    "ON.KEY",
    "ON.RECALC",
    "ON.SHEET",
    "ON.TIME",
    "ON.WINDOW",
    "RUN",
    "SEND.KEYS",
}
# `GET.CELL` is an XLM information function that can observe cell contents,
# formatting, dimensions, protection, and other state outside ordinary
# value/formula dependency semantics. Keep it in a dedicated stored-definition
# boundary. Direct worksheet calls remain outside this narrow boundary; the
# caller enables inspection only for formula-defined names and named LAMBDAs.
_FORMULA_DEFINED_XLM_GET_CELL_FUNCTIONS = {"GET.CELL"}
# These legacy XLM information calls can make a formula-defined name depend on
# workbook, application/workspace, or document state. Keep this intentionally
# small, explicit native set separate from the cell-information boundary.
# Direct worksheet calls remain outside the stored-definition boundary.
_FORMULA_DEFINED_XLM_ENVIRONMENT_INFORMATION_FUNCTIONS = {
    "GET.DOCUMENT",
    "GET.WORKBOOK",
    "GET.WORKSPACE",
}
# ``CELL`` and ``INFO`` can observe the current file, client, folder,
# selection, or other environment state rather than just visible cell
# precedents. ``SHEET`` and ``SHEETS`` can likewise observe the workbook tab
# catalog: sheet numbers, and an omitted SHEETS reference, depend on that
# catalog rather than normal formula precedents. Unlike the intentionally
# narrow XLM boundary above, inspect these native calls wherever a formula can
# be stored. The private marker carries one additional static fact through a
# named-formula chain: ``CELL`` was called without its optional reference.
# Microsoft documents that Excel can then use the currently selected cell at
# calculation time; FormulaFence inventories that surface without evaluating
# the call or simulating a selection.
_FORMULA_ENVIRONMENT_INFORMATION_FUNCTIONS = {"CELL", "INFO", "SHEET", "SHEETS"}
_FORMULA_ENVIRONMENT_INFORMATION_IMPLICIT_CELL_REFERENCE_MARKER = (
    "FORMULAFENCE_CELL_IMPLICIT_REFERENCE_MARKER"
)
_FORMULA_ENVIRONMENT_INFORMATION_IMPLICIT_SHEETS_REFERENCE_MARKER = (
    "FORMULAFENCE_SHEETS_IMPLICIT_REFERENCE_MARKER"
)
# A direct DDE worksheet formula is not a function call.  Excel's documented
# form is ``application|topic!item`` (for example,
# ``='Quote'|'NYSE'!ZAXX``), and some hostile samples use the same syntax to
# name a command processor.  Keep the lexical result opaque: it is only used
# to count and propagate a statically visible DDE-link surface, never to
# resolve a service, topic, item, server, command, or process.
_FORMULA_DDE_LINK_MARKER = "FORMULAFENCE_FORMULA_DDE_LINK"
_DDE_UNQUOTED_COMPONENT_DELIMITERS = frozenset(
    " \t\r\n'\"|!(),;{}+-*/^&=<>%"
)
_DDE_EXPRESSION_BOUNDARIES = frozenset("=+-*/^&,(;{@")
_DDE_ITEM_TERMINATORS = frozenset(" \t\r\n(),;{}+-*/^&=<>%")
# Excel's native function catalog includes a small but important set of dotted
# names.  A namespace separator is also how Office Add-in custom functions are
# displayed (for example, ``CONTOSO.ADD``), so keep these known native names out
# of the custom-function candidate ledger.  This is deliberately a local,
# stable list instead of depending on a version-specific third-party catalog.
_EXCEL_DOTTED_FUNCTIONS = {
    "BETA.DIST",
    "BETA.INV",
    "BINOM.DIST",
    "BINOM.DIST.RANGE",
    "BINOM.INV",
    "CEILING.MATH",
    "CEILING.PRECISE",
    "CHISQ.DIST",
    "CHISQ.DIST.RT",
    "CHISQ.INV",
    "CHISQ.INV.RT",
    "CHISQ.TEST",
    "CONFIDENCE.NORM",
    "CONFIDENCE.T",
    "COVARIANCE.P",
    "COVARIANCE.S",
    "ECMA.CEILING",
    "ERF.PRECISE",
    "ERFC.PRECISE",
    "ERROR.TYPE",
    "EXPON.DIST",
    "F.DIST",
    "F.DIST.RT",
    "F.INV",
    "F.INV.RT",
    "F.TEST",
    "FLOOR.MATH",
    "FLOOR.PRECISE",
    "FORECAST.ETS",
    "FORECAST.ETS.CONFINT",
    "FORECAST.ETS.SEASONALITY",
    "FORECAST.ETS.STAT",
    "FORECAST.LINEAR",
    "GAMMA.DIST",
    "GAMMA.INV",
    "GAMMALN.PRECISE",
    "GET.CELL",
    "GET.DOCUMENT",
    "GET.WORKBOOK",
    "GET.WORKSPACE",
    "HYPGEOM.DIST",
    "ISO.CEILING",
    "LOGNORM.DIST",
    "LOGNORM.INV",
    "MODE.MULT",
    "MODE.SNGL",
    "NEGBINOM.DIST",
    "NETWORKDAYS.INTL",
    "NORM.DIST",
    "NORM.INV",
    "NORM.S.DIST",
    "NORM.S.INV",
    "PERCENTILE.EXC",
    "PERCENTILE.INC",
    "PERCENTRANK.EXC",
    "PERCENTRANK.INC",
    "POISSON.DIST",
    "QUARTILE.EXC",
    "QUARTILE.INC",
    "RANK.AVG",
    "RANK.EQ",
    "REGISTER.ID",
    "SKEW.P",
    "STDEV.P",
    "STDEV.S",
    "T.DIST",
    "T.DIST.2T",
    "T.DIST.RT",
    "T.INV",
    "T.INV.2T",
    "T.TEST",
    "VAR.P",
    "VAR.S",
    "WEIBULL.DIST",
    "WORKDAY.INTL",
    "Z.TEST",
}

# A bare formula call that is not a workbook-defined LAMBDA can be resolved by
# Excel through VBA, a COM/Automation add-in, an XLL, or another registered
# runtime.  That is a useful review surface, but classifying ordinary native
# functions as candidates would make the boundary unusable.  Keep a stable
# snapshot rather than importing a third-party runtime catalog: Microsoft’s
# alphabetical worksheet-function reference is the primary source, with the
# formula-serialization/runtime-only names added explicitly below.  The source
# was reviewed on 2026-07-26:
# https://support.microsoft.com/en-us/office/excel-functions-alphabetical-b3944572-255d-4efb-bb96-c6d90033e188
#
# Dotted native functions are intentionally absent because the classifier below
# accepts bare identifiers only.  ``PY`` is listed separately because its own
# Microsoft reference documents a formula spelling that is not currently
# present in the alphabetical index. ``SINGLE`` and ``ANCHORARRAY`` are the
# OOXML-compatible function spellings used to serialize implicit intersection
# and spilled-array references; neither is a runtime-function candidate. The
# selected bare XLM spellings below are already handled by dedicated stored
# definition boundaries, so they are excluded instead of producing a duplicate
# generic UDF signal.
_EXCEL_UNQUALIFIED_NATIVE_FUNCTIONS = frozenset(
    """
    ABS ACCRINT ACCRINTM ACOS ACOSH ACOT ACOTH ADDRESS AGGREGATE AMORDEGRC AMORLINC AND
    ARABIC AREAS ARRAYTOTEXT ASC ASIN ASINH ATAN ATAN2 ATANH AVEDEV AVERAGE AVERAGEA
    AVERAGEIF AVERAGEIFS BAHTTEXT BASE BESSELI BESSELJ BESSELK BESSELY BETADIST BETAINV
    BIN2DEC BIN2HEX BIN2OCT BINOMDIST BITAND BITLSHIFT BITOR BITRSHIFT BITXOR BYCOL BYROW
    CALL CEILING CELL CHAR CHIDIST CHIINV CHITEST CHOOSE CHOOSECOLS CHOOSEROWS CLEAN CODE
    COLUMN COLUMNS COMBIN COMBINA COMPLEX CONCAT CONCATENATE CONFIDENCE CONVERT COPILOT
    CORREL COS COSH COT COTH COUNT COUNTA COUNTBLANK COUNTIF COUNTIFS COUPDAYBS COUPDAYS
    COUPDAYSNC COUPNCD COUPNUM COUPPCD COVAR CRITBINOM CSC CSCH CUBEKPIMEMBER CUBEMEMBER
    CUBEMEMBERPROPERTY CUBERANKEDMEMBER CUBESET CUBESETCOUNT CUBEVALUE CUMIPMT CUMPRINC
    DATE DATEDIF DATEVALUE DAVERAGE DAY DAYS DAYS360 DB DBCS DCOUNT DCOUNTA DDB DEC2BIN
    DEC2HEX DEC2OCT DECIMAL DEGREES DELTA DETECTLANGUAGE DEVSQ DGET DISC DMAX DMIN DOLLAR
    DOLLARDE DOLLARFR DPRODUCT DROP DSTDEV DSTDEVP DSUM DURATION DVAR DVARP EDATE EFFECT
    ENCODEURL EOMONTH ERF ERFC EUROCONVERT EVEN EXACT EXP EXPAND EXPONDIST FACT FACTDOUBLE
    FALSE FDIST FIELDVALUE FILTER FILTERXML FIND FINDB FINV FISHER FISHERINV FIXED FLOOR
    FORECAST FORMULATEXT FREQUENCY FTEST FV FVSCHEDULE GAMMA GAMMADIST GAMMAINV GAMMALN
    GAUSS GCD GEOMEAN GESTEP GETPIVOTDATA GROUPBY GROWTH HARMEAN HEX2BIN HEX2DEC HEX2OCT
    HLOOKUP HOUR HSTACK HYPERLINK HYPGEOMDIST IF IFERROR IFNA IFS IMABS IMAGE IMAGINARY
    IMARGUMENT IMCONJUGATE IMCOS IMCOSH IMCOT IMCSC IMCSCH IMDIV IMEXP IMLN IMLOG10 IMLOG2
    IMPOWER IMPRODUCT IMREAL IMSEC IMSECH IMSIN IMSINH IMSQRT IMSUB IMSUM IMTAN INDEX
    INDIRECT INFO INT INTERCEPT INTRATE IPMT IRR ISBLANK ISERR ISERROR ISEVEN ISFORMULA
    ISLOGICAL ISNA ISNONTEXT ISNUMBER ISODD ISOMITTED ISOWEEKNUM ISPMT ISREF ISTEXT JIS
    KURT LAMBDA LARGE LCM LEFT LEFTB LEN LENB LET LINEST LN LOG LOG10 LOGEST LOGINV
    LOGNORMDIST LOOKUP LOWER MAKEARRAY MAP MATCH MAX MAXA MAXIFS MDETERM MDURATION MEDIAN
    MID MIDB MIN MINA MINIFS MINUTE MINVERSE MIRR MMULT MOD MODE MONTH MROUND MULTINOMIAL
    MUNIT N NA NEGBINOMDIST NETWORKDAYS NOMINAL NORMDIST NORMINV NORMSDIST NORMSINV NOT
    NOW NPER NPV NUMBERVALUE OCT2BIN OCT2DEC OCT2HEX ODD ODDFPRICE ODDFYIELD ODDLPRICE
    ODDLYIELD OFFSET OR PDURATION PEARSON PERCENTILE PERCENTOF PERCENTRANK PERMUT
    PERMUTATIONA PHI PHONETIC PI PIVOTBY PMT POISSON POWER PPMT PRICE PRICEDISC PRICEMAT
    PROB PRODUCT PROPER PV QUARTILE QUOTIENT RADIANS RAND RANDARRAY RANDBETWEEN RANK RATE
    RECEIVED REDUCE REGEXEXTRACT REGEXREPLACE REGEXTEST REPLACE REPLACEB REPT RIGHT RIGHTB
    ROMAN ROUND ROUNDDOWN ROUNDUP ROW ROWS RRI RSQ RTD SCAN SEARCH SEARCHB SEC SECH SECOND
    SEQUENCE SERIESSUM SHEET SHEETS SIGN SIN SINH SKEW SLN SLOPE SMALL SORT SORTBY SQRT
    SQRTPI STANDARDIZE STDEV STDEVA STDEVP STDEVPA STEYX STOCKHISTORY SUBSTITUTE SUBTOTAL
    SUM SUMIF SUMIFS SUMPRODUCT SUMSQ SUMX2MY2 SUMX2PY2 SUMXMY2 SWITCH SYD T TAKE TAN TANH
    TBILLEQ TBILLPRICE TBILLYIELD TDIST TEXT TEXTAFTER TEXTBEFORE TEXTJOIN TEXTSPLIT TIME
    TIMEVALUE TINV TOCOL TODAY TOROW TRANSLATE TRANSPOSE TREND TRIM TRIMMEAN TRIMRANGE TRUE
    TRUNC TTEST TYPE UNICHAR UNICODE UNIQUE UPPER VALUE VALUETOTEXT VAR VARA VARP VARPA VDB
    VLOOKUP VSTACK WEBSERVICE WEEKDAY WEEKNUM WEIBULL WORKDAY WRAPCOLS WRAPROWS XIRR
    XLOOKUP XMATCH XNPV XOR YEAR YEARFRAC YIELD YIELDDISC YIELDMAT ZTEST
    PY SINGLE ANCHORARRAY EVALUATE EXEC EXECUTE REGISTER RUN SEND
    """.split()
)

_CELL_REFERENCE = re.compile(
    r"(?<![A-Z0-9_])(?P<column_absolute>\$?)(?P<column>[A-Z]{1,3})"
    r"(?P<row_absolute>\$?)(?P<row>[1-9][0-9]*)(?![A-Z0-9_])",
    re.IGNORECASE,
)
_LOCAL_IDENTIFIER = re.compile(r"[A-Z_][A-Z0-9_]*", re.IGNORECASE)
_A1_LOCAL_IDENTIFIER_CONFLICT = re.compile(r"[A-Z]{1,3}[1-9][0-9]*", re.IGNORECASE)
_R1C1_LOCAL_IDENTIFIER_CONFLICT = re.compile(
    r"R[1-9][0-9]*C[1-9][0-9]*", re.IGNORECASE
)
_SERIALIZED_LOCAL_PREFIXES = ("_xlpm.", "_xlop.")
_GROUP_TOKEN_TYPES = {"FUNC", "PAREN", "ARRAY"}
_WHITESPACE_TOKEN_TYPES = {"WSPACE", "WHITE-SPACE"}
_SPILL_REFERENCE_FUNCTION = "ANCHORARRAY"
_IMPLICIT_INTERSECTION_FUNCTION = "SINGLE"
# This private tokenizer-only wrapper preserves the semantic distinction of a
# literal ``@A1`` while avoiding an ambiguity with a user-authored SINGLE().
_LITERAL_IMPLICIT_INTERSECTION_FUNCTION = "_FORMULAFENCE_IMPLICIT_INTERSECTION"
# ``#`` is not understood by openpyxl's tokenizer, even though it is Excel's
# display syntax for a spilled-array reference. Keep the accepted grammar
# intentionally narrow: one internal A1 anchor, optionally sheet-qualified.
# External, 3-D, range, named, malformed, and implicit-intersection variants
# remain tokenizer coverage limits instead of being rewritten into a guess.
_LITERAL_SPILL_REFERENCE = re.compile(
    r"(?<![A-Z0-9_.'\"!:@$\[\]])"
    r"(?P<reference>(?:(?:'(?:[^']|'')*'|[A-Z0-9_.]+)!)?"
    r"\$?[A-Z]{1,3}\$?[1-9][0-9]*)"
    r"#(?=$|[ \t\r\n,;)}+\-*/^&=<>%])",
    re.IGNORECASE,
)
# Like ``#``, Excel's display-only implicit-intersection operator is outside
# openpyxl's reference grammar. Only rewrite a direct, internal A1 reference or
# A1 range. Table current-row syntax, named expressions, external references,
# 3-D spans, spill combinations, and malformed forms remain untouched so they
# can never be mistaken for a static dependency.
_LITERAL_IMPLICIT_INTERSECTION_REFERENCE = re.compile(
    r"(?<![A-Z0-9_.'\"!:@$\[\]])"
    r"@(?P<reference>(?:(?:'(?:[^']|'')*'|[A-Z0-9_.]+)!)?"
    r"\$?[A-Z]{1,3}\$?[1-9][0-9]*(?::\$?[A-Z]{1,3}\$?[1-9][0-9]*)?)"
    r"(?=$|[ \t\r\n,;)}+\-*/^&=<>%])",
    re.IGNORECASE,
)
# Excel's conditional ``IFS`` aggregates require their relevant range arguments
# to have the same dimensions. Keep this grammar much narrower than the general
# reference parser: it admits one internal A1 cell/range or whole-column range,
# with an optional simple sheet qualifier. Named expressions, structured
# references, external and 3-D references, full rows, unions, computed
# references, and display-only spill/intersection syntax intentionally remain
# outside this lint's static evidence boundary.
_DIRECT_STATIC_A1_RANGE_ARGUMENT = re.compile(
    r"^(?:(?:'(?:[^']|'')*'|[A-Z0-9_.]+)!)?"
    r"(?:\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6}"
    r"(?::\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6})?"
    r"|\$?[A-Z]{1,3}:\$?[A-Z]{1,3})$",
    re.IGNORECASE,
)
_CONDITIONAL_AGGREGATE_RANGE_SHAPE_FUNCTIONS = frozenset(
    {"SUMIFS", "COUNTIFS", "AVERAGEIFS", "MAXIFS", "MINIFS"}
)
_CONDITIONAL_AGGREGATE_VALUE_RANGE_FUNCTIONS = frozenset(
    {"SUMIFS", "AVERAGEIFS", "MAXIFS", "MINIFS"}
)
# MAXIFS and MINIFS are OOXML future functions.  Microsoft documents these two
# exact serialized spellings in the XLSX function list, so recognize them as
# their native counterparts without generally folding arbitrary namespaces.
_OOXML_FUTURE_CONDITIONAL_AGGREGATE_FUNCTION_NAMES = {
    "_XLFN.MAXIFS": "MAXIFS",
    "_XLFN.MINIFS": "MINIFS",
}


@dataclass(frozen=True)
class ParsedReference:
    """An explicit A1-style reference found in a formula token."""

    sheet: str | None
    min_column: int | None
    min_row: int | None
    max_column: int | None
    max_row: int | None
    raw: str
    is_external: bool = False

    @property
    def is_range(self) -> bool:
        return (
            self.min_column != self.max_column
            or self.min_row != self.max_row
        )


@dataclass(frozen=True)
class IndexedExternalWorkbookReference:
    """One package-indexed external A1 reference without a source path.

    Excel's ``[N]Sheet!A1`` spelling uses ``N`` as a one-based position in
    the workbook's external-link collection, rather than a workbook filename.
    Formula parsing cannot safely resolve that position by itself, so this
    private-lookup intermediary intentionally keeps only the bounded index and
    static A1 destination until package metadata supplies an exact target.
    """

    index: int
    sheet: str
    min_column: int
    min_row: int
    max_column: int
    max_row: int


@dataclass(frozen=True)
class IndexedExternalWorkbookThreeDReference:
    """One package-indexed external A1 span without a source path.

    The nonzero index is a position in the workbook's external-link
    collection. The source path remains unavailable until a separately
    validated package declaration supplies it; the two worksheet endpoints
    are retained privately for the same candidate-only portfolio bridge.
    """

    index: int
    first_sheet: str
    last_sheet: str
    min_column: int
    min_row: int
    max_column: int
    max_row: int


@dataclass(frozen=True)
class IndexedExternalWorkbookStructuredReference:
    """One package-indexed external table selector without a source path.

    A nonzero book index becomes meaningful only after the workbook package
    validates its external-link declaration. Keep the table identity and its
    exact static selector private until that separate candidate-only bridge
    supplies a source path.
    """

    index: int
    table_name: str
    table_reference: str


@dataclass(frozen=True)
class FormulaInspection:
    """Static formula coverage collected from one formula without evaluation."""

    references: tuple[ParsedReference, ...]
    unresolved_range_tokens: tuple[str, ...]
    dynamic_reference_functions: tuple[str, ...]
    external_workbook_references: tuple[ExternalWorkbookReference, ...] = ()
    external_workbook_three_d_references: tuple[
        ExternalWorkbookThreeDReference, ...
    ] = ()
    external_workbook_structured_references: tuple[
        ExternalWorkbookStructuredReference, ...
    ] = ()
    external_workbook_defined_name_references: tuple[
        ExternalWorkbookDefinedNameReference, ...
    ] = ()
    external_action_functions: tuple[str, ...] = ()
    formula_dde_link_markers: tuple[str, ...] = ()
    python_functions: tuple[str, ...] = ()
    office_custom_function_candidates: tuple[str, ...] = ()
    unqualified_runtime_function_candidates: tuple[str, ...] = ()
    worksheet_code_resource_registration_functions: tuple[str, ...] = ()
    formula_defined_xlm_registration_functions: tuple[str, ...] = ()
    formula_defined_xlm_evaluation_functions: tuple[str, ...] = ()
    formula_defined_xlm_action_functions: tuple[str, ...] = ()
    formula_defined_xlm_get_cell_functions: tuple[str, ...] = ()
    formula_defined_xlm_environment_information_functions: tuple[str, ...] = ()
    formula_environment_information_functions: tuple[str, ...] = ()
    formula_environment_information_implicit_cell_reference_count: int = 0
    formula_environment_information_implicit_sheets_reference_count: int = 0
    three_d_reference_tokens: tuple[str, ...] = ()
    tokenization_failed: bool = False
    spill_reference_tokens: tuple[str, ...] = ()
    implicit_intersection_tokens: tuple[str, ...] = ()

    @property
    def formula_environment_information_function_count(self) -> int:
        """Return native information calls without exposing their arguments."""
        return len(self.formula_environment_information_functions)

    @property
    def formula_dde_link_count(self) -> int:
        """Return statically visible direct DDE links without their endpoints."""
        return len(self.formula_dde_link_markers)

    @property
    def formula_environment_information_signal_values(self) -> tuple[str, ...]:
        """Return private propagation values for named-formula resolution.

        The marker is deliberately not part of
        ``formula_environment_information_functions`` so normal consumers see
        only native function names. They remain necessary internally to retain
        statically visible omitted-reference distinctions through a named formula
        or named LAMBDA invocation chain.
        """
        return (
            self.formula_environment_information_functions
            + (
                _FORMULA_ENVIRONMENT_INFORMATION_IMPLICIT_CELL_REFERENCE_MARKER,
            )
            * self.formula_environment_information_implicit_cell_reference_count
            + (
                _FORMULA_ENVIRONMENT_INFORMATION_IMPLICIT_SHEETS_REFERENCE_MARKER,
            )
            * self.formula_environment_information_implicit_sheets_reference_count
        )


@dataclass(frozen=True)
class StructuredTable:
    """Static metadata needed to resolve conservative Excel-table references."""

    name: str
    sheet: str
    ref: str
    columns: tuple[str, ...]
    header_row_count: int
    totals_row_count: int


def _last_unquoted_bang(value: str) -> int:
    """Find the separator in `'A sheet'!A1` while respecting Excel quote escapes."""
    quoted = False
    last = -1
    index = 0
    while index < len(value):
        character = value[index]
        if character == "'":
            if quoted and index + 1 < len(value) and value[index + 1] == "'":
                index += 2
                continue
            quoted = not quoted
        elif character == "!" and not quoted:
            last = index
        index += 1
    return last


def _split_sheet_reference(value: str) -> tuple[str | None, str]:
    separator = _last_unquoted_bang(value)
    if separator < 0:
        return None, value
    return value[:separator], value[separator + 1 :]


def _normalise_sheet_name(value: str | None) -> tuple[str | None, bool]:
    """Return a local sheet title, or mark explicit external workbook syntax."""
    if value is None:
        return None, False
    value = value.strip()
    if "[" in value and "]" in value:
        return None, True
    if len(value) >= 2 and value[0] == "'" and value[-1] == "'":
        value = value[1:-1].replace("''", "'")
    return value, False


def _relative_cell_reference(match: re.Match[str], origin: str) -> str:
    origin_row, origin_column = coordinate_to_tuple(origin)
    column = column_index_from_string(match.group("column").upper())
    row = int(match.group("row"))
    if match.group("row_absolute"):
        normalised_row = f"R{row}"
    else:
        normalised_row = f"R[{row - origin_row}]"
    if match.group("column_absolute"):
        normalised_column = f"C{column}"
    else:
        normalised_column = f"C[{column - origin_column}]"
    return f"{normalised_row}{normalised_column}"


def _normalise_range_token(value: str, origin: str) -> str:
    sheet, address = _split_sheet_reference(value)
    if sheet is not None:
        # Sheet names are case-insensitive in Excel. Preserve quote semantics only
        # where they are needed to make a fingerprint readable.
        clean_sheet, external = _normalise_sheet_name(sheet)
        if external:
            sheet_prefix = "[EXTERNAL]!"
        else:
            sheet_prefix = f"{(clean_sheet or '').casefold()}!"
    else:
        sheet_prefix = ""
    normalised_address = _CELL_REFERENCE.sub(
        lambda match: _relative_cell_reference(match, origin), address.upper()
    )
    return f"{sheet_prefix}{normalised_address}"


def formula_fingerprint(formula: str, origin: str) -> str:
    """Make copied formulas comparable without losing absolute-reference intent.

    FormulaFence does not attempt to parse every Excel function. It relies on
    openpyxl's tokenizer and only rewrites A1 references in `OPERAND/RANGE`
    tokens, leaving literals and function arguments untouched.
    """
    tokens, _, _ = _tokenize_formula(formula, preserve_literal_spill_operator=True)
    if tokens is None:  # pragma: no cover - defensive against malformed formulas
        return formula.strip()

    parts: list[str] = []
    for token in tokens:
        if _is_whitespace(token):
            continue
        if token.type == "OPERAND" and token.subtype == "RANGE":
            parts.append(_normalise_range_token(token.value, origin))
        else:
            parts.append(_fingerprint_token_value(token))
    return "".join(parts)


def parse_reference_token(value: str) -> ParsedReference | None:
    """Parse an A1 token such as `'Inputs'!$B$2:$B$20` without resolving names."""
    sheet, address = _split_sheet_reference(value)
    normalised_sheet, external = _normalise_sheet_name(sheet)
    if external:
        return ParsedReference(None, None, None, None, None, value, is_external=True)
    if normalised_sheet is not None and ":" in normalised_sheet:
        # A 3-D reference needs workbook tab order to resolve safely. Returning
        # None here prevents an invented ``Sheet1:Sheet3`` pseudo-sheet edge.
        return None

    # Table references and named ranges arrive as RANGE tokens too. Their
    # resolution is deliberately handled by inspect_formula, where a workbook
    # supplied name map is available.
    try:
        min_column, min_row, max_column, max_row = range_boundaries(address)
    except ValueError:
        return None

    min_column = min_column or 1
    max_column = max_column or MAX_EXCEL_COLUMN
    min_row = min_row or 1
    max_row = max_row or MAX_EXCEL_ROW
    return ParsedReference(
        normalised_sheet,
        min_column,
        min_row,
        max_column,
        max_row,
        value,
    )


def _parse_external_link_index(value: str) -> int | None:
    """Parse one bounded, nonzero external-link collection position."""
    # Bound parsing before converting an untrusted arbitrary-length decimal.
    if (
        len(value) > 10
        or not value.isascii()
        or not value.isdecimal()
        or value.startswith("0")
    ):
        return None
    index = int(value)
    return index if index <= 2_147_483_647 else None


def _unescape_external_reference_prefix(value: str) -> tuple[str, bool] | None:
    """Unquote one Excel external-reference prefix without accepting noise."""
    quoted = value.startswith("'")
    if not quoted:
        return (value, False) if "'" not in value else None
    if len(value) < 2 or not value.endswith("'"):
        return None
    unescaped: list[str] = []
    position = 1
    while position < len(value) - 1:
        character = value[position]
        if character == "'":
            if position + 1 < len(value) - 1 and value[position + 1] == "'":
                unescaped.append("'")
                position += 2
                continue
            return None
        unescaped.append(character)
        position += 1
    return "".join(unescaped), True


def _is_static_external_sheet_name(value: str, *, quoted: bool) -> bool:
    """Return whether one external sheet endpoint has an exact static spelling."""
    return not (
        not value
        or any(character in "[]\\?*/:" for character in value)
        or any(ord(character) < 32 or ord(character) == 127 for character in value)
        or (
            not quoted
            and any(
                character.isspace() or character in "'+-*/^&=<>%,;(){}!"
                for character in value
            )
        )
    )


def _split_static_external_sheet_range(
    value: str, *, quoted: bool
) -> tuple[str, str] | None:
    """Split two exact external worksheet endpoints without inferring either."""
    first_sheet, separator, last_sheet = value.partition(":")
    if (
        not separator
        or not first_sheet
        or not last_sheet
        or ":" in last_sheet
        or not _is_static_external_sheet_name(first_sheet, quoted=quoted)
        or not _is_static_external_sheet_name(last_sheet, quoted=quoted)
    ):
        return None
    return first_sheet, last_sheet


def _parse_external_link_indexed_sheet_prefix_parts(
    value: str,
) -> tuple[int, str, str, bool] | None:
    """Split an indexed external prefix before selecting single or 3-D sheets."""
    token = value.strip()
    if token.startswith("="):
        token = token[1:].strip()
    sheet_prefix, payload = _split_sheet_reference(token)
    if sheet_prefix is None:
        return None
    unescaped_prefix = _unescape_external_reference_prefix(sheet_prefix.strip())
    if unescaped_prefix is None:
        return None
    prefix, quoted = unescaped_prefix
    if not prefix.startswith("["):
        return None
    closing = prefix.find("]", 1)
    if closing < 2:
        return None
    index = _parse_external_link_index(prefix[1:closing])
    sheet = prefix[closing + 1 :]
    if index is None:
        return None
    return index, sheet, payload, quoted


def _parse_external_link_indexed_sheet_prefix(
    value: str,
) -> tuple[int, str, str] | None:
    """Split one strict ``[N]Sheet!`` prefix without assigning its payload.

    Excel reuses ``single-sheet-prefix`` for both an external A1 destination
    and an external sheet-local defined name.  Keep the quote and sheet-name
    checks in one place so the two parsers cannot disagree about which private
    package index or source scope they observed.
    """
    parsed = _parse_external_link_indexed_sheet_prefix_parts(value)
    if parsed is None:
        return None
    index, sheet, payload, quoted = parsed
    if not _is_static_external_sheet_name(sheet, quoted=quoted):
        return None
    return index, sheet, payload


def _parse_direct_external_workbook_sheet_prefix_parts(
    value: str,
) -> tuple[str, str, str, bool] | None:
    """Split one exact direct external prefix before selecting sheet form.

    Defined-name OOXML commonly stores a static target with no leading
    equals sign, while some producers retain one.  Accept both spellings, but
    only after the remaining value is still exactly one quoted or unquoted
    external literal.  The returned source path is private lookup data: its
    later portfolio resolver decides whether it is a safe in-root relative
    target.
    """
    token = value.strip()
    had_leading_equals = token.startswith("=")
    if had_leading_equals:
        token = token[1:].strip()
        # With an explicit formula marker, a static external literal begins
        # with its external-book prefix (or its enclosing apostrophe).  This
        # avoids treating e.g. ``=SUM(...)`` or unary-operator formulas as a
        # direct named-range target.
        if not token.startswith(("[", "'")):
            return None
    sheet_prefix, payload = _split_sheet_reference(token)
    if sheet_prefix is None:
        return None
    unescaped_prefix = _unescape_external_reference_prefix(sheet_prefix.strip())
    if unescaped_prefix is None:
        return None
    prefix, quoted = unescaped_prefix

    opening = prefix.find("[")
    closing = prefix.find("]", opening + 1)
    if opening < 0 or closing < opening + 2:
        return None
    source_prefix = prefix[:opening]
    workbook_name = prefix[opening + 1 : closing]
    sheet = prefix[closing + 1 :]
    if (
        not workbook_name
        or workbook_name != workbook_name.strip()
        or not sheet
        or "[" in source_prefix
        or "]" in source_prefix
        or any(character in "[]\\?*/:" for character in workbook_name)
        or any(
            ord(character) < 32 or ord(character) == 127
            for character in f"{source_prefix}{workbook_name}{sheet}"
        )
        # A path prefix is either absent or explicitly directory-like.  This
        # prevents a formula operator immediately before ``[Book]`` from being
        # reinterpreted as a filename prefix while retaining normal relative,
        # absolute, UNC, and URI-like source spellings for the later resolver
        # to reject or bound as appropriate.
        or (source_prefix and not source_prefix.endswith(("/", "\\")))
        or (
            not quoted
            and any(character in "'+-*/^&=<>%,;(){}!" for character in source_prefix)
        )
        # A numeric ``[N]`` prefix is a package external-link index, never a
        # direct filename. It must be resolved only after package validation.
        or (
            not source_prefix
            and workbook_name.isascii()
            and workbook_name.isdecimal()
        )
    ):
        return None
    return f"{source_prefix}{workbook_name}", sheet, payload, quoted


def _parse_direct_external_workbook_sheet_prefix(
    value: str,
) -> tuple[str, str, str] | None:
    """Split one strict direct external ``[Book]Sheet!`` literal."""
    parsed = _parse_direct_external_workbook_sheet_prefix_parts(value)
    if parsed is None:
        return None
    source_path, sheet, payload, quoted = parsed
    if not _is_static_external_sheet_name(sheet, quoted=quoted):
        return None
    return source_path, sheet, payload


def parse_external_workbook_reference(value: str) -> ExternalWorkbookReference | None:
    """Parse a direct external A1 token without resolving its filesystem path.

    Excel formulas may spell an external reference as ``[Book.xlsx]Sheet!A1``
    or include a relative/absolute source path before the workbook name. The
    path is intentionally retained only for a later, bounded portfolio-root
    resolver. External names, structured references, 3-D spans, and malformed
    A1 syntax return ``None`` rather than being approximated.
    """
    parsed_prefix = _parse_direct_external_workbook_sheet_prefix(value)
    if parsed_prefix is None:
        return None
    source_path, sheet, address = parsed_prefix
    try:
        min_column, min_row, max_column, max_row = range_boundaries(address)
    except ValueError:
        return None
    return ExternalWorkbookReference(
        source_path=source_path,
        sheet=sheet,
        min_column=min_column or 1,
        min_row=min_row or 1,
        max_column=max_column or MAX_EXCEL_COLUMN,
        max_row=max_row or MAX_EXCEL_ROW,
    )


def parse_external_link_indexed_workbook_reference(
    value: str,
) -> IndexedExternalWorkbookReference | None:
    """Parse Excel's package-indexed external A1 syntax without resolving it.

    Office stores an external cell reference as ``[N]Sheet!A1`` (or a quoted
    sheet spelling such as ``'[N]Input Sheet'!A1``), where nonzero ``N`` is a
    position in the workbook's external-link collection.  This helper accepts
    only one static A1 destination.  Names, 3-D spans, malformed quoting, and
    invalid sheet syntax stay outside the static portfolio graph rather than
    being approximated.
    """
    parsed_prefix = _parse_external_link_indexed_sheet_prefix(value)
    if parsed_prefix is None:
        return None
    index, sheet, address = parsed_prefix
    try:
        min_column, min_row, max_column, max_row = range_boundaries(address)
    except ValueError:
        return None
    return IndexedExternalWorkbookReference(
        index=index,
        sheet=sheet,
        min_column=min_column or 1,
        min_row=min_row or 1,
        max_column=max_column or MAX_EXCEL_COLUMN,
        max_row=max_row or MAX_EXCEL_ROW,
    )


def parse_external_workbook_three_d_reference(
    value: str,
) -> ExternalWorkbookThreeDReference | None:
    """Parse one exact direct external 3-D A1 reference without resolving it.

    A source workbook can use ``[Book.xlsx]Jan:Mar!A1`` (or its quoted,
    path-bearing equivalent) to select the same A1 destination across every
    worksheet between two endpoint tabs. This parser retains only static A1
    bounds and both endpoint identities. Candidate portfolio analysis later
    decides whether a private path resolves to an inspected source and whether
    its real worksheet order makes the span unambiguous.
    """
    parsed_prefix = _parse_direct_external_workbook_sheet_prefix_parts(value)
    if parsed_prefix is None:
        return None
    source_path, sheet_range, address, quoted = parsed_prefix
    sheet_endpoints = _split_static_external_sheet_range(sheet_range, quoted=quoted)
    if sheet_endpoints is None:
        return None
    try:
        min_column, min_row, max_column, max_row = range_boundaries(address)
    except ValueError:
        return None
    first_sheet, last_sheet = sheet_endpoints
    return ExternalWorkbookThreeDReference(
        source_path=source_path,
        first_sheet=first_sheet,
        last_sheet=last_sheet,
        min_column=min_column or 1,
        min_row=min_row or 1,
        max_column=max_column or MAX_EXCEL_COLUMN,
        max_row=max_row or MAX_EXCEL_ROW,
    )


def parse_external_link_indexed_workbook_three_d_reference(
    value: str,
) -> IndexedExternalWorkbookThreeDReference | None:
    """Parse one package-indexed external 3-D A1 reference without resolving it.

    Office uses the same ``[N]`` external-link collection index for a
    single-sheet prefix and a sheet range. Keep only a nonzero, bounded index,
    static A1 bounds, and exact endpoints until validated package metadata
    provides a candidate-only source path.
    """
    parsed_prefix = _parse_external_link_indexed_sheet_prefix_parts(value)
    if parsed_prefix is None:
        return None
    index, sheet_range, address, quoted = parsed_prefix
    sheet_endpoints = _split_static_external_sheet_range(sheet_range, quoted=quoted)
    if sheet_endpoints is None:
        return None
    try:
        min_column, min_row, max_column, max_row = range_boundaries(address)
    except ValueError:
        return None
    first_sheet, last_sheet = sheet_endpoints
    return IndexedExternalWorkbookThreeDReference(
        index=index,
        first_sheet=first_sheet,
        last_sheet=last_sheet,
        min_column=min_column or 1,
        min_row=min_row or 1,
        max_column=max_column or MAX_EXCEL_COLUMN,
        max_row=max_row or MAX_EXCEL_ROW,
    )


def _static_external_structured_table_reference(
    value: str,
) -> tuple[str, str] | None:
    """Return one static table identity and selector, without resolving it.

    The OOXML structured-reference grammar puts an external book prefix before
    a table name, not before a source worksheet.  Reuse the ordinary static
    table parser for the selector, but reject row-relative forms here: an
    external source table has no safe relationship to the consuming formula's
    row.  The original selector stays private until candidate portfolio
    analysis can resolve it against exactly one inspected source table.
    """
    # A bare ``[N]!InputRange`` is indistinguishable from the longstanding
    # package-indexed external defined-name form. Require a selector bracket
    # so the new table boundary never steals that established interpretation.
    if "[" not in value:
        return None
    parsed = _structured_reference_parts(value)
    if parsed is None:
        return None
    table_name, groups, _ = parsed
    if (
        not table_name
        or len(table_name) > 255
        or table_name != table_name.strip()
        or not _is_external_link_name_identity(table_name)
        or parse_reference_token(table_name) is not None
    ):
        return None
    if any(
        group.strip().casefold() in {"#this row", "@"}
        or group.strip().startswith("@")
        for group in groups
    ):
        return None
    return table_name, value.strip()


def _is_static_direct_external_workbook_path(
    source_path: str, *, quoted: bool
) -> bool:
    """Check a direct book-only path before later bounded resolution.

    This is syntax validation, not filesystem validation.  The portfolio
    resolver remains responsible for rejecting absolute paths, URIs, and any
    target outside its already inspected candidate set.  Require a supported
    workbook suffix here so a local ``Sheet!Table`` spelling cannot be
    reclassified as an external-table source merely because its sheet name is
    name-like.
    """
    if (
        not source_path
        or source_path != source_path.strip()
        or "[" in source_path
        or "]" in source_path
        or any(character in "?*" for character in source_path)
        or any(ord(character) < 32 or ord(character) == 127 for character in source_path)
        or not source_path.casefold().endswith((".xlsx", ".xlsm"))
    ):
        return False
    if not quoted and any(
        character in "'+-*/^&=<>%,;(){}!" for character in source_path
    ):
        return False
    return True


def _parse_direct_external_workbook_only_prefix(
    value: str,
) -> tuple[str, str] | None:
    """Split a direct book-only prefix used by an external table reference.

    Excel's direct table-link spelling is ``'path/book.xlsx'!Table[Column]``
    or ``'[book.xlsx]'!Table[Column]``.  Unlike external A1 syntax it has no
    source-sheet suffix.  Keeping those forms separate avoids inventing a
    sheet relationship for a table identifier and rejects the tempting but
    nonstandard ``[book.xlsx]Sheet!Table[...]`` form.
    """
    token = value.strip()
    if token.startswith("="):
        token = token[1:].strip()
        if not token.startswith(("[", "'")):
            return None
    prefix, payload = _split_sheet_reference(token)
    if prefix is None or not payload:
        return None
    unescaped_prefix = _unescape_external_reference_prefix(prefix.strip())
    if unescaped_prefix is None:
        return None
    raw_prefix, quoted = unescaped_prefix
    if any(ord(character) < 32 or ord(character) == 127 for character in raw_prefix):
        return None

    opening = raw_prefix.find("[")
    if opening >= 0:
        closing = raw_prefix.find("]", opening + 1)
        if closing < opening + 2 or closing != len(raw_prefix) - 1:
            return None
        source_prefix = raw_prefix[:opening]
        workbook_name = raw_prefix[opening + 1 : closing]
        if (
            not workbook_name
            or workbook_name != workbook_name.strip()
            or "[" in source_prefix
            or "]" in source_prefix
            or any(character in "[]\\?*/:" for character in workbook_name)
            or (source_prefix and not source_prefix.endswith(("/", "\\")))
            or (
                not quoted
                and any(
                    character in "'+-*/^&=<>%,;(){}!" for character in source_prefix
                )
            )
            or (
                not source_prefix
                and
                workbook_name.isascii()
                and workbook_name.isdecimal()
            )
        ):
            return None
        source_path = f"{source_prefix}{workbook_name}"
    else:
        # An unbracketed source book is accepted only when it is quoted. This
        # keeps a local name-like sheet prefix from being reinterpreted as a
        # path while retaining Excel's documented direct table-link spelling.
        if not quoted:
            return None
        source_path = raw_prefix

    if not _is_static_direct_external_workbook_path(source_path, quoted=quoted):
        return None
    return source_path, payload


def parse_external_workbook_structured_reference(
    value: str,
) -> ExternalWorkbookStructuredReference | None:
    """Parse one direct static external table selector without resolving it.

    A direct external structured reference uses a book-only prefix, for
    example ``'../inputs/source.xlsx'!Sales[#Data]``.  The source path, table
    identity, and selector are private candidate lookup data; no file is
    opened here and no table calculation is attempted.
    """
    parsed_prefix = _parse_direct_external_workbook_only_prefix(value)
    if parsed_prefix is None:
        return None
    source_path, selector = parsed_prefix
    parsed_selector = _static_external_structured_table_reference(selector)
    if parsed_selector is None:
        return None
    table_name, table_reference = parsed_selector
    return ExternalWorkbookStructuredReference(
        source_path=source_path,
        table_name=table_name,
        table_reference=table_reference,
    )


def parse_external_link_indexed_workbook_structured_reference(
    value: str,
) -> IndexedExternalWorkbookStructuredReference | None:
    """Parse one package-indexed external table selector without resolving it.

    The standards form ``[N]!Table[Column]`` uses an external-link collection
    position, not a filename.  Keep only a nonzero index and static table
    selector until raw package metadata validates one candidate source path.
    Sheet-qualified forms deliberately remain outside this parser because an
    external structured reference's table identifier is workbook-scoped.
    """
    token = value.strip()
    if token.startswith("="):
        token = token[1:].strip()
    prefix, selector = _split_sheet_reference(token)
    if prefix is None or not selector:
        return None
    unescaped_prefix = _unescape_external_reference_prefix(prefix.strip())
    if unescaped_prefix is None:
        return None
    raw_prefix, quoted = unescaped_prefix
    if quoted or not raw_prefix.startswith("[") or not raw_prefix.endswith("]"):
        return None
    index = _parse_external_link_index(raw_prefix[1:-1])
    if index is None:
        return None
    parsed_selector = _static_external_structured_table_reference(selector)
    if parsed_selector is None:
        return None
    table_name, table_reference = parsed_selector
    return IndexedExternalWorkbookStructuredReference(
        index=index,
        table_name=table_name,
        table_reference=table_reference,
    )


def parse_external_workbook_defined_name_reference(
    value: str,
) -> ExternalWorkbookDefinedNameReference | None:
    """Parse one direct external workbook-scoped name without resolving it.

    Excel permits ``=[Book.xlsx]InputRange`` alongside direct external A1
    syntax. This accepts only that workbook-scoped form. Sheet-qualified,
    structured, malformed, and otherwise ambiguous spellings remain outside
    the static portfolio graph instead of being approximated.
    """
    token = value.strip()
    had_leading_equals = token.startswith("=")
    if had_leading_equals:
        token = token[1:].strip()
        if not token.startswith(("[", "'")):
            return None
    if not token or _last_unquoted_bang(token) >= 0:
        return None
    unescaped_token = _unescape_external_reference_prefix(token)
    if unescaped_token is None:
        return None
    token, quoted = unescaped_token

    opening = token.find("[")
    closing = token.find("]", opening + 1)
    if opening < 0 or closing < opening + 2:
        return None
    source_prefix = token[:opening]
    workbook_name = token[opening + 1 : closing]
    name = token[closing + 1 :]
    if (
        not workbook_name
        or not name
        or workbook_name != workbook_name.strip()
        or "[" in source_prefix
        or "]" in source_prefix
        or any(character in "[]\\?*/:" for character in workbook_name)
        or any(
            ord(character) < 32 or ord(character) == 127
            for character in f"{source_prefix}{workbook_name}{name}"
        )
        or (source_prefix and not source_prefix.endswith(("/", "\\")))
        or (
            not quoted
            and any(character in "'+-*/^&=<>%,;(){}!" for character in source_prefix)
        )
        or len(name) > 255
        or name != name.strip()
        or name.startswith("\\")
        or not _is_external_link_name_identity(name)
        or parse_reference_token(name) is not None
    ):
        return None
    return ExternalWorkbookDefinedNameReference(
        source_path=f"{source_prefix}{workbook_name}",
        name_key=reference_lookup_key(name),
    )


def parse_external_workbook_sheet_defined_name_reference(
    value: str,
) -> ExternalWorkbookDefinedNameReference | None:
    """Parse one direct external sheet-local name without resolving it.

    ``[Book.xlsx]Data!LocalInput`` is distinct from both a workbook-scoped
    external name and an A1 reference: the source sheet selects that
    workbook's local defined-name scope.  Accept only one static name identity
    and a valid single sheet prefix.  Dynamic, 3-D, A1, structured, malformed,
    and package-indexed spellings stay outside this direct-path parser.
    """
    parsed_prefix = _parse_direct_external_workbook_sheet_prefix(value)
    if parsed_prefix is None:
        return None
    source_path, sheet, name = parsed_prefix
    if (
        not name
        or len(name) > 255
        or name != name.strip()
        or name.startswith("\\")
        or not _is_external_link_name_identity(name)
        or any(ord(character) < 32 or ord(character) == 127 for character in name)
        or parse_reference_token(name) is not None
    ):
        return None
    return ExternalWorkbookDefinedNameReference(
        source_path=source_path,
        name_key=reference_lookup_key(name),
        scope_sheet=sheet,
    )


def parse_workbook_defined_name_alias(value: str) -> str | None:
    """Parse one exact workbook-name alias without evaluating its formula.

    A defined-name formula can contain one ``name-reference``.  That is a
    useful, deterministic indirection layer for a workbook-scoped alias of an
    already validated external link, but it is not a licence to evaluate an
    arbitrary formula.  Accept both OOXML's common no-leading-``=`` spelling
    and writers which retain it; require the remainder to be exactly one
    unqualified, non-A1 name identity.  Sheet-local references, functions,
    operators, structured references, direct external literals, and malformed
    name text deliberately remain outside the static portfolio graph.
    """
    token = value.strip()
    if token.startswith("="):
        token = token[1:].strip()
    if (
        not token
        or len(token) > 255
        or token != token.strip()
        or token.startswith("\\")
        or any(character in "[]'!+-*/^&=<>%,;(){}" for character in token)
        or any(ord(character) < 32 or ord(character) == 127 for character in token)
        or not _is_external_link_name_identity(token)
        or parse_reference_token(token) is not None
    ):
        return None
    return reference_lookup_key(token)


def parse_external_link_indexed_defined_name_reference(
    value: str,
) -> tuple[int, str] | None:
    """Parse Excel's package-indexed external-name spelling without resolving it.

    OOXML can store an external name as ``[1]!InputRange``.  The number is
    not a filename: it is a one-based position in the workbook's
    ``externalReferences`` declaration.  Resolving that position requires
    raw package metadata, so this lexical helper deliberately returns only
    the bounded index and private normalized name identity.  Sheet/A1 forms,
    index zero (the current workbook), malformed tokens, and quoted/structured
    spellings stay outside this static path.
    """
    token = value.strip()
    if token.startswith("="):
        token = token[1:].strip()
    if not token or token[0] != "[" or "'" in token:
        return None
    closing = token.find("]", 1)
    if closing < 2 or closing + 1 >= len(token) or token[closing + 1] != "!":
        return None
    index_text = token[1:closing]
    index = _parse_external_link_index(index_text)
    if index is None:
        return None
    name = token[closing + 2 :]
    if (
        not name
        or len(name) > 255
        or name != name.strip()
        # Office does not create or load a backslash-prefixed defined name,
        # despite the broad ECMA grammar admitting a backslash start.
        or name.startswith("\\")
        or not _is_external_link_name_identity(name)
        or any(ord(character) < 32 or ord(character) == 127 for character in name)
        or parse_reference_token(name) is not None
    ):
        return None
    return index, reference_lookup_key(name)


def parse_external_link_indexed_sheet_defined_name_reference(
    value: str,
) -> tuple[int, str, str] | None:
    """Parse Excel's package-indexed sheet-local name without resolving it.

    A token such as ``[1]Data!LocalInput`` uses the external-link collection
    index and the source sheet's local defined-name scope.  It is deliberately
    accepted only when the payload is a bounded non-A1 name identity; callers
    must still validate the package mapping and candidate source workbook.
    """
    parsed_prefix = _parse_external_link_indexed_sheet_prefix(value)
    if parsed_prefix is None:
        return None
    index, sheet, name = parsed_prefix
    if (
        not name
        or len(name) > 255
        or name != name.strip()
        or name.startswith("\\")
        or not _is_external_link_name_identity(name)
        or any(ord(character) < 32 or ord(character) == 127 for character in name)
        or parse_reference_token(name) is not None
    ):
        return None
    return index, sheet, reference_lookup_key(name)


def _is_external_link_name_identity(value: str) -> bool:
    """Check the bounded name grammar used after Excel's ``[N]!`` prefix.

    The package-specific prefix is not a workbook filename.  Once it is
    removed, Office's external-name grammar admits a defined-name identity,
    not arbitrary formula punctuation.  Keep Unicode letters and combining
    marks usable while rejecting strings such as ``[1]!A+B`` or ``[1]!$Name``
    that could otherwise be mistaken for a name by a permissive tokenizer.
    """

    def is_start(character: str) -> bool:
        category = unicodedata.category(character)
        return (
            character in {"_", "\\"}
            or character.isalpha()
            or category == "Nl"
        )

    def is_continue(character: str) -> bool:
        return (
            is_start(character)
            or character.isdigit()
            or character == "."
            or unicodedata.category(character) in {"Mn", "Mc"}
        )

    return is_start(value[0]) and all(is_continue(character) for character in value[1:])


def _mask_double_quoted_strings(value: str) -> str:
    """Blank formula strings while preserving offsets for token-safe matching."""
    masked = list(value)
    in_string = False
    position = 0
    while position < len(value):
        character = value[position]
        if not in_string:
            if character == '"':
                masked[position] = " "
                in_string = True
            position += 1
            continue

        masked[position] = " "
        if character == '"':
            if position + 1 < len(value) and value[position + 1] == '"':
                masked[position + 1] = " "
                position += 2
                continue
            in_string = False
        position += 1
    return "".join(masked)


def _rewrite_literal_spill_references(
    formula: str, *, preserve_operator: bool = False
) -> tuple[str, tuple[str, ...]]:
    """Remove only static ``A1#`` operators so openpyxl can inspect the anchor.

    The graph edge is deliberately to the formula's anchor cell, not to an
    invented fixed spill extent. The caller retains the original spill token in
    ``FormulaInspection`` so profiles and policy can make the remaining dynamic
    shape and blocker behavior visible.
    """
    masked_formula = _mask_double_quoted_strings(formula)
    parts: list[str] = []
    literal_spill_tokens: list[str] = []
    cursor = 0
    for match in _LITERAL_SPILL_REFERENCE.finditer(masked_formula):
        reference = formula[match.start("reference") : match.end("reference")]
        if parse_reference_token(reference) is None:
            continue
        parts.append(formula[cursor : match.start()])
        if preserve_operator:
            # Keep a spill reference distinct from an ordinary direct cell in
            # formula fingerprints while still giving the underlying tokenizer
            # a grammar it understands. This is equivalent to the OOXML-style
            # ANCHORARRAY representation documented by XlsxWriter.
            parts.append(f"{_SPILL_REFERENCE_FUNCTION}({reference})")
        else:
            parts.append(reference)
        cursor = match.end()
        literal_spill_tokens.append(formula[match.start() : match.end()])
    if not literal_spill_tokens:
        return formula, ()
    parts.append(formula[cursor:])
    return "".join(parts), tuple(dict.fromkeys(literal_spill_tokens))


def _rewrite_literal_implicit_intersection_references(
    formula: str,
) -> tuple[str, tuple[str, ...]]:
    """Make direct ``@A1`` references tokenizable without guessing their scope.

    Excel normally serializes persisted implicit intersection as ``SINGLE()``,
    but the display syntax can still appear in authoring tools. The private
    wrapper lets the rest of the inspection retain the operator as a distinct
    semantic feature and, when an origin is available, select a direct static
    range precisely.
    """
    masked_formula = _mask_double_quoted_strings(formula)
    parts: list[str] = []
    literal_tokens: list[str] = []
    cursor = 0
    for match in _LITERAL_IMPLICIT_INTERSECTION_REFERENCE.finditer(masked_formula):
        reference = formula[match.start("reference") : match.end("reference")]
        if parse_reference_token(reference) is None:
            continue
        parts.append(formula[cursor : match.start()])
        parts.append(f"{_LITERAL_IMPLICIT_INTERSECTION_FUNCTION}({reference})")
        cursor = match.end()
        literal_tokens.append(formula[match.start() : match.end()])
    if not literal_tokens:
        return formula, ()
    parts.append(formula[cursor:])
    return "".join(parts), tuple(dict.fromkeys(literal_tokens))


def _tokenize_formula(
    formula: str, *, preserve_literal_spill_operator: bool = False
) -> tuple[tuple[object, ...] | None, tuple[str, ...], tuple[str, ...]]:
    """Tokenize after narrow spill and implicit-intersection compatibility passes."""
    tokenizer_formula, literal_implicit_intersection_tokens = (
        _rewrite_literal_implicit_intersection_references(formula)
    )
    tokenizer_formula, literal_spill_tokens = _rewrite_literal_spill_references(
        tokenizer_formula, preserve_operator=preserve_literal_spill_operator
    )
    try:
        return (
            tuple(Tokenizer(tokenizer_formula).items),
            literal_spill_tokens,
            literal_implicit_intersection_tokens,
        )
    except Exception:
        return None, literal_spill_tokens, literal_implicit_intersection_tokens


def resolve_3d_reference(
    value: str, sheet_order: Sequence[str] | None
) -> tuple[ParsedReference, ...] | None:
    """Expand a static Excel 3-D A1 reference across its tab-order endpoints.

    Excel treats ``Sales:Marketing!B3`` as the same cell on every worksheet
    from ``Sales`` through ``Marketing`` in workbook tab order. There is no
    safe single-sheet approximation, so unknown, external, malformed, or
    endpoint-missing references deliberately return ``None`` for coverage
    reporting instead of creating a fictitious sheet dependency.
    """
    if not sheet_order:
        return None
    sheet, address = _split_sheet_reference(value)
    normalised_sheet, external = _normalise_sheet_name(sheet)
    if external or normalised_sheet is None:
        return None
    first_sheet, separator, last_sheet = normalised_sheet.partition(":")
    if not separator or not first_sheet or not last_sheet or ":" in last_sheet:
        return None

    sheet_positions = {title.casefold(): position for position, title in enumerate(sheet_order)}
    first_position = sheet_positions.get(first_sheet.casefold())
    last_position = sheet_positions.get(last_sheet.casefold())
    if first_position is None or last_position is None or first_position > last_position:
        return None
    try:
        min_column, min_row, max_column, max_row = range_boundaries(address)
    except ValueError:
        return None

    min_column = min_column or 1
    max_column = max_column or MAX_EXCEL_COLUMN
    min_row = min_row or 1
    max_row = max_row or MAX_EXCEL_ROW
    return tuple(
        ParsedReference(
            sheet=sheet_order[position],
            min_column=min_column,
            min_row=min_row,
            max_column=max_column,
            max_row=max_row,
            raw=value,
        )
        for position in range(first_position, last_position + 1)
    )


def reference_lookup_key(value: str) -> str:
    """Normalize a name-like formula token for a case-insensitive lookup.

    This also canonicalizes a sheet-qualified local name such as
    `'Debt Schedule'!TaxRate` so it can be matched against workbook metadata.
    External tokens are deliberately returned unchanged apart from case because
    they are handled as explicit external references before name resolution.
    """
    token = value.strip()
    sheet, address = _split_sheet_reference(token)
    if sheet is None:
        return token.casefold()
    normalized_sheet, external = _normalise_sheet_name(sheet)
    if normalized_sheet is None or external:
        return token.casefold()
    return f"{normalized_sheet.casefold()}!{address.strip().casefold()}"


def _structured_reference_parts(
    value: str,
) -> tuple[str, tuple[str, ...], tuple[str, ...]] | None:
    """Parse a conservative subset of the bracket syntax after a table name."""
    token = value.strip()
    opening = token.find("[")
    if opening < 0:
        return (token, (), ()) if token else None
    if opening == 0:
        return None  # Unqualified table references need formula-table context.
    table_name = token[:opening].strip()
    if not table_name:
        return None
    selector_parts = _structured_selector_parts(token[opening:])
    if selector_parts is None:
        return None
    groups, separators = selector_parts
    return table_name, groups, separators


def _structured_selector_parts(
    value: str,
) -> tuple[tuple[str, ...], tuple[str, ...]] | None:
    """Parse one bracketed structured-reference selector conservatively.

    The nested-group reader is intentionally small, but it is needed for the
    normal ``[@[Column Name]]`` spelling: its inner column-name brackets are
    not a second selector. Bracket-escaped column names remain outside this
    subset and are left as explicit coverage gaps.
    """
    selector = value.strip()
    if not (selector.startswith("[") and selector.endswith("]")):
        return None
    inner = selector[1:-1].strip()
    if not inner:
        return None
    if not inner.startswith("["):
        return (inner,), ()

    groups: list[str] = []
    separators: list[str] = []
    position = 0
    while position < len(inner):
        while position < len(inner) and inner[position].isspace():
            position += 1
        if position >= len(inner) or inner[position] != "[":
            return None
        closing = _matching_closing_bracket(inner, position)
        if closing < 0:
            return None
        group = inner[position + 1 : closing].strip()
        if not group:
            return None
        groups.append(group)
        position = closing + 1
        while position < len(inner) and inner[position].isspace():
            position += 1
        if position == len(inner):
            break
        if inner[position] not in {",", ":"}:
            return None
        separators.append(inner[position])
        position += 1
    return tuple(groups), tuple(separators)


def _matching_closing_bracket(value: str, opening: int) -> int:
    """Return the matching close bracket for ordinary nested selector groups."""
    depth = 0
    for position in range(opening, len(value)):
        character = value[position]
        if character == "[":
            depth += 1
        elif character == "]":
            depth -= 1
            if depth == 0:
                return position
            if depth < 0:
                return -1
    return -1


def _unqualified_structured_reference_parts(
    value: str,
) -> tuple[tuple[str, ...], tuple[str, ...]] | None:
    """Parse a selector whose table is supplied by a formula-cell context."""
    return _structured_selector_parts(value)


def _unescape_structured_column_name(value: str) -> str:
    """Undo Excel's bracket/header escape prefix for the supported subset."""
    result = value.strip()
    for character in "[#'@":
        result = result.replace(f"'{character}", character)
    return result


def _structured_table_regions(
    table: StructuredTable, items: set[str]
) -> tuple[tuple[int, int], ...] | None:
    try:
        _, min_row, _, max_row = range_boundaries(table.ref)
    except ValueError:  # pragma: no cover - invalid table refs come from malformed OOXML
        return None
    height = max_row - min_row + 1
    header_rows = min(max(table.header_row_count, 0), height)
    totals_rows = min(max(table.totals_row_count, 0), height - header_rows)
    data_start = min_row + header_rows
    data_end = max_row - totals_rows

    if "#all" in items:
        if len(items) != 1:
            return None
        return ((min_row, max_row),)
    selected = items or {"#data"}
    regions: list[tuple[int, int]] = []
    if "#headers" in selected and header_rows:
        regions.append((min_row, data_start - 1))
    if "#data" in selected and data_start <= data_end:
        regions.append((data_start, data_end))
    if "#totals" in selected and totals_rows:
        regions.append((data_end + 1, max_row))
    return tuple(regions)


def _table_data_row_for_origin(
    table: StructuredTable, origin: tuple[str, str] | None
) -> int | None:
    """Return an origin row only when it lies in this table's data body."""
    if origin is None or table.sheet.casefold() != origin[0].casefold():
        return None
    try:
        row, _ = coordinate_to_tuple(origin[1])
        _, min_row, _, max_row = range_boundaries(table.ref)
    except ValueError:  # pragma: no cover - malformed OOXML cannot supply a usable context
        return None
    height = max_row - min_row + 1
    header_rows = min(max(table.header_row_count, 0), height)
    totals_rows = min(max(table.totals_row_count, 0), height - header_rows)
    data_start = min_row + header_rows
    data_end = max_row - totals_rows
    return row if data_start <= row <= data_end else None


def _origin_table(
    tables: Mapping[str, StructuredTable], origin: tuple[str, str] | None
) -> StructuredTable | None:
    """Find the sole table whose data cells contain the formula origin."""
    if origin is None:
        return None
    try:
        _, origin_column = coordinate_to_tuple(origin[1])
    except ValueError:  # pragma: no cover - generated cell coordinates are valid
        return None
    candidates: list[StructuredTable] = []
    seen: set[StructuredTable] = set()
    for table in tables.values():
        if table in seen:
            continue
        seen.add(table)
        if _table_data_row_for_origin(table, origin) is None:
            continue
        try:
            min_column, _, max_column, _ = range_boundaries(table.ref)
        except ValueError:  # pragma: no cover - malformed OOXML cannot supply a usable context
            continue
        if min_column <= origin_column <= max_column:
            candidates.append(table)
    return candidates[0] if len(candidates) == 1 else None


def _current_row_column_name(value: str) -> str | None:
    """Read the column part of one ``@Column`` selector group."""
    column = value.strip()[1:].strip()
    if column.startswith("[") and column.endswith("]"):
        column = column[1:-1].strip()
    return column or None


def _current_row_column_spans(
    table: StructuredTable,
    groups: tuple[str, ...],
    separators: tuple[str, ...],
    *,
    implicit: bool,
) -> tuple[tuple[int, int], ...] | None:
    """Resolve the narrow current-row column subset without evaluating Excel.

    Explicit forms are ``Table[@Column]`` and
    ``Table[[#This Row],[Column]]``. Inside a table data row, Excel also uses
    an unqualified ``[Column]`` as a current-row reference; that form is
    admitted only when ``implicit`` is true and the origin determines one table.
    """
    if not groups or len(separators) != len(groups) - 1:
        return None

    column_lookup = {
        _unescape_structured_column_name(column).casefold(): index
        for index, column in enumerate(table.columns)
    }
    column_positions: list[tuple[int, int]] = []
    current_markers: list[int] = []
    for position, group in enumerate(groups):
        raw_group = group.strip()
        normalized = raw_group.casefold()
        if normalized in {"#this row", "@"}:
            current_markers.append(position)
            continue
        if raw_group.startswith("@"):
            column_name = _current_row_column_name(raw_group)
            if column_name is None:
                return None
            current_markers.append(position)
        else:
            if normalized.startswith("#"):
                return None
            column_name = raw_group
        column_index = column_lookup.get(
            _unescape_structured_column_name(column_name).casefold()
        )
        if column_index is None:
            return None
        column_positions.append((position, column_index))

    if implicit:
        if current_markers or not column_positions:
            return None
    else:
        if len(current_markers) != 1:
            return None
        marker = current_markers[0]
        if marker > 0:
            return None
        if marker < len(separators) and separators[marker] != ",":
            return None
        if marker in {position for position, _ in column_positions} and len(groups) != 1:
            return None

    if not column_positions:
        # ``Table[#This Row]`` / ``Table[@]`` is the whole table row.
        return ((0, len(table.columns) - 1),) if table.columns else None

    column_separators = [
        separator
        for position, separator in enumerate(separators)
        if position not in current_markers
    ]
    if any(separator not in {",", ":"} for separator in column_separators):
        return None
    if ":" in column_separators:
        if len(column_positions) != 2 or column_separators != [":"]:
            return None
        indexes = [index for _, index in column_positions]
        return ((min(indexes), max(indexes)),)
    return tuple((index, index) for _, index in column_positions)


def _current_row_references(
    table: StructuredTable,
    row: int,
    column_spans: tuple[tuple[int, int], ...],
    raw: str,
) -> tuple[ParsedReference, ...] | None:
    try:
        min_column, _, _, _ = range_boundaries(table.ref)
    except ValueError:  # pragma: no cover - invalid table refs come from malformed OOXML
        return None
    return tuple(
        ParsedReference(
            sheet=table.sheet,
            min_column=min_column + start_column,
            min_row=row,
            max_column=min_column + end_column,
            max_row=row,
            raw=raw,
        )
        for start_column, end_column in column_spans
    )


def _resolve_current_row_structured_reference(
    value: str,
    tables: Mapping[str, StructuredTable],
    origin: tuple[str, str] | None,
) -> tuple[ParsedReference, ...] | None:
    """Resolve context-bound row selectors only for a table data-row formula."""
    parsed = _structured_reference_parts(value)
    if parsed is not None:
        table_name, groups, separators = parsed
        table = tables.get(table_name.casefold())
        if table is None:
            return None
        row = _table_data_row_for_origin(table, origin)
        if row is None:
            return None
        spans = _current_row_column_spans(table, groups, separators, implicit=False)
        if spans is None:
            return None
        return _current_row_references(table, row, spans, value)

    selector_parts = _unqualified_structured_reference_parts(value)
    table = _origin_table(tables, origin)
    if selector_parts is None or table is None:
        return None
    row = _table_data_row_for_origin(table, origin)
    if row is None:  # defensive: _origin_table establishes this already
        return None
    groups, separators = selector_parts
    explicit_current_row = any(
        group.strip().casefold() in {"#this row", "@"}
        or group.strip().startswith("@")
        for group in groups
    )
    spans = _current_row_column_spans(
        table, groups, separators, implicit=not explicit_current_row
    )
    if spans is None:
        return None
    return _current_row_references(table, row, spans, value)


def resolve_structured_reference(
    value: str,
    tables: Mapping[str, StructuredTable],
    origin: tuple[str, str] | None = None,
) -> tuple[ParsedReference, ...] | None:
    """Resolve static, fully qualified Excel-table references without evaluation.

    Supported forms include a table name, a single column, contiguous column
    ranges, and the ``#All``, ``#Data``, ``#Headers``, and ``#Totals`` item
    specifiers. When the formula-cell origin is in a table data row, it also
    resolves conservative ``@``/``#This Row`` forms and unqualified table
    columns. Complex bracket escaping remains an explicit coverage gap.
    """
    current_row_reference = _resolve_current_row_structured_reference(value, tables, origin)
    if current_row_reference is not None:
        return current_row_reference
    parsed = _structured_reference_parts(value)
    if parsed is None:
        return None
    table_name, groups, separators = parsed
    table = tables.get(table_name.casefold())
    if table is None:
        return None
    item_tokens = {"#all", "#data", "#headers", "#totals", "#this row"}
    items: set[str] = set()
    columns: list[int] = []
    column_lookup = {
        _unescape_structured_column_name(column).casefold(): index
        for index, column in enumerate(table.columns)
    }
    for group in groups:
        normalized = group.strip().casefold()
        if normalized in item_tokens:
            if normalized == "#this row":
                return None
            items.add(normalized)
            continue
        if normalized.startswith("#") or normalized.startswith("@"):
            return None
        column_index = column_lookup.get(_unescape_structured_column_name(group).casefold())
        if column_index is None:
            return None
        columns.append(column_index)

    try:
        min_column, _, max_column, _ = range_boundaries(table.ref)
    except ValueError:  # pragma: no cover - invalid table refs come from malformed OOXML
        return None
    table_width = max_column - min_column + 1
    if len(columns) > 2 and ":" in separators:
        return None
    if ":" in separators:
        if len(columns) != 2:
            return None
        column_spans = [(min(columns), max(columns))]
    elif columns:
        column_spans = [(column, column) for column in columns]
    else:
        column_spans = [(0, table_width - 1)]
    regions = _structured_table_regions(table, items)
    if regions is None:
        return None
    references = [
        ParsedReference(
            sheet=table.sheet,
            min_column=min_column + start_column,
            min_row=min_row,
            max_column=min_column + end_column,
            max_row=max_row,
            raw=value,
        )
        for start_column, end_column in column_spans
        for min_row, max_row in regions
    ]
    return tuple(dict.fromkeys(references))


def _is_group_open(token: object) -> bool:
    """Return whether an openpyxl token starts a nested expression group."""
    return (
        getattr(token, "type", None) in _GROUP_TOKEN_TYPES
        and getattr(token, "subtype", None) == "OPEN"
    )


def _is_group_close(token: object) -> bool:
    """Return whether an openpyxl token ends a nested expression group."""
    return (
        getattr(token, "type", None) in _GROUP_TOKEN_TYPES
        and getattr(token, "subtype", None) == "CLOSE"
    )


def _is_whitespace(token: object) -> bool:
    """Handle the whitespace labels emitted by supported openpyxl versions."""
    return getattr(token, "type", None) in _WHITESPACE_TOKEN_TYPES


def _matching_group_close(tokens: Sequence[object], opening: int, end: int) -> int | None:
    """Find a matching formula-token group close without parsing Excel values."""
    depth = 0
    for position in range(opening, end):
        if _is_group_open(tokens[position]):
            depth += 1
        elif _is_group_close(tokens[position]):
            depth -= 1
            if depth == 0:
                return position
    return None


def _function_argument_spans(
    tokens: Sequence[object], start: int, end: int
) -> tuple[tuple[int, int], ...]:
    """Split one function body into top-level comma-separated token spans."""
    spans: list[tuple[int, int]] = []
    argument_start = start
    depth = 0
    for position in range(start, end):
        token = tokens[position]
        if _is_group_open(token):
            depth += 1
        elif _is_group_close(token):
            depth -= 1
        elif (
            depth == 0
            and getattr(token, "type", None) == "SEP"
            and getattr(token, "subtype", None) == "ARG"
        ):
            spans.append((argument_start, position))
            argument_start = position + 1
    spans.append((argument_start, end))
    return tuple(spans)


def _local_scope_key(value: str) -> str:
    """Normalize a local name while hiding OOXML parameter-marker prefixes."""
    token = value.strip()
    normalized_token = token.casefold()
    for prefix in _SERIALIZED_LOCAL_PREFIXES:
        if normalized_token.startswith(prefix):
            identifier = token[len(prefix) :]
            if _LOCAL_IDENTIFIER.fullmatch(identifier):
                return identifier.casefold()
    return reference_lookup_key(token)


def _simple_local_identifier(
    tokens: Sequence[object],
    start: int,
    end: int,
    *,
    allow_serialized_local_prefix: bool = False,
) -> tuple[int, str] | None:
    """Return one conservative LET/LAMBDA identifier declaration token.

    Excel-compatible writers may serialize LAMBDA parameters or local values
    with ``_xlpm.`` or ``_xlop.`` prefixes. They are accepted only at a
    LET/LAMBDA declaration so an ordinary dotted defined name cannot be
    mistaken for a local variable.
    """
    meaningful = [
        position
        for position in range(start, end)
        if not _is_whitespace(tokens[position])
    ]
    if len(meaningful) != 1:
        return None
    position = meaningful[0]
    token = tokens[position]
    if not (
        getattr(token, "type", None) == "OPERAND"
        and getattr(token, "subtype", None) == "RANGE"
    ):
        return None
    value = str(getattr(token, "value", "")).strip()
    identifier = value
    if allow_serialized_local_prefix:
        normalized_value = value.casefold()
        for prefix in _SERIALIZED_LOCAL_PREFIXES:
            if normalized_value.startswith(prefix):
                identifier = value[len(prefix) :]
                break
    if not _LOCAL_IDENTIFIER.fullmatch(identifier):
        return None
    if (
        identifier.casefold() in {"r", "c"}
        or _A1_LOCAL_IDENTIFIER_CONFLICT.fullmatch(identifier)
        or _R1C1_LOCAL_IDENTIFIER_CONFLICT.fullmatch(identifier)
    ):
        return None
    return position, _local_scope_key(value)


def _function_name(token: object) -> str:
    """Normalize function names, including Excel's OOXML namespace prefixes."""
    value = str(getattr(token, "value", "")).rstrip("(").strip().upper()
    return value.rsplit(".", 1)[-1].lstrip("@")


def _unqualified_native_function_name(token: object) -> str | None:
    """Return an exact native function spelling without folding namespaces.

    ``_function_name`` intentionally folds namespace prefixes for broad
    inspection. Static function contracts must not do that: a custom
    ``Vendor.SUMPRODUCT`` call can have unrelated argument semantics. This
    helper recognizes only an unqualified spelling, optionally preceded by
    Excel's display-only ``@`` operator.
    """
    if not (
        getattr(token, "type", None) == "FUNC"
        and getattr(token, "subtype", None) == "OPEN"
    ):
        return None
    return str(getattr(token, "value", "")).rstrip("(").strip().upper().removeprefix("@")


def _native_conditional_aggregate_function_name(token: object) -> str | None:
    """Return one exact native conditional-aggregate spelling, if present.

    ``_function_name`` intentionally folds namespace prefixes for broad
    inspection. That would be unsafe here: a custom ``Vendor.SUMIFS`` call can
    have unrelated range semantics. This lint trusts Excel's unqualified native
    spelling, optionally preceded by the display-only ``@`` operator, plus the
    exact OOXML future-function serializations for ``MAXIFS`` and ``MINIFS``.
    """
    function_name = _unqualified_native_function_name(token)
    if function_name is None:
        return None
    if function_name in _CONDITIONAL_AGGREGATE_RANGE_SHAPE_FUNCTIONS:
        return function_name
    return _OOXML_FUTURE_CONDITIONAL_AGGREGATE_FUNCTION_NAMES.get(function_name)


def _direct_static_a1_range_shape(
    tokens: Sequence[object], start: int, end: int
) -> tuple[int, int] | None:
    """Return one bounded direct A1 range's ``(width, height)``, if exact."""
    meaningful = [
        position
        for position in range(start, end)
        if not _is_whitespace(tokens[position])
    ]
    if len(meaningful) != 1:
        return None
    token = tokens[meaningful[0]]
    if not (
        getattr(token, "type", None) == "OPERAND"
        and getattr(token, "subtype", None) == "RANGE"
    ):
        return None
    value = str(getattr(token, "value", "")).strip()
    if _DIRECT_STATIC_A1_RANGE_ARGUMENT.fullmatch(value) is None:
        return None
    reference = parse_reference_token(value)
    if (
        reference is None
        or reference.is_external
        or reference.min_column is None
        or reference.min_row is None
        or reference.max_column is None
        or reference.max_row is None
        or not (
            1
            <= reference.min_column
            <= reference.max_column
            <= MAX_EXCEL_COLUMN
            and 1
            <= reference.min_row
            <= reference.max_row
            <= MAX_EXCEL_ROW
        )
    ):
        return None
    return (
        reference.max_column - reference.min_column + 1,
        reference.max_row - reference.min_row + 1,
    )


def _direct_positive_integer_literal(
    tokens: Sequence[object], start: int, end: int
) -> str | None:
    """Return one direct positive integer literal without evaluating it.

    The normalized decimal string deliberately avoids converting an
    arbitrarily long formula token to ``int``. Callers can compare it against
    their bounded structural limit lexically. Unary signs, decimal notation,
    arithmetic, names, and references remain outside this narrow static
    contract.
    """
    meaningful = [
        position
        for position in range(start, end)
        if not _is_whitespace(tokens[position])
    ]
    if len(meaningful) != 1:
        return None
    token = tokens[meaningful[0]]
    if not (
        getattr(token, "type", None) == "OPERAND"
        and getattr(token, "subtype", None) == "NUMBER"
    ):
        return None
    value = str(getattr(token, "value", "")).strip()
    if not value.isascii() or not value.isdecimal():
        return None
    normalized = value.lstrip("0")
    return normalized or None


def _positive_integer_literal_exceeds(value: str, limit: int) -> bool:
    """Return whether normalized decimal ``value`` is greater than ``limit``."""
    limit_text = str(limit)
    return len(value) > len(limit_text) or (
        len(value) == len(limit_text) and value > limit_text
    )


def conditional_aggregate_range_shape_mismatches(
    formula: str,
) -> tuple[tuple[str, int], ...]:
    """Find direct static range-shape mismatches in conditional aggregates.

    Each returned pair contains a native function name and the number of its
    range arguments whose dimensions differ from that call's first range.
    This scans formula tokens only; it does not resolve names, inspect table
    identities, calculate a formula, or infer values. A call is deliberately
    ignored unless every relevant range argument is one direct, bounded,
    internal A1 cell/range or whole-column reference. It covers `SUMIFS`,
    `COUNTIFS`, `AVERAGEIFS`, `MAXIFS`, and `MINIFS`, including the exact
    OOXML ``_xlfn.MAXIFS`` and ``_xlfn.MINIFS`` serializations. All dynamic and
    otherwise ambiguous formulas remain outside the finding boundary.
    """
    tokens, _, _ = _tokenize_formula(
        formula,
        preserve_literal_spill_operator=True,
    )
    if tokens is None:
        return ()
    if any(
        getattr(token, "type", None) == "OPERAND"
        and getattr(token, "subtype", None) == "ERROR"
        and str(getattr(token, "value", "")).strip().upper() == "#REF!"
        for token in tokens
    ):
        return ()

    mismatches: list[tuple[str, int]] = []
    for position, token in enumerate(tokens):
        function_name = _native_conditional_aggregate_function_name(token)
        if function_name is None:
            continue
        closing = _matching_group_close(tokens, position, len(tokens))
        if closing is None:
            continue
        arguments = _function_argument_spans(tokens, position + 1, closing)
        if function_name in _CONDITIONAL_AGGREGATE_VALUE_RANGE_FUNCTIONS:
            if len(arguments) < 3 or len(arguments) % 2 == 0:
                continue
            range_argument_indexes = (0, *range(1, len(arguments), 2))
        else:
            if len(arguments) < 2 or len(arguments) % 2 != 0:
                continue
            range_argument_indexes = range(0, len(arguments), 2)

        shapes: list[tuple[int, int]] = []
        for argument_index in range_argument_indexes:
            start, end = arguments[argument_index]
            shape = _direct_static_a1_range_shape(tokens, start, end)
            if shape is None:
                break
            shapes.append(shape)
        else:
            mismatched_range_argument_count = sum(
                shape != shapes[0]
                for shape in shapes[1:]
            )
            if mismatched_range_argument_count:
                mismatches.append((function_name, mismatched_range_argument_count))
    return tuple(mismatches)


def sumproduct_range_shape_mismatches(formula: str) -> tuple[int, ...]:
    """Find direct static range-shape mismatches in native ``SUMPRODUCT`` calls.

    Each returned integer is the number of direct array arguments whose
    dimensions differ from that call's first array argument. This scans formula
    tokens only; it does not resolve names, inspect table identities, calculate
    a formula, or infer values. A call is deliberately ignored unless it has at
    least two arguments and every argument is one direct, bounded, internal A1
    cell/range or whole-column reference. It accepts only Excel's unqualified
    native ``SUMPRODUCT`` spelling, optionally preceded by ``@``. All dynamic
    and otherwise ambiguous formulas remain outside the finding boundary.
    """
    tokens, _, _ = _tokenize_formula(
        formula,
        preserve_literal_spill_operator=True,
    )
    if tokens is None:
        return ()
    if any(
        getattr(token, "type", None) == "OPERAND"
        and getattr(token, "subtype", None) == "ERROR"
        and str(getattr(token, "value", "")).strip().upper() == "#REF!"
        for token in tokens
    ):
        return ()

    mismatches: list[int] = []
    for position, token in enumerate(tokens):
        if _unqualified_native_function_name(token) != "SUMPRODUCT":
            continue
        closing = _matching_group_close(tokens, position, len(tokens))
        if closing is None:
            continue
        arguments = _function_argument_spans(tokens, position + 1, closing)
        if len(arguments) < 2:
            continue

        shapes: list[tuple[int, int]] = []
        for start, end in arguments:
            shape = _direct_static_a1_range_shape(tokens, start, end)
            if shape is None:
                break
            shapes.append(shape)
        else:
            mismatched_argument_count = sum(
                shape != shapes[0]
                for shape in shapes[1:]
            )
            if mismatched_argument_count:
                mismatches.append(mismatched_argument_count)
    return tuple(mismatches)


def mmult_dimension_mismatch_count(formula: str) -> int:
    """Count provable direct-static matrix dimension mismatches in ``MMULT``.

    This scans formula tokens only; it does not resolve names, inspect table
    identities, calculate a formula, or infer cell values. A call is
    deliberately ignored unless it has exactly two arguments and each is one
    direct, bounded, internal A1 cell/range or whole-column reference. It
    accepts only Excel's unqualified native ``MMULT`` spelling, optionally
    preceded by ``@``. A finding is returned only when the first argument's
    column count differs from the second argument's row count. All dynamic and
    otherwise ambiguous formulas remain outside the finding boundary.
    """
    tokens, _, _ = _tokenize_formula(
        formula,
        preserve_literal_spill_operator=True,
    )
    if tokens is None:
        return 0
    if any(
        getattr(token, "type", None) == "OPERAND"
        and getattr(token, "subtype", None) == "ERROR"
        and str(getattr(token, "value", "")).strip().upper() == "#REF!"
        for token in tokens
    ):
        return 0

    mismatch_count = 0
    for position, token in enumerate(tokens):
        if _unqualified_native_function_name(token) != "MMULT":
            continue
        closing = _matching_group_close(tokens, position, len(tokens))
        if closing is None:
            continue
        arguments = _function_argument_spans(tokens, position + 1, closing)
        if len(arguments) != 2:
            continue
        first_shape = _direct_static_a1_range_shape(tokens, *arguments[0])
        second_shape = _direct_static_a1_range_shape(tokens, *arguments[1])
        if first_shape is None or second_shape is None:
            continue
        first_column_count, _ = first_shape
        _, second_row_count = second_shape
        if first_column_count != second_row_count:
            mismatch_count += 1
    return mismatch_count


def lookup_return_index_mismatches(formula: str) -> tuple[str, ...]:
    """Find provable direct-static return-index errors in legacy lookups.

    This scans formula tokens only; it does not resolve names, inspect Table
    identities, calculate a formula, or infer values. A call is deliberately
    ignored unless it is an unqualified native ``VLOOKUP`` or ``HLOOKUP``
    (optionally preceded by ``@``), has exactly three or four arguments, uses
    one direct bounded internal A1 cell/range or whole-column table argument,
    and supplies a direct positive integer literal as its return index. A
    finding is returned only when a ``VLOOKUP`` index exceeds that table's
    width or an ``HLOOKUP`` index exceeds its height. All dynamic and otherwise
    ambiguous formulas remain outside the finding boundary.
    """
    tokens, _, _ = _tokenize_formula(
        formula,
        preserve_literal_spill_operator=True,
    )
    if tokens is None:
        return ()
    if any(
        getattr(token, "type", None) == "OPERAND"
        and getattr(token, "subtype", None) == "ERROR"
        and str(getattr(token, "value", "")).strip().upper() == "#REF!"
        for token in tokens
    ):
        return ()

    mismatches: list[str] = []
    for position, token in enumerate(tokens):
        function_name = _unqualified_native_function_name(token)
        if function_name not in {"VLOOKUP", "HLOOKUP"}:
            continue
        closing = _matching_group_close(tokens, position, len(tokens))
        if closing is None:
            continue
        arguments = _function_argument_spans(tokens, position + 1, closing)
        if len(arguments) not in {3, 4}:
            continue
        table_shape = _direct_static_a1_range_shape(tokens, *arguments[1])
        index_literal = _direct_positive_integer_literal(tokens, *arguments[2])
        if table_shape is None or index_literal is None:
            continue
        table_limit = table_shape[0] if function_name == "VLOOKUP" else table_shape[1]
        if _positive_integer_literal_exceeds(index_literal, table_limit):
            mismatches.append(function_name)
    return tuple(mismatches)


def _function_lookup_key(token: object) -> str:
    """Normalize a callable defined-name token without losing sheet qualification."""
    return reference_lookup_key(str(getattr(token, "value", "")).rstrip("(").strip())


def _office_custom_function_candidate(token: object) -> str | None:
    """Return a normalized namespaced custom-function candidate, if safe.

    Office Add-in custom functions are displayed as a namespace followed by a
    function name, but their manifest and JavaScript runtime are not embedded
    in a normal workbook.  This intentionally recognizes only the documented
    namespaced spelling, excludes OOXML compatibility prefixes and native
    dotted functions, and leaves every other UDF shape out of this boundary.
    """
    raw_name = str(getattr(token, "value", "")).rstrip("(").strip()
    candidate = raw_name.removeprefix("@")
    upper_candidate = candidate.upper()
    if (
        not candidate
        or upper_candidate.startswith(("_XLFN.", "_XLWS."))
        or candidate.startswith("_")
        or "." not in candidate
        or any(character in candidate for character in "![]'")
        or any(character.isspace() for character in candidate)
        or upper_candidate in _EXCEL_DOTTED_FUNCTIONS
    ):
        return None
    namespace, function_name = candidate.split(".", 1)
    if (
        not namespace
        or not function_name
        or not namespace[0].isalpha()
        or not function_name[0].isalpha()
        or any(not segment for segment in candidate.split("."))
        or any(
            not (character.isalnum() or character in "_.")
            for character in candidate
        )
    ):
        return None
    return upper_candidate


def _unqualified_runtime_function_candidate(token: object) -> str | None:
    """Return a bare callable that may resolve through an Excel runtime.

    A stored formula cannot prove whether an unknown bare identifier resolves
    to a VBA UDF, COM/Automation add-in, XLL, or another registered runtime.
    This classifier is therefore deliberately lexical and conservative: a
    documented native function, a qualified/dotted spelling, and every
    non-identifier are outside the candidate boundary. Callers additionally
    suppress workbook-defined names and local LET/LAMBDA bindings before using
    this result.
    """
    raw_name = str(getattr(token, "value", "")).rstrip("(").strip()
    candidate = raw_name.removeprefix("@")
    upper_candidate = candidate.upper()
    if (
        not candidate
        or "." in candidate
        or not _LOCAL_IDENTIFIER.fullmatch(candidate)
        or upper_candidate in _EXCEL_UNQUALIFIED_NATIVE_FUNCTIONS
    ):
        return None
    return upper_candidate


def _worksheet_code_resource_registration_function(token: object) -> str | None:
    """Return a documented worksheet code-resource registration call, if present.

    ``_function_name`` deliberately drops dotted namespace prefixes for broad
    native-function handling, so this narrow classifier works from the raw
    callable spelling instead.  A defined name with the same spelling is
    resolved before this classifier runs; FormulaFence will not assert that a
    user-defined callable is Excel's legacy registration primitive.
    """
    raw_name = str(getattr(token, "value", "")).rstrip("(").strip()
    normalized = raw_name.removeprefix("@").upper()
    if normalized in _WORKSHEET_CODE_RESOURCE_REGISTRATION_FUNCTIONS:
        return normalized
    return None


def _formula_defined_xlm_registration_function(token: object) -> str | None:
    """Return a defined-name-only XLM ``REGISTER`` call, if present.

    This raw-spelling classifier intentionally does not run for ordinary
    worksheet formula inspection. A caller can enable it only while examining
    a stored formula-defined name or named LAMBDA, and known defined callables
    are resolved before it runs so a user-defined ``REGISTER`` name is never
    asserted to be Excel's XLM primitive.
    """
    raw_name = str(getattr(token, "value", "")).rstrip("(").strip()
    normalized = raw_name.removeprefix("@").upper()
    if normalized in _FORMULA_DEFINED_XLM_REGISTRATION_FUNCTIONS:
        return normalized
    return None


def _formula_defined_xlm_evaluation_function(token: object) -> str | None:
    """Return a defined-name-only XLM `EVALUATE` call, if present.

    The raw-spelling classifier is deliberately opt-in and only used while
    examining a stored formula-defined name or named LAMBDA. Known defined
    callables are resolved before it runs, so a workbook-defined `EVALUATE`
    name is not asserted to be Excel's legacy expression-evaluation primitive.
    """
    raw_name = str(getattr(token, "value", "")).rstrip("(").strip()
    normalized = raw_name.removeprefix("@").upper()
    if normalized in _FORMULA_DEFINED_XLM_EVALUATION_FUNCTIONS:
        return normalized
    return None


def _formula_defined_xlm_action_function(token: object) -> str | None:
    """Return a selected defined-name-only XLM action call, if present.

    The raw-spelling classifier is deliberately opt-in and only used while
    examining a stored formula-defined name or named LAMBDA. Known defined
    callables are resolved before it runs, so a workbook-defined action-shaped
    name is never asserted to be an Excel legacy macro primitive.
    """
    raw_name = str(getattr(token, "value", "")).rstrip("(").strip()
    normalized = raw_name.removeprefix("@").upper()
    if normalized in _FORMULA_DEFINED_XLM_ACTION_FUNCTIONS:
        return normalized
    return None


def _formula_defined_xlm_get_cell_function(token: object) -> str | None:
    """Return a defined-name-only XLM `GET.CELL` call, if present.

    The raw-spelling classifier is deliberately opt-in and only used while
    examining a stored formula-defined name or named LAMBDA. Known defined
    callables are resolved before it runs, so a workbook-defined `GET.CELL`
    name is not asserted to be Excel's legacy information primitive.
    """
    raw_name = str(getattr(token, "value", "")).rstrip("(").strip()
    normalized = raw_name.removeprefix("@").upper()
    if normalized in _FORMULA_DEFINED_XLM_GET_CELL_FUNCTIONS:
        return normalized
    return None


def _formula_defined_xlm_environment_information_function(
    token: object,
) -> str | None:
    """Return an opt-in stored-definition XLM environment-information call.

    Known defined callables are resolved before this raw-spelling classifier
    runs, so a workbook-defined name cannot be asserted to be one of Excel's
    legacy native information primitives.
    """
    raw_name = str(getattr(token, "value", "")).rstrip("(").strip()
    normalized = raw_name.removeprefix("@").upper()
    if normalized in _FORMULA_DEFINED_XLM_ENVIRONMENT_INFORMATION_FUNCTIONS:
        return normalized
    return None


def _formula_environment_information_function(
    tokens: Sequence[object], position: int
) -> tuple[str, bool, bool] | None:
    """Return a native information call and statically visible omitted arguments.

    This only reads token structure.  It intentionally does not inspect the
    information type, calculate a formula, resolve a dynamic argument, or
    infer the active cell. A malformed CELL or SHEETS call still remains an
    inventory item, but cannot safely be asserted to have the documented
    omitted-reference behavior.
    """
    token = tokens[position]
    raw_name = str(getattr(token, "value", "")).rstrip("(").strip()
    normalized = raw_name.removeprefix("@").upper()
    if normalized not in _FORMULA_ENVIRONMENT_INFORMATION_FUNCTIONS:
        return None
    if normalized not in {"CELL", "SHEETS"}:
        return normalized, False, False

    closing = _matching_group_close(tokens, position, len(tokens))
    if closing is None:
        return normalized, False, False
    arguments = _function_argument_spans(tokens, position + 1, closing)
    has_single_nonempty_argument = len(arguments) == 1 and any(
        not _is_whitespace(tokens[index])
        for index in range(arguments[0][0], arguments[0][1])
    )
    has_nonempty_argument = any(
        not _is_whitespace(tokens[index])
        for start, end in arguments
        for index in range(start, end)
    )
    if normalized == "CELL":
        return normalized, has_single_nonempty_argument, False
    # Only the documented no-argument form is a statically reliable workbook
    # count.  Keep malformed multi-argument calls in the private inventory, but
    # do not treat e.g. ``SHEETS(,)`` as an omitted reference.
    return normalized, False, len(arguments) == 1 and not has_nonempty_argument


def _fingerprint_token_value(token: object) -> str:
    """Normalize OOXML spellings of the two dynamic-array compatibility functions."""
    if (
        getattr(token, "type", None) == "FUNC"
        and getattr(token, "subtype", None) == "OPEN"
    ):
        function_name = _function_name(token)
        if function_name == _SPILL_REFERENCE_FUNCTION:
            return f"{_SPILL_REFERENCE_FUNCTION}("
        if function_name in {
            _IMPLICIT_INTERSECTION_FUNCTION,
            _LITERAL_IMPLICIT_INTERSECTION_FUNCTION,
        }:
            return f"{_IMPLICIT_INTERSECTION_FUNCTION}("
    return str(getattr(token, "value", ""))


def lambda_parameter_count(formula: str) -> int | None:
    """Return a valid top-level LAMBDA definition's parameter count, if present.

    A workbook-defined LAMBDA is callable only when its full definition is one
    ``LAMBDA(...)`` expression. This intentionally rejects formulas that merely
    contain a lambda, malformed parameter lists, and trailing invocations so a
    caller cannot mistake an arbitrary defined formula for a custom function.
    """
    tokens, _, _ = _tokenize_formula(formula)
    if tokens is None:
        return _raw_top_level_lambda_parameter_count(formula)
    meaningful = [
        position
        for position, token in enumerate(tokens)
        if not _is_whitespace(token)
    ]
    if not meaningful:
        return None
    opening = meaningful[0]
    token = tokens[opening]
    if not (
        getattr(token, "type", None) == "FUNC"
        and getattr(token, "subtype", None) == "OPEN"
        and _function_name(token) == "LAMBDA"
    ):
        return None
    closing = _matching_group_close(tokens, opening, len(tokens))
    if closing is None or any(position > closing for position in meaningful):
        return None
    arguments = _function_argument_spans(tokens, opening + 1, closing)
    if not arguments:
        return None
    calculation_start, calculation_end = arguments[-1]
    if not any(
        not _is_whitespace(tokens[position])
        for position in range(calculation_start, calculation_end)
    ):
        return None
    parameters: set[str] = set()
    for start, end in arguments[:-1]:
        declaration = _simple_local_identifier(
            tokens, start, end, allow_serialized_local_prefix=True
        )
        if declaration is None or declaration[1] in parameters:
            return None
        parameters.add(declaration[1])
    return len(parameters)


def _raw_top_level_lambda_parameter_count(formula: str) -> int | None:
    """Recognize a complete named LAMBDA when tokenization is unavailable.

    Direct DDE formula syntax is outside openpyxl's grammar, but it can appear
    in the body of a stored named LAMBDA.  This fallback deliberately recognizes
    only a full top-level ``LAMBDA(...)`` expression and validates each
    parameter conservatively; it does not attempt to parse or evaluate its
    calculation body.  Normal tokenization remains authoritative whenever it
    succeeds.
    """
    match = re.match(r"^\s*=\s*(?:_xlfn\.)?LAMBDA\s*\(", formula, re.IGNORECASE)
    if match is None:
        return None
    opening = match.end() - 1
    depth = 1
    argument_starts = [opening + 1]
    argument_ends: list[int] = []
    index = opening + 1
    while index < len(formula):
        character = formula[index]
        if character == '"':
            index = _skip_excel_double_quoted_literal(formula, index)
            continue
        if character == "'":
            quoted_end = _single_quoted_component_end(formula, index)
            if quoted_end is None:
                return None
            index = quoted_end
            continue
        if character == "(":
            depth += 1
        elif character == ")":
            depth -= 1
            if depth == 0:
                argument_ends.append(index)
                if formula[index + 1 :].strip():
                    return None
                break
        elif character in {",", ";"} and depth == 1:
            argument_ends.append(index)
            argument_starts.append(index + 1)
        index += 1
    else:
        return None

    if not argument_ends or len(argument_starts) != len(argument_ends):
        return None
    body = formula[argument_starts[-1] : argument_ends[-1]].strip()
    if not body:
        return None
    parameters: set[str] = set()
    for start, end in zip(argument_starts[:-1], argument_ends[:-1], strict=True):
        parameter = formula[start:end].strip()
        serialised_prefix = next(
            (
                prefix
                for prefix in _SERIALIZED_LOCAL_PREFIXES
                if parameter.casefold().startswith(prefix)
            ),
            None,
        )
        if serialised_prefix is not None:
            parameter = parameter[len(serialised_prefix) :]
        if (
            not _LOCAL_IDENTIFIER.fullmatch(parameter)
            or _A1_LOCAL_IDENTIFIER_CONFLICT.fullmatch(parameter)
            or _R1C1_LOCAL_IDENTIFIER_CONFLICT.fullmatch(parameter)
        ):
            return None
        key = parameter.casefold()
        if key in parameters:
            return None
        parameters.add(key)
    return len(parameters)


def _local_variable_token_indexes(tokens: Sequence[object]) -> set[int]:
    """Find LET/LAMBDA declarations and lexical uses masquerading as ranges.

    openpyxl's tokenizer correctly preserves the argument structure but labels
    Excel's local names as ``OPERAND/RANGE``. This walk is deliberately narrow:
    malformed declarations fall back to ordinary inspection, where their tokens
    remain visible as unresolved rather than being suppressed.
    """
    local_indexes: set[int] = set()

    def visit(start: int, end: int, scope: frozenset[str]) -> None:
        position = start
        while position < end:
            token = tokens[position]
            if _is_group_open(token):
                closing = _matching_group_close(tokens, position, end)
                if closing is None:
                    position += 1
                    continue
                body_start = position + 1
                if getattr(token, "type", None) == "FUNC":
                    arguments = _function_argument_spans(tokens, body_start, closing)
                    name = _function_name(token)
                    local_callable = _local_scope_key(
                        str(getattr(token, "value", "")).rstrip("(")
                    ) in scope
                    if local_callable:
                        local_indexes.add(position)
                    if name == "LET" and not local_callable:
                        visit_let(arguments, scope)
                    elif name == "LAMBDA" and not local_callable:
                        visit_lambda(arguments, scope)
                    else:
                        visit(body_start, closing, scope)
                else:
                    visit(body_start, closing, scope)
                position = closing + 1
                continue
            if (
                getattr(token, "type", None) == "OPERAND"
                and getattr(token, "subtype", None) == "RANGE"
                and _local_scope_key(str(getattr(token, "value", ""))) in scope
            ):
                local_indexes.add(position)
            position += 1

    def visit_let(arguments: Sequence[tuple[int, int]], scope: frozenset[str]) -> None:
        if len(arguments) < 3 or len(arguments) % 2 == 0:
            for start, end in arguments:
                visit(start, end, scope)
            return
        declarations: list[tuple[int, str]] = []
        for start, end in arguments[:-1:2]:
            declaration = _simple_local_identifier(
                tokens, start, end, allow_serialized_local_prefix=True
            )
            if declaration is None:
                for fallback_start, fallback_end in arguments:
                    visit(fallback_start, fallback_end, scope)
                return
            declarations.append(declaration)

        active_scope = set(scope)
        for pair_index, (declaration_index, identifier) in enumerate(declarations):
            local_indexes.add(declaration_index)
            value_start, value_end = arguments[pair_index * 2 + 1]
            visit(value_start, value_end, frozenset(active_scope))
            active_scope.add(identifier)
        final_start, final_end = arguments[-1]
        visit(final_start, final_end, frozenset(active_scope))

    def visit_lambda(arguments: Sequence[tuple[int, int]], scope: frozenset[str]) -> None:
        if len(arguments) < 2:
            for start, end in arguments:
                visit(start, end, scope)
            return
        declarations: list[tuple[int, str]] = []
        for start, end in arguments[:-1]:
            declaration = _simple_local_identifier(
                tokens, start, end, allow_serialized_local_prefix=True
            )
            if declaration is None:
                for fallback_start, fallback_end in arguments:
                    visit(fallback_start, fallback_end, scope)
                return
            declarations.append(declaration)
        local_scope = set(scope)
        for declaration_index, identifier in declarations:
            local_indexes.add(declaration_index)
            local_scope.add(identifier)
        body_start, body_end = arguments[-1]
        visit(body_start, body_end, frozenset(local_scope))

    try:
        visit(0, len(tokens), frozenset())
    except Exception:  # pragma: no cover - conservative fallback for unknown token shapes
        return set()
    return local_indexes


def _implicit_intersection_selection(
    reference: ParsedReference, origin: tuple[str, str]
) -> ParsedReference | None:
    """Return the one static cell selected by a direct implicit intersection.

    Microsoft documents implicit intersection as selecting the input on the
    formula's row or column. A one-dimensional range can therefore be resolved
    exactly. For a two-dimensional range, selection is only certain when the
    formula cell lies within it. All other forms deliberately fall back to the
    normal conservative range edge instead of inventing a selected cell.
    """
    if reference.is_external or None in {
        reference.min_column,
        reference.min_row,
        reference.max_column,
        reference.max_row,
    }:
        return None
    try:
        origin_row, origin_column = coordinate_to_tuple(origin[1])
    except ValueError:
        return None

    min_column = reference.min_column
    min_row = reference.min_row
    max_column = reference.max_column
    max_row = reference.max_row
    if min_column == max_column and min_row == max_row:
        selected_column, selected_row = min_column, min_row
    elif min_column == max_column and min_row <= origin_row <= max_row:
        selected_column, selected_row = min_column, origin_row
    elif min_row == max_row and min_column <= origin_column <= max_column:
        selected_column, selected_row = origin_column, min_row
    elif (
        min_column <= origin_column <= max_column
        and min_row <= origin_row <= max_row
    ):
        selected_column, selected_row = origin_column, origin_row
    else:
        return None
    return ParsedReference(
        reference.sheet,
        selected_column,
        selected_row,
        selected_column,
        selected_row,
        raw=reference.raw,
        is_external=reference.is_external,
    )


def _implicit_intersection_reference_replacements(
    tokens: Sequence[object], origin: tuple[str, str] | None
) -> dict[int, ParsedReference]:
    """Map direct SINGLE() operands to precise static dependency cells."""
    if origin is None:
        return {}
    replacements: dict[int, ParsedReference] = {}
    for position, token in enumerate(tokens):
        if not (
            getattr(token, "type", None) == "FUNC"
            and getattr(token, "subtype", None) == "OPEN"
            and _function_name(token)
            in {
                _IMPLICIT_INTERSECTION_FUNCTION,
                _LITERAL_IMPLICIT_INTERSECTION_FUNCTION,
            }
        ):
            continue
        closing = _matching_group_close(tokens, position, len(tokens))
        if closing is None:
            continue
        arguments = _function_argument_spans(tokens, position + 1, closing)
        if len(arguments) != 1:
            continue
        start, end = arguments[0]
        meaningful = [
            index for index in range(start, end) if not _is_whitespace(tokens[index])
        ]
        if len(meaningful) != 1:
            continue
        reference_position = meaningful[0]
        reference_token = tokens[reference_position]
        if not (
            getattr(reference_token, "type", None) == "OPERAND"
            and getattr(reference_token, "subtype", None) == "RANGE"
        ):
            continue
        reference = parse_reference_token(str(getattr(reference_token, "value", "")))
        if reference is None:
            continue
        selected = _implicit_intersection_selection(reference, origin)
        if selected is not None:
            replacements[reference_position] = selected
    return replacements


def _skip_excel_double_quoted_literal(formula: str, start: int) -> int:
    """Return the first index after a double-quoted Excel string literal.

    Excel escapes a literal double quote by doubling it.  An unterminated
    literal deliberately consumes the remaining text: an unclosed string is
    not a sound place from which to infer a DDE expression.
    """
    index = start + 1
    while index < len(formula):
        if formula[index] != '"':
            index += 1
            continue
        if index + 1 < len(formula) and formula[index + 1] == '"':
            index += 2
            continue
        return index + 1
    return len(formula)


def _single_quoted_component_end(formula: str, start: int) -> int | None:
    """Return the end of an Excel single-quoted identifier, if complete."""
    index = start + 1
    while index < len(formula):
        if formula[index] != "'":
            index += 1
            continue
        if index + 1 < len(formula) and formula[index + 1] == "'":
            index += 2
            continue
        return index + 1
    return None


def _quoted_component_start(formula: str, end: int) -> int | None:
    """Return a complete quoted component ending at ``end``, if present."""
    index = 0
    while index < end:
        if formula[index] == '"':
            index = _skip_excel_double_quoted_literal(formula, index)
            continue
        if formula[index] != "'":
            index += 1
            continue
        quoted_end = _single_quoted_component_end(formula, index)
        if quoted_end is None:
            return None
        if quoted_end == end:
            return index if quoted_end > index + 2 else None
        index = quoted_end
    return None


def _dde_component_start(formula: str, end: int) -> int | None:
    """Return the start of a quoted or unquoted DDE component before ``end``."""
    if end <= 0:
        return None
    if formula[end - 1] == "'":
        return _quoted_component_start(formula, end)
    index = end
    while (
        index > 0
        and formula[index - 1] not in _DDE_UNQUOTED_COMPONENT_DELIMITERS
    ):
        index -= 1
    return index if index < end else None


def _dde_component_end(formula: str, start: int) -> int | None:
    """Return the end of one nonempty quoted or unquoted DDE component."""
    if start >= len(formula):
        return None
    if formula[start] == "'":
        quoted_end = _single_quoted_component_end(formula, start)
        if quoted_end is None or quoted_end == start + 2:
            return None
        return quoted_end
    index = start
    while (
        index < len(formula)
        and formula[index] not in _DDE_UNQUOTED_COMPONENT_DELIMITERS
    ):
        index += 1
    return index if index > start else None


def _has_dde_expression_boundary(formula: str, component_start: int) -> bool:
    """Return whether a DDE service starts at a conservative expression edge."""
    index = component_start - 1
    while index >= 0 and formula[index].isspace():
        index -= 1
    return index < 0 or formula[index] in _DDE_EXPRESSION_BOUNDARIES


def _formula_dde_link_markers(formula: str) -> tuple[str, ...]:
    """Return opaque markers for explicit direct DDE formula syntax.

    The scanner recognizes only an application, a pipe *outside* quoted text,
    a topic, and an exclamation mark: ``application|topic!item``.  It accepts
    Excel's single-quoted application/topic spelling and a missing item, since
    the latter is used by DDE command-style formulas.  Pipes in double-quoted
    literals and ordinary single-quoted sheet names are skipped, so a formula
    such as ``='cmd|/C calc'!A0`` is not misclassified.  This is lexical
    inventory only; no formula is evaluated and no DDE endpoint is contacted.
    """
    markers: list[str] = []
    index = 0
    while index < len(formula):
        character = formula[index]
        if character == '"':
            index = _skip_excel_double_quoted_literal(formula, index)
            continue
        if character == "'":
            quoted_end = _single_quoted_component_end(formula, index)
            index = quoted_end if quoted_end is not None else len(formula)
            continue
        if character != "|":
            index += 1
            continue

        service_start = _dde_component_start(formula, index)
        topic_start = index + 1
        topic_end = _dde_component_end(formula, topic_start)
        if (
            service_start is None
            or not _has_dde_expression_boundary(formula, service_start)
            or topic_end is None
            or topic_end >= len(formula)
            or formula[topic_end] != "!"
        ):
            index += 1
            continue

        item_start = topic_end + 1
        if item_start >= len(formula) or formula[item_start] in _DDE_ITEM_TERMINATORS:
            item_end = item_start
        else:
            item_end = _dde_component_end(formula, item_start)
            if (
                item_end is None
                or (
                    item_end < len(formula)
                    and formula[item_end] not in _DDE_ITEM_TERMINATORS
                )
            ):
                index += 1
                continue
        markers.append(_FORMULA_DDE_LINK_MARKER)
        index = max(item_end, index + 1)
    return tuple(markers)


def inspect_formula(
    formula: str,
    named_references: Mapping[str, Sequence[ParsedReference]] | None = None,
    structured_tables: Mapping[str, StructuredTable] | None = None,
    origin: tuple[str, str] | None = None,
    sheet_order: Sequence[str] | None = None,
    named_function_references: (
        Mapping[str, Sequence[ParsedReference] | None] | None
    ) = None,
    named_function_external_workbook_defined_name_references: (
        Mapping[str, Sequence[ExternalWorkbookDefinedNameReference]] | None
    ) = None,
    named_function_external_workbook_references: (
        Mapping[str, Sequence[ExternalWorkbookReference]] | None
    ) = None,
    named_function_external_workbook_three_d_references: (
        Mapping[str, Sequence[ExternalWorkbookThreeDReference]] | None
    ) = None,
    named_function_external_workbook_structured_references: (
        Mapping[str, Sequence[ExternalWorkbookStructuredReference]] | None
    ) = None,
    named_custom_function_candidates: Mapping[str, Sequence[str]] | None = None,
    named_function_custom_function_candidates: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_unqualified_runtime_function_candidates: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_function_unqualified_runtime_function_candidates: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_formula_external_action_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_function_formula_external_action_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_formula_dde_link_markers: Mapping[str, Sequence[str]] | None = None,
    named_function_formula_dde_link_markers: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_worksheet_code_resource_registration_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_function_worksheet_code_resource_registration_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_formula_defined_xlm_registration_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_function_formula_defined_xlm_registration_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_formula_defined_xlm_evaluation_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_function_formula_defined_xlm_evaluation_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_formula_defined_xlm_action_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_function_formula_defined_xlm_action_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_formula_defined_xlm_get_cell_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_function_formula_defined_xlm_get_cell_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_formula_defined_xlm_environment_information_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_function_formula_defined_xlm_environment_information_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_formula_environment_information_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_function_formula_environment_information_functions: (
        Mapping[str, Sequence[str]] | None
    ) = None,
    named_external_workbook_defined_name_references: (
        Mapping[str, Sequence[ExternalWorkbookDefinedNameReference]] | None
    ) = None,
    indexed_external_workbook_paths: Mapping[int, str] | None = None,
    named_external_workbook_references: (
        Mapping[str, Sequence[ExternalWorkbookReference]] | None
    ) = None,
    named_external_workbook_three_d_references: (
        Mapping[str, Sequence[ExternalWorkbookThreeDReference]] | None
    ) = None,
    named_external_workbook_structured_references: (
        Mapping[str, Sequence[ExternalWorkbookStructuredReference]] | None
    ) = None,
    *,
    inspect_formula_defined_xlm_registrations: bool = False,
    inspect_formula_defined_xlm_evaluations: bool = False,
    inspect_formula_defined_xlm_actions: bool = False,
    inspect_formula_defined_xlm_get_cell_calls: bool = False,
    inspect_formula_defined_xlm_environment_information_calls: bool = False,
) -> FormulaInspection:
    """Inspect static reference coverage while resolving known named ranges.

    A caller provides case-folded name and named-LAMBDA maps assembled from the
    workbook. It may also provide private, package-derived external-name and
    external-A1 maps; those maps never cause a filesystem lookup. Supported
    fully qualified table references are resolved from table metadata,
    context-bound row references require the formula origin, and 3-D references
    require workbook tab order. Other non-A1 tokens are returned explicitly
    instead of being silently omitted from the graph. A ``None`` named-function
    value records a known LAMBDA whose definition is not safe to expand, so its
    call remains a visible coverage gap.
    """
    direct_formula_dde_link_markers = _formula_dde_link_markers(formula)
    (
        tokens,
        literal_spill_tokens,
        literal_implicit_intersection_tokens,
    ) = _tokenize_formula(formula)
    if tokens is None:
        return FormulaInspection(
            (),
            (),
            (),
            formula_dde_link_markers=direct_formula_dde_link_markers,
            tokenization_failed=True,
            spill_reference_tokens=literal_spill_tokens,
            implicit_intersection_tokens=literal_implicit_intersection_tokens,
        )
    resolved_names = named_references or {}
    resolved_named_external_workbook_defined_names = (
        named_external_workbook_defined_name_references or {}
    )
    resolved_indexed_external_workbook_paths = indexed_external_workbook_paths or {}
    resolved_named_external_workbooks = named_external_workbook_references or {}
    resolved_named_external_workbook_three_d_references = (
        named_external_workbook_three_d_references or {}
    )
    resolved_named_external_workbook_structured_references = (
        named_external_workbook_structured_references or {}
    )
    resolved_named_functions = named_function_references or {}
    resolved_named_function_external_workbook_defined_names = (
        named_function_external_workbook_defined_name_references or {}
    )
    resolved_named_function_external_workbooks = (
        named_function_external_workbook_references or {}
    )
    resolved_named_function_external_workbook_three_d_references = (
        named_function_external_workbook_three_d_references or {}
    )
    resolved_named_function_external_workbook_structured_references = (
        named_function_external_workbook_structured_references or {}
    )
    resolved_named_custom_functions = named_custom_function_candidates or {}
    resolved_named_function_custom_functions = (
        named_function_custom_function_candidates or {}
    )
    resolved_named_unqualified_runtime_functions = (
        named_unqualified_runtime_function_candidates or {}
    )
    resolved_named_function_unqualified_runtime_functions = (
        named_function_unqualified_runtime_function_candidates or {}
    )
    resolved_named_formula_external_actions = (
        named_formula_external_action_functions or {}
    )
    resolved_named_function_formula_external_actions = (
        named_function_formula_external_action_functions or {}
    )
    resolved_named_formula_dde_links = named_formula_dde_link_markers or {}
    resolved_named_function_formula_dde_links = (
        named_function_formula_dde_link_markers or {}
    )
    resolved_named_worksheet_code_resource_registrations = (
        named_worksheet_code_resource_registration_functions or {}
    )
    resolved_named_function_worksheet_code_resource_registrations = (
        named_function_worksheet_code_resource_registration_functions or {}
    )
    resolved_named_formula_defined_xlm_registrations = (
        named_formula_defined_xlm_registration_functions or {}
    )
    resolved_named_function_formula_defined_xlm_registrations = (
        named_function_formula_defined_xlm_registration_functions or {}
    )
    resolved_named_formula_defined_xlm_evaluations = (
        named_formula_defined_xlm_evaluation_functions or {}
    )
    resolved_named_function_formula_defined_xlm_evaluations = (
        named_function_formula_defined_xlm_evaluation_functions or {}
    )
    resolved_named_formula_defined_xlm_actions = (
        named_formula_defined_xlm_action_functions or {}
    )
    resolved_named_function_formula_defined_xlm_actions = (
        named_function_formula_defined_xlm_action_functions or {}
    )
    resolved_named_formula_defined_xlm_get_cell_calls = (
        named_formula_defined_xlm_get_cell_functions or {}
    )
    resolved_named_function_formula_defined_xlm_get_cell_calls = (
        named_function_formula_defined_xlm_get_cell_functions or {}
    )
    resolved_named_formula_defined_xlm_environment_information_calls = (
        named_formula_defined_xlm_environment_information_functions or {}
    )
    resolved_named_function_formula_defined_xlm_environment_information_calls = (
        named_function_formula_defined_xlm_environment_information_functions or {}
    )
    resolved_named_formula_environment_information_calls = (
        named_formula_environment_information_functions or {}
    )
    resolved_named_function_formula_environment_information_calls = (
        named_function_formula_environment_information_functions or {}
    )
    resolved_tables = structured_tables or {}
    references: list[ParsedReference] = []
    unresolved_range_tokens: list[str] = []
    dynamic_reference_functions: list[str] = []
    external_workbook_references: list[ExternalWorkbookReference] = []
    external_workbook_three_d_references: list[ExternalWorkbookThreeDReference] = []
    external_workbook_structured_references: list[
        ExternalWorkbookStructuredReference
    ] = []
    external_workbook_defined_name_references: list[
        ExternalWorkbookDefinedNameReference
    ] = []
    external_action_functions: list[str] = []
    formula_dde_link_markers: list[str] = list(direct_formula_dde_link_markers)
    python_functions: list[str] = []
    office_custom_function_candidates: list[str] = []
    unqualified_runtime_function_candidates: list[str] = []
    worksheet_code_resource_registration_functions: list[str] = []
    formula_defined_xlm_registration_functions: list[str] = []
    formula_defined_xlm_evaluation_functions: list[str] = []
    formula_defined_xlm_action_functions: list[str] = []
    formula_defined_xlm_get_cell_functions: list[str] = []
    formula_defined_xlm_environment_information_functions: list[str] = []
    formula_environment_information_functions: list[str] = []
    formula_environment_information_implicit_cell_reference_count = 0
    formula_environment_information_implicit_sheets_reference_count = 0
    three_d_reference_tokens: list[str] = []
    spill_reference_tokens: list[str] = list(literal_spill_tokens)
    implicit_intersection_tokens: list[str] = list(literal_implicit_intersection_tokens)
    local_variable_indexes = _local_variable_token_indexes(tokens)
    implicit_intersection_replacements = _implicit_intersection_reference_replacements(
        tokens, origin
    )

    def extend_formula_environment_information_signals(
        signals: Sequence[str],
    ) -> None:
        nonlocal formula_environment_information_implicit_cell_reference_count
        nonlocal formula_environment_information_implicit_sheets_reference_count
        for signal in signals:
            if signal == _FORMULA_ENVIRONMENT_INFORMATION_IMPLICIT_CELL_REFERENCE_MARKER:
                formula_environment_information_implicit_cell_reference_count += 1
            elif signal == _FORMULA_ENVIRONMENT_INFORMATION_IMPLICIT_SHEETS_REFERENCE_MARKER:
                formula_environment_information_implicit_sheets_reference_count += 1
            else:
                formula_environment_information_functions.append(signal)

    def extend_named_external_endpoints(
        raw: str,
        *,
        workbook_references: Sequence[ExternalWorkbookReference] = (),
        three_d_references: Sequence[ExternalWorkbookThreeDReference] = (),
        structured_references: Sequence[ExternalWorkbookStructuredReference] = (),
        defined_name_references: Sequence[ExternalWorkbookDefinedNameReference] = (),
    ) -> bool:
        """Add every private endpoint retained by one resolved name or function.

        Formula-defined names may have multiple endpoint kinds, such as a
        static external A1 range and a table selector. Keep all of them rather
        than stopping at the first matching map; this remains dependency
        extraction, not evaluation.
        """
        if not (
            workbook_references
            or three_d_references
            or structured_references
            or defined_name_references
        ):
            return False
        references.append(
            ParsedReference(
                None,
                None,
                None,
                None,
                None,
                raw,
                is_external=True,
            )
        )
        external_workbook_references.extend(workbook_references)
        external_workbook_three_d_references.extend(three_d_references)
        external_workbook_structured_references.extend(structured_references)
        external_workbook_defined_name_references.extend(defined_name_references)
        return True

    for position, token in enumerate(tokens):
        if token.type == "OPERAND" and token.subtype == "RANGE":
            if position in local_variable_indexes:
                continue
            if token.value.strip().startswith("@"):
                implicit_intersection_tokens.append(token.value.strip())
            if selected_reference := implicit_intersection_replacements.get(position):
                references.append(selected_reference)
                continue
            if indexed_external_workbook_structured_reference := (
                parse_external_link_indexed_workbook_structured_reference(token.value)
            ):
                # The package index is only a declaration position, never a
                # source filename. Keep an explicit external ledger entry even
                # when package metadata cannot safely bind it to a candidate.
                references.append(
                    ParsedReference(
                        None,
                        None,
                        None,
                        None,
                        None,
                        token.value,
                        is_external=True,
                    )
                )
                if source_path := resolved_indexed_external_workbook_paths.get(
                    indexed_external_workbook_structured_reference.index
                ):
                    external_workbook_structured_references.append(
                        ExternalWorkbookStructuredReference(
                            source_path=source_path,
                            table_name=(
                                indexed_external_workbook_structured_reference.table_name
                            ),
                            table_reference=(
                                indexed_external_workbook_structured_reference.table_reference
                            ),
                        )
                    )
                continue
            if indexed_external_name := (
                parse_external_link_indexed_defined_name_reference(token.value)
            ):
                index, name_key = indexed_external_name
                if source_path := resolved_indexed_external_workbook_paths.get(index):
                    # Keep indexed external-name syntax in the ordinary
                    # external-reference ledger while retaining the package
                    # target and source name only in private portfolio data.
                    references.append(
                        ParsedReference(
                            None,
                            None,
                            None,
                            None,
                            None,
                            token.value,
                            is_external=True,
                        )
                    )
                    external_workbook_defined_name_references.append(
                        ExternalWorkbookDefinedNameReference(source_path, name_key)
                    )
                    continue
            if indexed_external_sheet_name := (
                parse_external_link_indexed_sheet_defined_name_reference(token.value)
            ):
                index, scope_sheet, name_key = indexed_external_sheet_name
                # A source sheet makes this an external sheet-local name, not
                # an A1 range or a workbook-scoped external name. Keep the
                # accounting edge even when the package index lacks a trusted
                # target; portfolio resolution itself remains candidate-only.
                references.append(
                    ParsedReference(
                        None,
                        None,
                        None,
                        None,
                        None,
                        token.value,
                        is_external=True,
                    )
                )
                if source_path := resolved_indexed_external_workbook_paths.get(index):
                    external_workbook_defined_name_references.append(
                        ExternalWorkbookDefinedNameReference(
                            source_path,
                            name_key,
                            scope_sheet,
                        )
                    )
                continue
            if indexed_external_workbook_three_d_reference := (
                parse_external_link_indexed_workbook_three_d_reference(token.value)
            ):
                # A package index never supplies a filename by itself. Retain
                # this private span only after the raw external-link metadata
                # has validated one exact candidate source spelling.
                references.append(
                    ParsedReference(
                        None,
                        None,
                        None,
                        None,
                        None,
                        token.value,
                        is_external=True,
                    )
                )
                if source_path := resolved_indexed_external_workbook_paths.get(
                    indexed_external_workbook_three_d_reference.index
                ):
                    external_workbook_three_d_references.append(
                        ExternalWorkbookThreeDReference(
                            source_path=source_path,
                            first_sheet=(
                                indexed_external_workbook_three_d_reference.first_sheet
                            ),
                            last_sheet=(
                                indexed_external_workbook_three_d_reference.last_sheet
                            ),
                            min_column=(
                                indexed_external_workbook_three_d_reference.min_column
                            ),
                            min_row=indexed_external_workbook_three_d_reference.min_row,
                            max_column=(
                                indexed_external_workbook_three_d_reference.max_column
                            ),
                            max_row=indexed_external_workbook_three_d_reference.max_row,
                        )
                    )
                continue
            if indexed_external_workbook_reference := (
                parse_external_link_indexed_workbook_reference(token.value)
            ):
                # The formula-level index has no filename semantics.  Only a
                # separately validated external-link declaration can provide
                # the private source spelling used by candidate portfolios.
                references.append(
                    ParsedReference(
                        None,
                        None,
                        None,
                        None,
                        None,
                        token.value,
                        is_external=True,
                    )
                )
                if source_path := resolved_indexed_external_workbook_paths.get(
                    indexed_external_workbook_reference.index
                ):
                    external_workbook_references.append(
                        ExternalWorkbookReference(
                            source_path=source_path,
                            sheet=indexed_external_workbook_reference.sheet,
                            min_column=indexed_external_workbook_reference.min_column,
                            min_row=indexed_external_workbook_reference.min_row,
                            max_column=indexed_external_workbook_reference.max_column,
                            max_row=indexed_external_workbook_reference.max_row,
                        )
                    )
                continue
            if direct_external_table_prefix := (
                _parse_direct_external_workbook_only_prefix(token.value)
            ):
                source_path, selector = direct_external_table_prefix
                # A valid book-only table prefix is explicit external syntax
                # even when its selector is row-relative or outside our static
                # subset. Keep it in the ordinary external ledger so its raw
                # source spelling does not become an unresolved-token payload.
                references.append(
                    ParsedReference(
                        None,
                        None,
                        None,
                        None,
                        None,
                        token.value,
                        is_external=True,
                    )
                )
                if parsed_selector := _static_external_structured_table_reference(
                    selector
                ):
                    table_name, table_reference = parsed_selector
                    # The static source table becomes a dependency only after
                    # candidate portfolio analysis resolves this private path
                    # and selector against one inspected table definition.
                    external_workbook_structured_references.append(
                        ExternalWorkbookStructuredReference(
                            source_path=source_path,
                            table_name=table_name,
                            table_reference=table_reference,
                        )
                    )
                continue
            reference = parse_reference_token(token.value)
            if reference is not None:
                references.append(reference)
                if reference.is_external:
                    if external_workbook_three_d_reference := (
                        parse_external_workbook_three_d_reference(token.value)
                    ):
                        external_workbook_three_d_references.append(
                            external_workbook_three_d_reference
                        )
                    elif external_workbook_reference := (
                        parse_external_workbook_reference(token.value)
                    ):
                        external_workbook_references.append(external_workbook_reference)
                    elif external_workbook_sheet_name_reference := (
                        parse_external_workbook_sheet_defined_name_reference(
                            token.value
                        )
                    ):
                        external_workbook_defined_name_references.append(
                            external_workbook_sheet_name_reference
                        )
                continue
            if external_workbook_defined_name_reference := (
                parse_external_workbook_defined_name_reference(token.value)
            ):
                # Keep the ordinary external-reference accounting consistent
                # with external A1 tokens while retaining the source name only
                # as private portfolio-resolution data.
                references.append(
                    ParsedReference(
                        None,
                        None,
                        None,
                        None,
                        None,
                        token.value,
                        is_external=True,
                    )
                )
                external_workbook_defined_name_references.append(
                    external_workbook_defined_name_reference
                )
                continue
            three_d_reference = resolve_3d_reference(token.value, sheet_order)
            if three_d_reference is not None:
                references.extend(three_d_reference)
                three_d_reference_tokens.append(token.value)
                continue
            named_key = reference_lookup_key(token.value)
            has_named_external_endpoints = extend_named_external_endpoints(
                token.value,
                workbook_references=resolved_named_external_workbooks.get(
                    named_key, ()
                ),
                three_d_references=(
                    resolved_named_external_workbook_three_d_references.get(
                        named_key, ()
                    )
                ),
                structured_references=(
                    resolved_named_external_workbook_structured_references.get(
                        named_key, ()
                    )
                ),
                defined_name_references=(
                    resolved_named_external_workbook_defined_names.get(named_key, ())
                ),
            )
            if named_key in resolved_names:
                references.extend(resolved_names[named_key])
                external_action_functions.extend(
                    resolved_named_formula_external_actions.get(named_key, ())
                )
                formula_dde_link_markers.extend(
                    resolved_named_formula_dde_links.get(named_key, ())
                )
                office_custom_function_candidates.extend(
                    resolved_named_custom_functions.get(named_key, ())
                )
                unqualified_runtime_function_candidates.extend(
                    resolved_named_unqualified_runtime_functions.get(named_key, ())
                )
                worksheet_code_resource_registration_functions.extend(
                    resolved_named_worksheet_code_resource_registrations.get(
                        named_key, ()
                    )
                )
                formula_defined_xlm_registration_functions.extend(
                    resolved_named_formula_defined_xlm_registrations.get(named_key, ())
                )
                formula_defined_xlm_evaluation_functions.extend(
                    resolved_named_formula_defined_xlm_evaluations.get(named_key, ())
                )
                formula_defined_xlm_action_functions.extend(
                    resolved_named_formula_defined_xlm_actions.get(named_key, ())
                )
                formula_defined_xlm_get_cell_functions.extend(
                    resolved_named_formula_defined_xlm_get_cell_calls.get(named_key, ())
                )
                formula_defined_xlm_environment_information_functions.extend(
                    resolved_named_formula_defined_xlm_environment_information_calls.get(
                        named_key, ()
                    )
                )
                extend_formula_environment_information_signals(
                    resolved_named_formula_environment_information_calls.get(
                        named_key, ()
                    )
                )
                continue
            if has_named_external_endpoints:
                continue
            named_formula_dde_links = resolved_named_formula_dde_links.get(
                named_key, ()
            )
            if named_external_actions := resolved_named_formula_external_actions.get(
                named_key
            ):
                external_action_functions.extend(named_external_actions)
            if named_formula_dde_links:
                formula_dde_link_markers.extend(named_formula_dde_links)
            if named_custom_functions := resolved_named_custom_functions.get(named_key):
                office_custom_function_candidates.extend(named_custom_functions)
            if named_unqualified_runtime_functions := (
                resolved_named_unqualified_runtime_functions.get(named_key)
            ):
                unqualified_runtime_function_candidates.extend(
                    named_unqualified_runtime_functions
                )
            if (
                named_worksheet_code_resource_registrations := (
                    resolved_named_worksheet_code_resource_registrations.get(named_key)
                )
            ):
                worksheet_code_resource_registration_functions.extend(
                    named_worksheet_code_resource_registrations
                )
            if named_formula_defined_xlm_registrations := (
                resolved_named_formula_defined_xlm_registrations.get(named_key)
            ):
                formula_defined_xlm_registration_functions.extend(
                    named_formula_defined_xlm_registrations
                )
            if named_formula_defined_xlm_evaluations := (
                resolved_named_formula_defined_xlm_evaluations.get(named_key)
            ):
                formula_defined_xlm_evaluation_functions.extend(
                    named_formula_defined_xlm_evaluations
                )
            if named_formula_defined_xlm_actions := (
                resolved_named_formula_defined_xlm_actions.get(named_key)
            ):
                formula_defined_xlm_action_functions.extend(
                    named_formula_defined_xlm_actions
                )
            if named_formula_defined_xlm_get_cell_calls := (
                resolved_named_formula_defined_xlm_get_cell_calls.get(named_key)
            ):
                formula_defined_xlm_get_cell_functions.extend(
                    named_formula_defined_xlm_get_cell_calls
                )
            if named_formula_defined_xlm_environment_information_calls := (
                resolved_named_formula_defined_xlm_environment_information_calls.get(
                    named_key
                )
            ):
                formula_defined_xlm_environment_information_functions.extend(
                    named_formula_defined_xlm_environment_information_calls
                )
            if named_formula_environment_information_calls := (
                resolved_named_formula_environment_information_calls.get(named_key)
            ):
                extend_formula_environment_information_signals(
                    named_formula_environment_information_calls
                )
            # A direct DDE name can be classified even though its raw syntax is
            # intentionally outside the ordinary static-reference graph.  The
            # dedicated private ledger retains that coverage boundary, so do
            # not also expose its defined-name identity as an unknown token.
            if named_formula_dde_links:
                continue
            table_reference = resolve_structured_reference(
                token.value, resolved_tables, origin
            )
            if table_reference is not None:
                references.extend(table_reference)
                continue
            unresolved_range_tokens.append(token.value)
        elif token.type == "FUNC" and token.subtype == "OPEN":
            raw_function_name = token.value.rstrip("(").strip()
            if position not in local_variable_indexes:
                function_key = _function_lookup_key(token)
                if function_key in resolved_named_functions:
                    function_references = resolved_named_functions[function_key]
                    named_function_dde_links = (
                        resolved_named_function_formula_dde_links.get(
                            function_key, ()
                        )
                    )
                    has_named_function_external_endpoints = (
                        extend_named_external_endpoints(
                            raw_function_name,
                            workbook_references=(
                                resolved_named_function_external_workbooks.get(
                                    function_key, ()
                                )
                            ),
                            three_d_references=(
                                resolved_named_function_external_workbook_three_d_references.get(
                                    function_key, ()
                                )
                            ),
                            structured_references=(
                                resolved_named_function_external_workbook_structured_references.get(
                                    function_key, ()
                                )
                            ),
                            defined_name_references=(
                                resolved_named_function_external_workbook_defined_names.get(
                                    function_key, ()
                                )
                            ),
                        )
                    )
                    if function_references is None:
                        if (
                            not named_function_dde_links
                            and not has_named_function_external_endpoints
                        ):
                            unresolved_range_tokens.append(
                                token.value.rstrip("(").strip()
                            )
                    else:
                        references.extend(function_references)
                    external_action_functions.extend(
                        resolved_named_function_formula_external_actions.get(
                            function_key, ()
                        )
                    )
                    formula_dde_link_markers.extend(named_function_dde_links)
                    office_custom_function_candidates.extend(
                        resolved_named_function_custom_functions.get(function_key, ())
                    )
                    unqualified_runtime_function_candidates.extend(
                        resolved_named_function_unqualified_runtime_functions.get(
                            function_key, ()
                        )
                    )
                    worksheet_code_resource_registration_functions.extend(
                        resolved_named_function_worksheet_code_resource_registrations.get(
                            function_key, ()
                        )
                    )
                    formula_defined_xlm_registration_functions.extend(
                        resolved_named_function_formula_defined_xlm_registrations.get(
                            function_key, ()
                        )
                    )
                    formula_defined_xlm_evaluation_functions.extend(
                        resolved_named_function_formula_defined_xlm_evaluations.get(
                            function_key, ()
                        )
                    )
                    formula_defined_xlm_action_functions.extend(
                        resolved_named_function_formula_defined_xlm_actions.get(
                            function_key, ()
                        )
                    )
                    formula_defined_xlm_get_cell_functions.extend(
                        resolved_named_function_formula_defined_xlm_get_cell_calls.get(
                            function_key, ()
                        )
                    )
                    formula_defined_xlm_environment_information_functions.extend(
                        resolved_named_function_formula_defined_xlm_environment_information_calls.get(
                            function_key, ()
                        )
                    )
                    extend_formula_environment_information_signals(
                        resolved_named_function_formula_environment_information_calls.get(
                            function_key, ()
                        )
                    )
                if (
                    function_key not in resolved_names
                    and function_key not in resolved_named_functions
                    and (
                        custom_function_candidate := _office_custom_function_candidate(
                            token
                        )
                    )
                    is not None
                ):
                    office_custom_function_candidates.append(custom_function_candidate)
                if (
                    function_key not in resolved_names
                    and function_key not in resolved_named_functions
                    and (
                        unqualified_runtime_function_candidate := (
                            _unqualified_runtime_function_candidate(token)
                        )
                    )
                    is not None
                ):
                    unqualified_runtime_function_candidates.append(
                        unqualified_runtime_function_candidate
                    )
                if (
                    function_key not in resolved_names
                    and function_key not in resolved_named_functions
                    and (
                        worksheet_code_resource_registration_function := (
                            _worksheet_code_resource_registration_function(token)
                        )
                    )
                    is not None
                ):
                    worksheet_code_resource_registration_functions.append(
                        worksheet_code_resource_registration_function
                    )
                if (
                    inspect_formula_defined_xlm_registrations
                    and function_key not in resolved_names
                    and function_key not in resolved_named_functions
                    and (
                        formula_defined_xlm_registration_function := (
                            _formula_defined_xlm_registration_function(token)
                        )
                    )
                    is not None
                ):
                    formula_defined_xlm_registration_functions.append(
                        formula_defined_xlm_registration_function
                    )
                if (
                    inspect_formula_defined_xlm_evaluations
                    and function_key not in resolved_names
                    and function_key not in resolved_named_functions
                    and (
                        formula_defined_xlm_evaluation_function := (
                            _formula_defined_xlm_evaluation_function(token)
                        )
                    )
                    is not None
                ):
                    formula_defined_xlm_evaluation_functions.append(
                        formula_defined_xlm_evaluation_function
                    )
                if (
                    inspect_formula_defined_xlm_actions
                    and function_key not in resolved_names
                    and function_key not in resolved_named_functions
                    and (
                        formula_defined_xlm_action_function := (
                            _formula_defined_xlm_action_function(token)
                        )
                    )
                    is not None
                ):
                    formula_defined_xlm_action_functions.append(
                        formula_defined_xlm_action_function
                    )
                if (
                    inspect_formula_defined_xlm_get_cell_calls
                    and function_key not in resolved_names
                    and function_key not in resolved_named_functions
                    and (
                        formula_defined_xlm_get_cell_function := (
                            _formula_defined_xlm_get_cell_function(token)
                        )
                    )
                    is not None
                ):
                    formula_defined_xlm_get_cell_functions.append(
                        formula_defined_xlm_get_cell_function
                    )
                if (
                    inspect_formula_defined_xlm_environment_information_calls
                    and function_key not in resolved_names
                    and function_key not in resolved_named_functions
                    and (
                        formula_defined_xlm_environment_information_function := (
                            _formula_defined_xlm_environment_information_function(
                                token
                            )
                        )
                    )
                    is not None
                ):
                    formula_defined_xlm_environment_information_functions.append(
                        formula_defined_xlm_environment_information_function
                    )
                if (
                    function_key not in resolved_names
                    and function_key not in resolved_named_functions
                    and (
                        formula_environment_information_function := (
                            _formula_environment_information_function(tokens, position)
                        )
                    )
                    is not None
                ):
                    (
                        function_name,
                        has_implicit_cell_reference,
                        has_implicit_sheets_reference,
                    ) = (
                        formula_environment_information_function
                    )
                    formula_environment_information_functions.append(function_name)
                    if has_implicit_cell_reference:
                        formula_environment_information_implicit_cell_reference_count += 1
                    if has_implicit_sheets_reference:
                        formula_environment_information_implicit_sheets_reference_count += 1
            function_name = _function_name(token)
            if function_name in _DYNAMIC_REFERENCE_FUNCTIONS:
                dynamic_reference_functions.append(function_name)
            if (
                position not in local_variable_indexes
                and function_key not in resolved_names
                and function_key not in resolved_named_functions
                and function_name in _EXTERNAL_ACTION_FUNCTIONS
            ):
                external_action_functions.append(function_name)
            if (
                position not in local_variable_indexes
                and function_name in _PYTHON_FUNCTIONS
            ):
                python_functions.append(function_name)
            if function_name == _SPILL_REFERENCE_FUNCTION:
                spill_reference_tokens.append(raw_function_name)
            if raw_function_name.startswith("@"):
                implicit_intersection_tokens.append(raw_function_name)
            elif function_name == _IMPLICIT_INTERSECTION_FUNCTION:
                implicit_intersection_tokens.append(raw_function_name)
    return FormulaInspection(
        references=tuple(references),
        unresolved_range_tokens=tuple(dict.fromkeys(unresolved_range_tokens)),
        dynamic_reference_functions=tuple(dict.fromkeys(dynamic_reference_functions)),
        external_workbook_references=tuple(dict.fromkeys(external_workbook_references)),
        external_workbook_three_d_references=tuple(
            dict.fromkeys(external_workbook_three_d_references)
        ),
        external_workbook_structured_references=tuple(
            dict.fromkeys(external_workbook_structured_references)
        ),
        external_workbook_defined_name_references=tuple(
            dict.fromkeys(external_workbook_defined_name_references)
        ),
        external_action_functions=tuple(external_action_functions),
        formula_dde_link_markers=tuple(formula_dde_link_markers),
        python_functions=tuple(python_functions),
        office_custom_function_candidates=tuple(office_custom_function_candidates),
        unqualified_runtime_function_candidates=tuple(
            unqualified_runtime_function_candidates
        ),
        worksheet_code_resource_registration_functions=tuple(
            worksheet_code_resource_registration_functions
        ),
        formula_defined_xlm_registration_functions=tuple(
            formula_defined_xlm_registration_functions
        ),
        formula_defined_xlm_evaluation_functions=tuple(
            formula_defined_xlm_evaluation_functions
        ),
        formula_defined_xlm_action_functions=tuple(
            formula_defined_xlm_action_functions
        ),
        formula_defined_xlm_get_cell_functions=tuple(
            formula_defined_xlm_get_cell_functions
        ),
        formula_defined_xlm_environment_information_functions=tuple(
            formula_defined_xlm_environment_information_functions
        ),
        formula_environment_information_functions=tuple(
            formula_environment_information_functions
        ),
        formula_environment_information_implicit_cell_reference_count=(
            formula_environment_information_implicit_cell_reference_count
        ),
        formula_environment_information_implicit_sheets_reference_count=(
            formula_environment_information_implicit_sheets_reference_count
        ),
        three_d_reference_tokens=tuple(dict.fromkeys(three_d_reference_tokens)),
        spill_reference_tokens=tuple(dict.fromkeys(spill_reference_tokens)),
        implicit_intersection_tokens=tuple(dict.fromkeys(implicit_intersection_tokens)),
    )


def extract_references(
    formula: str,
    named_references: Mapping[str, Sequence[ParsedReference]] | None = None,
    structured_tables: Mapping[str, StructuredTable] | None = None,
    origin: tuple[str, str] | None = None,
    sheet_order: Sequence[str] | None = None,
) -> list[ParsedReference]:
    """Return A1-style, supplied named-range, and static table references."""
    return list(
        inspect_formula(
            formula, named_references, structured_tables, origin, sheet_order
        ).references
    )


def has_broken_reference(formula: str) -> bool:
    """Return whether a tokenized formula contains an explicit ``#REF!`` operand.

    A raw substring would incorrectly classify ordinary text such as
    ``="#REF!"`` and a valid quoted worksheet name such as ``'#REF!'!A1``.
    Tokenization lets the caller distinguish the actual SpreadsheetML error
    operand without evaluating the formula.  If FormulaFence cannot tokenize a
    formula, the caller retains its separate tokenization-coverage boundary
    rather than guessing from raw text.
    """
    tokens, _, _ = _tokenize_formula(formula)
    return bool(
        tokens is not None
        and any(
            token.type == "OPERAND"
            and token.subtype == "ERROR"
            and token.value.casefold() == "#ref!"
            for token in tokens
        )
    )
