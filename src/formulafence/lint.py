"""Conservative, single-workbook formula linting."""

from __future__ import annotations

import re
from collections import defaultdict
from collections.abc import Iterator
from dataclasses import dataclass, field
from numbers import Real

from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, range_boundaries

from formulafence.formulas import (
    MAX_EXCEL_COLUMN,
    MAX_EXCEL_ROW,
    choose_literal_index_mismatch_count,
    conditional_aggregate_range_shape_mismatches,
    lookup_return_index_mismatches,
    mmult_dimension_mismatch_count,
    parse_reference_token,
    randbetween_literal_bound_mismatch_count,
    sumproduct_range_shape_mismatches,
)
from formulafence.models import (
    ArrayFormulaRange,
    CellKey,
    Finding,
    FormulaFenceError,
    FormulaLintReport,
    WorkbookSnapshot,
    display_location,
)

DEFAULT_MAX_FORMULA_PATTERN_FINDINGS = 10_000
DEFAULT_MAX_AGGREGATE_OMISSION_GAP_CELLS = 128

_MIN_AGGREGATE_OMISSION_GAP_CELLS = 2
_SIMPLE_NUMERIC_AGGREGATE = re.compile(
    r"^\s*=\s*(?P<function>SUM|AVERAGE|MIN|MAX|COUNT)\s*"
    r"\(\s*(?P<reference>[^()]+?)\s*\)\s*$",
    re.IGNORECASE,
)

_PATTERN_ORIENTATIONS = (
    ("row", 0, 1),
    ("column", 1, 0),
)

_IGNORED_ERROR_SUPPRESSION_FIELDS = (
    ("evaluation_error", "evaluation_error_count"),
    ("inconsistent_formula", "inconsistent_formula_count"),
    ("formula_range_omission", "formula_range_omission_count"),
    ("unlocked_formula", "unlocked_formula_count"),
    ("empty_cell_reference", "empty_cell_reference_count"),
    ("list_data_validation", "list_data_validation_count"),
    ("calculated_column", "calculated_column_count"),
    ("number_stored_as_text", "number_stored_as_text_count"),
    ("two_digit_text_year", "two_digit_text_year_count"),
)


@dataclass(frozen=True)
class _FormulaPatternEvidence:
    """One copied-peer pattern supporting a lint finding."""

    orientation: str
    preceding: CellKey
    following: CellKey
    supporting: CellKey


@dataclass
class _FormulaPatternCandidate:
    """One target with one or more independent local pattern signals."""

    kind: str
    evidence: list[_FormulaPatternEvidence] = field(default_factory=list)


@dataclass(frozen=True)
class _AggregateOmissionCandidate:
    """One bounded, contiguous numeric run immediately outside an aggregate."""

    function: str
    orientation: str
    referenced_range: str
    omitted_range: str
    omitted_cell_count: int


def _location_sort_key(location: CellKey) -> tuple[str, int, int]:
    sheet, coordinate = location
    row, column = coordinate_to_tuple(coordinate)
    return sheet.casefold(), row, column


def _offset_location(location: CellKey, row_delta: int, column_delta: int) -> CellKey | None:
    """Return one bounded relative cell coordinate without leaving Excel's grid."""
    sheet, coordinate = location
    row, column = coordinate_to_tuple(coordinate)
    target_row = row + row_delta
    target_column = column + column_delta
    if not (1 <= target_row <= MAX_EXCEL_ROW and 1 <= target_column <= MAX_EXCEL_COLUMN):
        return None
    return sheet, f"{get_column_letter(target_column)}{target_row}"


def _array_ranges_by_sheet(snapshot: WorkbookSnapshot) -> dict[str, tuple[ArrayFormulaRange, ...]]:
    """Group compact array ranges so ordinary pattern evidence never crosses one."""
    grouped: dict[str, list[ArrayFormulaRange]] = defaultdict(list)
    for array_range in (
        *snapshot.legacy_array_formula_ranges,
        *snapshot.dynamic_array_formula_ranges,
    ):
        grouped[array_range.sheet].append(array_range)
    return {
        sheet: tuple(ranges)
        for sheet, ranges in grouped.items()
    }


def _is_array_member(
    snapshot: WorkbookSnapshot,
    location: CellKey,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
) -> bool:
    """Return whether a cell is an array anchor or observed result member."""
    if (
        location in snapshot.dynamic_array_formula_cells
        or location in snapshot.unclassified_array_formula_cells
    ):
        return True
    return any(
        array_range.contains(location)
        for array_range in array_ranges_by_sheet.get(location[0], ())
    )


def _eligible_formula(
    snapshot: WorkbookSnapshot,
    location: CellKey,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
) -> bool:
    """Return whether one ordinary formula is safe local-copy evidence."""
    cell = snapshot.cells.get(location)
    return bool(
        cell is not None
        and cell.is_formula
        and cell.value_type == "formula"
        and cell.formula_fingerprint is not None
        and location not in snapshot.tokenization_failure_cells
        and not _is_array_member(snapshot, location, array_ranges_by_sheet)
    )


def _matches_table_calculated_column_formula(
    snapshot: WorkbookSnapshot,
    location: CellKey,
    formula_fingerprint: str,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
) -> bool:
    """Return whether one ordinary formula matches a Table master fingerprint."""
    return bool(
        _eligible_formula(snapshot, location, array_ranges_by_sheet)
        and snapshot.cells[location].formula_fingerprint == formula_fingerprint
    )


def _table_calculated_column_exception_candidates(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
    *,
    formula_pattern_locations: set[CellKey],
    max_formula_pattern_findings: int,
    existing_finding_count: int,
) -> tuple[tuple[CellKey, str], ...]:
    """Return isolated interior exceptions to stored Excel Table master formulas.

    A calculated-column declaration is stronger evidence than a generic copied
    formula pattern, but the first and last table rows can be intentional
    exceptions (for example, a running comparison with no prior row). This
    accepts only a cell strictly inside the data body when its immediate rows
    above and below are eligible formulas matching the same declared master.
    It indexes existing snapshot cells rather than walking declared table
    rectangles, so an oversized sparse table cannot turn linting into a grid
    scan. Array territory and uninspectable formula cells remain outside the
    boundary. The same shared finding cap as the other formula signals applies
    while candidates are retained.
    """
    cells_by_sheet_column: dict[tuple[str, int], list[tuple[int, CellKey]]] = (
        defaultdict(list)
    )
    for location in snapshot.cells:
        row, column = coordinate_to_tuple(location[1])
        cells_by_sheet_column[(location[0], column)].append((row, location))
    for locations in cells_by_sheet_column.values():
        locations.sort()

    candidates: dict[CellKey, str] = {}
    for table in sorted(
        snapshot.tables.values(),
        key=lambda table: (table.sheet.casefold(), table.name.casefold()),
    ):
        if not table.calculated_column_formulas:
            continue
        try:
            min_column, min_row, max_column, max_row = range_boundaries(table.ref)
        except ValueError:
            continue
        first_data_row = min_row + table.header_row_count
        last_data_row = max_row - table.totals_row_count
        if first_data_row + 2 > last_data_row:
            continue
        for declaration in table.calculated_column_formulas:
            column = min_column + declaration.column_index - 1
            if not min_column <= column <= max_column:
                continue
            matching_rows = [
                row
                for row, location in cells_by_sheet_column.get((table.sheet, column), ())
                if (
                    first_data_row <= row <= last_data_row
                    and _matches_table_calculated_column_formula(
                        snapshot,
                        location,
                        declaration.formula_fingerprint,
                        array_ranges_by_sheet,
                    )
                )
            ]
            for preceding_row, following_row in zip(
                matching_rows,
                matching_rows[1:],
                strict=False,
            ):
                if following_row != preceding_row + 2:
                    continue
                target_row = preceding_row + 1
                if not first_data_row < target_row < last_data_row:
                    continue
                location = table.sheet, f"{get_column_letter(column)}{target_row}"
                if location in formula_pattern_locations:
                    continue
                if _is_array_member(snapshot, location, array_ranges_by_sheet):
                    continue
                target = snapshot.cells.get(location)
                if target is None:
                    exception_kind = "blank"
                elif target.is_formula:
                    if (
                        location in snapshot.broken_references
                        or not _eligible_formula(
                            snapshot,
                            location,
                            array_ranges_by_sheet,
                        )
                    ):
                        continue
                    exception_kind = "formula_mismatch"
                elif target.cell_type == "error":
                    exception_kind = "stored_error_value"
                elif isinstance(target.value, str):
                    exception_kind = "text_value"
                else:
                    exception_kind = "non_formula_value"
                if location in candidates:
                    continue
                if (
                    existing_finding_count + len(candidates)
                    >= max_formula_pattern_findings
                ):
                    raise FormulaFenceError(
                        "Formula lint exceeds "
                        "max_formula_pattern_findings="
                        f"{max_formula_pattern_findings}."
                    )
                candidates[location] = exception_kind
    return tuple(
        sorted(candidates.items(), key=lambda candidate: _location_sort_key(candidate[0]))
    )


def _conditional_aggregate_range_shape_candidates(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
    *,
    max_formula_pattern_findings: int,
) -> tuple[tuple[CellKey, int, int], ...]:
    """Return formula cells with an unambiguous conditional-range mismatch.

    The formula helper accepts only native ``SUMIFS``, ``COUNTIFS``,
    ``AVERAGEIFS``, ``MAXIFS``, and ``MINIFS`` calls, plus the exact OOXML
    ``_xlfn`` serializations for the latter two, whose relevant arguments are
    direct, bounded internal A1 ranges. This layer adds workbook context:
    ordinary inspectable formula cells only, no array territory, and no
    explicit broken-reference operand. Findings retain only aggregate counts,
    never formula text, range spellings, or table identity.
    """
    candidates: list[tuple[CellKey, int, int]] = []
    for location in sorted(snapshot.cells, key=_location_sort_key):
        if (
            location in snapshot.broken_references
            or not _eligible_formula(snapshot, location, array_ranges_by_sheet)
        ):
            continue
        formula = snapshot.cells[location].formula
        if formula is None:
            continue
        mismatches = conditional_aggregate_range_shape_mismatches(formula)
        if not mismatches:
            continue
        if len(candidates) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        candidates.append(
            (
                location,
                len(mismatches),
                sum(mismatched_range_count for _, mismatched_range_count in mismatches),
            )
        )
    return tuple(candidates)


def _sumproduct_range_shape_candidates(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
    *,
    existing_finding_count: int,
    max_formula_pattern_findings: int,
) -> tuple[tuple[CellKey, int, int], ...]:
    """Return formula cells with an unambiguous ``SUMPRODUCT`` range mismatch.

    The formula helper accepts only native ``SUMPRODUCT`` calls (optionally
    preceded by ``@``) with at least two direct, bounded internal A1 range
    arguments. This layer adds workbook context: ordinary inspectable formula
    cells only, no array territory, and no explicit broken-reference operand.
    Findings retain only aggregate counts, never formula text, range spellings,
    or source sheet identity.
    """
    candidates: list[tuple[CellKey, int, int]] = []
    for location in sorted(snapshot.cells, key=_location_sort_key):
        if (
            location in snapshot.broken_references
            or not _eligible_formula(snapshot, location, array_ranges_by_sheet)
        ):
            continue
        formula = snapshot.cells[location].formula
        if formula is None:
            continue
        mismatches = sumproduct_range_shape_mismatches(formula)
        if not mismatches:
            continue
        if (
            existing_finding_count + len(candidates)
            >= max_formula_pattern_findings
        ):
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        candidates.append(
            (
                location,
                len(mismatches),
                sum(mismatches),
            )
        )
    return tuple(candidates)


def _mmult_dimension_mismatch_candidates(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
    *,
    existing_finding_count: int,
    max_formula_pattern_findings: int,
) -> tuple[tuple[CellKey, int], ...]:
    """Return formula cells with an unambiguous ``MMULT`` dimension mismatch.

    The formula helper accepts only native ``MMULT`` calls (optionally preceded
    by ``@``) with exactly two direct, bounded internal A1 range arguments.
    This layer adds workbook context: ordinary inspectable formula cells only,
    no array territory, and no explicit broken-reference operand. Findings
    retain only an aggregate count, never formula text, range spellings, or
    source sheet identity.
    """
    candidates: list[tuple[CellKey, int]] = []
    for location in sorted(snapshot.cells, key=_location_sort_key):
        if (
            location in snapshot.broken_references
            or not _eligible_formula(snapshot, location, array_ranges_by_sheet)
        ):
            continue
        formula = snapshot.cells[location].formula
        if formula is None:
            continue
        mismatch_count = mmult_dimension_mismatch_count(formula)
        if not mismatch_count:
            continue
        if (
            existing_finding_count + len(candidates)
            >= max_formula_pattern_findings
        ):
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        candidates.append((location, mismatch_count))
    return tuple(candidates)


def _lookup_return_index_candidates(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
    *,
    existing_finding_count: int,
    max_formula_pattern_findings: int,
) -> tuple[tuple[CellKey, int], ...]:
    """Return cells with a provable static legacy-lookup return-index error.

    The formula helper accepts only native ``VLOOKUP`` and ``HLOOKUP`` calls
    (optionally preceded by ``@``), a direct bounded internal A1 table range,
    and a direct positive integer return-index literal. This layer adds
    workbook context: ordinary inspectable formula cells only, no array
    territory, and no explicit broken-reference operand. Findings retain only
    an aggregate count, never formula text, range spellings, or source sheet
    identity.
    """
    candidates: list[tuple[CellKey, int]] = []
    for location in sorted(snapshot.cells, key=_location_sort_key):
        if (
            location in snapshot.broken_references
            or not _eligible_formula(snapshot, location, array_ranges_by_sheet)
        ):
            continue
        formula = snapshot.cells[location].formula
        if formula is None:
            continue
        mismatches = lookup_return_index_mismatches(formula)
        if not mismatches:
            continue
        if (
            existing_finding_count + len(candidates)
            >= max_formula_pattern_findings
        ):
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        candidates.append((location, len(mismatches)))
    return tuple(candidates)


def _choose_literal_index_candidates(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
    *,
    existing_finding_count: int,
    max_formula_pattern_findings: int,
) -> tuple[tuple[CellKey, int], ...]:
    """Return cells with a provable static ``CHOOSE`` literal-index error.

    The formula helper accepts only native ``CHOOSE`` calls (optionally
    preceded by ``@``) with a direct nonnegative decimal index and one through
    254 nonempty value arguments. This layer adds workbook context: ordinary
    inspectable formula cells only, no array territory, and no explicit
    broken-reference operand. Findings retain only aggregate counts, never
    formula text, value arguments, or source sheet identity.
    """
    candidates: list[tuple[CellKey, int]] = []
    for location in sorted(snapshot.cells, key=_location_sort_key):
        if (
            location in snapshot.broken_references
            or not _eligible_formula(snapshot, location, array_ranges_by_sheet)
        ):
            continue
        formula = snapshot.cells[location].formula
        if formula is None:
            continue
        mismatch_count = choose_literal_index_mismatch_count(formula)
        if not mismatch_count:
            continue
        if (
            existing_finding_count + len(candidates)
            >= max_formula_pattern_findings
        ):
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        candidates.append((location, mismatch_count))
    return tuple(candidates)


def _randbetween_literal_bound_candidates(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
    *,
    existing_finding_count: int,
    max_formula_pattern_findings: int,
) -> tuple[tuple[CellKey, int], ...]:
    """Return cells with provably inverted literal ``RANDBETWEEN`` bounds.

    The formula helper accepts only native ``RANDBETWEEN`` calls (optionally
    preceded by ``@``), exactly two arguments, and direct signed integer
    literals for both bounds. This layer adds workbook context: ordinary
    inspectable formula cells only, no array territory, and no explicit
    broken-reference operand. Findings retain only an aggregate count, never
    formula text, literal values, or source sheet identity.
    """
    candidates: list[tuple[CellKey, int]] = []
    for location in sorted(snapshot.cells, key=_location_sort_key):
        if (
            location in snapshot.broken_references
            or not _eligible_formula(snapshot, location, array_ranges_by_sheet)
        ):
            continue
        formula = snapshot.cells[location].formula
        if formula is None:
            continue
        mismatch_count = randbetween_literal_bound_mismatch_count(formula)
        if not mismatch_count:
            continue
        if (
            existing_finding_count + len(candidates)
            >= max_formula_pattern_findings
        ):
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        candidates.append((location, mismatch_count))
    return tuple(candidates)


def _range_intersects_array_territory(
    snapshot: WorkbookSnapshot,
    *,
    sheet: str,
    min_column: int,
    min_row: int,
    max_column: int,
    max_row: int,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
) -> bool:
    """Return whether a compact local range overlaps any declared array territory."""
    if any(
        min_column <= array_range.max_column
        and array_range.min_column <= max_column
        and min_row <= array_range.max_row
        and array_range.min_row <= max_row
        for array_range in array_ranges_by_sheet.get(sheet, ())
    ):
        return True
    for locations in (
        snapshot.dynamic_array_formula_cells,
        snapshot.unclassified_array_formula_cells,
    ):
        for location in locations:
            if location[0] != sheet:
                continue
            row, column = coordinate_to_tuple(location[1])
            if min_column <= column <= max_column and min_row <= row <= max_row:
                return True
    return False


def _is_numeric_literal(cell: object) -> bool:
    """Return whether one stored ordinary cell is a numeric aggregate input."""
    return bool(
        getattr(cell, "cell_type", None) == "value"
        and isinstance(getattr(cell, "value", None), Real)
        and not isinstance(getattr(cell, "value", None), bool)
    )


def _aggregate_omission_candidate(
    snapshot: WorkbookSnapshot,
    location: CellKey,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
    *,
    max_gap_cells: int,
) -> _AggregateOmissionCandidate | None:
    """Find one deliberately narrow numeric run omitted by a local aggregate.

    Excel's own error checker warns when an aggregate ends before immediately
    adjacent values.  This portable variant accepts only a pure single-range
    ``SUM``, ``AVERAGE``, ``MIN``, ``MAX``, or ``COUNT`` over a one-dimensional
    local range.  The candidate must sit on the same row or column after that
    range, with a short, fully populated run of numeric literals in between.
    It does not infer an intended formula or evaluate values.
    """
    cell = snapshot.cells[location]
    if not isinstance(cell.formula, str):  # narrowed by _eligible_formula
        return None
    match = _SIMPLE_NUMERIC_AGGREGATE.fullmatch(cell.formula)
    if match is None:
        return None
    reference = parse_reference_token(match.group("reference"))
    if (
        reference is None
        or reference.is_external
        or reference.min_column is None
        or reference.min_row is None
        or reference.max_column is None
        or reference.max_row is None
        or not reference.is_range
        or (
            reference.sheet is not None
            and reference.sheet.casefold() != location[0].casefold()
        )
    ):
        return None
    if (reference.min_column == reference.max_column) == (
        reference.min_row == reference.max_row
    ):
        return None

    sheet, coordinate = location
    formula_row, formula_column = coordinate_to_tuple(coordinate)
    if (
        reference.min_column == reference.max_column
        and formula_column == reference.min_column
        and formula_row > reference.max_row + 1
    ):
        orientation = "column"
        omitted_min_column = omitted_max_column = reference.min_column
        omitted_min_row = reference.max_row + 1
        omitted_max_row = formula_row - 1
    elif (
        reference.min_row == reference.max_row
        and formula_row == reference.min_row
        and formula_column > reference.max_column + 1
    ):
        orientation = "row"
        omitted_min_column = reference.max_column + 1
        omitted_max_column = formula_column - 1
        omitted_min_row = omitted_max_row = reference.min_row
    else:
        return None

    omitted_cell_count = (
        (omitted_max_column - omitted_min_column + 1)
        * (omitted_max_row - omitted_min_row + 1)
    )
    if not _MIN_AGGREGATE_OMISSION_GAP_CELLS <= omitted_cell_count <= max_gap_cells:
        return None
    if _range_intersects_array_territory(
        snapshot,
        sheet=sheet,
        min_column=reference.min_column,
        min_row=reference.min_row,
        max_column=reference.max_column,
        max_row=reference.max_row,
        array_ranges_by_sheet=array_ranges_by_sheet,
    ) or _range_intersects_array_territory(
        snapshot,
        sheet=sheet,
        min_column=omitted_min_column,
        min_row=omitted_min_row,
        max_column=omitted_max_column,
        max_row=omitted_max_row,
        array_ranges_by_sheet=array_ranges_by_sheet,
    ):
        return None

    if orientation == "column":
        omitted_locations = (
            (sheet, f"{get_column_letter(omitted_min_column)}{row}")
            for row in range(omitted_min_row, omitted_max_row + 1)
        )
    else:
        omitted_locations = (
            (sheet, f"{get_column_letter(column)}{omitted_min_row}")
            for column in range(omitted_min_column, omitted_max_column + 1)
        )
    if not all(
        (omitted := snapshot.cells.get(omitted_location)) is not None
        and _is_numeric_literal(omitted)
        for omitted_location in omitted_locations
    ):
        return None

    referenced_start = f"{get_column_letter(reference.min_column)}{reference.min_row}"
    referenced_end = f"{get_column_letter(reference.max_column)}{reference.max_row}"
    omitted_start = f"{get_column_letter(omitted_min_column)}{omitted_min_row}"
    omitted_end = f"{get_column_letter(omitted_max_column)}{omitted_max_row}"
    return _AggregateOmissionCandidate(
        function=match.group("function").upper(),
        orientation=orientation,
        referenced_range=referenced_start
        if referenced_start == referenced_end
        else f"{referenced_start}:{referenced_end}",
        omitted_range=(
            omitted_start
            if omitted_start == omitted_end
            else f"{omitted_start}:{omitted_end}"
        ),
        omitted_cell_count=omitted_cell_count,
    )


def _direct_unlocked_formula_locations(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
) -> tuple[CellKey, ...]:
    """Return ordinary formulas explicitly unlocked on an active worksheet.

    Excel's error checker warns about unlocked formula cells because they can be
    overwritten.  FormulaFence accepts only direct cell protection assignments
    on an actively protected worksheet.  It deliberately does not try to infer
    an effective style from row, column, default, or allowed-edit-range state:
    those broader cases need a complete precedence model before they are safe
    to present as an actionable review finding.
    """
    protected_sheets = {
        protection.sheet
        for protection in snapshot.sheet_protections
        if protection.enabled and protection.sheet_type == "worksheet"
    }
    locations: list[CellKey] = []
    for assignment in sorted(
        snapshot.cell_protection_assignments,
        key=lambda item: item.sort_key(),
    ):
        if (
            assignment.scope != "cell"
            or assignment.locked
            or assignment.sheet not in protected_sheets
        ):
            continue
        location = (assignment.sheet, assignment.target)
        if _eligible_formula(snapshot, location, array_ranges_by_sheet):
            locations.append(location)
    return tuple(locations)


def _has_incomplete_manual_calculation(snapshot: WorkbookSnapshot) -> bool:
    """Return whether stored formula results may be stale by explicit metadata.

    ``calcCompleted=false`` says that the workbook was not recalculated before
    it was saved.  FormulaFence reports that state only when the workbook also
    explicitly requests manual calculation, where Excel requires a user to
    request recalculation.  It is a configuration risk, not a claim that any
    particular formula result is mathematically wrong.
    """
    return bool(
        any(cell.is_formula for cell in snapshot.cells.values())
        and snapshot.calculation_settings.get("calcMode") == "manual"
        and snapshot.calculation_settings.get("calcCompleted") is False
    )


def _ignored_error_suppression_details(
    snapshot: WorkbookSnapshot,
) -> dict[str, object] | None:
    """Return safe evidence for recognized Excel error-checking suppressions.

    FormulaFence's raw OOXML reader retains target ranges only in a private
    signature because they can identify sensitive model locations. This lint
    reports only recognized warning categories and their aggregate counts; it
    does not decide whether a suppressed Excel prompt was correct or evaluate
    a formula.
    """
    controls = snapshot.ignored_error_controls
    suppressed_warning_counts = {
        name: count
        for name, field in _IGNORED_ERROR_SUPPRESSION_FIELDS
        if (count := getattr(controls, field))
    }
    if not suppressed_warning_counts:
        return None
    return {
        "suppressed_warning_counts": suppressed_warning_counts,
        "suppressed_warning_rule_count": controls.ignored_error_rule_count,
        "suppressed_warning_target_range_count": controls.target_range_count,
    }


def _direct_self_reference_locations(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
) -> tuple[CellKey, ...]:
    """Return exact static self references when OOXML iteration is disabled.

    The workbook ``iterate`` calculation property defaults to false.  A direct
    scalar dependency back to its own ordinary formula cell is therefore a
    circular reference Excel will not attempt to calculate by default.  This
    deliberately does not try to infer indirect cycles or interpret a range,
    dynamic reference, spill, or explicit-intersection expression: each needs
    more evaluation semantics than this lint promises.
    """
    if snapshot.calculation_settings.get("iterate") is True:
        return ()

    locations: list[CellKey] = []
    for location in sorted(snapshot.cells, key=_location_sort_key):
        if not _eligible_formula(snapshot, location, array_ranges_by_sheet):
            continue
        if (
            location in snapshot.dynamic_reference_functions
            or location in snapshot.spill_reference_tokens
            or location in snapshot.implicit_intersection_tokens
        ):
            continue
        if location in snapshot.reverse_dependencies.get(location, ()):
            locations.append(location)
    return tuple(locations)


def _multi_cell_static_cycle_locations(
    snapshot: WorkbookSnapshot,
    array_ranges_by_sheet: dict[str, tuple[ArrayFormulaRange, ...]],
    direct_self_reference_locations: set[CellKey],
    *,
    max_formula_pattern_findings: int,
    existing_finding_count: int,
) -> tuple[tuple[CellKey, int], ...]:
    """Return ordinary formula members of proven multi-cell static cycles.

    ``reverse_dependencies`` retains only resolved scalar A1 edges; ranges are
    held in a separate index. Strongly connected components are therefore an
    exact static circular-reference fact for the ordinary formulas represented
    here, not an attempted calculation. The stored edge direction is the
    reverse of formula evaluation, but reversing every edge leaves strongly
    connected components unchanged.

    The iterative Tarjan traversal deliberately reuses the existing graph
    rather than materializing an inverse or expanded range graph. Dynamic-
    reference, spill, explicit-intersection, three-dimensional, array, and
    tokenizer-failure territory stays outside this narrow signal. Direct self-reference
    members remain under FF087 and are omitted here to avoid duplicate prompts.
    """
    if snapshot.calculation_settings.get("iterate") is True:
        return ()

    locations = tuple(
        location
        for location in sorted(snapshot.cells, key=_location_sort_key)
        if (
            _eligible_formula(snapshot, location, array_ranges_by_sheet)
            and location not in snapshot.dynamic_reference_functions
            and location not in snapshot.three_d_reference_tokens
            and location not in snapshot.spill_reference_tokens
            and location not in snapshot.implicit_intersection_tokens
        )
    )
    if len(locations) < 2:
        return ()

    location_indexes = {location: index for index, location in enumerate(locations)}
    discovery_indexes = [-1] * len(locations)
    low_links = [0] * len(locations)
    on_component_stack = bytearray(len(locations))
    component_stack: list[int] = []
    findings: list[tuple[CellKey, int]] = []
    next_discovery_index = 0

    # ``reverse_dependencies`` maps a scalar source cell to its formula
    # dependents. Cyclic components are invariant under that reversal, so it
    # can serve as Tarjan's outgoing adjacency without allocating another graph.
    for root_index, root_location in enumerate(locations):
        if discovery_indexes[root_index] != -1:
            continue
        discovery_indexes[root_index] = next_discovery_index
        low_links[root_index] = next_discovery_index
        next_discovery_index += 1
        component_stack.append(root_index)
        on_component_stack[root_index] = 1
        traversal: list[tuple[int, Iterator[CellKey]]] = [
            (root_index, iter(snapshot.reverse_dependencies.get(root_location, ())))
        ]

        while traversal:
            current_index, dependents = traversal[-1]
            try:
                dependent_location = next(dependents)
            except StopIteration:
                traversal.pop()
                if traversal:
                    parent_index = traversal[-1][0]
                    low_links[parent_index] = min(
                        low_links[parent_index], low_links[current_index]
                    )
                if low_links[current_index] != discovery_indexes[current_index]:
                    continue

                component: list[int] = []
                while component_stack:
                    member_index = component_stack.pop()
                    on_component_stack[member_index] = 0
                    component.append(member_index)
                    if member_index == current_index:
                        break
                if len(component) < 2:
                    continue
                for member_index in component:
                    location = locations[member_index]
                    if location in direct_self_reference_locations:
                        continue
                    if (
                        existing_finding_count + len(findings)
                        >= max_formula_pattern_findings
                    ):
                        raise FormulaFenceError(
                            "Formula lint exceeds "
                            "max_formula_pattern_findings="
                            f"{max_formula_pattern_findings}."
                        ) from None
                    findings.append((location, len(component)))
                continue

            dependent_index = location_indexes.get(dependent_location)
            if dependent_index is None:
                continue
            if discovery_indexes[dependent_index] == -1:
                discovery_indexes[dependent_index] = next_discovery_index
                low_links[dependent_index] = next_discovery_index
                next_discovery_index += 1
                component_stack.append(dependent_index)
                on_component_stack[dependent_index] = 1
                traversal.append(
                    (
                        dependent_index,
                        iter(
                            snapshot.reverse_dependencies.get(
                                dependent_location,
                                (),
                            )
                        ),
                    )
                )
            elif on_component_stack[dependent_index]:
                low_links[current_index] = min(
                    low_links[current_index], discovery_indexes[dependent_index]
                )

    return tuple(
        sorted(findings, key=lambda finding: _location_sort_key(finding[0]))
    )


def _explicit_broken_reference_locations(snapshot: WorkbookSnapshot) -> tuple[CellKey, ...]:
    """Return formula locations whose tokenizer exposed a ``#REF!`` operand.

    Workbook loading records this ledger only for actual error operands, not
    for a matching string literal or quoted worksheet name.  The lint keeps
    that exact lexical fact separate from a cached formula result and does not
    evaluate the formula.
    """
    return tuple(sorted(snapshot.broken_references, key=_location_sort_key))


def _saved_broken_reference_result_locations(
    snapshot: WorkbookSnapshot,
) -> tuple[CellKey, ...]:
    """Return locations with a valid saved ``#REF!`` formula result.

    SpreadsheetML retains a formula's last calculated value beside its formula.
    The private cache reader recognizes an exact broken-reference error without
    keeping the error text. This lint reports that saved display state; it does
    not recalculate the formula or infer that the current result is unchanged.
    """
    return tuple(
        sorted(
            {
                entry.location
                for entry in snapshot.formula_cached_results.entries
                if entry.is_broken_reference_error
            },
            key=_location_sort_key,
        )
    )


def _candidate_message(kind: str) -> tuple[str, str, str]:
    """Return the rule, severity, and reviewer-facing message for a pattern kind."""
    if kind == "blank_gap":
        return (
            "FF082",
            "high",
            "A blank cell interrupts a stable copied-formula pattern.",
        )
    if kind == "error_gap":
        return (
            "FF082",
            "high",
            "A stored error cell interrupts a stable copied-formula pattern.",
        )
    if kind == "text_gap":
        return (
            "FF082",
            "low",
            "A text value interrupts a stable copied-formula pattern.",
        )
    if kind == "non_formula_gap":
        return (
            "FF082",
            "medium",
            "A non-formula value interrupts a stable copied-formula pattern.",
        )
    return (
        "FF083",
        "medium",
        "Formula differs from a stable copied-formula pattern.",
    )


def lint_snapshot(
    snapshot: WorkbookSnapshot,
    *,
    max_formula_pattern_findings: int = DEFAULT_MAX_FORMULA_PATTERN_FINDINGS,
    max_aggregate_omission_gap_cells: int = DEFAULT_MAX_AGGREGATE_OMISSION_GAP_CELLS,
) -> FormulaLintReport:
    """Find conservative formula patterns, protection, and calculation risks.

    A target is reported only when its immediate predecessor and successor have
    the same relative formula fingerprint *and* a third contiguous peer repeats
    that fingerprint. This intentionally misses short or ambiguous sequences
    rather than treating a general spreadsheet smell as an error. Separately,
    a pure local numeric aggregate is reported only when it stops before a
    short, contiguous numeric run on the same row or column. It also reports
    an explicitly unlocked formula on a protected sheet, an explicit incomplete
    manual-calculation state for a formula workbook, stored error-checking
    suppressions, isolated interior Excel Table calculated-column exceptions,
    direct static conditional-aggregate and ``SUMPRODUCT`` range-shape
    mismatches, direct static ``MMULT`` matrix-dimension mismatches, direct
    static legacy-lookup return-index mismatches, direct static ``CHOOSE``
    literal-index mismatches, direct literal ``RANDBETWEEN`` bound mismatches,
    direct and multi-cell static circular references while iteration is
    disabled, an explicit broken reference operand, and a saved
    broken-reference result. It never evaluates formulas, and rejects
    incomplete array metadata before claiming ordinary-cell coverage.
    """
    if max_formula_pattern_findings < 1:
        raise FormulaFenceError("max_formula_pattern_findings must be at least 1.")
    if max_aggregate_omission_gap_cells < _MIN_AGGREGATE_OMISSION_GAP_CELLS:
        raise FormulaFenceError(
            "max_aggregate_omission_gap_cells must be at least "
            f"{_MIN_AGGREGATE_OMISSION_GAP_CELLS}."
        )
    if not snapshot.array_formula_metadata_complete:
        raise FormulaFenceError(
            "Formula lint requires complete array-formula metadata."
        )

    array_ranges_by_sheet = _array_ranges_by_sheet(snapshot)
    conditional_aggregate_candidates = _conditional_aggregate_range_shape_candidates(
        snapshot,
        array_ranges_by_sheet,
        max_formula_pattern_findings=max_formula_pattern_findings,
    )
    sumproduct_candidates = _sumproduct_range_shape_candidates(
        snapshot,
        array_ranges_by_sheet,
        existing_finding_count=len(conditional_aggregate_candidates),
        max_formula_pattern_findings=max_formula_pattern_findings,
    )
    mmult_candidates = _mmult_dimension_mismatch_candidates(
        snapshot,
        array_ranges_by_sheet,
        existing_finding_count=(
            len(conditional_aggregate_candidates) + len(sumproduct_candidates)
        ),
        max_formula_pattern_findings=max_formula_pattern_findings,
    )
    lookup_candidates = _lookup_return_index_candidates(
        snapshot,
        array_ranges_by_sheet,
        existing_finding_count=(
            len(conditional_aggregate_candidates)
            + len(sumproduct_candidates)
            + len(mmult_candidates)
        ),
        max_formula_pattern_findings=max_formula_pattern_findings,
    )
    choose_candidates = _choose_literal_index_candidates(
        snapshot,
        array_ranges_by_sheet,
        existing_finding_count=(
            len(conditional_aggregate_candidates)
            + len(sumproduct_candidates)
            + len(mmult_candidates)
            + len(lookup_candidates)
        ),
        max_formula_pattern_findings=max_formula_pattern_findings,
    )
    randbetween_candidates = _randbetween_literal_bound_candidates(
        snapshot,
        array_ranges_by_sheet,
        existing_finding_count=(
            len(conditional_aggregate_candidates)
            + len(sumproduct_candidates)
            + len(mmult_candidates)
            + len(lookup_candidates)
            + len(choose_candidates)
        ),
        max_formula_pattern_findings=max_formula_pattern_findings,
    )
    structural_formula_locations = {
        location
        for location, _, _ in conditional_aggregate_candidates
    }
    structural_formula_locations.update(
        location for location, _, _ in sumproduct_candidates
    )
    structural_formula_locations.update(location for location, _ in mmult_candidates)
    structural_formula_locations.update(location for location, _ in lookup_candidates)
    structural_formula_locations.update(location for location, _ in choose_candidates)
    structural_formula_locations.update(
        location for location, _ in randbetween_candidates
    )
    candidates: dict[CellKey, _FormulaPatternCandidate] = {}

    for preceding_location in sorted(snapshot.cells, key=_location_sort_key):
        if not _eligible_formula(snapshot, preceding_location, array_ranges_by_sheet):
            continue
        preceding = snapshot.cells[preceding_location]
        fingerprint = preceding.formula_fingerprint
        assert fingerprint is not None  # narrowed by _eligible_formula

        for orientation, row_delta, column_delta in _PATTERN_ORIENTATIONS:
            target_location = _offset_location(
                preceding_location,
                row_delta,
                column_delta,
            )
            following_location = _offset_location(
                preceding_location,
                row_delta * 2,
                column_delta * 2,
            )
            if target_location is None or following_location is None:
                continue
            if not _eligible_formula(snapshot, following_location, array_ranges_by_sheet):
                continue
            following = snapshot.cells[following_location]
            if following.formula_fingerprint != fingerprint:
                continue
            if _is_array_member(snapshot, target_location, array_ranges_by_sheet):
                continue
            if target_location in structural_formula_locations:
                continue

            target = snapshot.cells.get(target_location)
            if target is None:
                kind = "blank_gap"
            elif target.is_formula:
                if not _eligible_formula(snapshot, target_location, array_ranges_by_sheet):
                    continue
                if target.formula_fingerprint == fingerprint:
                    continue
                kind = "formula_outlier"
            else:
                kind = (
                    "error_gap"
                    if target.cell_type == "error"
                    else "text_gap"
                    if isinstance(target.value, str)
                    else "non_formula_gap"
                )

            support_before = _offset_location(
                preceding_location,
                -row_delta,
                -column_delta,
            )
            support_after = _offset_location(
                following_location,
                row_delta,
                column_delta,
            )
            supporting_location = next(
                (
                    location
                    for location in (support_before, support_after)
                    if location is not None
                    and _eligible_formula(snapshot, location, array_ranges_by_sheet)
                    and snapshot.cells[location].formula_fingerprint == fingerprint
                ),
                None,
            )
            if supporting_location is None:
                continue

            candidate = candidates.get(target_location)
            if candidate is None:
                if (
                    len(candidates)
                    + len(conditional_aggregate_candidates)
                    + len(sumproduct_candidates)
                    + len(mmult_candidates)
                    + len(lookup_candidates)
                    + len(choose_candidates)
                    + len(randbetween_candidates)
                    >= max_formula_pattern_findings
                ):
                    raise FormulaFenceError(
                        "Formula lint exceeds "
                        f"max_formula_pattern_findings={max_formula_pattern_findings}."
                    )
                candidate = _FormulaPatternCandidate(kind=kind)
                candidates[target_location] = candidate
            elif candidate.kind != kind:  # pragma: no cover - one target has one cell kind
                continue
            candidate.evidence.append(
                _FormulaPatternEvidence(
                    orientation=orientation,
                    preceding=preceding_location,
                    following=following_location,
                    supporting=supporting_location,
                )
            )

    formula_pattern_locations = set(candidates)
    findings: list[Finding] = []
    for location in sorted(candidates, key=_location_sort_key):
        candidate = candidates[location]
        rule_id, severity, message = _candidate_message(candidate.kind)
        findings.append(
            Finding(
                rule_id=rule_id,
                severity=severity,
                message=message,
                location=location,
                details={
                    "pattern_kind": candidate.kind,
                    "pattern_evidence": [
                        {
                            "orientation": evidence.orientation,
                            "preceding_formula": display_location(evidence.preceding),
                            "following_formula": display_location(evidence.following),
                            "supporting_formula": display_location(evidence.supporting),
                        }
                        for evidence in candidate.evidence
                    ],
                },
            )
        )

    for (
        location,
        conditional_aggregate_call_count,
        mismatched_direct_range_argument_count,
    ) in conditional_aggregate_candidates:
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF093",
                severity="high",
                message=(
                    "A conditional aggregate uses direct static ranges with different "
                    "shapes."
                ),
                location=location,
                details={
                    "conditional_aggregate_call_count": (
                        conditional_aggregate_call_count
                    ),
                    "mismatched_direct_range_argument_count": (
                        mismatched_direct_range_argument_count
                    ),
                    "evidence_scope": "conditional_aggregate_direct_a1_ranges",
                },
            )
        )

    for (
        location,
        sumproduct_call_count,
        mismatched_direct_array_argument_count,
    ) in sumproduct_candidates:
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF094",
                severity="high",
                message="A SUMPRODUCT call uses direct static ranges with different shapes.",
                location=location,
                details={
                    "sumproduct_call_count": sumproduct_call_count,
                    "mismatched_direct_array_argument_count": (
                        mismatched_direct_array_argument_count
                    ),
                    "evidence_scope": "sumproduct_direct_a1_ranges",
                },
            )
        )

    for location, mmult_call_count in mmult_candidates:
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF095",
                severity="high",
                message=(
                    "An MMULT call uses direct static arrays with incompatible "
                    "matrix dimensions."
                ),
                location=location,
                details={
                    "mmult_call_count": mmult_call_count,
                    "incompatible_direct_matrix_pair_count": mmult_call_count,
                    "evidence_scope": "mmult_direct_a1_arrays",
                },
            )
        )

    for location, lookup_call_count in lookup_candidates:
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF096",
                severity="high",
                message=(
                    "A VLOOKUP or HLOOKUP call uses a literal return index outside "
                    "its direct static table range."
                ),
                location=location,
                details={
                    "lookup_call_count": lookup_call_count,
                    "out_of_range_literal_index_count": lookup_call_count,
                    "evidence_scope": "lookup_direct_a1_table_literal_index",
                },
            )
        )

    for location, choose_call_count in choose_candidates:
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF097",
                severity="high",
                message=(
                    "A CHOOSE call uses a literal index outside its available "
                    "value arguments."
                ),
                location=location,
                details={
                    "choose_call_count": choose_call_count,
                    "out_of_range_literal_index_count": choose_call_count,
                    "evidence_scope": "choose_literal_index_value_arity",
                },
            )
        )

    for location, randbetween_call_count in randbetween_candidates:
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF098",
                severity="high",
                message=(
                    "A RANDBETWEEN call uses direct literal bounds with the bottom "
                    "above the top."
                ),
                location=location,
                details={
                    "randbetween_call_count": randbetween_call_count,
                    "inverted_literal_bound_count": randbetween_call_count,
                    "evidence_scope": "randbetween_direct_signed_integer_bounds",
                },
            )
        )

    for location in sorted(snapshot.cells, key=_location_sort_key):
        if not _eligible_formula(snapshot, location, array_ranges_by_sheet):
            continue
        candidate = _aggregate_omission_candidate(
            snapshot,
            location,
            array_ranges_by_sheet,
            max_gap_cells=max_aggregate_omission_gap_cells,
        )
        if candidate is None:
            continue
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF084",
                severity="medium",
                message=(
                    "A simple numeric aggregate stops before contiguous adjacent "
                    "numeric cells."
                ),
                location=location,
                details={
                    "aggregate_function": candidate.function,
                    "orientation": candidate.orientation,
                    "referenced_range": display_location(
                        (location[0], candidate.referenced_range)
                    ),
                    "omitted_range": display_location((location[0], candidate.omitted_range)),
                    "omitted_cell_count": candidate.omitted_cell_count,
                },
            )
        )
    for location in _direct_unlocked_formula_locations(
        snapshot,
        array_ranges_by_sheet,
    ):
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF085",
                severity="medium",
                message=(
                    "A formula cell is explicitly unlocked on a protected worksheet."
                ),
                location=location,
                details={"protection_scope": "direct_cell"},
            )
        )
    if _has_incomplete_manual_calculation(snapshot):
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF086",
                severity="medium",
                message=(
                    "Workbook contains formulas and was saved with incomplete "
                    "manual calculation."
                ),
                details={
                    "calculation_mode": "manual",
                    "calculation_completed_before_save": False,
                },
            )
        )
    if details := _ignored_error_suppression_details(snapshot):
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF091",
                severity="medium",
                message=(
                    "Workbook suppresses Excel error-checking prompts; review warnings "
                    "may be hidden."
                ),
                details=details,
            )
        )
    for location, exception_kind in _table_calculated_column_exception_candidates(
        snapshot,
        array_ranges_by_sheet,
        formula_pattern_locations=formula_pattern_locations,
        max_formula_pattern_findings=max_formula_pattern_findings,
        existing_finding_count=len(findings),
    ):
        findings.append(
            Finding(
                rule_id="FF092",
                severity="medium",
                message=(
                    "An interior Excel Table cell differs from its declared "
                    "calculated-column formula."
                ),
                location=location,
                details={
                    "exception_kind": exception_kind,
                    "matching_adjacent_formula_peers": 2,
                    "evidence_scope": "table_calculated_column",
                },
            )
        )
    direct_self_reference_locations = _direct_self_reference_locations(
        snapshot,
        array_ranges_by_sheet,
    )
    for location in direct_self_reference_locations:
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF087",
                severity="high",
                message=(
                    "A formula directly references its own cell while calculation "
                    "iteration is disabled."
                ),
                location=location,
                details={
                    "calculation_iteration_enabled": False,
                    "reference_scope": "direct_static",
                },
            )
        )
    for location, cycle_member_count in _multi_cell_static_cycle_locations(
        snapshot,
        array_ranges_by_sheet,
        set(direct_self_reference_locations),
        max_formula_pattern_findings=max_formula_pattern_findings,
        existing_finding_count=len(findings),
    ):
        findings.append(
            Finding(
                rule_id="FF090",
                severity="high",
                message=(
                    "A formula participates in a static multi-cell circular reference "
                    "while calculation iteration is disabled."
                ),
                location=location,
                details={
                    "calculation_iteration_enabled": False,
                    "reference_scope": "multi_cell_static",
                    "cycle_member_count": cycle_member_count,
                },
            )
        )
    explicit_broken_reference_locations = _explicit_broken_reference_locations(snapshot)
    for location in explicit_broken_reference_locations:
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF088",
                severity="critical",
                message="Formula contains an explicit broken #REF! reference.",
                location=location,
            )
        )
    explicit_broken_reference_location_set = set(explicit_broken_reference_locations)
    for location in _saved_broken_reference_result_locations(snapshot):
        if location in explicit_broken_reference_location_set:
            continue
        if len(findings) >= max_formula_pattern_findings:
            raise FormulaFenceError(
                "Formula lint exceeds "
                f"max_formula_pattern_findings={max_formula_pattern_findings}."
            )
        findings.append(
            Finding(
                rule_id="FF089",
                severity="high",
                message="A formula's saved result is a broken-reference error.",
                location=location,
            )
        )
    findings.sort(
        key=lambda finding: (
            _location_sort_key(finding.location)
            if finding.location is not None
            else ("", 0, 0),
            finding.rule_id,
        )
    )
    return FormulaLintReport(workbook=snapshot, findings=findings)
