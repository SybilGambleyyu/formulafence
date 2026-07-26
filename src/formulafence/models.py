"""Data structures shared by FormulaFence's parser, diff, and policy layers."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import TYPE_CHECKING, Any, TypeAlias

if TYPE_CHECKING:
    from formulafence.formulas import ParsedReference

CellKey: TypeAlias = tuple[str, str]

SEVERITY_ORDER = {"note": 0, "low": 1, "medium": 2, "high": 3, "critical": 4}


@dataclass(frozen=True)
class ExternalWorkbookReference:
    """One static A1 reference to another workbook, retained for portfolio use.

    The source spelling can contain a user's local or network path, so it is
    deliberately private implementation data: portfolio renderers may expose
    only a safely resolved relative workbook identity.
    """

    source_path: str = field(repr=False)
    sheet: str
    min_column: int
    min_row: int
    max_column: int
    max_row: int

    @property
    def is_range(self) -> bool:
        return self.min_column != self.max_column or self.min_row != self.max_row


@dataclass(frozen=True)
class ExternalWorkbookThreeDReference:
    """One static A1 reference spanning worksheets in another workbook.

    The source spelling and both endpoint titles can disclose an author's
    private model layout, so all three remain private implementation data.
    Portfolio analysis may expand the span only after it has bound the source
    to an inspected candidate workbook and verified its worksheet tab order.
    """

    source_path: str = field(repr=False)
    first_sheet: str = field(repr=False)
    last_sheet: str = field(repr=False)
    min_column: int
    min_row: int
    max_column: int
    max_row: int

    @property
    def is_range(self) -> bool:
        return self.min_column != self.max_column or self.min_row != self.max_row


@dataclass(frozen=True)
class ExternalWorkbookDefinedNameReference:
    """One direct external defined-name reference to another workbook.

    Both the external source spelling and the source defined-name identity can
    disclose a model's private layout, so they remain private implementation
    data. ``scope_sheet`` distinguishes the Excel sheet-local form from a
    workbook-scoped name; it too remains private. Portfolio analysis may expose
    only a resolved candidate workbook and logical cell locations that it
    already inspected.
    """

    source_path: str = field(repr=False)
    name_key: str = field(repr=False)
    scope_sheet: str | None = field(default=None, repr=False)


def display_location(location: CellKey | None) -> str | None:
    """Return an Excel-friendly cell location."""
    if location is None:
        return None
    sheet, coordinate = location
    escaped_sheet = sheet.replace("'", "''")
    if any(character in sheet for character in " !'[]"):
        escaped_sheet = f"'{escaped_sheet}'"
    return f"{escaped_sheet}!{coordinate}"


def json_safe_value(value: Any) -> Any:
    """Convert an openpyxl value to a deterministic JSON-compatible representation."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


@dataclass(frozen=True)
class CellSnapshot:
    """The semantic content FormulaFence compares for one non-empty cell."""

    sheet: str
    coordinate: str
    cell_type: str
    value: Any
    value_type: str
    formula: str | None = None
    formula_fingerprint: str | None = None
    array_formula_kind: str | None = None
    array_formula_ref: str | None = None

    @property
    def location(self) -> CellKey:
        return (self.sheet, self.coordinate)

    @property
    def is_formula(self) -> bool:
        return self.cell_type == "formula"

    def to_dict(self) -> dict[str, Any]:
        return {
            "location": display_location(self.location),
            "cell_type": self.cell_type,
            "value": self.value,
            "value_type": self.value_type,
            "formula": self.formula,
            "formula_fingerprint": self.formula_fingerprint,
            "array_formula_kind": self.array_formula_kind,
            "array_formula_ref": self.array_formula_ref,
        }


@dataclass(frozen=True)
class SheetSnapshot:
    title: str
    state: str
    nonempty_cells: int
    formula_cells: int
    max_row: int
    max_column: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class TableSnapshot:
    """Inspectable Excel-table metadata that can change formula semantics."""

    name: str
    sheet: str
    ref: str
    columns: tuple[str, ...]
    header_row_count: int
    totals_row_count: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "sheet": self.sheet,
            "ref": self.ref,
            "columns": list(self.columns),
            "header_row_count": self.header_row_count,
            "totals_row_count": self.totals_row_count,
        }


@dataclass(frozen=True)
class DataValidationSnapshot:
    """One compact Excel data-validation control and its effective settings."""

    sheet: str
    ranges: tuple[str, ...]
    validation_type: str
    operator: str
    formula1: str | None
    formula2: str | None
    allow_blank: bool
    dropdown_hidden: bool
    prompts_disabled: bool
    show_input_message: bool
    show_error_message: bool
    error_style: str
    error_title: str | None
    error: str | None
    prompt_title: str | None
    prompt: str | None
    ime_mode: str

    @property
    def target_range_count(self) -> int:
        return len(self.ranges)

    @property
    def criteria_count(self) -> int:
        return sum(value is not None for value in (self.formula1, self.formula2))

    def sort_key(self) -> tuple[object, ...]:
        """Return a writer-order-independent key for deterministic inventories."""
        def optional(value: str | None) -> tuple[bool, str]:
            return value is not None, value or ""

        return (
            self.sheet.casefold(),
            self.ranges,
            self.validation_type,
            self.operator,
            optional(self.formula1),
            optional(self.formula2),
            self.allow_blank,
            self.dropdown_hidden,
            self.prompts_disabled,
            self.show_input_message,
            self.show_error_message,
            self.error_style,
            optional(self.error_title),
            optional(self.error),
            optional(self.prompt_title),
            optional(self.prompt),
            self.ime_mode,
        )

    def to_dict(self) -> dict[str, Any]:
        """Return full local review evidence, including criterion expressions."""
        return {
            "sheet": self.sheet,
            "ranges": [
                display_location((self.sheet, target_range))
                for target_range in self.ranges
            ],
            "type": self.validation_type,
            "operator": self.operator,
            "formula1": self.formula1,
            "formula2": self.formula2,
            "allow_blank": self.allow_blank,
            "dropdown_hidden": self.dropdown_hidden,
            "prompts_disabled": self.prompts_disabled,
            "show_input_message": self.show_input_message,
            "show_error_message": self.show_error_message,
            "error_style": self.error_style,
            "error_title": self.error_title,
            "error": self.error,
            "prompt_title": self.prompt_title,
            "prompt": self.prompt,
            "ime_mode": self.ime_mode,
        }

    def profile_dict(self) -> dict[str, Any]:
        """Return a data-minimising control inventory without messages or formulas."""
        return {
            "sheet": self.sheet,
            "ranges": [
                display_location((self.sheet, target_range))
                for target_range in self.ranges
            ],
            "type": self.validation_type,
            "operator": self.operator,
            "criteria_count": self.criteria_count,
            "allow_blank": self.allow_blank,
            "dropdown_hidden": self.dropdown_hidden,
            "prompts_disabled": self.prompts_disabled,
            "show_input_message": self.show_input_message,
            "show_error_message": self.show_error_message,
            "error_style": self.error_style,
            "has_error_alert_text": bool(self.error_title or self.error),
            "has_input_prompt_text": bool(self.prompt_title or self.prompt),
            "ime_mode": self.ime_mode,
        }


@dataclass(frozen=True)
class XmlFragmentSnapshot:
    """A deterministic, inspectable OOXML fragment used for control evidence."""

    tag: str
    attributes: tuple[tuple[str, str], ...] = ()
    text: str | None = None
    children: tuple[XmlFragmentSnapshot, ...] = ()

    def sort_key(self) -> tuple[object, ...]:
        return (
            self.tag,
            self.attributes,
            self.text is not None,
            self.text or "",
            tuple(child.sort_key() for child in self.children),
        )

    def to_dict(self) -> dict[str, Any]:
        result: dict[str, Any] = {"element": self.tag}
        if self.attributes:
            result["attributes"] = dict(self.attributes)
        if self.text is not None:
            result["text"] = self.text
        if self.children:
            result["children"] = [child.to_dict() for child in self.children]
        return result


@dataclass(frozen=True)
class ConditionalFormattingSnapshot:
    """One worksheet conditional-formatting rule and its effective semantics."""

    sheet: str
    ranges: tuple[str, ...]
    priority: int
    rule_type: str
    operator: str | None
    formulas: tuple[str, ...]
    stop_if_true: bool
    above_average: bool
    percent: bool
    bottom: bool
    rank: int | None
    std_dev: int | None
    equal_average: bool
    text: str | None
    time_period: str | None
    differential_style: XmlFragmentSnapshot | None
    color_scale: XmlFragmentSnapshot | None
    data_bar: XmlFragmentSnapshot | None
    icon_set: XmlFragmentSnapshot | None
    extensions: tuple[XmlFragmentSnapshot, ...] = ()

    @property
    def target_range_count(self) -> int:
        return len(self.ranges)

    @property
    def formula_count(self) -> int:
        return len(self.formulas)

    @property
    def formatting_kinds(self) -> tuple[str, ...]:
        kinds: list[str] = []
        if self.differential_style is not None:
            kinds.append("differential style")
        if self.color_scale is not None:
            kinds.append("color scale")
        if self.data_bar is not None:
            kinds.append("data bar")
        if self.icon_set is not None:
            kinds.append("icon set")
        return tuple(kinds)

    def sort_key(self) -> tuple[object, ...]:
        return (
            self.sheet.casefold(),
            self.priority,
            self.ranges,
            self.rule_type,
            self.operator or "",
            self.formulas,
            self.stop_if_true,
            self.above_average,
            self.percent,
            self.bottom,
            self.rank is not None,
            self.rank or 0,
            self.std_dev is not None,
            self.std_dev or 0,
            self.equal_average,
            self.text is not None,
            self.text or "",
            self.time_period or "",
            self.differential_style is not None,
            self.differential_style.sort_key() if self.differential_style is not None else (),
            self.color_scale is not None,
            self.color_scale.sort_key() if self.color_scale is not None else (),
            self.data_bar is not None,
            self.data_bar.sort_key() if self.data_bar is not None else (),
            self.icon_set is not None,
            self.icon_set.sort_key() if self.icon_set is not None else (),
            tuple(extension.sort_key() for extension in self.extensions),
        )

    def to_dict(self) -> dict[str, Any]:
        """Return full local evidence, including formulas and formatting fragments."""
        return {
            "sheet": self.sheet,
            "ranges": [
                display_location((self.sheet, target_range))
                for target_range in self.ranges
            ],
            "priority": self.priority,
            "type": self.rule_type,
            "operator": self.operator,
            "formulas": list(self.formulas),
            "stop_if_true": self.stop_if_true,
            "above_average": self.above_average,
            "percent": self.percent,
            "bottom": self.bottom,
            "rank": self.rank,
            "std_dev": self.std_dev,
            "equal_average": self.equal_average,
            "text": self.text,
            "time_period": self.time_period,
            "differential_style": (
                self.differential_style.to_dict()
                if self.differential_style is not None
                else None
            ),
            "color_scale": self.color_scale.to_dict() if self.color_scale is not None else None,
            "data_bar": self.data_bar.to_dict() if self.data_bar is not None else None,
            "icon_set": self.icon_set.to_dict() if self.icon_set is not None else None,
            "extensions": [extension.to_dict() for extension in self.extensions],
        }

    def profile_dict(self) -> dict[str, Any]:
        """Return reviewable control metadata without formulas or text criteria."""
        return {
            "sheet": self.sheet,
            "ranges": [
                display_location((self.sheet, target_range))
                for target_range in self.ranges
            ],
            "priority": self.priority,
            "type": self.rule_type,
            "operator": self.operator,
            "formula_count": self.formula_count,
            "has_text_criterion": self.text is not None,
            "stop_if_true": self.stop_if_true,
            "above_average": self.above_average,
            "percent": self.percent,
            "bottom": self.bottom,
            "rank": self.rank,
            "std_dev": self.std_dev,
            "equal_average": self.equal_average,
            "time_period": self.time_period,
            "formatting": list(self.formatting_kinds),
            "extension_count": len(self.extensions),
        }


@dataclass(frozen=True)
class ConditionalFormattingExtensionSnapshot:
    """An OOXML conditional-formatting extension openpyxl may not model."""

    sheet: str
    fragment: XmlFragmentSnapshot

    def sort_key(self) -> tuple[object, ...]:
        return self.sheet.casefold(), self.fragment.sort_key()

    def to_dict(self) -> dict[str, Any]:
        return {"sheet": self.sheet, "extension": self.fragment.to_dict()}

    def profile_dict(self) -> dict[str, Any]:
        return {"sheet": self.sheet, "element": self.fragment.tag}


@dataclass(frozen=True)
class ProtectionCredentialSnapshot:
    """Non-secret evidence that a protection credential is configured.

    OOXML stores legacy password verifiers and, in newer workbooks, password
    hashes and salts alongside the protection settings.  FormulaFence uses a
    private signature to compare that material but never serialises it into a
    profile, change report, or finding.
    """

    has_legacy_verifier: bool = False
    has_modern_verifier: bool = False
    algorithm: str | None = None
    spin_count: int | None = None
    signature: str | None = field(default=None, repr=False)

    @property
    def configured(self) -> bool:
        """Return whether any verifier material was present."""
        return self.has_legacy_verifier or self.has_modern_verifier

    def to_dict(self) -> dict[str, Any]:
        """Return safe verifier metadata without password material."""
        return {
            "configured": self.configured,
            "has_legacy_verifier": self.has_legacy_verifier,
            "has_modern_verifier": self.has_modern_verifier,
            "algorithm": self.algorithm,
            "spin_count": self.spin_count,
        }


@dataclass(frozen=True)
class ProtectionOpaqueMetadataSnapshot:
    """Private comparison evidence for unmodelled protection metadata."""

    count: int = 0
    signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return self.count > 0

    def to_dict(self) -> dict[str, Any]:
        """Return only the existence and amount of opaque metadata."""
        return {
            "present": self.present,
            "count": self.count,
        }


@dataclass(frozen=True)
class WorkbookProtectionSnapshot:
    """Workbook-structure and revision protection controls."""

    lock_structure: bool
    lock_windows: bool
    lock_revision: bool
    workbook_credential: ProtectionCredentialSnapshot = field(
        default_factory=ProtectionCredentialSnapshot
    )
    revisions_credential: ProtectionCredentialSnapshot = field(
        default_factory=ProtectionCredentialSnapshot
    )
    opaque_metadata: ProtectionOpaqueMetadataSnapshot = field(
        default_factory=ProtectionOpaqueMetadataSnapshot
    )

    @property
    def enabled(self) -> bool:
        """Return whether any workbook-level operation is locked."""
        return self.lock_structure or self.lock_windows or self.lock_revision

    def to_dict(self) -> dict[str, Any]:
        """Return safe, reviewable control settings."""
        return {
            "enabled": self.enabled,
            "lock_structure": self.lock_structure,
            "lock_windows": self.lock_windows,
            "lock_revision": self.lock_revision,
            "workbook_credential": self.workbook_credential.to_dict(),
            "revisions_credential": self.revisions_credential.to_dict(),
            "opaque_metadata": self.opaque_metadata.to_dict(),
        }


@dataclass(frozen=True)
class SheetProtectionSnapshot:
    """One worksheet, dialog-sheet, or chart-sheet protection declaration."""

    sheet: str
    sheet_type: str
    enabled: bool
    locked_actions: tuple[str, ...]
    credential: ProtectionCredentialSnapshot = field(
        default_factory=ProtectionCredentialSnapshot
    )
    opaque_metadata: ProtectionOpaqueMetadataSnapshot = field(
        default_factory=ProtectionOpaqueMetadataSnapshot
    )

    def sort_key(self) -> tuple[object, ...]:
        return self.sheet.casefold(), self.sheet_type, self.locked_actions

    def to_dict(self) -> dict[str, Any]:
        """Return safe, reviewable sheet-protection settings."""
        return {
            "sheet": self.sheet,
            "sheet_type": self.sheet_type,
            "enabled": self.enabled,
            "locked_actions": list(self.locked_actions),
            "credential": self.credential.to_dict(),
            "opaque_metadata": self.opaque_metadata.to_dict(),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class ProtectedRangeSnapshot:
    """A protected worksheet range without its password or identity material."""

    sheet: str
    ranges: tuple[str, ...]
    has_name: bool
    name_signature: str | None = field(default=None, repr=False)
    credential: ProtectionCredentialSnapshot = field(
        default_factory=ProtectionCredentialSnapshot
    )
    has_security_descriptor: bool = False
    security_descriptor_signature: str | None = field(default=None, repr=False)
    opaque_metadata: ProtectionOpaqueMetadataSnapshot = field(
        default_factory=ProtectionOpaqueMetadataSnapshot
    )

    @property
    def target_range_count(self) -> int:
        return len(self.ranges)

    def sort_key(self) -> tuple[object, ...]:
        return (
            self.sheet.casefold(),
            self.ranges,
            self.has_name,
            self.name_signature or "",
            self.credential.configured,
            self.has_security_descriptor,
        )

    def to_dict(self) -> dict[str, Any]:
        """Return safe range permissions without names or credential material."""
        return {
            "sheet": self.sheet,
            "ranges": [
                display_location((self.sheet, target_range))
                for target_range in self.ranges
            ],
            "has_name": self.has_name,
            "credential": self.credential.to_dict(),
            "has_security_descriptor": self.has_security_descriptor,
            "opaque_metadata": self.opaque_metadata.to_dict(),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class CellProtectionDefaultSnapshot:
    """Effective locked/hidden defaults from the workbook's base cell style."""

    locked: bool
    hidden: bool

    def to_dict(self) -> dict[str, bool]:
        return {"locked": self.locked, "hidden": self.hidden}


@dataclass(frozen=True)
class CellProtectionAssignmentSnapshot:
    """One direct cell, row, or column protection-style assignment."""

    sheet: str
    scope: str
    target: str
    locked: bool
    hidden: bool

    def sort_key(self) -> tuple[object, ...]:
        return self.sheet.casefold(), self.scope, self.target.casefold()

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "scope": self.scope,
            "target": self.target,
            "locked": self.locked,
            "hidden": self.hidden,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class ExternalDataOpaqueMetadataSnapshot:
    """Private comparison evidence for unmodelled external-data XML or mashups."""

    count: int = 0
    signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return self.count > 0

    def to_dict(self) -> dict[str, Any]:
        """Return only the existence and amount of opaque metadata."""
        return {
            "present": self.present,
            "count": self.count,
        }


@dataclass(frozen=True)
class ExternalDataRefreshSettingsSnapshot:
    """Workbook-wide controls that can cause external data to refresh."""

    update_links: str = "user_set"
    allow_refresh_query: bool = False
    refresh_all_connections: bool = False
    save_external_link_values: bool = True

    def to_dict(self) -> dict[str, Any]:
        """Return safe, reviewable workbook-wide external-data settings."""
        return {
            "update_links": self.update_links,
            "allow_refresh_query": self.allow_refresh_query,
            "refresh_all_connections": self.refresh_all_connections,
            "save_external_link_values": self.save_external_link_values,
        }


@dataclass(frozen=True)
class ExternalDataConnectionSnapshot:
    """One OOXML external-data connection without its source material."""

    connection_id: int | None
    source_type: str
    deleted: bool = False
    refresh_on_load: bool = False
    refresh_interval_minutes: int | None = None
    background: bool = False
    keep_alive: bool = False
    save_data: bool = False
    save_password: bool = False
    has_source_file: bool = False
    has_connection_file: bool = False
    only_use_connection_file: bool = False
    reconnection_method: str = "as_required"
    credential_method: str = "integrated"
    minimum_refreshable_version: int = 0
    has_single_sign_on_id: bool = False
    awaiting_initial_refresh: bool = False
    has_name: bool = False
    has_description: bool = False
    source_components: tuple[str, ...] = ()
    parameter_count: int = 0
    parameters_refresh_on_change: int = 0
    identity_signature: str | None = field(default=None, repr=False)
    source_configuration_signature: str | None = field(default=None, repr=False)
    opaque_metadata: ExternalDataOpaqueMetadataSnapshot = field(
        default_factory=ExternalDataOpaqueMetadataSnapshot
    )

    def sort_key(self) -> tuple[object, ...]:
        return (
            self.connection_id is None,
            self.connection_id if self.connection_id is not None else -1,
            self.identity_signature or "",
            self.source_configuration_signature or "",
        )

    def to_dict(self) -> dict[str, Any]:
        """Return safe connection controls without source strings or credentials."""
        return {
            "id": self.connection_id,
            "source_type": self.source_type,
            "deleted": self.deleted,
            "refresh_on_load": self.refresh_on_load,
            "refresh_interval_minutes": self.refresh_interval_minutes,
            "background": self.background,
            "keep_alive": self.keep_alive,
            "save_data": self.save_data,
            "save_password": self.save_password,
            "has_source_file": self.has_source_file,
            "has_connection_file": self.has_connection_file,
            "only_use_connection_file": self.only_use_connection_file,
            "reconnection_method": self.reconnection_method,
            "credential_method": self.credential_method,
            "minimum_refreshable_version": self.minimum_refreshable_version,
            "has_single_sign_on_id": self.has_single_sign_on_id,
            "awaiting_initial_refresh": self.awaiting_initial_refresh,
            "has_name": self.has_name,
            "has_description": self.has_description,
            "source_components": list(self.source_components),
            "parameter_count": self.parameter_count,
            "parameters_refresh_on_change": self.parameters_refresh_on_change,
            "opaque_metadata": self.opaque_metadata.to_dict(),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class QueryTableRefreshSnapshot:
    """One query-table refresh control set, tied to its worksheet safely."""

    sheet: str
    connection_id: int | None
    refresh_on_load: bool = False
    background_refresh: bool = True
    refresh_disabled: bool = False
    remove_data_on_save: bool = False
    fill_formulas: bool = False
    connection_edit_disabled: bool = False
    growth_behavior: str = "insert_delete"
    has_name: bool = False
    has_refresh_metadata: bool = False
    identity_signature: str | None = field(default=None, repr=False)
    opaque_metadata: ExternalDataOpaqueMetadataSnapshot = field(
        default_factory=ExternalDataOpaqueMetadataSnapshot
    )

    def sort_key(self) -> tuple[object, ...]:
        return (
            self.sheet.casefold(),
            self.connection_id is None,
            self.connection_id if self.connection_id is not None else -1,
            self.identity_signature or "",
        )

    def to_dict(self) -> dict[str, Any]:
        """Return safe query-table behavior without query names or data."""
        return {
            "sheet": self.sheet,
            "connection_id": self.connection_id,
            "refresh_on_load": self.refresh_on_load,
            "background_refresh": self.background_refresh,
            "refresh_disabled": self.refresh_disabled,
            "remove_data_on_save": self.remove_data_on_save,
            "fill_formulas": self.fill_formulas,
            "connection_edit_disabled": self.connection_edit_disabled,
            "growth_behavior": self.growth_behavior,
            "has_name": self.has_name,
            "has_refresh_metadata": self.has_refresh_metadata,
            "opaque_metadata": self.opaque_metadata.to_dict(),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class PivotCacheRefreshSnapshot:
    """One pivot-cache source and refresh control set without cached data."""

    # OOXML cache IDs are writer-assigned graph handles. Preserve the safe value
    # in a standalone profile, but do not make a harmless renumbering a source
    # or refresh-control change; the PivotTable package scanner compares the
    # normalized cache binding separately.
    cache_id: int | None = field(compare=False)
    source_type: str
    connection_id: int | None = None
    refresh_on_load: bool = False
    background_query: bool = False
    refresh_enabled: bool = True
    save_data: bool = True
    upgrade_on_refresh: bool = False
    source_configuration_signature: str | None = field(default=None, repr=False)
    opaque_metadata: ExternalDataOpaqueMetadataSnapshot = field(
        default_factory=ExternalDataOpaqueMetadataSnapshot
    )

    def sort_key(self) -> tuple[object, ...]:
        return (
            self.source_configuration_signature or "",
            self.source_type,
            self.connection_id is None,
            self.connection_id if self.connection_id is not None else -1,
            self.refresh_on_load,
            self.background_query,
            self.refresh_enabled,
            self.save_data,
            self.upgrade_on_refresh,
            self.opaque_metadata.signature or "",
        )

    def to_dict(self) -> dict[str, Any]:
        """Return safe pivot-cache controls without source or cache contents."""
        return {
            "cache_id": self.cache_id,
            "source_type": self.source_type,
            "connection_id": self.connection_id,
            "refresh_on_load": self.refresh_on_load,
            "background_query": self.background_query,
            "refresh_enabled": self.refresh_enabled,
            "save_data": self.save_data,
            "upgrade_on_refresh": self.upgrade_on_refresh,
            "opaque_metadata": self.opaque_metadata.to_dict(),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class ExternalLinkPackageSnapshot:
    """Safe aggregate of raw OOXML external-workbook, DDE, and OLE links."""

    external_link_count: int = 0
    external_workbook_count: int = 0
    dde_link_count: int = 0
    ole_link_count: int = 0
    unrecognized_link_count: int = 0
    external_workbook_sheet_count: int = 0
    external_defined_name_count: int = 0
    external_workbook_cached_sheet_count: int = 0
    external_workbook_cached_cell_count: int = 0
    external_workbook_cached_refresh_error_count: int = 0
    dde_item_count: int = 0
    dde_advise_item_count: int = 0
    dde_ole_item_count: int = 0
    dde_prefer_picture_item_count: int = 0
    dde_cached_value_count: int = 0
    ole_item_count: int = 0
    ole_advise_item_count: int = 0
    ole_icon_item_count: int = 0
    ole_prefer_picture_item_count: int = 0
    source_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    cached_material_signature: str | None = field(default=None, repr=False)
    opaque_metadata: ExternalDataOpaqueMetadataSnapshot = field(
        default_factory=ExternalDataOpaqueMetadataSnapshot
    )

    @property
    def present(self) -> bool:
        return self.external_link_count > 0

    def to_dict(self) -> dict[str, Any]:
        """Return structural link evidence without targets, names, or cached values."""
        return {
            "present": self.present,
            "external_link_count": self.external_link_count,
            "external_workbook_count": self.external_workbook_count,
            "dde_link_count": self.dde_link_count,
            "ole_link_count": self.ole_link_count,
            "unrecognized_link_count": self.unrecognized_link_count,
            "external_workbook_sheet_count": self.external_workbook_sheet_count,
            "external_defined_name_count": self.external_defined_name_count,
            "external_workbook_cached_sheet_count": self.external_workbook_cached_sheet_count,
            "external_workbook_cached_cell_count": self.external_workbook_cached_cell_count,
            "external_workbook_cached_refresh_error_count": (
                self.external_workbook_cached_refresh_error_count
            ),
            "dde_item_count": self.dde_item_count,
            "dde_advise_item_count": self.dde_advise_item_count,
            "dde_ole_item_count": self.dde_ole_item_count,
            "dde_prefer_picture_item_count": self.dde_prefer_picture_item_count,
            "dde_cached_value_count": self.dde_cached_value_count,
            "ole_item_count": self.ole_item_count,
            "ole_advise_item_count": self.ole_advise_item_count,
            "ole_icon_item_count": self.ole_icon_item_count,
            "ole_prefer_picture_item_count": self.ole_prefer_picture_item_count,
            "opaque_metadata": self.opaque_metadata.to_dict(),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class ExternalRelationshipSnapshot:
    """Safe aggregate of package-wide OPC relationships with external targets.

    An OPC relationship can point outside the workbook from any package part,
    including parts that a feature-specific reader does not understand. Private
    signatures retain source, type, target, and malformed-metadata evidence for
    comparison while public output exposes counts only.
    """

    external_relationship_part_count: int = 0
    external_relationship_source_count: int = 0
    external_relationship_count: int = 0
    external_hyperlink_relationship_count: int = 0
    external_image_relationship_count: int = 0
    external_other_relationship_count: int = 0
    unrecognized_relationship_count: int = 0
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.external_relationship_part_count
            or self.external_relationship_count
            or self.unrecognized_relationship_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural relationship evidence without endpoints or sources."""
        return {
            "present": self.present,
            "external_relationship_part_count": self.external_relationship_part_count,
            "external_relationship_source_count": self.external_relationship_source_count,
            "external_relationship_count": self.external_relationship_count,
            "external_hyperlink_relationship_count": (
                self.external_hyperlink_relationship_count
            ),
            "external_image_relationship_count": self.external_image_relationship_count,
            "external_other_relationship_count": self.external_other_relationship_count,
            "unrecognized_relationship_count": self.unrecognized_relationship_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FormulaExternalActionSnapshot:
    """Safe aggregate of formula calls that can cross workbook or host boundaries.

    ``HYPERLINK`` can be in-workbook, while ``WEBSERVICE``, ``IMAGE``, ``RTD``,
    ``STOCKHISTORY``, and the Cube family can use dynamic expressions or
    host-side data providers. FormulaFence inventories the known action and
    provider functions without evaluating an argument, resolving a destination,
    requesting content, querying a cube, or starting an automation server.
    Formula-defined names and named LAMBDAs can hold those calls too, so their
    relevant definitions are retained in a separate private signature. Private
    signatures retain cell and formula material so same-count endpoint,
    connection, query, or provider changes remain visible without exposing
    them. Private action-cell identities allow the diff to guard statically
    visible inputs.
    """

    formula_external_action_cell_count: int = 0
    action_defined_name_count: int = 0
    hyperlink_function_count: int = 0
    webservice_function_count: int = 0
    image_function_count: int = 0
    rtd_function_count: int = 0
    stockhistory_function_count: int = 0
    cube_function_count: int = 0
    action_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    action_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.formula_external_action_cell_count or self.action_defined_name_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return function counts without formulas, locations, or arguments."""
        return {
            "present": self.present,
            "formula_external_action_cell_count": self.formula_external_action_cell_count,
            "action_defined_name_count": self.action_defined_name_count,
            "hyperlink_function_count": self.hyperlink_function_count,
            "webservice_function_count": self.webservice_function_count,
            "image_function_count": self.image_function_count,
            "rtd_function_count": self.rtd_function_count,
            "stockhistory_function_count": self.stockhistory_function_count,
            "cube_function_count": self.cube_function_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FormulaDdeLinkSnapshot:
    """Private ledger for direct DDE-style worksheet and named formulas.

    Direct DDE formulas use an application/topic/item expression rather than a
    normal worksheet function.  FormulaFence recognizes only a conservative,
    lexical ``application|topic!item`` boundary (including terminal command
    forms with no item) and retains aggregate counts.
    The application, topic, item, source formula, cells, and defined-name
    identities remain in private signatures so same-count material changes are
    comparable without exposing or resolving an endpoint.  FormulaFence never
    evaluates a formula, launches or contacts a DDE server, or executes a
    command.  Raw OOXML external-link packages are covered separately because
    they can exist independently of direct formula syntax.
    """

    dde_formula_cell_count: int = 0
    dde_link_count: int = 0
    dde_defined_name_count: int = 0
    invocation_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    dde_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(self.dde_formula_cell_count or self.dde_defined_name_count)

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate counts without DDE endpoint or formula material."""
        return {
            "present": self.present,
            "dde_formula_cell_count": self.dde_formula_cell_count,
            "dde_link_count": self.dde_link_count,
            "dde_defined_name_count": self.dde_defined_name_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class PythonInExcelSnapshot:
    """Safe aggregate of Python-in-Excel code and its workbook bindings.

    Python-in-Excel can store executable code in the documented 2023
    ``xl/python.xml`` part and in a separately stored 2022
    ``xl/pythonScripts.xml`` package part. The public model deliberately
    retains only aggregate physical-part and script counts. Private signatures
    preserve code, environment, package binding, formula placeholders, and
    arguments for comparison; private formula-cell identities support
    static-input guarding.
    """

    python_part_count: int = 0
    python_formula_cell_count: int = 0
    python_function_count: int = 0
    python_script_count: int = 0
    python_environment_definition_count: int = 0
    python_initialization_count: int = 0
    unrecognized_python_in_excel_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    formula_signature: str | None = field(default=None, repr=False)
    python_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.python_part_count
            or self.python_formula_cell_count
            or self.unrecognized_python_in_excel_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return safe counts without source code, cell locations, or arguments."""
        return {
            "present": self.present,
            "python_part_count": self.python_part_count,
            "python_formula_cell_count": self.python_formula_cell_count,
            "python_function_count": self.python_function_count,
            "python_script_count": self.python_script_count,
            "python_environment_definition_count": (
                self.python_environment_definition_count
            ),
            "python_initialization_count": self.python_initialization_count,
            "unrecognized_python_in_excel_count": (
                self.unrecognized_python_in_excel_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class OfficeCustomFunctionSnapshot:
    """Safe aggregate of namespaced Office custom-function call candidates.

    An Office Add-in custom function is surfaced in Excel as a namespaced
    formula call, but its manifest, JavaScript, and remote runtime are not
    stored in the normal workbook package.  FormulaFence keeps only a private
    signature and cell identities for candidate calls, so the public model
    never reveals function names, namespaces, cells, formulas, or arguments.
    A candidate is a review signal rather than proof that an Office Add-in is
    installed or runnable.
    """

    namespaced_custom_function_formula_cell_count: int = 0
    namespaced_custom_function_call_count: int = 0
    namespaced_custom_function_namespace_count: int = 0
    call_signature: str | None = field(default=None, repr=False)
    call_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(self.namespaced_custom_function_formula_cell_count)

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate candidate counts without formula material."""
        return {
            "present": self.present,
            "namespaced_custom_function_formula_cell_count": (
                self.namespaced_custom_function_formula_cell_count
            ),
            "namespaced_custom_function_call_count": (
                self.namespaced_custom_function_call_count
            ),
            "namespaced_custom_function_namespace_count": (
                self.namespaced_custom_function_namespace_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class UnqualifiedRuntimeFunctionSnapshot:
    """Private ledger for unknown bare worksheet-call candidates.

    Excel can bind an unqualified function token to a VBA UDF, COM/Automation
    add-in, XLL, or another registered runtime. FormulaFence cannot establish
    which provider, if any, would resolve the call, so the public model keeps
    only aggregate counts. Private signatures retain the candidate formula and
    relevant formula-defined-name chain for comparison; cell identities support
    static-input guarding without publishing names, formulas, or arguments.
    """

    unqualified_runtime_function_formula_cell_count: int = 0
    unqualified_runtime_function_call_count: int = 0
    unqualified_runtime_function_defined_name_count: int = 0
    call_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    call_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.unqualified_runtime_function_formula_cell_count
            or self.unqualified_runtime_function_defined_name_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate candidate counts without formula material."""
        return {
            "present": self.present,
            "unqualified_runtime_function_formula_cell_count": (
                self.unqualified_runtime_function_formula_cell_count
            ),
            "unqualified_runtime_function_call_count": (
                self.unqualified_runtime_function_call_count
            ),
            "unqualified_runtime_function_defined_name_count": (
                self.unqualified_runtime_function_defined_name_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WorksheetCodeResourceRegistrationSnapshot:
    """Private ledger for stored worksheet/formula-defined ``REGISTER.ID`` calls.

    Microsoft documents ``REGISTER.ID`` as a worksheet function that registers
    a DLL or code resource when it has not already been registered.  Module
    names, procedure names, type strings, arguments, formula-defined names,
    and locations are sensitive implementation material, so public output
    keeps only counts.  Private signatures retain the stored expression and
    the relevant named-definition chain for comparisons; cell identities let
    the diff layer guard ordinary edits that statically reach a registration.
    """

    registration_formula_cell_count: int = 0
    register_id_function_count: int = 0
    registration_defined_name_count: int = 0
    call_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    registration_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.registration_formula_cell_count
            or self.registration_defined_name_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate counts without DLL/code-resource material."""
        return {
            "present": self.present,
            "registration_formula_cell_count": self.registration_formula_cell_count,
            "register_id_function_count": self.register_id_function_count,
            "registration_defined_name_count": self.registration_defined_name_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FormulaDefinedXlmRegistrationSnapshot:
    """Private ledger for XLM ``REGISTER`` calls stored in defined formulas.

    Excel's legacy ``REGISTER`` macro function can be stored in a defined-name
    formula or named LAMBDA. This boundary deliberately inventories only that
    documented stored-definition surface: direct worksheet formulas and raw
    XLM macro-sheet parts remain outside it. Module names, procedure names,
    type strings, arguments, formula-defined names, and locations are private.
    Private signatures preserve stored expressions and the relevant
    named-definition chain for comparison; cell identities let the diff layer
    guard ordinary edits that statically reach a registration invocation.
    """

    registration_formula_cell_count: int = 0
    register_function_count: int = 0
    registration_defined_name_count: int = 0
    invocation_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    registration_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.registration_formula_cell_count
            or self.registration_defined_name_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate counts without DLL/XLL or formula material."""
        return {
            "present": self.present,
            "registration_formula_cell_count": self.registration_formula_cell_count,
            "register_function_count": self.register_function_count,
            "registration_defined_name_count": self.registration_defined_name_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FormulaDefinedXlmEvaluationSnapshot:
    """Private ledger for XLM `EVALUATE` calls stored in defined formulas.

    Excel's legacy `EVALUATE` function parses a supplied text expression at
    calculation time. FormulaFence inventories only calls stored in a
    formula-defined name or named LAMBDA, preserving their stored syntax and
    definition chain privately. It records worksheet cells that statically
    invoke a stored evaluation, but does not evaluate its text, infer
    dependencies from the resulting expression, or expose formulas, arguments,
    locations, or defined-name identities in public output.
    """

    evaluation_formula_cell_count: int = 0
    evaluate_function_count: int = 0
    evaluation_defined_name_count: int = 0
    invocation_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    evaluation_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.evaluation_formula_cell_count
            or self.evaluation_defined_name_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate counts without expression or formula material."""
        return {
            "present": self.present,
            "evaluation_formula_cell_count": self.evaluation_formula_cell_count,
            "evaluate_function_count": self.evaluate_function_count,
            "evaluation_defined_name_count": self.evaluation_defined_name_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FormulaDefinedXlmActionSnapshot:
    """Private ledger for selected XLM actions stored in defined formulas.

    The selected legacy functions can dispatch a macro, program, DLL entry
    point, DDE command, or a future event handler. FormulaFence inventories
    only the function spelling stored in a formula-defined name or named
    LAMBDA and the worksheet cells that statically invoke it. It neither
    evaluates a formula nor resolves, exposes, or executes an action target,
    argument, handler, macro, program, DLL, or DDE command.
    """

    action_formula_cell_count: int = 0
    action_function_count: int = 0
    action_defined_name_count: int = 0
    invocation_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    action_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(self.action_formula_cell_count or self.action_defined_name_count)

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate counts without action or formula material."""
        return {
            "present": self.present,
            "action_formula_cell_count": self.action_formula_cell_count,
            "action_function_count": self.action_function_count,
            "action_defined_name_count": self.action_defined_name_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FormulaDefinedXlmGetCellSnapshot:
    """Private ledger for XLM `GET.CELL` calls stored in defined formulas.

    Excel's legacy `GET.CELL` information function can expose cell contents,
    formulas, display text, formatting, dimensions, protection, comments, and
    other workbook state. FormulaFence inventories only calls stored in a
    formula-defined name or named LAMBDA, preserving their stored syntax and
    definition chain privately. It records worksheet cells that statically
    invoke a stored call, but does not evaluate it, determine its requested
    information type, resolve a dynamic reference, or expose formulas,
    arguments, locations, or defined-name identities in public output.
    """

    get_cell_formula_cell_count: int = 0
    get_cell_function_count: int = 0
    get_cell_defined_name_count: int = 0
    invocation_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    get_cell_cells: frozenset[CellKey] = field(default_factory=frozenset, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.get_cell_formula_cell_count or self.get_cell_defined_name_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate counts without metadata or formula material."""
        return {
            "present": self.present,
            "get_cell_formula_cell_count": self.get_cell_formula_cell_count,
            "get_cell_function_count": self.get_cell_function_count,
            "get_cell_defined_name_count": self.get_cell_defined_name_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FormulaDefinedXlmEnvironmentInformationSnapshot:
    """Private ledger for selected XLM environment-information calls.

    Legacy GET.WORKBOOK, GET.WORKSPACE, and GET.DOCUMENT calls can observe
    workbook, application/workspace, or document state outside ordinary cell
    dependencies. FormulaFence inventories only calls stored in a
    formula-defined name or named LAMBDA, preserving their stored syntax and
    definition chain privately. It records worksheet cells that statically
    invoke a stored call, but does not evaluate it, determine its information
    type, resolve a dynamic reference, or expose formulas, arguments,
    locations, or defined-name identities in public output.
    """

    environment_information_formula_cell_count: int = 0
    environment_information_function_count: int = 0
    environment_information_defined_name_count: int = 0
    invocation_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    environment_information_cells: frozenset[CellKey] = field(
        default_factory=frozenset, repr=False
    )

    @property
    def present(self) -> bool:
        return bool(
            self.environment_information_formula_cell_count
            or self.environment_information_defined_name_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate counts without environment or formula material."""
        return {
            "present": self.present,
            "environment_information_formula_cell_count": (
                self.environment_information_formula_cell_count
            ),
            "environment_information_function_count": (
                self.environment_information_function_count
            ),
            "environment_information_defined_name_count": (
                self.environment_information_defined_name_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FormulaEnvironmentInformationSnapshot:
    """Private ledger for native workbook and environment information calls.

    Excel's ordinary CELL, INFO, SHEET, and SHEETS functions can observe
    workbook/file, client, or tab-catalog information outside visible cell
    precedents. FormulaFence inventories stored calls in worksheet formulas,
    formula-defined names, and named LAMBDAs. It separately aggregates CELL
    calls whose optional reference is omitted, because Excel may use the current
    selection at calculation time. It reports SHEET/SHEETS calls separately and
    separately aggregates SHEETS calls with an omitted reference, whose result
    is documented to depend on the workbook tab count. The ledger preserves
    material only in private signatures; it does not evaluate a formula,
    determine an information type, resolve dynamic arguments, or simulate file,
    client, workspace, selection, or workbook state.
    """

    environment_information_formula_cell_count: int = 0
    environment_information_function_count: int = 0
    environment_information_defined_name_count: int = 0
    implicit_cell_reference_function_count: int = 0
    implicit_sheets_reference_function_count: int = 0
    sheet_function_count: int = 0
    sheets_function_count: int = 0
    invocation_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    environment_information_cells: frozenset[CellKey] = field(
        default_factory=frozenset, repr=False
    )

    @property
    def present(self) -> bool:
        return bool(
            self.environment_information_formula_cell_count
            or self.environment_information_defined_name_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate counts without environment or formula material."""
        return {
            "present": self.present,
            "environment_information_formula_cell_count": (
                self.environment_information_formula_cell_count
            ),
            "environment_information_function_count": (
                self.environment_information_function_count
            ),
            "environment_information_defined_name_count": (
                self.environment_information_defined_name_count
            ),
            "implicit_cell_reference_function_count": (
                self.implicit_cell_reference_function_count
            ),
            "implicit_sheets_reference_function_count": (
                self.implicit_sheets_reference_function_count
            ),
            "sheet_function_count": self.sheet_function_count,
            "sheets_function_count": self.sheets_function_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class XlmMacroSheetSnapshot:
    """Safe aggregate of raw Excel 4.0 / XLM macro-sheet package parts.

    XLM automation lives in XML macro-sheet parts rather than in the VBA
    binary. The private signatures retain complete program, relationship, and
    direct internal related-part evidence for comparison, while ``to_dict``
    intentionally exposes only structural counts.
    """

    declared_macro_sheet_count: int = 0
    macro_sheet_count: int = 0
    international_macro_sheet_count: int = 0
    unrecognized_macro_sheet_count: int = 0
    hidden_macro_sheet_count: int = 0
    very_hidden_macro_sheet_count: int = 0
    formula_cell_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    internal_related_part_count: int = 0
    fingerprinted_related_part_count: int = 0
    uninspected_related_part_count: int = 0
    embedded_object_relationship_count: int = 0
    embedded_package_relationship_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    program_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)
    related_part_payload_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(self.declared_macro_sheet_count or self.macro_sheet_count)

    def to_dict(self) -> dict[str, Any]:
        """Return safe macro-sheet inventory without commands or endpoints."""
        return {
            "present": self.present,
            "declared_macro_sheet_count": self.declared_macro_sheet_count,
            "macro_sheet_count": self.macro_sheet_count,
            "international_macro_sheet_count": self.international_macro_sheet_count,
            "unrecognized_macro_sheet_count": self.unrecognized_macro_sheet_count,
            "hidden_macro_sheet_count": self.hidden_macro_sheet_count,
            "very_hidden_macro_sheet_count": self.very_hidden_macro_sheet_count,
            "formula_cell_count": self.formula_cell_count,
            "related_relationship_count": self.related_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "internal_related_part_count": self.internal_related_part_count,
            "fingerprinted_related_part_count": self.fingerprinted_related_part_count,
            "uninspected_related_part_count": self.uninspected_related_part_count,
            "embedded_object_relationship_count": (
                self.embedded_object_relationship_count
            ),
            "embedded_package_relationship_count": (
                self.embedded_package_relationship_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class XlmAutomaticMacroBindingSnapshot:
    """Safe aggregate of workbook-scoped XLM automatic-macro bindings.

    Legacy automatic macros are routed through special workbook defined names
    such as ``Auto_Open``. The private signature retains the name-to-cell
    binding material needed to identify a retarget even when aggregate counts
    stay stable; public output intentionally exposes only documented event
    counts.
    """

    automatic_macro_binding_count: int = 0
    auto_open_binding_count: int = 0
    auto_close_binding_count: int = 0
    auto_activate_binding_count: int = 0
    auto_deactivate_binding_count: int = 0
    binding_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(self.automatic_macro_binding_count)

    def to_dict(self) -> dict[str, Any]:
        """Return binding counts without names, targets, or formula text."""
        return {
            "present": self.present,
            "automatic_macro_binding_count": self.automatic_macro_binding_count,
            "auto_open_binding_count": self.auto_open_binding_count,
            "auto_close_binding_count": self.auto_close_binding_count,
            "auto_activate_binding_count": self.auto_activate_binding_count,
            "auto_deactivate_binding_count": self.auto_deactivate_binding_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class RibbonCustomizationSnapshot:
    """Safe aggregate of workbook-scoped Office RibbonX customization parts.

    Ribbon XML can bind controls to workbook callbacks while remaining outside
    the VBA project. Private signatures retain complete package and callback
    material for comparison, while ``to_dict`` intentionally exposes only
    structural counts.
    """

    declared_ribbon_part_count: int = 0
    ribbon_part_count: int = 0
    office_2010_ribbon_part_count: int = 0
    unrecognized_ribbon_part_count: int = 0
    control_count: int = 0
    callback_attribute_count: int = 0
    action_callback_count: int = 0
    image_relationship_count: int = 0
    external_relationship_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(self.declared_ribbon_part_count or self.ribbon_part_count)

    def to_dict(self) -> dict[str, Any]:
        """Return safe RibbonX inventory without names, labels, or callbacks."""
        return {
            "present": self.present,
            "declared_ribbon_part_count": self.declared_ribbon_part_count,
            "ribbon_part_count": self.ribbon_part_count,
            "office_2010_ribbon_part_count": self.office_2010_ribbon_part_count,
            "unrecognized_ribbon_part_count": self.unrecognized_ribbon_part_count,
            "control_count": self.control_count,
            "callback_attribute_count": self.callback_attribute_count,
            "action_callback_count": self.action_callback_count,
            "image_relationship_count": self.image_relationship_count,
            "external_relationship_count": self.external_relationship_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class OfficeWebAddinSnapshot:
    """Safe aggregate of document-linked and in-content Office Web Add-ins.

    OOXML task-pane parts, worksheet binding extensions, and DrawingML
    in-content frames can bind a workbook to an installed Office Web Add-in.
    Private signatures retain the add-in reference, property, binding, frame,
    and relationship material needed for comparison; ``to_dict`` intentionally
    exposes only structural counts.
    """

    declared_taskpane_part_count: int = 0
    taskpane_part_count: int = 0
    web_extension_part_count: int = 0
    unrecognized_part_count: int = 0
    taskpane_count: int = 0
    visible_taskpane_count: int = 0
    locked_taskpane_count: int = 0
    web_extension_reference_count: int = 0
    auto_show_taskpane_count: int = 0
    store_reference_count: int = 0
    alternate_reference_count: int = 0
    binding_count: int = 0
    snapshot_reference_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    worksheet_binding_sheet_count: int = 0
    worksheet_binding_count: int = 0
    in_content_drawing_part_count: int = 0
    in_content_web_extension_reference_count: int = 0
    in_content_web_extension_part_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    taskpane_signature: str | None = field(default=None, repr=False)
    web_extension_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)
    worksheet_binding_signature: str | None = field(default=None, repr=False)
    in_content_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.declared_taskpane_part_count
            or self.taskpane_part_count
            or self.web_extension_part_count
            or self.unrecognized_part_count
            or self.worksheet_binding_count
            or self.in_content_web_extension_reference_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return add-in inventory without identities, bindings, or endpoints."""
        return {
            "present": self.present,
            "declared_taskpane_part_count": self.declared_taskpane_part_count,
            "taskpane_part_count": self.taskpane_part_count,
            "web_extension_part_count": self.web_extension_part_count,
            "unrecognized_part_count": self.unrecognized_part_count,
            "taskpane_count": self.taskpane_count,
            "visible_taskpane_count": self.visible_taskpane_count,
            "locked_taskpane_count": self.locked_taskpane_count,
            "web_extension_reference_count": self.web_extension_reference_count,
            "auto_show_taskpane_count": self.auto_show_taskpane_count,
            "store_reference_count": self.store_reference_count,
            "alternate_reference_count": self.alternate_reference_count,
            "binding_count": self.binding_count,
            "snapshot_reference_count": self.snapshot_reference_count,
            "related_relationship_count": self.related_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "worksheet_binding_sheet_count": self.worksheet_binding_sheet_count,
            "worksheet_binding_count": self.worksheet_binding_count,
            "in_content_drawing_part_count": self.in_content_drawing_part_count,
            "in_content_web_extension_reference_count": (
                self.in_content_web_extension_reference_count
            ),
            "in_content_web_extension_part_count": (
                self.in_content_web_extension_part_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class ChartDefinitionSnapshot:
    """Safe aggregate of DrawingML chart definitions and presentation material.

    Legacy and Office 2016+ ChartEx parts sit behind worksheet or chartsheet
    drawing relationships, outside the ordinary cell grid. Private signatures
    retain chart formulas, formatting, cached points, ChartEx declarations,
    overlay-shape definitions, relationships, and bounded direct payload
    evidence for comparison; ``to_dict`` deliberately exposes only structural
    counts.
    """

    chart_host_sheet_count: int = 0
    chart_drawing_part_count: int = 0
    chart_reference_count: int = 0
    chart_part_count: int = 0
    chart_ex_reference_count: int = 0
    chart_ex_part_count: int = 0
    chart_ex_series_count: int = 0
    chart_ex_title_count: int = 0
    chart_ex_data_reference_count: int = 0
    chart_user_shape_part_count: int = 0
    chart_user_shape_count: int = 0
    chart_type_count: int = 0
    series_count: int = 0
    title_count: int = 0
    data_reference_count: int = 0
    numeric_data_reference_count: int = 0
    string_data_reference_count: int = 0
    literal_data_point_count: int = 0
    cached_data_point_count: int = 0
    pivot_source_count: int = 0
    external_data_reference_count: int = 0
    user_shape_reference_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    internal_related_part_count: int = 0
    fingerprinted_related_part_count: int = 0
    uninspected_related_part_count: int = 0
    unrecognized_part_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    cached_data_signature: str | None = field(default=None, repr=False)
    user_shape_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)
    related_part_payload_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.chart_host_sheet_count
            or self.chart_drawing_part_count
            or self.chart_reference_count
            or self.chart_part_count
            or self.chart_ex_reference_count
            or self.chart_ex_part_count
            or self.chart_user_shape_part_count
            or self.unrecognized_part_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return chart evidence without labels, formulas, values, or targets."""
        return {
            "present": self.present,
            "chart_host_sheet_count": self.chart_host_sheet_count,
            "chart_drawing_part_count": self.chart_drawing_part_count,
            "chart_reference_count": self.chart_reference_count,
            "chart_part_count": self.chart_part_count,
            "chart_ex_reference_count": self.chart_ex_reference_count,
            "chart_ex_part_count": self.chart_ex_part_count,
            "chart_ex_series_count": self.chart_ex_series_count,
            "chart_ex_title_count": self.chart_ex_title_count,
            "chart_ex_data_reference_count": self.chart_ex_data_reference_count,
            "chart_user_shape_part_count": self.chart_user_shape_part_count,
            "chart_user_shape_count": self.chart_user_shape_count,
            "chart_type_count": self.chart_type_count,
            "series_count": self.series_count,
            "title_count": self.title_count,
            "data_reference_count": self.data_reference_count,
            "numeric_data_reference_count": self.numeric_data_reference_count,
            "string_data_reference_count": self.string_data_reference_count,
            "literal_data_point_count": self.literal_data_point_count,
            "cached_data_point_count": self.cached_data_point_count,
            "pivot_source_count": self.pivot_source_count,
            "external_data_reference_count": self.external_data_reference_count,
            "user_shape_reference_count": self.user_shape_reference_count,
            "related_relationship_count": self.related_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "internal_related_part_count": self.internal_related_part_count,
            "fingerprinted_related_part_count": self.fingerprinted_related_part_count,
            "uninspected_related_part_count": self.uninspected_related_part_count,
            "unrecognized_part_count": self.unrecognized_part_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class PivotTableDefinitionSnapshot:
    """Safe aggregate of PivotTable view, cache-schema, and cached-data material.

    PivotTable views and their cache packages can change grouping, filtering,
    aggregation, and report presentation outside ordinary formula cells. Private
    signatures retain those definitions and cache payload evidence for comparison;
    ``to_dict`` deliberately exposes only structural counts.
    """

    pivot_table_sheet_count: int = 0
    pivot_table_part_count: int = 0
    pivot_cache_definition_part_count: int = 0
    pivot_cache_records_part_count: int = 0
    pivot_cache_binding_count: int = 0
    layout_location_count: int = 0
    pivot_field_count: int = 0
    row_field_count: int = 0
    column_field_count: int = 0
    page_field_count: int = 0
    data_field_count: int = 0
    filter_count: int = 0
    row_item_count: int = 0
    column_item_count: int = 0
    cache_field_count: int = 0
    shared_item_count: int = 0
    calculated_item_count: int = 0
    calculated_member_count: int = 0
    cache_record_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    fingerprinted_cache_record_part_count: int = 0
    uninspected_cache_record_part_count: int = 0
    unrecognized_part_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    layout_signature: str | None = field(default=None, repr=False)
    cache_definition_signature: str | None = field(default=None, repr=False)
    cached_shared_item_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)
    cache_record_payload_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.pivot_table_sheet_count
            or self.pivot_table_part_count
            or self.pivot_cache_definition_part_count
            or self.pivot_cache_records_part_count
            or self.unrecognized_part_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return PivotTable evidence without values, formulas, names, or targets."""
        return {
            "present": self.present,
            "pivot_table_sheet_count": self.pivot_table_sheet_count,
            "pivot_table_part_count": self.pivot_table_part_count,
            "pivot_cache_definition_part_count": self.pivot_cache_definition_part_count,
            "pivot_cache_records_part_count": self.pivot_cache_records_part_count,
            "pivot_cache_binding_count": self.pivot_cache_binding_count,
            "layout_location_count": self.layout_location_count,
            "pivot_field_count": self.pivot_field_count,
            "row_field_count": self.row_field_count,
            "column_field_count": self.column_field_count,
            "page_field_count": self.page_field_count,
            "data_field_count": self.data_field_count,
            "filter_count": self.filter_count,
            "row_item_count": self.row_item_count,
            "column_item_count": self.column_item_count,
            "cache_field_count": self.cache_field_count,
            "shared_item_count": self.shared_item_count,
            "calculated_item_count": self.calculated_item_count,
            "calculated_member_count": self.calculated_member_count,
            "cache_record_count": self.cache_record_count,
            "related_relationship_count": self.related_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "fingerprinted_cache_record_part_count": (
                self.fingerprinted_cache_record_part_count
            ),
            "uninspected_cache_record_part_count": self.uninspected_cache_record_part_count,
            "unrecognized_part_count": self.unrecognized_part_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class SlicerTimelineCacheSnapshot:
    """Safe aggregate of interactive slicer and Timeline filter-cache material.

    Slicer and Timeline caches can apply report filters to PivotTables or
    tables without changing an ordinary worksheet cell. Private signatures
    retain their declarations, filter state, and source bindings for comparison;
    ``to_dict`` deliberately exposes only structural counts.
    """

    slicer_cache_part_count: int = 0
    timeline_cache_part_count: int = 0
    slicer_workbook_binding_count: int = 0
    timeline_workbook_binding_count: int = 0
    slicer_pivot_cache_binding_count: int = 0
    slicer_table_binding_count: int = 0
    timeline_pivot_cache_binding_count: int = 0
    slicer_pivot_table_binding_count: int = 0
    timeline_pivot_table_binding_count: int = 0
    slicer_item_count: int = 0
    selected_slicer_item_count: int = 0
    timeline_state_count: int = 0
    timeline_filter_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    unrecognized_part_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    slicer_definition_signature: str | None = field(default=None, repr=False)
    timeline_definition_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.slicer_cache_part_count
            or self.timeline_cache_part_count
            or self.slicer_workbook_binding_count
            or self.timeline_workbook_binding_count
            or self.unrecognized_part_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return filter-cache evidence without names, values, or targets."""
        return {
            "present": self.present,
            "slicer_cache_part_count": self.slicer_cache_part_count,
            "timeline_cache_part_count": self.timeline_cache_part_count,
            "slicer_workbook_binding_count": self.slicer_workbook_binding_count,
            "timeline_workbook_binding_count": self.timeline_workbook_binding_count,
            "slicer_pivot_cache_binding_count": self.slicer_pivot_cache_binding_count,
            "slicer_table_binding_count": self.slicer_table_binding_count,
            "timeline_pivot_cache_binding_count": self.timeline_pivot_cache_binding_count,
            "slicer_pivot_table_binding_count": self.slicer_pivot_table_binding_count,
            "timeline_pivot_table_binding_count": self.timeline_pivot_table_binding_count,
            "slicer_item_count": self.slicer_item_count,
            "selected_slicer_item_count": self.selected_slicer_item_count,
            "timeline_state_count": self.timeline_state_count,
            "timeline_filter_count": self.timeline_filter_count,
            "related_relationship_count": self.related_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "unrecognized_part_count": self.unrecognized_part_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class PowerPivotDataModelSnapshot:
    """Safe aggregate of an embedded Power Pivot / Data Model package.

    An Excel Data Model can hold tables, relationships, and DAX calculations
    outside the ordinary cell grid. Private signatures retain its workbook
    declaration, relationship bindings, and bounded raw model bytes for
    comparison; ``to_dict`` deliberately exposes only structural counts.
    """

    data_model_part_count: int = 0
    workbook_binding_count: int = 0
    data_model_declaration_count: int = 0
    model_table_count: int = 0
    model_relationship_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    fingerprinted_data_part_count: int = 0
    uninspected_data_part_count: int = 0
    unrecognized_part_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)
    payload_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.data_model_part_count
            or self.workbook_binding_count
            or self.data_model_declaration_count
            or self.unrecognized_part_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return Data Model evidence without names, DAX, values, or targets."""
        return {
            "present": self.present,
            "data_model_part_count": self.data_model_part_count,
            "workbook_binding_count": self.workbook_binding_count,
            "data_model_declaration_count": self.data_model_declaration_count,
            "model_table_count": self.model_table_count,
            "model_relationship_count": self.model_relationship_count,
            "related_relationship_count": self.related_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "fingerprinted_data_part_count": self.fingerprinted_data_part_count,
            "uninspected_data_part_count": self.uninspected_data_part_count,
            "unrecognized_part_count": self.unrecognized_part_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WhatIfDataTableSnapshot:
    """Safe aggregate of Excel What-If Data Table definitions.

    What-If Data Tables are formula-bearing sensitivity engines. Their output
    ranges and input references can disclose model structure, so the private
    signature retains their canonical definitions while ``to_dict`` exposes
    only structural counts suitable for review artifacts.
    """

    data_table_count: int = 0
    one_variable_data_table_count: int = 0
    two_variable_data_table_count: int = 0
    one_variable_row_oriented_count: int = 0
    one_variable_column_oriented_count: int = 0
    declared_output_cell_count: int = 0
    recalculation_requested_count: int = 0
    deleted_input_reference_count: int = 0
    unrecognized_data_table_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return self.data_table_count > 0

    def to_dict(self) -> dict[str, Any]:
        """Return structural Data Table evidence without references or values."""
        return {
            "present": self.present,
            "data_table_count": self.data_table_count,
            "one_variable_data_table_count": self.one_variable_data_table_count,
            "two_variable_data_table_count": self.two_variable_data_table_count,
            "one_variable_row_oriented_count": self.one_variable_row_oriented_count,
            "one_variable_column_oriented_count": self.one_variable_column_oriented_count,
            "declared_output_cell_count": self.declared_output_cell_count,
            "recalculation_requested_count": self.recalculation_requested_count,
            "deleted_input_reference_count": self.deleted_input_reference_count,
            "unrecognized_data_table_count": self.unrecognized_data_table_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class ScenarioManagerSnapshot:
    """Safe aggregate of Excel Scenario Manager definitions.

    Scenario Manager stores alternate input sets, names, comments, user
    metadata, and summary references inside worksheet OOXML rather than normal
    cells. Those values can be commercially sensitive, so the private
    signature retains canonical definitions while ``to_dict`` exposes only
    structural counts suitable for review artifacts.
    """

    scenario_sheet_count: int = 0
    scenario_count: int = 0
    input_cell_count: int = 0
    locked_scenario_count: int = 0
    hidden_scenario_count: int = 0
    scenario_with_comment_count: int = 0
    scenario_with_user_count: int = 0
    summary_reference_count: int = 0
    current_scenario_selection_count: int = 0
    shown_scenario_selection_count: int = 0
    deleted_input_cell_count: int = 0
    undone_input_cell_count: int = 0
    formatted_input_cell_count: int = 0
    unrecognized_scenario_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(self.scenario_count or self.unrecognized_scenario_count)

    def to_dict(self) -> dict[str, Any]:
        """Return structural Scenario Manager evidence without values or references."""
        return {
            "present": self.present,
            "scenario_sheet_count": self.scenario_sheet_count,
            "scenario_count": self.scenario_count,
            "input_cell_count": self.input_cell_count,
            "locked_scenario_count": self.locked_scenario_count,
            "hidden_scenario_count": self.hidden_scenario_count,
            "scenario_with_comment_count": self.scenario_with_comment_count,
            "scenario_with_user_count": self.scenario_with_user_count,
            "summary_reference_count": self.summary_reference_count,
            "current_scenario_selection_count": self.current_scenario_selection_count,
            "shown_scenario_selection_count": self.shown_scenario_selection_count,
            "deleted_input_cell_count": self.deleted_input_cell_count,
            "undone_input_cell_count": self.undone_input_cell_count,
            "formatted_input_cell_count": self.formatted_input_cell_count,
            "unrecognized_scenario_count": self.unrecognized_scenario_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FilterVisibilitySnapshot:
    """Safe aggregate of Excel filter, sort, and visibility controls.

    AutoFilter criteria can contain sensitive customer, product, or financial
    values, while row/column identifiers and sort ranges can reveal report structure.
    The private signature retains canonical declarations for comparison; the
    public profile deliberately exposes counts only.
    """

    worksheet_auto_filter_count: int = 0
    table_auto_filter_count: int = 0
    filter_column_count: int = 0
    filter_criterion_count: int = 0
    sort_state_count: int = 0
    sort_condition_count: int = 0
    default_hidden_sheet_count: int = 0
    default_zero_height_sheet_count: int = 0
    default_zero_width_sheet_count: int = 0
    hidden_row_count: int = 0
    zero_height_row_count: int = 0
    outlined_row_count: int = 0
    collapsed_row_count: int = 0
    visible_row_override_count: int = 0
    hidden_column_count: int = 0
    zero_width_column_count: int = 0
    outlined_column_count: int = 0
    collapsed_column_count: int = 0
    unrecognized_control_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.worksheet_auto_filter_count
            or self.table_auto_filter_count
            or self.default_hidden_sheet_count
            or self.default_zero_height_sheet_count
            or self.default_zero_width_sheet_count
            or self.hidden_row_count
            or self.zero_height_row_count
            or self.outlined_row_count
            or self.collapsed_row_count
            or self.visible_row_override_count
            or self.hidden_column_count
            or self.zero_width_column_count
            or self.outlined_column_count
            or self.collapsed_column_count
            or self.unrecognized_control_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural visibility evidence without criteria or references."""
        return {
            "present": self.present,
            "worksheet_auto_filter_count": self.worksheet_auto_filter_count,
            "table_auto_filter_count": self.table_auto_filter_count,
            "filter_column_count": self.filter_column_count,
            "filter_criterion_count": self.filter_criterion_count,
            "sort_state_count": self.sort_state_count,
            "sort_condition_count": self.sort_condition_count,
            "default_hidden_sheet_count": self.default_hidden_sheet_count,
            "default_zero_height_sheet_count": self.default_zero_height_sheet_count,
            "default_zero_width_sheet_count": self.default_zero_width_sheet_count,
            "hidden_row_count": self.hidden_row_count,
            "zero_height_row_count": self.zero_height_row_count,
            "outlined_row_count": self.outlined_row_count,
            "collapsed_row_count": self.collapsed_row_count,
            "visible_row_override_count": self.visible_row_override_count,
            "hidden_column_count": self.hidden_column_count,
            "zero_width_column_count": self.zero_width_column_count,
            "outlined_column_count": self.outlined_column_count,
            "collapsed_column_count": self.collapsed_column_count,
            "unrecognized_control_count": self.unrecognized_control_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class IgnoredErrorSnapshot:
    """Safe aggregate of Excel's per-range ignored error-checking controls.

    Ignored-error ranges may identify sensitive model locations, while their
    declarations can suppress formula, evaluation, and validation warnings a
    reviewer would otherwise see. The private signature retains canonical
    ranges and flags; the public profile deliberately exposes only counts.
    """

    worksheet_count: int = 0
    standard_container_count: int = 0
    extension_container_count: int = 0
    ignored_error_rule_count: int = 0
    target_range_count: int = 0
    evaluation_error_count: int = 0
    inconsistent_formula_count: int = 0
    formula_range_omission_count: int = 0
    unlocked_formula_count: int = 0
    empty_cell_reference_count: int = 0
    list_data_validation_count: int = 0
    calculated_column_count: int = 0
    number_stored_as_text_count: int = 0
    two_digit_text_year_count: int = 0
    unrecognized_ignored_error_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.standard_container_count
            or self.extension_container_count
            or self.ignored_error_rule_count
            or self.unrecognized_ignored_error_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural evidence without ranges or error-control details."""
        return {
            "present": self.present,
            "worksheet_count": self.worksheet_count,
            "standard_container_count": self.standard_container_count,
            "extension_container_count": self.extension_container_count,
            "ignored_error_rule_count": self.ignored_error_rule_count,
            "target_range_count": self.target_range_count,
            "evaluation_error_count": self.evaluation_error_count,
            "inconsistent_formula_count": self.inconsistent_formula_count,
            "formula_range_omission_count": self.formula_range_omission_count,
            "unlocked_formula_count": self.unlocked_formula_count,
            "empty_cell_reference_count": self.empty_cell_reference_count,
            "list_data_validation_count": self.list_data_validation_count,
            "calculated_column_count": self.calculated_column_count,
            "number_stored_as_text_count": self.number_stored_as_text_count,
            "two_digit_text_year_count": self.two_digit_text_year_count,
            "unrecognized_ignored_error_count": self.unrecognized_ignored_error_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class NamedSheetViewSnapshot:
    """Safe aggregate of modern Excel Named Sheet View declarations.

    A Named Sheet View can preserve alternate AutoFilter criteria and sort
    rules outside ordinary worksheet cells. View names, IDs, criteria, target
    ranges, table bindings, and sort keys remain inside the private signature;
    the public profile exposes structural counts only.
    """

    worksheet_count: int = 0
    part_count: int = 0
    named_sheet_view_count: int = 0
    named_filter_count: int = 0
    column_filter_count: int = 0
    filter_criterion_count: int = 0
    sort_rule_count: int = 0
    sort_condition_count: int = 0
    unrecognized_named_sheet_view_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.part_count
            or self.named_sheet_view_count
            or self.named_filter_count
            or self.unrecognized_named_sheet_view_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural Named Sheet View evidence without private settings."""
        return {
            "present": self.present,
            "worksheet_count": self.worksheet_count,
            "part_count": self.part_count,
            "named_sheet_view_count": self.named_sheet_view_count,
            "named_filter_count": self.named_filter_count,
            "column_filter_count": self.column_filter_count,
            "filter_criterion_count": self.filter_criterion_count,
            "sort_rule_count": self.sort_rule_count,
            "sort_condition_count": self.sort_condition_count,
            "unrecognized_named_sheet_view_count": self.unrecognized_named_sheet_view_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class CustomWorkbookViewSnapshot:
    """Safe aggregate of legacy Excel Custom View declarations.

    A legacy Custom View stores a named alternate workbook display and print
    surface. Its linked per-sheet declarations can hide rows or columns,
    preserve filters, change worksheet display controls, and change print
    settings without touching the default worksheet state. View names, GUIDs,
    ranges, filters, page settings, header/footer text, and raw XML remain in
    private signatures; public evidence is intentionally structural only.
    """

    custom_workbook_view_count: int = 0
    custom_sheet_view_count: int = 0
    custom_view_sheet_count: int = 0
    hidden_row_or_column_view_count: int = 0
    filtered_view_count: int = 0
    print_setting_view_count: int = 0
    display_setting_view_count: int = 0
    unrecognized_custom_view_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    unrecognized_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.custom_workbook_view_count
            or self.custom_sheet_view_count
            or self.unrecognized_custom_view_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural Custom View evidence without private settings."""
        return {
            "present": self.present,
            "custom_workbook_view_count": self.custom_workbook_view_count,
            "custom_sheet_view_count": self.custom_sheet_view_count,
            "custom_view_sheet_count": self.custom_view_sheet_count,
            "hidden_row_or_column_view_count": (
                self.hidden_row_or_column_view_count
            ),
            "filtered_view_count": self.filtered_view_count,
            "print_setting_view_count": self.print_setting_view_count,
            "display_setting_view_count": self.display_setting_view_count,
            "unrecognized_custom_view_count": self.unrecognized_custom_view_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class TableStyleControlsSnapshot:
    """Safe aggregate of applied Excel Table Style presentation controls.

    A table's ``tableStyleInfo`` can change its visible headers, totals,
    banding, and emphasized columns while leaving every ordinary cell and
    structured reference intact. Workbook-local custom style definitions can
    likewise alter the presentation of every table that selects them. Style
    names, table identities, differential formats, colours, and raw XML remain
    in private signatures; public evidence is intentionally structural only.
    """

    table_style_info_count: int = 0
    styled_table_count: int = 0
    custom_table_style_count: int = 0
    custom_table_style_element_count: int = 0
    custom_style_applied_table_count: int = 0
    table_direct_dxf_assignment_count: int = 0
    table_direct_dxf_table_count: int = 0
    table_named_cell_style_assignment_count: int = 0
    row_striped_table_count: int = 0
    column_striped_table_count: int = 0
    emphasized_column_table_count: int = 0
    unrecognized_table_style_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    unrecognized_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.table_style_info_count
            or self.custom_table_style_count
            or self.table_direct_dxf_assignment_count
            or self.table_named_cell_style_assignment_count
            or self.unrecognized_table_style_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural Table Style evidence without private formatting."""
        return {
            "present": self.present,
            "table_style_info_count": self.table_style_info_count,
            "styled_table_count": self.styled_table_count,
            "custom_table_style_count": self.custom_table_style_count,
            "custom_table_style_element_count": (
                self.custom_table_style_element_count
            ),
            "custom_style_applied_table_count": (
                self.custom_style_applied_table_count
            ),
            "table_direct_dxf_assignment_count": (
                self.table_direct_dxf_assignment_count
            ),
            "table_direct_dxf_table_count": self.table_direct_dxf_table_count,
            "table_named_cell_style_assignment_count": (
                self.table_named_cell_style_assignment_count
            ),
            "row_striped_table_count": self.row_striped_table_count,
            "column_striped_table_count": self.column_striped_table_count,
            "emphasized_column_table_count": self.emphasized_column_table_count,
            "unrecognized_table_style_count": self.unrecognized_table_style_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class SharedWorkbookRevisionSnapshot:
    """Safe aggregate of legacy shared-workbook revision-history controls.

    Legacy shared workbooks can retain a private change history in revision
    header and log parts. Those records can contain prior values, cell
    locations, author names, timestamps, comments, and conflict-resolution
    details that are outside the current worksheet grid. Public evidence stays
    structural while private signatures retain the complete stored history for
    comparison.
    """

    revision_header_part_count: int = 0
    revision_header_count: int = 0
    revision_log_part_count: int = 0
    revision_log_entry_count: int = 0
    shared_workbook_enabled_count: int = 0
    track_revisions_enabled_count: int = 0
    revision_history_enabled_count: int = 0
    keep_change_history_enabled_count: int = 0
    revision_history_protected_count: int = 0
    unrecognized_shared_workbook_revision_count: int = 0
    header_signature: str | None = field(default=None, repr=False)
    log_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)
    unrecognized_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.revision_header_part_count
            or self.revision_header_count
            or self.revision_log_part_count
            or self.revision_log_entry_count
            or self.unrecognized_shared_workbook_revision_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural revision-history evidence without private records."""
        return {
            "present": self.present,
            "revision_header_part_count": self.revision_header_part_count,
            "revision_header_count": self.revision_header_count,
            "revision_log_part_count": self.revision_log_part_count,
            "revision_log_entry_count": self.revision_log_entry_count,
            "shared_workbook_enabled_count": self.shared_workbook_enabled_count,
            "track_revisions_enabled_count": self.track_revisions_enabled_count,
            "revision_history_enabled_count": self.revision_history_enabled_count,
            "keep_change_history_enabled_count": (
                self.keep_change_history_enabled_count
            ),
            "revision_history_protected_count": (
                self.revision_history_protected_count
            ),
            "unrecognized_shared_workbook_revision_count": (
                self.unrecognized_shared_workbook_revision_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class NumberFormatSnapshot:
    """Safe aggregate of cell, row, and column number-format controls.

    A number format can change what a reviewer sees without changing a stored
    value or formula: it can hide a value, scale it, render it as a date or
    percentage, or suppress text. Format codes and target locations may carry
    sensitive reporting context, so the private signature retains canonical
    assignments for comparison while public output exposes structural counts
    only.
    """

    default_format_override_count: int = 0
    cell_format_assignment_count: int = 0
    row_format_assignment_count: int = 0
    column_format_assignment_count: int = 0
    built_in_format_assignment_count: int = 0
    custom_format_assignment_count: int = 0
    unrecognized_number_format_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.default_format_override_count
            or self.cell_format_assignment_count
            or self.row_format_assignment_count
            or self.column_format_assignment_count
            or self.unrecognized_number_format_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural number-format evidence without codes or targets."""
        return {
            "present": self.present,
            "default_format_override_count": self.default_format_override_count,
            "cell_format_assignment_count": self.cell_format_assignment_count,
            "row_format_assignment_count": self.row_format_assignment_count,
            "column_format_assignment_count": self.column_format_assignment_count,
            "built_in_format_assignment_count": self.built_in_format_assignment_count,
            "custom_format_assignment_count": self.custom_format_assignment_count,
            "unrecognized_number_format_count": self.unrecognized_number_format_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FontSnapshot:
    """Safe aggregate of effective cell-font controls.

    A font can change the review surface without changing a stored value or
    formula: for example, text can be made less visible by changing its colour,
    size, effects, or face. Font definitions and targets can expose report
    context, so the private signature retains canonical effective assignments
    for comparison while public output exposes structural counts only.
    """

    default_font_definition_count: int = 0
    cell_font_assignment_count: int = 0
    row_font_assignment_count: int = 0
    column_font_assignment_count: int = 0
    unrecognized_font_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.default_font_definition_count
            or self.cell_font_assignment_count
            or self.row_font_assignment_count
            or self.column_font_assignment_count
            or self.unrecognized_font_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural font evidence without definitions or targets."""
        return {
            "present": self.present,
            "default_font_definition_count": self.default_font_definition_count,
            "cell_font_assignment_count": self.cell_font_assignment_count,
            "row_font_assignment_count": self.row_font_assignment_count,
            "column_font_assignment_count": self.column_font_assignment_count,
            "unrecognized_font_count": self.unrecognized_font_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FillSnapshot:
    """Safe aggregate of effective cell-fill controls.

    A fill can change the review surface without changing a stored value or
    formula: for example, a matching fill can make text or an error indicator
    much less visible, while patterns and gradients can alter visual meaning.
    Fill definitions and targets can expose report context, so the private
    signature retains canonical effective assignments for comparison while
    public output exposes structural counts only.
    """

    default_fill_definition_count: int = 0
    cell_fill_assignment_count: int = 0
    row_fill_assignment_count: int = 0
    column_fill_assignment_count: int = 0
    unrecognized_fill_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.default_fill_definition_count
            or self.cell_fill_assignment_count
            or self.row_fill_assignment_count
            or self.column_fill_assignment_count
            or self.unrecognized_fill_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural fill evidence without definitions or targets."""
        return {
            "present": self.present,
            "default_fill_definition_count": self.default_fill_definition_count,
            "cell_fill_assignment_count": self.cell_fill_assignment_count,
            "row_fill_assignment_count": self.row_fill_assignment_count,
            "column_fill_assignment_count": self.column_fill_assignment_count,
            "unrecognized_fill_count": self.unrecognized_fill_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class AlignmentSnapshot:
    """Safe aggregate of effective cell-alignment controls.

    Alignment can change how an unchanged value, warning, or formula appears:
    indentation, wrapping, shrinking, rotation, and horizontal or vertical
    placement can all alter a reviewer's visible surface. Definitions and
    targets can expose report context, so the private signature retains
    canonical effective assignments for comparison while public output exposes
    structural counts only.
    """

    default_alignment_definition_count: int = 0
    cell_alignment_assignment_count: int = 0
    row_alignment_assignment_count: int = 0
    column_alignment_assignment_count: int = 0
    unrecognized_alignment_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.default_alignment_definition_count
            or self.cell_alignment_assignment_count
            or self.row_alignment_assignment_count
            or self.column_alignment_assignment_count
            or self.unrecognized_alignment_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural alignment evidence without definitions or targets."""
        return {
            "present": self.present,
            "default_alignment_definition_count": (
                self.default_alignment_definition_count
            ),
            "cell_alignment_assignment_count": self.cell_alignment_assignment_count,
            "row_alignment_assignment_count": self.row_alignment_assignment_count,
            "column_alignment_assignment_count": (
                self.column_alignment_assignment_count
            ),
            "unrecognized_alignment_count": self.unrecognized_alignment_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class BorderSnapshot:
    """Safe aggregate of effective ordinary cell-border controls.

    A border can change the review and printed surface without changing a
    stored value or formula: it can remove a total-line, conceal a boundary,
    alter a warning's visual weight, or add a diagonal mark. Definitions and
    targets can expose report context, so the private signature retains
    canonical effective assignments for comparison while public output exposes
    structural counts only.
    """

    default_border_definition_count: int = 0
    cell_border_assignment_count: int = 0
    row_border_assignment_count: int = 0
    column_border_assignment_count: int = 0
    unrecognized_border_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    unrecognized_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.default_border_definition_count
            or self.cell_border_assignment_count
            or self.row_border_assignment_count
            or self.column_border_assignment_count
            or self.unrecognized_border_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural border evidence without definitions or targets."""
        return {
            "present": self.present,
            "default_border_definition_count": self.default_border_definition_count,
            "cell_border_assignment_count": self.cell_border_assignment_count,
            "row_border_assignment_count": self.row_border_assignment_count,
            "column_border_assignment_count": self.column_border_assignment_count,
            "unrecognized_border_count": self.unrecognized_border_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WorksheetDimensionSnapshot:
    """Safe aggregate of material worksheet dimension controls.

    Row heights, column widths, AutoFit state, baseline adjustments, and
    border-driven automatic height adjustments can truncate wrapped content,
    reframe a report, or alter its automatic pagination without changing a
    stored cell. Dimension values and row/column targets can reveal report
    structure, so private signatures retain canonical effective declarations
    while public output exposes only structural counts.
    """

    default_row_height_count: int = 0
    default_column_width_count: int = 0
    default_baseline_adjustment_sheet_count: int = 0
    default_border_adjustment_sheet_count: int = 0
    row_height_assignment_count: int = 0
    row_baseline_adjustment_count: int = 0
    row_border_adjustment_count: int = 0
    column_width_assignment_count: int = 0
    best_fit_column_assignment_count: int = 0
    unrecognized_dimension_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    unrecognized_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.default_row_height_count
            or self.default_column_width_count
            or self.default_baseline_adjustment_sheet_count
            or self.default_border_adjustment_sheet_count
            or self.row_height_assignment_count
            or self.row_baseline_adjustment_count
            or self.row_border_adjustment_count
            or self.column_width_assignment_count
            or self.best_fit_column_assignment_count
            or self.unrecognized_dimension_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural dimension evidence without values or targets."""
        return {
            "present": self.present,
            "default_row_height_count": self.default_row_height_count,
            "default_column_width_count": self.default_column_width_count,
            "default_baseline_adjustment_sheet_count": (
                self.default_baseline_adjustment_sheet_count
            ),
            "default_border_adjustment_sheet_count": (
                self.default_border_adjustment_sheet_count
            ),
            "row_height_assignment_count": self.row_height_assignment_count,
            "row_baseline_adjustment_count": self.row_baseline_adjustment_count,
            "row_border_adjustment_count": self.row_border_adjustment_count,
            "column_width_assignment_count": self.column_width_assignment_count,
            "best_fit_column_assignment_count": (
                self.best_fit_column_assignment_count
            ),
            "unrecognized_dimension_count": self.unrecognized_dimension_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WorksheetDisplaySnapshot:
    """Safe aggregate of material worksheet display controls.

    A worksheet view can hide zeroes, gridlines, row and column headers, page
    margins, or outline symbols without changing a cell. It can also show
    formulas, alter gridline colour, switch to a page-oriented view, or
    split/freeze a review surface. Private signatures retain canonical
    non-default declarations only; public output exposes structural counts
    without sheet names, target cells, or raw XML.
    """

    zero_hidden_view_count: int = 0
    formula_view_count: int = 0
    gridlines_hidden_view_count: int = 0
    custom_gridline_color_view_count: int = 0
    headers_hidden_view_count: int = 0
    outline_symbols_hidden_view_count: int = 0
    ruler_hidden_view_count: int = 0
    white_space_hidden_view_count: int = 0
    right_to_left_view_count: int = 0
    non_normal_view_count: int = 0
    split_or_frozen_pane_count: int = 0
    unrecognized_display_control_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.zero_hidden_view_count
            or self.formula_view_count
            or self.gridlines_hidden_view_count
            or self.custom_gridline_color_view_count
            or self.headers_hidden_view_count
            or self.outline_symbols_hidden_view_count
            or self.ruler_hidden_view_count
            or self.white_space_hidden_view_count
            or self.right_to_left_view_count
            or self.non_normal_view_count
            or self.split_or_frozen_pane_count
            or self.unrecognized_display_control_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural display evidence without sheet targets or XML."""
        return {
            "present": self.present,
            "zero_hidden_view_count": self.zero_hidden_view_count,
            "formula_view_count": self.formula_view_count,
            "gridlines_hidden_view_count": self.gridlines_hidden_view_count,
            "custom_gridline_color_view_count": (
                self.custom_gridline_color_view_count
            ),
            "headers_hidden_view_count": self.headers_hidden_view_count,
            "outline_symbols_hidden_view_count": self.outline_symbols_hidden_view_count,
            "ruler_hidden_view_count": self.ruler_hidden_view_count,
            "white_space_hidden_view_count": self.white_space_hidden_view_count,
            "right_to_left_view_count": self.right_to_left_view_count,
            "non_normal_view_count": self.non_normal_view_count,
            "split_or_frozen_pane_count": self.split_or_frozen_pane_count,
            "unrecognized_display_control_count": (
                self.unrecognized_display_control_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WorksheetPrintLayoutSnapshot:
    """Safe aggregate of material worksheet print-layout controls.

    A workbook can print a materially different surface without changing a
    cell: a print area can omit content, titles can repeat context, and page
    setup, margins, headers, footers, or manual breaks can alter the saved
    output. Private signatures retain declarations and header/footer text for
    comparison; public output exposes structural counts only.
    """

    print_area_definition_count: int = 0
    print_title_definition_count: int = 0
    print_gridlines_sheet_count: int = 0
    print_headings_sheet_count: int = 0
    horizontally_centered_print_sheet_count: int = 0
    vertically_centered_print_sheet_count: int = 0
    page_margin_sheet_count: int = 0
    page_setup_sheet_count: int = 0
    header_footer_sheet_count: int = 0
    manual_row_page_break_count: int = 0
    manual_column_page_break_count: int = 0
    unrecognized_print_layout_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    unrecognized_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.print_area_definition_count
            or self.print_title_definition_count
            or self.print_gridlines_sheet_count
            or self.print_headings_sheet_count
            or self.horizontally_centered_print_sheet_count
            or self.vertically_centered_print_sheet_count
            or self.page_margin_sheet_count
            or self.page_setup_sheet_count
            or self.header_footer_sheet_count
            or self.manual_row_page_break_count
            or self.manual_column_page_break_count
            or self.unrecognized_print_layout_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural print evidence without ranges, text, or XML."""
        return {
            "present": self.present,
            "print_area_definition_count": self.print_area_definition_count,
            "print_title_definition_count": self.print_title_definition_count,
            "print_gridlines_sheet_count": self.print_gridlines_sheet_count,
            "print_headings_sheet_count": self.print_headings_sheet_count,
            "horizontally_centered_print_sheet_count": (
                self.horizontally_centered_print_sheet_count
            ),
            "vertically_centered_print_sheet_count": (
                self.vertically_centered_print_sheet_count
            ),
            "page_margin_sheet_count": self.page_margin_sheet_count,
            "page_setup_sheet_count": self.page_setup_sheet_count,
            "header_footer_sheet_count": self.header_footer_sheet_count,
            "manual_row_page_break_count": self.manual_row_page_break_count,
            "manual_column_page_break_count": self.manual_column_page_break_count,
            "unrecognized_print_layout_count": self.unrecognized_print_layout_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WorkbookThemeSnapshot:
    """Safe aggregate of workbook-wide DrawingML theme controls.

    A Theme part can alter the colours, fonts, and effects used by themed
    cells, charts, and drawing objects without changing their local style
    references. Private signatures retain the package state for comparison
    without serialising theme XML, scheme names, colour values, font names,
    image payloads, relationship IDs, or targets into review artifacts.
    """

    theme_part_count: int = 0
    colour_scheme_count: int = 0
    font_scheme_count: int = 0
    format_scheme_count: int = 0
    theme_relationship_count: int = 0
    external_theme_relationship_count: int = 0
    theme_image_part_count: int = 0
    theme_image_relationship_count: int = 0
    external_theme_image_relationship_count: int = 0
    unrecognized_theme_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    image_payload_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.theme_part_count
            or self.colour_scheme_count
            or self.font_scheme_count
            or self.format_scheme_count
            or self.theme_relationship_count
            or self.external_theme_relationship_count
            or self.theme_image_part_count
            or self.theme_image_relationship_count
            or self.external_theme_image_relationship_count
            or self.unrecognized_theme_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate theme evidence without visual-control material."""
        return {
            "present": self.present,
            "theme_part_count": self.theme_part_count,
            "colour_scheme_count": self.colour_scheme_count,
            "font_scheme_count": self.font_scheme_count,
            "format_scheme_count": self.format_scheme_count,
            "theme_relationship_count": self.theme_relationship_count,
            "external_theme_relationship_count": (
                self.external_theme_relationship_count
            ),
            "theme_image_part_count": self.theme_image_part_count,
            "theme_image_relationship_count": self.theme_image_relationship_count,
            "external_theme_image_relationship_count": (
                self.external_theme_image_relationship_count
            ),
            "unrecognized_theme_count": self.unrecognized_theme_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class FormulaCachedResultEntry:
    """One private formula-result cache record keyed to its worksheet cell."""

    sheet: str
    coordinate: str
    result_type: str
    result_signature: str = field(repr=False)

    @property
    def location(self) -> CellKey:
        return (self.sheet, self.coordinate)


@dataclass(frozen=True)
class FormulaCachedResultSnapshot:
    """Safe aggregate of stored results attached to formula cells.

    SpreadsheetML persists a formula's last calculated result beside its formula
    text. The private entries retain only per-result digests so comparison can
    detect a changed displayed result without exposing values or locations.
    """

    formula_cell_count: int = 0
    cached_result_cell_count: int = 0
    missing_cached_result_cell_count: int = 0
    numeric_cached_result_count: int = 0
    string_cached_result_count: int = 0
    boolean_cached_result_count: int = 0
    error_cached_result_count: int = 0
    unrecognized_cached_result_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    entries: tuple[FormulaCachedResultEntry, ...] = field(default=(), repr=False)

    @property
    def present(self) -> bool:
        return bool(self.formula_cell_count or self.unrecognized_cached_result_count)

    def to_dict(self) -> dict[str, Any]:
        """Return structural cache evidence without values or target locations."""
        return {
            "present": self.present,
            "formula_cell_count": self.formula_cell_count,
            "cached_result_cell_count": self.cached_result_cell_count,
            "missing_cached_result_cell_count": self.missing_cached_result_cell_count,
            "numeric_cached_result_count": self.numeric_cached_result_count,
            "string_cached_result_count": self.string_cached_result_count,
            "boolean_cached_result_count": self.boolean_cached_result_count,
            "error_cached_result_count": self.error_cached_result_count,
            "unrecognized_cached_result_count": self.unrecognized_cached_result_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class RichTextRunEntry:
    """One private character-level presentation record keyed to a worksheet cell."""

    sheet: str
    coordinate: str
    text_signature: str = field(repr=False)
    style_sequence_signature: str | None = field(default=None, repr=False)
    style_layout_signature: str | None = field(default=None, repr=False)

    @property
    def location(self) -> CellKey:
        return (self.sheet, self.coordinate)


@dataclass(frozen=True)
class RichTextRunSnapshot:
    """Safe aggregate of character-level string presentation controls.

    SpreadsheetML can store a cell's text as runs with independent font
    properties, outside the cell-level style table. The private entries retain
    only one-way signatures so a formatting-only change can be detected without
    exposing the text, colours, run definitions, or cell locations.
    """

    shared_rich_text_item_count: int = 0
    shared_rich_text_cell_count: int = 0
    shared_rich_text_run_count: int = 0
    inline_rich_text_cell_count: int = 0
    inline_rich_text_run_count: int = 0
    phonetic_run_count: int = 0
    phonetic_property_count: int = 0
    unrecognized_rich_text_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    entries: tuple[RichTextRunEntry, ...] = field(default=(), repr=False)

    @property
    def rich_text_cell_count(self) -> int:
        return self.shared_rich_text_cell_count + self.inline_rich_text_cell_count

    @property
    def rich_text_run_count(self) -> int:
        return self.shared_rich_text_run_count + self.inline_rich_text_run_count

    @property
    def present(self) -> bool:
        return bool(
            self.rich_text_cell_count
            or self.rich_text_run_count
            or self.phonetic_run_count
            or self.phonetic_property_count
            or self.unrecognized_rich_text_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural rich-text evidence without content or targets."""
        return {
            "present": self.present,
            "shared_rich_text_item_count": self.shared_rich_text_item_count,
            "shared_rich_text_cell_count": self.shared_rich_text_cell_count,
            "shared_rich_text_run_count": self.shared_rich_text_run_count,
            "inline_rich_text_cell_count": self.inline_rich_text_cell_count,
            "inline_rich_text_run_count": self.inline_rich_text_run_count,
            "phonetic_run_count": self.phonetic_run_count,
            "phonetic_property_count": self.phonetic_property_count,
            "unrecognized_rich_text_count": self.unrecognized_rich_text_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class CellHyperlinkSnapshot:
    """Safe aggregate of worksheet cell hyperlink controls.

    A SpreadsheetML cell hyperlink can retain a private target, internal
    location, display override, or ScreenTip outside the ordinary cell value.
    Private signatures preserve those declarations and package bindings for
    comparison without serialising link targets, locations, text, or
    relationship identifiers into a profile or change report.
    """

    worksheet_hyperlink_sheet_count: int = 0
    hyperlink_count: int = 0
    hyperlink_with_location_count: int = 0
    hyperlink_with_display_count: int = 0
    hyperlink_with_tooltip_count: int = 0
    binding_relationship_count: int = 0
    external_relationship_count: int = 0
    unrecognized_cell_hyperlink_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(self.hyperlink_count or self.unrecognized_cell_hyperlink_count)

    def to_dict(self) -> dict[str, Any]:
        """Return structural hyperlink evidence without targets or locations."""
        return {
            "present": self.present,
            "worksheet_hyperlink_sheet_count": self.worksheet_hyperlink_sheet_count,
            "hyperlink_count": self.hyperlink_count,
            "hyperlink_with_location_count": self.hyperlink_with_location_count,
            "hyperlink_with_display_count": self.hyperlink_with_display_count,
            "hyperlink_with_tooltip_count": self.hyperlink_with_tooltip_count,
            "binding_relationship_count": self.binding_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "unrecognized_cell_hyperlink_count": (
                self.unrecognized_cell_hyperlink_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WorksheetSparklineSnapshot:
    """Safe aggregate of worksheet sparkline controls.

    Office 2010 x14:sparklineGroup extensions can change a compact visual
    summary without changing any ordinary cell value. Private signatures retain
    source formulas, destination cells, group controls, and colour definitions
    for comparison without serialising them into a profile or change report.
    """

    worksheet_sparkline_sheet_count: int = 0
    sparkline_group_count: int = 0
    sparkline_count: int = 0
    sparkline_with_source_count: int = 0
    group_date_axis_source_count: int = 0
    color_control_count: int = 0
    unrecognized_worksheet_sparkline_count: int = 0
    binding_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.sparkline_group_count
            or self.sparkline_count
            or self.unrecognized_worksheet_sparkline_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural sparkline evidence without formulas or locations."""
        return {
            "present": self.present,
            "worksheet_sparkline_sheet_count": self.worksheet_sparkline_sheet_count,
            "sparkline_group_count": self.sparkline_group_count,
            "sparkline_count": self.sparkline_count,
            "sparkline_with_source_count": self.sparkline_with_source_count,
            "group_date_axis_source_count": self.group_date_axis_source_count,
            "color_control_count": self.color_control_count,
            "unrecognized_worksheet_sparkline_count": (
                self.unrecognized_worksheet_sparkline_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class XmlMappingSnapshot:
    """Safe aggregate of XML-mapped workbook import/export controls.

    SpreadsheetML XML Maps can bind a schema and refresh behavior to table
    columns or individual cells. Private signatures retain schemas, map names,
    XPath expressions, target cells, and connection identities for comparison
    without serialising that operational data into a profile or change report.
    """

    xml_map_part_count: int = 0
    xml_schema_count: int = 0
    xml_map_count: int = 0
    xml_map_data_binding_count: int = 0
    xml_map_file_binding_count: int = 0
    xml_map_connection_binding_count: int = 0
    table_xml_binding_part_count: int = 0
    table_xml_binding_count: int = 0
    single_cell_xml_binding_sheet_count: int = 0
    single_cell_xml_binding_part_count: int = 0
    single_cell_xml_binding_count: int = 0
    single_cell_xml_connection_binding_count: int = 0
    unrecognized_xml_mapping_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    binding_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.xml_map_part_count
            or self.xml_map_count
            or self.table_xml_binding_part_count
            or self.table_xml_binding_count
            or self.single_cell_xml_binding_part_count
            or self.single_cell_xml_binding_count
            or self.unrecognized_xml_mapping_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural XML-map evidence without schema or binding material."""
        return {
            "present": self.present,
            "xml_map_part_count": self.xml_map_part_count,
            "xml_schema_count": self.xml_schema_count,
            "xml_map_count": self.xml_map_count,
            "xml_map_data_binding_count": self.xml_map_data_binding_count,
            "xml_map_file_binding_count": self.xml_map_file_binding_count,
            "xml_map_connection_binding_count": self.xml_map_connection_binding_count,
            "table_xml_binding_part_count": self.table_xml_binding_part_count,
            "table_xml_binding_count": self.table_xml_binding_count,
            "single_cell_xml_binding_sheet_count": (
                self.single_cell_xml_binding_sheet_count
            ),
            "single_cell_xml_binding_part_count": (
                self.single_cell_xml_binding_part_count
            ),
            "single_cell_xml_binding_count": self.single_cell_xml_binding_count,
            "single_cell_xml_connection_binding_count": (
                self.single_cell_xml_connection_binding_count
            ),
            "unrecognized_xml_mapping_count": self.unrecognized_xml_mapping_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class DigitalSignatureSnapshot:
    """Safe aggregate of package and VBA digital-signature controls.

    Office packages can carry XML package signatures, while a VBA project can
    carry one or more binary code-signature payloads. Private signatures retain
    the certificate, signed-reference, and payload material for comparison
    without serialising provenance or cryptographic data into reports.
    """

    package_signature_origin_count: int = 0
    package_xml_signature_count: int = 0
    package_signature_reference_count: int = 0
    package_signature_certificate_count: int = 0
    package_signature_certificate_part_count: int = 0
    package_signature_certificate_relationship_count: int = 0
    vba_project_signature_count: int = 0
    vba_project_signature_relationship_count: int = 0
    unrecognized_digital_signature_count: int = 0
    package_signature_signature: str | None = field(default=None, repr=False)
    vba_signature_payload_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.package_signature_origin_count
            or self.package_xml_signature_count
            or self.package_signature_certificate_part_count
            or self.package_signature_certificate_relationship_count
            or self.vba_project_signature_count
            or self.vba_project_signature_relationship_count
            or self.unrecognized_digital_signature_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate signature evidence without signer or certificate data."""
        return {
            "present": self.present,
            "package_signature_origin_count": self.package_signature_origin_count,
            "package_xml_signature_count": self.package_xml_signature_count,
            "package_signature_reference_count": (
                self.package_signature_reference_count
            ),
            "package_signature_certificate_count": (
                self.package_signature_certificate_count
            ),
            "package_signature_certificate_part_count": (
                self.package_signature_certificate_part_count
            ),
            "package_signature_certificate_relationship_count": (
                self.package_signature_certificate_relationship_count
            ),
            "vba_project_signature_count": self.vba_project_signature_count,
            "vba_project_signature_relationship_count": (
                self.vba_project_signature_relationship_count
            ),
            "unrecognized_digital_signature_count": (
                self.unrecognized_digital_signature_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class RichDataSnapshot:
    """Safe aggregate of Excel rich-data values and relationship controls.

    Rich data types can store entity values, supporting fields, web-image
    associations, and worksheet value-metadata bindings outside ordinary cell
    values. Private signatures retain those values and relationships for
    comparison without serialising entity IDs, provider data, property names,
    URLs, image references, or bound cells into reports.
    """

    rich_value_data_part_count: int = 0
    rich_value_structure_part_count: int = 0
    rich_value_type_part_count: int = 0
    rich_value_array_part_count: int = 0
    supporting_property_bag_part_count: int = 0
    supporting_property_bag_structure_part_count: int = 0
    rich_style_part_count: int = 0
    rich_value_web_image_part_count: int = 0
    rich_value_relationship_part_count: int = 0
    rich_value_count: int = 0
    rich_value_structure_count: int = 0
    linked_entity_structure_count: int = 0
    rich_value_array_count: int = 0
    supporting_property_bag_count: int = 0
    rich_value_metadata_binding_count: int = 0
    rich_value_bound_cell_count: int = 0
    web_image_count: int = 0
    web_image_relationship_count: int = 0
    external_web_image_relationship_count: int = 0
    rich_value_relationship_reference_count: int = 0
    external_rich_value_relationship_count: int = 0
    unrecognized_rich_data_count: int = 0
    definition_signature: str | None = field(default=None, repr=False)
    value_signature: str | None = field(default=None, repr=False)
    metadata_binding_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.rich_value_data_part_count
            or self.rich_value_structure_part_count
            or self.rich_value_type_part_count
            or self.rich_value_array_part_count
            or self.supporting_property_bag_part_count
            or self.supporting_property_bag_structure_part_count
            or self.rich_style_part_count
            or self.rich_value_web_image_part_count
            or self.rich_value_relationship_part_count
            or self.rich_value_count
            or self.rich_value_structure_count
            or self.rich_value_array_count
            or self.supporting_property_bag_count
            or self.rich_value_metadata_binding_count
            or self.rich_value_bound_cell_count
            or self.web_image_count
            or self.web_image_relationship_count
            or self.external_web_image_relationship_count
            or self.rich_value_relationship_reference_count
            or self.external_rich_value_relationship_count
            or self.unrecognized_rich_data_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate rich-data evidence without values or endpoints."""
        return {
            "present": self.present,
            "rich_value_data_part_count": self.rich_value_data_part_count,
            "rich_value_structure_part_count": self.rich_value_structure_part_count,
            "rich_value_type_part_count": self.rich_value_type_part_count,
            "rich_value_array_part_count": self.rich_value_array_part_count,
            "supporting_property_bag_part_count": (
                self.supporting_property_bag_part_count
            ),
            "supporting_property_bag_structure_part_count": (
                self.supporting_property_bag_structure_part_count
            ),
            "rich_style_part_count": self.rich_style_part_count,
            "rich_value_web_image_part_count": (
                self.rich_value_web_image_part_count
            ),
            "rich_value_relationship_part_count": (
                self.rich_value_relationship_part_count
            ),
            "rich_value_count": self.rich_value_count,
            "rich_value_structure_count": self.rich_value_structure_count,
            "linked_entity_structure_count": self.linked_entity_structure_count,
            "rich_value_array_count": self.rich_value_array_count,
            "supporting_property_bag_count": self.supporting_property_bag_count,
            "rich_value_metadata_binding_count": (
                self.rich_value_metadata_binding_count
            ),
            "rich_value_bound_cell_count": self.rich_value_bound_cell_count,
            "web_image_count": self.web_image_count,
            "web_image_relationship_count": self.web_image_relationship_count,
            "external_web_image_relationship_count": (
                self.external_web_image_relationship_count
            ),
            "rich_value_relationship_reference_count": (
                self.rich_value_relationship_reference_count
            ),
            "external_rich_value_relationship_count": (
                self.external_rich_value_relationship_count
            ),
            "unrecognized_rich_data_count": self.unrecognized_rich_data_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class CustomDataStoreSnapshot:
    """Safe aggregate of custom workbook data stores and properties.

    Excel add-ins can retain workbook-specific state in generic custom XML,
    custom binary data, and custom document properties outside ordinary cells.
    Private signatures compare that state without serialising property names,
    values, custom XML, storage identifiers, binary payloads, or targets into
    review artifacts.
    """

    custom_xml_part_count: int = 0
    custom_xml_property_part_count: int = 0
    custom_xml_schema_reference_count: int = 0
    custom_xml_relationship_count: int = 0
    external_custom_xml_relationship_count: int = 0
    custom_data_properties_part_count: int = 0
    custom_data_part_count: int = 0
    document_custom_property_part_count: int = 0
    document_custom_property_count: int = 0
    linked_document_custom_property_count: int = 0
    unrecognized_custom_data_store_count: int = 0
    custom_xml_signature: str | None = field(default=None, repr=False)
    custom_data_signature: str | None = field(default=None, repr=False)
    document_property_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.custom_xml_part_count
            or self.custom_xml_property_part_count
            or self.custom_xml_schema_reference_count
            or self.custom_xml_relationship_count
            or self.external_custom_xml_relationship_count
            or self.custom_data_properties_part_count
            or self.custom_data_part_count
            or self.document_custom_property_part_count
            or self.document_custom_property_count
            or self.linked_document_custom_property_count
            or self.unrecognized_custom_data_store_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return aggregate state-store evidence without stored content."""
        return {
            "present": self.present,
            "custom_xml_part_count": self.custom_xml_part_count,
            "custom_xml_property_part_count": self.custom_xml_property_part_count,
            "custom_xml_schema_reference_count": (
                self.custom_xml_schema_reference_count
            ),
            "custom_xml_relationship_count": self.custom_xml_relationship_count,
            "external_custom_xml_relationship_count": (
                self.external_custom_xml_relationship_count
            ),
            "custom_data_properties_part_count": (
                self.custom_data_properties_part_count
            ),
            "custom_data_part_count": self.custom_data_part_count,
            "document_custom_property_part_count": (
                self.document_custom_property_part_count
            ),
            "document_custom_property_count": self.document_custom_property_count,
            "linked_document_custom_property_count": (
                self.linked_document_custom_property_count
            ),
            "unrecognized_custom_data_store_count": (
                self.unrecognized_custom_data_store_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class LegacyCommentSnapshot:
    """Safe aggregate of legacy Excel note and placeholder controls.

    Legacy comments store author and rich-text note content outside worksheet
    cells. Their VML note shapes separately retain visibility, layout, and
    display controls. Private signatures preserve those declarations for
    comparison without serialising text, locations, identities, or VML.
    """

    worksheet_comment_sheet_count: int = 0
    comment_part_count: int = 0
    comment_author_count: int = 0
    comment_count: int = 0
    comment_with_text_count: int = 0
    rich_text_comment_count: int = 0
    phonetic_comment_count: int = 0
    comment_property_count: int = 0
    threaded_placeholder_count: int = 0
    worksheet_note_drawing_sheet_count: int = 0
    note_vml_drawing_part_count: int = 0
    note_shape_count: int = 0
    visible_note_shape_count: int = 0
    anchored_note_shape_count: int = 0
    binding_relationship_count: int = 0
    external_relationship_count: int = 0
    unrecognized_legacy_comment_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    note_shape_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.comment_part_count
            or self.comment_count
            or self.note_vml_drawing_part_count
            or self.note_shape_count
            or self.unrecognized_legacy_comment_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural note evidence without content or identities."""
        return {
            "present": self.present,
            "worksheet_comment_sheet_count": self.worksheet_comment_sheet_count,
            "comment_part_count": self.comment_part_count,
            "comment_author_count": self.comment_author_count,
            "comment_count": self.comment_count,
            "comment_with_text_count": self.comment_with_text_count,
            "rich_text_comment_count": self.rich_text_comment_count,
            "phonetic_comment_count": self.phonetic_comment_count,
            "comment_property_count": self.comment_property_count,
            "threaded_placeholder_count": self.threaded_placeholder_count,
            "worksheet_note_drawing_sheet_count": (
                self.worksheet_note_drawing_sheet_count
            ),
            "note_vml_drawing_part_count": self.note_vml_drawing_part_count,
            "note_shape_count": self.note_shape_count,
            "visible_note_shape_count": self.visible_note_shape_count,
            "anchored_note_shape_count": self.anchored_note_shape_count,
            "binding_relationship_count": self.binding_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "unrecognized_legacy_comment_count": (
                self.unrecognized_legacy_comment_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class ThreadedCommentSnapshot:
    """Safe aggregate of modern Excel threaded-comment controls.

    Threaded comments, replies, resolution state, mentions, and collaborator
    identity records live in OOXML package parts rather than worksheet cells.
    Private signatures retain their text, locations, timestamps, identity data,
    and graph bindings for comparison without serialising any of that material
    into a profile or change report.
    """

    worksheet_threaded_comment_sheet_count: int = 0
    threaded_comment_part_count: int = 0
    comment_thread_count: int = 0
    comment_count: int = 0
    reply_count: int = 0
    resolved_comment_count: int = 0
    comment_with_text_count: int = 0
    mention_count: int = 0
    mentioned_person_count: int = 0
    person_part_count: int = 0
    person_count: int = 0
    orphan_person_count: int = 0
    binding_relationship_count: int = 0
    external_relationship_count: int = 0
    unrecognized_threaded_comment_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    person_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.threaded_comment_part_count
            or self.person_part_count
            or self.comment_count
            or self.person_count
            or self.unrecognized_threaded_comment_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural comment evidence without content or identities."""
        return {
            "present": self.present,
            "worksheet_threaded_comment_sheet_count": (
                self.worksheet_threaded_comment_sheet_count
            ),
            "threaded_comment_part_count": self.threaded_comment_part_count,
            "comment_thread_count": self.comment_thread_count,
            "comment_count": self.comment_count,
            "reply_count": self.reply_count,
            "resolved_comment_count": self.resolved_comment_count,
            "comment_with_text_count": self.comment_with_text_count,
            "mention_count": self.mention_count,
            "mentioned_person_count": self.mentioned_person_count,
            "person_part_count": self.person_part_count,
            "person_count": self.person_count,
            "orphan_person_count": self.orphan_person_count,
            "binding_relationship_count": self.binding_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "unrecognized_threaded_comment_count": (
                self.unrecognized_threaded_comment_count
            ),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WorksheetDrawingShapeSnapshot:
    """Safe aggregate of non-chart Worksheet DrawingML drawing controls.

    Worksheet DrawingML can carry text boxes, shapes, connectors, and
    non-chart graphic frames outside the cell grid. Their private signatures
    retain anchors, presentation, connector attachments, SmartArt diagram
    material and direct Diagram Data image evidence, text, linked actions, and
    relationship semantics for comparison without exposing any of that material
    in a profile or change report.
    """

    worksheet_drawing_sheet_count: int = 0
    worksheet_drawing_part_count: int = 0
    shape_anchor_count: int = 0
    shape_count: int = 0
    connector_shape_count: int = 0
    connector_attachment_count: int = 0
    group_shape_count: int = 0
    graphic_frame_count: int = 0
    diagram_graphic_frame_count: int = 0
    diagram_data_part_count: int = 0
    diagram_layout_part_count: int = 0
    diagram_quick_style_part_count: int = 0
    diagram_colour_part_count: int = 0
    diagram_drawing_part_count: int = 0
    diagram_image_part_count: int = 0
    fingerprinted_diagram_image_part_count: int = 0
    uninspected_diagram_image_part_count: int = 0
    text_shape_count: int = 0
    text_paragraph_count: int = 0
    text_run_count: int = 0
    macro_assignment_count: int = 0
    text_link_count: int = 0
    hyperlink_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    unrecognized_graphic_frame_count: int = 0
    unrecognized_shape_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    diagram_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.worksheet_drawing_part_count
            or self.shape_count
            or self.connector_shape_count
            or self.group_shape_count
            or self.graphic_frame_count
            or self.unrecognized_graphic_frame_count
            or self.unrecognized_shape_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural DrawingML evidence without text, targets, or anchors."""
        return {
            "present": self.present,
            "worksheet_drawing_sheet_count": self.worksheet_drawing_sheet_count,
            "worksheet_drawing_part_count": self.worksheet_drawing_part_count,
            "shape_anchor_count": self.shape_anchor_count,
            "shape_count": self.shape_count,
            "connector_shape_count": self.connector_shape_count,
            "connector_attachment_count": self.connector_attachment_count,
            "group_shape_count": self.group_shape_count,
            "graphic_frame_count": self.graphic_frame_count,
            "diagram_graphic_frame_count": self.diagram_graphic_frame_count,
            "diagram_data_part_count": self.diagram_data_part_count,
            "diagram_layout_part_count": self.diagram_layout_part_count,
            "diagram_quick_style_part_count": self.diagram_quick_style_part_count,
            "diagram_colour_part_count": self.diagram_colour_part_count,
            "diagram_drawing_part_count": self.diagram_drawing_part_count,
            "diagram_image_part_count": self.diagram_image_part_count,
            "fingerprinted_diagram_image_part_count": (
                self.fingerprinted_diagram_image_part_count
            ),
            "uninspected_diagram_image_part_count": (
                self.uninspected_diagram_image_part_count
            ),
            "text_shape_count": self.text_shape_count,
            "text_paragraph_count": self.text_paragraph_count,
            "text_run_count": self.text_run_count,
            "macro_assignment_count": self.macro_assignment_count,
            "text_link_count": self.text_link_count,
            "hyperlink_count": self.hyperlink_count,
            "related_relationship_count": self.related_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "unrecognized_graphic_frame_count": (
                self.unrecognized_graphic_frame_count
            ),
            "unrecognized_shape_count": self.unrecognized_shape_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WorksheetImageSnapshot:
    """Safe aggregate of native worksheet image controls.

    Floating DrawingML pictures, worksheet backgrounds, and header/footer VML
    watermark images can alter a report without changing a cell. Private
    signatures retain their bindings, visual declarations, relationships, and
    bounded media fingerprints while public output exposes counts only.
    """

    worksheet_image_sheet_count: int = 0
    anchored_picture_count: int = 0
    anchored_picture_anchor_count: int = 0
    worksheet_background_image_count: int = 0
    header_footer_image_count: int = 0
    image_part_count: int = 0
    fingerprinted_image_part_count: int = 0
    uninspected_image_part_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    unrecognized_image_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    definition_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)
    image_payload_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.anchored_picture_count
            or self.worksheet_background_image_count
            or self.header_footer_image_count
            or self.image_part_count
            or self.unrecognized_image_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return structural image evidence without visual or package contents."""
        return {
            "present": self.present,
            "worksheet_image_sheet_count": self.worksheet_image_sheet_count,
            "anchored_picture_count": self.anchored_picture_count,
            "anchored_picture_anchor_count": self.anchored_picture_anchor_count,
            "worksheet_background_image_count": self.worksheet_background_image_count,
            "header_footer_image_count": self.header_footer_image_count,
            "image_part_count": self.image_part_count,
            "fingerprinted_image_part_count": self.fingerprinted_image_part_count,
            "uninspected_image_part_count": self.uninspected_image_part_count,
            "related_relationship_count": self.related_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "unrecognized_image_count": self.unrecognized_image_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class WorksheetEmbeddedControlSnapshot:
    """Safe aggregate of worksheet control and OLE package data.

    Worksheet controls can bind persisted ActiveX state, modern and legacy form-
    control formulas, OLE payloads, and external linked objects outside cells and
    the VBA project. Private signatures retain all control definitions,
    relationships, and safely fingerprinted direct payloads for comparison;
    ``to_dict`` deliberately exposes only structural counts.
    """

    control_sheet_count: int = 0
    worksheet_control_count: int = 0
    active_x_part_count: int = 0
    active_x_binary_reference_count: int = 0
    form_control_property_part_count: int = 0
    legacy_vml_drawing_part_count: int = 0
    legacy_vml_control_count: int = 0
    legacy_vml_macro_assignment_count: int = 0
    legacy_vml_cell_link_count: int = 0
    legacy_vml_source_range_count: int = 0
    legacy_vml_camera_source_range_count: int = 0
    control_macro_assignment_count: int = 0
    control_cell_link_count: int = 0
    control_source_range_count: int = 0
    form_control_formula_binding_count: int = 0
    ole_object_count: int = 0
    linked_ole_object_count: int = 0
    auto_load_ole_object_count: int = 0
    auto_update_ole_object_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    internal_related_part_count: int = 0
    fingerprinted_related_part_count: int = 0
    uninspected_related_part_count: int = 0
    unrecognized_part_count: int = 0
    declaration_signature: str | None = field(default=None, repr=False)
    control_definition_signature: str | None = field(default=None, repr=False)
    active_x_definition_signature: str | None = field(default=None, repr=False)
    form_control_property_signature: str | None = field(default=None, repr=False)
    legacy_vml_definition_signature: str | None = field(default=None, repr=False)
    legacy_vml_relationship_signature: str | None = field(default=None, repr=False)
    relationship_signature: str | None = field(default=None, repr=False)
    related_part_payload_signature: str | None = field(default=None, repr=False)

    @property
    def present(self) -> bool:
        return bool(
            self.control_sheet_count
            or self.worksheet_control_count
            or self.active_x_part_count
            or self.form_control_property_part_count
            or self.legacy_vml_drawing_part_count
            or self.legacy_vml_control_count
            or self.ole_object_count
            or self.unrecognized_part_count
        )

    def to_dict(self) -> dict[str, Any]:
        """Return safe control inventory without identities, formulas, or payloads."""
        return {
            "present": self.present,
            "control_sheet_count": self.control_sheet_count,
            "worksheet_control_count": self.worksheet_control_count,
            "active_x_part_count": self.active_x_part_count,
            "active_x_binary_reference_count": self.active_x_binary_reference_count,
            "form_control_property_part_count": self.form_control_property_part_count,
            "legacy_vml_drawing_part_count": self.legacy_vml_drawing_part_count,
            "legacy_vml_control_count": self.legacy_vml_control_count,
            "legacy_vml_macro_assignment_count": (
                self.legacy_vml_macro_assignment_count
            ),
            "legacy_vml_cell_link_count": self.legacy_vml_cell_link_count,
            "legacy_vml_source_range_count": self.legacy_vml_source_range_count,
            "legacy_vml_camera_source_range_count": (
                self.legacy_vml_camera_source_range_count
            ),
            "control_macro_assignment_count": self.control_macro_assignment_count,
            "control_cell_link_count": self.control_cell_link_count,
            "control_source_range_count": self.control_source_range_count,
            "form_control_formula_binding_count": self.form_control_formula_binding_count,
            "ole_object_count": self.ole_object_count,
            "linked_ole_object_count": self.linked_ole_object_count,
            "auto_load_ole_object_count": self.auto_load_ole_object_count,
            "auto_update_ole_object_count": self.auto_update_ole_object_count,
            "related_relationship_count": self.related_relationship_count,
            "external_relationship_count": self.external_relationship_count,
            "internal_related_part_count": self.internal_related_part_count,
            "fingerprinted_related_part_count": self.fingerprinted_related_part_count,
            "uninspected_related_part_count": self.uninspected_related_part_count,
            "unrecognized_part_count": self.unrecognized_part_count,
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class PowerQueryPermissionControlsSnapshot:
    """Safe aggregate controls from Data Mashup permission documents."""

    payload_count: int = 0
    parsed_count: int = 0
    firewall_enabled_count: int = 0
    future_packages_allowed_count: int = 0
    workbook_group_type_count: int = 0
    opaque_metadata: ExternalDataOpaqueMetadataSnapshot = field(
        default_factory=ExternalDataOpaqueMetadataSnapshot
    )
    signature: str | None = field(default=None, repr=False)

    def to_dict(self) -> dict[str, Any]:
        """Return permission-control counts without query or user identity material."""
        return {
            "payload_count": self.payload_count,
            "parsed_count": self.parsed_count,
            "firewall_enabled_count": self.firewall_enabled_count,
            "future_packages_allowed_count": self.future_packages_allowed_count,
            "workbook_group_type_count": self.workbook_group_type_count,
            "opaque_metadata": self.opaque_metadata.to_dict(),
        }


@dataclass(frozen=True)
class PowerQuerySnapshot:
    """Safe inventory of Data Mashup query definitions stored in custom XML."""

    mashup_count: int = 0
    parsed_mashup_count: int = 0
    formula_document_count: int = 0
    package_part_count: int = 0
    embedded_content_part_count: int = 0
    metadata_document_count: int = 0
    metadata_item_count: int = 0
    permission_controls: PowerQueryPermissionControlsSnapshot = field(
        default_factory=PowerQueryPermissionControlsSnapshot
    )
    permission_binding_count: int = 0
    formula_signature: str | None = field(default=None, repr=False)
    package_configuration_signature: str | None = field(default=None, repr=False)
    metadata_identity_signature: str | None = field(default=None, repr=False)
    metadata_control_signature: str | None = field(default=None, repr=False)
    opaque_metadata: ExternalDataOpaqueMetadataSnapshot = field(
        default_factory=ExternalDataOpaqueMetadataSnapshot
    )

    @property
    def present(self) -> bool:
        return self.mashup_count > 0

    def to_dict(self) -> dict[str, Any]:
        """Return structural Power Query evidence without M text or source material."""
        return {
            "present": self.present,
            "mashup_count": self.mashup_count,
            "parsed_mashup_count": self.parsed_mashup_count,
            "formula_document_count": self.formula_document_count,
            "package_part_count": self.package_part_count,
            "embedded_content_part_count": self.embedded_content_part_count,
            "metadata_document_count": self.metadata_document_count,
            "metadata_item_count": self.metadata_item_count,
            "permission_controls": self.permission_controls.to_dict(),
            "permission_binding_count": self.permission_binding_count,
            "opaque_metadata": self.opaque_metadata.to_dict(),
        }

    def profile_dict(self) -> dict[str, Any]:
        return self.to_dict()


@dataclass(frozen=True)
class RangeDependency:
    """A range used by a formula, stored without expanding large Excel ranges."""

    source_sheet: str
    min_column: int
    min_row: int
    max_column: int
    max_row: int
    dependent: CellKey

    def contains(self, location: CellKey) -> bool:
        sheet, coordinate = location
        if sheet != self.source_sheet:
            return False
        from openpyxl.utils.cell import column_index_from_string, coordinate_to_tuple

        row, column = coordinate_to_tuple(coordinate)
        if isinstance(column, str):  # defensive for older openpyxl versions
            column = column_index_from_string(column)
        return (
            self.min_column <= column <= self.max_column
            and self.min_row <= row <= self.max_row
        )

    def intersects(self, other: ArrayFormulaRange) -> bool:
        """Return whether this dependency reads any cell in an array output range."""
        return (
            self.source_sheet == other.sheet
            and self.min_column <= other.max_column
            and other.min_column <= self.max_column
            and self.min_row <= other.max_row
            and other.min_row <= self.max_row
        )


@dataclass(frozen=True)
class ArrayFormulaRange:
    """The declared result range from one array formula.

    Excel stores the formula only at ``anchor``. The remaining cells in ``ref``
    are result members, not independent formulas. For legacy CSE formulas the
    range is fixed; for dynamic arrays it is the currently observed extent.
    FormulaFence preserves either form compactly rather than creating one graph
    node per result cell.
    """

    sheet: str
    anchor: str
    ref: str
    min_column: int
    min_row: int
    max_column: int
    max_row: int

    @property
    def location(self) -> CellKey:
        return (self.sheet, self.anchor)

    @property
    def output_cell_count(self) -> int:
        return (self.max_column - self.min_column + 1) * (
            self.max_row - self.min_row + 1
        )

    @property
    def has_multiple_outputs(self) -> bool:
        return self.output_cell_count > 1

    def contains(self, location: CellKey) -> bool:
        sheet, coordinate = location
        if sheet != self.sheet:
            return False
        from openpyxl.utils.cell import column_index_from_string, coordinate_to_tuple

        row, column = coordinate_to_tuple(coordinate)
        if isinstance(column, str):  # defensive for older openpyxl versions
            column = column_index_from_string(column)
        return (
            self.min_column <= column <= self.max_column
            and self.min_row <= row <= self.max_row
        )

    def intersects_non_anchor(self, dependency: RangeDependency) -> bool:
        """Return whether a dependency intersects any result member after anchor."""
        if not dependency.intersects(self):
            return False
        min_column = max(self.min_column, dependency.min_column)
        max_column = min(self.max_column, dependency.max_column)
        min_row = max(self.min_row, dependency.min_row)
        max_row = min(self.max_row, dependency.max_row)
        if min_column != max_column or min_row != max_row:
            return True
        return (min_column, min_row) != (self.min_column, self.min_row)

    def to_dict(self) -> dict[str, Any]:
        return {
            "anchor": display_location(self.location),
            "ref": display_location((self.sheet, self.ref)),
            "output_cell_count": self.output_cell_count,
        }


@dataclass(frozen=True)
class DynamicArrayOutputReference:
    """One formula reading a non-anchor cell of an observed dynamic spill."""

    anchor: CellKey
    observed_ref: str

    def to_dict(self) -> dict[str, str]:
        sheet, _ = self.anchor
        return {
            "anchor": display_location(self.anchor) or "",
            "observed_range": display_location((sheet, self.observed_ref)) or "",
        }


@dataclass(frozen=True)
class Finding:
    rule_id: str
    severity: str
    message: str
    location: CellKey | None = None
    details: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        result = {
            "rule_id": self.rule_id,
            "severity": self.severity,
            "message": self.message,
            "location": display_location(self.location),
        }
        if self.details:
            result["details"] = self.details
        return result


@dataclass(frozen=True)
class Change:
    kind: str
    location: CellKey | None
    severity: str
    before: CellSnapshot | None = None
    after: CellSnapshot | None = None
    impact_count: int = 0
    impacted_cells: tuple[CellKey, ...] = ()
    details: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "kind": self.kind,
            "severity": self.severity,
            "location": display_location(self.location),
            "before": self.before.to_dict() if self.before else None,
            "after": self.after.to_dict() if self.after else None,
            "impact_count": self.impact_count,
            "impacted_cells": [display_location(cell) for cell in self.impacted_cells],
            "details": self.details,
        }


@dataclass
class WorkbookSnapshot:
    """Workbook semantics plus the dependency indexes used for impact analysis."""

    path: Path
    sha256: str
    file_type: str
    sheets: dict[str, SheetSnapshot]
    cells: dict[CellKey, CellSnapshot]
    reverse_dependencies: dict[CellKey, set[CellKey]]
    range_dependencies: list[RangeDependency]
    external_references: set[CellKey]
    broken_references: set[CellKey]
    defined_names: dict[str, str]
    macro_hash: str | None
    calculation_settings: dict[str, Any]
    parser_warnings: tuple[str, ...]
    unresolved_reference_tokens: dict[CellKey, tuple[str, ...]] = field(default_factory=dict)
    dynamic_reference_functions: dict[CellKey, tuple[str, ...]] = field(default_factory=dict)
    tables: dict[str, TableSnapshot] = field(default_factory=dict)
    data_validations: tuple[DataValidationSnapshot, ...] = ()
    conditional_formatting: tuple[ConditionalFormattingSnapshot, ...] = ()
    conditional_formatting_extensions: tuple[ConditionalFormattingExtensionSnapshot, ...] = ()
    workbook_protection: WorkbookProtectionSnapshot | None = None
    sheet_protections: tuple[SheetProtectionSnapshot, ...] = ()
    protected_ranges: tuple[ProtectedRangeSnapshot, ...] = ()
    cell_protection_default: CellProtectionDefaultSnapshot | None = None
    cell_protection_assignments: tuple[CellProtectionAssignmentSnapshot, ...] = ()
    external_data_refresh_settings: ExternalDataRefreshSettingsSnapshot = field(
        default_factory=ExternalDataRefreshSettingsSnapshot
    )
    external_data_connections: tuple[ExternalDataConnectionSnapshot, ...] = ()
    query_table_refresh_controls: tuple[QueryTableRefreshSnapshot, ...] = ()
    pivot_cache_refresh_controls: tuple[PivotCacheRefreshSnapshot, ...] = ()
    external_link_packages: ExternalLinkPackageSnapshot = field(
        default_factory=ExternalLinkPackageSnapshot
    )
    external_relationships: ExternalRelationshipSnapshot = field(
        default_factory=ExternalRelationshipSnapshot
    )
    formula_external_actions: FormulaExternalActionSnapshot = field(
        default_factory=FormulaExternalActionSnapshot
    )
    formula_dde_links: FormulaDdeLinkSnapshot = field(
        default_factory=FormulaDdeLinkSnapshot
    )
    python_in_excel: PythonInExcelSnapshot = field(
        default_factory=PythonInExcelSnapshot
    )
    office_custom_functions: OfficeCustomFunctionSnapshot = field(
        default_factory=OfficeCustomFunctionSnapshot
    )
    unqualified_runtime_functions: UnqualifiedRuntimeFunctionSnapshot = field(
        default_factory=UnqualifiedRuntimeFunctionSnapshot
    )
    worksheet_code_resource_registrations: WorksheetCodeResourceRegistrationSnapshot = (
        field(default_factory=WorksheetCodeResourceRegistrationSnapshot)
    )
    formula_defined_xlm_registrations: FormulaDefinedXlmRegistrationSnapshot = (
        field(default_factory=FormulaDefinedXlmRegistrationSnapshot)
    )
    formula_defined_xlm_evaluations: FormulaDefinedXlmEvaluationSnapshot = (
        field(default_factory=FormulaDefinedXlmEvaluationSnapshot)
    )
    formula_defined_xlm_actions: FormulaDefinedXlmActionSnapshot = field(
        default_factory=FormulaDefinedXlmActionSnapshot
    )
    formula_defined_xlm_get_cell_calls: FormulaDefinedXlmGetCellSnapshot = field(
        default_factory=FormulaDefinedXlmGetCellSnapshot
    )
    formula_defined_xlm_environment_information_calls: (
        FormulaDefinedXlmEnvironmentInformationSnapshot
    ) = field(default_factory=FormulaDefinedXlmEnvironmentInformationSnapshot)
    formula_environment_information_calls: FormulaEnvironmentInformationSnapshot = field(
        default_factory=FormulaEnvironmentInformationSnapshot
    )
    xlm_macro_sheets: XlmMacroSheetSnapshot = field(
        default_factory=XlmMacroSheetSnapshot
    )
    xlm_automatic_macro_bindings: XlmAutomaticMacroBindingSnapshot = field(
        default_factory=XlmAutomaticMacroBindingSnapshot
    )
    ribbon_customization: RibbonCustomizationSnapshot = field(
        default_factory=RibbonCustomizationSnapshot
    )
    office_web_addins: OfficeWebAddinSnapshot = field(
        default_factory=OfficeWebAddinSnapshot
    )
    pivot_table_definitions: PivotTableDefinitionSnapshot = field(
        default_factory=PivotTableDefinitionSnapshot
    )
    slicer_timeline_caches: SlicerTimelineCacheSnapshot = field(
        default_factory=SlicerTimelineCacheSnapshot
    )
    power_pivot_data_model: PowerPivotDataModelSnapshot = field(
        default_factory=PowerPivotDataModelSnapshot
    )
    what_if_data_tables: WhatIfDataTableSnapshot = field(
        default_factory=WhatIfDataTableSnapshot
    )
    scenario_manager: ScenarioManagerSnapshot = field(
        default_factory=ScenarioManagerSnapshot
    )
    filter_visibility_controls: FilterVisibilitySnapshot = field(
        default_factory=FilterVisibilitySnapshot
    )
    ignored_error_controls: IgnoredErrorSnapshot = field(
        default_factory=IgnoredErrorSnapshot
    )
    named_sheet_views: NamedSheetViewSnapshot = field(
        default_factory=NamedSheetViewSnapshot
    )
    custom_workbook_views: CustomWorkbookViewSnapshot = field(
        default_factory=CustomWorkbookViewSnapshot
    )
    table_style_controls: TableStyleControlsSnapshot = field(
        default_factory=TableStyleControlsSnapshot
    )
    shared_workbook_revisions: SharedWorkbookRevisionSnapshot = field(
        default_factory=SharedWorkbookRevisionSnapshot
    )
    number_format_controls: NumberFormatSnapshot = field(
        default_factory=NumberFormatSnapshot
    )
    font_controls: FontSnapshot = field(default_factory=FontSnapshot)
    fill_controls: FillSnapshot = field(default_factory=FillSnapshot)
    alignment_controls: AlignmentSnapshot = field(
        default_factory=AlignmentSnapshot
    )
    border_controls: BorderSnapshot = field(default_factory=BorderSnapshot)
    worksheet_dimension_controls: WorksheetDimensionSnapshot = field(
        default_factory=WorksheetDimensionSnapshot
    )
    worksheet_display_controls: WorksheetDisplaySnapshot = field(
        default_factory=WorksheetDisplaySnapshot
    )
    worksheet_print_layout_controls: WorksheetPrintLayoutSnapshot = field(
        default_factory=WorksheetPrintLayoutSnapshot
    )
    workbook_theme: WorkbookThemeSnapshot = field(
        default_factory=WorkbookThemeSnapshot
    )
    formula_cached_results: FormulaCachedResultSnapshot = field(
        default_factory=FormulaCachedResultSnapshot
    )
    rich_text_runs: RichTextRunSnapshot = field(default_factory=RichTextRunSnapshot)
    cell_hyperlinks: CellHyperlinkSnapshot = field(
        default_factory=CellHyperlinkSnapshot
    )
    worksheet_sparklines: WorksheetSparklineSnapshot = field(
        default_factory=WorksheetSparklineSnapshot
    )
    xml_mapping_controls: XmlMappingSnapshot = field(
        default_factory=XmlMappingSnapshot
    )
    digital_signatures: DigitalSignatureSnapshot = field(
        default_factory=DigitalSignatureSnapshot
    )
    rich_data: RichDataSnapshot = field(default_factory=RichDataSnapshot)
    custom_data_stores: CustomDataStoreSnapshot = field(
        default_factory=CustomDataStoreSnapshot
    )
    legacy_comments: LegacyCommentSnapshot = field(
        default_factory=LegacyCommentSnapshot
    )
    threaded_comments: ThreadedCommentSnapshot = field(
        default_factory=ThreadedCommentSnapshot
    )
    worksheet_drawing_shapes: WorksheetDrawingShapeSnapshot = field(
        default_factory=WorksheetDrawingShapeSnapshot
    )
    worksheet_images: WorksheetImageSnapshot = field(
        default_factory=WorksheetImageSnapshot
    )
    chart_definitions: ChartDefinitionSnapshot = field(
        default_factory=ChartDefinitionSnapshot
    )
    worksheet_embedded_controls: WorksheetEmbeddedControlSnapshot = field(
        default_factory=WorksheetEmbeddedControlSnapshot
    )
    power_query: PowerQuerySnapshot = field(default_factory=PowerQuerySnapshot)
    # ``SHEET`` and ``SHEETS`` observe all workbook tabs, including chart,
    # macro, and dialog sheets that are intentionally outside the cell reader.
    # Keep this raw OOXML catalog private: the formula-information profile only
    # needs aggregate counts, while ordinary sheet inventory remains separate.
    workbook_tab_order: tuple[str, ...] = field(default_factory=tuple, repr=False)
    workbook_tab_order_complete: bool = False
    # The raw subset whose relationships identify ordinary worksheets. This is
    # private because external 3-D portfolio resolution uses it only to prove
    # the workbook reader retained every potentially referenced worksheet.
    worksheet_tab_order: tuple[str, ...] = field(default_factory=tuple, repr=False)
    worksheet_tab_order_complete: bool = False
    sheet_order: tuple[str, ...] = ()
    three_d_reference_tokens: dict[CellKey, tuple[str, ...]] = field(default_factory=dict)
    spill_reference_tokens: dict[CellKey, tuple[str, ...]] = field(default_factory=dict)
    implicit_intersection_tokens: dict[CellKey, tuple[str, ...]] = field(
        default_factory=dict
    )
    legacy_array_formula_ranges: tuple[ArrayFormulaRange, ...] = ()
    dynamic_array_formula_cells: set[CellKey] = field(default_factory=set)
    dynamic_array_formula_ranges: tuple[ArrayFormulaRange, ...] = ()
    dynamic_array_output_references: dict[
        CellKey, tuple[DynamicArrayOutputReference, ...]
    ] = field(default_factory=dict)
    # Formula spellings for external workbooks can disclose an author's local
    # filesystem or network layout. Keep them private; portfolio comparison
    # resolves only a narrow, relative subset against supplied input roots.
    external_workbook_references: dict[
        CellKey, tuple[ExternalWorkbookReference, ...]
    ] = field(default_factory=dict, repr=False)
    # Static external 3-D A1 spans stay separate from single-sheet ranges.
    # Their private endpoints are expanded only against the exact worksheet
    # order of an already inspected candidate source workbook.
    external_workbook_three_d_references: dict[
        CellKey, tuple[ExternalWorkbookThreeDReference, ...]
    ] = field(default_factory=dict, repr=False)
    # Direct external workbook-defined names are retained separately from
    # external A1 ranges. Both their source spelling and the name itself stay
    # private, and portfolio comparison resolves only statically expanded,
    # candidate-name indexes below.
    external_workbook_defined_name_references: dict[
        CellKey, tuple[ExternalWorkbookDefinedNameReference, ...]
    ] = field(default_factory=dict, repr=False)
    static_global_defined_name_references: dict[
        str, tuple[ParsedReference, ...]
    ] = field(default_factory=dict, repr=False)
    static_local_defined_name_references: dict[
        str, dict[str, tuple[ParsedReference, ...]]
    ] = field(default_factory=dict, repr=False)
    unclassified_array_formula_cells: set[CellKey] = field(default_factory=set)
    array_formula_output_dependents: dict[CellKey, set[CellKey]] = field(
        default_factory=dict
    )
    tokenization_failure_cells: set[CellKey] = field(default_factory=set)

    def direct_dependents(self, location: CellKey) -> set[CellKey]:
        dependents = set(self.reverse_dependencies.get(location, set()))
        dependents.update(self.array_formula_output_dependents.get(location, set()))
        for dependency in self.range_dependencies:
            if dependency.contains(location):
                dependents.add(dependency.dependent)
        return dependents

    def summary(self) -> dict[str, Any]:
        return {
            "path": str(self.path),
            "sha256": self.sha256,
            "file_type": self.file_type,
            "sheet_count": len(self.sheets),
            "nonempty_cells": len(self.cells),
            "formula_cells": sum(1 for cell in self.cells.values() if cell.is_formula),
            "formula_cached_result_cell_count": (
                self.formula_cached_results.cached_result_cell_count
            ),
            "formula_missing_cached_result_cell_count": (
                self.formula_cached_results.missing_cached_result_cell_count
            ),
            "has_formula_cached_results": self.formula_cached_results.present,
            "rich_text_cell_count": self.rich_text_runs.rich_text_cell_count,
            "rich_text_run_count": self.rich_text_runs.rich_text_run_count,
            "phonetic_run_count": self.rich_text_runs.phonetic_run_count,
            "has_rich_text_runs": self.rich_text_runs.present,
            "cell_hyperlink_count": self.cell_hyperlinks.hyperlink_count,
            "cell_hyperlink_external_relationship_count": (
                self.cell_hyperlinks.external_relationship_count
            ),
            "has_cell_hyperlinks": self.cell_hyperlinks.present,
            "worksheet_sparkline_group_count": (
                self.worksheet_sparklines.sparkline_group_count
            ),
            "worksheet_sparkline_count": self.worksheet_sparklines.sparkline_count,
            "has_worksheet_sparklines": self.worksheet_sparklines.present,
            "xml_map_count": self.xml_mapping_controls.xml_map_count,
            "xml_map_table_binding_count": (
                self.xml_mapping_controls.table_xml_binding_count
            ),
            "xml_map_single_cell_binding_count": (
                self.xml_mapping_controls.single_cell_xml_binding_count
            ),
            "has_xml_mapping_controls": self.xml_mapping_controls.present,
            "package_xml_signature_count": (
                self.digital_signatures.package_xml_signature_count
            ),
            "package_signature_certificate_part_count": (
                self.digital_signatures.package_signature_certificate_part_count
            ),
            "vba_project_signature_count": (
                self.digital_signatures.vba_project_signature_count
            ),
            "has_digital_signatures": self.digital_signatures.present,
            "rich_value_count": self.rich_data.rich_value_count,
            "rich_value_bound_cell_count": self.rich_data.rich_value_bound_cell_count,
            "rich_data_external_relationship_count": (
                self.rich_data.external_web_image_relationship_count
                + self.rich_data.external_rich_value_relationship_count
            ),
            "has_rich_data": self.rich_data.present,
            "custom_xml_part_count": self.custom_data_stores.custom_xml_part_count,
            "custom_data_part_count": self.custom_data_stores.custom_data_part_count,
            "document_custom_property_count": (
                self.custom_data_stores.document_custom_property_count
            ),
            "has_custom_data_stores": self.custom_data_stores.present,
            "legacy_comment_count": self.legacy_comments.comment_count,
            "legacy_comment_author_count": self.legacy_comments.comment_author_count,
            "legacy_comment_note_shape_count": self.legacy_comments.note_shape_count,
            "has_legacy_comments": self.legacy_comments.present,
            "threaded_comment_count": self.threaded_comments.comment_count,
            "threaded_comment_thread_count": self.threaded_comments.comment_thread_count,
            "threaded_comment_reply_count": self.threaded_comments.reply_count,
            "threaded_comment_person_count": self.threaded_comments.person_count,
            "has_threaded_comments": self.threaded_comments.present,
            "worksheet_drawing_shape_count": self.worksheet_drawing_shapes.shape_count,
            "worksheet_drawing_connector_shape_count": (
                self.worksheet_drawing_shapes.connector_shape_count
            ),
            "worksheet_drawing_connector_attachment_count": (
                self.worksheet_drawing_shapes.connector_attachment_count
            ),
            "worksheet_drawing_text_shape_count": (
                self.worksheet_drawing_shapes.text_shape_count
            ),
            "worksheet_drawing_graphic_frame_count": (
                self.worksheet_drawing_shapes.graphic_frame_count
            ),
            "worksheet_drawing_diagram_frame_count": (
                self.worksheet_drawing_shapes.diagram_graphic_frame_count
            ),
            "has_worksheet_drawing_shapes": self.worksheet_drawing_shapes.present,
            "worksheet_anchored_picture_count": (
                self.worksheet_images.anchored_picture_count
            ),
            "worksheet_background_image_count": (
                self.worksheet_images.worksheet_background_image_count
            ),
            "worksheet_header_footer_image_count": (
                self.worksheet_images.header_footer_image_count
            ),
            "has_worksheet_images": self.worksheet_images.present,
            "what_if_data_table_count": self.what_if_data_tables.data_table_count,
            "what_if_data_table_output_cell_count": (
                self.what_if_data_tables.declared_output_cell_count
            ),
            "has_what_if_data_tables": self.what_if_data_tables.present,
            "scenario_manager_sheet_count": self.scenario_manager.scenario_sheet_count,
            "scenario_manager_scenario_count": self.scenario_manager.scenario_count,
            "scenario_manager_input_cell_count": self.scenario_manager.input_cell_count,
            "has_scenario_manager": self.scenario_manager.present,
            "filter_visibility_auto_filter_count": (
                self.filter_visibility_controls.worksheet_auto_filter_count
                + self.filter_visibility_controls.table_auto_filter_count
            ),
            "filter_visibility_hidden_row_count": (
                self.filter_visibility_controls.hidden_row_count
            ),
            "filter_visibility_zero_height_row_count": (
                self.filter_visibility_controls.zero_height_row_count
            ),
            "filter_visibility_hidden_column_count": (
                self.filter_visibility_controls.hidden_column_count
            ),
            "filter_visibility_zero_width_column_count": (
                self.filter_visibility_controls.zero_width_column_count
            ),
            "filter_visibility_default_zero_height_sheet_count": (
                self.filter_visibility_controls.default_zero_height_sheet_count
            ),
            "filter_visibility_default_zero_width_sheet_count": (
                self.filter_visibility_controls.default_zero_width_sheet_count
            ),
            "has_filter_visibility_controls": self.filter_visibility_controls.present,
            "ignored_error_rule_count": self.ignored_error_controls.ignored_error_rule_count,
            "ignored_error_target_range_count": self.ignored_error_controls.target_range_count,
            "has_ignored_error_controls": self.ignored_error_controls.present,
            "named_sheet_view_count": self.named_sheet_views.named_sheet_view_count,
            "named_sheet_view_filter_count": self.named_sheet_views.named_filter_count,
            "has_named_sheet_views": self.named_sheet_views.present,
            "custom_workbook_view_count": (
                self.custom_workbook_views.custom_workbook_view_count
            ),
            "custom_sheet_view_count": (
                self.custom_workbook_views.custom_sheet_view_count
            ),
            "custom_view_sheet_count": (
                self.custom_workbook_views.custom_view_sheet_count
            ),
            "has_custom_workbook_views": self.custom_workbook_views.present,
            "table_style_info_count": (
                self.table_style_controls.table_style_info_count
            ),
            "custom_table_style_count": (
                self.table_style_controls.custom_table_style_count
            ),
            "table_direct_dxf_assignment_count": (
                self.table_style_controls.table_direct_dxf_assignment_count
            ),
            "table_named_cell_style_assignment_count": (
                self.table_style_controls.table_named_cell_style_assignment_count
            ),
            "has_table_style_controls": self.table_style_controls.present,
            "revision_header_count": (
                self.shared_workbook_revisions.revision_header_count
            ),
            "revision_log_entry_count": (
                self.shared_workbook_revisions.revision_log_entry_count
            ),
            "has_shared_workbook_revisions": (
                self.shared_workbook_revisions.present
            ),
            "number_format_assignment_count": (
                self.number_format_controls.default_format_override_count
                + self.number_format_controls.cell_format_assignment_count
                + self.number_format_controls.row_format_assignment_count
                + self.number_format_controls.column_format_assignment_count
            ),
            "number_format_custom_assignment_count": (
                self.number_format_controls.custom_format_assignment_count
            ),
            "has_number_format_controls": self.number_format_controls.present,
            "font_assignment_count": (
                self.font_controls.default_font_definition_count
                + self.font_controls.cell_font_assignment_count
                + self.font_controls.row_font_assignment_count
                + self.font_controls.column_font_assignment_count
            ),
            "has_font_controls": self.font_controls.present,
            "fill_assignment_count": (
                self.fill_controls.default_fill_definition_count
                + self.fill_controls.cell_fill_assignment_count
                + self.fill_controls.row_fill_assignment_count
                + self.fill_controls.column_fill_assignment_count
            ),
            "has_fill_controls": self.fill_controls.present,
            "alignment_assignment_count": (
                self.alignment_controls.cell_alignment_assignment_count
                + self.alignment_controls.row_alignment_assignment_count
                + self.alignment_controls.column_alignment_assignment_count
            ),
            "has_alignment_controls": self.alignment_controls.present,
            "border_assignment_count": (
                self.border_controls.default_border_definition_count
                + self.border_controls.cell_border_assignment_count
                + self.border_controls.row_border_assignment_count
                + self.border_controls.column_border_assignment_count
            ),
            "has_border_controls": self.border_controls.present,
            "worksheet_dimension_control_count": (
                self.worksheet_dimension_controls.default_row_height_count
                + self.worksheet_dimension_controls.default_column_width_count
                + self.worksheet_dimension_controls.default_baseline_adjustment_sheet_count
                + self.worksheet_dimension_controls.default_border_adjustment_sheet_count
                + self.worksheet_dimension_controls.row_height_assignment_count
                + self.worksheet_dimension_controls.row_baseline_adjustment_count
                + self.worksheet_dimension_controls.row_border_adjustment_count
                + self.worksheet_dimension_controls.column_width_assignment_count
                + self.worksheet_dimension_controls.best_fit_column_assignment_count
            ),
            "has_worksheet_dimension_controls": (
                self.worksheet_dimension_controls.present
            ),
            "worksheet_display_control_count": (
                self.worksheet_display_controls.zero_hidden_view_count
                + self.worksheet_display_controls.formula_view_count
                + self.worksheet_display_controls.gridlines_hidden_view_count
                + self.worksheet_display_controls.custom_gridline_color_view_count
                + self.worksheet_display_controls.headers_hidden_view_count
                + self.worksheet_display_controls.outline_symbols_hidden_view_count
                + self.worksheet_display_controls.ruler_hidden_view_count
                + self.worksheet_display_controls.white_space_hidden_view_count
                + self.worksheet_display_controls.right_to_left_view_count
                + self.worksheet_display_controls.non_normal_view_count
                + self.worksheet_display_controls.split_or_frozen_pane_count
            ),
            "has_worksheet_display_controls": (
                self.worksheet_display_controls.present
            ),
            "worksheet_print_layout_control_count": (
                self.worksheet_print_layout_controls.print_area_definition_count
                + self.worksheet_print_layout_controls.print_title_definition_count
                + self.worksheet_print_layout_controls.print_gridlines_sheet_count
                + self.worksheet_print_layout_controls.print_headings_sheet_count
                + self.worksheet_print_layout_controls.horizontally_centered_print_sheet_count
                + self.worksheet_print_layout_controls.vertically_centered_print_sheet_count
                + self.worksheet_print_layout_controls.page_margin_sheet_count
                + self.worksheet_print_layout_controls.page_setup_sheet_count
                + self.worksheet_print_layout_controls.header_footer_sheet_count
                + self.worksheet_print_layout_controls.manual_row_page_break_count
                + self.worksheet_print_layout_controls.manual_column_page_break_count
            ),
            "has_worksheet_print_layout_controls": (
                self.worksheet_print_layout_controls.present
            ),
            "workbook_theme_part_count": self.workbook_theme.theme_part_count,
            "workbook_theme_image_part_count": self.workbook_theme.theme_image_part_count,
            "has_workbook_theme": self.workbook_theme.present,
            "defined_names": len(self.defined_names),
            "table_count": len(self.tables),
            "data_validation_rules": len(self.data_validations),
            "data_validation_target_ranges": sum(
                validation.target_range_count for validation in self.data_validations
            ),
            "conditional_formatting_rules": len(self.conditional_formatting),
            "conditional_formatting_target_ranges": sum(
                rule.target_range_count for rule in self.conditional_formatting
            ),
            "conditional_formatting_extensions": len(
                self.conditional_formatting_extensions
            ),
            "workbook_protection_enabled": bool(
                self.workbook_protection and self.workbook_protection.enabled
            ),
            "sheet_protection_controls": len(self.sheet_protections),
            "protected_sheet_count": sum(
                protection.enabled for protection in self.sheet_protections
            ),
            "protected_range_count": len(self.protected_ranges),
            "protected_range_target_ranges": sum(
                protected_range.target_range_count
                for protected_range in self.protected_ranges
            ),
            "cell_protection_assignment_count": len(
                self.cell_protection_assignments
            ),
            "external_data_connection_count": len(self.external_data_connections),
            "external_data_connections_refresh_on_load": sum(
                connection.refresh_on_load
                for connection in self.external_data_connections
            ),
            "query_table_refresh_control_count": len(
                self.query_table_refresh_controls
            ),
            "query_tables_refresh_on_load": sum(
                control.refresh_on_load
                for control in self.query_table_refresh_controls
            ),
            "pivot_cache_refresh_control_count": len(
                self.pivot_cache_refresh_controls
            ),
            "pivot_caches_refresh_on_load": sum(
                control.refresh_on_load
                for control in self.pivot_cache_refresh_controls
            ),
            "external_link_package_count": self.external_link_packages.external_link_count,
            "external_workbook_link_count": self.external_link_packages.external_workbook_count,
            "dde_link_count": self.external_link_packages.dde_link_count,
            "ole_link_count": self.external_link_packages.ole_link_count,
            "package_external_relationship_count": (
                self.external_relationships.external_relationship_count
            ),
            "package_external_relationship_source_count": (
                self.external_relationships.external_relationship_source_count
            ),
            "package_external_hyperlink_relationship_count": (
                self.external_relationships.external_hyperlink_relationship_count
            ),
            "package_external_image_relationship_count": (
                self.external_relationships.external_image_relationship_count
            ),
            "has_external_relationships": self.external_relationships.present,
            "formula_external_action_cell_count": (
                self.formula_external_actions.formula_external_action_cell_count
            ),
            "formula_external_action_defined_name_count": (
                self.formula_external_actions.action_defined_name_count
            ),
            "formula_hyperlink_function_count": (
                self.formula_external_actions.hyperlink_function_count
            ),
            "formula_webservice_function_count": (
                self.formula_external_actions.webservice_function_count
            ),
            "formula_image_function_count": (
                self.formula_external_actions.image_function_count
            ),
            "formula_rtd_function_count": (
                self.formula_external_actions.rtd_function_count
            ),
            "formula_stockhistory_function_count": (
                self.formula_external_actions.stockhistory_function_count
            ),
            "formula_cube_function_count": (
                self.formula_external_actions.cube_function_count
            ),
            "has_formula_external_actions": self.formula_external_actions.present,
            "formula_dde_link_formula_cell_count": (
                self.formula_dde_links.dde_formula_cell_count
            ),
            "formula_dde_link_count": self.formula_dde_links.dde_link_count,
            "formula_dde_link_defined_name_count": (
                self.formula_dde_links.dde_defined_name_count
            ),
            "has_formula_dde_links": self.formula_dde_links.present,
            "python_in_excel_part_count": self.python_in_excel.python_part_count,
            "python_in_excel_formula_cell_count": (
                self.python_in_excel.python_formula_cell_count
            ),
            "python_in_excel_function_count": (
                self.python_in_excel.python_function_count
            ),
            "python_in_excel_script_count": self.python_in_excel.python_script_count,
            "python_in_excel_environment_definition_count": (
                self.python_in_excel.python_environment_definition_count
            ),
            "python_in_excel_initialization_count": (
                self.python_in_excel.python_initialization_count
            ),
            "has_python_in_excel": self.python_in_excel.present,
            "namespaced_custom_function_formula_cell_count": (
                self.office_custom_functions.namespaced_custom_function_formula_cell_count
            ),
            "namespaced_custom_function_call_count": (
                self.office_custom_functions.namespaced_custom_function_call_count
            ),
            "namespaced_custom_function_namespace_count": (
                self.office_custom_functions.namespaced_custom_function_namespace_count
            ),
            "has_namespaced_custom_function_calls": (
                self.office_custom_functions.present
            ),
            "unqualified_runtime_function_formula_cell_count": (
                self.unqualified_runtime_functions.unqualified_runtime_function_formula_cell_count
            ),
            "unqualified_runtime_function_call_count": (
                self.unqualified_runtime_functions.unqualified_runtime_function_call_count
            ),
            "unqualified_runtime_function_defined_name_count": (
                self.unqualified_runtime_functions.unqualified_runtime_function_defined_name_count
            ),
            "has_unqualified_runtime_function_calls": (
                self.unqualified_runtime_functions.present
            ),
            "worksheet_code_resource_registration_formula_cell_count": (
                self.worksheet_code_resource_registrations.registration_formula_cell_count
            ),
            "worksheet_code_resource_register_id_function_count": (
                self.worksheet_code_resource_registrations.register_id_function_count
            ),
            "worksheet_code_resource_registration_defined_name_count": (
                self.worksheet_code_resource_registrations.registration_defined_name_count
            ),
            "has_worksheet_code_resource_registrations": (
                self.worksheet_code_resource_registrations.present
            ),
            "formula_defined_xlm_registration_formula_cell_count": (
                self.formula_defined_xlm_registrations.registration_formula_cell_count
            ),
            "formula_defined_xlm_register_function_count": (
                self.formula_defined_xlm_registrations.register_function_count
            ),
            "formula_defined_xlm_registration_defined_name_count": (
                self.formula_defined_xlm_registrations.registration_defined_name_count
            ),
            "has_formula_defined_xlm_registrations": (
                self.formula_defined_xlm_registrations.present
            ),
            "formula_defined_xlm_evaluation_formula_cell_count": (
                self.formula_defined_xlm_evaluations.evaluation_formula_cell_count
            ),
            "formula_defined_xlm_evaluate_function_count": (
                self.formula_defined_xlm_evaluations.evaluate_function_count
            ),
            "formula_defined_xlm_evaluation_defined_name_count": (
                self.formula_defined_xlm_evaluations.evaluation_defined_name_count
            ),
            "has_formula_defined_xlm_evaluations": (
                self.formula_defined_xlm_evaluations.present
            ),
            "formula_defined_xlm_action_formula_cell_count": (
                self.formula_defined_xlm_actions.action_formula_cell_count
            ),
            "formula_defined_xlm_action_function_count": (
                self.formula_defined_xlm_actions.action_function_count
            ),
            "formula_defined_xlm_action_defined_name_count": (
                self.formula_defined_xlm_actions.action_defined_name_count
            ),
            "has_formula_defined_xlm_actions": (
                self.formula_defined_xlm_actions.present
            ),
            "formula_defined_xlm_get_cell_formula_cell_count": (
                self.formula_defined_xlm_get_cell_calls.get_cell_formula_cell_count
            ),
            "formula_defined_xlm_get_cell_function_count": (
                self.formula_defined_xlm_get_cell_calls.get_cell_function_count
            ),
            "formula_defined_xlm_get_cell_defined_name_count": (
                self.formula_defined_xlm_get_cell_calls.get_cell_defined_name_count
            ),
            "has_formula_defined_xlm_get_cell_calls": (
                self.formula_defined_xlm_get_cell_calls.present
            ),
            "formula_defined_xlm_environment_information_formula_cell_count": (
                self.formula_defined_xlm_environment_information_calls.environment_information_formula_cell_count
            ),
            "formula_defined_xlm_environment_information_function_count": (
                self.formula_defined_xlm_environment_information_calls.environment_information_function_count
            ),
            "formula_defined_xlm_environment_information_defined_name_count": (
                self.formula_defined_xlm_environment_information_calls.environment_information_defined_name_count
            ),
            "has_formula_defined_xlm_environment_information_calls": (
                self.formula_defined_xlm_environment_information_calls.present
            ),
            "formula_environment_information_formula_cell_count": (
                self.formula_environment_information_calls.environment_information_formula_cell_count
            ),
            "formula_environment_information_function_count": (
                self.formula_environment_information_calls.environment_information_function_count
            ),
            "formula_environment_information_defined_name_count": (
                self.formula_environment_information_calls.environment_information_defined_name_count
            ),
            "formula_environment_information_implicit_cell_reference_function_count": (
                self.formula_environment_information_calls.implicit_cell_reference_function_count
            ),
            "formula_environment_information_implicit_sheets_reference_function_count": (
                self.formula_environment_information_calls.implicit_sheets_reference_function_count
            ),
            "formula_sheet_function_count": (
                self.formula_environment_information_calls.sheet_function_count
            ),
            "formula_sheets_function_count": (
                self.formula_environment_information_calls.sheets_function_count
            ),
            "has_formula_environment_information_calls": (
                self.formula_environment_information_calls.present
            ),
            "xlm_macro_sheet_count": self.xlm_macro_sheets.macro_sheet_count,
            "xlm_macro_formula_cell_count": self.xlm_macro_sheets.formula_cell_count,
            "xlm_related_part_payload_count": (
                self.xlm_macro_sheets.fingerprinted_related_part_count
            ),
            "has_xlm_macro_sheets": self.xlm_macro_sheets.present,
            "xlm_automatic_macro_binding_count": (
                self.xlm_automatic_macro_bindings.automatic_macro_binding_count
            ),
            "xlm_auto_open_binding_count": (
                self.xlm_automatic_macro_bindings.auto_open_binding_count
            ),
            "xlm_auto_close_binding_count": (
                self.xlm_automatic_macro_bindings.auto_close_binding_count
            ),
            "xlm_auto_activate_binding_count": (
                self.xlm_automatic_macro_bindings.auto_activate_binding_count
            ),
            "xlm_auto_deactivate_binding_count": (
                self.xlm_automatic_macro_bindings.auto_deactivate_binding_count
            ),
            "has_xlm_automatic_macro_bindings": (
                self.xlm_automatic_macro_bindings.present
            ),
            "ribbon_customization_part_count": self.ribbon_customization.ribbon_part_count,
            "ribbon_callback_attribute_count": (
                self.ribbon_customization.callback_attribute_count
            ),
            "has_ribbon_customization": self.ribbon_customization.present,
            "office_web_addin_taskpane_part_count": (
                self.office_web_addins.taskpane_part_count
            ),
            "office_web_addin_web_extension_part_count": (
                self.office_web_addins.web_extension_part_count
            ),
            "office_web_addin_auto_show_taskpane_count": (
                self.office_web_addins.auto_show_taskpane_count
            ),
            "office_web_addin_worksheet_binding_count": (
                self.office_web_addins.worksheet_binding_count
            ),
            "office_web_addin_in_content_reference_count": (
                self.office_web_addins.in_content_web_extension_reference_count
            ),
            "has_office_web_addins": self.office_web_addins.present,
            "pivot_table_sheet_count": (
                self.pivot_table_definitions.pivot_table_sheet_count
            ),
            "pivot_table_part_count": self.pivot_table_definitions.pivot_table_part_count,
            "pivot_cache_definition_part_count": (
                self.pivot_table_definitions.pivot_cache_definition_part_count
            ),
            "pivot_cache_record_count": self.pivot_table_definitions.cache_record_count,
            "has_pivot_table_definitions": self.pivot_table_definitions.present,
            "slicer_cache_part_count": self.slicer_timeline_caches.slicer_cache_part_count,
            "timeline_cache_part_count": self.slicer_timeline_caches.timeline_cache_part_count,
            "selected_slicer_item_count": (
                self.slicer_timeline_caches.selected_slicer_item_count
            ),
            "has_slicer_timeline_caches": self.slicer_timeline_caches.present,
            "power_pivot_data_model_part_count": (
                self.power_pivot_data_model.data_model_part_count
            ),
            "power_pivot_data_model_table_count": (
                self.power_pivot_data_model.model_table_count
            ),
            "has_power_pivot_data_model": self.power_pivot_data_model.present,
            "chart_host_sheet_count": self.chart_definitions.chart_host_sheet_count,
            "chart_drawing_part_count": self.chart_definitions.chart_drawing_part_count,
            "chart_part_count": self.chart_definitions.chart_part_count,
            "chart_series_count": self.chart_definitions.series_count,
            "chart_cached_data_point_count": (
                self.chart_definitions.cached_data_point_count
            ),
            "chart_ex_part_count": self.chart_definitions.chart_ex_part_count,
            "chart_ex_series_count": self.chart_definitions.chart_ex_series_count,
            "has_chart_definitions": self.chart_definitions.present,
            "worksheet_embedded_control_sheet_count": (
                self.worksheet_embedded_controls.control_sheet_count
            ),
            "worksheet_active_x_part_count": (
                self.worksheet_embedded_controls.active_x_part_count
            ),
            "worksheet_legacy_vml_drawing_part_count": (
                self.worksheet_embedded_controls.legacy_vml_drawing_part_count
            ),
            "worksheet_legacy_vml_control_count": (
                self.worksheet_embedded_controls.legacy_vml_control_count
            ),
            "worksheet_ole_object_count": (
                self.worksheet_embedded_controls.ole_object_count
            ),
            "has_worksheet_embedded_controls": self.worksheet_embedded_controls.present,
            "power_query_mashup_count": self.power_query.mashup_count,
            "power_query_formula_document_count": self.power_query.formula_document_count,
            "power_query_metadata_item_count": self.power_query.metadata_item_count,
            "has_vba": self.macro_hash is not None,
            "external_reference_cells": len(self.external_references),
            "broken_reference_cells": len(self.broken_references),
            "unresolved_reference_cells": len(self.unresolved_reference_tokens),
            "dynamic_reference_cells": len(self.dynamic_reference_functions),
            "three_d_reference_cells": len(self.three_d_reference_tokens),
            "spill_reference_cells": len(self.spill_reference_tokens),
            "implicit_intersection_cells": len(self.implicit_intersection_tokens),
            "legacy_array_formula_cells": len(self.legacy_array_formula_ranges),
            "legacy_array_formula_output_ranges": sum(
                array_range.has_multiple_outputs
                for array_range in self.legacy_array_formula_ranges
            ),
            "legacy_array_formula_output_cells": sum(
                array_range.output_cell_count
                for array_range in self.legacy_array_formula_ranges
            ),
            "dynamic_array_formula_cells": len(self.dynamic_array_formula_cells),
            "dynamic_array_observed_output_ranges": len(
                self.dynamic_array_formula_ranges
            ),
            "dynamic_array_output_reference_cells": len(
                self.dynamic_array_output_references
            ),
            "unclassified_array_formula_cells": len(
                self.unclassified_array_formula_cells
            ),
            "tokenization_failure_cells": len(self.tokenization_failure_cells),
            "parser_warning_count": len(self.parser_warnings),
        }


@dataclass
class DiffReport:
    before: WorkbookSnapshot
    after: WorkbookSnapshot
    changes: list[Change]
    findings: list[Finding]

    def severity_counts(self) -> dict[str, int]:
        counts = {severity: 0 for severity in SEVERITY_ORDER}
        for finding in self.findings:
            counts[finding.severity] = counts.get(finding.severity, 0) + 1
        return {severity: count for severity, count in counts.items() if count}

    @property
    def highest_severity(self) -> str:
        if not self.findings:
            return "note"
        return max(self.findings, key=lambda finding: SEVERITY_ORDER[finding.severity]).severity

    def to_dict(self, extra_findings: Iterable[Finding] = ()) -> dict[str, Any]:
        findings = [*self.findings, *extra_findings]
        counts = {severity: 0 for severity in SEVERITY_ORDER}
        for finding in findings:
            counts[finding.severity] = counts.get(finding.severity, 0) + 1
        highest = (
            max(findings, key=lambda finding: SEVERITY_ORDER[finding.severity]).severity
            if findings
            else "note"
        )
        return {
            "schema_version": "1.0",
            "before": self.before.summary(),
            "after": self.after.summary(),
            "summary": {
                "change_count": len(self.changes),
                "finding_count": len(findings),
                "highest_severity": highest,
                "findings_by_severity": {
                    severity: count for severity, count in counts.items() if count
                },
            },
            "changes": [change.to_dict() for change in self.changes],
            "findings": [finding.to_dict() for finding in findings],
        }


class FormulaFenceError(Exception):
    """Base class for errors that should be presented cleanly by the CLI."""


class WorkbookLoadError(FormulaFenceError):
    """The workbook could not be safely inspected."""


class PolicyError(FormulaFenceError):
    """The policy file is invalid or cannot be interpreted."""
