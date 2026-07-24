"""Policy-as-code validation for FormulaFence reports."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl.utils.cell import coordinate_to_tuple

from formulafence.formulas import ParsedReference, parse_reference_token
from formulafence.models import CellKey, DiffReport, Finding, PolicyError

_RULE_FIELDS = {
    "no_formula_to_value",
    "no_new_external_links",
    "no_new_broken_references",
    "no_macro_changes",
    "no_new_parser_warnings",
    "no_new_unresolved_references",
    "no_new_dynamic_references",
    "no_new_spill_references",
    "no_new_dynamic_array_output_references",
    "no_new_implicit_intersections",
    "no_array_formula_semantics_changes",
    "no_new_tokenization_failures",
    "no_table_definition_changes",
    "no_data_validation_changes",
    "no_conditional_formatting_changes",
    "no_protection_changes",
    "no_external_data_connection_changes",
    "no_3d_reference_scope_changes",
    "no_sheet_visibility_changes",
    "max_changed_formulas",
    "max_downstream_impact",
}
_TOP_LEVEL_FIELDS = {"version", "rules", "protected_cells", "allowed_changes"}


@dataclass(frozen=True)
class CellSelector:
    sheet: str
    min_column: int
    min_row: int
    max_column: int
    max_row: int
    source: str

    def matches(self, location: CellKey) -> bool:
        sheet, coordinate = location
        row, column = coordinate_to_tuple(coordinate)
        return (
            sheet.casefold() == self.sheet.casefold()
            and self.min_column <= column <= self.max_column
            and self.min_row <= row <= self.max_row
        )


@dataclass(frozen=True)
class Policy:
    version: int
    no_formula_to_value: bool = False
    no_new_external_links: bool = False
    no_new_broken_references: bool = False
    no_macro_changes: bool = False
    no_new_parser_warnings: bool = False
    no_new_unresolved_references: bool = False
    no_new_dynamic_references: bool = False
    no_new_spill_references: bool = False
    no_new_dynamic_array_output_references: bool = False
    no_new_implicit_intersections: bool = False
    no_array_formula_semantics_changes: bool = False
    no_new_tokenization_failures: bool = False
    no_table_definition_changes: bool = False
    no_data_validation_changes: bool = False
    no_conditional_formatting_changes: bool = False
    no_protection_changes: bool = False
    no_external_data_connection_changes: bool = False
    no_3d_reference_scope_changes: bool = False
    no_sheet_visibility_changes: bool = False
    max_changed_formulas: int | None = None
    max_downstream_impact: int | None = None
    protected_cells: tuple[CellSelector, ...] = ()
    allowed_changes: tuple[CellSelector, ...] = ()


DEFAULT_POLICY = """# FormulaFence policy: commit this file beside a material workbook.
version: 1

rules:
  no_formula_to_value: true
  no_new_external_links: true
  no_new_broken_references: true
  no_macro_changes: true
  no_new_parser_warnings: true
  no_new_unresolved_references: true
  no_new_dynamic_references: true
  no_new_spill_references: true
  no_new_dynamic_array_output_references: true
  no_new_implicit_intersections: true
  no_array_formula_semantics_changes: true
  no_new_tokenization_failures: true
  no_table_definition_changes: true
  no_data_validation_changes: true
  no_conditional_formatting_changes: true
  no_protection_changes: true
  no_external_data_connection_changes: true
  no_3d_reference_scope_changes: true
  max_changed_formulas: 20
  max_downstream_impact: 100

# Exact cells or ranges that must not be changed.
protected_cells: []

# Optional. When non-empty, every changed cell must fall in one of these ranges.
allowed_changes: []
"""


def _parse_selector(value: object, field_name: str) -> CellSelector:
    if not isinstance(value, str) or not value.strip():
        raise PolicyError(f"{field_name} entries must be non-empty strings like Dashboard!B12")
    reference: ParsedReference | None = parse_reference_token(value)
    if (
        reference is None
        or reference.is_external
        or reference.sheet is None
        or reference.min_column is None
        or reference.min_row is None
        or reference.max_column is None
        or reference.max_row is None
    ):
        raise PolicyError(
            f"Invalid {field_name} selector {value!r}; use a sheet-qualified A1 cell or range"
        )
    return CellSelector(
        sheet=reference.sheet,
        min_column=reference.min_column,
        min_row=reference.min_row,
        max_column=reference.max_column,
        max_row=reference.max_row,
        source=value,
    )


def _parse_selectors(value: object, field_name: str) -> tuple[CellSelector, ...]:
    if value is None:
        return ()
    if not isinstance(value, list):
        raise PolicyError(f"{field_name} must be a YAML list")
    return tuple(_parse_selector(item, field_name) for item in value)


def _boolean_rule(rules: dict[str, Any], name: str) -> bool:
    value = rules.get(name, False)
    if type(value) is not bool:  # bool is an int subclass; reject accidental 0/1 configs.
        raise PolicyError(f"rules.{name} must be true or false")
    return value


def _integer_rule(rules: dict[str, Any], name: str) -> int | None:
    value = rules.get(name)
    if value is None:
        return None
    if type(value) is not int or value < 0:
        raise PolicyError(f"rules.{name} must be a non-negative integer")
    return value


def parse_policy(data: object) -> Policy:
    if not isinstance(data, dict):
        raise PolicyError("Policy root must be a mapping")
    unknown_fields = set(data) - _TOP_LEVEL_FIELDS
    if unknown_fields:
        raise PolicyError(f"Unknown policy fields: {', '.join(sorted(unknown_fields))}")
    version = data.get("version")
    if version != 1:
        raise PolicyError("Policy must declare version: 1")
    rules = data.get("rules", {})
    if not isinstance(rules, dict):
        raise PolicyError("rules must be a mapping")
    unknown_rules = set(rules) - _RULE_FIELDS
    if unknown_rules:
        raise PolicyError(f"Unknown rules: {', '.join(sorted(unknown_rules))}")

    return Policy(
        version=version,
        no_formula_to_value=_boolean_rule(rules, "no_formula_to_value"),
        no_new_external_links=_boolean_rule(rules, "no_new_external_links"),
        no_new_broken_references=_boolean_rule(rules, "no_new_broken_references"),
        no_macro_changes=_boolean_rule(rules, "no_macro_changes"),
        no_new_parser_warnings=_boolean_rule(rules, "no_new_parser_warnings"),
        no_new_unresolved_references=_boolean_rule(rules, "no_new_unresolved_references"),
        no_new_dynamic_references=_boolean_rule(rules, "no_new_dynamic_references"),
        no_new_spill_references=_boolean_rule(rules, "no_new_spill_references"),
        no_new_dynamic_array_output_references=_boolean_rule(
            rules, "no_new_dynamic_array_output_references"
        ),
        no_new_implicit_intersections=_boolean_rule(
            rules, "no_new_implicit_intersections"
        ),
        no_array_formula_semantics_changes=_boolean_rule(
            rules, "no_array_formula_semantics_changes"
        ),
        no_new_tokenization_failures=_boolean_rule(
            rules, "no_new_tokenization_failures"
        ),
        no_table_definition_changes=_boolean_rule(rules, "no_table_definition_changes"),
        no_data_validation_changes=_boolean_rule(rules, "no_data_validation_changes"),
        no_conditional_formatting_changes=_boolean_rule(
            rules, "no_conditional_formatting_changes"
        ),
        no_protection_changes=_boolean_rule(rules, "no_protection_changes"),
        no_external_data_connection_changes=_boolean_rule(
            rules, "no_external_data_connection_changes"
        ),
        no_3d_reference_scope_changes=_boolean_rule(
            rules, "no_3d_reference_scope_changes"
        ),
        no_sheet_visibility_changes=_boolean_rule(rules, "no_sheet_visibility_changes"),
        max_changed_formulas=_integer_rule(rules, "max_changed_formulas"),
        max_downstream_impact=_integer_rule(rules, "max_downstream_impact"),
        protected_cells=_parse_selectors(data.get("protected_cells"), "protected_cells"),
        allowed_changes=_parse_selectors(data.get("allowed_changes"), "allowed_changes"),
    )


def load_policy(path: str | Path) -> Policy:
    policy_path = Path(path)
    if not policy_path.exists() or not policy_path.is_file():
        raise PolicyError(f"Policy does not exist or is not a file: {policy_path}")
    try:
        data = yaml.safe_load(policy_path.read_text(encoding="utf-8"))
    except (OSError, yaml.YAMLError) as error:
        raise PolicyError(f"Could not read policy {policy_path}: {error}") from error
    return parse_policy(data)


def _rule_triggered(report: DiffReport, rule_id: str) -> list[Finding]:
    return [finding for finding in report.findings if finding.rule_id == rule_id]


def evaluate_policy(report: DiffReport, policy: Policy) -> list[Finding]:
    """Return every policy violation; callers decide whether to fail the build."""
    violations: list[Finding] = []

    if policy.no_formula_to_value:
        for change in report.changes:
            if change.kind == "formula_to_value":
                violations.append(
                    Finding(
                        "FFP001",
                        "high",
                        "Policy forbids replacing a formula with a value.",
                        change.location,
                    )
                )
    if policy.no_new_external_links:
        for finding in _rule_triggered(report, "FF004"):
            violations.append(
                Finding(
                    "FFP002",
                    "high",
                    "Policy forbids new external-workbook references.",
                    finding.location,
                )
            )
    if policy.no_new_broken_references:
        for finding in _rule_triggered(report, "FF003"):
            violations.append(
                Finding(
                    "FFP003",
                    "critical",
                    "Policy forbids new broken #REF! references.",
                    finding.location,
                )
            )
    if policy.no_macro_changes and _rule_triggered(report, "FF005"):
        violations.append(
            Finding("FFP004", "critical", "Policy forbids changes to the VBA macro payload.")
        )
    if policy.no_new_parser_warnings:
        for finding in _rule_triggered(report, "FF010"):
            violations.append(
                Finding(
                    "FFP010",
                    "high",
                    "Policy forbids new unsupported-workbook coverage warnings.",
                    details=finding.details,
                )
            )
    if policy.no_new_unresolved_references:
        for finding in _rule_triggered(report, "FF011"):
            violations.append(
                Finding(
                    "FFP011",
                    "high",
                    "Policy forbids newly unresolvable formula references.",
                    finding.location,
                    details=finding.details,
                )
            )
    if policy.no_new_dynamic_references:
        for finding in _rule_triggered(report, "FF012"):
            violations.append(
                Finding(
                    "FFP012",
                    "high",
                    "Policy forbids new dynamic reference functions.",
                    finding.location,
                    details=finding.details,
                )
            )
    if policy.no_new_spill_references:
        for finding in _rule_triggered(report, "FF015"):
            violations.append(
                Finding(
                    "FFP015",
                    "high",
                    "Policy forbids new dynamic-array spill references.",
                    finding.location,
                    details=finding.details,
                )
            )
    if policy.no_new_dynamic_array_output_references:
        for finding in _rule_triggered(report, "FF019"):
            violations.append(
                Finding(
                    "FFP019",
                    "high",
                    (
                        "Policy forbids newly observed dynamic-array output-member "
                        "references."
                    ),
                    finding.location,
                    details=finding.details,
                )
            )
    if policy.no_new_implicit_intersections:
        for finding in _rule_triggered(report, "FF017"):
            violations.append(
                Finding(
                    "FFP017",
                    "high",
                    "Policy forbids new explicit implicit-intersection operators.",
                    finding.location,
                    details=finding.details,
                )
            )
    if policy.no_array_formula_semantics_changes:
        for finding in _rule_triggered(report, "FF018"):
            violations.append(
                Finding(
                    "FFP018",
                    "high",
                    "Policy forbids array-formula mode or fixed-output-range changes.",
                    finding.location,
                    details=finding.details,
                )
            )
    if policy.no_new_tokenization_failures:
        for finding in _rule_triggered(report, "FF016"):
            violations.append(
                Finding(
                    "FFP016",
                    "high",
                    "Policy forbids formulas that FormulaFence cannot tokenize.",
                    finding.location,
                    details=finding.details,
                )
            )
    if policy.no_table_definition_changes:
        for finding in _rule_triggered(report, "FF013"):
            violations.append(
                Finding(
                    "FFP013",
                    "high",
                    "Policy forbids changes to Excel-table definitions.",
                    details=finding.details,
                )
            )
    if policy.no_data_validation_changes:
        for finding in _rule_triggered(report, "FF020"):
            violations.append(
                Finding(
                    "FFP020",
                    "high",
                    "Policy forbids changes to data-validation controls.",
                    details=finding.details,
                )
            )
    if policy.no_conditional_formatting_changes:
        for finding in _rule_triggered(report, "FF021"):
            violations.append(
                Finding(
                    "FFP021",
                    "high",
                    "Policy forbids changes to conditional-formatting controls.",
                    details=finding.details,
                )
            )
    if policy.no_protection_changes:
        for finding in _rule_triggered(report, "FF022"):
            violations.append(
                Finding(
                    "FFP022",
                    "high",
                    "Policy forbids changes to workbook, sheet, or cell protection controls.",
                    details=finding.details,
                )
            )
    if policy.no_external_data_connection_changes:
        for finding in _rule_triggered(report, "FF023"):
            violations.append(
                Finding(
                    "FFP023",
                    "high",
                    "Policy forbids changes to external-data connections and refresh controls.",
                    details=finding.details,
                )
            )
    if policy.no_3d_reference_scope_changes:
        for finding in _rule_triggered(report, "FF014"):
            violations.append(
                Finding(
                    "FFP014",
                    "high",
                    "Policy forbids changes to static 3-D reference scope.",
                    finding.location,
                    details=finding.details,
                )
            )
    if policy.no_sheet_visibility_changes:
        for finding in _rule_triggered(report, "FF007"):
            violations.append(
                Finding(
                    "FFP005",
                    "high",
                    "Policy forbids changes to sheet visibility.",
                    details=finding.details,
                )
            )

    for change in report.changes:
        if change.location is None:
            continue
        if any(selector.matches(change.location) for selector in policy.protected_cells):
            violations.append(
                Finding(
                    "FFP006",
                    "high",
                    "Policy protects this cell or range from changes.",
                    change.location,
                )
            )
        if policy.allowed_changes and not any(
            selector.matches(change.location) for selector in policy.allowed_changes
        ):
            violations.append(
                Finding(
                    "FFP007",
                    "high",
                    "Policy permits changes only inside allowed_changes ranges.",
                    change.location,
                )
            )

    formula_changes = sum(
        1
        for change in report.changes
        if (change.before is not None and change.before.is_formula)
        or (change.after is not None and change.after.is_formula)
    )
    if policy.max_changed_formulas is not None and formula_changes > policy.max_changed_formulas:
        violations.append(
            Finding(
                "FFP008",
                "high",
                (
                    "Changed formula count exceeds policy limit "
                    f"({formula_changes} > {policy.max_changed_formulas})."
                ),
                details={"actual": formula_changes, "limit": policy.max_changed_formulas},
            )
        )
    if policy.max_downstream_impact is not None:
        for change in report.changes:
            if change.location is None or change.impact_count <= policy.max_downstream_impact:
                continue
            violations.append(
                Finding(
                    "FFP009",
                    "high",
                    (
                        "Change exceeds the downstream-impact limit "
                        f"({change.impact_count} > {policy.max_downstream_impact})."
                    ),
                    change.location,
                    details={
                        "actual": change.impact_count,
                        "limit": policy.max_downstream_impact,
                        "impact_paths": change.details.get("impact_paths", []),
                    },
                )
            )
    return violations
