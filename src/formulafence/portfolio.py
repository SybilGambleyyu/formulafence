"""Bounded, fail-closed comparison of workbook portfolios.

Portfolio comparison deliberately uses a relative path as the identity of a
workbook.  It does not guess that two differently named files are a rename:
that would make a review less trustworthy when a file was intentionally
replaced.  Every supported workbook is inspected independently, so one
malformed archive becomes visible evidence instead of erasing the rest of a
portfolio report.
"""

from __future__ import annotations

import stat
from collections import Counter, defaultdict, deque
from collections.abc import Iterable
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any

from openpyxl.utils.cell import coordinate_to_tuple

from formulafence.diff import compare_snapshots
from formulafence.formulas import (
    ParsedReference,
    StructuredTable,
    resolve_structured_reference,
)
from formulafence.models import (
    SEVERITY_ORDER,
    CellKey,
    Change,
    DiffReport,
    ExternalWorkbookStructuredReference,
    Finding,
    FormulaFenceError,
    WorkbookSnapshot,
    display_location,
)
from formulafence.policy import (
    Policy,
    evaluate_policy,
    evaluate_portfolio_link_policy,
    evaluate_portfolio_membership_policy,
)
from formulafence.workbook import WorkbookSourceIdentity, load_snapshot

_SUPPORTED_SUFFIXES = frozenset({".xlsx", ".xlsm"})
_UNSUPPORTED_EXCEL_SUFFIXES = frozenset(
    {".xls", ".xlsb", ".xlt", ".xltx", ".xltm", ".xlam", ".ods"}
)
_DEFAULT_MAX_WORKBOOKS = 512
DEFAULT_MAX_INVENTORY_ENTRIES = 32_768
DEFAULT_MAX_LINK_IMPACT = 100_000
_LINK_IMPACT_SAMPLE_LIMIT = 10

PortfolioNode = tuple[str, CellKey]


class PortfolioError(FormulaFenceError):
    """The supplied portfolio cannot be safely compared."""


@dataclass(frozen=True)
class _PortfolioWorkbookSource:
    """A workbook path coupled to the regular file observed during inventory."""

    path: Path
    identity: WorkbookSourceIdentity


def _path_sort_key(value: str) -> tuple[str, str]:
    return value.casefold(), value


def _location_sort_key(location: CellKey) -> tuple[str, str, int, int, str]:
    sheet, coordinate = location
    row, column = coordinate_to_tuple(coordinate)
    return sheet.casefold(), sheet, row, column, coordinate


def _node_sort_key(node: PortfolioNode) -> tuple[str, str, str, str, int, int, str]:
    workbook, location = node
    return (*_path_sort_key(workbook), *_location_sort_key(location))


@dataclass(frozen=True)
class _ExternalPortfolioDependency:
    """A safely resolved external range pointing to one candidate formula."""

    source_workbook: str
    source_sheet: str
    min_column: int
    min_row: int
    max_column: int
    max_row: int
    dependent: PortfolioNode

    def contains(self, location: CellKey) -> bool:
        sheet, coordinate = location
        if sheet.casefold() != self.source_sheet.casefold():
            return False
        row, column = coordinate_to_tuple(coordinate)
        return (
            self.min_column <= column <= self.max_column
            and self.min_row <= row <= self.max_row
        )


@dataclass(frozen=True)
class _PortfolioImpactGraph:
    """Candidate-only local and safely resolved cross-workbook dependency graph."""

    snapshots: dict[str, WorkbookSnapshot]
    external_dependents: dict[
        tuple[str, str], tuple[_ExternalPortfolioDependency, ...]
    ]

    def direct_dependents(self, node: PortfolioNode) -> tuple[PortfolioNode, ...]:
        workbook, location = node
        snapshot = self.snapshots[workbook]
        dependents: set[PortfolioNode] = {
            (workbook, dependent) for dependent in snapshot.direct_dependents(location)
        }
        for dependency in self.external_dependents.get(
            (workbook, location[0].casefold()), ()
        ):
            if dependency.contains(location):
                dependents.add(dependency.dependent)
        return tuple(sorted(dependents, key=_node_sort_key))


@dataclass(frozen=True)
class _PortfolioImpactTraversal:
    """Bounded source-to-downstream evidence for changed candidate cells."""

    paths_by_root: dict[PortfolioNode, tuple[tuple[PortfolioNode, ...], ...]]
    incomplete_root: PortfolioNode | None = None


def _safe_relative_path(path: Path, root: Path) -> str:
    """Return a portable relative identity or fail before scanning a file."""
    try:
        relative = path.relative_to(root).as_posix()
    except ValueError as error:  # pragma: no cover - defensive against races
        raise PortfolioError("Workbook escaped its supplied portfolio root.") from error
    if "\\" in relative or any(
        ord(character) < 32 or ord(character) == 127 for character in relative
    ):
        raise PortfolioError(
            "Portfolio workbook paths must not contain control characters or backslashes."
        )
    return relative


def _resolve_directory(path: str | Path, label: str) -> Path:
    supplied = Path(path)
    try:
        resolved = supplied.resolve(strict=True)
    except OSError as error:
        raise PortfolioError(
            f"Could not resolve {label} portfolio directory: {supplied}"
        ) from error
    if not resolved.is_dir():
        raise PortfolioError(f"{label.capitalize()} portfolio is not a directory: {supplied}")
    return resolved


def _discover_portfolio_workbook_sources(
    root: str | Path,
    *,
    label: str,
    max_workbooks: int = _DEFAULT_MAX_WORKBOOKS,
    max_inventory_entries: int = DEFAULT_MAX_INVENTORY_ENTRIES,
) -> dict[str, _PortfolioWorkbookSource]:
    """Return supported inventoried workbook sources keyed by relative path.

    Office lock files (``~$*.xlsx``) are transient and intentionally ignored.
    Every other known spreadsheet extension outside FormulaFence's `.xlsx` /
    `.xlsm` contract is an explicit error rather than a silent coverage hole.
    """
    if max_workbooks < 1:
        raise PortfolioError("max_workbooks must be at least 1.")
    if max_inventory_entries < 1:
        raise PortfolioError("max_inventory_entries must be at least 1.")

    resolved_root = _resolve_directory(root, label)
    try:
        candidates: list[Path] = []
        for candidate in resolved_root.rglob("*"):
            if len(candidates) >= max_inventory_entries:
                raise PortfolioError(
                    f"{label.capitalize()} portfolio contains more than "
                    f"{max_inventory_entries} filesystem entries, exceeding "
                    f"max_inventory_entries={max_inventory_entries}."
                )
            candidates.append(candidate)
    except OSError as error:
        raise PortfolioError(f"Could not inventory {label} portfolio directory.") from error
    candidates.sort(key=lambda item: _path_sort_key(item.as_posix()))

    workbooks: dict[str, _PortfolioWorkbookSource] = {}
    casefolded_paths: dict[str, str] = {}
    unsupported: list[str] = []
    for candidate in candidates:
        try:
            candidate_stat = candidate.stat(follow_symlinks=False)
            if stat.S_ISLNK(candidate_stat.st_mode):
                relative = _safe_relative_path(candidate, resolved_root)
                raise PortfolioError(
                    f"Refusing symlinked path in {label} portfolio: {relative}"
                )
            if not stat.S_ISREG(candidate_stat.st_mode):
                continue
        except OSError as error:
            raise PortfolioError(f"Could not inspect a path in {label} portfolio.") from error

        relative = _safe_relative_path(candidate, resolved_root)
        if candidate.name.startswith("~$"):
            continue
        suffix = candidate.suffix.casefold()
        if suffix in _UNSUPPORTED_EXCEL_SUFFIXES:
            unsupported.append(relative)
            continue
        if suffix not in _SUPPORTED_SUFFIXES:
            continue

        try:
            candidate.resolve(strict=True).relative_to(resolved_root)
        except (OSError, ValueError) as error:
            raise PortfolioError(
                f"Workbook escaped its supplied {label} portfolio root: {relative}"
            ) from error

        portable_key = relative.casefold()
        previous = casefolded_paths.get(portable_key)
        if previous is not None:
            raise PortfolioError(
                "Portfolio contains paths that differ only by case: "
                f"{previous} and {relative}"
            )
        casefolded_paths[portable_key] = relative
        workbooks[relative] = _PortfolioWorkbookSource(
            path=candidate,
            identity=WorkbookSourceIdentity(
                device=candidate_stat.st_dev,
                inode=candidate_stat.st_ino,
                changed_at_ns=candidate_stat.st_ctime_ns,
                size=candidate_stat.st_size,
            ),
        )

    if unsupported:
        shown = ", ".join(unsupported[:5])
        remainder = "" if len(unsupported) <= 5 else f" (+{len(unsupported) - 5} more)"
        raise PortfolioError(
            f"{label.capitalize()} portfolio contains unsupported spreadsheet files: "
            f"{shown}{remainder}"
        )
    if len(workbooks) > max_workbooks:
        raise PortfolioError(
            f"{label.capitalize()} portfolio contains {len(workbooks)} supported workbooks, "
            f"exceeding max_workbooks={max_workbooks}."
        )
    return workbooks


def discover_workbooks(
    root: str | Path,
    *,
    label: str,
    max_workbooks: int = _DEFAULT_MAX_WORKBOOKS,
    max_inventory_entries: int = DEFAULT_MAX_INVENTORY_ENTRIES,
) -> dict[str, Path]:
    """Return supported workbook files keyed by a stable relative path.

    This public inventory view intentionally exposes paths for callers that
    only need to inspect membership. Portfolio comparison keeps the matching
    private file identities and validates them when it later opens each file.
    """
    sources = _discover_portfolio_workbook_sources(
        root,
        label=label,
        max_workbooks=max_workbooks,
        max_inventory_entries=max_inventory_entries,
    )
    return {relative: source.path for relative, source in sources.items()}


def _resolve_relative_external_workbook(
    consumer_workbook: str,
    source_path: str,
    candidate_paths: dict[str, str],
) -> str | None:
    """Resolve only an exact, safely relative external path in this portfolio.

    Excel stores several link spellings, including paths relative to the
    consuming workbook.  This resolver is intentionally narrower than Excel:
    it never follows a path on disk, expands a URI, or searches by basename.
    A spelling must normalize to an already-inspected candidate relative path.
    """
    normalized = source_path.replace("\\", "/")
    if (
        not normalized
        or normalized.startswith("/")
        or ":" in normalized
        or any(ord(character) < 32 or ord(character) == 127 for character in normalized)
    ):
        return None

    components = consumer_workbook.split("/")[:-1]
    for component in normalized.split("/"):
        if component == ".":
            continue
        if not component:
            return None
        if component == "..":
            if not components:
                return None
            components.pop()
            continue
        if "[" in component or "]" in component:
            return None
        components.append(component)

    if not components or Path(components[-1]).suffix.casefold() not in _SUPPORTED_SUFFIXES:
        return None
    return candidate_paths.get("/".join(components).casefold())


def _canonical_sheet_name(snapshot: WorkbookSnapshot, requested_sheet: str) -> str | None:
    """Return one exact candidate sheet identity without guessing collisions."""
    matches = [
        title for title in snapshot.sheets if title.casefold() == requested_sheet.casefold()
    ]
    return matches[0] if len(matches) == 1 else None


def _canonical_three_d_sheet_span(
    snapshot: WorkbookSnapshot, first_sheet: str, last_sheet: str
) -> tuple[str, ...]:
    """Return one bounded source worksheet span, or nothing when it is ambiguous.

    Excel's 3-D form selects every worksheet between two endpoint tabs. The
    source snapshot carries both the worksheet order and a raw OOXML tab
    catalog captured from the candidate itself, not a filename-derived
    approximation. Require that complete catalog's ordinary-worksheet subset to
    agree exactly with the loaded order, then require unique endpoint and span
    identities, a forward order, and an inspected worksheet for every included
    tab before creating any cross-workbook graph edge.
    """
    if (
        not snapshot.workbook_tab_order_complete
        or not snapshot.worksheet_tab_order_complete
        or not snapshot.sheet_order
    ):
        return ()
    sheet_positions: dict[str, int] = {}
    for position, title in enumerate(snapshot.sheet_order):
        key = title.casefold()
        if key in sheet_positions:
            return ()
        sheet_positions[key] = position
    raw_worksheet_keys = tuple(title.casefold() for title in snapshot.worksheet_tab_order)
    if len(raw_worksheet_keys) != len(set(raw_worksheet_keys)):
        return ()
    if raw_worksheet_keys != tuple(title.casefold() for title in snapshot.sheet_order):
        return ()
    canonical_first = _canonical_sheet_name(snapshot, first_sheet)
    canonical_last = _canonical_sheet_name(snapshot, last_sheet)
    if canonical_first is None or canonical_last is None:
        return ()
    first_position = sheet_positions.get(canonical_first.casefold())
    last_position = sheet_positions.get(canonical_last.casefold())
    if (
        first_position is None
        or last_position is None
        or first_position > last_position
    ):
        return ()
    span = snapshot.sheet_order[first_position : last_position + 1]
    if any(_canonical_sheet_name(snapshot, title) != title for title in span):
        return ()
    return span


def _canonical_external_table_references(
    snapshot: WorkbookSnapshot,
    reference: ExternalWorkbookStructuredReference,
) -> tuple[ParsedReference, ...]:
    """Resolve one external table selector only against one source snapshot.

    A table identifier is workbook-scoped.  It is safe to create candidate
    graph edges only when the inspected source snapshot contains exactly one
    case-insensitive table-name match, its worksheet identity is unambiguous,
    and the existing static structured-reference resolver can turn the
    selector into fixed source cells.  Current-row and unsupported selectors
    return no references rather than approximating a relationship.
    """
    tables = [
        table
        for table in snapshot.tables.values()
        if table.name.casefold() == reference.table_name.casefold()
    ]
    if len(tables) != 1:
        return ()
    table = tables[0]
    source_sheet = _canonical_sheet_name(snapshot, table.sheet)
    if source_sheet is None:
        return ()
    source_table = StructuredTable(
        name=table.name,
        sheet=source_sheet,
        ref=table.ref,
        columns=table.columns,
        header_row_count=table.header_row_count,
        totals_row_count=table.totals_row_count,
    )
    resolved = resolve_structured_reference(
        reference.table_reference,
        {table.name.casefold(): source_table},
    )
    if resolved is None:
        return ()
    return tuple(
        source_reference
        for source_reference in resolved
        if source_reference.sheet == source_sheet
        and None
        not in {
            source_reference.min_column,
            source_reference.min_row,
            source_reference.max_column,
            source_reference.max_row,
        }
    )


def _build_candidate_impact_graph(
    entries: Iterable[PortfolioWorkbookReport],
) -> _PortfolioImpactGraph:
    """Build the in-portfolio candidate graph without resolving any external file."""
    snapshots = {
        entry.path: entry.after
        for entry in entries
        if entry.candidate_present and entry.after is not None
    }
    candidate_paths = {path.casefold(): path for path in snapshots}
    external_dependents: defaultdict[
        tuple[str, str], list[_ExternalPortfolioDependency]
    ] = defaultdict(list)

    for dependent_workbook in sorted(snapshots, key=_path_sort_key):
        snapshot = snapshots[dependent_workbook]
        for dependent_location in sorted(
            snapshot.external_workbook_references, key=_location_sort_key
        ):
            for reference in snapshot.external_workbook_references[dependent_location]:
                source_workbook = _resolve_relative_external_workbook(
                    dependent_workbook,
                    reference.source_path,
                    candidate_paths,
                )
                if source_workbook is None:
                    continue
                source_snapshot = snapshots[source_workbook]
                source_sheet = _canonical_sheet_name(source_snapshot, reference.sheet)
                if source_sheet is None:
                    continue
                dependency = _ExternalPortfolioDependency(
                    source_workbook=source_workbook,
                    source_sheet=source_sheet,
                    min_column=reference.min_column,
                    min_row=reference.min_row,
                    max_column=reference.max_column,
                    max_row=reference.max_row,
                    dependent=(dependent_workbook, dependent_location),
                )
                external_dependents[
                    (source_workbook, source_sheet.casefold())
                ].append(dependency)

        for dependent_location in sorted(
            snapshot.external_workbook_three_d_references, key=_location_sort_key
        ):
            for reference in snapshot.external_workbook_three_d_references[
                dependent_location
            ]:
                source_workbook = _resolve_relative_external_workbook(
                    dependent_workbook,
                    reference.source_path,
                    candidate_paths,
                )
                if source_workbook is None:
                    continue
                source_snapshot = snapshots[source_workbook]
                for source_sheet in _canonical_three_d_sheet_span(
                    source_snapshot,
                    reference.first_sheet,
                    reference.last_sheet,
                ):
                    dependency = _ExternalPortfolioDependency(
                        source_workbook=source_workbook,
                        source_sheet=source_sheet,
                        min_column=reference.min_column,
                        min_row=reference.min_row,
                        max_column=reference.max_column,
                        max_row=reference.max_row,
                        dependent=(dependent_workbook, dependent_location),
                    )
                    external_dependents[
                        (source_workbook, source_sheet.casefold())
                    ].append(dependency)

        for dependent_location in sorted(
            snapshot.external_workbook_structured_references,
            key=_location_sort_key,
        ):
            for reference in snapshot.external_workbook_structured_references[
                dependent_location
            ]:
                source_workbook = _resolve_relative_external_workbook(
                    dependent_workbook,
                    reference.source_path,
                    candidate_paths,
                )
                if source_workbook is None:
                    continue
                source_snapshot = snapshots[source_workbook]
                for source_reference in _canonical_external_table_references(
                    source_snapshot,
                    reference,
                ):
                    source_sheet = source_reference.sheet
                    if source_sheet is None:  # defensive; helper validates bounds/sheet
                        continue
                    dependency = _ExternalPortfolioDependency(
                        source_workbook=source_workbook,
                        source_sheet=source_sheet,
                        min_column=source_reference.min_column,
                        min_row=source_reference.min_row,
                        max_column=source_reference.max_column,
                        max_row=source_reference.max_row,
                        dependent=(dependent_workbook, dependent_location),
                    )
                    external_dependents[
                        (source_workbook, source_sheet.casefold())
                    ].append(dependency)

        for dependent_location in sorted(
            snapshot.external_workbook_defined_name_references,
            key=_location_sort_key,
        ):
            for reference in snapshot.external_workbook_defined_name_references[
                dependent_location
            ]:
                source_workbook = _resolve_relative_external_workbook(
                    dependent_workbook,
                    reference.source_path,
                    candidate_paths,
                )
                if source_workbook is None:
                    continue
                source_snapshot = snapshots[source_workbook]
                if reference.scope_sheet is None:
                    source_references = (
                        source_snapshot.static_global_defined_name_references.get(
                            reference.name_key, ()
                        )
                    )
                else:
                    source_scope = _canonical_sheet_name(
                        source_snapshot, reference.scope_sheet
                    )
                    if source_scope is None:
                        continue
                    # Explicit external sheet qualification selects only that
                    # source sheet's local-name scope. Never guess a global
                    # fallback or another local definition with the same key.
                    source_references = (
                        source_snapshot.static_local_defined_name_references.get(
                            source_scope.casefold(), {}
                        ).get(reference.name_key, ())
                    )
                for source_reference in source_references:
                    if (
                        source_reference.sheet is None
                        or None
                        in {
                            source_reference.min_column,
                            source_reference.min_row,
                            source_reference.max_column,
                            source_reference.max_row,
                        }
                    ):
                        continue
                    source_sheet = _canonical_sheet_name(
                        source_snapshot, source_reference.sheet
                    )
                    if source_sheet is None:
                        continue
                    dependency = _ExternalPortfolioDependency(
                        source_workbook=source_workbook,
                        source_sheet=source_sheet,
                        min_column=source_reference.min_column,
                        min_row=source_reference.min_row,
                        max_column=source_reference.max_column,
                        max_row=source_reference.max_row,
                        dependent=(dependent_workbook, dependent_location),
                    )
                    external_dependents[
                        (source_workbook, source_sheet.casefold())
                    ].append(dependency)

    return _PortfolioImpactGraph(
        snapshots=snapshots,
        external_dependents={
            key: tuple(
                sorted(
                    dependencies,
                    key=lambda dependency: (
                        *_node_sort_key(dependency.dependent),
                        dependency.min_column,
                        dependency.min_row,
                        dependency.max_column,
                        dependency.max_row,
                    ),
                )
            )
            for key, dependencies in external_dependents.items()
        },
    )


def _changed_candidate_roots(
    entries: Iterable[PortfolioWorkbookReport],
) -> tuple[PortfolioNode, ...]:
    """Return candidate cells whose semantic change can have downstream impact."""
    roots = {
        (entry.path, change.location)
        for entry in entries
        if entry.after is not None
        for change in entry.changes
        if change.location is not None
    }
    return tuple(sorted(roots, key=_node_sort_key))


def _materialize_impact_paths(
    roots: tuple[PortfolioNode, ...],
    nodes_by_root: dict[PortfolioNode, set[PortfolioNode]],
    parents: dict[tuple[PortfolioNode, PortfolioNode], tuple[PortfolioNode, PortfolioNode]],
) -> dict[PortfolioNode, tuple[tuple[PortfolioNode, ...], ...]]:
    """Reconstruct deterministic shortest paths after the bounded graph walk."""
    paths_by_root: dict[PortfolioNode, tuple[tuple[PortfolioNode, ...], ...]] = {}
    for root in roots:
        cross_workbook_nodes = sorted(
            (
                node
                for node in nodes_by_root.get(root, set())
                if node[0] != root[0]
            ),
            key=_node_sort_key,
        )
        paths: list[tuple[PortfolioNode, ...]] = []
        for node in cross_workbook_nodes:
            state = (root, node)
            path: list[PortfolioNode] = []
            while True:
                path.append(state[1])
                parent = parents.get(state)
                if parent is None:
                    break
                state = parent
            paths.append(tuple(reversed(path)))
        if paths:
            paths_by_root[root] = tuple(paths)
    return paths_by_root


def _traverse_cross_workbook_impacts(
    graph: _PortfolioImpactGraph,
    roots: tuple[PortfolioNode, ...],
    *,
    max_link_impact: int,
) -> _PortfolioImpactTraversal:
    """Follow candidate dependency states once, stopping before the configured cap.

    A state is a ``(changed-source, reachable-node)`` pair.  Counting pairs
    keeps evidence correct when two changed cells independently reach the same
    downstream formula, while the single global cap bounds the full portfolio
    run rather than each workbook in isolation.
    """
    queue: deque[tuple[PortfolioNode, PortfolioNode]] = deque()
    nodes_by_root: defaultdict[PortfolioNode, set[PortfolioNode]] = defaultdict(set)
    parents: dict[
        tuple[PortfolioNode, PortfolioNode], tuple[PortfolioNode, PortfolioNode]
    ] = {}
    state_count = 0

    for root in roots:
        if state_count >= max_link_impact:
            return _PortfolioImpactTraversal(
                _materialize_impact_paths(roots, nodes_by_root, parents), root
            )
        nodes_by_root[root].add(root)
        queue.append((root, root))
        state_count += 1

    while queue:
        root, node = queue.popleft()
        for dependent in graph.direct_dependents(node):
            if dependent in nodes_by_root[root]:
                continue
            if state_count >= max_link_impact:
                return _PortfolioImpactTraversal(
                    _materialize_impact_paths(roots, nodes_by_root, parents), root
                )
            nodes_by_root[root].add(dependent)
            parents[(root, dependent)] = (root, node)
            queue.append((root, dependent))
            state_count += 1

    return _PortfolioImpactTraversal(
        _materialize_impact_paths(roots, nodes_by_root, parents)
    )


def _cross_workbook_impact_finding(
    root: PortfolioNode,
    paths: tuple[tuple[PortfolioNode, ...], ...],
) -> Finding:
    """Render only audited relative identities, never the original link spelling."""
    _, root_location = root
    return Finding(
        "FF079",
        "high",
        "A changed workbook cell has statically reachable formulas in another portfolio workbook.",
        root_location,
        details={
            "impacted_workbook_count": len({path[-1][0] for path in paths}),
            "impacted_formula_count": len(paths),
            "sample_impacts": [
                {
                    "workbook": path[-1][0],
                    "location": display_location(path[-1][1]),
                    "path": [
                        {
                            "workbook": workbook,
                            "location": display_location(location),
                        }
                        for workbook, location in path
                    ],
                }
                for path in paths[:_LINK_IMPACT_SAMPLE_LIMIT]
            ],
        },
    )


def _cross_workbook_impact_limit_finding(
    root: PortfolioNode, max_link_impact: int
) -> Finding:
    """Make a capped traversal explicit instead of claiming complete evidence."""
    _, root_location = root
    return Finding(
        "FF080",
        "critical",
        (
            "Cross-workbook impact analysis reached its configured bound; "
            "portfolio impact evidence is incomplete."
        ),
        root_location,
        details={"max_link_impact": max_link_impact},
    )


def _add_cross_workbook_impact_evidence(
    entries: tuple[PortfolioWorkbookReport, ...],
    policy: Policy | None,
    *,
    max_link_impact: int,
) -> tuple[PortfolioWorkbookReport, ...]:
    """Attach bounded candidate-graph evidence to changed source workbooks."""
    graph = _build_candidate_impact_graph(entries)
    roots = _changed_candidate_roots(entries)
    if not graph.snapshots or not roots:
        return entries

    traversal = _traverse_cross_workbook_impacts(
        graph,
        roots,
        max_link_impact=max_link_impact,
    )
    raw_findings_by_workbook: defaultdict[str, list[Finding]] = defaultdict(list)
    policy_findings_by_workbook: defaultdict[str, list[Finding]] = defaultdict(list)
    for root, paths in traversal.paths_by_root.items():
        finding = _cross_workbook_impact_finding(root, paths)
        raw_findings_by_workbook[root[0]].append(finding)
        if policy is not None:
            policy_findings_by_workbook[root[0]].extend(
                evaluate_portfolio_link_policy((finding,), policy)
            )
    if traversal.incomplete_root is not None:
        root = traversal.incomplete_root
        raw_findings_by_workbook[root[0]].append(
            _cross_workbook_impact_limit_finding(root, max_link_impact)
        )

    return tuple(
        replace(
            entry,
            status=(
                "changed"
                if entry.status == "unchanged"
                and (
                    raw_findings_by_workbook[entry.path]
                    or policy_findings_by_workbook[entry.path]
                )
                else entry.status
            ),
            portfolio_findings=(
                *entry.portfolio_findings,
                *raw_findings_by_workbook[entry.path],
            ),
            policy_findings=(
                *entry.policy_findings,
                *policy_findings_by_workbook[entry.path],
            ),
        )
        for entry in entries
    )


def _safe_workbook_summary(snapshot: WorkbookSnapshot | None, path: str) -> dict[str, Any] | None:
    if snapshot is None:
        return None
    summary = snapshot.summary()
    # A directory report must stay portable and must not leak a worker's absolute
    # filesystem layout through the nested single-workbook summary.
    summary["path"] = path
    return summary


def _summary_for(
    changes: Iterable[Change],
    raw_findings: Iterable[Finding],
    policy_findings: Iterable[Finding],
) -> dict[str, Any]:
    changes = tuple(changes)
    raw_findings = tuple(raw_findings)
    policy_findings = tuple(policy_findings)
    findings = (*raw_findings, *policy_findings)
    counts = Counter(finding.severity for finding in findings)
    highest = (
        max(findings, key=lambda finding: SEVERITY_ORDER[finding.severity]).severity
        if findings
        else "note"
    )
    return {
        "change_count": len(changes),
        "finding_count": len(findings),
        "raw_finding_count": len(raw_findings),
        "policy_finding_count": len(policy_findings),
        "highest_severity": highest,
        "findings_by_severity": {
            severity: counts[severity] for severity in SEVERITY_ORDER if counts[severity]
        },
    }


@dataclass(frozen=True)
class PortfolioWorkbookReport:
    """One relative-path workbook result within a portfolio report."""

    path: str
    status: str
    baseline_present: bool
    candidate_present: bool
    before: WorkbookSnapshot | None = None
    after: WorkbookSnapshot | None = None
    report: DiffReport | None = None
    standalone_changes: tuple[Change, ...] = ()
    standalone_findings: tuple[Finding, ...] = ()
    portfolio_findings: tuple[Finding, ...] = ()
    policy_findings: tuple[Finding, ...] = ()

    @property
    def changes(self) -> tuple[Change, ...]:
        if self.report is not None:
            return tuple(self.report.changes)
        return self.standalone_changes

    @property
    def raw_findings(self) -> tuple[Finding, ...]:
        if self.report is not None:
            return (*self.report.findings, *self.portfolio_findings)
        return (*self.standalone_findings, *self.portfolio_findings)

    @property
    def findings(self) -> tuple[Finding, ...]:
        return (*self.raw_findings, *self.policy_findings)

    @property
    def incomplete(self) -> bool:
        return self.status == "unreadable" or any(
            finding.rule_id == "FF080" for finding in self.portfolio_findings
        )

    def to_dict(self) -> dict[str, Any]:
        if self.report is not None:
            payload = self.report.to_dict(
                (*self.portfolio_findings, *self.policy_findings)
            )
            payload["before"]["path"] = self.path
            payload["after"]["path"] = self.path
            payload["summary"]["raw_finding_count"] = len(self.raw_findings)
            payload["summary"]["policy_finding_count"] = len(self.policy_findings)
        else:
            payload = {
                "before": _safe_workbook_summary(self.before, self.path),
                "after": _safe_workbook_summary(self.after, self.path),
                "summary": _summary_for(
                    self.standalone_changes,
                    self.raw_findings,
                    self.policy_findings,
                ),
                "changes": [change.to_dict() for change in self.standalone_changes],
                "findings": [finding.to_dict() for finding in self.findings],
            }
        return {
            "path": self.path,
            "status": self.status,
            "baseline_present": self.baseline_present,
            "candidate_present": self.candidate_present,
            **payload,
        }


@dataclass(frozen=True)
class PortfolioReport:
    """A deterministic inventory and comparison report for two directories."""

    baseline_workbook_count: int
    candidate_workbook_count: int
    workbooks: tuple[PortfolioWorkbookReport, ...]

    @property
    def incomplete(self) -> bool:
        return any(entry.incomplete for entry in self.workbooks)

    @property
    def policy_findings(self) -> tuple[Finding, ...]:
        return tuple(
            finding
            for entry in self.workbooks
            for finding in entry.policy_findings
        )

    def severities(self) -> list[str]:
        """Return every change/finding severity for ``--fail-on`` handling."""
        return [
            *(change.severity for entry in self.workbooks for change in entry.changes),
            *(finding.severity for entry in self.workbooks for finding in entry.findings),
        ]

    def to_dict(self) -> dict[str, Any]:
        entries = [entry.to_dict() for entry in self.workbooks]
        findings = [finding for entry in self.workbooks for finding in entry.findings]
        cross_workbook_findings = [
            finding
            for entry in self.workbooks
            for finding in entry.raw_findings
            if finding.rule_id == "FF079"
        ]
        cross_workbook_analysis_incomplete = any(
            finding.rule_id == "FF080"
            for entry in self.workbooks
            for finding in entry.raw_findings
        )
        counts = Counter(finding.severity for finding in findings)
        highest = (
            max(findings, key=lambda finding: SEVERITY_ORDER[finding.severity]).severity
            if findings
            else "note"
        )
        status_counts = Counter(entry.status for entry in self.workbooks)
        return {
            "schema_version": "1.0",
            "report_type": "portfolio",
            "before": {"workbook_count": self.baseline_workbook_count},
            "after": {"workbook_count": self.candidate_workbook_count},
            "summary": {
                "matched_workbook_count": sum(
                    entry.baseline_present and entry.candidate_present
                    for entry in self.workbooks
                ),
                "unchanged_workbook_count": status_counts["unchanged"],
                "changed_workbook_count": status_counts["changed"],
                "added_workbook_count": sum(
                    not entry.baseline_present and entry.candidate_present
                    for entry in self.workbooks
                ),
                "removed_workbook_count": sum(
                    entry.baseline_present and not entry.candidate_present
                    for entry in self.workbooks
                ),
                "unreadable_workbook_count": status_counts["unreadable"],
                "change_count": sum(len(entry.changes) for entry in self.workbooks),
                "finding_count": len(findings),
                "policy_finding_count": len(self.policy_findings),
                "cross_workbook_impact_source_count": len(cross_workbook_findings),
                "cross_workbook_impacted_formula_count": sum(
                    finding.details["impacted_formula_count"]
                    for finding in cross_workbook_findings
                ),
                "cross_workbook_impact_incomplete": cross_workbook_analysis_incomplete,
                "highest_severity": highest,
                "findings_by_severity": {
                    severity: counts[severity] for severity in SEVERITY_ORDER if counts[severity]
                },
                "incomplete": self.incomplete,
            },
            "workbooks": entries,
        }


def _membership_evidence(
    status: str, policy: Policy | None
) -> tuple[tuple[Change, ...], tuple[Finding, ...], tuple[Finding, ...]]:
    """Return the evidence for a known added or removed relative path."""
    change_kind = "workbook_added" if status == "added" else "workbook_removed"
    raw_finding = Finding(
        "FF077",
        "high",
        (
            "Workbook portfolio membership changed; the workbook was "
            f"{status} and was not matched to a peer for semantic comparison."
        ),
        details={"portfolio_change": status},
    )
    policy_findings = (
        tuple(evaluate_portfolio_membership_policy((raw_finding,), policy))
        if policy is not None
        else ()
    )
    return (
        (Change(change_kind, None, "high", details={"portfolio_change": status}),),
        (raw_finding,),
        policy_findings,
    )


def _membership_entry(
    path: str,
    *,
    before: WorkbookSnapshot | None,
    after: WorkbookSnapshot | None,
    policy: Policy | None,
) -> PortfolioWorkbookReport:
    status = "added" if after is not None else "removed"
    changes, findings, policy_findings = _membership_evidence(status, policy)
    return PortfolioWorkbookReport(
        path=path,
        status=status,
        baseline_present=before is not None,
        candidate_present=after is not None,
        before=before,
        after=after,
        standalone_changes=changes,
        standalone_findings=findings,
        policy_findings=policy_findings,
    )


def _unreadable_entry(
    path: str,
    *,
    baseline_present: bool,
    candidate_present: bool,
    before: WorkbookSnapshot | None,
    after: WorkbookSnapshot | None,
    unreadable_sides: tuple[str, ...],
    policy: Policy | None,
) -> PortfolioWorkbookReport:
    membership_changes: tuple[Change, ...] = ()
    membership_findings: tuple[Finding, ...] = ()
    membership_policy_findings: tuple[Finding, ...] = ()
    if baseline_present != candidate_present:
        membership_status = "added" if candidate_present else "removed"
        (
            membership_changes,
            membership_findings,
            membership_policy_findings,
        ) = _membership_evidence(membership_status, policy)
    return PortfolioWorkbookReport(
        path=path,
        status="unreadable",
        baseline_present=baseline_present,
        candidate_present=candidate_present,
        before=before,
        after=after,
        standalone_changes=membership_changes,
        standalone_findings=(
            *membership_findings,
            Finding(
                "FF078",
                "critical",
                "Workbook could not be inspected; portfolio comparison is incomplete.",
                details={"unreadable_sides": list(unreadable_sides)},
            ),
        ),
        policy_findings=membership_policy_findings,
    )


def _load_portfolio_workbook(
    source: _PortfolioWorkbookSource | None,
) -> tuple[WorkbookSnapshot | None, bool]:
    if source is None:
        return None, False
    try:
        return (
            load_snapshot(
                source.path,
                expected_source_identity=source.identity,
            ),
            False,
        )
    except Exception:  # noqa: BLE001 - malformed workbooks must not erase portfolio evidence
        return None, True


def compare_portfolios(
    baseline_directory: str | Path,
    candidate_directory: str | Path,
    *,
    policy: Policy | None = None,
    max_workbooks: int = _DEFAULT_MAX_WORKBOOKS,
    max_inventory_entries: int = DEFAULT_MAX_INVENTORY_ENTRIES,
    max_link_impact: int = DEFAULT_MAX_LINK_IMPACT,
) -> PortfolioReport:
    """Compare every workbook at the same relative path in two directories.

    The scan is deliberately sequential and bounded.  A folder-wide run should
    be deterministic, memory-safe for CI, and explicit about unsupported,
    unreadable, or cross-workbook impact material rather than silently treating
    it as unchanged.
    """
    if max_link_impact < 1:
        raise PortfolioError("max_link_impact must be at least 1.")
    baseline = _discover_portfolio_workbook_sources(
        baseline_directory,
        label="baseline",
        max_workbooks=max_workbooks,
        max_inventory_entries=max_inventory_entries,
    )
    candidate = _discover_portfolio_workbook_sources(
        candidate_directory,
        label="candidate",
        max_workbooks=max_workbooks,
        max_inventory_entries=max_inventory_entries,
    )
    if not baseline and not candidate:
        raise PortfolioError(
            "No supported .xlsx or .xlsm workbooks were found in either portfolio."
        )

    entries: list[PortfolioWorkbookReport] = []
    for path in sorted(set(baseline) | set(candidate), key=_path_sort_key):
        baseline_path = baseline.get(path)
        candidate_path = candidate.get(path)
        before, before_unreadable = _load_portfolio_workbook(baseline_path)
        after, after_unreadable = _load_portfolio_workbook(candidate_path)
        if before_unreadable or after_unreadable:
            unreadable_sides = tuple(
                side
                for side, unreadable in (
                    ("baseline", before_unreadable),
                    ("candidate", after_unreadable),
                )
                if unreadable
            )
            entries.append(
                _unreadable_entry(
                    path,
                    baseline_present=baseline_path is not None,
                    candidate_present=candidate_path is not None,
                    before=before,
                    after=after,
                    unreadable_sides=unreadable_sides,
                    policy=policy,
                )
            )
            continue
        if before is None or after is None:
            entries.append(_membership_entry(path, before=before, after=after, policy=policy))
            continue

        report = compare_snapshots(before, after)
        policy_findings = tuple(evaluate_policy(report, policy)) if policy is not None else ()
        status = "changed" if report.changes or report.findings or policy_findings else "unchanged"
        entries.append(
            PortfolioWorkbookReport(
                path=path,
                status=status,
                baseline_present=True,
                candidate_present=True,
                before=before,
                after=after,
                report=report,
                policy_findings=policy_findings,
            )
        )

    entries = list(
        _add_cross_workbook_impact_evidence(
            tuple(entries),
            policy,
            max_link_impact=max_link_impact,
        )
    )
    return PortfolioReport(
        baseline_workbook_count=len(baseline),
        candidate_workbook_count=len(candidate),
        workbooks=tuple(entries),
    )
