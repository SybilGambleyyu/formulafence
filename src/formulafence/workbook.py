"""Safe, non-evaluating workbook inspection and dependency indexing."""

from __future__ import annotations

import base64
import binascii
import hashlib
import io
import os
import posixpath
import re
import shutil
import struct
import tempfile
import warnings
from collections import defaultdict
from collections.abc import Mapping
from dataclasses import dataclass, replace
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

from formulafence.formulas import (
    ParsedReference,
    StructuredTable,
    formula_fingerprint,
    has_broken_reference,
    inspect_formula,
    lambda_parameter_count,
    parse_reference_token,
    reference_lookup_key,
)
from formulafence.models import (
    ArrayFormulaRange,
    CellKey,
    CellProtectionAssignmentSnapshot,
    CellProtectionDefaultSnapshot,
    CellSnapshot,
    ChartDefinitionSnapshot,
    ConditionalFormattingExtensionSnapshot,
    ConditionalFormattingSnapshot,
    DataValidationSnapshot,
    DynamicArrayOutputReference,
    ExternalDataConnectionSnapshot,
    ExternalDataOpaqueMetadataSnapshot,
    ExternalDataRefreshSettingsSnapshot,
    ExternalLinkPackageSnapshot,
    OfficeWebAddinSnapshot,
    PivotCacheRefreshSnapshot,
    PivotTableDefinitionSnapshot,
    PowerQueryPermissionControlsSnapshot,
    PowerQuerySnapshot,
    ProtectedRangeSnapshot,
    ProtectionCredentialSnapshot,
    ProtectionOpaqueMetadataSnapshot,
    QueryTableRefreshSnapshot,
    RangeDependency,
    RibbonCustomizationSnapshot,
    SheetProtectionSnapshot,
    SheetSnapshot,
    TableSnapshot,
    WorkbookLoadError,
    WorkbookProtectionSnapshot,
    WorkbookSnapshot,
    WorksheetEmbeddedControlSnapshot,
    XlmMacroSheetSnapshot,
    XmlFragmentSnapshot,
    display_location,
    json_safe_value,
)

_SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
_CALCULATION_FIELDS = (
    "calcMode",
    "fullCalcOnLoad",
    "refMode",
    "iterate",
    "iterateCount",
    "iterateDelta",
    "fullPrecision",
    "calcCompleted",
    "calcOnSave",
    "concurrentCalc",
    "concurrentManualCount",
    "forceFullCalc",
)
_SPREADSHEETML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_PACKAGE_RELATIONSHIP_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_DOCUMENT_RELATIONSHIP_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_DYNAMIC_ARRAY_NS = "http://schemas.microsoft.com/office/spreadsheetml/2017/dynamicarray"
_OFFICE_2010_SPREADSHEET_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
_EXCEL_2006_MAIN_NS = "http://schemas.microsoft.com/office/excel/2006/main"
_DATA_MASHUP_NS = "http://schemas.microsoft.com/DataMashup"
_MARKUP_COMPATIBILITY_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
_XML_NAMESPACE_PREFIXES = {
    _SPREADSHEETML_NS: "",
    _OFFICE_2010_SPREADSHEET_NS: "x14:",
    _EXCEL_2006_MAIN_NS: "xm:",
}
_GUID_PATTERN = re.compile(
    r"\{[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}\}"
)
_CUSTOM_XML_ITEM_PATTERN = re.compile(r"^customXml/item(?:\d+)?\.xml$", re.IGNORECASE)
_EXTERNAL_LINK_PART_PATTERN = re.compile(
    r"^xl/externalLinks/externalLink(?:\d+)?\.xml$", re.IGNORECASE
)
_EXTERNAL_LINK_MAX_PART_BYTES = 16 * 1024 * 1024
_XLM_MACRO_SHEET_PART_PATTERN = re.compile(r"^xl/macrosheets/[^/]+\.xml$", re.IGNORECASE)
_XLM_MACRO_SHEET_MAX_PART_BYTES = 16 * 1024 * 1024
_XLM_RELATED_PART_MAX_BYTES = 32 * 1024 * 1024
_XLM_RELATED_PART_TOTAL_MAX_BYTES = 64 * 1024 * 1024
_XLM_RELATED_PART_TOTAL_MAX_COUNT = 256
_XLM_RELATED_PART_HASH_CHUNK_BYTES = 1024 * 1024
_XLM_MACRO_SHEET_RELATIONSHIPS = {
    "http://schemas.microsoft.com/office/2006/relationships/xlMacrosheet": "macro",
    "http://schemas.microsoft.com/office/2006/relationships/xlIntlMacrosheet": (
        "international"
    ),
}
_XLM_MACRO_SHEET_CONTENT_TYPES = {
    "application/vnd.ms-excel.macrosheet+xml": "macro",
    "application/vnd.ms-excel.intlmacrosheet+xml": "international",
}
_RIBBON_CUSTOM_UI_PART_PATTERN = re.compile(
    r"^customUI/[^/]+\.xml$", re.IGNORECASE
)
_RIBBON_CUSTOM_UI_MAX_PART_BYTES = 16 * 1024 * 1024
_RIBBON_CUSTOM_UI_TOTAL_MAX_BYTES = 32 * 1024 * 1024
_RIBBON_CUSTOM_UI_TOTAL_MAX_COUNT = 8
_RIBBON_CUSTOM_UI_NAMESPACES = {
    "http://schemas.microsoft.com/office/2006/01/customui": "2007",
    "http://schemas.microsoft.com/office/2007/10/customui": "2010",
    "http://schemas.microsoft.com/office/2009/07/customui": "2010",
}
_RIBBON_CUSTOM_UI_RELATIONSHIPS = {
    "http://schemas.microsoft.com/office/2006/relationships/ui/extensibility": "2007",
    "http://schemas.microsoft.com/office/2007/relationships/ui/extensibility": "2010",
}
_WEB_EXTENSION_TASKPANES_PART_PATTERN = re.compile(
    r"^xl/webextensions/taskpanes(?:\d+)?\.xml$", re.IGNORECASE
)
_WEB_EXTENSION_PART_PATTERN = re.compile(
    r"^xl/webextensions/webextension(?:\d+)?\.xml$", re.IGNORECASE
)
_WEB_EXTENSION_MAX_PART_BYTES = 16 * 1024 * 1024
_WEB_EXTENSION_TOTAL_MAX_BYTES = 32 * 1024 * 1024
_WEB_EXTENSION_TOTAL_MAX_COUNT = 64
_WEB_EXTENSION_TASKPANES_NS = (
    "http://schemas.microsoft.com/office/webextensions/taskpanes/2010/11"
)
_WEB_EXTENSION_NS = "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
_WEB_EXTENSION_TASKPANES_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2011/relationships/webextensiontaskpanes"
)
_WEB_EXTENSION_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2011/relationships/webextension"
)
_WORKSHEET_EMBEDDED_CONTROL_MAX_XML_PART_BYTES = 16 * 1024 * 1024
_WORKSHEET_EMBEDDED_CONTROL_TOTAL_XML_MAX_BYTES = 64 * 1024 * 1024
_WORKSHEET_EMBEDDED_CONTROL_TOTAL_XML_MAX_COUNT = 512
_WORKSHEET_EMBEDDED_CONTROL_RELATED_PART_MAX_BYTES = 32 * 1024 * 1024
_WORKSHEET_EMBEDDED_CONTROL_RELATED_PART_TOTAL_MAX_BYTES = 64 * 1024 * 1024
_WORKSHEET_EMBEDDED_CONTROL_RELATED_PART_TOTAL_MAX_COUNT = 512
_WORKSHEET_EMBEDDED_CONTROL_HASH_CHUNK_BYTES = 1024 * 1024
_WORKSHEET_ACTIVEX_PART_PATTERN = re.compile(
    r"^xl/activeX/[^/]+\.xml$", re.IGNORECASE
)
_WORKSHEET_CONTROL_PROPERTY_PART_PATTERN = re.compile(
    r"^xl/ctrlProps/[^/]+\.xml$", re.IGNORECASE
)
_WORKSHEET_LEGACY_VML_PART_PATTERN = re.compile(
    r"^xl/drawings/[^/]+\.vml$", re.IGNORECASE
)
_CHART_PART_PATTERN = re.compile(r"^xl/charts/[^/]+\.xml$", re.IGNORECASE)
_CHART_MAX_XML_PART_BYTES = 16 * 1024 * 1024
_CHART_TOTAL_XML_MAX_BYTES = 64 * 1024 * 1024
_CHART_TOTAL_XML_MAX_COUNT = 512
_CHART_RELATED_PART_MAX_BYTES = 32 * 1024 * 1024
_CHART_RELATED_PART_TOTAL_MAX_BYTES = 64 * 1024 * 1024
_CHART_RELATED_PART_TOTAL_MAX_COUNT = 512
_CHART_RELATED_PART_HASH_CHUNK_BYTES = 1024 * 1024
_PIVOT_TABLE_PART_PATTERN = re.compile(r"^xl/pivotTables/[^/]+\.xml$", re.IGNORECASE)
_PIVOT_CACHE_DEFINITION_PART_PATTERN = re.compile(
    r"^xl/pivotCache/pivotCacheDefinition[^/]*\.xml$", re.IGNORECASE
)
_PIVOT_CACHE_RECORDS_PART_PATTERN = re.compile(
    r"^xl/pivotCache/pivotCacheRecords[^/]*\.xml$", re.IGNORECASE
)
_PIVOT_MAX_XML_PART_BYTES = 16 * 1024 * 1024
_PIVOT_TOTAL_XML_MAX_BYTES = 64 * 1024 * 1024
_PIVOT_TOTAL_XML_MAX_COUNT = 512
_PIVOT_CACHE_RECORD_MAX_BYTES = 32 * 1024 * 1024
_PIVOT_CACHE_RECORD_TOTAL_MAX_BYTES = 64 * 1024 * 1024
_PIVOT_CACHE_RECORD_TOTAL_MAX_COUNT = 512
_PIVOT_CACHE_RECORD_HASH_CHUNK_BYTES = 1024 * 1024
_ACTIVEX_NS = "http://schemas.microsoft.com/office/2006/activeX"
_VML_NS = "urn:schemas-microsoft-com:vml"
_VML_OFFICE_NS = "urn:schemas-microsoft-com:office:office"
_VML_EXCEL_NS = "urn:schemas-microsoft-com:office:excel"
_DRAWINGML_SPREADSHEET_NS = (
    "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
)
_DRAWINGML_CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_DRAWINGML_CHART_DRAWING_NS = (
    "http://schemas.openxmlformats.org/drawingml/2006/chartDrawing"
)
_WORKSHEET_CONTROL_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/control"
_WORKSHEET_CTRLPROP_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/ctrlProp"
_WORKSHEET_OLE_OBJECT_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/oleObject"
_WORKSHEET_EMBEDDED_PACKAGE_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/package"
_WORKSHEET_VML_DRAWING_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/vmlDrawing"
_WORKSHEET_DRAWING_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/drawing"
_CHART_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/chart"
_CHART_USER_SHAPES_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/chartUserShapes"
_PIVOT_TABLE_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/pivotTable"
_PIVOT_CACHE_DEFINITION_RELATIONSHIP = (
    f"{_DOCUMENT_RELATIONSHIP_NS}/pivotCacheDefinition"
)
_PIVOT_CACHE_RECORDS_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIP_NS}/pivotCacheRecords"
_WORKSHEET_ACTIVEX_BINARY_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2006/relationships/activeXControlBinary"
)
_CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
_CHART_RELATIONSHIP_ATTRIBUTES = frozenset(
    {
        f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id",
        f"{{{_DOCUMENT_RELATIONSHIP_NS}}}embed",
        f"{{{_DOCUMENT_RELATIONSHIP_NS}}}link",
    }
)
_PIVOT_RELATIONSHIP_ATTRIBUTES = frozenset(
    {
        f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id",
        f"{{{_DOCUMENT_RELATIONSHIP_NS}}}embed",
        f"{{{_DOCUMENT_RELATIONSHIP_NS}}}link",
    }
)
_PIVOT_CACHE_DEFINITION_VOLATILE_ATTRIBUTES = frozenset(
    {
        "invalid",
        "saveData",
        "refreshOnLoad",
        "optimizeMemory",
        "enableRefresh",
        "refreshedBy",
        "refreshedDate",
        "refreshedDateIso",
        "backgroundQuery",
        "missingItemsLimit",
        "createdVersion",
        "refreshedVersion",
        "minRefreshableVersion",
        "recordCount",
        "upgradeOnRefresh",
        "tupleCache",
        "supportSubquery",
        "supportAdvancedDrill",
    }
)
_CHART_CACHE_ELEMENT_NAMES = frozenset({"numCache", "strCache", "multiLvlStrCache"})
_CHART_LITERAL_ELEMENT_NAMES = frozenset({"numLit", "strLit", "multiLvlStrLit"})
_CHART_TYPE_ELEMENT_NAMES = frozenset(
    {
        "area3DChart",
        "areaChart",
        "bar3DChart",
        "barChart",
        "bubbleChart",
        "doughnutChart",
        "line3DChart",
        "lineChart",
        "ofPieChart",
        "pie3DChart",
        "pieChart",
        "radarChart",
        "scatterChart",
        "stockChart",
        "surface3DChart",
        "surfaceChart",
    }
)


@dataclass(frozen=True)
class _ArrayFormulaMetadata:
    """Raw OOXML metadata needed to distinguish CSE from dynamic arrays."""

    dynamic_cells: set[CellKey]
    unclassified_cells: set[CellKey]
    scanned_sheets: set[str]
    complete: bool
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _ArrayFormulaClassification:
    """The conservative array-formula classifications for one loaded workbook."""

    kinds: dict[CellKey, str]
    refs: dict[CellKey, str]
    legacy_ranges: tuple[ArrayFormulaRange, ...]
    dynamic_cells: set[CellKey]
    dynamic_ranges: tuple[ArrayFormulaRange, ...]
    unclassified_cells: set[CellKey]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _ConditionalFormattingMetadata:
    """Raw worksheet conditional-formatting controls and extension evidence."""

    rules: tuple[ConditionalFormattingSnapshot, ...]
    extensions: tuple[ConditionalFormattingExtensionSnapshot, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _ProtectionMetadata:
    """Raw OOXML protection controls retained before libraries can omit them."""

    workbook_protection: WorkbookProtectionSnapshot | None
    sheet_protections: tuple[SheetProtectionSnapshot, ...]
    protected_ranges: tuple[ProtectedRangeSnapshot, ...]
    cell_protection_default: CellProtectionDefaultSnapshot | None
    cell_protection_assignments: tuple[CellProtectionAssignmentSnapshot, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _ExternalDataMetadata:
    """Raw OOXML external-data controls omitted by the workbook reader."""

    refresh_settings: ExternalDataRefreshSettingsSnapshot
    connections: tuple[ExternalDataConnectionSnapshot, ...]
    query_tables: tuple[QueryTableRefreshSnapshot, ...]
    pivot_caches: tuple[PivotCacheRefreshSnapshot, ...]
    external_link_packages: ExternalLinkPackageSnapshot
    power_query: PowerQuerySnapshot
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _XlmMacroMetadata:
    """Raw XLM macro-sheet evidence retained before the workbook reader omits it."""

    macro_sheets: XlmMacroSheetSnapshot
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _RibbonCustomizationMetadata:
    """Raw Office RibbonX evidence retained outside the workbook reader."""

    customization: RibbonCustomizationSnapshot
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _OfficeWebAddinMetadata:
    """Raw Office Web Add-in task-pane evidence outside the workbook reader."""

    addins: OfficeWebAddinSnapshot
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _ChartDefinitionMetadata:
    """Raw chart evidence retained before the workbook reader can omit it."""

    charts: ChartDefinitionSnapshot
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _PivotTableMetadata:
    """Raw PivotTable evidence retained before the workbook reader can omit it."""

    pivot_tables: PivotTableDefinitionSnapshot
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _WorksheetEmbeddedControlMetadata:
    """Raw worksheet control evidence retained before the workbook reader omits it."""

    controls: WorksheetEmbeddedControlSnapshot
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _XlmRawRelationship:
    """One private package relationship, including the original target material."""

    relationship_id: str | None
    relationship_type: str
    target: str | None
    target_mode: str
    safe_target: str | None

    def semantic_key(self) -> tuple[str, str, str]:
        """Return canonical target semantics without writer-chosen identifiers."""
        target = self.target or ""
        if self.target_mode.casefold() == "internal" and self.safe_target is not None:
            target = self.safe_target
        return (
            self.relationship_type,
            self.target_mode.casefold(),
            target,
        )


@dataclass(frozen=True)
class _RibbonRawRelationship:
    """One private RibbonX package relationship and its canonical target."""

    relationship_id: str | None
    relationship_type: str
    target: str | None
    target_mode: str
    safe_target: str | None

    def semantic_key(self) -> tuple[str, str, str]:
        """Return relationship semantics while ignoring arbitrary identifiers."""
        target = self.target or ""
        if self.target_mode.casefold() == "internal" and self.safe_target is not None:
            target = self.safe_target
        return (
            self.relationship_type,
            self.target_mode.casefold(),
            target,
        )


@dataclass(frozen=True)
class _OfficeWebAddinRawRelationship:
    """One private Office Web Add-in relationship and canonical target."""

    relationship_id: str | None
    relationship_type: str
    target: str | None
    target_mode: str
    safe_target: str | None

    def semantic_key(self) -> tuple[str, str, str]:
        """Return relationship semantics while ignoring arbitrary identifiers."""
        target = self.target or ""
        if self.target_mode.casefold() == "internal" and self.safe_target is not None:
            target = self.safe_target
        return (
            self.relationship_type,
            self.target_mode.casefold(),
            target,
        )


@dataclass(frozen=True)
class _WorksheetControlRawRelationship:
    """One private worksheet-control relationship and canonical target."""

    relationship_id: str | None
    relationship_type: str
    target: str | None
    target_mode: str
    safe_target: str | None

    def semantic_key(self) -> tuple[str, str, str]:
        """Return relationship semantics while ignoring arbitrary identifiers."""
        target = self.target or ""
        if self.target_mode.casefold() == "internal" and self.safe_target is not None:
            target = self.safe_target
        return (
            self.relationship_type,
            self.target_mode.casefold(),
            target,
        )


@dataclass(frozen=True)
class _ChartRawRelationship:
    """One private chart-package relationship and its canonical target."""

    relationship_id: str | None
    relationship_type: str
    target: str | None
    target_mode: str
    safe_target: str | None

    def semantic_key(self) -> tuple[str, str, str]:
        """Return relationship semantics while ignoring arbitrary identifiers."""
        target = self.target or ""
        if self.target_mode.casefold() == "internal" and self.safe_target is not None:
            target = self.safe_target
        return (
            self.relationship_type,
            self.target_mode.casefold(),
            target,
        )


@dataclass(frozen=True)
class _PivotRawRelationship:
    """One private PivotTable-package relationship and its canonical target."""

    relationship_id: str | None
    relationship_type: str
    target: str | None
    target_mode: str
    safe_target: str | None

    def semantic_key(self) -> tuple[str, str, str]:
        """Return relationship semantics while ignoring arbitrary identifiers."""
        target = self.target or ""
        if self.target_mode.casefold() == "internal" and self.safe_target is not None:
            target = self.safe_target
        return (
            self.relationship_type,
            self.target_mode.casefold(),
            target,
        )


@dataclass
class _RibbonCustomizationBudget:
    """Bound total custom-UI bytes read across one RibbonX package scan."""

    remaining_bytes: int = _RIBBON_CUSTOM_UI_TOTAL_MAX_BYTES
    remaining_parts: int = _RIBBON_CUSTOM_UI_TOTAL_MAX_COUNT


@dataclass
class _OfficeWebAddinBudget:
    """Bound total task-pane and web-extension XML bytes in one package scan."""

    remaining_bytes: int = _WEB_EXTENSION_TOTAL_MAX_BYTES
    remaining_parts: int = _WEB_EXTENSION_TOTAL_MAX_COUNT


@dataclass
class _ChartXmlBudget:
    """Bound chart, drawing, and overlay XML bytes in one package scan."""

    remaining_bytes: int = _CHART_TOTAL_XML_MAX_BYTES
    remaining_parts: int = _CHART_TOTAL_XML_MAX_COUNT


@dataclass
class _ChartRelatedPartBudget:
    """Bound direct chart-presentation payload bytes in one package scan."""

    remaining_bytes: int = _CHART_RELATED_PART_TOTAL_MAX_BYTES
    remaining_parts: int = _CHART_RELATED_PART_TOTAL_MAX_COUNT


@dataclass
class _PivotXmlBudget:
    """Bound PivotTable and cache-definition XML bytes in one package scan."""

    remaining_bytes: int = _PIVOT_TOTAL_XML_MAX_BYTES
    remaining_parts: int = _PIVOT_TOTAL_XML_MAX_COUNT


@dataclass
class _PivotCacheRecordBudget:
    """Bound raw PivotTable cache-record bytes in one package scan."""

    remaining_bytes: int = _PIVOT_CACHE_RECORD_TOTAL_MAX_BYTES
    remaining_parts: int = _PIVOT_CACHE_RECORD_TOTAL_MAX_COUNT


@dataclass
class _WorksheetEmbeddedControlXmlBudget:
    """Bound worksheet-control XML bytes read across one package scan."""

    remaining_bytes: int = _WORKSHEET_EMBEDDED_CONTROL_TOTAL_XML_MAX_BYTES
    remaining_parts: int = _WORKSHEET_EMBEDDED_CONTROL_TOTAL_XML_MAX_COUNT


@dataclass
class _WorksheetEmbeddedControlRelatedPartBudget:
    """Bound direct embedded-control payload bytes across one package scan."""

    remaining_bytes: int = _WORKSHEET_EMBEDDED_CONTROL_RELATED_PART_TOTAL_MAX_BYTES
    remaining_parts: int = _WORKSHEET_EMBEDDED_CONTROL_RELATED_PART_TOTAL_MAX_COUNT


@dataclass
class _XlmRelatedPartBudget:
    """Bound total related-part bytes read across one XLM package scan."""

    remaining_bytes: int = _XLM_RELATED_PART_TOTAL_MAX_BYTES
    remaining_parts: int = _XLM_RELATED_PART_TOTAL_MAX_COUNT


@dataclass(frozen=True)
class _XlmRelatedPartPayloadInspection:
    """Private fingerprint result for direct, internal XLM related parts."""

    internal_part_count: int = 0
    fingerprinted_part_count: int = 0
    uninspected_part_count: int = 0
    payload_signature: str | None = None


@dataclass(frozen=True)
class _XlmMacroSheetInspection:
    """Private parsed state for one candidate XLM macro-sheet part."""

    member: str
    formula_cell_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    internal_related_part_count: int = 0
    fingerprinted_related_part_count: int = 0
    uninspected_related_part_count: int = 0
    embedded_object_relationship_count: int = 0
    embedded_package_relationship_count: int = 0
    inspected: bool = False
    program_signature: str | None = None
    relationship_signature: str | None = None
    related_part_payload_signature: str | None = None


@dataclass(frozen=True)
class _RibbonPartInspection:
    """Private parsed state for one RibbonX customization part."""

    member: str
    office_2010: bool = False
    control_count: int = 0
    callback_attribute_count: int = 0
    action_callback_count: int = 0
    image_relationship_count: int = 0
    external_relationship_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _OfficeWebAddinTaskpaneInspection:
    """Private parsed state for one Office Web Add-in task-pane part."""

    member: str
    taskpane_count: int = 0
    visible_taskpane_count: int = 0
    locked_taskpane_count: int = 0
    web_extension_reference_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    declared_web_extension_members: tuple[str, ...] = ()
    unresolved_binding_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _OfficeWebAddinExtensionInspection:
    """Private parsed state for one Office Web Add-in definition part."""

    member: str
    auto_show_taskpane_count: int = 0
    store_reference_count: int = 0
    alternate_reference_count: int = 0
    binding_count: int = 0
    snapshot_reference_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    unresolved_snapshot_reference_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _WorksheetControlRelatedPartPayloadInspection:
    """Private fingerprint result for direct embedded-control payload parts."""

    internal_part_count: int = 0
    fingerprinted_part_count: int = 0
    uninspected_part_count: int = 0
    payload_signature: str | None = None


@dataclass(frozen=True)
class _ChartRelatedPartPayloadInspection:
    """Private fingerprint result for direct chart-presentation payloads."""

    internal_part_count: int = 0
    fingerprinted_part_count: int = 0
    uninspected_part_count: int = 0
    payload_signature: str | None = None


@dataclass(frozen=True)
class _PivotCacheRecordPayloadInspection:
    """Private fingerprint result for raw PivotTable cache-record parts."""

    record_part_count: int = 0
    fingerprinted_part_count: int = 0
    uninspected_part_count: int = 0
    payload_signature: str | None = None


@dataclass(frozen=True)
class _ChartDrawingInspection:
    """Private chart bindings discovered in one worksheet/chart-sheet drawing."""

    member: str
    present: bool = False
    chart_reference_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    chart_members: tuple[str, ...] = ()
    unrecognized_count: int = 0
    inspected: bool = False
    declaration_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _ChartPartInspection:
    """Private parsed state for one standard DrawingML chart part."""

    member: str
    chart_type_count: int = 0
    series_count: int = 0
    title_count: int = 0
    data_reference_count: int = 0
    numeric_data_reference_count: int = 0
    string_data_reference_count: int = 0
    literal_data_point_count: int = 0
    cached_data_point_count: int = 0
    pivot_source_count: int = 0
    external_data_reference_count: int = 0
    user_shape_reference_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    user_shape_members: tuple[str, ...] = ()
    payload_members: tuple[str, ...] = ()
    unresolved_payload_entries: tuple[tuple[str, str], ...] = ()
    unrecognized_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    cached_data_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _ChartUserShapeInspection:
    """Private parsed state for one chart-overlay drawing part."""

    member: str
    shape_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    payload_members: tuple[str, ...] = ()
    unresolved_payload_entries: tuple[tuple[str, str], ...] = ()
    unrecognized_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _PivotTablePartInspection:
    """Private parsed state for one PivotTable view-definition part."""

    member: str
    cache_id: str | None = None
    layout_location_count: int = 0
    pivot_field_count: int = 0
    row_field_count: int = 0
    column_field_count: int = 0
    page_field_count: int = 0
    data_field_count: int = 0
    filter_count: int = 0
    row_item_count: int = 0
    column_item_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    cache_definition_members: tuple[str, ...] = ()
    unrecognized_count: int = 0
    inspected: bool = False
    layout_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _PivotCacheDefinitionInspection:
    """Private parsed state for one PivotTable cache-definition part."""

    member: str
    cache_field_count: int = 0
    shared_item_count: int = 0
    calculated_item_count: int = 0
    calculated_member_count: int = 0
    cache_record_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    cache_record_members: tuple[str, ...] = ()
    unrecognized_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    cached_shared_item_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _WorksheetControlSheetInspection:
    """Private parsed worksheet controls and OLE declarations for one sheet."""

    member: str
    present: bool = False
    worksheet_control_count: int = 0
    control_macro_assignment_count: int = 0
    control_cell_link_count: int = 0
    control_source_range_count: int = 0
    ole_object_count: int = 0
    linked_ole_object_count: int = 0
    auto_load_ole_object_count: int = 0
    auto_update_ole_object_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    active_x_members: tuple[str, ...] = ()
    control_property_members: tuple[str, ...] = ()
    legacy_vml_members: tuple[str, ...] = ()
    legacy_vml_relationships: tuple[_WorksheetControlRawRelationship, ...] = ()
    legacy_vml_unrecognized_count: int = 0
    payload_members: tuple[str, ...] = ()
    unresolved_payload_entries: tuple[tuple[str, str], ...] = ()
    unrecognized_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _WorksheetControlActiveXInspection:
    """Private parsed state for one worksheet ActiveX persistence part."""

    member: str
    binary_reference_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    payload_members: tuple[str, ...] = ()
    unresolved_payload_entries: tuple[tuple[str, str], ...] = ()
    unrecognized_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _WorksheetControlPropertyInspection:
    """Private parsed state for one form-control properties part."""

    member: str
    formula_binding_count: int = 0
    cell_link_count: int = 0
    source_range_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    payload_members: tuple[str, ...] = ()
    unresolved_payload_entries: tuple[tuple[str, str], ...] = ()
    unrecognized_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _WorksheetVmlDrawingInspection:
    """Private parsed state for one legacy worksheet VML drawing part."""

    member: str
    present: bool = False
    control_count: int = 0
    macro_assignment_count: int = 0
    cell_link_count: int = 0
    source_range_count: int = 0
    camera_source_range_count: int = 0
    related_relationship_count: int = 0
    external_relationship_count: int = 0
    unrecognized_count: int = 0
    inspected: bool = False
    definition_signature: str | None = None
    relationship_signature: str | None = None


@dataclass(frozen=True)
class _ExternalLinkPartInspection:
    """Private parsed state for one externalLink OOXML part."""

    member: str = ""
    kind: str = "unrecognized"
    external_workbook_sheet_count: int = 0
    external_defined_name_count: int = 0
    external_workbook_cached_sheet_count: int = 0
    external_workbook_cached_cell_count: int = 0
    external_workbook_cached_refresh_error_count: int = 0
    dde_item_count: int = 0
    dde_advise_item_count: int = 0
    dde_ole_item_count: int = 0
    dde_prefer_picture_item_count: int = 0
    dde_cached_value_count: int = 0
    ole_item_count: int = 0
    ole_advise_item_count: int = 0
    ole_icon_item_count: int = 0
    ole_prefer_picture_item_count: int = 0
    source_signature: str | None = None
    definition_signature: str | None = None
    cached_material_signature: str | None = None
    opaque_entries: tuple[tuple[str, str], ...] = ()


@dataclass(frozen=True)
class _PowerQueryMashupInspection:
    """Private parsed state for one Data Mashup custom XML payload."""

    parsed: bool = False
    formula_document_count: int = 0
    package_part_count: int = 0
    embedded_content_part_count: int = 0
    metadata_document_count: int = 0
    metadata_item_count: int = 0
    permission_payload_count: int = 0
    permission_parsed_count: int = 0
    firewall_enabled_count: int = 0
    future_packages_allowed_count: int = 0
    workbook_group_type_count: int = 0
    permission_binding_present: bool = False
    formula_signature: str | None = None
    package_configuration_signature: str | None = None
    metadata_identity_signature: str | None = None
    metadata_control_signature: str | None = None
    permission_signature: str | None = None
    opaque_entries: tuple[tuple[str, str], ...] = ()
    permission_opaque_entries: tuple[tuple[str, str], ...] = ()


@dataclass(frozen=True)
class _PackageRelationship:
    """One package relationship with an already-safe internal target."""

    relationship_id: str | None
    relationship_type: str
    target: str | None
    target_mode: str


@dataclass(frozen=True)
class _StyleProtection:
    """Effective protection from one cell XF, plus whether it is explicit."""

    locked: bool
    hidden: bool
    explicit: bool = False


_SHEET_PROTECTION_ACTIONS = (
    ("objects", "objects", False),
    ("scenarios", "scenarios", False),
    ("formatCells", "format_cells", True),
    ("formatColumns", "format_columns", True),
    ("formatRows", "format_rows", True),
    ("insertColumns", "insert_columns", True),
    ("insertRows", "insert_rows", True),
    ("insertHyperlinks", "insert_hyperlinks", True),
    ("deleteColumns", "delete_columns", True),
    ("deleteRows", "delete_rows", True),
    ("selectLockedCells", "select_locked_cells", False),
    ("sort", "sort", True),
    ("autoFilter", "auto_filter", True),
    ("pivotTables", "pivot_tables", True),
    ("selectUnlockedCells", "select_unlocked_cells", False),
)
_CHART_SHEET_PROTECTION_ACTIONS = (
    ("content", "content", False),
    ("objects", "objects", False),
)
_EXTERNAL_CONNECTION_TYPES = {
    1: "odbc",
    2: "dao",
    3: "file_database",
    4: "web_query",
    5: "ole_db",
    6: "text",
    7: "ado_recordset",
    8: "dsp",
}
_EXTERNAL_RECONNECTION_METHODS = {
    1: "as_required",
    2: "always",
    3: "never",
}
_EXTERNAL_CREDENTIAL_METHODS = {
    "integrated": "integrated",
    "none": "none",
    "stored": "stored",
    "prompt": "prompt",
}
_QUERY_TABLE_GROWTH_BEHAVIORS = {
    "insertClear": "insert_clear",
    "insertDelete": "insert_delete",
    "overwriteClear": "overwrite_clear",
}
_PIVOT_CACHE_SOURCE_TYPES = {
    "worksheet": "worksheet",
    "external": "external",
    "consolidation": "consolidation",
    "scenario": "scenario",
}
_POWER_QUERY_STABLE_METADATA_TYPES = frozenset(
    {
        "AddedToDataModel",
        "BufferNextRefresh",
        "FillEnabled",
        "FillObjectType",
        "FillTarget",
        "FillTargetNameCustomized",
        "FillToDataModelEnabled",
        "IsFunctionQuery",
        "IsPrivate",
        "IsRelationshipDetectionEnabled",
        "NameUpdatedAfterFill",
        "QueryGroups",
        "Relationships",
        "ResultType",
    }
)
_POWER_QUERY_VOLATILE_METADATA_TYPES = frozenset(
    {
        "FillColumnNames",
        "FillColumnTypes",
        "FillCount",
        "FillErrorCode",
        "FillErrorCount",
        "FillErrorMessage",
        "FilledCompleteResultToWorksheet",
        "FillLastUpdated",
        "FillStatus",
        "PublishedPackageID",
        "PublishedPackageLastModifiedAt",
        "QueryGroupID",
        "QueryID",
        "RecoveryTargetColumn",
        "RecoveryTargetRow",
        "RecoveryTargetSheet",
        "RelationshipInfoContainer",
    }
)
_POWER_QUERY_FORMULA_MEMBER = "Formulas/Section1.m"
_POWER_QUERY_MAX_FORMULA_BYTES = 16 * 1024 * 1024


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_handle:
        for block in iter(lambda: file_handle.read(1_048_576), b""):
            digest.update(block)
    return digest.hexdigest()


def _vba_hash(path: Path) -> str | None:
    """Hash the macro payload without loading or executing it."""
    try:
        with ZipFile(path) as archive:
            try:
                payload = archive.read("xl/vbaProject.bin")
            except KeyError:
                return None
    except BadZipFile as error:
        raise WorkbookLoadError(f"{path} is not a valid Office Open XML workbook") from error
    return hashlib.sha256(payload).hexdigest()


def _xml_root(archive: ZipFile, member: str) -> ElementTree.Element:
    """Read one OOXML part without accepting document type declarations."""
    return _xml_root_from_payload(archive.read(member))


def _xml_root_from_payload(payload: bytes) -> ElementTree.Element:
    """Parse untrusted XML payload without accepting document type declarations."""
    if b"<!DOCTYPE" in payload or b"<!ENTITY" in payload:
        raise ValueError("OOXML metadata contains a document type declaration")
    return ElementTree.fromstring(payload)


def _normalise_package_target(target: str) -> str | None:
    """Turn a workbook relationship target into a safe ZIP member name."""
    return _normalise_part_target("xl/workbook.xml", target)


def _normalise_part_target(source_member: str, target: str) -> str | None:
    """Resolve one internal OOXML relationship target inside the ZIP package."""
    candidate = (
        target.lstrip("/")
        if target.startswith("/")
        else posixpath.join(posixpath.dirname(source_member), target)
    )
    normalised = posixpath.normpath(candidate)
    if normalised == ".." or normalised.startswith("../"):
        return None
    return normalised


def _relationship_part_path(source_member: str) -> str:
    """Return the OPC relationship-part member for one package part."""
    return posixpath.join(
        posixpath.dirname(source_member),
        "_rels",
        f"{posixpath.basename(source_member)}.rels",
    )


def _package_relationships(
    archive: ZipFile,
    source_member: str,
) -> tuple[_PackageRelationship, ...]:
    """Read package relationships without following external targets."""
    relationships = _xml_root(archive, _relationship_part_path(source_member))
    relationship_tag = f"{{{_PACKAGE_RELATIONSHIP_NS}}}Relationship"
    parsed: list[_PackageRelationship] = []
    for relationship in relationships.findall(relationship_tag):
        target_mode = relationship.get("TargetMode", "Internal")
        target = relationship.get("Target")
        parsed.append(
            _PackageRelationship(
                relationship_id=relationship.get("Id"),
                relationship_type=relationship.get("Type", ""),
                target=(
                    _normalise_part_target(source_member, target)
                    if target is not None and target_mode.casefold() == "internal"
                    else None
                ),
                target_mode=target_mode,
            )
        )
    return tuple(parsed)


def _sheet_xml_parts(archive: ZipFile) -> dict[str, tuple[str, str]]:
    """Map workbook sheet titles to safe OOXML parts and their sheet kind."""
    workbook = _xml_root(archive, "xl/workbook.xml")
    relationship_targets: dict[str, tuple[str, str]] = {}
    for relationship in _package_relationships(archive, "xl/workbook.xml"):
        sheet_type = relationship.relationship_type.rsplit("/", maxsplit=1)[-1]
        if (
            sheet_type not in {"worksheet", "chartsheet", "dialogsheet"}
            or not relationship.relationship_id
            or relationship.target is None
        ):
            continue
        relationship_targets[relationship.relationship_id] = (
            relationship.target,
            sheet_type,
        )

    sheet_parts: dict[str, tuple[str, str]] = {}
    sheet_tag = f"{{{_SPREADSHEETML_NS}}}sheet"
    relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
    for sheet in workbook.iter(sheet_tag):
        title = sheet.get("name")
        relationship_id = sheet.get(relationship_id_attribute)
        if not title or not relationship_id:
            continue
        if part := relationship_targets.get(relationship_id):
            sheet_parts[title] = part
    return sheet_parts


def _worksheet_xml_paths(archive: ZipFile) -> dict[str, str]:
    """Map standard worksheet titles to their OOXML worksheet parts."""
    return {
        title: member
        for title, (member, sheet_type) in _sheet_xml_parts(archive).items()
        if sheet_type == "worksheet"
    }


def _xml_local_name(tag: str) -> str:
    """Return an XML local name without discarding a namespace elsewhere."""
    return tag.rsplit("}", maxsplit=1)[-1]


def _xml_namespace(tag: str) -> str | None:
    """Return an XML namespace, if ``tag`` is namespace-qualified."""
    if not tag.startswith("{"):
        return None
    return tag[1:].split("}", maxsplit=1)[0]


def _xml_display_name(tag: str) -> str:
    """Render a qualified XML name deterministically for local review evidence."""
    if not tag.startswith("{"):
        return tag
    namespace, local_name = tag[1:].split("}", maxsplit=1)
    if prefix := _XML_NAMESPACE_PREFIXES.get(namespace):
        return f"{prefix}{local_name}"
    if namespace == _SPREADSHEETML_NS:
        return local_name
    return f"{{{namespace}}}{local_name}"


def _normalise_guid(value: str) -> str:
    """Remove writer-specific GUID noise while retaining extension structure."""
    return _GUID_PATTERN.sub("{GUID}", value)


def _normalise_conditional_formula(value: str) -> str:
    """Accept the optional leading-equals spelling used by workbook writers."""
    return value[1:] if value.startswith("=") else value


def _is_conditional_guid_link(
    element: ElementTree.Element,
    attribute: str | None = None,
) -> bool:
    """Return whether an x14 GUID only links a base rule to its extension.

    An ``ext`` element's ``uri`` is a semantic extension-type identifier and
    must never be normalised away. The known ``x14:id`` / ``x14:cfRule@id``
    pair, by contrast, is writer-generated linkage between equivalent base and
    extension rule records.
    """
    if _xml_namespace(element.tag) != _OFFICE_2010_SPREADSHEET_NS:
        return False
    local_name = _xml_local_name(element.tag)
    if attribute is None:
        return local_name == "id"
    return local_name == "cfRule" and _xml_local_name(attribute) == "id"


def _xml_fragment(
    element: ElementTree.Element,
    *,
    normalise_guids: bool = False,
    normalise_formulas: bool = False,
) -> XmlFragmentSnapshot:
    """Capture an XML subtree with deterministic names and attribute ordering."""
    local_name = _xml_local_name(element.tag)
    formula_value = normalise_formulas and local_name in {"f", "formula"}
    value_object = normalise_formulas and local_name == "cfvo" and element.get("type") == "formula"

    def normalise_value(
        value: str,
        *,
        formula: bool = False,
        guid_link: bool = False,
    ) -> str:
        if formula:
            value = _normalise_conditional_formula(value)
        return _normalise_guid(value) if normalise_guids and guid_link else value

    attributes = tuple(
        sorted(
            (
                _xml_display_name(attribute),
                normalise_value(
                    value,
                    formula=value_object and _xml_local_name(attribute) == "val",
                    guid_link=_is_conditional_guid_link(element, attribute),
                ),
            )
            for attribute, value in element.attrib.items()
        )
    )
    children = tuple(
        _xml_fragment(
            child,
            normalise_guids=normalise_guids,
            normalise_formulas=normalise_formulas,
        )
        for child in element
    )
    text = element.text
    if children and text is not None and not text.strip():
        text = None
    if text is not None:
        text = normalise_value(
            text,
            formula=formula_value,
            guid_link=_is_conditional_guid_link(element),
        )
    return XmlFragmentSnapshot(
        tag=_xml_display_name(element.tag),
        attributes=attributes,
        text=text,
        children=children,
    )


def _conditional_ranges(value: str | None) -> tuple[str, ...]:
    """Return an order-independent, compact inventory of an ``sqref`` value."""
    return tuple(sorted({part for part in (value or "").split() if part}, key=str.casefold))


def _conditional_bool(
    value: str | None,
    default: bool,
    attribute: str,
    warnings: set[str],
) -> bool:
    """Read one OOXML boolean and preserve the schema default when omitted."""
    if value is None:
        return default
    lowered = value.casefold()
    if lowered in {"1", "true"}:
        return True
    if lowered in {"0", "false"}:
        return False
    warnings.add(
        "FormulaFence could not interpret a conditional-formatting "
        f"{attribute} boolean; the schema default was used."
    )
    return default


def _conditional_int(
    element: ElementTree.Element,
    attribute: str,
    warnings: set[str],
) -> int | None:
    """Read an optional integer attribute without making malformed XML look safe."""
    value = element.get(attribute)
    if value is None:
        return None
    try:
        return int(value)
    except ValueError:
        warnings.add(
            "FormulaFence could not interpret a conditional-formatting "
            f"{attribute} integer; the affected rule has a coverage gap."
        )
        return None


def _differential_style_fragments(archive: ZipFile) -> tuple[XmlFragmentSnapshot, ...]:
    """Resolve conditional-format ``dxfId`` values without trusting their order."""
    try:
        styles = _xml_root(archive, "xl/styles.xml")
    except KeyError:
        return ()
    dxfs = styles.find(f"{{{_SPREADSHEETML_NS}}}dxfs")
    if dxfs is None:
        return ()
    return tuple(
        _xml_fragment(dxf, normalise_formulas=True)
        for dxf in dxfs.findall(f"{{{_SPREADSHEETML_NS}}}dxf")
    )


def _conditional_rule_snapshot(
    sheet: str,
    ranges: tuple[str, ...],
    rule: ElementTree.Element,
    differential_styles: tuple[XmlFragmentSnapshot, ...],
    warnings: set[str],
) -> ConditionalFormattingSnapshot:
    """Read one base OOXML ``cfRule`` without calculating its formula."""
    raw_priority = _conditional_int(rule, "priority", warnings)
    if raw_priority is None or raw_priority <= 0:
        warnings.add(
            "FormulaFence could not read a valid conditional-formatting priority; "
            "the affected rule has a coverage gap."
        )
        raw_priority = 0

    rule_type = rule.get("type") or "unknown"
    if rule_type == "unknown":
        warnings.add(
            "FormulaFence found a conditional-formatting rule without a type; "
            "the affected rule has a coverage gap."
        )
    formula_tag = f"{{{_SPREADSHEETML_NS}}}formula"
    formulas = tuple(
        _normalise_conditional_formula(formula.text or "")
        for formula in rule.findall(formula_tag)
    )
    differential_style: XmlFragmentSnapshot | None = None
    dxf_id = rule.get("dxfId")
    if dxf_id is not None:
        parsed_dxf_id = _conditional_int(rule, "dxfId", warnings)
        if parsed_dxf_id is None or not 0 <= parsed_dxf_id < len(differential_styles):
            warnings.add(
                "FormulaFence could not resolve a conditional-formatting differential "
                "style; the affected rule has a coverage gap."
            )
            differential_style = XmlFragmentSnapshot(
                "unresolved-dxf", (("id", dxf_id),)
            )
        else:
            differential_style = differential_styles[parsed_dxf_id]

    def component(name: str) -> XmlFragmentSnapshot | None:
        found = rule.find(f"{{{_SPREADSHEETML_NS}}}{name}")
        return (
            _xml_fragment(found, normalise_formulas=True)
            if found is not None
            else None
        )

    extension_list = rule.find(f"{{{_SPREADSHEETML_NS}}}extLst")
    extensions = (
        tuple(
            _xml_fragment(
                extension,
                normalise_guids=True,
                normalise_formulas=True,
            )
            for extension in extension_list
        )
        if extension_list is not None
        else ()
    )
    known_attributes = {
        "type",
        "priority",
        "dxfId",
        "stopIfTrue",
        "aboveAverage",
        "percent",
        "bottom",
        "operator",
        "text",
        "timePeriod",
        "rank",
        "stdDev",
        "equalAverage",
    }
    unmodelled_attributes = tuple(
        sorted(
            (
                _xml_display_name(attribute),
                value,
            )
            for attribute, value in rule.attrib.items()
            if _xml_local_name(attribute) not in known_attributes
        )
    )
    if unmodelled_attributes:
        warnings.add(
            "FormulaFence found unmodelled conditional-formatting rule attributes; "
            "their raw structure is retained for review."
        )
        extensions += (
            XmlFragmentSnapshot("unmodelled-cf-rule-attributes", unmodelled_attributes),
        )
    known_children = {"formula", "colorScale", "dataBar", "iconSet", "extLst"}
    unmodelled_children = tuple(
        _xml_fragment(child, normalise_guids=True, normalise_formulas=True)
        for child in rule
        if _xml_local_name(child.tag) not in known_children
    )
    if unmodelled_children:
        warnings.add(
            "FormulaFence found unmodelled conditional-formatting rule children; "
            "their raw structure is retained for review."
        )
        extensions += unmodelled_children

    return ConditionalFormattingSnapshot(
        sheet=sheet,
        ranges=ranges,
        priority=raw_priority,
        rule_type=rule_type,
        operator=rule.get("operator"),
        formulas=formulas,
        stop_if_true=_conditional_bool(rule.get("stopIfTrue"), False, "stopIfTrue", warnings),
        above_average=_conditional_bool(
            rule.get("aboveAverage"), True, "aboveAverage", warnings
        ),
        percent=_conditional_bool(rule.get("percent"), False, "percent", warnings),
        bottom=_conditional_bool(rule.get("bottom"), False, "bottom", warnings),
        rank=_conditional_int(rule, "rank", warnings),
        std_dev=_conditional_int(rule, "stdDev", warnings),
        equal_average=_conditional_bool(
            rule.get("equalAverage"), False, "equalAverage", warnings
        ),
        text=rule.get("text"),
        time_period=rule.get("timePeriod"),
        differential_style=differential_style,
        color_scale=component("colorScale"),
        data_bar=component("dataBar"),
        icon_set=component("iconSet"),
        extensions=extensions,
    )


def _worksheet_conditional_formatting_extensions(
    sheet: str,
    worksheet: ElementTree.Element,
) -> tuple[ConditionalFormattingExtensionSnapshot, ...]:
    """Keep worksheet-level x14 conditional-formatting extensions inspectable."""
    extension_snapshots: list[ConditionalFormattingExtensionSnapshot] = []
    extension_list_tag = f"{{{_SPREADSHEETML_NS}}}extLst"
    for extension_list in worksheet.findall(extension_list_tag):
        for extension in extension_list:
            if not any(
                _xml_local_name(element.tag)
                in {"conditionalFormattings", "conditionalFormatting", "cfRule"}
                for element in extension.iter()
            ):
                continue
            extension_snapshots.append(
                ConditionalFormattingExtensionSnapshot(
                    sheet=sheet,
                    fragment=_xml_fragment(
                        extension,
                        normalise_guids=True,
                        normalise_formulas=True,
                    ),
                )
            )
    return tuple(
        sorted(extension_snapshots, key=ConditionalFormattingExtensionSnapshot.sort_key)
    )


def _conditional_formatting_metadata(path: Path) -> _ConditionalFormattingMetadata:
    """Read conditional formatting from OOXML before a library can discard extensions."""
    rules_by_sheet: dict[str, list[tuple[int, int, ConditionalFormattingSnapshot]]] = (
        defaultdict(list)
    )
    extensions: list[ConditionalFormattingExtensionSnapshot] = []
    warnings: set[str] = set()
    try:
        with ZipFile(path) as archive:
            differential_styles = _differential_style_fragments(archive)
            conditional_formatting_tag = f"{{{_SPREADSHEETML_NS}}}conditionalFormatting"
            rule_tag = f"{{{_SPREADSHEETML_NS}}}cfRule"
            for sheet, member in _worksheet_xml_paths(archive).items():
                worksheet = _xml_root(archive, member)
                source_order = 0
                for conditional_formatting in worksheet.findall(conditional_formatting_tag):
                    ranges = _conditional_ranges(conditional_formatting.get("sqref"))
                    if not ranges:
                        warnings.add(
                            "FormulaFence found conditional formatting without target ranges; "
                            "the affected control has a coverage gap."
                        )
                        continue
                    for rule in conditional_formatting.findall(rule_tag):
                        snapshot = _conditional_rule_snapshot(
                            sheet,
                            ranges,
                            rule,
                            differential_styles,
                            warnings,
                        )
                        rules_by_sheet[sheet].append(
                            (snapshot.priority, source_order, snapshot)
                        )
                        source_order += 1
                extensions.extend(
                    _worksheet_conditional_formatting_extensions(sheet, worksheet)
                )
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError, ValueError) as error:
        return _ConditionalFormattingMetadata(
            rules=(),
            extensions=(),
            warnings=(
                "FormulaFence could not inspect conditional-formatting OOXML "
                f"({type(error).__name__}); conditional-formatting controls were not compared.",
            ),
        )

    rules: list[ConditionalFormattingSnapshot] = []
    for _sheet, entries in rules_by_sheet.items():
        priorities = [priority for priority, _, _ in entries]
        if len(set(priorities)) != len(priorities):
            warnings.add(
                "FormulaFence found duplicate conditional-formatting priorities on a "
                "worksheet; the affected controls have a coverage gap."
            )
        for normalized_priority, (_, _, snapshot) in enumerate(
            sorted(entries, key=lambda item: item[:2]),
            start=1,
        ):
            rules.append(replace(snapshot, priority=normalized_priority))
    return _ConditionalFormattingMetadata(
        rules=tuple(sorted(rules, key=ConditionalFormattingSnapshot.sort_key)),
        extensions=tuple(
            sorted(extensions, key=ConditionalFormattingExtensionSnapshot.sort_key)
        ),
        warnings=tuple(sorted(warnings)),
    )


def _protection_bool(
    value: str | None,
    default: bool,
    attribute: str,
    warnings: set[str],
) -> bool:
    """Read a protection boolean while retaining its schema default."""
    if value is None:
        return default
    lowered = value.casefold()
    if lowered in {"1", "true", "on"}:
        return True
    if lowered in {"0", "false", "off"}:
        return False
    warnings.add(
        "FormulaFence could not interpret a protection "
        f"{attribute} boolean; the schema default was used."
    )
    return default


def _protection_int(
    element: ElementTree.Element,
    attribute: str,
    warnings: set[str],
    *,
    context: str,
) -> int | None:
    """Read an optional non-negative protection integer conservatively."""
    value = element.get(attribute)
    if value is None:
        return None
    try:
        parsed = int(value)
    except ValueError:
        warnings.add(
            "FormulaFence could not interpret a "
            f"{context} {attribute} integer; the affected protection has a coverage gap."
        )
        return None
    if parsed < 0:
        warnings.add(
            "FormulaFence found a negative "
            f"{context} {attribute} integer; the affected protection has a coverage gap."
        )
        return None
    return parsed


def _private_protection_signature(
    entries: tuple[tuple[str, str], ...],
) -> str | None:
    """Hash sensitive comparison material without retaining it in output models."""
    if not entries:
        return None
    digest = hashlib.sha256()
    for name, value in entries:
        digest.update(name.encode("utf-8"))
        digest.update(b"\0")
        digest.update(value.encode("utf-8"))
        digest.update(b"\0")
    return digest.hexdigest()


def _merge_opaque_protection_metadata(
    *metadata: ProtectionOpaqueMetadataSnapshot,
) -> ProtectionOpaqueMetadataSnapshot:
    """Combine opaque fragments while leaving their contents private."""
    present = tuple(
        item for item in metadata if item.count and item.signature is not None
    )
    if not present:
        return ProtectionOpaqueMetadataSnapshot()
    return ProtectionOpaqueMetadataSnapshot(
        count=sum(item.count for item in present),
        signature=_private_protection_signature(
            tuple(
                (str(index), item.signature or "")
                for index, item in enumerate(present, start=1)
            )
        ),
    )


def _opaque_protection_metadata(
    element: ElementTree.Element,
    *,
    known_attributes: set[str],
    known_children: set[str] = frozenset(),
) -> ProtectionOpaqueMetadataSnapshot:
    """Fingerprint unmodelled protection XML without exposing its contents."""
    entries: list[tuple[str, str]] = []
    for attribute, value in element.attrib.items():
        if _xml_local_name(attribute) not in known_attributes:
            entries.append((f"attribute:{_xml_display_name(attribute)}", value))
    for child in element:
        if _xml_local_name(child.tag) in known_children:
            continue
        entries.append(
            (
                f"child:{_xml_display_name(child.tag)}",
                repr(_xml_fragment(child).sort_key()),
            )
        )
    entries.sort()
    return ProtectionOpaqueMetadataSnapshot(
        count=len(entries),
        signature=_private_protection_signature(tuple(entries)),
    )


def _protection_credential_snapshot(
    element: ElementTree.Element,
    *,
    legacy_attribute: str,
    algorithm_attribute: str,
    hash_attribute: str,
    salt_attribute: str,
    spin_count_attribute: str,
    context: str,
    warnings: set[str],
) -> ProtectionCredentialSnapshot:
    """Record safe verifier metadata while hashing the actual verifier values."""
    attribute_names = (
        legacy_attribute,
        algorithm_attribute,
        hash_attribute,
        salt_attribute,
        spin_count_attribute,
    )
    signature_entries = tuple(
        (attribute, value)
        for attribute in attribute_names
        if (value := element.get(attribute)) is not None
    )
    modern_attributes = (
        algorithm_attribute,
        hash_attribute,
        salt_attribute,
        spin_count_attribute,
    )
    has_modern_verifier = any(element.get(attribute) is not None for attribute in modern_attributes)
    if has_modern_verifier and not (
        element.get(hash_attribute) is not None and element.get(salt_attribute) is not None
    ):
        warnings.add(
            "FormulaFence found incomplete modern "
            f"{context} verifier metadata; its presence is compared but has a coverage gap."
        )
    return ProtectionCredentialSnapshot(
        has_legacy_verifier=element.get(legacy_attribute) is not None,
        has_modern_verifier=has_modern_verifier,
        algorithm=element.get(algorithm_attribute),
        spin_count=_protection_int(
            element,
            spin_count_attribute,
            warnings,
            context=context,
        ),
        signature=_private_protection_signature(signature_entries),
    )


def _workbook_protection_snapshot(
    workbook: ElementTree.Element,
    warnings: set[str],
) -> WorkbookProtectionSnapshot | None:
    """Read workbook protection without passing verifier material to openpyxl."""
    element = workbook.find(f"{{{_SPREADSHEETML_NS}}}workbookProtection")
    if element is None:
        return None
    workbook_credential = _protection_credential_snapshot(
        element,
        legacy_attribute="workbookPassword",
        algorithm_attribute="workbookAlgorithmName",
        hash_attribute="workbookHashValue",
        salt_attribute="workbookSaltValue",
        spin_count_attribute="workbookSpinCount",
        context="workbook-protection",
        warnings=warnings,
    )
    revisions_credential = _protection_credential_snapshot(
        element,
        legacy_attribute="revisionsPassword",
        algorithm_attribute="revisionsAlgorithmName",
        hash_attribute="revisionsHashValue",
        salt_attribute="revisionsSaltValue",
        spin_count_attribute="revisionsSpinCount",
        context="revision-protection",
        warnings=warnings,
    )
    return WorkbookProtectionSnapshot(
        lock_structure=_protection_bool(
            element.get("lockStructure"), False, "lockStructure", warnings
        ),
        lock_windows=_protection_bool(
            element.get("lockWindows"), False, "lockWindows", warnings
        ),
        lock_revision=_protection_bool(
            element.get("lockRevision"), False, "lockRevision", warnings
        ),
        workbook_credential=workbook_credential,
        revisions_credential=revisions_credential,
        opaque_metadata=_opaque_protection_metadata(
            element,
            known_attributes={
                "lockStructure",
                "lockWindows",
                "lockRevision",
                "workbookPassword",
                "workbookAlgorithmName",
                "workbookHashValue",
                "workbookSaltValue",
                "workbookSpinCount",
                "revisionsPassword",
                "revisionsAlgorithmName",
                "revisionsHashValue",
                "revisionsSaltValue",
                "revisionsSpinCount",
            },
        ),
    )


def _sheet_protection_snapshot(
    sheet: str,
    sheet_type: str,
    root: ElementTree.Element,
    warnings: set[str],
) -> SheetProtectionSnapshot | None:
    """Read one sheet-protection declaration with effective action defaults."""
    element = root.find(f"{{{_SPREADSHEETML_NS}}}sheetProtection")
    if element is None:
        return None
    actions = (
        _CHART_SHEET_PROTECTION_ACTIONS
        if sheet_type == "chartsheet"
        else _SHEET_PROTECTION_ACTIONS
    )
    enabled = (
        any(
            _protection_bool(element.get(attribute), default, attribute, warnings)
            for attribute, _name, default in actions
        )
        if sheet_type == "chartsheet"
        else _protection_bool(element.get("sheet"), False, "sheet", warnings)
    )
    locked_actions = (
        tuple(
            name
            for attribute, name, default in actions
            if _protection_bool(element.get(attribute), default, attribute, warnings)
        )
        if enabled
        else ()
    )
    credential = _protection_credential_snapshot(
        element,
        legacy_attribute="password",
        algorithm_attribute="algorithmName",
        hash_attribute="hashValue",
        salt_attribute="saltValue",
        spin_count_attribute="spinCount",
        context="sheet-protection",
        warnings=warnings,
    )
    known_attributes = {
        *(attribute for attribute, _name, _default in actions),
        "password",
        "algorithmName",
        "hashValue",
        "saltValue",
        "spinCount",
    }
    if sheet_type == "chartsheet":
        known_attributes.add("content")
    else:
        known_attributes.add("sheet")
    return SheetProtectionSnapshot(
        sheet=sheet,
        sheet_type=sheet_type,
        enabled=enabled,
        locked_actions=locked_actions,
        credential=credential,
        opaque_metadata=_opaque_protection_metadata(
            element,
            known_attributes=known_attributes,
        ),
    )


def _protected_range_snapshots(
    sheet: str,
    root: ElementTree.Element,
    warnings: set[str],
) -> tuple[tuple[ProtectedRangeSnapshot, ...], ProtectionOpaqueMetadataSnapshot]:
    """Read protected ranges while redacting names and security descriptors."""
    container = root.find(f"{{{_SPREADSHEETML_NS}}}protectedRanges")
    if container is None:
        return (), ProtectionOpaqueMetadataSnapshot()
    snapshots: list[ProtectedRangeSnapshot] = []
    range_tag = f"{{{_SPREADSHEETML_NS}}}protectedRange"
    for protected_range in container.findall(range_tag):
        ranges = _conditional_ranges(protected_range.get("sqref"))
        if not ranges:
            warnings.add(
                "FormulaFence found a protected range without target references; "
                "the affected permission has a coverage gap."
            )
        name = protected_range.get("name")
        if name is None:
            warnings.add(
                "FormulaFence found a protected range without its required name; "
                "the affected permission has a coverage gap."
            )
        security_descriptor = protected_range.get("securityDescriptor")
        credential = _protection_credential_snapshot(
            protected_range,
            legacy_attribute="password",
            algorithm_attribute="algorithmName",
            hash_attribute="hashValue",
            salt_attribute="saltValue",
            spin_count_attribute="spinCount",
            context="protected-range",
            warnings=warnings,
        )
        snapshots.append(
            ProtectedRangeSnapshot(
                sheet=sheet,
                ranges=ranges,
                has_name=name is not None,
                name_signature=_private_protection_signature(
                    (("name", name),) if name is not None else ()
                ),
                credential=credential,
                has_security_descriptor=security_descriptor is not None,
                security_descriptor_signature=_private_protection_signature(
                    (("securityDescriptor", security_descriptor),)
                    if security_descriptor is not None
                    else ()
                ),
                opaque_metadata=_opaque_protection_metadata(
                    protected_range,
                    known_attributes={
                        "name",
                        "sqref",
                        "password",
                        "algorithmName",
                        "hashValue",
                        "saltValue",
                        "spinCount",
                        "securityDescriptor",
                    },
                ),
            )
        )
    return (
        tuple(sorted(snapshots, key=ProtectedRangeSnapshot.sort_key)),
        _opaque_protection_metadata(
            container,
            known_attributes=set(),
            known_children={"protectedRange"},
        ),
    )


def _style_index(
    element: ElementTree.Element,
    attribute: str,
    warnings: set[str],
    *,
    context: str,
) -> int | None:
    """Read a zero-based style index without silently accepting malformed XML."""
    value = element.get(attribute)
    if value is None:
        return None
    try:
        parsed = int(value)
    except ValueError:
        warnings.add(
            "FormulaFence could not interpret a "
            f"{context} style index; the affected cell-protection assignment has a coverage gap."
        )
        return None
    if parsed < 0:
        warnings.add(
            "FormulaFence found a negative "
            f"{context} style index; the affected cell-protection assignment has a coverage gap."
        )
        return None
    return parsed


def _xf_protection(
    xf: ElementTree.Element,
    inherited: _StyleProtection,
    warnings: set[str],
    *,
    context: str,
) -> _StyleProtection:
    """Resolve the protection part of one XF without inspecting formatting data."""
    protection = xf.find(f"{{{_SPREADSHEETML_NS}}}protection")
    if protection is None:
        return inherited
    if (
        any(
            _xml_local_name(attribute) not in {"locked", "hidden"}
            for attribute in protection.attrib
        )
        or list(protection)
    ):
        warnings.add(
            "FormulaFence found unmodelled cell-protection style metadata; "
            "the affected assignment has a coverage gap."
        )
    applies = _protection_bool(
        xf.get("applyProtection"),
        True,
        "applyProtection",
        warnings,
    )
    if not applies:
        return inherited
    return _StyleProtection(
        locked=_protection_bool(
            protection.get("locked"), True, f"{context} locked", warnings
        ),
        hidden=_protection_bool(
            protection.get("hidden"), False, f"{context} hidden", warnings
        ),
        explicit=True,
    )


def _styles_with_protection(
    archive: ZipFile,
    warnings: set[str],
) -> tuple[_StyleProtection, tuple[_StyleProtection, ...]]:
    """Return the base default and effective cell-XF protection table."""
    default = _StyleProtection(locked=True, hidden=False)
    try:
        styles = _xml_root(archive, "xl/styles.xml")
    except KeyError:
        return default, (default,)
    style_xfs = styles.find(f"{{{_SPREADSHEETML_NS}}}cellStyleXfs")
    base_styles: list[_StyleProtection] = []
    if style_xfs is not None:
        for xf in style_xfs.findall(f"{{{_SPREADSHEETML_NS}}}xf"):
            base_styles.append(
                _xf_protection(xf, default, warnings, context="base-cell-style")
            )
    if not base_styles:
        base_styles.append(default)

    cell_xfs = styles.find(f"{{{_SPREADSHEETML_NS}}}cellXfs")
    effective_styles: list[_StyleProtection] = []
    if cell_xfs is not None:
        for xf in cell_xfs.findall(f"{{{_SPREADSHEETML_NS}}}xf"):
            xf_id = _style_index(xf, "xfId", warnings, context="cell-XF")
            inherited = default
            if xf_id is not None:
                if xf_id >= len(base_styles):
                    warnings.add(
                        "FormulaFence found a cell-XF with an unknown base style; "
                        "the affected cell-protection assignment has a coverage gap."
                    )
                else:
                    inherited = base_styles[xf_id]
            effective_styles.append(
                _xf_protection(xf, inherited, warnings, context="cell-style")
            )
    if not effective_styles:
        effective_styles.append(default)
    return effective_styles[0], tuple(effective_styles)


def _style_is_protection_relevant(
    style: _StyleProtection,
    default: _StyleProtection,
) -> bool:
    """Return whether a styled record can alter visible protection behavior."""
    return style.explicit or (style.locked, style.hidden) != (default.locked, default.hidden)


def _style_for_assignment(
    element: ElementTree.Element,
    attribute: str,
    styles: tuple[_StyleProtection, ...],
    warnings: set[str],
    *,
    context: str,
) -> _StyleProtection | None:
    """Resolve a direct style reference, returning ``None`` when absent/invalid."""
    style_index = _style_index(element, attribute, warnings, context=context)
    if style_index is None:
        return None
    if style_index >= len(styles):
        warnings.add(
            "FormulaFence found a "
            f"{context} style index outside the workbook style table; "
            "the affected cell-protection assignment has a coverage gap."
        )
        return None
    return styles[style_index]


def _column_span(
    element: ElementTree.Element,
    warnings: set[str],
) -> tuple[int, int, str] | None:
    """Return a compact column span from a raw ``col`` record."""
    start = _protection_int(element, "min", warnings, context="column-protection")
    end = _protection_int(element, "max", warnings, context="column-protection")
    if start is None or end is None or start < 1 or end < start or end > 16_384:
        warnings.add(
            "FormulaFence could not interpret a column-protection span; "
            "the affected assignment has a coverage gap."
        )
        return None
    try:
        start_column = get_column_letter(start)
        end_column = get_column_letter(end)
    except ValueError:
        warnings.add(
            "FormulaFence found a column-protection span outside Excel's column range; "
            "the affected assignment has a coverage gap."
        )
        return None
    return start, end, f"{start_column}:{end_column}"


def _cell_protection_assignments(
    archive: ZipFile,
    sheet_parts: Mapping[str, tuple[str, str]],
    sheet_protections: tuple[SheetProtectionSnapshot, ...],
    warnings: set[str],
) -> tuple[
    CellProtectionDefaultSnapshot | None,
    tuple[CellProtectionAssignmentSnapshot, ...],
]:
    """Read sparse direct cell/row/column protection assignments from OOXML.

    The inventory deliberately retains raw assignment scopes instead of
    expanding rows, columns, or styled rectangles into cells. It records only
    normal protected sheets, because locked/hidden cell styles are inactive on
    an unprotected sheet.
    """
    protected_sheets = {
        protection.sheet
        for protection in sheet_protections
        if protection.enabled and protection.sheet_type in {"worksheet", "dialogsheet"}
    }
    if not protected_sheets:
        return None, ()
    default, styles = _styles_with_protection(archive, warnings)
    assignments: set[CellProtectionAssignmentSnapshot] = set()
    row_tag = f"{{{_SPREADSHEETML_NS}}}row"
    cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
    column_tag = f"{{{_SPREADSHEETML_NS}}}col"
    for sheet in sorted(protected_sheets, key=str.casefold):
        part = sheet_parts.get(sheet)
        if part is None:
            warnings.add(
                "FormulaFence could not locate OOXML for a protected worksheet; "
                "cell-protection assignments have a coverage gap."
            )
            continue
        member, _sheet_type = part
        worksheet = _xml_root(archive, member)
        raw_rows: list[tuple[int, _StyleProtection]] = []
        raw_columns: list[tuple[int, int, str, _StyleProtection]] = []
        for row in worksheet.iter(row_tag):
            style = _style_for_assignment(
                row, "s", styles, warnings, context="row-protection"
            )
            if style is None:
                continue
            row_number = _protection_int(row, "r", warnings, context="row-protection")
            if row_number is None or row_number < 1:
                warnings.add(
                    "FormulaFence could not interpret a row-protection target; "
                    "the affected assignment has a coverage gap."
                )
                continue
            raw_rows.append((row_number, style))
        for column in worksheet.iter(column_tag):
            style = _style_for_assignment(
                column, "style", styles, warnings, context="column-protection"
            )
            if style is None or (span := _column_span(column, warnings)) is None:
                continue
            start_column, end_column, target = span
            raw_columns.append((start_column, end_column, target, style))

        relevant_rows = {
            row_number
            for row_number, style in raw_rows
            if _style_is_protection_relevant(style, default)
        }
        relevant_columns = [
            column_target
            for _start_column, _end_column, column_target, style in raw_columns
            if _style_is_protection_relevant(style, default)
        ]
        for row_number, style in raw_rows:
            if not _style_is_protection_relevant(style, default) and not relevant_columns:
                continue
            assignments.add(
                CellProtectionAssignmentSnapshot(
                    sheet=sheet,
                    scope="row",
                    target=str(row_number),
                    locked=style.locked,
                    hidden=style.hidden,
                )
            )
        for _start_column, _end_column, target, style in raw_columns:
            if not _style_is_protection_relevant(style, default) and not relevant_rows:
                continue
            assignments.add(
                CellProtectionAssignmentSnapshot(
                    sheet=sheet,
                    scope="column",
                    target=target,
                    locked=style.locked,
                    hidden=style.hidden,
                )
            )

        for cell in worksheet.iter(cell_tag):
            style = _style_for_assignment(
                cell, "s", styles, warnings, context="cell-protection"
            )
            if style is None:
                continue
            coordinate = cell.get("r")
            if not coordinate:
                warnings.add(
                    "FormulaFence found a styled cell without a coordinate; "
                    "the affected cell-protection assignment has a coverage gap."
                )
                continue
            match = re.fullmatch(r"([A-Za-z]+)([1-9][0-9]*)", coordinate)
            if match is None:
                warnings.add(
                    "FormulaFence could not interpret a styled cell coordinate; "
                    "the affected cell-protection assignment has a coverage gap."
                )
                continue
            column_letters, raw_row_number = match.groups()
            row_number = int(raw_row_number)
            try:
                column_number = column_index_from_string(column_letters)
            except ValueError:
                warnings.add(
                    "FormulaFence could not interpret a styled cell column; "
                    "the affected cell-protection assignment has a coverage gap."
                )
                continue
            intersects_relevant_column = any(
                start_column <= column_number <= end_column
                for start_column, end_column, _target, style in raw_columns
                if _style_is_protection_relevant(style, default)
            )
            if not (
                _style_is_protection_relevant(style, default)
                or row_number in relevant_rows
                or intersects_relevant_column
            ):
                continue
            assignments.add(
                CellProtectionAssignmentSnapshot(
                    sheet=sheet,
                    scope="cell",
                    target=coordinate.upper(),
                    locked=style.locked,
                    hidden=style.hidden,
                )
            )
    return (
        CellProtectionDefaultSnapshot(locked=default.locked, hidden=default.hidden),
        tuple(sorted(assignments, key=CellProtectionAssignmentSnapshot.sort_key)),
    )


def _protection_metadata(path: Path) -> _ProtectionMetadata:
    """Read protection controls directly from OOXML before a library drops data."""
    warnings: set[str] = set()
    try:
        with ZipFile(path) as archive:
            workbook = _xml_root(archive, "xl/workbook.xml")
            sheet_parts = _sheet_xml_parts(archive)
            workbook_protection = _workbook_protection_snapshot(workbook, warnings)
            sheet_protections: dict[str, SheetProtectionSnapshot] = {}
            protected_ranges: list[ProtectedRangeSnapshot] = []
            for sheet, (member, sheet_type) in sheet_parts.items():
                root = _xml_root(archive, member)
                sheet_protection = _sheet_protection_snapshot(
                    sheet, sheet_type, root, warnings
                )
                ranges: tuple[ProtectedRangeSnapshot, ...] = ()
                range_container_opaque = ProtectionOpaqueMetadataSnapshot()
                if sheet_type != "chartsheet":
                    ranges, range_container_opaque = _protected_range_snapshots(
                        sheet, root, warnings
                    )
                    protected_ranges.extend(ranges)
                if sheet_protection is not None:
                    sheet_protections[sheet] = replace(
                        sheet_protection,
                        opaque_metadata=_merge_opaque_protection_metadata(
                            sheet_protection.opaque_metadata,
                            range_container_opaque,
                        ),
                    )
                elif range_container_opaque.present:
                    sheet_protections[sheet] = SheetProtectionSnapshot(
                        sheet=sheet,
                        sheet_type=sheet_type,
                        enabled=False,
                        locked_actions=(),
                        opaque_metadata=range_container_opaque,
                    )
            sorted_sheet_protections = tuple(
                sorted(sheet_protections.values(), key=SheetProtectionSnapshot.sort_key)
            )
            cell_default, cell_assignments = _cell_protection_assignments(
                archive,
                sheet_parts,
                sorted_sheet_protections,
                warnings,
            )
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError, ValueError) as error:
        return _ProtectionMetadata(
            workbook_protection=None,
            sheet_protections=(),
            protected_ranges=(),
            cell_protection_default=None,
            cell_protection_assignments=(),
            warnings=(
                "FormulaFence could not inspect protection OOXML "
                f"({type(error).__name__}); protection controls were not compared.",
            ),
        )
    return _ProtectionMetadata(
        workbook_protection=workbook_protection,
        sheet_protections=sorted_sheet_protections,
        protected_ranges=tuple(
            sorted(protected_ranges, key=ProtectedRangeSnapshot.sort_key)
        ),
        cell_protection_default=cell_default,
        cell_protection_assignments=cell_assignments,
        warnings=tuple(sorted(warnings)),
    )


def _external_data_bool(
    value: str | None,
    default: bool,
    attribute: str,
    warnings: set[str],
    *,
    context: str,
) -> bool:
    """Read an OOXML external-data boolean while retaining its default."""
    if value is None:
        return default
    lowered = value.casefold()
    if lowered in {"1", "true", "on"}:
        return True
    if lowered in {"0", "false", "off"}:
        return False
    warnings.add(
        "FormulaFence could not interpret an external-data "
        f"{context} {attribute} boolean; the schema default was used."
    )
    return default


def _external_data_unsigned_int(
    element: ElementTree.Element,
    attribute: str,
    default: int | None,
    warnings: set[str],
    *,
    context: str,
    maximum: int = 4_294_967_295,
) -> int | None:
    """Read a non-negative OOXML integer without accepting malformed values."""
    value = element.get(attribute)
    if value is None:
        return default
    try:
        parsed = int(value)
    except ValueError:
        warnings.add(
            "FormulaFence could not interpret an external-data "
            f"{context} {attribute} integer; the affected control has a coverage gap."
        )
        return default
    if not 0 <= parsed <= maximum:
        warnings.add(
            "FormulaFence found an out-of-range external-data "
            f"{context} {attribute} integer; the affected control has a coverage gap."
        )
        return default
    return parsed


def _external_data_enum(
    value: str | None,
    default: str,
    values: Mapping[str, str],
    warnings: set[str],
    *,
    context: str,
    attribute: str,
) -> str:
    """Normalize a safe OOXML enumeration without exposing invalid values."""
    if value is None:
        return default
    if normalised := values.get(value):
        return normalised
    warnings.add(
        "FormulaFence found an unrecognized external-data "
        f"{context} {attribute} value; the affected control has a coverage gap."
    )
    return "unrecognized"


def _private_external_data_signature(
    entries: tuple[tuple[str, str], ...],
) -> str | None:
    """Hash source or identity material without retaining it in output models."""
    if not entries:
        return None
    digest = hashlib.sha256()
    for name, value in entries:
        digest.update(name.encode("utf-8"))
        digest.update(b"\0")
        digest.update(value.encode("utf-8"))
        digest.update(b"\0")
    return digest.hexdigest()


def _external_data_opaque_metadata(
    element: ElementTree.Element,
    *,
    known_attributes: frozenset[str],
    known_children: frozenset[str] = frozenset(),
) -> ExternalDataOpaqueMetadataSnapshot:
    """Fingerprint unknown external-data XML without serializing its contents."""
    entries: list[tuple[str, str]] = []
    for attribute, value in element.attrib.items():
        if _xml_local_name(attribute) not in known_attributes:
            entries.append((f"attribute:{_xml_display_name(attribute)}", value))
    for child in element:
        if _xml_local_name(child.tag) in known_children:
            continue
        entries.append(
            (
                f"child:{_xml_display_name(child.tag)}",
                repr(_xml_fragment(child).sort_key()),
            )
        )
    entries.sort()
    return ExternalDataOpaqueMetadataSnapshot(
        count=len(entries),
        signature=_private_external_data_signature(tuple(entries)),
    )


def _private_payload_signature(payload: bytes) -> str:
    """Hash opaque bytes without retaining them in any reviewable model."""
    return hashlib.sha256(payload).hexdigest()


def _power_query_opaque_metadata(
    entries: list[tuple[str, str]],
) -> ExternalDataOpaqueMetadataSnapshot:
    """Aggregate private Data Mashup evidence without serializing its material."""
    entries.sort()
    return ExternalDataOpaqueMetadataSnapshot(
        count=len(entries),
        signature=_private_external_data_signature(tuple(entries)),
    )


def _power_query_boolean(
    value: str | None,
    default: bool,
    element_name: str,
    warnings: set[str],
) -> bool:
    """Read a Data Mashup boolean while retaining the documented default."""
    if value is None:
        return default
    lowered = value.strip().casefold()
    if lowered in {"1", "true"}:
        return True
    if lowered in {"0", "false"}:
        return False
    warnings.add(
        "FormulaFence could not interpret a Power Query "
        f"{element_name} boolean; the documented default was used."
    )
    return default


def _power_query_fields(
    payload: bytes,
    member: str,
    warnings: set[str],
    opaque_entries: list[tuple[str, str]],
) -> tuple[int, tuple[bytes, bytes, bytes, bytes]] | None:
    """Read the versioned, length-prefixed Data Mashup binary stream safely."""
    if len(payload) < 4:
        warnings.add(
            "FormulaFence found a truncated Power Query Data Mashup stream; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append((f"{member}:truncated", _private_payload_signature(payload)))
        return None
    version = struct.unpack_from("<I", payload, 0)[0]
    if version != 0:
        warnings.add(
            "FormulaFence found an unrecognized Power Query Data Mashup stream version; "
            "the affected query controls have a coverage gap."
        )
    cursor = 4
    fields: list[bytes] = []
    for field_name in ("package parts", "permissions", "metadata", "permission bindings"):
        if cursor + 4 > len(payload):
            warnings.add(
                "FormulaFence found a truncated Power Query Data Mashup "
                f"{field_name} length; the affected query controls have a coverage gap."
            )
            opaque_entries.append(
                (f"{member}:truncated-{field_name}", _private_payload_signature(payload))
            )
            return None
        length = struct.unpack_from("<I", payload, cursor)[0]
        cursor += 4
        if length > len(payload) - cursor:
            warnings.add(
                "FormulaFence found an out-of-range Power Query Data Mashup "
                f"{field_name} length; the affected query controls have a coverage gap."
            )
            opaque_entries.append(
                (f"{member}:out-of-range-{field_name}", _private_payload_signature(payload))
            )
            return None
        fields.append(payload[cursor : cursor + length])
        cursor += length
    if cursor != len(payload):
        warnings.add(
            "FormulaFence found trailing bytes in a Power Query Data Mashup stream; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:trailing-bytes", _private_payload_signature(payload[cursor:]))
        )
    return version, (fields[0], fields[1], fields[2], fields[3])


def _power_query_package_inventory(
    package_parts: bytes,
    member: str,
    warnings: set[str],
    opaque_entries: list[tuple[str, str]],
) -> tuple[int, int, int, str | None, str | None]:
    """Fingerprint logical query-package parts without exposing any part contents."""
    try:
        with ZipFile(io.BytesIO(package_parts)) as package:
            entries = package.infolist()
            formula_entries = [
                entry for entry in entries if entry.filename == _POWER_QUERY_FORMULA_MEMBER
            ]
            if len(formula_entries) > 1:
                warnings.add(
                    "FormulaFence found multiple Power Query Section1.m formula parts; "
                    "the affected query controls have a coverage gap."
                )
            formula_material: list[tuple[str, str]] = []
            configuration_material: list[tuple[str, str]] = []
            for entry in entries:
                label = f"{member}:{entry.filename}"
                if entry.file_size > _POWER_QUERY_MAX_FORMULA_BYTES:
                    warnings.add(
                        "FormulaFence did not fully read an oversized Power Query package part; "
                        "the affected query controls have a coverage gap."
                    )
                    signature = _private_external_data_signature(
                        (
                            ("size", str(entry.file_size)),
                            ("crc", str(entry.CRC)),
                            ("compressed_size", str(entry.compress_size)),
                        )
                    )
                    opaque_entries.append((f"{label}:oversized", signature or ""))
                else:
                    signature = _private_payload_signature(package.read(entry))
                if entry.filename == _POWER_QUERY_FORMULA_MEMBER:
                    formula_material.append((label, signature or ""))
                else:
                    configuration_material.append((label, signature or ""))
            embedded_content_count = sum(
                entry.filename.startswith("Content/") for entry in entries
            )
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a Power Query package-part stream "
            f"({type(error).__name__}); the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:package-parts", _private_payload_signature(package_parts))
        )
        return 0, 0, 0, None, None
    return (
        len(entries),
        len(formula_entries),
        embedded_content_count,
        _private_external_data_signature(tuple(sorted(formula_material))),
        _private_external_data_signature(tuple(sorted(configuration_material))),
    )


def _power_query_metadata_content_count(
    payload: bytes,
    member: str,
    warnings: set[str],
    opaque_entries: list[tuple[str, str]],
) -> int:
    """Count metadata-side embedded parts without reading their sensitive contents."""
    if not payload:
        return 0
    try:
        with ZipFile(io.BytesIO(payload)) as archive:
            return len(archive.infolist())
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect Power Query metadata embedded content "
            f"({type(error).__name__}); the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:metadata-content", _private_payload_signature(payload))
        )
        return 0


def _power_query_metadata_inventory(
    payload: bytes,
    member: str,
    warnings: set[str],
    opaque_entries: list[tuple[str, str]],
) -> tuple[int, int, int, str | None, str | None]:
    """Read stable query metadata while intentionally ignoring refresh-result noise."""
    if not payload:
        return 0, 0, 0, None, None
    if len(payload) < 8:
        warnings.add(
            "FormulaFence found a truncated Power Query metadata stream; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append((f"{member}:truncated-metadata", _private_payload_signature(payload)))
        return 0, 0, 0, None, None

    version, xml_length = struct.unpack_from("<II", payload, 0)
    if version != 0:
        warnings.add(
            "FormulaFence found an unrecognized Power Query metadata version; "
            "the affected query controls have a coverage gap."
        )
    cursor = 8
    if xml_length > len(payload) - cursor:
        warnings.add(
            "FormulaFence found an out-of-range Power Query metadata XML length; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:metadata-xml-length", _private_payload_signature(payload))
        )
        return 0, 0, 0, None, None
    metadata_xml = payload[cursor : cursor + xml_length]
    cursor += xml_length
    if cursor + 4 > len(payload):
        warnings.add(
            "FormulaFence found a truncated Power Query metadata content length; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:metadata-content-length", _private_payload_signature(payload))
        )
        return 0, 0, 0, None, None
    content_length = struct.unpack_from("<I", payload, cursor)[0]
    cursor += 4
    if content_length > len(payload) - cursor:
        warnings.add(
            "FormulaFence found an out-of-range Power Query metadata content length; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:metadata-content-length", _private_payload_signature(payload))
        )
        return 0, 0, 0, None, None
    embedded_content = payload[cursor : cursor + content_length]
    cursor += content_length
    if cursor != len(payload):
        warnings.add(
            "FormulaFence found trailing Power Query metadata bytes; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:metadata-trailing-bytes", _private_payload_signature(payload[cursor:]))
        )

    embedded_content_count = _power_query_metadata_content_count(
        embedded_content,
        member,
        warnings,
        opaque_entries,
    )
    try:
        root = _xml_root_from_payload(metadata_xml)
    except (ElementTree.ParseError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect Power Query metadata XML "
            f"({type(error).__name__}); the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:metadata-xml", _private_payload_signature(metadata_xml))
        )
        return 0, 0, embedded_content_count, None, None
    if (
        _xml_local_name(root.tag) != "LocalPackageMetadataFile"
        or _xml_namespace(root.tag) not in {None, _DATA_MASHUP_NS}
    ):
        warnings.add(
            "FormulaFence found Power Query metadata with an unexpected root; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:metadata-root", _private_payload_signature(metadata_xml))
        )
        return 0, 0, embedded_content_count, None, None

    def named_children(element: ElementTree.Element, name: str) -> list[ElementTree.Element]:
        return [
            child
            for child in element
            if _xml_local_name(child.tag) == name
            and _xml_namespace(child.tag) in {None, _DATA_MASHUP_NS}
        ]

    def named_child(element: ElementTree.Element, name: str) -> ElementTree.Element | None:
        children = named_children(element, name)
        return children[0] if children else None

    items_containers = named_children(root, "Items")
    if len(items_containers) != 1:
        warnings.add(
            "FormulaFence found an unexpected Power Query metadata Items container; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (f"{member}:metadata-items", _private_payload_signature(metadata_xml))
        )
        return 1, 0, embedded_content_count, None, None
    unknown_root_children = [
        child
        for child in root
        if _xml_local_name(child.tag) != "Items"
    ]
    for child in unknown_root_children:
        opaque_entries.append(
            (
                f"{member}:metadata-root-child",
                _private_payload_signature(ElementTree.tostring(child, encoding="utf-8")),
            )
        )

    identity_entries: list[tuple[str, str]] = [("metadata-version", str(version))]
    control_entries: list[tuple[str, str]] = []
    item_count = 0
    for item_index, item in enumerate(named_children(items_containers[0], "Item")):
        item_count += 1
        item_location = named_child(item, "ItemLocation")
        stable_entries = named_child(item, "StableEntries")
        if item_location is None or stable_entries is None:
            warnings.add(
                "FormulaFence found incomplete Power Query query metadata; "
                "the affected query controls have a coverage gap."
            )
            opaque_entries.append(
                (
                    f"{member}:metadata-item:{item_index}",
                    _private_payload_signature(ElementTree.tostring(item, encoding="utf-8")),
                )
            )
            continue
        item_type_element = named_child(item_location, "ItemType")
        item_path_element = named_child(item_location, "ItemPath")
        if item_type_element is None or item_path_element is None:
            warnings.add(
                "FormulaFence found Power Query metadata without an item location; "
                "the affected query controls have a coverage gap."
            )
            opaque_entries.append(
                (
                    f"{member}:metadata-location:{item_index}",
                    _private_payload_signature(
                        ElementTree.tostring(item_location, encoding="utf-8")
                    ),
                )
            )
            continue
        item_type = item_type_element.text or ""
        item_path = item_path_element.text or ""
        if not item_type:
            warnings.add(
                "FormulaFence found Power Query metadata without an item type; "
                "the affected query controls have a coverage gap."
            )
            opaque_entries.append(
                (
                    f"{member}:metadata-item-type:{item_index}",
                    _private_payload_signature(
                        ElementTree.tostring(item_location, encoding="utf-8")
                    ),
                )
            )
            continue
        item_key = f"{item_type}\0{item_path}"
        identity_entries.append((f"item:{item_index}", item_key))
        unknown_item_children = [
            child
            for child in item
            if _xml_local_name(child.tag) not in {"ItemLocation", "StableEntries"}
        ]
        for child in unknown_item_children:
            opaque_entries.append(
                (
                    f"{member}:metadata-item-child:{item_index}",
                    _private_payload_signature(ElementTree.tostring(child, encoding="utf-8")),
                )
            )
        for entry_index, entry in enumerate(named_children(stable_entries, "Entry")):
            entry_type = entry.get("Type")
            value = entry.get("Value")
            if entry_type is None or value is None:
                warnings.add(
                    "FormulaFence found incomplete Power Query metadata entry; "
                    "the affected query controls have a coverage gap."
                )
                opaque_entries.append(
                    (
                        f"{member}:metadata-entry:{item_index}:{entry_index}",
                        _private_payload_signature(ElementTree.tostring(entry, encoding="utf-8")),
                    )
                )
                continue
            if entry_type in _POWER_QUERY_STABLE_METADATA_TYPES:
                control_entries.append(
                    (f"{item_key}\0{entry_type}", value)
                )
            elif entry_type not in _POWER_QUERY_VOLATILE_METADATA_TYPES:
                opaque_entries.append(
                    (
                        f"{member}:unknown-metadata-entry:{item_index}:{entry_index}",
                        _private_external_data_signature(
                            (("item", item_key), ("type", entry_type), ("value", value))
                        )
                        or "",
                    )
                )
        unknown_stable_children = [
            child
            for child in stable_entries
            if _xml_local_name(child.tag) != "Entry"
        ]
        for child in unknown_stable_children:
            opaque_entries.append(
                (
                    f"{member}:metadata-stable-child:{item_index}",
                    _private_payload_signature(ElementTree.tostring(child, encoding="utf-8")),
                )
            )
    return (
        1,
        item_count,
        embedded_content_count,
        _private_external_data_signature(tuple(sorted(identity_entries))),
        _private_external_data_signature(tuple(sorted(control_entries))),
    )


def _power_query_permission_inventory(
    payload: bytes,
    member: str,
    warnings: set[str],
    opaque_entries: list[tuple[str, str]],
) -> tuple[int, int, int, int, int, str | None]:
    """Read safe formula-firewall controls without retaining user-bound payloads."""
    if not payload:
        return 0, 0, 0, 0, 0, None
    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect Power Query permissions "
            f"({type(error).__name__}); the affected query controls have a coverage gap."
        )
        opaque_entries.append((f"{member}:permissions", _private_payload_signature(payload)))
        return 1, 0, 0, 0, 0, None
    if _xml_local_name(root.tag) != "PermissionList":
        warnings.add(
            "FormulaFence found Power Query permissions with an unexpected root; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append((f"{member}:permissions-root", _private_payload_signature(payload)))
        return 1, 0, 0, 0, 0, None

    children = {_xml_local_name(child.tag): child for child in root}
    firewall_enabled = _power_query_boolean(
        children.get("FirewallEnabled").text
        if children.get("FirewallEnabled") is not None
        else None,
        True,
        "FirewallEnabled",
        warnings,
    )
    future_packages_allowed = _power_query_boolean(
        children.get("CanEvaluateFuturePackages").text
        if children.get("CanEvaluateFuturePackages") is not None
        else None,
        False,
        "CanEvaluateFuturePackages",
        warnings,
    )
    workbook_group_type = children.get("WorkbookGroupType")
    xsi_nil = "{http://www.w3.org/2001/XMLSchema-instance}nil"
    has_workbook_group_type = bool(
        workbook_group_type is not None
        and workbook_group_type.get(xsi_nil, "").casefold() not in {"1", "true"}
        and (workbook_group_type.text or "").strip()
    )
    signature_entries: list[tuple[str, str]] = [
        ("firewall", str(firewall_enabled)),
        ("future-packages", str(future_packages_allowed)),
    ]
    if has_workbook_group_type and workbook_group_type is not None:
        signature_entries.append(("workbook-group-type", workbook_group_type.text or ""))
    for child in root:
        if _xml_local_name(child.tag) not in {
            "CanEvaluateFuturePackages",
            "FirewallEnabled",
            "WorkbookGroupType",
        }:
            opaque_entries.append(
                (
                    f"{member}:permissions-child",
                    _private_payload_signature(ElementTree.tostring(child, encoding="utf-8")),
                )
            )
    for attribute, value in root.attrib.items():
        opaque_entries.append(
            (
                f"{member}:permissions-attribute:{_xml_display_name(attribute)}",
                value,
            )
        )
    return (
        1,
        1,
        int(firewall_enabled),
        int(future_packages_allowed),
        int(has_workbook_group_type),
        _private_external_data_signature(tuple(sorted(signature_entries))),
    )


def _power_query_mashup_inspection(
    root: ElementTree.Element,
    member: str,
    warnings: set[str],
) -> _PowerQueryMashupInspection:
    """Inspect one Data Mashup part while keeping formulas and sources private."""
    opaque_entries: list[tuple[str, str]] = []
    permission_opaque_entries: list[tuple[str, str]] = []
    for attribute, value in root.attrib.items():
        if _xml_local_name(attribute) != "sqmid":
            opaque_entries.append(
                (f"{member}:attribute:{_xml_display_name(attribute)}", value)
            )
    if tuple(root):
        warnings.add(
            "FormulaFence found child XML inside a Power Query Data Mashup root; "
            "the affected query controls have a coverage gap."
        )
        for child in root:
            opaque_entries.append(
                (
                    f"{member}:root-child",
                    _private_payload_signature(ElementTree.tostring(child, encoding="utf-8")),
                )
            )
    encoded_payload = "".join((root.text or "").split())
    if not encoded_payload:
        warnings.add(
            "FormulaFence found an empty Power Query Data Mashup payload; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append((f"{member}:empty-payload", ""))
        return _PowerQueryMashupInspection(opaque_entries=tuple(opaque_entries))
    try:
        payload = base64.b64decode(encoded_payload.encode("ascii"), validate=True)
    except (UnicodeEncodeError, binascii.Error, ValueError):
        warnings.add(
            "FormulaFence could not decode a Power Query Data Mashup payload; "
            "the affected query controls have a coverage gap."
        )
        opaque_entries.append(
            (
                f"{member}:invalid-base64",
                _private_external_data_signature((("payload", encoded_payload),)) or "",
            )
        )
        return _PowerQueryMashupInspection(opaque_entries=tuple(opaque_entries))

    fields = _power_query_fields(payload, member, warnings, opaque_entries)
    if fields is None:
        return _PowerQueryMashupInspection(opaque_entries=tuple(opaque_entries))
    version, (package_parts, permissions, metadata, permission_bindings) = fields
    (
        package_part_count,
        formula_document_count,
        package_embedded_content_count,
        formula_signature,
        package_configuration_signature,
    ) = _power_query_package_inventory(
        package_parts,
        member,
        warnings,
        opaque_entries,
    )
    (
        metadata_document_count,
        metadata_item_count,
        metadata_embedded_content_count,
        metadata_identity_signature,
        metadata_control_signature,
    ) = _power_query_metadata_inventory(metadata, member, warnings, opaque_entries)
    (
        permission_payload_count,
        permission_parsed_count,
        firewall_enabled_count,
        future_packages_allowed_count,
        workbook_group_type_count,
        permission_signature,
    ) = _power_query_permission_inventory(
        permissions,
        member,
        warnings,
        permission_opaque_entries,
    )
    package_signature_entries = [("stream-version", str(version))]
    if package_configuration_signature is not None:
        package_signature_entries.append(("logical-package", package_configuration_signature))
    return _PowerQueryMashupInspection(
        parsed=True,
        formula_document_count=formula_document_count,
        package_part_count=package_part_count,
        embedded_content_part_count=(
            package_embedded_content_count + metadata_embedded_content_count
        ),
        metadata_document_count=metadata_document_count,
        metadata_item_count=metadata_item_count,
        permission_payload_count=permission_payload_count,
        permission_parsed_count=permission_parsed_count,
        firewall_enabled_count=firewall_enabled_count,
        future_packages_allowed_count=future_packages_allowed_count,
        workbook_group_type_count=workbook_group_type_count,
        permission_binding_present=bool(permission_bindings),
        formula_signature=formula_signature,
        package_configuration_signature=_private_external_data_signature(
            tuple(package_signature_entries)
        ),
        metadata_identity_signature=metadata_identity_signature,
        metadata_control_signature=metadata_control_signature,
        permission_signature=permission_signature,
        opaque_entries=tuple(opaque_entries),
        permission_opaque_entries=tuple(permission_opaque_entries),
    )


def _power_query_snapshot(
    archive: ZipFile,
    warnings: set[str],
) -> PowerQuerySnapshot:
    """Inventory all Data Mashup custom XML parts without exposing their payloads."""
    inspections: list[tuple[str, _PowerQueryMashupInspection]] = []
    for member in sorted(archive.namelist(), key=str.casefold):
        if not _CUSTOM_XML_ITEM_PATTERN.fullmatch(member):
            continue
        try:
            root = _xml_root(archive, member)
        except (KeyError, ElementTree.ParseError, ValueError) as error:
            warnings.add(
                "FormulaFence could not inspect a custom XML part while looking for "
                f"Power Query definitions ({type(error).__name__}); coverage may be incomplete."
            )
            continue
        if (
            _xml_local_name(root.tag) != "DataMashup"
            or _xml_namespace(root.tag) != _DATA_MASHUP_NS
        ):
            continue
        inspections.append((member, _power_query_mashup_inspection(root, member, warnings)))
    if not inspections:
        return PowerQuerySnapshot()

    opaque_entries = [
        entry
        for _, inspection in inspections
        for entry in inspection.opaque_entries
    ]
    permission_opaque_entries = [
        entry
        for _, inspection in inspections
        for entry in inspection.permission_opaque_entries
    ]

    def aggregate_signature(attribute: str) -> str | None:
        material = [
            (member, value)
            for member, inspection in inspections
            if (value := getattr(inspection, attribute)) is not None
        ]
        return _private_external_data_signature(tuple(material))

    permission_controls = PowerQueryPermissionControlsSnapshot(
        payload_count=sum(inspection.permission_payload_count for _, inspection in inspections),
        parsed_count=sum(inspection.permission_parsed_count for _, inspection in inspections),
        firewall_enabled_count=sum(
            inspection.firewall_enabled_count for _, inspection in inspections
        ),
        future_packages_allowed_count=sum(
            inspection.future_packages_allowed_count for _, inspection in inspections
        ),
        workbook_group_type_count=sum(
            inspection.workbook_group_type_count for _, inspection in inspections
        ),
        opaque_metadata=_power_query_opaque_metadata(permission_opaque_entries),
        signature=aggregate_signature("permission_signature"),
    )
    return PowerQuerySnapshot(
        mashup_count=len(inspections),
        parsed_mashup_count=sum(inspection.parsed for _, inspection in inspections),
        formula_document_count=sum(
            inspection.formula_document_count for _, inspection in inspections
        ),
        package_part_count=sum(inspection.package_part_count for _, inspection in inspections),
        embedded_content_part_count=sum(
            inspection.embedded_content_part_count for _, inspection in inspections
        ),
        metadata_document_count=sum(
            inspection.metadata_document_count for _, inspection in inspections
        ),
        metadata_item_count=sum(inspection.metadata_item_count for _, inspection in inspections),
        permission_controls=permission_controls,
        permission_binding_count=sum(
            inspection.permission_binding_present for _, inspection in inspections
        ),
        formula_signature=aggregate_signature("formula_signature"),
        package_configuration_signature=aggregate_signature("package_configuration_signature"),
        metadata_identity_signature=aggregate_signature("metadata_identity_signature"),
        metadata_control_signature=aggregate_signature("metadata_control_signature"),
        opaque_metadata=_power_query_opaque_metadata(opaque_entries),
    )


def _external_link_part_root(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    *,
    context: str,
) -> ElementTree.Element | None:
    """Read one bounded external-link XML part without exposing its content."""
    try:
        if archive.getinfo(member).file_size > _EXTERNAL_LINK_MAX_PART_BYTES:
            warnings.add(
                "FormulaFence did not fully read an oversized external-link package part; "
                "the affected external-link controls have a coverage gap."
            )
            return None
    except KeyError:
        warnings.add(
            "FormulaFence could not locate an external-link package part; "
            "the affected external-link controls were not compared."
        )
        return None
    return _external_data_part_root(archive, member, warnings, context=context)


def _external_link_opaque_entries(
    element: ElementTree.Element,
    *,
    context: str,
    known_attributes: frozenset[str] = frozenset(),
    known_children: frozenset[str] = frozenset(),
    known_namespaces: frozenset[str] = frozenset({_SPREADSHEETML_NS}),
) -> list[tuple[str, str]]:
    """Privately retain unsupported external-link XML without serializing it."""
    entries: list[tuple[str, str]] = []
    for attribute, value in element.attrib.items():
        if attribute == f"{{{_MARKUP_COMPATIBILITY_NS}}}Ignorable":
            continue
        if _xml_local_name(attribute) in known_attributes:
            continue
        entries.append(
            (
                f"{context}:attribute:{_xml_display_name(attribute)}",
                value,
            )
        )
    for child in element:
        if (
            _xml_local_name(child.tag) in known_children
            and _xml_namespace(child.tag) in known_namespaces
        ):
            continue
        entries.append(
            (
                f"{context}:child:{_xml_display_name(child.tag)}",
                repr(_xml_fragment(child).sort_key()),
            )
        )
    return entries


def _external_link_relationship_signature(
    archive: ZipFile,
    member: str,
    relationship_id: str | None,
    warnings: set[str],
    opaque_entries: list[tuple[str, str]],
) -> str | None:
    """Fingerprint one link's package target without retaining its location."""
    if not relationship_id:
        warnings.add(
            "FormulaFence found an external-link definition without its required "
            "relationship id; the affected external-link control has a coverage gap."
        )
        return None
    relationships = _external_link_part_root(
        archive,
        _relationship_part_path(member),
        warnings,
        context="external-link relationship",
    )
    if relationships is None:
        return None
    relationship_tag = f"{{{_PACKAGE_RELATIONSHIP_NS}}}Relationship"
    matches = [
        relationship
        for relationship in relationships.findall(relationship_tag)
        if relationship.get("Id") == relationship_id
    ]
    if not matches:
        warnings.add(
            "FormulaFence could not locate an external-link target relationship; "
            "the affected external-link control has a coverage gap."
        )
        return None
    if len(matches) > 1:
        warnings.add(
            "FormulaFence found repeated external-link target relationships; "
            "the affected external-link control has a coverage gap."
        )
    material: list[tuple[str, str]] = []
    for relationship in matches:
        target = relationship.get("Target")
        if target is None:
            warnings.add(
                "FormulaFence found an external-link target relationship without a target; "
                "the affected external-link control has a coverage gap."
            )
            target = ""
        material.extend(
            [
                ("type", relationship.get("Type", "")),
                ("target_mode", relationship.get("TargetMode", "Internal").casefold()),
                ("target", target),
            ]
        )
        opaque_entries.extend(
            _external_link_opaque_entries(
                relationship,
                context="relationship",
                known_attributes=frozenset({"Id", "Type", "Target", "TargetMode"}),
                known_children=frozenset(),
                known_namespaces=frozenset(),
            )
        )
    material.sort()
    return _private_external_data_signature(tuple(material))


def _external_link_item_flags(
    item: ElementTree.Element,
    warnings: set[str],
    *,
    context: str,
) -> tuple[bool, bool, bool]:
    """Read safe DDE/OLE item behavior flags with their schema defaults."""
    return (
        _external_data_bool(
            item.get("advise"), False, "advise", warnings, context=context
        ),
        _external_data_bool(item.get("ole"), False, "ole", warnings, context=context),
        _external_data_bool(
            item.get("preferPic"), False, "preferPic", warnings, context=context
        ),
    )


def _external_link_optional_child(
    element: ElementTree.Element,
    name: str,
    warnings: set[str],
    opaque_entries: list[tuple[str, str]],
    *,
    context: str,
) -> ElementTree.Element | None:
    """Return one schema-singleton child while retaining repeated XML privately."""
    matches = [
        child
        for child in element
        if _xml_namespace(child.tag) == _SPREADSHEETML_NS
        and _xml_local_name(child.tag) == name
    ]
    if len(matches) > 1:
        warnings.add(
            "FormulaFence found repeated "
            f"{name} containers in an external-link {context}; "
            "the affected controls have a coverage gap."
        )
        opaque_entries.extend(
            (
                f"{context}:repeated-{name}:{index}",
                repr(_xml_fragment(child).sort_key()),
            )
            for index, child in enumerate(matches[1:], start=1)
        )
    return matches[0] if matches else None


def _external_link_part_inspection(
    archive: ZipFile,
    member: str,
    root: ElementTree.Element,
    warnings: set[str],
) -> _ExternalLinkPartInspection:
    """Read one external-link definition while keeping all endpoint material private."""
    opaque_entries = _external_link_opaque_entries(
        root,
        context="external-link",
        known_children=frozenset({"externalBook", "ddeLink", "oleLink"}),
    )
    recognised_children = [
        child
        for child in root
        if _xml_namespace(child.tag) == _SPREADSHEETML_NS
        and _xml_local_name(child.tag) in {"externalBook", "ddeLink", "oleLink"}
    ]
    if len(recognised_children) != 1:
        warnings.add(
            "FormulaFence found an external-link part without exactly one supported link "
            "definition; the affected external-link controls have a coverage gap."
        )
        opaque_entries.append(
            ("external-link:ambiguous-definition", repr(_xml_fragment(root).sort_key()))
        )
        return _ExternalLinkPartInspection(
            member=member,
            opaque_entries=tuple(opaque_entries),
        )

    element = recognised_children[0]
    kind = _xml_local_name(element.tag)
    source_entries: list[tuple[str, str]] = [("kind", kind)]
    definition_entries: list[tuple[str, str]] = []
    cached_entries: list[tuple[str, str]] = []
    standard = _SPREADSHEETML_NS
    relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"

    if kind == "externalBook":
        relationship_signature = _external_link_relationship_signature(
            archive,
            member,
            element.get(relationship_id_attribute),
            warnings,
            opaque_entries,
        )
        if relationship_signature is not None:
            source_entries.append(("target", relationship_signature))
        opaque_entries.extend(
            _external_link_opaque_entries(
                element,
                context="external-workbook",
                known_attributes=frozenset({"id"}),
                known_children=frozenset({"sheetNames", "definedNames", "sheetDataSet"}),
            )
        )
        sheet_names = _external_link_optional_child(
            element,
            "sheetNames",
            warnings,
            opaque_entries,
            context="external-workbook definition",
        )
        sheet_name_count = 0
        if sheet_names is not None:
            opaque_entries.extend(
                _external_link_opaque_entries(
                    sheet_names,
                    context="external-workbook-sheet-names",
                    known_children=frozenset({"sheetName"}),
                )
            )
            for index, sheet_name in enumerate(
                sheet_names.findall(f"{{{standard}}}sheetName")
            ):
                value = sheet_name.get("val")
                if value is None:
                    warnings.add(
                        "FormulaFence found an external-workbook sheet name without a value; "
                        "the affected external-link control has a coverage gap."
                    )
                    value = ""
                definition_entries.append((f"sheet-name:{index}", value))
                opaque_entries.extend(
                    _external_link_opaque_entries(
                        sheet_name,
                        context="external-workbook-sheet-name",
                        known_attributes=frozenset({"val"}),
                    )
                )
                sheet_name_count += 1
        defined_names = _external_link_optional_child(
            element,
            "definedNames",
            warnings,
            opaque_entries,
            context="external-workbook definition",
        )
        defined_name_count = 0
        if defined_names is not None:
            opaque_entries.extend(
                _external_link_opaque_entries(
                    defined_names,
                    context="external-workbook-defined-names",
                    known_children=frozenset({"definedName"}),
                )
            )
            for index, defined_name in enumerate(
                defined_names.findall(f"{{{standard}}}definedName")
            ):
                if defined_name.get("name") is None:
                    warnings.add(
                        "FormulaFence found an external defined name without a name; "
                        "the affected external-link control has a coverage gap."
                    )
                definition_entries.append(
                    (
                        f"defined-name:{index}",
                        repr(_xml_fragment(defined_name).sort_key()),
                    )
                )
                opaque_entries.extend(
                    _external_link_opaque_entries(
                        defined_name,
                        context="external-workbook-defined-name",
                        known_attributes=frozenset({"name", "refersTo", "sheetId"}),
                    )
                )
                defined_name_count += 1
        sheet_data_set = _external_link_optional_child(
            element,
            "sheetDataSet",
            warnings,
            opaque_entries,
            context="external-workbook definition",
        )
        cached_sheet_count = 0
        cached_cell_count = 0
        cached_refresh_error_count = 0
        if sheet_data_set is not None:
            opaque_entries.extend(
                _external_link_opaque_entries(
                    sheet_data_set,
                    context="external-workbook-cache",
                    known_children=frozenset({"sheetData"}),
                )
            )
            for index, sheet_data in enumerate(
                sheet_data_set.findall(f"{{{standard}}}sheetData")
            ):
                if sheet_data.get("sheetId") is None:
                    warnings.add(
                        "FormulaFence found cached external-workbook data without a sheet id; "
                        "the affected external-link control has a coverage gap."
                    )
                refresh_error = _external_data_bool(
                    sheet_data.get("refreshError"),
                    False,
                    "refreshError",
                    warnings,
                    context="external-workbook cache",
                )
                cached_refresh_error_count += int(refresh_error)
                cells = tuple(sheet_data.iter(f"{{{standard}}}cell"))
                cached_cell_count += len(cells)
                cached_entries.append(
                    (f"external-workbook-cache:{index}", repr(_xml_fragment(sheet_data).sort_key()))
                )
                opaque_entries.extend(
                    _external_link_opaque_entries(
                        sheet_data,
                        context="external-workbook-cache-sheet",
                        known_attributes=frozenset({"sheetId", "refreshError"}),
                        known_children=frozenset({"row"}),
                    )
                )
                cached_sheet_count += 1
        return _ExternalLinkPartInspection(
            member=member,
            kind="external_workbook",
            external_workbook_sheet_count=sheet_name_count,
            external_defined_name_count=defined_name_count,
            external_workbook_cached_sheet_count=cached_sheet_count,
            external_workbook_cached_cell_count=cached_cell_count,
            external_workbook_cached_refresh_error_count=cached_refresh_error_count,
            source_signature=_private_external_data_signature(tuple(sorted(source_entries))),
            definition_signature=_private_external_data_signature(
                tuple(sorted(definition_entries))
            ),
            cached_material_signature=_private_external_data_signature(
                tuple(sorted(cached_entries))
            ),
            opaque_entries=tuple(opaque_entries),
        )

    if kind == "ddeLink":
        service = element.get("ddeService")
        topic = element.get("ddeTopic")
        if service is None or topic is None:
            warnings.add(
                "FormulaFence found a DDE link without its required service or topic; "
                "the affected external-link control has a coverage gap."
            )
        source_entries.extend(
            [("service", service or ""), ("topic", topic or "")]
        )
        opaque_entries.extend(
            _external_link_opaque_entries(
                element,
                context="dde-link",
                known_attributes=frozenset({"ddeService", "ddeTopic"}),
                known_children=frozenset({"ddeItems"}),
            )
        )
        items = _external_link_optional_child(
            element,
            "ddeItems",
            warnings,
            opaque_entries,
            context="DDE definition",
        )
        item_count = advise_count = ole_count = prefer_picture_count = cached_value_count = 0
        if items is not None:
            opaque_entries.extend(
                _external_link_opaque_entries(
                    items,
                    context="dde-items",
                    known_children=frozenset({"ddeItem"}),
                )
            )
            for index, item in enumerate(items.findall(f"{{{standard}}}ddeItem")):
                advise, ole, prefer_picture = _external_link_item_flags(
                    item, warnings, context="DDE item"
                )
                advise_count += int(advise)
                ole_count += int(ole)
                prefer_picture_count += int(prefer_picture)
                definition_entries.append((f"dde-item:{index}", item.get("name", "0")))
                values = _external_link_optional_child(
                    item,
                    "values",
                    warnings,
                    opaque_entries,
                    context="DDE item",
                )
                if values is not None:
                    cached_value_count += len(values.findall(f"{{{standard}}}value"))
                    cached_entries.append(
                        (f"dde-values:{index}", repr(_xml_fragment(values).sort_key()))
                    )
                opaque_entries.extend(
                    _external_link_opaque_entries(
                        item,
                        context="dde-item",
                        known_attributes=frozenset({"name", "ole", "advise", "preferPic"}),
                        known_children=frozenset({"values"}),
                    )
                )
                item_count += 1
        return _ExternalLinkPartInspection(
            member=member,
            kind="dde",
            dde_item_count=item_count,
            dde_advise_item_count=advise_count,
            dde_ole_item_count=ole_count,
            dde_prefer_picture_item_count=prefer_picture_count,
            dde_cached_value_count=cached_value_count,
            source_signature=_private_external_data_signature(tuple(sorted(source_entries))),
            definition_signature=_private_external_data_signature(
                tuple(sorted(definition_entries))
            ),
            cached_material_signature=_private_external_data_signature(
                tuple(sorted(cached_entries))
            ),
            opaque_entries=tuple(opaque_entries),
        )

    relationship_signature = _external_link_relationship_signature(
        archive,
        member,
        element.get(relationship_id_attribute),
        warnings,
        opaque_entries,
    )
    if relationship_signature is not None:
        source_entries.append(("target", relationship_signature))
    program_id = element.get("progId")
    if program_id is None:
        warnings.add(
            "FormulaFence found an OLE link without its required program id; "
            "the affected external-link control has a coverage gap."
        )
    source_entries.append(("program-id", program_id or ""))
    opaque_entries.extend(
        _external_link_opaque_entries(
            element,
            context="ole-link",
            known_attributes=frozenset({"id", "progId"}),
            known_children=frozenset({"oleItems"}),
        )
    )
    items = _external_link_optional_child(
        element,
        "oleItems",
        warnings,
        opaque_entries,
        context="OLE definition",
    )
    item_count = advise_count = icon_count = prefer_picture_count = 0
    if items is not None:
        opaque_entries.extend(
            _external_link_opaque_entries(
                items,
                context="ole-items",
                known_children=frozenset({"oleItem"}),
            )
        )
        for index, item in enumerate(items.findall(f"{{{standard}}}oleItem")):
            advise, _ole, prefer_picture = _external_link_item_flags(
                item, warnings, context="OLE item"
            )
            icon = _external_data_bool(
                item.get("icon"), False, "icon", warnings, context="OLE item"
            )
            advise_count += int(advise)
            icon_count += int(icon)
            prefer_picture_count += int(prefer_picture)
            definition_entries.append((f"ole-item:{index}", item.get("name", "")))
            opaque_entries.extend(
                _external_link_opaque_entries(
                    item,
                    context="ole-item",
                    known_attributes=frozenset({"name", "icon", "advise", "preferPic"}),
                )
            )
            item_count += 1
    return _ExternalLinkPartInspection(
        member=member,
        kind="ole",
        ole_item_count=item_count,
        ole_advise_item_count=advise_count,
        ole_icon_item_count=icon_count,
        ole_prefer_picture_item_count=prefer_picture_count,
        source_signature=_private_external_data_signature(tuple(sorted(source_entries))),
        definition_signature=_private_external_data_signature(tuple(sorted(definition_entries))),
        cached_material_signature=_private_external_data_signature(tuple(sorted(cached_entries))),
        opaque_entries=tuple(opaque_entries),
    )


def _external_link_packages_snapshot(
    archive: ZipFile,
    workbook: ElementTree.Element,
    relationships: tuple[_PackageRelationship, ...],
    warnings: set[str],
) -> ExternalLinkPackageSnapshot:
    """Inventory externalLink OOXML parts without following their endpoints."""
    external_relationships = [
        relationship
        for relationship in relationships
        if relationship.relationship_type.rsplit("/", maxsplit=1)[-1] == "externalLink"
    ]
    relationships_by_id = {
        relationship.relationship_id: relationship
        for relationship in external_relationships
        if relationship.relationship_id
    }
    external_references = workbook.find(f"{{{_SPREADSHEETML_NS}}}externalReferences")
    reference_ids: list[str] = []
    declaration_opaque_entries: list[tuple[str, str]] = []
    if external_references is not None:
        declaration_opaque_entries.extend(
            _external_link_opaque_entries(
                external_references,
                context="external-link-declarations",
                known_children=frozenset({"externalReference"}),
            )
        )
        identifier_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
        for reference in external_references.findall(
            f"{{{_SPREADSHEETML_NS}}}externalReference"
        ):
            declaration_opaque_entries.extend(
                _external_link_opaque_entries(
                    reference,
                    context="external-link-declaration",
                    known_attributes=frozenset({"id"}),
                )
            )
            if (relationship_id := reference.get(identifier_attribute)) is None:
                warnings.add(
                    "FormulaFence found an external-link declaration without a relationship id; "
                    "the affected external-link controls have a coverage gap."
                )
                continue
            reference_ids.append(relationship_id)
            if relationship_id not in relationships_by_id:
                warnings.add(
                    "FormulaFence could not locate an external-link declaration relationship; "
                    "the affected external-link controls were not compared."
                )
    if len(reference_ids) != len(set(reference_ids)):
        warnings.add(
            "FormulaFence found repeated external-link declaration relationships; "
            "the affected external-link controls have a coverage gap."
        )
    relationship_ids = [
        relationship.relationship_id
        for relationship in external_relationships
        if relationship.relationship_id is not None
    ]
    if len(relationship_ids) != len(set(relationship_ids)):
        warnings.add(
            "FormulaFence found duplicate external-link package relationship ids; "
            "the affected external-link controls have a coverage gap."
        )
    declaration_entries: list[tuple[str, str]] = []
    for index, relationship_id in enumerate(reference_ids):
        relationship = relationships_by_id.get(relationship_id)
        if relationship is None:
            declaration_entries.append((f"external-reference:{index}", "missing"))
            continue
        declaration_entries.append(
            (
                f"external-reference:{index}",
                repr(
                    (
                        relationship.relationship_type,
                        relationship.target_mode,
                        relationship.target,
                    )
                ),
            )
        )
    referenced_relationship_ids = set(reference_ids)
    unreferenced_relationships = sorted(
        (
            (
                relationship.relationship_type,
                relationship.target_mode,
                relationship.target,
            )
            for relationship in external_relationships
            if relationship.relationship_id not in referenced_relationship_ids
        ),
        key=repr,
    )
    for index, relationship in enumerate(unreferenced_relationships):
        declaration_entries.append(
            (f"unreferenced-workbook-relationship:{index}", repr(relationship))
        )
    declaration_signature = _private_external_data_signature(
        tuple(declaration_entries)
    )

    members: set[str] = set()
    declared_targets: list[str] = []
    for relationship in external_relationships:
        if relationship.target is None:
            warnings.add(
                "FormulaFence found an external-link declaration without a safe internal part "
                "target; the affected external-link controls were not compared."
            )
            continue
        members.add(relationship.target)
        declared_targets.append(relationship.target)
        if relationship.relationship_id not in reference_ids:
            warnings.add(
                "FormulaFence found an external-link package relationship not declared by the "
                "workbook; the affected external-link controls have a coverage gap."
            )
    if len(declared_targets) != len(set(declared_targets)):
        warnings.add(
            "FormulaFence found multiple external-link relationships targeting one package "
            "part; the affected external-link controls have a coverage gap."
        )
    discovered_members = {
        member for member in archive.namelist() if _EXTERNAL_LINK_PART_PATTERN.fullmatch(member)
    }
    for _member in discovered_members - members:
        warnings.add(
            "FormulaFence found an external-link package part not declared by the workbook; "
            "the affected external-link controls have a coverage gap."
        )
    members.update(discovered_members)

    inspections: list[_ExternalLinkPartInspection] = []
    for member in sorted(members, key=str.casefold):
        root = _external_link_part_root(
            archive,
            member,
            warnings,
            context="external-link package",
        )
        if root is None:
            continue
        if (
            _xml_local_name(root.tag) != "externalLink"
            or _xml_namespace(root.tag) != _SPREADSHEETML_NS
        ):
            warnings.add(
                "FormulaFence found an external-link package part with an unexpected root; "
                "the affected external-link controls were not compared."
            )
            continue
        inspections.append(_external_link_part_inspection(archive, member, root, warnings))
    if not members:
        return ExternalLinkPackageSnapshot()
    inspected_members = {inspection.member for inspection in inspections}
    uninspected_member_count = len(members - inspected_members)

    def aggregate_signature(attribute: str) -> str | None:
        material = sorted(
            (inspection.member, value)
            for inspection in inspections
            if (value := getattr(inspection, attribute)) is not None
        )
        return _private_external_data_signature(tuple(material))

    opaque_entries = [
        (f"workbook:{entry[0]}", entry[1])
        for entry in declaration_opaque_entries
    ] + [
        (f"{inspection.member}:{entry[0]}", entry[1])
        for inspection in inspections
        for entry in inspection.opaque_entries
    ]
    opaque_entries.sort()
    part_source_signature = aggregate_signature("source_signature")
    source_signature = _private_external_data_signature(
        tuple(
            entry
            for entry in (
                ("declarations", declaration_signature),
                ("parts", part_source_signature),
            )
            if entry[1] is not None
        )
    )
    return ExternalLinkPackageSnapshot(
        external_link_count=len(members),
        external_workbook_count=sum(
            inspection.kind == "external_workbook" for inspection in inspections
        ),
        dde_link_count=sum(inspection.kind == "dde" for inspection in inspections),
        ole_link_count=sum(inspection.kind == "ole" for inspection in inspections),
        unrecognized_link_count=sum(
            inspection.kind == "unrecognized" for inspection in inspections
        )
        + uninspected_member_count,
        external_workbook_sheet_count=sum(
            inspection.external_workbook_sheet_count for inspection in inspections
        ),
        external_defined_name_count=sum(
            inspection.external_defined_name_count for inspection in inspections
        ),
        external_workbook_cached_sheet_count=sum(
            inspection.external_workbook_cached_sheet_count for inspection in inspections
        ),
        external_workbook_cached_cell_count=sum(
            inspection.external_workbook_cached_cell_count for inspection in inspections
        ),
        external_workbook_cached_refresh_error_count=sum(
            inspection.external_workbook_cached_refresh_error_count
            for inspection in inspections
        ),
        dde_item_count=sum(inspection.dde_item_count for inspection in inspections),
        dde_advise_item_count=sum(
            inspection.dde_advise_item_count for inspection in inspections
        ),
        dde_ole_item_count=sum(
            inspection.dde_ole_item_count for inspection in inspections
        ),
        dde_prefer_picture_item_count=sum(
            inspection.dde_prefer_picture_item_count for inspection in inspections
        ),
        dde_cached_value_count=sum(
            inspection.dde_cached_value_count for inspection in inspections
        ),
        ole_item_count=sum(inspection.ole_item_count for inspection in inspections),
        ole_advise_item_count=sum(
            inspection.ole_advise_item_count for inspection in inspections
        ),
        ole_icon_item_count=sum(
            inspection.ole_icon_item_count for inspection in inspections
        ),
        ole_prefer_picture_item_count=sum(
            inspection.ole_prefer_picture_item_count for inspection in inspections
        ),
        source_signature=source_signature,
        definition_signature=aggregate_signature("definition_signature"),
        cached_material_signature=aggregate_signature("cached_material_signature"),
        opaque_metadata=ExternalDataOpaqueMetadataSnapshot(
            count=len(opaque_entries),
            signature=_private_external_data_signature(tuple(opaque_entries)),
        ),
    )


def _external_data_workbook_settings(
    workbook: ElementTree.Element,
    warnings: set[str],
) -> ExternalDataRefreshSettingsSnapshot:
    """Read workbook-wide refresh controls directly from ``workbookPr``."""
    properties = workbook.find(f"{{{_SPREADSHEETML_NS}}}workbookPr")
    if properties is None:
        return ExternalDataRefreshSettingsSnapshot()
    return ExternalDataRefreshSettingsSnapshot(
        update_links=_external_data_enum(
            properties.get("updateLinks"),
            "user_set",
            {"userSet": "user_set", "never": "never", "always": "always"},
            warnings,
            context="workbook properties",
            attribute="updateLinks",
        ),
        allow_refresh_query=_external_data_bool(
            properties.get("allowRefreshQuery"),
            False,
            "allowRefreshQuery",
            warnings,
            context="workbook properties",
        ),
        refresh_all_connections=_external_data_bool(
            properties.get("refreshAllConnections"),
            False,
            "refreshAllConnections",
            warnings,
            context="workbook properties",
        ),
        save_external_link_values=_external_data_bool(
            properties.get("saveExternalLinkValues"),
            True,
            "saveExternalLinkValues",
            warnings,
            context="workbook properties",
        ),
    )


def _connection_source_type(
    element: ElementTree.Element,
    warnings: set[str],
) -> str:
    """Return the safe source-type label for one connection."""
    raw_value = element.get("type")
    if raw_value is None:
        return "unspecified"
    source_type = _external_data_unsigned_int(
        element,
        "type",
        None,
        warnings,
        context="connection",
    )
    if source_type is None:
        return "unrecognized"
    if normalised := _EXTERNAL_CONNECTION_TYPES.get(source_type):
        return normalised
    warnings.add(
        "FormulaFence found an unrecognized external-data connection type; "
        "the affected source has a coverage gap."
    )
    return "unrecognized"


def _connection_reconnection_method(
    element: ElementTree.Element,
    warnings: set[str],
) -> str:
    """Normalize a connection-file reload policy without exposing raw values."""
    raw_value = element.get("reconnectionMethod")
    if raw_value is None:
        return "as_required"
    method = _external_data_unsigned_int(
        element,
        "reconnectionMethod",
        None,
        warnings,
        context="connection",
    )
    if method is None:
        return "unrecognized"
    if normalised := _EXTERNAL_RECONNECTION_METHODS.get(method):
        return normalised
    warnings.add(
        "FormulaFence found an unrecognized external-data connection "
        "reconnection method; the affected control has a coverage gap."
    )
    return "unrecognized"


def _connection_identity_signature(element: ElementTree.Element) -> str | None:
    """Privately retain connection name and description changes."""
    return _private_external_data_signature(
        tuple(
            (attribute, value)
            for attribute in ("name", "description")
            if (value := element.get(attribute)) is not None
        )
    )


def _connection_source_configuration_signature(
    element: ElementTree.Element,
) -> str | None:
    """Privately retain source paths, credentials, and query configuration."""
    entries: list[tuple[str, str]] = [
        (attribute, value)
        for attribute in ("sourceFile", "odcFile", "singleSignOnId")
        if (value := element.get(attribute)) is not None
    ]
    for index, child in enumerate(element):
        if _xml_local_name(child.tag) not in {
            "dbPr",
            "olapPr",
            "webPr",
            "textPr",
            "parameters",
        }:
            continue
        entries.append(
            (
                f"child:{index}:{_xml_display_name(child.tag)}",
                repr(_xml_fragment(child).sort_key()),
            )
        )
    return _private_external_data_signature(tuple(entries))


def _connection_snapshot(
    element: ElementTree.Element,
    warnings: set[str],
) -> ExternalDataConnectionSnapshot:
    """Read one connection while keeping all source material private."""
    deleted = _external_data_bool(
        element.get("deleted"),
        False,
        "deleted",
        warnings,
        context="connection",
    )
    if element.get("id") is None and not deleted:
        warnings.add(
            "FormulaFence found an external-data connection without its required id; "
            "the affected connection has a coverage gap."
        )
    connection_id = _external_data_unsigned_int(
        element,
        "id",
        None,
        warnings,
        context="connection",
    )
    interval = _external_data_unsigned_int(
        element,
        "interval",
        0,
        warnings,
        context="connection",
    )
    parameter_elements = [
        child for child in element if _xml_local_name(child.tag) == "parameters"
    ]
    if len(parameter_elements) > 1:
        warnings.add(
            "FormulaFence found multiple external-data parameter containers; "
            "the affected connection has a coverage gap."
        )
    parameter_count = 0
    parameters_refresh_on_change = 0
    parameter_tag = f"{{{_SPREADSHEETML_NS}}}parameter"
    for parameters in parameter_elements:
        for parameter in parameters.findall(parameter_tag):
            parameter_count += 1
            if _external_data_bool(
                parameter.get("refreshOnChange"),
                False,
                "refreshOnChange",
                warnings,
                context="connection parameter",
            ):
                parameters_refresh_on_change += 1
    source_components: list[str] = []
    for element_name, component in (
        ("dbPr", "database"),
        ("olapPr", "olap"),
        ("webPr", "web_query"),
        ("textPr", "text_import"),
        ("parameters", "parameters"),
    ):
        count = sum(_xml_local_name(child.tag) == element_name for child in element)
        if count:
            source_components.append(component)
        if count > 1:
            warnings.add(
                "FormulaFence found repeated external-data connection source metadata; "
                "the affected connection has a coverage gap."
            )
    return ExternalDataConnectionSnapshot(
        connection_id=connection_id,
        source_type=_connection_source_type(element, warnings),
        deleted=deleted,
        refresh_on_load=_external_data_bool(
            element.get("refreshOnLoad"),
            False,
            "refreshOnLoad",
            warnings,
            context="connection",
        ),
        refresh_interval_minutes=interval if interval else None,
        background=_external_data_bool(
            element.get("background"),
            False,
            "background",
            warnings,
            context="connection",
        ),
        keep_alive=_external_data_bool(
            element.get("keepAlive"),
            False,
            "keepAlive",
            warnings,
            context="connection",
        ),
        save_data=_external_data_bool(
            element.get("saveData"),
            False,
            "saveData",
            warnings,
            context="connection",
        ),
        save_password=_external_data_bool(
            element.get("savePassword"),
            False,
            "savePassword",
            warnings,
            context="connection",
        ),
        has_source_file=element.get("sourceFile") is not None,
        has_connection_file=element.get("odcFile") is not None,
        only_use_connection_file=_external_data_bool(
            element.get("onlyUseConnectionFile"),
            False,
            "onlyUseConnectionFile",
            warnings,
            context="connection",
        ),
        reconnection_method=_connection_reconnection_method(element, warnings),
        credential_method=_external_data_enum(
            element.get("credentials"),
            "integrated",
            _EXTERNAL_CREDENTIAL_METHODS,
            warnings,
            context="connection",
            attribute="credentials",
        ),
        minimum_refreshable_version=(
            _external_data_unsigned_int(
                element,
                "minRefreshableVersion",
                0,
                warnings,
                context="connection",
                maximum=255,
            )
            or 0
        ),
        has_single_sign_on_id=element.get("singleSignOnId") is not None,
        awaiting_initial_refresh=_external_data_bool(
            element.get("new"),
            False,
            "new",
            warnings,
            context="connection",
        ),
        has_name=element.get("name") is not None,
        has_description=element.get("description") is not None,
        source_components=tuple(source_components),
        parameter_count=parameter_count,
        parameters_refresh_on_change=parameters_refresh_on_change,
        identity_signature=_connection_identity_signature(element),
        source_configuration_signature=_connection_source_configuration_signature(element),
        opaque_metadata=_external_data_opaque_metadata(
            element,
            known_attributes=frozenset(
                {
                    "id",
                    "sourceFile",
                    "odcFile",
                    "keepAlive",
                    "interval",
                    "name",
                    "description",
                    "type",
                    "reconnectionMethod",
                    "refreshedVersion",
                    "minRefreshableVersion",
                    "savePassword",
                    "new",
                    "deleted",
                    "onlyUseConnectionFile",
                    "background",
                    "refreshOnLoad",
                    "saveData",
                    "credentials",
                    "singleSignOnId",
                }
            ),
            known_children=frozenset(
                {"dbPr", "olapPr", "webPr", "textPr", "parameters"}
            ),
        ),
    )


def _external_data_part_root(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    *,
    context: str,
) -> ElementTree.Element | None:
    """Read one external-data part without allowing a bad sibling to hide all controls."""
    try:
        return _xml_root(archive, member)
    except (KeyError, ElementTree.ParseError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect "
            f"{context} OOXML ({type(error).__name__}); the affected controls were not compared."
        )
        return None


def _external_data_part_relationships(
    archive: ZipFile,
    source_member: str,
    warnings: set[str],
    *,
    context: str,
) -> tuple[_PackageRelationship, ...]:
    """Read relationships for one external-data owner part safely."""
    try:
        return _package_relationships(archive, source_member)
    except (KeyError, ElementTree.ParseError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect "
            f"{context} relationships ({type(error).__name__}); "
            "the affected controls were not compared."
        )
        return ()


def _connection_snapshots(
    archive: ZipFile,
    relationships: tuple[_PackageRelationship, ...],
    warnings: set[str],
) -> tuple[ExternalDataConnectionSnapshot, ...]:
    """Read every workbook-connected ``connections`` part safely."""
    connection_parts = [
        relationship
        for relationship in relationships
        if relationship.relationship_type.rsplit("/", maxsplit=1)[-1] == "connections"
    ]
    if len(connection_parts) > 1:
        warnings.add(
            "FormulaFence found multiple external-data Connections parts; "
            "the workbook has a coverage gap."
        )
    snapshots: list[ExternalDataConnectionSnapshot] = []
    connection_tag = f"{{{_SPREADSHEETML_NS}}}connection"
    for relationship in connection_parts:
        if relationship.target is None:
            warnings.add(
                "FormulaFence found a Connections relationship without a safe internal target; "
                "the affected controls were not compared."
            )
            continue
        root = _external_data_part_root(
            archive,
            relationship.target,
            warnings,
            context="Connections",
        )
        if root is None:
            continue
        if (
            _xml_local_name(root.tag) != "connections"
            or _xml_namespace(root.tag) != _SPREADSHEETML_NS
        ):
            warnings.add(
                "FormulaFence found a workbook Connections part with an unexpected root; "
                "the affected controls were not compared."
            )
            continue
        if root.attrib or any(
            _xml_local_name(child.tag) != "connection" for child in root
        ):
            warnings.add(
                "FormulaFence found unmodelled Connections-container metadata; "
                "the affected controls have a coverage gap."
            )
        snapshots.extend(
            _connection_snapshot(connection, warnings)
            for connection in root.findall(connection_tag)
        )
    identifiers = [
        snapshot.connection_id
        for snapshot in snapshots
        if snapshot.connection_id is not None
    ]
    if len(identifiers) != len(set(identifiers)):
        warnings.add(
            "FormulaFence found duplicate external-data connection ids; "
            "the affected controls have a coverage gap."
        )
    return tuple(sorted(snapshots, key=ExternalDataConnectionSnapshot.sort_key))


def _query_table_snapshot(
    sheet: str,
    element: ElementTree.Element,
    warnings: set[str],
) -> QueryTableRefreshSnapshot:
    """Read query-table refresh behavior without names or result metadata."""
    if element.get("connectionId") is None:
        warnings.add(
            "FormulaFence found a query table without its required connection id; "
            "the affected control has a coverage gap."
        )
    connection_id = _external_data_unsigned_int(
        element,
        "connectionId",
        None,
        warnings,
        context="query table",
    )
    return QueryTableRefreshSnapshot(
        sheet=sheet,
        connection_id=connection_id,
        refresh_on_load=_external_data_bool(
            element.get("refreshOnLoad"),
            False,
            "refreshOnLoad",
            warnings,
            context="query table",
        ),
        background_refresh=_external_data_bool(
            element.get("backgroundRefresh"),
            True,
            "backgroundRefresh",
            warnings,
            context="query table",
        ),
        refresh_disabled=_external_data_bool(
            element.get("disableRefresh"),
            False,
            "disableRefresh",
            warnings,
            context="query table",
        ),
        remove_data_on_save=_external_data_bool(
            element.get("removeDataOnSave"),
            False,
            "removeDataOnSave",
            warnings,
            context="query table",
        ),
        fill_formulas=_external_data_bool(
            element.get("fillFormulas"),
            False,
            "fillFormulas",
            warnings,
            context="query table",
        ),
        connection_edit_disabled=_external_data_bool(
            element.get("disableEdit"),
            False,
            "disableEdit",
            warnings,
            context="query table",
        ),
        growth_behavior=_external_data_enum(
            element.get("growShrinkType"),
            "insert_delete",
            _QUERY_TABLE_GROWTH_BEHAVIORS,
            warnings,
            context="query table",
            attribute="growShrinkType",
        ),
        has_name=element.get("name") is not None,
        has_refresh_metadata=any(
            _xml_local_name(child.tag) == "queryTableRefresh" for child in element
        ),
        identity_signature=_private_external_data_signature(
            (("name", element.get("name")),)
            if element.get("name") is not None
            else ()
        ),
        opaque_metadata=_external_data_opaque_metadata(
            element,
            known_attributes=frozenset(
                {
                    "name",
                    "headers",
                    "rowNumbers",
                    "disableRefresh",
                    "backgroundRefresh",
                    "firstBackgroundRefresh",
                    "refreshOnLoad",
                    "growShrinkType",
                    "fillFormulas",
                    "removeDataOnSave",
                    "disableEdit",
                    "preserveFormatting",
                    "adjustColumnWidth",
                    "intermediate",
                    "connectionId",
                    "autoFormatId",
                    "applyNumberFormats",
                    "applyBorderFormats",
                    "applyFontFormats",
                    "applyPatternFormats",
                    "applyAlignmentFormats",
                    "applyWidthHeightFormats",
                }
            ),
            known_children=frozenset({"queryTableRefresh"}),
        ),
    )


def _query_table_refresh_snapshots(
    archive: ZipFile,
    sheet_parts: Mapping[str, tuple[str, str]],
    warnings: set[str],
) -> tuple[QueryTableRefreshSnapshot, ...]:
    """Read query tables linked directly from worksheets or through Excel tables."""
    snapshots: list[QueryTableRefreshSnapshot] = []
    seen_parts: set[tuple[str, str]] = set()

    def append_query_table(sheet: str, target: str | None, *, context: str) -> None:
        if target is None:
            warnings.add(
                "FormulaFence found a query-table relationship without a safe internal target; "
                "the affected controls were not compared."
            )
            return
        part_key = (sheet, target)
        if part_key in seen_parts:
            return
        seen_parts.add(part_key)
        root = _external_data_part_root(
            archive,
            target,
            warnings,
            context=context,
        )
        if root is None:
            return
        if (
            _xml_local_name(root.tag) != "queryTable"
            or _xml_namespace(root.tag) != _SPREADSHEETML_NS
        ):
            warnings.add(
                "FormulaFence found a query-table part with an unexpected root; "
                "the affected controls were not compared."
            )
            return
        snapshots.append(_query_table_snapshot(sheet, root, warnings))

    for sheet, (sheet_member, sheet_type) in sheet_parts.items():
        if sheet_type != "worksheet":
            continue
        relationships_member = _relationship_part_path(sheet_member)
        if relationships_member not in archive.namelist():
            continue
        relationships = _external_data_part_relationships(
            archive,
            sheet_member,
            warnings,
            context="worksheet query-table",
        )
        for relationship in relationships:
            relationship_kind = relationship.relationship_type.rsplit("/", maxsplit=1)[-1]
            if relationship_kind == "queryTable":
                append_query_table(
                    sheet,
                    relationship.target,
                    context="worksheet query-table",
                )
                continue
            if relationship_kind != "table":
                continue
            if relationship.target is None:
                warnings.add(
                    "FormulaFence found a worksheet table relationship without a safe internal "
                    "target; linked query-table controls were not compared."
                )
                continue
            table_root = _external_data_part_root(
                archive,
                relationship.target,
                warnings,
                context="worksheet table",
            )
            if table_root is None:
                continue
            if (
                _xml_local_name(table_root.tag) != "table"
                or _xml_namespace(table_root.tag) != _SPREADSHEETML_NS
            ):
                warnings.add(
                    "FormulaFence found a worksheet table part with an unexpected root; "
                    "linked query-table controls were not compared."
                )
                continue
            table_relationships_member = _relationship_part_path(relationship.target)
            if table_relationships_member not in archive.namelist():
                continue
            table_relationships = _external_data_part_relationships(
                archive,
                relationship.target,
                warnings,
                context="worksheet table query-table",
            )
            for table_relationship in table_relationships:
                if table_relationship.relationship_type.rsplit("/", maxsplit=1)[-1] != "queryTable":
                    continue
                append_query_table(
                    sheet,
                    table_relationship.target,
                    context="table query-table",
                )
    return tuple(sorted(snapshots, key=QueryTableRefreshSnapshot.sort_key))


def _pivot_cache_source_type(
    cache_source: ElementTree.Element,
    warnings: set[str],
) -> str:
    """Normalize a pivot cache source type without exposing source values."""
    source_type = cache_source.get("type")
    if source_type is None:
        warnings.add(
            "FormulaFence found a pivot cache source without its required type; "
            "the affected control has a coverage gap."
        )
        return "unrecognized"
    return _external_data_enum(
        source_type,
        "unrecognized",
        _PIVOT_CACHE_SOURCE_TYPES,
        warnings,
        context="pivot cache source",
        attribute="type",
    )


def _pivot_cache_snapshot(
    cache_id: int | None,
    element: ElementTree.Element,
    warnings: set[str],
) -> PivotCacheRefreshSnapshot:
    """Read one pivot-cache source and refresh setting set safely."""
    cache_source = element.find(f"{{{_SPREADSHEETML_NS}}}cacheSource")
    if cache_source is None:
        warnings.add(
            "FormulaFence found a pivot cache without a source definition; "
            "the affected control has a coverage gap."
        )
        source_type = "unrecognized"
        connection_id = None
        source_configuration_signature = None
    else:
        source_type = _pivot_cache_source_type(cache_source, warnings)
        if source_type == "external" and cache_source.get("connectionId") is None:
            warnings.add(
                "FormulaFence found an external pivot cache without a connection id; "
                "the affected control has a coverage gap."
            )
        connection_id = _external_data_unsigned_int(
            cache_source,
            "connectionId",
            None,
            warnings,
            context="pivot cache source",
        )
        source_configuration_signature = _private_external_data_signature(
            (("cacheSource", repr(_xml_fragment(cache_source).sort_key())),)
        )
    return PivotCacheRefreshSnapshot(
        cache_id=cache_id,
        source_type=source_type,
        connection_id=connection_id,
        refresh_on_load=_external_data_bool(
            element.get("refreshOnLoad"),
            False,
            "refreshOnLoad",
            warnings,
            context="pivot cache",
        ),
        background_query=_external_data_bool(
            element.get("backgroundQuery"),
            False,
            "backgroundQuery",
            warnings,
            context="pivot cache",
        ),
        refresh_enabled=_external_data_bool(
            element.get("enableRefresh"),
            True,
            "enableRefresh",
            warnings,
            context="pivot cache",
        ),
        save_data=_external_data_bool(
            element.get("saveData"),
            True,
            "saveData",
            warnings,
            context="pivot cache",
        ),
        upgrade_on_refresh=_external_data_bool(
            element.get("upgradeOnRefresh"),
            False,
            "upgradeOnRefresh",
            warnings,
            context="pivot cache",
        ),
        source_configuration_signature=source_configuration_signature,
        opaque_metadata=_external_data_opaque_metadata(
            element,
            known_attributes=frozenset(
                {
                    "invalid",
                    "saveData",
                    "refreshOnLoad",
                    "optimizeMemory",
                    "enableRefresh",
                    "refreshedBy",
                    "refreshedDate",
                    "backgroundQuery",
                    "missingItemsLimit",
                    "createdVersion",
                    "refreshedVersion",
                    "minRefreshableVersion",
                    "recordCount",
                    "upgradeOnRefresh",
                    "tupleCache",
                    "supportSubquery",
                    "supportAdvancedDrill",
                    "id",
                }
            ),
            known_children=frozenset(
                {
                    "cacheSource",
                    "cacheFields",
                    "cacheHierarchies",
                    "kpis",
                    "tupleCache",
                    "calculatedItems",
                    "calculatedMembers",
                    "dimensions",
                    "measureGroups",
                    "maps",
                }
            ),
        ),
    )


def _pivot_cache_refresh_snapshots(
    archive: ZipFile,
    workbook: ElementTree.Element,
    relationships: tuple[_PackageRelationship, ...],
    warnings: set[str],
) -> tuple[PivotCacheRefreshSnapshot, ...]:
    """Read pivot-cache refresh controls linked from the workbook."""
    relationship_targets = {
        relationship.relationship_id: relationship.target
        for relationship in relationships
        if relationship.relationship_id
        and relationship.relationship_type.rsplit("/", maxsplit=1)[-1]
        == "pivotCacheDefinition"
    }
    cache_container = workbook.find(f"{{{_SPREADSHEETML_NS}}}pivotCaches")
    if cache_container is None:
        return ()
    cache_tag = f"{{{_SPREADSHEETML_NS}}}pivotCache"
    relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
    snapshots: list[PivotCacheRefreshSnapshot] = []
    for cache in cache_container.findall(cache_tag):
        if cache.get("cacheId") is None:
            warnings.add(
                "FormulaFence found a pivot-cache declaration without its required cache id; "
                "the affected control has a coverage gap."
            )
        cache_id = _external_data_unsigned_int(
            cache,
            "cacheId",
            None,
            warnings,
            context="pivot-cache declaration",
        )
        relationship_id = cache.get(relationship_id_attribute)
        if not relationship_id or relationship_id not in relationship_targets:
            warnings.add(
                "FormulaFence could not locate a pivot-cache definition relationship; "
                "the affected controls were not compared."
            )
            continue
        target = relationship_targets[relationship_id]
        if target is None:
            warnings.add(
                "FormulaFence found a pivot-cache definition without a safe internal target; "
                "the affected controls were not compared."
            )
            continue
        root = _external_data_part_root(
            archive,
            target,
            warnings,
            context="pivot-cache definition",
        )
        if root is None:
            continue
        if (
            _xml_local_name(root.tag) != "pivotCacheDefinition"
            or _xml_namespace(root.tag) != _SPREADSHEETML_NS
        ):
            warnings.add(
                "FormulaFence found a pivot-cache definition with an unexpected root; "
                "the affected controls were not compared."
            )
            continue
        snapshots.append(_pivot_cache_snapshot(cache_id, root, warnings))
    identifiers = [snapshot.cache_id for snapshot in snapshots if snapshot.cache_id is not None]
    if len(identifiers) != len(set(identifiers)):
        warnings.add(
            "FormulaFence found duplicate pivot-cache ids; "
            "the affected controls have a coverage gap."
        )
    return tuple(sorted(snapshots, key=PivotCacheRefreshSnapshot.sort_key))


def _external_data_metadata(path: Path) -> _ExternalDataMetadata:
    """Read external-data refresh controls before the workbook reader drops them."""
    warnings: set[str] = set()
    default_settings = ExternalDataRefreshSettingsSnapshot()
    default_external_link_packages = ExternalLinkPackageSnapshot()
    default_power_query = PowerQuerySnapshot()
    try:
        with ZipFile(path) as archive:
            workbook = _external_data_part_root(
                archive,
                "xl/workbook.xml",
                warnings,
                context="workbook",
            )
            if workbook is None:
                return _ExternalDataMetadata(
                    refresh_settings=default_settings,
                    connections=(),
                    query_tables=(),
                    pivot_caches=(),
                    external_link_packages=default_external_link_packages,
                    power_query=default_power_query,
                    warnings=tuple(sorted(warnings)),
                )
            refresh_settings = _external_data_workbook_settings(workbook, warnings)
            workbook_relationships = _external_data_part_relationships(
                archive,
                "xl/workbook.xml",
                warnings,
                context="workbook",
            )
            connections = _connection_snapshots(
                archive,
                workbook_relationships,
                warnings,
            )
            try:
                sheet_parts = _sheet_xml_parts(archive)
            except (KeyError, ElementTree.ParseError, ValueError) as error:
                warnings.add(
                    "FormulaFence could not map worksheet OOXML for query-table inspection "
                    f"({type(error).__name__}); query-table controls were not compared."
                )
                sheet_parts = {}
            query_tables = _query_table_refresh_snapshots(
                archive,
                sheet_parts,
                warnings,
            )
            pivot_caches = _pivot_cache_refresh_snapshots(
                archive,
                workbook,
                workbook_relationships,
                warnings,
            )
            external_link_packages = _external_link_packages_snapshot(
                archive,
                workbook,
                workbook_relationships,
                warnings,
            )
            power_query = _power_query_snapshot(archive, warnings)
    except (BadZipFile, OSError, ValueError) as error:
        return _ExternalDataMetadata(
            refresh_settings=default_settings,
            connections=(),
            query_tables=(),
            pivot_caches=(),
            external_link_packages=default_external_link_packages,
            power_query=default_power_query,
            warnings=(
                "FormulaFence could not inspect external-data OOXML "
                f"({type(error).__name__}); external-data controls were not compared.",
            ),
        )
    return _ExternalDataMetadata(
        refresh_settings=refresh_settings,
        connections=connections,
        query_tables=query_tables,
        pivot_caches=pivot_caches,
        external_link_packages=external_link_packages,
        power_query=power_query,
        warnings=tuple(sorted(warnings)),
    )


def _normalise_content_type_part_name(part_name: str) -> str | None:
    """Turn an OPC content-type part name into a safe ZIP member name."""
    normalised = posixpath.normpath(part_name.lstrip("/"))
    if normalised in {"", ".", ".."} or normalised.startswith("../"):
        return None
    return normalised


def _xlm_macro_content_type_parts(
    archive: ZipFile,
    warnings: set[str],
) -> dict[str, set[str]]:
    """Find XLM macro-sheet parts declared through OPC content types."""
    try:
        root = _xml_root(archive, "[Content_Types].xml")
    except (KeyError, ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect XLM macro-sheet content types "
            f"({type(error).__name__}); affected macro-sheet controls may be incomplete."
        )
        return {}
    if (
        _xml_local_name(root.tag) != "Types"
        or _xml_namespace(root.tag) != _CONTENT_TYPES_NS
    ):
        warnings.add(
            "FormulaFence found an unexpected OPC content-types root while inspecting "
            "XLM macro sheets; affected macro-sheet controls may be incomplete."
        )
        return {}

    parts: dict[str, set[str]] = defaultdict(set)
    override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
    for override in root.findall(override_tag):
        kind = _XLM_MACRO_SHEET_CONTENT_TYPES.get(override.get("ContentType", ""))
        if kind is None:
            continue
        part_name = override.get("PartName")
        member = _normalise_content_type_part_name(part_name) if part_name else None
        if member is None:
            warnings.add(
                "FormulaFence found an XLM macro-sheet content type without a safe part "
                "name; the affected macro-sheet controls were not compared."
            )
            continue
        parts[member].add(kind)
    return dict(parts)


def _xlm_raw_relationships(
    archive: ZipFile,
    source_member: str,
    warnings: set[str],
    *,
    context: str,
    missing_is_warning: bool = False,
) -> tuple[_XlmRawRelationship, ...]:
    """Read private relationship endpoints without following any target."""
    relationship_member = _relationship_part_path(source_member)
    try:
        root = _xml_root(archive, relationship_member)
    except KeyError:
        if missing_is_warning:
            warnings.add(
                "FormulaFence could not locate "
                f"{context} relationships while inspecting XLM macro sheets; "
                "affected macro-sheet controls may be incomplete."
            )
        return ()
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect "
            f"{context} relationships for XLM macro sheets ({type(error).__name__}); "
            "affected macro-sheet controls were not compared."
        )
        return ()
    if (
        _xml_local_name(root.tag) != "Relationships"
        or _xml_namespace(root.tag) != _PACKAGE_RELATIONSHIP_NS
    ):
        warnings.add(
            "FormulaFence found an unexpected "
            f"{context} relationship root while inspecting XLM macro sheets; "
            "affected macro-sheet controls were not compared."
        )
        return ()

    relationship_tag = f"{{{_PACKAGE_RELATIONSHIP_NS}}}Relationship"
    relationships: list[_XlmRawRelationship] = []
    for relationship in root.findall(relationship_tag):
        target = relationship.get("Target")
        target_mode = relationship.get("TargetMode", "Internal")
        safe_target = (
            _normalise_part_target(source_member, target)
            if target is not None and target_mode.casefold() == "internal"
            else None
        )
        relationships.append(
            _XlmRawRelationship(
                relationship_id=relationship.get("Id"),
                relationship_type=relationship.get("Type", ""),
                target=target,
                target_mode=target_mode,
                safe_target=safe_target,
            )
        )
    if len(relationships) != len(root):
        warnings.add(
            "FormulaFence found unmodelled relationship XML while inspecting XLM macro "
            "sheets; affected macro-sheet controls may be incomplete."
        )
    return tuple(relationships)


def _xlm_relationship_signature(
    relationships: tuple[_XlmRawRelationship, ...],
) -> str | None:
    """Fingerprint relationship semantics while ignoring arbitrary identifiers."""
    material = sorted(relationship.semantic_key() for relationship in relationships)
    return _private_external_data_signature(
        tuple(
            (f"relationship:{index}", repr(relationship))
            for index, relationship in enumerate(material)
        )
    )


def _xlm_related_part_payloads(
    archive: ZipFile,
    relationships: tuple[_XlmRawRelationship, ...],
    warnings: set[str],
    budget: _XlmRelatedPartBudget,
) -> _XlmRelatedPartPayloadInspection:
    """Fingerprint direct internal XLM relationship targets without parsing them.

    Macro-sheet XML can point to OLE objects, embedded packages, drawings, and
    other package parts. Hash their bytes only after resolving a safe internal
    target; never follow an external target or parse an embedded payload.
    """
    members: set[str] = set()
    unresolved_entries: list[tuple[str, str]] = []
    for relationship in relationships:
        if relationship.target_mode.casefold() != "internal":
            continue
        if relationship.safe_target is None:
            warnings.add(
                "FormulaFence found an XLM macro-sheet internal relationship without "
                "a safe related-part target; the affected controls have a coverage gap."
            )
            unresolved_entries.append(
                (
                    "unsafe-target",
                    repr((relationship.relationship_type, relationship.target)),
                )
            )
            continue
        members.add(relationship.safe_target)

    entries = list(unresolved_entries)
    fingerprinted_part_count = 0
    uninspected_part_count = len(unresolved_entries)
    for member in sorted(members, key=str.casefold):
        if budget.remaining_parts == 0:
            warnings.add(
                "FormulaFence reached its bounded XLM macro-sheet related-part count "
                "budget; the affected controls have a coverage gap."
            )
            entries.append(("part-count-budget-exhausted", member))
            uninspected_part_count += 1
            continue
        budget.remaining_parts -= 1
        try:
            info = archive.getinfo(member)
        except KeyError:
            warnings.add(
                "FormulaFence could not locate an XLM macro-sheet internal related "
                "part; the affected controls were not compared."
            )
            entries.append(("missing-part", member))
            uninspected_part_count += 1
            continue
        metadata = repr((member, info.file_size, info.compress_size, info.CRC))
        if info.file_size > _XLM_RELATED_PART_MAX_BYTES:
            warnings.add(
                "FormulaFence did not fully read an oversized XLM macro-sheet related "
                "part; the affected controls have a coverage gap."
            )
            entries.append(("oversized-part", metadata))
            uninspected_part_count += 1
            continue
        if info.file_size > budget.remaining_bytes:
            warnings.add(
                "FormulaFence reached its bounded XLM macro-sheet related-part read "
                "budget; the affected controls have a coverage gap."
            )
            entries.append(("read-budget-exhausted", metadata))
            uninspected_part_count += 1
            continue
        budget.remaining_bytes -= info.file_size
        digest = hashlib.sha256()
        bytes_read = 0
        try:
            with archive.open(info) as related_part:
                while chunk := related_part.read(_XLM_RELATED_PART_HASH_CHUNK_BYTES):
                    bytes_read += len(chunk)
                    if bytes_read > info.file_size:
                        raise ValueError("related part exceeded its declared size")
                    digest.update(chunk)
            if bytes_read != info.file_size:
                raise ValueError("related part did not match its declared size")
        except (BadZipFile, OSError, RuntimeError, ValueError) as error:
            warnings.add(
                "FormulaFence could not fingerprint an XLM macro-sheet internal related "
                f"part ({type(error).__name__}); the affected controls were not compared."
            )
            entries.append(("unreadable-part", metadata))
            uninspected_part_count += 1
            continue
        entries.append(("payload", repr((member, digest.hexdigest()))))
        fingerprinted_part_count += 1

    entries.sort()
    return _XlmRelatedPartPayloadInspection(
        internal_part_count=len(members) + len(unresolved_entries),
        fingerprinted_part_count=fingerprinted_part_count,
        uninspected_part_count=uninspected_part_count,
        payload_signature=_private_external_data_signature(tuple(entries)),
    )


def _xlm_macro_fragment(
    element: ElementTree.Element,
    relationship_semantics: Mapping[str, tuple[str, str, str]],
) -> tuple[object, ...]:
    """Canonicalize macro XML while resolving relationship ids privately.

    Relationship identifiers are writer-chosen.  Replacing each ``r:id`` with
    its relationship semantic tuple makes an identifier-only rewrite compare
    equal, but retains the target/type change that would alter macro behavior.
    """
    relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
    attributes: list[tuple[str, str]] = []
    for attribute, value in element.attrib.items():
        if attribute == relationship_attribute:
            relationship = relationship_semantics.get(value)
            resolved = (
                ("relationship", relationship)
                if relationship is not None
                else ("missing-relationship", value)
            )
            attributes.append((_xml_display_name(attribute), repr(resolved)))
        else:
            attributes.append((_xml_display_name(attribute), value))
    children = tuple(
        _xlm_macro_fragment(child, relationship_semantics) for child in element
    )
    text = element.text
    if children and text is not None and not text.strip():
        text = None
    return (
        _xml_display_name(element.tag),
        tuple(sorted(attributes)),
        text,
        children,
    )


def _xlm_macro_formula_cell_count(root: ElementTree.Element) -> int:
    """Count macro-sheet cells with a formula without evaluating the formula."""
    return sum(
        any(_xml_local_name(child.tag) == "f" for child in cell)
        for cell in root.iter()
        if _xml_local_name(cell.tag) == "c"
    )


def _xlm_macro_part_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    related_part_budget: _XlmRelatedPartBudget,
) -> _XlmMacroSheetInspection:
    """Fingerprint one XLM macro-sheet part and its relationships privately."""
    try:
        info = archive.getinfo(member)
    except KeyError:
        warnings.add(
            "FormulaFence could not locate an XLM macro-sheet package part; "
            "the affected macro-sheet controls were not compared."
        )
        return _XlmMacroSheetInspection(
            member=member,
            program_signature=_private_external_data_signature(
                (("missing-member", member),)
            ),
        )
    if info.file_size > _XLM_MACRO_SHEET_MAX_PART_BYTES:
        warnings.add(
            "FormulaFence did not fully read an oversized XLM macro-sheet part; "
            "the affected macro-sheet controls have a coverage gap."
        )
        return _XlmMacroSheetInspection(
            member=member,
            program_signature=_private_external_data_signature(
                (
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            ),
        )
    payload: bytes | None = None
    try:
        payload = archive.read(member)
        root = _xml_root_from_payload(payload)
    except (KeyError, ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect an XLM macro-sheet XML part "
            f"({type(error).__name__}); the affected macro-sheet controls were not compared."
        )
        signature = (
            _private_payload_signature(payload)
            if payload is not None
            else _private_external_data_signature(
                (
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            )
        )
        return _XlmMacroSheetInspection(
            member=member,
            program_signature=signature,
        )
    if (
        _xml_local_name(root.tag) != "macrosheet"
        or _xml_namespace(root.tag) != _EXCEL_2006_MAIN_NS
    ):
        warnings.add(
            "FormulaFence found an XLM macro-sheet part with an unexpected root; "
            "the affected macro-sheet controls were not compared."
        )
        return _XlmMacroSheetInspection(
            member=member,
            program_signature=_private_payload_signature(payload),
        )

    relationships = _xlm_raw_relationships(
        archive,
        member,
        warnings,
        context="XLM macro-sheet",
    )
    relationships_by_id: dict[str, list[_XlmRawRelationship]] = defaultdict(list)
    for relationship in relationships:
        if relationship.relationship_id:
            relationships_by_id[relationship.relationship_id].append(relationship)
    relationship_semantics: dict[str, tuple[str, str, str]] = {}
    for relationship_id, matches in relationships_by_id.items():
        semantics = sorted(match.semantic_key() for match in matches)
        if len(semantics) > 1:
            warnings.add(
                "FormulaFence found duplicate XLM macro-sheet relationship ids; "
                "the affected macro-sheet controls have a coverage gap."
            )
        relationship_semantics[relationship_id] = semantics[0]

    referenced_relationship_ids = {
        value
        for element in root.iter()
        if (value := element.get(f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id")) is not None
    }
    if referenced_relationship_ids - set(relationship_semantics):
        warnings.add(
            "FormulaFence found an XLM macro-sheet relationship reference without a "
            "matching relationship; the affected macro-sheet controls have a coverage gap."
        )
    relationship_kinds = [
        relationship.relationship_type.rsplit("/", maxsplit=1)[-1]
        for relationship in relationships
    ]
    related_part_payloads = _xlm_related_part_payloads(
        archive,
        relationships,
        warnings,
        related_part_budget,
    )
    formula_cell_count = _xlm_macro_formula_cell_count(root)
    try:
        program_signature = _private_external_data_signature(
            (("macro-sheet", repr(_xlm_macro_fragment(root, relationship_semantics))),)
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested XLM macro-sheet "
            "part; the affected macro-sheet controls were not compared."
        )
        return _XlmMacroSheetInspection(
            member=member,
            formula_cell_count=formula_cell_count,
            related_relationship_count=len(relationships),
            external_relationship_count=sum(
                relationship.target_mode.casefold() != "internal"
                for relationship in relationships
            ),
            internal_related_part_count=related_part_payloads.internal_part_count,
            fingerprinted_related_part_count=(
                related_part_payloads.fingerprinted_part_count
            ),
            uninspected_related_part_count=related_part_payloads.uninspected_part_count,
            embedded_object_relationship_count=relationship_kinds.count("oleObject"),
            embedded_package_relationship_count=relationship_kinds.count("package"),
            program_signature=_private_payload_signature(payload),
            relationship_signature=_xlm_relationship_signature(relationships),
            related_part_payload_signature=related_part_payloads.payload_signature,
        )
    return _XlmMacroSheetInspection(
        member=member,
        formula_cell_count=formula_cell_count,
        related_relationship_count=len(relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in relationships
        ),
        internal_related_part_count=related_part_payloads.internal_part_count,
        fingerprinted_related_part_count=related_part_payloads.fingerprinted_part_count,
        uninspected_related_part_count=related_part_payloads.uninspected_part_count,
        embedded_object_relationship_count=relationship_kinds.count("oleObject"),
        embedded_package_relationship_count=relationship_kinds.count("package"),
        inspected=True,
        program_signature=program_signature,
        relationship_signature=_xlm_relationship_signature(relationships),
        related_part_payload_signature=related_part_payloads.payload_signature,
    )


def _xlm_macro_metadata(path: Path) -> _XlmMacroMetadata:
    """Inspect XLM macro sheets before a workbook library can omit their code."""
    warnings: set[str] = set()
    default = XlmMacroSheetSnapshot()
    try:
        with ZipFile(path) as archive:
            try:
                workbook = _xml_root(archive, "xl/workbook.xml")
            except (KeyError, ElementTree.ParseError, ValueError) as error:
                warnings.add(
                    "FormulaFence could not inspect workbook OOXML for XLM macro sheets "
                    f"({type(error).__name__}); affected macro-sheet controls were not compared."
                )
                return _XlmMacroMetadata(default, tuple(sorted(warnings)))
            if (
                _xml_local_name(workbook.tag) != "workbook"
                or _xml_namespace(workbook.tag) != _SPREADSHEETML_NS
            ):
                warnings.add(
                    "FormulaFence found an unexpected workbook root while inspecting XLM "
                    "macro sheets; affected macro-sheet controls were not compared."
                )
                return _XlmMacroMetadata(default, tuple(sorted(warnings)))

            workbook_relationships = _xlm_raw_relationships(
                archive,
                "xl/workbook.xml",
                warnings,
                context="workbook",
                missing_is_warning=True,
            )
            macro_relationships = tuple(
                relationship
                for relationship in workbook_relationships
                if relationship.relationship_type in _XLM_MACRO_SHEET_RELATIONSHIPS
            )
            workbook_relationships_by_id: dict[
                str, list[_XlmRawRelationship]
            ] = defaultdict(list)
            for relationship in workbook_relationships:
                if relationship.relationship_id:
                    workbook_relationships_by_id[relationship.relationship_id].append(
                        relationship
                    )
            workbook_relationship_semantics = {
                relationship_id: sorted(
                    relationship.semantic_key() for relationship in matches
                )[0]
                for relationship_id, matches in workbook_relationships_by_id.items()
            }
            content_type_parts = _xlm_macro_content_type_parts(archive, warnings)
            discovered_members = {
                entry.filename
                for entry in archive.infolist()
                if _XLM_MACRO_SHEET_PART_PATTERN.fullmatch(entry.filename)
            }

            candidate_kinds: dict[str, set[str]] = defaultdict(set)
            declaration_entries: list[tuple[str, str]] = []
            declared_targets: set[str] = set()
            relationships_by_id: dict[str, list[_XlmRawRelationship]] = defaultdict(list)
            for relationship in macro_relationships:
                kind = _XLM_MACRO_SHEET_RELATIONSHIPS[relationship.relationship_type]
                declaration_entries.append(
                    ("workbook-relationship", repr((kind, relationship.semantic_key())))
                )
                if relationship.relationship_id:
                    relationships_by_id[relationship.relationship_id].append(relationship)
                else:
                    warnings.add(
                        "FormulaFence found an XLM macro-sheet workbook relationship without "
                        "an id; the affected macro-sheet controls have a coverage gap."
                    )
                if relationship.safe_target is None:
                    warnings.add(
                        "FormulaFence found an XLM macro-sheet relationship without a safe "
                        "internal part target; the affected macro-sheet controls were not compared."
                    )
                    continue
                candidate_kinds[relationship.safe_target].add(kind)
                declared_targets.add(relationship.safe_target)

            duplicate_relationship_ids = {
                relationship_id
                for relationship_id, matches in relationships_by_id.items()
                if len(matches) > 1
            }
            if duplicate_relationship_ids:
                warnings.add(
                    "FormulaFence found duplicate XLM macro-sheet workbook relationship ids; "
                    "the affected macro-sheet controls have a coverage gap."
                )

            for member, kinds in content_type_parts.items():
                candidate_kinds[member].update(kinds)
                declaration_entries.append(
                    ("content-type", repr((member, tuple(sorted(kinds)))))
                )
                if member not in declared_targets:
                    warnings.add(
                        "FormulaFence found an XLM macro-sheet content-type part not declared "
                        "by the workbook; the affected macro-sheet controls have a coverage gap."
                    )
            for member in discovered_members:
                candidate_kinds.setdefault(member, set())
                if member not in declared_targets:
                    warnings.add(
                        "FormulaFence found an XLM macro-sheet package part not declared by "
                        "the workbook; the affected macro-sheet controls have a coverage gap."
                    )

            binding_entries: list[tuple[str, str]] = []
            bound_relationship_ids: set[str] = set()
            binding_occurrences: dict[str, int] = defaultdict(int)
            hidden_macro_sheet_count = 0
            very_hidden_macro_sheet_count = 0
            sheets = workbook.find(f"{{{_SPREADSHEETML_NS}}}sheets")
            relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
            if sheets is not None:
                for index, sheet in enumerate(sheets.findall(f"{{{_SPREADSHEETML_NS}}}sheet")):
                    relationship_id = sheet.get(relationship_id_attribute)
                    matches = relationships_by_id.get(relationship_id or "", [])
                    if not matches:
                        continue
                    if len(matches) != 1:
                        warnings.add(
                            "FormulaFence found a workbook sheet bound to repeated XLM macro-sheet "
                            "relationships; the affected macro-sheet controls have a coverage gap."
                        )
                    relationship = matches[0]
                    state = sheet.get("state", "visible")
                    if state not in {"visible", "hidden", "veryHidden"}:
                        warnings.add(
                            "FormulaFence found an unrecognized XLM macro-sheet visibility state; "
                            "the affected macro-sheet controls have a coverage gap."
                        )
                    hidden_macro_sheet_count += int(state == "hidden")
                    very_hidden_macro_sheet_count += int(state == "veryHidden")
                    if sheet.get("name") is None:
                        warnings.add(
                            "FormulaFence found an XLM macro-sheet workbook declaration without "
                            "a name; the affected macro-sheet controls have a coverage gap."
                        )
                    binding_entries.append(
                        (
                            f"sheet-binding:{index}",
                            repr(
                                _xlm_macro_fragment(
                                    sheet, workbook_relationship_semantics
                                )
                            ),
                        )
                    )
                    if relationship_id:
                        bound_relationship_ids.add(relationship_id)
                        binding_occurrences[relationship_id] += 1
            if any(count > 1 for count in binding_occurrences.values()):
                warnings.add(
                    "FormulaFence found multiple workbook sheets bound to one XLM macro-sheet "
                    "relationship; the affected macro-sheet controls have a coverage gap."
                )
            for relationship in macro_relationships:
                if (
                    relationship.relationship_id is None
                    or relationship.relationship_id not in bound_relationship_ids
                ):
                    warnings.add(
                        "FormulaFence found an XLM macro-sheet relationship not bound to a "
                        "workbook sheet; the affected macro-sheet controls have a coverage gap."
                    )
                    binding_entries.append(
                        (
                            "unbound-workbook-relationship",
                            repr(relationship.semantic_key()),
                        )
                    )

            inspections: list[_XlmMacroSheetInspection] = []
            unrecognized_macro_sheet_members: set[str] = set()
            international_macro_sheet_count = 0
            related_part_budget = _XlmRelatedPartBudget()
            for member in sorted(candidate_kinds, key=str.casefold):
                kinds = candidate_kinds[member]
                if len(kinds) != 1:
                    warnings.add(
                        "FormulaFence could not identify one XLM macro-sheet part's declared "
                        "kind; the affected macro-sheet controls have a coverage gap."
                    )
                    unrecognized_macro_sheet_members.add(member)
                elif "international" in kinds:
                    international_macro_sheet_count += 1
                inspection = _xlm_macro_part_inspection(
                    archive,
                    member,
                    warnings,
                    related_part_budget,
                )
                inspections.append(inspection)
                if not inspection.inspected:
                    unrecognized_macro_sheet_members.add(member)

            def aggregate_signature(attribute: str) -> str | None:
                material = sorted(
                    (inspection.member, value)
                    for inspection in inspections
                    if (value := getattr(inspection, attribute)) is not None
                )
                return _private_external_data_signature(tuple(material))

            declaration_entries.extend(binding_entries)
            declaration_entries.sort()
            snapshot = XlmMacroSheetSnapshot(
                declared_macro_sheet_count=len(macro_relationships),
                macro_sheet_count=len(candidate_kinds),
                international_macro_sheet_count=international_macro_sheet_count,
                unrecognized_macro_sheet_count=len(unrecognized_macro_sheet_members),
                hidden_macro_sheet_count=hidden_macro_sheet_count,
                very_hidden_macro_sheet_count=very_hidden_macro_sheet_count,
                formula_cell_count=sum(
                    inspection.formula_cell_count for inspection in inspections
                ),
                related_relationship_count=sum(
                    inspection.related_relationship_count for inspection in inspections
                ),
                external_relationship_count=sum(
                    inspection.external_relationship_count for inspection in inspections
                ),
                internal_related_part_count=sum(
                    inspection.internal_related_part_count for inspection in inspections
                ),
                fingerprinted_related_part_count=sum(
                    inspection.fingerprinted_related_part_count
                    for inspection in inspections
                ),
                uninspected_related_part_count=sum(
                    inspection.uninspected_related_part_count for inspection in inspections
                ),
                embedded_object_relationship_count=sum(
                    inspection.embedded_object_relationship_count
                    for inspection in inspections
                ),
                embedded_package_relationship_count=sum(
                    inspection.embedded_package_relationship_count
                    for inspection in inspections
                ),
                declaration_signature=_private_external_data_signature(
                    tuple(declaration_entries)
                ),
                program_signature=aggregate_signature("program_signature"),
                relationship_signature=aggregate_signature("relationship_signature"),
                related_part_payload_signature=aggregate_signature(
                    "related_part_payload_signature"
                ),
            )
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        return _XlmMacroMetadata(
            default,
            (
                "FormulaFence could not inspect XLM macro-sheet OOXML "
                f"({type(error).__name__}); affected macro-sheet controls were not compared.",
            ),
        )
    return _XlmMacroMetadata(snapshot, tuple(sorted(warnings)))


def _ribbon_raw_relationships(
    archive: ZipFile,
    source_member: str,
    warnings: set[str],
    *,
    context: str,
    missing_is_warning: bool = False,
) -> tuple[_RibbonRawRelationship, ...]:
    """Read RibbonX relationships without following any package target."""
    relationship_member = _relationship_part_path(source_member)
    try:
        root = _xml_root(archive, relationship_member)
    except KeyError:
        if missing_is_warning:
            warnings.add(
                "FormulaFence could not locate "
                f"{context} relationships while inspecting RibbonX customization; "
                "affected controls may be incomplete."
            )
        return ()
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect "
            f"{context} relationships for RibbonX customization ({type(error).__name__}); "
            "affected controls were not compared."
        )
        return ()
    if (
        _xml_local_name(root.tag) != "Relationships"
        or _xml_namespace(root.tag) != _PACKAGE_RELATIONSHIP_NS
    ):
        warnings.add(
            "FormulaFence found an unexpected "
            f"{context} relationship root while inspecting RibbonX customization; "
            "affected controls were not compared."
        )
        return ()

    relationship_tag = f"{{{_PACKAGE_RELATIONSHIP_NS}}}Relationship"
    relationships: list[_RibbonRawRelationship] = []
    for relationship in root.findall(relationship_tag):
        target = relationship.get("Target")
        target_mode = relationship.get("TargetMode", "Internal")
        safe_target = (
            _normalise_part_target(source_member, target)
            if target is not None and target_mode.casefold() == "internal"
            else None
        )
        relationships.append(
            _RibbonRawRelationship(
                relationship_id=relationship.get("Id"),
                relationship_type=relationship.get("Type", ""),
                target=target,
                target_mode=target_mode,
                safe_target=safe_target,
            )
        )
    if len(relationships) != len(root):
        warnings.add(
            "FormulaFence found unmodelled relationship XML while inspecting RibbonX "
            "customization; affected controls may be incomplete."
        )
    return tuple(relationships)


def _ribbon_relationship_signature(
    relationships: tuple[_RibbonRawRelationship, ...],
) -> str | None:
    """Fingerprint RibbonX relationship semantics without identifier churn."""
    material = sorted(relationship.semantic_key() for relationship in relationships)
    return _private_external_data_signature(
        tuple(
            (f"relationship:{index}", repr(relationship))
            for index, relationship in enumerate(material)
        )
    )


def _ribbon_fragment(
    element: ElementTree.Element,
    relationship_semantics: Mapping[str, tuple[str, str, str]],
) -> tuple[object, ...]:
    """Canonicalize RibbonX XML while resolving private relationship identifiers.

    A custom UI part can point to embedded images using a writer-chosen
    relationship id. Resolving an ``r:id`` or an ``image`` value only when it
    matches a known relationship keeps ID-only package rewrites out of diffs,
    without hiding a changed relationship endpoint.
    """
    relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
    attributes: list[tuple[str, str]] = []
    for attribute, value in element.attrib.items():
        is_relationship_reference = attribute == relationship_attribute or (
            _xml_local_name(attribute) == "image" and value in relationship_semantics
        )
        if is_relationship_reference:
            relationship = relationship_semantics.get(value)
            resolved = (
                ("relationship", relationship)
                if relationship is not None
                else ("missing-relationship", value)
            )
            attributes.append((_xml_display_name(attribute), repr(resolved)))
        else:
            attributes.append((_xml_display_name(attribute), value))
    children = tuple(_ribbon_fragment(child, relationship_semantics) for child in element)
    text = element.text
    if children and text is not None and not text.strip():
        text = None
    return (
        _xml_display_name(element.tag),
        tuple(sorted(attributes)),
        text,
        children,
    )


def _ribbon_control_counts(root: ElementTree.Element) -> tuple[int, int, int]:
    """Count RibbonX controls and callback attributes without retaining names."""
    control_count = 0
    callback_attribute_count = 0
    action_callback_count = 0
    control_id_attributes = {"id", "idMso", "idQ"}
    for element in root.iter():
        if element is not root and any(
            _xml_local_name(attribute) in control_id_attributes
            for attribute in element.attrib
        ):
            control_count += 1
        for attribute in element.attrib:
            name = _xml_local_name(attribute)
            if name.startswith(("on", "get")) or name == "loadImage":
                callback_attribute_count += 1
            if name == "onAction":
                action_callback_count += 1
    return control_count, callback_attribute_count, action_callback_count


def _ribbon_part_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    budget: _RibbonCustomizationBudget,
) -> _RibbonPartInspection:
    """Fingerprint one bounded RibbonX part and its image relationships privately."""
    if budget.remaining_parts == 0:
        warnings.add(
            "FormulaFence reached its bounded RibbonX customization part count budget; "
            "the affected controls have a coverage gap."
        )
        return _RibbonPartInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (("part-count-budget-exhausted", member),)
            ),
        )
    budget.remaining_parts -= 1
    try:
        info = archive.getinfo(member)
    except KeyError:
        warnings.add(
            "FormulaFence could not locate a RibbonX customization package part; "
            "the affected controls were not compared."
        )
        return _RibbonPartInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (("missing-member", member),)
            ),
        )
    if info.file_size > _RIBBON_CUSTOM_UI_MAX_PART_BYTES:
        warnings.add(
            "FormulaFence did not fully read an oversized RibbonX customization part; "
            "the affected controls have a coverage gap."
        )
        return _RibbonPartInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            ),
        )
    if info.file_size > budget.remaining_bytes:
        warnings.add(
            "FormulaFence reached its bounded RibbonX customization part read budget; "
            "the affected controls have a coverage gap."
        )
        return _RibbonPartInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (
                    ("read-budget-exhausted", member),
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            ),
        )
    budget.remaining_bytes -= info.file_size
    payload: bytes | None = None
    try:
        payload = archive.read(member)
        root = _xml_root_from_payload(payload)
    except (KeyError, ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a RibbonX customization XML part "
            f"({type(error).__name__}); the affected controls were not compared."
        )
        signature = (
            _private_payload_signature(payload)
            if payload is not None
            else _private_external_data_signature(
                (
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            )
        )
        return _RibbonPartInspection(member=member, definition_signature=signature)

    namespace = _xml_namespace(root.tag)
    if (
        _xml_local_name(root.tag) != "customUI"
        or namespace not in _RIBBON_CUSTOM_UI_NAMESPACES
    ):
        warnings.add(
            "FormulaFence found a RibbonX customization part with an unexpected root; "
            "the affected controls were not compared."
        )
        return _RibbonPartInspection(
            member=member,
            definition_signature=_private_payload_signature(payload),
        )

    relationships = _ribbon_raw_relationships(
        archive,
        member,
        warnings,
        context="RibbonX customization",
    )
    relationships_by_id: dict[str, list[_RibbonRawRelationship]] = defaultdict(list)
    for relationship in relationships:
        if relationship.relationship_id:
            relationships_by_id[relationship.relationship_id].append(relationship)
        else:
            warnings.add(
                "FormulaFence found a RibbonX customization relationship without an id; "
                "the affected controls have a coverage gap."
            )
    relationship_semantics: dict[str, tuple[str, str, str]] = {}
    for relationship_id, matches in relationships_by_id.items():
        semantics = sorted(match.semantic_key() for match in matches)
        if len(semantics) > 1:
            warnings.add(
                "FormulaFence found duplicate RibbonX customization relationship ids; "
                "the affected controls have a coverage gap."
            )
        relationship_semantics[relationship_id] = semantics[0]

    referenced_relationship_ids = {
        value
        for element in root.iter()
        if (value := element.get(f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id")) is not None
    }
    if referenced_relationship_ids - set(relationship_semantics):
        warnings.add(
            "FormulaFence found a RibbonX customization relationship reference without a "
            "matching relationship; the affected controls have a coverage gap."
        )
    control_count, callback_attribute_count, action_callback_count = _ribbon_control_counts(
        root
    )
    try:
        definition_signature = _private_external_data_signature(
            (("ribbon", repr(_ribbon_fragment(root, relationship_semantics))),)
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested RibbonX "
            "customization part; the affected controls were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        inspected = False
    else:
        inspected = True
    return _RibbonPartInspection(
        member=member,
        office_2010=_RIBBON_CUSTOM_UI_NAMESPACES[namespace] == "2010",
        control_count=control_count,
        callback_attribute_count=callback_attribute_count,
        action_callback_count=action_callback_count,
        image_relationship_count=sum(
            relationship.relationship_type.rsplit("/", maxsplit=1)[-1].casefold()
            == "image"
            for relationship in relationships
        ),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in relationships
        ),
        inspected=inspected,
        definition_signature=definition_signature,
        relationship_signature=_ribbon_relationship_signature(relationships),
    )


def _ribbon_customization_metadata(path: Path) -> _RibbonCustomizationMetadata:
    """Inspect Office RibbonX callbacks before the workbook reader omits the parts."""
    warnings: set[str] = set()
    default = RibbonCustomizationSnapshot()
    try:
        with ZipFile(path) as archive:
            package_relationships = _ribbon_raw_relationships(
                archive,
                "",
                warnings,
                context="package",
                missing_is_warning=True,
            )
            ribbon_relationships = tuple(
                relationship
                for relationship in package_relationships
                if relationship.relationship_type in _RIBBON_CUSTOM_UI_RELATIONSHIPS
            )
            candidate_kinds: dict[str, set[str]] = defaultdict(set)
            declaration_entries: list[tuple[str, str]] = []
            declared_targets: set[str] = set()
            unresolved_declaration_count = 0
            relationships_by_id: dict[str, list[_RibbonRawRelationship]] = defaultdict(list)
            for relationship in ribbon_relationships:
                kind = _RIBBON_CUSTOM_UI_RELATIONSHIPS[relationship.relationship_type]
                declaration_entries.append(
                    ("package-relationship", repr((kind, relationship.semantic_key())))
                )
                if relationship.relationship_id:
                    relationships_by_id[relationship.relationship_id].append(relationship)
                else:
                    warnings.add(
                        "FormulaFence found a RibbonX package relationship without an id; "
                        "the affected controls have a coverage gap."
                    )
                if relationship.safe_target is None:
                    warnings.add(
                        "FormulaFence found a RibbonX package relationship without a safe "
                        "internal customization target; the affected controls were not compared."
                    )
                    unresolved_declaration_count += 1
                    continue
                candidate_kinds[relationship.safe_target].add(kind)
                declared_targets.add(relationship.safe_target)

            if any(len(matches) > 1 for matches in relationships_by_id.values()):
                warnings.add(
                    "FormulaFence found duplicate RibbonX package relationship ids; "
                    "the affected controls have a coverage gap."
                )

            declared_version_counts: dict[str, int] = defaultdict(int)
            for relationship in ribbon_relationships:
                declared_version_counts[
                    _RIBBON_CUSTOM_UI_RELATIONSHIPS[relationship.relationship_type]
                ] += 1
            if any(count > 1 for count in declared_version_counts.values()):
                warnings.add(
                    "FormulaFence found repeated RibbonX package declarations for one "
                    "customization version; the affected controls have a coverage gap."
                )

            discovered_members = {
                entry.filename
                for entry in archive.infolist()
                if _RIBBON_CUSTOM_UI_PART_PATTERN.fullmatch(entry.filename)
            }
            for member in discovered_members:
                candidate_kinds.setdefault(member, set())
                if member not in declared_targets:
                    warnings.add(
                        "FormulaFence found a RibbonX customization package part not declared "
                        "by the package; the affected controls have a coverage gap."
                    )

            inspections: list[_RibbonPartInspection] = []
            unrecognized_members: set[str] = set()
            customization_budget = _RibbonCustomizationBudget()
            for member in sorted(candidate_kinds, key=str.casefold):
                kinds = candidate_kinds[member]
                declaration_entries.append(
                    ("package-part", repr((member, tuple(sorted(kinds)))))
                )
                if len(kinds) != 1:
                    warnings.add(
                        "FormulaFence could not identify one RibbonX customization part's "
                        "declared version; the affected controls have a coverage gap."
                    )
                    unrecognized_members.add(member)
                inspection = _ribbon_part_inspection(
                    archive,
                    member,
                    warnings,
                    customization_budget,
                )
                inspections.append(inspection)
                if not inspection.inspected:
                    unrecognized_members.add(member)
                    continue
                declared_kind = next(iter(kinds), None)
                inspected_kind = "2010" if inspection.office_2010 else "2007"
                if declared_kind is not None and declared_kind != inspected_kind:
                    warnings.add(
                        "FormulaFence found a RibbonX customization declaration whose version "
                        "does not match its XML root; the affected controls have a coverage gap."
                    )
                    unrecognized_members.add(member)

            def aggregate_signature(attribute: str) -> str | None:
                material = sorted(
                    (inspection.member, value)
                    for inspection in inspections
                    if (value := getattr(inspection, attribute)) is not None
                )
                return _private_external_data_signature(tuple(material))

            declaration_entries.sort()
            snapshot = RibbonCustomizationSnapshot(
                declared_ribbon_part_count=len(ribbon_relationships),
                ribbon_part_count=len(candidate_kinds),
                office_2010_ribbon_part_count=sum(
                    inspection.office_2010 for inspection in inspections
                ),
                unrecognized_ribbon_part_count=(
                    len(unrecognized_members) + unresolved_declaration_count
                ),
                control_count=sum(inspection.control_count for inspection in inspections),
                callback_attribute_count=sum(
                    inspection.callback_attribute_count for inspection in inspections
                ),
                action_callback_count=sum(
                    inspection.action_callback_count for inspection in inspections
                ),
                image_relationship_count=sum(
                    inspection.image_relationship_count for inspection in inspections
                ),
                external_relationship_count=sum(
                    inspection.external_relationship_count for inspection in inspections
                ),
                declaration_signature=_private_external_data_signature(
                    tuple(declaration_entries)
                ),
                definition_signature=aggregate_signature("definition_signature"),
                relationship_signature=aggregate_signature("relationship_signature"),
            )
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        return _RibbonCustomizationMetadata(
            default,
            (
                "FormulaFence could not inspect RibbonX customization OOXML "
                f"({type(error).__name__}); affected controls were not compared.",
            ),
        )
    return _RibbonCustomizationMetadata(snapshot, tuple(sorted(warnings)))


def _office_web_addin_raw_relationships(
    archive: ZipFile,
    source_member: str,
    warnings: set[str],
    *,
    context: str,
    missing_is_warning: bool = False,
) -> tuple[_OfficeWebAddinRawRelationship, ...]:
    """Read task-pane relationships without opening any package target."""
    relationship_member = _relationship_part_path(source_member)
    try:
        root = _xml_root(archive, relationship_member)
    except KeyError:
        if missing_is_warning:
            warnings.add(
                "FormulaFence could not locate "
                f"{context} relationships while inspecting Office Web Add-ins; "
                "affected controls may be incomplete."
            )
        return ()
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect "
            f"{context} relationships for Office Web Add-ins ({type(error).__name__}); "
            "affected controls were not compared."
        )
        return ()
    if (
        _xml_local_name(root.tag) != "Relationships"
        or _xml_namespace(root.tag) != _PACKAGE_RELATIONSHIP_NS
    ):
        warnings.add(
            "FormulaFence found an unexpected "
            f"{context} relationship root while inspecting Office Web Add-ins; "
            "affected controls were not compared."
        )
        return ()

    relationship_tag = f"{{{_PACKAGE_RELATIONSHIP_NS}}}Relationship"
    relationships: list[_OfficeWebAddinRawRelationship] = []
    for relationship in root.findall(relationship_tag):
        target = relationship.get("Target")
        target_mode = relationship.get("TargetMode", "Internal")
        safe_target = (
            _normalise_part_target(source_member, target)
            if target is not None and target_mode.casefold() == "internal"
            else None
        )
        relationships.append(
            _OfficeWebAddinRawRelationship(
                relationship_id=relationship.get("Id"),
                relationship_type=relationship.get("Type", ""),
                target=target,
                target_mode=target_mode,
                safe_target=safe_target,
            )
        )
    if len(relationships) != len(root):
        warnings.add(
            "FormulaFence found unmodelled relationship XML while inspecting Office Web "
            "Add-ins; affected controls may be incomplete."
        )
    return tuple(relationships)


def _office_web_addin_relationship_signature(
    relationships: tuple[_OfficeWebAddinRawRelationship, ...],
) -> str | None:
    """Fingerprint relationship semantics without writer-chosen identifiers."""
    material = sorted(relationship.semantic_key() for relationship in relationships)
    return _private_external_data_signature(
        tuple(
            (f"relationship:{index}", repr(relationship))
            for index, relationship in enumerate(material)
        )
    )


def _office_web_addin_relationship_semantics(
    relationships: tuple[_OfficeWebAddinRawRelationship, ...],
    warnings: set[str],
    *,
    context: str,
) -> dict[str, tuple[str, str, str]]:
    """Map relationship IDs to stable private semantics for XML canonicalization."""
    relationships_by_id: dict[str, list[_OfficeWebAddinRawRelationship]] = defaultdict(
        list
    )
    for relationship in relationships:
        if relationship.relationship_id:
            relationships_by_id[relationship.relationship_id].append(relationship)
        else:
            warnings.add(
                "FormulaFence found an Office Web Add-in "
                f"{context} relationship without an id; affected controls have a coverage gap."
            )

    semantics_by_id: dict[str, tuple[str, str, str]] = {}
    for relationship_id, matches in relationships_by_id.items():
        semantics = sorted(match.semantic_key() for match in matches)
        if len(matches) > 1:
            warnings.add(
                "FormulaFence found duplicate Office Web Add-in "
                f"{context} relationship ids; affected controls have a coverage gap."
            )
        semantics_by_id[relationship_id] = semantics[0]
    return semantics_by_id


def _office_web_addin_fragment(
    element: ElementTree.Element,
    relationship_semantics: Mapping[str, tuple[str, str, str]],
) -> tuple[object, ...]:
    """Canonicalize Web Add-in XML while resolving relationship-ID churn privately."""
    relationship_attributes = {
        f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id",
        f"{{{_DOCUMENT_RELATIONSHIP_NS}}}embed",
        f"{{{_DOCUMENT_RELATIONSHIP_NS}}}link",
    }
    attributes: list[tuple[str, str]] = []
    for attribute, value in element.attrib.items():
        if attribute in relationship_attributes:
            relationship = relationship_semantics.get(value)
            resolved = (
                ("relationship", relationship)
                if relationship is not None
                else ("missing-relationship", value)
            )
            attributes.append((_xml_display_name(attribute), repr(resolved)))
        else:
            attributes.append((_xml_display_name(attribute), value))
    children = tuple(
        _office_web_addin_fragment(child, relationship_semantics) for child in element
    )
    text = element.text
    if children and text is not None and not text.strip():
        text = None
    return (
        _xml_display_name(element.tag),
        tuple(sorted(attributes)),
        text,
        children,
    )


def _office_web_addin_boolean(
    value: str | None,
    default: bool,
    warnings: set[str],
    *,
    context: str,
    attribute: str,
) -> bool:
    """Parse an OOXML boolean while preserving malformed material privately."""
    if value is None:
        return default
    normalized = value.strip().casefold()
    if normalized in {"1", "true"}:
        return True
    if normalized in {"0", "false"}:
        return False
    warnings.add(
        "FormulaFence found an unrecognized Office Web Add-in "
        f"{context} {attribute} value; affected controls have a coverage gap."
    )
    return default


def _office_web_addin_taskpane_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    budget: _OfficeWebAddinBudget,
) -> _OfficeWebAddinTaskpaneInspection:
    """Fingerprint one bounded task-pane part and its extension bindings privately."""
    if budget.remaining_parts == 0:
        warnings.add(
            "FormulaFence reached its bounded Office Web Add-in part count budget; "
            "the affected controls have a coverage gap."
        )
        return _OfficeWebAddinTaskpaneInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (("part-count-budget-exhausted", member),)
            ),
        )
    budget.remaining_parts -= 1
    try:
        info = archive.getinfo(member)
    except KeyError:
        warnings.add(
            "FormulaFence could not locate an Office Web Add-in task-pane package part; "
            "the affected controls were not compared."
        )
        return _OfficeWebAddinTaskpaneInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (("missing-member", member),)
            ),
        )
    if info.file_size > _WEB_EXTENSION_MAX_PART_BYTES:
        warnings.add(
            "FormulaFence did not fully read an oversized Office Web Add-in package part; "
            "the affected controls have a coverage gap."
        )
        return _OfficeWebAddinTaskpaneInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            ),
        )
    if info.file_size > budget.remaining_bytes:
        warnings.add(
            "FormulaFence reached its bounded Office Web Add-in part read budget; "
            "the affected controls have a coverage gap."
        )
        return _OfficeWebAddinTaskpaneInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (
                    ("read-budget-exhausted", member),
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            ),
        )
    budget.remaining_bytes -= info.file_size
    payload: bytes | None = None
    try:
        payload = archive.read(member)
        root = _xml_root_from_payload(payload)
    except (KeyError, ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect an Office Web Add-in task-pane XML part "
            f"({type(error).__name__}); the affected controls were not compared."
        )
        signature = (
            _private_payload_signature(payload)
            if payload is not None
            else _private_external_data_signature(
                (
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            )
        )
        return _OfficeWebAddinTaskpaneInspection(
            member=member,
            definition_signature=signature,
        )

    if (
        _xml_local_name(root.tag) != "taskpanes"
        or _xml_namespace(root.tag) != _WEB_EXTENSION_TASKPANES_NS
    ):
        warnings.add(
            "FormulaFence found an Office Web Add-in task-pane part with an unexpected "
            "root; the affected controls were not compared."
        )
        return _OfficeWebAddinTaskpaneInspection(
            member=member,
            definition_signature=_private_payload_signature(payload),
        )

    relationships = _office_web_addin_raw_relationships(
        archive,
        member,
        warnings,
        context="task-pane",
    )
    relationship_semantics = _office_web_addin_relationship_semantics(
        relationships,
        warnings,
        context="task-pane",
    )
    taskpane_tag = f"{{{_WEB_EXTENSION_TASKPANES_NS}}}taskpane"
    web_extension_ref_tags = {
        f"{{{_WEB_EXTENSION_TASKPANES_NS}}}webextensionref",
        f"{{{_WEB_EXTENSION_TASKPANES_NS}}}webextension",
    }
    relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
    taskpanes = root.findall(taskpane_tag)
    if len(taskpanes) != len(root):
        warnings.add(
            "FormulaFence found unmodelled XML in an Office Web Add-in task-pane part; "
            "affected controls may be incomplete."
        )

    declared_web_extension_members: set[str] = set()
    referenced_relationship_ids: set[str] = set()
    unresolved_binding_count = 0
    visible_taskpane_count = 0
    locked_taskpane_count = 0
    web_extension_reference_count = 0
    for taskpane in taskpanes:
        visible_taskpane_count += _office_web_addin_boolean(
            taskpane.get("visibility"),
            False,
            warnings,
            context="task-pane",
            attribute="visibility",
        )
        locked_taskpane_count += _office_web_addin_boolean(
            taskpane.get("locked"),
            False,
            warnings,
            context="task-pane",
            attribute="locked",
        )
        references = [
            child for child in taskpane if child.tag in web_extension_ref_tags
        ]
        web_extension_reference_count += len(references)
        if not references:
            warnings.add(
                "FormulaFence found an Office Web Add-in task-pane without a web-extension "
                "reference; affected controls have a coverage gap."
            )
            unresolved_binding_count += 1
        for reference in references:
            relationship_id = reference.get(relationship_id_attribute)
            if not relationship_id:
                warnings.add(
                    "FormulaFence found an Office Web Add-in task-pane reference without a "
                    "relationship id; affected controls have a coverage gap."
                )
                unresolved_binding_count += 1
                continue
            referenced_relationship_ids.add(relationship_id)
            semantic = relationship_semantics.get(relationship_id)
            if semantic is None:
                warnings.add(
                    "FormulaFence found an Office Web Add-in task-pane reference without a "
                    "matching relationship; affected controls have a coverage gap."
                )
                unresolved_binding_count += 1
                continue
            if semantic[0] != _WEB_EXTENSION_RELATIONSHIP:
                warnings.add(
                    "FormulaFence found an Office Web Add-in task-pane reference with an "
                    "unexpected relationship type; affected controls have a coverage gap."
                )
                unresolved_binding_count += 1
                continue
            relationship = next(
                (
                    candidate
                    for candidate in relationships
                    if candidate.relationship_id == relationship_id
                    and candidate.semantic_key() == semantic
                ),
                None,
            )
            if relationship is None or relationship.safe_target is None:
                warnings.add(
                    "FormulaFence found an Office Web Add-in task-pane reference without a "
                    "safe internal extension target; affected controls were not compared."
                )
                unresolved_binding_count += 1
                continue
            declared_web_extension_members.add(relationship.safe_target)

    for relationship in relationships:
        if relationship.relationship_type != _WEB_EXTENSION_RELATIONSHIP:
            continue
        if relationship.safe_target is None:
            warnings.add(
                "FormulaFence found an Office Web Add-in extension relationship without a "
                "safe internal target; affected controls were not compared."
            )
            unresolved_binding_count += 1
            continue
        declared_web_extension_members.add(relationship.safe_target)
        if (
            relationship.relationship_id is None
            or relationship.relationship_id not in referenced_relationship_ids
        ):
            warnings.add(
                "FormulaFence found an Office Web Add-in extension relationship not bound "
                "by a task pane; affected controls have a coverage gap."
            )
            unresolved_binding_count += 1

    try:
        definition_signature = _private_external_data_signature(
            (
                (
                    "taskpanes",
                    repr(_office_web_addin_fragment(root, relationship_semantics)),
                ),
            )
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested Office Web Add-in "
            "task-pane part; the affected controls were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        inspected = False
    else:
        inspected = True
    return _OfficeWebAddinTaskpaneInspection(
        member=member,
        taskpane_count=len(taskpanes),
        visible_taskpane_count=visible_taskpane_count,
        locked_taskpane_count=locked_taskpane_count,
        web_extension_reference_count=web_extension_reference_count,
        related_relationship_count=len(relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in relationships
        ),
        declared_web_extension_members=tuple(
            sorted(declared_web_extension_members, key=str.casefold)
        ),
        unresolved_binding_count=unresolved_binding_count,
        inspected=inspected,
        definition_signature=definition_signature,
        relationship_signature=_office_web_addin_relationship_signature(relationships),
    )


def _office_web_addin_extension_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    budget: _OfficeWebAddinBudget,
) -> _OfficeWebAddinExtensionInspection:
    """Fingerprint one bounded Office Web Add-in definition part privately."""
    if budget.remaining_parts == 0:
        warnings.add(
            "FormulaFence reached its bounded Office Web Add-in part count budget; "
            "the affected controls have a coverage gap."
        )
        return _OfficeWebAddinExtensionInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (("part-count-budget-exhausted", member),)
            ),
        )
    budget.remaining_parts -= 1
    try:
        info = archive.getinfo(member)
    except KeyError:
        warnings.add(
            "FormulaFence could not locate an Office Web Add-in definition package part; "
            "the affected controls were not compared."
        )
        return _OfficeWebAddinExtensionInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (("missing-member", member),)
            ),
        )
    if info.file_size > _WEB_EXTENSION_MAX_PART_BYTES:
        warnings.add(
            "FormulaFence did not fully read an oversized Office Web Add-in package part; "
            "the affected controls have a coverage gap."
        )
        return _OfficeWebAddinExtensionInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            ),
        )
    if info.file_size > budget.remaining_bytes:
        warnings.add(
            "FormulaFence reached its bounded Office Web Add-in part read budget; "
            "the affected controls have a coverage gap."
        )
        return _OfficeWebAddinExtensionInspection(
            member=member,
            definition_signature=_private_external_data_signature(
                (
                    ("read-budget-exhausted", member),
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            ),
        )
    budget.remaining_bytes -= info.file_size
    payload: bytes | None = None
    try:
        payload = archive.read(member)
        root = _xml_root_from_payload(payload)
    except (KeyError, ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect an Office Web Add-in definition XML part "
            f"({type(error).__name__}); the affected controls were not compared."
        )
        signature = (
            _private_payload_signature(payload)
            if payload is not None
            else _private_external_data_signature(
                (
                    ("size", str(info.file_size)),
                    ("compressed-size", str(info.compress_size)),
                    ("crc", str(info.CRC)),
                )
            )
        )
        return _OfficeWebAddinExtensionInspection(
            member=member,
            definition_signature=signature,
        )

    if (
        _xml_local_name(root.tag) != "webextension"
        or _xml_namespace(root.tag) != _WEB_EXTENSION_NS
    ):
        warnings.add(
            "FormulaFence found an Office Web Add-in definition part with an unexpected "
            "root; the affected controls were not compared."
        )
        return _OfficeWebAddinExtensionInspection(
            member=member,
            definition_signature=_private_payload_signature(payload),
        )

    relationships = _office_web_addin_raw_relationships(
        archive,
        member,
        warnings,
        context="definition",
    )
    relationship_semantics = _office_web_addin_relationship_semantics(
        relationships,
        warnings,
        context="definition",
    )
    property_tag = f"{{{_WEB_EXTENSION_NS}}}property"
    reference_tag = f"{{{_WEB_EXTENSION_NS}}}reference"
    alternate_references_tag = f"{{{_WEB_EXTENSION_NS}}}alternateReferences"
    alternate_reference_tags = {
        reference_tag,
        f"{{{_WEB_EXTENSION_NS}}}alternateReference",
    }
    binding_tag = f"{{{_WEB_EXTENSION_NS}}}binding"
    snapshot_tag = f"{{{_WEB_EXTENSION_NS}}}snapshot"
    embedded_relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}embed"

    auto_show_taskpane_count = 0
    for property_element in root.iter(property_tag):
        if property_element.get("name") != "Office.AutoShowTaskpaneWithDocument":
            continue
        auto_show_taskpane_count += _office_web_addin_boolean(
            property_element.get("value"),
            False,
            warnings,
            context="definition property",
            attribute="value",
        )
    snapshot_reference_count = 0
    unresolved_snapshot_reference_count = 0
    for snapshot in root.iter(snapshot_tag):
        relationship_id = snapshot.get(embedded_relationship_attribute)
        if relationship_id is None:
            continue
        snapshot_reference_count += 1
        if relationship_id not in relationship_semantics:
            warnings.add(
                "FormulaFence found an Office Web Add-in snapshot reference without a "
                "matching relationship; affected controls have a coverage gap."
            )
            unresolved_snapshot_reference_count += 1

    try:
        definition_signature = _private_external_data_signature(
            (
                (
                    "webextension",
                    repr(_office_web_addin_fragment(root, relationship_semantics)),
                ),
            )
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested Office Web Add-in "
            "definition part; the affected controls were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        inspected = False
    else:
        inspected = True
    return _OfficeWebAddinExtensionInspection(
        member=member,
        auto_show_taskpane_count=auto_show_taskpane_count,
        store_reference_count=len(root.findall(reference_tag)),
        alternate_reference_count=sum(
            1
            for container in root.iter(alternate_references_tag)
            for child in container
            if child.tag in alternate_reference_tags
        ),
        binding_count=sum(1 for _ in root.iter(binding_tag)),
        snapshot_reference_count=snapshot_reference_count,
        related_relationship_count=len(relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in relationships
        ),
        unresolved_snapshot_reference_count=unresolved_snapshot_reference_count,
        inspected=inspected,
        definition_signature=definition_signature,
        relationship_signature=_office_web_addin_relationship_signature(relationships),
    )


def _office_web_addin_metadata(path: Path) -> _OfficeWebAddinMetadata:
    """Inspect task-pane Office Web Add-ins before the workbook reader omits them."""
    warnings: set[str] = set()
    default = OfficeWebAddinSnapshot()
    try:
        with ZipFile(path) as archive:
            budget = _OfficeWebAddinBudget()
            workbook_relationships = _office_web_addin_raw_relationships(
                archive,
                "xl/workbook.xml",
                warnings,
                context="workbook",
                missing_is_warning=True,
            )
            taskpane_relationships = tuple(
                relationship
                for relationship in workbook_relationships
                if relationship.relationship_type == _WEB_EXTENSION_TASKPANES_RELATIONSHIP
            )
            candidate_taskpane_members: dict[str, set[str]] = defaultdict(set)
            declaration_entries: list[tuple[str, str]] = []
            declared_taskpane_members: set[str] = set()
            unresolved_declaration_count = 0
            relationship_ids: set[str] = set()
            for relationship in taskpane_relationships:
                declaration_entries.append(
                    ("workbook-relationship", repr(relationship.semantic_key()))
                )
                if relationship.relationship_id is None:
                    warnings.add(
                        "FormulaFence found an Office Web Add-in workbook relationship without "
                        "an id; affected controls have a coverage gap."
                    )
                elif relationship.relationship_id in relationship_ids:
                    warnings.add(
                        "FormulaFence found duplicate Office Web Add-in workbook relationship "
                        "ids; affected controls have a coverage gap."
                    )
                relationship_ids.add(relationship.relationship_id or "")
                if relationship.safe_target is None:
                    warnings.add(
                        "FormulaFence found an Office Web Add-in workbook relationship without "
                        "a safe internal task-pane target; affected controls were not compared."
                    )
                    unresolved_declaration_count += 1
                    continue
                candidate_taskpane_members[relationship.safe_target].add("workbook")
                declared_taskpane_members.add(relationship.safe_target)

            discovered_taskpane_members = {
                entry.filename
                for entry in archive.infolist()
                if _WEB_EXTENSION_TASKPANES_PART_PATTERN.fullmatch(entry.filename)
            }
            for member in discovered_taskpane_members:
                candidate_taskpane_members.setdefault(member, set())
                if member not in declared_taskpane_members:
                    warnings.add(
                        "FormulaFence found an Office Web Add-in task-pane package part not "
                        "declared by the workbook; affected controls have a coverage gap."
                    )

            taskpane_inspections: list[_OfficeWebAddinTaskpaneInspection] = []
            unrecognized_members: set[str] = set()
            unresolved_binding_count = unresolved_declaration_count
            candidate_web_extension_members: dict[str, set[str]] = defaultdict(set)
            declared_web_extension_members: set[str] = set()
            for member in sorted(candidate_taskpane_members, key=str.casefold):
                sources = candidate_taskpane_members[member]
                declaration_entries.append(
                    ("taskpane-part", repr((member, tuple(sorted(sources)))))
                )
                inspection = _office_web_addin_taskpane_inspection(
                    archive,
                    member,
                    warnings,
                    budget,
                )
                taskpane_inspections.append(inspection)
                if not inspection.inspected:
                    unrecognized_members.add(member)
                    continue
                unresolved_binding_count += inspection.unresolved_binding_count
                for web_extension_member in inspection.declared_web_extension_members:
                    candidate_web_extension_members[web_extension_member].add(member)
                    declared_web_extension_members.add(web_extension_member)

            discovered_web_extension_members = {
                entry.filename
                for entry in archive.infolist()
                if _WEB_EXTENSION_PART_PATTERN.fullmatch(entry.filename)
            }
            for member in discovered_web_extension_members:
                candidate_web_extension_members.setdefault(member, set())
                if member not in declared_web_extension_members:
                    warnings.add(
                        "FormulaFence found an Office Web Add-in definition package part not "
                        "declared by a task pane; affected controls have a coverage gap."
                    )

            extension_inspections: list[_OfficeWebAddinExtensionInspection] = []
            for member in sorted(candidate_web_extension_members, key=str.casefold):
                sources = candidate_web_extension_members[member]
                declaration_entries.append(
                    ("web-extension-part", repr((member, tuple(sorted(sources)))))
                )
                inspection = _office_web_addin_extension_inspection(
                    archive,
                    member,
                    warnings,
                    budget,
                )
                extension_inspections.append(inspection)
                if not inspection.inspected:
                    unrecognized_members.add(member)
                    continue
                unresolved_binding_count += inspection.unresolved_snapshot_reference_count

            def aggregate_signature(
                inspections: list[object], attribute: str
            ) -> str | None:
                material = sorted(
                    (inspection.member, value)
                    for inspection in inspections
                    if (value := getattr(inspection, attribute)) is not None
                )
                return _private_external_data_signature(tuple(material))

            declaration_entries.sort()
            all_inspections: list[object] = [
                *taskpane_inspections,
                *extension_inspections,
            ]
            snapshot = OfficeWebAddinSnapshot(
                declared_taskpane_part_count=len(taskpane_relationships),
                taskpane_part_count=len(candidate_taskpane_members),
                web_extension_part_count=len(candidate_web_extension_members),
                unrecognized_part_count=(
                    len(unrecognized_members) + unresolved_binding_count
                ),
                taskpane_count=sum(
                    inspection.taskpane_count for inspection in taskpane_inspections
                ),
                visible_taskpane_count=sum(
                    inspection.visible_taskpane_count for inspection in taskpane_inspections
                ),
                locked_taskpane_count=sum(
                    inspection.locked_taskpane_count for inspection in taskpane_inspections
                ),
                web_extension_reference_count=sum(
                    inspection.web_extension_reference_count
                    for inspection in taskpane_inspections
                ),
                auto_show_taskpane_count=sum(
                    inspection.auto_show_taskpane_count
                    for inspection in extension_inspections
                ),
                store_reference_count=sum(
                    inspection.store_reference_count for inspection in extension_inspections
                ),
                alternate_reference_count=sum(
                    inspection.alternate_reference_count
                    for inspection in extension_inspections
                ),
                binding_count=sum(
                    inspection.binding_count for inspection in extension_inspections
                ),
                snapshot_reference_count=sum(
                    inspection.snapshot_reference_count
                    for inspection in extension_inspections
                ),
                related_relationship_count=sum(
                    inspection.related_relationship_count for inspection in all_inspections
                ),
                external_relationship_count=sum(
                    inspection.external_relationship_count for inspection in all_inspections
                ),
                declaration_signature=_private_external_data_signature(
                    tuple(declaration_entries)
                ),
                taskpane_signature=aggregate_signature(
                    taskpane_inspections, "definition_signature"
                ),
                web_extension_signature=aggregate_signature(
                    extension_inspections, "definition_signature"
                ),
                relationship_signature=aggregate_signature(
                    all_inspections, "relationship_signature"
                ),
            )
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        return _OfficeWebAddinMetadata(
            default,
            (
                "FormulaFence could not inspect Office Web Add-in task-pane OOXML "
                f"({type(error).__name__}); affected controls were not compared.",
            ),
        )
    return _OfficeWebAddinMetadata(snapshot, tuple(sorted(warnings)))


def _chart_raw_relationships(
    archive: ZipFile,
    source_member: str,
    warnings: set[str],
    *,
    context: str,
) -> tuple[_ChartRawRelationship, ...]:
    """Read chart-package relationships without opening their targets."""
    relationship_member = _relationship_part_path(source_member)
    try:
        root = _xml_root(archive, relationship_member)
    except KeyError:
        return ()
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect "
            f"{context} relationships for chart inspection "
            f"({type(error).__name__}); affected charts were not compared."
        )
        return ()
    if (
        _xml_local_name(root.tag) != "Relationships"
        or _xml_namespace(root.tag) != _PACKAGE_RELATIONSHIP_NS
    ):
        warnings.add(
            "FormulaFence found an unexpected "
            f"{context} relationship root while inspecting charts; affected charts "
            "were not compared."
        )
        return ()

    relationship_tag = f"{{{_PACKAGE_RELATIONSHIP_NS}}}Relationship"
    relationships: list[_ChartRawRelationship] = []
    for relationship in root.findall(relationship_tag):
        target = relationship.get("Target")
        target_mode = relationship.get("TargetMode", "Internal")
        safe_target = (
            _normalise_part_target(source_member, target)
            if target is not None and target_mode.casefold() == "internal"
            else None
        )
        relationships.append(
            _ChartRawRelationship(
                relationship_id=relationship.get("Id"),
                relationship_type=relationship.get("Type", ""),
                target=target,
                target_mode=target_mode,
                safe_target=safe_target,
            )
        )
    if len(relationships) != len(root):
        warnings.add(
            "FormulaFence found unmodelled relationship XML while inspecting charts; "
            "affected charts may be incomplete."
        )
    return tuple(relationships)


def _chart_relationship_signature(
    relationships: tuple[_ChartRawRelationship, ...],
) -> str | None:
    """Fingerprint chart relationship semantics without identifier churn."""
    material = sorted(relationship.semantic_key() for relationship in relationships)
    return _private_external_data_signature(
        tuple(
            (f"relationship:{index}", repr(relationship))
            for index, relationship in enumerate(material)
        )
    )


def _chart_relationship_semantics(
    relationships: tuple[_ChartRawRelationship, ...],
    warnings: set[str],
    *,
    context: str,
) -> dict[str, tuple[str, str, str]]:
    """Resolve private chart relationship identifiers into stable semantics."""
    relationships_by_id: dict[str, list[_ChartRawRelationship]] = defaultdict(list)
    for relationship in relationships:
        if relationship.relationship_id:
            relationships_by_id[relationship.relationship_id].append(relationship)
        else:
            warnings.add(
                "FormulaFence found a chart "
                f"{context} relationship without an id; affected charts have a coverage gap."
            )
    semantics: dict[str, tuple[str, str, str]] = {}
    for relationship_id, matches in relationships_by_id.items():
        values = sorted(match.semantic_key() for match in matches)
        if len(values) > 1:
            warnings.add(
                "FormulaFence found duplicate chart "
                f"{context} relationship ids; affected charts have a coverage gap."
            )
        semantics[relationship_id] = values[0]
    return semantics


def _chart_xml_fragment(
    element: ElementTree.Element,
    relationship_semantics: Mapping[str, tuple[str, str, str]],
    *,
    omit_cached_data: bool = False,
) -> tuple[object, ...]:
    """Canonicalize private chart XML while resolving relationship identifiers."""
    attributes: list[tuple[str, str]] = []
    for attribute, value in element.attrib.items():
        if attribute in _CHART_RELATIONSHIP_ATTRIBUTES:
            relationship = relationship_semantics.get(value)
            resolved = (
                ("relationship", relationship)
                if relationship is not None
                else ("missing-relationship", value)
            )
            attributes.append((_xml_display_name(attribute), repr(resolved)))
        else:
            attributes.append((_xml_display_name(attribute), value))
    children = tuple(
        _chart_xml_fragment(
            child,
            relationship_semantics,
            omit_cached_data=omit_cached_data,
        )
        for child in element
        if not (
            omit_cached_data
            and _xml_namespace(child.tag) == _DRAWINGML_CHART_NS
            and _xml_local_name(child.tag) in _CHART_CACHE_ELEMENT_NAMES
        )
    )
    text = element.text
    if children and text is not None and not text.strip():
        text = None
    return (
        _xml_display_name(element.tag),
        tuple(sorted(attributes)),
        text,
        children,
    )


def _chart_xml_payload(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    budget: _ChartXmlBudget,
) -> tuple[bytes | None, str | None]:
    """Read one bounded chart XML part without following relationship targets."""
    if budget.remaining_parts == 0:
        warnings.add(
            "FormulaFence reached its bounded chart XML part count budget; affected "
            "charts have a coverage gap."
        )
        return None, _private_external_data_signature(
            (("part-count-budget-exhausted", member),)
        )
    budget.remaining_parts -= 1
    try:
        info = archive.getinfo(member)
    except KeyError:
        warnings.add(
            "FormulaFence could not locate a chart XML package part; affected charts "
            "were not compared."
        )
        return None, _private_external_data_signature((("missing-member", member),))
    metadata = repr((member, info.file_size, info.compress_size, info.CRC))
    if info.file_size > _CHART_MAX_XML_PART_BYTES:
        warnings.add(
            "FormulaFence did not fully read an oversized chart XML part; affected "
            "charts have a coverage gap."
        )
        return None, _private_external_data_signature((("oversized-part", metadata),))
    if info.file_size > budget.remaining_bytes:
        warnings.add(
            "FormulaFence reached its bounded chart XML read budget; affected charts "
            "have a coverage gap."
        )
        return None, _private_external_data_signature(
            (("read-budget-exhausted", metadata),)
        )
    budget.remaining_bytes -= info.file_size
    try:
        return archive.read(member), None
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not read a chart XML package part "
            f"({type(error).__name__}); affected charts were not compared."
        )
        return None, _private_external_data_signature((("unreadable-part", metadata),))


def _chart_related_part_payloads(
    archive: ZipFile,
    members: set[str],
    unresolved_entries: list[tuple[str, str]],
    warnings: set[str],
    budget: _ChartRelatedPartBudget,
) -> _ChartRelatedPartPayloadInspection:
    """Fingerprint bounded direct chart-presentation payloads without opening them."""
    entries = list(unresolved_entries)
    fingerprinted_part_count = 0
    uninspected_part_count = len(unresolved_entries)
    for member in sorted(members, key=str.casefold):
        if budget.remaining_parts == 0:
            warnings.add(
                "FormulaFence reached its bounded chart related-part count budget; "
                "affected charts have a coverage gap."
            )
            entries.append(("part-count-budget-exhausted", member))
            uninspected_part_count += 1
            continue
        budget.remaining_parts -= 1
        try:
            info = archive.getinfo(member)
        except KeyError:
            warnings.add(
                "FormulaFence could not locate a chart related part; affected charts "
                "were not compared."
            )
            entries.append(("missing-part", member))
            uninspected_part_count += 1
            continue
        metadata = repr((member, info.file_size, info.compress_size, info.CRC))
        if info.file_size > _CHART_RELATED_PART_MAX_BYTES:
            warnings.add(
                "FormulaFence did not fully read an oversized chart related part; "
                "affected charts have a coverage gap."
            )
            entries.append(("oversized-part", metadata))
            uninspected_part_count += 1
            continue
        if info.file_size > budget.remaining_bytes:
            warnings.add(
                "FormulaFence reached its bounded chart related-part read budget; "
                "affected charts have a coverage gap."
            )
            entries.append(("read-budget-exhausted", metadata))
            uninspected_part_count += 1
            continue
        budget.remaining_bytes -= info.file_size
        digest = hashlib.sha256()
        bytes_read = 0
        try:
            with archive.open(info) as payload:
                while chunk := payload.read(_CHART_RELATED_PART_HASH_CHUNK_BYTES):
                    bytes_read += len(chunk)
                    if bytes_read > info.file_size:
                        raise ValueError("payload exceeded its declared size")
                    digest.update(chunk)
            if bytes_read != info.file_size:
                raise ValueError("payload did not match its declared size")
        except (BadZipFile, OSError, RuntimeError, ValueError) as error:
            warnings.add(
                "FormulaFence could not fingerprint a chart related part "
                f"({type(error).__name__}); affected charts were not compared."
            )
            entries.append(("unreadable-part", metadata))
            uninspected_part_count += 1
            continue
        entries.append(("payload", repr((member, digest.hexdigest()))))
        fingerprinted_part_count += 1

    entries.sort()
    return _ChartRelatedPartPayloadInspection(
        internal_part_count=len(members) + len(unresolved_entries),
        fingerprinted_part_count=fingerprinted_part_count,
        uninspected_part_count=uninspected_part_count,
        payload_signature=_private_external_data_signature(tuple(entries)),
    )


def _chart_drawing_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    xml_budget: _ChartXmlBudget,
) -> _ChartDrawingInspection:
    """Discover DrawingML chart bindings without interpreting worksheet cells."""
    relationships = _chart_raw_relationships(
        archive,
        member,
        warnings,
        context="worksheet or chartsheet drawing",
    )
    relationship_semantics = _chart_relationship_semantics(
        relationships,
        warnings,
        context="worksheet or chartsheet drawing",
    )
    relationship_by_id: dict[str, _ChartRawRelationship] = {}
    for relationship in relationships:
        if relationship.relationship_id and relationship.relationship_id not in relationship_by_id:
            relationship_by_id[relationship.relationship_id] = relationship

    payload, fallback_signature = _chart_xml_payload(
        archive,
        member,
        warnings,
        xml_budget,
    )
    if payload is None:
        return _ChartDrawingInspection(
            member=member,
            unrecognized_count=1,
            declaration_signature=fallback_signature,
            relationship_signature=_chart_relationship_signature(relationships),
        )
    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a worksheet or chartsheet drawing XML part "
            f"({type(error).__name__}); affected charts were not compared."
        )
        return _ChartDrawingInspection(
            member=member,
            unrecognized_count=1,
            declaration_signature=_private_payload_signature(payload),
            relationship_signature=_chart_relationship_signature(relationships),
        )
    if (
        _xml_local_name(root.tag) != "wsDr"
        or _xml_namespace(root.tag) != _DRAWINGML_SPREADSHEET_NS
    ):
        warnings.add(
            "FormulaFence found a worksheet or chartsheet drawing part with an unexpected "
            "root; affected charts were not compared."
        )
        return _ChartDrawingInspection(
            member=member,
            unrecognized_count=1,
            declaration_signature=_private_payload_signature(payload),
            relationship_signature=_chart_relationship_signature(relationships),
        )

    chart_tag = f"{{{_DRAWINGML_CHART_NS}}}chart"
    relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
    chart_nodes = list(root.iter(chart_tag))
    selected_relationships: list[_ChartRawRelationship] = []
    referenced_relationship_ids: set[str] = set()
    chart_members: set[str] = set()
    unrecognized_count = 0

    for chart in chart_nodes:
        relationship_id = chart.get(relationship_attribute)
        if relationship_id is None:
            warnings.add(
                "FormulaFence found a DrawingML chart reference without a relationship id; "
                "affected charts have a coverage gap."
            )
            unrecognized_count += 1
            continue
        referenced_relationship_ids.add(relationship_id)
        relationship = relationship_by_id.get(relationship_id)
        if relationship is None:
            warnings.add(
                "FormulaFence found a DrawingML chart reference without a matching "
                "relationship; affected charts have a coverage gap."
            )
            unrecognized_count += 1
            continue
        selected_relationships.append(relationship)
        if relationship.relationship_type != _CHART_RELATIONSHIP:
            warnings.add(
                "FormulaFence found a DrawingML chart reference with an unexpected "
                "relationship type; affected charts have a coverage gap."
            )
            unrecognized_count += 1
            continue
        if relationship.target_mode.casefold() != "internal" or relationship.safe_target is None:
            warnings.add(
                "FormulaFence found a DrawingML chart reference without a safe internal "
                "target; affected charts were not compared."
            )
            unrecognized_count += 1
            continue
        chart_members.add(relationship.safe_target)

    for relationship in relationships:
        if relationship.relationship_type != _CHART_RELATIONSHIP:
            continue
        if (
            relationship.relationship_id is None
            or relationship.relationship_id not in referenced_relationship_ids
        ):
            warnings.add(
                "FormulaFence found a chart relationship not bound by an inspected drawing; "
                "affected charts have a coverage gap."
            )
            unrecognized_count += 1

    if not chart_nodes and not any(
        relationship.relationship_type == _CHART_RELATIONSHIP for relationship in relationships
    ):
        return _ChartDrawingInspection(member=member, inspected=True)
    try:
        declaration_signature = _private_external_data_signature(
            tuple(
                (
                    f"chart-reference:{index}",
                    repr(_chart_xml_fragment(chart, relationship_semantics)),
                )
                for index, chart in enumerate(chart_nodes)
            )
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested DrawingML chart "
            "binding; affected charts were not compared."
        )
        declaration_signature = _private_payload_signature(payload)
        inspected = False
        unrecognized_count += 1
    else:
        inspected = True
    return _ChartDrawingInspection(
        member=member,
        present=bool(chart_nodes or any(
            relationship.relationship_type == _CHART_RELATIONSHIP
            for relationship in relationships
        )),
        chart_reference_count=len(chart_nodes),
        related_relationship_count=len(selected_relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in selected_relationships
        ),
        chart_members=tuple(sorted(chart_members, key=str.casefold)),
        unrecognized_count=unrecognized_count,
        inspected=inspected,
        declaration_signature=declaration_signature,
        relationship_signature=_chart_relationship_signature(tuple(selected_relationships)),
    )


def _chart_part_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    xml_budget: _ChartXmlBudget,
) -> _ChartPartInspection:
    """Inspect one standard chart part without evaluating series formulas."""
    relationships = _chart_raw_relationships(archive, member, warnings, context="chart")
    relationship_semantics = _chart_relationship_semantics(
        relationships,
        warnings,
        context="chart",
    )
    relationship_by_id: dict[str, _ChartRawRelationship] = {}
    for relationship in relationships:
        if relationship.relationship_id and relationship.relationship_id not in relationship_by_id:
            relationship_by_id[relationship.relationship_id] = relationship

    payload, fallback_signature = _chart_xml_payload(
        archive,
        member,
        warnings,
        xml_budget,
    )
    if payload is None:
        return _ChartPartInspection(
            member=member,
            unrecognized_count=1,
            definition_signature=fallback_signature,
            relationship_signature=_chart_relationship_signature(relationships),
        )
    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a chart XML part "
            f"({type(error).__name__}); affected charts were not compared."
        )
        return _ChartPartInspection(
            member=member,
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_chart_relationship_signature(relationships),
        )
    if (
        _xml_local_name(root.tag) != "chartSpace"
        or _xml_namespace(root.tag) != _DRAWINGML_CHART_NS
    ):
        warnings.add(
            "FormulaFence found a chart part with an unexpected root; affected charts "
            "were not compared."
        )
        return _ChartPartInspection(
            member=member,
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_chart_relationship_signature(relationships),
        )

    relationship_ids = {
        value
        for element in root.iter()
        for attribute in _CHART_RELATIONSHIP_ATTRIBUTES
        if (value := element.get(attribute)) is not None
    }
    user_shapes_tag = f"{{{_DRAWINGML_CHART_NS}}}userShapes"
    relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
    user_shape_relationship_ids = {
        relationship_id
        for element in root.iter(user_shapes_tag)
        if (relationship_id := element.get(relationship_attribute)) is not None
    }
    selected_relationships: list[_ChartRawRelationship] = []
    user_shape_members: set[str] = set()
    payload_members: set[str] = set()
    unresolved_payload_entries: list[tuple[str, str]] = []
    unrecognized_count = 0

    for relationship_id in relationship_ids:
        relationship = relationship_by_id.get(relationship_id)
        if relationship is None:
            warnings.add(
                "FormulaFence found a chart relationship reference without a matching "
                "relationship; affected charts have a coverage gap."
            )
            unrecognized_count += 1
            continue
        selected_relationships.append(relationship)
        is_user_shape = relationship_id in user_shape_relationship_ids
        if is_user_shape and relationship.relationship_type != _CHART_USER_SHAPES_RELATIONSHIP:
            warnings.add(
                "FormulaFence found a chart overlay reference with an unexpected "
                "relationship type; affected charts have a coverage gap."
            )
            unrecognized_count += 1
            continue
        if relationship.target_mode.casefold() != "internal":
            if is_user_shape:
                warnings.add(
                    "FormulaFence found a chart overlay reference without an internal target; "
                    "affected charts were not compared."
                )
                unrecognized_count += 1
            continue
        if relationship.safe_target is None:
            warnings.add(
                "FormulaFence found a chart relationship without a safe internal target; "
                "affected charts were not compared."
            )
            unresolved_payload_entries.append(
                ("unsafe-related-target", repr(relationship.semantic_key()))
            )
            unrecognized_count += 1
            continue
        if is_user_shape:
            user_shape_members.add(relationship.safe_target)
        else:
            payload_members.add(relationship.safe_target)

    for relationship in relationships:
        if (
            relationship.relationship_id is None
            or relationship.relationship_id not in relationship_ids
        ):
            warnings.add(
                "FormulaFence found a chart relationship not bound by inspected chart XML; "
                "affected charts have a coverage gap."
            )
            unrecognized_count += 1

    cache_nodes = [
        element
        for element in root.iter()
        if _xml_namespace(element.tag) == _DRAWINGML_CHART_NS
        and _xml_local_name(element.tag) in _CHART_CACHE_ELEMENT_NAMES
    ]
    literal_containers = [
        element
        for element in root.iter()
        if _xml_namespace(element.tag) == _DRAWINGML_CHART_NS
        and _xml_local_name(element.tag) in _CHART_LITERAL_ELEMENT_NAMES
    ]
    point_tag = f"{{{_DRAWINGML_CHART_NS}}}pt"
    try:
        definition_signature = _private_external_data_signature(
            (
                (
                    "chart",
                    repr(
                        _chart_xml_fragment(
                            root,
                            relationship_semantics,
                            omit_cached_data=True,
                        )
                    ),
                ),
            )
        )
        cached_data_signature = _private_external_data_signature(
            tuple(
                (f"cache:{index}", repr(_chart_xml_fragment(cache, relationship_semantics)))
                for index, cache in enumerate(cache_nodes)
            )
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested chart part; "
            "affected charts were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        cached_data_signature = None
        inspected = False
        unrecognized_count += 1
    else:
        inspected = True
    return _ChartPartInspection(
        member=member,
        chart_type_count=sum(
            _xml_namespace(element.tag) == _DRAWINGML_CHART_NS
            and _xml_local_name(element.tag) in _CHART_TYPE_ELEMENT_NAMES
            for element in root.iter()
        ),
        series_count=sum(
            element.tag == f"{{{_DRAWINGML_CHART_NS}}}ser" for element in root.iter()
        ),
        title_count=sum(
            element.tag == f"{{{_DRAWINGML_CHART_NS}}}title" for element in root.iter()
        ),
        data_reference_count=sum(
            element.tag == f"{{{_DRAWINGML_CHART_NS}}}f" for element in root.iter()
        ),
        numeric_data_reference_count=sum(
            element.tag == f"{{{_DRAWINGML_CHART_NS}}}numRef" for element in root.iter()
        ),
        string_data_reference_count=sum(
            element.tag
            in {
                f"{{{_DRAWINGML_CHART_NS}}}strRef",
                f"{{{_DRAWINGML_CHART_NS}}}multiLvlStrRef",
            }
            for element in root.iter()
        ),
        literal_data_point_count=sum(
            sum(point.tag == point_tag for point in container.iter())
            for container in literal_containers
        ),
        cached_data_point_count=sum(
            sum(point.tag == point_tag for point in cache.iter()) for cache in cache_nodes
        ),
        pivot_source_count=sum(
            element.tag == f"{{{_DRAWINGML_CHART_NS}}}pivotSource" for element in root.iter()
        ),
        external_data_reference_count=sum(
            element.tag == f"{{{_DRAWINGML_CHART_NS}}}externalData" for element in root.iter()
        ),
        user_shape_reference_count=len(user_shape_relationship_ids),
        related_relationship_count=len(selected_relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in selected_relationships
        ),
        user_shape_members=tuple(sorted(user_shape_members, key=str.casefold)),
        payload_members=tuple(sorted(payload_members, key=str.casefold)),
        unresolved_payload_entries=tuple(unresolved_payload_entries),
        unrecognized_count=unrecognized_count,
        inspected=inspected,
        definition_signature=definition_signature,
        cached_data_signature=cached_data_signature,
        relationship_signature=_chart_relationship_signature(tuple(selected_relationships)),
    )


def _chart_user_shape_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    xml_budget: _ChartXmlBudget,
) -> _ChartUserShapeInspection:
    """Inspect one chart overlay part without rendering shapes or opening media."""
    relationships = _chart_raw_relationships(
        archive,
        member,
        warnings,
        context="chart overlay",
    )
    relationship_semantics = _chart_relationship_semantics(
        relationships,
        warnings,
        context="chart overlay",
    )
    relationship_by_id: dict[str, _ChartRawRelationship] = {}
    for relationship in relationships:
        if relationship.relationship_id and relationship.relationship_id not in relationship_by_id:
            relationship_by_id[relationship.relationship_id] = relationship

    payload, fallback_signature = _chart_xml_payload(
        archive,
        member,
        warnings,
        xml_budget,
    )
    if payload is None:
        return _ChartUserShapeInspection(
            member=member,
            unrecognized_count=1,
            definition_signature=fallback_signature,
            relationship_signature=_chart_relationship_signature(relationships),
        )
    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a chart overlay XML part "
            f"({type(error).__name__}); affected charts were not compared."
        )
        return _ChartUserShapeInspection(
            member=member,
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_chart_relationship_signature(relationships),
        )
    if (
        _xml_local_name(root.tag) != "userShapes"
        or _xml_namespace(root.tag) != _DRAWINGML_CHART_DRAWING_NS
    ):
        warnings.add(
            "FormulaFence found a chart overlay part with an unexpected root; affected "
            "charts were not compared."
        )
        return _ChartUserShapeInspection(
            member=member,
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_chart_relationship_signature(relationships),
        )

    relationship_ids = {
        value
        for element in root.iter()
        for attribute in _CHART_RELATIONSHIP_ATTRIBUTES
        if (value := element.get(attribute)) is not None
    }
    selected_relationships: list[_ChartRawRelationship] = []
    payload_members: set[str] = set()
    unresolved_payload_entries: list[tuple[str, str]] = []
    unrecognized_count = 0
    for relationship_id in relationship_ids:
        relationship = relationship_by_id.get(relationship_id)
        if relationship is None:
            warnings.add(
                "FormulaFence found a chart overlay relationship reference without a "
                "matching relationship; affected charts have a coverage gap."
            )
            unrecognized_count += 1
            continue
        selected_relationships.append(relationship)
        if relationship.target_mode.casefold() != "internal":
            continue
        if relationship.safe_target is None:
            warnings.add(
                "FormulaFence found a chart overlay relationship without a safe internal "
                "target; affected charts were not compared."
            )
            unresolved_payload_entries.append(
                ("unsafe-overlay-target", repr(relationship.semantic_key()))
            )
            unrecognized_count += 1
            continue
        payload_members.add(relationship.safe_target)

    for relationship in relationships:
        if (
            relationship.relationship_id is None
            or relationship.relationship_id not in relationship_ids
        ):
            warnings.add(
                "FormulaFence found a chart overlay relationship not bound by inspected "
                "overlay XML; affected charts have a coverage gap."
            )
            unrecognized_count += 1
    try:
        definition_signature = _private_external_data_signature(
            (("chart-overlay", repr(_chart_xml_fragment(root, relationship_semantics))),)
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested chart overlay; "
            "affected charts were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        inspected = False
        unrecognized_count += 1
    else:
        inspected = True
    shape_names = {"sp", "pic", "cxnSp", "graphicFrame", "grpSp"}
    return _ChartUserShapeInspection(
        member=member,
        shape_count=sum(
            _xml_namespace(element.tag) == _DRAWINGML_CHART_DRAWING_NS
            and _xml_local_name(element.tag) in shape_names
            for element in root.iter()
        ),
        related_relationship_count=len(selected_relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in selected_relationships
        ),
        payload_members=tuple(sorted(payload_members, key=str.casefold)),
        unresolved_payload_entries=tuple(unresolved_payload_entries),
        unrecognized_count=unrecognized_count,
        inspected=inspected,
        definition_signature=definition_signature,
        relationship_signature=_chart_relationship_signature(tuple(selected_relationships)),
    )


def _chart_definition_metadata(path: Path) -> _ChartDefinitionMetadata:
    """Inspect chart presentation parts before the workbook reader can omit them.

    The scan is package-only: it does not calculate a series formula, render a
    chart, open related media or embedded data, or follow external targets.
    """
    warnings: set[str] = set()
    default = ChartDefinitionSnapshot()
    try:
        with ZipFile(path) as archive:
            try:
                sheet_parts = _sheet_xml_parts(archive)
            except (KeyError, ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
                return _ChartDefinitionMetadata(
                    default,
                    (
                        "FormulaFence could not map worksheet OOXML for chart inspection "
                        f"({type(error).__name__}); affected charts were not compared.",
                    ),
                )

            xml_budget = _ChartXmlBudget()
            related_part_budget = _ChartRelatedPartBudget()
            drawing_sources: dict[str, set[tuple[str, str, str]]] = defaultdict(set)
            unresolved_drawing_entries: list[tuple[str, str]] = []
            unrecognized_declaration_count = 0
            for sheet, (member, sheet_kind) in sorted(
                sheet_parts.items(),
                key=lambda item: item[0].casefold(),
            ):
                if sheet_kind not in {"worksheet", "chartsheet"}:
                    continue
                relationships = _chart_raw_relationships(
                    archive,
                    member,
                    warnings,
                    context="worksheet or chartsheet",
                )
                for relationship in relationships:
                    if relationship.relationship_type != _WORKSHEET_DRAWING_RELATIONSHIP:
                        continue
                    if (
                        relationship.target_mode.casefold() != "internal"
                        or relationship.safe_target is None
                    ):
                        warnings.add(
                            "FormulaFence found a worksheet or chartsheet drawing relationship "
                            "without a safe internal target; affected charts were not compared."
                        )
                        unresolved_drawing_entries.append(
                            (member, repr(relationship.semantic_key()))
                        )
                        unrecognized_declaration_count += 1
                        continue
                    drawing_sources[relationship.safe_target].add(
                        (sheet, sheet_kind, member)
                    )

            drawing_inspections: list[_ChartDrawingInspection] = []
            chart_sources: dict[str, set[str]] = defaultdict(set)
            declaration_entries: list[tuple[str, str]] = []
            for drawing_member in sorted(drawing_sources, key=str.casefold):
                sources = tuple(
                    sorted(drawing_sources[drawing_member], key=lambda item: item[0].casefold())
                )
                inspection = _chart_drawing_inspection(
                    archive,
                    drawing_member,
                    warnings,
                    xml_budget,
                )
                drawing_inspections.append(inspection)
                if not inspection.present:
                    continue
                declaration_entries.append(
                    ("chart-drawing-part", repr((drawing_member, sources)))
                )
                for chart_member in inspection.chart_members:
                    chart_sources[chart_member].add(drawing_member)

            orphan_chart_part_count = 0
            discovered_chart_members = {
                entry.filename
                for entry in archive.infolist()
                if _CHART_PART_PATTERN.fullmatch(entry.filename)
            }
            for chart_member in discovered_chart_members:
                chart_sources.setdefault(chart_member, set())
                if chart_member not in {
                    member
                    for inspection in drawing_inspections
                    for member in inspection.chart_members
                }:
                    warnings.add(
                        "FormulaFence found a chart package part not declared by an inspected "
                        "worksheet or chartsheet drawing; affected charts have a coverage gap."
                    )
                    orphan_chart_part_count += 1

            chart_inspections: list[_ChartPartInspection] = []
            user_shape_sources: dict[str, set[str]] = defaultdict(set)
            payload_members: set[str] = set()
            unresolved_payload_entries: list[tuple[str, str]] = []
            for chart_member in sorted(chart_sources, key=str.casefold):
                sources = tuple(sorted(chart_sources[chart_member], key=str.casefold))
                declaration_entries.append(("chart-part", repr((chart_member, sources))))
                inspection = _chart_part_inspection(
                    archive,
                    chart_member,
                    warnings,
                    xml_budget,
                )
                chart_inspections.append(inspection)
                for user_shape_member in inspection.user_shape_members:
                    user_shape_sources[user_shape_member].add(chart_member)
                payload_members.update(inspection.payload_members)
                unresolved_payload_entries.extend(inspection.unresolved_payload_entries)

            user_shape_inspections: list[_ChartUserShapeInspection] = []
            for user_shape_member in sorted(user_shape_sources, key=str.casefold):
                sources = tuple(sorted(user_shape_sources[user_shape_member], key=str.casefold))
                declaration_entries.append(
                    ("chart-overlay-part", repr((user_shape_member, sources)))
                )
                inspection = _chart_user_shape_inspection(
                    archive,
                    user_shape_member,
                    warnings,
                    xml_budget,
                )
                user_shape_inspections.append(inspection)
                payload_members.update(inspection.payload_members)
                unresolved_payload_entries.extend(inspection.unresolved_payload_entries)

            payload_inspection = _chart_related_part_payloads(
                archive,
                payload_members,
                unresolved_payload_entries,
                warnings,
                related_part_budget,
            )

            def aggregate_signature(
                inspections: list[object], attribute: str
            ) -> str | None:
                material = sorted(
                    (inspection.member, value)
                    for inspection in inspections
                    if (value := getattr(inspection, attribute)) is not None
                )
                return _private_external_data_signature(tuple(material))

            chart_host_sheets = {
                sheet
                for drawing_member, sources in drawing_sources.items()
                if any(
                    inspection.member == drawing_member and inspection.present
                    for inspection in drawing_inspections
                )
                for sheet, _sheet_kind, _member in sources
            }
            declaration_entries.extend(
                ("unresolved-sheet-drawing-relationship", entry)
                for entry in unresolved_drawing_entries
            )
            declaration_entries.sort()
            all_inspections: list[object] = [
                *drawing_inspections,
                *chart_inspections,
                *user_shape_inspections,
            ]
            snapshot = ChartDefinitionSnapshot(
                chart_host_sheet_count=len(chart_host_sheets),
                chart_drawing_part_count=sum(
                    inspection.present for inspection in drawing_inspections
                ),
                chart_reference_count=sum(
                    inspection.chart_reference_count for inspection in drawing_inspections
                ),
                chart_part_count=len(chart_sources),
                chart_user_shape_part_count=len(user_shape_sources),
                chart_user_shape_count=sum(
                    inspection.shape_count for inspection in user_shape_inspections
                ),
                chart_type_count=sum(
                    inspection.chart_type_count for inspection in chart_inspections
                ),
                series_count=sum(
                    inspection.series_count for inspection in chart_inspections
                ),
                title_count=sum(
                    inspection.title_count for inspection in chart_inspections
                ),
                data_reference_count=sum(
                    inspection.data_reference_count for inspection in chart_inspections
                ),
                numeric_data_reference_count=sum(
                    inspection.numeric_data_reference_count
                    for inspection in chart_inspections
                ),
                string_data_reference_count=sum(
                    inspection.string_data_reference_count
                    for inspection in chart_inspections
                ),
                literal_data_point_count=sum(
                    inspection.literal_data_point_count for inspection in chart_inspections
                ),
                cached_data_point_count=sum(
                    inspection.cached_data_point_count for inspection in chart_inspections
                ),
                pivot_source_count=sum(
                    inspection.pivot_source_count for inspection in chart_inspections
                ),
                external_data_reference_count=sum(
                    inspection.external_data_reference_count
                    for inspection in chart_inspections
                ),
                user_shape_reference_count=sum(
                    inspection.user_shape_reference_count
                    for inspection in chart_inspections
                ),
                related_relationship_count=sum(
                    inspection.related_relationship_count for inspection in all_inspections
                ),
                external_relationship_count=sum(
                    inspection.external_relationship_count for inspection in all_inspections
                ),
                internal_related_part_count=payload_inspection.internal_part_count,
                fingerprinted_related_part_count=payload_inspection.fingerprinted_part_count,
                uninspected_related_part_count=payload_inspection.uninspected_part_count,
                unrecognized_part_count=(
                    unrecognized_declaration_count
                    + orphan_chart_part_count
                    + sum(inspection.unrecognized_count for inspection in all_inspections)
                ),
                declaration_signature=_private_external_data_signature(
                    tuple(declaration_entries)
                ),
                definition_signature=aggregate_signature(
                    list(chart_inspections), "definition_signature"
                ),
                cached_data_signature=aggregate_signature(
                    list(chart_inspections), "cached_data_signature"
                ),
                user_shape_signature=aggregate_signature(
                    list(user_shape_inspections), "definition_signature"
                ),
                relationship_signature=aggregate_signature(
                    all_inspections, "relationship_signature"
                ),
                related_part_payload_signature=payload_inspection.payload_signature,
            )
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        return _ChartDefinitionMetadata(
            default,
            (
                "FormulaFence could not inspect chart OOXML "
                f"({type(error).__name__}); affected charts were not compared.",
            ),
        )
    return _ChartDefinitionMetadata(snapshot, tuple(sorted(warnings)))


def _pivot_raw_relationships(
    archive: ZipFile,
    source_member: str,
    warnings: set[str],
    *,
    context: str,
) -> tuple[_PivotRawRelationship, ...]:
    """Read PivotTable-package relationships without opening their targets."""
    relationship_member = _relationship_part_path(source_member)
    try:
        root = _xml_root(archive, relationship_member)
    except KeyError:
        return ()
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect "
            f"{context} relationships for PivotTable inspection "
            f"({type(error).__name__}); affected PivotTables were not compared."
        )
        return ()
    if (
        _xml_local_name(root.tag) != "Relationships"
        or _xml_namespace(root.tag) != _PACKAGE_RELATIONSHIP_NS
    ):
        warnings.add(
            "FormulaFence found an unexpected "
            f"{context} relationship root while inspecting PivotTables; affected "
            "PivotTables were not compared."
        )
        return ()

    relationship_tag = f"{{{_PACKAGE_RELATIONSHIP_NS}}}Relationship"
    relationships: list[_PivotRawRelationship] = []
    for relationship in root.findall(relationship_tag):
        target = relationship.get("Target")
        target_mode = relationship.get("TargetMode", "Internal")
        safe_target = (
            _normalise_part_target(source_member, target)
            if target is not None and target_mode.casefold() == "internal"
            else None
        )
        relationships.append(
            _PivotRawRelationship(
                relationship_id=relationship.get("Id"),
                relationship_type=relationship.get("Type", ""),
                target=target,
                target_mode=target_mode,
                safe_target=safe_target,
            )
        )
    if len(relationships) != len(root):
        warnings.add(
            "FormulaFence found unmodelled relationship XML while inspecting PivotTables; "
            "affected PivotTables may be incomplete."
        )
    return tuple(relationships)


def _pivot_relationship_signature(
    relationships: tuple[_PivotRawRelationship, ...],
) -> str | None:
    """Fingerprint PivotTable relationship semantics without identifier churn."""
    material = sorted(relationship.semantic_key() for relationship in relationships)
    return _private_external_data_signature(
        tuple(
            (f"relationship:{index}", repr(relationship))
            for index, relationship in enumerate(material)
        )
    )


def _pivot_relationship_semantics(
    relationships: tuple[_PivotRawRelationship, ...],
    warnings: set[str],
    *,
    context: str,
) -> dict[str, tuple[str, str, str]]:
    """Resolve PivotTable relationship identifiers into stable private semantics."""
    relationships_by_id: dict[str, list[_PivotRawRelationship]] = defaultdict(list)
    for relationship in relationships:
        if relationship.relationship_id:
            relationships_by_id[relationship.relationship_id].append(relationship)
        else:
            warnings.add(
                "FormulaFence found a PivotTable "
                f"{context} relationship without an id; affected PivotTables have a coverage gap."
            )
    semantics: dict[str, tuple[str, str, str]] = {}
    for relationship_id, matches in relationships_by_id.items():
        values = sorted(match.semantic_key() for match in matches)
        if len(values) > 1:
            warnings.add(
                "FormulaFence found duplicate PivotTable "
                f"{context} relationship ids; affected PivotTables have a coverage gap."
            )
        semantics[relationship_id] = values[0]
    return semantics


def _pivot_xml_fragment(
    element: ElementTree.Element,
    relationship_semantics: Mapping[str, tuple[str, str, str]],
    *,
    cache_definition_members_by_id: Mapping[str, str] | None = None,
    omit_cache_data: bool = False,
) -> tuple[object, ...]:
    """Canonicalize private PivotTable XML while resolving relationship identifiers."""
    local_name = _xml_local_name(element.tag)
    attributes: list[tuple[str, str]] = []
    for attribute, value in element.attrib.items():
        attribute_name = _xml_local_name(attribute)
        if attribute in _PIVOT_RELATIONSHIP_ATTRIBUTES:
            relationship = relationship_semantics.get(value)
            resolved = (
                ("relationship", relationship)
                if relationship is not None
                else ("missing-relationship", value)
            )
            attributes.append((_xml_display_name(attribute), repr(resolved)))
            continue
        if (
            local_name == "pivotTableDefinition"
            and attribute_name == "cacheId"
            and cache_definition_members_by_id is not None
        ):
            cache_definition_member = cache_definition_members_by_id.get(value)
            resolved = (
                ("cache-definition", cache_definition_member)
                if cache_definition_member is not None
                else ("missing-cache-definition", value)
            )
            attributes.append((_xml_display_name(attribute), repr(resolved)))
            continue
        if (
            omit_cache_data
            and local_name == "pivotCacheDefinition"
            and attribute_name in _PIVOT_CACHE_DEFINITION_VOLATILE_ATTRIBUTES
        ):
            continue
        attributes.append((_xml_display_name(attribute), value))
    children = tuple(
        _pivot_xml_fragment(
            child,
            relationship_semantics,
            cache_definition_members_by_id=cache_definition_members_by_id,
            omit_cache_data=omit_cache_data,
        )
        for child in element
        if not (
            omit_cache_data
            and _xml_namespace(child.tag) == _SPREADSHEETML_NS
            and _xml_local_name(child.tag) in {"cacheSource", "sharedItems"}
        )
    )
    text = element.text
    if children and text is not None and not text.strip():
        text = None
    return (
        _xml_display_name(element.tag),
        tuple(sorted(attributes)),
        text,
        children,
    )


def _pivot_xml_payload(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    budget: _PivotXmlBudget,
) -> tuple[bytes | None, str | None]:
    """Read one bounded PivotTable XML part without following relationship targets."""
    if budget.remaining_parts == 0:
        warnings.add(
            "FormulaFence reached its bounded PivotTable XML part count budget; affected "
            "PivotTables have a coverage gap."
        )
        return None, _private_external_data_signature(
            (("part-count-budget-exhausted", member),)
        )
    budget.remaining_parts -= 1
    try:
        info = archive.getinfo(member)
    except KeyError:
        warnings.add(
            "FormulaFence could not locate a PivotTable XML package part; affected "
            "PivotTables were not compared."
        )
        return None, _private_external_data_signature((("missing-member", member),))
    metadata = repr((member, info.file_size, info.compress_size, info.CRC))
    if info.file_size > _PIVOT_MAX_XML_PART_BYTES:
        warnings.add(
            "FormulaFence did not fully read an oversized PivotTable XML part; affected "
            "PivotTables have a coverage gap."
        )
        return None, _private_external_data_signature((("oversized-part", metadata),))
    if info.file_size > budget.remaining_bytes:
        warnings.add(
            "FormulaFence reached its bounded PivotTable XML read budget; affected "
            "PivotTables have a coverage gap."
        )
        return None, _private_external_data_signature(
            (("read-budget-exhausted", metadata),)
        )
    budget.remaining_bytes -= info.file_size
    try:
        return archive.read(member), None
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not read a PivotTable XML package part "
            f"({type(error).__name__}); affected PivotTables were not compared."
        )
        return None, _private_external_data_signature((("unreadable-part", metadata),))


def _pivot_cache_record_payloads(
    archive: ZipFile,
    members: set[str],
    unresolved_entries: list[tuple[str, str]],
    warnings: set[str],
    budget: _PivotCacheRecordBudget,
) -> _PivotCacheRecordPayloadInspection:
    """Fingerprint bounded raw PivotTable cache records without parsing values."""
    entries = list(unresolved_entries)
    fingerprinted_part_count = 0
    uninspected_part_count = len(unresolved_entries)
    for member in sorted(members, key=str.casefold):
        if budget.remaining_parts == 0:
            warnings.add(
                "FormulaFence reached its bounded PivotTable cache-record part count budget; "
                "affected PivotTables have a coverage gap."
            )
            entries.append(("part-count-budget-exhausted", member))
            uninspected_part_count += 1
            continue
        budget.remaining_parts -= 1
        try:
            info = archive.getinfo(member)
        except KeyError:
            warnings.add(
                "FormulaFence could not locate a PivotTable cache-record part; affected "
                "PivotTables were not compared."
            )
            entries.append(("missing-part", member))
            uninspected_part_count += 1
            continue
        metadata = repr((member, info.file_size, info.compress_size, info.CRC))
        if info.file_size > _PIVOT_CACHE_RECORD_MAX_BYTES:
            warnings.add(
                "FormulaFence did not fully read an oversized PivotTable cache-record part; "
                "affected PivotTables have a coverage gap."
            )
            entries.append(("oversized-part", metadata))
            uninspected_part_count += 1
            continue
        if info.file_size > budget.remaining_bytes:
            warnings.add(
                "FormulaFence reached its bounded PivotTable cache-record read budget; "
                "affected PivotTables have a coverage gap."
            )
            entries.append(("read-budget-exhausted", metadata))
            uninspected_part_count += 1
            continue
        budget.remaining_bytes -= info.file_size
        digest = hashlib.sha256()
        bytes_read = 0
        try:
            with archive.open(info) as payload:
                while chunk := payload.read(_PIVOT_CACHE_RECORD_HASH_CHUNK_BYTES):
                    bytes_read += len(chunk)
                    if bytes_read > info.file_size:
                        raise ValueError("payload exceeded its declared size")
                    digest.update(chunk)
            if bytes_read != info.file_size:
                raise ValueError("payload did not match its declared size")
        except (BadZipFile, OSError, RuntimeError, ValueError) as error:
            warnings.add(
                "FormulaFence could not fingerprint a PivotTable cache-record part "
                f"({type(error).__name__}); affected PivotTables were not compared."
            )
            entries.append(("unreadable-part", metadata))
            uninspected_part_count += 1
            continue
        entries.append(("payload", repr((member, digest.hexdigest()))))
        fingerprinted_part_count += 1

    entries.sort()
    return _PivotCacheRecordPayloadInspection(
        record_part_count=len(members) + len(unresolved_entries),
        fingerprinted_part_count=fingerprinted_part_count,
        uninspected_part_count=uninspected_part_count,
        payload_signature=_private_external_data_signature(tuple(entries)),
    )


def _pivot_table_part_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    xml_budget: _PivotXmlBudget,
    cache_definition_members_by_id: Mapping[str, tuple[str, ...]],
) -> _PivotTablePartInspection:
    """Inspect one PivotTable view definition without calculating its report."""
    relationships = _pivot_raw_relationships(
        archive,
        member,
        warnings,
        context="PivotTable",
    )
    relationship_semantics = _pivot_relationship_semantics(
        relationships,
        warnings,
        context="PivotTable",
    )
    payload, fallback_signature = _pivot_xml_payload(
        archive,
        member,
        warnings,
        xml_budget,
    )
    if payload is None:
        return _PivotTablePartInspection(
            member=member,
            unrecognized_count=1,
            layout_signature=fallback_signature,
            relationship_signature=_pivot_relationship_signature(relationships),
        )
    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a PivotTable XML part "
            f"({type(error).__name__}); affected PivotTables were not compared."
        )
        return _PivotTablePartInspection(
            member=member,
            unrecognized_count=1,
            layout_signature=_private_payload_signature(payload),
            relationship_signature=_pivot_relationship_signature(relationships),
        )
    if (
        _xml_local_name(root.tag) != "pivotTableDefinition"
        or _xml_namespace(root.tag) != _SPREADSHEETML_NS
    ):
        warnings.add(
            "FormulaFence found a PivotTable part with an unexpected root; affected "
            "PivotTables were not compared."
        )
        return _PivotTablePartInspection(
            member=member,
            unrecognized_count=1,
            layout_signature=_private_payload_signature(payload),
            relationship_signature=_pivot_relationship_signature(relationships),
        )

    cache_id = root.get("cacheId")
    unrecognized_count = 0
    if cache_id is None:
        warnings.add(
            "FormulaFence found a PivotTable definition without a cache id; affected "
            "PivotTables have a coverage gap."
        )
        unrecognized_count += 1
    cache_relationships = tuple(
        relationship
        for relationship in relationships
        if relationship.relationship_type == _PIVOT_CACHE_DEFINITION_RELATIONSHIP
    )
    cache_definition_members: set[str] = set()
    for relationship in cache_relationships:
        if relationship.target_mode.casefold() != "internal" or relationship.safe_target is None:
            warnings.add(
                "FormulaFence found a PivotTable cache-definition relationship without a "
                "safe internal target; affected PivotTables were not compared."
            )
            unrecognized_count += 1
            continue
        cache_definition_members.add(relationship.safe_target)
    if not cache_relationships:
        warnings.add(
            "FormulaFence found a PivotTable definition without a cache-definition "
            "relationship; affected PivotTables have a coverage gap."
        )
        unrecognized_count += 1
    for relationship in relationships:
        if relationship.relationship_type == _PIVOT_CACHE_DEFINITION_RELATIONSHIP:
            continue
        warnings.add(
            "FormulaFence found an unmodelled PivotTable relationship; affected "
            "PivotTables have a coverage gap."
        )
        unrecognized_count += 1
    expected_cache_members = (
        set(cache_definition_members_by_id.get(cache_id, ())) if cache_id is not None else set()
    )
    if cache_id is not None and not expected_cache_members:
        warnings.add(
            "FormulaFence could not bind a PivotTable cache id to a workbook cache "
            "declaration; affected PivotTables have a coverage gap."
        )
        unrecognized_count += 1
    elif expected_cache_members and cache_definition_members != expected_cache_members:
        warnings.add(
            "FormulaFence found a PivotTable cache relationship that disagrees with its "
            "workbook cache declaration; affected PivotTables have a coverage gap."
        )
        unrecognized_count += 1

    def child_count(container_name: str, child_name: str) -> int:
        container = root.find(f"{{{_SPREADSHEETML_NS}}}{container_name}")
        if container is None:
            return 0
        return len(container.findall(f"{{{_SPREADSHEETML_NS}}}{child_name}"))

    try:
        layout_signature = _private_external_data_signature(
            (
                (
                    "pivot-table",
                    repr(
                        _pivot_xml_fragment(
                            root,
                            relationship_semantics,
                            cache_definition_members_by_id={
                                identifier: members[0]
                                for identifier, members in cache_definition_members_by_id.items()
                                if len(members) == 1
                            },
                        )
                    ),
                ),
            )
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested PivotTable part; "
            "affected PivotTables were not compared."
        )
        layout_signature = _private_payload_signature(payload)
        inspected = False
        unrecognized_count += 1
    else:
        inspected = True
    return _PivotTablePartInspection(
        member=member,
        cache_id=cache_id,
        layout_location_count=len(root.findall(f"{{{_SPREADSHEETML_NS}}}location")),
        pivot_field_count=child_count("pivotFields", "pivotField"),
        row_field_count=child_count("rowFields", "field"),
        column_field_count=child_count("colFields", "field"),
        page_field_count=child_count("pageFields", "pageField"),
        data_field_count=child_count("dataFields", "dataField"),
        filter_count=child_count("filters", "filter"),
        row_item_count=child_count("rowItems", "i"),
        column_item_count=child_count("colItems", "i"),
        related_relationship_count=len(relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in relationships
        ),
        cache_definition_members=tuple(
            sorted(cache_definition_members, key=str.casefold)
        ),
        unrecognized_count=unrecognized_count,
        inspected=inspected,
        layout_signature=layout_signature,
        relationship_signature=_pivot_relationship_signature(relationships),
    )


def _pivot_cache_definition_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    xml_budget: _PivotXmlBudget,
) -> _PivotCacheDefinitionInspection:
    """Inspect one PivotTable cache definition without opening cache-record values."""
    relationships = _pivot_raw_relationships(
        archive,
        member,
        warnings,
        context="PivotTable cache definition",
    )
    relationship_semantics = _pivot_relationship_semantics(
        relationships,
        warnings,
        context="PivotTable cache definition",
    )
    relationship_by_id: dict[str, _PivotRawRelationship] = {}
    for relationship in relationships:
        if relationship.relationship_id and relationship.relationship_id not in relationship_by_id:
            relationship_by_id[relationship.relationship_id] = relationship

    payload, fallback_signature = _pivot_xml_payload(
        archive,
        member,
        warnings,
        xml_budget,
    )
    if payload is None:
        return _PivotCacheDefinitionInspection(
            member=member,
            unrecognized_count=1,
            definition_signature=fallback_signature,
            relationship_signature=_pivot_relationship_signature(relationships),
        )
    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a PivotTable cache-definition XML part "
            f"({type(error).__name__}); affected PivotTables were not compared."
        )
        return _PivotCacheDefinitionInspection(
            member=member,
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_pivot_relationship_signature(relationships),
        )
    if (
        _xml_local_name(root.tag) != "pivotCacheDefinition"
        or _xml_namespace(root.tag) != _SPREADSHEETML_NS
    ):
        warnings.add(
            "FormulaFence found a PivotTable cache-definition part with an unexpected root; "
            "affected PivotTables were not compared."
        )
        return _PivotCacheDefinitionInspection(
            member=member,
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_pivot_relationship_signature(relationships),
        )

    unrecognized_count = 0
    relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
    record_relationship_id = root.get(relationship_attribute)
    cache_record_members: set[str] = set()
    if record_relationship_id is not None:
        relationship = relationship_by_id.get(record_relationship_id)
        if relationship is None:
            warnings.add(
                "FormulaFence found a PivotTable cache-record reference without a matching "
                "relationship; affected PivotTables have a coverage gap."
            )
            unrecognized_count += 1
        elif relationship.relationship_type != _PIVOT_CACHE_RECORDS_RELATIONSHIP:
            warnings.add(
                "FormulaFence found a PivotTable cache-record reference with an unexpected "
                "relationship type; affected PivotTables have a coverage gap."
            )
            unrecognized_count += 1
        elif (
            relationship.target_mode.casefold() != "internal"
            or relationship.safe_target is None
        ):
            warnings.add(
                "FormulaFence found a PivotTable cache-record reference without a safe "
                "internal target; affected PivotTables were not compared."
            )
            unrecognized_count += 1
        else:
            cache_record_members.add(relationship.safe_target)
    for relationship in relationships:
        if relationship.relationship_type != _PIVOT_CACHE_RECORDS_RELATIONSHIP:
            warnings.add(
                "FormulaFence found an unmodelled PivotTable cache-definition relationship; "
                "affected PivotTables have a coverage gap."
            )
            unrecognized_count += 1
            continue
        if relationship.relationship_id != record_relationship_id:
            warnings.add(
                "FormulaFence found a PivotTable cache-record relationship not bound by its "
                "cache definition; affected PivotTables have a coverage gap."
            )
            unrecognized_count += 1

    cache_record_count = 0
    record_count_value = root.get("recordCount")
    if record_count_value is not None:
        try:
            cache_record_count = int(record_count_value)
            if cache_record_count < 0:
                raise ValueError("negative record count")
        except ValueError:
            warnings.add(
                "FormulaFence could not interpret a PivotTable cache record count; affected "
                "PivotTables have a coverage gap."
            )
            cache_record_count = 0
            unrecognized_count += 1
    cache_fields = [
        element
        for element in root.iter(f"{{{_SPREADSHEETML_NS}}}cacheField")
    ]
    shared_items = [
        element
        for element in root.iter(f"{{{_SPREADSHEETML_NS}}}sharedItems")
    ]
    try:
        definition_signature = _private_external_data_signature(
            (
                (
                    "pivot-cache-definition",
                    repr(
                        _pivot_xml_fragment(
                            root,
                            relationship_semantics,
                            omit_cache_data=True,
                        )
                    ),
                ),
            )
        )
        cached_shared_item_signature = _private_external_data_signature(
            tuple(
                [("record-count", record_count_value or "")]
                + [
                    (
                        f"shared-items:{index}",
                        repr(_pivot_xml_fragment(items, relationship_semantics)),
                    )
                    for index, items in enumerate(shared_items)
                ]
            )
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested PivotTable cache "
            "definition; affected PivotTables were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        cached_shared_item_signature = None
        inspected = False
        unrecognized_count += 1
    else:
        inspected = True
    return _PivotCacheDefinitionInspection(
        member=member,
        cache_field_count=len(cache_fields),
        shared_item_count=sum(len(list(items)) for items in shared_items),
        calculated_item_count=sum(
            element.tag == f"{{{_SPREADSHEETML_NS}}}calculatedItem" for element in root.iter()
        ),
        calculated_member_count=sum(
            element.tag == f"{{{_SPREADSHEETML_NS}}}calculatedMember" for element in root.iter()
        ),
        cache_record_count=cache_record_count,
        related_relationship_count=len(relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in relationships
        ),
        cache_record_members=tuple(sorted(cache_record_members, key=str.casefold)),
        unrecognized_count=unrecognized_count,
        inspected=inspected,
        definition_signature=definition_signature,
        cached_shared_item_signature=cached_shared_item_signature,
        relationship_signature=_pivot_relationship_signature(relationships),
    )


def _pivot_table_metadata(path: Path) -> _PivotTableMetadata:
    """Inspect PivotTable views and cache packages before the reader omits them.

    The scan is package-only: it neither refreshes a cache nor renders a report.
    Cache source and refresh controls are already compared by the external-data
    scanner; this scanner instead protects PivotTable presentation, cache schema,
    shared-item, and bounded cache-record material.
    """
    warnings: set[str] = set()
    default = PivotTableDefinitionSnapshot()
    try:
        with ZipFile(path) as archive:
            try:
                workbook_root = _xml_root(archive, "xl/workbook.xml")
                sheet_parts = _sheet_xml_parts(archive)
            except (
                KeyError,
                ElementTree.ParseError,
                OSError,
                RuntimeError,
                ValueError,
            ) as error:
                return _PivotTableMetadata(
                    default,
                    (
                        "FormulaFence could not map workbook OOXML for PivotTable inspection "
                        f"({type(error).__name__}); affected PivotTables were not compared.",
                    ),
                )

            xml_budget = _PivotXmlBudget()
            record_budget = _PivotCacheRecordBudget()
            workbook_relationships = _pivot_raw_relationships(
                archive,
                "xl/workbook.xml",
                warnings,
                context="workbook",
            )
            workbook_relationships_by_id: dict[
                str, list[_PivotRawRelationship]
            ] = defaultdict(list)
            for relationship in workbook_relationships:
                if relationship.relationship_id:
                    workbook_relationships_by_id[relationship.relationship_id].append(
                        relationship
                    )

            cache_definition_members_by_id: dict[str, set[str]] = defaultdict(set)
            declaration_entries: list[tuple[str, str]] = []
            unrecognized_declaration_count = 0
            declared_workbook_relationship_ids: set[str] = set()
            declared_cache_ids: set[str] = set()
            cache_container = workbook_root.find(f"{{{_SPREADSHEETML_NS}}}pivotCaches")
            if cache_container is not None:
                for cache in cache_container.findall(f"{{{_SPREADSHEETML_NS}}}pivotCache"):
                    cache_id = cache.get("cacheId")
                    relationship_id = cache.get(f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id")
                    if cache_id is None:
                        warnings.add(
                            "FormulaFence found a PivotTable cache declaration without a cache "
                            "id; affected PivotTables have a coverage gap."
                        )
                        unrecognized_declaration_count += 1
                    elif cache_id in declared_cache_ids:
                        warnings.add(
                            "FormulaFence found duplicate PivotTable cache ids; affected "
                            "PivotTables have a coverage gap."
                        )
                        unrecognized_declaration_count += 1
                    else:
                        declared_cache_ids.add(cache_id)
                    if relationship_id is None:
                        warnings.add(
                            "FormulaFence found a PivotTable cache declaration without a "
                            "relationship; affected PivotTables have a coverage gap."
                        )
                        unrecognized_declaration_count += 1
                        continue
                    declared_workbook_relationship_ids.add(relationship_id)
                    relationship_matches = workbook_relationships_by_id.get(
                        relationship_id,
                        [],
                    )
                    if not relationship_matches:
                        warnings.add(
                            "FormulaFence found a PivotTable cache declaration without a "
                            "matching workbook relationship; affected PivotTables have a "
                            "coverage gap."
                        )
                        unrecognized_declaration_count += 1
                        continue
                    if len(relationship_matches) > 1:
                        warnings.add(
                            "FormulaFence found duplicate workbook relationship ids for a "
                            "PivotTable cache; affected PivotTables have a coverage gap."
                        )
                        unrecognized_declaration_count += 1
                        continue
                    relationship = relationship_matches[0]
                    if relationship.relationship_type != _PIVOT_CACHE_DEFINITION_RELATIONSHIP:
                        warnings.add(
                            "FormulaFence found a PivotTable cache declaration with an "
                            "unexpected relationship type; affected PivotTables have a "
                            "coverage gap."
                        )
                        unrecognized_declaration_count += 1
                        continue
                    if (
                        relationship.target_mode.casefold() != "internal"
                        or relationship.safe_target is None
                    ):
                        warnings.add(
                            "FormulaFence found a PivotTable cache declaration without a safe "
                            "internal target; affected PivotTables were not compared."
                        )
                        unrecognized_declaration_count += 1
                        continue
                    declaration_entries.append(
                        ("workbook-cache-definition", relationship.safe_target)
                    )
                    if cache_id is not None:
                        cache_definition_members_by_id[cache_id].add(
                            relationship.safe_target
                        )

            for relationship in workbook_relationships:
                if relationship.relationship_type != _PIVOT_CACHE_DEFINITION_RELATIONSHIP:
                    continue
                if relationship.relationship_id in declared_workbook_relationship_ids:
                    continue
                warnings.add(
                    "FormulaFence found a workbook PivotTable cache-definition relationship "
                    "not bound by a cache declaration; affected PivotTables have a coverage "
                    "gap."
                )
                unrecognized_declaration_count += 1
            for _cache_id, members in cache_definition_members_by_id.items():
                if len(members) <= 1:
                    continue
                warnings.add(
                    "FormulaFence found one PivotTable cache id bound to multiple cache "
                    "definitions; affected PivotTables have a coverage gap."
                )
                unrecognized_declaration_count += 1

            pivot_table_sources: dict[str, set[tuple[str, str]]] = defaultdict(set)
            unresolved_sheet_entries: list[tuple[str, str]] = []
            for sheet, (member, sheet_kind) in sorted(
                sheet_parts.items(),
                key=lambda item: item[0].casefold(),
            ):
                if sheet_kind != "worksheet":
                    continue
                relationships = _pivot_raw_relationships(
                    archive,
                    member,
                    warnings,
                    context="worksheet",
                )
                for relationship in relationships:
                    if relationship.relationship_type != _PIVOT_TABLE_RELATIONSHIP:
                        continue
                    if (
                        relationship.target_mode.casefold() != "internal"
                        or relationship.safe_target is None
                    ):
                        warnings.add(
                            "FormulaFence found a worksheet PivotTable relationship without a "
                            "safe internal target; affected PivotTables were not compared."
                        )
                        unresolved_sheet_entries.append(
                            (member, repr(relationship.semantic_key()))
                        )
                        unrecognized_declaration_count += 1
                        continue
                    pivot_table_sources[relationship.safe_target].add((sheet, member))

            referenced_pivot_table_members = set(pivot_table_sources)
            for entry in archive.infolist():
                member = entry.filename
                if not _PIVOT_TABLE_PART_PATTERN.fullmatch(member):
                    continue
                pivot_table_sources.setdefault(member, set())
                if member in referenced_pivot_table_members:
                    continue
                warnings.add(
                    "FormulaFence found a PivotTable package part not declared by an "
                    "inspected worksheet; affected PivotTables have a coverage gap."
                )
                unrecognized_declaration_count += 1

            cache_members_by_id = {
                cache_id: tuple(sorted(members, key=str.casefold))
                for cache_id, members in cache_definition_members_by_id.items()
            }
            pivot_table_inspections: list[_PivotTablePartInspection] = []
            cache_definition_sources: dict[str, set[str]] = defaultdict(set)
            for _cache_id, members in cache_members_by_id.items():
                for member in members:
                    cache_definition_sources[member].add("workbook")
            for member in sorted(pivot_table_sources, key=str.casefold):
                sources = tuple(
                    sorted(pivot_table_sources[member], key=lambda item: item[0].casefold())
                )
                declaration_entries.append(("pivot-table-part", repr((member, sources))))
                inspection = _pivot_table_part_inspection(
                    archive,
                    member,
                    warnings,
                    xml_budget,
                    cache_members_by_id,
                )
                pivot_table_inspections.append(inspection)
                for cache_member in inspection.cache_definition_members:
                    cache_definition_sources[cache_member].add(member)

            referenced_cache_definition_members = set(cache_definition_sources)
            for entry in archive.infolist():
                member = entry.filename
                if not _PIVOT_CACHE_DEFINITION_PART_PATTERN.fullmatch(member):
                    continue
                cache_definition_sources.setdefault(member, set())
                if member in referenced_cache_definition_members:
                    continue
                warnings.add(
                    "FormulaFence found a PivotTable cache-definition package part not bound "
                    "by an inspected declaration; affected PivotTables have a coverage gap."
                )
                unrecognized_declaration_count += 1

            cache_definition_inspections: list[_PivotCacheDefinitionInspection] = []
            cache_record_members: set[str] = set()
            for member in sorted(cache_definition_sources, key=str.casefold):
                sources = tuple(sorted(cache_definition_sources[member], key=str.casefold))
                declaration_entries.append(
                    ("pivot-cache-definition-part", repr((member, sources)))
                )
                inspection = _pivot_cache_definition_inspection(
                    archive,
                    member,
                    warnings,
                    xml_budget,
                )
                cache_definition_inspections.append(inspection)
                cache_record_members.update(inspection.cache_record_members)

            referenced_cache_record_members = set(cache_record_members)
            for entry in archive.infolist():
                member = entry.filename
                if not _PIVOT_CACHE_RECORDS_PART_PATTERN.fullmatch(member):
                    continue
                cache_record_members.add(member)
                if member in referenced_cache_record_members:
                    continue
                warnings.add(
                    "FormulaFence found a PivotTable cache-record package part not bound by "
                    "an inspected cache definition; affected PivotTables have a coverage gap."
                )
                unrecognized_declaration_count += 1

            cache_record_payloads = _pivot_cache_record_payloads(
                archive,
                cache_record_members,
                [],
                warnings,
                record_budget,
            )

            def aggregate_signature(
                inspections: list[object], attribute: str
            ) -> str | None:
                material = sorted(
                    (inspection.member, value)
                    for inspection in inspections
                    if (value := getattr(inspection, attribute)) is not None
                )
                return _private_external_data_signature(tuple(material))

            pivot_table_sheets = {
                sheet
                for member, sources in pivot_table_sources.items()
                if any(inspection.member == member for inspection in pivot_table_inspections)
                for sheet, _source_member in sources
            }
            declaration_entries.extend(
                ("unresolved-worksheet-pivot-relationship", entry)
                for entry in unresolved_sheet_entries
            )
            declaration_entries.sort()
            all_inspections: list[object] = [
                *pivot_table_inspections,
                *cache_definition_inspections,
            ]
            snapshot = PivotTableDefinitionSnapshot(
                pivot_table_sheet_count=len(pivot_table_sheets),
                pivot_table_part_count=len(pivot_table_sources),
                pivot_cache_definition_part_count=len(cache_definition_sources),
                pivot_cache_records_part_count=cache_record_payloads.record_part_count,
                pivot_cache_binding_count=sum(
                    len(inspection.cache_definition_members)
                    for inspection in pivot_table_inspections
                ),
                layout_location_count=sum(
                    inspection.layout_location_count for inspection in pivot_table_inspections
                ),
                pivot_field_count=sum(
                    inspection.pivot_field_count for inspection in pivot_table_inspections
                ),
                row_field_count=sum(
                    inspection.row_field_count for inspection in pivot_table_inspections
                ),
                column_field_count=sum(
                    inspection.column_field_count for inspection in pivot_table_inspections
                ),
                page_field_count=sum(
                    inspection.page_field_count for inspection in pivot_table_inspections
                ),
                data_field_count=sum(
                    inspection.data_field_count for inspection in pivot_table_inspections
                ),
                filter_count=sum(
                    inspection.filter_count for inspection in pivot_table_inspections
                ),
                row_item_count=sum(
                    inspection.row_item_count for inspection in pivot_table_inspections
                ),
                column_item_count=sum(
                    inspection.column_item_count for inspection in pivot_table_inspections
                ),
                cache_field_count=sum(
                    inspection.cache_field_count
                    for inspection in cache_definition_inspections
                ),
                shared_item_count=sum(
                    inspection.shared_item_count
                    for inspection in cache_definition_inspections
                ),
                calculated_item_count=sum(
                    inspection.calculated_item_count
                    for inspection in cache_definition_inspections
                ),
                calculated_member_count=sum(
                    inspection.calculated_member_count
                    for inspection in cache_definition_inspections
                ),
                cache_record_count=sum(
                    inspection.cache_record_count
                    for inspection in cache_definition_inspections
                ),
                related_relationship_count=sum(
                    inspection.related_relationship_count for inspection in all_inspections
                ),
                external_relationship_count=sum(
                    inspection.external_relationship_count for inspection in all_inspections
                ),
                fingerprinted_cache_record_part_count=(
                    cache_record_payloads.fingerprinted_part_count
                ),
                uninspected_cache_record_part_count=(
                    cache_record_payloads.uninspected_part_count
                ),
                unrecognized_part_count=(
                    unrecognized_declaration_count
                    + sum(inspection.unrecognized_count for inspection in all_inspections)
                ),
                declaration_signature=_private_external_data_signature(
                    tuple(declaration_entries)
                ),
                layout_signature=aggregate_signature(
                    list(pivot_table_inspections), "layout_signature"
                ),
                cache_definition_signature=aggregate_signature(
                    list(cache_definition_inspections), "definition_signature"
                ),
                cached_shared_item_signature=aggregate_signature(
                    list(cache_definition_inspections), "cached_shared_item_signature"
                ),
                relationship_signature=aggregate_signature(
                    all_inspections, "relationship_signature"
                ),
                cache_record_payload_signature=cache_record_payloads.payload_signature,
            )
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        return _PivotTableMetadata(
            default,
            (
                "FormulaFence could not inspect PivotTable OOXML "
                f"({type(error).__name__}); affected PivotTables were not compared.",
            ),
        )
    return _PivotTableMetadata(snapshot, tuple(sorted(warnings)))


def _pivot_reader_cache_record_replacements(
    path: Path,
) -> tuple[dict[str, bytes], tuple[str, ...]]:
    """Build safe cache-definition overlays for the ordinary workbook reader.

    openpyxl eagerly parses complete PivotTable cache-record streams even though
    FormulaFence never needs those values to index cells. Replace only the
    temporary reader copy's cache-record bindings after the package scanner has
    fingerprinted the original payload under its explicit limits.
    """
    replacements: dict[str, bytes] = {}
    reader_warnings: set[str] = set()
    try:
        with ZipFile(path) as archive:
            xml_budget = _PivotXmlBudget()
            cache_definition_members = {
                entry.filename
                for entry in archive.infolist()
                if _PIVOT_CACHE_DEFINITION_PART_PATTERN.fullmatch(entry.filename)
            }
            workbook_relationships = _pivot_raw_relationships(
                archive,
                "xl/workbook.xml",
                set(),
                context="workbook",
            )
            cache_definition_members.update(
                relationship.safe_target
                for relationship in workbook_relationships
                if (
                    relationship.relationship_type == _PIVOT_CACHE_DEFINITION_RELATIONSHIP
                    and relationship.safe_target is not None
                )
            )
            for member in sorted(cache_definition_members, key=str.casefold):
                payload, _fallback_signature = _pivot_xml_payload(
                    archive,
                    member,
                    set(),
                    xml_budget,
                )
                if payload is None:
                    reader_warnings.add(
                        "FormulaFence could not isolate a bounded PivotTable cache "
                        "definition from the underlying workbook reader."
                    )
                    continue
                try:
                    cache_definition = _xml_root_from_payload(payload)
                except (ElementTree.ParseError, OSError, RuntimeError, ValueError):
                    continue
                if (
                    _xml_local_name(cache_definition.tag) != "pivotCacheDefinition"
                    or _xml_namespace(cache_definition.tag) != _SPREADSHEETML_NS
                ):
                    continue

                relationship_member = _relationship_part_path(member)
                relationship_payload, _fallback_signature = _pivot_xml_payload(
                    archive,
                    relationship_member,
                    set(),
                    xml_budget,
                )
                if relationship_payload is None:
                    continue
                try:
                    relationships = _xml_root_from_payload(relationship_payload)
                except (ElementTree.ParseError, OSError, RuntimeError, ValueError):
                    continue
                if (
                    _xml_local_name(relationships.tag) != "Relationships"
                    or _xml_namespace(relationships.tag) != _PACKAGE_RELATIONSHIP_NS
                ):
                    continue
                relationship_tag = f"{{{_PACKAGE_RELATIONSHIP_NS}}}Relationship"
                record_relationships = [
                    relationship
                    for relationship in relationships.findall(relationship_tag)
                    if relationship.get("Type") == _PIVOT_CACHE_RECORDS_RELATIONSHIP
                ]
                if not record_relationships:
                    continue
                record_relationship_id_attribute = (
                    f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
                )
                cache_definition.attrib.pop(record_relationship_id_attribute, None)
                for relationship in record_relationships:
                    relationships.remove(relationship)
                replacements[member] = ElementTree.tostring(
                    cache_definition,
                    encoding="utf-8",
                    xml_declaration=True,
                )
                replacements[relationship_member] = ElementTree.tostring(
                    relationships,
                    encoding="utf-8",
                    xml_declaration=True,
                )
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        reader_warnings.add(
            "FormulaFence could not prepare a PivotTable-safe workbook-reader copy "
            f"({type(error).__name__})."
        )
    return replacements, tuple(sorted(reader_warnings))


def _openpyxl_safe_source(path: Path) -> tuple[Path, Path | None, tuple[str, ...]]:
    """Return a temporary source that prevents unbounded PivotTable record reads."""
    replacements, reader_warnings = _pivot_reader_cache_record_replacements(path)
    if not replacements:
        return path, None, reader_warnings

    temporary_path: Path | None = None
    warnings_for_reader = set(reader_warnings)
    try:
        descriptor, temporary_name = tempfile.mkstemp(
            prefix="formulafence-pivot-reader-",
            suffix=path.suffix,
        )
        os.close(descriptor)
        temporary_path = Path(temporary_name)
        shutil.copyfile(path, temporary_path)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            with ZipFile(temporary_path, "a", compression=ZIP_DEFLATED) as archive:
                for member, payload in sorted(replacements.items(), key=lambda item: item[0]):
                    archive.writestr(member, payload)
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
        warnings_for_reader.add(
            "FormulaFence could not isolate PivotTable cache records before the "
            f"underlying workbook reader ran ({type(error).__name__})."
        )
        return path, None, tuple(sorted(warnings_for_reader))
    return temporary_path, temporary_path, tuple(sorted(warnings_for_reader))


def _worksheet_control_raw_relationships(
    archive: ZipFile,
    source_member: str,
    warnings: set[str],
    *,
    context: str,
    missing_is_warning: bool = False,
) -> tuple[_WorksheetControlRawRelationship, ...]:
    """Read worksheet-control relationships without opening any package target."""
    relationship_member = _relationship_part_path(source_member)
    try:
        root = _xml_root(archive, relationship_member)
    except KeyError:
        if missing_is_warning:
            warnings.add(
                "FormulaFence could not locate "
                f"{context} relationships while inspecting worksheet embedded controls; "
                "affected controls may be incomplete."
            )
        return ()
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect "
            f"{context} relationships for worksheet embedded controls "
            f"({type(error).__name__}); affected controls were not compared."
        )
        return ()
    if (
        _xml_local_name(root.tag) != "Relationships"
        or _xml_namespace(root.tag) != _PACKAGE_RELATIONSHIP_NS
    ):
        warnings.add(
            "FormulaFence found an unexpected "
            f"{context} relationship root while inspecting worksheet embedded controls; "
            "affected controls were not compared."
        )
        return ()

    relationship_tag = f"{{{_PACKAGE_RELATIONSHIP_NS}}}Relationship"
    relationships: list[_WorksheetControlRawRelationship] = []
    for relationship in root.findall(relationship_tag):
        target = relationship.get("Target")
        target_mode = relationship.get("TargetMode", "Internal")
        safe_target = (
            _normalise_part_target(source_member, target)
            if target is not None and target_mode.casefold() == "internal"
            else None
        )
        relationships.append(
            _WorksheetControlRawRelationship(
                relationship_id=relationship.get("Id"),
                relationship_type=relationship.get("Type", ""),
                target=target,
                target_mode=target_mode,
                safe_target=safe_target,
            )
        )
    if len(relationships) != len(root):
        warnings.add(
            "FormulaFence found unmodelled relationship XML while inspecting worksheet "
            "embedded controls; affected controls may be incomplete."
        )
    return tuple(relationships)


def _worksheet_control_relationship_signature(
    relationships: tuple[_WorksheetControlRawRelationship, ...],
) -> str | None:
    """Fingerprint control relationship semantics without identifier churn."""
    material = sorted(relationship.semantic_key() for relationship in relationships)
    return _private_external_data_signature(
        tuple(
            (f"relationship:{index}", repr(relationship))
            for index, relationship in enumerate(material)
        )
    )


def _worksheet_control_relationship_semantics(
    relationships: tuple[_WorksheetControlRawRelationship, ...],
    warnings: set[str],
    *,
    context: str,
) -> dict[str, tuple[str, str, str]]:
    """Resolve private relationship identifiers while preserving malformed evidence."""
    relationships_by_id: dict[str, list[_WorksheetControlRawRelationship]] = defaultdict(
        list
    )
    for relationship in relationships:
        if relationship.relationship_id:
            relationships_by_id[relationship.relationship_id].append(relationship)
        else:
            warnings.add(
                "FormulaFence found a worksheet embedded-control "
                f"{context} relationship without an id; affected controls have a coverage gap."
            )
    semantics: dict[str, tuple[str, str, str]] = {}
    for relationship_id, matches in relationships_by_id.items():
        values = sorted(match.semantic_key() for match in matches)
        if len(values) > 1:
            warnings.add(
                "FormulaFence found duplicate worksheet embedded-control "
                f"{context} relationship ids; affected controls have a coverage gap."
            )
        semantics[relationship_id] = values[0]
    return semantics


def _worksheet_control_fragment(
    element: ElementTree.Element,
    relationship_semantics: Mapping[str, tuple[str, str, str]],
    *,
    relationship_attributes: frozenset[str] | None = None,
) -> tuple[object, ...]:
    """Canonicalize private control XML while resolving relationship identifiers."""
    if relationship_attributes is None:
        relationship_attributes = frozenset(
            {f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"}
        )
    attributes: list[tuple[str, str]] = []
    for attribute, value in element.attrib.items():
        if attribute in relationship_attributes:
            relationship = relationship_semantics.get(value)
            resolved = (
                ("relationship", relationship)
                if relationship is not None
                else ("missing-relationship", value)
            )
            attributes.append((_xml_display_name(attribute), repr(resolved)))
        else:
            attributes.append((_xml_display_name(attribute), value))
    children = tuple(
        _worksheet_control_fragment(
            child,
            relationship_semantics,
            relationship_attributes=relationship_attributes,
        )
        for child in element
    )
    text = element.text
    if children and text is not None and not text.strip():
        text = None
    return (
        _xml_display_name(element.tag),
        tuple(sorted(attributes)),
        text,
        children,
    )


def _worksheet_control_boolean(
    value: str | None,
    default: bool,
    warnings: set[str],
    *,
    context: str,
    attribute: str,
) -> bool:
    """Read an OOXML boolean while retaining invalid values privately."""
    if value is None:
        return default
    lowered = value.casefold()
    if lowered in {"1", "true", "on"}:
        return True
    if lowered in {"0", "false", "off"}:
        return False
    warnings.add(
        "FormulaFence could not interpret a worksheet embedded-control "
        f"{context} {attribute} boolean; the affected control has a coverage gap."
    )
    return default


def _worksheet_control_ole_auto_update(
    value: str | None,
    warnings: set[str],
) -> bool:
    """Return whether an OLE link requests automatic updates."""
    if value is None or value.casefold() == "oncall":
        return False
    if value.casefold() == "always":
        return True
    warnings.add(
        "FormulaFence found an unrecognized worksheet embedded-control OLE update "
        "setting; the affected control has a coverage gap."
    )
    return False


def _worksheet_control_xml_payload(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    budget: _WorksheetEmbeddedControlXmlBudget,
) -> tuple[bytes | None, str | None]:
    """Read one bounded control XML part without parsing it or following targets."""
    if budget.remaining_parts == 0:
        warnings.add(
            "FormulaFence reached its bounded worksheet embedded-control XML part count "
            "budget; the affected controls have a coverage gap."
        )
        return None, _private_external_data_signature(
            (("part-count-budget-exhausted", member),)
        )
    budget.remaining_parts -= 1
    try:
        info = archive.getinfo(member)
    except KeyError:
        warnings.add(
            "FormulaFence could not locate a worksheet embedded-control XML package part; "
            "the affected controls were not compared."
        )
        return None, _private_external_data_signature((("missing-member", member),))
    metadata = repr((member, info.file_size, info.compress_size, info.CRC))
    if info.file_size > _WORKSHEET_EMBEDDED_CONTROL_MAX_XML_PART_BYTES:
        warnings.add(
            "FormulaFence did not fully read an oversized worksheet embedded-control XML "
            "part; the affected controls have a coverage gap."
        )
        return None, _private_external_data_signature((("oversized-part", metadata),))
    if info.file_size > budget.remaining_bytes:
        warnings.add(
            "FormulaFence reached its bounded worksheet embedded-control XML read budget; "
            "the affected controls have a coverage gap."
        )
        return None, _private_external_data_signature(
            (("read-budget-exhausted", metadata),)
        )
    budget.remaining_bytes -= info.file_size
    try:
        return archive.read(member), None
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not read a worksheet embedded-control XML package part "
            f"({type(error).__name__}); the affected controls were not compared."
        )
        return None, _private_external_data_signature((("unreadable-part", metadata),))


def _worksheet_control_related_part_payloads(
    archive: ZipFile,
    members: set[str],
    unresolved_entries: list[tuple[str, str]],
    warnings: set[str],
    budget: _WorksheetEmbeddedControlRelatedPartBudget,
) -> _WorksheetControlRelatedPartPayloadInspection:
    """Fingerprint bounded direct control payloads without opening their contents.

    ActiveX binary data and OLE/package payloads are arbitrary formats. The
    scanner therefore resolves only safe internal targets and hashes their
    bytes; it never executes, parses, or follows embedded or external content.
    """
    entries = list(unresolved_entries)
    fingerprinted_part_count = 0
    uninspected_part_count = len(unresolved_entries)
    for member in sorted(members, key=str.casefold):
        if budget.remaining_parts == 0:
            warnings.add(
                "FormulaFence reached its bounded worksheet embedded-control payload "
                "part count budget; the affected controls have a coverage gap."
            )
            entries.append(("part-count-budget-exhausted", member))
            uninspected_part_count += 1
            continue
        budget.remaining_parts -= 1
        try:
            info = archive.getinfo(member)
        except KeyError:
            warnings.add(
                "FormulaFence could not locate a worksheet embedded-control payload part; "
                "the affected controls were not compared."
            )
            entries.append(("missing-part", member))
            uninspected_part_count += 1
            continue
        metadata = repr((member, info.file_size, info.compress_size, info.CRC))
        if info.file_size > _WORKSHEET_EMBEDDED_CONTROL_RELATED_PART_MAX_BYTES:
            warnings.add(
                "FormulaFence did not fully read an oversized worksheet embedded-control "
                "payload part; the affected controls have a coverage gap."
            )
            entries.append(("oversized-part", metadata))
            uninspected_part_count += 1
            continue
        if info.file_size > budget.remaining_bytes:
            warnings.add(
                "FormulaFence reached its bounded worksheet embedded-control payload "
                "read budget; the affected controls have a coverage gap."
            )
            entries.append(("read-budget-exhausted", metadata))
            uninspected_part_count += 1
            continue
        budget.remaining_bytes -= info.file_size
        digest = hashlib.sha256()
        bytes_read = 0
        try:
            with archive.open(info) as payload:
                while chunk := payload.read(_WORKSHEET_EMBEDDED_CONTROL_HASH_CHUNK_BYTES):
                    bytes_read += len(chunk)
                    if bytes_read > info.file_size:
                        raise ValueError("payload exceeded its declared size")
                    digest.update(chunk)
            if bytes_read != info.file_size:
                raise ValueError("payload did not match its declared size")
        except (BadZipFile, OSError, RuntimeError, ValueError) as error:
            warnings.add(
                "FormulaFence could not fingerprint a worksheet embedded-control payload "
                f"part ({type(error).__name__}); the affected controls were not compared."
            )
            entries.append(("unreadable-part", metadata))
            uninspected_part_count += 1
            continue
        entries.append(("payload", repr((member, digest.hexdigest()))))
        fingerprinted_part_count += 1

    entries.sort()
    return _WorksheetControlRelatedPartPayloadInspection(
        internal_part_count=len(members) + len(unresolved_entries),
        fingerprinted_part_count=fingerprinted_part_count,
        uninspected_part_count=uninspected_part_count,
        payload_signature=_private_external_data_signature(tuple(entries)),
    )


def _worksheet_control_sheet_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    xml_budget: _WorksheetEmbeddedControlXmlBudget,
) -> _WorksheetControlSheetInspection:
    """Inspect one worksheet's control bindings without reading cells or payloads."""
    relationships = _worksheet_control_raw_relationships(
        archive,
        member,
        warnings,
        context="worksheet",
    )
    relationship_semantics = _worksheet_control_relationship_semantics(
        relationships,
        warnings,
        context="worksheet",
    )
    relationship_by_id: dict[str, _WorksheetControlRawRelationship] = {}
    for relationship in relationships:
        if (
            relationship.relationship_id
            and relationship.relationship_id not in relationship_by_id
        ):
            relationship_by_id[relationship.relationship_id] = relationship

    active_x_members: set[str] = set()
    control_property_members: set[str] = set()
    legacy_vml_members: set[str] = set()
    payload_members: set[str] = set()
    unresolved_payload_entries: list[tuple[str, str]] = []
    legacy_vml_relationships = tuple(
        relationship
        for relationship in relationships
        if relationship.relationship_type == _WORKSHEET_VML_DRAWING_RELATIONSHIP
    )
    legacy_vml_unrecognized_count = 0
    for relationship in legacy_vml_relationships:
        if relationship.safe_target is None:
            warnings.add(
                "FormulaFence found a legacy VML drawing relationship without a safe "
                "internal target; the affected controls were not compared."
            )
            legacy_vml_unrecognized_count += 1
        else:
            legacy_vml_members.add(relationship.safe_target)
    legacy_vml_kwargs = {
        "legacy_vml_members": tuple(sorted(legacy_vml_members, key=str.casefold)),
        "legacy_vml_relationships": legacy_vml_relationships,
        "legacy_vml_unrecognized_count": legacy_vml_unrecognized_count,
    }
    relevant_relationship_types = {
        _WORKSHEET_CONTROL_RELATIONSHIP,
        _WORKSHEET_CTRLPROP_RELATIONSHIP,
        _WORKSHEET_OLE_OBJECT_RELATIONSHIP,
        _WORKSHEET_EMBEDDED_PACKAGE_RELATIONSHIP,
        _WORKSHEET_ACTIVEX_BINARY_RELATIONSHIP,
    }
    for relationship in relationships:
        if relationship.relationship_type == _WORKSHEET_CONTROL_RELATIONSHIP:
            if relationship.safe_target is not None:
                active_x_members.add(relationship.safe_target)
        elif relationship.relationship_type == _WORKSHEET_CTRLPROP_RELATIONSHIP:
            if relationship.safe_target is not None:
                control_property_members.add(relationship.safe_target)
        elif (
            relationship.relationship_type
            in {
                _WORKSHEET_OLE_OBJECT_RELATIONSHIP,
                _WORKSHEET_EMBEDDED_PACKAGE_RELATIONSHIP,
                _WORKSHEET_ACTIVEX_BINARY_RELATIONSHIP,
            }
            and relationship.target_mode.casefold() == "internal"
        ):
            if relationship.safe_target is not None:
                payload_members.add(relationship.safe_target)
            else:
                unresolved_payload_entries.append(
                    (
                        "unsafe-direct-payload-target",
                        repr((relationship.relationship_type, relationship.target)),
                    )
                )

    relevant_relationships = tuple(
        relationship
        for relationship in relationships
        if relationship.relationship_type in relevant_relationship_types
    )
    # A conforming worksheet control or OLE object is relationship-bound. Do
    # not consume the bounded XML-read budget for ordinary worksheets that
    # have no relevant relationship declaration at all: large data sheets are
    # common, and no valid control chain can start there.
    if not relevant_relationships:
        return _WorksheetControlSheetInspection(member=member, **legacy_vml_kwargs)
    payload, fallback_signature = _worksheet_control_xml_payload(
        archive,
        member,
        warnings,
        xml_budget,
    )
    if payload is None:
        return _WorksheetControlSheetInspection(
            member=member,
            present=bool(relevant_relationships),
            active_x_members=tuple(sorted(active_x_members, key=str.casefold)),
            control_property_members=tuple(
                sorted(control_property_members, key=str.casefold)
            ),
            **legacy_vml_kwargs,
            payload_members=tuple(sorted(payload_members, key=str.casefold)),
            unresolved_payload_entries=tuple(unresolved_payload_entries),
            unrecognized_count=int(bool(relevant_relationships)),
            definition_signature=fallback_signature,
            relationship_signature=_worksheet_control_relationship_signature(
                relevant_relationships
            ),
        )

    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a worksheet embedded-control XML part "
            f"({type(error).__name__}); the affected controls were not compared."
        )
        return _WorksheetControlSheetInspection(
            member=member,
            present=bool(relevant_relationships),
            active_x_members=tuple(sorted(active_x_members, key=str.casefold)),
            control_property_members=tuple(
                sorted(control_property_members, key=str.casefold)
            ),
            **legacy_vml_kwargs,
            payload_members=tuple(sorted(payload_members, key=str.casefold)),
            unresolved_payload_entries=tuple(unresolved_payload_entries),
            unrecognized_count=int(bool(relevant_relationships)),
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_worksheet_control_relationship_signature(
                relevant_relationships
            ),
        )
    if (
        _xml_local_name(root.tag) != "worksheet"
        or _xml_namespace(root.tag) != _SPREADSHEETML_NS
    ):
        warnings.add(
            "FormulaFence found a worksheet embedded-control part with an unexpected root; "
            "the affected controls were not compared."
        )
        return _WorksheetControlSheetInspection(
            member=member,
            present=bool(relevant_relationships),
            active_x_members=tuple(sorted(active_x_members, key=str.casefold)),
            control_property_members=tuple(
                sorted(control_property_members, key=str.casefold)
            ),
            **legacy_vml_kwargs,
            payload_members=tuple(sorted(payload_members, key=str.casefold)),
            unresolved_payload_entries=tuple(unresolved_payload_entries),
            unrecognized_count=int(bool(relevant_relationships)),
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_worksheet_control_relationship_signature(
                relevant_relationships
            ),
        )

    controls_tag = f"{{{_SPREADSHEETML_NS}}}controls"
    control_tag = f"{{{_SPREADSHEETML_NS}}}control"
    control_properties_tag = f"{{{_SPREADSHEETML_NS}}}controlPr"
    ole_objects_tag = f"{{{_SPREADSHEETML_NS}}}oleObjects"
    ole_object_tag = f"{{{_SPREADSHEETML_NS}}}oleObject"
    object_properties_tag = f"{{{_SPREADSHEETML_NS}}}objectPr"
    relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"

    controls_containers = list(root.iter(controls_tag))
    ole_objects_containers = list(root.iter(ole_objects_tag))
    raw_controls = [
        control
        for container in controls_containers
        for control in container.iter(control_tag)
    ]
    controls: list[ElementTree.Element] = []
    seen_control_keys: set[tuple[str, ...]] = set()
    for index, control in enumerate(raw_controls):
        relationship_id = control.get(relationship_attribute)
        if relationship_id:
            key = ("relationship", relationship_id)
        else:
            key = (
                "unbound",
                str(index),
                control.get("shapeId", ""),
                control.get("name", ""),
            )
        if key in seen_control_keys:
            continue
        seen_control_keys.add(key)
        controls.append(control)
    ole_objects = [
        ole_object
        for container in ole_objects_containers
        for ole_object in container
        if ole_object.tag == ole_object_tag
    ]

    selected_relationship_ids: set[str] = set()
    unrecognized_count = 0

    def relationship_for(
        relationship_id: str,
        *,
        context: str,
    ) -> _WorksheetControlRawRelationship | None:
        nonlocal unrecognized_count
        selected_relationship_ids.add(relationship_id)
        if relationship_id not in relationship_semantics:
            warnings.add(
                "FormulaFence found a worksheet embedded-control "
                f"{context} reference without a matching relationship; affected controls "
                "have a coverage gap."
            )
            unrecognized_count += 1
            return None
        return relationship_by_id.get(relationship_id)

    control_macro_assignment_count = 0
    control_cell_link_count = 0
    control_source_range_count = 0
    for control in controls:
        relationship_id = control.get(relationship_attribute)
        if relationship_id is None:
            warnings.add(
                "FormulaFence found a worksheet control without a relationship id; "
                "the affected controls have a coverage gap."
            )
            unrecognized_count += 1
        elif relationship := relationship_for(relationship_id, context="control"):
            if relationship.relationship_type == _WORKSHEET_CONTROL_RELATIONSHIP:
                if relationship.safe_target is None:
                    warnings.add(
                        "FormulaFence found a worksheet control without a safe internal "
                        "ActiveX target; the affected controls were not compared."
                    )
                    unrecognized_count += 1
                else:
                    active_x_members.add(relationship.safe_target)
            elif relationship.relationship_type == _WORKSHEET_CTRLPROP_RELATIONSHIP:
                if relationship.safe_target is None:
                    warnings.add(
                        "FormulaFence found a worksheet control without a safe internal "
                        "form-control properties target; the affected controls were not compared."
                    )
                    unrecognized_count += 1
                else:
                    control_property_members.add(relationship.safe_target)
            else:
                warnings.add(
                    "FormulaFence found a worksheet control with an unexpected relationship "
                    "type; the affected controls have a coverage gap."
                )
                unrecognized_count += 1

        for properties in control:
            if properties.tag != control_properties_tag:
                continue
            control_macro_assignment_count += int(properties.get("macro") is not None)
            control_cell_link_count += int(properties.get("linkedCell") is not None)
            control_source_range_count += int(
                properties.get("listFillRange") is not None
            )
            properties_relationship_id = properties.get(relationship_attribute)
            if properties_relationship_id is None:
                continue
            if relationship := relationship_for(
                properties_relationship_id,
                context="control properties",
            ):
                # An inline controlPr can bind an image/presentation relationship
                # for an ActiveX control as well as a ctrlProp package part for a
                # form control. Retain every relationship semantically, but only
                # traverse the documented ctrlProp target type.
                if relationship.relationship_type == _WORKSHEET_CTRLPROP_RELATIONSHIP:
                    if relationship.safe_target is None:
                        warnings.add(
                            "FormulaFence found worksheet control properties without a safe "
                            "internal target; the affected controls were not compared."
                        )
                        unrecognized_count += 1
                    else:
                        control_property_members.add(relationship.safe_target)
                elif relationship.relationship_type == _WORKSHEET_CONTROL_RELATIONSHIP:
                    warnings.add(
                        "FormulaFence found worksheet control properties with a control-part "
                        "relationship; the affected controls have a coverage gap."
                    )
                    unrecognized_count += 1

    linked_ole_object_count = 0
    auto_load_ole_object_count = 0
    auto_update_ole_object_count = 0
    for ole_object in ole_objects:
        relationship_id = ole_object.get(relationship_attribute)
        relationship: _WorksheetControlRawRelationship | None = None
        if relationship_id is None:
            warnings.add(
                "FormulaFence found a worksheet OLE object without a relationship id; "
                "the affected controls have a coverage gap."
            )
            unrecognized_count += 1
        else:
            relationship = relationship_for(relationship_id, context="OLE object")
            if relationship is not None and relationship.relationship_type not in {
                _WORKSHEET_OLE_OBJECT_RELATIONSHIP,
                _WORKSHEET_EMBEDDED_PACKAGE_RELATIONSHIP,
            }:
                warnings.add(
                    "FormulaFence found a worksheet OLE object with an unexpected relationship "
                    "type; the affected controls have a coverage gap."
                )
                unrecognized_count += 1
        linked_ole_object_count += int(
            ole_object.get("link") is not None
            or (
                relationship is not None
                and relationship.target_mode.casefold() != "internal"
            )
        )
        auto_load_ole_object_count += _worksheet_control_boolean(
            ole_object.get("autoLoad"),
            False,
            warnings,
            context="OLE object",
            attribute="autoLoad",
        )
        auto_update_ole_object_count += _worksheet_control_ole_auto_update(
            ole_object.get("oleUpdate"),
            warnings,
        )
        for properties in ole_object.iter(object_properties_tag):
            properties_relationship_id = properties.get(relationship_attribute)
            if properties_relationship_id is not None:
                relationship_for(
                    properties_relationship_id,
                    context="OLE presentation",
                )

    for relationship in relevant_relationships:
        if (
            relationship.relationship_id is None
            or relationship.relationship_id not in selected_relationship_ids
        ):
            warnings.add(
                "FormulaFence found a worksheet embedded-control relationship not bound "
                "by inspected worksheet markup; the affected controls have a coverage gap."
            )
            unrecognized_count += 1

    selected_relationships = tuple(
        relationship
        for relationship in relationships
        if (
            relationship.relationship_id in selected_relationship_ids
            or relationship.relationship_type in relevant_relationship_types
        )
    )
    fragments: list[tuple[str, str]] = []
    try:
        fragments.extend(
            (
                f"controls:{index}",
                repr(_worksheet_control_fragment(container, relationship_semantics)),
            )
            for index, container in enumerate(controls_containers)
        )
        fragments.extend(
            (
                f"ole-objects:{index}",
                repr(_worksheet_control_fragment(container, relationship_semantics)),
            )
            for index, container in enumerate(ole_objects_containers)
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested worksheet "
            "embedded-control part; the affected controls were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        inspected = False
        unrecognized_count += 1
    else:
        definition_signature = _private_external_data_signature(tuple(fragments))
        inspected = True
    return _WorksheetControlSheetInspection(
        member=member,
        present=bool(controls_containers or ole_objects_containers or relevant_relationships),
        worksheet_control_count=len(controls),
        control_macro_assignment_count=control_macro_assignment_count,
        control_cell_link_count=control_cell_link_count,
        control_source_range_count=control_source_range_count,
        ole_object_count=len(ole_objects),
        linked_ole_object_count=linked_ole_object_count,
        auto_load_ole_object_count=auto_load_ole_object_count,
        auto_update_ole_object_count=auto_update_ole_object_count,
        related_relationship_count=len(selected_relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in selected_relationships
        ),
        active_x_members=tuple(sorted(active_x_members, key=str.casefold)),
        control_property_members=tuple(
            sorted(control_property_members, key=str.casefold)
        ),
        **legacy_vml_kwargs,
        payload_members=tuple(sorted(payload_members, key=str.casefold)),
        unresolved_payload_entries=tuple(unresolved_payload_entries),
        unrecognized_count=unrecognized_count,
        inspected=inspected,
        definition_signature=definition_signature,
        relationship_signature=_worksheet_control_relationship_signature(
            selected_relationships
        ),
    )


def _worksheet_control_activex_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    xml_budget: _WorksheetEmbeddedControlXmlBudget,
) -> _WorksheetControlActiveXInspection:
    """Inspect one ActiveX persistence XML part without interpreting its binary data."""
    relationships = _worksheet_control_raw_relationships(
        archive,
        member,
        warnings,
        context="ActiveX control",
    )
    relationship_semantics = _worksheet_control_relationship_semantics(
        relationships,
        warnings,
        context="ActiveX control",
    )
    relationship_by_id: dict[str, _WorksheetControlRawRelationship] = {}
    for relationship in relationships:
        if (
            relationship.relationship_id
            and relationship.relationship_id not in relationship_by_id
        ):
            relationship_by_id[relationship.relationship_id] = relationship

    payload_members: set[str] = set()
    unresolved_payload_entries: list[tuple[str, str]] = []
    for relationship in relationships:
        if relationship.relationship_type != _WORKSHEET_ACTIVEX_BINARY_RELATIONSHIP:
            continue
        if relationship.target_mode.casefold() != "internal":
            continue
        if relationship.safe_target is not None:
            payload_members.add(relationship.safe_target)
        else:
            unresolved_payload_entries.append(
                (
                    "unsafe-activex-binary-target",
                    repr((relationship.relationship_type, relationship.target)),
                )
            )

    payload, fallback_signature = _worksheet_control_xml_payload(
        archive,
        member,
        warnings,
        xml_budget,
    )
    if payload is None:
        return _WorksheetControlActiveXInspection(
            member=member,
            related_relationship_count=len(relationships),
            external_relationship_count=sum(
                relationship.target_mode.casefold() != "internal"
                for relationship in relationships
            ),
            payload_members=tuple(sorted(payload_members, key=str.casefold)),
            unresolved_payload_entries=tuple(unresolved_payload_entries),
            unrecognized_count=1,
            definition_signature=fallback_signature,
            relationship_signature=_worksheet_control_relationship_signature(
                relationships
            ),
        )
    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect an ActiveX control XML part "
            f"({type(error).__name__}); the affected controls were not compared."
        )
        return _WorksheetControlActiveXInspection(
            member=member,
            related_relationship_count=len(relationships),
            external_relationship_count=sum(
                relationship.target_mode.casefold() != "internal"
                for relationship in relationships
            ),
            payload_members=tuple(sorted(payload_members, key=str.casefold)),
            unresolved_payload_entries=tuple(unresolved_payload_entries),
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_worksheet_control_relationship_signature(
                relationships
            ),
        )
    if _xml_local_name(root.tag) != "ocx" or _xml_namespace(root.tag) != _ACTIVEX_NS:
        warnings.add(
            "FormulaFence found an ActiveX control part with an unexpected root; "
            "the affected controls were not compared."
        )
        return _WorksheetControlActiveXInspection(
            member=member,
            related_relationship_count=len(relationships),
            external_relationship_count=sum(
                relationship.target_mode.casefold() != "internal"
                for relationship in relationships
            ),
            payload_members=tuple(sorted(payload_members, key=str.casefold)),
            unresolved_payload_entries=tuple(unresolved_payload_entries),
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_worksheet_control_relationship_signature(
                relationships
            ),
        )

    relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id"
    referenced_relationship_ids = {
        relationship_id
        for element in root.iter()
        if (relationship_id := element.get(relationship_attribute)) is not None
    }
    unrecognized_count = 0
    for relationship_id in referenced_relationship_ids:
        semantic = relationship_semantics.get(relationship_id)
        if semantic is None:
            warnings.add(
                "FormulaFence found an ActiveX control binary reference without a matching "
                "relationship; the affected controls have a coverage gap."
            )
            unrecognized_count += 1
            continue
        if semantic[0] != _WORKSHEET_ACTIVEX_BINARY_RELATIONSHIP:
            warnings.add(
                "FormulaFence found an ActiveX control reference with an unexpected "
                "relationship type; the affected controls have a coverage gap."
            )
            unrecognized_count += 1
            continue
        relationship = relationship_by_id.get(relationship_id)
        if relationship is None or relationship.safe_target is None:
            warnings.add(
                "FormulaFence found an ActiveX control binary reference without a safe "
                "internal target; the affected controls were not compared."
            )
            unrecognized_count += 1
            continue
        payload_members.add(relationship.safe_target)

    for relationship in relationships:
        if relationship.relationship_type != _WORKSHEET_ACTIVEX_BINARY_RELATIONSHIP:
            warnings.add(
                "FormulaFence found an unexpected ActiveX control relationship type; "
                "the affected controls have a coverage gap."
            )
            unrecognized_count += 1
            continue
        if (
            relationship.relationship_id is None
            or relationship.relationship_id not in referenced_relationship_ids
        ):
            warnings.add(
                "FormulaFence found an ActiveX binary relationship not bound by the "
                "control definition; the affected controls have a coverage gap."
            )
            unrecognized_count += 1

    try:
        definition_signature = _private_external_data_signature(
            (("activex", repr(_worksheet_control_fragment(root, relationship_semantics))),)
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested ActiveX control "
            "part; the affected controls were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        inspected = False
        unrecognized_count += 1
    else:
        inspected = True
    return _WorksheetControlActiveXInspection(
        member=member,
        binary_reference_count=len(referenced_relationship_ids),
        related_relationship_count=len(relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in relationships
        ),
        payload_members=tuple(sorted(payload_members, key=str.casefold)),
        unresolved_payload_entries=tuple(unresolved_payload_entries),
        unrecognized_count=unrecognized_count,
        inspected=inspected,
        definition_signature=definition_signature,
        relationship_signature=_worksheet_control_relationship_signature(relationships),
    )


def _worksheet_control_property_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    xml_budget: _WorksheetEmbeddedControlXmlBudget,
) -> _WorksheetControlPropertyInspection:
    """Inspect one form-control properties part without disclosing formulas or labels."""
    relationships = _worksheet_control_raw_relationships(
        archive,
        member,
        warnings,
        context="form-control properties",
    )
    relationship_semantics = _worksheet_control_relationship_semantics(
        relationships,
        warnings,
        context="form-control properties",
    )
    payload_members: set[str] = set()
    unresolved_payload_entries: list[tuple[str, str]] = []
    for relationship in relationships:
        if relationship.target_mode.casefold() != "internal":
            continue
        if relationship.safe_target is not None:
            payload_members.add(relationship.safe_target)
        else:
            unresolved_payload_entries.append(
                (
                    "unsafe-control-properties-target",
                    repr((relationship.relationship_type, relationship.target)),
                )
            )

    payload, fallback_signature = _worksheet_control_xml_payload(
        archive,
        member,
        warnings,
        xml_budget,
    )
    if payload is None:
        return _WorksheetControlPropertyInspection(
            member=member,
            related_relationship_count=len(relationships),
            external_relationship_count=sum(
                relationship.target_mode.casefold() != "internal"
                for relationship in relationships
            ),
            payload_members=tuple(sorted(payload_members, key=str.casefold)),
            unresolved_payload_entries=tuple(unresolved_payload_entries),
            unrecognized_count=1,
            definition_signature=fallback_signature,
            relationship_signature=_worksheet_control_relationship_signature(
                relationships
            ),
        )
    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a form-control properties XML part "
            f"({type(error).__name__}); the affected controls were not compared."
        )
        return _WorksheetControlPropertyInspection(
            member=member,
            related_relationship_count=len(relationships),
            external_relationship_count=sum(
                relationship.target_mode.casefold() != "internal"
                for relationship in relationships
            ),
            payload_members=tuple(sorted(payload_members, key=str.casefold)),
            unresolved_payload_entries=tuple(unresolved_payload_entries),
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_worksheet_control_relationship_signature(
                relationships
            ),
        )
    if (
        _xml_local_name(root.tag) != "formControlPr"
        or _xml_namespace(root.tag) != _OFFICE_2010_SPREADSHEET_NS
    ):
        warnings.add(
            "FormulaFence found a form-control properties part with an unexpected root; "
            "the affected controls were not compared."
        )
        return _WorksheetControlPropertyInspection(
            member=member,
            related_relationship_count=len(relationships),
            external_relationship_count=sum(
                relationship.target_mode.casefold() != "internal"
                for relationship in relationships
            ),
            payload_members=tuple(sorted(payload_members, key=str.casefold)),
            unresolved_payload_entries=tuple(unresolved_payload_entries),
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_worksheet_control_relationship_signature(
                relationships
            ),
        )

    unrecognized_count = 0
    if relationships:
        warnings.add(
            "FormulaFence found relationships from a form-control properties part; "
            "the affected controls have a coverage gap."
        )
        unrecognized_count += len(relationships)
    formula_attributes = ("fmlaGroup", "fmlaLink", "fmlaRange", "fmlaTxbx")
    formula_binding_count = sum(
        root.get(attribute) is not None for attribute in formula_attributes
    )
    cell_link_count = sum(
        root.get(attribute) is not None
        for attribute in ("fmlaGroup", "fmlaLink", "fmlaTxbx")
    )
    source_range_count = int(root.get("fmlaRange") is not None)
    try:
        definition_signature = _private_external_data_signature(
            (
                (
                    "form-control-properties",
                    repr(_worksheet_control_fragment(root, relationship_semantics)),
                ),
            )
        )
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested form-control "
            "properties part; the affected controls were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        inspected = False
        unrecognized_count += 1
    else:
        inspected = True
    return _WorksheetControlPropertyInspection(
        member=member,
        formula_binding_count=formula_binding_count,
        cell_link_count=cell_link_count,
        source_range_count=source_range_count,
        related_relationship_count=len(relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in relationships
        ),
        payload_members=tuple(sorted(payload_members, key=str.casefold)),
        unresolved_payload_entries=tuple(unresolved_payload_entries),
        unrecognized_count=unrecognized_count,
        inspected=inspected,
        definition_signature=definition_signature,
        relationship_signature=_worksheet_control_relationship_signature(relationships),
    )


def _worksheet_vml_drawing_inspection(
    archive: ZipFile,
    member: str,
    warnings: set[str],
    xml_budget: _WorksheetEmbeddedControlXmlBudget,
) -> _WorksheetVmlDrawingInspection:
    """Inspect legacy VML controls without rendering drawings or evaluating bindings.

    VML drawing parts also carry ordinary comment notes. Only ``ClientData`` nodes
    whose ``ObjectType`` is not ``Note`` enter this control inventory, so comment
    layout and note text do not create control findings.
    """
    relationships = _worksheet_control_raw_relationships(
        archive,
        member,
        warnings,
        context="legacy VML drawing",
    )
    relationship_semantics = _worksheet_control_relationship_semantics(
        relationships,
        warnings,
        context="legacy VML drawing",
    )
    relationship_by_id: dict[str, _WorksheetControlRawRelationship] = {}
    for relationship in relationships:
        if (
            relationship.relationship_id
            and relationship.relationship_id not in relationship_by_id
        ):
            relationship_by_id[relationship.relationship_id] = relationship

    payload, fallback_signature = _worksheet_control_xml_payload(
        archive,
        member,
        warnings,
        xml_budget,
    )
    if payload is None:
        return _WorksheetVmlDrawingInspection(
            member=member,
            present=True,
            unrecognized_count=1,
            definition_signature=fallback_signature,
            relationship_signature=_worksheet_control_relationship_signature(
                relationships
            ),
        )
    try:
        root = _xml_root_from_payload(payload)
    except (ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        warnings.add(
            "FormulaFence could not inspect a legacy VML control XML part "
            f"({type(error).__name__}); the affected controls were not compared."
        )
        return _WorksheetVmlDrawingInspection(
            member=member,
            present=True,
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_worksheet_control_relationship_signature(
                relationships
            ),
        )
    if _xml_local_name(root.tag) != "xml" or _xml_namespace(root.tag) not in {
        None,
        _VML_NS,
    }:
        warnings.add(
            "FormulaFence found a legacy VML control part with an unexpected root; "
            "the affected controls were not compared."
        )
        return _WorksheetVmlDrawingInspection(
            member=member,
            present=True,
            unrecognized_count=1,
            definition_signature=_private_payload_signature(payload),
            relationship_signature=_worksheet_control_relationship_signature(
                relationships
            ),
        )

    client_data_tag = f"{{{_VML_EXCEL_NS}}}ClientData"
    client_data_nodes = list(root.iter(client_data_tag))
    parent_by_child = {
        child: parent
        for parent in root.iter()
        for child in parent
    }
    controls: list[ElementTree.Element] = []
    control_containers: list[ElementTree.Element] = []
    seen_containers: set[int] = set()
    unrecognized_count = 0
    for client_data in client_data_nodes:
        object_type = (client_data.get("ObjectType") or "").strip()
        if object_type.casefold() == "note":
            continue
        if not object_type:
            warnings.add(
                "FormulaFence found legacy VML ClientData without an ObjectType; "
                "the affected controls have a coverage gap."
            )
            unrecognized_count += 1
        controls.append(client_data)
        container = parent_by_child.get(client_data, client_data)
        if id(container) not in seen_containers:
            seen_containers.add(id(container))
            control_containers.append(container)

    if not controls:
        return _WorksheetVmlDrawingInspection(member=member, inspected=True)

    relationship_attributes = frozenset(
        {
            f"{{{_DOCUMENT_RELATIONSHIP_NS}}}id",
            f"{{{_VML_OFFICE_NS}}}relid",
        }
    )
    referenced_relationship_ids = {
        relationship_id
        for container in control_containers
        for element in container.iter()
        for relationship_attribute in relationship_attributes
        if (relationship_id := element.get(relationship_attribute)) is not None
    }
    selected_relationships: list[_WorksheetControlRawRelationship] = []
    for relationship_id in sorted(referenced_relationship_ids):
        semantic = relationship_semantics.get(relationship_id)
        if semantic is None:
            warnings.add(
                "FormulaFence found a legacy VML control relationship reference without "
                "a matching relationship; the affected controls have a coverage gap."
            )
            unrecognized_count += 1
            continue
        relationship = relationship_by_id.get(relationship_id)
        if relationship is not None:
            selected_relationships.append(relationship)

    macro_tag = f"{{{_VML_EXCEL_NS}}}FmlaMacro"
    cell_link_tags = {
        f"{{{_VML_EXCEL_NS}}}FmlaGroup",
        f"{{{_VML_EXCEL_NS}}}FmlaLink",
        f"{{{_VML_EXCEL_NS}}}FmlaTxbx",
    }
    source_range_tag = f"{{{_VML_EXCEL_NS}}}FmlaRange"
    camera_source_range_tag = f"{{{_VML_EXCEL_NS}}}FmlaPict"
    try:
        fragments = tuple(
            (
                f"legacy-vml-control:{index}",
                repr(
                    _worksheet_control_fragment(
                        container,
                        relationship_semantics,
                        relationship_attributes=relationship_attributes,
                    )
                ),
            )
            for index, container in enumerate(control_containers)
        )
        definition_signature = _private_external_data_signature(fragments)
    except RecursionError:
        warnings.add(
            "FormulaFence could not fully traverse an excessively nested legacy VML "
            "control part; the affected controls were not compared."
        )
        definition_signature = _private_payload_signature(payload)
        inspected = False
        unrecognized_count += 1
    else:
        inspected = True
    return _WorksheetVmlDrawingInspection(
        member=member,
        present=True,
        control_count=len(controls),
        macro_assignment_count=sum(
            sum(child.tag == macro_tag for child in client_data)
            for client_data in controls
        ),
        cell_link_count=sum(
            sum(child.tag in cell_link_tags for child in client_data)
            for client_data in controls
        ),
        source_range_count=sum(
            sum(child.tag == source_range_tag for child in client_data)
            for client_data in controls
        ),
        camera_source_range_count=sum(
            sum(child.tag == camera_source_range_tag for child in client_data)
            for client_data in controls
        ),
        related_relationship_count=len(selected_relationships),
        external_relationship_count=sum(
            relationship.target_mode.casefold() != "internal"
            for relationship in selected_relationships
        ),
        unrecognized_count=unrecognized_count,
        inspected=inspected,
        definition_signature=definition_signature,
        relationship_signature=_worksheet_control_relationship_signature(
            tuple(selected_relationships)
        ),
    )


def _worksheet_embedded_control_metadata(
    path: Path,
) -> _WorksheetEmbeddedControlMetadata:
    """Inspect worksheet controls and OLE data before load time.

    This scan is intentionally package-only: it does not initialize controls,
    deserialize OLE data, open embedded packages, render VML, or follow external
    targets.
    """
    warnings: set[str] = set()
    default = WorksheetEmbeddedControlSnapshot()
    try:
        with ZipFile(path) as archive:
            try:
                sheet_parts = _sheet_xml_parts(archive)
            except (KeyError, ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
                return _WorksheetEmbeddedControlMetadata(
                    default,
                    (
                        "FormulaFence could not map worksheet OOXML for embedded-control "
                        f"inspection ({type(error).__name__}); affected controls were not "
                        "compared.",
                    ),
                )

            xml_budget = _WorksheetEmbeddedControlXmlBudget()
            payload_budget = _WorksheetEmbeddedControlRelatedPartBudget()
            sheet_inspections: list[_WorksheetControlSheetInspection] = []
            active_x_sources: dict[str, set[str]] = defaultdict(set)
            control_property_sources: dict[str, set[str]] = defaultdict(set)
            legacy_vml_sources: dict[str, set[str]] = defaultdict(set)
            legacy_vml_declarations: dict[
                str, list[tuple[str, tuple[str, str, str]]]
            ] = defaultdict(list)
            declared_active_x_members: set[str] = set()
            declared_control_property_members: set[str] = set()
            declared_legacy_vml_members: set[str] = set()
            payload_members: set[str] = set()
            unresolved_payload_entries: list[tuple[str, str]] = []
            declaration_entries: list[tuple[str, str]] = []
            unresolved_legacy_vml_declarations: list[tuple[str, str]] = []

            for _, (member, sheet_kind) in sorted(
                sheet_parts.items(),
                key=lambda item: item[0].casefold(),
            ):
                if sheet_kind != "worksheet":
                    continue
                inspection = _worksheet_control_sheet_inspection(
                    archive,
                    member,
                    warnings,
                    xml_budget,
                )
                sheet_inspections.append(inspection)
                if inspection.present:
                    declaration_entries.append(
                        (
                            "worksheet-control-part",
                            repr(
                                (
                                    member,
                                    inspection.active_x_members,
                                    inspection.control_property_members,
                                    inspection.payload_members,
                                )
                            ),
                        )
                    )
                for active_x_member in inspection.active_x_members:
                    active_x_sources[active_x_member].add(member)
                    declared_active_x_members.add(active_x_member)
                for control_property_member in inspection.control_property_members:
                    control_property_sources[control_property_member].add(member)
                    declared_control_property_members.add(control_property_member)
                for legacy_vml_member in inspection.legacy_vml_members:
                    legacy_vml_sources[legacy_vml_member].add(member)
                    declared_legacy_vml_members.add(legacy_vml_member)
                for relationship in inspection.legacy_vml_relationships:
                    if relationship.safe_target is None:
                        unresolved_legacy_vml_declarations.append(
                            (member, repr(relationship.semantic_key()))
                        )
                    else:
                        legacy_vml_declarations[relationship.safe_target].append(
                            (member, relationship.semantic_key())
                        )
                payload_members.update(inspection.payload_members)
                unresolved_payload_entries.extend(inspection.unresolved_payload_entries)

            orphan_part_count = 0
            discovered_active_x_members = {
                entry.filename
                for entry in archive.infolist()
                if _WORKSHEET_ACTIVEX_PART_PATTERN.fullmatch(entry.filename)
            }
            for member in discovered_active_x_members:
                active_x_sources.setdefault(member, set())
                if member not in declared_active_x_members:
                    warnings.add(
                        "FormulaFence found an ActiveX control package part not declared "
                        "by an inspected worksheet; the affected controls have a coverage gap."
                    )
                    orphan_part_count += 1

            discovered_control_property_members = {
                entry.filename
                for entry in archive.infolist()
                if _WORKSHEET_CONTROL_PROPERTY_PART_PATTERN.fullmatch(entry.filename)
            }
            for member in discovered_control_property_members:
                control_property_sources.setdefault(member, set())
                if member not in declared_control_property_members:
                    warnings.add(
                        "FormulaFence found a form-control properties package part not declared "
                        "by an inspected worksheet; the affected controls have a coverage gap."
                    )
                    orphan_part_count += 1

            discovered_legacy_vml_members = {
                entry.filename
                for entry in archive.infolist()
                if _WORKSHEET_LEGACY_VML_PART_PATTERN.fullmatch(entry.filename)
            }
            for member in discovered_legacy_vml_members:
                legacy_vml_sources.setdefault(member, set())
                if member not in declared_legacy_vml_members:
                    warnings.add(
                        "FormulaFence found a legacy VML drawing package part not declared "
                        "by an inspected worksheet; the affected controls have a coverage gap."
                    )
                    orphan_part_count += 1

            active_x_inspections: list[_WorksheetControlActiveXInspection] = []
            for member in sorted(active_x_sources, key=str.casefold):
                sources = tuple(sorted(active_x_sources[member], key=str.casefold))
                declaration_entries.append(
                    ("activex-part", repr((member, sources)))
                )
                inspection = _worksheet_control_activex_inspection(
                    archive,
                    member,
                    warnings,
                    xml_budget,
                )
                active_x_inspections.append(inspection)
                payload_members.update(inspection.payload_members)
                unresolved_payload_entries.extend(inspection.unresolved_payload_entries)

            control_property_inspections: list[_WorksheetControlPropertyInspection] = []
            for member in sorted(control_property_sources, key=str.casefold):
                sources = tuple(sorted(control_property_sources[member], key=str.casefold))
                declaration_entries.append(
                    ("form-control-properties-part", repr((member, sources)))
                )
                inspection = _worksheet_control_property_inspection(
                    archive,
                    member,
                    warnings,
                    xml_budget,
                )
                control_property_inspections.append(inspection)
                payload_members.update(inspection.payload_members)
                unresolved_payload_entries.extend(inspection.unresolved_payload_entries)

            legacy_vml_inspections: list[_WorksheetVmlDrawingInspection] = []
            for member in sorted(legacy_vml_sources, key=str.casefold):
                inspection = _worksheet_vml_drawing_inspection(
                    archive,
                    member,
                    warnings,
                    xml_budget,
                )
                legacy_vml_inspections.append(inspection)
                if inspection.present:
                    declarations = tuple(
                        sorted(
                            legacy_vml_declarations.get(member, []),
                            key=repr,
                        )
                    )
                    declaration_entries.append(
                        ("legacy-vml-drawing-part", repr((member, declarations)))
                    )
            declaration_entries.extend(
                ("legacy-vml-worksheet-relationship", declaration)
                for declaration in unresolved_legacy_vml_declarations
            )

            payload_inspection = _worksheet_control_related_part_payloads(
                archive,
                payload_members,
                unresolved_payload_entries,
                warnings,
                payload_budget,
            )

            def aggregate_signature(
                inspections: list[object],
                attribute: str,
            ) -> str | None:
                material = sorted(
                    (inspection.member, value)
                    for inspection in inspections
                    if (value := getattr(inspection, attribute)) is not None
                )
                return _private_external_data_signature(tuple(material))

            control_sheet_members = {
                inspection.member
                for inspection in sheet_inspections
                if inspection.present
            }
            active_legacy_vml_members = {
                inspection.member
                for inspection in legacy_vml_inspections
                if inspection.present
            }
            active_legacy_vml_source_relationships = [
                relationship
                for inspection in sheet_inspections
                for relationship in inspection.legacy_vml_relationships
                if relationship.safe_target in active_legacy_vml_members
            ]
            for inspection in legacy_vml_inspections:
                if inspection.present:
                    control_sheet_members.update(legacy_vml_sources[inspection.member])
            all_inspections: list[object] = [
                *sheet_inspections,
                *active_x_inspections,
                *control_property_inspections,
                *legacy_vml_inspections,
            ]
            declaration_entries.sort()
            snapshot = WorksheetEmbeddedControlSnapshot(
                control_sheet_count=len(control_sheet_members),
                worksheet_control_count=sum(
                    inspection.worksheet_control_count for inspection in sheet_inspections
                ),
                active_x_part_count=len(active_x_sources),
                active_x_binary_reference_count=sum(
                    inspection.binary_reference_count
                    for inspection in active_x_inspections
                ),
                form_control_property_part_count=len(control_property_sources),
                legacy_vml_drawing_part_count=sum(
                    inspection.present for inspection in legacy_vml_inspections
                ),
                legacy_vml_control_count=sum(
                    inspection.control_count for inspection in legacy_vml_inspections
                ),
                legacy_vml_macro_assignment_count=sum(
                    inspection.macro_assignment_count
                    for inspection in legacy_vml_inspections
                ),
                legacy_vml_cell_link_count=sum(
                    inspection.cell_link_count for inspection in legacy_vml_inspections
                ),
                legacy_vml_source_range_count=sum(
                    inspection.source_range_count
                    for inspection in legacy_vml_inspections
                ),
                legacy_vml_camera_source_range_count=sum(
                    inspection.camera_source_range_count
                    for inspection in legacy_vml_inspections
                ),
                control_macro_assignment_count=sum(
                    inspection.control_macro_assignment_count
                    for inspection in sheet_inspections
                )
                + sum(
                    inspection.macro_assignment_count
                    for inspection in legacy_vml_inspections
                ),
                control_cell_link_count=sum(
                    inspection.control_cell_link_count for inspection in sheet_inspections
                )
                + sum(
                    inspection.cell_link_count
                    for inspection in control_property_inspections
                )
                + sum(
                    inspection.cell_link_count for inspection in legacy_vml_inspections
                ),
                control_source_range_count=sum(
                    inspection.control_source_range_count
                    for inspection in sheet_inspections
                )
                + sum(
                    inspection.source_range_count
                    for inspection in control_property_inspections
                )
                + sum(
                    inspection.source_range_count
                    for inspection in legacy_vml_inspections
                ),
                form_control_formula_binding_count=sum(
                    inspection.formula_binding_count
                    for inspection in control_property_inspections
                ),
                ole_object_count=sum(
                    inspection.ole_object_count for inspection in sheet_inspections
                ),
                linked_ole_object_count=sum(
                    inspection.linked_ole_object_count for inspection in sheet_inspections
                ),
                auto_load_ole_object_count=sum(
                    inspection.auto_load_ole_object_count for inspection in sheet_inspections
                ),
                auto_update_ole_object_count=sum(
                    inspection.auto_update_ole_object_count
                    for inspection in sheet_inspections
                ),
                related_relationship_count=sum(
                    inspection.related_relationship_count for inspection in all_inspections
                )
                + len(active_legacy_vml_source_relationships),
                external_relationship_count=sum(
                    inspection.external_relationship_count for inspection in all_inspections
                )
                + sum(
                    relationship.target_mode.casefold() != "internal"
                    for relationship in active_legacy_vml_source_relationships
                ),
                internal_related_part_count=payload_inspection.internal_part_count,
                fingerprinted_related_part_count=payload_inspection.fingerprinted_part_count,
                uninspected_related_part_count=payload_inspection.uninspected_part_count,
                unrecognized_part_count=orphan_part_count
                + sum(inspection.unrecognized_count for inspection in all_inspections)
                + sum(
                    inspection.legacy_vml_unrecognized_count
                    for inspection in sheet_inspections
                ),
                declaration_signature=_private_external_data_signature(
                    tuple(declaration_entries)
                ),
                control_definition_signature=aggregate_signature(
                    list(sheet_inspections),
                    "definition_signature",
                ),
                active_x_definition_signature=aggregate_signature(
                    list(active_x_inspections),
                    "definition_signature",
                ),
                form_control_property_signature=aggregate_signature(
                    list(control_property_inspections),
                    "definition_signature",
                ),
                legacy_vml_definition_signature=aggregate_signature(
                    list(legacy_vml_inspections),
                    "definition_signature",
                ),
                legacy_vml_relationship_signature=aggregate_signature(
                    list(legacy_vml_inspections),
                    "relationship_signature",
                ),
                relationship_signature=aggregate_signature(
                    all_inspections,
                    "relationship_signature",
                ),
                related_part_payload_signature=payload_inspection.payload_signature,
            )
    except (BadZipFile, OSError, RuntimeError, ValueError) as error:
        return _WorksheetEmbeddedControlMetadata(
            default,
            (
                "FormulaFence could not inspect worksheet embedded-control OOXML "
                f"({type(error).__name__}); affected controls were not compared.",
            ),
        )
    return _WorksheetEmbeddedControlMetadata(snapshot, tuple(sorted(warnings)))


def _dynamic_metadata_indexes(archive: ZipFile) -> set[int]:
    """Return the one-based cell-metadata indexes marked as dynamic arrays."""
    try:
        metadata = _xml_root(archive, "xl/metadata.xml")
    except KeyError:
        return set()

    def tag(name: str) -> str:
        return f"{{{_SPREADSHEETML_NS}}}{name}"

    metadata_types = [
        metadata_type.get("name", "")
        for metadata_type in metadata.findall(f"./{tag('metadataTypes')}/{tag('metadataType')}")
    ]
    dynamic_future_indexes: dict[str, set[int]] = {}
    dynamic_properties_tag = f"{{{_DYNAMIC_ARRAY_NS}}}dynamicArrayProperties"
    for future_metadata in metadata.findall(tag("futureMetadata")):
        name = future_metadata.get("name")
        if not name:
            continue
        dynamic_indexes: set[int] = set()
        for index, metadata_block in enumerate(future_metadata.findall(tag("bk"))):
            if any(
                element.tag == dynamic_properties_tag
                and element.get("fDynamic", "").casefold() in {"1", "true"}
                for element in metadata_block.iter()
            ):
                dynamic_indexes.add(index)
        if dynamic_indexes:
            dynamic_future_indexes[name] = dynamic_indexes

    cell_metadata = metadata.find(tag("cellMetadata"))
    if cell_metadata is None:
        return set()
    dynamic_cell_indexes: set[int] = set()
    for cell_index, metadata_block in enumerate(cell_metadata.findall(tag("bk")), start=1):
        for record in metadata_block.findall(tag("rc")):
            try:
                type_index = int(record.get("t", ""))
                value_index = int(record.get("v", ""))
            except ValueError:
                continue
            if not 1 <= type_index <= len(metadata_types):
                continue
            type_name = metadata_types[type_index - 1]
            if value_index in dynamic_future_indexes.get(type_name, set()):
                dynamic_cell_indexes.add(cell_index)
                break
    return dynamic_cell_indexes


def _array_formula_metadata(path: Path) -> _ArrayFormulaMetadata:
    """Inspect raw markers that openpyxl intentionally does not expose."""
    dynamic_cells: set[CellKey] = set()
    unclassified_cells: set[CellKey] = set()
    scanned_sheets: set[str] = set()
    try:
        with ZipFile(path) as archive:
            dynamic_indexes = _dynamic_metadata_indexes(archive)
            for sheet, member in _worksheet_xml_paths(archive).items():
                worksheet = _xml_root(archive, member)
                scanned_sheets.add(sheet)
                for cell in worksheet.iter(f"{{{_SPREADSHEETML_NS}}}c"):
                    formula = cell.find(f"{{{_SPREADSHEETML_NS}}}f")
                    coordinate = cell.get("r")
                    if (
                        formula is None
                        or formula.get("t") != "array"
                        or not coordinate
                    ):
                        continue
                    try:
                        metadata_index = int(cell.get("cm", "0"))
                    except ValueError:
                        unclassified_cells.add((sheet, coordinate))
                        continue
                    if metadata_index <= 0:
                        continue
                    if metadata_index in dynamic_indexes:
                        dynamic_cells.add((sheet, coordinate))
                    else:
                        unclassified_cells.add((sheet, coordinate))
    except (BadZipFile, ElementTree.ParseError, OSError, ValueError) as error:
        return _ArrayFormulaMetadata(
            dynamic_cells=set(),
            unclassified_cells=set(),
            scanned_sheets=set(),
            complete=False,
            warnings=(
                "FormulaFence could not inspect array-formula OOXML metadata "
                f"({type(error).__name__}); fixed CSE output aliases were not traced.",
            ),
        )
    return _ArrayFormulaMetadata(
        dynamic_cells=dynamic_cells,
        unclassified_cells=unclassified_cells,
        scanned_sheets=scanned_sheets,
        complete=True,
        warnings=(),
    )


def _array_formula_range(sheet: str, cell: object) -> ArrayFormulaRange | None:
    """Return one canonical local output range when an array anchor is valid."""
    value = getattr(cell, "value", None)
    ref = getattr(value, "ref", None)
    if not isinstance(ref, str) or not ref.strip():
        return None
    try:
        min_column, min_row, max_column, max_row = range_boundaries(ref)
    except ValueError:
        return None
    if None in {min_column, min_row, max_column, max_row}:
        return None
    anchor = f"{get_column_letter(min_column)}{min_row}"
    if getattr(cell, "coordinate", "").upper() != anchor:
        return None
    endpoint = f"{get_column_letter(max_column)}{max_row}"
    canonical_ref = anchor if anchor == endpoint else f"{anchor}:{endpoint}"
    return ArrayFormulaRange(
        sheet=sheet,
        anchor=anchor,
        ref=canonical_ref,
        min_column=min_column,
        min_row=min_row,
        max_column=max_column,
        max_row=max_row,
    )


def _is_array_formula(cell: object) -> bool:
    """Return whether openpyxl exposed the cell as an OOXML array formula."""
    value = getattr(cell, "value", None)
    return (
        getattr(cell, "data_type", None) == "f"
        and isinstance(getattr(value, "ref", None), str)
        and _formula_text(value) is not None
    )


def _classify_array_formulas(
    workbook: object, metadata: _ArrayFormulaMetadata
) -> _ArrayFormulaClassification:
    """Classify array formulas without treating dynamic extents as fixed."""
    kinds: dict[CellKey, str] = {}
    refs: dict[CellKey, str] = {}
    legacy_ranges: list[ArrayFormulaRange] = []
    dynamic_cells: set[CellKey] = set()
    dynamic_ranges: list[ArrayFormulaRange] = []
    unclassified_cells: set[CellKey] = set()
    warnings = set(metadata.warnings)
    for worksheet in getattr(workbook, "worksheets", ()):
        for cell in worksheet._cells.values():  # noqa: SLF001 - sparse workbook safety
            if not _is_array_formula(cell):
                continue
            location = (worksheet.title, cell.coordinate)
            if not metadata.complete or worksheet.title not in metadata.scanned_sheets:
                kinds[location] = "unclassified"
                unclassified_cells.add(location)
                continue
            if location in metadata.dynamic_cells:
                kinds[location] = "dynamic"
                dynamic_cells.add(location)
                array_range = _array_formula_range(worksheet.title, cell)
                if array_range is None:
                    warnings.add(
                        "FormulaFence could not read one or more observed dynamic-array "
                        "output ranges; non-anchor output aliases were not traced for "
                        "those anchors."
                    )
                    continue
                refs[location] = array_range.ref
                dynamic_ranges.append(array_range)
                continue
            if location in metadata.unclassified_cells:
                kinds[location] = "unclassified"
                unclassified_cells.add(location)
                continue
            array_range = _array_formula_range(worksheet.title, cell)
            if array_range is None:
                kinds[location] = "unclassified"
                unclassified_cells.add(location)
                continue
            kinds[location] = "legacy_cse"
            refs[location] = array_range.ref
            legacy_ranges.append(array_range)

    if unclassified_cells:
        warnings.add(
            "FormulaFence could not classify one or more OOXML array formulas; "
            "fixed CSE output aliases were not traced for those cells."
        )
    return _ArrayFormulaClassification(
        kinds=kinds,
        refs=refs,
        legacy_ranges=tuple(legacy_ranges),
        dynamic_cells=dynamic_cells,
        dynamic_ranges=tuple(dynamic_ranges),
        unclassified_cells=unclassified_cells,
        warnings=tuple(sorted(warnings)),
    )


def _array_formula_output_dependents(
    legacy_ranges: tuple[ArrayFormulaRange, ...],
    reverse_dependencies: Mapping[CellKey, set[CellKey]],
    range_dependencies: list[RangeDependency],
) -> dict[CellKey, set[CellKey]]:
    """Link a CSE anchor to formulas that read any fixed result member.

    The output range is never expanded into millions of virtual cells. Instead
    each anchor receives direct aliases only to the already-known formula cells
    whose static references intersect that compact range.
    """
    ranges_by_sheet: dict[str, list[ArrayFormulaRange]] = defaultdict(list)
    for array_range in legacy_ranges:
        if array_range.has_multiple_outputs:
            ranges_by_sheet[array_range.sheet].append(array_range)

    dependents: dict[CellKey, set[CellKey]] = defaultdict(set)
    for source, formula_cells in reverse_dependencies.items():
        for array_range in ranges_by_sheet.get(source[0], ()):
            if array_range.contains(source):
                dependents[array_range.location].update(formula_cells)
    for dependency in range_dependencies:
        for array_range in ranges_by_sheet.get(dependency.source_sheet, ()):
            if dependency.intersects(array_range):
                dependents[array_range.location].add(dependency.dependent)
    return {anchor: cells for anchor, cells in dependents.items() if cells}


def _dynamic_array_output_dependents(
    dynamic_ranges: tuple[ArrayFormulaRange, ...],
    reverse_dependencies: Mapping[CellKey, set[CellKey]],
    range_dependencies: list[RangeDependency],
) -> tuple[
    dict[CellKey, set[CellKey]],
    dict[CellKey, tuple[DynamicArrayOutputReference, ...]],
]:
    """Link observed dynamic spill members to their current formula consumers.

    Dynamic-array output ranges may resize during recalculation, so this is not
    a fixed-range assertion. It records only a formula that currently reads at
    least one non-anchor member, without materializing individual spill cells.
    """
    ranges_by_sheet: dict[str, list[ArrayFormulaRange]] = defaultdict(list)
    for array_range in dynamic_ranges:
        if array_range.has_multiple_outputs:
            ranges_by_sheet[array_range.sheet].append(array_range)

    dependents: dict[CellKey, set[CellKey]] = defaultdict(set)
    references: dict[CellKey, set[DynamicArrayOutputReference]] = defaultdict(set)

    def add_reference(array_range: ArrayFormulaRange, dependent: CellKey) -> None:
        if dependent == array_range.location:
            return
        dependents[array_range.location].add(dependent)
        references[dependent].add(
            DynamicArrayOutputReference(
                anchor=array_range.location,
                observed_ref=array_range.ref,
            )
        )

    for source, formula_cells in reverse_dependencies.items():
        for array_range in ranges_by_sheet.get(source[0], ()):
            if source != array_range.location and array_range.contains(source):
                for dependent in formula_cells:
                    add_reference(array_range, dependent)
    for dependency in range_dependencies:
        for array_range in ranges_by_sheet.get(dependency.source_sheet, ()):
            if array_range.intersects_non_anchor(dependency):
                add_reference(array_range, dependency.dependent)

    return (
        {anchor: cells for anchor, cells in dependents.items() if cells},
        {
            dependent: tuple(
                sorted(
                    values,
                    key=lambda reference: (
                        reference.anchor[0].casefold(),
                        reference.anchor[1],
                        reference.observed_ref,
                    ),
                )
            )
            for dependent, values in references.items()
            if values
        },
    )


def _formula_text(value: object) -> str | None:
    if isinstance(value, str):
        return value
    # openpyxl represents array formulas as an object in recent versions.
    text = getattr(value, "text", None)
    return text if isinstance(text, str) else None


def _cell_snapshot(
    sheet: str,
    cell: object,
    *,
    array_formula_kind: str | None = None,
    array_formula_ref: str | None = None,
) -> CellSnapshot:
    coordinate = cell.coordinate
    value = cell.value
    data_type = cell.data_type
    formula = _formula_text(value) if data_type == "f" else None
    if formula is not None:
        return CellSnapshot(
            sheet=sheet,
            coordinate=coordinate,
            cell_type="formula",
            value=formula,
            value_type="formula",
            formula=formula,
            formula_fingerprint=formula_fingerprint(formula, coordinate),
            array_formula_kind=array_formula_kind,
            array_formula_ref=array_formula_ref,
        )
    cell_type = "error" if data_type == "e" else "value"
    return CellSnapshot(
        sheet=sheet,
        coordinate=coordinate,
        cell_type=cell_type,
        value=json_safe_value(value),
        value_type=type(value).__name__,
    )


def _calculation_settings(workbook: object) -> dict[str, object]:
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return {}
    return {
        field: value
        for field in _CALCULATION_FIELDS
        if (value := getattr(calculation, field, None)) is not None
    }


def _definition_text(definition: object) -> str:
    attr_text = getattr(definition, "attr_text", None)
    return str(attr_text if attr_text is not None else definition)


def _defined_names(workbook: object) -> dict[str, str]:
    """Inventory workbook and sheet-scoped names with unambiguous local keys."""
    names = getattr(workbook, "defined_names", {})
    result: dict[str, str] = {}
    try:
        items = names.items()
    except AttributeError:
        items = ()
    for name, definition in items:
        if getattr(definition, "localSheetId", None) is None:
            result[str(name)] = _definition_text(definition)
    for worksheet in getattr(workbook, "worksheets", ()):
        worksheet_names = getattr(worksheet, "defined_names", {})
        try:
            worksheet_items = worksheet_names.items()
        except AttributeError:
            continue
        for name, definition in worksheet_items:
            result[f"{worksheet.title}!{name}"] = _definition_text(definition)
    return result


def _named_destination_reference(
    name: str, sheet: str, coordinate: str
) -> ParsedReference | None:
    """Convert one workbook-defined-name destination into a static range."""
    parsed = parse_reference_token(f"{sheet}!{coordinate}")
    if parsed is None:
        return None
    return ParsedReference(
        parsed.sheet,
        parsed.min_column,
        parsed.min_row,
        parsed.max_column,
        parsed.max_row,
        raw=name,
        is_external=parsed.is_external,
    )


def _definition_references(name: str, definition: object) -> tuple[ParsedReference, ...]:
    """Return only static destinations from an ordinary defined name."""
    try:
        destinations = list(definition.destinations)
    except Exception:  # pragma: no cover - malformed name syntax is workbook-specific
        return ()
    return tuple(
        reference
        for sheet, coordinate in destinations
        if (reference := _named_destination_reference(name, sheet, coordinate)) is not None
    )


@dataclass(frozen=True)
class _FormulaDefinedName:
    """A defined name whose OOXML text is an Excel formula expression."""

    key: str
    formula: str
    scope: str | None


def _formula_defined_name(
    name: str, definition: object, scope: str | None
) -> _FormulaDefinedName | None:
    """Return a formula-valued defined name, excluding ordinary destinations.

    OOXML stores the text of a defined-name formula without its leading equals
    sign, while older writer output commonly retains one. ``DefinedName.type``
    lets us preserve ordinary A1 destinations and normalize either formula
    spelling before sending it through the formula tokenizer.
    """
    attr_text = getattr(definition, "attr_text", None)
    if not isinstance(attr_text, str):
        return None
    formula = attr_text.strip()
    if not formula:
        return None
    if not formula.startswith("="):
        try:
            definition_type = getattr(definition, "type", None)
        except Exception:  # pragma: no cover - malformed name text is workbook-specific
            definition_type = None
        if definition_type == "RANGE":
            return None
        formula = f"={formula}"
    return _FormulaDefinedName(reference_lookup_key(name), formula, scope)


def _qualified_name_key(sheet: str, name: str) -> str:
    """Build a lookup key for a sheet-local name without quote ambiguity."""
    escaped_sheet = sheet.replace("'", "''")
    return reference_lookup_key(f"'{escaped_sheet}'!{name}")


def _named_reference_maps(
    workbook: object,
    structured_tables: Mapping[str, StructuredTable],
    sheet_order: tuple[str, ...],
) -> tuple[
    dict[str, tuple[ParsedReference, ...]],
    dict[str, dict[str, tuple[ParsedReference, ...]]],
    dict[str, tuple[ParsedReference, ...] | None],
    dict[str, dict[str, tuple[ParsedReference, ...] | None]],
]:
    """Build direct, formula-defined, and callable named-LAMBDA maps.

    Formula-valued names are expanded only when every dependency is statically
    visible and internal. Relative references, dynamic functions, unresolved
    tokens, recursive LAMBDAs, external links, and 3-D spans remain unresolved
    at a use site instead of producing a guessed graph edge.
    """
    workbook_names = getattr(workbook, "defined_names", {})
    global_references: dict[str, tuple[ParsedReference, ...]] = {}
    global_formulas: dict[str, _FormulaDefinedName] = {}
    global_lambdas: dict[str, _FormulaDefinedName] = {}
    try:
        workbook_items = workbook_names.items()
    except AttributeError:
        workbook_items = ()
    for name, definition in workbook_items:
        if getattr(definition, "localSheetId", None) is not None:
            continue
        formula_definition = _formula_defined_name(str(name), definition, None)
        if formula_definition is not None:
            target = (
                global_lambdas
                if lambda_parameter_count(formula_definition.formula) is not None
                else global_formulas
            )
            target[formula_definition.key] = formula_definition
            continue
        references = _definition_references(str(name), definition)
        if references:
            global_references[reference_lookup_key(str(name))] = references

    local_references: dict[str, dict[str, tuple[ParsedReference, ...]]] = {}
    local_formulas: dict[str, dict[str, _FormulaDefinedName]] = {}
    local_lambdas: dict[str, dict[str, _FormulaDefinedName]] = {}
    sheet_titles: dict[str, str] = {}
    for worksheet in getattr(workbook, "worksheets", ()):
        scope = worksheet.title.casefold()
        sheet_titles[scope] = worksheet.title
        worksheet_names = getattr(worksheet, "defined_names", {})
        try:
            worksheet_items = worksheet_names.items()
        except AttributeError:
            continue
        sheet_references: dict[str, tuple[ParsedReference, ...]] = {}
        for name, definition in worksheet_items:
            formula_definition = _formula_defined_name(str(name), definition, scope)
            if formula_definition is not None:
                target = (
                    local_lambdas
                    if lambda_parameter_count(formula_definition.formula) is not None
                    else local_formulas
                )
                target.setdefault(scope, {})[formula_definition.key] = formula_definition
                continue
            references = _definition_references(str(name), definition)
            if not references:
                continue
            sheet_references[reference_lookup_key(str(name))] = references
        if sheet_references:
            local_references[scope] = sheet_references

    qualified_formulas = {
        _qualified_name_key(sheet_titles[scope], key): definition
        for scope, definitions in local_formulas.items()
        for key, definition in definitions.items()
    }
    qualified_lambdas = {
        _qualified_name_key(sheet_titles[scope], key): definition
        for scope, definitions in local_lambdas.items()
        for key, definition in definitions.items()
    }
    resolved_definitions: dict[tuple[str | None, str], tuple[ParsedReference, ...]] = {}
    resolving: set[tuple[str | None, str]] = set()
    failed: set[tuple[str | None, str]] = set()

    def identity_for(definition: _FormulaDefinedName) -> tuple[str | None, str]:
        return definition.scope, definition.key

    def cached_references(
        definition: _FormulaDefinedName,
    ) -> tuple[ParsedReference, ...] | None:
        return resolved_definitions.get(identity_for(definition))

    def has_cached_references(definition: _FormulaDefinedName) -> bool:
        return identity_for(definition) in resolved_definitions

    def visible_references(scope: str | None) -> dict[str, tuple[ParsedReference, ...]]:
        references = dict(global_references)
        for key, definition in global_formulas.items():
            if cached := cached_references(definition):
                references[key] = cached
            elif has_cached_references(definition):
                references[key] = ()
        for local_scope, values in local_references.items():
            for key, local_values in values.items():
                references[_qualified_name_key(sheet_titles[local_scope], key)] = local_values
        for local_scope, definitions in local_formulas.items():
            for key, definition in definitions.items():
                if cached := cached_references(definition):
                    references[_qualified_name_key(sheet_titles[local_scope], key)] = cached
                elif has_cached_references(definition):
                    references[_qualified_name_key(sheet_titles[local_scope], key)] = ()
        if scope is not None:
            references.update(local_references.get(scope, {}))
            for key, definition in local_formulas.get(scope, {}).items():
                if cached := cached_references(definition):
                    references[key] = cached
                elif has_cached_references(definition):
                    references[key] = ()
        return references

    def visible_named_functions(
        scope: str | None,
    ) -> dict[str, tuple[ParsedReference, ...] | None]:
        functions: dict[str, tuple[ParsedReference, ...] | None] = {
            key: cached_references(definition)
            for key, definition in global_lambdas.items()
        }
        for local_scope, definitions in local_lambdas.items():
            for key, definition in definitions.items():
                functions[_qualified_name_key(sheet_titles[local_scope], key)] = (
                    cached_references(definition)
                )
        if scope is not None:
            for key, definition in local_lambdas.get(scope, {}).items():
                functions[key] = cached_references(definition)
        return functions

    def named_definition_for(
        token: str, scope: str | None
    ) -> _FormulaDefinedName | None:
        key = reference_lookup_key(token)
        if qualified := qualified_formulas.get(key):
            return qualified
        if qualified := qualified_lambdas.get(key):
            return qualified
        if "!" in key:
            return None
        if scope is not None and (local := local_formulas.get(scope, {}).get(key)):
            return local
        if scope is not None and (local := local_lambdas.get(scope, {}).get(key)):
            return local
        return global_formulas.get(key) or global_lambdas.get(key)

    def resolve_definition(
        definition: _FormulaDefinedName,
    ) -> tuple[ParsedReference, ...] | None:
        identity = identity_for(definition)
        if has_cached_references(definition):
            return cached_references(definition)
        if identity in failed or identity in resolving:
            return None

        resolving.add(identity)
        resolved_references: tuple[ParsedReference, ...] | None = None
        try:
            inspection = inspect_formula(
                definition.formula,
                named_references=visible_references(definition.scope),
                structured_tables=structured_tables,
                sheet_order=sheet_order,
                named_function_references=visible_named_functions(definition.scope),
            )
            can_expand = not (
                has_broken_reference(definition.formula)
                or inspection.tokenization_failed
                or inspection.dynamic_reference_functions
                or inspection.three_d_reference_tokens
                or inspection.spill_reference_tokens
                or inspection.implicit_intersection_tokens
            )
            if can_expand:
                for token in inspection.unresolved_range_tokens:
                    dependency = named_definition_for(token, definition.scope)
                    if dependency is None or resolve_definition(dependency) is None:
                        can_expand = False
                        break
            if can_expand:
                inspection = inspect_formula(
                    definition.formula,
                    named_references=visible_references(definition.scope),
                    structured_tables=structured_tables,
                    sheet_order=sheet_order,
                    named_function_references=visible_named_functions(definition.scope),
                )
                if (
                    inspection.unresolved_range_tokens
                    or inspection.tokenization_failed
                    or inspection.dynamic_reference_functions
                    or inspection.three_d_reference_tokens
                    or inspection.spill_reference_tokens
                    or inspection.implicit_intersection_tokens
                    or any(
                        reference.is_external or reference.sheet is None
                        for reference in inspection.references
                    )
                ):
                    can_expand = False
            if can_expand:
                resolved_references = tuple(dict.fromkeys(inspection.references))
        finally:
            resolving.remove(identity)

        if resolved_references is None:
            failed.add(identity)
            return None
        resolved_definitions[identity] = resolved_references
        return resolved_references

    for definition in global_formulas.values():
        resolve_definition(definition)
    for definition in global_lambdas.values():
        resolve_definition(definition)
    for definitions in local_formulas.values():
        for definition in definitions.values():
            resolve_definition(definition)
    for definitions in local_lambdas.values():
        for definition in definitions.values():
            resolve_definition(definition)

    global_result = dict(global_references)
    for key, definition in global_formulas.items():
        if cached := cached_references(definition):
            global_result[key] = cached
        elif has_cached_references(definition):
            global_result[key] = ()
    for scope, values in local_references.items():
        for key, references in values.items():
            global_result[_qualified_name_key(sheet_titles[scope], key)] = references
    for scope, definitions in local_formulas.items():
        for key, definition in definitions.items():
            if cached := cached_references(definition):
                global_result[_qualified_name_key(sheet_titles[scope], key)] = cached
            elif has_cached_references(definition):
                global_result[_qualified_name_key(sheet_titles[scope], key)] = ()

    local_result: dict[str, dict[str, tuple[ParsedReference, ...]]] = {}
    for scope in set(local_references) | set(local_formulas):
        references = dict(local_references.get(scope, {}))
        for key, definition in local_formulas.get(scope, {}).items():
            if cached := cached_references(definition):
                references[key] = cached
            elif has_cached_references(definition):
                references[key] = ()
        if references:
            local_result[scope] = references

    global_function_result: dict[str, tuple[ParsedReference, ...] | None] = {
        key: cached_references(definition)
        for key, definition in global_lambdas.items()
    }
    for scope, definitions in local_lambdas.items():
        for key, definition in definitions.items():
            global_function_result[_qualified_name_key(sheet_titles[scope], key)] = (
                cached_references(definition)
            )

    local_function_result: dict[
        str, dict[str, tuple[ParsedReference, ...] | None]
    ] = {}
    for scope, definitions in local_lambdas.items():
        functions = {
            key: cached_references(definition) for key, definition in definitions.items()
        }
        if functions:
            local_function_result[scope] = functions
    return global_result, local_result, global_function_result, local_function_result


def _table_columns(
    worksheet: object,
    table: object,
    min_column: int,
    min_row: int,
    max_column: int,
    header_row_count: int,
) -> tuple[str, ...]:
    """Read table column labels, falling back to the inspectable header cells."""
    table_columns = tuple(getattr(table, "tableColumns", ()) or ())
    names = tuple(str(getattr(column, "name", "")) for column in table_columns)
    width = max_column - min_column + 1
    if len(names) == width and all(names):
        return names
    if header_row_count:
        return tuple(
            str(value) if (value := worksheet.cell(min_row, column).value) is not None else ""
            for column in range(min_column, max_column + 1)
        )
    return ()


def _table_snapshots(workbook: object) -> dict[str, TableSnapshot]:
    """Inventory Excel-table definitions that affect structured references."""
    result: dict[str, TableSnapshot] = {}
    for worksheet in getattr(workbook, "worksheets", ()):
        table_list = getattr(worksheet, "tables", {})
        try:
            table_values = table_list.values()
        except AttributeError:
            continue
        for table in table_values:
            name = str(
                getattr(table, "displayName", None)
                or getattr(table, "name", None)
                or ""
            )
            ref = getattr(table, "ref", None)
            if not name or not isinstance(ref, str):
                continue
            try:
                min_column, min_row, max_column, max_row = range_boundaries(ref)
            except ValueError:
                continue
            height = max_row - min_row + 1
            header_rows = min(max(int(getattr(table, "headerRowCount", 1) or 0), 0), height)
            totals_rows = min(
                max(int(getattr(table, "totalsRowCount", 0) or 0), 0), height - header_rows
            )
            result[name] = TableSnapshot(
                name=name,
                sheet=worksheet.title,
                ref=ref,
                columns=_table_columns(
                    worksheet,
                    table,
                    min_column,
                    min_row,
                    max_column,
                    header_rows,
                ),
                header_row_count=header_rows,
                totals_row_count=totals_rows,
            )
    return result


def _validation_formula(value: object) -> str | None:
    """Normalize the optional leading equals sign used by workbook writers."""
    if not isinstance(value, str):
        return None
    return value[1:] if value.startswith("=") else value


def _validation_ranges(validation: object) -> tuple[str, ...]:
    """Return a compact, writer-order-independent validation target inventory."""
    sqref = getattr(validation, "sqref", None)
    try:
        ranges = tuple(str(target_range) for target_range in sqref.ranges)
    except (AttributeError, TypeError):
        raw = str(sqref or "")
        ranges = tuple(raw.split())
    return tuple(sorted(set(ranges), key=str.casefold))


def _validation_bool(validation: object, attribute: str, default: bool) -> bool:
    """Read one OOXML boolean while applying its schema default."""
    value = getattr(validation, attribute, None)
    return default if value is None else bool(value)


def _data_validation_snapshots(workbook: object) -> tuple[DataValidationSnapshot, ...]:
    """Inventory data-entry controls without expanding their applied ranges.

    OOXML defaults are normalized because writers commonly omit values such as
    ``operator=between`` and ``errorStyle=stop`` while other writers serialize
    them explicitly. The control's formulas and messages remain available in a
    local diff but are omitted from profile output.
    """
    snapshots: list[DataValidationSnapshot] = []
    for worksheet in getattr(workbook, "worksheets", ()):
        validation_container = getattr(worksheet, "data_validations", None)
        validation_list = getattr(validation_container, "dataValidation", ())
        prompts_disabled = _validation_bool(validation_container, "disablePrompts", False)
        for validation in validation_list or ():
            ranges = _validation_ranges(validation)
            if not ranges:
                continue
            snapshots.append(
                DataValidationSnapshot(
                    sheet=worksheet.title,
                    ranges=ranges,
                    validation_type=str(getattr(validation, "type", None) or "none"),
                    operator=str(getattr(validation, "operator", None) or "between"),
                    formula1=_validation_formula(getattr(validation, "formula1", None)),
                    formula2=_validation_formula(getattr(validation, "formula2", None)),
                    allow_blank=_validation_bool(validation, "allowBlank", False),
                    dropdown_hidden=_validation_bool(validation, "showDropDown", False),
                    prompts_disabled=prompts_disabled,
                    show_input_message=_validation_bool(
                        validation, "showInputMessage", False
                    ),
                    show_error_message=_validation_bool(
                        validation, "showErrorMessage", False
                    ),
                    error_style=str(getattr(validation, "errorStyle", None) or "stop"),
                    error_title=getattr(validation, "errorTitle", None),
                    error=getattr(validation, "error", None),
                    prompt_title=getattr(validation, "promptTitle", None),
                    prompt=getattr(validation, "prompt", None),
                    ime_mode=str(getattr(validation, "imeMode", None) or "noControl"),
                )
            )
    # ``sqref`` is only a compact list of targets for a rule. Writers may keep
    # identical rules together or split them into separate ``dataValidation``
    # elements, with no change to what Excel applies to any target cell.
    # Canonicalize that grouping before comparing workbooks.
    grouped_ranges: dict[DataValidationSnapshot, set[str]] = defaultdict(set)
    for snapshot in snapshots:
        grouped_ranges[replace(snapshot, ranges=())].update(snapshot.ranges)
    return tuple(
        sorted(
            (
                replace(
                    snapshot,
                    ranges=tuple(sorted(ranges, key=str.casefold)),
                )
                for snapshot, ranges in grouped_ranges.items()
            ),
            key=DataValidationSnapshot.sort_key,
        )
    )


def _structured_table_map(tables: dict[str, TableSnapshot]) -> dict[str, StructuredTable]:
    """Translate stable table inventory records into formula-resolution metadata."""
    return {
        name.casefold(): StructuredTable(
            name=table.name,
            sheet=table.sheet,
            ref=table.ref,
            columns=table.columns,
            header_row_count=table.header_row_count,
            totals_row_count=table.totals_row_count,
        )
        for name, table in tables.items()
    }


def load_snapshot(path: str | Path) -> WorkbookSnapshot:
    """Load a workbook as a semantic snapshot without evaluating its contents."""
    source = Path(path)
    if not source.exists() or not source.is_file():
        raise WorkbookLoadError(f"Workbook does not exist or is not a file: {source}")
    if source.suffix.lower() not in _SUPPORTED_SUFFIXES:
        supported = ", ".join(sorted(_SUPPORTED_SUFFIXES))
        raise WorkbookLoadError(
            f"Unsupported workbook type {source.suffix!r}; supported types: {supported}"
        )

    xlm_macro_metadata = _xlm_macro_metadata(source)
    ribbon_customization_metadata = _ribbon_customization_metadata(source)
    office_web_addin_metadata = _office_web_addin_metadata(source)
    pivot_table_metadata = _pivot_table_metadata(source)
    chart_definition_metadata = _chart_definition_metadata(source)
    worksheet_embedded_control_metadata = _worksheet_embedded_control_metadata(source)
    reader_source, temporary_reader_source, reader_source_warnings = _openpyxl_safe_source(
        source
    )
    try:
        with warnings.catch_warnings(record=True) as caught_warnings:
            warnings.simplefilter("always")
            workbook = load_workbook(
                reader_source,
                read_only=False,
                data_only=False,
                keep_vba=False,
                keep_links=False,
                rich_text=False,
            )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as error:
        raise WorkbookLoadError(f"Could not read workbook {source}: {error}") from error
    finally:
        if temporary_reader_source is not None:
            try:
                temporary_reader_source.unlink(missing_ok=True)
            except OSError:
                pass
    parser_warnings = {str(warning.message) for warning in caught_warnings}
    parser_warnings.update(reader_source_warnings)
    parser_warnings.update(xlm_macro_metadata.warnings)
    parser_warnings.update(ribbon_customization_metadata.warnings)
    parser_warnings.update(office_web_addin_metadata.warnings)
    parser_warnings.update(pivot_table_metadata.warnings)
    parser_warnings.update(chart_definition_metadata.warnings)
    parser_warnings.update(worksheet_embedded_control_metadata.warnings)
    has_array_formulas = any(
        _is_array_formula(cell)
        for worksheet in workbook.worksheets
        for cell in worksheet._cells.values()  # noqa: SLF001 - sparse workbook safety
    )
    if has_array_formulas:
        array_formula_classification = _classify_array_formulas(
            workbook, _array_formula_metadata(source)
        )
    else:
        array_formula_classification = _ArrayFormulaClassification(
            kinds={},
            refs={},
            legacy_ranges=(),
            dynamic_cells=set(),
            dynamic_ranges=(),
            unclassified_cells=set(),
            warnings=(),
        )
    parser_warnings.update(array_formula_classification.warnings)
    conditional_formatting_metadata = _conditional_formatting_metadata(source)
    parser_warnings.update(conditional_formatting_metadata.warnings)
    protection_metadata = _protection_metadata(source)
    parser_warnings.update(protection_metadata.warnings)
    external_data_metadata = _external_data_metadata(source)
    parser_warnings.update(external_data_metadata.warnings)

    sheets: dict[str, SheetSnapshot] = {}
    cells: dict[CellKey, CellSnapshot] = {}
    reverse_dependencies: dict[CellKey, set[CellKey]] = defaultdict(set)
    range_dependencies: list[RangeDependency] = []
    external_references: set[CellKey] = set()
    broken_references: set[CellKey] = set()
    unresolved_reference_tokens: dict[CellKey, tuple[str, ...]] = {}
    dynamic_reference_functions: dict[CellKey, tuple[str, ...]] = {}
    three_d_reference_tokens: dict[CellKey, tuple[str, ...]] = {}
    spill_reference_tokens: dict[CellKey, tuple[str, ...]] = {}
    implicit_intersection_tokens: dict[CellKey, tuple[str, ...]] = {}
    tokenization_failure_cells: set[CellKey] = set()
    tables = _table_snapshots(workbook)
    data_validations = _data_validation_snapshots(workbook)
    structured_tables = _structured_table_map(tables)
    sheet_order = tuple(worksheet.title for worksheet in workbook.worksheets)
    (
        global_named_references,
        local_named_references,
        global_named_functions,
        local_named_functions,
    ) = _named_reference_maps(workbook, structured_tables, sheet_order)

    for worksheet in workbook.worksheets:
        named_references = {
            **global_named_references,
            **local_named_references.get(worksheet.title.casefold(), {}),
        }
        named_functions = {
            **global_named_functions,
            **local_named_functions.get(worksheet.title.casefold(), {}),
        }
        nonempty_cells = 0
        formula_cells = 0
        # _cells lets us avoid traversing a sheet's whole used rectangle when a
        # workbook has a sparse, accidentally enormous dimension.
        worksheet_cells = sorted(
            worksheet._cells.values(),  # noqa: SLF001 - needed for sparse workbook safety
            key=lambda current: (current.row, current.column),
        )
        for cell in worksheet_cells:
            if cell.value is None:
                continue
            location = (worksheet.title, cell.coordinate)
            snapshot = _cell_snapshot(
                worksheet.title,
                cell,
                array_formula_kind=array_formula_classification.kinds.get(location),
                array_formula_ref=array_formula_classification.refs.get(location),
            )
            cells[snapshot.location] = snapshot
            nonempty_cells += 1
            if not snapshot.is_formula or snapshot.formula is None:
                continue

            formula_cells += 1
            if has_broken_reference(snapshot.formula):
                broken_references.add(snapshot.location)
            inspection = inspect_formula(
                snapshot.formula,
                named_references=named_references,
                structured_tables=structured_tables,
                origin=snapshot.location,
                sheet_order=sheet_order,
                named_function_references=named_functions,
            )
            if inspection.unresolved_range_tokens:
                unresolved_reference_tokens[snapshot.location] = inspection.unresolved_range_tokens
            if inspection.tokenization_failed:
                tokenization_failure_cells.add(snapshot.location)
            if inspection.dynamic_reference_functions:
                dynamic_reference_functions[snapshot.location] = (
                    inspection.dynamic_reference_functions
                )
            if inspection.three_d_reference_tokens:
                three_d_reference_tokens[snapshot.location] = inspection.three_d_reference_tokens
            if inspection.spill_reference_tokens:
                spill_reference_tokens[snapshot.location] = inspection.spill_reference_tokens
            if inspection.implicit_intersection_tokens:
                implicit_intersection_tokens[snapshot.location] = (
                    inspection.implicit_intersection_tokens
                )
            for reference in inspection.references:
                if reference.is_external:
                    external_references.add(snapshot.location)
                    continue
                if None in {
                    reference.min_column,
                    reference.min_row,
                    reference.max_column,
                    reference.max_row,
                }:
                    continue
                source_sheet = reference.sheet or worksheet.title
                if reference.is_range:
                    range_dependencies.append(
                        RangeDependency(
                            source_sheet=source_sheet,
                            min_column=reference.min_column,
                            min_row=reference.min_row,
                            max_column=reference.max_column,
                            max_row=reference.max_row,
                            dependent=snapshot.location,
                        )
                    )
                else:
                    source_coordinate = (
                        f"{get_column_letter(reference.min_column)}{reference.min_row}"
                    )
                    reverse_dependencies[(source_sheet, source_coordinate)].add(snapshot.location)

        sheets[worksheet.title] = SheetSnapshot(
            title=worksheet.title,
            state=worksheet.sheet_state,
            nonempty_cells=nonempty_cells,
            formula_cells=formula_cells,
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
        )

    legacy_array_formula_output_dependents = _array_formula_output_dependents(
        array_formula_classification.legacy_ranges,
        reverse_dependencies,
        range_dependencies,
    )
    (
        dynamic_array_formula_output_dependents,
        dynamic_array_output_references,
    ) = _dynamic_array_output_dependents(
        array_formula_classification.dynamic_ranges,
        reverse_dependencies,
        range_dependencies,
    )
    array_formula_output_dependents = {
        anchor: set(legacy_array_formula_output_dependents.get(anchor, set()))
        | dynamic_array_formula_output_dependents.get(anchor, set())
        for anchor in (
            set(legacy_array_formula_output_dependents)
            | set(dynamic_array_formula_output_dependents)
        )
    }

    return WorkbookSnapshot(
        path=source,
        sha256=sha256_file(source),
        file_type=source.suffix.lower().lstrip("."),
        sheets=sheets,
        cells=cells,
        reverse_dependencies=dict(reverse_dependencies),
        range_dependencies=range_dependencies,
        external_references=external_references,
        broken_references=broken_references,
        unresolved_reference_tokens=unresolved_reference_tokens,
        dynamic_reference_functions=dynamic_reference_functions,
        three_d_reference_tokens=three_d_reference_tokens,
        spill_reference_tokens=spill_reference_tokens,
        implicit_intersection_tokens=implicit_intersection_tokens,
        legacy_array_formula_ranges=array_formula_classification.legacy_ranges,
        dynamic_array_formula_cells=array_formula_classification.dynamic_cells,
        dynamic_array_formula_ranges=array_formula_classification.dynamic_ranges,
        dynamic_array_output_references=dynamic_array_output_references,
        unclassified_array_formula_cells=array_formula_classification.unclassified_cells,
        array_formula_output_dependents=array_formula_output_dependents,
        tokenization_failure_cells=tokenization_failure_cells,
        tables=tables,
        data_validations=data_validations,
        conditional_formatting=conditional_formatting_metadata.rules,
        conditional_formatting_extensions=conditional_formatting_metadata.extensions,
        workbook_protection=protection_metadata.workbook_protection,
        sheet_protections=protection_metadata.sheet_protections,
        protected_ranges=protection_metadata.protected_ranges,
        cell_protection_default=protection_metadata.cell_protection_default,
        cell_protection_assignments=protection_metadata.cell_protection_assignments,
        external_data_refresh_settings=external_data_metadata.refresh_settings,
        external_data_connections=external_data_metadata.connections,
        query_table_refresh_controls=external_data_metadata.query_tables,
        pivot_cache_refresh_controls=external_data_metadata.pivot_caches,
        external_link_packages=external_data_metadata.external_link_packages,
        xlm_macro_sheets=xlm_macro_metadata.macro_sheets,
        ribbon_customization=ribbon_customization_metadata.customization,
        office_web_addins=office_web_addin_metadata.addins,
        pivot_table_definitions=pivot_table_metadata.pivot_tables,
        chart_definitions=chart_definition_metadata.charts,
        worksheet_embedded_controls=worksheet_embedded_control_metadata.controls,
        power_query=external_data_metadata.power_query,
        sheet_order=sheet_order,
        defined_names=_defined_names(workbook),
        macro_hash=_vba_hash(source),
        calculation_settings=_calculation_settings(workbook),
        parser_warnings=tuple(sorted(parser_warnings)),
    )


def profile_snapshot(snapshot: WorkbookSnapshot) -> dict[str, object]:
    """Return a data-minimising inventory suitable for a safe review artifact."""
    return {
        "schema_version": "1.0",
        "workbook": snapshot.summary(),
        "sheets": [sheet.to_dict() for sheet in snapshot.sheets.values()],
        "tables": [
            snapshot.tables[name].to_dict()
            for name in sorted(snapshot.tables, key=str.casefold)
        ],
        "data_validations": [
            validation.profile_dict() for validation in snapshot.data_validations
        ],
        "conditional_formatting": [
            rule.profile_dict() for rule in snapshot.conditional_formatting
        ],
        "conditional_formatting_extensions": [
            extension.profile_dict()
            for extension in snapshot.conditional_formatting_extensions
        ],
        "workbook_protection": (
            snapshot.workbook_protection.to_dict()
            if snapshot.workbook_protection is not None
            else None
        ),
        "sheet_protections": [
            protection.profile_dict() for protection in snapshot.sheet_protections
        ],
        "protected_ranges": [
            protected_range.profile_dict()
            for protected_range in snapshot.protected_ranges
        ],
        "cell_protection_default": (
            snapshot.cell_protection_default.to_dict()
            if snapshot.cell_protection_default is not None
            else None
        ),
        "cell_protection_assignments": [
            assignment.profile_dict()
            for assignment in snapshot.cell_protection_assignments
        ],
        "external_data_refresh_settings": snapshot.external_data_refresh_settings.to_dict(),
        "external_data_connections": [
            connection.profile_dict() for connection in snapshot.external_data_connections
        ],
        "query_table_refresh_controls": [
            control.profile_dict() for control in snapshot.query_table_refresh_controls
        ],
        "pivot_cache_refresh_controls": [
            control.profile_dict() for control in snapshot.pivot_cache_refresh_controls
        ],
        "external_link_packages": snapshot.external_link_packages.profile_dict(),
        "xlm_macro_sheets": snapshot.xlm_macro_sheets.profile_dict(),
        "ribbon_customization": snapshot.ribbon_customization.profile_dict(),
        "office_web_addins": snapshot.office_web_addins.profile_dict(),
        "pivot_table_definitions": snapshot.pivot_table_definitions.profile_dict(),
        "chart_definitions": snapshot.chart_definitions.profile_dict(),
        "worksheet_embedded_controls": snapshot.worksheet_embedded_controls.profile_dict(),
        "power_query": snapshot.power_query.profile_dict(),
        "defined_names": sorted(snapshot.defined_names),
        "calculation_settings": snapshot.calculation_settings,
        "features": {
            "external_reference_cells": [
                f"{sheet}!{coordinate}"
                for sheet, coordinate in sorted(snapshot.external_references)
            ],
            "broken_reference_cells": [
                f"{sheet}!{coordinate}"
                for sheet, coordinate in sorted(snapshot.broken_references)
            ],
            "has_vba": snapshot.macro_hash is not None,
            "has_xlm_macro_sheets": snapshot.xlm_macro_sheets.present,
            "has_ribbon_customization": snapshot.ribbon_customization.present,
            "has_office_web_addins": snapshot.office_web_addins.present,
            "has_pivot_table_definitions": snapshot.pivot_table_definitions.present,
            "has_chart_definitions": snapshot.chart_definitions.present,
            "has_worksheet_embedded_controls": snapshot.worksheet_embedded_controls.present,
            "parser_warnings": list(snapshot.parser_warnings),
            "unresolved_reference_cells": [
                {
                    "location": display_location(location),
                    "tokens": list(tokens),
                }
                for location, tokens in sorted(snapshot.unresolved_reference_tokens.items())
            ],
            "dynamic_reference_cells": [
                {
                    "location": display_location(location),
                    "functions": list(functions),
                }
                for location, functions in sorted(snapshot.dynamic_reference_functions.items())
            ],
            "three_d_reference_cells": [
                {
                    "location": display_location(location),
                    "tokens": list(tokens),
                }
                for location, tokens in sorted(snapshot.three_d_reference_tokens.items())
            ],
            "spill_reference_cells": [
                {
                    "location": display_location(location),
                    "tokens": list(tokens),
                }
                for location, tokens in sorted(snapshot.spill_reference_tokens.items())
            ],
            "implicit_intersection_cells": [
                {
                    "location": display_location(location),
                    "tokens": list(tokens),
                }
                for location, tokens in sorted(
                    snapshot.implicit_intersection_tokens.items()
                )
            ],
            "legacy_array_formula_ranges": [
                array_range.to_dict()
                for array_range in sorted(
                    snapshot.legacy_array_formula_ranges,
                    key=lambda array_range: (array_range.sheet.casefold(), array_range.anchor),
                )
            ],
            "dynamic_array_formula_cells": [
                display_location(location)
                for location in sorted(snapshot.dynamic_array_formula_cells)
            ],
            "dynamic_array_observed_output_ranges": [
                array_range.to_dict()
                for array_range in sorted(
                    snapshot.dynamic_array_formula_ranges,
                    key=lambda array_range: (array_range.sheet.casefold(), array_range.anchor),
                )
            ],
            "dynamic_array_output_reference_cells": [
                {
                    "location": display_location(location),
                    "references": [reference.to_dict() for reference in references],
                }
                for location, references in sorted(
                    snapshot.dynamic_array_output_references.items()
                )
            ],
            "unclassified_array_formula_cells": [
                display_location(location)
                for location in sorted(snapshot.unclassified_array_formula_cells)
            ],
            "tokenization_failure_cells": [
                display_location(location)
                for location in sorted(snapshot.tokenization_failure_cells)
            ],
        },
    }
