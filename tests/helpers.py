from __future__ import annotations

import base64
import io
import struct
from collections.abc import Callable
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chartsheet.protection import ChartsheetProtection
from openpyxl.comments import Comment
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.table import Table

_DATA_MASHUP_NS = "http://schemas.microsoft.com/DataMashup"
_PACKAGE_RELATIONSHIPS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_DOCUMENT_RELATIONSHIPS_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_STRICT_DOCUMENT_RELATIONSHIPS_NS = (
    "http://purl.oclc.org/ooxml/officeDocument/relationships"
)
_CUSTOM_XML_DATA_PROPERTIES_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/customXmlDataProps"
)
_CUSTOM_DATA_PROPERTIES_NS = (
    "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
)
_CUSTOM_DOCUMENT_PROPERTIES_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
)
_DOCUMENT_PROPERTY_TYPES_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"
)
_SPREADSHEETML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_STRICT_SPREADSHEETML_NS = "http://purl.oclc.org/ooxml/spreadsheetml/main"
_CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
_XML_SCHEMA_NS = "http://www.w3.org/2001/XMLSchema"
_OFFICE_2010_SPREADSHEET_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
_OFFICE_2010_AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
_OFFICE_2013_SPREADSHEET_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
_OFFICE_2014_REVISION_NS = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
_OFFICE_2016_REVISION9_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision9"
_OFFICE_2016_REVISION10_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10"
_EXCEL_2006_MAIN_NS = "http://schemas.microsoft.com/office/excel/2006/main"
_NAMED_SHEET_VIEW_NS = "http://schemas.microsoft.com/office/spreadsheetml/2019/namedsheetviews"
_XML_MAP_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/xmlMaps"
_TABLE_SINGLE_CELLS_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/tableSingleCells"
_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIP = (
    f"{_DOCUMENT_RELATIONSHIPS_NS}/revisionHeaders"
)
_SHARED_WORKBOOK_REVISION_LOG_RELATIONSHIP = (
    f"{_DOCUMENT_RELATIONSHIPS_NS}/revisionLog"
)
_XML_DIGITAL_SIGNATURE_NS = "http://www.w3.org/2000/09/xmldsig#"
_DIGITAL_SIGNATURE_ORIGIN_RELATIONSHIP = (
    f"{_PACKAGE_RELATIONSHIPS_NS}/digital-signature/origin"
)
_DIGITAL_SIGNATURE_SIGNATURE_RELATIONSHIP = (
    f"{_PACKAGE_RELATIONSHIPS_NS}/digital-signature/signature"
)
_DIGITAL_SIGNATURE_CERTIFICATE_RELATIONSHIP = (
    f"{_PACKAGE_RELATIONSHIPS_NS}/digital-signature/certificate"
)
_VBA_PROJECT_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2006/relationships/vbaProject"
)
_VBA_PROJECT_SIGNATURE_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2006/relationships/vbaProjectSignature"
)
_RICH_DATA_NS = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"
_RICH_DATA_2_NS = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata2"
_RICH_DATA_WEB_IMAGE_NS = (
    "http://schemas.microsoft.com/office/spreadsheetml/2020/richdatawebimage"
)
_RICH_VALUE_REL_NS = (
    "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel"
)
_PYTHON_IN_EXCEL_NS = (
    "http://schemas.microsoft.com/office/spreadsheetml/2023/python"
)
_PYTHON_IN_EXCEL_CONTENT_TYPE = "application/vnd.ms-excel.python+xml"
_PYTHON_IN_EXCEL_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2023/09/relationships/Python"
)
_PYTHON_IN_EXCEL_SCRIPTS_NS = (
    "http://schemas.microsoft.com/office/spreadsheetml/2022/pythonscript"
)
_PYTHON_IN_EXCEL_SCRIPTS_CONTENT_TYPE = (
    "application/vnd.ms-excel.pythonscripts+xml"
)
_PYTHON_IN_EXCEL_SCRIPTS_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2022/03/relationships/PythonScripts"
)
_RICH_DATA_METADATA_EXTENSION_URI = "{3E2802C4-A4D2-4D8B-9148-E3BE6C30E623}"
_DRAWINGML_MAIN_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_DRAWINGML_STRICT_MAIN_NS = "http://purl.oclc.org/ooxml/drawingml/main"
_DRAWINGML_CHART_EX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
_MARKUP_COMPATIBILITY_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
_RICH_DATA_RELATIONSHIPS = {
    "rich-value-data": (
        "http://schemas.microsoft.com/office/2017/06/relationships/rdRichValue"
    ),
    "rich-value-structure": (
        "http://schemas.microsoft.com/office/2017/06/relationships/"
        "rdRichValueStructure"
    ),
    "rich-value-types": (
        "http://schemas.microsoft.com/office/2017/06/relationships/rdRichValueTypes"
    ),
    "rich-value-array": (
        "http://schemas.microsoft.com/office/2017/06/relationships/rdArray"
    ),
    "supporting-property-bag": (
        "http://schemas.microsoft.com/office/2017/06/relationships/"
        "rdSupportingPropertyBag"
    ),
    "supporting-property-bag-structure": (
        "http://schemas.microsoft.com/office/2017/06/relationships/"
        "rdSupportingPropertyBagStructure"
    ),
    "rich-styles": (
        "http://schemas.microsoft.com/office/2017/06/relationships/richStyles"
    ),
    "rich-value-web-image": (
        "http://schemas.microsoft.com/office/2020/07/relationships/"
        "rdRichValueWebImage"
    ),
    "rich-value-relationships": (
        "http://schemas.microsoft.com/office/2022/10/relationships/richValueRel"
    ),
}


def make_model(path: Path) -> Path:
    """Create a small, multi-sheet financial-model-shaped fixture."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Revenue"
    inputs["B2"] = 100
    inputs["B3"] = 150
    inputs["B4"] = 175

    model = workbook.create_sheet("Model")
    model["A1"] = "Calculated revenue"
    model["B2"] = "=Inputs!B2*2"
    model["B3"] = "=Inputs!B3*2"
    model["B4"] = "=Inputs!B4*2"
    model["C2"] = "=B2+10"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Headline output"
    dashboard["B12"] = "=Model!C2"

    control = workbook.create_sheet("Control")
    control["A1"] = "Do not hide or alter casually"
    workbook.defined_names.add(DefinedName("HeadlineOutput", attr_text="Dashboard!$B$12"))
    workbook.save(path)
    return path


def make_formula_external_action_model(path: Path) -> Path:
    """Create stored formulas with known external-action function surfaces."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Stored formula action controls"
    inputs["B2"] = '=HYPERLINK("https://private.example.test/PRIVATE-LINK-BASELINE", "Open")'
    inputs["B3"] = '=WEBSERVICE("https://private.example.test/PRIVATE-WEBSERVICE")'
    inputs["B4"] = '=IMAGE("https://private.example.test/PRIVATE-IMAGE.png", "Image")'
    inputs["B5"] = '=RTD("PRIVATE.FormulaFence.Provider", "PRIVATE-SERVER", "Topic")'
    inputs["B6"] = '=_xlfn.IMAGE("https://private.example.test/PRIVATE-NAMESPACED-IMAGE.png")'
    inputs["B7"] = '=HYPERLINK("#Inputs!A1", "Internal")'
    inputs["B8"] = (
        '=IF(TRUE,HYPERLINK("https://private.example.test/PRIVATE-SECOND-LINK", "One"),'
        'HYPERLINK("https://private.example.test/PRIVATE-THIRD-LINK", "Two"))'
    )
    inputs["A9"] = "https://private.example.test/PRIVATE-REFERENCED-LINK-BASELINE"
    inputs["B9"] = '=HYPERLINK(A9, "Referenced")'
    workbook.save(path)
    return path


def change_formula_external_action_target(path: Path) -> Path:
    """Change a formula action destination without changing its call inventory."""
    workbook = load_workbook(path)
    workbook["Inputs"]["B2"] = (
        '=HYPERLINK("https://private.example.test/PRIVATE-LINK-CANDIDATE", "Open")'
    )
    workbook.save(path)
    return path


def change_formula_external_action_input(path: Path) -> Path:
    """Change a statically referenced action input without editing its formula."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "https://private.example.test/PRIVATE-REFERENCED-LINK-CANDIDATE"
    workbook.save(path)
    return path


def make_formula_external_data_provider_model(path: Path) -> Path:
    """Create provider-backed formulas, including all documented Cube functions."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Formula external data-provider controls"
    inputs["A2"] = "XNAS:PRIVATE-STOCK-BASELINE"
    inputs["B2"] = (
        '=_xlfn.STOCKHISTORY(A2,DATE(2024,1,1),DATE(2024,1,31),0,0,1)'
    )
    inputs["B3"] = (
        '=CUBEVALUE("PRIVATE-CUBE-CONNECTION-BASELINE","[Measures].[PRIVATE-REVENUE]")'
    )
    inputs["B4"] = '=CUBEMEMBER("PRIVATE-CUBE-CONNECTION","[Date].[Calendar].&[2024]")'
    inputs["B5"] = (
        '=CUBEMEMBERPROPERTY("PRIVATE-CUBE-CONNECTION",B4,"PRIVATE-MEMBER-PROPERTY")'
    )
    inputs["B6"] = (
        '=CUBESET("PRIVATE-CUBE-CONNECTION","[Product].[Category].Members",'
        '"PRIVATE-CUBE-CAPTION")'
    )
    inputs["B7"] = '=CUBERANKEDMEMBER("PRIVATE-CUBE-CONNECTION",B6,1)'
    inputs["B8"] = "=CUBESETCOUNT(B6)"
    inputs["B9"] = '=CUBEKPIMEMBER("PRIVATE-CUBE-CONNECTION","PRIVATE-KPI",1)'
    inputs["B10"] = "=FENCE.MARKET(A2)"
    inputs["B11"] = "=FENCE.CHAIN(A2)"
    inputs["B12"] = "=FENCE.CUBE"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.MARKET",
            attr_text=(
                '=LAMBDA(stock,_xlfn.STOCKHISTORY(stock,DATE(2024,1,1),'
                "DATE(2024,1,31),0,0,1))"
            ),
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.CHAIN",
            attr_text="=LAMBDA(stock,FENCE.MARKET(stock))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.CUBE",
            attr_text=(
                '=CUBEVALUE("PRIVATE-NAMED-CUBE-CONNECTION-BASELINE",'
                '"[Measures].[PRIVATE-NAMED-REVENUE]")'
            ),
        )
    )
    workbook.save(path)
    return path


def change_formula_external_data_provider_target(path: Path) -> Path:
    """Retarget one stored Cube connection without changing its call inventory."""
    workbook = load_workbook(path)
    workbook["Inputs"]["B3"] = (
        '=CUBEVALUE("PRIVATE-CUBE-CONNECTION-CANDIDATE","[Measures].[PRIVATE-REVENUE]")'
    )
    workbook.save(path)
    return path


def change_formula_external_data_provider_definition(path: Path) -> Path:
    """Retarget a private formula-defined Cube query without moving callers."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.CUBE"]
    if definition.attr_text != (
        '=CUBEVALUE("PRIVATE-NAMED-CUBE-CONNECTION-BASELINE",'
        '"[Measures].[PRIVATE-NAMED-REVENUE]")'
    ):
        raise ValueError("Fixture does not contain the expected named Cube query")
    definition.attr_text = (
        '=CUBEVALUE("PRIVATE-NAMED-CUBE-CONNECTION-CANDIDATE",'
        '"[Measures].[PRIVATE-NAMED-REVENUE]")'
    )
    workbook.save(path)
    return path


def change_formula_external_data_provider_input(path: Path) -> Path:
    """Change a statically visible STOCKHISTORY input without editing a call."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A2"] = "XNAS:PRIVATE-STOCK-CANDIDATE"
    workbook.save(path)
    return path


def make_named_formula_external_action_model(path: Path) -> Path:
    """Create action calls reached through names and named LAMBDA bodies."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Named formula external-action controls"
    inputs["A9"] = "https://private.example.test/PRIVATE-NAMED-ACTION-INPUT-BASELINE"
    inputs["B2"] = "=FENCE.WRAPPER(A9)"
    inputs["B3"] = "=FENCE.DIRECT"
    inputs["B4"] = "=FENCE.CHAIN(A9)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.WRAPPER",
            attr_text=(
                '=LAMBDA(value,HYPERLINK(value,"PRIVATE-NAMED-ACTION-LABEL-BASELINE"))'
            ),
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.CHAIN",
            attr_text="=LAMBDA(value,FENCE.WRAPPER(value))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.DIRECT",
            attr_text="=WEBSERVICE(Inputs!$A$9)",
        )
    )
    workbook.save(path)
    return path


def change_named_formula_external_action_definition(path: Path) -> Path:
    """Rewrite a private action definition without changing its callers."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.WRAPPER"]
    if definition.attr_text != (
        '=LAMBDA(value,HYPERLINK(value,"PRIVATE-NAMED-ACTION-LABEL-BASELINE"))'
    ):
        raise ValueError("Fixture does not contain the expected named action")
    definition.attr_text = (
        '=LAMBDA(value,HYPERLINK(value,"PRIVATE-NAMED-ACTION-LABEL-CANDIDATE"))'
    )
    workbook.save(path)
    return path


def change_named_formula_external_action_input(path: Path) -> Path:
    """Change an input used through named formula-action definitions."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = (
        "https://private.example.test/PRIVATE-NAMED-ACTION-INPUT-CANDIDATE"
    )
    workbook.save(path)
    return path


def make_office_custom_function_model(path: Path) -> Path:
    """Create formula calls shaped like documented Office Add-in functions.

    The add-in manifest and JavaScript runtime intentionally remain outside this
    controlled workbook: FormulaFence can only inventory the stored namespaced
    calls. Native dotted formulas and a workbook-defined LAMBDA verify that the
    candidate scope does not treat every dotted callable as an Office Add-in.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Namespaced custom-function controls"
    inputs["A9"] = "PRIVATE-CUSTOM-FUNCTION-INPUT-BASELINE"
    inputs["A10"] = 45_000
    inputs["B2"] = (
        '=CONTOSO.GETMARKETDATA(A9,"PRIVATE-CUSTOM-FUNCTION-QUERY-BASELINE")'
    )
    inputs["B3"] = "=CONTOSO.ADD(A9,200)"
    inputs["B4"] = '=FENCE.GETRISK(A9,"PRIVATE-CUSTOM-FUNCTION-RISK-BASELINE")'
    inputs["B5"] = (
        '=CONTOSO.GETMARKETDATA(A9,"PRIVATE-CUSTOM-FUNCTION-SECOND-BASELINE")'
        "+CONTOSO.ADD(A9,1)"
    )
    inputs["B6"] = "=ECMA.CEILING(4.1,1)"
    inputs["B7"] = "=WORKDAY.INTL(A10,1)"
    inputs["B8"] = "=LOCAL.RATE(A9)"
    workbook.defined_names.add(
        DefinedName("LOCAL.RATE", attr_text="=LAMBDA(value,value)")
    )
    workbook.save(path)
    return path


def change_office_custom_function_call(path: Path) -> Path:
    """Change one custom-function call without moving public candidate counts."""
    workbook = load_workbook(path)
    formula = workbook["Inputs"]["B2"].value
    if not isinstance(formula, str) or "GETMARKETDATA" not in formula:
        raise ValueError("Fixture does not contain the expected custom-function call")
    workbook["Inputs"]["B2"] = formula.replace("GETMARKETDATA", "GETRISKDATA")
    workbook.save(path)
    return path


def change_office_custom_function_input(path: Path) -> Path:
    """Change a static source cell without rewriting a namespaced call."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-CUSTOM-FUNCTION-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_named_office_custom_function_model(path: Path) -> Path:
    """Create custom calls reached through names and named LAMBDA bodies.

    The chain deliberately exercises both forms of indirection. ``FENCE.CHAIN``
    requires fixed-point propagation through one named LAMBDA to reach the
    stored ``CONTOSO`` call, while ``FENCE.DIRECT`` is a formula-defined name
    with a statically resolvable source cell.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Named custom-function controls"
    inputs["A9"] = "PRIVATE-NAMED-CUSTOM-FUNCTION-INPUT-BASELINE"
    inputs["B2"] = "=FENCE.WRAPPER(A9)"
    inputs["B3"] = "=FENCE.DIRECT"
    inputs["B4"] = "=FENCE.CHAIN(A9)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.WRAPPER",
            attr_text="=LAMBDA(value,CONTOSO.GETDATA(value)+CONTOSO.GETDATA(value))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.CHAIN",
            attr_text="=LAMBDA(value,FENCE.WRAPPER(value))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.DIRECT",
            attr_text="=CONTOSO.GETDATA(Inputs!$A$9)",
        )
    )
    workbook.save(path)
    return path


def change_named_office_custom_function_definition(path: Path) -> Path:
    """Rewrite a hidden named-LAMBDA callable without changing its callers."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.WRAPPER"]
    if definition.attr_text != (
        "=LAMBDA(value,CONTOSO.GETDATA(value)+CONTOSO.GETDATA(value))"
    ):
        raise ValueError("Fixture does not contain the expected named custom call")
    definition.attr_text = (
        "=LAMBDA(value,CONTOSO.GETRISK(value)+CONTOSO.GETRISK(value))"
    )
    workbook.save(path)
    return path


def change_named_office_custom_function_input(path: Path) -> Path:
    """Change an input used through a named custom-function definition."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-NAMED-CUSTOM-FUNCTION-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_unqualified_runtime_function_model(path: Path) -> Path:
    """Create unknown bare calls without installing or invoking a provider.

    The workbook deliberately contains only stored formula text. It exercises
    the conservative candidate boundary while keeping a range of native and
    workbook-defined calls nearby as non-candidate controls.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Unqualified runtime-function controls"
    inputs["A9"] = "PRIVATE-RUNTIME-FUNCTION-INPUT-BASELINE"
    inputs["A10"] = 45_000
    inputs["B2"] = '=PRIVATEUDF(A9,"PRIVATE-RUNTIME-FUNCTION-QUERY-BASELINE")'
    inputs["B3"] = "=RISKUDF(A9)"
    inputs["B4"] = (
        '=PRIVATEUDF(A9,"PRIVATE-RUNTIME-FUNCTION-SECOND-BASELINE")+RISKUDF(A9)'
    )
    inputs["B5"] = "=SUM(A10)"
    inputs["B6"] = "=XLOOKUP(A10,A10,A10)"
    inputs["B7"] = "=VSTACK(A10,A10)"
    inputs["B8"] = '=FIELDVALUE(A10,"Price")'
    inputs["B9"] = "=LOCALFUN(A9)"
    inputs["B10"] = "=LET(LOCALUDF,LAMBDA(value,value),LOCALUDF(A9))"
    workbook.defined_names.add(
        DefinedName("LOCALFUN", attr_text="=LAMBDA(value,value)")
    )
    workbook.save(path)
    return path


def change_unqualified_runtime_function_call(path: Path) -> Path:
    """Rewrite an unknown callable without changing public candidate counts."""
    workbook = load_workbook(path)
    formula = workbook["Inputs"]["B2"].value
    if not isinstance(formula, str) or "PRIVATEUDF" not in formula:
        raise ValueError("Fixture does not contain the expected runtime function")
    workbook["Inputs"]["B2"] = formula.replace("PRIVATEUDF", "UPDATEDUDF")
    workbook.save(path)
    return path


def change_unqualified_runtime_function_input(path: Path) -> Path:
    """Change a static source without rewriting an unknown callable."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-RUNTIME-FUNCTION-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_named_unqualified_runtime_function_model(path: Path) -> Path:
    """Create unknown bare calls reached through names and named LAMBDAs."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Named unqualified runtime-function controls"
    inputs["A9"] = "PRIVATE-NAMED-RUNTIME-FUNCTION-INPUT-BASELINE"
    inputs["B2"] = "=FENCE.WRAPPER(A9)"
    inputs["B3"] = "=FENCE.DIRECT"
    inputs["B4"] = "=FENCE.CHAIN(A9)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.WRAPPER",
            attr_text="=LAMBDA(value,PRIVATEUDF(value)+PRIVATEUDF(value))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.CHAIN",
            attr_text="=LAMBDA(value,FENCE.WRAPPER(value))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.DIRECT",
            attr_text="=PRIVATEUDF(Inputs!$A$9)",
        )
    )
    workbook.save(path)
    return path


def change_named_unqualified_runtime_function_definition(path: Path) -> Path:
    """Rewrite a hidden unknown call without moving its callers."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.WRAPPER"]
    if definition.attr_text != "=LAMBDA(value,PRIVATEUDF(value)+PRIVATEUDF(value))":
        raise ValueError("Fixture does not contain the expected named runtime function")
    definition.attr_text = "=LAMBDA(value,UPDATEDUDF(value)+UPDATEDUDF(value))"
    workbook.save(path)
    return path


def change_named_unqualified_runtime_function_input(path: Path) -> Path:
    """Change a static source used through named runtime-function definitions."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-NAMED-RUNTIME-FUNCTION-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_worksheet_code_resource_registration_model(path: Path) -> Path:
    """Create inert stored ``REGISTER.ID`` expressions for boundary tests.

    The workbook is never opened in Excel. Private module/procedure sentinels
    exist solely to prove FormulaFence keeps code-resource registration material
    out of its public profile, findings, and SARIF output.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Worksheet code-resource registration controls"
    inputs["A9"] = "PRIVATE-REGISTRATION-MODULE-BASELINE"
    inputs["A10"] = "PRIVATE-REGISTRATION-PROCEDURE-BASELINE"
    inputs["B2"] = '=REGISTER.ID(A9,A10,"J!")'
    inputs["B3"] = (
        '=REGISTER.ID("PRIVATE-REGISTRATION-MODULE-LITERAL-BASELINE",'
        '"PRIVATE-REGISTRATION-PROCEDURE-LITERAL-BASELINE","J!")'
    )
    inputs["B4"] = '=REGISTER.ID(A9,A10,"J!")'
    workbook.save(path)
    return path


def change_worksheet_code_resource_registration_call(path: Path) -> Path:
    """Rewrite private registration material without changing public counts."""
    workbook = load_workbook(path)
    formula = workbook["Inputs"]["B3"].value
    if not isinstance(formula, str) or "LITERAL-BASELINE" not in formula:
        raise ValueError("Fixture does not contain the expected registration call")
    workbook["Inputs"]["B3"] = formula.replace("LITERAL-BASELINE", "LITERAL-CANDIDATE")
    workbook.save(path)
    return path


def change_worksheet_code_resource_registration_input(path: Path) -> Path:
    """Change a static module source without rewriting a registration formula."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-REGISTRATION-MODULE-CANDIDATE"
    workbook.save(path)
    return path


def make_named_worksheet_code_resource_registration_model(path: Path) -> Path:
    """Create registrations reached through formula names and named LAMBDAs."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Named worksheet code-resource registration controls"
    inputs["A9"] = "PRIVATE-NAMED-REGISTRATION-MODULE-BASELINE"
    inputs["A10"] = "PRIVATE-NAMED-REGISTRATION-PROCEDURE-BASELINE"
    inputs["B2"] = "=FENCE.REGISTER(A9,A10)"
    inputs["B3"] = "=FENCE.DIRECT"
    inputs["B4"] = "=FENCE.CHAIN(A9,A10)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.REGISTER",
            attr_text='=LAMBDA(module,procedure,REGISTER.ID(module,procedure,"J!"))',
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.CHAIN",
            attr_text="=LAMBDA(module,procedure,FENCE.REGISTER(module,procedure))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.DIRECT",
            attr_text='=REGISTER.ID(Inputs!$A$9,Inputs!$A$10,"J!")',
        )
    )
    workbook.save(path)
    return path


def change_named_worksheet_code_resource_registration_definition(path: Path) -> Path:
    """Change a named registration type string without changing public counts."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.REGISTER"]
    if definition.attr_text != (
        '=LAMBDA(module,procedure,REGISTER.ID(module,procedure,"J!"))'
    ):
        raise ValueError("Fixture does not contain the expected named registration")
    definition.attr_text = '=LAMBDA(module,procedure,REGISTER.ID(module,procedure,"K!"))'
    workbook.save(path)
    return path


def change_named_worksheet_code_resource_registration_input(path: Path) -> Path:
    """Change a static input used by named registration expressions."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-NAMED-REGISTRATION-MODULE-CANDIDATE"
    workbook.save(path)
    return path


def make_formula_defined_xlm_registration_model(path: Path) -> Path:
    """Create inert XLM ``REGISTER`` calls stored only in defined formulas.

    The workbook is never opened in Excel. Private module/procedure sentinels
    exist solely to prove FormulaFence keeps formula-defined XLM registration
    material out of public profiles, findings, and SARIF output.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Formula-defined XLM registration controls"
    inputs["A9"] = "PRIVATE-XLM-REGISTRATION-MODULE-BASELINE"
    inputs["A10"] = "PRIVATE-XLM-REGISTRATION-PROCEDURE-BASELINE"
    inputs["B2"] = "=FENCE.XLM.REGISTER(A9,A10)"
    inputs["B3"] = "=FENCE.XLM.DIRECT"
    inputs["B4"] = "=FENCE.XLM.CHAIN(A9,A10)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.REGISTER",
            attr_text='=LAMBDA(module,procedure,REGISTER(module,procedure,"J!"))',
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.CHAIN",
            attr_text="=LAMBDA(module,procedure,FENCE.XLM.REGISTER(module,procedure))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.DIRECT",
            attr_text='=REGISTER(Inputs!$A$9,Inputs!$A$10,"J!")',
        )
    )
    workbook.save(path)
    return path


def change_formula_defined_xlm_registration_definition(path: Path) -> Path:
    """Change a hidden XLM registration type string without changing counts."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.XLM.REGISTER"]
    if definition.attr_text != (
        '=LAMBDA(module,procedure,REGISTER(module,procedure,"J!"))'
    ):
        raise ValueError("Fixture does not contain the expected XLM registration")
    definition.attr_text = '=LAMBDA(module,procedure,REGISTER(module,procedure,"K!"))'
    workbook.save(path)
    return path


def change_formula_defined_xlm_registration_input(path: Path) -> Path:
    """Change a static input used by a formula-defined XLM registration."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-XLM-REGISTRATION-MODULE-CANDIDATE"
    workbook.save(path)
    return path


def make_formula_defined_xlm_evaluation_model(path: Path) -> Path:
    """Create inert XLM `EVALUATE` calls stored only in defined formulas.

    The workbook is never opened in Excel. The controlled expression text
    proves FormulaFence compares stored dynamic-evaluation material privately
    without attempting to evaluate or parse it as an Excel formula.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Formula-defined XLM evaluation controls"
    inputs["A9"] = "PRIVATE-XLM-EVALUATE-EXPRESSION-BASELINE"
    inputs["B2"] = "=FENCE.XLM.EVALUATE(A9)"
    inputs["B3"] = "=FENCE.XLM.EVALUATE.DIRECT"
    inputs["B4"] = "=FENCE.XLM.EVALUATE.CHAIN(A9)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.EVALUATE",
            attr_text="=LAMBDA(expression,EVALUATE(expression))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.EVALUATE.CHAIN",
            attr_text="=LAMBDA(expression,FENCE.XLM.EVALUATE(expression))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.EVALUATE.DIRECT",
            attr_text="=EVALUATE(Inputs!$A$9)",
        )
    )
    workbook.save(path)
    return path


def change_formula_defined_xlm_evaluation_definition(path: Path) -> Path:
    """Change hidden expression material without changing public counts."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.XLM.EVALUATE"]
    if definition.attr_text != "=LAMBDA(expression,EVALUATE(expression))":
        raise ValueError("Fixture does not contain the expected XLM evaluation")
    definition.attr_text = (
        '=LAMBDA(expression,EVALUATE("PRIVATE-XLM-EVALUATE-DEFINITION-CANDIDATE"))'
    )
    workbook.save(path)
    return path


def change_formula_defined_xlm_evaluation_input(path: Path) -> Path:
    """Change a static text input used by a formula-defined XLM evaluation."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-XLM-EVALUATE-EXPRESSION-CANDIDATE"
    workbook.save(path)
    return path


def make_formula_dde_link_model(path: Path) -> Path:
    """Create inert direct DDE syntax in cells and stored name definitions.

    The fixture is never opened in Excel.  Its opaque service/topic/item text
    lets FormulaFence prove that it inventories and compares formula material
    without launching, resolving, or contacting a DDE server.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Direct DDE formula controls"
    inputs["A9"] = "PRIVATE-DDE-INPUT-BASELINE"
    inputs["B2"] = "=DDE.TEST|'PRIVATE-DDE-DIRECT-TOPIC-BASELINE'!PRIVATE_ITEM"
    inputs["B3"] = "=FENCE.DDE.DIRECT"
    inputs["B4"] = "=FENCE.DDE.CHAIN"
    inputs["B5"] = "=FENCE.DDE.LAMBDA(A9)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.DDE.DIRECT",
            attr_text="=DDE.TEST|'PRIVATE-DDE-NAMED-TOPIC-BASELINE'!PRIVATE_ITEM",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.DDE.CHAIN",
            attr_text="=FENCE.DDE.DIRECT",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.DDE.LAMBDA",
            attr_text=(
                "=LAMBDA(payload,DDE.TEST|'PRIVATE-DDE-LAMBDA-TOPIC-BASELINE'!payload)"
            ),
        )
    )
    workbook.save(path)
    return path


def change_formula_dde_link_definition(path: Path) -> Path:
    """Change a stored DDE topic without changing public aggregate counts."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.DDE.DIRECT"]
    baseline = "=DDE.TEST|'PRIVATE-DDE-NAMED-TOPIC-BASELINE'!PRIVATE_ITEM"
    if definition.attr_text != baseline:
        raise ValueError("Fixture does not contain the expected DDE definition")
    definition.attr_text = "=DDE.TEST|'PRIVATE-DDE-NAMED-TOPIC-CANDIDATE'!PRIVATE_ITEM"
    workbook.save(path)
    return path


def change_formula_dde_link_input(path: Path) -> Path:
    """Change a visible argument to an invoking DDE named LAMBDA."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-DDE-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_formula_defined_xlm_action_model(path: Path) -> Path:
    """Create inert XLM action calls stored only in defined formulas.

    The workbook is never opened in Excel. The stored literals exercise static
    comparison of selected macro and event-dispatch calls without resolving or
    executing a macro, program, DLL entry point, DDE command, or event handler.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Formula-defined XLM action controls"
    inputs["A9"] = "PRIVATE-XLM-ACTION-INPUT-BASELINE"
    inputs["B2"] = "=FENCE.XLM.ACTION(A9)"
    inputs["B3"] = "=FENCE.XLM.ACTION.DIRECT"
    inputs["B4"] = "=FENCE.XLM.ACTION.CHAIN(A9)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.ACTION",
            attr_text=(
                '=LAMBDA(payload,EXEC(payload)+RUN("PRIVATE-XLM-ACTION-MACRO"))'
            ),
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.ACTION.CHAIN",
            attr_text="=LAMBDA(payload,FENCE.XLM.ACTION(payload))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.ACTION.DIRECT",
            attr_text='=ON.TIME(NOW(),"PRIVATE-XLM-ACTION-EVENT")',
        )
    )
    workbook.save(path)
    return path


def change_formula_defined_xlm_action_definition(path: Path) -> Path:
    """Change a stored XLM action target without changing public counts."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.XLM.ACTION"]
    baseline = '=LAMBDA(payload,EXEC(payload)+RUN("PRIVATE-XLM-ACTION-MACRO"))'
    if definition.attr_text != baseline:
        raise ValueError("Fixture does not contain the expected XLM action")
    definition.attr_text = (
        '=LAMBDA(payload,EXEC(payload)+RUN("PRIVATE-XLM-ACTION-MACRO-CANDIDATE"))'
    )
    workbook.save(path)
    return path


def change_formula_defined_xlm_action_input(path: Path) -> Path:
    """Change a static input used by a formula-defined XLM action."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-XLM-ACTION-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_formula_defined_xlm_get_cell_model(path: Path) -> Path:
    """Create inert XLM `GET.CELL` calls stored only in defined formulas.

    The workbook is never opened in Excel. The controlled formula-defined
    expressions prove FormulaFence compares stored information-call material
    privately without evaluating it or simulating the referenced cell state.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Formula-defined XLM GET.CELL controls"
    inputs["A9"] = "PRIVATE-XLM-GET-CELL-INPUT-BASELINE"
    inputs["B2"] = "=FENCE.XLM.GET.CELL(A9)"
    inputs["B3"] = "=FENCE.XLM.GET.CELL.DIRECT"
    inputs["B4"] = "=FENCE.XLM.GET.CELL.CHAIN(A9)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.GET.CELL",
            attr_text="=LAMBDA(reference,GET.CELL(7,reference))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.GET.CELL.CHAIN",
            attr_text="=LAMBDA(reference,FENCE.XLM.GET.CELL(reference))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.GET.CELL.DIRECT",
            attr_text="=GET.CELL(53,Inputs!$A$9)",
        )
    )
    workbook.save(path)
    return path


def change_formula_defined_xlm_get_cell_definition(path: Path) -> Path:
    """Change hidden XLM information-call material without changing counts."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.XLM.GET.CELL"]
    if definition.attr_text != "=LAMBDA(reference,GET.CELL(7,reference))":
        raise ValueError("Fixture does not contain the expected XLM GET.CELL call")
    definition.attr_text = "=LAMBDA(reference,GET.CELL(8,reference))"
    workbook.save(path)
    return path


def change_formula_defined_xlm_get_cell_input(path: Path) -> Path:
    """Change a static input used by a formula-defined XLM GET.CELL call."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-XLM-GET-CELL-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_formula_defined_xlm_environment_information_model(path: Path) -> Path:
    """Create inert legacy environment-information calls in defined formulas.

    The workbook is never opened in Excel. The controlled stored formulas prove
    FormulaFence inventories and compares private call material without
    evaluating a call or simulating its workbook, workspace, or document state.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Formula-defined XLM environment-information controls"
    inputs["A9"] = "PRIVATE-XLM-ENVIRONMENT-INPUT-BASELINE"
    inputs["B2"] = "=FENCE.XLM.GET.WORKBOOK(A9)"
    inputs["B3"] = "=FENCE.XLM.GET.WORKSPACE"
    inputs["B4"] = "=FENCE.XLM.GET.DOCUMENT(A9)"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.GET.WORKBOOK",
            attr_text="=LAMBDA(workbook_name,GET.WORKBOOK(4,workbook_name))",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.GET.WORKSPACE",
            attr_text="=GET.WORKSPACE(2)",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.XLM.GET.DOCUMENT",
            attr_text="=LAMBDA(document_name,GET.DOCUMENT(37,document_name))",
        )
    )
    workbook.save(path)
    return path


def change_formula_defined_xlm_environment_information_definition(path: Path) -> Path:
    """Change private XLM environment-information material without new counts."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.XLM.GET.WORKBOOK"]
    if definition.attr_text != "=LAMBDA(workbook_name,GET.WORKBOOK(4,workbook_name))":
        raise ValueError(
            "Fixture does not contain the expected XLM environment-information call"
        )
    definition.attr_text = "=LAMBDA(workbook_name,GET.WORKBOOK(1,workbook_name))"
    workbook.save(path)
    return path


def change_formula_defined_xlm_environment_information_input(path: Path) -> Path:
    """Change a static input used by a formula-defined XLM environment call."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-XLM-ENVIRONMENT-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_formula_environment_information_model(path: Path) -> Path:
    """Create native CELL/INFO calls without opening the workbook in Excel.

    The fixture keeps location, information type, inputs, and defined-name
    identities private. It proves FormulaFence inventories stored syntax and
    static edges without evaluating a call or simulating a file/client/current
    selection environment.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Native CELL and INFO environment-information controls"
    inputs["A9"] = "PRIVATE-NATIVE-ENVIRONMENT-INPUT-BASELINE"
    inputs["B2"] = '=CELL("filename")'
    inputs["B3"] = '=CELL("type",A9)'
    inputs["B4"] = '=INFO("directory")'
    inputs["B5"] = "=FENCE.NATIVE.ENVIRONMENT(A9)"
    inputs["B6"] = "=FENCE.NATIVE.DIRECT"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.NATIVE.ENVIRONMENT",
            attr_text='=LAMBDA(value,CELL("filename")+INFO("system")+value)',
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.NATIVE.DIRECT",
            attr_text='=CELL("filename")',
        )
    )
    workbook.save(path)
    return path


def change_formula_environment_information_definition(path: Path) -> Path:
    """Change private native CELL/INFO material without changing counts."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.NATIVE.ENVIRONMENT"]
    expected = '=LAMBDA(value,CELL("filename")+INFO("system")+value)'
    if definition.attr_text != expected:
        raise ValueError("Fixture does not contain the expected native environment call")
    definition.attr_text = '=LAMBDA(value,CELL("filename")+INFO("osversion")+value)'
    workbook.save(path)
    return path


def change_formula_environment_information_input(path: Path) -> Path:
    """Change a static input used by a native CELL/INFO formula call."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-NATIVE-ENVIRONMENT-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_formula_workbook_structure_information_model(path: Path) -> Path:
    """Create native SHEET/SHEETS calls without calculating workbook state.

    The fixture covers direct calls, explicit references, a named LAMBDA, and a
    formula-defined name. Its tab titles and static input sentinel must remain
    private in FormulaFence's public profile, diff, Markdown, and SARIF output.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    workbook.create_sheet("Model")
    workbook.create_sheet("Report")
    inputs["A1"] = "Native SHEET and SHEETS workbook-structure controls"
    inputs["A9"] = "PRIVATE-WORKBOOK-TAB-INPUT-BASELINE"
    inputs["B2"] = "=SHEET()"
    inputs["B3"] = "=SHEETS()"
    inputs["B4"] = "=SHEET(Inputs!$A$1)"
    inputs["B5"] = "=SHEETS(Inputs!$A$1)"
    inputs["B6"] = "=FENCE.TAB.INFORMATION(A9)"
    inputs["B7"] = "=FENCE.TAB.DIRECT"
    workbook.defined_names.add(
        DefinedName(
            "FENCE.TAB.INFORMATION",
            attr_text="=LAMBDA(value,SHEET()+SHEETS()+value)",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "FENCE.TAB.DIRECT",
            attr_text="=SHEETS()",
        )
    )
    workbook.save(path)
    return path


def change_formula_workbook_structure_information_definition(path: Path) -> Path:
    """Change private SHEET material while retaining public call counts."""
    workbook = load_workbook(path)
    definition = workbook.defined_names["FENCE.TAB.INFORMATION"]
    expected = "=LAMBDA(value,SHEET()+SHEETS()+value)"
    if definition.attr_text != expected:
        raise ValueError("Fixture does not contain the expected workbook-tab call")
    definition.attr_text = "=LAMBDA(value,SHEET(Inputs!$A$1)+SHEETS()+value)"
    workbook.save(path)
    return path


def change_formula_workbook_structure_information_input(path: Path) -> Path:
    """Change a static input used by a named SHEET/SHEETS formula call."""
    workbook = load_workbook(path)
    workbook["Inputs"]["A9"] = "PRIVATE-WORKBOOK-TAB-INPUT-CANDIDATE"
    workbook.save(path)
    return path


def make_python_in_excel_model(path: Path) -> Path:
    """Create a workbook with stored Python-in-Excel package code.

    ``openpyxl`` does not write Python-in-Excel parts, so the controlled
    package-only declarations below mirror Excel's workbook relationship and
    content-type wiring. The source-code sentinels must remain package-private:
    tests assert that FormulaFence never reports them.
    """
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Python in Excel controls"
    inputs["A9"] = 7
    inputs["B2"] = "=_xlfn._xlws.PY(0,0,A9)"
    inputs["B3"] = "=_xlws.PY(1,1,A9)"
    workbook.save(path)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        python = ElementTree.Element(f"{{{_PYTHON_IN_EXCEL_NS}}}python")
        environment = ElementTree.SubElement(
            python,
            f"{{{_PYTHON_IN_EXCEL_NS}}}environmentDefinition",
            {"id": "{11111111-2222-3333-4444-555555555555}"},
        )
        initialization = ElementTree.SubElement(
            environment,
            f"{{{_PYTHON_IN_EXCEL_NS}}}initialization",
        )
        initialization_code = ElementTree.SubElement(
            initialization,
            f"{{{_PYTHON_IN_EXCEL_NS}}}code",
        )
        initialization_code.text = "PRIVATE-PYTHON-INIT-BASELINE"
        scripts = ElementTree.SubElement(
            python,
            f"{{{_PYTHON_IN_EXCEL_NS}}}pythonScripts",
        )
        for code in (
            "PRIVATE-PYTHON-SCRIPT-BASELINE = xl(\"A9\")",
            "PRIVATE-PYTHON-SCRIPT-SECOND = xl(\"A9\")",
        ):
            script = ElementTree.SubElement(
                scripts,
                f"{{{_PYTHON_IN_EXCEL_NS}}}pythonScript",
            )
            script_code = ElementTree.SubElement(
                script,
                f"{{{_PYTHON_IN_EXCEL_NS}}}code",
            )
            script_code.text = code
        contents["xl/python.xml"] = serialize(python)

        relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            relationships,
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship",
            {
                "Id": "rIdFencePython",
                "Type": _PYTHON_IN_EXCEL_RELATIONSHIP,
                "Target": "python.xml",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = serialize(relationships)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        ElementTree.SubElement(
            content_types,
            f"{{{_CONTENT_TYPES_NS}}}Override",
            {
                "PartName": "/xl/python.xml",
                "ContentType": _PYTHON_IN_EXCEL_CONTENT_TYPE,
            },
        )
        contents["[Content_Types].xml"] = serialize(content_types)

    return _rewrite_archive(path, mutate, ".python-in-excel-model.tmp.xlsx")


def add_python_in_excel_scripts_compatibility_part(path: Path) -> Path:
    """Add Excel's separately stored 2022 PythonScripts package contract."""
    def mutate(contents: dict[str, bytes]) -> None:
        if "xl/pythonScripts.xml" in contents:
            raise ValueError("Fixture already contains a PythonScripts package part")

        scripts = ElementTree.Element(
            f"{{{_PYTHON_IN_EXCEL_SCRIPTS_NS}}}pythonScripts"
        )
        for code in (
            "PRIVATE-PYTHON-SCRIPTS-BASELINE = xl(\"A9\")",
            "PRIVATE-PYTHON-SCRIPTS-SECOND = xl(\"A9\")",
        ):
            script = ElementTree.SubElement(
                scripts,
                f"{{{_PYTHON_IN_EXCEL_SCRIPTS_NS}}}pythonScript",
            )
            script_code = ElementTree.SubElement(
                script,
                f"{{{_PYTHON_IN_EXCEL_SCRIPTS_NS}}}code",
            )
            script_code.text = code
        contents["xl/pythonScripts.xml"] = ElementTree.tostring(
            scripts,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            relationships,
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship",
            {
                "Id": "rIdFencePythonScripts",
                "Type": _PYTHON_IN_EXCEL_SCRIPTS_RELATIONSHIP,
                "Target": "pythonScripts.xml",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        ElementTree.SubElement(
            content_types,
            f"{{{_CONTENT_TYPES_NS}}}Override",
            {
                "PartName": "/xl/pythonScripts.xml",
                "ContentType": _PYTHON_IN_EXCEL_SCRIPTS_CONTENT_TYPE,
            },
        )
        contents["[Content_Types].xml"] = ElementTree.tostring(
            content_types,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".python-in-excel-scripts-model.tmp.xlsx")


def make_python_in_excel_scripts_model(path: Path) -> Path:
    """Create a workbook that uses only the 2022 PythonScripts contract."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "PythonScripts compatibility controls"
    inputs["A9"] = 7
    inputs["B2"] = "=_xlfn._xlws.PY(0,0,A9)"
    workbook.save(path)
    return add_python_in_excel_scripts_compatibility_part(path)


def change_python_in_excel_scripts_script(path: Path) -> Path:
    """Change stored 2022 PythonScripts code without rewriting its PY formula."""
    def mutate(contents: dict[str, bytes]) -> None:
        root = ElementTree.fromstring(contents["xl/pythonScripts.xml"])
        script = root.find(
            f"{{{_PYTHON_IN_EXCEL_SCRIPTS_NS}}}pythonScript/"
            f"{{{_PYTHON_IN_EXCEL_SCRIPTS_NS}}}code"
        )
        if script is None:
            raise ValueError("Fixture does not contain a PythonScripts script")
        script.text = "PRIVATE-PYTHON-SCRIPTS-CANDIDATE = xl(\"A9\")"
        contents["xl/pythonScripts.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".python-in-excel-scripts-script.tmp.xlsx")


def renumber_python_in_excel_scripts_relationship_identifier(path: Path) -> Path:
    """Change only the non-semantic 2022 PythonScripts relationship ID."""
    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
                )
                if current.get("Type") == _PYTHON_IN_EXCEL_SCRIPTS_RELATIONSHIP
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a PythonScripts relationship")
        relationship.set("Id", "rIdFencePythonScriptsRenumbered")
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(
        path,
        mutate,
        ".python-in-excel-scripts-relationship-id.tmp.xlsx",
    )


def corrupt_python_in_excel_scripts_package(path: Path) -> Path:
    """Make the 2022 PythonScripts material malformed for coverage tests."""
    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/pythonScripts.xml"] = b"<pythonScripts"

    return _rewrite_archive(path, mutate, ".python-in-excel-scripts-corrupt.tmp.xlsx")


def mismatch_python_in_excel_scripts_relationship_contract(path: Path) -> Path:
    """Bind a 2022 PythonScripts target with the incompatible 2023 relation."""
    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
                )
                if current.get("Type") == _PYTHON_IN_EXCEL_SCRIPTS_RELATIONSHIP
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a PythonScripts relationship")
        relationship.set("Type", _PYTHON_IN_EXCEL_RELATIONSHIP)
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(
        path,
        mutate,
        ".python-in-excel-scripts-mismatched-contract.tmp.xlsx",
    )


def change_python_in_excel_script(path: Path) -> Path:
    """Change stored Python code without changing a PY formula placeholder."""
    def mutate(contents: dict[str, bytes]) -> None:
        root = ElementTree.fromstring(contents["xl/python.xml"])
        script = root.find(
            f"{{{_PYTHON_IN_EXCEL_NS}}}pythonScripts/"
            f"{{{_PYTHON_IN_EXCEL_NS}}}pythonScript/"
            f"{{{_PYTHON_IN_EXCEL_NS}}}code"
        )
        if script is None:
            raise ValueError("Fixture does not contain a Python-in-Excel script")
        script.text = "PRIVATE-PYTHON-SCRIPT-CANDIDATE = xl(\"A9\")"
        contents["xl/python.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".python-in-excel-script.tmp.xlsx")


def change_python_in_excel_input(path: Path) -> Path:
    """Change a static PY source cell without rewriting package declarations."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = _inputs_worksheet_root(contents)
        cell = next(
            (
                current
                for current in worksheet.iter(f"{{{_SPREADSHEETML_NS}}}c")
                if current.get("r") == "A9"
            ),
            None,
        )
        if cell is None:
            raise ValueError("Fixture does not contain Python-in-Excel input A9")
        value = cell.find(f"{{{_SPREADSHEETML_NS}}}v")
        if value is None:
            raise ValueError("Fixture Python-in-Excel input A9 has no numeric value")
        value.text = "8"
        _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".python-in-excel-input.tmp.xlsx")


def change_python_in_excel_formula_binding(path: Path) -> Path:
    """Change the PY script binding without altering stored package code."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = _inputs_worksheet_root(contents)
        cell = next(
            (
                current
                for current in worksheet.iter(f"{{{_SPREADSHEETML_NS}}}c")
                if current.get("r") == "B2"
            ),
            None,
        )
        formula = (
            cell.find(f"{{{_SPREADSHEETML_NS}}}f") if cell is not None else None
        )
        if formula is None or formula.text is None:
            raise ValueError("Fixture does not contain Python-in-Excel formula B2")
        formula.text = formula.text.replace("PY(0,0,A9)", "PY(1,0,A9)")
        _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".python-in-excel-binding.tmp.xlsx")


def renumber_python_in_excel_relationship_identifier(path: Path) -> Path:
    """Change only a non-semantic workbook relationship identifier."""
    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
                )
                if current.get("Type") == _PYTHON_IN_EXCEL_RELATIONSHIP
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a Python-in-Excel relationship")
        relationship.set("Id", "rIdFencePythonRenumbered")
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".python-in-excel-relationship-id.tmp.xlsx")


def corrupt_python_in_excel_package(path: Path) -> Path:
    """Make stored Python metadata malformed to exercise coverage reporting."""
    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/python.xml"] = b"<python"

    return _rewrite_archive(path, mutate, ".python-in-excel-corrupt.tmp.xlsx")


def make_legacy_comment_model(path: Path) -> Path:
    """Create a real Excel Note whose data lives outside worksheet cells."""
    make_model(path)
    workbook = load_workbook(path)
    note = Comment(
        "PRIVATE-LEGACY-NOTE-BASELINE",
        "Private Legacy Note Author",
    )
    note.width = 240
    note.height = 120
    workbook["Inputs"]["A1"].comment = note
    workbook.save(path)
    return path


def _cell_hyperlink_part_names(contents: dict[str, bytes]) -> tuple[str, str]:
    """Return the worksheet and relationship members used by cell hyperlinks."""
    worksheet_member = "xl/worksheets/sheet1.xml"
    relationships_member = _relationship_member(worksheet_member)
    if (
        worksheet_member not in contents
        or relationships_member not in contents
    ):
        raise ValueError("Fixture does not contain a worksheet cell hyperlink package")
    return worksheet_member, relationships_member


def make_cell_hyperlink_model(path: Path) -> Path:
    """Create private external and in-workbook cell hyperlink declarations."""
    make_model(path)
    workbook = load_workbook(path)
    worksheet = workbook["Inputs"]
    worksheet["A1"] = "Open approved source"
    worksheet["A1"].hyperlink = "https://approved.example.test/PRIVATE-LINK-BASELINE"
    worksheet["A1"].hyperlink.tooltip = "PRIVATE-EXTERNAL-LINK-TOOLTIP"
    worksheet["A1"].style = "Hyperlink"
    worksheet["A2"] = "Jump to source"
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member, _relationships_member = _cell_hyperlink_part_names(contents)
        worksheet_root = ElementTree.fromstring(contents[worksheet_member])
        hyperlinks_tag = f"{{{_SPREADSHEETML_NS}}}hyperlinks"
        hyperlink_tag = f"{{{_SPREADSHEETML_NS}}}hyperlink"
        hyperlinks = worksheet_root.find(hyperlinks_tag)
        if hyperlinks is None:
            hyperlinks = ElementTree.SubElement(worksheet_root, hyperlinks_tag)
        ElementTree.SubElement(
            hyperlinks,
            hyperlink_tag,
            {
                "ref": "A2",
                "location": "'Inputs'!A1",
                "display": "PRIVATE-INTERNAL-LINK-DISPLAY",
                "tooltip": "PRIVATE-INTERNAL-LINK-TOOLTIP",
            },
        )
        contents[worksheet_member] = ElementTree.tostring(
            worksheet_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".cell-hyperlink-model.tmp.xlsx")


def _cell_hyperlink_element(
    worksheet: ElementTree.Element,
    reference: str,
) -> ElementTree.Element:
    """Return one standard hyperlink fixture declaration by its private ref."""
    hyperlink_tag = f"{{{_SPREADSHEETML_NS}}}hyperlink"
    hyperlink = next(
        (
            current
            for current in worksheet.iter(hyperlink_tag)
            if current.get("ref") == reference
        ),
        None,
    )
    if hyperlink is None:
        raise ValueError(f"Fixture does not contain cell hyperlink {reference}")
    return hyperlink


def change_cell_hyperlink_target(path: Path) -> Path:
    """Change an external cell hyperlink target without changing its cell value."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _worksheet_member, relationships_member = _cell_hyperlink_part_names(contents)
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if (current.get("Type") or "").endswith("/hyperlink")
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a hyperlink relationship")
        relationship.set("Target", "https://review.example.test/PRIVATE-LINK-CANDIDATE")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".cell-hyperlink-target.tmp.xlsx")


def change_cell_hyperlink_tooltip(path: Path) -> Path:
    """Change only a private ScreenTip on an ordinary cell hyperlink."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member, _relationships_member = _cell_hyperlink_part_names(contents)
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        _cell_hyperlink_element(worksheet, "A1").set(
            "tooltip",
            "PRIVATE-EXTERNAL-LINK-TOOLTIP-CANDIDATE",
        )
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".cell-hyperlink-tooltip.tmp.xlsx")


def change_cell_hyperlink_display(path: Path) -> Path:
    """Change a private display override without changing the cell's value."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member, _relationships_member = _cell_hyperlink_part_names(contents)
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        _cell_hyperlink_element(worksheet, "A2").set(
            "display",
            "PRIVATE-INTERNAL-LINK-DISPLAY-CANDIDATE",
        )
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".cell-hyperlink-display.tmp.xlsx")


def change_cell_hyperlink_location(path: Path) -> Path:
    """Retarget an in-workbook cell hyperlink without changing visible text."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member, _relationships_member = _cell_hyperlink_part_names(contents)
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        _cell_hyperlink_element(worksheet, "A2").set(
            "location",
            "'Inputs'!B2",
        )
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".cell-hyperlink-location.tmp.xlsx")


def renumber_cell_hyperlink_identifiers(path: Path) -> Path:
    """Rewrite a package relationship ID and optional writer UID consistently."""
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    revision = _OFFICE_2014_REVISION_NS

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member, relationships_member = _cell_hyperlink_part_names(contents)
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        relationship_attribute = f"{{{document_relationships}}}id"
        hyperlink = _cell_hyperlink_element(worksheet, "A1")
        old_identifier = hyperlink.get(relationship_attribute)
        if old_identifier is None:
            raise ValueError("Fixture hyperlink is missing a relationship identifier")
        hyperlink.set(relationship_attribute, "rIdFenceCellHyperlink")
        hyperlink.set(f"{{{revision}}}uid", "{AAAAAAAA-AAAA-AAAA-AAAA-AAAAAAAAAAAA}")
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if current.get("Id") == old_identifier
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture relationship identifier is missing")
        relationship.set("Id", "rIdFenceCellHyperlink")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".cell-hyperlink-identifiers.tmp.xlsx")


def rewrite_cell_hyperlink_as_revision_declaration(path: Path) -> Path:
    """Move one stored link into Excel's Office 2016 revision namespace."""
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    revision = _OFFICE_2014_REVISION_NS

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member, _relationships_member = _cell_hyperlink_part_names(contents)
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        hyperlinks_tag = f"{{{_SPREADSHEETML_NS}}}hyperlinks"
        ext_list_tag = f"{{{_SPREADSHEETML_NS}}}extLst"
        ext_tag = f"{{{_SPREADSHEETML_NS}}}ext"
        hyperlinks = worksheet.find(hyperlinks_tag)
        if hyperlinks is None:
            raise ValueError("Fixture does not contain a hyperlink container")
        original = _cell_hyperlink_element(worksheet, "A1")
        relationship_attribute = f"{{{document_relationships}}}id"
        relationship_id = original.get(relationship_attribute)
        if relationship_id is None:
            raise ValueError("Fixture hyperlink is missing a relationship identifier")
        hyperlinks.remove(original)
        revision_link = ElementTree.Element(
            f"{{{revision}}}hyperlink",
            {
                "ref": "A1",
                relationship_attribute: relationship_id,
                "tooltip": "PRIVATE-EXTERNAL-LINK-TOOLTIP",
            },
        )
        ext_list = worksheet.find(ext_list_tag)
        if ext_list is None:
            ext_list = ElementTree.SubElement(worksheet, ext_list_tag)
        extension = ElementTree.SubElement(
            ext_list,
            ext_tag,
            {"uri": "{C11C4282-8D77-4AC7-B5A1-96BCB966B2A4}"},
        )
        extension.append(revision_link)
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".cell-hyperlink-revision.tmp.xlsx")


def corrupt_cell_hyperlink_reference(path: Path) -> Path:
    """Inject a malformed reference that the ordinary reader must not parse."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member, _relationships_member = _cell_hyperlink_part_names(contents)
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        _cell_hyperlink_element(worksheet, "A1").set(
            "ref",
            "PRIVATE-NOT-A-CELL-REFERENCE",
        )
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".cell-hyperlink-corrupt.tmp.xlsx")


def unbind_cell_hyperlink_relationship(path: Path) -> Path:
    """Remove a referenced hyperlink relationship to exercise fail-closed parsing."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member, relationships_member = _cell_hyperlink_part_names(contents)
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        relationship_id = _cell_hyperlink_element(worksheet, "A1").get(
            relationship_attribute
        )
        if relationship_id is None:
            raise ValueError("Fixture hyperlink is missing a relationship identifier")
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if current.get("Id") == relationship_id
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture relationship identifier is missing")
        relationships.remove(relationship)
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".cell-hyperlink-unbound.tmp.xlsx")


def _worksheet_sparkline_group(
    worksheet: ElementTree.Element,
) -> ElementTree.Element:
    """Return the first x14 sparkline group from the ordinary Inputs sheet."""
    group_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparklineGroup"
    group = next(worksheet.iter(group_tag), None)
    if group is None:
        raise ValueError("Fixture does not contain a worksheet sparkline group")
    return group


def make_worksheet_sparkline_model(path: Path) -> Path:
    """Create Office 2010 worksheet sparklines outside ordinary cell values."""
    make_model(path)

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        extension_list_tag = f"{{{_SPREADSHEETML_NS}}}extLst"
        extension_tag = f"{{{_SPREADSHEETML_NS}}}ext"
        group_container_tag = (
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparklineGroups"
        )
        group_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparklineGroup"
        sparklines_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparklines"
        sparkline_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparkline"
        formula_tag = f"{{{_EXCEL_2006_MAIN_NS}}}f"
        destination_tag = f"{{{_EXCEL_2006_MAIN_NS}}}sqref"
        extension_list = worksheet.find(extension_list_tag)
        if extension_list is None:
            extension_list = ElementTree.SubElement(worksheet, extension_list_tag)
        extension = ElementTree.SubElement(
            extension_list,
            extension_tag,
            {"uri": "{05C60535-1F16-4FD2-B633-F4F36F0B64E0}"},
        )
        groups = ElementTree.SubElement(extension, group_container_tag)
        group = ElementTree.SubElement(
            groups,
            group_tag,
            {
                "type": "line",
                "dateAxis": "1",
                "displayEmptyCellsAs": "span",
                "displayHidden": "1",
                "displayXAxis": "1",
                "first": "1",
                "high": "1",
                "last": "1",
                "low": "1",
                "markers": "1",
                "negative": "1",
                "rightToLeft": "1",
                "maxAxisType": "custom",
                "manualMax": "100",
                "minAxisType": "custom",
                "manualMin": "0",
                "lineWeight": "0.75",
            },
        )
        ElementTree.SubElement(
            group,
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}colorSeries",
            {"rgb": "FF112233"},
        )
        ElementTree.SubElement(
            group,
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}colorMarkers",
            {"rgb": "FF445566"},
        )
        ElementTree.SubElement(group, formula_tag).text = "Inputs!$A$2:$A$4"
        sparklines = ElementTree.SubElement(group, sparklines_tag)
        first = ElementTree.SubElement(sparklines, sparkline_tag)
        ElementTree.SubElement(first, formula_tag).text = "Inputs!$B$2:$B$4"
        ElementTree.SubElement(first, destination_tag).text = "$F$1"
        second = ElementTree.SubElement(sparklines, sparkline_tag)
        ElementTree.SubElement(second, formula_tag).text = "Inputs!$B$2:$B$4"
        ElementTree.SubElement(second, destination_tag).text = "$F$2"
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-sparkline-model.tmp.xlsx")


def make_cell_hyperlink_sparkline_model(path: Path) -> Path:
    """Create one worksheet carrying both raw hyperlinks and sparklines."""
    make_cell_hyperlink_model(path)

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        extension_list_tag = f"{{{_SPREADSHEETML_NS}}}extLst"
        extension_tag = f"{{{_SPREADSHEETML_NS}}}ext"
        group_container_tag = (
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparklineGroups"
        )
        group_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparklineGroup"
        sparklines_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparklines"
        sparkline_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparkline"
        formula_tag = f"{{{_EXCEL_2006_MAIN_NS}}}f"
        destination_tag = f"{{{_EXCEL_2006_MAIN_NS}}}sqref"
        extension_list = worksheet.find(extension_list_tag)
        if extension_list is None:
            extension_list = ElementTree.SubElement(worksheet, extension_list_tag)
        extension = ElementTree.SubElement(
            extension_list,
            extension_tag,
            {"uri": "{05C60535-1F16-4FD2-B633-F4F36F0B64E0}"},
        )
        groups = ElementTree.SubElement(extension, group_container_tag)
        group = ElementTree.SubElement(groups, group_tag, {"type": "line"})
        sparklines = ElementTree.SubElement(group, sparklines_tag)
        sparkline = ElementTree.SubElement(sparklines, sparkline_tag)
        ElementTree.SubElement(sparkline, formula_tag).text = "Inputs!$B$2:$B$4"
        ElementTree.SubElement(sparkline, destination_tag).text = "F1"
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(
        path,
        mutate,
        ".cell-hyperlink-sparkline-model.tmp.xlsx",
    )


def change_worksheet_sparkline_source(path: Path) -> Path:
    """Retarget a private sparkline source without changing ordinary cell values."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        formula_tag = f"{{{_EXCEL_2006_MAIN_NS}}}f"
        sparklines_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparklines"
        sparklines = _worksheet_sparkline_group(worksheet).find(sparklines_tag)
        if sparklines is None:
            raise ValueError("Fixture sparkline group has no sparklines")
        source = next(sparklines.iter(formula_tag), None)
        if source is None:
            raise ValueError("Fixture sparkline is missing a source formula")
        source.text = "Inputs!$B$3:$B$5"
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-sparkline-source.tmp.xlsx")


def change_worksheet_sparkline_presentation(path: Path) -> Path:
    """Change sparkline style controls without changing source cells or values."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        group = _worksheet_sparkline_group(worksheet)
        group.set("type", "column")
        colour = group.find(
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}colorSeries"
        )
        if colour is None:
            raise ValueError("Fixture sparkline group is missing its series colour")
        colour.set("rgb", "FF778899")
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-sparkline-style.tmp.xlsx")


def reorder_worksheet_sparklines(path: Path) -> Path:
    """Reorder equivalent declarations to exercise order-independent comparison."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        sparklines_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}sparklines"
        sparklines = _worksheet_sparkline_group(worksheet).find(sparklines_tag)
        if sparklines is None or len(sparklines) < 2:
            raise ValueError("Fixture needs two sparkline declarations to reorder")
        children = list(sparklines)
        for child in children:
            sparklines.remove(child)
        for child in reversed(children):
            sparklines.append(child)
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-sparkline-reorder.tmp.xlsx")


def normalize_worksheet_sparkline_control_spelling(path: Path) -> Path:
    """Rewrite equivalent range, Boolean, numeric, and colour spellings."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        formula_tag = f"{{{_EXCEL_2006_MAIN_NS}}}f"
        destination_tag = f"{{{_EXCEL_2006_MAIN_NS}}}sqref"
        group = _worksheet_sparkline_group(worksheet)
        group.set("dateAxis", "true")
        group.set("displayHidden", "true")
        group.set("markers", "true")
        group.set("manualMax", "100.0")
        group.set("manualMin", "0.0")
        group.set("lineWeight", "0.750")
        colour = group.find(
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}colorSeries"
        )
        if colour is None:
            raise ValueError("Fixture sparkline group is missing its series colour")
        colour.set("rgb", "ff112233")
        formulas = list(group.iter(formula_tag))
        if len(formulas) != 3:
            raise ValueError("Fixture has an unexpected sparkline formula count")
        formulas[0].text = "inputs!A2:A4"
        formulas[1].text = "inputs!B2:B4"
        formulas[2].text = "inputs!B2:B4"
        destinations = list(group.iter(destination_tag))
        if len(destinations) != 2:
            raise ValueError("Fixture has an unexpected sparkline destination count")
        destinations[0].text = "F1"
        destinations[1].text = "F2"
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-sparkline-normalize.tmp.xlsx")


def corrupt_worksheet_sparkline_destination(path: Path) -> Path:
    """Inject an invalid private sqref that the ordinary reader must not parse."""
    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        destination_tag = f"{{{_EXCEL_2006_MAIN_NS}}}sqref"
        destination = next(worksheet.iter(destination_tag), None)
        if destination is None:
            raise ValueError("Fixture sparkline is missing a destination")
        destination.text = "PRIVATE-NOT-A-SPARKLINE-CELL"
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-sparkline-corrupt.tmp.xlsx")


def _legacy_comment_part_names(contents: dict[str, bytes]) -> tuple[str, str, str]:
    """Return the comments, VML, and worksheet relationship members for a Note."""
    comment_members = sorted(
        member
        for member in contents
        if member.startswith("xl/comments/") and member.endswith(".xml")
    )
    drawing_members = sorted(
        member
        for member in contents
        if member.startswith("xl/drawings/") and member.endswith(".vml")
    )
    worksheet_relationships = _relationship_member("xl/worksheets/sheet1.xml")
    if (
        len(comment_members) != 1
        or len(drawing_members) != 1
        or worksheet_relationships not in contents
    ):
        raise ValueError("Fixture does not contain one legacy Excel Note package")
    return comment_members[0], drawing_members[0], worksheet_relationships


def change_legacy_comment_text(path: Path) -> Path:
    """Change private conventional Note text without changing worksheet cells."""
    def mutate(contents: dict[str, bytes]) -> None:
        comment_member, _drawing_member, _relationships_member = (
            _legacy_comment_part_names(contents)
        )
        root = ElementTree.fromstring(contents[comment_member])
        comment = next(root.iter(f"{{{_SPREADSHEETML_NS}}}comment"), None)
        if comment is None:
            raise ValueError("Fixture does not contain a legacy comment")
        text = comment.find(f"{{{_SPREADSHEETML_NS}}}text")
        value = (
            text.find(f"{{{_SPREADSHEETML_NS}}}t") if text is not None else None
        )
        if value is None:
            raise ValueError("Fixture does not contain plain legacy comment text")
        value.text = "PRIVATE-LEGACY-NOTE-CANDIDATE"
        contents[comment_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-comment-text.tmp.xlsx")


def change_legacy_note_visibility(path: Path) -> Path:
    """Reveal a Note VML shape without changing its corresponding cell."""
    vml = "urn:schemas-microsoft-com:vml"
    vml_excel = "urn:schemas-microsoft-com:office:excel"

    def mutate(contents: dict[str, bytes]) -> None:
        _comment_member, drawing_member, _relationships_member = (
            _legacy_comment_part_names(contents)
        )
        root = ElementTree.fromstring(contents[drawing_member])
        parent_by_child = {
            child: parent
            for parent in root.iter()
            for child in parent
        }
        note_data = next(
            (
                client_data
                for client_data in root.iter(f"{{{vml_excel}}}ClientData")
                if (client_data.get("ObjectType") or "").casefold() == "note"
            ),
            None,
        )
        if note_data is None:
            raise ValueError("Fixture does not contain a Note VML shape")
        shape = parent_by_child.get(note_data)
        if shape is None or shape.tag != f"{{{vml}}}shape":
            raise ValueError("Fixture Note VML shape is malformed")
        declarations = [
            declaration.strip()
            for declaration in (shape.get("style") or "").split(";")
            if declaration.strip()
        ]
        updated = False
        for index, declaration in enumerate(declarations):
            name, separator, _value = declaration.partition(":")
            if separator and name.strip().casefold() == "visibility":
                declarations[index] = "visibility:visible"
                updated = True
        if not updated:
            declarations.append("visibility:visible")
        shape.set("style", ";".join(declarations))
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-comment-visibility.tmp.xlsx")


def externalize_legacy_comment_relationship(path: Path) -> Path:
    """Make a conventional comments relationship unsafe without opening it."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _comment_member, _drawing_member, relationships_member = (
            _legacy_comment_part_names(contents)
        )
        root = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            (
                item
                for item in root.findall(f"{{{package_relationships}}}Relationship")
                if (item.get("Type") or "").endswith("/comments")
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a comments relationship")
        relationship.set("Target", "https://example.invalid/private-legacy-note")
        relationship.set("TargetMode", "External")
        contents[relationships_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-comment-external.tmp.xlsx")


def rebind_legacy_comment_relationship(path: Path) -> Path:
    """Point a Note at an equivalent but distinct comments package part."""
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        comment_member, _drawing_member, relationships_member = (
            _legacy_comment_part_names(contents)
        )
        replacement_member = "xl/comments/comment2.xml"
        contents[replacement_member] = contents.pop(comment_member)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override = next(
            (
                item
                for item in types.findall(f"{{{content_types}}}Override")
                if item.get("PartName") == f"/{comment_member}"
            ),
            None,
        )
        if override is None:
            raise ValueError("Fixture does not declare its comments package part")
        override.set("PartName", f"/{replacement_member}")
        contents["[Content_Types].xml"] = ElementTree.tostring(
            types,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            (
                item
                for item in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if (item.get("Type") or "").endswith("/comments")
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a comments relationship")
        relationship.set("Target", "../comments/comment2.xml")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-comment-rebind.tmp.xlsx")


def externalize_legacy_note_vml_relationship(path: Path) -> Path:
    """Make a Note VML relationship unsafe without opening its target."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _comment_member, _drawing_member, relationships_member = (
            _legacy_comment_part_names(contents)
        )
        root = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            (
                item
                for item in root.findall(f"{{{package_relationships}}}Relationship")
                if (item.get("Type") or "").endswith("/vmlDrawing")
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a Note VML relationship")
        relationship.set("Target", "https://example.invalid/private-legacy-note-vml")
        relationship.set("TargetMode", "External")
        contents[relationships_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-note-vml-external.tmp.xlsx")


def renumber_legacy_comment_identifiers(path: Path) -> Path:
    """Rewrite volatile Note shape and relationship identifiers consistently."""
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    vml = "urn:schemas-microsoft-com:vml"
    vml_excel = "urn:schemas-microsoft-com:office:excel"
    vml_office = "urn:schemas-microsoft-com:office:office"

    def mutate(contents: dict[str, bytes]) -> None:
        comment_member, drawing_member, relationships_member = (
            _legacy_comment_part_names(contents)
        )
        comment_root = ElementTree.fromstring(contents[comment_member])
        for index, comment in enumerate(
            comment_root.iter(f"{{{_SPREADSHEETML_NS}}}comment"),
            start=900,
        ):
            comment.set("shapeId", str(index))
        contents[comment_member] = ElementTree.tostring(
            comment_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        drawing_root = ElementTree.fromstring(contents[drawing_member])
        parent_by_child = {
            child: parent
            for parent in drawing_root.iter()
            for child in parent
        }
        for index, client_data in enumerate(
            drawing_root.iter(f"{{{vml_excel}}}ClientData"),
            start=900,
        ):
            if (client_data.get("ObjectType") or "").casefold() != "note":
                continue
            shape = parent_by_child.get(client_data)
            if shape is None or shape.tag != f"{{{vml}}}shape":
                raise ValueError("Fixture Note VML shape is malformed")
            shape.set("id", f"_x0000_s{index}")
            shape.set(f"{{{vml_office}}}spid", f"_x0000_s{index}")
        contents[drawing_member] = ElementTree.tostring(
            drawing_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships_root = ElementTree.fromstring(contents[relationships_member])
        relationship_ids: dict[str, str] = {}
        for relationship in relationships_root.findall(
            f"{{{package_relationships}}}Relationship"
        ):
            relationship_type = relationship.get("Type") or ""
            if relationship_type.endswith("/comments"):
                old_identifier = relationship.get("Id")
                if old_identifier:
                    relationship_ids[old_identifier] = "rIdFenceCommentRenumbered"
                    relationship.set("Id", "rIdFenceCommentRenumbered")
            elif relationship_type.endswith("/vmlDrawing"):
                old_identifier = relationship.get("Id")
                if old_identifier:
                    relationship_ids[old_identifier] = "rIdFenceNoteVmlRenumbered"
                    relationship.set("Id", "rIdFenceNoteVmlRenumbered")
        if len(relationship_ids) != 2:
            raise ValueError("Fixture does not contain both Note relationships")
        contents[relationships_member] = ElementTree.tostring(
            relationships_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet = _inputs_worksheet_root(contents)
        relationship_attribute = f"{{{document_relationships}}}id"
        for legacy_drawing in worksheet.iter(
            f"{{{_SPREADSHEETML_NS}}}legacyDrawing"
        ):
            old_identifier = legacy_drawing.get(relationship_attribute)
            if old_identifier in relationship_ids:
                legacy_drawing.set(relationship_attribute, relationship_ids[old_identifier])
        _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".legacy-comment-id.tmp.xlsx")


def corrupt_legacy_comment_root(path: Path) -> Path:
    """Replace a comments root to exercise legacy-note fail-closed coverage."""
    def mutate(contents: dict[str, bytes]) -> None:
        comment_member, _drawing_member, _relationships_member = (
            _legacy_comment_part_names(contents)
        )
        root = ElementTree.fromstring(contents[comment_member])
        root.tag = "notComments"
        contents[comment_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-comment-corrupt.tmp.xlsx")


def _digital_signature_xml_root(contents: dict[str, bytes]) -> ElementTree.Element:
    """Return the one package XML signature fixture envelope."""
    try:
        return ElementTree.fromstring(contents["_xmlsignatures/sig1.xml"])
    except KeyError as error:
        raise ValueError("Fixture does not contain a package XML signature") from error


def _digital_signature_relationship(
    relationships: ElementTree.Element,
    relationship_type: str,
) -> ElementTree.Element:
    """Return one digital-signature fixture relationship by its type."""
    relationship = next(
        (
            current
            for current in relationships.findall(
                f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            )
            if current.get("Type") == relationship_type
        ),
        None,
    )
    if relationship is None:
        raise ValueError(f"Fixture does not contain {relationship_type!r} relationship")
    return relationship


def _digital_signature_content_type_override(
    content_types: ElementTree.Element,
    member: str,
) -> ElementTree.Element:
    """Return one fixture content-type override by its package member."""
    override = next(
        (
            current
            for current in content_types.findall(f"{{{_CONTENT_TYPES_NS}}}Override")
            if current.get("PartName") == f"/{member}"
        ),
        None,
    )
    if override is None:
        raise ValueError(f"Fixture does not contain content type for {member!r}")
    return override


def make_digital_signature_model(path: Path) -> Path:
    """Create an inspectable package and VBA signature fixture.

    The intentionally private signature, digest, certificate, and binary values
    are structurally shaped but not cryptographically valid. FormulaFence only
    inventories and compares the envelopes; it must never claim validation.
    """
    make_model(path)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(
        content_types: ElementTree.Element,
        member: str,
        content_type: str,
    ) -> None:
        ElementTree.SubElement(
            content_types,
            f"{{{_CONTENT_TYPES_NS}}}Override",
            {"PartName": f"/{member}", "ContentType": content_type},
        )

    def mutate(contents: dict[str, bytes]) -> None:
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        signature = _XML_DIGITAL_SIGNATURE_NS

        root_relationships = ElementTree.fromstring(contents["_rels/.rels"])
        ElementTree.SubElement(
            root_relationships,
            relationship_tag,
            {
                "Id": "rIdFencePackageSignatureOrigin",
                "Type": _DIGITAL_SIGNATURE_ORIGIN_RELATIONSHIP,
                "Target": "_xmlsignatures/origin.sigs",
            },
        )
        contents["_rels/.rels"] = serialize(root_relationships)

        origin_member = "_xmlsignatures/origin.sigs"
        signature_member = "_xmlsignatures/sig1.xml"
        certificate_member = "package/services/digital-signature/certificate/cert1.cer"
        contents[origin_member] = b""
        origin_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            origin_relationships,
            relationship_tag,
            {
                "Id": "rIdFencePackageXmlSignature",
                "Type": _DIGITAL_SIGNATURE_SIGNATURE_RELATIONSHIP,
                "Target": "sig1.xml",
            },
        )
        contents["_xmlsignatures/_rels/origin.sigs.rels"] = serialize(
            origin_relationships
        )

        envelope = ElementTree.Element(f"{{{signature}}}Signature")
        signed_info = ElementTree.SubElement(envelope, f"{{{signature}}}SignedInfo")
        ElementTree.SubElement(
            signed_info,
            f"{{{signature}}}CanonicalizationMethod",
            {"Algorithm": "http://www.w3.org/TR/2001/REC-xml-c14n-20010315"},
        )
        ElementTree.SubElement(
            signed_info,
            f"{{{signature}}}SignatureMethod",
            {"Algorithm": "http://www.w3.org/2001/04/xmldsig-more#rsa-sha256"},
        )
        reference = ElementTree.SubElement(
            signed_info,
            f"{{{signature}}}Reference",
            {"URI": "/xl/workbook.xml"},
        )
        ElementTree.SubElement(
            reference,
            f"{{{signature}}}DigestMethod",
            {"Algorithm": "http://www.w3.org/2001/04/xmlenc#sha256"},
        )
        ElementTree.SubElement(reference, f"{{{signature}}}DigestValue").text = (
            "PRIVATE-PACKAGE-DIGEST-BASELINE"
        )
        ElementTree.SubElement(envelope, f"{{{signature}}}SignatureValue").text = (
            "PRIVATE-PACKAGE-SIGNATURE-BASELINE"
        )
        key_info = ElementTree.SubElement(envelope, f"{{{signature}}}KeyInfo")
        certificate_data = ElementTree.SubElement(key_info, f"{{{signature}}}X509Data")
        ElementTree.SubElement(
            certificate_data,
            f"{{{signature}}}X509Certificate",
        ).text = "PRIVATE-SIGNER-CERTIFICATE-BASELINE"
        contents[signature_member] = serialize(envelope)
        certificate_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            certificate_relationships,
            relationship_tag,
            {
                "Id": "rIdFencePackageSignatureCertificate",
                "Type": _DIGITAL_SIGNATURE_CERTIFICATE_RELATIONSHIP,
                "Target": "../package/services/digital-signature/certificate/cert1.cer",
            },
        )
        contents["_xmlsignatures/_rels/sig1.xml.rels"] = serialize(
            certificate_relationships
        )
        contents[certificate_member] = b"PRIVATE-CERTIFICATE-PART-BASELINE"

        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceVbaProject",
                "Type": _VBA_PROJECT_RELATIONSHIP,
                "Target": "vbaProject.bin",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)
        contents["xl/vbaProject.bin"] = b"PRIVATE-VBA-PROJECT-BASELINE"

        vba_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        for suffix in ("", "Agile", "V3"):
            member = f"xl/vbaProjectSignature{suffix}.bin"
            ElementTree.SubElement(
                vba_relationships,
                relationship_tag,
                {
                    "Id": f"rIdFenceVbaSignature{suffix or 'Classic'}",
                    "Type": f"{_VBA_PROJECT_SIGNATURE_RELATIONSHIP}{suffix}",
                    "Target": member.rsplit("/", maxsplit=1)[-1],
                },
            )
            contents[member] = (
                f"PRIVATE-VBA-SIGNATURE-{suffix or 'CLASSIC'}-BASELINE".encode()
            )
        contents["xl/_rels/vbaProject.bin.rels"] = serialize(vba_relationships)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        if not any(
            current.get("Extension", "").casefold() == "sigs"
            for current in content_types.findall(f"{{{_CONTENT_TYPES_NS}}}Default")
        ):
            ElementTree.SubElement(
                content_types,
                f"{{{_CONTENT_TYPES_NS}}}Default",
                {
                    "Extension": "sigs",
                    "ContentType": (
                        "application/vnd.openxmlformats-package.digital-signature-origin"
                    ),
                },
            )
        add_override(
            content_types,
            signature_member,
            "application/vnd.openxmlformats-package.digital-signature-xmlsignature+xml",
        )
        add_override(
            content_types,
            certificate_member,
            "application/vnd.openxmlformats-package.digital-signature-certificate",
        )
        add_override(
            content_types,
            "xl/vbaProject.bin",
            "application/vnd.ms-office.vbaProject",
        )
        for suffix in ("", "Agile", "V3"):
            add_override(
                content_types,
                f"xl/vbaProjectSignature{suffix}.bin",
                f"application/vnd.ms-office.vbaProjectSignature{suffix}",
            )
        contents["[Content_Types].xml"] = serialize(content_types)

    return _rewrite_archive(path, mutate, ".digital-signature-model.tmp.xlsx")


def change_package_signature_reference(path: Path) -> Path:
    """Change a package signed reference without changing a workbook cell."""
    def mutate(contents: dict[str, bytes]) -> None:
        root = _digital_signature_xml_root(contents)
        reference = root.find(
            f"{{{_XML_DIGITAL_SIGNATURE_NS}}}SignedInfo/"
            f"{{{_XML_DIGITAL_SIGNATURE_NS}}}Reference"
        )
        if reference is None:
            raise ValueError("Fixture package signature does not contain a reference")
        reference.set("URI", "/xl/worksheets/sheet1.xml")
        contents["_xmlsignatures/sig1.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".package-signature-reference.tmp.xlsx")


def change_package_signature_certificate_payload(path: Path) -> Path:
    """Change a package certificate part without changing a workbook cell."""
    def mutate(contents: dict[str, bytes]) -> None:
        contents["package/services/digital-signature/certificate/cert1.cer"] = (
            b"PRIVATE-CERTIFICATE-PART-CANDIDATE"
        )

    return _rewrite_archive(path, mutate, ".package-signature-certificate.tmp.xlsx")


def change_vba_project_signature_payload(path: Path) -> Path:
    """Change only the Agile VBA code-signature binary payload."""
    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/vbaProjectSignatureAgile.bin"] = (
            b"PRIVATE-VBA-SIGNATURE-AGILE-CANDIDATE"
        )

    return _rewrite_archive(path, mutate, ".vba-signature-payload.tmp.xlsx")


def rebind_package_signature_relationship(path: Path) -> Path:
    """Move an XML signature part and retarget the origin relationship."""
    def mutate(contents: dict[str, bytes]) -> None:
        old_member = "_xmlsignatures/sig1.xml"
        new_member = "_xmlsignatures/sig2.xml"
        try:
            contents[new_member] = contents.pop(old_member)
        except KeyError as error:
            raise ValueError("Fixture does not contain an XML signature part") from error
        relationships_member = "_xmlsignatures/_rels/origin.sigs.rels"
        relationships = ElementTree.fromstring(contents[relationships_member])
        _digital_signature_relationship(
            relationships,
            _DIGITAL_SIGNATURE_SIGNATURE_RELATIONSHIP,
        ).set("Target", "sig2.xml")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        _digital_signature_content_type_override(content_types, old_member).set(
            "PartName",
            f"/{new_member}",
        )
        contents["[Content_Types].xml"] = ElementTree.tostring(
            content_types,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".package-signature-rebind.tmp.xlsx")


def normalize_digital_signature_control_spelling(path: Path) -> Path:
    """Rewrite equivalent relationship IDs, targets, and base64 whitespace."""
    def mutate(contents: dict[str, bytes]) -> None:
        root_relationships = ElementTree.fromstring(contents["_rels/.rels"])
        _digital_signature_relationship(
            root_relationships,
            _DIGITAL_SIGNATURE_ORIGIN_RELATIONSHIP,
        ).set("Id", "rIdFencePackageSignatureOriginRenumbered")
        contents["_rels/.rels"] = ElementTree.tostring(
            root_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships_member = "_xmlsignatures/_rels/origin.sigs.rels"
        origin_relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = _digital_signature_relationship(
            origin_relationships,
            _DIGITAL_SIGNATURE_SIGNATURE_RELATIONSHIP,
        )
        relationship.set("Id", "rIdFencePackageXmlSignatureRenumbered")
        relationship.set("Target", "./sig1.xml")
        contents[relationships_member] = ElementTree.tostring(
            origin_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        root = _digital_signature_xml_root(contents)
        for local_name in ("DigestValue", "SignatureValue", "X509Certificate"):
            element = next(
                root.iter(f"{{{_XML_DIGITAL_SIGNATURE_NS}}}{local_name}"),
                None,
            )
            if element is None or element.text is None:
                raise ValueError(f"Fixture package signature does not contain {local_name}")
            midpoint = len(element.text) // 2
            element.text = f"  {element.text[:midpoint]}\n{element.text[midpoint:]}  "
        contents["_xmlsignatures/sig1.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".digital-signature-normalize.tmp.xlsx")


def externalize_package_signature_relationship(path: Path) -> Path:
    """Make a package-signature relationship unsafe without following it."""
    def mutate(contents: dict[str, bytes]) -> None:
        relationships_member = "_xmlsignatures/_rels/origin.sigs.rels"
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = _digital_signature_relationship(
            relationships,
            _DIGITAL_SIGNATURE_SIGNATURE_RELATIONSHIP,
        )
        relationship.set(
            "Target",
            "https://private.example.test/PRIVATE-PACKAGE-SIGNATURE",
        )
        relationship.set("TargetMode", "External")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".package-signature-external.tmp.xlsx")


def corrupt_package_signature_root(path: Path) -> Path:
    """Replace an XMLDSIG root to exercise fail-closed package inspection."""
    def mutate(contents: dict[str, bytes]) -> None:
        root = _digital_signature_xml_root(contents)
        root.tag = "privateUnexpectedSignature"
        contents["_xmlsignatures/sig1.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".package-signature-corrupt.tmp.xlsx")


def _rich_data_part_root(
    contents: dict[str, bytes],
    member: str,
) -> ElementTree.Element:
    """Return one raw rich-data fixture part."""
    try:
        return ElementTree.fromstring(contents[member])
    except KeyError as error:
        raise ValueError(f"Fixture does not contain rich-data part {member!r}") from error


def _rich_data_relationship_root(
    contents: dict[str, bytes],
    member: str,
) -> ElementTree.Element:
    """Return the relationship part for one raw rich-data fixture part."""
    return _rich_data_part_root(contents, _relationship_member(member))


def make_rich_data_model(path: Path) -> Path:
    """Create a rich-data package with private entity and endpoint sentinels."""
    make_model(path)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(
        content_types: ElementTree.Element,
        member: str,
        content_type: str,
    ) -> None:
        ElementTree.SubElement(
            content_types,
            f"{{{_CONTENT_TYPES_NS}}}Override",
            {"PartName": f"/{member}", "ContentType": content_type},
        )

    def relationship_root() -> ElementTree.Element:
        return ElementTree.Element(f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships")

    def mutate(contents: dict[str, bytes]) -> None:
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"

        worksheet = _inputs_worksheet_root(contents)
        input_cell = next(
            (
                cell
                for cell in worksheet.iter(f"{{{_SPREADSHEETML_NS}}}c")
                if cell.get("r") == "B2"
            ),
            None,
        )
        if input_cell is None:
            raise ValueError("Fixture Inputs worksheet does not contain B2")
        input_cell.set("vm", "1")
        _save_inputs_worksheet(contents, worksheet)

        metadata = ElementTree.Element(f"{{{_SPREADSHEETML_NS}}}metadata")
        metadata_types = ElementTree.SubElement(
            metadata,
            f"{{{_SPREADSHEETML_NS}}}metadataTypes",
            {"count": "1"},
        )
        ElementTree.SubElement(
            metadata_types,
            f"{{{_SPREADSHEETML_NS}}}metadataType",
            {"name": "XLRICHVALUE"},
        )
        future_metadata = ElementTree.SubElement(
            metadata,
            f"{{{_SPREADSHEETML_NS}}}futureMetadata",
            {"name": "XLRICHVALUE", "count": "2"},
        )
        for index in range(2):
            block = ElementTree.SubElement(
                future_metadata,
                f"{{{_SPREADSHEETML_NS}}}bk",
            )
            extensions = ElementTree.SubElement(
                block,
                f"{{{_SPREADSHEETML_NS}}}extLst",
            )
            extension = ElementTree.SubElement(
                extensions,
                f"{{{_SPREADSHEETML_NS}}}ext",
                {"uri": _RICH_DATA_METADATA_EXTENSION_URI},
            )
            ElementTree.SubElement(
                extension,
                f"{{{_RICH_DATA_NS}}}rvb",
                {"i": str(index)},
            )
        value_metadata = ElementTree.SubElement(
            metadata,
            f"{{{_SPREADSHEETML_NS}}}valueMetadata",
            {"count": "2"},
        )
        for index in range(2):
            block = ElementTree.SubElement(
                value_metadata,
                f"{{{_SPREADSHEETML_NS}}}bk",
            )
            ElementTree.SubElement(
                block,
                f"{{{_SPREADSHEETML_NS}}}rc",
                {"t": "1", "v": str(index)},
            )
        contents["xl/metadata.xml"] = serialize(metadata)

        rich_values = ElementTree.Element(
            f"{{{_RICH_DATA_NS}}}rvData",
            {"count": "2"},
        )
        for index, value in enumerate(
            ("PRIVATE-RICH-ENTITY-BASELINE", "PRIVATE-RICH-ENTITY-SECOND"),
        ):
            rich_value = ElementTree.SubElement(
                rich_values,
                f"{{{_RICH_DATA_NS}}}rv",
                {"s": str(index)},
            )
            ElementTree.SubElement(
                rich_value,
                f"{{{_RICH_DATA_NS}}}v",
            ).text = value
        contents["xl/richData/rdrichvalue.xml"] = serialize(rich_values)

        structures = ElementTree.Element(
            f"{{{_RICH_DATA_NS}}}rvStructures",
            {"count": "2"},
        )
        entity_structure = ElementTree.SubElement(
            structures,
            f"{{{_RICH_DATA_NS}}}s",
            {"t": "_linkedentity2"},
        )
        ElementTree.SubElement(
            entity_structure,
            f"{{{_RICH_DATA_NS}}}k",
            {"n": "PRIVATE-RICH-FIELD", "t": "s"},
        )
        ElementTree.SubElement(
            structures,
            f"{{{_RICH_DATA_NS}}}s",
            {"t": "_formattednumber"},
        )
        contents["xl/richData/rdrichvaluestructure.xml"] = serialize(structures)

        types = ElementTree.Element(f"{{{_RICH_DATA_2_NS}}}rvTypesInfo")
        ElementTree.SubElement(types, f"{{{_RICH_DATA_2_NS}}}types")
        contents["xl/richData/rdRichValueTypes.xml"] = serialize(types)

        arrays = ElementTree.Element(
            f"{{{_RICH_DATA_2_NS}}}arrayData",
            {"count": "1"},
        )
        array = ElementTree.SubElement(
            arrays,
            f"{{{_RICH_DATA_2_NS}}}a",
            {"r": "1"},
        )
        ElementTree.SubElement(
            array,
            f"{{{_RICH_DATA_2_NS}}}v",
            {"t": "s"},
        ).text = "PRIVATE-RICH-ARRAY"
        contents["xl/richData/rdarray.xml"] = serialize(arrays)

        supporting_bags = ElementTree.Element(
            f"{{{_RICH_DATA_2_NS}}}supportingPropertyBags"
        )
        bag_arrays = ElementTree.SubElement(
            supporting_bags,
            f"{{{_RICH_DATA_2_NS}}}spbArrays",
            {"count": "1"},
        )
        bag = ElementTree.SubElement(
            bag_arrays,
            f"{{{_RICH_DATA_2_NS}}}a",
            {"count": "1"},
        )
        ElementTree.SubElement(
            bag,
            f"{{{_RICH_DATA_2_NS}}}v",
            {"t": "s"},
        ).text = "PRIVATE-RICH-PROPERTY"
        contents["xl/richData/rdsupportingpropertybag.xml"] = serialize(
            supporting_bags
        )

        bag_structures = ElementTree.Element(
            f"{{{_RICH_DATA_2_NS}}}spbStructures"
        )
        ElementTree.SubElement(
            bag_structures,
            f"{{{_RICH_DATA_2_NS}}}s",
            {"n": "PRIVATE-RICH-BAG-STRUCTURE"},
        )
        contents["xl/richData/rdsupportingpropertybagstructure.xml"] = serialize(
            bag_structures
        )

        styles = ElementTree.Element(f"{{{_RICH_DATA_2_NS}}}richStyleSheet")
        ElementTree.SubElement(styles, f"{{{_RICH_DATA_2_NS}}}key")
        contents["xl/richData/richStyles.xml"] = serialize(styles)

        web_images = ElementTree.Element(
            f"{{{_RICH_DATA_WEB_IMAGE_NS}}}webImagesSrd"
        )
        web_image = ElementTree.SubElement(
            web_images,
            f"{{{_RICH_DATA_WEB_IMAGE_NS}}}webImageSrd",
        )
        ElementTree.SubElement(
            web_image,
            f"{{{_RICH_DATA_WEB_IMAGE_NS}}}address",
            {relationship_attribute: "rIdFenceRichImage"},
        )
        ElementTree.SubElement(
            web_image,
            f"{{{_RICH_DATA_WEB_IMAGE_NS}}}moreImagesAddress",
            {relationship_attribute: "rIdFenceRichMoreImages"},
        )
        web_image_member = "xl/richData/rdRichValueWebImage.xml"
        contents[web_image_member] = serialize(web_images)
        web_image_relationships = relationship_root()
        ElementTree.SubElement(
            web_image_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceRichImage",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/hyperlink",
                "Target": "https://private.example.test/PRIVATE-RICH-IMAGE-BASELINE",
                "TargetMode": "External",
            },
        )
        ElementTree.SubElement(
            web_image_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceRichMoreImages",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/hyperlink",
                "Target": (
                    "https://private.example.test/PRIVATE-RICH-MORE-IMAGES-BASELINE"
                ),
                "TargetMode": "External",
            },
        )
        contents[_relationship_member(web_image_member)] = serialize(
            web_image_relationships
        )

        rich_value_relationships = ElementTree.Element(
            f"{{{_RICH_VALUE_REL_NS}}}richValueRels"
        )
        ElementTree.SubElement(
            rich_value_relationships,
            f"{{{_RICH_VALUE_REL_NS}}}richValueRel",
            {relationship_attribute: "rIdFenceRichRelationship"},
        )
        rich_value_relationship_member = "xl/richData/richValueRel.xml"
        contents[rich_value_relationship_member] = serialize(rich_value_relationships)
        rich_value_relationship_targets = relationship_root()
        ElementTree.SubElement(
            rich_value_relationship_targets,
            relationship_tag,
            {
                "Id": "rIdFenceRichRelationship",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/hyperlink",
                "Target": (
                    "https://private.example.test/PRIVATE-RICH-RELATIONSHIP-BASELINE"
                ),
                "TargetMode": "External",
            },
        )
        contents[_relationship_member(rich_value_relationship_member)] = serialize(
            rich_value_relationship_targets
        )

        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceRichMetadata",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/sheetMetadata",
                "Target": "metadata.xml",
            },
        )
        rich_parts = {
            "rich-value-data": (
                "xl/richData/rdrichvalue.xml",
                "application/vnd.ms-excel.rdrichvalue+xml",
            ),
            "rich-value-structure": (
                "xl/richData/rdrichvaluestructure.xml",
                "application/vnd.ms-excel.rdrichvaluestructure+xml",
            ),
            "rich-value-types": (
                "xl/richData/rdRichValueTypes.xml",
                "application/vnd.ms-excel.rdrichvaluetypes+xml",
            ),
            "rich-value-array": (
                "xl/richData/rdarray.xml",
                "application/vnd.ms-excel.rdarray+xml",
            ),
            "supporting-property-bag": (
                "xl/richData/rdsupportingpropertybag.xml",
                "application/vnd.ms-excel.rdsupportingpropertybag+xml",
            ),
            "supporting-property-bag-structure": (
                "xl/richData/rdsupportingpropertybagstructure.xml",
                "application/vnd.ms-excel.rdsupportingpropertybagstructure+xml",
            ),
            "rich-styles": (
                "xl/richData/richStyles.xml",
                "application/vnd.ms-excel.richstyles+xml",
            ),
            "rich-value-web-image": (
                web_image_member,
                "application/vnd.ms-excel.rdrichvaluewebimage+xml",
            ),
            "rich-value-relationships": (
                rich_value_relationship_member,
                "application/vnd.ms-excel.richvaluerel+xml",
            ),
        }
        for index, (category, (member, _content_type)) in enumerate(
            rich_parts.items(),
            start=1,
        ):
            ElementTree.SubElement(
                workbook_relationships,
                relationship_tag,
                {
                    "Id": f"rIdFenceRichData{index}",
                    "Type": _RICH_DATA_RELATIONSHIPS[category],
                    "Target": member.removeprefix("xl/"),
                },
            )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        add_override(
            content_types,
            "xl/metadata.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheetMetadata+xml",
        )
        for member, content_type in rich_parts.values():
            add_override(content_types, member, content_type)
        contents["[Content_Types].xml"] = serialize(content_types)

    return _rewrite_archive(path, mutate, ".rich-data-model.tmp.xlsx")


def change_rich_data_value(path: Path) -> Path:
    """Change one provider-backed rich value without changing a worksheet cell."""

    def mutate(contents: dict[str, bytes]) -> None:
        member = "xl/richData/rdrichvalue.xml"
        root = _rich_data_part_root(contents, member)
        value = root.find(f"{{{_RICH_DATA_NS}}}rv/{{{_RICH_DATA_NS}}}v")
        if value is None:
            raise ValueError("Fixture does not contain a rich value")
        value.text = "PRIVATE-RICH-ENTITY-CANDIDATE"
        contents[member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".rich-data-value.tmp.xlsx")


def change_rich_data_binding(path: Path) -> Path:
    """Rebind a worksheet rich value metadata record without changing the cell."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = _inputs_worksheet_root(contents)
        cell = next(
            (
                current
                for current in worksheet.iter(f"{{{_SPREADSHEETML_NS}}}c")
                if current.get("r") == "B2"
            ),
            None,
        )
        if cell is None:
            raise ValueError("Fixture Inputs worksheet does not contain B2")
        cell.set("vm", "2")
        _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".rich-data-binding.tmp.xlsx")


def change_rich_data_web_image_target(path: Path) -> Path:
    """Retarget a rich web image without contacting its endpoint."""

    def mutate(contents: dict[str, bytes]) -> None:
        member = _relationship_member("xl/richData/rdRichValueWebImage.xml")
        relationships = _rich_data_part_root(contents, member)
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
                )
                if current.get("Id") == "rIdFenceRichImage"
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a rich web-image relationship")
        relationship.set(
            "Target",
            "https://private.example.test/PRIVATE-RICH-IMAGE-CANDIDATE",
        )
        contents[member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".rich-data-web-image.tmp.xlsx")


def normalize_rich_data_relationship_ids(path: Path) -> Path:
    """Renumber and reorder rich-data relationships without changing targets."""

    def rewrite_relationship_ids(
        contents: dict[str, bytes],
        member: str,
        replacements: dict[str, str],
    ) -> None:
        relationship_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        relationship_root = _rich_data_relationship_root(contents, member)
        relationships = list(
            relationship_root.findall(f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship")
        )
        for relationship in relationships:
            old_identifier = relationship.get("Id")
            if old_identifier in replacements:
                relationship.set("Id", replacements[old_identifier])
        for relationship in relationships:
            relationship_root.remove(relationship)
        for relationship in reversed(relationships):
            relationship_root.append(relationship)
        contents[_relationship_member(member)] = ElementTree.tostring(
            relationship_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        root = _rich_data_part_root(contents, member)
        for element in root.iter():
            identifier = element.get(relationship_attribute)
            if identifier in replacements:
                element.set(relationship_attribute, replacements[identifier])
        contents[member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    def mutate(contents: dict[str, bytes]) -> None:
        rewrite_relationship_ids(
            contents,
            "xl/richData/rdRichValueWebImage.xml",
            {
                "rIdFenceRichImage": "rIdFenceRenumberedImage",
                "rIdFenceRichMoreImages": "rIdFenceRenumberedMoreImages",
            },
        )
        rewrite_relationship_ids(
            contents,
            "xl/richData/richValueRel.xml",
            {
                "rIdFenceRichRelationship": "rIdFenceRenumberedRelationship",
            },
        )

    return _rewrite_archive(path, mutate, ".rich-data-relationship-ids.tmp.xlsx")


def corrupt_rich_data_value_root(path: Path) -> Path:
    """Corrupt a rich value root to exercise fail-closed inspection."""

    def mutate(contents: dict[str, bytes]) -> None:
        member = "xl/richData/rdrichvalue.xml"
        root = _rich_data_part_root(contents, member)
        root.tag = "privateUnexpectedRichData"
        contents[member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".rich-data-corrupt.tmp.xlsx")


def make_custom_data_store_model(path: Path) -> Path:
    """Create generic custom XML, binary data, and custom-property stores."""
    make_model(path)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(
        content_types: ElementTree.Element,
        member: str,
        content_type: str,
    ) -> None:
        ElementTree.SubElement(
            content_types,
            f"{{{_CONTENT_TYPES_NS}}}Override",
            {"PartName": f"/{member}", "ContentType": content_type},
        )

    def mutate(contents: dict[str, bytes]) -> None:
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"

        root_relationships = ElementTree.fromstring(contents["_rels/.rels"])
        ElementTree.SubElement(
            root_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceCustomXml",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/customXml",
                "Target": "customXml/item1.xml",
            },
        )
        ElementTree.SubElement(
            root_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceCustomDocumentProperties",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/custom-properties",
                "Target": "docProps/custom.xml",
            },
        )
        contents["_rels/.rels"] = serialize(root_relationships)

        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceCustomDataProperties",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/customDataProps",
                "Target": "customData/itemProps1.xml",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        custom_xml = ElementTree.Element("{urn:formulafence:test}reviewState")
        ElementTree.SubElement(
            custom_xml,
            "{urn:formulafence:test}reviewGate",
        ).text = "PRIVATE-CUSTOM-XML-BASELINE"
        contents["customXml/item1.xml"] = serialize(custom_xml)
        custom_xml_properties = ElementTree.Element(
            f"{{{_CUSTOM_XML_DATA_PROPERTIES_NS}}}datastoreItem",
            {f"{{{_CUSTOM_XML_DATA_PROPERTIES_NS}}}itemID": "{PRIVATE-CUSTOM-XML-ID}"},
        )
        schema_references = ElementTree.SubElement(
            custom_xml_properties,
            f"{{{_CUSTOM_XML_DATA_PROPERTIES_NS}}}schemaRefs",
        )
        ElementTree.SubElement(
            schema_references,
            f"{{{_CUSTOM_XML_DATA_PROPERTIES_NS}}}schemaRef",
            {"uri": "urn:formulafence:private-custom-xml-schema"},
        )
        contents["customXml/itemProps1.xml"] = serialize(custom_xml_properties)
        custom_xml_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            custom_xml_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceCustomXmlProperties",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/customXmlProps",
                "Target": "itemProps1.xml",
            },
        )
        contents["customXml/_rels/item1.xml.rels"] = serialize(custom_xml_relationships)

        custom_data_properties = ElementTree.Element(
            f"{{{_CUSTOM_DATA_PROPERTIES_NS}}}datastoreItem",
            {"id": "{PRIVATE-CUSTOM-DATA-ID}"},
        )
        contents["xl/customData/itemProps1.xml"] = serialize(custom_data_properties)
        custom_data_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            custom_data_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceCustomDataPayload",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/customData",
                "Target": "item1.bin",
            },
        )
        contents["xl/customData/_rels/itemProps1.xml.rels"] = serialize(
            custom_data_relationships
        )
        contents["xl/customData/item1.bin"] = b"PRIVATE-CUSTOM-DATA-BASELINE"

        document_properties = ElementTree.Element(
            f"{{{_CUSTOM_DOCUMENT_PROPERTIES_NS}}}Properties"
        )
        document_property = ElementTree.SubElement(
            document_properties,
            f"{{{_CUSTOM_DOCUMENT_PROPERTIES_NS}}}property",
            {
                "fmtid": "{D5CDD505-2E9C-101B-9397-08002B2CF9AE}",
                "pid": "2",
                "name": "PRIVATE-CUSTOM-DOCUMENT-PROPERTY-NAME",
            },
        )
        ElementTree.SubElement(
            document_property,
            f"{{{_DOCUMENT_PROPERTY_TYPES_NS}}}lpwstr",
        ).text = "PRIVATE-CUSTOM-DOCUMENT-PROPERTY-BASELINE"
        contents["docProps/custom.xml"] = serialize(document_properties)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        add_override(
            content_types,
            "customXml/itemProps1.xml",
            "application/vnd.openxmlformats-officedocument.customXmlProperties+xml",
        )
        add_override(
            content_types,
            "xl/customData/itemProps1.xml",
            "application/vnd.openxmlformats-officedocument.customDataProperties+xml",
        )
        add_override(content_types, "xl/customData/item1.bin", "application/binary")
        add_override(
            content_types,
            "docProps/custom.xml",
            "application/vnd.openxmlformats-officedocument.custom-properties+xml",
        )
        contents["[Content_Types].xml"] = serialize(content_types)

    return _rewrite_archive(path, mutate, ".custom-data-store-model.tmp.xlsx")


def change_custom_xml_data_store_value(path: Path) -> Path:
    """Change persisted generic custom XML without changing a worksheet cell."""
    def mutate(contents: dict[str, bytes]) -> None:
        root = ElementTree.fromstring(contents["customXml/item1.xml"])
        gate = root.find("{urn:formulafence:test}reviewGate")
        if gate is None:
            raise ValueError("Fixture custom XML does not contain its review gate")
        gate.text = "PRIVATE-CUSTOM-XML-CANDIDATE"
        contents["customXml/item1.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".custom-xml-data-store-change.tmp.xlsx")


def change_custom_data_payload(path: Path) -> Path:
    """Change opaque custom binary data without changing a worksheet cell."""
    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/customData/item1.bin"] = b"PRIVATE-CUSTOM-DATA-CANDIDATE"

    return _rewrite_archive(path, mutate, ".custom-data-payload-change.tmp.xlsx")


def change_custom_document_property_value(path: Path) -> Path:
    """Change a private custom document property without changing a worksheet cell."""
    def mutate(contents: dict[str, bytes]) -> None:
        root = ElementTree.fromstring(contents["docProps/custom.xml"])
        value = root.find(
            f"{{{_CUSTOM_DOCUMENT_PROPERTIES_NS}}}property/"
            f"{{{_DOCUMENT_PROPERTY_TYPES_NS}}}lpwstr"
        )
        if value is None:
            raise ValueError("Fixture does not contain a custom document property")
        value.text = "PRIVATE-CUSTOM-DOCUMENT-PROPERTY-CANDIDATE"
        contents["docProps/custom.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".custom-document-property-change.tmp.xlsx")


def normalize_custom_data_store_identifiers(path: Path) -> Path:
    """Rewrite nonsemantic relationship and document-property IDs."""
    def mutate(contents: dict[str, bytes]) -> None:
        for member in (
            "_rels/.rels",
            "xl/_rels/workbook.xml.rels",
            "customXml/_rels/item1.xml.rels",
            "xl/customData/_rels/itemProps1.xml.rels",
        ):
            root = ElementTree.fromstring(contents[member])
            relationships = list(
                root.findall(f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship")
            )
            for index, relationship in enumerate(relationships, start=801):
                if (relationship.get("Id") or "").startswith("rIdFence"):
                    relationship.set("Id", f"rIdFenceCustomStoreNormalised{index}")
            contents[member] = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )

        document_properties = ElementTree.fromstring(contents["docProps/custom.xml"])
        property_element = document_properties.find(
            f"{{{_CUSTOM_DOCUMENT_PROPERTIES_NS}}}property"
        )
        if property_element is None:
            raise ValueError("Fixture does not contain a custom document property")
        property_element.set("pid", "901")
        contents["docProps/custom.xml"] = ElementTree.tostring(
            document_properties,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".custom-data-store-ids.tmp.xlsx")


def change_custom_data_store_storage_identifiers(path: Path) -> Path:
    """Change private storage identities that can bind persisted add-in state."""
    def mutate(contents: dict[str, bytes]) -> None:
        custom_xml_properties = ElementTree.fromstring(
            contents["customXml/itemProps1.xml"]
        )
        custom_xml_properties.set(
            f"{{{_CUSTOM_XML_DATA_PROPERTIES_NS}}}itemID",
            "{PRIVATE-CUSTOM-XML-ID-CANDIDATE}",
        )
        contents["customXml/itemProps1.xml"] = ElementTree.tostring(
            custom_xml_properties,
            encoding="utf-8",
            xml_declaration=True,
        )

        custom_data_properties = ElementTree.fromstring(
            contents["xl/customData/itemProps1.xml"]
        )
        custom_data_properties.set("id", "{PRIVATE-CUSTOM-DATA-ID-CANDIDATE}")
        contents["xl/customData/itemProps1.xml"] = ElementTree.tostring(
            custom_data_properties,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".custom-data-store-storage-id-change.tmp.xlsx")


def corrupt_custom_data_properties_root(path: Path) -> Path:
    """Corrupt custom-data properties to exercise fail-closed coverage."""
    def mutate(contents: dict[str, bytes]) -> None:
        root = ElementTree.fromstring(contents["xl/customData/itemProps1.xml"])
        root.tag = "privateUnexpectedCustomDataProperties"
        contents["xl/customData/itemProps1.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".custom-data-store-corrupt.tmp.xlsx")


def _workbook_theme_member(contents: dict[str, bytes]) -> str:
    """Return the workbook-bound Theme part used by the fixture."""
    relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
    relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
    relationship = next(
        (
            current
            for current in relationships.findall(relationship_tag)
            if (current.get("Type") or "").casefold().endswith("/theme")
        ),
        None,
    )
    if relationship is None or not relationship.get("Target"):
        raise ValueError("Fixture does not contain a workbook Theme relationship")
    target = relationship.get("Target") or ""
    member = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
    if member not in contents:
        raise ValueError("Fixture Theme relationship target is missing")
    return member


def make_workbook_theme_image_model(path: Path) -> Path:
    """Create a real Theme-part image relationship without changing cells."""
    make_model(path)

    def mutate(contents: dict[str, bytes]) -> None:
        theme_member = _workbook_theme_member(contents)
        theme = ElementTree.fromstring(contents[theme_member])
        theme.set("name", "PRIVATE-THEME-SCHEME-BASELINE")
        format_scheme = theme.find(
            f".//{{{_DRAWINGML_MAIN_NS}}}fmtScheme/"
            f"{{{_DRAWINGML_MAIN_NS}}}fillStyleLst"
        )
        if format_scheme is None:
            raise ValueError("Fixture Theme does not contain a fill style list")
        blip_fill = ElementTree.SubElement(
            format_scheme,
            f"{{{_DRAWINGML_MAIN_NS}}}blipFill",
            {"rotateWithShape": "1"},
        )
        ElementTree.SubElement(
            blip_fill,
            f"{{{_DRAWINGML_MAIN_NS}}}blip",
            {
                f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}embed": (
                    "rIdFenceThemeImage"
                )
            },
        )
        stretch = ElementTree.SubElement(
            blip_fill,
            f"{{{_DRAWINGML_MAIN_NS}}}stretch",
        )
        ElementTree.SubElement(stretch, f"{{{_DRAWINGML_MAIN_NS}}}fillRect")
        contents[theme_member] = ElementTree.tostring(
            theme,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_member = _relationship_member(theme_member)
        relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            relationships,
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship",
            {
                "Id": "rIdFenceThemeImage",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/image",
                "Target": "media/fence-theme-image.bin",
            },
        )
        contents[relationship_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/theme/media/fence-theme-image.bin"] = (
            b"PRIVATE-THEME-IMAGE-BASELINE"
        )

    return _rewrite_archive(path, mutate, ".workbook-theme-image-model.tmp.xlsx")


def make_strict_workbook_theme_image_model(path: Path) -> Path:
    """Create a strict-namespace Theme with a direct image relationship."""
    make_workbook_theme_image_model(path)

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        theme_member = _workbook_theme_member(contents)
        theme = ElementTree.fromstring(contents[theme_member])
        for element in theme.iter():
            element.tag = strict_name(
                element.tag,
                _DRAWINGML_MAIN_NS,
                _DRAWINGML_STRICT_MAIN_NS,
            )
            attributes = {
                strict_name(
                    name,
                    _DOCUMENT_RELATIONSHIPS_NS,
                    _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                ): value
                for name, value in element.attrib.items()
            }
            element.attrib.clear()
            element.attrib.update(attributes)
        contents[theme_member] = ElementTree.tostring(
            theme,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        for relationship in workbook_relationships.findall(relationship_tag):
            if (relationship.get("Type") or "").casefold().endswith("/theme"):
                relationship.set(
                    "Type",
                    f"{_STRICT_DOCUMENT_RELATIONSHIPS_NS}/theme",
                )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            workbook_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_member = _relationship_member(theme_member)
        theme_relationships = ElementTree.fromstring(contents[relationship_member])
        for relationship in theme_relationships.findall(relationship_tag):
            if (relationship.get("Type") or "").casefold().endswith("/image"):
                relationship.set(
                    "Type",
                    f"{_STRICT_DOCUMENT_RELATIONSHIPS_NS}/image",
                )
        contents[relationship_member] = ElementTree.tostring(
            theme_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".strict-workbook-theme-image-model.tmp.xlsx")


def change_workbook_theme_colour(path: Path) -> Path:
    """Change a Theme colour without changing a stored local cell style."""

    def mutate(contents: dict[str, bytes]) -> None:
        theme_member = _workbook_theme_member(contents)
        theme = ElementTree.fromstring(contents[theme_member])
        accent = theme.find(f".//{{{_DRAWINGML_MAIN_NS}}}accent1")
        if accent is None or len(accent) != 1:
            raise ValueError("Fixture Theme does not contain accent1")
        colour = accent[0]
        if "val" in colour.attrib:
            colour.set("val", "C00000")
        elif "lastClr" in colour.attrib:
            colour.set("lastClr", "C00000")
        else:
            raise ValueError("Fixture Theme accent1 does not contain a colour value")
        contents[theme_member] = ElementTree.tostring(
            theme,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".workbook-theme-colour-change.tmp.xlsx")


def change_workbook_theme_image_payload(path: Path) -> Path:
    """Change a direct Theme image payload without changing a worksheet cell."""

    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/theme/media/fence-theme-image.bin"] = (
            b"PRIVATE-THEME-IMAGE-CANDIDATE"
        )

    return _rewrite_archive(path, mutate, ".workbook-theme-image-change.tmp.xlsx")


def normalize_workbook_theme_relationship_identifiers(path: Path) -> Path:
    """Rewrite only writer-selected Theme relationship identifiers."""

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        theme_relationship = next(
            (
                current
                for current in workbook_relationships.findall(relationship_tag)
                if (current.get("Type") or "").casefold().endswith("/theme")
            ),
            None,
        )
        if theme_relationship is None:
            raise ValueError("Fixture does not contain a workbook Theme relationship")
        theme_relationship.set("Id", "rIdFenceRenumberedTheme")
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            workbook_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        theme_member = _workbook_theme_member(contents)
        relationship_member = _relationship_member(theme_member)
        theme_relationships = ElementTree.fromstring(contents[relationship_member])
        image_relationship = next(
            (
                current
                for current in theme_relationships.findall(relationship_tag)
                if (current.get("Type") or "").casefold().endswith("/image")
            ),
            None,
        )
        if image_relationship is None:
            raise ValueError("Fixture Theme does not contain an image relationship")
        old_identifier = image_relationship.get("Id")
        if not old_identifier:
            raise ValueError("Fixture Theme image relationship does not have an Id")
        new_identifier = "rIdFenceRenumberedThemeImage"
        image_relationship.set("Id", new_identifier)
        contents[relationship_member] = ElementTree.tostring(
            theme_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        theme = ElementTree.fromstring(contents[theme_member])
        embed_attributes = (
            f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}embed",
            f"{{{_STRICT_DOCUMENT_RELATIONSHIPS_NS}}}embed",
        )
        blip_and_attribute = next(
            (
                (current, attribute)
                for current in theme.iter()
                for attribute in embed_attributes
                if current.get(attribute) == old_identifier
            ),
            None,
        )
        if blip_and_attribute is None:
            raise ValueError("Fixture Theme does not contain an image binding")
        blip, embed_attribute = blip_and_attribute
        blip.set(embed_attribute, new_identifier)
        contents[theme_member] = ElementTree.tostring(
            theme,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".workbook-theme-ids.tmp.xlsx")


def corrupt_workbook_theme_root(path: Path) -> Path:
    """Corrupt Theme XML to exercise fail-closed coverage behavior."""

    def mutate(contents: dict[str, bytes]) -> None:
        theme_member = _workbook_theme_member(contents)
        theme = ElementTree.fromstring(contents[theme_member])
        theme.tag = "privateUnexpectedWorkbookTheme"
        contents[theme_member] = ElementTree.tostring(
            theme,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".workbook-theme-corrupt.tmp.xlsx")


def make_table_model(path: Path) -> Path:
    """Create a small workbook with static Excel-table references."""
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(["Amount", "Rate", "Value"])
    data.append([10, 0.1, 1])
    data.append([20, 0.2, 4])
    data.append([30, 0.3, 9])
    data.add_table(Table(displayName="Sales", ref="A1:C4"))

    report = workbook.create_sheet("Report")
    report["A1"] = "Table-driven outputs"
    report["B2"] = "=SUM(Sales[Amount])"
    report["B3"] = "=SUM(Sales[[#Data],[Amount]:[Rate]])"
    report["B4"] = "=ROWS(Sales[#All])"
    report["B5"] = "=COUNTA(Sales[[#Headers],[#Data],[Rate]])"
    workbook.save(path)
    return path


def _table_style_control_table_member(contents: dict[str, bytes]) -> str:
    """Return the one ordinary table part used by the Table Style fixture."""
    members = sorted(
        member
        for member in contents
        if member.startswith("xl/tables/") and member.endswith(".xml")
    )
    if len(members) != 1:
        raise ValueError("Fixture needs exactly one table package part")
    return members[0]


def _table_style_control_style_container(
    styles: ElementTree.Element,
) -> ElementTree.Element:
    """Return the fixture's ``tableStyles`` container."""
    container = styles.find(f"{{{_SPREADSHEETML_NS}}}tableStyles")
    if container is None:
        raise ValueError("Fixture styles XML does not contain tableStyles")
    return container


def _table_style_control_dxfs(styles: ElementTree.Element) -> ElementTree.Element:
    """Return or create the fixture's differential-format container."""
    dxfs = styles.find(f"{{{_SPREADSHEETML_NS}}}dxfs")
    if dxfs is not None:
        return dxfs
    dxfs = ElementTree.Element(f"{{{_SPREADSHEETML_NS}}}dxfs", {"count": "0"})
    table_styles = _table_style_control_style_container(styles)
    styles.insert(list(styles).index(table_styles), dxfs)
    return dxfs


def make_table_style_control_model(path: Path) -> Path:
    """Create an applied private custom Table Style backed by three Dxfs."""
    make_table_model(path)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_fill_dxf(dxfs: ElementTree.Element, color: str) -> None:
        dxf = ElementTree.SubElement(dxfs, f"{{{_SPREADSHEETML_NS}}}dxf")
        fill = ElementTree.SubElement(dxf, f"{{{_SPREADSHEETML_NS}}}fill")
        pattern = ElementTree.SubElement(
            fill,
            f"{{{_SPREADSHEETML_NS}}}patternFill",
            {"patternType": "solid"},
        )
        ElementTree.SubElement(
            pattern,
            f"{{{_SPREADSHEETML_NS}}}fgColor",
            {"rgb": color},
        )
        ElementTree.SubElement(
            pattern,
            f"{{{_SPREADSHEETML_NS}}}bgColor",
            {"indexed": "64"},
        )

    def mutate(contents: dict[str, bytes]) -> None:
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        dxfs = _table_style_control_dxfs(styles)
        add_fill_dxf(dxfs, "FF113355")
        add_fill_dxf(dxfs, "FF557799")
        add_fill_dxf(dxfs, "FF99BBDD")
        dxfs.set("count", str(len(dxfs)))

        table_styles = _table_style_control_style_container(styles)
        table_styles.set("count", "1")
        custom_style = ElementTree.SubElement(
            table_styles,
            f"{{{_SPREADSHEETML_NS}}}tableStyle",
            {
                "name": "PRIVATE-FINANCE-TABLE-STYLE",
                "pivot": "false",
                "table": "true",
                f"{{{_OFFICE_2016_REVISION9_NS}}}uid": (
                    "{01234567-89AB-CDEF-0123-456789ABCDEF}"
                ),
            },
        )
        ElementTree.SubElement(
            custom_style,
            f"{{{_SPREADSHEETML_NS}}}tableStyleElement",
            {"type": "wholeTable", "dxfId": "0"},
        )
        ElementTree.SubElement(
            custom_style,
            f"{{{_SPREADSHEETML_NS}}}tableStyleElement",
            {"type": "headerRow", "dxfId": "1"},
        )
        ElementTree.SubElement(
            custom_style,
            f"{{{_SPREADSHEETML_NS}}}tableStyleElement",
            {"type": "firstRowStripe", "size": "1", "dxfId": "2"},
        )
        # A Pivot-only region is intentionally present to verify the ordinary
        # worksheet Table Style inventory does not absorb PivotTable controls.
        ElementTree.SubElement(
            custom_style,
            f"{{{_SPREADSHEETML_NS}}}tableStyleElement",
            {"type": "firstSubtotalRow", "dxfId": "0"},
        )
        contents["xl/styles.xml"] = serialize(styles)

        table_member = _table_style_control_table_member(contents)
        table = ElementTree.fromstring(contents[table_member])
        table_columns = table.find(f"{{{_SPREADSHEETML_NS}}}tableColumns")
        if table_columns is None:
            raise ValueError("Fixture table does not contain tableColumns")
        first_column = table_columns.find(f"{{{_SPREADSHEETML_NS}}}tableColumn")
        if first_column is None:
            raise ValueError("Fixture table does not contain a table column")
        first_column.set("dataDxfId", "2")
        first_column.set("dataCellStyle", "PRIVATE-FINANCE-DATA-STYLE")
        ElementTree.SubElement(
            table,
            f"{{{_SPREADSHEETML_NS}}}tableStyleInfo",
            {
                "name": "PRIVATE-FINANCE-TABLE-STYLE",
                "showFirstColumn": "false",
                "showLastColumn": "true",
                "showRowStripes": "true",
                "showColumnStripes": "false",
            },
        )
        contents[table_member] = serialize(table)

    return _rewrite_archive(path, mutate, ".table-style-controls.tmp.xlsx")


def change_table_style_control(path: Path) -> Path:
    """Change an applied Table Style toggle without modifying a table range."""

    def mutate(contents: dict[str, bytes]) -> None:
        table_member = _table_style_control_table_member(contents)
        table = ElementTree.fromstring(contents[table_member])
        style_info = table.find(f"{{{_SPREADSHEETML_NS}}}tableStyleInfo")
        if style_info is None:
            raise ValueError("Fixture table does not contain tableStyleInfo")
        style_info.set("showRowStripes", "0")
        style_info.set("showColumnStripes", "1")
        contents[table_member] = ElementTree.tostring(
            table,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".table-style-toggle.tmp.xlsx")


def change_table_style_definition(path: Path) -> Path:
    """Change a custom Table Style Dxf without modifying table cells or refs."""

    def mutate(contents: dict[str, bytes]) -> None:
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        colors = list(styles.iter(f"{{{_SPREADSHEETML_NS}}}fgColor"))
        if len(colors) < 2:
            raise ValueError("Fixture styles XML does not contain custom Table Style Dxfs")
        colors[1].set("rgb", "FFCC8844")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".table-style-definition.tmp.xlsx")


def change_table_direct_dxf_assignment(path: Path) -> Path:
    """Change a table-column Dxf binding without changing any table content."""

    def mutate(contents: dict[str, bytes]) -> None:
        table_member = _table_style_control_table_member(contents)
        table = ElementTree.fromstring(contents[table_member])
        column = table.find(
            f"{{{_SPREADSHEETML_NS}}}tableColumns/"
            f"{{{_SPREADSHEETML_NS}}}tableColumn"
        )
        if column is None:
            raise ValueError("Fixture table does not contain a table column")
        column.set("dataDxfId", "1")
        contents[table_member] = ElementTree.tostring(
            table,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".table-direct-dxf.tmp.xlsx")


def normalize_table_style_control_writer_noise(path: Path) -> Path:
    """Rewrite equivalent Table Style spelling and Dxf indexes."""

    def mutate(contents: dict[str, bytes]) -> None:
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        dxfs = _table_style_control_dxfs(styles)
        original_dxfs = list(dxfs)
        if len(original_dxfs) != 3:
            raise ValueError("Fixture styles XML needs exactly three custom Table Style Dxfs")
        # Writer A may reorder differential formats while consistently changing
        # every corresponding dxfId. The semantic style is unchanged.
        reorder = (2, 0, 1)
        old_to_new = {old: new for new, old in enumerate(reorder)}
        for child in original_dxfs:
            dxfs.remove(child)
        for old_index in reorder:
            dxfs.append(original_dxfs[old_index])
        for element in styles.iter(f"{{{_SPREADSHEETML_NS}}}tableStyleElement"):
            raw_dxf_id = element.get("dxfId")
            if raw_dxf_id is not None:
                element.set("dxfId", str(old_to_new[int(raw_dxf_id)]))
        style = next(styles.iter(f"{{{_SPREADSHEETML_NS}}}tableStyle"))
        style.set("name", "private-finance-table-style")
        style.set("pivot", "0")
        style.set("table", "1")
        style.set(
            f"{{{_OFFICE_2016_REVISION9_NS}}}uid",
            "{11111111-2222-3333-4444-555555555555}",
        )
        stripe = next(
            element
            for element in styles.iter(f"{{{_SPREADSHEETML_NS}}}tableStyleElement")
            if element.get("type") == "firstRowStripe"
        )
        stripe.set("size", "01")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

        table_member = _table_style_control_table_member(contents)
        table = ElementTree.fromstring(contents[table_member])
        style_info = table.find(f"{{{_SPREADSHEETML_NS}}}tableStyleInfo")
        if style_info is None:
            raise ValueError("Fixture table does not contain tableStyleInfo")
        style_info.set("name", "private-finance-table-style")
        style_info.set("showFirstColumn", "0")
        style_info.set("showLastColumn", "1")
        style_info.set("showRowStripes", "1")
        style_info.attrib.pop("showColumnStripes", None)
        for element in table.iter():
            for attribute in (
                "headerRowDxfId",
                "dataDxfId",
                "totalsRowDxfId",
                "headerRowBorderDxfId",
                "tableBorderDxfId",
                "totalsRowBorderDxfId",
            ):
                raw_dxf_id = element.get(attribute)
                if raw_dxf_id is not None:
                    element.set(attribute, str(old_to_new[int(raw_dxf_id)]))
        column = table.find(
            f"{{{_SPREADSHEETML_NS}}}tableColumns/"
            f"{{{_SPREADSHEETML_NS}}}tableColumn"
        )
        if column is None:
            raise ValueError("Fixture table does not contain a table column")
        column.set("dataCellStyle", "private-finance-data-style")
        contents[table_member] = ElementTree.tostring(
            table,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".table-style-noise.tmp.xlsx")


def make_strict_table_style_control_model(path: Path) -> Path:
    """Create a Strict SpreadsheetML Table Style fixture."""
    make_table_style_control_model(path)

    def strict_name(name: str) -> str:
        prefix = f"{{{_SPREADSHEETML_NS}}}"
        if name.startswith(prefix):
            return f"{{{_STRICT_SPREADSHEETML_NS}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        for member in (
            "xl/styles.xml",
            _table_style_control_table_member(contents),
        ):
            root = ElementTree.fromstring(contents[member])
            for element in root.iter():
                element.tag = strict_name(element.tag)
            contents[member] = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".strict-table-style-controls.tmp.xlsx")


def corrupt_table_style_control(path: Path) -> Path:
    """Inject malformed private Table Style metadata for fail-closed coverage."""

    def mutate(contents: dict[str, bytes]) -> None:
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        style_element = next(
            styles.iter(f"{{{_SPREADSHEETML_NS}}}tableStyleElement")
        )
        style_element.set("dxfId", "4294967296")
        style_element.set("privateControl", "PRIVATE-TABLE-STYLE-CONTROL")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )
        table_member = _table_style_control_table_member(contents)
        table = ElementTree.fromstring(contents[table_member])
        column = table.find(
            f"{{{_SPREADSHEETML_NS}}}tableColumns/"
            f"{{{_SPREADSHEETML_NS}}}tableColumn"
        )
        if column is None:
            raise ValueError("Fixture table does not contain a table column")
        column.set("dataDxfId", "4294967296")
        contents[table_member] = ElementTree.tostring(
            table,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".table-style-corrupt.tmp.xlsx")


_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER = "xl/revisions/revisionHeaders.xml"
_SHARED_WORKBOOK_REVISION_LOG_MEMBER = "xl/revisions/revisionLog1.xml"
_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER = (
    "xl/revisions/_rels/revisionHeaders.xml.rels"
)


def _shared_workbook_revision_override(
    content_types: ElementTree.Element,
    member: str,
    content_type: str,
) -> None:
    """Register one private revision part with the fixture package."""
    if any(
        current.get("PartName") == f"/{member}"
        for current in content_types.findall(f"{{{_CONTENT_TYPES_NS}}}Override")
    ):
        raise ValueError(f"Fixture already contains revision content type for {member!r}")
    ElementTree.SubElement(
        content_types,
        f"{{{_CONTENT_TYPES_NS}}}Override",
        {"PartName": f"/{member}", "ContentType": content_type},
    )


def make_shared_workbook_revision_model(path: Path) -> Path:
    """Create a workbook with a private legacy shared-workbook audit trail.

    ``openpyxl`` does not author the obsolete shared-workbook revision parts,
    so this fixture deliberately injects a relationship-backed header and log
    after writing an ordinary model. The cell history contains private values,
    identities, dates, and identifiers that reporting must never expose.
    """
    make_model(path)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        _shared_workbook_revision_override(
            content_types,
            _SHARED_WORKBOOK_REVISION_HEADERS_MEMBER,
            "application/vnd.openxmlformats-officedocument.spreadsheetml."
            "revisionHeaders+xml",
        )
        _shared_workbook_revision_override(
            content_types,
            _SHARED_WORKBOOK_REVISION_LOG_MEMBER,
            "application/vnd.openxmlformats-officedocument.spreadsheetml."
            "revisionLog+xml",
        )
        contents["[Content_Types].xml"] = serialize(content_types)

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdPrivateRevisionHeaders",
                "Type": _SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIP,
                "Target": "revisions/revisionHeaders.xml",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        headers = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}headers",
            {
                "diskRevisions": "1",
                "exclusive": "0",
                "history": "true",
                "keepChangeHistory": "true",
                "protected": "true",
                "shared": "1",
                "trackRevisions": "true",
                "preserveHistory": "30",
                "revisionId": "7",
                "version": "+2",
            },
        )
        ElementTree.SubElement(
            headers,
            f"{{{_SPREADSHEETML_NS}}}header",
            {
                "guid": "{11111111-2222-3333-4444-555555555555}",
                "dateTime": "2026-07-26T01:02:03Z",
                "maxSheetId": "4",
                "userName": "PRIVATE-REVISION-AUTHOR",
                f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": "rIdPrivateRevisionLog",
            },
        )
        contents[_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER] = serialize(headers)

        header_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            header_relationships,
            relationship_tag,
            {
                "Id": "rIdPrivateRevisionLog",
                "Type": _SHARED_WORKBOOK_REVISION_LOG_RELATIONSHIP,
                "Target": "revisionLog1.xml",
            },
        )
        contents[_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER] = serialize(
            header_relationships
        )

        revisions = ElementTree.Element(f"{{{_SPREADSHEETML_NS}}}revisions")
        changed_cells = ElementTree.SubElement(
            revisions,
            f"{{{_SPREADSHEETML_NS}}}rcc",
            {"rId": "1", "sId": "1"},
        )
        old_cell = ElementTree.SubElement(
            changed_cells,
            f"{{{_SPREADSHEETML_NS}}}oc",
            {"r": "B2"},
        )
        ElementTree.SubElement(
            old_cell,
            f"{{{_SPREADSHEETML_NS}}}v",
        ).text = "PRIVATE-OLD-REVISION-VALUE"
        new_cell = ElementTree.SubElement(
            changed_cells,
            f"{{{_SPREADSHEETML_NS}}}nc",
            {"r": "B2"},
        )
        ElementTree.SubElement(
            new_cell,
            f"{{{_SPREADSHEETML_NS}}}v",
        ).text = "PRIVATE-NEW-REVISION-VALUE"
        ElementTree.SubElement(
            revisions,
            f"{{{_SPREADSHEETML_NS}}}rrc",
            {
                "rId": "2",
                "sId": "1",
                "ref": "A1:A1048576",
                "action": "insertCol",
            },
        )
        ElementTree.SubElement(
            revisions,
            f"{{{_SPREADSHEETML_NS}}}rfmt",
            {"sheetId": "1", "sqref": "B2"},
        )
        contents[_SHARED_WORKBOOK_REVISION_LOG_MEMBER] = serialize(revisions)

    return _rewrite_archive(path, mutate, ".shared-workbook-revisions.tmp.xlsx")


def change_shared_workbook_revision_log(path: Path) -> Path:
    """Change private historic data without touching visible workbook cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        revisions = ElementTree.fromstring(
            contents[_SHARED_WORKBOOK_REVISION_LOG_MEMBER]
        )
        old_value = next(
            (
                current
                for current in revisions.iter(f"{{{_SPREADSHEETML_NS}}}v")
                if current.text == "PRIVATE-OLD-REVISION-VALUE"
            ),
            None,
        )
        if old_value is None:
            raise ValueError("Fixture revision log does not contain its private old value")
        old_value.text = "CANDIDATE-PRIVATE-OLD-REVISION-VALUE"
        contents[_SHARED_WORKBOOK_REVISION_LOG_MEMBER] = ElementTree.tostring(
            revisions,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".shared-workbook-revision-log.tmp.xlsx")


def change_shared_workbook_revision_controls(path: Path) -> Path:
    """Change retention and protection controls without modifying history cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        headers = ElementTree.fromstring(
            contents[_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER]
        )
        headers.set("keepChangeHistory", "false")
        headers.set("protected", "0")
        contents[_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER] = ElementTree.tostring(
            headers,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".shared-workbook-revision-controls.tmp.xlsx")


def normalize_shared_workbook_revision_writer_noise(path: Path) -> Path:
    """Use equivalent scalar spellings and relationship identifiers."""

    def mutate(contents: dict[str, bytes]) -> None:
        headers = ElementTree.fromstring(
            contents[_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER]
        )
        headers.set("diskRevisions", "true")
        headers.set("exclusive", "false")
        headers.set("history", "1")
        headers.set("keepChangeHistory", "1")
        headers.set("protected", "1")
        headers.set("shared", "true")
        headers.set("trackRevisions", "1")
        headers.set("preserveHistory", "030")
        headers.set("revisionId", "0007")
        headers.set("version", "2")
        header = headers.find(f"{{{_SPREADSHEETML_NS}}}header")
        if header is None:
            raise ValueError("Fixture revision headers does not contain a header")
        identifier_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        header.set(identifier_attribute, "rIdWriterRevisionLog")
        contents[_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER] = ElementTree.tostring(
            headers,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships = ElementTree.fromstring(
            contents[_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER]
        )
        relationship = relationships.find(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        )
        if relationship is None:
            raise ValueError("Fixture revision headers does not contain a relationship")
        relationship.set("Id", "rIdWriterRevisionLog")
        contents[_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER] = (
            ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )
        )

    return _rewrite_archive(path, mutate, ".shared-workbook-revision-noise.tmp.xlsx")


def make_strict_shared_workbook_revision_model(path: Path) -> Path:
    """Create a Strict SpreadsheetML legacy shared-workbook revision fixture."""
    make_shared_workbook_revision_model(path)

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        for member in (
            _SHARED_WORKBOOK_REVISION_HEADERS_MEMBER,
            _SHARED_WORKBOOK_REVISION_LOG_MEMBER,
        ):
            root = ElementTree.fromstring(contents[member])
            for element in root.iter():
                element.tag = strict_name(
                    element.tag,
                    _SPREADSHEETML_NS,
                    _STRICT_SPREADSHEETML_NS,
                )
                replacements = {
                    attribute: strict_name(
                        attribute,
                        _DOCUMENT_RELATIONSHIPS_NS,
                        _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                    )
                    for attribute in element.attrib
                    if attribute.startswith(f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}")
                }
                for old_name, new_name in replacements.items():
                    element.attrib[new_name] = element.attrib.pop(old_name)
            contents[member] = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        for relationship in workbook_relationships.findall(relationship_tag):
            if relationship.get("Type") == _SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIP:
                relationship.set(
                    "Type",
                    f"{_STRICT_DOCUMENT_RELATIONSHIPS_NS}/revisionHeaders",
                )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            workbook_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        header_relationships = ElementTree.fromstring(
            contents[_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER]
        )
        for relationship in header_relationships.findall(relationship_tag):
            if relationship.get("Type") == _SHARED_WORKBOOK_REVISION_LOG_RELATIONSHIP:
                relationship.set(
                    "Type",
                    f"{_STRICT_DOCUMENT_RELATIONSHIPS_NS}/revisionLog",
                )
        contents[_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER] = (
            ElementTree.tostring(
                header_relationships,
                encoding="utf-8",
                xml_declaration=True,
            )
        )

    return _rewrite_archive(path, mutate, ".strict-shared-workbook-revisions.tmp.xlsx")


def corrupt_shared_workbook_revision(path: Path) -> Path:
    """Inject an unrecognized private revision control to fail closed."""

    def mutate(contents: dict[str, bytes]) -> None:
        headers = ElementTree.fromstring(
            contents[_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER]
        )
        headers.set("privateRevisionControl", "PRIVATE-REVISION-CONTROL")
        contents[_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER] = ElementTree.tostring(
            headers,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".shared-workbook-revision-corrupt.tmp.xlsx")


def _xml_mapping_table_member(contents: dict[str, bytes]) -> str:
    """Return the one ordinary table part used by the XML-map fixture."""
    members = sorted(
        member
        for member in contents
        if member.startswith("xl/tables/") and member.endswith(".xml")
    )
    if len(members) != 1:
        raise ValueError("Fixture needs exactly one table package part")
    return members[0]


def _xml_mapping_map_root(contents: dict[str, bytes]) -> ElementTree.Element:
    """Return the custom XML Maps root in a raw XML-mapping fixture."""
    try:
        payload = contents["xl/xmlMaps.xml"]
    except KeyError as error:
        raise ValueError("Fixture does not contain an XML Maps package part") from error
    return ElementTree.fromstring(payload)


def _xml_mapping_table_binding(
    table: ElementTree.Element,
) -> ElementTree.Element:
    """Return the fixture's mapped table-column declaration."""
    binding = table.find(f".//{{{_SPREADSHEETML_NS}}}xmlColumnPr")
    if binding is None:
        raise ValueError("Fixture does not contain an XML-mapped table column")
    return binding


def _xml_mapping_single_cell_root(contents: dict[str, bytes]) -> ElementTree.Element:
    """Return the fixture's raw single-cell XML mapping table."""
    try:
        payload = contents["xl/singleCellTables/singleCellTable1.xml"]
    except KeyError as error:
        raise ValueError("Fixture does not contain a single-cell XML mapping part") from error
    return ElementTree.fromstring(payload)


def make_xml_mapping_model(path: Path) -> Path:
    """Create a valid XML-map package with table and single-cell bindings.

    The declarations use intentionally private schema, map, XPath, and binding
    names so tests can verify that raw OOXML comparison never emits them.
    """
    make_table_model(path)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        spreadsheet = _SPREADSHEETML_NS

        table_member = _xml_mapping_table_member(contents)
        table = ElementTree.fromstring(contents[table_member])
        table.set("tableType", "xml")
        table.set("connectionId", "7")
        columns = table.find(f"{{{spreadsheet}}}tableColumns")
        if columns is None:
            raise ValueError("Fixture table does not contain table columns")
        first_column = columns.find(f"{{{spreadsheet}}}tableColumn")
        if first_column is None:
            raise ValueError("Fixture table does not contain a table column")
        ElementTree.SubElement(
            first_column,
            f"{{{spreadsheet}}}xmlColumnPr",
            {
                "mapId": "1",
                "xpath": "/private:PrivateRoot/private:Record/private:Amount",
                "denormalized": "false",
                "xmlDataType": "string",
            },
        )
        contents[table_member] = serialize(table)

        map_info = ElementTree.Element(
            f"{{{spreadsheet}}}MapInfo",
            {"SelectionNamespaces": "xmlns:private='urn:formulafence:test:private'"},
        )
        schema = ElementTree.SubElement(
            map_info,
            f"{{{spreadsheet}}}Schema",
            {"ID": "PRIVATE-XML-SCHEMA"},
        )
        schema_definition = ElementTree.SubElement(
            schema,
            f"{{{_XML_SCHEMA_NS}}}schema",
            {"targetNamespace": "urn:formulafence:test:private"},
        )
        ElementTree.SubElement(
            schema_definition,
            f"{{{_XML_SCHEMA_NS}}}element",
            {"name": "PrivateRoot"},
        )
        xml_map = ElementTree.SubElement(
            map_info,
            f"{{{spreadsheet}}}Map",
            {
                "ID": "1",
                "Name": "PRIVATE-XML-MAP",
                "RootElement": "PrivateRoot",
                "SchemaID": "PRIVATE-XML-SCHEMA",
                "ShowImportExportValidationErrors": "false",
                "AutoFit": "true",
                "Append": "false",
                "PreserveSortAFLayout": "true",
                "PreserveFormat": "true",
            },
        )
        ElementTree.SubElement(
            xml_map,
            f"{{{spreadsheet}}}DataBinding",
            {
                "DataBindingName": "PRIVATE-XML-DATA-BINDING",
                "FileBinding": "true",
                "ConnectionID": "7",
                "FileBindingName": "PRIVATE-XML-BINDING-FILE",
                "DataBindingLoadMode": "1",
            },
        )
        contents["xl/xmlMaps.xml"] = serialize(map_info)

        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceXmlMaps",
                "Type": _XML_MAP_RELATIONSHIP,
                "Target": "xmlMaps.xml",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        single_cells = ElementTree.Element(f"{{{spreadsheet}}}singleXmlCells")
        single_cell = ElementTree.SubElement(
            single_cells,
            f"{{{spreadsheet}}}singleXmlCell",
            {
                "id": "1",
                "r": "B2",
                "connectionId": "7",
            },
        )
        cell_properties = ElementTree.SubElement(
            single_cell,
            f"{{{spreadsheet}}}xmlCellPr",
            {
                "id": "1",
                "uniqueName": "PRIVATE-XML-SINGLE-CELL",
            },
        )
        ElementTree.SubElement(
            cell_properties,
            f"{{{spreadsheet}}}xmlPr",
            {
                "mapId": "1",
                "xpath": "/private:PrivateRoot/private:Header/private:AsOf",
                "xmlDataType": "date",
            },
        )
        single_cell_member = "xl/singleCellTables/singleCellTable1.xml"
        contents[single_cell_member] = serialize(single_cells)

        worksheet_relationships_member = _relationship_member(
            "xl/worksheets/sheet1.xml"
        )
        worksheet_relationships = ElementTree.fromstring(
            contents[worksheet_relationships_member]
        )
        ElementTree.SubElement(
            worksheet_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceXmlMappingSingleCells",
                "Type": _TABLE_SINGLE_CELLS_RELATIONSHIP,
                "Target": "../singleCellTables/singleCellTable1.xml",
            },
        )
        contents[worksheet_relationships_member] = serialize(worksheet_relationships)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        if not any(
            override.get("PartName") == f"/{single_cell_member}"
            for override in content_types.findall(override_tag)
        ):
            ElementTree.SubElement(
                content_types,
                override_tag,
                {
                    "PartName": f"/{single_cell_member}",
                    "ContentType": "application/vnd.ms-excel.tableSingleCells",
                },
            )
        contents["[Content_Types].xml"] = serialize(content_types)

    return _rewrite_archive(path, mutate, ".xml-mapping-model.tmp.xlsx")


def change_xml_mapping_xpath(path: Path) -> Path:
    """Retarget a mapped table field without changing table cells or formulas."""

    def mutate(contents: dict[str, bytes]) -> None:
        table_member = _xml_mapping_table_member(contents)
        table = ElementTree.fromstring(contents[table_member])
        _xml_mapping_table_binding(table).set(
            "xpath",
            "/private:PrivateRoot/private:Record/private:CandidateAmount",
        )
        contents[table_member] = ElementTree.tostring(
            table,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xml-mapping-xpath.tmp.xlsx")


def change_xml_mapping_refresh_behavior(path: Path) -> Path:
    """Change map refresh behavior without changing mapping targets or cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        map_info = _xml_mapping_map_root(contents)
        xml_map = map_info.find(f"{{{_SPREADSHEETML_NS}}}Map")
        if xml_map is None:
            raise ValueError("Fixture does not contain an XML map declaration")
        xml_map.set("Append", "true")
        data_binding = xml_map.find(f"{{{_SPREADSHEETML_NS}}}DataBinding")
        if data_binding is None:
            raise ValueError("Fixture XML map does not contain a data binding")
        data_binding.set("DataBindingLoadMode", "2")
        contents["xl/xmlMaps.xml"] = ElementTree.tostring(
            map_info,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xml-mapping-refresh.tmp.xlsx")


def rebind_xml_mapping_relationship(path: Path) -> Path:
    """Move an XML Maps part and retarget its workbook relationship."""

    def mutate(contents: dict[str, bytes]) -> None:
        source_member = "xl/xmlMaps.xml"
        replacement_member = "xl/xmlMaps2.xml"
        try:
            contents[replacement_member] = contents.pop(source_member)
        except KeyError as error:
            raise ValueError("Fixture does not contain an XML Maps package part") from error
        relationships_member = "xl/_rels/workbook.xml.rels"
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            (
                item
                for item in relationships.findall(
                    f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
                )
                if item.get("Type") == _XML_MAP_RELATIONSHIP
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not declare an XML Maps relationship")
        relationship.set("Target", "xmlMaps2.xml")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xml-mapping-rebind.tmp.xlsx")


def externalize_xml_mapping_relationship(path: Path) -> Path:
    """Make an XML Maps relationship unsafe without following its private target."""

    def mutate(contents: dict[str, bytes]) -> None:
        relationships_member = "xl/_rels/workbook.xml.rels"
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            (
                item
                for item in relationships.findall(
                    f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
                )
                if item.get("Type") == _XML_MAP_RELATIONSHIP
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not declare an XML Maps relationship")
        relationship.set(
            "Target",
            "https://private.example.test/PRIVATE-XML-MAP-RELATIONSHIP",
        )
        relationship.set("TargetMode", "External")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xml-mapping-external.tmp.xlsx")


def normalize_xml_mapping_control_spelling(path: Path) -> Path:
    """Rewrite equivalent XML-map Booleans and unsigned integers."""

    def mutate(contents: dict[str, bytes]) -> None:
        map_info = _xml_mapping_map_root(contents)
        xml_map = map_info.find(f"{{{_SPREADSHEETML_NS}}}Map")
        if xml_map is None:
            raise ValueError("Fixture does not contain an XML map declaration")
        xml_map.attrib.update(
            {
                "ID": "0001",
                "ShowImportExportValidationErrors": "0",
                "AutoFit": "1",
                "Append": "0",
                "PreserveSortAFLayout": "1",
                "PreserveFormat": "1",
            }
        )
        data_binding = xml_map.find(f"{{{_SPREADSHEETML_NS}}}DataBinding")
        if data_binding is None:
            raise ValueError("Fixture XML map does not contain a data binding")
        data_binding.attrib.update(
            {
                "FileBinding": "1",
                "ConnectionID": "0007",
                "DataBindingLoadMode": "01",
            }
        )
        contents["xl/xmlMaps.xml"] = ElementTree.tostring(
            map_info,
            encoding="utf-8",
            xml_declaration=True,
        )

        table_member = _xml_mapping_table_member(contents)
        table = ElementTree.fromstring(contents[table_member])
        table.set("connectionId", "0007")
        _xml_mapping_table_binding(table).attrib.update(
            {
                "mapId": "0001",
                "denormalized": "0",
            }
        )
        contents[table_member] = ElementTree.tostring(
            table,
            encoding="utf-8",
            xml_declaration=True,
        )

        single_cells = _xml_mapping_single_cell_root(contents)
        single_cell = single_cells.find(f"{{{_SPREADSHEETML_NS}}}singleXmlCell")
        if single_cell is None:
            raise ValueError("Fixture does not contain an XML-mapped single cell")
        single_cell.set("connectionId", "0007")
        xml_properties = single_cell.find(
            f"{{{_SPREADSHEETML_NS}}}xmlCellPr/{{{_SPREADSHEETML_NS}}}xmlPr"
        )
        if xml_properties is None:
            raise ValueError("Fixture single cell does not contain XML properties")
        xml_properties.set("mapId", "0001")
        contents["xl/singleCellTables/singleCellTable1.xml"] = ElementTree.tostring(
            single_cells,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xml-mapping-normalize.tmp.xlsx")


def corrupt_xml_mapping_single_cell_reference(path: Path) -> Path:
    """Inject an invalid raw XML-map cell target to exercise fail-closed handling."""

    def mutate(contents: dict[str, bytes]) -> None:
        single_cells = _xml_mapping_single_cell_root(contents)
        single_cell = single_cells.find(f"{{{_SPREADSHEETML_NS}}}singleXmlCell")
        if single_cell is None:
            raise ValueError("Fixture does not contain an XML-mapped single cell")
        single_cell.set("r", "PRIVATE-NOT-AN-XML-MAP-CELL")
        contents["xl/singleCellTables/singleCellTable1.xml"] = ElementTree.tostring(
            single_cells,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xml-mapping-corrupt.tmp.xlsx")


def make_data_validation_model(path: Path, *, reverse_status_targets: bool = False) -> Path:
    """Create an input workbook whose data-entry controls carry business rules."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Controlled entries"
    inputs["B1"] = "Status"
    inputs["B2"] = "Draft"
    inputs["C1"] = "Approved amount"
    inputs["C2"] = 25
    inputs["D1"] = "Secondary status"
    inputs["D2"] = "Review"

    limits = workbook.create_sheet("Limits")
    limits["A1"] = "Approved statuses"
    limits["A2"] = "Draft"
    limits["A3"] = "Review"
    limits["A4"] = "Approved"
    limits["B1"] = "Minimum"
    limits["B2"] = 10
    limits["B3"] = 100

    status = DataValidation(
        type="list",
        formula1="=Limits!$A$2:$A$4",
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Invalid status",
        error="Choose an approved status.",
        promptTitle="Approved status",
        prompt="Choose a documented status.",
    )
    if reverse_status_targets:
        status.add("D2")
        status.add("B2:B100")
    else:
        status.add("B2:B100")
        status.add("D2")
    inputs.add_data_validation(status)

    amount = DataValidation(
        type="decimal",
        operator="between",
        formula1="=Limits!$B$2",
        formula2="=Limits!$B$3",
        allow_blank=False,
        showErrorMessage=True,
        errorStyle="warning",
        errorTitle="Outside approved range",
        error="Use an amount within the approved range.",
    )
    amount.add("C2:C100")
    inputs.add_data_validation(amount)
    workbook.save(path)
    return path


def make_conditional_formatting_model(path: Path) -> Path:
    """Create visual review controls with overlapping precedence and builtins."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Controlled metric"
    for row, value in enumerate((10, -5, 75, 120), start=2):
        inputs.cell(row, 1, value)

    red_fill = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
    red_font = Font(color="FF9C0006")
    green_fill = PatternFill(fill_type="solid", fgColor="FFC6EFCE")
    green_font = Font(color="FF006100")
    inputs.conditional_formatting.add(
        "A2:A100",
        FormulaRule(
            formula=["$A2<0"],
            stopIfTrue=True,
            fill=red_fill,
            font=red_font,
        ),
    )
    inputs.conditional_formatting.add(
        "A2:A100",
        CellIsRule(
            operator="greaterThan",
            formula=["100"],
            fill=green_fill,
            font=green_font,
        ),
    )
    inputs.conditional_formatting.add(
        "B2:B100",
        ColorScaleRule(
            start_type="min",
            start_color="FFF8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFFEB84",
            end_type="max",
            end_color="FF63BE7B",
        ),
    )
    inputs.conditional_formatting.add(
        "C2:C100",
        DataBarRule(
            start_type="min",
            start_value=None,
            end_type="max",
            end_value=None,
            color="FF638EC6",
        ),
    )
    inputs.conditional_formatting.add(
        "D2:D100",
        IconSetRule(
            "3TrafficLights1",
            "percent",
            [0, 33, 67],
            showValue=None,
            percent=None,
            reverse=None,
        ),
    )
    workbook.save(path)
    return path


def _rewrite_archive(path: Path, mutate: Callable[[dict[str, bytes]], None], suffix: str) -> Path:
    """Apply a small raw OOXML test mutation without changing ZIP member order."""
    with ZipFile(path) as archive:
        contents = {
            entry.filename: archive.read(entry.filename) for entry in archive.infolist()
        }
    mutate(contents)
    staging = path.with_suffix(suffix)
    with ZipFile(staging, "w", compression=ZIP_DEFLATED) as archive:
        for name, content in contents.items():
            archive.writestr(name, content)
    staging.replace(path)
    return path


def _inputs_worksheet_root(contents: dict[str, bytes]) -> ElementTree.Element:
    """Return the fixture's first worksheet XML root."""
    return ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])


def _save_inputs_worksheet(contents: dict[str, bytes], root: ElementTree.Element) -> None:
    contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True,
    )


def _zip_payload(parts: dict[str, bytes]) -> bytes:
    """Build one in-memory OPC-like package for a Data Mashup test fixture."""
    payload = io.BytesIO()
    with ZipFile(payload, "w", compression=ZIP_DEFLATED) as archive:
        for name, value in parts.items():
            archive.writestr(name, value)
    return payload.getvalue()


def _synthetic_power_query_mashup(
    *,
    formula: str,
    fill_enabled: bool = True,
    firewall_enabled: bool = True,
    volatile_timestamp: str = "2026-07-24T12:00:00.0000000",
    sqmid: str = "11111111-2222-3333-4444-555555555555",
    permission_binding: bytes = b"\x00",
) -> bytes:
    """Return a valid, deliberately confidential Data Mashup XML payload."""
    package_parts = _zip_payload(
        {
            "[Content_Types].xml": b"<Types/>",
            "Config/Package.xml": (
                b"<Package>private-baseline-package-config</Package>"
            ),
            "Formulas/Section1.m": formula.encode("utf-8"),
            "Content/11111111-2222-3333-4444-555555555555": (
                b"private-baseline-embedded-content"
            ),
        }
    )
    metadata_xml = f"""<?xml version="1.0" encoding="utf-8"?>
<LocalPackageMetadataFile xmlns="{_DATA_MASHUP_NS}">
  <Items>
    <Item>
      <ItemLocation>
        <ItemType>Formula</ItemType>
        <ItemPath>Section1/Private revenue query</ItemPath>
      </ItemLocation>
      <StableEntries>
        <Entry Type="FillEnabled" Value="l{int(fill_enabled)}" />
        <Entry Type="FillTarget" Value="sprivate-baseline-target" />
        <Entry Type="FillLastUpdated" Value="d{volatile_timestamp}" />
        <Entry Type="FillErrorMessage" Value="sprivate-refresh-message" />
      </StableEntries>
    </Item>
  </Items>
</LocalPackageMetadataFile>
""".encode()
    metadata_content = _zip_payload({})
    metadata = (
        struct.pack("<II", 0, len(metadata_xml))
        + metadata_xml
        + struct.pack("<I", len(metadata_content))
        + metadata_content
    )
    permissions = f"""<?xml version="1.0" encoding="utf-8"?>
<PermissionList xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <CanEvaluateFuturePackages>false</CanEvaluateFuturePackages>
  <FirewallEnabled>{str(firewall_enabled).lower()}</FirewallEnabled>
  <WorkbookGroupType xsi:nil="true" />
</PermissionList>
""".encode()
    stream = struct.pack("<I", 0)
    for field in (package_parts, permissions, metadata, permission_binding):
        stream += struct.pack("<I", len(field)) + field
    root = ElementTree.Element(
        f"{{{_DATA_MASHUP_NS}}}DataMashup",
        {"sqmid": sqmid},
    )
    root.text = base64.b64encode(stream).decode("ascii")
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _power_query_formula(*, source: str) -> str:
    """Build one M document whose source material must never be emitted."""
    return f'''section Section1;

shared PrivateRevenueQuery = let
    Source = Web.Contents("{source}"),
    Result = Source
in
    Result;
'''


def make_power_query_model(path: Path) -> Path:
    """Create a raw-OOXML workbook containing a Power Query Data Mashup."""
    make_model(path)
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    content_types_namespace = (
        "http://schemas.openxmlformats.org/package/2006/content-types"
    )

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(root: ElementTree.Element, part_name: str, content_type: str) -> None:
        override_tag = f"{{{content_types_namespace}}}Override"
        if any(item.get("PartName") == part_name for item in root.findall(override_tag)):
            return
        ElementTree.SubElement(
            root,
            override_tag,
            {"PartName": part_name, "ContentType": content_type},
        )

    def mutate(contents: dict[str, bytes]) -> None:
        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        add_override(
            content_types,
            "/xl/tables/table1.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml",
        )
        add_override(
            content_types,
            "/xl/queryTables/queryTable1.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.queryTable+xml",
        )
        contents["[Content_Types].xml"] = serialize(content_types)

        root_relationships = ElementTree.fromstring(contents["_rels/.rels"])
        relationship_tag = f"{{{package_relationships}}}Relationship"
        ElementTree.SubElement(
            root_relationships,
            relationship_tag,
            {
                "Id": "rIdFencePowerQueryCustomXml",
                "Type": f"{document_relationships}/customXml",
                "Target": "customXml/item1.xml",
            },
        )
        contents["_rels/.rels"] = serialize(root_relationships)

        worksheet = _inputs_worksheet_root(contents)
        table_parts = ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}tableParts",
            {"count": "1"},
        )
        ElementTree.SubElement(
            table_parts,
            f"{{{spreadsheet}}}tablePart",
            {f"{{{document_relationships}}}id": "rIdFencePowerQueryTable"},
        )
        _save_inputs_worksheet(contents, worksheet)

        worksheet_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            worksheet_relationships,
            relationship_tag,
            {
                "Id": "rIdFencePowerQueryTable",
                "Type": f"{document_relationships}/table",
                "Target": "../tables/table1.xml",
            },
        )
        contents["xl/worksheets/_rels/sheet1.xml.rels"] = serialize(
            worksheet_relationships
        )

        table = ElementTree.Element(
            f"{{{spreadsheet}}}table",
            {
                "id": "1",
                "name": "ResultTable",
                "displayName": "ResultTable",
                "ref": "A1:B2",
            },
        )
        ElementTree.SubElement(table, f"{{{spreadsheet}}}autoFilter", {"ref": "A1:B2"})
        columns = ElementTree.SubElement(
            table,
            f"{{{spreadsheet}}}tableColumns",
            {"count": "2"},
        )
        ElementTree.SubElement(
            columns,
            f"{{{spreadsheet}}}tableColumn",
            {"id": "1", "name": "Metric"},
        )
        ElementTree.SubElement(
            columns,
            f"{{{spreadsheet}}}tableColumn",
            {"id": "2", "name": "Value"},
        )
        ElementTree.SubElement(
            table,
            f"{{{spreadsheet}}}tableStyleInfo",
            {
                "name": "TableStyleMedium2",
                "showFirstColumn": "0",
                "showLastColumn": "0",
                "showRowStripes": "1",
                "showColumnStripes": "0",
            },
        )
        contents["xl/tables/table1.xml"] = serialize(table)
        table_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            table_relationships,
            relationship_tag,
            {
                "Id": "rIdFencePowerQueryQueryTable",
                "Type": f"{document_relationships}/queryTable",
                "Target": "../queryTables/queryTable1.xml",
            },
        )
        contents["xl/tables/_rels/table1.xml.rels"] = serialize(table_relationships)

        query_table = ElementTree.Element(
            f"{{{spreadsheet}}}queryTable",
            {
                "name": "ResultTable_ExternalData_1",
                "connectionId": "1",
                "refreshOnLoad": "1",
                "backgroundRefresh": "1",
                "disableRefresh": "0",
                "removeDataOnSave": "0",
                "fillFormulas": "0",
                "disableEdit": "1",
                "growShrinkType": "insertClear",
            },
        )
        contents["xl/queryTables/queryTable1.xml"] = serialize(query_table)
        contents["customXml/item1.xml"] = _synthetic_power_query_mashup(
            formula=_power_query_formula(
                source="https://private.example/private-baseline-power-query-token"
            )
        )

    return _rewrite_archive(path, mutate, ".power-query.tmp.xlsx")


def change_power_query_controls(path: Path) -> Path:
    """Change private M material plus safe Power Query execution controls."""

    def mutate(contents: dict[str, bytes]) -> None:
        contents["customXml/item1.xml"] = _synthetic_power_query_mashup(
            formula=_power_query_formula(
                source="https://private.example/private-candidate-power-query-token"
            ),
            fill_enabled=False,
            firewall_enabled=False,
            sqmid="aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee",
        )

    return _rewrite_archive(path, mutate, ".power-query-change.tmp.xlsx")


def change_power_query_refresh_noise(path: Path) -> Path:
    """Change volatile refresh state and user-bound binding without semantic drift."""

    def mutate(contents: dict[str, bytes]) -> None:
        contents["customXml/item1.xml"] = _synthetic_power_query_mashup(
            formula=_power_query_formula(
                source="https://private.example/private-baseline-power-query-token"
            ),
            volatile_timestamp="2026-07-24T13:45:00.0000000",
            sqmid="99999999-8888-7777-6666-555555555555",
            permission_binding=b"synthetic-user-bound-permission-binding",
        )

    return _rewrite_archive(path, mutate, ".power-query-noise.tmp.xlsx")


def make_power_pivot_data_model(path: Path) -> Path:
    """Create a harmless relationship-backed Power Pivot/Data Model fixture.

    ``item.data`` is intentionally opaque: FormulaFence must fingerprint its
    bounded bytes without trying to deserialize an Analysis Services backup or
    exposing its contents. The controlled package is never opened in Office.
    """
    make_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    data_model = _OFFICE_2013_SPREADSHEET_NS

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        extensions = ElementTree.SubElement(workbook, f"{{{spreadsheet}}}extLst")
        extension = ElementTree.SubElement(
            extensions,
            f"{{{spreadsheet}}}ext",
            {"uri": "{FCE2AD5D-F65C-4FA6-A056-5C36A1767C68}"},
        )
        model = ElementTree.SubElement(
            extension,
            f"{{{data_model}}}dataModel",
            {"minVersionLoad": "1"},
        )
        model_tables = ElementTree.SubElement(model, f"{{{data_model}}}modelTables")
        ElementTree.SubElement(
            model_tables,
            f"{{{data_model}}}modelTable",
            {
                "id": "PrivateRevenue_{11111111-1111-1111-1111-111111111111}",
                "name": "Private baseline revenue table",
                "connection": "Private baseline data connection",
            },
        )
        ElementTree.SubElement(
            model_tables,
            f"{{{data_model}}}modelTable",
            {
                "id": "PrivateCalendar_{22222222-2222-2222-2222-222222222222}",
                "name": "Private baseline calendar table",
                "connection": "Private baseline calendar connection",
            },
        )
        model_relationships = ElementTree.SubElement(
            model,
            f"{{{data_model}}}modelRelationships",
        )
        ElementTree.SubElement(
            model_relationships,
            f"{{{data_model}}}modelRelationship",
            {
                "fromTable": "Private baseline revenue table",
                "fromColumn": "Private baseline calendar key",
                "toTable": "Private baseline calendar table",
                "toColumn": "Private baseline date key",
            },
        )
        contents["xl/workbook.xml"] = serialize(workbook)

        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            workbook_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFencePowerPivotData",
                "Type": f"{document_relationships}/powerPivotData",
                "Target": "model/item.data",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        default_tag = f"{{{content_types}}}Default"
        if not any(item.get("Extension") == "data" for item in types.findall(default_tag)):
            ElementTree.SubElement(
                types,
                default_tag,
                {
                    "Extension": "data",
                    "ContentType": "application/vnd.openxmlformats-officedocument.model+data",
                },
            )
        contents["[Content_Types].xml"] = serialize(types)
        contents["xl/model/item.data"] = b"private baseline Power Pivot data model payload"

    return _rewrite_archive(path, mutate, ".power-pivot-data-model.tmp.xlsx")


def change_power_pivot_data_model_payload(path: Path) -> Path:
    """Change only the private raw embedded Data Model payload."""

    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/model/item.data"] = b"private candidate Power Pivot data model payload"

    return _rewrite_archive(path, mutate, ".power-pivot-data-payload.tmp.xlsx")


def change_power_pivot_data_model_declaration(path: Path) -> Path:
    """Change a private model relationship without touching raw payload bytes."""
    data_model = _OFFICE_2013_SPREADSHEET_NS

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        relationship = next(
            workbook.iter(f"{{{data_model}}}modelRelationship")
        )
        relationship.set("toColumn", "Private candidate calendar key")
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".power-pivot-data-declaration.tmp.xlsx")


def renumber_power_pivot_data_model_relationship(path: Path) -> Path:
    """Rewrite the workbook binding ID while retaining its semantic target."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        relationship = next(
            item
            for item in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if item.get("Id") == "rIdFencePowerPivotData"
        )
        relationship.set("Id", "rIdFenceRenumberedPowerPivotData")
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".power-pivot-data-renumber.tmp.xlsx")


def rewrite_power_pivot_data_model_internal_target_spelling(path: Path) -> Path:
    """Use an equivalent relative spelling for the embedded model target."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        relationship = next(
            item
            for item in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if item.get("Id") == "rIdFencePowerPivotData"
        )
        relationship.set("Target", "./model/../model/item.data")
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".power-pivot-data-target.tmp.xlsx")


def set_power_pivot_data_model_equivalent_guids(path: Path) -> Path:
    """Regenerate writer IDs that do not alter model table relationships."""
    data_model = _OFFICE_2013_SPREADSHEET_NS

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        tables = tuple(workbook.iter(f"{{{data_model}}}modelTable"))
        tables[0].set("id", "PrivateRevenue_{AAAAAAAA-AAAA-AAAA-AAAA-AAAAAAAAAAAA}")
        tables[1].set("id", "PrivateCalendar_{BBBBBBBB-BBBB-BBBB-BBBB-BBBBBBBBBBBB}")
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".power-pivot-data-guid.tmp.xlsx")


def rebind_power_pivot_data_model(path: Path) -> Path:
    """Move the embedded Data Model to a different package member."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        relationship = next(
            item
            for item in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if item.get("Id") == "rIdFencePowerPivotData"
        )
        relationship.set("Target", "model/item2.data")
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/model/item2.data"] = contents.pop("xl/model/item.data")

    return _rewrite_archive(path, mutate, ".power-pivot-data-rebind.tmp.xlsx")


def externalize_power_pivot_data_model(path: Path) -> Path:
    """Turn the workbook Data Model binding into an external redaction fixture."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        relationship = next(
            item
            for item in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if item.get("Id") == "rIdFencePowerPivotData"
        )
        relationship.set("Target", "https://example.invalid/private-power-pivot-data-model")
        relationship.set("TargetMode", "External")
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".power-pivot-data-external.tmp.xlsx")


def remove_power_pivot_data_model_workbook_binding(path: Path) -> Path:
    """Leave Data Model metadata and payload orphaned from the workbook."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        relationship = next(
            item
            for item in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if item.get("Id") == "rIdFencePowerPivotData"
        )
        relationships.remove(relationship)
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".power-pivot-data-unbound.tmp.xlsx")


def add_power_pivot_data_model_direct_relationship(path: Path) -> Path:
    """Add an unexpected direct model-part relationship for coverage tests."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFencePowerPivotRelated",
                "Type": f"{document_relationships}/image",
                "Target": "private-model-preview.png",
            },
        )
        contents["xl/model/_rels/item.data.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/model/private-model-preview.png"] = b"private model preview payload"

    return _rewrite_archive(path, mutate, ".power-pivot-data-related.tmp.xlsx")


def make_external_data_refresh_model(path: Path) -> Path:
    """Create a raw-OOXML fixture covering every external-data refresh layer."""
    make_model(path)
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    document_relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    content_types_namespace = (
        "http://schemas.openxmlformats.org/package/2006/content-types"
    )

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(
        root: ElementTree.Element,
        part_name: str,
        content_type: str,
    ) -> None:
        override_tag = f"{{{content_types_namespace}}}Override"
        if any(item.get("PartName") == part_name for item in root.findall(override_tag)):
            return
        ElementTree.SubElement(
            root,
            override_tag,
            {"PartName": part_name, "ContentType": content_type},
        )

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        workbook_properties = workbook.find(f"{{{spreadsheet}}}workbookPr")
        if workbook_properties is None:
            workbook_properties = ElementTree.Element(f"{{{spreadsheet}}}workbookPr")
            workbook.insert(0, workbook_properties)
        workbook_properties.attrib.update(
            {
                "updateLinks": "always",
                "allowRefreshQuery": "1",
                "refreshAllConnections": "1",
                "saveExternalLinkValues": "0",
            }
        )
        pivot_caches = workbook.find(f"{{{spreadsheet}}}pivotCaches")
        if pivot_caches is None:
            pivot_caches = ElementTree.Element(f"{{{spreadsheet}}}pivotCaches")
            calc_properties = workbook.find(f"{{{spreadsheet}}}calcPr")
            insertion_index = (
                list(workbook).index(calc_properties)
                if calc_properties is not None
                else len(workbook)
            )
            workbook.insert(insertion_index, pivot_caches)
        ElementTree.SubElement(
            pivot_caches,
            f"{{{spreadsheet}}}pivotCache",
            {"cacheId": "7", f"{{{document_relationships}}}id": "rIdFencePivot"},
        )
        contents["xl/workbook.xml"] = serialize(workbook)

        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        relationship_tag = f"{{{package_relationships}}}Relationship"
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceConnections",
                "Type": f"{document_relationships}/connections",
                "Target": "connections.xml",
            },
        )
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdFencePivot",
                "Type": f"{document_relationships}/pivotCacheDefinition",
                "Target": "pivotCache/pivotCacheDefinition1.xml",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        add_override(
            content_types,
            "/xl/connections.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml",
        )
        add_override(
            content_types,
            "/xl/queryTables/queryTable1.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.queryTable+xml",
        )
        add_override(
            content_types,
            "/xl/pivotCache/pivotCacheDefinition1.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml",
        )
        contents["[Content_Types].xml"] = serialize(content_types)

        connection_root = ElementTree.Element(f"{{{spreadsheet}}}connections")
        database_connection = ElementTree.SubElement(
            connection_root,
            f"{{{spreadsheet}}}connection",
            {
                "id": "1",
                "name": "synthetic confidential revenue connection",
                "description": "synthetic confidential connection description",
                "type": "5",
                "refreshedVersion": "6",
                "minRefreshableVersion": "3",
                "sourceFile": "C:/private/synthetic-revenue-source.accdb",
                "odcFile": "https://private.example/synthetic-revenue.odc",
                "keepAlive": "1",
                "interval": "60",
                "savePassword": "1",
                "new": "1",
                "onlyUseConnectionFile": "1",
                "reconnectionMethod": "2",
                "background": "1",
                "refreshOnLoad": "1",
                "saveData": "0",
                "credentials": "stored",
                "singleSignOnId": "synthetic-private-sso-identifier",
            },
        )
        ElementTree.SubElement(
            database_connection,
            f"{{{spreadsheet}}}dbPr",
            {
                "connection": "Provider=synthetic;Password=private-baseline-password",
                "command": "select * from confidential_revenue",
                "commandType": "2",
            },
        )
        parameters = ElementTree.SubElement(
            database_connection,
            f"{{{spreadsheet}}}parameters",
            {"count": "1"},
        )
        ElementTree.SubElement(
            parameters,
            f"{{{spreadsheet}}}parameter",
            {
                "name": "synthetic confidential parameter",
                "parameterType": "cell",
                "cell": "Inputs!$B$2",
                "refreshOnChange": "1",
                "string": "synthetic confidential parameter value",
            },
        )
        connection_extensions = ElementTree.SubElement(
            database_connection,
            f"{{{spreadsheet}}}extLst",
        )
        ElementTree.SubElement(
            connection_extensions,
            f"{{{spreadsheet}}}ext",
            {"uri": "urn:synthetic:private-connection-extension"},
        ).text = "synthetic private connection extension payload"
        web_connection = ElementTree.SubElement(
            connection_root,
            f"{{{spreadsheet}}}connection",
            {
                "id": "2",
                "name": "synthetic confidential web connection",
                "type": "4",
                "refreshedVersion": "6",
            },
        )
        ElementTree.SubElement(
            web_connection,
            f"{{{spreadsheet}}}webPr",
            {"url": "https://private.example/synthetic-web-query"},
        )
        contents["xl/connections.xml"] = serialize(connection_root)

        query_table = ElementTree.Element(
            f"{{{spreadsheet}}}queryTable",
            {
                "name": "synthetic confidential query table",
                "connectionId": "1",
                "refreshOnLoad": "1",
                "backgroundRefresh": "0",
                "disableRefresh": "0",
                "removeDataOnSave": "1",
                "fillFormulas": "1",
                "disableEdit": "1",
                "growShrinkType": "overwriteClear",
            },
        )
        query_extensions = ElementTree.SubElement(query_table, f"{{{spreadsheet}}}extLst")
        ElementTree.SubElement(
            query_extensions,
            f"{{{spreadsheet}}}ext",
            {"uri": "urn:synthetic:private-query-extension"},
        ).text = "synthetic private query extension payload"
        contents["xl/queryTables/queryTable1.xml"] = serialize(query_table)

        sheet_relationships = ElementTree.Element(f"{{{package_relationships}}}Relationships")
        ElementTree.SubElement(
            sheet_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceQuery",
                "Type": f"{document_relationships}/queryTable",
                "Target": "../queryTables/queryTable1.xml",
            },
        )
        contents["xl/worksheets/_rels/sheet1.xml.rels"] = serialize(sheet_relationships)

        pivot_cache_definition = ElementTree.Element(
            f"{{{spreadsheet}}}pivotCacheDefinition",
            {
                "refreshOnLoad": "1",
                "backgroundQuery": "1",
                "enableRefresh": "0",
                "saveData": "0",
                "upgradeOnRefresh": "1",
                "refreshedBy": "synthetic confidential refresh identity",
            },
        )
        pivot_source = ElementTree.SubElement(
            pivot_cache_definition,
            f"{{{spreadsheet}}}cacheSource",
            {"type": "external", "connectionId": "2"},
        )
        ElementTree.SubElement(
            pivot_source,
            f"{{{spreadsheet}}}extLst",
        ).text = "synthetic private pivot source payload"
        pivot_extensions = ElementTree.SubElement(
            pivot_cache_definition,
            f"{{{spreadsheet}}}extLst",
        )
        ElementTree.SubElement(
            pivot_extensions,
            f"{{{spreadsheet}}}ext",
            {"uri": "urn:synthetic:private-pivot-extension"},
        ).text = "synthetic private pivot extension payload"
        contents["xl/pivotCache/pivotCacheDefinition1.xml"] = serialize(
            pivot_cache_definition
        )

    return _rewrite_archive(path, mutate, ".external-data.tmp.xlsx")


def change_external_data_refresh_controls(path: Path) -> Path:
    """Change safe controls and private source material in the raw fixture."""
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        workbook_properties = workbook.find(f"{{{spreadsheet}}}workbookPr")
        if workbook_properties is None:
            raise ValueError("Fixture does not contain workbook properties")
        workbook_properties.attrib.update(
            {
                "updateLinks": "never",
                "allowRefreshQuery": "0",
                "refreshAllConnections": "0",
                "saveExternalLinkValues": "1",
            }
        )
        contents["xl/workbook.xml"] = serialize(workbook)

        connections = ElementTree.fromstring(contents["xl/connections.xml"])
        connection = connections.find(f"{{{spreadsheet}}}connection")
        if connection is None:
            raise ValueError("Fixture does not contain a connection")
        connection.attrib.update(
            {
                "name": "changed synthetic confidential revenue connection",
                "sourceFile": "C:/private/changed-synthetic-source.accdb",
                "odcFile": "https://private.example/changed-synthetic.odc",
                "singleSignOnId": "changed-synthetic-private-sso-identifier",
                "interval": "15",
                "refreshOnLoad": "0",
                "savePassword": "0",
            }
        )
        database_properties = connection.find(f"{{{spreadsheet}}}dbPr")
        if database_properties is None:
            raise ValueError("Fixture does not contain database connection properties")
        database_properties.set(
            "connection", "Provider=changed;Password=private-candidate-password"
        )
        contents["xl/connections.xml"] = serialize(connections)

        query_table = ElementTree.fromstring(contents["xl/queryTables/queryTable1.xml"])
        query_table.attrib.update(
            {
                "name": "changed synthetic confidential query table",
                "connectionId": "2",
                "refreshOnLoad": "0",
                "fillFormulas": "0",
            }
        )
        contents["xl/queryTables/queryTable1.xml"] = serialize(query_table)

        pivot_cache = ElementTree.fromstring(
            contents["xl/pivotCache/pivotCacheDefinition1.xml"]
        )
        pivot_cache.attrib.update(
            {
                "refreshOnLoad": "0",
                "enableRefresh": "1",
                "saveData": "1",
            }
        )
        pivot_source = pivot_cache.find(f"{{{spreadsheet}}}cacheSource")
        if pivot_source is None:
            raise ValueError("Fixture does not contain a pivot-cache source")
        pivot_source.set("connectionId", "1")
        contents["xl/pivotCache/pivotCacheDefinition1.xml"] = serialize(pivot_cache)

    return _rewrite_archive(path, mutate, ".external-data-change.tmp.xlsx")


def set_external_data_connection_defaults(path: Path, *, explicit: bool) -> Path:
    """Toggle omitted versus explicit defaults on the fixture's web connection."""
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    defaults = {
        "keepAlive": "0",
        "interval": "0",
        "reconnectionMethod": "1",
        "minRefreshableVersion": "0",
        "savePassword": "0",
        "new": "0",
        "deleted": "0",
        "onlyUseConnectionFile": "0",
        "background": "0",
        "refreshOnLoad": "0",
        "saveData": "0",
        "credentials": "integrated",
    }

    def mutate(contents: dict[str, bytes]) -> None:
        connections = ElementTree.fromstring(contents["xl/connections.xml"])
        connection_tag = f"{{{spreadsheet}}}connection"
        connection = next(
            (
                item
                for item in connections.findall(connection_tag)
                if item.get("id") == "2"
            ),
            None,
        )
        if connection is None:
            raise ValueError("Fixture does not contain the web connection")
        for attribute, value in defaults.items():
            if explicit:
                connection.set(attribute, value)
            else:
                connection.attrib.pop(attribute, None)
        contents["xl/connections.xml"] = ElementTree.tostring(
            connections,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".external-data-defaults.tmp.xlsx")


def make_external_link_package_model(path: Path) -> Path:
    """Create external-workbook, DDE, and OLE packages with private payloads."""
    make_model(path)
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    document_relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    content_types_namespace = (
        "http://schemas.openxmlformats.org/package/2006/content-types"
    )

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(
        root: ElementTree.Element,
        part_name: str,
        content_type: str,
    ) -> None:
        override_tag = f"{{{content_types_namespace}}}Override"
        if any(item.get("PartName") == part_name for item in root.findall(override_tag)):
            return
        ElementTree.SubElement(
            root,
            override_tag,
            {"PartName": part_name, "ContentType": content_type},
        )

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        external_references = ElementTree.Element(f"{{{spreadsheet}}}externalReferences")
        for relationship_id in (
            "rIdFenceExternalWorkbook",
            "rIdFenceDde",
            "rIdFenceOle",
        ):
            ElementTree.SubElement(
                external_references,
                f"{{{spreadsheet}}}externalReference",
                {f"{{{document_relationships}}}id": relationship_id},
            )
        calc_properties = workbook.find(f"{{{spreadsheet}}}calcPr")
        insertion_index = (
            list(workbook).index(calc_properties)
            if calc_properties is not None
            else len(workbook)
        )
        workbook.insert(insertion_index, external_references)
        contents["xl/workbook.xml"] = serialize(workbook)

        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        relationship_tag = f"{{{package_relationships}}}Relationship"
        for relationship_id, target in (
            ("rIdFenceExternalWorkbook", "externalLinks/externalLink1.xml"),
            ("rIdFenceDde", "externalLinks/externalLink2.xml"),
            ("rIdFenceOle", "externalLinks/externalLink3.xml"),
        ):
            ElementTree.SubElement(
                workbook_relationships,
                relationship_tag,
                {
                    "Id": relationship_id,
                    "Type": f"{document_relationships}/externalLink",
                    "Target": target,
                },
            )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        for number in (1, 2, 3):
            add_override(
                content_types,
                f"/xl/externalLinks/externalLink{number}.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml",
            )
        contents["[Content_Types].xml"] = serialize(content_types)

        external_book_root = ElementTree.Element(f"{{{spreadsheet}}}externalLink")
        external_book = ElementTree.SubElement(
            external_book_root,
            f"{{{spreadsheet}}}externalBook",
            {f"{{{document_relationships}}}id": "rIdFenceExternalTarget"},
        )
        sheet_names = ElementTree.SubElement(external_book, f"{{{spreadsheet}}}sheetNames")
        for sheet_name in ("private external baseline sheet", "private external scenario"):
            ElementTree.SubElement(
                sheet_names,
                f"{{{spreadsheet}}}sheetName",
                {"val": sheet_name},
            )
        defined_names = ElementTree.SubElement(
            external_book, f"{{{spreadsheet}}}definedNames"
        )
        ElementTree.SubElement(
            defined_names,
            f"{{{spreadsheet}}}definedName",
            {
                "name": "private external baseline defined name",
                "refersTo": "'private external baseline sheet'!$A$1:$A$2",
                "sheetId": "0",
            },
        )
        sheet_data_set = ElementTree.SubElement(
            external_book, f"{{{spreadsheet}}}sheetDataSet"
        )
        sheet_data = ElementTree.SubElement(
            sheet_data_set,
            f"{{{spreadsheet}}}sheetData",
            {"sheetId": "0", "refreshError": "1"},
        )
        row = ElementTree.SubElement(sheet_data, f"{{{spreadsheet}}}row", {"r": "1"})
        for coordinate, value in (
            ("A1", "private external baseline cached value"),
            ("B1", "private external baseline cached amount"),
        ):
            cell = ElementTree.SubElement(
                row,
                f"{{{spreadsheet}}}cell",
                {"r": coordinate, "t": "str"},
            )
            ElementTree.SubElement(cell, f"{{{spreadsheet}}}v").text = value
        contents["xl/externalLinks/externalLink1.xml"] = serialize(external_book_root)

        external_book_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            external_book_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceExternalTarget",
                "Type": f"{document_relationships}/externalLinkPath",
                "Target": "file:///private/baseline/external-workbook.xlsx",
                "TargetMode": "External",
            },
        )
        contents["xl/externalLinks/_rels/externalLink1.xml.rels"] = serialize(
            external_book_relationships
        )

        dde_root = ElementTree.Element(
            f"{{{spreadsheet}}}externalLink",
        )
        dde_link = ElementTree.SubElement(
            dde_root,
            f"{{{spreadsheet}}}ddeLink",
            {
                "ddeService": "private-baseline-dde-service",
                "ddeTopic": "private-baseline-dde-topic",
            },
        )
        dde_items = ElementTree.SubElement(dde_link, f"{{{spreadsheet}}}ddeItems")
        first_dde_item = ElementTree.SubElement(
            dde_items,
            f"{{{spreadsheet}}}ddeItem",
            {
                "name": "private-baseline-dde-item",
                "ole": "1",
                "advise": "1",
                "preferPic": "0",
            },
        )
        dde_values = ElementTree.SubElement(
            first_dde_item,
            f"{{{spreadsheet}}}values",
            {"rows": "1", "cols": "2"},
        )
        for value in ("private-baseline-dde-value", "private-baseline-dde-second-value"):
            dde_value = ElementTree.SubElement(dde_values, f"{{{spreadsheet}}}value")
            ElementTree.SubElement(dde_value, f"{{{spreadsheet}}}val").text = value
        ElementTree.SubElement(
            dde_items,
            f"{{{spreadsheet}}}ddeItem",
            {
                "name": "private-baseline-dde-picture-item",
                "preferPic": "1",
            },
        )
        contents["xl/externalLinks/externalLink2.xml"] = serialize(dde_root)

        ole_root = ElementTree.Element(f"{{{spreadsheet}}}externalLink")
        ole_link = ElementTree.SubElement(
            ole_root,
            f"{{{spreadsheet}}}oleLink",
            {
                f"{{{document_relationships}}}id": "rIdFenceOleTarget",
                "progId": "private.baseline.ole.program",
            },
        )
        ole_items = ElementTree.SubElement(ole_link, f"{{{spreadsheet}}}oleItems")
        ElementTree.SubElement(
            ole_items,
            f"{{{spreadsheet}}}oleItem",
            {
                "name": "private-baseline-ole-item",
                "icon": "1",
                "advise": "1",
                "preferPic": "0",
            },
        )
        ElementTree.SubElement(
            ole_items,
            f"{{{spreadsheet}}}oleItem",
            {
                "name": "private-baseline-ole-picture-item",
                "preferPic": "1",
            },
        )
        extensions = ElementTree.SubElement(ole_root, f"{{{spreadsheet}}}extLst")
        ElementTree.SubElement(
            extensions,
            f"{{{spreadsheet}}}ext",
            {"uri": "urn:private:baseline:external-link-extension"},
        ).text = "private baseline external-link extension payload"
        contents["xl/externalLinks/externalLink3.xml"] = serialize(ole_root)

        ole_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            ole_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceOleTarget",
                "Type": f"{document_relationships}/oleObject",
                "Target": "file:///private/baseline/external-object.bin",
                "TargetMode": "External",
            },
        )
        contents["xl/externalLinks/_rels/externalLink3.xml.rels"] = serialize(
            ole_relationships
        )

    return _rewrite_archive(path, mutate, ".external-link.tmp.xlsx")


def make_indexed_external_workbook_name_link_model(
    path: Path,
    *,
    target_paths: tuple[str, ...] = ("../inputs/source.xlsx",),
    source_name: str = "PrivateInputAlias",
    link_index: int = 1,
    consumer_alias_local_sheet_id: int | None = None,
    include_direct_indexed_formula: bool = True,
    consumer_formula_alias: bool = False,
    consumer_static_formula_name: str | None = None,
    consumer_static_lambda_name: str | None = None,
    external_reference: str | None = None,
    consumer_alias_name: str = "PackageExternalInput",
) -> Path:
    """Create a consumer with direct and workbook-scoped package links.

    The caller may provide a direct ``[N]!Name`` or ``[N]Sheet!A1`` spelling.
    Excel uses ``N`` as the document order of ``externalReference`` declarations.
    Relationships are deliberately emitted in reverse order so portfolio tests
    prove that a reader uses the declaration sequence rather than ZIP-part names
    or relationship order.
    """
    if not target_paths or not 1 <= link_index <= len(target_paths):
        raise ValueError("link_index must identify one supplied external target")
    if external_reference is None:
        external_reference = f"[{link_index}]!{source_name}"
    path.parent.mkdir(parents=True, exist_ok=True)
    make_model(path)
    workbook = load_workbook(path)
    consumer_formula_name = (
        "PackageExternalFormulaAlias"
        if consumer_formula_alias
        else consumer_alias_name
    )
    workbook["Model"]["D2"] = f"=SUM({consumer_formula_name})"
    if include_direct_indexed_formula:
        workbook["Model"]["E2"] = f"=SUM({external_reference})"
    workbook.defined_names.add(
        DefinedName(
            consumer_alias_name,
            attr_text=external_reference,
            localSheetId=consumer_alias_local_sheet_id,
        )
    )
    if consumer_formula_alias:
        workbook.defined_names.add(
            DefinedName(
                "PackageExternalFormulaAlias",
                attr_text=f"={consumer_alias_name}",
            )
        )
    if consumer_static_formula_name is not None:
        workbook["Model"]["F2"] = f"=SUM({consumer_static_formula_name})"
        workbook.defined_names.add(
            DefinedName(
                consumer_static_formula_name,
                attr_text=f"=SUM({consumer_formula_name})",
            )
        )
    if consumer_static_lambda_name is not None:
        workbook["Model"]["G2"] = f"=SUM({consumer_static_lambda_name}(7))"
        workbook.defined_names.add(
            DefinedName(
                consumer_static_lambda_name,
                attr_text=f"=LAMBDA(value,SUM(value,{consumer_formula_name}))",
            )
        )
    workbook.save(path)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_root = ElementTree.fromstring(contents["xl/workbook.xml"])
        external_references = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}externalReferences"
        )
        for number in range(1, len(target_paths) + 1):
            ElementTree.SubElement(
                external_references,
                f"{{{_SPREADSHEETML_NS}}}externalReference",
                {f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": f"rIdFenceExternal{number}"},
            )
        calculation = workbook_root.find(f"{{{_SPREADSHEETML_NS}}}calcPr")
        workbook_root.insert(
            list(workbook_root).index(calculation)
            if calculation is not None
            else len(workbook_root),
            external_references,
        )
        contents["xl/workbook.xml"] = serialize(workbook_root)

        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        for number in range(len(target_paths), 0, -1):
            ElementTree.SubElement(
                workbook_relationships,
                relationship_tag,
                {
                    "Id": f"rIdFenceExternal{number}",
                    "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/externalLink",
                    "Target": f"externalLinks/externalLink{number}.xml",
                },
            )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        for number, target_path in enumerate(target_paths, start=1):
            ElementTree.SubElement(
                content_types,
                override_tag,
                {
                    "PartName": f"/xl/externalLinks/externalLink{number}.xml",
                    "ContentType": (
                        "application/vnd.openxmlformats-officedocument.spreadsheetml."
                        "externalLink+xml"
                    ),
                },
            )
            external_link = ElementTree.Element(
                f"{{{_SPREADSHEETML_NS}}}externalLink"
            )
            ElementTree.SubElement(
                external_link,
                f"{{{_SPREADSHEETML_NS}}}externalBook",
                {f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": "rIdFenceExternalTarget"},
            )
            contents[f"xl/externalLinks/externalLink{number}.xml"] = serialize(
                external_link
            )

            external_relationships = ElementTree.Element(
                f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
            )
            ElementTree.SubElement(
                external_relationships,
                relationship_tag,
                {
                    "Id": "rIdFenceExternalTarget",
                    "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/externalLinkPath",
                    "Target": target_path,
                    "TargetMode": "External",
                },
            )
            contents[
                f"xl/externalLinks/_rels/externalLink{number}.xml.rels"
            ] = serialize(external_relationships)
        contents["[Content_Types].xml"] = serialize(content_types)

    return _rewrite_archive(path, mutate, ".indexed-external-name.tmp.xlsx")


def make_indexed_external_workbook_a1_link_model(
    path: Path,
    *,
    target_paths: tuple[str, ...] = ("../inputs/source.xlsx",),
    source_sheet: str = "Data",
    source_range: str = "$B$2:$B$4",
    link_index: int = 1,
    consumer_alias_local_sheet_id: int | None = None,
    include_direct_indexed_formula: bool = True,
    consumer_formula_alias: bool = False,
) -> Path:
    """Create direct and aliased ``[N]Sheet!A1`` package-link fixtures."""
    escaped_sheet = source_sheet.replace("'", "''")
    needs_quotes = any(
        character.isspace() or character in "'+-*/^&=<>%,;(){}!"
        for character in source_sheet
    )
    external_sheet = (
        f"'[{link_index}]{escaped_sheet}'"
        if needs_quotes
        else f"[{link_index}]{source_sheet}"
    )
    return make_indexed_external_workbook_name_link_model(
        path,
        target_paths=target_paths,
        link_index=link_index,
        consumer_alias_local_sheet_id=consumer_alias_local_sheet_id,
        include_direct_indexed_formula=include_direct_indexed_formula,
        consumer_formula_alias=consumer_formula_alias,
        external_reference=f"{external_sheet}!{source_range}",
        consumer_alias_name="PackageExternalCell",
    )


def make_indexed_external_workbook_sheet_defined_name_link_model(
    path: Path,
    *,
    target_paths: tuple[str, ...] = ("../inputs/source.xlsx",),
    source_sheet: str = "Data",
    source_name: str = "PrivateLocalInput",
    link_index: int = 1,
    consumer_alias_local_sheet_id: int | None = None,
    include_direct_indexed_formula: bool = True,
    consumer_formula_alias: bool = False,
) -> Path:
    """Create direct and aliased ``[N]Sheet!LocalName`` package links."""
    escaped_sheet = source_sheet.replace("'", "''")
    needs_quotes = any(
        character.isspace() or character in "'+-*/^&=<>%,;(){}!"
        for character in source_sheet
    )
    external_sheet = (
        f"'[{link_index}]{escaped_sheet}'"
        if needs_quotes
        else f"[{link_index}]{source_sheet}"
    )
    return make_indexed_external_workbook_name_link_model(
        path,
        target_paths=target_paths,
        link_index=link_index,
        consumer_alias_local_sheet_id=consumer_alias_local_sheet_id,
        include_direct_indexed_formula=include_direct_indexed_formula,
        consumer_formula_alias=consumer_formula_alias,
        external_reference=f"{external_sheet}!{source_name}",
        consumer_alias_name="PackageExternalSheetName",
    )


def change_external_link_package_controls(path: Path) -> Path:
    """Change external source, definition, cache, and opaque package material."""
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        external_book_relationships = ElementTree.fromstring(
            contents["xl/externalLinks/_rels/externalLink1.xml.rels"]
        )
        relationship = external_book_relationships.find(
            f"{{{package_relationships}}}Relationship"
        )
        if relationship is None:
            raise ValueError("Fixture does not contain an external-workbook target")
        relationship.set("Target", "file:///private/candidate/external-workbook.xlsx")
        contents["xl/externalLinks/_rels/externalLink1.xml.rels"] = serialize(
            external_book_relationships
        )

        external_book_root = ElementTree.fromstring(
            contents["xl/externalLinks/externalLink1.xml"]
        )
        defined_name = external_book_root.find(
            f"./{{{spreadsheet}}}externalBook/{{{spreadsheet}}}definedNames/"
            f"{{{spreadsheet}}}definedName"
        )
        cached_value = external_book_root.find(
            f"./{{{spreadsheet}}}externalBook/{{{spreadsheet}}}sheetDataSet/"
            f"{{{spreadsheet}}}sheetData/{{{spreadsheet}}}row/"
            f"{{{spreadsheet}}}cell/{{{spreadsheet}}}v"
        )
        if defined_name is None or cached_value is None:
            raise ValueError("Fixture does not contain external-workbook definition data")
        defined_name.set("refersTo", "'private external scenario'!$B$5")
        cached_value.text = "private external candidate cached value"
        contents["xl/externalLinks/externalLink1.xml"] = serialize(external_book_root)

        dde_root = ElementTree.fromstring(contents["xl/externalLinks/externalLink2.xml"])
        dde_link = dde_root.find(f"{{{spreadsheet}}}ddeLink")
        first_dde_item = dde_root.find(
            f"./{{{spreadsheet}}}ddeLink/{{{spreadsheet}}}ddeItems/"
            f"{{{spreadsheet}}}ddeItem"
        )
        dde_value = dde_root.find(
            f"./{{{spreadsheet}}}ddeLink/{{{spreadsheet}}}ddeItems/"
            f"{{{spreadsheet}}}ddeItem/{{{spreadsheet}}}values/"
            f"{{{spreadsheet}}}value/{{{spreadsheet}}}val"
        )
        if dde_link is None or first_dde_item is None or dde_value is None:
            raise ValueError("Fixture does not contain DDE controls")
        dde_link.set("ddeTopic", "private-candidate-dde-topic")
        first_dde_item.set("name", "private-candidate-dde-item")
        first_dde_item.set("advise", "0")
        dde_value.text = "private-candidate-dde-value"
        contents["xl/externalLinks/externalLink2.xml"] = serialize(dde_root)

        ole_root = ElementTree.fromstring(contents["xl/externalLinks/externalLink3.xml"])
        ole_link = ole_root.find(f"{{{spreadsheet}}}oleLink")
        first_ole_item = ole_root.find(
            f"./{{{spreadsheet}}}oleLink/{{{spreadsheet}}}oleItems/"
            f"{{{spreadsheet}}}oleItem"
        )
        extension = ole_root.find(f"./{{{spreadsheet}}}extLst/{{{spreadsheet}}}ext")
        if ole_link is None or first_ole_item is None or extension is None:
            raise ValueError("Fixture does not contain OLE controls")
        ole_link.set("progId", "private.candidate.ole.program")
        first_ole_item.set("name", "private-candidate-ole-item")
        first_ole_item.set("icon", "0")
        extension.set("uri", "urn:private:candidate:external-link-extension")
        extension.text = "private candidate external-link extension payload"
        contents["xl/externalLinks/externalLink3.xml"] = serialize(ole_root)

    return _rewrite_archive(path, mutate, ".external-link-change.tmp.xlsx")


def rebind_external_link_declaration(path: Path) -> Path:
    """Point one workbook declaration at a different existing externalLink part."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        relationship_tag = f"{{{package_relationships}}}Relationship"
        external_workbook = next(
            (
                relationship
                for relationship in relationships.findall(relationship_tag)
                if relationship.get("Id") == "rIdFenceExternalWorkbook"
            ),
            None,
        )
        if external_workbook is None:
            raise ValueError("Fixture does not contain the external-workbook declaration")
        external_workbook.set("Target", "externalLinks/externalLink2.xml")
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".external-link-rebind.tmp.xlsx")


def renumber_external_link_declaration_relationships(path: Path) -> Path:
    """Rewrite only arbitrary workbook external-link relationship identifiers."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    document_relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    replacements = {
        "rIdFenceExternalWorkbook": "rIdFenceRenumberedWorkbook",
        "rIdFenceDde": "rIdFenceRenumberedDde",
        "rIdFenceOle": "rIdFenceRenumberedOle",
    }

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        relationship_tag = f"{{{package_relationships}}}Relationship"
        for relationship in relationships.findall(relationship_tag):
            if replacement := replacements.get(relationship.get("Id")):
                relationship.set("Id", replacement)
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        reference_tag = (
            "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            "externalReference"
        )
        relationship_id_attribute = f"{{{document_relationships}}}id"
        for reference in workbook.findall(f".//{reference_tag}"):
            if replacement := replacements.get(reference.get(relationship_id_attribute)):
                reference.set(relationship_id_attribute, replacement)
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".external-link-renumber.tmp.xlsx")


def make_external_relationship_model(path: Path) -> Path:
    """Create opaque package-wide external relationships outside known bindings."""
    make_model(path)
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    workbook_relationships_member = "xl/_rels/workbook.xml.rels"
    worksheet_member = "xl/worksheets/sheet1.xml"
    worksheet_relationships_member = _relationship_member(worksheet_member)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        if workbook_relationships_member not in contents:
            raise ValueError("Fixture does not contain workbook relationships")
        relationships_tag = f"{{{package_relationships}}}Relationships"
        relationship_tag = f"{{{package_relationships}}}Relationship"
        workbook_relationships = ElementTree.fromstring(
            contents[workbook_relationships_member]
        )
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceOpaqueExternal",
                "Type": "https://private.example.test/relationships/opaque-external",
                "Target": "https://private.example.test/PRIVATE-PACKAGE-EXTERNAL-BASELINE",
                "TargetMode": "External",
            },
        )
        contents[workbook_relationships_member] = serialize(workbook_relationships)

        if worksheet_member not in contents:
            raise ValueError("Fixture does not contain the expected worksheet part")
        worksheet_relationships = (
            ElementTree.fromstring(contents[worksheet_relationships_member])
            if worksheet_relationships_member in contents
            else ElementTree.Element(relationships_tag)
        )
        ElementTree.SubElement(
            worksheet_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceOpaqueHyperlink",
                "Type": f"{document_relationships}/hyperlink",
                "Target": "https://private.example.test/PRIVATE-PACKAGE-HYPERLINK-BASELINE",
                "TargetMode": "External",
            },
        )
        ElementTree.SubElement(
            worksheet_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceOpaqueImage",
                "Type": f"{document_relationships}/image",
                "Target": "https://private.example.test/PRIVATE-PACKAGE-IMAGE-BASELINE",
                "TargetMode": "External",
            },
        )
        contents[worksheet_relationships_member] = serialize(worksheet_relationships)

    return _rewrite_archive(path, mutate, ".external-relationship-model.tmp.xlsx")


def change_external_relationship_target(path: Path) -> Path:
    """Retarget an opaque external relationship without changing workbook cells."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        member = "xl/_rels/workbook.xml.rels"
        relationships = ElementTree.fromstring(contents[member])
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if current.get("Id") == "rIdFenceOpaqueExternal"
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain the opaque external relationship")
        relationship.set(
            "Target",
            "https://private.example.test/PRIVATE-PACKAGE-EXTERNAL-CANDIDATE",
        )
        contents[member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".external-relationship-target.tmp.xlsx")


def renumber_external_relationship_identifiers(path: Path) -> Path:
    """Rewrite arbitrary relationship identifiers without changing endpoints."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    replacements = {
        "rIdFenceOpaqueExternal": "rIdFenceOpaqueExternalRenumbered",
        "rIdFenceOpaqueHyperlink": "rIdFenceOpaqueHyperlinkRenumbered",
        "rIdFenceOpaqueImage": "rIdFenceOpaqueImageRenumbered",
    }

    def mutate(contents: dict[str, bytes]) -> None:
        for member in (
            "xl/_rels/workbook.xml.rels",
            _relationship_member("xl/worksheets/sheet1.xml"),
        ):
            relationships = ElementTree.fromstring(contents[member])
            for relationship in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            ):
                if replacement := replacements.get(relationship.get("Id")):
                    relationship.set("Id", replacement)
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".external-relationship-renumber.tmp.xlsx")


def corrupt_external_relationship_metadata(path: Path) -> Path:
    """Make one external relationship unrecognizable without exposing its target."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        member = _relationship_member("xl/worksheets/sheet1.xml")
        relationships = ElementTree.fromstring(contents[member])
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if current.get("Id") == "rIdFenceOpaqueHyperlink"
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain the opaque hyperlink relationship")
        relationship.set("privateUnknownAttribute", "PRIVATE-OPAQUE-ATTRIBUTE")
        contents[member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".external-relationship-corrupt.tmp.xlsx")


def duplicate_external_link_definition(path: Path) -> Path:
    """Add a second supported definition to exercise the fail-closed parser path."""
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    def mutate(contents: dict[str, bytes]) -> None:
        root = ElementTree.fromstring(contents["xl/externalLinks/externalLink1.xml"])
        external_book = root.find(f"{{{spreadsheet}}}externalBook")
        if external_book is None:
            raise ValueError("Fixture does not contain an external-workbook definition")
        root.append(
            ElementTree.fromstring(
                ElementTree.tostring(external_book, encoding="utf-8")
            )
        )
        contents["xl/externalLinks/externalLink1.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".external-link-duplicate.tmp.xlsx")


def duplicate_indexed_external_link_part_binding(path: Path) -> Path:
    """Bind two indexed declarations to one link part to make it ambiguous."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if current.get("Id") == "rIdFenceExternal2"
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture needs a second indexed external-link declaration")
        relationship.set("Target", "externalLinks/externalLink1.xml")
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".indexed-external-link-rebind.tmp.xlsx")


def duplicate_external_link_sheet_names(path: Path) -> Path:
    """Repeat a schema-singleton workbook child to test coverage reporting."""
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    def mutate(contents: dict[str, bytes]) -> None:
        root = ElementTree.fromstring(contents["xl/externalLinks/externalLink1.xml"])
        external_book = root.find(f"{{{spreadsheet}}}externalBook")
        if external_book is None:
            raise ValueError("Fixture does not contain an external-workbook definition")
        sheet_names = external_book.find(f"{{{spreadsheet}}}sheetNames")
        if sheet_names is None:
            raise ValueError("Fixture does not contain external-workbook sheet names")
        external_book.append(
            ElementTree.fromstring(ElementTree.tostring(sheet_names, encoding="utf-8"))
        )
        contents["xl/externalLinks/externalLink1.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".external-link-repeated-child.tmp.xlsx")


def make_xlm_macro_sheet_model(path: Path, *, international: bool = False) -> Path:
    """Create a harmless raw XLM macro-sheet fixture that readers commonly omit.

    The macro formula strings are never opened in Excel or evaluated. They are
    only private OOXML test material for FormulaFence's package scanner.
    """
    make_model(path)
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    document_relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    content_types_namespace = (
        "http://schemas.openxmlformats.org/package/2006/content-types"
    )
    macro_namespace = "http://schemas.microsoft.com/office/excel/2006/main"
    macro_member = (
        "xl/macrosheets/intlsheet1.xml"
        if international
        else "xl/macrosheets/sheet1.xml"
    )
    macro_target = macro_member.removeprefix("xl/")
    macro_kind = "xlIntlMacrosheet" if international else "xlMacrosheet"
    macro_content_type = (
        "application/vnd.ms-excel.intlmacrosheet+xml"
        if international
        else "application/vnd.ms-excel.macrosheet+xml"
    )

    def serialize(root: ElementTree.Element) -> bytes:
        namespace = root.tag[1:].split("}", maxsplit=1)[0]
        if namespace in {
            spreadsheet,
            package_relationships,
            content_types_namespace,
            macro_namespace,
        }:
            ElementTree.register_namespace("", namespace)
        ElementTree.register_namespace("r", document_relationships)
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        sheets = workbook.find(f"{{{spreadsheet}}}sheets")
        if sheets is None:
            raise ValueError("Fixture does not contain workbook sheets")
        sheet_ids = [int(sheet.get("sheetId", "0")) for sheet in sheets]
        ElementTree.SubElement(
            sheets,
            f"{{{spreadsheet}}}sheet",
            {
                "name": "Macro Automation",
                "sheetId": str(max(sheet_ids) + 1),
                "state": "veryHidden",
                f"{{{document_relationships}}}id": "rIdFenceXlmMacro",
            },
        )
        contents["xl/workbook.xml"] = serialize(workbook)

        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            workbook_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceXlmMacro",
                "Type": (
                    "http://schemas.microsoft.com/office/2006/relationships/"
                    f"{macro_kind}"
                ),
                "Target": macro_target,
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override_tag = f"{{{content_types_namespace}}}Override"
        for override in content_types.findall(override_tag):
            if override.get("PartName") == "/xl/workbook.xml":
                override.set(
                    "ContentType",
                    "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                )
        ElementTree.SubElement(
            content_types,
            override_tag,
            {"PartName": f"/{macro_member}", "ContentType": macro_content_type},
        )
        contents["[Content_Types].xml"] = serialize(content_types)

        macro_sheet = ElementTree.Element(f"{{{macro_namespace}}}macrosheet")
        sheet_data = ElementTree.SubElement(
            macro_sheet, f"{{{macro_namespace}}}sheetData"
        )
        for row_number, formula, value in (
            (1, 'PRIVATE.XLM("private-baseline-xl-command-argument")', None),
            (2, "RETURN()", "private-baseline-xl-cell-value"),
        ):
            row = ElementTree.SubElement(
                sheet_data,
                f"{{{macro_namespace}}}row",
                {"r": str(row_number)},
            )
            cell = ElementTree.SubElement(
                row,
                f"{{{macro_namespace}}}c",
                {"r": f"A{row_number}"},
            )
            ElementTree.SubElement(cell, f"{{{macro_namespace}}}f").text = formula
            if value is not None:
                ElementTree.SubElement(cell, f"{{{macro_namespace}}}v").text = value
        ole_objects = ElementTree.SubElement(
            macro_sheet, f"{{{macro_namespace}}}oleObjects"
        )
        ElementTree.SubElement(
            ole_objects,
            f"{{{macro_namespace}}}oleObject",
            {
                f"{{{document_relationships}}}id": "rIdFenceEmbeddedObject",
                "progId": "private.baseline.xlm.embedded.object",
            },
        )
        ElementTree.SubElement(
            ole_objects,
            f"{{{macro_namespace}}}oleObject",
            {
                f"{{{document_relationships}}}id": "rIdFenceExternalObject",
                "progId": "private.baseline.xlm.linked.object",
            },
        )
        contents[macro_member] = serialize(macro_sheet)

        macro_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        relationship_tag = f"{{{package_relationships}}}Relationship"
        for relationship_id, relationship_type, target, target_mode in (
            (
                "rIdFenceEmbeddedObject",
                f"{document_relationships}/oleObject",
                "../embeddings/private-baseline-xl-object.bin",
                None,
            ),
            (
                "rIdFenceExternalObject",
                f"{document_relationships}/oleObject",
                "file:///private/baseline-xl-linked-object.bin",
                "External",
            ),
            (
                "rIdFenceEmbeddedPackage",
                f"{document_relationships}/package",
                "../embeddings/private-baseline-xl-package.bin",
                None,
            ),
        ):
            attributes = {
                "Id": relationship_id,
                "Type": relationship_type,
                "Target": target,
            }
            if target_mode is not None:
                attributes["TargetMode"] = target_mode
            ElementTree.SubElement(macro_relationships, relationship_tag, attributes)
        contents[
            "xl/macrosheets/_rels/" + macro_member.rsplit("/", maxsplit=1)[-1] + ".rels"
        ] = serialize(macro_relationships)
        contents["xl/embeddings/private-baseline-xl-object.bin"] = (
            b"private baseline embedded XLM object payload"
        )
        contents["xl/embeddings/private-baseline-xl-package.bin"] = (
            b"private baseline embedded XLM package payload"
        )

    return _rewrite_archive(path, mutate, ".xlm-macro-sheet.tmp.xlsx")


def add_xlm_automatic_macro_binding(
    path: Path,
    *,
    name: str = "Auto_Open",
    target: str = "'Macro Automation'!$A$1",
    local_sheet_id: int | None = None,
) -> Path:
    """Add one raw workbook defined name for an automatic-macro fixture.

    The target remains harmless static OOXML test material. This helper never
    opens the workbook in Excel or asks a workbook reader to interpret legacy
    macro dispatch.
    """
    spreadsheet = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        defined_names = workbook.find(f"{{{spreadsheet}}}definedNames")
        if defined_names is None:
            defined_names = ElementTree.Element(f"{{{spreadsheet}}}definedNames")
            sheets = workbook.find(f"{{{spreadsheet}}}sheets")
            if sheets is None:
                raise ValueError("Fixture does not contain workbook sheets")
            workbook.insert(list(workbook).index(sheets) + 1, defined_names)
        attributes = {"name": name}
        if local_sheet_id is not None:
            attributes["localSheetId"] = str(local_sheet_id)
        ElementTree.SubElement(
            defined_names,
            f"{{{spreadsheet}}}definedName",
            attributes,
        ).text = target
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xlm-automatic-macro.tmp.xlsx")


def duplicate_xlm_macro_sheet_workbook_relationship(path: Path) -> Path:
    """Duplicate a macro-sheet workbook relationship to exercise fail-closed binding."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        relationship_tag = f"{{{package_relationships}}}Relationship"
        macro_relationship = next(
            (
                relationship
                for relationship in relationships.findall(relationship_tag)
                if relationship.get("Type", "").rsplit("/", maxsplit=1)[-1]
                in {"xlMacrosheet", "xlIntlMacrosheet"}
            ),
            None,
        )
        if macro_relationship is None:
            raise ValueError("Fixture does not contain an XLM macro-sheet relationship")
        relationships.append(
            ElementTree.fromstring(
                ElementTree.tostring(macro_relationship, encoding="utf-8")
            )
        )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xlm-automatic-macro-duplicate.tmp.xlsx")


def change_xlm_macro_sheet_controls(path: Path) -> Path:
    """Change private XLM code, binding, and related-part material."""
    macro_namespace = "http://schemas.microsoft.com/office/excel/2006/main"
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    macro_member = "xl/macrosheets/sheet1.xml"

    def mutate(contents: dict[str, bytes]) -> None:
        macro_sheet = ElementTree.fromstring(contents[macro_member])
        formula = macro_sheet.find(
            f"./{{{macro_namespace}}}sheetData/{{{macro_namespace}}}row/"
            f"{{{macro_namespace}}}c/{{{macro_namespace}}}f"
        )
        cell_value = macro_sheet.find(
            f"./{{{macro_namespace}}}sheetData/{{{macro_namespace}}}row[2]/"
            f"{{{macro_namespace}}}c/{{{macro_namespace}}}v"
        )
        embedded_object = macro_sheet.find(
            f"./{{{macro_namespace}}}oleObjects/{{{macro_namespace}}}oleObject"
        )
        if formula is None or cell_value is None or embedded_object is None:
            raise ValueError("Fixture does not contain expected XLM macro-sheet controls")
        formula.text = 'PRIVATE.XLM("private-candidate-xl-command-argument")'
        cell_value.text = "private-candidate-xl-cell-value"
        embedded_object.set("progId", "private.candidate.xlm.embedded.object")
        extensions = ElementTree.SubElement(macro_sheet, f"{{{macro_namespace}}}extLst")
        ElementTree.SubElement(
            extensions,
            f"{{{macro_namespace}}}ext",
            {"uri": "urn:private:candidate:xlm-extension"},
        ).text = "private candidate XLM extension payload"
        contents[macro_member] = ElementTree.tostring(
            macro_sheet,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships_name = "xl/macrosheets/_rels/sheet1.xml.rels"
        relationships = ElementTree.fromstring(contents[relationships_name])
        relationship_tag = f"{{{package_relationships}}}Relationship"
        for relationship in relationships.findall(relationship_tag):
            if relationship.get("Id") == "rIdFenceEmbeddedObject":
                relationship.set(
                    "Target", "../embeddings/private-candidate-xl-object.bin"
                )
            elif relationship.get("Id") == "rIdFenceExternalObject":
                relationship.set(
                    "Target", "file:///private/candidate-xl-linked-object.bin"
                )
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/embeddings/private-candidate-xl-object.bin"] = (
            b"private candidate embedded XLM object payload"
        )

        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        sheet = next(
            (
                item
                for item in workbook.findall(
                    ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
                )
                if item.get("name") == "Macro Automation"
            ),
            None,
        )
        if sheet is None:
            raise ValueError("Fixture does not contain XLM macro-sheet declaration")
        sheet.set("state", "hidden")
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xlm-macro-sheet-change.tmp.xlsx")


def change_xlm_macro_sheet_related_part_payload(path: Path) -> Path:
    """Change only a private internal payload reached from an XLM macro sheet."""

    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/embeddings/private-baseline-xl-object.bin"] = (
            b"private candidate XLM related-part payload only"
        )

    return _rewrite_archive(path, mutate, ".xlm-related-payload-change.tmp.xlsx")


def remove_xlm_macro_sheet_related_part_payload(path: Path) -> Path:
    """Remove one internal XLM related payload to exercise fail-closed coverage."""

    def mutate(contents: dict[str, bytes]) -> None:
        contents.pop("xl/embeddings/private-baseline-xl-object.bin")

    return _rewrite_archive(path, mutate, ".xlm-related-payload-missing.tmp.xlsx")


def renumber_xlm_macro_sheet_relationships(path: Path) -> Path:
    """Rewrite arbitrary XLM relationship ids while retaining their bindings."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    document_relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    replacements = {
        "rIdFenceXlmMacro": "rIdFenceRenumberedXlmMacro",
        "rIdFenceEmbeddedObject": "rIdFenceRenumberedEmbeddedObject",
        "rIdFenceExternalObject": "rIdFenceRenumberedExternalObject",
        "rIdFenceEmbeddedPackage": "rIdFenceRenumberedEmbeddedPackage",
    }

    def mutate(contents: dict[str, bytes]) -> None:
        relationship_tag = f"{{{package_relationships}}}Relationship"
        for name in (
            "xl/_rels/workbook.xml.rels",
            "xl/macrosheets/_rels/sheet1.xml.rels",
        ):
            relationships = ElementTree.fromstring(contents[name])
            for relationship in relationships.findall(relationship_tag):
                if replacement := replacements.get(relationship.get("Id")):
                    relationship.set("Id", replacement)
            contents[name] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

        relationship_id_attribute = f"{{{document_relationships}}}id"
        for name in ("xl/workbook.xml", "xl/macrosheets/sheet1.xml"):
            root = ElementTree.fromstring(contents[name])
            for element in root.iter():
                if replacement := replacements.get(element.get(relationship_id_attribute)):
                    element.set(relationship_id_attribute, replacement)
            contents[name] = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".xlm-macro-sheet-renumber.tmp.xlsx")


def rewrite_xlm_macro_sheet_internal_target_spelling(path: Path) -> Path:
    """Rewrite an XLM internal target using an equivalent relative path."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        relationships_name = "xl/_rels/workbook.xml.rels"
        relationships = ElementTree.fromstring(contents[relationships_name])
        macro_relationship = next(
            (
                relationship
                for relationship in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if relationship.get("Type", "").endswith("/xlMacrosheet")
            ),
            None,
        )
        if macro_relationship is None:
            raise ValueError("Fixture does not contain an XLM macro-sheet relationship")
        macro_relationship.set(
            "Target", "./macrosheets/../macrosheets/sheet1.xml"
        )
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xlm-macro-sheet-target.tmp.xlsx")


def corrupt_xlm_macro_sheet_root(path: Path) -> Path:
    """Replace the macro-sheet root with an unexpected element for fail-closed tests."""
    macro_namespace = "http://schemas.microsoft.com/office/excel/2006/main"

    def mutate(contents: dict[str, bytes]) -> None:
        macro_sheet = ElementTree.fromstring(contents["xl/macrosheets/sheet1.xml"])
        macro_sheet.tag = f"{{{macro_namespace}}}notMacroSheet"
        contents["xl/macrosheets/sheet1.xml"] = ElementTree.tostring(
            macro_sheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".xlm-macro-sheet-corrupt.tmp.xlsx")


def make_ribbon_customization_model(
    path: Path,
    *,
    office_2010: bool = False,
    compatibility_namespace: bool = False,
) -> Path:
    """Create a harmless raw RibbonX fixture outside ordinary workbook XML.

    The callback names and labels are synthetic private material. The fixture
    is never opened by Office; FormulaFence only reads its bounded package
    parts before the workbook reader can omit them.
    """
    make_model(path)
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    document_relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    if compatibility_namespace and not office_2010:
        raise ValueError("RibbonX compatibility namespace requires office_2010=True")
    custom_ui_namespace = (
        "http://schemas.microsoft.com/office/2007/10/customui"
        if compatibility_namespace
        else (
            "http://schemas.microsoft.com/office/2009/07/customui"
            if office_2010
            else "http://schemas.microsoft.com/office/2006/01/customui"
        )
    )
    custom_ui_member = (
        "customUI/customUI14.xml" if office_2010 else "customUI/customUI.xml"
    )
    customization_relationship = (
        "http://schemas.microsoft.com/office/2007/relationships/ui/extensibility"
        if office_2010
        else "http://schemas.microsoft.com/office/2006/relationships/ui/extensibility"
    )
    image_relationship = f"{document_relationships}/image"

    def serialize(root: ElementTree.Element) -> bytes:
        namespace = root.tag[1:].split("}", maxsplit=1)[0]
        if namespace in {package_relationships, custom_ui_namespace}:
            ElementTree.register_namespace("", namespace)
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        root_relationships = ElementTree.fromstring(contents["_rels/.rels"])
        ElementTree.SubElement(
            root_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceRibbonCustomization",
                "Type": customization_relationship,
                "Target": f"/{custom_ui_member}",
            },
        )
        contents["_rels/.rels"] = serialize(root_relationships)

        custom_ui = ElementTree.Element(
            f"{{{custom_ui_namespace}}}customUI",
            {"onLoad": "PrivateBaselineRibbonLoad"},
        )
        ribbon = ElementTree.SubElement(custom_ui, f"{{{custom_ui_namespace}}}ribbon")
        tabs = ElementTree.SubElement(ribbon, f"{{{custom_ui_namespace}}}tabs")
        tab = ElementTree.SubElement(
            tabs,
            f"{{{custom_ui_namespace}}}tab",
            {"id": "FenceTab", "label": "private baseline ribbon tab"},
        )
        group = ElementTree.SubElement(
            tab,
            f"{{{custom_ui_namespace}}}group",
            {"id": "FenceGroup", "label": "private baseline ribbon group"},
        )
        ElementTree.SubElement(
            group,
            f"{{{custom_ui_namespace}}}button",
            {
                "id": "FenceAction",
                "label": "private baseline ribbon action",
                "onAction": "PrivateBaselineRibbonAction",
                "image": "rIdFenceRibbonImage",
            },
        )
        contents[custom_ui_member] = serialize(custom_ui)

        image_member = "customUI/images/private-baseline-ribbon.png"
        part_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            part_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceRibbonImage",
                "Type": image_relationship,
                "Target": "images/private-baseline-ribbon.png",
            },
        )
        relationship_member = (
            "customUI/_rels/"
            f"{custom_ui_member.rsplit('/', maxsplit=1)[-1]}.rels"
        )
        contents[relationship_member] = serialize(part_relationships)
        contents[image_member] = b"private baseline RibbonX image payload"

    return _rewrite_archive(path, mutate, ".ribbon-customization.tmp.xlsx")


def change_ribbon_customization_callback(path: Path) -> Path:
    """Change only a private RibbonX callback name."""

    def mutate(contents: dict[str, bytes]) -> None:
        custom_ui_member = next(
            name
            for name in contents
            if name.startswith("customUI/")
            and name.endswith(".xml")
            and "/_rels/" not in name
        )
        custom_ui = ElementTree.fromstring(contents[custom_ui_member])
        button = next(
            (
                element
                for element in custom_ui.iter()
                if element.tag.rsplit("}", maxsplit=1)[-1] == "button"
            ),
            None,
        )
        if button is None:
            raise ValueError("Fixture does not contain a RibbonX button")
        button.set("onAction", "PrivateCandidateRibbonAction")
        contents[custom_ui_member] = ElementTree.tostring(
            custom_ui,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ribbon-callback-change.tmp.xlsx")


def change_ribbon_customization_controls(path: Path) -> Path:
    """Change private RibbonX callbacks, labels, and image relationship material."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        custom_ui_member = next(
            name
            for name in contents
            if name.startswith("customUI/")
            and name.endswith(".xml")
            and "/_rels/" not in name
        )
        custom_ui = ElementTree.fromstring(contents[custom_ui_member])
        button = next(
            (
                element
                for element in custom_ui.iter()
                if element.tag.rsplit("}", maxsplit=1)[-1] == "button"
            ),
            None,
        )
        if button is None:
            raise ValueError("Fixture does not contain a RibbonX button")
        custom_ui.set("onLoad", "PrivateCandidateRibbonLoad")
        button.set("label", "private candidate ribbon action")
        button.set("onAction", "PrivateCandidateRibbonAction")
        contents[custom_ui_member] = ElementTree.tostring(
            custom_ui,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_member = (
            "customUI/_rels/"
            f"{custom_ui_member.rsplit('/', maxsplit=1)[-1]}.rels"
        )
        relationships = ElementTree.fromstring(contents[relationship_member])
        relationship = next(
            (
                item
                for item in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if item.get("Id") == "rIdFenceRibbonImage"
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a RibbonX image relationship")
        relationship.set("Target", "images/private-candidate-ribbon.png")
        contents[relationship_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["customUI/images/private-candidate-ribbon.png"] = (
            b"private candidate RibbonX image payload"
        )

    return _rewrite_archive(path, mutate, ".ribbon-customization-change.tmp.xlsx")


def renumber_ribbon_customization_relationships(path: Path) -> Path:
    """Rewrite RibbonX package IDs while retaining their semantic bindings."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    replacements = {
        "rIdFenceRibbonCustomization": "rIdFenceRenumberedRibbonCustomization",
        "rIdFenceRibbonImage": "rIdFenceRenumberedRibbonImage",
    }

    def mutate(contents: dict[str, bytes]) -> None:
        root_relationships = ElementTree.fromstring(contents["_rels/.rels"])
        for relationship in root_relationships.findall(
            f"{{{package_relationships}}}Relationship"
        ):
            if replacement := replacements.get(relationship.get("Id")):
                relationship.set("Id", replacement)
        contents["_rels/.rels"] = ElementTree.tostring(
            root_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        custom_ui_member = next(
            name
            for name in contents
            if name.startswith("customUI/")
            and name.endswith(".xml")
            and "/_rels/" not in name
        )
        custom_ui = ElementTree.fromstring(contents[custom_ui_member])
        for element in custom_ui.iter():
            if element.get("image") == "rIdFenceRibbonImage":
                element.set("image", "rIdFenceRenumberedRibbonImage")
        contents[custom_ui_member] = ElementTree.tostring(
            custom_ui,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_member = (
            "customUI/_rels/"
            f"{custom_ui_member.rsplit('/', maxsplit=1)[-1]}.rels"
        )
        relationships = ElementTree.fromstring(contents[relationship_member])
        for relationship in relationships.findall(
            f"{{{package_relationships}}}Relationship"
        ):
            if replacement := replacements.get(relationship.get("Id")):
                relationship.set("Id", replacement)
        contents[relationship_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ribbon-customization-renumber.tmp.xlsx")


def rewrite_ribbon_customization_internal_target_spelling(path: Path) -> Path:
    """Rewrite a RibbonX package target using an equivalent relative path."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(contents["_rels/.rels"])
        customization = next(
            (
                relationship
                for relationship in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if relationship.get("Type", "").endswith("/ui/extensibility")
            ),
            None,
        )
        if customization is None:
            raise ValueError("Fixture does not contain a RibbonX package declaration")
        customization.set("Target", "./customUI/../customUI/customUI.xml")
        contents["_rels/.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ribbon-customization-target.tmp.xlsx")


def corrupt_ribbon_customization_root(path: Path) -> Path:
    """Replace the RibbonX root with an unexpected element for coverage tests."""

    def mutate(contents: dict[str, bytes]) -> None:
        custom_ui_member = next(
            name
            for name in contents
            if name.startswith("customUI/")
            and name.endswith(".xml")
            and "/_rels/" not in name
        )
        custom_ui = ElementTree.fromstring(contents[custom_ui_member])
        namespace = custom_ui.tag[1:].split("}", maxsplit=1)[0]
        custom_ui.tag = f"{{{namespace}}}notCustomUI"
        contents[custom_ui_member] = ElementTree.tostring(
            custom_ui,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ribbon-customization-corrupt.tmp.xlsx")


def make_office_web_addin_model(
    path: Path,
    *,
    taskpane_reference_element: str = "webextension",
    worksheet_binding: bool = False,
) -> Path:
    """Create a harmless Office Web Add-in task-pane package fixture.

    The synthetic extension is never opened by Office. FormulaFence only reads
    its bounded OOXML task-pane and web-extension parts before the workbook
    reader can omit them.
    """
    if taskpane_reference_element not in {"webextension", "webextensionref"}:
        raise ValueError("Unsupported task-pane web-extension reference element")
    make_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    document_relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    taskpanes_namespace = (
        "http://schemas.microsoft.com/office/webextensions/taskpanes/2010/11"
    )
    web_extension_namespace = (
        "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
    )
    taskpane_relationship = (
        "http://schemas.microsoft.com/office/2011/relationships/webextensiontaskpanes"
    )
    web_extension_relationship = (
        "http://schemas.microsoft.com/office/2011/relationships/webextension"
    )
    worksheet_web_extension_uri = "{F7C9EE02-42E1-4005-9D12-6889AFFD525C}"
    image_relationship = f"{document_relationships}/image"

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_relationships = ElementTree.fromstring(
            contents["xl/_rels/workbook.xml.rels"]
        )
        ElementTree.SubElement(
            workbook_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceWebTaskpanes",
                "Type": taskpane_relationship,
                "Target": "webextensions/taskpanes.xml",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        for member, content_type in (
            (
                "/xl/webextensions/taskpanes.xml",
                "application/vnd.ms-office.webextensiontaskpanes+xml",
            ),
            (
                "/xl/webextensions/webextension1.xml",
                "application/vnd.ms-office.webextension+xml",
            ),
        ):
            ElementTree.SubElement(
                types,
                f"{{{content_types}}}Override",
                {"PartName": member, "ContentType": content_type},
            )
        contents["[Content_Types].xml"] = serialize(types)

        taskpanes = ElementTree.Element(f"{{{taskpanes_namespace}}}taskpanes")
        taskpane = ElementTree.SubElement(
            taskpanes,
            f"{{{taskpanes_namespace}}}taskpane",
            {
                "dockstate": "right",
                "visibility": "1",
                "width": "350",
                "row": "4",
                "locked": "1",
            },
        )
        ElementTree.SubElement(
            taskpane,
            f"{{{taskpanes_namespace}}}{taskpane_reference_element}",
            {f"{{{document_relationships}}}id": "rIdFenceTaskpaneExtension"},
        )
        contents["xl/webextensions/taskpanes.xml"] = serialize(taskpanes)

        taskpane_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            taskpane_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceTaskpaneExtension",
                "Type": web_extension_relationship,
                "Target": "webextension1.xml",
            },
        )
        contents["xl/webextensions/_rels/taskpanes.xml.rels"] = serialize(
            taskpane_relationships
        )

        web_extension = ElementTree.Element(
            f"{{{web_extension_namespace}}}webextension",
            {"id": "{11111111-1111-1111-1111-111111111111}"},
        )
        ElementTree.SubElement(
            web_extension,
            f"{{{web_extension_namespace}}}reference",
            {
                "id": "PrivateBaselineAddin",
                "version": "1.0.0.0",
                "store": "private-baseline-manifest.xml",
                "storeType": "Filesystem",
            },
        )
        alternate_references = ElementTree.SubElement(
            web_extension,
            f"{{{web_extension_namespace}}}alternateReferences",
        )
        ElementTree.SubElement(
            alternate_references,
            f"{{{web_extension_namespace}}}reference",
            {
                "id": "PrivateFallbackAddin",
                "version": "1.0.0.0",
                "store": "private-fallback-manifest.xml",
                "storeType": "Filesystem",
            },
        )
        properties = ElementTree.SubElement(
            web_extension,
            f"{{{web_extension_namespace}}}properties",
        )
        ElementTree.SubElement(
            properties,
            f"{{{web_extension_namespace}}}property",
            {"name": "Office.AutoShowTaskpaneWithDocument", "value": "true"},
        )
        ElementTree.SubElement(
            properties,
            f"{{{web_extension_namespace}}}property",
            {"name": "PrivateAddinBehavior", "value": "private baseline behavior"},
        )
        bindings = ElementTree.SubElement(
            web_extension,
            f"{{{web_extension_namespace}}}bindings",
        )
        ElementTree.SubElement(
            bindings,
            f"{{{web_extension_namespace}}}binding",
            {
                "id": "PrivateBaselineBinding",
                "type": "table",
                "appref": "PrivateBaselineTable",
            },
        )
        ElementTree.SubElement(
            web_extension,
            f"{{{web_extension_namespace}}}snapshot",
            {f"{{{document_relationships}}}embed": "rIdFenceSnapshot"},
        )
        contents["xl/webextensions/webextension1.xml"] = serialize(web_extension)

        extension_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            extension_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceSnapshot",
                "Type": image_relationship,
                "Target": "snapshots/private-baseline.png",
            },
        )
        ElementTree.SubElement(
            extension_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceExternalPreview",
                "Type": image_relationship,
                "Target": "https://private.example.invalid/addin-preview.png",
                "TargetMode": "External",
            },
        )
        contents["xl/webextensions/_rels/webextension1.xml.rels"] = serialize(
            extension_relationships
        )
        contents["xl/webextensions/snapshots/private-baseline.png"] = (
            b"private baseline add-in snapshot"
        )

        if worksheet_binding:
            worksheet = _inputs_worksheet_root(contents)
            extension_list = ElementTree.SubElement(
                worksheet,
                f"{{{_SPREADSHEETML_NS}}}extLst",
            )
            extension = ElementTree.SubElement(
                extension_list,
                f"{{{_SPREADSHEETML_NS}}}ext",
                {"uri": worksheet_web_extension_uri},
            )
            bindings = ElementTree.SubElement(
                extension,
                f"{{{_OFFICE_2013_SPREADSHEET_NS}}}webExtensions",
            )
            binding = ElementTree.SubElement(
                bindings,
                f"{{{_OFFICE_2013_SPREADSHEET_NS}}}webExtension",
                {"appRef": "PrivateBaselineTable"},
            )
            ElementTree.SubElement(
                binding,
                f"{{{_EXCEL_2006_MAIN_NS}}}f",
            ).text = "Inputs!$B$2:$B$4"
            _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".office-web-addin.tmp.xlsx")


def change_office_web_addin_auto_show(path: Path) -> Path:
    """Disable only the synthetic document auto-show request."""
    web_extension_namespace = (
        "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        web_extension = ElementTree.fromstring(
            contents["xl/webextensions/webextension1.xml"]
        )
        auto_show = next(
            (
                property_element
                for property_element in web_extension.iter(
                    f"{{{web_extension_namespace}}}property"
                )
                if property_element.get("name")
                == "Office.AutoShowTaskpaneWithDocument"
            ),
            None,
        )
        if auto_show is None:
            raise ValueError("Fixture does not contain an auto-show property")
        auto_show.set("value", "false")
        contents["xl/webextensions/webextension1.xml"] = ElementTree.tostring(
            web_extension,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".office-web-addin-autoshow.tmp.xlsx")


def change_office_web_addin_worksheet_binding(path: Path) -> Path:
    """Change only the private local range bound to a worksheet add-in."""
    formula_tag = f"{{{_EXCEL_2006_MAIN_NS}}}f"

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = _inputs_worksheet_root(contents)
        formula = next(worksheet.iter(formula_tag), None)
        if formula is None:
            raise ValueError("Fixture does not contain a worksheet add-in binding")
        formula.text = "Inputs!$B$2:$B$3"
        _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".office-web-addin-worksheet-change.tmp.xlsx")


def corrupt_office_web_addin_worksheet_binding(path: Path) -> Path:
    """Remove a required worksheet appRef so add-in parsing fails closed."""
    binding_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}webExtension"

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = _inputs_worksheet_root(contents)
        binding = next(worksheet.iter(binding_tag), None)
        if binding is None:
            raise ValueError("Fixture does not contain a worksheet add-in binding")
        binding.attrib.pop("appRef", None)
        _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".office-web-addin-worksheet-corrupt.tmp.xlsx")


def change_office_web_addin_controls(path: Path) -> Path:
    """Change private task-pane configuration and a snapshot relationship target."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    taskpanes_namespace = (
        "http://schemas.microsoft.com/office/webextensions/taskpanes/2010/11"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        taskpanes = ElementTree.fromstring(contents["xl/webextensions/taskpanes.xml"])
        taskpane = taskpanes.find(f"{{{taskpanes_namespace}}}taskpane")
        if taskpane is None:
            raise ValueError("Fixture does not contain a task pane")
        taskpane.set("visibility", "0")
        taskpane.set("width", "475")
        contents["xl/webextensions/taskpanes.xml"] = ElementTree.tostring(
            taskpanes,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships_name = "xl/webextensions/_rels/webextension1.xml.rels"
        relationships = ElementTree.fromstring(contents[relationships_name])
        snapshot_relationship = next(
            (
                relationship
                for relationship in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if relationship.get("Id") == "rIdFenceSnapshot"
            ),
            None,
        )
        if snapshot_relationship is None:
            raise ValueError("Fixture does not contain a snapshot relationship")
        snapshot_relationship.set("Target", "snapshots/private-candidate.png")
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/webextensions/snapshots/private-candidate.png"] = (
            b"private candidate add-in snapshot"
        )

    return _rewrite_archive(path, mutate, ".office-web-addin-change.tmp.xlsx")


def renumber_office_web_addin_relationships(path: Path) -> Path:
    """Rewrite task-pane relationship IDs while preserving semantic bindings."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    document_relationships = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    taskpanes_namespace = (
        "http://schemas.microsoft.com/office/webextensions/taskpanes/2010/11"
    )
    web_extension_namespace = (
        "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
    )
    replacements = {
        "rIdFenceWebTaskpanes": "rIdFenceRenumberedWebTaskpanes",
        "rIdFenceTaskpaneExtension": "rIdFenceRenumberedTaskpaneExtension",
        "rIdFenceSnapshot": "rIdFenceRenumberedSnapshot",
    }

    def replace_relationship_ids(root: ElementTree.Element) -> None:
        for relationship in root.findall(f"{{{package_relationships}}}Relationship"):
            if replacement := replacements.get(relationship.get("Id")):
                relationship.set("Id", replacement)

    def mutate(contents: dict[str, bytes]) -> None:
        for member in (
            "xl/_rels/workbook.xml.rels",
            "xl/webextensions/_rels/taskpanes.xml.rels",
            "xl/webextensions/_rels/webextension1.xml.rels",
        ):
            relationships = ElementTree.fromstring(contents[member])
            replace_relationship_ids(relationships)
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

        taskpanes = ElementTree.fromstring(contents["xl/webextensions/taskpanes.xml"])
        for reference in taskpanes.iter():
            if reference.tag not in {
                f"{{{taskpanes_namespace}}}webextension",
                f"{{{taskpanes_namespace}}}webextensionref",
            }:
                continue
            if replacement := replacements.get(
                reference.get(f"{{{document_relationships}}}id")
            ):
                reference.set(f"{{{document_relationships}}}id", replacement)
        contents["xl/webextensions/taskpanes.xml"] = ElementTree.tostring(
            taskpanes,
            encoding="utf-8",
            xml_declaration=True,
        )

        web_extension = ElementTree.fromstring(
            contents["xl/webextensions/webextension1.xml"]
        )
        for snapshot in web_extension.iter(f"{{{web_extension_namespace}}}snapshot"):
            if replacement := replacements.get(
                snapshot.get(f"{{{document_relationships}}}embed")
            ):
                snapshot.set(f"{{{document_relationships}}}embed", replacement)
        contents["xl/webextensions/webextension1.xml"] = ElementTree.tostring(
            web_extension,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".office-web-addin-renumber.tmp.xlsx")


def rewrite_office_web_addin_internal_target_spelling(path: Path) -> Path:
    """Use equivalent relative target spellings in the Office Web Add-in chain."""
    package_relationships = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_relationships_name = "xl/_rels/workbook.xml.rels"
        workbook_relationships = ElementTree.fromstring(
            contents[workbook_relationships_name]
        )
        taskpane_relationship = next(
            (
                relationship
                for relationship in workbook_relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if relationship.get("Id") == "rIdFenceWebTaskpanes"
            ),
            None,
        )
        if taskpane_relationship is None:
            raise ValueError("Fixture does not contain a workbook task-pane relationship")
        taskpane_relationship.set(
            "Target", "./webextensions/../webextensions/taskpanes.xml"
        )
        contents[workbook_relationships_name] = ElementTree.tostring(
            workbook_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        taskpane_relationships_name = "xl/webextensions/_rels/taskpanes.xml.rels"
        taskpane_relationships = ElementTree.fromstring(
            contents[taskpane_relationships_name]
        )
        extension_relationship = next(
            (
                relationship
                for relationship in taskpane_relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if relationship.get("Id") == "rIdFenceTaskpaneExtension"
            ),
            None,
        )
        if extension_relationship is None:
            raise ValueError("Fixture does not contain a task-pane extension relationship")
        extension_relationship.set("Target", "./webextension1.xml")
        contents[taskpane_relationships_name] = ElementTree.tostring(
            taskpane_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        extension_relationships_name = "xl/webextensions/_rels/webextension1.xml.rels"
        extension_relationships = ElementTree.fromstring(
            contents[extension_relationships_name]
        )
        snapshot_relationship = next(
            (
                relationship
                for relationship in extension_relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if relationship.get("Id") == "rIdFenceSnapshot"
            ),
            None,
        )
        if snapshot_relationship is None:
            raise ValueError("Fixture does not contain a snapshot relationship")
        snapshot_relationship.set(
            "Target", "./snapshots/../snapshots/private-baseline.png"
        )
        contents[extension_relationships_name] = ElementTree.tostring(
            extension_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".office-web-addin-target.tmp.xlsx")


def corrupt_office_web_addin_definition_root(path: Path) -> Path:
    """Replace the extension root with an unexpected element for coverage tests."""
    web_extension_namespace = (
        "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        web_extension = ElementTree.fromstring(
            contents["xl/webextensions/webextension1.xml"]
        )
        web_extension.tag = f"{{{web_extension_namespace}}}notWebExtension"
        contents["xl/webextensions/webextension1.xml"] = ElementTree.tostring(
            web_extension,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".office-web-addin-corrupt.tmp.xlsx")


def make_in_content_office_web_addin_model(path: Path) -> Path:
    """Create an Office Web Add-in hosted in a worksheet DrawingML frame.

    Modern Excel workbooks can host a Web Extension in the active branch of
    ``mc:AlternateContent`` and retain a native-picture preview as a fallback.
    This fixture deliberately has no workbook task-pane declaration, exercising
    the distinct in-content relationship graph that FormulaFence follows.
    """
    make_model(path)
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_main = _DRAWINGML_MAIN_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    markup_compatibility = _MARKUP_COMPATIBILITY_NS
    web_extension = (
        "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
    )
    web_extension_relationship = (
        "http://schemas.microsoft.com/office/2011/relationships/webextension"
    )
    baseline_png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9JdcAAAAAASUVORK5CYII="
    )

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def marker(
        parent: ElementTree.Element,
        name: str,
        *,
        column: int,
        row: int,
    ) -> None:
        point = ElementTree.SubElement(parent, f"{{{drawing}}}{name}")
        ElementTree.SubElement(point, f"{{{drawing}}}col").text = str(column)
        ElementTree.SubElement(point, f"{{{drawing}}}colOff").text = "0"
        ElementTree.SubElement(point, f"{{{drawing}}}row").text = str(row)
        ElementTree.SubElement(point, f"{{{drawing}}}rowOff").text = "0"

    def fallback_picture(parent: ElementTree.Element) -> None:
        picture = ElementTree.SubElement(parent, f"{{{drawing}}}pic")
        nonvisual = ElementTree.SubElement(picture, f"{{{drawing}}}nvPicPr")
        ElementTree.SubElement(
            nonvisual,
            f"{{{drawing}}}cNvPr",
            {
                "id": "7301",
                "name": "PRIVATE-IN-CONTENT-ADDIN-PREVIEW",
                "descr": "PRIVATE-IN-CONTENT-ADDIN-PREVIEW-DESCRIPTION",
            },
        )
        ElementTree.SubElement(nonvisual, f"{{{drawing}}}cNvPicPr")
        fill = ElementTree.SubElement(picture, f"{{{drawing}}}blipFill")
        ElementTree.SubElement(
            fill,
            f"{{{drawing_main}}}blip",
            {f"{{{document_relationships}}}embed": "rIdFenceInContentPreview"},
        )
        stretch = ElementTree.SubElement(fill, f"{{{drawing_main}}}stretch")
        ElementTree.SubElement(stretch, f"{{{drawing_main}}}fillRect")
        properties = ElementTree.SubElement(picture, f"{{{drawing}}}spPr")
        transform = ElementTree.SubElement(properties, f"{{{drawing_main}}}xfrm")
        ElementTree.SubElement(transform, f"{{{drawing_main}}}off", {"x": "0", "y": "0"})
        ElementTree.SubElement(
            transform,
            f"{{{drawing_main}}}ext",
            {"cx": "1828800", "cy": "1371600"},
        )
        geometry = ElementTree.SubElement(
            properties,
            f"{{{drawing_main}}}prstGeom",
            {"prst": "rect"},
        )
        ElementTree.SubElement(geometry, f"{{{drawing_main}}}avLst")

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member = "xl/drawings/drawing1.xml"
        worksheet_member = "xl/worksheets/sheet1.xml"
        preview_member = "xl/media/private-in-content-addin-preview.png"
        definition_member = "xl/webextensions/webextension1.xml"

        drawing_root = ElementTree.Element(f"{{{drawing}}}wsDr")
        anchor = ElementTree.SubElement(
            drawing_root,
            f"{{{drawing}}}twoCellAnchor",
            {"editAs": "oneCell"},
        )
        marker(anchor, "from", column=2, row=3)
        marker(anchor, "to", column=8, row=12)
        alternate_content = ElementTree.SubElement(
            anchor,
            f"{{{markup_compatibility}}}AlternateContent",
        )
        choice = ElementTree.SubElement(
            alternate_content,
            f"{{{markup_compatibility}}}Choice",
            {"Requires": "we"},
        )
        frame = ElementTree.SubElement(choice, f"{{{drawing}}}graphicFrame")
        nonvisual = ElementTree.SubElement(frame, f"{{{drawing}}}nvGraphicFramePr")
        ElementTree.SubElement(
            nonvisual,
            f"{{{drawing}}}cNvPr",
            {
                "id": "7300",
                "name": "PRIVATE-IN-CONTENT-ADDIN-FRAME",
                "descr": "PRIVATE-IN-CONTENT-ADDIN-FRAME-DESCRIPTION",
            },
        )
        ElementTree.SubElement(nonvisual, f"{{{drawing}}}cNvGraphicFramePr")
        transform = ElementTree.SubElement(frame, f"{{{drawing}}}xfrm")
        ElementTree.SubElement(transform, f"{{{drawing_main}}}off", {"x": "0", "y": "0"})
        ElementTree.SubElement(
            transform,
            f"{{{drawing_main}}}ext",
            {"cx": "1828800", "cy": "1371600"},
        )
        graphic = ElementTree.SubElement(frame, f"{{{drawing_main}}}graphic")
        graphic_data = ElementTree.SubElement(
            graphic,
            f"{{{drawing_main}}}graphicData",
            {"uri": web_extension},
        )
        ElementTree.SubElement(
            graphic_data,
            f"{{{web_extension}}}webextensionref",
            {f"{{{document_relationships}}}id": "rIdFenceInContentExtension"},
        )
        fallback = ElementTree.SubElement(
            alternate_content,
            f"{{{markup_compatibility}}}Fallback",
        )
        fallback_picture(fallback)
        ElementTree.SubElement(anchor, f"{{{drawing}}}clientData")
        contents[drawing_member] = serialize(drawing_root)

        drawing_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        for relationship_id, relationship_type, target in (
            (
                "rIdFenceInContentExtension",
                web_extension_relationship,
                "../webextensions/webextension1.xml",
            ),
            (
                "rIdFenceInContentPreview",
                f"{document_relationships}/image",
                "../media/private-in-content-addin-preview.png",
            ),
        ):
            ElementTree.SubElement(
                drawing_relationships,
                f"{{{package_relationships}}}Relationship",
                {"Id": relationship_id, "Type": relationship_type, "Target": target},
            )
        contents[_relationship_member(drawing_member)] = serialize(drawing_relationships)

        definition = ElementTree.Element(
            f"{{{web_extension}}}webextension",
            {"id": "{22222222-2222-2222-2222-222222222222}"},
        )
        ElementTree.SubElement(
            definition,
            f"{{{web_extension}}}reference",
            {
                "id": "PrivateInContentAddin",
                "version": "1.0.0.0",
                "store": "private-in-content-manifest.xml",
                "storeType": "Filesystem",
            },
        )
        properties = ElementTree.SubElement(definition, f"{{{web_extension}}}properties")
        ElementTree.SubElement(
            properties,
            f"{{{web_extension}}}property",
            {"name": "PrivateInContentBehavior", "value": "private in-content behavior"},
        )
        contents[definition_member] = serialize(definition)
        contents[preview_member] = baseline_png

        worksheet = _inputs_worksheet_root(contents)
        ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}drawing",
            {f"{{{document_relationships}}}id": "rIdFenceInContentDrawing"},
        )
        _save_inputs_worksheet(contents, worksheet)
        worksheet_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            worksheet_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceInContentDrawing",
                "Type": f"{document_relationships}/drawing",
                "Target": "../drawings/drawing1.xml",
            },
        )
        contents[_relationship_member(worksheet_member)] = serialize(worksheet_relationships)

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        if not any(
            element.get("Extension") == "png"
            for element in content_types.findall(default_tag)
        ):
            ElementTree.SubElement(
                content_types,
                default_tag,
                {"Extension": "png", "ContentType": "image/png"},
            )
        for part_name, content_type in (
            (
                "/xl/drawings/drawing1.xml",
                "application/vnd.openxmlformats-officedocument.drawing+xml",
            ),
            (
                "/xl/webextensions/webextension1.xml",
                "application/vnd.ms-office.webextension+xml",
            ),
        ):
            ElementTree.SubElement(
                content_types,
                override_tag,
                {"PartName": part_name, "ContentType": content_type},
            )
        contents["[Content_Types].xml"] = serialize(content_types)

    return _rewrite_archive(path, mutate, ".in-content-office-web-addin.tmp.xlsx")


def change_in_content_office_web_addin_anchor(path: Path) -> Path:
    """Move only an in-content add-in frame's private anchor."""
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member = "xl/drawings/drawing1.xml"
        root = ElementTree.fromstring(contents[drawing_member])
        end = next(root.iter(f"{{{drawing}}}to"), None)
        if end is None:
            raise ValueError("Fixture does not contain an in-content add-in anchor")
        row = end.find(f"{{{drawing}}}row")
        if row is None:
            raise ValueError("Fixture in-content add-in anchor has no ending row")
        row.text = "14"
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".in-content-office-web-addin-anchor.tmp.xlsx")


def change_in_content_office_web_addin_preview_payload(path: Path) -> Path:
    """Replace only the static fallback-preview image bytes."""
    candidate_png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAIAAACQd1PeAAAADUlEQVR42mP8z8DwHwAFgAI/ScLkYQAAAABJRU5ErkJggg=="
    )

    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/media/private-in-content-addin-preview.png"] = candidate_png

    return _rewrite_archive(path, mutate, ".in-content-office-web-addin-preview.tmp.xlsx")


def renumber_in_content_office_web_addin_identifiers(path: Path) -> Path:
    """Renumber volatile frame and drawing relationship IDs consistently."""
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    web_extension = (
        "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member = "xl/drawings/drawing1.xml"
        root = ElementTree.fromstring(contents[drawing_member])
        for index, properties in enumerate(
            root.iter(f"{{{drawing}}}cNvPr"),
            start=9500,
        ):
            properties.set("id", str(index))
        reference = next(root.iter(f"{{{web_extension}}}webextensionref"), None)
        if reference is None:
            raise ValueError("Fixture does not contain an in-content add-in reference")
        reference.set(
            f"{{{document_relationships}}}id",
            "rIdFenceRenumberedInContentExtension",
        )
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships_member = _relationship_member(drawing_member)
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            (
                current
                for current in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if current.get("Id") == "rIdFenceInContentExtension"
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain an in-content add-in relationship")
        relationship.set("Id", "rIdFenceRenumberedInContentExtension")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".in-content-office-web-addin-ids.tmp.xlsx")


def corrupt_in_content_office_web_addin_reference(path: Path) -> Path:
    """Remove the frame relationship ID so in-content parsing fails closed."""
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    web_extension = (
        "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member = "xl/drawings/drawing1.xml"
        root = ElementTree.fromstring(contents[drawing_member])
        reference = next(root.iter(f"{{{web_extension}}}webextensionref"), None)
        if reference is None:
            raise ValueError("Fixture does not contain an in-content add-in reference")
        reference.attrib.pop(f"{{{document_relationships}}}id", None)
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".in-content-office-web-addin-corrupt.tmp.xlsx")


def make_worksheet_embedded_control_model(
    path: Path,
    *,
    alternate_content: bool = False,
) -> Path:
    """Create harmless raw OOXML ActiveX, form-control, and OLE test material.

    The arbitrary binary bytes are never opened by an Office application. They
    exist only to prove FormulaFence's bounded, non-executing package scanner.
    """
    make_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    form_control_namespace = (
        "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
    )
    activex = "http://schemas.microsoft.com/office/2006/activeX"
    markup_compatibility = "http://schemas.openxmlformats.org/markup-compatibility/2006"
    control_relationship = f"{document_relationships}/control"
    control_properties_relationship = f"{document_relationships}/ctrlProp"
    ole_object_relationship = f"{document_relationships}/oleObject"
    image_relationship = f"{document_relationships}/image"
    activex_binary_relationship = (
        "http://schemas.microsoft.com/office/2006/relationships/activeXControlBinary"
    )

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(
        root: ElementTree.Element,
        part_name: str,
        content_type: str,
    ) -> None:
        ElementTree.SubElement(
            root,
            f"{{{content_types}}}Override",
            {"PartName": part_name, "ContentType": content_type},
        )

    def mutate(contents: dict[str, bytes]) -> None:
        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        add_override(
            types,
            "/xl/activeX/activeX1.xml",
            "application/vnd.ms-office.activeX+xml",
        )
        add_override(
            types,
            "/xl/ctrlProps/ctrlProp1.xml",
            "application/vnd.ms-excel.controlproperties+xml",
        )
        contents["[Content_Types].xml"] = serialize(types)

        worksheet = _inputs_worksheet_root(contents)
        controls = ElementTree.SubElement(worksheet, f"{{{spreadsheet}}}controls")
        active_x_control = ElementTree.Element(
            f"{{{spreadsheet}}}control",
            {
                "shapeId": "1025",
                "name": "PrivateBaselineCommandButton",
                f"{{{document_relationships}}}id": "rIdFenceActiveX",
            },
        )
        ElementTree.SubElement(
            active_x_control,
            f"{{{spreadsheet}}}controlPr",
            {
                "defaultSize": "0",
                f"{{{document_relationships}}}id": "rIdFenceControlPresentation",
            },
        )
        form_control_element = ElementTree.Element(
            f"{{{spreadsheet}}}control",
            {
                "shapeId": "1028",
                "name": "PrivateBaselineFormControl",
                f"{{{document_relationships}}}id": "rIdFenceControlProperties",
            },
        )
        ElementTree.SubElement(
            form_control_element,
            f"{{{spreadsheet}}}controlPr",
            {
                "macro": "PrivateBaselineControlMacro",
                "linkedCell": "Inputs!$B$2",
                "listFillRange": "Inputs!$B$2:$B$4",
            },
        )
        if alternate_content:
            alternate = ElementTree.SubElement(
                controls,
                f"{{{markup_compatibility}}}AlternateContent",
            )
            choice = ElementTree.SubElement(
                alternate,
                f"{{{markup_compatibility}}}Choice",
                {"Requires": "x14"},
            )
            choice.append(active_x_control)
            fallback = ElementTree.SubElement(
                alternate,
                f"{{{markup_compatibility}}}Fallback",
            )
            ElementTree.SubElement(
                fallback,
                f"{{{spreadsheet}}}control",
                {
                    "shapeId": "1025",
                    "name": "PrivateBaselineCommandButton",
                    f"{{{document_relationships}}}id": "rIdFenceActiveX",
                },
            )
            controls.append(form_control_element)
        else:
            controls.extend((active_x_control, form_control_element))

        ole_objects = ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}oleObjects",
        )
        embedded_ole = ElementTree.SubElement(
            ole_objects,
            f"{{{spreadsheet}}}oleObject",
            {
                "progId": "Private.Baseline.Embedded.Object",
                "shapeId": "1026",
                "autoLoad": "1",
                f"{{{document_relationships}}}id": "rIdFenceEmbeddedOle",
            },
        )
        ElementTree.SubElement(
            embedded_ole,
            f"{{{spreadsheet}}}objectPr",
            {f"{{{document_relationships}}}id": "rIdFenceOlePresentation"},
        )
        ElementTree.SubElement(
            ole_objects,
            f"{{{spreadsheet}}}oleObject",
            {
                "progId": "Private.Baseline.Linked.Object",
                "shapeId": "1027",
                "link": "private-baseline-link-name",
                "oleUpdate": "always",
                f"{{{document_relationships}}}id": "rIdFenceLinkedOle",
            },
        )
        _save_inputs_worksheet(contents, worksheet)

        worksheet_relationships = ElementTree.fromstring(
            contents.get(
                "xl/worksheets/_rels/sheet1.xml.rels",
                (
                    b'<?xml version="1.0" encoding="UTF-8"?>'
                    b'<Relationships xmlns="http://schemas.openxmlformats.org/'
                    b'package/2006/relationships"/>'
                ),
            )
        )
        relationship_tag = f"{{{package_relationships}}}Relationship"
        for relationship_id, relationship_type, target, target_mode in (
            (
                "rIdFenceActiveX",
                control_relationship,
                "../activeX/activeX1.xml",
                None,
            ),
            (
                "rIdFenceControlProperties",
                control_properties_relationship,
                "../ctrlProps/ctrlProp1.xml",
                None,
            ),
            (
                "rIdFenceControlPresentation",
                image_relationship,
                "../media/private-baseline-control.png",
                None,
            ),
            (
                "rIdFenceEmbeddedOle",
                ole_object_relationship,
                "../embeddings/private-baseline-ole.bin",
                None,
            ),
            (
                "rIdFenceLinkedOle",
                ole_object_relationship,
                "file:///private/baseline-linked-ole.bin",
                "External",
            ),
            (
                "rIdFenceOlePresentation",
                image_relationship,
                "../media/private-baseline-ole.png",
                None,
            ),
        ):
            attributes = {
                "Id": relationship_id,
                "Type": relationship_type,
                "Target": target,
            }
            if target_mode is not None:
                attributes["TargetMode"] = target_mode
            ElementTree.SubElement(worksheet_relationships, relationship_tag, attributes)
        contents["xl/worksheets/_rels/sheet1.xml.rels"] = serialize(
            worksheet_relationships
        )

        activex_root = ElementTree.Element(
            f"{{{activex}}}ocx",
            {
                "classid": "{11111111-1111-1111-1111-111111111111}",
                "license": "private-baseline-activex-license",
                "persistence": "persistStreamInit",
                f"{{{document_relationships}}}id": "rIdFenceActiveXBinary",
            },
        )
        contents["xl/activeX/activeX1.xml"] = serialize(activex_root)
        activex_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            activex_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceActiveXBinary",
                "Type": activex_binary_relationship,
                "Target": "activeX1.bin",
            },
        )
        contents["xl/activeX/_rels/activeX1.xml.rels"] = serialize(
            activex_relationships
        )

        form_control_root = ElementTree.Element(
            f"{{{form_control_namespace}}}formControlPr",
            {
                "objectType": "Drop",
                "fmlaGroup": "Inputs!$B$2",
                "fmlaLink": "Inputs!$B$3",
                "fmlaRange": "Inputs!$B$2:$B$4",
                "fmlaTxbx": "Inputs!$A$1",
            },
        )
        contents["xl/ctrlProps/ctrlProp1.xml"] = serialize(form_control_root)
        contents["xl/activeX/activeX1.bin"] = b"private baseline ActiveX binary payload"
        contents["xl/embeddings/private-baseline-ole.bin"] = (
            b"private baseline embedded OLE payload"
        )
        contents["xl/media/private-baseline-ole.png"] = b"private OLE presentation"
        contents["xl/media/private-baseline-control.png"] = b"private control presentation"

    return _rewrite_archive(path, mutate, ".worksheet-embedded-control.tmp.xlsx")


def change_worksheet_embedded_control_controls(path: Path) -> Path:
    """Change private worksheet-control, ActiveX, and OLE configuration material."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = _inputs_worksheet_root(contents)
        control_properties = next(
            (
                properties
                for properties in worksheet.iter(f"{{{spreadsheet}}}controlPr")
                if properties.get("macro") is not None
            ),
            None,
        )
        embedded_ole = worksheet.find(
            f".//{{{spreadsheet}}}oleObject"
        )
        if control_properties is None or embedded_ole is None:
            raise ValueError("Fixture does not contain worksheet embedded controls")
        control_properties.set("macro", "PrivateCandidateControlMacro")
        control_properties.set("linkedCell", "Inputs!$B$4")
        embedded_ole.set("autoLoad", "0")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

        activex_root = ElementTree.fromstring(contents["xl/activeX/activeX1.xml"])
        activex_root.set("classid", "{22222222-2222-2222-2222-222222222222}")
        activex_root.set("license", "private-candidate-activex-license")
        contents["xl/activeX/activeX1.xml"] = ElementTree.tostring(
            activex_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        form_control = ElementTree.fromstring(contents["xl/ctrlProps/ctrlProp1.xml"])
        form_control.set("fmlaRange", "Inputs!$B$3:$B$4")
        contents["xl/ctrlProps/ctrlProp1.xml"] = ElementTree.tostring(
            form_control,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships_name = "xl/worksheets/_rels/sheet1.xml.rels"
        relationships = ElementTree.fromstring(contents[relationships_name])
        relationship = next(
            (
                item
                for item in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if item.get("Id") == "rIdFenceEmbeddedOle"
            ),
            None,
        )
        if relationship is None:
            raise ValueError("Fixture does not contain an embedded OLE relationship")
        relationship.set("Target", "../embeddings/private-candidate-ole.bin")
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/embeddings/private-candidate-ole.bin"] = (
            b"private candidate embedded OLE payload"
        )

    return _rewrite_archive(path, mutate, ".worksheet-embedded-control-change.tmp.xlsx")


def change_worksheet_embedded_control_payload(path: Path) -> Path:
    """Change only an embedded OLE payload to exercise private byte fingerprinting."""

    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/embeddings/private-baseline-ole.bin"] = (
            b"private candidate changed embedded OLE payload"
        )

    return _rewrite_archive(path, mutate, ".worksheet-embedded-control-payload.tmp.xlsx")


def renumber_worksheet_embedded_control_relationships(path: Path) -> Path:
    """Rewrite only writer-chosen relationship IDs across the control chain."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    replacements = {
        "rIdFenceActiveX": "rIdFenceRenumberedActiveX",
        "rIdFenceControlProperties": "rIdFenceRenumberedControlProperties",
        "rIdFenceControlPresentation": "rIdFenceRenumberedControlPresentation",
        "rIdFenceEmbeddedOle": "rIdFenceRenumberedEmbeddedOle",
        "rIdFenceLinkedOle": "rIdFenceRenumberedLinkedOle",
        "rIdFenceOlePresentation": "rIdFenceRenumberedOlePresentation",
        "rIdFenceActiveXBinary": "rIdFenceRenumberedActiveXBinary",
    }

    def replace_relationship_ids(root: ElementTree.Element) -> None:
        for relationship in root.findall(f"{{{package_relationships}}}Relationship"):
            if replacement := replacements.get(relationship.get("Id")):
                relationship.set("Id", replacement)

    def replace_xml_ids(root: ElementTree.Element) -> None:
        relationship_attribute = f"{{{document_relationships}}}id"
        for element in root.iter():
            if replacement := replacements.get(element.get(relationship_attribute)):
                element.set(relationship_attribute, replacement)

    def mutate(contents: dict[str, bytes]) -> None:
        for member in (
            "xl/worksheets/_rels/sheet1.xml.rels",
            "xl/activeX/_rels/activeX1.xml.rels",
        ):
            relationships = ElementTree.fromstring(contents[member])
            replace_relationship_ids(relationships)
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )
        for member in ("xl/worksheets/sheet1.xml", "xl/activeX/activeX1.xml"):
            root = ElementTree.fromstring(contents[member])
            replace_xml_ids(root)
            contents[member] = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".worksheet-embedded-control-renumber.tmp.xlsx")


def rewrite_worksheet_embedded_control_internal_target_spelling(path: Path) -> Path:
    """Use equivalent relative spellings without changing control semantics."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def set_target(
        relationships: ElementTree.Element,
        relationship_id: str,
        target: str,
    ) -> None:
        relationship = next(
            (
                item
                for item in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if item.get("Id") == relationship_id
            ),
            None,
        )
        if relationship is None:
            raise ValueError(f"Fixture does not contain relationship {relationship_id}")
        relationship.set("Target", target)

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_relationships_name = "xl/worksheets/_rels/sheet1.xml.rels"
        worksheet_relationships = ElementTree.fromstring(
            contents[worksheet_relationships_name]
        )
        set_target(
            worksheet_relationships,
            "rIdFenceActiveX",
            "../activeX/./activeX1.xml",
        )
        set_target(
            worksheet_relationships,
            "rIdFenceControlProperties",
            "../ctrlProps/./ctrlProp1.xml",
        )
        set_target(
            worksheet_relationships,
            "rIdFenceControlPresentation",
            "../media/./private-baseline-control.png",
        )
        set_target(
            worksheet_relationships,
            "rIdFenceEmbeddedOle",
            "../embeddings/./private-baseline-ole.bin",
        )
        contents[worksheet_relationships_name] = ElementTree.tostring(
            worksheet_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        activex_relationships_name = "xl/activeX/_rels/activeX1.xml.rels"
        activex_relationships = ElementTree.fromstring(
            contents[activex_relationships_name]
        )
        set_target(
            activex_relationships,
            "rIdFenceActiveXBinary",
            "./activeX1.bin",
        )
        contents[activex_relationships_name] = ElementTree.tostring(
            activex_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-embedded-control-target.tmp.xlsx")


def corrupt_worksheet_embedded_control_activex_root(path: Path) -> Path:
    """Replace an ActiveX root with unexpected XML for fail-closed coverage tests."""
    activex = "http://schemas.microsoft.com/office/2006/activeX"

    def mutate(contents: dict[str, bytes]) -> None:
        root = ElementTree.fromstring(contents["xl/activeX/activeX1.xml"])
        root.tag = f"{{{activex}}}notOcx"
        contents["xl/activeX/activeX1.xml"] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-embedded-control-corrupt.tmp.xlsx")


def make_legacy_vml_control_model(path: Path) -> Path:
    """Create raw VML form-control material alongside a harmless comment note.

    The VML macros and bindings are inert XML strings. They exercise the scanner
    without being opened by Excel or executed during tests.
    """
    make_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    vml = "urn:schemas-microsoft-com:vml"
    vml_office = "urn:schemas-microsoft-com:office:office"
    vml_excel = "urn:schemas-microsoft-com:office:excel"
    vml_drawing_relationship = f"{document_relationships}/vmlDrawing"
    image_relationship = f"{document_relationships}/image"

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_vml_default(root: ElementTree.Element) -> None:
        default_tag = f"{{{content_types}}}Default"
        if any(item.get("Extension") == "vml" for item in root.findall(default_tag)):
            return
        ElementTree.SubElement(
            root,
            default_tag,
            {
                "Extension": "vml",
                "ContentType": "application/vnd.openxmlformats-officedocument.vmlDrawing",
            },
        )

    def formula_child(
        parent: ElementTree.Element,
        name: str,
        value: str,
    ) -> None:
        ElementTree.SubElement(parent, f"{{{vml_excel}}}{name}").text = value

    def mutate(contents: dict[str, bytes]) -> None:
        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        add_vml_default(types)
        contents["[Content_Types].xml"] = serialize(types)

        worksheet = _inputs_worksheet_root(contents)
        ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}legacyDrawing",
            {f"{{{document_relationships}}}id": "rIdFenceLegacyVmlDrawing"},
        )
        _save_inputs_worksheet(contents, worksheet)

        relationships = ElementTree.fromstring(
            contents.get(
                "xl/worksheets/_rels/sheet1.xml.rels",
                (
                    b'<?xml version="1.0" encoding="UTF-8"?>'
                    b'<Relationships xmlns="http://schemas.openxmlformats.org/'
                    b'package/2006/relationships"/>'
                ),
            )
        )
        relationship_tag = f"{{{package_relationships}}}Relationship"
        ElementTree.SubElement(
            relationships,
            relationship_tag,
            {
                "Id": "rIdFenceLegacyVmlDrawing",
                "Type": vml_drawing_relationship,
                "Target": "../drawings/vmlDrawing1.vml",
            },
        )
        contents["xl/worksheets/_rels/sheet1.xml.rels"] = serialize(relationships)

        drawing = ElementTree.Element("xml")
        note_shape = ElementTree.SubElement(
            drawing,
            f"{{{vml}}}shape",
            {"id": "_x0000_s1024", "type": "#_x0000_t202"},
        )
        ElementTree.SubElement(note_shape, f"{{{vml}}}textbox").text = (
            "Private legacy VML note text"
        )
        ElementTree.SubElement(
            note_shape,
            f"{{{vml_excel}}}ClientData",
            {"ObjectType": "Note"},
        )

        button_shape = ElementTree.SubElement(
            drawing,
            f"{{{vml}}}shape",
            {
                "id": "_x0000_s1025",
                "type": "#_x0000_t201",
                f"{{{vml_office}}}button": "t",
            },
        )
        ElementTree.SubElement(
            button_shape,
            f"{{{vml}}}imagedata",
            {f"{{{vml_office}}}relid": "rIdFenceLegacyVmlPresentation"},
        )
        ElementTree.SubElement(button_shape, f"{{{vml}}}textbox").text = (
            "Private legacy VML button caption"
        )
        button_data = ElementTree.SubElement(
            button_shape,
            f"{{{vml_excel}}}ClientData",
            {"ObjectType": "Button"},
        )
        formula_child(button_data, "FmlaMacro", "[0]!PrivateLegacyVmlMacro")

        dropdown_shape = ElementTree.SubElement(
            drawing,
            f"{{{vml}}}shape",
            {"id": "_x0000_s1026", "type": "#_x0000_t201"},
        )
        dropdown_data = ElementTree.SubElement(
            dropdown_shape,
            f"{{{vml_excel}}}ClientData",
            {"ObjectType": "Drop"},
        )
        formula_child(dropdown_data, "FmlaLink", "Inputs!$B$2")
        formula_child(dropdown_data, "FmlaRange", "Inputs!$B$2:$B$4")
        formula_child(dropdown_data, "FmlaTxbx", "Inputs!$A$1")

        camera_shape = ElementTree.SubElement(
            drawing,
            f"{{{vml}}}shape",
            {"id": "_x0000_s1027", "type": "#_x0000_t201"},
        )
        camera_data = ElementTree.SubElement(
            camera_shape,
            f"{{{vml_excel}}}ClientData",
            {"ObjectType": "Pict"},
        )
        formula_child(camera_data, "FmlaPict", "Inputs!$B$2:$B$4")

        group_shape = ElementTree.SubElement(
            drawing,
            f"{{{vml}}}shape",
            {"id": "_x0000_s1028", "type": "#_x0000_t201"},
        )
        group_data = ElementTree.SubElement(
            group_shape,
            f"{{{vml_excel}}}ClientData",
            {"ObjectType": "GBox"},
        )
        formula_child(group_data, "FmlaGroup", "Inputs!$B$3")
        contents["xl/drawings/vmlDrawing1.vml"] = serialize(drawing)

        drawing_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            drawing_relationships,
            relationship_tag,
            {
                "Id": "rIdFenceLegacyVmlPresentation",
                "Type": image_relationship,
                "Target": "../media/private-legacy-vml.png",
            },
        )
        contents["xl/drawings/_rels/vmlDrawing1.vml.rels"] = serialize(
            drawing_relationships
        )
        contents["xl/media/private-legacy-vml.png"] = b"private legacy VML presentation"

    return _rewrite_archive(path, mutate, ".legacy-vml-control.tmp.xlsx")


def make_legacy_vml_note_model(path: Path) -> Path:
    """Create a worksheet VML drawing that contains only an ordinary note."""
    make_legacy_vml_control_model(path)
    vml_excel = "urn:schemas-microsoft-com:office:excel"

    def mutate(contents: dict[str, bytes]) -> None:
        drawing = ElementTree.fromstring(contents["xl/drawings/vmlDrawing1.vml"])
        parent_by_child = {
            child: parent
            for parent in drawing.iter()
            for child in parent
        }
        for client_data in list(drawing.iter(f"{{{vml_excel}}}ClientData")):
            if client_data.get("ObjectType") == "Note":
                continue
            parent = parent_by_child.get(client_data)
            if parent is not None:
                parent_by_child.get(parent, drawing).remove(parent)
        contents["xl/drawings/vmlDrawing1.vml"] = ElementTree.tostring(
            drawing,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-vml-note-only.tmp.xlsx")


def change_legacy_vml_control_controls(path: Path) -> Path:
    """Change private legacy VML control definitions and presentation target."""
    vml_excel = "urn:schemas-microsoft-com:office:excel"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        drawing = ElementTree.fromstring(contents["xl/drawings/vmlDrawing1.vml"])
        button = next(
            (
                item
                for item in drawing.iter(f"{{{vml_excel}}}ClientData")
                if item.get("ObjectType") == "Button"
            ),
            None,
        )
        dropdown = next(
            (
                item
                for item in drawing.iter(f"{{{vml_excel}}}ClientData")
                if item.get("ObjectType") == "Drop"
            ),
            None,
        )
        camera = next(
            (
                item
                for item in drawing.iter(f"{{{vml_excel}}}ClientData")
                if item.get("ObjectType") == "Pict"
            ),
            None,
        )
        if button is None or dropdown is None or camera is None:
            raise ValueError("Fixture does not contain legacy VML controls")
        button.find(f"{{{vml_excel}}}FmlaMacro").text = "[0]!PrivateCandidateVmlMacro"
        dropdown.find(f"{{{vml_excel}}}FmlaLink").text = "Inputs!$B$4"
        dropdown.find(f"{{{vml_excel}}}FmlaRange").text = "Inputs!$B$3:$B$4"
        camera.find(f"{{{vml_excel}}}FmlaPict").text = "Inputs!$B$3:$B$4"
        contents["xl/drawings/vmlDrawing1.vml"] = ElementTree.tostring(
            drawing,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships_name = "xl/drawings/_rels/vmlDrawing1.vml.rels"
        relationships = ElementTree.fromstring(contents[relationships_name])
        relationship = relationships.find(
            f"{{{package_relationships}}}Relationship"
        )
        if relationship is None:
            raise ValueError("Fixture does not contain a legacy VML presentation relationship")
        relationship.set("Target", "../media/private-candidate-legacy-vml.png")
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/media/private-candidate-legacy-vml.png"] = (
            b"private candidate legacy VML presentation"
        )

    return _rewrite_archive(path, mutate, ".legacy-vml-control-change.tmp.xlsx")


def change_legacy_vml_note(path: Path) -> Path:
    """Change a comment note stored beside VML controls without touching a control."""
    vml = "urn:schemas-microsoft-com:vml"
    vml_excel = "urn:schemas-microsoft-com:office:excel"

    def mutate(contents: dict[str, bytes]) -> None:
        drawing = ElementTree.fromstring(contents["xl/drawings/vmlDrawing1.vml"])
        note = next(
            (
                item
                for item in drawing.iter(f"{{{vml_excel}}}ClientData")
                if item.get("ObjectType") == "Note"
            ),
            None,
        )
        if note is None:
            raise ValueError("Fixture does not contain a legacy VML note")
        parent_by_child = {
            child: parent
            for parent in drawing.iter()
            for child in parent
        }
        shape = parent_by_child.get(note)
        textbox = shape.find(f"{{{vml}}}textbox") if shape is not None else None
        if textbox is None:
            raise ValueError("Fixture does not contain a legacy VML note textbox")
        textbox.text = "Private candidate legacy VML note text"
        contents["xl/drawings/vmlDrawing1.vml"] = ElementTree.tostring(
            drawing,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-vml-note-change.tmp.xlsx")


def renumber_legacy_vml_control_relationships(path: Path) -> Path:
    """Rewrite arbitrary relationship identifiers across a legacy VML chain."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    vml = "urn:schemas-microsoft-com:vml"
    vml_office = "urn:schemas-microsoft-com:office:office"
    replacements = {
        "rIdFenceLegacyVmlDrawing": "rIdFenceRenumberedLegacyVmlDrawing",
        "rIdFenceLegacyVmlPresentation": "rIdFenceRenumberedLegacyVmlPresentation",
    }

    def replace_relationship_ids(root: ElementTree.Element) -> None:
        for relationship in root.findall(f"{{{package_relationships}}}Relationship"):
            if replacement := replacements.get(relationship.get("Id")):
                relationship.set("Id", replacement)

    def mutate(contents: dict[str, bytes]) -> None:
        for member in (
            "xl/worksheets/_rels/sheet1.xml.rels",
            "xl/drawings/_rels/vmlDrawing1.vml.rels",
        ):
            relationships = ElementTree.fromstring(contents[member])
            replace_relationship_ids(relationships)
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )
        worksheet = _inputs_worksheet_root(contents)
        legacy_drawing = worksheet.find(f"{{{_SPREADSHEETML_NS}}}legacyDrawing")
        if legacy_drawing is None:
            raise ValueError("Fixture does not contain a legacy VML worksheet binding")
        relationship_attribute = f"{{{document_relationships}}}id"
        legacy_drawing.set(
            relationship_attribute,
            replacements[legacy_drawing.get(relationship_attribute)],
        )
        _save_inputs_worksheet(contents, worksheet)

        drawing = ElementTree.fromstring(contents["xl/drawings/vmlDrawing1.vml"])
        relationship_attribute = f"{{{vml_office}}}relid"
        for image in drawing.iter(f"{{{vml}}}imagedata"):
            if replacement := replacements.get(image.get(relationship_attribute)):
                image.set(relationship_attribute, replacement)
        contents["xl/drawings/vmlDrawing1.vml"] = ElementTree.tostring(
            drawing,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-vml-control-renumber.tmp.xlsx")


def rewrite_legacy_vml_control_internal_target_spelling(path: Path) -> Path:
    """Use equivalent relative target spellings in a legacy VML control chain."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_relationships_name = "xl/worksheets/_rels/sheet1.xml.rels"
        worksheet_relationships = ElementTree.fromstring(
            contents[worksheet_relationships_name]
        )
        worksheet_relationship = worksheet_relationships.find(
            f"{{{package_relationships}}}Relationship"
        )
        if worksheet_relationship is None:
            raise ValueError("Fixture does not contain a legacy VML worksheet relationship")
        worksheet_relationship.set("Target", "../drawings/./vmlDrawing1.vml")
        contents[worksheet_relationships_name] = ElementTree.tostring(
            worksheet_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        drawing_relationships_name = "xl/drawings/_rels/vmlDrawing1.vml.rels"
        drawing_relationships = ElementTree.fromstring(contents[drawing_relationships_name])
        drawing_relationship = drawing_relationships.find(
            f"{{{package_relationships}}}Relationship"
        )
        if drawing_relationship is None:
            raise ValueError("Fixture does not contain a legacy VML presentation relationship")
        drawing_relationship.set("Target", "../media/./private-legacy-vml.png")
        contents[drawing_relationships_name] = ElementTree.tostring(
            drawing_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-vml-control-target.tmp.xlsx")


def corrupt_legacy_vml_control_root(path: Path) -> Path:
    """Replace a legacy VML root with unexpected XML for coverage tests."""

    def mutate(contents: dict[str, bytes]) -> None:
        drawing = ElementTree.fromstring(contents["xl/drawings/vmlDrawing1.vml"])
        drawing.tag = "notVmlXml"
        contents["xl/drawings/vmlDrawing1.vml"] = ElementTree.tostring(
            drawing,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-vml-control-corrupt.tmp.xlsx")


def _chart_fixture_part_names(contents: dict[str, bytes]) -> tuple[str, str, str]:
    """Return the chart fixture's drawing, chart, and overlay package members."""
    drawing_member = next(
        name
        for name in contents
        if name.startswith("xl/drawings/drawing")
        and name.endswith(".xml")
        and "/_rels/" not in name
    )
    chart_member = next(
        name
        for name in contents
        if name.startswith("xl/charts/") and name.endswith(".xml")
    )
    overlay_member = "xl/drawings/chartDrawing1.xml"
    if overlay_member not in contents:
        raise ValueError("Fixture does not contain a chart overlay part")
    return drawing_member, chart_member, overlay_member


def _relationship_member(member: str) -> str:
    """Return the OOXML relationship member beside one fixture part."""
    directory, filename = member.rsplit("/", maxsplit=1)
    return f"{directory}/_rels/{filename}.rels"


def _pivot_fixture_part_names(contents: dict[str, bytes]) -> tuple[str, str, str]:
    """Return the fixture's PivotTable, cache-definition, and record parts."""
    pivot_table_member = next(
        name
        for name in contents
        if name.startswith("xl/pivotTables/") and name.endswith(".xml")
    )
    cache_definition_member = next(
        name
        for name in contents
        if name.startswith("xl/pivotCache/pivotCacheDefinition") and name.endswith(".xml")
    )
    cache_records_member = next(
        name
        for name in contents
        if name.startswith("xl/pivotCache/pivotCacheRecords") and name.endswith(".xml")
    )
    return pivot_table_member, cache_definition_member, cache_records_member


def make_pivot_table_definition_model(path: Path) -> Path:
    """Create a harmless OOXML PivotTable graph with private redaction sentinels.

    openpyxl preserves PivotTable packages but deliberately does not create or
    edit them. This controlled fixture adds a small valid graph directly so the
    package-only FormulaFence inspection can be tested without Office software.
    """
    make_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    slicer = _OFFICE_2010_SPREADSHEET_NS

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        pivot_caches = ElementTree.SubElement(workbook, f"{{{spreadsheet}}}pivotCaches")
        ElementTree.SubElement(
            pivot_caches,
            f"{{{spreadsheet}}}pivotCache",
            {
                "cacheId": "7",
                f"{{{document_relationships}}}id": "rIdFenceWorkbookPivotCache",
            },
        )
        contents["xl/workbook.xml"] = serialize(workbook)

        workbook_relationships_name = _relationship_member("xl/workbook.xml")
        workbook_relationships = ElementTree.fromstring(
            contents[workbook_relationships_name]
        )
        ElementTree.SubElement(
            workbook_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceWorkbookPivotCache",
                "Type": f"{document_relationships}/pivotCacheDefinition",
                "Target": "pivotCache/pivotCacheDefinition1.xml",
            },
        )
        contents[workbook_relationships_name] = serialize(workbook_relationships)

        worksheet = _inputs_worksheet_root(contents)
        pivot_table_parts = ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}pivotTableParts",
            {"count": "1"},
        )
        ElementTree.SubElement(
            pivot_table_parts,
            f"{{{spreadsheet}}}pivotTablePart",
            {f"{{{document_relationships}}}id": "rIdFencePivotTable"},
        )
        _save_inputs_worksheet(contents, worksheet)

        worksheet_relationships_name = _relationship_member("xl/worksheets/sheet1.xml")
        worksheet_relationships = (
            ElementTree.fromstring(contents[worksheet_relationships_name])
            if worksheet_relationships_name in contents
            else ElementTree.Element(f"{{{package_relationships}}}Relationships")
        )
        ElementTree.SubElement(
            worksheet_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFencePivotTable",
                "Type": f"{document_relationships}/pivotTable",
                "Target": "../pivotTables/pivotTable1.xml",
            },
        )
        contents[worksheet_relationships_name] = serialize(worksheet_relationships)

        pivot_table_member = "xl/pivotTables/pivotTable1.xml"
        pivot_table = ElementTree.Element(
            f"{{{spreadsheet}}}pivotTableDefinition",
            {
                "name": "Private baseline pivot report",
                "cacheId": "7",
                "dataCaption": "Private baseline total",
            },
        )
        ElementTree.SubElement(
            pivot_table,
            f"{{{spreadsheet}}}location",
            {
                "ref": "A8:B10",
                "firstHeaderRow": "1",
                "firstDataRow": "2",
                "firstDataCol": "1",
            },
        )
        pivot_fields = ElementTree.SubElement(
            pivot_table,
            f"{{{spreadsheet}}}pivotFields",
            {"count": "1"},
        )
        pivot_field = ElementTree.SubElement(
            pivot_fields,
            f"{{{spreadsheet}}}pivotField",
            {"axis": "axisRow", "showAll": "0"},
        )
        items = ElementTree.SubElement(
            pivot_field,
            f"{{{spreadsheet}}}items",
            {"count": "2"},
        )
        ElementTree.SubElement(items, f"{{{spreadsheet}}}item", {"x": "0"})
        ElementTree.SubElement(items, f"{{{spreadsheet}}}item", {"x": "1"})
        row_fields = ElementTree.SubElement(
            pivot_table,
            f"{{{spreadsheet}}}rowFields",
            {"count": "1"},
        )
        ElementTree.SubElement(row_fields, f"{{{spreadsheet}}}field", {"x": "0"})
        row_items = ElementTree.SubElement(
            pivot_table,
            f"{{{spreadsheet}}}rowItems",
            {"count": "1"},
        )
        row_item = ElementTree.SubElement(row_items, f"{{{spreadsheet}}}i")
        ElementTree.SubElement(row_item, f"{{{spreadsheet}}}x", {"v": "0"})
        data_fields = ElementTree.SubElement(
            pivot_table,
            f"{{{spreadsheet}}}dataFields",
            {"count": "1"},
        )
        ElementTree.SubElement(
            data_fields,
            f"{{{spreadsheet}}}dataField",
            {"name": "Private baseline sum", "fld": "0", "subtotal": "sum"},
        )
        contents[pivot_table_member] = serialize(pivot_table)

        pivot_table_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            pivot_table_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFencePivotCache",
                "Type": f"{document_relationships}/pivotCacheDefinition",
                "Target": "../pivotCache/pivotCacheDefinition1.xml",
            },
        )
        contents[_relationship_member(pivot_table_member)] = serialize(
            pivot_table_relationships
        )

        cache_definition_member = "xl/pivotCache/pivotCacheDefinition1.xml"
        cache_definition = ElementTree.Element(
            f"{{{spreadsheet}}}pivotCacheDefinition",
            {
                "recordCount": "2",
                "refreshOnLoad": "0",
                "saveData": "1",
                f"{{{document_relationships}}}id": "rIdFencePivotRecords",
            },
        )
        cache_source = ElementTree.SubElement(
            cache_definition,
            f"{{{spreadsheet}}}cacheSource",
            {"type": "worksheet"},
        )
        ElementTree.SubElement(
            cache_source,
            f"{{{spreadsheet}}}worksheetSource",
            {"ref": "A2:B3", "sheet": "Private Pivot Source"},
        )
        cache_fields = ElementTree.SubElement(
            cache_definition,
            f"{{{spreadsheet}}}cacheFields",
            {"count": "1"},
        )
        cache_field = ElementTree.SubElement(
            cache_fields,
            f"{{{spreadsheet}}}cacheField",
            {"name": "Private baseline category", "numFmtId": "0"},
        )
        shared_items = ElementTree.SubElement(
            cache_field,
            f"{{{spreadsheet}}}sharedItems",
            {"count": "2"},
        )
        ElementTree.SubElement(
            shared_items,
            f"{{{spreadsheet}}}s",
            {"v": "Private baseline item one"},
        )
        ElementTree.SubElement(
            shared_items,
            f"{{{spreadsheet}}}s",
            {"v": "Private baseline item two"},
        )
        cache_extensions = ElementTree.SubElement(
            cache_definition,
            f"{{{spreadsheet}}}extLst",
        )
        cache_extension = ElementTree.SubElement(
            cache_extensions,
            f"{{{spreadsheet}}}ext",
            {"uri": "{725AE2AE-9491-48BE-B2B4-4EB974FC3084}"},
        )
        ElementTree.SubElement(
            cache_extension,
            f"{{{slicer}}}pivotCacheDefinition",
            {"pivotCacheId": "7"},
        )
        contents[cache_definition_member] = serialize(cache_definition)

        cache_definition_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            cache_definition_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFencePivotRecords",
                "Type": f"{document_relationships}/pivotCacheRecords",
                "Target": "pivotCacheRecords1.xml",
            },
        )
        contents[_relationship_member(cache_definition_member)] = serialize(
            cache_definition_relationships
        )

        cache_records_member = "xl/pivotCache/pivotCacheRecords1.xml"
        cache_records = ElementTree.Element(
            f"{{{spreadsheet}}}pivotCacheRecords",
            {"count": "2"},
        )
        for value in ("0", "1"):
            record = ElementTree.SubElement(cache_records, f"{{{spreadsheet}}}r")
            ElementTree.SubElement(record, f"{{{spreadsheet}}}x", {"v": value})
        contents[cache_records_member] = serialize(cache_records)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override_tag = f"{{{content_types}}}Override"
        for part_name, content_type in (
            (
                "/xl/pivotTables/pivotTable1.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml."
                "pivotTable+xml",
            ),
            (
                "/xl/pivotCache/pivotCacheDefinition1.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml."
                "pivotCacheDefinition+xml",
            ),
            (
                "/xl/pivotCache/pivotCacheRecords1.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml."
                "pivotCacheRecords+xml",
            ),
        ):
            ElementTree.SubElement(
                types,
                override_tag,
                {"PartName": part_name, "ContentType": content_type},
            )
        contents["[Content_Types].xml"] = serialize(types)

    return _rewrite_archive(path, mutate, ".pivot-table-definition.tmp.xlsx")


def change_pivot_table_definition_material(path: Path) -> Path:
    """Change private PivotTable layout, cache schema, items, and record values."""
    spreadsheet = _SPREADSHEETML_NS

    def mutate(contents: dict[str, bytes]) -> None:
        pivot_table_member, cache_definition_member, cache_records_member = (
            _pivot_fixture_part_names(contents)
        )
        pivot_table = ElementTree.fromstring(contents[pivot_table_member])
        pivot_table.set("dataCaption", "Private candidate total")
        location = pivot_table.find(f"{{{spreadsheet}}}location")
        if location is None:
            raise ValueError("Fixture does not contain a PivotTable location")
        location.set("ref", "D8:E10")
        contents[pivot_table_member] = ElementTree.tostring(
            pivot_table,
            encoding="utf-8",
            xml_declaration=True,
        )

        cache_definition = ElementTree.fromstring(contents[cache_definition_member])
        cache_field = next(cache_definition.iter(f"{{{spreadsheet}}}cacheField"))
        cache_field.set("name", "Private candidate category")
        shared_item = next(cache_definition.iter(f"{{{spreadsheet}}}s"))
        shared_item.set("v", "Private candidate item")
        contents[cache_definition_member] = ElementTree.tostring(
            cache_definition,
            encoding="utf-8",
            xml_declaration=True,
        )

        cache_records = ElementTree.fromstring(contents[cache_records_member])
        cache_value = next(cache_records.iter(f"{{{spreadsheet}}}x"))
        cache_value.set("v", "999")
        contents[cache_records_member] = ElementTree.tostring(
            cache_records,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".pivot-table-definition-change.tmp.xlsx")


def change_pivot_table_refresh_control(path: Path) -> Path:
    """Change only a refresh setting owned by the FF023 external-data guard."""
    spreadsheet = _SPREADSHEETML_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _pivot_table_member, cache_definition_member, _cache_records_member = (
            _pivot_fixture_part_names(contents)
        )
        cache_definition = ElementTree.fromstring(contents[cache_definition_member])
        if cache_definition.tag != f"{{{spreadsheet}}}pivotCacheDefinition":
            raise ValueError("Fixture does not contain a PivotTable cache definition")
        cache_definition.set("refreshOnLoad", "1")
        contents[cache_definition_member] = ElementTree.tostring(
            cache_definition,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".pivot-refresh-control.tmp.xlsx")


def rebind_pivot_table_cache_records(path: Path) -> Path:
    """Move the safe cache-record relationship to a distinct internal part."""
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _pivot_table_member, cache_definition_member, cache_records_member = (
            _pivot_fixture_part_names(contents)
        )
        cache_relationships_name = _relationship_member(cache_definition_member)
        cache_relationships = ElementTree.fromstring(contents[cache_relationships_name])
        cache_relationship = next(
            relationship
            for relationship in cache_relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if relationship.get("Id") == "rIdFencePivotRecords"
        )
        replacement_member = "xl/pivotCache/pivotCacheRecords2.xml"
        cache_relationship.set("Target", "pivotCacheRecords2.xml")
        contents[cache_relationships_name] = ElementTree.tostring(
            cache_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents[replacement_member] = contents.pop(cache_records_member)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override_tag = f"{{{content_types}}}Override"
        for override in types.findall(override_tag):
            if override.get("PartName") == f"/{cache_records_member}":
                override.set("PartName", f"/{replacement_member}")
        contents["[Content_Types].xml"] = ElementTree.tostring(
            types,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".pivot-cache-record-rebind.tmp.xlsx")


def externalize_pivot_table_cache_record_relationship(path: Path) -> Path:
    """Turn the direct cache-record target into an external redaction fixture."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _pivot_table_member, cache_definition_member, _cache_records_member = (
            _pivot_fixture_part_names(contents)
        )
        relationships_name = _relationship_member(cache_definition_member)
        relationships = ElementTree.fromstring(contents[relationships_name])
        relationship = next(
            relationship
            for relationship in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if relationship.get("Id") == "rIdFencePivotRecords"
        )
        relationship.set("Target", "https://example.invalid/private-pivot-cache-records")
        relationship.set("TargetMode", "External")
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".pivot-external-target.tmp.xlsx")


def renumber_pivot_table_relationships(path: Path) -> Path:
    """Rewrite PivotTable relationship IDs while retaining the same graph."""
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    replacements = {
        "rIdFenceWorkbookPivotCache": "rIdFenceRenumberedWorkbookPivotCache",
        "rIdFencePivotTable": "rIdFenceRenumberedPivotTable",
        "rIdFencePivotCache": "rIdFenceRenumberedPivotCache",
        "rIdFencePivotRecords": "rIdFenceRenumberedPivotRecords",
    }

    def replace_ids(root: ElementTree.Element) -> None:
        for relationship in root.findall(f"{{{package_relationships}}}Relationship"):
            if replacement := replacements.get(relationship.get("Id")):
                relationship.set("Id", replacement)

    def mutate(contents: dict[str, bytes]) -> None:
        pivot_table_member, cache_definition_member, _cache_records_member = (
            _pivot_fixture_part_names(contents)
        )
        for member in (
            _relationship_member("xl/workbook.xml"),
            _relationship_member("xl/worksheets/sheet1.xml"),
            _relationship_member(pivot_table_member),
            _relationship_member(cache_definition_member),
        ):
            relationships = ElementTree.fromstring(contents[member])
            replace_ids(relationships)
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        pivot_cache = next(workbook.iter(f"{{{spreadsheet}}}pivotCache"))
        pivot_cache.set(
            f"{{{document_relationships}}}id",
            replacements["rIdFenceWorkbookPivotCache"],
        )
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet = _inputs_worksheet_root(contents)
        pivot_table_part = next(worksheet.iter(f"{{{spreadsheet}}}pivotTablePart"))
        pivot_table_part.set(
            f"{{{document_relationships}}}id",
            replacements["rIdFencePivotTable"],
        )
        _save_inputs_worksheet(contents, worksheet)

        cache_definition = ElementTree.fromstring(contents[cache_definition_member])
        cache_definition.set(
            f"{{{document_relationships}}}id",
            replacements["rIdFencePivotRecords"],
        )
        contents[cache_definition_member] = ElementTree.tostring(
            cache_definition,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".pivot-relationship-renumber.tmp.xlsx")


def renumber_pivot_table_cache_id(path: Path) -> Path:
    """Renumber the workbook/cache-view identifier without changing its binding."""
    spreadsheet = _SPREADSHEETML_NS

    def mutate(contents: dict[str, bytes]) -> None:
        pivot_table_member, _cache_definition_member, _cache_records_member = (
            _pivot_fixture_part_names(contents)
        )
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        pivot_cache = next(workbook.iter(f"{{{spreadsheet}}}pivotCache"))
        pivot_cache.set("cacheId", "101")
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

        pivot_table = ElementTree.fromstring(contents[pivot_table_member])
        pivot_table.set("cacheId", "101")
        contents[pivot_table_member] = ElementTree.tostring(
            pivot_table,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".pivot-cache-id-renumber.tmp.xlsx")


def rewrite_pivot_table_internal_target_spelling(path: Path) -> Path:
    """Use equivalent relative target spellings across the PivotTable graph."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        pivot_table_member, cache_definition_member, _cache_records_member = (
            _pivot_fixture_part_names(contents)
        )
        targets = {
            _relationship_member("xl/workbook.xml"): "./pivotCache/./pivotCacheDefinition1.xml",
            _relationship_member("xl/worksheets/sheet1.xml"): "../pivotTables/./pivotTable1.xml",
            _relationship_member(pivot_table_member): "../pivotCache/./pivotCacheDefinition1.xml",
            _relationship_member(cache_definition_member): "./pivotCacheRecords1.xml",
        }
        for member, target in targets.items():
            relationships = ElementTree.fromstring(contents[member])
            relationship = next(
                relationship
                for relationship in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if relationship.get("Type", "").startswith(document_relationships)
                and relationship.get("TargetMode", "Internal").casefold() == "internal"
                and "pivot" in relationship.get("Type", "").casefold()
            )
            relationship.set("Target", target)
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".pivot-target-spelling.tmp.xlsx")


def corrupt_pivot_table_definition_root(path: Path) -> Path:
    """Replace a PivotTable root with unexpected XML for coverage tests."""
    def mutate(contents: dict[str, bytes]) -> None:
        pivot_table_member, _cache_definition_member, _cache_records_member = (
            _pivot_fixture_part_names(contents)
        )
        pivot_table = ElementTree.fromstring(contents[pivot_table_member])
        pivot_table.tag = "notPivotTableDefinition"
        contents[pivot_table_member] = ElementTree.tostring(
            pivot_table,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".pivot-definition-corrupt.tmp.xlsx")


def _slicer_timeline_fixture_part_names(contents: dict[str, bytes]) -> tuple[str, str, str]:
    """Return controlled slicer-cache and Timeline-cache package members."""
    pivot_slicer_member = "xl/slicerCaches/slicerCache1.xml"
    table_slicer_member = "xl/slicerCaches/slicerCache2.xml"
    timeline_member = "xl/timelineCaches/timelineCache1.xml"
    if not {
        pivot_slicer_member,
        table_slicer_member,
        timeline_member,
    } <= contents.keys():
        raise ValueError("Fixture does not contain its slicer/Timeline cache parts")
    return pivot_slicer_member, table_slicer_member, timeline_member


def make_slicer_timeline_cache_model(path: Path) -> Path:
    """Create harmless slicer and Timeline cache material with private sentinels.

    The package is intentionally constructed at the OOXML layer because the
    workbook reader preserves neither cache definition. It is used only for
    static comparison and redaction tests, never opened in Office software.
    """
    make_pivot_table_definition_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    slicer = _OFFICE_2010_SPREADSHEET_NS
    timeline = _OFFICE_2013_SPREADSHEET_NS
    revision10 = _OFFICE_2016_REVISION10_NS

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        extensions = ElementTree.SubElement(workbook, f"{{{spreadsheet}}}extLst")
        slicer_extension = ElementTree.SubElement(
            extensions,
            f"{{{spreadsheet}}}ext",
            {"uri": "{BBE1A952-AA13-448E-AADC-164F8A28A991}"},
        )
        slicer_caches = ElementTree.SubElement(
            slicer_extension,
            f"{{{slicer}}}slicerCaches",
        )
        ElementTree.SubElement(
            slicer_caches,
            f"{{{slicer}}}slicerCache",
            {f"{{{document_relationships}}}id": "rIdFencePivotSlicer"},
        )
        table_slicer_extension = ElementTree.SubElement(
            extensions,
            f"{{{spreadsheet}}}ext",
            {"uri": "{46BE6895-7355-4A93-B00E-2C351335B9C9}"},
        )
        table_slicer_caches = ElementTree.SubElement(
            table_slicer_extension,
            f"{{{timeline}}}slicerCaches",
        )
        ElementTree.SubElement(
            table_slicer_caches,
            f"{{{slicer}}}slicerCache",
            {f"{{{document_relationships}}}id": "rIdFenceTableSlicer"},
        )
        timeline_extension = ElementTree.SubElement(
            extensions,
            f"{{{spreadsheet}}}ext",
            {"uri": "{7E03D99C-DC04-49D9-9315-930204A7B6E9}"},
        )
        timeline_refs = ElementTree.SubElement(
            timeline_extension,
            f"{{{timeline}}}timelineCacheRefs",
        )
        ElementTree.SubElement(
            timeline_refs,
            f"{{{timeline}}}timelineCacheRef",
            {f"{{{document_relationships}}}id": "rIdFenceTimeline"},
        )
        contents["xl/workbook.xml"] = serialize(workbook)

        workbook_relationships_name = _relationship_member("xl/workbook.xml")
        workbook_relationships = ElementTree.fromstring(
            contents[workbook_relationships_name]
        )
        for relationship_id, relationship_type, target in (
            (
                "rIdFencePivotSlicer",
                "http://schemas.microsoft.com/office/2007/relationships/slicerCache",
                "slicerCaches/slicerCache1.xml",
            ),
            (
                "rIdFenceTableSlicer",
                "http://schemas.microsoft.com/office/2007/relationships/slicerCache",
                "slicerCaches/slicerCache2.xml",
            ),
            (
                "rIdFenceTimeline",
                "http://schemas.microsoft.com/office/2010/relationships/TimelineCache",
                "timelineCaches/timelineCache1.xml",
            ),
        ):
            ElementTree.SubElement(
                workbook_relationships,
                f"{{{package_relationships}}}Relationship",
                {
                    "Id": relationship_id,
                    "Type": relationship_type,
                    "Target": target,
                },
            )
        contents[workbook_relationships_name] = serialize(workbook_relationships)

        pivot_slicer_member = "xl/slicerCaches/slicerCache1.xml"
        pivot_slicer = ElementTree.Element(
            f"{{{slicer}}}slicerCacheDefinition",
            {
                "name": "Private baseline revenue slicer",
                "sourceName": "Private baseline business unit",
            },
        )
        pivot_slicer_data = ElementTree.SubElement(
            pivot_slicer,
            f"{{{slicer}}}data",
        )
        tabular_slicer = ElementTree.SubElement(
            pivot_slicer_data,
            f"{{{slicer}}}tabular",
            {"pivotCacheId": "7"},
        )
        pivot_slicer_items = ElementTree.SubElement(
            tabular_slicer,
            f"{{{slicer}}}items",
            {"count": "2"},
        )
        ElementTree.SubElement(
            pivot_slicer_items,
            f"{{{slicer}}}i",
            {"x": "0", "s": "1"},
        )
        ElementTree.SubElement(
            pivot_slicer_items,
            f"{{{slicer}}}i",
            {"x": "1", "s": "0"},
        )
        pivot_slicer_tables = ElementTree.SubElement(
            pivot_slicer,
            f"{{{slicer}}}pivotTables",
            {"count": "1"},
        )
        ElementTree.SubElement(
            pivot_slicer_tables,
            f"{{{slicer}}}pivotTable",
            {"tabId": "1", "name": "Private baseline pivot report"},
        )
        contents[pivot_slicer_member] = serialize(pivot_slicer)

        table_slicer_member = "xl/slicerCaches/slicerCache2.xml"
        table_slicer = ElementTree.Element(
            f"{{{slicer}}}slicerCacheDefinition",
            {
                "name": "Private baseline table slicer",
                "sourceName": "Private baseline table field",
            },
        )
        table_slicer_extensions = ElementTree.SubElement(
            table_slicer,
            f"{{{slicer}}}extLst",
        )
        table_slicer_extension = ElementTree.SubElement(
            table_slicer_extensions,
            f"{{{slicer}}}ext",
            {"uri": "{2F2917AC-EB37-4324-AD4E-5DD8C200BD13}"},
        )
        ElementTree.SubElement(
            table_slicer_extension,
            f"{{{timeline}}}tableSlicerCache",
            {"tableId": "41", "column": "1"},
        )
        contents[table_slicer_member] = serialize(table_slicer)

        timeline_member = "xl/timelineCaches/timelineCache1.xml"
        timeline_cache = ElementTree.Element(
            f"{{{timeline}}}timelineCacheDefinition",
            {
                "name": "Private baseline sales timeline",
                "sourceName": "Private baseline transaction date",
                f"{{{revision10}}}uid": "{D75EB4A3-A42F-45D9-8E99-7C310D8EF203}",
            },
        )
        timeline_pivot_tables = ElementTree.SubElement(
            timeline_cache,
            f"{{{timeline}}}pivotTables",
            {"count": "1"},
        )
        ElementTree.SubElement(
            timeline_pivot_tables,
            f"{{{timeline}}}pivotTable",
            {"name": "Private baseline pivot report"},
        )
        ElementTree.SubElement(
            timeline_cache,
            f"{{{timeline}}}state",
            {
                "minimalRefreshVersion": "6",
                "lastRefreshVersion": "6",
                "pivotCacheId": "7",
                "filterType": "dateBetween",
                "singleRangeFilterState": "1",
            },
        )
        timeline_state = timeline_cache.find(f"{{{timeline}}}state")
        if timeline_state is None:
            raise ValueError("Fixture did not create a Timeline state")
        ElementTree.SubElement(
            timeline_state,
            f"{{{timeline}}}selection",
            {
                "startDate": "2024-01-01T00:00:00Z",
                "endDate": "2024-03-31T00:00:00Z",
            },
        )
        ElementTree.SubElement(
            timeline_state,
            f"{{{timeline}}}bounds",
            {
                "startDate": "2023-01-01T00:00:00Z",
                "endDate": "2024-12-31T00:00:00Z",
            },
        )
        contents[timeline_member] = serialize(timeline_cache)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override_tag = f"{{{content_types}}}Override"
        for part_name, content_type in (
            ("/xl/slicerCaches/slicerCache1.xml", "application/vnd.ms-excel.slicerCache+xml"),
            ("/xl/slicerCaches/slicerCache2.xml", "application/vnd.ms-excel.slicerCache+xml"),
            ("/xl/timelineCaches/timelineCache1.xml", "application/vnd.ms-excel.TimelineCache+xml"),
        ):
            ElementTree.SubElement(
                types,
                override_tag,
                {"PartName": part_name, "ContentType": content_type},
            )
        contents["[Content_Types].xml"] = serialize(types)

    return _rewrite_archive(path, mutate, ".slicer-timeline-cache.tmp.xlsx")


def change_slicer_timeline_filter_material(path: Path) -> Path:
    """Change private slicer selection and Timeline filter state material."""
    slicer = _OFFICE_2010_SPREADSHEET_NS
    timeline = _OFFICE_2013_SPREADSHEET_NS

    def mutate(contents: dict[str, bytes]) -> None:
        pivot_slicer_member, _table_slicer_member, timeline_member = (
            _slicer_timeline_fixture_part_names(contents)
        )
        pivot_slicer = ElementTree.fromstring(contents[pivot_slicer_member])
        selected_item = next(
            item
            for item in pivot_slicer.iter(f"{{{slicer}}}i")
            if item.get("s") == "1"
        )
        selected_item.set("s", "0")
        contents[pivot_slicer_member] = ElementTree.tostring(
            pivot_slicer,
            encoding="utf-8",
            xml_declaration=True,
        )

        timeline_cache = ElementTree.fromstring(contents[timeline_member])
        state = timeline_cache.find(f"{{{timeline}}}state")
        if state is None:
            raise ValueError("Fixture does not contain a Timeline state")
        state.set("filterType", "dateEqual")
        selection = state.find(f"{{{timeline}}}selection")
        if selection is None:
            raise ValueError("Fixture does not contain a Timeline selection")
        selection.set("startDate", "2024-05-14T00:00:00Z")
        selection.set("endDate", "2024-05-14T00:00:00Z")
        contents[timeline_member] = ElementTree.tostring(
            timeline_cache,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-filter-change.tmp.xlsx")


def set_slicer_timeline_equivalent_defaults(path: Path) -> Path:
    """Spell out optional slicer defaults and regenerate the Timeline UID."""
    slicer = _OFFICE_2010_SPREADSHEET_NS
    timeline = _OFFICE_2013_SPREADSHEET_NS
    revision10 = _OFFICE_2016_REVISION10_NS

    def mutate(contents: dict[str, bytes]) -> None:
        pivot_slicer_member, table_slicer_member, timeline_member = (
            _slicer_timeline_fixture_part_names(contents)
        )
        pivot_slicer = ElementTree.fromstring(contents[pivot_slicer_member])
        tabular = next(pivot_slicer.iter(f"{{{slicer}}}tabular"))
        tabular.set("sortOrder", "ascending")
        tabular.set("customListSort", "1")
        tabular.set("showMissing", "true")
        tabular.set("crossFilter", "showItemsWithDataAtTop")
        for item in pivot_slicer.iter(f"{{{slicer}}}i"):
            if item.get("s") == "0":
                item.set("s", "false")
            item.set("nd", "0")
        contents[pivot_slicer_member] = ElementTree.tostring(
            pivot_slicer,
            encoding="utf-8",
            xml_declaration=True,
        )

        table_slicer = ElementTree.fromstring(contents[table_slicer_member])
        table_cache = next(table_slicer.iter(f"{{{timeline}}}tableSlicerCache"))
        table_cache.set("sortOrder", "ascending")
        table_cache.set("customListSort", "true")
        table_cache.set("crossFilter", "showItemsWithDataAtTop")
        contents[table_slicer_member] = ElementTree.tostring(
            table_slicer,
            encoding="utf-8",
            xml_declaration=True,
        )

        timeline_cache = ElementTree.fromstring(contents[timeline_member])
        timeline_cache.set(
            f"{{{revision10}}}uid", "{E5E42E16-64AA-4734-B3B0-E0B82821C9DD}"
        )
        contents[timeline_member] = ElementTree.tostring(
            timeline_cache,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-defaults.tmp.xlsx")


def rebind_slicer_timeline_cache(path: Path) -> Path:
    """Move one slicer cache to a distinct package target for binding tests."""
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        pivot_slicer_member, _table_slicer_member, _timeline_member = (
            _slicer_timeline_fixture_part_names(contents)
        )
        replacement_member = "xl/slicerCaches/slicerCache3.xml"
        relationships_name = _relationship_member("xl/workbook.xml")
        relationships = ElementTree.fromstring(contents[relationships_name])
        relationship = next(
            current
            for current in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if current.get("Id") == "rIdFencePivotSlicer"
        )
        relationship.set("Target", "slicerCaches/slicerCache3.xml")
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents[replacement_member] = contents.pop(pivot_slicer_member)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override_tag = f"{{{content_types}}}Override"
        for override in types.findall(override_tag):
            if override.get("PartName") == f"/{pivot_slicer_member}":
                override.set("PartName", f"/{replacement_member}")
        contents["[Content_Types].xml"] = ElementTree.tostring(
            types,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-rebind.tmp.xlsx")


def renumber_slicer_timeline_relationships(path: Path) -> Path:
    """Rewrite workbook relationship IDs while retaining filter-cache bindings."""
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    slicer = _OFFICE_2010_SPREADSHEET_NS
    timeline = _OFFICE_2013_SPREADSHEET_NS
    replacements = {
        "rIdFencePivotSlicer": "rIdFenceRenumberedPivotSlicer",
        "rIdFenceTableSlicer": "rIdFenceRenumberedTableSlicer",
        "rIdFenceTimeline": "rIdFenceRenumberedTimeline",
    }

    def mutate(contents: dict[str, bytes]) -> None:
        relationships_name = _relationship_member("xl/workbook.xml")
        relationships = ElementTree.fromstring(contents[relationships_name])
        for relationship in relationships.findall(
            f"{{{package_relationships}}}Relationship"
        ):
            if replacement := replacements.get(relationship.get("Id")):
                relationship.set("Id", replacement)
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        for element in workbook.iter():
            if element.tag in {
                f"{{{slicer}}}slicerCache",
                f"{{{timeline}}}slicerCache",
                f"{{{timeline}}}timelineCacheRef",
            }:
                relationship_id = element.get(f"{{{document_relationships}}}id")
                if replacement := replacements.get(relationship_id):
                    element.set(f"{{{document_relationships}}}id", replacement)
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-relationship-renumber.tmp.xlsx")


def use_slicer_timeline_2011_relationship_type(path: Path) -> Path:
    """Use a widely emitted Timeline cache relationship compatibility form."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships_name = _relationship_member("xl/workbook.xml")
        relationships = ElementTree.fromstring(contents[relationships_name])
        relationship = next(
            current
            for current in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if current.get("Id") == "rIdFenceTimeline"
        )
        relationship.set(
            "Type",
            "http://schemas.microsoft.com/office/2011/relationships/timelineCache",
        )
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-2011-relationship.tmp.xlsx")


def renumber_slicer_timeline_pivot_cache_id(path: Path) -> Path:
    """Renumber the common Pivot cache ID without changing its target binding."""
    slicer = _OFFICE_2010_SPREADSHEET_NS
    timeline = _OFFICE_2013_SPREADSHEET_NS

    def mutate(contents: dict[str, bytes]) -> None:
        pivot_slicer_member, _table_slicer_member, timeline_member = (
            _slicer_timeline_fixture_part_names(contents)
        )
        _pivot_table_member, cache_definition_member, _cache_records_member = (
            _pivot_fixture_part_names(contents)
        )

        cache_definition = ElementTree.fromstring(contents[cache_definition_member])
        extension = next(
            cache_definition.iter(f"{{{slicer}}}pivotCacheDefinition")
        )
        extension.set("pivotCacheId", "101")
        contents[cache_definition_member] = ElementTree.tostring(
            cache_definition,
            encoding="utf-8",
            xml_declaration=True,
        )

        pivot_slicer = ElementTree.fromstring(contents[pivot_slicer_member])
        tabular = next(pivot_slicer.iter(f"{{{slicer}}}tabular"))
        tabular.set("pivotCacheId", "101")
        contents[pivot_slicer_member] = ElementTree.tostring(
            pivot_slicer,
            encoding="utf-8",
            xml_declaration=True,
        )

        timeline_cache = ElementTree.fromstring(contents[timeline_member])
        state = timeline_cache.find(f"{{{timeline}}}state")
        if state is None:
            raise ValueError("Fixture does not contain a Timeline state")
        state.set("pivotCacheId", "101")
        contents[timeline_member] = ElementTree.tostring(
            timeline_cache,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-cache-id-renumber.tmp.xlsx")


def break_slicer_timeline_pivot_cache_binding(path: Path) -> Path:
    """Leave slicer/Timeline IDs stale after changing their PivotCache extension ID."""
    slicer = _OFFICE_2010_SPREADSHEET_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _pivot_table_member, cache_definition_member, _cache_records_member = (
            _pivot_fixture_part_names(contents)
        )
        cache_definition = ElementTree.fromstring(contents[cache_definition_member])
        extension = next(
            cache_definition.iter(f"{{{slicer}}}pivotCacheDefinition")
        )
        extension.set("pivotCacheId", "303")
        contents[cache_definition_member] = ElementTree.tostring(
            cache_definition,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-broken-pivot-cache-id.tmp.xlsx")


def rewrite_slicer_timeline_internal_target_spelling(path: Path) -> Path:
    """Use equivalent target spellings for workbook filter-cache relationships."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships_name = _relationship_member("xl/workbook.xml")
        relationships = ElementTree.fromstring(contents[relationships_name])
        for relationship in relationships.findall(
            f"{{{package_relationships}}}Relationship"
        ):
            if relationship.get("Id") == "rIdFencePivotSlicer":
                relationship.set("Target", "./slicerCaches/./slicerCache1.xml")
            elif relationship.get("Id") == "rIdFenceTableSlicer":
                relationship.set("Target", "./slicerCaches/./slicerCache2.xml")
            elif relationship.get("Id") == "rIdFenceTimeline":
                relationship.set("Target", "./timelineCaches/./timelineCache1.xml")
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-target-spelling.tmp.xlsx")


def externalize_slicer_timeline_cache_relationship(path: Path) -> Path:
    """Turn one slicer cache target external without exposing its address."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships_name = _relationship_member("xl/workbook.xml")
        relationships = ElementTree.fromstring(contents[relationships_name])
        relationship = next(
            current
            for current in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if current.get("Id") == "rIdFencePivotSlicer"
        )
        relationship.set("Target", "https://example.invalid/private-slicer-cache")
        relationship.set("TargetMode", "External")
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-external-target.tmp.xlsx")


def corrupt_slicer_timeline_cache_root(path: Path) -> Path:
    """Replace one cache root with unexpected XML for coverage tests."""
    def mutate(contents: dict[str, bytes]) -> None:
        pivot_slicer_member, _table_slicer_member, _timeline_member = (
            _slicer_timeline_fixture_part_names(contents)
        )
        root = ElementTree.fromstring(contents[pivot_slicer_member])
        root.tag = "notSlicerCacheDefinition"
        contents[pivot_slicer_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".slicer-timeline-corrupt.tmp.xlsx")


def make_chart_definition_model(path: Path) -> Path:
    """Create a harmless chart package with cache and overlay test material.

    The fixture is only used to prove FormulaFence's static package inspection.
    It is never opened in an Office application and all private strings and
    payloads are deliberate redaction sentinels.
    """
    make_model(path)

    def add_chart(workbook: Workbook) -> None:
        inputs = workbook["Inputs"]
        inputs["A2"] = "January"
        inputs["A3"] = "February"
        inputs["A4"] = "March"
        chart = BarChart()
        chart.title = "Private baseline chart title"
        chart.add_data(
            Reference(inputs, min_col=2, min_row=2, max_row=4),
            titles_from_data=False,
        )
        chart.set_categories(Reference(inputs, min_col=1, min_row=2, max_row=4))
        inputs.add_chart(chart, "D2")

    rewrite(path, add_chart)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    chart_namespace = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    chart_drawing_namespace = (
        "http://schemas.openxmlformats.org/drawingml/2006/chartDrawing"
    )
    drawing_namespace = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member = next(
            name
            for name in contents
            if name.startswith("xl/drawings/drawing")
            and name.endswith(".xml")
            and "/_rels/" not in name
        )
        chart_member = next(
            name
            for name in contents
            if name.startswith("xl/charts/") and name.endswith(".xml")
        )
        chart = ElementTree.fromstring(contents[chart_member])
        series = next(chart.iter(f"{{{chart_namespace}}}ser"))
        numeric_reference = next(series.iter(f"{{{chart_namespace}}}numRef"))
        numeric_cache = ElementTree.SubElement(
            numeric_reference,
            f"{{{chart_namespace}}}numCache",
        )
        ElementTree.SubElement(
            numeric_cache,
            f"{{{chart_namespace}}}ptCount",
            {"val": "3"},
        )
        for index, value in enumerate(("100", "150", "175")):
            point = ElementTree.SubElement(
                numeric_cache,
                f"{{{chart_namespace}}}pt",
                {"idx": str(index)},
            )
            ElementTree.SubElement(point, f"{{{chart_namespace}}}v").text = value

        category = ElementTree.Element(f"{{{chart_namespace}}}cat")
        string_reference = ElementTree.SubElement(
            category,
            f"{{{chart_namespace}}}strRef",
        )
        ElementTree.SubElement(
            string_reference,
            f"{{{chart_namespace}}}f",
        ).text = "Inputs!$A$2:$A$4"
        string_cache = ElementTree.SubElement(
            string_reference,
            f"{{{chart_namespace}}}strCache",
        )
        ElementTree.SubElement(
            string_cache,
            f"{{{chart_namespace}}}ptCount",
            {"val": "3"},
        )
        for index, value in enumerate(("January", "February", "March")):
            point = ElementTree.SubElement(
                string_cache,
                f"{{{chart_namespace}}}pt",
                {"idx": str(index)},
            )
            ElementTree.SubElement(point, f"{{{chart_namespace}}}v").text = value
        value = next(series.iter(f"{{{chart_namespace}}}val"))
        series.insert(list(series).index(value), category)
        ElementTree.SubElement(
            chart,
            f"{{{chart_namespace}}}userShapes",
            {f"{{{document_relationships}}}id": "rIdFenceChartOverlay"},
        )
        contents[chart_member] = serialize(chart)

        overlay_member = "xl/drawings/chartDrawing1.xml"
        overlay = ElementTree.Element(f"{{{chart_drawing_namespace}}}userShapes")
        anchor = ElementTree.SubElement(
            overlay,
            f"{{{chart_drawing_namespace}}}relSizeAnchor",
        )
        shape = ElementTree.SubElement(anchor, f"{{{chart_drawing_namespace}}}sp")
        ElementTree.SubElement(
            shape,
            f"{{{chart_drawing_namespace}}}cNvPr",
            {"id": "1", "name": "Private baseline chart annotation"},
        )
        text = ElementTree.SubElement(shape, f"{{{drawing_namespace}}}t")
        text.text = "Private baseline overlay text"
        ElementTree.SubElement(
            shape,
            f"{{{drawing_namespace}}}blip",
            {f"{{{document_relationships}}}embed": "rIdFenceChartImage"},
        )
        contents[overlay_member] = serialize(overlay)
        contents["xl/media/private-chart-overlay-baseline.png"] = (
            b"private baseline chart overlay image"
        )

        chart_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            chart_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceChartOverlay",
                "Type": f"{document_relationships}/chartUserShapes",
                "Target": "../drawings/chartDrawing1.xml",
            },
        )
        contents[_relationship_member(chart_member)] = serialize(chart_relationships)

        overlay_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            overlay_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceChartImage",
                "Type": f"{document_relationships}/image",
                "Target": "../media/private-chart-overlay-baseline.png",
            },
        )
        contents[_relationship_member(overlay_member)] = serialize(overlay_relationships)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override_tag = f"{{{content_types}}}Override"
        ElementTree.SubElement(
            types,
            override_tag,
            {
                "PartName": "/xl/drawings/chartDrawing1.xml",
                "ContentType": (
                    "application/vnd.openxmlformats-officedocument.drawingml."
                    "chartshapes+xml"
                ),
            },
        )
        contents["[Content_Types].xml"] = serialize(types)

        # Keep this fixture's worksheet drawing relationship and its DrawingML
        # chart binding explicit so relationship-ID normalization has real paths.
        drawing_relationships_name = _relationship_member(drawing_member)
        if drawing_relationships_name not in contents:
            raise ValueError("Fixture does not contain a chart drawing relationship part")

    return _rewrite_archive(path, mutate, ".chart-definition.tmp.xlsx")


def change_chart_definition_material(path: Path) -> Path:
    """Change only private chart, overlay, and related-image material."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    chart_namespace = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    drawing_namespace = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, chart_member, overlay_member = _chart_fixture_part_names(contents)
        chart = ElementTree.fromstring(contents[chart_member])
        formula = next(chart.iter(f"{{{chart_namespace}}}f"))
        formula.text = "Inputs!$B$3:$B$4"
        title_text = next(chart.iter(f"{{{drawing_namespace}}}t"))
        title_text.text = "Private candidate chart title"
        contents[chart_member] = serialize(chart)

        overlay = ElementTree.fromstring(contents[overlay_member])
        overlay_text = next(overlay.iter(f"{{{drawing_namespace}}}t"))
        overlay_text.text = "Private candidate overlay text"
        contents[overlay_member] = serialize(overlay)

        relationships_name = _relationship_member(overlay_member)
        relationships = ElementTree.fromstring(contents[relationships_name])
        image = next(
            relationship
            for relationship in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if relationship.get("Id") == "rIdFenceChartImage"
        )
        image.set("Target", "../media/private-chart-overlay-candidate.png")
        contents[relationships_name] = serialize(relationships)
        contents["xl/media/private-chart-overlay-candidate.png"] = (
            b"private candidate chart overlay image"
        )

    return _rewrite_archive(path, mutate, ".chart-definition-change.tmp.xlsx")


def set_chart_formula_external_workbook_target(path: Path, formula: str) -> Path:
    """Replace one chart formula with a controlled external-workbook target."""
    chart_namespace = "http://schemas.openxmlformats.org/drawingml/2006/chart"

    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, chart_member, _overlay_member = _chart_fixture_part_names(contents)
        chart = ElementTree.fromstring(contents[chart_member])
        chart_formula = next(chart.iter(f"{{{chart_namespace}}}f"))
        chart_formula.text = formula
        contents[chart_member] = ElementTree.tostring(
            chart,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".chart-external-workbook-target.tmp.xlsx")


def change_chart_cached_data(path: Path) -> Path:
    """Change one chart cache value while leaving its definition intact."""
    chart_namespace = "http://schemas.openxmlformats.org/drawingml/2006/chart"

    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, chart_member, _overlay_member = _chart_fixture_part_names(contents)
        chart = ElementTree.fromstring(contents[chart_member])
        value = next(chart.iter(f"{{{chart_namespace}}}numCache"))
        point_value = next(value.iter(f"{{{chart_namespace}}}v"))
        point_value.text = "999"
        contents[chart_member] = ElementTree.tostring(
            chart,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".chart-cache-change.tmp.xlsx")


def externalize_chart_overlay_relationship(path: Path) -> Path:
    """Turn an overlay image relation into a harmless external-target fixture."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, _chart_member, overlay_member = _chart_fixture_part_names(contents)
        relationships_name = _relationship_member(overlay_member)
        relationships = ElementTree.fromstring(contents[relationships_name])
        image = next(
            relationship
            for relationship in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if relationship.get("Id") == "rIdFenceChartImage"
        )
        image.set("Target", "https://example.invalid/private-chart-overlay.png")
        image.set("TargetMode", "External")
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".chart-external-target.tmp.xlsx")


def renumber_chart_relationships(path: Path) -> Path:
    """Rewrite chart relationship IDs while retaining the same semantic graph."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    chart_namespace = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    drawing_namespace = "http://schemas.openxmlformats.org/drawingml/2006/main"
    spreadsheet = _SPREADSHEETML_NS
    replacements = {
        "rId1": "rIdFenceRenumberedChart",
        "rIdFenceChartOverlay": "rIdFenceRenumberedChartOverlay",
        "rIdFenceChartImage": "rIdFenceRenumberedChartImage",
    }

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def replace_ids(root: ElementTree.Element) -> None:
        for relationship in root.findall(f"{{{package_relationships}}}Relationship"):
            if replacement := replacements.get(relationship.get("Id")):
                relationship.set("Id", replacement)

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, chart_member, overlay_member = _chart_fixture_part_names(contents)
        for member in (
            _relationship_member(drawing_member),
            _relationship_member(chart_member),
            _relationship_member(overlay_member),
        ):
            relationships = ElementTree.fromstring(contents[member])
            replace_ids(relationships)
            contents[member] = serialize(relationships)

        drawing = ElementTree.fromstring(contents[drawing_member])
        chart_reference = next(drawing.iter(f"{{{chart_namespace}}}chart"))
        chart_reference.set(
            f"{{{document_relationships}}}id",
            replacements["rId1"],
        )
        contents[drawing_member] = serialize(drawing)

        chart = ElementTree.fromstring(contents[chart_member])
        overlay_reference = next(chart.iter(f"{{{chart_namespace}}}userShapes"))
        overlay_reference.set(
            f"{{{document_relationships}}}id",
            replacements["rIdFenceChartOverlay"],
        )
        contents[chart_member] = serialize(chart)

        overlay = ElementTree.fromstring(contents[overlay_member])
        blip = next(overlay.iter(f"{{{drawing_namespace}}}blip"))
        blip.set(
            f"{{{document_relationships}}}embed",
            replacements["rIdFenceChartImage"],
        )
        contents[overlay_member] = serialize(overlay)

        worksheet = _inputs_worksheet_root(contents)
        drawing_reference = worksheet.find(f"{{{spreadsheet}}}drawing")
        if drawing_reference is not None and drawing_reference.get(
            f"{{{document_relationships}}}id"
        ) in replacements:
            drawing_reference.set(
                f"{{{document_relationships}}}id",
                replacements[drawing_reference.get(f"{{{document_relationships}}}id")],
            )
        _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".chart-relationship-renumber.tmp.xlsx")


def rewrite_chart_internal_target_spelling(path: Path) -> Path:
    """Use equivalent relative target spellings across the chart relationship chain."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, chart_member, overlay_member = _chart_fixture_part_names(contents)
        targets = {
            _relationship_member("xl/worksheets/sheet1.xml"): "../drawings/./"
            + drawing_member.rsplit("/", maxsplit=1)[-1],
            _relationship_member(drawing_member): "../charts/./"
            + chart_member.rsplit("/", maxsplit=1)[-1],
            _relationship_member(chart_member): "../drawings/./"
            + overlay_member.rsplit("/", maxsplit=1)[-1],
            _relationship_member(overlay_member): "../media/./private-chart-overlay-baseline.png",
        }
        for member, target in targets.items():
            relationships = ElementTree.fromstring(contents[member])
            relationship = next(
                relationship
                for relationship in relationships.findall(
                    f"{{{package_relationships}}}Relationship"
                )
                if relationship.get("TargetMode", "Internal").casefold() == "internal"
            )
            relationship.set("Target", target)
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".chart-target-spelling.tmp.xlsx")


def corrupt_chart_definition_root(path: Path) -> Path:
    """Replace a chart root with unexpected XML for coverage tests."""

    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, chart_member, _overlay_member = _chart_fixture_part_names(contents)
        chart = ElementTree.fromstring(contents[chart_member])
        chart.tag = "notChartSpace"
        contents[chart_member] = ElementTree.tostring(
            chart,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".chart-definition-corrupt.tmp.xlsx")


def _extended_chart_fixture_part_names(contents: dict[str, bytes]) -> tuple[str, str]:
    """Return the drawing and ChartEx members from the controlled fixture."""
    drawing_member = "xl/drawings/drawing1.xml"
    chart_member = "xl/charts/chartEx1.xml"
    required = {
        drawing_member,
        chart_member,
        _relationship_member(drawing_member),
        _relationship_member(chart_member),
    }
    if not required <= set(contents):
        raise ValueError("Fixture does not contain an extended-chart package")
    return drawing_member, chart_member


def make_extended_chart_definition_model(path: Path) -> Path:
    """Create an Office 2016+ ChartEx chain behind an Excel-style fallback.

    The fixture mirrors real Excel packages: a ``cx:chart`` binding is nested
    in ``mc:AlternateContent`` and points to a ChartEx part with fixed chart
    style and colour relationships.  All values are deliberate redaction
    sentinels and the package is only used for static inspection tests.
    """
    make_model(path)
    content_types = _CONTENT_TYPES_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    chart_ex = _DRAWINGML_CHART_EX_NS
    chart_style = "http://schemas.microsoft.com/office/drawing/2012/chartStyle"
    chart_colour_style = "http://schemas.microsoft.com/office/drawing/2012/chartColorStyle"

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def marker(
        parent: ElementTree.Element,
        name: str,
        *,
        column: int,
        row: int,
    ) -> None:
        point = ElementTree.SubElement(parent, f"{{{drawing}}}{name}")
        ElementTree.SubElement(point, f"{{{drawing}}}col").text = str(column)
        ElementTree.SubElement(point, f"{{{drawing}}}colOff").text = "0"
        ElementTree.SubElement(point, f"{{{drawing}}}row").text = str(row)
        ElementTree.SubElement(point, f"{{{drawing}}}rowOff").text = "0"

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        drawing_member = "xl/drawings/drawing1.xml"
        chart_member = "xl/charts/chartEx1.xml"

        worksheet = ElementTree.fromstring(contents[worksheet_member])
        ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}drawing",
            {f"{{{document_relationships}}}id": "rIdFenceChartExDrawing"},
        )
        contents[worksheet_member] = serialize(worksheet)

        worksheet_relationships_member = _relationship_member(worksheet_member)
        worksheet_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            worksheet_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceChartExDrawing",
                "Type": f"{document_relationships}/drawing",
                "Target": "../drawings/drawing1.xml",
            },
        )
        contents[worksheet_relationships_member] = serialize(worksheet_relationships)

        drawing_root = ElementTree.Element(f"{{{drawing}}}wsDr")
        anchor = ElementTree.SubElement(drawing_root, f"{{{drawing}}}twoCellAnchor")
        marker(anchor, "from", column=4, row=1)
        marker(anchor, "to", column=11, row=16)
        alternate_content = ElementTree.SubElement(
            anchor,
            f"{{{_MARKUP_COMPATIBILITY_NS}}}AlternateContent",
        )
        choice = ElementTree.SubElement(
            alternate_content,
            f"{{{_MARKUP_COMPATIBILITY_NS}}}Choice",
            {"Requires": "cx1"},
        )
        graphic_frame = ElementTree.SubElement(choice, f"{{{drawing}}}graphicFrame")
        nonvisual = ElementTree.SubElement(
            graphic_frame,
            f"{{{drawing}}}nvGraphicFramePr",
        )
        ElementTree.SubElement(
            nonvisual,
            f"{{{drawing}}}cNvPr",
            {
                "id": "9101",
                "name": "PRIVATE-CHARTEX-NAME",
                "descr": "PRIVATE-CHARTEX-DESCRIPTION",
            },
        )
        ElementTree.SubElement(nonvisual, f"{{{drawing}}}cNvGraphicFramePr")
        transform = ElementTree.SubElement(graphic_frame, f"{{{drawing}}}xfrm")
        ElementTree.SubElement(
            transform,
            f"{{{_DRAWINGML_MAIN_NS}}}off",
            {"x": "0", "y": "0"},
        )
        ElementTree.SubElement(
            transform,
            f"{{{_DRAWINGML_MAIN_NS}}}ext",
            {"cx": "4572000", "cy": "2743200"},
        )
        graphic = ElementTree.SubElement(graphic_frame, f"{{{_DRAWINGML_MAIN_NS}}}graphic")
        graphic_data = ElementTree.SubElement(
            graphic,
            f"{{{_DRAWINGML_MAIN_NS}}}graphicData",
            {"uri": chart_ex},
        )
        ElementTree.SubElement(
            graphic_data,
            f"{{{chart_ex}}}chart",
            {f"{{{document_relationships}}}id": "rIdFenceChartEx"},
        )
        fallback = ElementTree.SubElement(
            alternate_content,
            f"{{{_MARKUP_COMPATIBILITY_NS}}}Fallback",
        )
        fallback_shape = ElementTree.SubElement(fallback, f"{{{drawing}}}sp")
        fallback_nonvisual = ElementTree.SubElement(
            fallback_shape,
            f"{{{drawing}}}nvSpPr",
        )
        ElementTree.SubElement(
            fallback_nonvisual,
            f"{{{drawing}}}cNvPr",
            {"id": "9102", "name": "PRIVATE-CHARTEX-FALLBACK"},
        )
        ElementTree.SubElement(fallback_nonvisual, f"{{{drawing}}}cNvSpPr")
        fallback_text = ElementTree.SubElement(
            fallback_shape,
            f"{{{drawing}}}txBody",
        )
        paragraph = ElementTree.SubElement(fallback_text, f"{{{_DRAWINGML_MAIN_NS}}}p")
        run = ElementTree.SubElement(paragraph, f"{{{_DRAWINGML_MAIN_NS}}}r")
        ElementTree.SubElement(run, f"{{{_DRAWINGML_MAIN_NS}}}t").text = (
            "PRIVATE-CHARTEX-FALLBACK-TEXT"
        )
        ElementTree.SubElement(anchor, f"{{{drawing}}}clientData")
        contents[drawing_member] = serialize(drawing_root)

        drawing_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            drawing_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceChartEx",
                "Type": "http://schemas.microsoft.com/office/2014/relationships/chartEx",
                "Target": "../charts/chartEx1.xml",
            },
        )
        contents[_relationship_member(drawing_member)] = serialize(drawing_relationships)

        chart_space = ElementTree.Element(f"{{{chart_ex}}}chartSpace")
        chart_data = ElementTree.SubElement(chart_space, f"{{{chart_ex}}}chartData")
        data = ElementTree.SubElement(chart_data, f"{{{chart_ex}}}data", {"id": "0"})
        category = ElementTree.SubElement(data, f"{{{chart_ex}}}strDim", {"type": "cat"})
        ElementTree.SubElement(category, f"{{{chart_ex}}}f").text = (
            "PRIVATE-CHARTEX-CATEGORIES"
        )
        values = ElementTree.SubElement(data, f"{{{chart_ex}}}numDim", {"type": "val"})
        ElementTree.SubElement(values, f"{{{chart_ex}}}f").text = "PRIVATE-CHARTEX-VALUES"
        chart = ElementTree.SubElement(chart_space, f"{{{chart_ex}}}chart")
        title = ElementTree.SubElement(chart, f"{{{chart_ex}}}title")
        title_properties = ElementTree.SubElement(title, f"{{{chart_ex}}}txPr")
        title_paragraph = ElementTree.SubElement(
            title_properties,
            f"{{{_DRAWINGML_MAIN_NS}}}p",
        )
        title_run = ElementTree.SubElement(title_paragraph, f"{{{_DRAWINGML_MAIN_NS}}}r")
        ElementTree.SubElement(title_run, f"{{{_DRAWINGML_MAIN_NS}}}t").text = (
            "PRIVATE-CHARTEX-TITLE"
        )
        plot_area = ElementTree.SubElement(chart, f"{{{chart_ex}}}plotArea")
        region = ElementTree.SubElement(plot_area, f"{{{chart_ex}}}plotAreaRegion")
        ElementTree.SubElement(
            region,
            f"{{{chart_ex}}}series",
            {"layoutId": "sunburst", "uniqueId": "{AAAAAAAA-AAAA-AAAA-AAAA-AAAAAAAAAAAA}"},
        )
        contents[chart_member] = serialize(chart_space)

        style_member = "xl/charts/style1.xml"
        style = ElementTree.Element(f"{{{chart_style}}}chartStyle", {"id": "201"})
        ElementTree.SubElement(style, f"{{{chart_style}}}title")
        contents[style_member] = serialize(style)
        colour_member = "xl/charts/colors1.xml"
        colours = ElementTree.Element(
            f"{{{chart_colour_style}}}colorStyle",
            {"meth": "cycle", "id": "10"},
        )
        contents[colour_member] = serialize(colours)

        chart_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            chart_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceChartExStyle",
                "Type": "http://schemas.microsoft.com/office/2011/relationships/chartStyle",
                "Target": "style1.xml",
            },
        )
        ElementTree.SubElement(
            chart_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceChartExColours",
                "Type": "http://schemas.microsoft.com/office/2011/relationships/chartColorStyle",
                "Target": "colors1.xml",
            },
        )
        contents[_relationship_member(chart_member)] = serialize(chart_relationships)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        override_tag = f"{{{content_types}}}Override"
        for part_name, content_type in (
            (
                "/xl/drawings/drawing1.xml",
                "application/vnd.openxmlformats-officedocument.drawing+xml",
            ),
            ("/xl/charts/chartEx1.xml", "application/vnd.ms-office.chartex+xml"),
            ("/xl/charts/style1.xml", "application/vnd.ms-office.chartstyle+xml"),
            (
                "/xl/charts/colors1.xml",
                "application/vnd.ms-office.chartcolorstyle+xml",
            ),
        ):
            ElementTree.SubElement(
                types,
                override_tag,
                {"PartName": part_name, "ContentType": content_type},
            )
        contents["[Content_Types].xml"] = serialize(types)

    return _rewrite_archive(path, mutate, ".extended-chart-definition.tmp.xlsx")


def change_extended_chart_definition_material(path: Path) -> Path:
    """Change private ChartEx XML without editing ordinary worksheet cells."""
    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, chart_member = _extended_chart_fixture_part_names(contents)
        chart = ElementTree.fromstring(contents[chart_member])
        formula = next(chart.iter(f"{{{_DRAWINGML_CHART_EX_NS}}}f"))
        formula.text = "PRIVATE-CANDIDATE-CHARTEX-VALUES"
        title = next(chart.iter(f"{{{_DRAWINGML_MAIN_NS}}}t"))
        title.text = "PRIVATE-CANDIDATE-CHARTEX-TITLE"
        contents[chart_member] = ElementTree.tostring(
            chart,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".extended-chart-definition-change.tmp.xlsx")


def change_extended_chart_style_payload(path: Path) -> Path:
    """Change only a direct ChartEx style payload."""
    chart_style = "http://schemas.microsoft.com/office/drawing/2012/chartStyle"

    def mutate(contents: dict[str, bytes]) -> None:
        style_member = "xl/charts/style1.xml"
        style = ElementTree.fromstring(contents[style_member])
        if style.tag != f"{{{chart_style}}}chartStyle":
            raise ValueError("Fixture does not contain a ChartEx style part")
        style.set("id", "202")
        contents[style_member] = ElementTree.tostring(
            style,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".extended-chart-style-change.tmp.xlsx")


def renumber_extended_chart_relationships(path: Path) -> Path:
    """Rewrite the ChartEx drawing relationship ID without changing its graph."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, _chart_member = _extended_chart_fixture_part_names(contents)
        relationships_member = _relationship_member(drawing_member)
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            item
            for item in relationships.findall(f"{{{package_relationships}}}Relationship")
            if item.get("Id") == "rIdFenceChartEx"
        )
        relationship.set("Id", "rIdFenceRenumberedChartEx")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        drawing = ElementTree.fromstring(contents[drawing_member])
        chart_reference = next(drawing.iter(f"{{{_DRAWINGML_CHART_EX_NS}}}chart"))
        chart_reference.set(
            f"{{{document_relationships}}}id",
            "rIdFenceRenumberedChartEx",
        )
        contents[drawing_member] = ElementTree.tostring(
            drawing,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".extended-chart-id-renumber.tmp.xlsx")


def corrupt_extended_chart_definition_root(path: Path) -> Path:
    """Replace the ChartEx root to exercise fail-closed inspection."""
    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, chart_member = _extended_chart_fixture_part_names(contents)
        chart = ElementTree.fromstring(contents[chart_member])
        chart.tag = "notChartExSpace"
        contents[chart_member] = ElementTree.tostring(
            chart,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".extended-chart-corrupt.tmp.xlsx")


def externalize_extended_chart_relationship(path: Path) -> Path:
    """Turn the ChartEx drawing binding into a harmless external target."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, _chart_member = _extended_chart_fixture_part_names(contents)
        relationships_member = _relationship_member(drawing_member)
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            item
            for item in relationships.findall(f"{{{package_relationships}}}Relationship")
            if item.get("Id") == "rIdFenceChartEx"
        )
        relationship.set("Target", "https://example.invalid/private-chartex")
        relationship.set("TargetMode", "External")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".extended-chart-external.tmp.xlsx")


def remove_extended_chart_direct_relationship_id(path: Path) -> Path:
    """Remove a required ChartEx direct-relationship identifier for coverage tests."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, chart_member = _extended_chart_fixture_part_names(contents)
        relationships_member = _relationship_member(chart_member)
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            item
            for item in relationships.findall(f"{{{package_relationships}}}Relationship")
            if item.get("Id") == "rIdFenceChartExStyle"
        )
        relationship.attrib.pop("Id")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".extended-chart-missing-id.tmp.xlsx")


def add_unsupported_extended_chart_relationship(path: Path) -> Path:
    """Add a direct ChartEx edge outside FormulaFence's bounded relationship set."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, chart_member = _extended_chart_fixture_part_names(contents)
        relationships_member = _relationship_member(chart_member)
        relationships = ElementTree.fromstring(contents[relationships_member])
        ElementTree.SubElement(
            relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceUnsupportedChartEx",
                "Type": "https://example.invalid/relationships/private-chartex",
                "Target": "private-chart-ex.xml",
            },
        )
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/charts/private-chart-ex.xml"] = b"private chart ex payload"

    return _rewrite_archive(path, mutate, ".extended-chart-unsupported.tmp.xlsx")


def make_protection_model(path: Path, *, include_chartsheet: bool = False) -> Path:
    """Create a workbook with operational protection and sparse style controls."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Controlled input"
    inputs["B2"] = 10
    inputs["C2"] = "=B2*2"
    inputs["D2"] = "Column default"
    inputs["B2"].protection = Protection(locked=False)
    inputs["C2"].protection = Protection(hidden=True)
    inputs.row_dimensions[5].protection = Protection(hidden=True)
    inputs.column_dimensions["D"].protection = Protection(locked=False)
    inputs.protection.sheet = True
    inputs.protection.formatCells = False
    inputs.protection.sort = False
    inputs.protection.autoFilter = False
    inputs.protection.selectLockedCells = True
    inputs.protection.set_password("synthetic-sheet-password")

    workbook.security.lockStructure = True
    workbook.security.set_workbook_password("synthetic-workbook-password")
    if include_chartsheet:
        chart = BarChart()
        chart.add_data(
            Reference(inputs, min_col=2, min_row=2, max_row=2), titles_from_data=False
        )
        dashboard = workbook.create_chartsheet("Dashboard")
        dashboard.add_chart(chart)
        dashboard.sheetProtection = ChartsheetProtection(
            content=True,
            objects=True,
            password="synthetic-chart-password",
        )
    workbook.save(path)
    return add_protected_range(path)


def add_protected_range(
    path: Path,
    *,
    name: str = "Synthetic approved inputs",
    sqref: str = "B2:B5",
    password: str = "A1B2",
    security_descriptor: str = "synthetic-security-descriptor",
) -> Path:
    """Add an OOXML protected range that openpyxl intentionally does not model."""
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    def mutate(contents: dict[str, bytes]) -> None:
        root = _inputs_worksheet_root(contents)
        protected_ranges = root.find(f"{{{namespace}}}protectedRanges")
        if protected_ranges is None:
            protected_ranges = ElementTree.Element(f"{{{namespace}}}protectedRanges")
            protection = root.find(f"{{{namespace}}}sheetProtection")
            insertion_index = (
                list(root).index(protection) + 1
                if protection is not None
                else len(root)
            )
            root.insert(insertion_index, protected_ranges)
        protected_range = ElementTree.SubElement(
            protected_ranges,
            f"{{{namespace}}}protectedRange",
        )
        protected_range.set("name", name)
        protected_range.set("sqref", sqref)
        protected_range.set("password", password)
        protected_range.set("securityDescriptor", security_descriptor)
        _save_inputs_worksheet(contents, root)

    return _rewrite_archive(path, mutate, ".protected-range.tmp.xlsx")


def set_sheet_protection_defaults(path: Path, *, explicit: bool) -> Path:
    """Toggle equivalent omitted versus explicit normal-sheet protection defaults."""
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    defaults = {
        "objects": "0",
        "scenarios": "0",
        "formatCells": "1",
        "formatColumns": "1",
        "formatRows": "1",
        "insertColumns": "1",
        "insertRows": "1",
        "insertHyperlinks": "1",
        "deleteColumns": "1",
        "deleteRows": "1",
        "selectLockedCells": "0",
        "sort": "1",
        "autoFilter": "1",
        "pivotTables": "1",
        "selectUnlockedCells": "0",
    }

    def mutate(contents: dict[str, bytes]) -> None:
        root = _inputs_worksheet_root(contents)
        protection = root.find(f"{{{namespace}}}sheetProtection")
        if protection is None:
            raise ValueError("Fixture does not contain sheet protection")
        for attribute, value in defaults.items():
            if explicit:
                protection.set(attribute, value)
            else:
                protection.attrib.pop(attribute, None)
        _save_inputs_worksheet(contents, root)

    return _rewrite_archive(path, mutate, ".protection-defaults.tmp.xlsx")


def set_sheet_protection_modern_verifier(path: Path, hash_value: str) -> Path:
    """Set synthetic SHA-512 verifier fields for redaction and equality tests."""
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    def mutate(contents: dict[str, bytes]) -> None:
        root = _inputs_worksheet_root(contents)
        protection = root.find(f"{{{namespace}}}sheetProtection")
        if protection is None:
            raise ValueError("Fixture does not contain sheet protection")
        protection.attrib.pop("password", None)
        protection.set("algorithmName", "SHA-512")
        protection.set("hashValue", hash_value)
        protection.set("saltValue", "c3ludGhldGljLXNhbHQ=")
        protection.set("spinCount", "100000")
        _save_inputs_worksheet(contents, root)

    return _rewrite_archive(path, mutate, ".modern-verifier.tmp.xlsx")


def change_protected_range(
    path: Path,
    *,
    sqref: str | None = None,
    name: str | None = None,
    password: str | None = None,
    security_descriptor: str | None = None,
) -> Path:
    """Change one synthetic protected range's non-secret test properties."""
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    def mutate(contents: dict[str, bytes]) -> None:
        root = _inputs_worksheet_root(contents)
        protected_range = root.find(
            f"{{{namespace}}}protectedRanges/{{{namespace}}}protectedRange"
        )
        if protected_range is None:
            raise ValueError("Fixture does not contain a protected range")
        for attribute, value in (
            ("sqref", sqref),
            ("name", name),
            ("password", password),
            ("securityDescriptor", security_descriptor),
        ):
            if value is not None:
                protected_range.set(attribute, value)
        _save_inputs_worksheet(contents, root)

    return _rewrite_archive(path, mutate, ".protected-range-change.tmp.xlsx")


def reorder_conditional_differential_styles(path: Path) -> Path:
    """Swap ``dxfs`` and their rule ids without changing any visual rule."""
    with ZipFile(path) as archive:
        contents = {
            entry.filename: archive.read(entry.filename) for entry in archive.infolist()
        }
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    styles = ElementTree.fromstring(contents["xl/styles.xml"])
    dxfs = styles.find(f"{{{namespace}}}dxfs")
    if dxfs is None or len(dxfs) < 2:
        raise ValueError("Fixture needs at least two conditional differential styles")
    first, second = dxfs[:2]
    dxfs[:2] = [second, first]
    contents["xl/styles.xml"] = ElementTree.tostring(
        styles,
        encoding="utf-8",
        xml_declaration=True,
    )
    rule_tag = f"{{{namespace}}}cfRule"
    for name, content in tuple(contents.items()):
        if not name.startswith("xl/worksheets/"):
            continue
        worksheet = ElementTree.fromstring(content)
        for rule in worksheet.iter(rule_tag):
            if rule.get("dxfId") == "0":
                rule.set("dxfId", "1")
            elif rule.get("dxfId") == "1":
                rule.set("dxfId", "0")
        contents[name] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )
    staging = path.with_suffix(".dxf-order.tmp.xlsx")
    with ZipFile(staging, "w", compression=ZIP_DEFLATED) as archive:
        for name, content in contents.items():
            archive.writestr(name, content)
    staging.replace(path)
    return path


def add_conditional_formatting_databar_extension(
    path: Path,
    *,
    guid: str,
    axis_color: str,
    worksheet_extension_uri: str = "{78C0D931-6437-407D-A8EE-F0AAD7539E65}",
) -> Path:
    """Add a minimal Excel-2010 data-bar extension that openpyxl does not model."""
    with ZipFile(path) as archive:
        contents = {
            entry.filename: archive.read(entry.filename) for entry in archive.infolist()
        }
    main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    x14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
    xm = "http://schemas.microsoft.com/office/excel/2006/main"
    rule_tag = f"{{{main}}}cfRule"
    data_bar_rule: ElementTree.Element | None = None
    worksheet_name: str | None = None
    for name, content in tuple(contents.items()):
        if not name.startswith("xl/worksheets/"):
            continue
        worksheet = ElementTree.fromstring(content)
        data_bar_rule = next(
            (rule for rule in worksheet.iter(rule_tag) if rule.get("type") == "dataBar"),
            None,
        )
        if data_bar_rule is None:
            continue
        worksheet_name = name
        base_extensions = ElementTree.SubElement(data_bar_rule, f"{{{main}}}extLst")
        base_extension = ElementTree.SubElement(
            base_extensions,
            f"{{{main}}}ext",
            {"uri": "{B025F937-C7B1-47D3-B67F-A62EFF666E3E}"},
        )
        ElementTree.SubElement(base_extension, f"{{{x14}}}id").text = guid

        worksheet_extensions = ElementTree.SubElement(worksheet, f"{{{main}}}extLst")
        worksheet_extension = ElementTree.SubElement(
            worksheet_extensions,
            f"{{{main}}}ext",
            {"uri": worksheet_extension_uri},
        )
        conditional_formattings = ElementTree.SubElement(
            worksheet_extension,
            f"{{{x14}}}conditionalFormattings",
        )
        conditional_formatting = ElementTree.SubElement(
            conditional_formattings,
            f"{{{x14}}}conditionalFormatting",
        )
        extension_rule = ElementTree.SubElement(
            conditional_formatting,
            f"{{{x14}}}cfRule",
            {"type": "dataBar", "id": guid},
        )
        extension_data_bar = ElementTree.SubElement(
            extension_rule,
            f"{{{x14}}}dataBar",
            {"minLength": "0", "maxLength": "100", "axisPosition": "middle"},
        )
        ElementTree.SubElement(extension_data_bar, f"{{{x14}}}cfvo", {"type": "autoMin"})
        ElementTree.SubElement(extension_data_bar, f"{{{x14}}}cfvo", {"type": "autoMax"})
        ElementTree.SubElement(extension_data_bar, f"{{{x14}}}axisColor", {"rgb": axis_color})
        ElementTree.SubElement(conditional_formatting, f"{{{xm}}}sqref").text = "C2:C100"
        contents[name] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )
        break
    if data_bar_rule is None or worksheet_name is None:
        raise ValueError("Could not find the fixture data-bar rule")
    staging = path.with_suffix(".conditional-extension.tmp.xlsx")
    with ZipFile(staging, "w", compression=ZIP_DEFLATED) as archive:
        for name, content in contents.items():
            archive.writestr(name, content)
    staging.replace(path)
    return path


def make_current_row_table_model(path: Path) -> Path:
    """Create calculated-column formulas using three current-row spellings."""
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(["Sales Amount", "Rate", "Value"])
    data.append([10, 0.1, "=[@[Sales Amount]]*[@Rate]"])
    data.append([20, 0.2, "=[Sales Amount]*[Rate]"])
    data.append([30, 0.3, "=Sales[[#This Row],[Sales Amount]]*Sales[@Rate]"])
    data.add_table(Table(displayName="Sales", ref="A1:C4"))
    data["E1"] = "Adjacent current-row output"
    data["E2"] = "=Sales[[#This Row],[Sales Amount]]"

    report = workbook.create_sheet("Report")
    report["A1"] = "Current-row table output"
    report["B2"] = "=SUM(Sales[Value])"
    workbook.save(path)
    return path


def make_three_d_model(path: Path) -> Path:
    """Create a period-stack workbook with one static 3-D summary formula."""
    workbook = Workbook()
    january = workbook.active
    january.title = "Jan"
    january["A1"] = "Period input"
    january["B2"] = 10

    for title, amount in (("Feb", 20), ("Mar", 30)):
        period = workbook.create_sheet(title)
        period["A1"] = "Period input"
        period["B2"] = amount

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "3-D consolidation"
    summary["B2"] = "=SUM(Jan:Mar!B2)"
    workbook.save(path)
    return path


def make_spill_model(path: Path) -> Path:
    """Create literal and OOXML-style dynamic-array spill-reference callers."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Dynamic-array anchors"
    inputs["B2"] = "=SEQUENCE(3)"
    inputs["B3"] = "=SEQUENCE(2)"

    model = workbook.create_sheet("Model")
    model["A1"] = "Spill-driven calculations"
    model["B2"] = "=SUM(Inputs!B2#)"
    model["B3"] = "=COUNTA(_xlfn.ANCHORARRAY(Inputs!B3))"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Spill-driven output"
    dashboard["B2"] = "=Model!B2"
    workbook.save(path)
    return path


def make_implicit_intersection_model(path: Path) -> Path:
    """Create persisted SINGLE() and literal @ consumers of a static input range."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Implicit-intersection inputs"
    inputs["B2"] = 10
    inputs["B3"] = 20
    inputs["B4"] = 30

    model = workbook.create_sheet("Model")
    model["A1"] = "Implicit-intersection calculations"
    model["B2"].value = ArrayFormula(
        ref="B2", text="=_xlfn.SINGLE(Inputs!B2:B4)"
    )
    model["B3"] = "=@Inputs!B2:B4"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Implicit-intersection output"
    dashboard["B2"] = "=Model!B2"
    workbook.save(path)
    return path


def make_legacy_array_model(path: Path, array_ref: str = "B1:B3") -> Path:
    """Create a fixed CSE array whose result members feed ordinary formulas."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "a"
    inputs["A2"] = "bb"
    inputs["A3"] = "ccc"

    model = workbook.create_sheet("Model")
    model["A1"] = "Legacy CSE array result"
    model["B1"].value = ArrayFormula(ref=array_ref, text="=LEN(Inputs!A1:A3)")
    model["C2"] = "=B2*10"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Array output consumers"
    dashboard["B2"] = "=SUM(Model!B2:B3)"
    workbook.save(path)
    return path


def make_what_if_data_table_model(path: Path) -> Path:
    """Create controlled one- and two-variable What-If Data Table masters."""
    workbook = Workbook()
    sensitivity = workbook.active
    sensitivity.title = "Sensitivity"
    sensitivity["A1"] = "Controlled scenario analysis"
    sensitivity["B2"] = 0.05
    sensitivity["B3"] = 0.1
    sensitivity["B4"] = 0.15
    sensitivity["D2"] = "=B2*100"
    sensitivity["F2"] = "=B2*200"
    sensitivity["K2"] = "=B2*B3*1000"
    sensitivity["D3"] = DataTableFormula(
        ref="D3:D6",
        dt2D=False,
        dtr=False,
        r1="B2",
        ca=True,
    )
    sensitivity["F3"] = DataTableFormula(
        ref="F3:I3",
        dt2D=False,
        dtr=True,
        r1="B2",
    )
    sensitivity["K3"] = DataTableFormula(
        ref="K3:M5",
        dt2D=True,
        r1="B2",
        r2="B3",
    )
    workbook.save(path)
    return path


def make_scenario_manager_model(
    path: Path, *, duplicate_names_on_second_sheet: bool = False
) -> Path:
    """Create a controlled raw-OOXML Excel Scenario Manager fixture."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Scenario-controlled inputs"
    inputs["B2"] = 0.1
    inputs["B3"] = 125
    inputs["D2"] = "Scenario summary result one"
    inputs["D3"] = "Scenario summary result two"
    if duplicate_names_on_second_sheet:
        workbook.create_sheet("Second Inputs")
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        scenarios = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}scenarios",
            {"current": "1", "show": "0", "sqref": "D2 D3"},
        )
        upside = ElementTree.SubElement(
            scenarios,
            f"{{{_SPREADSHEETML_NS}}}scenario",
            {
                "name": "Private Upside",
                "locked": "1",
                "count": "2",
                "user": "private-scenario-owner",
                "comment": "PRIVATE-SCENARIO-COMMENT",
            },
        )
        ElementTree.SubElement(
            upside,
            f"{{{_SPREADSHEETML_NS}}}inputCells",
            {"r": "B2", "val": "PRIVATE-UPSIDERATE", "numFmtId": "10"},
        )
        ElementTree.SubElement(
            upside,
            f"{{{_SPREADSHEETML_NS}}}inputCells",
            {"r": "B3", "val": "PRIVATE-UPSIDECOST"},
        )
        downside = ElementTree.SubElement(
            scenarios,
            f"{{{_SPREADSHEETML_NS}}}scenario",
            {
                "name": "Private Downside",
                "hidden": "1",
                "count": "2",
                "comment": "PRIVATE-DOWNSIDE-COMMENT",
            },
        )
        ElementTree.SubElement(
            downside,
            f"{{{_SPREADSHEETML_NS}}}inputCells",
            {"r": "B2", "val": "PRIVATE-DOWNSIDERATE"},
        )
        ElementTree.SubElement(
            downside,
            f"{{{_SPREADSHEETML_NS}}}inputCells",
            {"r": "B3", "val": "PRIVATE-DOWNSIDECOST"},
        )
        sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
        sheet_data_index = next(
            index
            for index, child in enumerate(worksheet)
            if child.tag == sheet_data_tag
        )
        worksheet.insert(sheet_data_index + 1, scenarios)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )
        if duplicate_names_on_second_sheet:
            second_worksheet = ElementTree.fromstring(
                contents["xl/worksheets/sheet2.xml"]
            )
            second_sheet_data_index = next(
                index
                for index, child in enumerate(second_worksheet)
                if child.tag == sheet_data_tag
            )
            second_worksheet.insert(
                second_sheet_data_index + 1,
                ElementTree.fromstring(
                    ElementTree.tostring(scenarios, encoding="utf-8")
                ),
            )
            contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
                second_worksheet,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".scenario-manager.tmp.xlsx")


def change_scenario_manager_input_value(path: Path) -> Path:
    """Change a private stored Scenario Manager input without touching cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        input_cell_tag = f"{{{_SPREADSHEETML_NS}}}inputCells"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        input_cell = next(worksheet.iter(input_cell_tag))
        input_cell.set("val", "CANDIDATE-PRIVATE-SCENARIO-VALUE")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".scenario-manager-change.tmp.xlsx")


def normalize_scenario_manager_reference_spelling(path: Path) -> Path:
    """Use equivalent OOXML references, integers, and Boolean spellings."""

    def mutate(contents: dict[str, bytes]) -> None:
        scenarios_tag = f"{{{_SPREADSHEETML_NS}}}scenarios"
        scenario_tag = f"{{{_SPREADSHEETML_NS}}}scenario"
        input_cell_tag = f"{{{_SPREADSHEETML_NS}}}inputCells"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        scenarios = worksheet.find(scenarios_tag)
        if scenarios is None:
            raise ValueError("Could not find Scenario Manager fixture")
        scenarios.set("current", "01")
        scenarios.set("show", "00")
        scenarios.set("sqref", "$D$2   $D$3")
        for index, scenario in enumerate(scenarios.findall(scenario_tag)):
            scenario.set("count", "02")
            scenario.set("locked", "true" if index == 0 else "false")
            scenario.set("hidden", "false" if index == 0 else "true")
            for input_cell in scenario.findall(input_cell_tag):
                reference = input_cell.get("r")
                if reference is None:
                    raise ValueError("Scenario Manager fixture input lacks a reference")
                input_cell.set("r", f"${reference[0]}${reference[1:]}")
                input_cell.set("deleted", "false")
                input_cell.set("undone", "0")
                if input_cell.get("numFmtId") is not None:
                    input_cell.set("numFmtId", "010")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".scenario-manager-noise.tmp.xlsx")


def corrupt_scenario_manager_input(path: Path) -> Path:
    """Inject an unsupported Scenario Manager input to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        input_cell_tag = f"{{{_SPREADSHEETML_NS}}}inputCells"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        input_cell = next(worksheet.iter(input_cell_tag))
        input_cell.set("r", "PrivateInputSheet!B2")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".scenario-manager-corrupt.tmp.xlsx")


def make_filter_visibility_model(path: Path) -> Path:
    """Create raw OOXML filter and visibility controls with private criteria."""
    workbook = Workbook()
    report = workbook.active
    report.title = "Report"
    report.append(["Department", "Region", "Amount", "Notes", "Visible total"])
    report.append(["Operations", "North", 100, "ordinary", "=SUBTOTAL(109,C2:C5)"])
    report.append(["Sales", "South", 200, "ordinary", "=AGGREGATE(9,5,C2:C5)"])
    report.append(["Finance", "North", 300, "ordinary", None])
    report.append(["Legal", "West", 400, "ordinary", None])

    table_sheet = workbook.create_sheet("Table Report")
    table_sheet.append(["Segment", "Value", "Status"])
    table_sheet.append(["Private A", 10, "Open"])
    table_sheet.append(["Private B", 20, "Open"])
    table_sheet.append(["Private C", 30, "Closed"])
    table_sheet.append(["Private D", 40, "Closed"])
    table_sheet.add_table(Table(displayName="ReportTable", ref="A1:C5"))

    default_hidden = workbook.create_sheet("Default Hidden")
    default_hidden["A1"] = "Visible override under a hidden-by-default sheet"
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
        filter_column_tag = f"{{{_SPREADSHEETML_NS}}}filterColumn"
        filters_tag = f"{{{_SPREADSHEETML_NS}}}filters"
        filter_tag = f"{{{_SPREADSHEETML_NS}}}filter"
        custom_filters_tag = f"{{{_SPREADSHEETML_NS}}}customFilters"
        custom_filter_tag = f"{{{_SPREADSHEETML_NS}}}customFilter"
        sort_state_tag = f"{{{_SPREADSHEETML_NS}}}sortState"
        sort_condition_tag = f"{{{_SPREADSHEETML_NS}}}sortCondition"
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
        sheet_format_tag = f"{{{_SPREADSHEETML_NS}}}sheetFormatPr"
        row_tag = f"{{{_SPREADSHEETML_NS}}}row"

        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        auto_filter = ElementTree.Element(auto_filter_tag, {"ref": "A1:C5"})
        worksheet_column = ElementTree.SubElement(
            auto_filter,
            filter_column_tag,
            {"colId": "1", "hiddenButton": "0", "showButton": "true"},
        )
        worksheet_filters = ElementTree.SubElement(
            worksheet_column,
            filters_tag,
            {"blank": "false", "calendarType": "none"},
        )
        ElementTree.SubElement(
            worksheet_filters,
            filter_tag,
            {"val": "PRIVATE-WORKSHEET-REGION"},
        )
        worksheet_sort = ElementTree.SubElement(
            auto_filter,
            sort_state_tag,
            {"ref": "A1:C5", "caseSensitive": "false"},
        )
        ElementTree.SubElement(
            worksheet_sort,
            sort_condition_tag,
            {"ref": "C2:C5", "descending": "1", "customList": "PRIVATE-SORT-LIST"},
        )
        sheet_data_index = next(
            index
            for index, child in enumerate(report_xml)
            if child.tag == sheet_data_tag
        )
        report_xml.insert(sheet_data_index + 1, auto_filter)
        columns = ElementTree.Element(cols_tag)
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "2", "max": "5", "hidden": "1", "outlineLevel": "1"},
        )
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "2", "max": "5", "width": "11", "customWidth": "1"},
        )
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "3", "max": "3", "hidden": "false", "outlineLevel": "2"},
        )
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "6", "max": "6", "collapsed": "true"},
        )
        report_xml.insert(sheet_data_index, columns)
        rows = {row.get("r"): row for row in report_xml.iter(row_tag)}
        rows["3"].set("hidden", "1")
        rows["3"].set("outlineLevel", "1")
        rows["4"].set("hidden", "1")
        rows["4"].set("outlineLevel", "2")
        rows["4"].set("collapsed", "1")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

        table_xml = ElementTree.fromstring(contents["xl/tables/table1.xml"])
        table_auto_filter = table_xml.find(auto_filter_tag)
        if table_auto_filter is None:
            table_auto_filter = ElementTree.Element(auto_filter_tag, {"ref": "A1:C5"})
            table_columns_tag = f"{{{_SPREADSHEETML_NS}}}tableColumns"
            table_columns_index = next(
                index
                for index, child in enumerate(table_xml)
                if child.tag == table_columns_tag
            )
            table_xml.insert(table_columns_index, table_auto_filter)
        table_column = ElementTree.SubElement(
            table_auto_filter,
            filter_column_tag,
            {"colId": "0"},
        )
        table_filters = ElementTree.SubElement(
            table_column,
            custom_filters_tag,
            {"and": "true"},
        )
        ElementTree.SubElement(
            table_filters,
            custom_filter_tag,
            {"operator": "equal", "val": "PRIVATE-TABLE-SEGMENT"},
        )
        table_sort = ElementTree.SubElement(table_auto_filter, sort_state_tag, {"ref": "A1:C5"})
        ElementTree.SubElement(
            table_sort,
            sort_condition_tag,
            {"ref": "B2:B5", "descending": "false"},
        )
        contents["xl/tables/table1.xml"] = ElementTree.tostring(
            table_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

        default_hidden_xml = ElementTree.fromstring(contents["xl/worksheets/sheet3.xml"])
        sheet_format = default_hidden_xml.find(sheet_format_tag)
        if sheet_format is None:
            sheet_format = ElementTree.Element(sheet_format_tag)
            default_hidden_xml.insert(0, sheet_format)
        sheet_format.set("zeroHeight", "true")
        default_hidden_row = next(default_hidden_xml.iter(row_tag))
        default_hidden_row.set("hidden", "false")
        contents["xl/worksheets/sheet3.xml"] = ElementTree.tostring(
            default_hidden_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".filter-visibility.tmp.xlsx")


def make_zero_dimension_visibility_model(path: Path) -> Path:
    """Create a workbook whose rows and columns can be concealed by size alone."""
    workbook = Workbook()
    report = workbook.active
    report.title = "Dimension Report"
    report.append(["Metric", "Private amount", "Visible amount"])
    report.append(["Revenue", 1200, 700])
    report.append(["Costs", 750, 350])
    report.append(["Profit", "=B2-B3", "=C2-C3"])
    workbook.save(path)
    return path


def change_zero_dimension_visibility_controls(path: Path) -> Path:
    """Hide one populated row and column through raw zero-size declarations."""

    def mutate(contents: dict[str, bytes]) -> None:
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
        row_tag = f"{{{_SPREADSHEETML_NS}}}row"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = report_xml.find(cols_tag)
        if columns is None:
            columns = ElementTree.Element(cols_tag)
            sheet_data_index = next(
                index
                for index, child in enumerate(report_xml)
                if child.tag == sheet_data_tag
            )
            report_xml.insert(sheet_data_index, columns)
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "2", "max": "3", "width": "0", "customWidth": "1"},
        )
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "3", "max": "3", "width": "12", "customWidth": "1"},
        )
        row = next(row for row in report_xml.iter(row_tag) if row.get("r") == "3")
        row.set("ht", "0")
        row.set("customHeight", "true")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".zero-dimension-visibility.tmp.xlsx")


def change_default_zero_dimension_visibility_controls(path: Path) -> Path:
    """Set worksheet-default row and column dimensions to Excel's hidden size."""

    def mutate(contents: dict[str, bytes]) -> None:
        sheet_format_tag = f"{{{_SPREADSHEETML_NS}}}sheetFormatPr"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        sheet_format = report_xml.find(sheet_format_tag)
        if sheet_format is None:
            sheet_format = ElementTree.Element(sheet_format_tag)
            report_xml.insert(0, sheet_format)
        sheet_format.set("defaultRowHeight", "0")
        sheet_format.set("defaultColWidth", "0")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".default-zero-dimension-visibility.tmp.xlsx")


def add_ordinary_dimension_resize(path: Path) -> Path:
    """Change positive dimensions, which remain outside the concealment boundary."""

    def mutate(contents: dict[str, bytes]) -> None:
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
        row_tag = f"{{{_SPREADSHEETML_NS}}}row"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = report_xml.find(cols_tag)
        if columns is None:
            columns = ElementTree.Element(cols_tag)
            sheet_data_index = next(
                index
                for index, child in enumerate(report_xml)
                if child.tag == sheet_data_tag
            )
            report_xml.insert(sheet_data_index, columns)
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "2", "max": "2", "width": "24", "customWidth": "1"},
        )
        row = next(row for row in report_xml.iter(row_tag) if row.get("r") == "3")
        row.set("ht", "22")
        row.set("customHeight", "1")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ordinary-dimension-resize.tmp.xlsx")


def make_worksheet_dimension_model(path: Path) -> Path:
    """Create material positive sizing controls without changing cell values."""
    make_zero_dimension_visibility_model(path)

    def mutate(contents: dict[str, bytes]) -> None:
        sheet_format_tag = f"{{{_SPREADSHEETML_NS}}}sheetFormatPr"
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
        row_tag = f"{{{_SPREADSHEETML_NS}}}row"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        sheet_format = worksheet.find(sheet_format_tag)
        if sheet_format is None:
            sheet_format = ElementTree.Element(sheet_format_tag)
            worksheet.insert(0, sheet_format)
        sheet_format.set("baseColWidth", "10")
        sheet_format.set("defaultColWidth", "17.5")
        sheet_format.set("defaultRowHeight", "21")
        sheet_format.set("customHeight", "false")
        sheet_format.set("thickTop", "true")
        sheet_format.set("thickBottom", "false")

        columns = worksheet.find(cols_tag)
        if columns is None:
            columns = ElementTree.Element(cols_tag)
            sheet_data_index = next(
                index
                for index, child in enumerate(worksheet)
                if child.tag == sheet_data_tag
            )
            worksheet.insert(sheet_data_index, columns)
        ElementTree.SubElement(
            columns,
            col_tag,
            {
                "min": "2",
                "max": "3",
                "width": "24",
                "bestFit": "1",
                "customWidth": "true",
            },
        )
        row = next(current for current in worksheet.iter(row_tag) if current.get("r") == "3")
        row.set("ht", "22")
        row.set("customHeight", "true")
        automatic_row = next(
            current for current in worksheet.iter(row_tag) if current.get("r") == "4"
        )
        automatic_row.set("thickTop", "1")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-dimension.tmp.xlsx")


def change_worksheet_dimension_controls(path: Path) -> Path:
    """Change raw positive dimensions and AutoFit without editing cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        namespace = worksheet.tag.removeprefix("{").split("}", maxsplit=1)[0]
        sheet_format_tag = f"{{{namespace}}}sheetFormatPr"
        col_tag = f"{{{namespace}}}col"
        row_tag = f"{{{namespace}}}row"
        sheet_format = worksheet.find(sheet_format_tag)
        if sheet_format is None:
            raise ValueError("Could not find worksheet dimension sheet format")
        sheet_format.set("defaultRowHeight", "23")
        sheet_format.set("defaultColWidth", "18.25")
        sheet_format.set("thickBottom", "true")
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "2" and current.get("max") == "3"
        )
        column.set("width", "31")
        column.set("bestFit", "false")
        row = next(current for current in worksheet.iter(row_tag) if current.get("r") == "3")
        row.set("ht", "27")
        automatic_row = next(
            current for current in worksheet.iter(row_tag) if current.get("r") == "4"
        )
        automatic_row.set("thickTop", "false")
        automatic_row.set("thickBot", "true")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-dimension-change.tmp.xlsx")


def normalize_worksheet_dimension_control_spelling(path: Path) -> Path:
    """Use equivalent numeric, Boolean, and layered-range encodings."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        namespace = worksheet.tag.removeprefix("{").split("}", maxsplit=1)[0]
        sheet_format_tag = f"{{{namespace}}}sheetFormatPr"
        cols_tag = f"{{{namespace}}}cols"
        col_tag = f"{{{namespace}}}col"
        row_tag = f"{{{namespace}}}row"
        sheet_format = worksheet.find(sheet_format_tag)
        columns = worksheet.find(cols_tag)
        if sheet_format is None or columns is None:
            raise ValueError("Could not find worksheet dimension fixture")
        sheet_format.set("baseColWidth", "010")
        sheet_format.set("defaultColWidth", "17.50")
        sheet_format.set("defaultRowHeight", "2.1e1")
        sheet_format.set("customHeight", "0")
        sheet_format.set("thickTop", "1")
        sheet_format.set("thickBottom", "0")
        column = next(
            current
            for current in columns.findall(col_tag)
            if current.get("min") == "2" and current.get("max") == "3"
        )
        attributes = dict(column.attrib)
        columns.remove(column)
        for minimum, maximum in (("2", "2"), ("3", "3")):
            ElementTree.SubElement(
                columns,
                col_tag,
                {
                    **attributes,
                    "min": minimum,
                    "max": maximum,
                    "width": "2.4e1",
                    "bestFit": "true",
                    "customWidth": "1",
                },
            )
        row = next(current for current in worksheet.iter(row_tag) if current.get("r") == "3")
        row.set("ht", "2.20e1")
        row.set("customHeight", "1")
        automatic_row = next(
            current for current in worksheet.iter(row_tag) if current.get("r") == "4"
        )
        automatic_row.set("thickTop", "true")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-dimension-noise.tmp.xlsx")


def normalize_worksheet_dimension_inert_declarations(path: Path) -> Path:
    """Add valid inert writer hints to fixed-height dimension controls."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        namespace = worksheet.tag.removeprefix("{").split("}", maxsplit=1)[0]
        col_tag = f"{{{namespace}}}col"
        row_tag = f"{{{namespace}}}row"
        column = next(current for current in worksheet.iter(col_tag) if current.get("min") == "2")
        column.set("customWidth", "false")
        row = next(current for current in worksheet.iter(row_tag) if current.get("r") == "3")
        # Fixed custom heights supersede thick-border auto-adjustment.
        row.set("thickTop", "true")
        row.set("thickBot", "true")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-dimension-inert.tmp.xlsx")


def add_worksheet_dimension_baseline_adjustments(path: Path) -> Path:
    """Add Office 2010 baseline controls with custom-height side effects."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        namespace = worksheet.tag.removeprefix("{").split("}", maxsplit=1)[0]
        sheet_format_tag = f"{{{namespace}}}sheetFormatPr"
        row_tag = f"{{{namespace}}}row"
        dy_descent_attribute = f"{{{_OFFICE_2010_AC_NS}}}dyDescent"
        sheet_format = worksheet.find(sheet_format_tag)
        if sheet_format is None:
            raise ValueError("Could not find worksheet dimension sheet format")
        sheet_format.set(dy_descent_attribute, "0.25")
        row = next(current for current in worksheet.iter(row_tag) if current.get("r") == "3")
        row.set(dy_descent_attribute, "0.375")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-dimension-baseline.tmp.xlsx")


def corrupt_worksheet_dimension_control(path: Path) -> Path:
    """Inject malformed raw sizing metadata for fail-closed coverage tests."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        namespace = worksheet.tag.removeprefix("{").split("}", maxsplit=1)[0]
        col_tag = f"{{{namespace}}}col"
        row_tag = f"{{{namespace}}}row"
        column = next(current for current in worksheet.iter(col_tag) if current.get("min") == "2")
        column.set("width", "PRIVATE-INVALID-DIMENSION")
        row = next(current for current in worksheet.iter(row_tag) if current.get("r") == "3")
        row.set("ht", "NaN")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-dimension-corrupt.tmp.xlsx")


def make_strict_worksheet_dimension_model(path: Path) -> Path:
    """Create a strict-SpreadsheetML worksheet-dimension fixture."""
    make_worksheet_dimension_model(path)

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_member = "xl/workbook.xml"
        workbook = ElementTree.fromstring(contents[workbook_member])
        for element in workbook.iter():
            element.tag = strict_name(
                element.tag,
                _SPREADSHEETML_NS,
                _STRICT_SPREADSHEETML_NS,
            )
            attributes = {
                strict_name(
                    name,
                    _DOCUMENT_RELATIONSHIPS_NS,
                    _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                ): value
                for name, value in element.attrib.items()
            }
            element.attrib.clear()
            element.attrib.update(attributes)
        contents[workbook_member] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        for relationship in relationships.findall(relationship_tag):
            if (relationship.get("Type") or "").casefold().endswith("/worksheet"):
                relationship.set(
                    "Type",
                    f"{_STRICT_DOCUMENT_RELATIONSHIPS_NS}/worksheet",
                )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        for member in sorted(
            name
            for name in contents
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ):
            worksheet = ElementTree.fromstring(contents[member])
            for element in worksheet.iter():
                element.tag = strict_name(
                    element.tag,
                    _SPREADSHEETML_NS,
                    _STRICT_SPREADSHEETML_NS,
                )
            contents[member] = ElementTree.tostring(
                worksheet,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".strict-worksheet-dimension.tmp.xlsx")


def normalize_zero_dimension_visibility_control_spelling(path: Path) -> Path:
    """Use equivalent zero-size spellings without changing effective visibility."""

    def mutate(contents: dict[str, bytes]) -> None:
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        row_tag = f"{{{_SPREADSHEETML_NS}}}row"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = report_xml.find(cols_tag)
        if columns is None:
            raise ValueError("Could not find zero-dimension column fixture")
        zero_width_column = next(
            column
            for column in columns.findall(col_tag)
            if column.get("min") == "2" and column.get("max") == "3"
        )
        zero_width_column.set("width", "-0.0")
        row = next(row for row in report_xml.iter(row_tag) if row.get("r") == "3")
        row.set("ht", "0e-4")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".zero-dimension-visibility-noise.tmp.xlsx")


def corrupt_zero_dimension_visibility_controls(path: Path) -> Path:
    """Inject malformed sizes to ensure visibility parsing fails closed."""

    def mutate(contents: dict[str, bytes]) -> None:
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
        row_tag = f"{{{_SPREADSHEETML_NS}}}row"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = report_xml.find(cols_tag)
        if columns is None:
            columns = ElementTree.Element(cols_tag)
            sheet_data_index = next(
                index
                for index, child in enumerate(report_xml)
                if child.tag == sheet_data_tag
            )
            report_xml.insert(sheet_data_index, columns)
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "2", "max": "2", "width": "0_0", "customWidth": "1"},
        )
        row = next(row for row in report_xml.iter(row_tag) if row.get("r") == "3")
        row.set("ht", "NaN")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".zero-dimension-visibility-corrupt.tmp.xlsx")


def corrupt_default_zero_dimension_visibility_controls(path: Path) -> Path:
    """Inject malformed worksheet-default sizes to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        sheet_format_tag = f"{{{_SPREADSHEETML_NS}}}sheetFormatPr"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        sheet_format = report_xml.find(sheet_format_tag)
        if sheet_format is None:
            sheet_format = ElementTree.Element(sheet_format_tag)
            report_xml.insert(0, sheet_format)
        sheet_format.set("defaultRowHeight", "-1")
        sheet_format.set("defaultColWidth", "256")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(
        path,
        mutate,
        ".default-zero-dimension-visibility-corrupt.tmp.xlsx",
    )


def change_filter_visibility_criterion(path: Path) -> Path:
    """Change a worksheet filter member without touching a cell or formula."""

    def mutate(contents: dict[str, bytes]) -> None:
        filter_tag = f"{{{_SPREADSHEETML_NS}}}filter"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        next(report_xml.iter(filter_tag)).set("val", "CANDIDATE-PRIVATE-WORKSHEET-REGION")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".filter-visibility-criterion.tmp.xlsx")


def change_filter_visibility_hidden_row(path: Path) -> Path:
    """Reveal one raw hidden row without changing cell content."""

    def mutate(contents: dict[str, bytes]) -> None:
        row_tag = f"{{{_SPREADSHEETML_NS}}}row"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        row = next(row for row in report_xml.iter(row_tag) if row.get("r") == "3")
        row.set("hidden", "false")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".filter-visibility-row.tmp.xlsx")


def change_filter_visibility_hidden_column(path: Path) -> Path:
    """Reveal an effective hidden column range without changing cell content."""

    def mutate(contents: dict[str, bytes]) -> None:
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = report_xml.find(cols_tag)
        if columns is None:
            raise ValueError("Could not find column-control fixture")
        hidden_range = next(
            column
            for column in columns.findall(col_tag)
            if column.get("min") == "2" and column.get("max") == "5"
        )
        hidden_range.set("hidden", "false")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".filter-visibility-column.tmp.xlsx")


def change_table_filter_visibility_criterion(path: Path) -> Path:
    """Change a table-part filter member without changing table cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        custom_filter_tag = f"{{{_SPREADSHEETML_NS}}}customFilter"
        table_xml = ElementTree.fromstring(contents["xl/tables/table1.xml"])
        next(table_xml.iter(custom_filter_tag)).set("val", "CANDIDATE-PRIVATE-TABLE-SEGMENT")
        contents["xl/tables/table1.xml"] = ElementTree.tostring(
            table_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".table-filter-visibility-criterion.tmp.xlsx")


def normalize_filter_visibility_control_spelling(path: Path) -> Path:
    """Use equivalent local-reference, Boolean, and unsigned-integer spellings."""

    def mutate(contents: dict[str, bytes]) -> None:
        auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
        filter_column_tag = f"{{{_SPREADSHEETML_NS}}}filterColumn"
        filters_tag = f"{{{_SPREADSHEETML_NS}}}filters"
        sort_state_tag = f"{{{_SPREADSHEETML_NS}}}sortState"
        sort_condition_tag = f"{{{_SPREADSHEETML_NS}}}sortCondition"
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        row_tag = f"{{{_SPREADSHEETML_NS}}}row"
        sheet_format_tag = f"{{{_SPREADSHEETML_NS}}}sheetFormatPr"

        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        auto_filter = report_xml.find(auto_filter_tag)
        if auto_filter is None:
            raise ValueError("Could not find worksheet AutoFilter fixture")
        auto_filter.set("ref", "$a$1:$c$5")
        filter_column = auto_filter.find(filter_column_tag)
        if filter_column is None:
            raise ValueError("Could not find worksheet filter-column fixture")
        filter_column.set("colId", "01")
        filter_column.set("hiddenButton", "false")
        filter_column.set("showButton", "1")
        filters = filter_column.find(filters_tag)
        if filters is None:
            raise ValueError("Could not find worksheet filters fixture")
        filters.set("blank", "0")
        worksheet_sort = auto_filter.find(sort_state_tag)
        if worksheet_sort is None:
            raise ValueError("Could not find worksheet sort-state fixture")
        worksheet_sort.set("ref", "$A$1:$C$5")
        worksheet_sort.set("caseSensitive", "0")
        worksheet_condition = worksheet_sort.find(sort_condition_tag)
        if worksheet_condition is None:
            raise ValueError("Could not find worksheet sort-condition fixture")
        worksheet_condition.set("ref", "$c$2:$c$5")
        worksheet_condition.set("descending", "true")
        rows = {row.get("r"): row for row in report_xml.iter(row_tag)}
        rows["3"].set("hidden", "true")
        rows["3"].set("outlineLevel", "01")
        rows["4"].set("hidden", "1")
        rows["4"].set("outlineLevel", "002")
        rows["4"].set("collapsed", "true")
        columns = report_xml.find(cols_tag)
        if columns is None:
            raise ValueError("Could not find column-control fixture")
        for column in list(columns):
            columns.remove(column)
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "02", "max": "02", "hidden": "true", "outlineLevel": "01"},
        )
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "003", "max": "003", "hidden": "1", "outlineLevel": "1"},
        )
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "04", "max": "05", "hidden": "true", "outlineLevel": "01"},
        )
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "3", "max": "3", "hidden": "0", "outlineLevel": "02"},
        )
        ElementTree.SubElement(
            columns,
            col_tag,
            {"min": "006", "max": "006", "collapsed": "1"},
        )
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

        table_xml = ElementTree.fromstring(contents["xl/tables/table1.xml"])
        table_auto_filter = table_xml.find(auto_filter_tag)
        if table_auto_filter is None:
            raise ValueError("Could not find table AutoFilter fixture")
        table_auto_filter.set("ref", "$A$1:$C$5")
        table_column = table_auto_filter.find(filter_column_tag)
        if table_column is None:
            raise ValueError("Could not find table filter-column fixture")
        table_column.set("colId", "00")
        table_sort = table_auto_filter.find(sort_state_tag)
        if table_sort is None:
            raise ValueError("Could not find table sort-state fixture")
        table_sort.set("ref", "$a$1:$c$5")
        table_condition = table_sort.find(sort_condition_tag)
        if table_condition is None:
            raise ValueError("Could not find table sort-condition fixture")
        table_condition.set("ref", "$B$2:$B$5")
        table_condition.set("descending", "0")
        contents["xl/tables/table1.xml"] = ElementTree.tostring(
            table_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

        default_hidden_xml = ElementTree.fromstring(contents["xl/worksheets/sheet3.xml"])
        sheet_format = default_hidden_xml.find(sheet_format_tag)
        if sheet_format is None:
            raise ValueError("Could not find default-hidden sheet-format fixture")
        sheet_format.set("zeroHeight", "1")
        default_hidden_row = next(default_hidden_xml.iter(row_tag))
        default_hidden_row.set("hidden", "0")
        contents["xl/worksheets/sheet3.xml"] = ElementTree.tostring(
            default_hidden_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".filter-visibility-noise.tmp.xlsx")


def corrupt_filter_visibility_control(path: Path) -> Path:
    """Inject an invalid filter-column identifier to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
        filter_column_tag = f"{{{_SPREADSHEETML_NS}}}filterColumn"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        auto_filter = report_xml.find(auto_filter_tag)
        if auto_filter is None:
            raise ValueError("Could not find worksheet AutoFilter fixture")
        filter_column = auto_filter.find(filter_column_tag)
        if filter_column is None:
            raise ValueError("Could not find worksheet filter-column fixture")
        filter_column.set("colId", "4294967296")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".filter-visibility-corrupt.tmp.xlsx")


def corrupt_filter_visibility_column_control(path: Path) -> Path:
    """Inject an out-of-bounds column range to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        report_xml = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = report_xml.find(cols_tag)
        if columns is None:
            raise ValueError("Could not find column-control fixture")
        hidden_range = next(
            column
            for column in columns.findall(col_tag)
            if column.get("min") == "2" and column.get("max") == "5"
        )
        hidden_range.set("max", "16385")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            report_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".filter-visibility-column-corrupt.tmp.xlsx")


def make_number_format_model(path: Path) -> Path:
    """Create display-only number-format controls with private format codes."""
    workbook = Workbook()
    report = workbook.active
    report.title = "Number Format Report"
    report["A1"] = "Built-in display"
    report["A2"] = 1234.5
    report["A2"].number_format = "0.00"
    report["B1"] = "Private literal display"
    report["B2"] = 1234567.89
    report["B2"].number_format = '"PRIVATE-BASELINE-NUMBER-FORMAT"'
    report["C1"] = "Hidden display"
    report["C2"] = 0.125
    report["C2"].number_format = ";;;"
    report["D1"] = "Formula without a direct style"
    report["D2"] = "=B2*C2"
    report.row_dimensions[4].number_format = '0.0,," M"'
    report.column_dimensions["D"].number_format = "0.0%"
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "4"
        )
        # A column style applies to unallocated/new cells. Use a short span so
        # the raw scanner exercises effective-range canonicalization without
        # claiming that the existing formula cell adopts the column default.
        column.set("max", "5")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".number-format.tmp.xlsx")


def change_number_format_code(path: Path) -> Path:
    """Change a private direct-cell format code without touching its value."""

    def mutate(contents: dict[str, bytes]) -> None:
        num_fmt_tag = f"{{{_SPREADSHEETML_NS}}}numFmt"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        number_format = next(
            current
            for current in styles.iter(num_fmt_tag)
            if current.get("formatCode") == '"PRIVATE-BASELINE-NUMBER-FORMAT"'
        )
        number_format.set("formatCode", '"CANDIDATE-PRIVATE-NUMBER-FORMAT"')
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".number-format-change.tmp.xlsx")


def change_number_format_default_style(path: Path) -> Path:
    """Change the workbook's base cell format without touching cell records."""

    def mutate(contents: dict[str, bytes]) -> None:
        cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        cell_xfs = styles.find(cell_xfs_tag)
        if cell_xfs is None:
            raise ValueError("Could not find number-format cell-XF fixture")
        default_xf = next(cell_xfs.iter(xf_tag))
        default_xf.set("numFmtId", "165")
        default_xf.set("applyNumberFormat", "true")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".number-format-default-change.tmp.xlsx")


def normalize_number_format_control_spelling(path: Path) -> Path:
    """Renumber custom formats and split an equivalent column-style range."""

    def mutate(contents: dict[str, bytes]) -> None:
        num_fmt_tag = f"{{{_SPREADSHEETML_NS}}}numFmt"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        custom_formats = [
            current
            for current in styles.iter(num_fmt_tag)
            if (identifier := current.get("numFmtId")) is not None
            and int(identifier) >= 164
        ]
        remapping = {
            current.get("numFmtId"): str(246 + index)
            for index, current in enumerate(custom_formats)
            if current.get("numFmtId") is not None
        }
        for number_format in custom_formats:
            identifier = number_format.get("numFmtId")
            if identifier is not None:
                number_format.set("numFmtId", remapping[identifier])
        for xf in styles.iter(xf_tag):
            identifier = xf.get("numFmtId")
            if identifier in remapping:
                xf.set("numFmtId", remapping[identifier])
                xf.set("applyNumberFormat", "true")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = worksheet.find(cols_tag)
        if columns is None:
            raise ValueError("Could not find number-format column fixture")
        column = next(
            current
            for current in columns.findall(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        attributes = dict(column.attrib)
        columns.remove(column)
        for minimum, maximum in (("4", "4"), ("5", "5")):
            split_attributes = {**attributes, "min": minimum, "max": maximum}
            ElementTree.SubElement(columns, col_tag, split_attributes)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".number-format-noise.tmp.xlsx")


def normalize_number_format_inheritance(path: Path) -> Path:
    """Move one direct format into its base XF without changing its effect."""

    def mutate(contents: dict[str, bytes]) -> None:
        num_fmt_tag = f"{{{_SPREADSHEETML_NS}}}numFmt"
        cell_style_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellStyleXfs"
        cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        custom_format = next(
            current
            for current in styles.iter(num_fmt_tag)
            if current.get("formatCode") == '"PRIVATE-BASELINE-NUMBER-FORMAT"'
        )
        identifier = custom_format.get("numFmtId")
        if identifier is None:
            raise ValueError("Could not find private number-format identifier")
        cell_style_xfs = styles.find(cell_style_xfs_tag)
        cell_xfs = styles.find(cell_xfs_tag)
        if cell_style_xfs is None or cell_xfs is None:
            raise ValueError("Could not find number-format XF fixture")
        ElementTree.SubElement(
            cell_style_xfs,
            xf_tag,
            {
                "numFmtId": identifier,
                "fontId": "0",
                "fillId": "0",
                "borderId": "0",
                "applyNumberFormat": "true",
            },
        )
        direct_xf = next(
            current
            for current in cell_xfs.findall(xf_tag)
            if current.get("numFmtId") == identifier
        )
        direct_xf.attrib.pop("numFmtId", None)
        direct_xf.set("xfId", "1")
        direct_xf.set("applyNumberFormat", "false")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".number-format-inheritance.tmp.xlsx")


def corrupt_number_format_column_control(path: Path) -> Path:
    """Inject an out-of-bounds format-style span to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        column.set("max", "16385")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".number-format-corrupt.tmp.xlsx")


def corrupt_number_format_definition(path: Path) -> Path:
    """Leave a direct custom format pointing at a missing format definition."""

    def mutate(contents: dict[str, bytes]) -> None:
        num_fmt_tag = f"{{{_SPREADSHEETML_NS}}}numFmt"
        cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        number_format = next(
            current
            for current in styles.iter(num_fmt_tag)
            if current.get("formatCode") == '"PRIVATE-BASELINE-NUMBER-FORMAT"'
        )
        identifier = number_format.get("numFmtId")
        cell_xfs = styles.find(cell_xfs_tag)
        if identifier is None or cell_xfs is None:
            raise ValueError("Could not find number-format definition fixture")
        direct_xf = next(
            current
            for current in cell_xfs.findall(xf_tag)
            if current.get("numFmtId") == identifier
        )
        direct_xf.set("numFmtId", "999987")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".number-format-missing-definition.tmp.xlsx")


def make_font_model(path: Path) -> Path:
    """Create display-only font controls with private names and colours."""
    workbook = Workbook()
    report = workbook.active
    report.title = "Font Report"
    report["A1"] = "Default display"
    report["A2"] = 1234.5
    report["B1"] = "Private direct font"
    report["B2"] = 1234567.89
    report["B2"].font = Font(
        name="PRIVATE-BASELINE-FONT",
        color="FF112233",
        bold=True,
    )
    report["C1"] = "Hidden-looking display"
    report["C2"] = 0.125
    report["C2"].font = Font(name="PRIVATE-WHITE-FONT", color="FFFFFFFF")
    report["D1"] = "Formula without a direct font"
    report["D2"] = "=B2*C2"
    report.row_dimensions[4].font = Font(
        name="PRIVATE-ROW-FONT",
        color="FF445566",
        italic=True,
    )
    report.column_dimensions["D"].font = Font(
        name="PRIVATE-COLUMN-FONT",
        color="FF778899",
        underline="single",
    )
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "4"
        )
        # Column fonts are OOXML defaults for unallocated/new cells. Keep the
        # fixture span short so range canonicalization is testable without
        # claiming that an allocated formula cell adopts the default.
        column.set("max", "5")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".font.tmp.xlsx")


def change_font_definition(path: Path) -> Path:
    """Change a private direct-cell font without touching its value."""

    def mutate(contents: dict[str, bytes]) -> None:
        font_tag = f"{{{_SPREADSHEETML_NS}}}font"
        name_tag = f"{{{_SPREADSHEETML_NS}}}name"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        font = next(
            current
            for current in styles.iter(font_tag)
            if (name := current.find(name_tag)) is not None
            and name.get("val") == "PRIVATE-BASELINE-FONT"
        )
        color = font.find(color_tag)
        if color is None:
            raise ValueError("Could not find private font colour fixture")
        color.attrib.clear()
        color.set("rgb", "FFABCDEF")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".font-change.tmp.xlsx")


def change_default_font_definition(path: Path) -> Path:
    """Change the default font record without touching any cell record."""

    def mutate(contents: dict[str, bytes]) -> None:
        fonts_tag = f"{{{_SPREADSHEETML_NS}}}fonts"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fonts = styles.find(fonts_tag)
        if fonts is None or not list(fonts):
            raise ValueError("Could not find default font fixture")
        default_font = list(fonts)[0]
        color = default_font.find(color_tag)
        if color is None:
            color = ElementTree.SubElement(default_font, color_tag)
        color.attrib.clear()
        color.set("rgb", "FFFAFAFA")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".font-default-change.tmp.xlsx")


def normalize_font_control_spelling(path: Path) -> Path:
    """Renumber fonts, normalize booleans, and split one column-style range."""

    def mutate(contents: dict[str, bytes]) -> None:
        fonts_tag = f"{{{_SPREADSHEETML_NS}}}fonts"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        bold_tag = f"{{{_SPREADSHEETML_NS}}}b"
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fonts = styles.find(fonts_tag)
        if fonts is None:
            raise ValueError("Could not find font fixture")
        original_fonts = list(fonts)
        remapping = {
            str(index): str(len(original_fonts) - index - 1)
            for index in range(len(original_fonts))
        }
        for font in original_fonts:
            fonts.remove(font)
        for font in reversed(original_fonts):
            # Writer child order is not semantic. Also spell a Boolean that
            # XlsxWriter emits without ``val`` in a different valid form.
            children = list(font)
            for child in children:
                font.remove(child)
            for child in reversed(children):
                font.append(child)
            bold = font.find(bold_tag)
            if bold is not None:
                bold.set("val", "true")
            fonts.append(font)
        for xf in styles.iter(xf_tag):
            identifier = xf.get("fontId")
            if identifier in remapping:
                xf.set("fontId", remapping[identifier])
                xf.set("applyFont", "1")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = worksheet.find(cols_tag)
        if columns is None:
            raise ValueError("Could not find font column fixture")
        column = next(
            current
            for current in columns.findall(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        attributes = dict(column.attrib)
        columns.remove(column)
        for minimum, maximum in (("4", "4"), ("5", "5")):
            split_attributes = {**attributes, "min": minimum, "max": maximum}
            ElementTree.SubElement(columns, col_tag, split_attributes)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".font-noise.tmp.xlsx")


def normalize_font_inheritance(path: Path) -> Path:
    """Move one direct font into its base XF without changing its effect."""

    def mutate(contents: dict[str, bytes]) -> None:
        fonts_tag = f"{{{_SPREADSHEETML_NS}}}fonts"
        font_tag = f"{{{_SPREADSHEETML_NS}}}font"
        name_tag = f"{{{_SPREADSHEETML_NS}}}name"
        cell_style_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellStyleXfs"
        cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fonts = styles.find(fonts_tag)
        cell_style_xfs = styles.find(cell_style_xfs_tag)
        cell_xfs = styles.find(cell_xfs_tag)
        if fonts is None or cell_style_xfs is None or cell_xfs is None:
            raise ValueError("Could not find font XF fixture")
        font_index = next(
            index
            for index, current in enumerate(fonts.findall(font_tag))
            if (name := current.find(name_tag)) is not None
            and name.get("val") == "PRIVATE-BASELINE-FONT"
        )
        ElementTree.SubElement(
            cell_style_xfs,
            xf_tag,
            {
                "numFmtId": "0",
                "fontId": str(font_index),
                "fillId": "0",
                "borderId": "0",
                "applyFont": "true",
            },
        )
        direct_xf = next(
            current
            for current in cell_xfs.findall(xf_tag)
            if current.get("fontId") == str(font_index)
        )
        direct_xf.attrib.pop("fontId", None)
        direct_xf.set("xfId", "1")
        direct_xf.set("applyFont", "false")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".font-inheritance.tmp.xlsx")


def corrupt_font_column_control(path: Path) -> Path:
    """Inject an out-of-bounds font-style span to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        column.set("max", "16385")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".font-corrupt.tmp.xlsx")


def corrupt_font_definition(path: Path) -> Path:
    """Leave a direct font assignment pointing at a missing font definition."""

    def mutate(contents: dict[str, bytes]) -> None:
        fonts_tag = f"{{{_SPREADSHEETML_NS}}}fonts"
        font_tag = f"{{{_SPREADSHEETML_NS}}}font"
        name_tag = f"{{{_SPREADSHEETML_NS}}}name"
        cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fonts = styles.find(fonts_tag)
        cell_xfs = styles.find(cell_xfs_tag)
        if fonts is None or cell_xfs is None:
            raise ValueError("Could not find font definition fixture")
        font_index = next(
            index
            for index, current in enumerate(fonts.findall(font_tag))
            if (name := current.find(name_tag)) is not None
            and name.get("val") == "PRIVATE-BASELINE-FONT"
        )
        direct_xf = next(
            current
            for current in cell_xfs.findall(xf_tag)
            if current.get("fontId") == str(font_index)
        )
        direct_xf.set("fontId", "999999999")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".font-missing-definition.tmp.xlsx")


def make_fill_model(path: Path) -> Path:
    """Create display-only fill controls with private colours and gradients."""
    workbook = Workbook()
    report = workbook.active
    report.title = "Fill Report"
    report["A1"] = "Default display"
    report["A2"] = 1234.5
    report["B1"] = "Private direct fill"
    report["B2"] = 1234567.89
    report["B2"].fill = PatternFill(
        fill_type="solid",
        fgColor="FF112233",
        bgColor="FF445566",
    )
    report["C1"] = "Hidden-looking display"
    report["C2"] = 0.125
    report["C2"].fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
    report["D1"] = "Gradient display"
    report["D2"] = "=B2*C2"
    report["D2"].fill = GradientFill(
        type="linear",
        degree=45,
        stop=("FF102030", "FF405060"),
    )
    report.row_dimensions[4].fill = PatternFill(
        fill_type="solid",
        fgColor="FF667788",
    )
    report.column_dimensions["D"].fill = PatternFill(
        fill_type="darkGrid",
        fgColor="FF99AABB",
        bgColor="FFCCDDEE",
    )
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "4"
        )
        # Column fills are OOXML defaults for unallocated/new cells. Keep the
        # fixture span short so range canonicalization is testable without
        # claiming that an allocated formula cell adopts the default.
        column.set("max", "5")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".fill.tmp.xlsx")


def change_fill_definition(path: Path) -> Path:
    """Change a private direct-cell fill without touching its value."""

    def mutate(contents: dict[str, bytes]) -> None:
        fill_tag = f"{{{_SPREADSHEETML_NS}}}fill"
        pattern_fill_tag = f"{{{_SPREADSHEETML_NS}}}patternFill"
        fg_color_tag = f"{{{_SPREADSHEETML_NS}}}fgColor"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fill = next(
            current
            for current in styles.iter(fill_tag)
            if (pattern_fill := current.find(pattern_fill_tag)) is not None
            and (colour := pattern_fill.find(fg_color_tag)) is not None
            and colour.get("rgb") == "FF112233"
        )
        pattern_fill = fill.find(pattern_fill_tag)
        if pattern_fill is None:
            raise ValueError("Could not find private fill pattern fixture")
        colour = pattern_fill.find(fg_color_tag)
        if colour is None:
            raise ValueError("Could not find private fill colour fixture")
        colour.attrib.clear()
        colour.set("rgb", "FFABCDEF")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".fill-change.tmp.xlsx")


def change_gradient_fill_definition(path: Path) -> Path:
    """Change a gradient direction without touching any cell value or formula."""

    def mutate(contents: dict[str, bytes]) -> None:
        gradient_fill_tag = f"{{{_SPREADSHEETML_NS}}}gradientFill"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        gradient = next(
            current
            for current in styles.iter(gradient_fill_tag)
            if current.get("degree") == "45"
        )
        gradient.set("degree", "90")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".gradient-fill-change.tmp.xlsx")


def change_default_fill_definition(path: Path) -> Path:
    """Change the default fill record without touching any cell record."""

    def mutate(contents: dict[str, bytes]) -> None:
        fills_tag = f"{{{_SPREADSHEETML_NS}}}fills"
        pattern_fill_tag = f"{{{_SPREADSHEETML_NS}}}patternFill"
        fg_color_tag = f"{{{_SPREADSHEETML_NS}}}fgColor"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fills = styles.find(fills_tag)
        if fills is None or not list(fills):
            raise ValueError("Could not find default fill fixture")
        default_fill = list(fills)[0]
        pattern_fill = default_fill.find(pattern_fill_tag)
        if pattern_fill is None:
            pattern_fill = ElementTree.SubElement(default_fill, pattern_fill_tag)
        pattern_fill.set("patternType", "solid")
        colour = pattern_fill.find(fg_color_tag)
        if colour is None:
            colour = ElementTree.SubElement(pattern_fill, fg_color_tag)
        colour.attrib.clear()
        colour.set("rgb", "FFFAFAFA")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".fill-default-change.tmp.xlsx")


def normalize_fill_control_spelling(path: Path) -> Path:
    """Renumber fills, reorder pattern colours, and split one column span."""

    def mutate(contents: dict[str, bytes]) -> None:
        fills_tag = f"{{{_SPREADSHEETML_NS}}}fills"
        pattern_fill_tag = f"{{{_SPREADSHEETML_NS}}}patternFill"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fills = styles.find(fills_tag)
        if fills is None:
            raise ValueError("Could not find fill fixture")
        original_fills = list(fills)
        remapping = {
            str(index): str(len(original_fills) - index - 1)
            for index in range(len(original_fills))
        }
        for fill in original_fills:
            fills.remove(fill)
        for fill in reversed(original_fills):
            # The schema orders foreground/background colours, but their order
            # carries no rendering meaning once their names are known.
            pattern_fill = fill.find(pattern_fill_tag)
            if pattern_fill is not None:
                children = list(pattern_fill)
                for child in children:
                    pattern_fill.remove(child)
                for child in reversed(children):
                    pattern_fill.append(child)
            fills.append(fill)
        for xf in styles.iter(xf_tag):
            identifier = xf.get("fillId")
            if identifier in remapping:
                xf.set("fillId", remapping[identifier])
                xf.set("applyFill", "1")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = worksheet.find(cols_tag)
        if columns is None:
            raise ValueError("Could not find fill column fixture")
        column = next(
            current
            for current in columns.findall(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        attributes = dict(column.attrib)
        columns.remove(column)
        for minimum, maximum in (("4", "4"), ("5", "5")):
            split_attributes = {**attributes, "min": minimum, "max": maximum}
            ElementTree.SubElement(columns, col_tag, split_attributes)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".fill-noise.tmp.xlsx")


def normalize_fill_inert_pattern_declarations(path: Path) -> Path:
    """Change valid colour declarations that do not affect the rendered fill."""

    def mutate(contents: dict[str, bytes]) -> None:
        fills_tag = f"{{{_SPREADSHEETML_NS}}}fills"
        fill_tag = f"{{{_SPREADSHEETML_NS}}}fill"
        pattern_fill_tag = f"{{{_SPREADSHEETML_NS}}}patternFill"
        fg_color_tag = f"{{{_SPREADSHEETML_NS}}}fgColor"
        bg_color_tag = f"{{{_SPREADSHEETML_NS}}}bgColor"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fills = styles.find(fills_tag)
        if fills is None or not list(fills):
            raise ValueError("Could not find fill inert-declaration fixture")
        default_fill = list(fills)[0]
        default_pattern = default_fill.find(pattern_fill_tag)
        if default_pattern is None:
            raise ValueError("Could not find default no-fill pattern fixture")
        ElementTree.SubElement(default_pattern, fg_color_tag, {"rgb": "FF010203"})
        private_fill = next(
            current
            for current in fills.findall(fill_tag)
            if (pattern_fill := current.find(pattern_fill_tag)) is not None
            and (colour := pattern_fill.find(fg_color_tag)) is not None
            and colour.get("rgb") == "FF112233"
        )
        private_pattern = private_fill.find(pattern_fill_tag)
        if private_pattern is None:
            raise ValueError("Could not find private solid fill fixture")
        background = private_pattern.find(bg_color_tag)
        if background is None:
            background = ElementTree.SubElement(private_pattern, bg_color_tag)
        background.attrib.clear()
        background.set("rgb", "FFABCDEF")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".fill-inert-noise.tmp.xlsx")


def normalize_fill_inheritance(path: Path) -> Path:
    """Move one direct fill into its base XF without changing its effect."""

    def mutate(contents: dict[str, bytes]) -> None:
        fills_tag = f"{{{_SPREADSHEETML_NS}}}fills"
        fill_tag = f"{{{_SPREADSHEETML_NS}}}fill"
        pattern_fill_tag = f"{{{_SPREADSHEETML_NS}}}patternFill"
        fg_color_tag = f"{{{_SPREADSHEETML_NS}}}fgColor"
        cell_style_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellStyleXfs"
        cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fills = styles.find(fills_tag)
        cell_style_xfs = styles.find(cell_style_xfs_tag)
        cell_xfs = styles.find(cell_xfs_tag)
        if fills is None or cell_style_xfs is None or cell_xfs is None:
            raise ValueError("Could not find fill XF fixture")
        fill_index = next(
            index
            for index, current in enumerate(fills.findall(fill_tag))
            if (pattern_fill := current.find(pattern_fill_tag)) is not None
            and (colour := pattern_fill.find(fg_color_tag)) is not None
            and colour.get("rgb") == "FF112233"
        )
        ElementTree.SubElement(
            cell_style_xfs,
            xf_tag,
            {
                "numFmtId": "0",
                "fontId": "0",
                "fillId": str(fill_index),
                "borderId": "0",
                "applyFill": "true",
            },
        )
        direct_xf = next(
            current
            for current in cell_xfs.findall(xf_tag)
            if current.get("fillId") == str(fill_index)
        )
        direct_xf.attrib.pop("fillId", None)
        direct_xf.set("xfId", "1")
        direct_xf.set("applyFill", "false")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".fill-inheritance.tmp.xlsx")


def corrupt_fill_column_control(path: Path) -> Path:
    """Inject an out-of-bounds fill-style span to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        column.set("max", "16385")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".fill-corrupt.tmp.xlsx")


def corrupt_fill_definition(path: Path) -> Path:
    """Leave a direct fill assignment pointing at a missing fill definition."""

    def mutate(contents: dict[str, bytes]) -> None:
        fills_tag = f"{{{_SPREADSHEETML_NS}}}fills"
        fill_tag = f"{{{_SPREADSHEETML_NS}}}fill"
        pattern_fill_tag = f"{{{_SPREADSHEETML_NS}}}patternFill"
        fg_color_tag = f"{{{_SPREADSHEETML_NS}}}fgColor"
        cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        fills = styles.find(fills_tag)
        cell_xfs = styles.find(cell_xfs_tag)
        if fills is None or cell_xfs is None:
            raise ValueError("Could not find fill definition fixture")
        fill_index = next(
            index
            for index, current in enumerate(fills.findall(fill_tag))
            if (pattern_fill := current.find(pattern_fill_tag)) is not None
            and (colour := pattern_fill.find(fg_color_tag)) is not None
            and colour.get("rgb") == "FF112233"
        )
        direct_xf = next(
            current
            for current in cell_xfs.findall(xf_tag)
            if current.get("fillId") == str(fill_index)
        )
        direct_xf.set("fillId", "999999999")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".fill-missing-definition.tmp.xlsx")


def make_alignment_model(path: Path) -> Path:
    """Create display-only alignment controls with private presentation values."""
    workbook = Workbook()
    report = workbook.active
    report.title = "Alignment Report"
    report["A1"] = "Default display"
    report["A2"] = 1234.5
    report["B1"] = "Private indented display"
    report["B2"] = 1234567.89
    report["B2"].alignment = Alignment(
        horizontal="left",
        indent=157,
        wrapText=True,
    )
    report["C1"] = "Private rotated display"
    report["C2"] = "PRIVATE-ROTATED-REVIEW-TEXT"
    report["C2"].alignment = Alignment(
        horizontal="right",
        textRotation=255,
        shrinkToFit=True,
        readingOrder=2,
    )
    report["D1"] = "Column-default display"
    report["D2"] = "=B2*A2"
    report["A4"] = "Row-default display"
    report.row_dimensions[4].alignment = Alignment(
        vertical="top",
        wrapText=True,
    )
    report.column_dimensions["D"].alignment = Alignment(
        horizontal="center",
        relativeIndent=7,
    )
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "4"
        )
        # Column alignment is a default for unallocated/new cells. Keep the
        # fixture span short so range canonicalization is testable without
        # claiming an allocated formula cell adopts that default.
        column.set("max", "5")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".alignment.tmp.xlsx")


def _alignment_fixture_xf(styles: ElementTree.Element) -> ElementTree.Element:
    """Return the direct indented alignment XF used by alignment fixtures."""
    cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
    xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
    alignment_tag = f"{{{_SPREADSHEETML_NS}}}alignment"
    cell_xfs = styles.find(cell_xfs_tag)
    if cell_xfs is None:
        raise ValueError("Could not find alignment XF fixture")
    direct_xf = next(
        (
            current
            for current in cell_xfs.findall(xf_tag)
            if (
                (alignment := current.find(alignment_tag)) is not None
                and alignment.get("horizontal") == "left"
                and alignment.get("indent") in {"157", "157.0"}
            )
        ),
        None,
    )
    if direct_xf is None:
        raise ValueError("Could not find direct alignment fixture")
    return direct_xf


def change_alignment_definition(path: Path) -> Path:
    """Change a private direct-cell alignment without touching its value."""

    def mutate(contents: dict[str, bytes]) -> None:
        alignment_tag = f"{{{_SPREADSHEETML_NS}}}alignment"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        alignment = _alignment_fixture_xf(styles).find(alignment_tag)
        if alignment is None:
            raise ValueError("Could not find direct alignment definition")
        alignment.set("indent", "158")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".alignment-change.tmp.xlsx")


def change_default_alignment_definition(path: Path) -> Path:
    """Change the default cell alignment without editing any cell record."""

    def mutate(contents: dict[str, bytes]) -> None:
        cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        alignment_tag = f"{{{_SPREADSHEETML_NS}}}alignment"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        cell_xfs = styles.find(cell_xfs_tag)
        if cell_xfs is None or not cell_xfs.findall(xf_tag):
            raise ValueError("Could not find default alignment XF fixture")
        default_xf = cell_xfs.findall(xf_tag)[0]
        alignment = default_xf.find(alignment_tag)
        if alignment is None:
            alignment = ElementTree.SubElement(default_xf, alignment_tag)
        alignment.set("horizontal", "center")
        default_xf.set("applyAlignment", "true")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".alignment-default-change.tmp.xlsx")


def normalize_alignment_control_spelling(path: Path) -> Path:
    """Use equivalent defaults, Boolean spellings, and column range splitting."""

    def mutate(contents: dict[str, bytes]) -> None:
        alignment_tag = f"{{{_SPREADSHEETML_NS}}}alignment"
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        direct_alignment = _alignment_fixture_xf(styles).find(alignment_tag)
        if direct_alignment is None:
            raise ValueError("Could not find direct alignment definition")
        direct_alignment.set("vertical", "bottom")
        direct_alignment.set("textRotation", "000")
        direct_alignment.set("wrapText", "true")
        direct_alignment.set("indent", "00157")
        direct_alignment.set("relativeIndent", "+000")
        direct_alignment.set("shrinkToFit", "false")
        direct_alignment.set("justifyLastLine", "0")
        direct_alignment.set("readingOrder", "0")
        # mergeCell is valid but inert compatibility material.
        direct_alignment.set("mergeCell", "1")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = worksheet.find(cols_tag)
        if columns is None:
            raise ValueError("Could not find alignment column fixture")
        column = next(
            current
            for current in columns.findall(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        attributes = dict(column.attrib)
        columns.remove(column)
        for minimum, maximum in (("4", "4"), ("5", "5")):
            split_attributes = {**attributes, "min": minimum, "max": maximum}
            ElementTree.SubElement(columns, col_tag, split_attributes)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".alignment-noise.tmp.xlsx")


def normalize_alignment_inheritance(path: Path) -> Path:
    """Move one direct alignment into its base XF without changing its effect."""

    def mutate(contents: dict[str, bytes]) -> None:
        cell_style_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellStyleXfs"
        alignment_tag = f"{{{_SPREADSHEETML_NS}}}alignment"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        cell_style_xfs = styles.find(cell_style_xfs_tag)
        if cell_style_xfs is None:
            raise ValueError("Could not find alignment base XF fixture")
        direct_xf = _alignment_fixture_xf(styles)
        direct_alignment = direct_xf.find(alignment_tag)
        if direct_alignment is None:
            raise ValueError("Could not find direct alignment definition")
        base_index = len(cell_style_xfs.findall(xf_tag))
        base_xf = ElementTree.SubElement(
            cell_style_xfs,
            xf_tag,
            {
                "numFmtId": "0",
                "fontId": "0",
                "fillId": "0",
                "borderId": "0",
                "applyAlignment": "1",
            },
        )
        base_xf.append(
            ElementTree.fromstring(
                ElementTree.tostring(direct_alignment, encoding="utf-8")
            )
        )
        cell_style_xfs.set("count", str(base_index + 1))
        direct_xf.remove(direct_alignment)
        direct_xf.set("xfId", str(base_index))
        direct_xf.set("applyAlignment", "false")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".alignment-inheritance.tmp.xlsx")


def corrupt_alignment_column_control(path: Path) -> Path:
    """Inject an out-of-bounds alignment-style span for fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        column.set("max", "16385")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".alignment-corrupt.tmp.xlsx")


def corrupt_alignment_definition(path: Path) -> Path:
    """Inject invalid reading-order metadata into a used alignment definition."""

    def mutate(contents: dict[str, bytes]) -> None:
        alignment_tag = f"{{{_SPREADSHEETML_NS}}}alignment"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        alignment = _alignment_fixture_xf(styles).find(alignment_tag)
        if alignment is None:
            raise ValueError("Could not find direct alignment definition")
        # openpyxl only checks that readingOrder is nonnegative. SpreadsheetML
        # limits it to 0, 1, or 2, so this remains readable for a fail-closed
        # FormulaFence coverage test.
        alignment.set("readingOrder", "424242")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".alignment-malformed.tmp.xlsx")


def make_border_model(path: Path) -> Path:
    """Create private ordinary cell-border controls without changing values."""
    workbook = Workbook()
    report = workbook.active
    report.title = "Border Report"
    report["A1"] = "Default presentation"
    report["A2"] = 1234.5
    report["B1"] = "Private total boundary"
    report["B2"] = 1234567.89
    report["B2"].border = Border(
        left=Side(style="thin", color="FF112233"),
        bottom=Side(style="double", color="FF445566"),
        diagonal=Side(style="dashed", color="FF778899"),
        diagonalUp=True,
        outline=False,
    )
    report["C1"] = "Private warning boundary"
    report["C2"] = "PRIVATE-BORDER-REVIEW-TEXT"
    report["C2"].border = Border(
        top=Side(style="medium", color="FF99AABB"),
        right=Side(style="thick", color="FFCCDDEE"),
    )
    report["D1"] = "Column-default boundary"
    report["D2"] = "=B2*A2"
    report["A4"] = "Row-default boundary"
    report.row_dimensions[4].border = Border(
        top=Side(style="thin", color="FF102030")
    )
    report.column_dimensions["D"].border = Border(
        left=Side(style="mediumDashed", color="FF304050")
    )
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "4"
        )
        # Column borders are defaults for unallocated/new cells. Keep the span
        # short so range canonicalization is testable without claiming an
        # allocated formula cell adopts that default.
        column.set("max", "5")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".border.tmp.xlsx")


def make_strict_border_model(path: Path) -> Path:
    """Create a strict-SpreadsheetML ordinary-border fixture."""
    make_border_model(path)

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_member = "xl/workbook.xml"
        workbook = ElementTree.fromstring(contents[workbook_member])
        for element in workbook.iter():
            element.tag = strict_name(
                element.tag,
                _SPREADSHEETML_NS,
                _STRICT_SPREADSHEETML_NS,
            )
            attributes = {
                strict_name(
                    name,
                    _DOCUMENT_RELATIONSHIPS_NS,
                    _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                ): value
                for name, value in element.attrib.items()
            }
            element.attrib.clear()
            element.attrib.update(attributes)
        contents[workbook_member] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        for relationship in relationships.findall(relationship_tag):
            if (relationship.get("Type") or "").casefold().endswith("/worksheet"):
                relationship.set(
                    "Type",
                    f"{_STRICT_DOCUMENT_RELATIONSHIPS_NS}/worksheet",
                )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        for member in sorted(
            name
            for name in contents
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ):
            worksheet = ElementTree.fromstring(contents[member])
            for element in worksheet.iter():
                element.tag = strict_name(
                    element.tag,
                    _SPREADSHEETML_NS,
                    _STRICT_SPREADSHEETML_NS,
                )
            contents[member] = ElementTree.tostring(
                worksheet,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".strict-border.tmp.xlsx")


def _border_fixture_definition(styles: ElementTree.Element) -> ElementTree.Element:
    """Return the direct double-bottom border used by border fixtures."""
    borders_tag = f"{{{_SPREADSHEETML_NS}}}borders"
    border_tag = f"{{{_SPREADSHEETML_NS}}}border"
    bottom_tag = f"{{{_SPREADSHEETML_NS}}}bottom"
    borders = styles.find(borders_tag)
    if borders is None:
        raise ValueError("Could not find border fixture definitions")
    definition = next(
        (
            border
            for border in borders.findall(border_tag)
            if (bottom := border.find(bottom_tag)) is not None
            and bottom.get("style") == "double"
        ),
        None,
    )
    if definition is None:
        raise ValueError("Could not find direct border fixture definition")
    return definition


def _border_fixture_xf(styles: ElementTree.Element) -> ElementTree.Element:
    """Return the direct double-bottom border XF used by border fixtures."""
    borders_tag = f"{{{_SPREADSHEETML_NS}}}borders"
    border_tag = f"{{{_SPREADSHEETML_NS}}}border"
    cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
    xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
    definition = _border_fixture_definition(styles)
    borders = styles.find(borders_tag)
    if borders is None:
        raise ValueError("Could not find border fixture definitions")
    border_index = next(
        (
            index
            for index, border in enumerate(borders.findall(border_tag))
            if border is definition
        ),
        None,
    )
    cell_xfs = styles.find(cell_xfs_tag)
    if border_index is None or cell_xfs is None:
        raise ValueError("Could not find border XF fixture")
    direct_xf = next(
        (
            current
            for current in cell_xfs.findall(xf_tag)
            if current.get("borderId") == str(border_index)
        ),
        None,
    )
    if direct_xf is None:
        raise ValueError("Could not find direct border XF fixture")
    return direct_xf


def change_border_definition(path: Path) -> Path:
    """Change a private direct-cell border without touching its value."""

    def mutate(contents: dict[str, bytes]) -> None:
        bottom_tag = f"{{{_SPREADSHEETML_NS}}}bottom"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        bottom = _border_fixture_definition(styles).find(bottom_tag)
        if bottom is None:
            raise ValueError("Could not find direct border definition")
        color = bottom.find(color_tag)
        if color is None:
            color = ElementTree.SubElement(bottom, color_tag)
        color.attrib.clear()
        color.set("rgb", "FF556677")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".border-change.tmp.xlsx")


def change_border_logical_start_side(path: Path) -> Path:
    """Add a raw Office 2010 logical start border without changing a cell value."""

    def mutate(contents: dict[str, bytes]) -> None:
        start_tag = f"{{{_SPREADSHEETML_NS}}}start"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        definition = _border_fixture_definition(styles)
        start = definition.find(start_tag)
        if start is None:
            start = ElementTree.SubElement(definition, start_tag)
        start.attrib.clear()
        start.set("style", "mediumDashDot")
        color = start.find(color_tag)
        if color is None:
            color = ElementTree.SubElement(start, color_tag)
        color.attrib.clear()
        color.set("rgb", "FFABCDEF")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".border-start-change.tmp.xlsx")


def change_default_border_definition(path: Path) -> Path:
    """Change the default cell border without editing any cell record."""

    def mutate(contents: dict[str, bytes]) -> None:
        borders_tag = f"{{{_SPREADSHEETML_NS}}}borders"
        border_tag = f"{{{_SPREADSHEETML_NS}}}border"
        top_tag = f"{{{_SPREADSHEETML_NS}}}top"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        cell_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        borders = styles.find(borders_tag)
        cell_xfs = styles.find(cell_xfs_tag)
        if borders is None or cell_xfs is None or not cell_xfs.findall(xf_tag):
            raise ValueError("Could not find default border fixture")
        default_border = borders.findall(border_tag)[0]
        top = default_border.find(top_tag)
        if top is None:
            top = ElementTree.SubElement(default_border, top_tag)
        top.set("style", "medium")
        color = top.find(color_tag)
        if color is None:
            color = ElementTree.SubElement(top, color_tag)
        color.attrib.clear()
        color.set("rgb", "FF8899AA")
        default_xf = cell_xfs.findall(xf_tag)[0]
        default_xf.set("borderId", "0")
        default_xf.set("applyBorder", "true")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".border-default-change.tmp.xlsx")


def normalize_border_control_spelling(path: Path) -> Path:
    """Use equivalent Boolean, colour, no-border, and range spellings."""

    def mutate(contents: dict[str, bytes]) -> None:
        right_tag = f"{{{_SPREADSHEETML_NS}}}right"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        cols_tag = f"{{{_SPREADSHEETML_NS}}}cols"
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        definition = _border_fixture_definition(styles)
        definition.set("diagonalUp", "true")
        definition.set("outline", "false")
        bottom = definition.find(f"{{{_SPREADSHEETML_NS}}}bottom")
        if bottom is None:
            raise ValueError("Could not find direct border bottom")
        color = bottom.find(color_tag)
        if color is None or color.get("rgb") is None:
            raise ValueError("Could not find direct border colour")
        color.set("rgb", color.get("rgb").lower())
        right = definition.find(right_tag)
        if right is None:
            right = ElementTree.SubElement(definition, right_tag)
        right.set("style", "none")
        right_color = right.find(color_tag)
        if right_color is None:
            right_color = ElementTree.SubElement(right, color_tag)
        right_color.attrib.clear()
        right_color.set("rgb", "FFDECADE")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        columns = worksheet.find(cols_tag)
        if columns is None:
            raise ValueError("Could not find border column fixture")
        column = next(
            current
            for current in columns.findall(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        attributes = dict(column.attrib)
        columns.remove(column)
        for minimum, maximum in (("4", "4"), ("5", "5")):
            split_attributes = {**attributes, "min": minimum, "max": maximum}
            ElementTree.SubElement(columns, col_tag, split_attributes)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".border-noise.tmp.xlsx")


def normalize_border_inert_declarations(path: Path) -> Path:
    """Add valid but visually inert default-border declarations."""

    def mutate(contents: dict[str, bytes]) -> None:
        borders_tag = f"{{{_SPREADSHEETML_NS}}}borders"
        border_tag = f"{{{_SPREADSHEETML_NS}}}border"
        diagonal_tag = f"{{{_SPREADSHEETML_NS}}}diagonal"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        borders = styles.find(borders_tag)
        if borders is None or not borders.findall(border_tag):
            raise ValueError("Could not find default border fixture")
        default_border = borders.findall(border_tag)[0]
        default_border.set("outline", "false")
        default_border.set("diagonalUp", "true")
        diagonal = default_border.find(diagonal_tag)
        if diagonal is None:
            diagonal = ElementTree.SubElement(default_border, diagonal_tag)
        diagonal.attrib.clear()
        color = diagonal.find(color_tag)
        if color is None:
            color = ElementTree.SubElement(diagonal, color_tag)
        color.attrib.clear()
        color.set("rgb", "FFBADC0D")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".border-inert-noise.tmp.xlsx")


def normalize_border_inheritance(path: Path) -> Path:
    """Move one direct border into its base XF without changing its effect."""

    def mutate(contents: dict[str, bytes]) -> None:
        cell_style_xfs_tag = f"{{{_SPREADSHEETML_NS}}}cellStyleXfs"
        xf_tag = f"{{{_SPREADSHEETML_NS}}}xf"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        cell_style_xfs = styles.find(cell_style_xfs_tag)
        if cell_style_xfs is None:
            raise ValueError("Could not find border base XF fixture")
        direct_xf = _border_fixture_xf(styles)
        border_id = direct_xf.get("borderId")
        if border_id is None:
            raise ValueError("Could not find direct border ID")
        base_index = len(cell_style_xfs.findall(xf_tag))
        ElementTree.SubElement(
            cell_style_xfs,
            xf_tag,
            {
                "numFmtId": "0",
                "fontId": "0",
                "fillId": "0",
                "borderId": border_id,
                "applyBorder": "1",
            },
        )
        cell_style_xfs.set("count", str(base_index + 1))
        direct_xf.set("xfId", str(base_index))
        direct_xf.set("applyBorder", "false")
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".border-inheritance.tmp.xlsx")


def corrupt_border_column_control(path: Path) -> Path:
    """Inject an out-of-bounds border-style span for fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        col_tag = f"{{{_SPREADSHEETML_NS}}}col"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        column = next(
            current
            for current in worksheet.iter(col_tag)
            if current.get("min") == "4" and current.get("max") == "5"
        )
        column.set("max", "16385")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".border-corrupt.tmp.xlsx")


def corrupt_border_definition(path: Path) -> Path:
    """Inject an unknown border property into a used definition."""

    def mutate(contents: dict[str, bytes]) -> None:
        bottom_tag = f"{{{_SPREADSHEETML_NS}}}bottom"
        styles = ElementTree.fromstring(contents["xl/styles.xml"])
        bottom = _border_fixture_definition(styles).find(bottom_tag)
        if bottom is None:
            raise ValueError("Could not find direct border definition")
        # Keep the package readable by ordinary spreadsheet readers while
        # exercising FormulaFence's fail-closed handling of a future/unknown
        # border property.
        bottom.set(
            "{urn:formulafence:private-border-test}privateBorderControl",
            "PRIVATE-INVALID-BORDER-METADATA",
        )
        contents["xl/styles.xml"] = ElementTree.tostring(
            styles,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".border-malformed.tmp.xlsx")


def make_worksheet_display_model(path: Path) -> Path:
    """Create private worksheet-view controls without changing stored cells."""
    workbook = Workbook()
    review = workbook.active
    review.title = "Private Display Review"
    review["A1"] = "PRIVATE-DISPLAY-HEADER"
    review["B2"] = 0
    review["C4"] = "PRIVATE-DISPLAY-FOCUS"
    review["D5"] = "=B2+1"
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        sheet_views_tag = f"{{{_SPREADSHEETML_NS}}}sheetViews"
        sheet_view_tag = f"{{{_SPREADSHEETML_NS}}}sheetView"
        pane_tag = f"{{{_SPREADSHEETML_NS}}}pane"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        sheet_views = worksheet.find(sheet_views_tag)
        if sheet_views is None:
            sheet_views = ElementTree.Element(sheet_views_tag)
            worksheet.insert(0, sheet_views)
        sheet_view = sheet_views.find(sheet_view_tag)
        if sheet_view is None:
            sheet_view = ElementTree.SubElement(
                sheet_views,
                sheet_view_tag,
                {"workbookViewId": "0"},
            )
        sheet_view.set("showZeros", "0")
        sheet_view.set("showFormulas", "true")
        sheet_view.set("showGridLines", "false")
        sheet_view.set("defaultGridColor", "false")
        sheet_view.set("colorId", "63")
        sheet_view.set("showRowColHeaders", "0")
        sheet_view.set("showOutlineSymbols", "false")
        sheet_view.set("showRuler", "0")
        sheet_view.set("showWhiteSpace", "false")
        sheet_view.set("rightToLeft", "1")
        sheet_view.set("view", "pageLayout")
        pane = sheet_view.find(pane_tag)
        if pane is None:
            pane = ElementTree.SubElement(sheet_view, pane_tag)
        pane.attrib.clear()
        pane.attrib.update(
            {
                "xSplit": "2.0",
                "ySplit": "3",
                "topLeftCell": "C4",
                "activePane": "bottomRight",
                "state": "frozen",
            }
        )
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-display.tmp.xlsx")


def make_strict_worksheet_display_model(path: Path) -> Path:
    """Create a strict-SpreadsheetML worksheet-display fixture."""
    make_worksheet_display_model(path)

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_member = "xl/workbook.xml"
        workbook = ElementTree.fromstring(contents[workbook_member])
        for element in workbook.iter():
            element.tag = strict_name(
                element.tag,
                _SPREADSHEETML_NS,
                _STRICT_SPREADSHEETML_NS,
            )
            attributes = {
                strict_name(
                    name,
                    _DOCUMENT_RELATIONSHIPS_NS,
                    _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                ): value
                for name, value in element.attrib.items()
            }
            element.attrib.clear()
            element.attrib.update(attributes)
        contents[workbook_member] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        for relationship in relationships.findall(relationship_tag):
            if (relationship.get("Type") or "").casefold().endswith("/worksheet"):
                relationship.set(
                    "Type",
                    f"{_STRICT_DOCUMENT_RELATIONSHIPS_NS}/worksheet",
                )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        for member in sorted(
            name
            for name in contents
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ):
            worksheet = ElementTree.fromstring(contents[member])
            for element in worksheet.iter():
                element.tag = strict_name(
                    element.tag,
                    _SPREADSHEETML_NS,
                    _STRICT_SPREADSHEETML_NS,
                )
            contents[member] = ElementTree.tostring(
                worksheet,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".strict-worksheet-display.tmp.xlsx")


def _worksheet_display_sheet_view(
    contents: dict[str, bytes],
) -> ElementTree.Element:
    """Return the raw worksheet display view used by display-control fixtures."""
    sheet_view_tag = f"{{{_SPREADSHEETML_NS}}}sheetView"
    worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
    sheet_view = worksheet.find(f".//{sheet_view_tag}")
    if sheet_view is None:
        raise ValueError("Could not find worksheet display fixture")
    return sheet_view


def change_worksheet_display_controls(path: Path) -> Path:
    """Reveal stored zeroes without editing their cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        sheet_view = _worksheet_display_sheet_view(contents)
        sheet_view.set("showZeros", "true")
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        replacement = worksheet.find(
            f".//{{{_SPREADSHEETML_NS}}}sheetView"
        )
        if replacement is None:
            raise ValueError("Could not replace worksheet display fixture")
        replacement.attrib.clear()
        replacement.attrib.update(sheet_view.attrib)
        for child in list(replacement):
            replacement.remove(child)
        for child in sheet_view:
            replacement.append(child)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-display-change.tmp.xlsx")


def change_strict_worksheet_display_controls(path: Path) -> Path:
    """Reveal stored zeroes in a strict-SpreadsheetML worksheet view."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        sheet_view = next(
            (
                element
                for element in worksheet.iter()
                if element.tag == f"{{{_STRICT_SPREADSHEETML_NS}}}sheetView"
            ),
            None,
        )
        if sheet_view is None:
            raise ValueError("Could not find strict worksheet display fixture")
        sheet_view.set("showZeros", "true")
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".strict-worksheet-display-change.tmp.xlsx")


def normalize_worksheet_display_control_spelling(path: Path) -> Path:
    """Use equivalent Boolean, decimal, and navigation spellings."""

    def mutate(contents: dict[str, bytes]) -> None:
        sheet_view = _worksheet_display_sheet_view(contents)
        pane_tag = f"{{{_SPREADSHEETML_NS}}}pane"
        selection_tag = f"{{{_SPREADSHEETML_NS}}}selection"
        sheet_view.set("showZeros", "false")
        sheet_view.set("showFormulas", "1")
        sheet_view.set("showGridLines", "0")
        sheet_view.set("defaultGridColor", "0")
        sheet_view.set("colorId", "0063")
        sheet_view.set("showRowColHeaders", "false")
        sheet_view.set("showOutlineSymbols", "0")
        sheet_view.set("showRuler", "false")
        sheet_view.set("showWhiteSpace", "0")
        sheet_view.set("rightToLeft", "true")
        sheet_view.set("topLeftCell", "Z999")
        sheet_view.set("zoomScale", "15")
        pane = sheet_view.find(pane_tag)
        if pane is None:
            raise ValueError("Could not find worksheet display pane fixture")
        pane.set("xSplit", "002.000")
        pane.set("ySplit", "+003.0")
        pane.set("topLeftCell", "Z999")
        selection = ElementTree.SubElement(
            sheet_view,
            selection_tag,
            {"activeCell": "Z999", "sqref": "Z999"},
        )
        selection.tail = "\n"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        replacement = worksheet.find(
            f".//{{{_SPREADSHEETML_NS}}}sheetView"
        )
        if replacement is None:
            raise ValueError("Could not replace worksheet display fixture")
        replacement.attrib.clear()
        replacement.attrib.update(sheet_view.attrib)
        for child in list(replacement):
            replacement.remove(child)
        for child in sheet_view:
            replacement.append(child)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-display-noise.tmp.xlsx")


def corrupt_worksheet_display_control(path: Path) -> Path:
    """Inject a negative raw pane split for fail-closed display parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        sheet_view = _worksheet_display_sheet_view(contents)
        pane_tag = f"{{{_SPREADSHEETML_NS}}}pane"
        pane = sheet_view.find(pane_tag)
        if pane is None:
            raise ValueError("Could not find worksheet display pane fixture")
        pane.set("xSplit", "-987.5")
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        replacement = worksheet.find(
            f".//{{{_SPREADSHEETML_NS}}}sheetView"
        )
        if replacement is None:
            raise ValueError("Could not replace worksheet display fixture")
        replacement.attrib.clear()
        replacement.attrib.update(sheet_view.attrib)
        for child in list(replacement):
            replacement.remove(child)
        for child in sheet_view:
            replacement.append(child)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-display-malformed.tmp.xlsx")


def make_worksheet_print_layout_model(path: Path) -> Path:
    """Create private, persisted worksheet print controls without cell edits."""
    workbook = Workbook()
    review = workbook.active
    review.title = "PRIVATE-PRINT-SHEET"
    review["A1"] = "PRIVATE-PRINT-HEADER"
    review["B2"] = 125
    review["C4"] = "PRIVATE-PRINT-FOOTER"
    review["D5"] = "=B2+1"
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_member = "xl/workbook.xml"
        workbook_xml = ElementTree.fromstring(contents[workbook_member])
        defined_names_tag = f"{{{_SPREADSHEETML_NS}}}definedNames"
        defined_name_tag = f"{{{_SPREADSHEETML_NS}}}definedName"
        defined_names = workbook_xml.find(defined_names_tag)
        if defined_names is None:
            defined_names = ElementTree.Element(defined_names_tag)
            workbook_xml.append(defined_names)
        ElementTree.SubElement(
            defined_names,
            defined_name_tag,
            {"name": "_xlnm.Print_Area", "localSheetId": "0"},
        ).text = "'PRIVATE-PRINT-SHEET'!$A$1:$D$5"
        ElementTree.SubElement(
            defined_names,
            defined_name_tag,
            {"name": "_xlnm.Print_Titles", "localSheetId": "0"},
        ).text = "'PRIVATE-PRINT-SHEET'!$1:$2,'PRIVATE-PRINT-SHEET'!$A:$B"
        contents[workbook_member] = ElementTree.tostring(
            workbook_xml,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        sheet_properties_tag = f"{{{_SPREADSHEETML_NS}}}sheetPr"
        page_setup_properties_tag = f"{{{_SPREADSHEETML_NS}}}pageSetUpPr"
        print_options_tag = f"{{{_SPREADSHEETML_NS}}}printOptions"
        page_margins_tag = f"{{{_SPREADSHEETML_NS}}}pageMargins"
        page_setup_tag = f"{{{_SPREADSHEETML_NS}}}pageSetup"
        header_footer_tag = f"{{{_SPREADSHEETML_NS}}}headerFooter"
        odd_header_tag = f"{{{_SPREADSHEETML_NS}}}oddHeader"
        odd_footer_tag = f"{{{_SPREADSHEETML_NS}}}oddFooter"
        row_breaks_tag = f"{{{_SPREADSHEETML_NS}}}rowBreaks"
        col_breaks_tag = f"{{{_SPREADSHEETML_NS}}}colBreaks"
        break_tag = f"{{{_SPREADSHEETML_NS}}}brk"

        sheet_properties = worksheet.find(sheet_properties_tag)
        if sheet_properties is None:
            sheet_properties = ElementTree.Element(sheet_properties_tag)
            worksheet.insert(0, sheet_properties)
        page_setup_properties = sheet_properties.find(page_setup_properties_tag)
        if page_setup_properties is None:
            page_setup_properties = ElementTree.SubElement(
                sheet_properties,
                page_setup_properties_tag,
            )
        page_setup_properties.set("fitToPage", "true")

        print_options = ElementTree.SubElement(
            worksheet,
            print_options_tag,
            {
                "gridLines": "1",
                "gridLinesSet": "true",
                "headings": "true",
                "horizontalCentered": "1",
                "verticalCentered": "true",
            },
        )
        print_options.tail = "\n"

        page_margins = worksheet.find(page_margins_tag)
        if page_margins is None:
            page_margins = ElementTree.SubElement(worksheet, page_margins_tag)
        page_margins.attrib.clear()
        page_margins.attrib.update(
            {
                "left": "1.25",
                "right": "0.75",
                "top": "1",
                "bottom": "1",
                "header": "0.5",
                "footer": "0.5",
            }
        )

        ElementTree.SubElement(
            worksheet,
            page_setup_tag,
            {
                "orientation": "landscape",
                "paperSize": "9",
                "scale": "80",
                "fitToWidth": "2",
                "fitToHeight": "3",
                "blackAndWhite": "true",
                "draft": "true",
                "cellComments": "atEnd",
                "errors": "blank",
                "useFirstPageNumber": "true",
                "firstPageNumber": "5",
                "usePrinterDefaults": "false",
                "paperHeight": "297mm",
                "paperWidth": "210mm",
            },
        )
        header_footer = ElementTree.SubElement(worksheet, header_footer_tag)
        ElementTree.SubElement(header_footer, odd_header_tag).text = (
            "&CPRIVATE-PRINT-HEADER-TEXT"
        )
        ElementTree.SubElement(header_footer, odd_footer_tag).text = (
            "&RPRIVATE-PRINT-FOOTER-TEXT"
        )
        row_breaks = ElementTree.SubElement(
            worksheet,
            row_breaks_tag,
            {"count": "1", "manualBreakCount": "1"},
        )
        ElementTree.SubElement(
            row_breaks,
            break_tag,
            {"id": "10", "min": "0", "max": "16383", "man": "true"},
        )
        col_breaks = ElementTree.SubElement(
            worksheet,
            col_breaks_tag,
            {"count": "1", "manualBreakCount": "1"},
        )
        ElementTree.SubElement(
            col_breaks,
            break_tag,
            {"id": "4", "min": "0", "max": "1048575", "man": "1"},
        )
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-print-layout.tmp.xlsx")


def _worksheet_print_layout_parts(
    contents: dict[str, bytes],
) -> tuple[ElementTree.Element, ElementTree.Element]:
    """Return the transitional workbook and worksheet print-layout fixture roots."""
    return (
        ElementTree.fromstring(contents["xl/workbook.xml"]),
        ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"]),
    )


def change_worksheet_print_layout_controls(path: Path) -> Path:
    """Change stored print declarations while leaving worksheet cells intact."""

    def mutate(contents: dict[str, bytes]) -> None:
        workbook, worksheet = _worksheet_print_layout_parts(contents)
        defined_name_tag = f"{{{_SPREADSHEETML_NS}}}definedName"
        for definition in workbook.iter(defined_name_tag):
            if definition.get("name") == "_xlnm.Print_Area":
                definition.text = "'PRIVATE-PRINT-SHEET'!$A$1:$B$2"
        print_options = worksheet.find(f"{{{_SPREADSHEETML_NS}}}printOptions")
        page_margins = worksheet.find(f"{{{_SPREADSHEETML_NS}}}pageMargins")
        page_setup = worksheet.find(f"{{{_SPREADSHEETML_NS}}}pageSetup")
        header = worksheet.find(
            f"{{{_SPREADSHEETML_NS}}}headerFooter/{{{_SPREADSHEETML_NS}}}oddHeader"
        )
        row_break = worksheet.find(
            f"{{{_SPREADSHEETML_NS}}}rowBreaks/{{{_SPREADSHEETML_NS}}}brk"
        )
        if any(
            value is None
            for value in (print_options, page_margins, page_setup, header, row_break)
        ):
            raise ValueError("Could not find worksheet print-layout fixture")
        print_options.set("gridLines", "false")
        page_margins.set("left", "2.75")
        page_setup.set("orientation", "portrait")
        page_setup.set("scale", "75")
        header.text = "&CPRIVATE-PRINT-HEADER-CANDIDATE"
        row_break.set("id", "25")
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-print-layout-change.tmp.xlsx")


def normalize_worksheet_print_layout_control_spelling(path: Path) -> Path:
    """Use equivalent print-layout spelling, decimal, and ordering noise."""

    def mutate(contents: dict[str, bytes]) -> None:
        workbook, worksheet = _worksheet_print_layout_parts(contents)
        for definition in workbook.iter(f"{{{_SPREADSHEETML_NS}}}definedName"):
            definition.text = f"  {definition.text or ''}  "
        print_options = worksheet.find(f"{{{_SPREADSHEETML_NS}}}printOptions")
        page_margins = worksheet.find(f"{{{_SPREADSHEETML_NS}}}pageMargins")
        page_setup = worksheet.find(f"{{{_SPREADSHEETML_NS}}}pageSetup")
        page_properties = worksheet.find(
            f"{{{_SPREADSHEETML_NS}}}sheetPr/{{{_SPREADSHEETML_NS}}}pageSetUpPr"
        )
        header_footer = worksheet.find(f"{{{_SPREADSHEETML_NS}}}headerFooter")
        row_breaks = worksheet.find(f"{{{_SPREADSHEETML_NS}}}rowBreaks")
        col_breaks = worksheet.find(f"{{{_SPREADSHEETML_NS}}}colBreaks")
        row_break = worksheet.find(
            f"{{{_SPREADSHEETML_NS}}}rowBreaks/{{{_SPREADSHEETML_NS}}}brk"
        )
        col_break = worksheet.find(
            f"{{{_SPREADSHEETML_NS}}}colBreaks/{{{_SPREADSHEETML_NS}}}brk"
        )
        if any(
            value is None
            for value in (
                print_options,
                page_margins,
                page_setup,
                page_properties,
                header_footer,
                row_breaks,
                col_breaks,
                row_break,
                col_break,
            )
        ):
            raise ValueError("Could not normalize worksheet print-layout fixture")
        print_options.attrib.update(
            {
                "gridLines": "true",
                "gridLinesSet": "1",
                "headings": "1",
                "horizontalCentered": "true",
                "verticalCentered": "1",
            }
        )
        page_margins.attrib.update(
            {
                "left": "1.2500",
                "right": "0.750",
                "top": "1.00",
                "bottom": "1.0",
                "header": ".500",
                "footer": "0.50",
            }
        )
        page_setup.attrib.update(
            {
                "paperSize": "009",
                "scale": "080",
                "fitToWidth": "002",
                "fitToHeight": "003",
                "blackAndWhite": "1",
                "draft": "1",
                "useFirstPageNumber": "true",
                "firstPageNumber": "005",
                "usePrinterDefaults": "0",
                "paperHeight": "297.0mm",
                "paperWidth": "210.00mm",
            }
        )
        page_properties.set("autoPageBreaks", "true")
        header_footer.set("alignWithMargins", "1")
        header_footer.set("scaleWithDoc", "true")
        header_footer.set("differentFirst", "0")
        header_footer.set("differentOddEven", "false")
        row_breaks.attrib.update({"count": "01", "manualBreakCount": "001"})
        row_break.set("id", "0010")
        row_break.set("min", "000")
        row_break.set("max", "016383")
        row_break.set("man", "1")
        col_breaks.attrib.update({"count": "01", "manualBreakCount": "001"})
        col_break.set("id", "0004")
        col_break.set("min", "000")
        col_break.set("max", "01048575")
        col_break.set("man", "true")
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-print-layout-noise.tmp.xlsx")


def normalize_worksheet_print_layout_inert_controls(
    path: Path,
    *,
    automatic_break_id: int,
    fit_to_height: int,
    fit_to_page: bool,
    fit_to_width: int,
    first_page_number: int,
    scale: int,
    show_auto_page_breaks: bool,
    inactive_header_suffix: str,
) -> Path:
    """Vary saved declarations that cannot affect this fixture's print output."""

    def mutate(contents: dict[str, bytes]) -> None:
        _, worksheet = _worksheet_print_layout_parts(contents)
        page_setup = worksheet.find(f"{{{_SPREADSHEETML_NS}}}pageSetup")
        page_properties = worksheet.find(
            f"{{{_SPREADSHEETML_NS}}}sheetPr/{{{_SPREADSHEETML_NS}}}pageSetUpPr"
        )
        header_footer = worksheet.find(f"{{{_SPREADSHEETML_NS}}}headerFooter")
        row_breaks = worksheet.find(f"{{{_SPREADSHEETML_NS}}}rowBreaks")
        if any(
            value is None
            for value in (page_setup, page_properties, header_footer, row_breaks)
        ):
            raise ValueError("Could not normalize inert worksheet print-layout controls")
        page_setup.set("useFirstPageNumber", "false")
        page_setup.set("firstPageNumber", str(first_page_number))
        page_setup.set("scale", str(scale))
        page_setup.set("fitToWidth", str(fit_to_width))
        page_setup.set("fitToHeight", str(fit_to_height))
        page_properties.set("fitToPage", "true" if fit_to_page else "false")
        page_properties.set(
            "autoPageBreaks", "true" if show_auto_page_breaks else "false"
        )
        header_footer.set("differentFirst", "false")
        header_footer.set("differentOddEven", "false")
        ElementTree.SubElement(
            header_footer,
            f"{{{_SPREADSHEETML_NS}}}evenHeader",
        ).text = f"&CPRIVATE-INACTIVE-EVEN-{inactive_header_suffix}"
        ElementTree.SubElement(
            header_footer,
            f"{{{_SPREADSHEETML_NS}}}firstFooter",
        ).text = f"&RPRIVATE-INACTIVE-FIRST-{inactive_header_suffix}"
        ElementTree.SubElement(
            row_breaks,
            f"{{{_SPREADSHEETML_NS}}}brk",
            {
                "id": str(automatic_break_id),
                "min": "0",
                "max": "16383",
                "man": "false",
            },
        )
        row_breaks.set("count", "2")
        row_breaks.set("manualBreakCount", "1")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-print-layout-inert.tmp.xlsx")


def corrupt_worksheet_print_layout_control(path: Path) -> Path:
    """Inject an invalid print scale for fail-closed print-layout parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        _, worksheet = _worksheet_print_layout_parts(contents)
        page_setup = worksheet.find(f"{{{_SPREADSHEETML_NS}}}pageSetup")
        if page_setup is None:
            raise ValueError("Could not find worksheet print-layout page setup")
        page_setup.set("scale", "987654321")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-print-layout-malformed.tmp.xlsx")


def make_strict_worksheet_print_layout_model(path: Path) -> Path:
    """Create a strict-SpreadsheetML worksheet print-layout fixture."""
    make_worksheet_print_layout_model(path)

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_member = "xl/workbook.xml"
        workbook = ElementTree.fromstring(contents[workbook_member])
        for element in workbook.iter():
            element.tag = strict_name(
                element.tag,
                _SPREADSHEETML_NS,
                _STRICT_SPREADSHEETML_NS,
            )
            attributes = {
                strict_name(
                    name,
                    _DOCUMENT_RELATIONSHIPS_NS,
                    _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                ): value
                for name, value in element.attrib.items()
            }
            element.attrib.clear()
            element.attrib.update(attributes)
        contents[workbook_member] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        for relationship in relationships.findall(relationship_tag):
            if (relationship.get("Type") or "").casefold().endswith("/worksheet"):
                relationship.set(
                    "Type",
                    f"{_STRICT_DOCUMENT_RELATIONSHIPS_NS}/worksheet",
                )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        for member in sorted(
            name
            for name in contents
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ):
            worksheet = ElementTree.fromstring(contents[member])
            for element in worksheet.iter():
                element.tag = strict_name(
                    element.tag,
                    _SPREADSHEETML_NS,
                    _STRICT_SPREADSHEETML_NS,
                )
            contents[member] = ElementTree.tostring(
                worksheet,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".strict-worksheet-print-layout.tmp.xlsx")


def change_strict_worksheet_print_layout_controls(path: Path) -> Path:
    """Change a strict-SpreadsheetML saved print control without cell edits."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(contents[worksheet_member])
        page_setup = worksheet.find(
            f"{{{_STRICT_SPREADSHEETML_NS}}}pageSetup"
        )
        if page_setup is None:
            raise ValueError("Could not find strict worksheet print-layout setup")
        page_setup.set("orientation", "portrait")
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".strict-worksheet-print-layout-change.tmp.xlsx")


def _formula_cached_result_cell(
    worksheet: ElementTree.Element,
    coordinate: str,
) -> ElementTree.Element:
    """Return one raw Report formula cell from the formula-cache fixture."""
    cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
    cell = next(
        (
            current
            for current in worksheet.iter(cell_tag)
            if current.get("r") == coordinate
        ),
        None,
    )
    if cell is None:
        raise ValueError(f"Could not find formula-cache fixture cell {coordinate}")
    return cell


def _set_formula_cached_result(
    cell: ElementTree.Element,
    result_type: str | None,
    value: str | None,
) -> None:
    """Set one raw formula result while preserving the formula expression."""
    value_tag = f"{{{_SPREADSHEETML_NS}}}v"
    if result_type is None:
        cell.attrib.pop("t", None)
    else:
        cell.set("t", result_type)
    cached_value = cell.find(value_tag)
    if cached_value is None:
        cached_value = ElementTree.SubElement(cell, value_tag)
    cached_value.text = value


def make_formula_cached_result_model(path: Path) -> Path:
    """Create a manual-calculation workbook with varied private formula caches."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = 100

    report = workbook.create_sheet("Report")
    report["A1"] = "Stored formula result review"
    report["B2"] = "=Inputs!A1*2"
    report["B3"] = '="PRIVATE-CACHED-STRING"'
    report["B4"] = "=TRUE"
    report["B5"] = "=1/0"
    report["B6"] = "=Inputs!A1+1"
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        calc_pr_tag = f"{{{_SPREADSHEETML_NS}}}calcPr"
        workbook_root = ElementTree.fromstring(contents["xl/workbook.xml"])
        calc_pr = workbook_root.find(calc_pr_tag)
        if calc_pr is None:
            calc_pr = ElementTree.SubElement(workbook_root, calc_pr_tag)
        calc_pr.set("calcMode", "manual")
        calc_pr.set("fullCalcOnLoad", "0")
        calc_pr.set("forceFullCalc", "0")
        calc_pr.set("calcOnSave", "0")
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet2.xml"])
        _set_formula_cached_result(
            _formula_cached_result_cell(worksheet, "B2"),
            None,
            "200",
        )
        _set_formula_cached_result(
            _formula_cached_result_cell(worksheet, "B3"),
            "str",
            "PRIVATE-CACHED-STRING",
        )
        _set_formula_cached_result(
            _formula_cached_result_cell(worksheet, "B4"),
            "b",
            "1",
        )
        _set_formula_cached_result(
            _formula_cached_result_cell(worksheet, "B5"),
            "e",
            "#DIV/0!",
        )
        _set_formula_cached_result(
            _formula_cached_result_cell(worksheet, "B6"),
            None,
            None,
        )
        contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".formula-cached-result.tmp.xlsx")


def change_formula_cached_result(path: Path) -> Path:
    """Change only a private formula cache, not a formula or visible input."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet2.xml"])
        _set_formula_cached_result(
            _formula_cached_result_cell(worksheet, "B2"),
            None,
            "999999",
        )
        contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".formula-cached-result-change.tmp.xlsx")


def change_formula_cached_result_with_visible_precedent(path: Path) -> Path:
    """Update a visible input and the matching downstream cached results."""

    def mutate(contents: dict[str, bytes]) -> None:
        value_tag = f"{{{_SPREADSHEETML_NS}}}v"
        inputs = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        input_cell = _formula_cached_result_cell(inputs, "A1")
        input_value = input_cell.find(value_tag)
        if input_value is None:
            input_value = ElementTree.SubElement(input_cell, value_tag)
        input_value.text = "101"
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            inputs,
            encoding="utf-8",
            xml_declaration=True,
        )

        report = ElementTree.fromstring(contents["xl/worksheets/sheet2.xml"])
        _set_formula_cached_result(
            _formula_cached_result_cell(report, "B2"),
            None,
            "202",
        )
        _set_formula_cached_result(
            _formula_cached_result_cell(report, "B6"),
            None,
            "102",
        )
        contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
            report,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".formula-cached-result-recalc.tmp.xlsx")


def normalize_formula_cached_result_spelling(path: Path) -> Path:
    """Use equivalent numeric and Boolean cache spellings."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet2.xml"])
        _set_formula_cached_result(
            _formula_cached_result_cell(worksheet, "B2"),
            None,
            "2.00E2",
        )
        _set_formula_cached_result(
            _formula_cached_result_cell(worksheet, "B4"),
            "b",
            "true",
        )
        contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".formula-cached-result-noise.tmp.xlsx")


def corrupt_formula_cached_result(path: Path) -> Path:
    """Inject an invalid numeric cache to exercise fail-closed coverage."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet2.xml"])
        _set_formula_cached_result(
            _formula_cached_result_cell(worksheet, "B2"),
            None,
            "PRIVATE-NOT-A-NUMBER",
        )
        contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".formula-cached-result-corrupt.tmp.xlsx")


def _threaded_comment_part_names(contents: dict[str, bytes]) -> tuple[str, str, str]:
    """Return the package members used by the threaded-comment fixture."""
    threaded_member = "xl/threadedComments/threadedComment1.xml"
    person_member = "xl/persons/person.xml"
    worksheet_relationships = _relationship_member("xl/worksheets/sheet1.xml")
    if (
        threaded_member not in contents
        or person_member not in contents
        or worksheet_relationships not in contents
    ):
        raise ValueError("Fixture does not contain threaded-comment package parts")
    return threaded_member, person_member, worksheet_relationships


def make_threaded_comment_model(path: Path) -> Path:
    """Create raw modern-comment data outside ordinary worksheet cells."""
    make_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    threaded_comments = (
        "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
    )
    people = threaded_comments

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        threaded_member = "xl/threadedComments/threadedComment1.xml"
        person_member = "xl/persons/person.xml"
        reviewer_id = "{11111111-1111-1111-1111-111111111111}"
        approver_id = "{22222222-2222-2222-2222-222222222222}"
        root_comment_id = "{33333333-3333-3333-3333-333333333333}"
        reply_comment_id = "{44444444-4444-4444-4444-444444444444}"

        threaded_root = ElementTree.Element(f"{{{threaded_comments}}}ThreadedComments")
        root_comment = ElementTree.SubElement(
            threaded_root,
            f"{{{threaded_comments}}}threadedComment",
            {
                "ref": "A1",
                "dT": "2026-07-24T00:00:00Z",
                "personId": reviewer_id,
                "id": root_comment_id,
                "done": "0",
            },
        )
        ElementTree.SubElement(root_comment, f"{{{threaded_comments}}}text").text = (
            "PRIVATE-THREADED-COMMENT-BASELINE"
        )
        mentions = ElementTree.SubElement(root_comment, f"{{{threaded_comments}}}mentions")
        ElementTree.SubElement(
            mentions,
            f"{{{threaded_comments}}}mention",
            {
                "mentionpersonId": approver_id,
                "mentionId": "{55555555-5555-5555-5555-555555555555}",
                "startIndex": "0",
                "length": "7",
            },
        )
        reply = ElementTree.SubElement(
            threaded_root,
            f"{{{threaded_comments}}}threadedComment",
            {
                "ref": "A1",
                "dT": "2026-07-24T00:01:00Z",
                "personId": approver_id,
                "id": reply_comment_id,
                "parentId": root_comment_id,
                "done": "1",
            },
        )
        ElementTree.SubElement(reply, f"{{{threaded_comments}}}text").text = (
            "PRIVATE-THREADED-REPLY-BASELINE"
        )
        contents[threaded_member] = serialize(threaded_root)

        people_root = ElementTree.Element(f"{{{people}}}personList")
        ElementTree.SubElement(
            people_root,
            f"{{{people}}}person",
            {
                "displayName": "Private Reviewer",
                "userId": "private-reviewer@example.invalid",
                "providerId": "private-provider",
                "id": reviewer_id,
            },
        )
        ElementTree.SubElement(
            people_root,
            f"{{{people}}}person",
            {
                "displayName": "Private Approver",
                "userId": "private-approver@example.invalid",
                "providerId": "private-provider",
                "id": approver_id,
            },
        )
        contents[person_member] = serialize(people_root)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        for part_name, content_type in (
            (threaded_member, "application/vnd.ms-excel.threadedcomments+xml"),
            (person_member, "application/vnd.ms-excel.person+xml"),
        ):
            ElementTree.SubElement(
                types,
                f"{{{content_types}}}Override",
                {"PartName": f"/{part_name}", "ContentType": content_type},
            )
        contents["[Content_Types].xml"] = serialize(types)

        workbook_relationships_name = _relationship_member("xl/workbook.xml")
        workbook_relationships = ElementTree.fromstring(contents[workbook_relationships_name])
        ElementTree.SubElement(
            workbook_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceThreadedPerson",
                "Type": "http://schemas.microsoft.com/office/2017/10/relationships/person",
                "Target": "persons/person.xml",
            },
        )
        contents[workbook_relationships_name] = serialize(workbook_relationships)

        worksheet_relationships_name = _relationship_member("xl/worksheets/sheet1.xml")
        if worksheet_relationships_name in contents:
            worksheet_relationships = ElementTree.fromstring(
                contents[worksheet_relationships_name]
            )
        else:
            worksheet_relationships = ElementTree.Element(
                f"{{{package_relationships}}}Relationships"
            )
        ElementTree.SubElement(
            worksheet_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceThreadedComment",
                "Type": (
                    "http://schemas.microsoft.com/office/2017/10/relationships/"
                    "threadedComment"
                ),
                "Target": "../threadedComments/threadedComment1.xml",
            },
        )
        contents[worksheet_relationships_name] = serialize(worksheet_relationships)

    return _rewrite_archive(path, mutate, ".threaded-comment.tmp.xlsx")


def make_legacy_threaded_placeholder_model(path: Path) -> Path:
    """Create an Excel threaded comment with its conventional Note placeholder."""
    make_legacy_comment_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    threaded_comments = (
        "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
    )

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        root_identifier = "{33333333-3333-3333-3333-333333333333}"
        reviewer_identifier = "{11111111-1111-1111-1111-111111111111}"
        comment_member, _drawing_member, worksheet_relationships_name = (
            _legacy_comment_part_names(contents)
        )
        comments_root = ElementTree.fromstring(contents[comment_member])
        author = next(
            comments_root.iter(f"{{{_SPREADSHEETML_NS}}}author"),
            None,
        )
        comment = next(
            comments_root.iter(f"{{{_SPREADSHEETML_NS}}}comment"),
            None,
        )
        if author is None or comment is None:
            raise ValueError("Fixture does not contain a conventional Note placeholder")
        author.text = f"tc={root_identifier}"
        comment.set("guid", root_identifier)
        contents[comment_member] = serialize(comments_root)

        threaded_member = "xl/threadedComments/threadedComment1.xml"
        person_member = "xl/persons/person.xml"
        threaded_root = ElementTree.Element(f"{{{threaded_comments}}}ThreadedComments")
        threaded_comment = ElementTree.SubElement(
            threaded_root,
            f"{{{threaded_comments}}}threadedComment",
            {
                "ref": "A1",
                "dT": "2026-07-24T00:00:00Z",
                "personId": reviewer_identifier,
                "id": root_identifier,
                "done": "0",
            },
        )
        ElementTree.SubElement(
            threaded_comment,
            f"{{{threaded_comments}}}text",
        ).text = "PRIVATE-THREADED-PLACEHOLDER-THREAD"
        contents[threaded_member] = serialize(threaded_root)

        people_root = ElementTree.Element(f"{{{threaded_comments}}}personList")
        ElementTree.SubElement(
            people_root,
            f"{{{threaded_comments}}}person",
            {
                "displayName": "Private Placeholder Reviewer",
                "userId": "private-placeholder@example.invalid",
                "providerId": "private-provider",
                "id": reviewer_identifier,
            },
        )
        contents[person_member] = serialize(people_root)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        for part_name, content_type in (
            (threaded_member, "application/vnd.ms-excel.threadedcomments+xml"),
            (person_member, "application/vnd.ms-excel.person+xml"),
        ):
            ElementTree.SubElement(
                types,
                f"{{{content_types}}}Override",
                {"PartName": f"/{part_name}", "ContentType": content_type},
            )
        contents["[Content_Types].xml"] = serialize(types)

        workbook_relationships_name = _relationship_member("xl/workbook.xml")
        workbook_relationships = ElementTree.fromstring(
            contents[workbook_relationships_name]
        )
        ElementTree.SubElement(
            workbook_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFencePlaceholderPerson",
                "Type": (
                    "http://schemas.microsoft.com/office/2017/10/relationships/person"
                ),
                "Target": "persons/person.xml",
            },
        )
        contents[workbook_relationships_name] = serialize(workbook_relationships)

        worksheet_relationships = ElementTree.fromstring(
            contents[worksheet_relationships_name]
        )
        ElementTree.SubElement(
            worksheet_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFencePlaceholderThread",
                "Type": (
                    "http://schemas.microsoft.com/office/2017/10/relationships/"
                    "threadedComment"
                ),
                "Target": "../threadedComments/threadedComment1.xml",
            },
        )
        contents[worksheet_relationships_name] = serialize(worksheet_relationships)

    return _rewrite_archive(path, mutate, ".legacy-threaded-placeholder.tmp.xlsx")


def renumber_legacy_threaded_placeholder_identifiers(path: Path) -> Path:
    """Rekey a threaded-comment placeholder and its thread consistently."""
    threaded_comments = (
        "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        new_identifier = "{AAAAAAAA-AAAA-AAAA-AAAA-AAAAAAAAAAAA}"
        comment_member, _drawing_member, _relationships_member = (
            _legacy_comment_part_names(contents)
        )
        comments_root = ElementTree.fromstring(contents[comment_member])
        author = next(
            comments_root.iter(f"{{{_SPREADSHEETML_NS}}}author"),
            None,
        )
        comment = next(
            comments_root.iter(f"{{{_SPREADSHEETML_NS}}}comment"),
            None,
        )
        if author is None or comment is None:
            raise ValueError("Fixture does not contain a conventional Note placeholder")
        author.text = f"tc={new_identifier}"
        comment.set("guid", new_identifier)
        contents[comment_member] = ElementTree.tostring(
            comments_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        threaded_member, _person_member, _worksheet_relationships = (
            _threaded_comment_part_names(contents)
        )
        threaded_root = ElementTree.fromstring(contents[threaded_member])
        threaded_comment = next(
            threaded_root.iter(f"{{{threaded_comments}}}threadedComment"),
            None,
        )
        if threaded_comment is None:
            raise ValueError("Fixture does not contain a threaded placeholder comment")
        threaded_comment.set("id", new_identifier)
        contents[threaded_member] = ElementTree.tostring(
            threaded_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-threaded-placeholder-id.tmp.xlsx")


def lowercase_legacy_threaded_placeholder_identifiers(path: Path) -> Path:
    """Use lowercase GUID spelling that strict workbook readers reject."""
    threaded_comments = (
        "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        new_identifier = "{aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa}"
        comment_member, _drawing_member, _relationships_member = (
            _legacy_comment_part_names(contents)
        )
        comments_root = ElementTree.fromstring(contents[comment_member])
        author = next(
            comments_root.iter(f"{{{_SPREADSHEETML_NS}}}author"),
            None,
        )
        comment = next(
            comments_root.iter(f"{{{_SPREADSHEETML_NS}}}comment"),
            None,
        )
        if author is None or comment is None:
            raise ValueError("Fixture does not contain a conventional Note placeholder")
        author.text = f"tc={new_identifier}"
        comment.set("guid", new_identifier)
        contents[comment_member] = ElementTree.tostring(
            comments_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        threaded_member, _person_member, _worksheet_relationships = (
            _threaded_comment_part_names(contents)
        )
        threaded_root = ElementTree.fromstring(contents[threaded_member])
        threaded_comment = next(
            threaded_root.iter(f"{{{threaded_comments}}}threadedComment"),
            None,
        )
        if threaded_comment is None:
            raise ValueError("Fixture does not contain a threaded placeholder comment")
        threaded_comment.set("id", new_identifier)
        contents[threaded_member] = ElementTree.tostring(
            threaded_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-threaded-placeholder-lower.tmp.xlsx")


def change_legacy_placeholder_author_context(path: Path) -> Path:
    """Add author context around a placeholder token without changing its GUID."""
    def mutate(contents: dict[str, bytes]) -> None:
        comment_member, _drawing_member, _relationships_member = (
            _legacy_comment_part_names(contents)
        )
        comments_root = ElementTree.fromstring(contents[comment_member])
        author = next(
            comments_root.iter(f"{{{_SPREADSHEETML_NS}}}author"),
            None,
        )
        if author is None or not author.text:
            raise ValueError("Fixture does not contain a placeholder author")
        author.text = f"Private context {author.text}"
        contents[comment_member] = ElementTree.tostring(
            comments_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".legacy-placeholder-author.tmp.xlsx")


def change_threaded_comment_reply(path: Path) -> Path:
    """Change only a private threaded reply body."""
    threaded_comments = (
        "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
    )

    def mutate(contents: dict[str, bytes]) -> None:
        threaded_member, _person_member, _worksheet_relationships = (
            _threaded_comment_part_names(contents)
        )
        root = ElementTree.fromstring(contents[threaded_member])
        comments = list(root.findall(f"{{{threaded_comments}}}threadedComment"))
        comments[-1].find(f"{{{threaded_comments}}}text").text = (
            "PRIVATE-THREADED-REPLY-CANDIDATE"
        )
        contents[threaded_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".threaded-comment-change.tmp.xlsx")


def change_threaded_comment_person_identity(path: Path) -> Path:
    """Change only private collaborator identity material."""
    people = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"

    def mutate(contents: dict[str, bytes]) -> None:
        _threaded_member, person_member, _worksheet_relationships = (
            _threaded_comment_part_names(contents)
        )
        root = ElementTree.fromstring(contents[person_member])
        person = next(root.iter(f"{{{people}}}person"))
        person.set("displayName", "Private Reviewer Candidate")
        contents[person_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".threaded-comment-person-change.tmp.xlsx")


def externalize_threaded_comment_relationship(path: Path) -> Path:
    """Make the worksheet comment relationship unsafe without following it."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _threaded_member, _person_member, worksheet_relationships_name = (
            _threaded_comment_part_names(contents)
        )
        root = ElementTree.fromstring(contents[worksheet_relationships_name])
        relationship = next(
            item
            for item in root.findall(f"{{{package_relationships}}}Relationship")
            if item.get("Id") == "rIdFenceThreadedComment"
        )
        relationship.set("Target", "https://example.invalid/private-threaded-comment")
        relationship.set("TargetMode", "External")
        contents[worksheet_relationships_name] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".threaded-comment-external.tmp.xlsx")


def renumber_threaded_comment_identifiers(path: Path) -> Path:
    """Rewrite raw person, comment, and package identifiers consistently."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    threaded_comments = (
        "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
    )
    people = threaded_comments

    def mutate(contents: dict[str, bytes]) -> None:
        threaded_member, person_member, worksheet_relationships_name = (
            _threaded_comment_part_names(contents)
        )
        person_root = ElementTree.fromstring(contents[person_member])
        person_mapping: dict[str, str] = {}
        for index, person in enumerate(
            person_root.iter(f"{{{people}}}person"),
            start=100,
        ):
            old_identifier = person.get("id")
            new_identifier = f"{{aaaaaaaa-aaaa-aaaa-aaaa-{index:012d}}}"
            if old_identifier is None:
                raise ValueError("Fixture person is missing its identifier")
            person_mapping[old_identifier] = new_identifier
            person.set("id", new_identifier)
        contents[person_member] = ElementTree.tostring(
            person_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        threaded_root = ElementTree.fromstring(contents[threaded_member])
        comment_mapping: dict[str, str] = {}
        comments = list(threaded_root.findall(f"{{{threaded_comments}}}threadedComment"))
        for index, comment in enumerate(comments, start=200):
            old_identifier = comment.get("id")
            new_identifier = f"{{bbbbbbbb-bbbb-bbbb-bbbb-{index:012d}}}"
            if old_identifier is None:
                raise ValueError("Fixture comment is missing its identifier")
            comment_mapping[old_identifier] = new_identifier
            comment.set("id", new_identifier)
        for comment in comments:
            person_identifier = comment.get("personId")
            if person_identifier:
                comment.set("personId", person_mapping[person_identifier])
            parent_identifier = comment.get("parentId")
            if parent_identifier:
                comment.set("parentId", comment_mapping[parent_identifier])
            for mention_index, mention in enumerate(
                comment.iter(f"{{{threaded_comments}}}mention"),
                start=300,
            ):
                mention_identifier = mention.get("mentionpersonId")
                if mention_identifier:
                    mention.set(
                        "mentionpersonId",
                        person_mapping[mention_identifier],
                    )
                mention.set(
                    "mentionId",
                    f"{{cccccccc-cccc-cccc-cccc-{mention_index:012d}}}",
                )
        contents[threaded_member] = ElementTree.tostring(
            threaded_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet_relationships = ElementTree.fromstring(
            contents[worksheet_relationships_name]
        )
        relationship = next(
            item
            for item in worksheet_relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if item.get("Id") == "rIdFenceThreadedComment"
        )
        relationship.set("Id", "rIdFenceThreadedCommentRenumbered")
        contents[worksheet_relationships_name] = ElementTree.tostring(
            worksheet_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        workbook_relationships_name = _relationship_member("xl/workbook.xml")
        workbook_relationships = ElementTree.fromstring(contents[workbook_relationships_name])
        relationship = next(
            item
            for item in workbook_relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if item.get("Id") == "rIdFenceThreadedPerson"
        )
        relationship.set("Id", "rIdFenceThreadedPersonRenumbered")
        contents[workbook_relationships_name] = ElementTree.tostring(
            workbook_relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".threaded-comment-id.tmp.xlsx")


def corrupt_threaded_comment_root(path: Path) -> Path:
    """Replace the threaded-comment root to exercise fail-closed coverage."""
    def mutate(contents: dict[str, bytes]) -> None:
        threaded_member, _person_member, _worksheet_relationships = (
            _threaded_comment_part_names(contents)
        )
        root = ElementTree.fromstring(contents[threaded_member])
        root.tag = "notThreadedComments"
        contents[threaded_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".threaded-comment-corrupt.tmp.xlsx")


def _worksheet_drawing_shape_part_names(
    contents: dict[str, bytes],
) -> tuple[str, str]:
    """Return the drawing and relationship members for the shape fixture."""
    drawing_member = "xl/drawings/drawing1.xml"
    relationship_member = _relationship_member(drawing_member)
    if drawing_member not in contents or relationship_member not in contents:
        raise ValueError("Fixture does not contain Worksheet DrawingML shape parts")
    return drawing_member, relationship_member


def make_worksheet_drawing_shape_model(path: Path) -> Path:
    """Create raw worksheet text-box and grouped-shape controls for inspection."""
    make_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_main = "http://schemas.openxmlformats.org/drawingml/2006/main"
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def marker(
        parent: ElementTree.Element,
        name: str,
        *,
        column: int,
        row: int,
    ) -> None:
        point = ElementTree.SubElement(parent, f"{{{drawing}}}{name}")
        ElementTree.SubElement(point, f"{{{drawing}}}col").text = str(column)
        ElementTree.SubElement(point, f"{{{drawing}}}colOff").text = "0"
        ElementTree.SubElement(point, f"{{{drawing}}}row").text = str(row)
        ElementTree.SubElement(point, f"{{{drawing}}}rowOff").text = "0"

    def text_body(parent: ElementTree.Element, text: str, colour: str) -> None:
        body = ElementTree.SubElement(parent, f"{{{drawing}}}txBody")
        ElementTree.SubElement(body, f"{{{drawing_main}}}bodyPr")
        ElementTree.SubElement(body, f"{{{drawing_main}}}lstStyle")
        paragraph = ElementTree.SubElement(body, f"{{{drawing_main}}}p")
        run = ElementTree.SubElement(paragraph, f"{{{drawing_main}}}r")
        properties = ElementTree.SubElement(run, f"{{{drawing_main}}}rPr")
        fill = ElementTree.SubElement(properties, f"{{{drawing_main}}}solidFill")
        ElementTree.SubElement(fill, f"{{{drawing_main}}}srgbClr", {"val": colour})
        ElementTree.SubElement(run, f"{{{drawing_main}}}t").text = text

    def shape(
        parent: ElementTree.Element,
        *,
        identifier: str,
        name: str,
        text: str,
        colour: str,
        macro: str | None = None,
        text_link: str | None = None,
        hyperlink: bool = False,
    ) -> ElementTree.Element:
        attributes: dict[str, str] = {}
        if macro is not None:
            attributes["macro"] = macro
        if text_link is not None:
            attributes["textlink"] = text_link
        current = ElementTree.SubElement(parent, f"{{{drawing}}}sp", attributes)
        non_visual = ElementTree.SubElement(current, f"{{{drawing}}}nvSpPr")
        properties = ElementTree.SubElement(
            non_visual,
            f"{{{drawing}}}cNvPr",
            {
                "id": identifier,
                "name": name,
                "descr": "Private worksheet shape description",
            },
        )
        if hyperlink:
            ElementTree.SubElement(
                properties,
                f"{{{drawing_main}}}hlinkClick",
                {f"{{{document_relationships}}}id": "rIdFenceShapeLink"},
            )
        ElementTree.SubElement(non_visual, f"{{{drawing}}}cNvSpPr")
        shape_properties = ElementTree.SubElement(current, f"{{{drawing}}}spPr")
        fill = ElementTree.SubElement(shape_properties, f"{{{drawing_main}}}solidFill")
        ElementTree.SubElement(fill, f"{{{drawing_main}}}srgbClr", {"val": "FFFFFF"})
        text_body(current, text, colour)
        return current

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member = "xl/drawings/drawing1.xml"
        drawing_root = ElementTree.Element(f"{{{drawing}}}wsDr")

        warning_anchor = ElementTree.SubElement(
            drawing_root,
            f"{{{drawing}}}twoCellAnchor",
            {"editAs": "oneCell"},
        )
        marker(warning_anchor, "from", column=1, row=1)
        marker(warning_anchor, "to", column=5, row=4)
        shape(
            warning_anchor,
            identifier="1025",
            name="Private worksheet warning shape",
            text="PRIVATE-SHAPE-DO-NOT-APPROVE",
            colour="000000",
            macro="PrivateWorksheetShapeMacro",
            text_link="=Inputs!$B$2",
            hyperlink=True,
        )
        ElementTree.SubElement(warning_anchor, f"{{{drawing}}}clientData")

        group_anchor = ElementTree.SubElement(
            drawing_root,
            f"{{{drawing}}}oneCellAnchor",
        )
        marker(group_anchor, "from", column=7, row=2)
        ElementTree.SubElement(
            group_anchor,
            f"{{{drawing}}}ext",
            {"cx": "1828800", "cy": "548640"},
        )
        group = ElementTree.SubElement(group_anchor, f"{{{drawing}}}grpSp")
        ElementTree.SubElement(group, f"{{{drawing}}}nvGrpSpPr")
        ElementTree.SubElement(group, f"{{{drawing}}}grpSpPr")
        shape(
            group,
            identifier="1026",
            name="Private grouped shape",
            text="PRIVATE-GROUP-SHAPE-TEXT",
            colour="334455",
        )
        ElementTree.SubElement(group_anchor, f"{{{drawing}}}clientData")
        contents[drawing_member] = serialize(drawing_root)

        relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceShapeLink",
                "Type": f"{document_relationships}/hyperlink",
                "Target": "https://example.invalid/private-worksheet-shape-target",
                "TargetMode": "External",
            },
        )
        contents[_relationship_member(drawing_member)] = serialize(relationships)

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}drawing",
            {f"{{{document_relationships}}}id": "rIdFenceWorksheetDrawing"},
        )
        contents["xl/worksheets/sheet1.xml"] = serialize(worksheet)
        worksheet_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            worksheet_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceWorksheetDrawing",
                "Type": f"{document_relationships}/drawing",
                "Target": "../drawings/drawing1.xml",
            },
        )
        contents[_relationship_member("xl/worksheets/sheet1.xml")] = serialize(
            worksheet_relationships
        )

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        ElementTree.SubElement(
            types,
            f"{{{content_types}}}Override",
            {
                "PartName": "/xl/drawings/drawing1.xml",
                "ContentType": (
                    "application/vnd.openxmlformats-officedocument.drawing+xml"
                ),
            },
        )
        contents["[Content_Types].xml"] = serialize(types)

    return _rewrite_archive(path, mutate, ".worksheet-drawing-shape.tmp.xlsx")


def change_worksheet_drawing_shape_presentation(path: Path) -> Path:
    """Make the primary shape text less visible without changing any cells."""
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_main = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, _relationship_member_name = _worksheet_drawing_shape_part_names(
            contents
        )
        root = ElementTree.fromstring(contents[drawing_member])
        shape = next(root.iter(f"{{{drawing}}}sp"))
        text_body = shape.find(f"{{{drawing}}}txBody")
        if text_body is None:
            raise ValueError("Fixture does not contain a worksheet shape text body")
        colour = next(text_body.iter(f"{{{drawing_main}}}srgbClr"))
        colour.set("val", "FFFFFF")
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-drawing-shape-change.tmp.xlsx")


def change_worksheet_drawing_shape_hyperlink(path: Path) -> Path:
    """Change only a private external hyperlink target for one shape."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _drawing_member, relationships_name = _worksheet_drawing_shape_part_names(
            contents
        )
        relationships = ElementTree.fromstring(contents[relationships_name])
        hyperlink = next(
            relationship
            for relationship in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if relationship.get("Id") == "rIdFenceShapeLink"
        )
        hyperlink.set(
            "Target",
            "https://example.invalid/private-worksheet-shape-candidate",
        )
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-drawing-shape-link.tmp.xlsx")


def renumber_worksheet_drawing_shape_identifiers(path: Path) -> Path:
    """Rewrite harmless nonvisual and relationship IDs without changing semantics."""
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_main = "http://schemas.openxmlformats.org/drawingml/2006/main"
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, relationships_name = _worksheet_drawing_shape_part_names(
            contents
        )
        root = ElementTree.fromstring(contents[drawing_member])
        for index, current in enumerate(
            root.iter(f"{{{drawing}}}cNvPr"),
            start=900,
        ):
            current.set("id", str(index))
        hyperlink = next(root.iter(f"{{{drawing_main}}}hlinkClick"))
        hyperlink.set(
            f"{{{document_relationships}}}id",
            "rIdFenceShapeLinkRenumbered",
        )
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships = ElementTree.fromstring(contents[relationships_name])
        relationship = next(
            current
            for current in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if current.get("Id") == "rIdFenceShapeLink"
        )
        relationship.set("Id", "rIdFenceShapeLinkRenumbered")
        contents[relationships_name] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-drawing-shape-id.tmp.xlsx")


def corrupt_worksheet_drawing_shape_root(path: Path) -> Path:
    """Replace the shape drawing root to exercise fail-closed coverage."""
    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, _relationship_member_name = _worksheet_drawing_shape_part_names(
            contents
        )
        root = ElementTree.fromstring(contents[drawing_member])
        root.tag = "notWorksheetDrawing"
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-drawing-shape-corrupt.tmp.xlsx")


def _worksheet_smartart_part_names(contents: dict[str, bytes]) -> tuple[str, str]:
    """Return the drawing and relationship members for the SmartArt fixture."""
    drawing_member = "xl/drawings/drawing1.xml"
    relationship_member = _relationship_member(drawing_member)
    if drawing_member not in contents or relationship_member not in contents:
        raise ValueError("Fixture does not contain Worksheet DrawingML SmartArt parts")
    return drawing_member, relationship_member


def make_worksheet_smartart_model(path: Path) -> Path:
    """Create a raw Worksheet DrawingML SmartArt package outside the cell grid."""
    make_model(path)
    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_main = "http://schemas.openxmlformats.org/drawingml/2006/main"
    diagram = "http://schemas.openxmlformats.org/drawingml/2006/diagram"
    diagram_drawing = "http://schemas.microsoft.com/office/drawing/2008/diagram"
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def marker(
        parent: ElementTree.Element,
        name: str,
        *,
        column: int,
        row: int,
    ) -> None:
        point = ElementTree.SubElement(parent, f"{{{drawing}}}{name}")
        ElementTree.SubElement(point, f"{{{drawing}}}col").text = str(column)
        ElementTree.SubElement(point, f"{{{drawing}}}colOff").text = "0"
        ElementTree.SubElement(point, f"{{{drawing}}}row").text = str(row)
        ElementTree.SubElement(point, f"{{{drawing}}}rowOff").text = "0"

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member = "xl/drawings/drawing1.xml"
        drawing_root = ElementTree.Element(f"{{{drawing}}}wsDr")
        anchor = ElementTree.SubElement(
            drawing_root,
            f"{{{drawing}}}twoCellAnchor",
            {"editAs": "oneCell"},
        )
        marker(anchor, "from", column=2, row=3)
        marker(anchor, "to", column=8, row=12)
        frame = ElementTree.SubElement(
            anchor,
            f"{{{drawing}}}graphicFrame",
            {"macro": "PrivateSmartArtMacro"},
        )
        nonvisual = ElementTree.SubElement(frame, f"{{{drawing}}}nvGraphicFramePr")
        ElementTree.SubElement(
            nonvisual,
            f"{{{drawing}}}cNvPr",
            {
                "id": "501",
                "name": "PRIVATE-SMARTART-NAME",
                "descr": "PRIVATE-SMARTART-DESCRIPTION",
            },
        )
        ElementTree.SubElement(nonvisual, f"{{{drawing}}}cNvGraphicFramePr")
        transform = ElementTree.SubElement(frame, f"{{{drawing}}}xfrm")
        ElementTree.SubElement(transform, f"{{{drawing_main}}}off", {"x": "0", "y": "0"})
        ElementTree.SubElement(
            transform,
            f"{{{drawing_main}}}ext",
            {"cx": "0", "cy": "0"},
        )
        graphic = ElementTree.SubElement(frame, f"{{{drawing_main}}}graphic")
        graphic_data = ElementTree.SubElement(
            graphic,
            f"{{{drawing_main}}}graphicData",
            {"uri": diagram},
        )
        ElementTree.SubElement(
            graphic_data,
            f"{{{diagram}}}relIds",
            {
                f"{{{document_relationships}}}dm": "rIdFenceDiagramData",
                f"{{{document_relationships}}}lo": "rIdFenceDiagramLayout",
                f"{{{document_relationships}}}qs": "rIdFenceDiagramStyle",
                f"{{{document_relationships}}}cs": "rIdFenceDiagramColours",
            },
        )
        ElementTree.SubElement(anchor, f"{{{drawing}}}clientData")
        contents[drawing_member] = serialize(drawing_root)

        relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        for relationship_id, relationship_type, target in (
            (
                "rIdFenceDiagramData",
                f"{document_relationships}/diagramData",
                "../diagrams/data1.xml",
            ),
            (
                "rIdFenceDiagramLayout",
                f"{document_relationships}/diagramLayout",
                "../diagrams/layout1.xml",
            ),
            (
                "rIdFenceDiagramStyle",
                f"{document_relationships}/diagramQuickStyle",
                "../diagrams/quickStyle1.xml",
            ),
            (
                "rIdFenceDiagramColours",
                f"{document_relationships}/diagramColors",
                "../diagrams/colors1.xml",
            ),
            (
                "rIdFenceDiagramRendering",
                "http://schemas.microsoft.com/office/2007/relationships/diagramDrawing",
                "../diagrams/drawing1.xml",
            ),
        ):
            ElementTree.SubElement(
                relationships,
                f"{{{package_relationships}}}Relationship",
                {"Id": relationship_id, "Type": relationship_type, "Target": target},
            )
        contents[_relationship_member(drawing_member)] = serialize(relationships)

        data_root = ElementTree.Element(f"{{{diagram}}}dataModel")
        points = ElementTree.SubElement(data_root, f"{{{diagram}}}ptLst")
        point = ElementTree.SubElement(
            points,
            f"{{{diagram}}}pt",
            {"modelId": "{AAAAAAAA-AAAA-AAAA-AAAA-AAAAAAAAAAAA}", "type": "doc"},
        )
        text = ElementTree.SubElement(point, f"{{{diagram}}}t")
        text.text = "PRIVATE-SMARTART-DO-NOT-APPROVE"
        contents["xl/diagrams/data1.xml"] = serialize(data_root)

        layout_root = ElementTree.Element(
            f"{{{diagram}}}layoutDef",
            {"uniqueId": "{BBBBBBBB-BBBB-BBBB-BBBB-BBBBBBBBBBBB}"},
        )
        ElementTree.SubElement(layout_root, f"{{{diagram}}}title").text = (
            "PRIVATE-SMARTART-LAYOUT"
        )
        contents["xl/diagrams/layout1.xml"] = serialize(layout_root)

        quick_style_root = ElementTree.Element(
            f"{{{diagram}}}styleDef",
            {"uniqueId": "{CCCCCCCC-CCCC-CCCC-CCCC-CCCCCCCCCCCC}"},
        )
        ElementTree.SubElement(quick_style_root, f"{{{diagram}}}title").text = (
            "PRIVATE-SMARTART-STYLE"
        )
        contents["xl/diagrams/quickStyle1.xml"] = serialize(quick_style_root)

        colours_root = ElementTree.Element(
            f"{{{diagram}}}colorsDef",
            {"uniqueId": "{DDDDDDDD-DDDD-DDDD-DDDD-DDDDDDDDDDDD}"},
        )
        ElementTree.SubElement(colours_root, f"{{{diagram}}}title").text = (
            "PRIVATE-SMARTART-COLOURS"
        )
        contents["xl/diagrams/colors1.xml"] = serialize(colours_root)

        rendering_root = ElementTree.Element(f"{{{diagram_drawing}}}drawing")
        ElementTree.SubElement(rendering_root, f"{{{diagram_drawing}}}sp").text = (
            "PRIVATE-SMARTART-RENDERING"
        )
        contents["xl/diagrams/drawing1.xml"] = serialize(rendering_root)

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}drawing",
            {f"{{{document_relationships}}}id": "rIdFenceWorksheetDrawing"},
        )
        contents["xl/worksheets/sheet1.xml"] = serialize(worksheet)
        worksheet_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            worksheet_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceWorksheetDrawing",
                "Type": f"{document_relationships}/drawing",
                "Target": "../drawings/drawing1.xml",
            },
        )
        contents[_relationship_member("xl/worksheets/sheet1.xml")] = serialize(
            worksheet_relationships
        )

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        for part_name, content_type in (
            (
                "/xl/drawings/drawing1.xml",
                "application/vnd.openxmlformats-officedocument.drawing+xml",
            ),
            (
                "/xl/diagrams/data1.xml",
                "application/vnd.openxmlformats-officedocument.drawingml.diagramData+xml",
            ),
            (
                "/xl/diagrams/layout1.xml",
                "application/vnd.openxmlformats-officedocument.drawingml.diagramLayout+xml",
            ),
            (
                "/xl/diagrams/quickStyle1.xml",
                "application/vnd.openxmlformats-officedocument.drawingml.diagramStyle+xml",
            ),
            (
                "/xl/diagrams/colors1.xml",
                "application/vnd.openxmlformats-officedocument.drawingml.diagramColors+xml",
            ),
            (
                "/xl/diagrams/drawing1.xml",
                "application/vnd.ms-office.drawingml.diagramDrawing+xml",
            ),
        ):
            ElementTree.SubElement(
                types,
                f"{{{content_types}}}Override",
                {"PartName": part_name, "ContentType": content_type},
            )
        contents["[Content_Types].xml"] = serialize(types)

    return _rewrite_archive(path, mutate, ".worksheet-smartart.tmp.xlsx")


def make_worksheet_smartart_image_model(path: Path) -> Path:
    """Add a direct Diagram Data image relationship to the SmartArt fixture."""
    make_worksheet_smartart_model(path)
    drawing_main = _DRAWINGML_MAIN_NS
    diagram = "http://schemas.openxmlformats.org/drawingml/2006/diagram"
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    baseline_png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9JdcAAAAAASUVORK5CYII="
    )

    def mutate(contents: dict[str, bytes]) -> None:
        data_member = "xl/diagrams/data1.xml"
        data = ElementTree.fromstring(contents[data_member])
        points = next(data.iter(f"{{{diagram}}}ptLst"))
        image_point = ElementTree.SubElement(
            points,
            f"{{{diagram}}}pt",
            {
                "modelId": "{EEEEEEEE-EEEE-EEEE-EEEE-EEEEEEEEEEEE}",
                "type": "pres",
            },
        )
        shape_properties = ElementTree.SubElement(image_point, f"{{{diagram}}}spPr")
        image_fill = ElementTree.SubElement(
            shape_properties,
            f"{{{drawing_main}}}blipFill",
            {"rotWithShape": "false"},
        )
        ElementTree.SubElement(
            image_fill,
            f"{{{drawing_main}}}blip",
            {f"{{{document_relationships}}}embed": "rIdFenceDiagramImage"},
        )
        stretch = ElementTree.SubElement(image_fill, f"{{{drawing_main}}}stretch")
        ElementTree.SubElement(stretch, f"{{{drawing_main}}}fillRect")
        contents[data_member] = ElementTree.tostring(
            data,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceDiagramImage",
                "Type": f"{document_relationships}/image",
                "Target": "../media/private-smartart-diagram.png",
            },
        )
        contents[_relationship_member(data_member)] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )
        contents["xl/media/private-smartart-diagram.png"] = baseline_png

        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"
        if not any(
            element.get("Extension") == "png"
            for element in content_types.findall(default_tag)
        ):
            ElementTree.SubElement(
                content_types,
                default_tag,
                {"Extension": "png", "ContentType": "image/png"},
            )
        contents["[Content_Types].xml"] = ElementTree.tostring(
            content_types,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-smartart-image.tmp.xlsx")


def change_worksheet_smartart_diagram_image_payload(path: Path) -> Path:
    """Change only direct SmartArt Diagram Data image bytes outside cells."""
    def mutate(contents: dict[str, bytes]) -> None:
        member = "xl/media/private-smartart-diagram.png"
        if member not in contents:
            raise ValueError("Fixture does not contain a SmartArt Diagram Data image")
        contents[member] = b"PRIVATE-SMARTART-DIAGRAM-IMAGE-CANDIDATE"

    return _rewrite_archive(path, mutate, ".worksheet-smartart-image-change.tmp.xlsx")


def remove_worksheet_smartart_diagram_image_payload(path: Path) -> Path:
    """Remove a declared Diagram Data image target to exercise coverage handling."""

    def mutate(contents: dict[str, bytes]) -> None:
        member = "xl/media/private-smartart-diagram.png"
        if member not in contents:
            raise ValueError("Fixture does not contain a SmartArt Diagram Data image")
        contents.pop(member)

    return _rewrite_archive(path, mutate, ".worksheet-smartart-image-missing.tmp.xlsx")


def renumber_worksheet_smartart_diagram_image_relationship(path: Path) -> Path:
    """Rewrite a Diagram Data image relationship ID without changing its graph."""
    data_member = "xl/diagrams/data1.xml"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships_member = _relationship_member(data_member)
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            item
            for item in relationships.findall(f"{{{package_relationships}}}Relationship")
            if item.get("Id") == "rIdFenceDiagramImage"
        )
        relationship.set("Id", "rIdFenceDiagramImageRenumbered")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        data = ElementTree.fromstring(contents[data_member])
        image = next(data.iter(f"{{{_DRAWINGML_MAIN_NS}}}blip"))
        image.set(
            f"{{{document_relationships}}}embed",
            "rIdFenceDiagramImageRenumbered",
        )
        contents[data_member] = ElementTree.tostring(
            data,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-smartart-image-id.tmp.xlsx")


def externalize_worksheet_smartart_diagram_image_relationship(path: Path) -> Path:
    """Turn a Diagram Data image relationship into an external coverage gap."""
    data_member = "xl/diagrams/data1.xml"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships_member = _relationship_member(data_member)
        relationships = ElementTree.fromstring(contents[relationships_member])
        relationship = next(
            item
            for item in relationships.findall(f"{{{package_relationships}}}Relationship")
            if item.get("Id") == "rIdFenceDiagramImage"
        )
        relationship.set("Target", "https://example.invalid/private-smartart-image")
        relationship.set("TargetMode", "External")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-smartart-image-external.tmp.xlsx")


def change_worksheet_smartart_data(path: Path) -> Path:
    """Change only private SmartArt data, leaving every worksheet cell unchanged."""

    def mutate(contents: dict[str, bytes]) -> None:
        member = "xl/diagrams/data1.xml"
        root = ElementTree.fromstring(contents[member])
        diagram = root.tag[1:].split("}", maxsplit=1)[0]
        text = next(root.iter(f"{{{diagram}}}t"))
        text.text = "PRIVATE-SMARTART-CANDIDATE-REVIEW-STATE"
        contents[member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-smartart-data.tmp.xlsx")


def change_worksheet_smartart_graphic_frame_uri(path: Path) -> Path:
    """Turn a SmartArt frame into an unsupported non-chart graphic frame."""
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, _relationships_member = _worksheet_smartart_part_names(contents)
        root = ElementTree.fromstring(contents[drawing_member])
        graphic_data = next(root.iter(f"{{{drawing}}}graphicData"))
        graphic_data.set("uri", "https://example.invalid/formulafence/unknown-graphic")
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-unknown-graphic-frame.tmp.xlsx")


def renumber_worksheet_smartart_identifiers(path: Path) -> Path:
    """Rewrite harmless DrawingML and relationship IDs without changing SmartArt."""
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    mapping = {
        "rIdFenceDiagramData": "rIdFenceDiagramDataRenumbered",
        "rIdFenceDiagramLayout": "rIdFenceDiagramLayoutRenumbered",
        "rIdFenceDiagramStyle": "rIdFenceDiagramStyleRenumbered",
        "rIdFenceDiagramColours": "rIdFenceDiagramColoursRenumbered",
        "rIdFenceDiagramRendering": "rIdFenceDiagramRenderingRenumbered",
    }

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, relationships_member = _worksheet_smartart_part_names(contents)
        root = ElementTree.fromstring(contents[drawing_member])
        next(root.iter(f"{{{drawing}}}cNvPr")).set("id", "9501")
        for element in root.iter():
            for attribute, value in tuple(element.attrib.items()):
                if value in mapping:
                    element.set(attribute, mapping[value])
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationships = ElementTree.fromstring(contents[relationships_member])
        for relationship in relationships.findall(
            f"{{{package_relationships}}}Relationship"
        ):
            if relationship.get("Id") in mapping:
                relationship.set("Id", mapping[relationship.get("Id")])
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-smartart-id.tmp.xlsx")


def corrupt_worksheet_smartart_relationships(path: Path) -> Path:
    """Remove a required diagram colour relationship to exercise coverage evidence."""
    diagram = "http://schemas.openxmlformats.org/drawingml/2006/diagram"
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, _relationships_member = _worksheet_smartart_part_names(contents)
        root = ElementTree.fromstring(contents[drawing_member])
        relationship_ids = next(root.iter(f"{{{diagram}}}relIds"))
        relationship_ids.attrib.pop(f"{{{document_relationships}}}cs")
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-smartart-corrupt.tmp.xlsx")


def corrupt_worksheet_smartart_component_root(path: Path) -> Path:
    """Replace a diagram component root to exercise fail-closed validation."""
    diagram = "http://schemas.openxmlformats.org/drawingml/2006/diagram"

    def mutate(contents: dict[str, bytes]) -> None:
        member = "xl/diagrams/data1.xml"
        root = ElementTree.fromstring(contents[member])
        root.tag = f"{{{diagram}}}unexpectedDataModel"
        contents[member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-smartart-component-root.tmp.xlsx")


def add_worksheet_smartart_component_relationship(path: Path) -> Path:
    """Add an unfollowed component relationship to exercise the scan boundary."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceSmartArtComponentTarget",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/hyperlink",
                "Target": "https://example.invalid/private-smartart-target",
                "TargetMode": "External",
            },
        )
        contents[_relationship_member("xl/diagrams/data1.xml")] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-smartart-component-rels.tmp.xlsx")


def _make_strict_worksheet_smartart_model(path: Path, *, with_image: bool) -> Path:
    """Create a Strict SpreadsheetML SmartArt fixture, optionally with an image."""
    if with_image:
        make_worksheet_smartart_image_model(path)
    else:
        make_worksheet_smartart_model(path)
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    strict_drawing = "http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing"
    diagram = "http://schemas.openxmlformats.org/drawingml/2006/diagram"
    strict_diagram = "http://purl.oclc.org/ooxml/drawingml/diagram"

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, drawing_relationships_member = _worksheet_smartart_part_names(
            contents
        )
        worksheet_member = "xl/worksheets/sheet1.xml"
        for member, source_namespace, target_namespace in (
            (worksheet_member, _SPREADSHEETML_NS, _STRICT_SPREADSHEETML_NS),
            (drawing_member, drawing, strict_drawing),
            ("xl/diagrams/data1.xml", diagram, strict_diagram),
            ("xl/diagrams/layout1.xml", diagram, strict_diagram),
            ("xl/diagrams/quickStyle1.xml", diagram, strict_diagram),
            ("xl/diagrams/colors1.xml", diagram, strict_diagram),
        ):
            root = ElementTree.fromstring(contents[member])
            for element in root.iter():
                element.tag = strict_name(
                    element.tag,
                    source_namespace,
                    target_namespace,
                )
                element.tag = strict_name(
                    element.tag,
                    _DRAWINGML_MAIN_NS,
                    _DRAWINGML_STRICT_MAIN_NS,
                )
                attributes = {
                    strict_name(
                        name,
                        _DOCUMENT_RELATIONSHIPS_NS,
                        _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                    ): value
                    for name, value in element.attrib.items()
                }
                element.attrib.clear()
                element.attrib.update(attributes)
                if (
                    element.tag == f"{{{_DRAWINGML_STRICT_MAIN_NS}}}graphicData"
                    and element.get("uri") == diagram
                ):
                    element.set("uri", strict_diagram)
            contents[member] = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationship_members = [
            _relationship_member(worksheet_member),
            drawing_relationships_member,
        ]
        if with_image:
            relationship_members.append(_relationship_member("xl/diagrams/data1.xml"))
        for member in relationship_members:
            relationships = ElementTree.fromstring(contents[member])
            for relationship in relationships.findall(relationship_tag):
                relationship_type = relationship.get("Type")
                if relationship_type and relationship_type.startswith(
                    _DOCUMENT_RELATIONSHIPS_NS
                ):
                    relationship.set(
                        "Type",
                        relationship_type.replace(
                            _DOCUMENT_RELATIONSHIPS_NS,
                            _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                            1,
                        ),
                    )
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".strict-worksheet-smartart.tmp.xlsx")


def make_strict_worksheet_smartart_model(path: Path) -> Path:
    """Create a Strict SpreadsheetML variant of the SmartArt fixture."""
    return _make_strict_worksheet_smartart_model(path, with_image=False)


def make_strict_worksheet_smartart_image_model(path: Path) -> Path:
    """Create a Strict SpreadsheetML SmartArt fixture with a Diagram Data image."""
    return _make_strict_worksheet_smartart_model(path, with_image=True)


def make_worksheet_drawing_connector_model(
    path: Path,
    *,
    attached: bool = True,
    grouped: bool = False,
) -> Path:
    """Create an anchored DrawingML connector with optional grouping and attachments."""
    make_worksheet_drawing_shape_model(path)
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_main = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def marker(
        parent: ElementTree.Element,
        name: str,
        *,
        column: int,
        row: int,
    ) -> None:
        point = ElementTree.SubElement(parent, f"{{{drawing}}}{name}")
        ElementTree.SubElement(point, f"{{{drawing}}}col").text = str(column)
        ElementTree.SubElement(point, f"{{{drawing}}}colOff").text = "0"
        ElementTree.SubElement(point, f"{{{drawing}}}row").text = str(row)
        ElementTree.SubElement(point, f"{{{drawing}}}rowOff").text = "0"

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, _relationships_member = _worksheet_drawing_shape_part_names(
            contents
        )
        root = ElementTree.fromstring(contents[drawing_member])
        anchor = ElementTree.SubElement(
            root,
            f"{{{drawing}}}twoCellAnchor",
            {"editAs": "oneCell"},
        )
        marker(anchor, "from", column=3, row=6)
        marker(anchor, "to", column=9, row=6)
        connector_parent = anchor
        if grouped:
            group = ElementTree.SubElement(anchor, f"{{{drawing}}}grpSp")
            group_nonvisual = ElementTree.SubElement(group, f"{{{drawing}}}nvGrpSpPr")
            ElementTree.SubElement(
                group_nonvisual,
                f"{{{drawing}}}cNvPr",
                {
                    "id": "1028",
                    "name": "PRIVATE-WORKFLOW-CONNECTOR-GROUP",
                },
            )
            ElementTree.SubElement(group_nonvisual, f"{{{drawing}}}cNvGrpSpPr")
            ElementTree.SubElement(group, f"{{{drawing}}}grpSpPr")
            connector_parent = group
        connector = ElementTree.SubElement(connector_parent, f"{{{drawing}}}cxnSp")
        nonvisual = ElementTree.SubElement(connector, f"{{{drawing}}}nvCxnSpPr")
        ElementTree.SubElement(
            nonvisual,
            f"{{{drawing}}}cNvPr",
            {
                "id": "1027",
                "name": "PRIVATE-WORKFLOW-CONNECTOR-NAME",
                "descr": "PRIVATE-WORKFLOW-CONNECTOR-DESCRIPTION",
            },
        )
        connection_properties = ElementTree.SubElement(
            nonvisual,
            f"{{{drawing}}}cNvCxnSpPr",
        )
        if attached:
            ElementTree.SubElement(
                connection_properties,
                f"{{{drawing_main}}}stCxn",
                {"id": "1025", "idx": "0"},
            )
            ElementTree.SubElement(
                connection_properties,
                f"{{{drawing_main}}}endCxn",
                {"id": "1026", "idx": "1"},
            )
        shape_properties = ElementTree.SubElement(connector, f"{{{drawing}}}spPr")
        geometry = ElementTree.SubElement(
            shape_properties,
            f"{{{drawing_main}}}prstGeom",
            {"prst": "straightConnector1"},
        )
        ElementTree.SubElement(geometry, f"{{{drawing_main}}}avLst")
        line = ElementTree.SubElement(
            shape_properties,
            f"{{{drawing_main}}}ln",
            {"w": "12700"},
        )
        fill = ElementTree.SubElement(line, f"{{{drawing_main}}}solidFill")
        ElementTree.SubElement(fill, f"{{{drawing_main}}}srgbClr", {"val": "112233"})
        ElementTree.SubElement(line, f"{{{drawing_main}}}headEnd", {"type": "triangle"})
        ElementTree.SubElement(anchor, f"{{{drawing}}}clientData")
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-drawing-connector.tmp.xlsx")


def _worksheet_drawing_connector_element(
    contents: dict[str, bytes],
) -> tuple[str, ElementTree.Element, ElementTree.Element]:
    """Return the raw drawing root and connector from the connector fixture."""
    drawing_member, _relationships_member = _worksheet_drawing_shape_part_names(contents)
    root = ElementTree.fromstring(contents[drawing_member])
    connector = next(
        (
            element
            for element in root.iter()
            if element.tag.rsplit("}", maxsplit=1)[-1] == "cxnSp"
        ),
        None,
    )
    if connector is None:
        raise ValueError("Fixture does not contain a Worksheet DrawingML connector")
    return drawing_member, root, connector


def change_worksheet_drawing_connector_presentation(path: Path) -> Path:
    """Change only a connector's private line presentation."""

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, root, connector = _worksheet_drawing_connector_element(contents)
        drawing_main = (
            _DRAWINGML_STRICT_MAIN_NS
            if connector.tag.startswith(
                "{http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing}"
            )
            else _DRAWINGML_MAIN_NS
        )
        colour = next(connector.iter(f"{{{drawing_main}}}srgbClr"), None)
        if colour is None:
            raise ValueError("Fixture does not contain a connector line colour")
        colour.set("val", "FFFFFF")
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-drawing-connector-change.tmp.xlsx")


def change_worksheet_drawing_connector_attachment(path: Path) -> Path:
    """Reattach a connector endpoint without changing any workbook cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, root, connector = _worksheet_drawing_connector_element(contents)
        drawing = connector.tag[1:].split("}", maxsplit=1)[0]
        drawing_main = (
            _DRAWINGML_STRICT_MAIN_NS
            if drawing == "http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing"
            else _DRAWINGML_MAIN_NS
        )
        connection_properties = next(
            connector.iter(f"{{{drawing}}}cNvCxnSpPr"),
            None,
        )
        if connection_properties is None:
            raise ValueError("Fixture does not contain connector connection properties")
        endpoint = connection_properties.find(f"{{{drawing_main}}}endCxn")
        if endpoint is None:
            raise ValueError("Fixture does not contain a connector end attachment")
        endpoint.set("id", "1025")
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-drawing-connector-attach.tmp.xlsx")


def renumber_worksheet_drawing_connector_identifiers(path: Path) -> Path:
    """Rewrite drawing IDs and matching connector references without a semantic change."""

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, root, connector = _worksheet_drawing_connector_element(contents)
        drawing = connector.tag[1:].split("}", maxsplit=1)[0]
        drawing_main = (
            _DRAWINGML_STRICT_MAIN_NS
            if drawing == "http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing"
            else _DRAWINGML_MAIN_NS
        )
        identifier_map: dict[str, str] = {}
        for index, nonvisual in enumerate(root.iter(f"{{{drawing}}}cNvPr"), start=900):
            previous = nonvisual.get("id")
            if previous is None:
                raise ValueError("Fixture has a nonvisual property without an identifier")
            replacement = str(index)
            identifier_map[previous] = replacement
            nonvisual.set("id", replacement)
        for endpoint_name in ("stCxn", "endCxn"):
            for endpoint in root.iter(f"{{{drawing_main}}}{endpoint_name}"):
                previous = endpoint.get("id")
                if previous not in identifier_map:
                    raise ValueError("Fixture connector references an unknown identifier")
                endpoint.set("id", identifier_map[previous])
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-drawing-connector-id.tmp.xlsx")


def corrupt_worksheet_drawing_connector_attachment(path: Path) -> Path:
    """Point a connector endpoint to an absent drawing object for fail-closed coverage."""

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, root, connector = _worksheet_drawing_connector_element(contents)
        drawing = connector.tag[1:].split("}", maxsplit=1)[0]
        drawing_main = (
            _DRAWINGML_STRICT_MAIN_NS
            if drawing == "http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing"
            else _DRAWINGML_MAIN_NS
        )
        connection_properties = next(
            connector.iter(f"{{{drawing}}}cNvCxnSpPr"),
            None,
        )
        if connection_properties is None:
            raise ValueError("Fixture does not contain connector connection properties")
        endpoint = connection_properties.find(f"{{{drawing_main}}}endCxn")
        if endpoint is None:
            raise ValueError("Fixture does not contain a connector end attachment")
        endpoint.set("id", "999999")
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-drawing-connector-corrupt.tmp.xlsx")


def make_strict_worksheet_drawing_connector_model(path: Path) -> Path:
    """Create a strict-namespace variant of the connector fixture."""
    make_worksheet_drawing_connector_model(path)
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    strict_drawing = "http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing"

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        drawing_member, drawing_relationships_member = _worksheet_drawing_shape_part_names(
            contents
        )
        worksheet_member = "xl/worksheets/sheet1.xml"
        for member, source_namespace, target_namespace in (
            (worksheet_member, _SPREADSHEETML_NS, _STRICT_SPREADSHEETML_NS),
            (drawing_member, drawing, strict_drawing),
        ):
            root = ElementTree.fromstring(contents[member])
            for element in root.iter():
                element.tag = strict_name(
                    element.tag,
                    source_namespace,
                    target_namespace,
                )
                element.tag = strict_name(
                    element.tag,
                    _DRAWINGML_MAIN_NS,
                    _DRAWINGML_STRICT_MAIN_NS,
                )
                attributes = {
                    strict_name(
                        name,
                        _DOCUMENT_RELATIONSHIPS_NS,
                        _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                    ): value
                    for name, value in element.attrib.items()
                }
                element.attrib.clear()
                element.attrib.update(attributes)
            contents[member] = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        for member in (
            _relationship_member(worksheet_member),
            drawing_relationships_member,
        ):
            relationships = ElementTree.fromstring(contents[member])
            for relationship in relationships.findall(relationship_tag):
                relationship_type = relationship.get("Type")
                if relationship_type and relationship_type.startswith(
                    _DOCUMENT_RELATIONSHIPS_NS
                ):
                    relationship.set(
                        "Type",
                        relationship_type.replace(
                            _DOCUMENT_RELATIONSHIPS_NS,
                            _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                            1,
                        ),
                    )
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".strict-worksheet-drawing-connector.tmp.xlsx")


def _worksheet_image_part_names(
    contents: dict[str, bytes],
) -> tuple[str, str, str, str, str]:
    """Return the raw members used by the native worksheet-image fixture."""
    worksheet_member = "xl/worksheets/sheet1.xml"
    drawing_member = "xl/drawings/drawing1.xml"
    vml_member = "xl/drawings/vmlDrawing1.vml"
    drawing_relationships = _relationship_member(drawing_member)
    vml_relationships = _relationship_member(vml_member)
    required = {
        worksheet_member,
        drawing_member,
        vml_member,
        drawing_relationships,
        vml_relationships,
        _relationship_member(worksheet_member),
    }
    if not required <= set(contents):
        raise ValueError("Fixture does not contain native worksheet image parts")
    return (
        worksheet_member,
        drawing_member,
        drawing_relationships,
        vml_member,
        vml_relationships,
    )


def make_worksheet_image_model(path: Path) -> Path:
    """Create anchored, background, and header/footer images outside cells."""
    make_model(path)
    content_types = _CONTENT_TYPES_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    spreadsheet = _SPREADSHEETML_NS
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_main = _DRAWINGML_MAIN_NS
    vml = "urn:schemas-microsoft-com:vml"
    vml_office = "urn:schemas-microsoft-com:office:office"
    baseline_png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9JdcAAAAAASUVORK5CYII="
    )

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def anchor_marker(
        parent: ElementTree.Element,
        name: str,
        *,
        column: int,
        row: int,
    ) -> None:
        marker = ElementTree.SubElement(parent, f"{{{drawing}}}{name}")
        ElementTree.SubElement(marker, f"{{{drawing}}}col").text = str(column)
        ElementTree.SubElement(marker, f"{{{drawing}}}colOff").text = "0"
        ElementTree.SubElement(marker, f"{{{drawing}}}row").text = str(row)
        ElementTree.SubElement(marker, f"{{{drawing}}}rowOff").text = "0"

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        drawing_member = "xl/drawings/drawing1.xml"
        vml_member = "xl/drawings/vmlDrawing1.vml"

        drawing_root = ElementTree.Element(f"{{{drawing}}}wsDr")
        anchor = ElementTree.SubElement(
            drawing_root,
            f"{{{drawing}}}twoCellAnchor",
            {"editAs": "oneCell"},
        )
        anchor_marker(anchor, "from", column=1, row=1)
        anchor_marker(anchor, "to", column=6, row=8)
        picture = ElementTree.SubElement(anchor, f"{{{drawing}}}pic")
        non_visual = ElementTree.SubElement(picture, f"{{{drawing}}}nvPicPr")
        ElementTree.SubElement(
            non_visual,
            f"{{{drawing}}}cNvPr",
            {
                "id": "4101",
                "name": "PRIVATE-ANCHORED-IMAGE-NAME",
                "descr": "PRIVATE-ANCHORED-IMAGE-DESCRIPTION",
            },
        )
        ElementTree.SubElement(non_visual, f"{{{drawing}}}cNvPicPr")
        blip_fill = ElementTree.SubElement(picture, f"{{{drawing}}}blipFill")
        ElementTree.SubElement(
            blip_fill,
            f"{{{drawing_main}}}blip",
            {f"{{{document_relationships}}}embed": "rIdFenceAnchoredImage"},
        )
        stretch = ElementTree.SubElement(blip_fill, f"{{{drawing_main}}}stretch")
        ElementTree.SubElement(stretch, f"{{{drawing_main}}}fillRect")
        picture_properties = ElementTree.SubElement(picture, f"{{{drawing}}}spPr")
        transform = ElementTree.SubElement(picture_properties, f"{{{drawing_main}}}xfrm")
        ElementTree.SubElement(
            transform,
            f"{{{drawing_main}}}off",
            {"x": "0", "y": "0"},
        )
        ElementTree.SubElement(
            transform,
            f"{{{drawing_main}}}ext",
            {"cx": "1828800", "cy": "1371600"},
        )
        geometry = ElementTree.SubElement(
            picture_properties,
            f"{{{drawing_main}}}prstGeom",
            {"prst": "rect"},
        )
        ElementTree.SubElement(geometry, f"{{{drawing_main}}}avLst")
        ElementTree.SubElement(anchor, f"{{{drawing}}}clientData")
        contents[drawing_member] = serialize(drawing_root)

        drawing_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            drawing_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceAnchoredImage",
                "Type": f"{document_relationships}/image",
                "Target": "../media/private-anchored.png",
            },
        )
        contents[_relationship_member(drawing_member)] = serialize(drawing_relationships)

        vml_root = ElementTree.Element("xml")
        shape_type = ElementTree.SubElement(
            vml_root,
            f"{{{vml}}}shapetype",
            {
                "id": "_x0000_t75",
                "coordsize": "21600,21600",
                f"{{{vml_office}}}spt": "75",
            },
        )
        ElementTree.SubElement(shape_type, f"{{{vml}}}path", {"gradientshapeok": "t"})
        watermark = ElementTree.SubElement(
            vml_root,
            f"{{{vml}}}shape",
            {
                "id": "PrivateHeaderFooterWatermark",
                "type": "#_x0000_t75",
                "style": (
                    "position:absolute;margin-left:0;margin-top:0;"
                    "width:120pt;height:40pt;z-index:1"
                ),
                f"{{{vml_office}}}spid": "_x0000_s4101",
            },
        )
        ElementTree.SubElement(
            watermark,
            f"{{{vml}}}imagedata",
            {
                f"{{{vml_office}}}relid": "rIdFenceWatermarkImage",
                f"{{{vml_office}}}title": "PRIVATE-WATERMARK-IMAGE-TITLE",
            },
        )
        contents[vml_member] = serialize(vml_root)

        vml_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        ElementTree.SubElement(
            vml_relationships,
            f"{{{package_relationships}}}Relationship",
            {
                "Id": "rIdFenceWatermarkImage",
                "Type": f"{document_relationships}/image",
                "Target": "../media/private-watermark.png",
            },
        )
        contents[_relationship_member(vml_member)] = serialize(vml_relationships)

        worksheet = ElementTree.fromstring(contents[worksheet_member])
        ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}drawing",
            {f"{{{document_relationships}}}id": "rIdFenceWorksheetDrawing"},
        )
        ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}picture",
            {f"{{{document_relationships}}}id": "rIdFenceBackgroundImage"},
        )
        header_footer = ElementTree.SubElement(worksheet, f"{{{spreadsheet}}}headerFooter")
        ElementTree.SubElement(header_footer, f"{{{spreadsheet}}}oddHeader").text = "&C&G"
        ElementTree.SubElement(
            worksheet,
            f"{{{spreadsheet}}}legacyDrawingHF",
            {f"{{{document_relationships}}}id": "rIdFenceHeaderFooterImage"},
        )
        contents[worksheet_member] = serialize(worksheet)

        worksheet_relationships = ElementTree.Element(
            f"{{{package_relationships}}}Relationships"
        )
        for identifier, relationship_type, target in (
            (
                "rIdFenceWorksheetDrawing",
                f"{document_relationships}/drawing",
                "../drawings/drawing1.xml",
            ),
            (
                "rIdFenceBackgroundImage",
                f"{document_relationships}/image",
                "../media/private-background.png",
            ),
            (
                "rIdFenceHeaderFooterImage",
                f"{document_relationships}/vmlDrawing",
                "../drawings/vmlDrawing1.vml",
            ),
        ):
            ElementTree.SubElement(
                worksheet_relationships,
                f"{{{package_relationships}}}Relationship",
                {"Id": identifier, "Type": relationship_type, "Target": target},
            )
        contents[_relationship_member(worksheet_member)] = serialize(worksheet_relationships)

        contents["xl/media/private-anchored.png"] = baseline_png
        contents["xl/media/private-background.png"] = baseline_png
        contents["xl/media/private-watermark.png"] = baseline_png

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        default_tag = f"{{{content_types}}}Default"
        override_tag = f"{{{content_types}}}Override"
        ElementTree.SubElement(
            types,
            default_tag,
            {"Extension": "png", "ContentType": "image/png"},
        )
        for part_name, content_type in (
            (
                "/xl/drawings/drawing1.xml",
                "application/vnd.openxmlformats-officedocument.drawing+xml",
            ),
            (
                "/xl/drawings/vmlDrawing1.vml",
                "application/vnd.openxmlformats-officedocument.vmlDrawing",
            ),
        ):
            ElementTree.SubElement(
                types,
                override_tag,
                {"PartName": part_name, "ContentType": content_type},
            )
        contents["[Content_Types].xml"] = serialize(types)

    return _rewrite_archive(path, mutate, ".worksheet-image-model.tmp.xlsx")


def change_worksheet_image_presentation(path: Path) -> Path:
    """Move one anchored image without changing worksheet cells or image bytes."""
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"

    def mutate(contents: dict[str, bytes]) -> None:
        _worksheet, drawing_member, _drawing_rels, _vml, _vml_rels = (
            _worksheet_image_part_names(contents)
        )
        root = ElementTree.fromstring(contents[drawing_member])
        marker = next(root.iter(f"{{{drawing}}}from"))
        column_offset = marker.find(f"{{{drawing}}}colOff")
        if column_offset is None:
            raise ValueError("Fixture picture anchor is missing its column offset")
        column_offset.text = "457200"
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-image-presentation.tmp.xlsx")


def change_worksheet_image_payload(path: Path) -> Path:
    """Replace one valid native image payload without changing declarations."""
    candidate_png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAIAAACQd1PeAAAADUlEQVR42mP8z8DwHwAFgAI/ScLkYQAAAABJRU5ErkJggg=="
    )

    def mutate(contents: dict[str, bytes]) -> None:
        contents["xl/media/private-anchored.png"] = candidate_png

    return _rewrite_archive(path, mutate, ".worksheet-image-payload.tmp.xlsx")


def externalize_worksheet_image_relationship(path: Path) -> Path:
    """Make an anchored picture target external without ever following it."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS

    def mutate(contents: dict[str, bytes]) -> None:
        _worksheet, _drawing, drawing_relationships, _vml, _vml_relationships = (
            _worksheet_image_part_names(contents)
        )
        relationships = ElementTree.fromstring(contents[drawing_relationships])
        image_relationship = next(
            relationship
            for relationship in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if relationship.get("Id") == "rIdFenceAnchoredImage"
        )
        image_relationship.set(
            "Target",
            "https://example.invalid/PRIVATE-EXTERNAL-WORKSHEET-IMAGE",
        )
        image_relationship.set("TargetMode", "External")
        contents[drawing_relationships] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-image-external.tmp.xlsx")


def rebind_worksheet_image_source_relationship(path: Path) -> Path:
    """Point a picture declaration at an equivalent new DrawingML part."""
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    content_types = _CONTENT_TYPES_NS

    def mutate(contents: dict[str, bytes]) -> None:
        (
            worksheet_member,
            drawing_member,
            drawing_relationships_member,
            _vml_member,
            _vml_relationships_member,
        ) = _worksheet_image_part_names(contents)
        rebound_drawing_member = "xl/drawings/drawing2.xml"
        contents[rebound_drawing_member] = contents[drawing_member]
        contents[_relationship_member(rebound_drawing_member)] = contents[
            drawing_relationships_member
        ]

        relationships_member = _relationship_member(worksheet_member)
        relationships = ElementTree.fromstring(contents[relationships_member])
        drawing_relationship = next(
            relationship
            for relationship in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            )
            if relationship.get("Id") == "rIdFenceWorksheetDrawing"
        )
        drawing_relationship.set("Target", "../drawings/drawing2.xml")
        contents[relationships_member] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        ElementTree.SubElement(
            types,
            f"{{{content_types}}}Override",
            {
                "PartName": "/xl/drawings/drawing2.xml",
                "ContentType": (
                    "application/vnd.openxmlformats-officedocument.drawing+xml"
                ),
            },
        )
        contents["[Content_Types].xml"] = ElementTree.tostring(
            types,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-image-rebind.tmp.xlsx")


def renumber_worksheet_image_identifiers(path: Path) -> Path:
    """Rewrite volatile drawing, VML, and relationship IDs consistently."""
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_main = _DRAWINGML_MAIN_NS
    document_relationships = _DOCUMENT_RELATIONSHIPS_NS
    package_relationships = _PACKAGE_RELATIONSHIPS_NS
    vml = "urn:schemas-microsoft-com:vml"
    vml_office = "urn:schemas-microsoft-com:office:office"

    def mutate(contents: dict[str, bytes]) -> None:
        (
            worksheet_member,
            drawing_member,
            drawing_relationships_member,
            vml_member,
            vml_relationships_member,
        ) = _worksheet_image_part_names(contents)
        worksheet_mapping = {
            "rIdFenceWorksheetDrawing": "rIdFenceWorksheetDrawingRenumbered",
            "rIdFenceBackgroundImage": "rIdFenceBackgroundImageRenumbered",
            "rIdFenceHeaderFooterImage": "rIdFenceHeaderFooterImageRenumbered",
        }
        drawing_mapping = {"rIdFenceAnchoredImage": "rIdFenceAnchoredImageRenumbered"}
        vml_mapping = {"rIdFenceWatermarkImage": "rIdFenceWatermarkImageRenumbered"}

        worksheet = ElementTree.fromstring(contents[worksheet_member])
        for element in worksheet:
            relationship_id = element.get(f"{{{document_relationships}}}id")
            if relationship_id in worksheet_mapping:
                element.set(
                    f"{{{document_relationships}}}id",
                    worksheet_mapping[relationship_id],
                )
        contents[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

        drawing_root = ElementTree.fromstring(contents[drawing_member])
        for index, non_visual in enumerate(
            drawing_root.iter(f"{{{drawing}}}cNvPr"),
            start=9100,
        ):
            non_visual.set("id", str(index))
        blip = next(drawing_root.iter(f"{{{drawing_main}}}blip"))
        blip.set(
            f"{{{document_relationships}}}embed",
            drawing_mapping["rIdFenceAnchoredImage"],
        )
        contents[drawing_member] = ElementTree.tostring(
            drawing_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        vml_root = ElementTree.fromstring(contents[vml_member])
        shape = next(vml_root.iter(f"{{{vml}}}shape"))
        shape.set("id", "PrivateHeaderFooterWatermarkRenumbered")
        shape.set(f"{{{vml_office}}}spid", "_x0000_s9101")
        image_data = next(vml_root.iter(f"{{{vml}}}imagedata"))
        image_data.set(
            f"{{{vml_office}}}relid",
            vml_mapping["rIdFenceWatermarkImage"],
        )
        contents[vml_member] = ElementTree.tostring(
            vml_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        for member, mapping in (
            (_relationship_member(worksheet_member), worksheet_mapping),
            (drawing_relationships_member, drawing_mapping),
            (vml_relationships_member, vml_mapping),
        ):
            relationships = ElementTree.fromstring(contents[member])
            for relationship in relationships.findall(
                f"{{{package_relationships}}}Relationship"
            ):
                identifier = relationship.get("Id")
                if identifier in mapping:
                    relationship.set("Id", mapping[identifier])
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".worksheet-image-ids.tmp.xlsx")


def corrupt_worksheet_image_drawing_root(path: Path) -> Path:
    """Replace an anchored-picture root to exercise fail-closed coverage."""
    def mutate(contents: dict[str, bytes]) -> None:
        _worksheet, drawing_member, _drawing_rels, _vml, _vml_rels = (
            _worksheet_image_part_names(contents)
        )
        root = ElementTree.fromstring(contents[drawing_member])
        root.tag = "notWorksheetDrawing"
        contents[drawing_member] = ElementTree.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".worksheet-image-corrupt.tmp.xlsx")


def make_strict_worksheet_image_model(path: Path) -> Path:
    """Create a strict-namespace variant of the native image fixture."""
    make_worksheet_image_model(path)
    drawing = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    strict_drawing = "http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing"

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        (
            worksheet_member,
            drawing_member,
            drawing_relationships_member,
            vml_member,
            vml_relationships_member,
        ) = _worksheet_image_part_names(contents)
        for member, source_namespace, target_namespace in (
            (worksheet_member, _SPREADSHEETML_NS, _STRICT_SPREADSHEETML_NS),
            (drawing_member, drawing, strict_drawing),
        ):
            root = ElementTree.fromstring(contents[member])
            for element in root.iter():
                element.tag = strict_name(
                    element.tag,
                    source_namespace,
                    target_namespace,
                )
                element.tag = strict_name(
                    element.tag,
                    _DRAWINGML_MAIN_NS,
                    _DRAWINGML_STRICT_MAIN_NS,
                )
                attributes = {
                    strict_name(
                        name,
                        _DOCUMENT_RELATIONSHIPS_NS,
                        _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                    ): value
                    for name, value in element.attrib.items()
                }
                element.attrib.clear()
                element.attrib.update(attributes)
            contents[member] = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        for member in (
            _relationship_member(worksheet_member),
            drawing_relationships_member,
            vml_relationships_member,
        ):
            relationships = ElementTree.fromstring(contents[member])
            for relationship in relationships.findall(relationship_tag):
                relationship_type = relationship.get("Type")
                if relationship_type and relationship_type.startswith(
                    _DOCUMENT_RELATIONSHIPS_NS
                ):
                    relationship.set(
                        "Type",
                        relationship_type.replace(
                            _DOCUMENT_RELATIONSHIPS_NS,
                            _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                            1,
                        ),
                    )
            contents[member] = ElementTree.tostring(
                relationships,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".strict-worksheet-image-model.tmp.xlsx")


def _rich_text_shared_string_root(
    contents: dict[str, bytes],
) -> ElementTree.Element:
    """Return the raw shared-string table used by the rich-text fixture."""
    return ElementTree.fromstring(contents["xl/sharedStrings.xml"])


def make_rich_text_run_model(path: Path) -> Path:
    """Create shared and inline strings whose character formatting matters."""
    workbook = Workbook()
    review = workbook.active
    review.title = "Review"
    review["A1"] = "PRIVATE-RICH-HEADER"
    review["A3"] = "Decision: DO NOT APPROVE"
    review["B3"] = "Inline: PRIVATE-INLINE-HOLD"
    review["C3"] = "=1+1"
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
        value_tag = f"{{{_SPREADSHEETML_NS}}}v"
        inline_string_tag = f"{{{_SPREADSHEETML_NS}}}is"
        shared_table_tag = f"{{{_SPREADSHEETML_NS}}}sst"
        shared_item_tag = f"{{{_SPREADSHEETML_NS}}}si"
        run_tag = f"{{{_SPREADSHEETML_NS}}}r"
        run_properties_tag = f"{{{_SPREADSHEETML_NS}}}rPr"
        text_tag = f"{{{_SPREADSHEETML_NS}}}t"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        size_tag = f"{{{_SPREADSHEETML_NS}}}sz"
        font_tag = f"{{{_SPREADSHEETML_NS}}}rFont"
        worksheet = _inputs_worksheet_root(contents)

        def fixture_cell(coordinate: str) -> ElementTree.Element:
            return next(
                current
                for current in worksheet.iter(cell_tag)
                if current.get("r") == coordinate
            )

        def clear_children(cell: ElementTree.Element) -> None:
            for child in list(cell):
                cell.remove(child)

        def rich_run(
            parent: ElementTree.Element,
            text: str,
            *,
            color: str | None = None,
        ) -> ElementTree.Element:
            run = ElementTree.SubElement(parent, run_tag)
            if color is not None:
                properties = ElementTree.SubElement(run, run_properties_tag)
                ElementTree.SubElement(properties, size_tag, {"val": "11"})
                ElementTree.SubElement(properties, color_tag, {"rgb": color})
                ElementTree.SubElement(
                    properties,
                    font_tag,
                    {"val": "PRIVATE-RICH-RUN-FONT"},
                )
            text_element = ElementTree.SubElement(run, text_tag)
            if text.endswith(" "):
                text_element.set(
                    "{http://www.w3.org/XML/1998/namespace}space",
                    "preserve",
                )
            text_element.text = text
            return run

        shared = ElementTree.Element(
            shared_table_tag,
            {"count": "1", "uniqueCount": "1"},
        )
        shared_item = ElementTree.SubElement(shared, shared_item_tag)
        rich_run(shared_item, "Decision: ")
        rich_run(shared_item, "DO NOT APPROVE", color="FF000000")
        contents["xl/sharedStrings.xml"] = ElementTree.tostring(
            shared,
            encoding="utf-8",
            xml_declaration=True,
        )

        shared_cell = fixture_cell("A3")
        shared_cell.attrib.clear()
        shared_cell.set("r", "A3")
        shared_cell.set("t", "s")
        clear_children(shared_cell)
        ElementTree.SubElement(shared_cell, value_tag).text = "0"

        inline_cell = fixture_cell("B3")
        inline_cell.attrib.clear()
        inline_cell.set("r", "B3")
        inline_cell.set("t", "inlineStr")
        clear_children(inline_cell)
        inline_string = ElementTree.SubElement(inline_cell, inline_string_tag)
        rich_run(inline_string, "Inline: ")
        rich_run(inline_string, "PRIVATE-INLINE-HOLD", color="FF334455")
        _save_inputs_worksheet(contents, worksheet)

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        ElementTree.SubElement(
            relationships,
            relationship_tag,
            {
                "Id": "rIdFormulaFenceRichTextSharedStrings",
                "Type": (
                    "http://schemas.openxmlformats.org/officeDocument/2006/"
                    "relationships/sharedStrings"
                ),
                "Target": "sharedStrings.xml",
            },
        )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        content_types_namespace = (
            "http://schemas.openxmlformats.org/package/2006/content-types"
        )
        override_tag = f"{{{content_types_namespace}}}Override"
        content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
        ElementTree.SubElement(
            content_types,
            override_tag,
            {
                "PartName": "/xl/sharedStrings.xml",
                "ContentType": (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sharedStrings+xml"
                ),
            },
        )
        contents["[Content_Types].xml"] = ElementTree.tostring(
            content_types,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".rich-text-run.tmp.xlsx")


def change_rich_text_run_color(path: Path) -> Path:
    """Hide the shared-string warning with a character-level colour change."""

    def mutate(contents: dict[str, bytes]) -> None:
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        shared = _rich_text_shared_string_root(contents)
        color = next(current for current in shared.iter(color_tag))
        color.set("rgb", "FFFFFFFF")
        contents["xl/sharedStrings.xml"] = ElementTree.tostring(
            shared,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".rich-text-run-colour.tmp.xlsx")


def change_inline_rich_text_run_color(path: Path) -> Path:
    """Hide an inline warning without changing its concatenated cell text."""

    def mutate(contents: dict[str, bytes]) -> None:
        cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        worksheet = _inputs_worksheet_root(contents)
        inline_cell = next(
            current
            for current in worksheet.iter(cell_tag)
            if current.get("r") == "B3"
        )
        color = next(current for current in inline_cell.iter(color_tag))
        color.set("rgb", "FFFFFFFF")
        _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".inline-rich-text-run-colour.tmp.xlsx")


def change_rich_text_run_boundary(path: Path) -> Path:
    """Move a shared warning's styled boundary while leaving its text unchanged."""

    def mutate(contents: dict[str, bytes]) -> None:
        run_tag = f"{{{_SPREADSHEETML_NS}}}r"
        text_tag = f"{{{_SPREADSHEETML_NS}}}t"
        shared = _rich_text_shared_string_root(contents)
        runs = list(shared.iter(run_tag))
        runs[0].find(text_tag).text = "Decision: D"
        runs[1].find(text_tag).text = "O NOT APPROVE"
        contents["xl/sharedStrings.xml"] = ElementTree.tostring(
            shared,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".rich-text-run-boundary.tmp.xlsx")


def change_rich_text_run_text_only(path: Path) -> Path:
    """Edit a rich string's displayed text without changing its run controls."""

    def mutate(contents: dict[str, bytes]) -> None:
        run_tag = f"{{{_SPREADSHEETML_NS}}}r"
        text_tag = f"{{{_SPREADSHEETML_NS}}}t"
        shared = _rich_text_shared_string_root(contents)
        runs = list(shared.iter(run_tag))
        runs[1].find(text_tag).text = "APPROVE WITH CHANGES"
        contents["xl/sharedStrings.xml"] = ElementTree.tostring(
            shared,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".rich-text-run-text.tmp.xlsx")


def rewrite_shared_rich_text_as_inline(path: Path) -> Path:
    """Store equivalent shared rich text inline to exercise storage normalization."""

    def mutate(contents: dict[str, bytes]) -> None:
        cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
        inline_string_tag = f"{{{_SPREADSHEETML_NS}}}is"
        shared_item_tag = f"{{{_SPREADSHEETML_NS}}}si"
        worksheet = _inputs_worksheet_root(contents)
        shared_cell = next(
            current
            for current in worksheet.iter(cell_tag)
            if current.get("r") == "A3"
        )
        shared = _rich_text_shared_string_root(contents)
        shared_item = next(shared.iter(shared_item_tag))
        inline_string = ElementTree.Element(inline_string_tag)
        for child in shared_item:
            inline_string.append(
                ElementTree.fromstring(ElementTree.tostring(child, encoding="utf-8"))
            )
        shared_cell.attrib.clear()
        shared_cell.set("r", "A3")
        shared_cell.set("t", "inlineStr")
        for child in list(shared_cell):
            shared_cell.remove(child)
        shared_cell.append(inline_string)
        _save_inputs_worksheet(contents, worksheet)

    return _rewrite_archive(path, mutate, ".rich-text-shared-to-inline.tmp.xlsx")


def normalize_rich_text_run_property_spelling(path: Path) -> Path:
    """Use equivalent rich-run ordering, color case, and Boolean spelling."""

    def mutate(contents: dict[str, bytes]) -> None:
        bold_tag = f"{{{_SPREADSHEETML_NS}}}b"
        color_tag = f"{{{_SPREADSHEETML_NS}}}color"
        run_properties_tag = f"{{{_SPREADSHEETML_NS}}}rPr"
        shared = _rich_text_shared_string_root(contents)
        properties = next(current for current in shared.iter(run_properties_tag))
        color = next(current for current in properties.iter(color_tag))
        color.set("rgb", "ff000000")
        ElementTree.SubElement(properties, bold_tag, {"val": "false"})
        children = list(properties)
        for child in children:
            properties.remove(child)
        for child in reversed(children):
            properties.append(child)
        contents["xl/sharedStrings.xml"] = ElementTree.tostring(
            shared,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".rich-text-run-noise.tmp.xlsx")


def corrupt_rich_text_run(path: Path) -> Path:
    """Inject unsupported private run metadata for fail-closed coverage."""

    def mutate(contents: dict[str, bytes]) -> None:
        run_properties_tag = f"{{{_SPREADSHEETML_NS}}}rPr"
        shared = _rich_text_shared_string_root(contents)
        properties = next(current for current in shared.iter(run_properties_tag))
        properties.set(
            "{urn:formulafence:test}privateUnsupported",
            "PRIVATE-UNSUPPORTED-RUN-CONTROL",
        )
        contents["xl/sharedStrings.xml"] = ElementTree.tostring(
            shared,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".rich-text-run-corrupt.tmp.xlsx")


def make_custom_workbook_view_model(path: Path) -> Path:
    """Create legacy Excel Custom Views with private alternate worksheet state."""
    workbook = Workbook()
    report = workbook.active
    report.title = "Private Custom View Report"
    report.append(["Region", "Status", "Amount"])
    report.append(["North", "Open", 100])
    report.append(["South", "Hold", 200])
    audit = workbook.create_sheet("Private Custom View Audit")
    audit.append(["Metric", "Value"])
    audit.append(["Control total", 300])
    workbook.save(path)

    view_a_guid = "{01234567-89AB-CDEF-0123-456789ABCDEF}"
    view_b_guid = "{89ABCDEF-0123-4567-89AB-CDEF01234567}"

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_standard_view(
        container: ElementTree.Element,
        *,
        guid: str,
        material: bool = False,
        scale: str | None = None,
    ) -> None:
        attributes = {"guid": guid}
        if material:
            attributes.update(
                {
                    "showFormulas": "1",
                    "showGridLines": "0",
                    "hiddenRows": "true",
                    "hiddenColumns": "1",
                    "printArea": "1",
                    "filter": "true",
                    "showAutoFilter": "0",
                    "topLeftCell": "B2",
                }
            )
        if scale is not None:
            attributes["scale"] = scale
        view = ElementTree.SubElement(
            container,
            f"{{{_SPREADSHEETML_NS}}}customSheetView",
            attributes,
        )
        if not material:
            return
        ElementTree.SubElement(
            view,
            f"{{{_SPREADSHEETML_NS}}}pane",
            {"xSplit": "1", "state": "frozen"},
        )
        ElementTree.SubElement(
            view,
            f"{{{_SPREADSHEETML_NS}}}selection",
            {"activeCell": "B2", "sqref": "B2"},
        )
        ElementTree.SubElement(
            view,
            f"{{{_SPREADSHEETML_NS}}}pageMargins",
            {
                "left": "0.3",
                "right": "0.3",
                "top": "0.5",
                "bottom": "0.5",
                "header": "0.2",
                "footer": "0.2",
            },
        )
        auto_filter = ElementTree.SubElement(
            view,
            f"{{{_SPREADSHEETML_NS}}}autoFilter",
            {"ref": "A1:C3"},
        )
        filter_column = ElementTree.SubElement(
            auto_filter,
            f"{{{_SPREADSHEETML_NS}}}filterColumn",
            {"colId": "1"},
        )
        filters = ElementTree.SubElement(
            filter_column,
            f"{{{_SPREADSHEETML_NS}}}filters",
        )
        ElementTree.SubElement(
            filters,
            f"{{{_SPREADSHEETML_NS}}}filter",
            {"val": "PRIVATE-CUSTOM-VIEW-FILTER"},
        )

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_xml = ElementTree.fromstring(contents["xl/workbook.xml"])
        workbook_views = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}customWorkbookViews"
        )
        ElementTree.SubElement(
            workbook_views,
            f"{{{_SPREADSHEETML_NS}}}customWorkbookView",
            {
                "name": "PRIVATE Executive View",
                "guid": view_a_guid,
                "windowWidth": "20000",
                "windowHeight": "12000",
                "activeSheetId": "1",
                "includeHiddenRowCol": "1",
                "includePrintSettings": "true",
                "showObjects": "none",
                "showComments": "commNone",
                "showFormulaBar": "0",
                "showSheetTabs": "false",
            },
        )
        ElementTree.SubElement(
            workbook_views,
            f"{{{_SPREADSHEETML_NS}}}customWorkbookView",
            {
                "name": "PRIVATE Print View",
                "guid": view_b_guid,
                "windowWidth": "20000",
                "windowHeight": "12000",
                "activeSheetId": "2",
                "includeHiddenRowCol": "true",
                "includePrintSettings": "1",
                "showObjects": "placeholders",
            },
        )
        workbook_xml.append(workbook_views)
        contents["xl/workbook.xml"] = serialize(workbook_xml)

        for member, material, scale in (
            ("xl/worksheets/sheet1.xml", True, "80"),
            ("xl/worksheets/sheet2.xml", False, None),
        ):
            worksheet = ElementTree.fromstring(contents[member])
            custom_sheet_views = ElementTree.Element(
                f"{{{_SPREADSHEETML_NS}}}customSheetViews"
            )
            add_standard_view(
                custom_sheet_views,
                guid=view_a_guid,
                material=material,
            )
            add_standard_view(
                custom_sheet_views,
                guid=view_b_guid,
                scale=scale,
            )
            sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
            sheet_data = worksheet.find(sheet_data_tag)
            if sheet_data is None:
                raise ValueError("Could not locate Custom View fixture sheetData")
            worksheet.insert(list(worksheet).index(sheet_data) + 1, custom_sheet_views)
            contents[member] = serialize(worksheet)

    return _rewrite_archive(path, mutate, ".custom-workbook-views.tmp.xlsx")


def change_custom_workbook_view_filter(path: Path) -> Path:
    """Change a private legacy Custom View filter without touching cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        filter_tag = f"{{{_SPREADSHEETML_NS}}}filter"
        next(worksheet.iter(filter_tag)).set(
            "val", "CANDIDATE-PRIVATE-CUSTOM-VIEW-FILTER"
        )
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".custom-workbook-view-filter.tmp.xlsx")


def normalize_custom_workbook_view_identifiers(path: Path) -> Path:
    """Rewrite Custom View GUIDs and use equivalent scalar spellings."""

    replacements = {
        "{01234567-89AB-CDEF-0123-456789ABCDEF}": (
            "{11111111-2222-3333-4444-555555555555}"
        ),
        "{89ABCDEF-0123-4567-89AB-CDEF01234567}": (
            "{66666666-7777-8888-9999-AAAAAAAAAAAA}"
        ),
    }

    def mutate(contents: dict[str, bytes]) -> None:
        for member in ("xl/workbook.xml", "xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml"):
            root = ElementTree.fromstring(contents[member])
            for element in root.iter():
                if element.get("guid") in replacements:
                    element.set("guid", replacements[element.get("guid")])
            if member == "xl/workbook.xml":
                sheets = list(root.iter(f"{{{_SPREADSHEETML_NS}}}sheet"))
                sheets[0].set("sheetId", "17")
                sheets[1].set("sheetId", "42")
                views = list(
                    root.iter(f"{{{_SPREADSHEETML_NS}}}customWorkbookView")
                )
                views[0].set("activeSheetId", "017")
                views[0].set("includeHiddenRowCol", "true")
                views[0].set("showFormulaBar", "false")
                views[0].set("showSheetTabs", "0")
                views[1].set("activeSheetId", "042")
                views[1].set("includePrintSettings", "true")
            elif member == "xl/worksheets/sheet1.xml":
                views = list(root.iter(f"{{{_SPREADSHEETML_NS}}}customSheetView"))
                views[0].set("showFormulas", "true")
                views[0].set("showGridLines", "false")
                views[0].set("hiddenRows", "1")
                views[0].set("hiddenColumns", "true")
                views[0].set("printArea", "true")
                views[0].set("filter", "1")
                views[1].set("scale", "080")
            contents[member] = ElementTree.tostring(
                root,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".custom-workbook-view-noise.tmp.xlsx")


def make_strict_custom_workbook_view_model(path: Path) -> Path:
    """Create a Strict SpreadsheetML Custom View fixture."""
    make_custom_workbook_view_model(path)

    def strict_name(name: str, source_namespace: str, target_namespace: str) -> str:
        prefix = f"{{{source_namespace}}}"
        if name.startswith(prefix):
            return f"{{{target_namespace}}}{name[len(prefix):]}"
        return name

    def mutate(contents: dict[str, bytes]) -> None:
        workbook_member = "xl/workbook.xml"
        workbook = ElementTree.fromstring(contents[workbook_member])
        for element in workbook.iter():
            element.tag = strict_name(
                element.tag,
                _SPREADSHEETML_NS,
                _STRICT_SPREADSHEETML_NS,
            )
            attributes = {
                strict_name(
                    name,
                    _DOCUMENT_RELATIONSHIPS_NS,
                    _STRICT_DOCUMENT_RELATIONSHIPS_NS,
                ): value
                for name, value in element.attrib.items()
            }
            element.attrib.clear()
            element.attrib.update(attributes)
        contents[workbook_member] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
        for relationship in relationships.findall(relationship_tag):
            if (relationship.get("Type") or "").casefold().endswith("/worksheet"):
                relationship.set(
                    "Type",
                    f"{_STRICT_DOCUMENT_RELATIONSHIPS_NS}/worksheet",
                )
        contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
            relationships,
            encoding="utf-8",
            xml_declaration=True,
        )

        for member in sorted(
            name
            for name in contents
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ):
            worksheet = ElementTree.fromstring(contents[member])
            for element in worksheet.iter():
                element.tag = strict_name(
                    element.tag,
                    _SPREADSHEETML_NS,
                    _STRICT_SPREADSHEETML_NS,
                )
            contents[member] = ElementTree.tostring(
                worksheet,
                encoding="utf-8",
                xml_declaration=True,
            )

    return _rewrite_archive(path, mutate, ".strict-custom-workbook-view.tmp.xlsx")


def corrupt_custom_workbook_view_control(path: Path) -> Path:
    """Inject malformed Custom View scalar metadata for fail-closed coverage."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        next(worksheet.iter(f"{{{_SPREADSHEETML_NS}}}customSheetView")).set(
            "scale", "4294967296"
        )
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".custom-workbook-view-corrupt.tmp.xlsx")


def unbind_custom_workbook_view(path: Path) -> Path:
    """Break one Custom View GUID association without exposing sheet data."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet2.xml"])
        views = list(worksheet.iter(f"{{{_SPREADSHEETML_NS}}}customSheetView"))
        views[1].set("guid", "{BBBBBBBB-CCCC-DDDD-EEEE-FFFFFFFFFFFF}")
        contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".custom-workbook-view-unbound.tmp.xlsx")


def make_chart_sheet_custom_workbook_view_model(path: Path) -> Path:
    """Create a Custom View that includes a chart sheet's alternate print state."""
    workbook = Workbook()
    data = workbook.active
    data.title = "Chart View Data"
    data.append(["Period", "Amount"])
    data.append(["January", 10])
    data.append(["February", 20])
    chart = BarChart()
    chart.add_data(Reference(data, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chartsheet = workbook.create_chartsheet("Private Custom View Chart")
    chartsheet.add_chart(chart)
    workbook.save(path)

    guid = "{13572468-2468-1357-2468-135724681357}"

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(contents: dict[str, bytes]) -> None:
        sheet_tag = f"{{{_SPREADSHEETML_NS}}}sheet"
        workbook_xml = ElementTree.fromstring(contents["xl/workbook.xml"])
        sheets = list(workbook_xml.iter(sheet_tag))
        if len(sheets) != 2 or sheets[1].get("sheetId") is None:
            raise ValueError("Could not locate chart-sheet Custom View fixture sheets")
        custom_workbook_views = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}customWorkbookViews"
        )
        ElementTree.SubElement(
            custom_workbook_views,
            f"{{{_SPREADSHEETML_NS}}}customWorkbookView",
            {
                "name": "PRIVATE Chart Print View",
                "guid": guid,
                "windowWidth": "20000",
                "windowHeight": "12000",
                "activeSheetId": sheets[1].get("sheetId"),
                "includePrintSettings": "true",
                "showObjects": "none",
            },
        )
        workbook_xml.append(custom_workbook_views)
        contents["xl/workbook.xml"] = serialize(workbook_xml)

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        worksheet_views = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}customSheetViews"
        )
        ElementTree.SubElement(
            worksheet_views,
            f"{{{_SPREADSHEETML_NS}}}customSheetView",
            {"guid": guid},
        )
        sheet_data = worksheet.find(f"{{{_SPREADSHEETML_NS}}}sheetData")
        if sheet_data is None:
            raise ValueError("Could not locate chart Custom View fixture sheetData")
        worksheet.insert(list(worksheet).index(sheet_data) + 1, worksheet_views)
        contents["xl/worksheets/sheet1.xml"] = serialize(worksheet)

        chart_root = ElementTree.fromstring(contents["xl/chartsheets/sheet1.xml"])
        chart_views = ElementTree.Element(f"{{{_SPREADSHEETML_NS}}}customSheetViews")
        chart_view = ElementTree.SubElement(
            chart_views,
            f"{{{_SPREADSHEETML_NS}}}customSheetView",
            {"guid": guid, "scale": "75", "zoomToFit": "true"},
        )
        ElementTree.SubElement(
            chart_view,
            f"{{{_SPREADSHEETML_NS}}}pageMargins",
            {
                "left": "0.3",
                "right": "0.3",
                "top": "0.5",
                "bottom": "0.5",
                "header": "0.2",
                "footer": "0.2",
            },
        )
        ElementTree.SubElement(
            chart_view,
            f"{{{_SPREADSHEETML_NS}}}pageSetup",
            {"orientation": "landscape"},
        )
        sheet_views = chart_root.find(f"{{{_SPREADSHEETML_NS}}}sheetViews")
        if sheet_views is None:
            raise ValueError("Could not locate chart Custom View sheetViews")
        chart_root.insert(list(chart_root).index(sheet_views) + 1, chart_views)
        contents["xl/chartsheets/sheet1.xml"] = serialize(chart_root)

    return _rewrite_archive(path, mutate, ".chart-custom-workbook-view.tmp.xlsx")


def change_chart_sheet_custom_workbook_view_scale(path: Path) -> Path:
    """Change a chart sheet's alternate Custom View scale without touching charts."""

    def mutate(contents: dict[str, bytes]) -> None:
        chart_root = ElementTree.fromstring(contents["xl/chartsheets/sheet1.xml"])
        view = next(
            chart_root.iter(f"{{{_SPREADSHEETML_NS}}}customSheetView")
        )
        view.set("scale", "80")
        contents["xl/chartsheets/sheet1.xml"] = ElementTree.tostring(
            chart_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".chart-custom-workbook-view-scale.tmp.xlsx")


def corrupt_chart_sheet_custom_workbook_view_page_setup(path: Path) -> Path:
    """Inject malformed chart-sheet Custom View print metadata."""

    def mutate(contents: dict[str, bytes]) -> None:
        chart_root = ElementTree.fromstring(contents["xl/chartsheets/sheet1.xml"])
        page_setup = next(chart_root.iter(f"{{{_SPREADSHEETML_NS}}}pageSetup"))
        page_setup.set("orientation", "PRIVATE-INVALID-CHART-ORIENTATION")
        contents["xl/chartsheets/sheet1.xml"] = ElementTree.tostring(
            chart_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".chart-custom-workbook-view-corrupt.tmp.xlsx")


def make_empty_chart_sheet_custom_view_container_model(path: Path) -> Path:
    """Create the schema-valid empty chart-sheet Custom Views container case."""
    make_chart_sheet_custom_workbook_view_model(path)

    def remove_custom_container(root: ElementTree.Element, name: str) -> None:
        for child in list(root):
            if child.tag == f"{{{_SPREADSHEETML_NS}}}{name}":
                root.remove(child)

    def mutate(contents: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(contents["xl/workbook.xml"])
        remove_custom_container(workbook, "customWorkbookViews")
        contents["xl/workbook.xml"] = ElementTree.tostring(
            workbook,
            encoding="utf-8",
            xml_declaration=True,
        )

        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        remove_custom_container(worksheet, "customSheetViews")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

        chart_root = ElementTree.fromstring(contents["xl/chartsheets/sheet1.xml"])
        custom_views = next(
            chart_root.iter(f"{{{_SPREADSHEETML_NS}}}customSheetViews")
        )
        for view in list(custom_views):
            custom_views.remove(view)
        contents["xl/chartsheets/sheet1.xml"] = ElementTree.tostring(
            chart_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".empty-chart-custom-view.tmp.xlsx")


def make_named_sheet_view_model(path: Path, *, table_owned: bool = False) -> Path:
    """Create modern alternate filter/sort views with private OOXML settings."""
    workbook = Workbook()
    report = workbook.active
    report.title = "Named View Report"
    report.append(["Region", "Department", "Amount"])
    report.append(["North", "Operations", 100])
    report.append(["South", "Sales", 200])
    report.append(["North", "Finance", 300])
    report.append(["West", "Legal", 400])
    if table_owned:
        report.add_table(Table(displayName="NamedViewTable", ref="A1:C5"))
    workbook.save(path)

    content_types = "http://schemas.openxmlformats.org/package/2006/content-types"
    named_view_relationship = (
        "http://schemas.microsoft.com/office/2019/04/relationships/namedSheetView"
    )

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_view(
        root: ElementTree.Element,
        *,
        name: str,
        view_id: str,
        filter_id: str,
        region: str,
        sort_ref: str,
    ) -> None:
        view = ElementTree.SubElement(
            root,
            f"{{{_NAMED_SHEET_VIEW_NS}}}namedSheetView",
            {"name": name, "id": view_id},
        )
        named_filter = ElementTree.SubElement(
            view,
            f"{{{_NAMED_SHEET_VIEW_NS}}}nsvFilter",
            {
                "filterId": filter_id,
                "ref": "A1:C5",
                "tableId": "1" if table_owned else "0",
            },
        )
        column_filter = ElementTree.SubElement(
            named_filter,
            f"{{{_NAMED_SHEET_VIEW_NS}}}columnFilter",
            {"colId": "1", "id": "{11111111-1111-1111-1111-111111111111}"},
        )
        filter_column = ElementTree.SubElement(
            column_filter,
            f"{{{_NAMED_SHEET_VIEW_NS}}}filter",
            {"colId": "1"},
        )
        filters = ElementTree.SubElement(
            filter_column,
            f"{{{_SPREADSHEETML_NS}}}filters",
            {"blank": "false", "calendarType": "none"},
        )
        ElementTree.SubElement(
            filters,
            f"{{{_SPREADSHEETML_NS}}}filter",
            {"val": region},
        )
        sort_rules = ElementTree.SubElement(
            named_filter,
            f"{{{_NAMED_SHEET_VIEW_NS}}}sortRules",
            {"caseSensitive": "false", "sortMethod": "none"},
        )
        sort_rule = ElementTree.SubElement(
            sort_rules,
            f"{{{_NAMED_SHEET_VIEW_NS}}}sortRule",
            {"colId": "2", "id": "{22222222-2222-2222-2222-222222222222}"},
        )
        ElementTree.SubElement(
            sort_rule,
            f"{{{_NAMED_SHEET_VIEW_NS}}}sortCondition",
            {
                "ref": sort_ref,
                "descending": "false",
                "customList": "PRIVATE-NAMED-VIEW-SORT-LIST",
            },
        )

    def mutate(contents: dict[str, bytes]) -> None:
        if table_owned:
            table_auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
            table_xml = ElementTree.fromstring(contents["xl/tables/table1.xml"])
            table_auto_filter = table_xml.find(table_auto_filter_tag)
            if table_auto_filter is None:
                raise ValueError("Could not find Named Sheet View table AutoFilter fixture")
            table_auto_filter.set(
                f"{{{_OFFICE_2014_REVISION_NS}}}uid",
                "{00000000-0001-0000-0000-000000000000}",
            )
            contents["xl/tables/table1.xml"] = serialize(table_xml)
        else:
            worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
            sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
            auto_filter = ElementTree.Element(
                f"{{{_SPREADSHEETML_NS}}}autoFilter",
                {
                    "ref": "A1:C5",
                    f"{{{_OFFICE_2014_REVISION_NS}}}uid": (
                        "{00000000-0001-0000-0000-000000000000}"
                    ),
                },
            )
            sheet_data_index = next(
                index for index, child in enumerate(worksheet) if child.tag == sheet_data_tag
            )
            worksheet.insert(sheet_data_index + 1, auto_filter)
            contents["xl/worksheets/sheet1.xml"] = serialize(worksheet)

        relationship_member = "xl/worksheets/_rels/sheet1.xml.rels"
        relationships = (
            ElementTree.fromstring(contents[relationship_member])
            if relationship_member in contents
            else ElementTree.Element(f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships")
        )
        ElementTree.SubElement(
            relationships,
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship",
            {
                "Id": "rIdFenceNamedSheetView",
                "Type": named_view_relationship,
                "Target": "../namedSheetViews/namedSheetView1.xml",
            },
        )
        contents[relationship_member] = serialize(relationships)

        named_views = ElementTree.Element(
            f"{{{_NAMED_SHEET_VIEW_NS}}}namedSheetViews"
        )
        filter_id = (
            "{00000000-0002-0000-0000-000000000000}"
            if table_owned
            else "{00000000-0001-0000-0000-000000000000}"
        )
        add_view(
            named_views,
            name="Private baseline review",
            view_id="{33333333-3333-3333-3333-333333333333}",
            filter_id=filter_id,
            region="PRIVATE-NAMED-VIEW-REGION",
            sort_ref="C2:C5",
        )
        add_view(
            named_views,
            name="Private alternate review",
            view_id="{44444444-4444-4444-4444-444444444444}",
            filter_id=filter_id,
            region="PRIVATE-ALTERNATE-NAMED-VIEW-REGION",
            sort_ref="C2:C5",
        )
        contents["xl/namedSheetViews/namedSheetView1.xml"] = serialize(named_views)

        types = ElementTree.fromstring(contents["[Content_Types].xml"])
        ElementTree.SubElement(
            types,
            f"{{{content_types}}}Override",
            {
                "PartName": "/xl/namedSheetViews/namedSheetView1.xml",
                "ContentType": "application/vnd.ms-excel.namedsheetviews+xml",
            },
        )
        contents["[Content_Types].xml"] = serialize(types)

    return _rewrite_archive(path, mutate, ".named-sheet-views.tmp.xlsx")


def change_named_sheet_view_criterion(path: Path) -> Path:
    """Change an alternate-view filter member without touching cells or formulas."""

    def mutate(contents: dict[str, bytes]) -> None:
        named_views = ElementTree.fromstring(
            contents["xl/namedSheetViews/namedSheetView1.xml"]
        )
        standard_filter_tag = f"{{{_SPREADSHEETML_NS}}}filter"
        next(named_views.iter(standard_filter_tag)).set(
            "val",
            "CANDIDATE-PRIVATE-NAMED-VIEW-REGION",
        )
        contents["xl/namedSheetViews/namedSheetView1.xml"] = ElementTree.tostring(
            named_views,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".named-sheet-view-criterion.tmp.xlsx")


def normalize_named_sheet_view_control_spelling(path: Path) -> Path:
    """Use equivalent GUID, reference, Boolean, and integer spellings."""

    def mutate(contents: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
        auto_filter = worksheet.find(auto_filter_tag)
        if auto_filter is None:
            raise ValueError("Could not find Named Sheet View AutoFilter fixture")
        replacement_filter_id = "{55555555-5555-5555-5555-555555555555}"
        auto_filter.set("ref", "$a$1:$c$5")
        auto_filter.set(
            f"{{{_OFFICE_2014_REVISION_NS}}}uid",
            replacement_filter_id,
        )
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

        named_views = ElementTree.fromstring(
            contents["xl/namedSheetViews/namedSheetView1.xml"]
        )
        for view_index, view in enumerate(
        named_views.findall(f"{{{_NAMED_SHEET_VIEW_NS}}}namedSheetView"),
            start=6,
        ):
            view.set(
                "id",
                f"{{{str(view_index) * 8}-5555-5555-5555-555555555555}}",
            )
            named_filter = view.find(f"{{{_NAMED_SHEET_VIEW_NS}}}nsvFilter")
            if named_filter is None:
                raise ValueError("Could not find Named Sheet View filter fixture")
            named_filter.set("filterId", replacement_filter_id)
            named_filter.set("ref", "$A$1:$C$5")
            named_filter.set("tableId", "00")
            column_filter = named_filter.find(
                f"{{{_NAMED_SHEET_VIEW_NS}}}columnFilter"
            )
            if column_filter is None:
                raise ValueError("Could not find Named Sheet View column-filter fixture")
            column_filter.set("colId", "01")
            inner_filter = column_filter.find(f"{{{_NAMED_SHEET_VIEW_NS}}}filter")
            if inner_filter is None:
                raise ValueError("Could not find Named Sheet View inner-filter fixture")
            inner_filter.set("colId", "01")
            filters = inner_filter.find(f"{{{_SPREADSHEETML_NS}}}filters")
            if filters is None:
                raise ValueError("Could not find Named Sheet View filters fixture")
            filters.set("blank", "0")
            sort_rules = named_filter.find(f"{{{_NAMED_SHEET_VIEW_NS}}}sortRules")
            if sort_rules is None:
                raise ValueError("Could not find Named Sheet View sort-rules fixture")
            sort_rules.set("caseSensitive", "0")
            sort_rule = sort_rules.find(f"{{{_NAMED_SHEET_VIEW_NS}}}sortRule")
            if sort_rule is None:
                raise ValueError("Could not find Named Sheet View sort-rule fixture")
            sort_rule.set("colId", "02")
            condition = sort_rule.find(f"{{{_NAMED_SHEET_VIEW_NS}}}sortCondition")
            if condition is None:
                raise ValueError("Could not find Named Sheet View sort-condition fixture")
            condition.set("ref", "$c$2:$c$5")
            condition.set("descending", "0")
        contents["xl/namedSheetViews/namedSheetView1.xml"] = ElementTree.tostring(
            named_views,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".named-sheet-view-noise.tmp.xlsx")


def corrupt_named_sheet_view_control(path: Path) -> Path:
    """Inject an invalid alternate-view column ID to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        named_views = ElementTree.fromstring(
            contents["xl/namedSheetViews/namedSheetView1.xml"]
        )
        column_filter = next(
            named_views.iter(f"{{{_NAMED_SHEET_VIEW_NS}}}columnFilter")
        )
        column_filter.set("colId", "4294967296")
        contents["xl/namedSheetViews/namedSheetView1.xml"] = ElementTree.tostring(
            named_views,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".named-sheet-view-corrupt.tmp.xlsx")


def make_ignored_error_model(path: Path) -> Path:
    """Create standard and Office 2010 ignored-error controls outside cells."""
    workbook = Workbook()
    model = workbook.active
    model.title = "Error Review"
    model["A1"] = "Review surface"
    model["B2"] = "=1/0"
    model["B3"] = "=NA()"
    model["C2"] = "000123"
    model["C3"] = "000456"
    model["D2"] = "=SUM(B2:B3)"
    checks = workbook.create_sheet("Extension Review")
    checks["E2"] = "=1/0"
    workbook.save(path)

    def mutate(contents: dict[str, bytes]) -> None:
        ignored_errors_tag = f"{{{_SPREADSHEETML_NS}}}ignoredErrors"
        ignored_error_tag = f"{{{_SPREADSHEETML_NS}}}ignoredError"
        sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
        ext_list_tag = f"{{{_SPREADSHEETML_NS}}}extLst"
        extension_tag = f"{{{_SPREADSHEETML_NS}}}ext"
        x14_ignored_errors_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}ignoredErrors"
        x14_ignored_error_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}ignoredError"

        standard_sheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        standard_controls = ElementTree.Element(ignored_errors_tag)
        ElementTree.SubElement(
            standard_controls,
            ignored_error_tag,
            {"sqref": "B2 B3", "evalError": "1", "formula": "true"},
        )
        ElementTree.SubElement(
            standard_controls,
            ignored_error_tag,
            {"sqref": "C2:C3", "numberStoredAsText": "1"},
        )
        ElementTree.SubElement(
            standard_controls,
            ignored_error_tag,
            {
                "sqref": "D2",
                "formulaRange": "1",
                "emptyCellReference": "true",
                "listDataValidation": "1",
            },
        )
        sheet_data_index = next(
            index
            for index, child in enumerate(standard_sheet)
            if child.tag == sheet_data_tag
        )
        standard_sheet.insert(sheet_data_index + 1, standard_controls)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            standard_sheet,
            encoding="utf-8",
            xml_declaration=True,
        )

        extension_sheet = ElementTree.fromstring(contents["xl/worksheets/sheet2.xml"])
        extension_list = ElementTree.Element(ext_list_tag)
        extension = ElementTree.SubElement(
            extension_list,
            extension_tag,
            {"uri": "{01252117-D84E-4E92-8308-4BE1C098FCBB}"},
        )
        extension_controls = ElementTree.SubElement(extension, x14_ignored_errors_tag)
        ElementTree.SubElement(
            extension_controls,
            x14_ignored_error_tag,
            {
                "sqref": "E2",
                "unlockedFormula": "1",
                "calculatedColumn": "true",
                "twoDigitTextYear": "1",
            },
        )
        extension_sheet.append(extension_list)
        contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
            extension_sheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ignored-error.tmp.xlsx")


def change_ignored_error_target(path: Path) -> Path:
    """Change a private standard ignored-error target without touching cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        ignored_error_tag = f"{{{_SPREADSHEETML_NS}}}ignoredError"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        rule = next(
            element
            for element in worksheet.iter(ignored_error_tag)
            if element.get("numberStoredAsText") is not None
        )
        rule.set("sqref", "C4:C5")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ignored-error-target.tmp.xlsx")


def change_ignored_error_extension_target(path: Path) -> Path:
    """Change a private Office 2010 ignored-error target without cell edits."""

    def mutate(contents: dict[str, bytes]) -> None:
        ignored_error_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}ignoredError"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet2.xml"])
        next(worksheet.iter(ignored_error_tag)).set("sqref", "E3")
        contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ignored-error-extension-target.tmp.xlsx")


def normalize_ignored_error_control_spelling(path: Path) -> Path:
    """Use equivalent A1, Boolean, and ordering spellings for ignored errors."""

    def mutate(contents: dict[str, bytes]) -> None:
        ignored_error_tag = f"{{{_SPREADSHEETML_NS}}}ignoredError"
        x14_ignored_error_tag = f"{{{_OFFICE_2010_SPREADSHEET_NS}}}ignoredError"
        standard_sheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        rules = list(standard_sheet.iter(ignored_error_tag))
        if len(rules) != 3:
            raise ValueError("Could not find every standard ignored-error fixture rule")
        rules[0].set("sqref", "$b$3 $B$2")
        rules[0].set("evalError", "true")
        rules[0].set("formula", "1")
        rules[1].set("sqref", "$c$2:$C$3")
        rules[1].set("numberStoredAsText", "true")
        rules[2].set("sqref", "$d$2")
        rules[2].set("formulaRange", "true")
        rules[2].set("emptyCellReference", "1")
        rules[2].set("listDataValidation", "true")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            standard_sheet,
            encoding="utf-8",
            xml_declaration=True,
        )

        extension_sheet = ElementTree.fromstring(contents["xl/worksheets/sheet2.xml"])
        extension_rule = next(extension_sheet.iter(x14_ignored_error_tag))
        extension_rule.set("sqref", "$e$2")
        extension_rule.set("unlockedFormula", "true")
        extension_rule.set("calculatedColumn", "1")
        extension_rule.set("twoDigitTextYear", "true")
        contents["xl/worksheets/sheet2.xml"] = ElementTree.tostring(
            extension_sheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ignored-error-noise.tmp.xlsx")


def corrupt_ignored_error_control(path: Path) -> Path:
    """Inject an unsupported ignored-error target to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        ignored_error_tag = f"{{{_SPREADSHEETML_NS}}}ignoredError"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        next(worksheet.iter(ignored_error_tag)).set(
            "sqref", "PrivateIgnoredErrorSheet!B2"
        )
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ignored-error-corrupt.tmp.xlsx")


def duplicate_ignored_error_container(path: Path) -> Path:
    """Duplicate a standard container to exercise malformed-container coverage."""

    def mutate(contents: dict[str, bytes]) -> None:
        ignored_errors_tag = f"{{{_SPREADSHEETML_NS}}}ignoredErrors"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        controls = worksheet.find(ignored_errors_tag)
        if controls is None:
            raise ValueError("Could not find standard ignored-error fixture controls")
        worksheet.append(
            ElementTree.fromstring(
                ElementTree.tostring(controls, encoding="utf-8"),
            )
        )
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".ignored-error-duplicate.tmp.xlsx")


def change_what_if_data_table_input(path: Path) -> Path:
    """Change one private Data Table input reference without touching cells."""

    def mutate(contents: dict[str, bytes]) -> None:
        formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        formula = next(
            element
            for element in worksheet.iter(formula_tag)
            if element.get("t") == "dataTable"
        )
        formula.set("r1", "B4")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".what-if-data-table-change.tmp.xlsx")


def normalize_what_if_data_table_reference_spelling(path: Path) -> Path:
    """Use equivalent absolute/lowercase OOXML reference spellings."""

    def mutate(contents: dict[str, bytes]) -> None:
        formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        formulas = [
            element
            for element in worksheet.iter(formula_tag)
            if element.get("t") == "dataTable"
        ]
        if len(formulas) != 3:
            raise ValueError("Could not find every What-If Data Table fixture declaration")
        formulas[0].set("ref", "$d$3:$d$6")
        formulas[0].set("r1", "$b$2")
        formulas[0].set("dt2D", "0")
        formulas[0].set("dtr", "false")
        formulas[0].set("ca", "true")
        formulas[1].set("ref", "$f$3:$i$3")
        formulas[1].set("r1", "$b$2")
        formulas[1].set("dt2D", "false")
        formulas[1].set("dtr", "1")
        formulas[2].set("ref", "$k$3:$m$5")
        formulas[2].set("r1", "$b$2")
        formulas[2].set("r2", "$b$3")
        formulas[2].set("dt2D", "true")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".what-if-data-table-noise.tmp.xlsx")


def delete_what_if_data_table_input(path: Path) -> Path:
    """Mark the first table's input reference deleted in its raw definition."""

    def mutate(contents: dict[str, bytes]) -> None:
        formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        formula = next(
            element
            for element in worksheet.iter(formula_tag)
            if element.get("t") == "dataTable"
        )
        formula.set("del1", "1")
        formula.attrib.pop("r1", None)
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".what-if-data-table-deleted-input.tmp.xlsx")


def corrupt_what_if_data_table_input(path: Path) -> Path:
    """Inject an unsupported private input reference to exercise fail-closed parsing."""

    def mutate(contents: dict[str, bytes]) -> None:
        formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        formula = next(
            element
            for element in worksheet.iter(formula_tag)
            if element.get("t") == "dataTable"
        )
        formula.set("r1", "PrivateInputSheet!B2")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".what-if-data-table-corrupt.tmp.xlsx")


def overlap_what_if_data_table_outputs(path: Path) -> Path:
    """Expand one master range across another master to test fail-closed overlap."""

    def mutate(contents: dict[str, bytes]) -> None:
        formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
        worksheet = ElementTree.fromstring(contents["xl/worksheets/sheet1.xml"])
        formula = next(
            element
            for element in worksheet.iter(formula_tag)
            if element.get("t") == "dataTable"
        )
        formula.set("ref", "D3:F6")
        contents["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    return _rewrite_archive(path, mutate, ".what-if-data-table-overlap.tmp.xlsx")


def mark_array_formula_dynamic(path: Path, anchor: str = "B1") -> Path:
    """Add the OOXML dynamic-array metadata that openpyxl does not write.

    The small fixture mirrors the publicly documented XlsxWriter serialization:
    an array formula anchor has ``cm=1`` and cell metadata resolves that index
    to an ``XLDAPR`` record with ``fDynamic=1``.
    """
    anchor_bytes = anchor.encode("ascii")
    with ZipFile(path) as archive:
        contents = {
            entry.filename: archive.read(entry.filename) for entry in archive.infolist()
        }

    worksheet_name: str | None = None
    needle = b'<c r="' + anchor_bytes + b'"><f t="array"'
    replacement = b'<c r="' + anchor_bytes + b'" cm="1"><f t="array"'
    for name, content in contents.items():
        if not name.startswith("xl/worksheets/") or needle not in content:
            continue
        if worksheet_name is not None:
            raise ValueError(f"More than one array formula anchor found for {anchor}")
        contents[name] = content.replace(needle, replacement, 1)
        worksheet_name = name
    if worksheet_name is None:
        raise ValueError(f"Could not find array formula anchor {anchor}")

    metadata = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<metadata xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'xmlns:xda="http://schemas.microsoft.com/office/spreadsheetml/2017/dynamicarray">'
        b'<metadataTypes count="1"><metadataType name="XLDAPR" '
        b'minSupportedVersion="120000" copy="1" pasteAll="1" pasteValues="1" '
        b'merge="1" splitFirst="1" rowColShift="1" clearFormats="1" clearComments="1" '
        b'assign="1" coerce="1" cellMeta="1"/></metadataTypes>'
        b'<futureMetadata name="XLDAPR" count="1"><bk><extLst>'
        b'<ext uri="{bdbb8cdc-fa1e-496e-a857-3c3f30c029c3}">'
        b'<xda:dynamicArrayProperties fDynamic="1" fCollapsed="0"/>'
        b'</ext></extLst></bk></futureMetadata><cellMetadata count="1"><bk>'
        b'<rc t="1" v="0"/></bk></cellMetadata></metadata>'
    )
    content_types = contents["[Content_Types].xml"]
    contents["[Content_Types].xml"] = content_types.replace(
        b"</Types>",
        b'<Override PartName="/xl/metadata.xml" '
        b'ContentType="application/vnd.openxmlformats-officedocument.'
        b'spreadsheetml.sheetMetadata+xml"/></Types>',
        1,
    )
    relationships_name = "xl/_rels/workbook.xml.rels"
    relationships = contents[relationships_name]
    contents[relationships_name] = relationships.replace(
        b"</Relationships>",
        b'<Relationship Id="rIdFormulaFenceDynamicArrayMetadata" '
        b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/'
        b'sheetMetadata" Target="metadata.xml"/></Relationships>',
        1,
    )
    contents["xl/metadata.xml"] = metadata

    staging = path.with_suffix(".dynamic.tmp.xlsx")
    with ZipFile(staging, "w", compression=ZIP_DEFLATED) as archive:
        for name, content in contents.items():
            archive.writestr(name, content)
    staging.replace(path)
    return path


def mark_array_formula_unclassified(path: Path, anchor: str = "B1") -> Path:
    """Give an array formula an unresolvable OOXML cell-metadata marker."""
    anchor_bytes = anchor.encode("ascii")
    needle = b'<c r="' + anchor_bytes + b'"><f t="array"'
    replacement = b'<c r="' + anchor_bytes + b'" cm="1"><f t="array"'
    staging = path.with_suffix(".array-metadata.tmp.xlsx")
    with ZipFile(path) as source, ZipFile(staging, "w", compression=ZIP_DEFLATED) as target:
        replaced = False
        for entry in source.infolist():
            content = source.read(entry.filename)
            if (
                not replaced
                and entry.filename.startswith("xl/worksheets/")
                and needle in content
            ):
                content = content.replace(needle, replacement, 1)
                replaced = True
            target.writestr(entry.filename, content)
    if not replaced:
        staging.unlink()
        raise ValueError(f"Could not find array formula anchor {anchor}")
    staging.replace(path)
    return path


def make_named_formula_model(path: Path) -> Path:
    """Create global and local names whose definitions contain formulas."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Named-formula inputs"
    inputs["B2"] = 0.1
    inputs["B3"] = 100
    inputs["B4"] = 7

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Named-formula outputs"
    summary["B2"] = "=DiscountedValue"
    summary["B3"] = "=StaticRate"
    summary["B4"] = "=LocalMetric"

    report = workbook.create_sheet("Report")
    report["A1"] = "Qualified local-name output"
    report["B2"] = "=Summary!LocalMetric"

    workbook.defined_names.add(DefinedName("TaxRate", attr_text="=Inputs!$B$2"))
    workbook.defined_names.add(
        DefinedName(
            "DiscountedValue",
            attr_text="=LET(rate,TaxRate,amount,Inputs!$B$3,amount*(1-rate))",
        )
    )
    workbook.defined_names.add(DefinedName("StaticRate", attr_text="=0.05"))
    workbook.defined_names.add(
        DefinedName("LocalMetric", attr_text="=Inputs!$B$4*2", localSheetId=1)
    )
    workbook.save(path)
    return path


def make_let_model(path: Path) -> Path:
    """Create a model whose calculation uses lexical LET variables."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "LET inputs"
    inputs["B2"] = 0.1
    inputs["B3"] = 100

    model = workbook.create_sheet("Model")
    model["A1"] = "LET calculation"
    model["B2"] = "=LET(rate,Inputs!B2,amount,Inputs!B3,amount*(1-rate))"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "LET output"
    dashboard["B2"] = "=Model!B2"
    workbook.save(path)
    return path


def make_named_lambda_model(path: Path) -> Path:
    """Create reusable named LAMBDAs plus a formula-defined name that calls one."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Named LAMBDA inputs"
    inputs["B2"] = 0.1
    inputs["B3"] = 100
    inputs["B4"] = 32

    model = workbook.create_sheet("Model")
    model["A1"] = "Named LAMBDA calculations"
    model["B2"] = "=ToCelsius(Inputs!B4)"
    model["B3"] = "=AdjustedCelsius(Inputs!B4)"
    model["B4"] = "=NamedAdjustedValue"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Named LAMBDA output"
    dashboard["B2"] = "=Model!B4"

    workbook.defined_names.add(
        DefinedName(
            "ToCelsius",
            attr_text="_xlfn.LAMBDA(_xlpm.temp,(5/9)*(_xlpm.temp-32)+Inputs!$B$2)",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "AdjustedCelsius",
            attr_text="_xlfn.LAMBDA(_xlpm.temp,ToCelsius(_xlpm.temp)+Inputs!$B$3)",
        )
    )
    workbook.defined_names.add(
        DefinedName("NamedAdjustedValue", attr_text="AdjustedCelsius(Inputs!$B$4)")
    )
    workbook.save(path)
    return path


def make_scoped_named_lambda_model(path: Path) -> Path:
    """Create global and worksheet-local LAMBDAs with the same callable name."""
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["B2"] = 2
    inputs["B3"] = 3

    model = workbook.create_sheet("Model")
    model["A2"] = 10
    model["B2"] = "=Scale(A2)"

    report = workbook.create_sheet("Report")
    report["A2"] = 10
    report["B2"] = "=Scale(A2)"
    report["B3"] = "=Model!Scale(A2)"

    workbook.defined_names.add(
        DefinedName("Scale", attr_text="=LAMBDA(value,value*Inputs!$B$2)")
    )
    model.defined_names.add(
        DefinedName(
            "Scale",
            attr_text="=LAMBDA(value,value*Inputs!$B$3)",
            localSheetId=1,
        )
    )
    workbook.save(path)
    return path


def rewrite(path: Path, mutate: Callable[[Workbook], None]) -> Path:
    """Load, mutate, and save a fixture in place."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    mutate(workbook)
    workbook.save(path)
    return path
