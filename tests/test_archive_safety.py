from __future__ import annotations

import binascii
import hashlib
import io
import shutil
import stat
import struct
import zipfile
from copy import deepcopy
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

import formulafence.workbook as workbook_module
from formulafence.cli import main
from formulafence.models import WorkbookLoadError
from formulafence.workbook import load_snapshot

from .helpers import (
    make_custom_workbook_view_model,
    make_external_data_refresh_model,
    make_external_link_package_model,
    make_model,
    make_protection_model,
    make_rich_text_run_model,
    make_strict_custom_workbook_view_model,
    make_strict_worksheet_drawing_connector_model,
    make_strict_worksheet_print_layout_model,
    make_table_model,
    make_worksheet_drawing_shape_model,
)


def _append_member(
    path: Path,
    name: str,
    payload: bytes,
    *,
    compression: int = ZIP_DEFLATED,
) -> None:
    with ZipFile(path, "a", compression=compression) as archive:
        archive.writestr(name, payload)


def _named_formula_fanout_workbook(
    path: Path,
    *,
    input_count: int = 3,
    caller_count: int = 4,
) -> Path:
    """Create a compact model whose one name expands at every caller cell."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    for row in range(1, input_count + 1):
        worksheet.cell(row=row, column=1, value=row)
    workbook.defined_names.add(
        DefinedName(
            "Fanout",
            attr_text=(
                "=SUM("
                + ",".join(f"Model!$A${row}" for row in range(1, input_count + 1))
                + ")"
            ),
        )
    )
    for row in range(1, caller_count + 1):
        worksheet.cell(row=row, column=2, value="=Fanout")
    workbook.save(path)
    return path


def _replace_member(path: Path, name: str, payload: bytes) -> None:
    """Replace one test package part without creating a duplicate ZIP entry."""
    staging = path.with_suffix(".replacement.xlsx")
    with ZipFile(path) as archive:
        contents = {
            member.filename: archive.read(member)
            for member in archive.infolist()
        }
    contents[name] = payload
    with ZipFile(staging, "w", compression=ZIP_DEFLATED) as archive:
        for member_name, member_payload in contents.items():
            archive.writestr(member_name, member_payload)
    staging.replace(path)


def _write_relationship_part_elements(
    path: Path,
    member_name: str,
    count: int,
    *,
    nested: bool = False,
) -> None:
    """Add package-relationship XML entries without using a workbook reader."""
    relationship_namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
    with ZipFile(path) as archive:
        payload = archive.read(member_name) if member_name in archive.namelist() else None
    root = (
        ElementTree.fromstring(payload)
        if payload is not None
        else ElementTree.Element(f"{{{relationship_namespace}}}Relationships")
    )
    namespace = root.tag.partition("}")[0].removeprefix("{")
    parent = root
    if nested:
        parent = ElementTree.SubElement(root, f"{{{namespace}}}opaque")
    for _ in range(count):
        ElementTree.SubElement(parent, f"{{{namespace}}}Relationship")
    serialized = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
    if payload is None:
        _append_member(path, member_name, serialized)
    else:
        _replace_member(path, member_name, serialized)


def _relationship_part_element_count(path: Path, member_name: str) -> int:
    """Count one relationship part's complete XML tree for exact limit tests."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member_name))
    return sum(1 for _ in root.iter())


def _append_worksheet_drawing_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    member_name: str = "xl/drawings/drawing1.xml",
) -> None:
    """Add opaque DrawingML descendants without using a workbook reader."""
    with ZipFile(path) as archive:
        drawing = archive.read(member_name)
    root = ElementTree.fromstring(drawing)
    namespace = "urn:formulafence:archive-safety"
    parent = root
    if nested:
        parent = ElementTree.SubElement(parent, f"{{{namespace}}}opaqueContainer")
    for _ in range(count):
        ElementTree.SubElement(parent, f"{{{namespace}}}opaque")
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _worksheet_drawing_xml_element_count(path: Path, member_name: str) -> int:
    """Count the complete tree of one DrawingML part for exact-limit tests."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member_name))
    return sum(1 for _ in root.iter())


def _append_table_definition_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    member_name: str = "xl/tables/table1.xml",
) -> None:
    """Add opaque table-definition descendants without a workbook reader."""
    with ZipFile(path) as archive:
        table = archive.read(member_name)
    closing = b"</table>"
    if not table.endswith(closing):
        raise ValueError("table fixture XML has no closing table tag")
    namespace = b"urn:formulafence:archive-safety"
    entries = b"<ff:opaque/>" * count
    if nested:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + b"<ff:nested>"
            + entries
            + b"</ff:nested></ff:container>"
        )
    else:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + entries
            + b"</ff:container>"
        )
    _replace_member(
        path,
        member_name,
        table[: -len(closing)] + inserted + closing,
    )


def _table_definition_xml_element_count(path: Path, member_name: str) -> int:
    """Count the complete tree of one table definition for exact-limit tests."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member_name))
    return sum(1 for _ in root.iter())


def _append_shared_string_root_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    member_name: str = "xl/sharedStrings.xml",
) -> None:
    """Add ignored root children without constructing a shared-string tree."""
    with ZipFile(path) as archive:
        shared_strings = archive.read(member_name)
    closing_offset = shared_strings.rfind(b"</")
    if closing_offset < 0:
        raise ValueError("shared-string fixture XML has no closing root tag")
    namespace = b"urn:formulafence:archive-safety"
    entries = b"<ff:opaque/>" * count
    if nested:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + b"<ff:nested>"
            + entries
            + b"</ff:nested></ff:container>"
        )
    else:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + entries
            + b"</ff:container>"
        )
    _replace_member(
        path,
        member_name,
        shared_strings[:closing_offset] + inserted + shared_strings[closing_offset:],
    )


def _append_worksheet_opaque_root_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    member_name: str = "xl/worksheets/sheet1.xml",
) -> int:
    """Append a direct non-Worksheet child without building a worksheet tree."""
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    closing_offset = worksheet.rfind(b"</")
    if closing_offset < 0:
        raise ValueError("worksheet fixture XML has no closing root tag")
    namespace = b"urn:formulafence:archive-safety"
    entries = b"<ff:opaque/>" * count
    if nested:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + b"<ff:nested>"
            + entries
            + b"</ff:nested></ff:container>"
        )
        element_count = count + 2
    else:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + entries
            + b"</ff:container>"
        )
        element_count = count + 1
    _replace_member(
        path,
        member_name,
        worksheet[:closing_offset] + inserted + worksheet[closing_offset:],
    )
    return element_count


def _append_worksheet_extension_list_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    within_sheet_properties: bool = False,
    member_name: str = "xl/worksheets/sheet1.xml",
) -> int:
    """Append one extension tree while preserving the Worksheet namespace."""
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    closing_offset = worksheet.rfind(b"</")
    if closing_offset < 0:
        raise ValueError("worksheet fixture XML has no closing root tag")
    worksheet_namespace = (
        ElementTree.fromstring(worksheet).tag.partition("}")[0].removeprefix("{").encode()
    )
    if not worksheet_namespace:
        raise ValueError("worksheet fixture XML has no namespace")
    namespace = b"urn:formulafence:archive-safety"
    entries = b"<ff:opaque/>" * count
    if nested:
        contents = b"<ff:container><ff:nested>" + entries + b"</ff:nested></ff:container>"
        element_count = count + 4
    else:
        contents = entries
        element_count = count + 2
    extension = (
        b'<ss:ext xmlns:ff="'
        + namespace
        + b'" uri="{1F4A6F6A-EB4A-4C41-9C9E-9231F8EAF001}">'
        + contents
        + b"</ss:ext>"
    )
    inserted = (
        b'<ss:sheetPr xmlns:ss="'
        + worksheet_namespace
        + b'"><ss:extLst>'
        + extension
        + b"</ss:extLst></ss:sheetPr>"
        if within_sheet_properties
        else b'<ss:extLst xmlns:ss="'
        + worksheet_namespace
        + b'">'
        + extension
        + b"</ss:extLst>"
    )
    _replace_member(
        path,
        member_name,
        worksheet[:closing_offset] + inserted + worksheet[closing_offset:],
    )
    return element_count


def _append_workbook_opaque_root_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    member_name: str = "xl/workbook.xml",
) -> int:
    """Append a direct non-Workbook child without building a workbook model."""
    with ZipFile(path) as archive:
        workbook = archive.read(member_name)
    closing_offset = workbook.rfind(b"</")
    if closing_offset < 0:
        raise ValueError("workbook fixture XML has no closing root tag")
    namespace = b"urn:formulafence:archive-safety"
    entries = b"<ff:opaque/>" * count
    if nested:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + b"<ff:nested>"
            + entries
            + b"</ff:nested></ff:container>"
        )
        element_count = count + 2
    else:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + entries
            + b"</ff:container>"
        )
        element_count = count + 1
    _replace_member(
        path,
        member_name,
        workbook[:closing_offset] + inserted + workbook[closing_offset:],
    )
    return element_count


def _append_workbook_extension_list_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    within_book_view: bool = False,
    extension_namespace: str | None = None,
    member_name: str = "xl/workbook.xml",
) -> int:
    """Append a workbook extension tree while preserving its namespace."""
    with ZipFile(path) as archive:
        workbook = archive.read(member_name)
    root = ElementTree.fromstring(workbook)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    if not namespace:
        raise ValueError("workbook fixture XML has no namespace")
    extension_namespace = extension_namespace or namespace
    extension_list = ElementTree.Element(f"{{{extension_namespace}}}extLst")
    extension = ElementTree.SubElement(
        extension_list,
        f"{{{extension_namespace}}}ext",
        {"uri": "{1F4A6F6A-EB4A-4C41-9C9E-9231F8EAF005}"},
    )
    if nested:
        parent = ElementTree.SubElement(
            extension,
            "{urn:formulafence:archive-safety}container",
        )
        parent = ElementTree.SubElement(
            parent,
            "{urn:formulafence:archive-safety}nested",
        )
        element_count = count + 4
    else:
        parent = extension
        element_count = count + 2
    for _ in range(count):
        ElementTree.SubElement(parent, "{urn:formulafence:archive-safety}opaque")
    if within_book_view:
        book_views = next(
            (child for child in root if child.tag == f"{{{namespace}}}bookViews"),
            None,
        )
        if book_views is None or not list(book_views):
            raise ValueError("workbook fixture has no book view")
        book_views[0].append(extension_list)
    else:
        root.append(extension_list)
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )
    return element_count


def _replace_workbook_root_namespace(path: Path, namespace: str) -> None:
    """Change only the workbook root namespace for parser-selection coverage."""
    member_name = "xl/workbook.xml"
    with ZipFile(path) as archive:
        workbook = archive.read(member_name)
    root = ElementTree.fromstring(workbook)
    root.tag = f"{{{namespace}}}{root.tag.rsplit('}', maxsplit=1)[-1]}"
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_non_grid_sheet_extension_list_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    member_name: str,
) -> None:
    """Append a non-grid-sheet extension tree without materializing that sheet."""
    with ZipFile(path) as archive:
        sheet = archive.read(member_name)
    closing_offset = sheet.rfind(b"</")
    if closing_offset < 0:
        raise ValueError("non-grid sheet fixture XML has no closing root tag")
    sheet_namespace = (
        ElementTree.fromstring(sheet)
        .tag.partition("}")[0]
        .removeprefix("{")
        .encode()
    )
    if not sheet_namespace:
        raise ValueError("non-grid sheet fixture XML has no namespace")
    namespace = b"urn:formulafence:archive-safety"
    entries = b"<ff:opaque/>" * count
    contents = (
        b"<ff:container><ff:nested>" + entries + b"</ff:nested></ff:container>"
        if nested
        else entries
    )
    inserted = (
        b'<ss:extLst xmlns:ss="'
        + sheet_namespace
        + b'">'
        + b'<ss:ext xmlns:ff="'
        + namespace
        + b'" uri="{1F4A6F6A-EB4A-4C41-9C9E-9231F8EAF003}">'
        + contents
        + b"</ss:ext></ss:extLst>"
    )
    _replace_member(
        path,
        member_name,
        sheet[:closing_offset] + inserted + sheet[closing_offset:],
    )


def _non_grid_sheet_xml_element_count(path: Path, member_name: str) -> int:
    """Count one non-grid-sheet tree for exact structural-boundary tests."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member_name))
    return sum(1 for _ in root.iter())


def _append_non_grid_sheet_opaque_root_xml_elements(
    path: Path,
    count: int,
    *,
    member_name: str,
) -> None:
    """Append foreign direct root content without materializing that sheet."""
    with ZipFile(path) as archive:
        sheet = archive.read(member_name)
    closing_offset = sheet.rfind(b"</")
    if closing_offset < 0:
        raise ValueError("non-grid sheet fixture XML has no closing root tag")
    inserted = (
        b'<ff:container xmlns:ff="urn:formulafence:archive-safety">'
        + (b"<ff:opaque/>" * count)
        + b"</ff:container>"
    )
    _replace_member(
        path,
        member_name,
        sheet[:closing_offset] + inserted + sheet[closing_offset:],
    )


def _convert_chartsheet_to_dialogsheet(path: Path, *, sheet_number: int = 1) -> None:
    """Turn one synthetic chart tab into a minimal relationship-selected dialog tab."""
    old_member = f"xl/chartsheets/sheet{sheet_number}.xml"
    new_member = f"xl/dialogsheets/sheet{sheet_number}.xml"
    old_relationship_member = f"xl/chartsheets/_rels/sheet{sheet_number}.xml.rels"
    with ZipFile(path) as archive:
        contents = {
            member.filename: archive.read(member)
            for member in archive.infolist()
        }
    contents.pop(old_member)
    contents.pop(old_relationship_member, None)
    relationship_tag = (
        "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
    )
    relationships = ElementTree.fromstring(contents["xl/_rels/workbook.xml.rels"])
    relationship = next(
        (
            candidate
            for candidate in relationships.findall(relationship_tag)
            if candidate.get("Type", "").endswith("/chartsheet")
            and candidate.get("Target", "").endswith(
                f"/chartsheets/sheet{sheet_number}.xml"
            )
        ),
        None,
    )
    if relationship is None:
        raise ValueError("synthetic chart sheet relationship was not found")
    relationship.set(
        "Type",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/dialogsheet",
    )
    relationship.set("Target", f"/xl/dialogsheets/sheet{sheet_number}.xml")
    contents["xl/_rels/workbook.xml.rels"] = ElementTree.tostring(
        relationships,
        encoding="utf-8",
        xml_declaration=True,
    )
    override_tag = (
        "{http://schemas.openxmlformats.org/package/2006/content-types}Override"
    )
    content_types = ElementTree.fromstring(contents["[Content_Types].xml"])
    override = next(
        (
            candidate
            for candidate in content_types.findall(override_tag)
            if candidate.get("PartName") == f"/{old_member}"
        ),
        None,
    )
    if override is None:
        raise ValueError("synthetic chart sheet content type was not found")
    override.set("PartName", f"/{new_member}")
    override.set(
        "ContentType",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.dialogsheet+xml",
    )
    contents["[Content_Types].xml"] = ElementTree.tostring(
        content_types,
        encoding="utf-8",
        xml_declaration=True,
    )
    contents[new_member] = (
        b'<?xml version="1.0" encoding="utf-8"?>'
        b'<dialogsheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        b"</dialogsheet>"
    )
    with ZipFile(path, "w", compression=ZIP_DEFLATED) as archive:
        for member_name, payload in contents.items():
            archive.writestr(member_name, payload)


def _make_non_grid_sheet_model(path: Path, sheet_kind: str) -> tuple[Path, str]:
    """Create a normal workbook with one relationship-selected non-grid sheet."""
    make_protection_model(path, include_chartsheet=True)
    if sheet_kind == "chartsheet":
        return path, "xl/chartsheets/sheet1.xml"
    if sheet_kind == "dialogsheet":
        _convert_chartsheet_to_dialogsheet(path)
        return path, "xl/dialogsheets/sheet1.xml"
    raise ValueError(f"unknown non-grid sheet kind: {sheet_kind}")


def _redirect_shared_string_relationship(path: Path, target: str) -> None:
    """Point the transitional shared-string relationship at a safe test part."""
    member_name = "xl/_rels/workbook.xml.rels"
    relationship_tag = (
        "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
    )
    with ZipFile(path) as archive:
        relationships = archive.read(member_name)
    root = ElementTree.fromstring(relationships)
    relationship = next(
        (
            candidate
            for candidate in root.findall(relationship_tag)
            if candidate.get("Type")
            == workbook_module._OOXML_SHARED_STRINGS_RELATIONSHIP
        ),
        None,
    )
    if relationship is None:
        raise ValueError("shared-string fixture has no workbook relationship")
    relationship.set("Target", target)
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_shared_string_item_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    member_name: str = "xl/sharedStrings.xml",
) -> None:
    """Add opaque descendants to the first shared string item."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member_name))
    item_tag = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"
    item = root.find(item_tag)
    if item is None:
        raise ValueError("shared-string fixture XML has no item")
    namespace = "urn:formulafence:archive-safety"
    parent = item
    if nested:
        parent = ElementTree.SubElement(parent, f"{{{namespace}}}container")
        parent = ElementTree.SubElement(parent, f"{{{namespace}}}nested")
    for _ in range(count):
        ElementTree.SubElement(parent, f"{{{namespace}}}opaque")
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _shared_string_item_xml_element_count(path: Path, member_name: str) -> int:
    """Count the first complete shared string item for exact-limit tests."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member_name))
    item_tag = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"
    item = root.find(item_tag)
    if item is None:
        raise ValueError("shared-string fixture XML has no item")
    return sum(1 for _ in item.iter())


def _complex_shared_string_xml_element_count(path: Path, member_name: str) -> int:
    """Mirror the semantic preflight's complex shared-string-item accounting."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member_name))
    item_tag = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"
    text_tag = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"
    total = 0
    for item in root.findall(item_tag):
        descendants = tuple(item.iter())
        if any(element.tag != text_tag for element in descendants[1:]):
            total += len(descendants)
    return total


def _duplicate_first_shared_string_item(path: Path, member_name: str) -> None:
    """Append one equivalent rich item to exercise the aggregate guard."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member_name))
    item_tag = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"
    item = root.find(item_tag)
    if item is None:
        raise ValueError("shared-string fixture XML has no item")
    root.append(deepcopy(item))
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _add_worksheet_table_relationship_target(
    path: Path,
    *,
    source_member: str = "xl/tables/table1.xml",
    target_member: str = "xl/private/table-definition.xml",
) -> None:
    """Add a noncanonical table target to the ordinary worksheet relationships."""
    relationship_member = "xl/worksheets/_rels/sheet1.xml.rels"
    with ZipFile(path) as archive:
        table = archive.read(source_member)
        relationships = archive.read(relationship_member)
    _append_member(path, target_member, table)
    root = ElementTree.fromstring(relationships)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    ElementTree.SubElement(
        root,
        f"{{{namespace}}}Relationship",
        {
            "Id": "rIdFormulaFenceTableDefinition",
            "Type": (
                "http://schemas.openxmlformats.org/officeDocument/2006/"
                "relationships/table"
            ),
            "Target": "../private/table-definition.xml",
        },
    )
    _replace_member(
        path,
        relationship_member,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _add_worksheet_drawing_relationship_target(
    path: Path,
    *,
    source_member: str = "xl/drawings/drawing1.xml",
    target_member: str = "xl/drawings/drawing2.xml",
) -> None:
    """Add a second raw DrawingML target to the selected worksheet relationship part."""
    relationship_member = "xl/worksheets/_rels/sheet1.xml.rels"
    with ZipFile(path) as archive:
        drawing = archive.read(source_member)
        relationships = archive.read(relationship_member)
    _append_member(path, target_member, drawing)
    root = ElementTree.fromstring(relationships)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    ElementTree.SubElement(
        root,
        f"{{{namespace}}}Relationship",
        {
            "Id": "rIdFormulaFenceSecondDrawing",
            "Type": (
                "http://schemas.openxmlformats.org/officeDocument/2006/"
                "relationships/drawing"
            ),
            "Target": "../drawings/drawing2.xml",
        },
    )
    _replace_member(
        path,
        relationship_member,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_workbook_sheet_declarations(
    path: Path,
    count: int,
    *,
    namespace_prefix: str | None = None,
) -> None:
    """Repeat one declared sheet to exercise pre-reader sheet amplification."""
    member_name = "xl/workbook.xml"
    with ZipFile(path) as archive:
        workbook = archive.read(member_name)
    start = workbook.index(b"<sheet ")
    end = workbook.index(b"/>", start) + 2
    declaration = workbook[start:end]
    if namespace_prefix is not None:
        namespace = b"urn:formulafence:archive-safety"
        workbook = workbook.replace(
            b"<workbook ",
            b"<workbook xmlns:" + namespace_prefix.encode() + b'=\"' + namespace + b'\" ',
            1,
        )
        declaration = declaration.replace(
            b"<sheet ",
            b"<" + namespace_prefix.encode() + b":sheet ",
            1,
        )
    _replace_member(
        path,
        member_name,
        workbook.replace(b"</sheets>", declaration * count + b"</sheets>"),
    )


def _append_workbook_defined_name_declarations(
    path: Path,
    count: int,
    *,
    namespace_prefix: str | None = None,
) -> None:
    """Append unique declared names without relying on an object-model writer."""
    member_name = "xl/workbook.xml"
    with ZipFile(path) as archive:
        workbook = archive.read(member_name)
    tag_name = "definedName"
    if namespace_prefix is not None:
        namespace = b"urn:formulafence:archive-safety"
        workbook = workbook.replace(
            b"<workbook ",
            b"<workbook xmlns:" + namespace_prefix.encode() + b'=\"' + namespace + b'\" ',
            1,
        )
        tag_name = f"{namespace_prefix}:definedName"
    declarations = b"".join(
        (
            f'<{tag_name} name="FormulaFenceAuditDefinedName{index:06d}">'
            f"Inputs!$B$2</{tag_name}>"
        ).encode()
        for index in range(count)
    )
    _replace_member(
        path,
        member_name,
        workbook.replace(
            b"</definedNames>",
            declarations + b"</definedNames>",
        ),
    )


def _append_workbook_relationship_catalog_declarations(
    path: Path,
    catalog_name: str,
    count: int,
    *,
    alternate_namespace: str | None = None,
) -> None:
    """Repeat one relationship-backed workbook declaration in a test package."""
    member_name = "xl/workbook.xml"
    with ZipFile(path) as archive:
        workbook = archive.read(member_name)
    root = ElementTree.fromstring(workbook)
    container = next(
        (
            child
            for child in root
            if child.tag.rsplit("}", maxsplit=1)[-1] == catalog_name
        ),
        None,
    )
    if container is None:
        raise ValueError(f"workbook fixture has no {catalog_name!r} catalog")
    declaration = next(iter(container), None)
    if declaration is None:
        raise ValueError(f"workbook fixture has an empty {catalog_name!r} catalog")
    if alternate_namespace is not None:
        declaration = deepcopy(declaration)
        declaration.tag = (
            f"{{{alternate_namespace}}}" + declaration.tag.rsplit("}", maxsplit=1)[-1]
        )
    for _ in range(count):
        container.append(deepcopy(declaration))
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_workbook_package_catalog_declarations(
    path: Path,
    catalog_name: str,
    child_name: str,
    count: int,
    *,
    attributes: dict[str, str] | None = None,
    alternate_namespace: str | None = None,
) -> None:
    """Add repeated workbook-package catalog entries without using a reader."""
    member_name = "xl/workbook.xml"
    with ZipFile(path) as archive:
        workbook = archive.read(member_name)
    root = ElementTree.fromstring(workbook)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_catalog_name = (
        f"{{{namespace}}}{catalog_name}" if namespace else catalog_name
    )
    qualified_child_name = f"{{{namespace}}}{child_name}" if namespace else child_name
    container = next(
        (
            child
            for child in root
            if child.tag.rsplit("}", maxsplit=1)[-1] == catalog_name
        ),
        None,
    )
    if container is None:
        container = ElementTree.Element(qualified_catalog_name)
        root.insert(0, container)
    declaration = ElementTree.Element(qualified_child_name)
    for name, value in (attributes or {}).items():
        declaration.set(name, value)
    if alternate_namespace is not None:
        declaration.tag = f"{{{alternate_namespace}}}{child_name}"
    for _ in range(count):
        container.append(deepcopy(declaration))
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_custom_sheet_view_declarations(
    path: Path,
    count: int,
    *,
    alternate_namespace: str | None = None,
    container_namespace: str | None = None,
    member_name: str = "xl/worksheets/sheet1.xml",
) -> None:
    """Repeat direct sheet Custom View entries for the preflight boundary."""
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    container_namespace = container_namespace or namespace
    qualified_container_name = (
        f"{{{container_namespace}}}customSheetViews"
        if container_namespace
        else "customSheetViews"
    )
    qualified_child_name = (
        f"{{{namespace}}}customSheetView" if namespace else "customSheetView"
    )
    container = next(
        (
            child
            for child in root
            if child.tag.rsplit("}", maxsplit=1)[-1] == "customSheetViews"
        ),
        None,
    )
    if container is None:
        container = ElementTree.Element(qualified_container_name)
        root.insert(0, container)
    declaration = ElementTree.Element(qualified_child_name)
    declaration.set("guid", "{11111111-1111-1111-1111-111111111111}")
    if alternate_namespace is not None:
        declaration.tag = f"{{{alternate_namespace}}}customSheetView"
    for _ in range(count):
        container.append(deepcopy(declaration))
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_merged_cell_declarations(
    path: Path,
    references: tuple[str, ...],
    *,
    alternate_namespace: str | None = None,
) -> None:
    """Add direct worksheet merge declarations without invoking a workbook reader."""
    member_name = "xl/worksheets/sheet1.xml"
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_container_name = (
        f"{{{namespace}}}mergeCells" if namespace else "mergeCells"
    )
    qualified_child_name = f"{{{namespace}}}mergeCell" if namespace else "mergeCell"
    container = next(
        (
            child
            for child in root
            if child.tag.rsplit("}", maxsplit=1)[-1] == "mergeCells"
        ),
        None,
    )
    if container is None:
        container = ElementTree.Element(qualified_container_name)
        root.insert(0, container)
    for reference in references:
        declaration = ElementTree.Element(qualified_child_name, {"ref": reference})
        if alternate_namespace is not None:
            declaration.tag = f"{{{alternate_namespace}}}mergeCell"
        container.append(declaration)
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_row_dimension_declarations(
    path: Path,
    count: int,
    *,
    member_name: str = "xl/worksheets/sheet1.xml",
    attributes: dict[str, str] | None = None,
    alternate_namespace: str | None = None,
) -> None:
    """Append empty rows that can make the workbook reader retain dimensions."""
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_row_name = f"{{{namespace}}}row" if namespace else "row"
    if alternate_namespace is not None:
        qualified_row_name = f"{{{alternate_namespace}}}row"
    sheet_data = next(
        (
            child
            for child in root
            if child.tag.rsplit("}", maxsplit=1)[-1] == "sheetData"
        ),
        None,
    )
    if sheet_data is None:
        raise ValueError("worksheet fixture has no sheetData element")
    row_attributes = (
        {"ht": "15", "customHeight": "1"}
        if attributes is None
        else attributes
    )
    for index in range(count):
        declaration_attributes = {"r": str(10_000 + index), **row_attributes}
        ElementTree.SubElement(sheet_data, qualified_row_name, declaration_attributes)
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_column_dimension_declarations(
    path: Path,
    count: int,
    *,
    member_name: str = "xl/worksheets/sheet1.xml",
    attributes: dict[str, str] | None = None,
    alternate_namespace: str | None = None,
) -> None:
    """Append reader-visible column declarations without an object-model writer."""
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_container_name = f"{{{namespace}}}cols" if namespace else "cols"
    qualified_column_name = f"{{{namespace}}}col" if namespace else "col"
    if alternate_namespace is not None:
        qualified_column_name = f"{{{alternate_namespace}}}col"
    columns = next(
        (child for child in root if child.tag == qualified_container_name),
        None,
    )
    if columns is None:
        columns = ElementTree.Element(qualified_container_name)
        sheet_data_index = next(
            index
            for index, child in enumerate(root)
            if child.tag.rsplit("}", maxsplit=1)[-1] == "sheetData"
        )
        root.insert(sheet_data_index, columns)
    column_attributes = {"min": "1", "max": "1"}
    if attributes is None:
        column_attributes.update({"width": "10", "customWidth": "1"})
    else:
        column_attributes.update(attributes)
    for _ in range(count):
        ElementTree.SubElement(columns, qualified_column_name, column_attributes)
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_column_dimension_containers(
    path: Path,
    count: int,
    *,
    member_name: str = "xl/worksheets/sheet1.xml",
    alternate_namespace: str | None = None,
) -> None:
    """Append direct column containers that raw dimension scanners must retain."""
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_container_name = f"{{{namespace}}}cols" if namespace else "cols"
    if alternate_namespace is not None:
        qualified_container_name = f"{{{alternate_namespace}}}cols"
    sheet_data_index = next(
        index
        for index, child in enumerate(root)
        if child.tag.rsplit("}", maxsplit=1)[-1] == "sheetData"
    )
    for _ in range(count):
        root.insert(sheet_data_index, ElementTree.Element(qualified_container_name))
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_page_break_declarations(
    path: Path,
    count: int,
    *,
    axis: str = "row",
    member_name: str = "xl/worksheets/sheet1.xml",
    child_name: str = "brk",
    alternate_namespace: str | None = None,
) -> None:
    """Append direct print-break records without invoking a workbook reader."""
    container_name = {"row": "rowBreaks", "column": "colBreaks"}.get(axis)
    if container_name is None:
        raise ValueError(f"Unsupported page-break axis {axis!r}")
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_container_name = (
        f"{{{namespace}}}{container_name}" if namespace else container_name
    )
    qualified_child_name = f"{{{namespace}}}{child_name}" if namespace else child_name
    if alternate_namespace is not None:
        qualified_child_name = f"{{{alternate_namespace}}}{child_name}"
    container = next(
        (child for child in root if child.tag == qualified_container_name),
        None,
    )
    if container is None:
        container = ElementTree.Element(qualified_container_name)
        _insert_worksheet_child_after_sheet_data(root, container)
    for _ in range(count):
        ElementTree.SubElement(
            container,
            qualified_child_name,
            {"id": "10", "min": "0", "max": "16383", "man": "1"},
        )
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_page_break_containers(
    path: Path,
    count: int,
    *,
    axis: str = "row",
    member_name: str = "xl/worksheets/sheet1.xml",
    alternate_namespace: str | None = None,
) -> None:
    """Append direct break containers that raw print scanners must retain."""
    container_name = {"row": "rowBreaks", "column": "colBreaks"}.get(axis)
    if container_name is None:
        raise ValueError(f"Unsupported page-break axis {axis!r}")
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_container_name = (
        f"{{{namespace}}}{container_name}" if namespace else container_name
    )
    if alternate_namespace is not None:
        qualified_container_name = f"{{{alternate_namespace}}}{container_name}"
    for _ in range(count):
        _insert_worksheet_child_after_sheet_data(
            root,
            ElementTree.Element(qualified_container_name),
        )
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _custom_sheet_view_page_break_target(
    root: ElementTree.Element,
    *,
    custom_sheet_view_namespace: str | None,
) -> tuple[str, ElementTree.Element]:
    """Return one fixture Custom View, optionally through an opaque child path."""
    namespace = root.tag.partition("}")[0].removeprefix("{")
    container = next(
        (
            child
            for child in root
            if child.tag.rsplit("}", maxsplit=1)[-1] == "customSheetViews"
        ),
        None,
    )
    if container is None:
        raise ValueError("Could not locate the Custom View fixture container")
    if custom_sheet_view_namespace is not None:
        view = ElementTree.Element(
            f"{{{custom_sheet_view_namespace}}}customSheetView",
            {"guid": "{11111111-1111-1111-1111-111111111111}"},
        )
        container.append(view)
        return namespace, view
    view = next(
        (
            child
            for child in reversed(list(container))
            if child.tag.rsplit("}", maxsplit=1)[-1] == "customSheetView"
        ),
        None,
    )
    if view is None:
        raise ValueError("Could not locate the Custom View fixture declaration")
    return namespace, view


def _append_custom_sheet_view_page_break_declarations(
    path: Path,
    count: int,
    *,
    axis: str = "row",
    member_name: str = "xl/worksheets/sheet1.xml",
    child_name: str = "brk",
    custom_sheet_view_namespace: str | None = None,
    break_container_namespace: str | None = None,
    break_child_namespace: str | None = None,
) -> None:
    """Append direct Custom View break records without a workbook reader."""
    container_name = {"row": "rowBreaks", "column": "colBreaks"}.get(axis)
    if container_name is None:
        raise ValueError(f"Unsupported page-break axis {axis!r}")
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace, view = _custom_sheet_view_page_break_target(
        root,
        custom_sheet_view_namespace=custom_sheet_view_namespace,
    )
    container_namespace = break_container_namespace or namespace
    child_namespace = break_child_namespace or container_namespace
    qualified_container_name = (
        f"{{{container_namespace}}}{container_name}"
        if container_namespace
        else container_name
    )
    qualified_child_name = (
        f"{{{child_namespace}}}{child_name}" if child_namespace else child_name
    )
    container = next(
        (child for child in view if child.tag == qualified_container_name),
        None,
    )
    if container is None:
        container = ElementTree.Element(qualified_container_name)
        view.append(container)
    for _ in range(count):
        ElementTree.SubElement(
            container,
            qualified_child_name,
            {"id": "10", "min": "0", "max": "16383", "man": "1"},
        )
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_custom_sheet_view_page_break_containers(
    path: Path,
    count: int,
    *,
    axis: str = "row",
    member_name: str = "xl/worksheets/sheet1.xml",
    custom_sheet_view_namespace: str | None = None,
    break_container_namespace: str | None = None,
) -> None:
    """Append Custom View break containers without a workbook reader."""
    container_name = {"row": "rowBreaks", "column": "colBreaks"}.get(axis)
    if container_name is None:
        raise ValueError(f"Unsupported page-break axis {axis!r}")
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace, view = _custom_sheet_view_page_break_target(
        root,
        custom_sheet_view_namespace=custom_sheet_view_namespace,
    )
    container_namespace = break_container_namespace or namespace
    qualified_container_name = (
        f"{{{container_namespace}}}{container_name}"
        if container_namespace
        else container_name
    )
    for _ in range(count):
        view.append(ElementTree.Element(qualified_container_name))
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_custom_sheet_view_opaque_descendants(
    path: Path,
    count: int,
    *,
    member_name: str = "xl/worksheets/sheet1.xml",
    custom_sheet_view_namespace: str | None = None,
    nested: bool = False,
) -> None:
    """Append raw Custom View descendants without invoking a workbook reader."""
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    _namespace, view = _custom_sheet_view_page_break_target(
        root,
        custom_sheet_view_namespace=custom_sheet_view_namespace,
    )
    opaque_namespace = "urn:formulafence:custom-view-subtree"
    parent = view
    if nested:
        parent = ElementTree.SubElement(view, f"{{{opaque_namespace}}}opaque")
    for _ in range(count):
        ElementTree.SubElement(parent, f"{{{opaque_namespace}}}entry")
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _make_custom_sheet_view_subtree_model(path: Path) -> Path:
    """Create one otherwise-empty standard Custom View for subtree tests."""
    workbook = make_model(path)
    _append_custom_sheet_view_declarations(workbook, 1)
    return workbook


def _insert_worksheet_child_after_sheet_data(
    root: ElementTree.Element,
    child: ElementTree.Element,
) -> None:
    """Insert a worksheet child in a reader-compatible location for a fixture."""
    sheet_data_index = next(
        (
            index
            for index, current in enumerate(root)
            if current.tag.rsplit("}", maxsplit=1)[-1] == "sheetData"
        ),
        None,
    )
    root.insert(0 if sheet_data_index is None else sheet_data_index + 1, child)


def _append_data_validation_declarations(
    path: Path,
    references: tuple[str, ...],
    *,
    alternate_namespace: str | None = None,
    formula: str | None = None,
) -> None:
    """Add direct validation declarations without invoking a workbook reader."""
    member_name = "xl/worksheets/sheet1.xml"
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_container_name = (
        f"{{{namespace}}}dataValidations" if namespace else "dataValidations"
    )
    qualified_child_name = (
        f"{{{namespace}}}dataValidation" if namespace else "dataValidation"
    )
    qualified_formula_name = f"{{{namespace}}}formula1" if namespace else "formula1"
    container = next(
        (
            child
            for child in root
            if child.tag.rsplit("}", maxsplit=1)[-1] == "dataValidations"
        ),
        None,
    )
    if container is None:
        container = ElementTree.Element(qualified_container_name)
        _insert_worksheet_child_after_sheet_data(root, container)
    for reference in references:
        declaration = ElementTree.Element(
            qualified_child_name,
            {"type": "whole", "sqref": reference},
        )
        if alternate_namespace is not None:
            declaration.tag = f"{{{alternate_namespace}}}dataValidation"
        if formula is not None:
            ElementTree.SubElement(declaration, qualified_formula_name).text = formula
        container.append(declaration)
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_conditional_formatting_declarations(
    path: Path,
    references: tuple[str, ...],
    *,
    rule_namespaces: tuple[str | None, ...] = (None,),
    formula: str | None = None,
) -> None:
    """Add direct conditional-formatting declarations without a reader."""
    member_name = "xl/worksheets/sheet1.xml"
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_container_name = (
        f"{{{namespace}}}conditionalFormatting" if namespace else "conditionalFormatting"
    )
    qualified_rule_name = f"{{{namespace}}}cfRule" if namespace else "cfRule"
    qualified_formula_name = f"{{{namespace}}}formula" if namespace else "formula"
    priority = 1
    for reference in references:
        container = ElementTree.Element(qualified_container_name, {"sqref": reference})
        for rule_namespace in rule_namespaces:
            rule_name = (
                qualified_rule_name
                if rule_namespace is None
                else f"{{{rule_namespace}}}cfRule"
            )
            rule = ElementTree.SubElement(
                container,
                rule_name,
                {"type": "expression", "priority": str(priority)},
            )
            priority += 1
            if formula is not None:
                ElementTree.SubElement(rule, qualified_formula_name).text = formula
        _insert_worksheet_child_after_sheet_data(root, container)
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_scenario_manager_containers(
    path: Path,
    references: tuple[str, ...],
    *,
    scenario_count: int = 1,
    input_cell_count: int = 1,
    alternate_namespace: str | None = None,
) -> None:
    """Add Scenario Manager declarations without invoking a workbook reader."""
    member_name = "xl/worksheets/sheet1.xml"
    with ZipFile(path) as archive:
        worksheet = archive.read(member_name)
    root = ElementTree.fromstring(worksheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified_container_name = f"{{{namespace}}}scenarios" if namespace else "scenarios"
    qualified_scenario_name = f"{{{namespace}}}scenario" if namespace else "scenario"
    qualified_input_name = f"{{{namespace}}}inputCells" if namespace else "inputCells"
    scenario_name = (
        qualified_scenario_name
        if alternate_namespace is None
        else f"{{{alternate_namespace}}}scenario"
    )
    input_name = (
        qualified_input_name
        if alternate_namespace is None
        else f"{{{alternate_namespace}}}inputCells"
    )
    for container_index, reference in enumerate(references):
        container = ElementTree.Element(qualified_container_name, {"sqref": reference})
        for scenario_index in range(scenario_count):
            scenario = ElementTree.SubElement(
                container,
                scenario_name,
                {
                    "name": f"FormulaFence audit {container_index}-{scenario_index}",
                    "locked": "0",
                    "hidden": "0",
                    "count": str(input_cell_count),
                },
            )
            for _ in range(input_cell_count):
                ElementTree.SubElement(scenario, input_name, {"r": "A1", "val": "1"})
        _insert_worksheet_child_after_sheet_data(root, container)
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _append_stylesheet_catalog_declarations(
    path: Path,
    parent_names: tuple[str, ...],
    child_name: str,
    count: int,
    *,
    alternate_parent_namespace: str | None = None,
) -> int:
    """Append reader-visible style records and return their direct-child count."""
    member_name = "xl/styles.xml"
    with ZipFile(path) as archive:
        styles = archive.read(member_name)
    root = ElementTree.fromstring(styles)
    namespace = root.tag.partition("}")[0].removeprefix("{")

    def qualified(name: str) -> str:
        return f"{{{namespace}}}{name}" if namespace else name

    parent = root
    for parent_name in parent_names:
        child = next(
            (
                current
                for current in parent
                if current.tag.rsplit("}", maxsplit=1)[-1] == parent_name
            ),
            None,
        )
        if child is None:
            child = ElementTree.Element(qualified(parent_name))
            parent.append(child)
        parent = child
    if alternate_parent_namespace is not None:
        parent.tag = f"{{{alternate_parent_namespace}}}{parent_names[-1]}"

    for index in range(count):
        declaration = ElementTree.Element(qualified(child_name))
        if child_name == "numFmt":
            declaration.attrib.update(
                {"numFmtId": str(164 + index), "formatCode": "0.000"}
            )
        elif child_name == "font":
            ElementTree.SubElement(declaration, qualified("name"), {"val": "Arial"})
        elif child_name == "fill":
            ElementTree.SubElement(
                declaration,
                qualified("patternFill"),
                {"patternType": "none"},
            )
        elif child_name == "xf":
            declaration.attrib.update(
                {"numFmtId": "0", "fontId": "0", "fillId": "0", "borderId": "0"}
            )
        elif child_name == "cellStyle":
            declaration.attrib.update(
                {"name": f"FormulaFence audit {index}", "xfId": "0"}
            )
        elif child_name in {"rgbColor", "color"}:
            declaration.set("rgb", "FF000000")
        elif child_name == "tableStyle":
            declaration.set("name", f"FormulaFence audit {index}")
        elif child_name == "tableStyleElement":
            declaration.set("type", "wholeTable")
        elif child_name == "ext":
            declaration.set("uri", f"urn:formulafence:archive-safety:{index}")
        elif child_name == "patternFill":
            declaration.set("patternType", "none")
        elif child_name == "stop":
            declaration.set("position", str(index))
        parent.append(declaration)

    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )
    return len(parent)


def _append_stylesheet_start_tag_attributes(
    path: Path,
    count: int,
    *,
    marker: bytes,
    after: bytes | None = None,
    member_name: str = "xl/styles.xml",
) -> int:
    """Append distinct attributes to one raw style XML start tag."""
    with ZipFile(path) as archive:
        stylesheet = archive.read(member_name)
    search_start = stylesheet.index(after) if after is not None else 0
    start = stylesheet.index(marker, search_start)
    end = stylesheet.index(b">", start)
    insertion = end - 1 if stylesheet[end - 1 : end] == b"/" else end
    attributes = b"".join(
        f' ff{index:08x}="x"'.encode() for index in range(count)
    )
    _replace_member(
        path,
        member_name,
        stylesheet[:insertion] + attributes + stylesheet[insertion:],
    )
    return end - start + len(attributes) + 1


def _append_stylesheet_opaque_root_text(
    path: Path,
    character_count: int,
    *,
    member_name: str = "xl/styles.xml",
) -> None:
    """Append an ignored stylesheet-root text node without building a style model."""
    with ZipFile(path) as archive:
        stylesheet = archive.read(member_name)
    closing_offset = stylesheet.rfind(b"</")
    if closing_offset < 0:
        raise ValueError("style fixture XML has no closing root tag")
    opaque_text = (
        b'<ff:opaque xmlns:ff="urn:formulafence:archive-safety">'
        + (b"x" * character_count)
        + b"</ff:opaque>"
    )
    _replace_member(
        path,
        member_name,
        stylesheet[:closing_offset] + opaque_text + stylesheet[closing_offset:],
    )


def _append_stylesheet_opaque_root_markup(
    path: Path,
    markup: bytes,
    *,
    member_name: str = "xl/styles.xml",
) -> None:
    """Append raw ignored XML markup without materializing a style model."""
    with ZipFile(path) as archive:
        stylesheet = archive.read(member_name)
    closing_offset = stylesheet.rfind(b"</")
    if closing_offset < 0:
        raise ValueError("style fixture XML has no closing root tag")
    _replace_member(
        path,
        member_name,
        stylesheet[:closing_offset] + markup + stylesheet[closing_offset:],
    )


def _append_stylesheet_opaque_root_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    member_name: str = "xl/styles.xml",
) -> int:
    """Append a direct non-Stylesheet child without building a style model."""
    with ZipFile(path) as archive:
        stylesheet = archive.read(member_name)
    closing_offset = stylesheet.rfind(b"</")
    if closing_offset < 0:
        raise ValueError("stylesheet fixture XML has no closing root tag")
    namespace = b"urn:formulafence:archive-safety"
    entries = b"<ff:opaque/>" * count
    if nested:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + b"<ff:nested>"
            + entries
            + b"</ff:nested></ff:container>"
        )
        element_count = count + 2
    else:
        inserted = (
            b'<ff:container xmlns:ff="'
            + namespace
            + b'">'
            + entries
            + b"</ff:container>"
        )
        element_count = count + 1
    _replace_member(
        path,
        member_name,
        stylesheet[:closing_offset] + inserted + stylesheet[closing_offset:],
    )
    return element_count


def _append_stylesheet_extension_list_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    within_cell_style: bool = False,
    extension_namespace: str | None = None,
    member_name: str = "xl/styles.xml",
) -> int:
    """Append a stylesheet extension tree while preserving its namespace."""
    with ZipFile(path) as archive:
        stylesheet = archive.read(member_name)
    root = ElementTree.fromstring(stylesheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    if not namespace:
        raise ValueError("stylesheet fixture XML has no namespace")
    extension_namespace = extension_namespace or namespace
    extension_list = ElementTree.Element(f"{{{extension_namespace}}}extLst")
    extension = ElementTree.SubElement(
        extension_list,
        f"{{{extension_namespace}}}ext",
        {"uri": "{1F4A6F6A-EB4A-4C41-9C9E-9231F8EAF007}"},
    )
    if nested:
        parent = ElementTree.SubElement(
            extension,
            "{urn:formulafence:archive-safety}container",
        )
        parent = ElementTree.SubElement(
            parent,
            "{urn:formulafence:archive-safety}nested",
        )
        element_count = count + 4
    else:
        parent = extension
        element_count = count + 2
    for _ in range(count):
        ElementTree.SubElement(parent, "{urn:formulafence:archive-safety}opaque")
    if within_cell_style:
        cell_formats = next(
            (child for child in root if child.tag == f"{{{namespace}}}cellXfs"),
            None,
        )
        if cell_formats is None or not list(cell_formats):
            raise ValueError("stylesheet fixture has no cell style")
        cell_formats[0].append(extension_list)
    else:
        root.append(extension_list)
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )
    return element_count


def _append_stylesheet_opaque_catalog_xml_elements(
    path: Path,
    count: int,
    *,
    nested: bool = False,
    member_name: str = "xl/styles.xml",
) -> int:
    """Append an ignored direct child inside the named ``cellXfs`` catalog."""
    with ZipFile(path) as archive:
        stylesheet = archive.read(member_name)
    closing_offset = stylesheet.find(b"</cellXfs>")
    if closing_offset < 0:
        raise ValueError("stylesheet fixture XML has no cellXfs container")
    entries = b"<ff:opaque/>" * count
    if nested:
        inserted = (
            b'<ff:container xmlns:ff="urn:formulafence:archive-safety">'
            + b"<ff:nested>"
            + entries
            + b"</ff:nested></ff:container>"
        )
        element_count = count + 2
    else:
        inserted = (
            b'<ff:container xmlns:ff="urn:formulafence:archive-safety">'
            + entries
            + b"</ff:container>"
        )
        element_count = count + 1
    _replace_member(
        path,
        member_name,
        stylesheet[:closing_offset] + inserted + stylesheet[closing_offset:],
    )
    return element_count


def _append_cell_style_xml_elements(
    path: Path,
    count: int,
    *,
    known_alignment: bool = False,
    nested: bool = False,
    member_name: str = "xl/styles.xml",
) -> int:
    """Append retained XML beneath one materialized ``cellXfs`` record."""
    with ZipFile(path) as archive:
        stylesheet = archive.read(member_name)
    root = ElementTree.fromstring(stylesheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    cell_formats = next(
        (child for child in root if child.tag == f"{{{namespace}}}cellXfs"),
        None,
    )
    if cell_formats is None or not list(cell_formats):
        raise ValueError("stylesheet fixture has no cell style")
    cell_style = cell_formats[0]
    if known_alignment:
        for _ in range(count):
            ElementTree.SubElement(cell_style, f"{{{namespace}}}alignment")
        element_count = count
    else:
        parent = ElementTree.SubElement(
            cell_style,
            "{urn:formulafence:archive-safety}container",
        )
        if nested:
            parent = ElementTree.SubElement(
                parent,
                "{urn:formulafence:archive-safety}nested",
            )
            element_count = count + 2
        else:
            element_count = count + 1
        for _ in range(count):
            ElementTree.SubElement(parent, "{urn:formulafence:archive-safety}opaque")
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )
    return element_count


def _replace_stylesheet_root_namespace(path: Path, namespace: str) -> None:
    """Change only the stylesheet root namespace for parser-selection coverage."""
    member_name = "xl/styles.xml"
    with ZipFile(path) as archive:
        stylesheet = archive.read(member_name)
    root = ElementTree.fromstring(stylesheet)
    root.tag = f"{{{namespace}}}{root.tag.rsplit('}', maxsplit=1)[-1]}"
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _replace_stylesheet_root_local_name(path: Path, local_name: str) -> None:
    """Change the root local name without changing its child inventory."""
    member_name = "xl/styles.xml"
    with ZipFile(path) as archive:
        stylesheet = archive.read(member_name)
    root = ElementTree.fromstring(stylesheet)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    root.tag = f"{{{namespace}}}{local_name}" if namespace else local_name
    _replace_member(
        path,
        member_name,
        ElementTree.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _last_central_directory_offset(contents: bytes | bytearray) -> int:
    offset = contents.rfind(b"PK\x01\x02")
    assert offset >= 0
    return offset


def _reject_before_workbook_readers(monkeypatch: pytest.MonkeyPatch, path: Path) -> str:
    def unexpected_reader(*args, **kwargs):
        raise AssertionError("a workbook reader ran before archive preflight rejected input")

    monkeypatch.setattr(workbook_module, "_workbook_tab_order_metadata", unexpected_reader)
    with pytest.raises(WorkbookLoadError, match="safety preflight") as error:
        load_snapshot(path)
    return str(error.value)


_WORKBOOK_AUXILIARY_CATALOG_CASES = (
    (
        "bookViews",
        "workbookView",
        {},
        "_OOXML_READER_MAX_WORKBOOK_BOOK_VIEW_COUNT",
        "workbook book-view declarations",
    ),
    (
        "customWorkbookViews",
        "customWorkbookView",
        {
            "name": "FormulaFence audit",
            "guid": "{11111111-1111-1111-1111-111111111111}",
            "windowWidth": "100",
            "windowHeight": "100",
            "activeSheetId": "0",
        },
        "_OOXML_READER_MAX_WORKBOOK_CUSTOM_VIEW_COUNT",
        "workbook custom-view declarations",
    ),
    (
        "functionGroups",
        "functionGroup",
        {"name": "FormulaFence audit"},
        "_OOXML_READER_MAX_WORKBOOK_FUNCTION_GROUP_COUNT",
        "workbook function-group declarations",
    ),
    (
        "smartTagTypes",
        "smartTagType",
        {
            "namespaceUri": "urn:formulafence:archive-safety",
            "name": "audit",
            "url": "https://example.invalid/formulafence",
        },
        "_OOXML_READER_MAX_WORKBOOK_SMART_TAG_TYPE_COUNT",
        "workbook smart-tag type declarations",
    ),
    (
        "webPublishObjects",
        "webPublishObject",
        {"id": "1", "divId": "audit", "destinationFile": "audit.html"},
        "_OOXML_READER_MAX_WORKBOOK_WEB_PUBLISH_OBJECT_COUNT",
        "workbook web-publish-object declarations",
    ),
)


_STYLESHEET_CATALOG_CASES = (
    (
        ("numFmts",),
        "numFmt",
        "_OOXML_READER_MAX_NUMBER_FORMAT_COUNT",
        "number-format records",
    ),
    (("fonts",), "font", "_OOXML_READER_MAX_FONT_COUNT", "font records"),
    (("fills",), "fill", "_OOXML_READER_MAX_FILL_COUNT", "fill records"),
    (
        ("fills", "fill"),
        "patternFill",
        "_OOXML_READER_MAX_FILL_CHILD_COUNT",
        "fill child records",
    ),
    (
        ("fills", "fill", "gradientFill"),
        "stop",
        "_OOXML_READER_MAX_GRADIENT_FILL_STOP_COUNT",
        "gradient-fill stops",
    ),
    (("borders",), "border", "_OOXML_READER_MAX_BORDER_COUNT", "border records"),
    (
        ("cellStyleXfs",),
        "xf",
        "_OOXML_READER_MAX_BASE_CELL_STYLE_COUNT",
        "base cell styles",
    ),
    (
        ("cellXfs",),
        "xf",
        "_OOXML_READER_MAX_CELL_STYLE_COUNT",
        "cell styles",
    ),
    (
        ("cellStyles",),
        "cellStyle",
        "_OOXML_READER_MAX_NAMED_CELL_STYLE_COUNT",
        "named cell styles",
    ),
    (
        ("dxfs",),
        "dxf",
        "_OOXML_READER_MAX_DIFFERENTIAL_STYLE_COUNT",
        "differential styles",
    ),
    (
        ("colors", "indexedColors"),
        "rgbColor",
        "_OOXML_READER_MAX_STYLE_COLOR_COUNT",
        "stylesheet colour records",
    ),
    (
        ("colors", "mruColors"),
        "color",
        "_OOXML_READER_MAX_STYLE_COLOR_COUNT",
        "stylesheet colour records",
    ),
    (
        ("tableStyles",),
        "tableStyle",
        "_OOXML_READER_MAX_TABLE_STYLE_COUNT",
        "table styles",
    ),
    (
        ("tableStyles", "tableStyle"),
        "tableStyleElement",
        "_OOXML_READER_MAX_TABLE_STYLE_ELEMENT_COUNT",
        "table-style elements",
    ),
    (
        ("extLst",),
        "ext",
        "_OOXML_READER_MAX_STYLE_EXTENSION_COUNT",
        "stylesheet extension records",
    ),
)


def test_archive_preflight_accepts_an_ordinary_workbook(tmp_path: Path) -> None:
    workbook = make_model(tmp_path / "ordinary.xlsx")

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"
    assert set(snapshot.sheets) == {"Inputs", "Model", "Dashboard", "Control"}


def test_archive_preflight_accepts_a_valid_zip64_workbook(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(zipfile, "ZIP64_LIMIT", 1)
    workbook = make_model(tmp_path / "zip64.xlsx")

    assert bytes((80, 75, 6, 6)) in workbook.read_bytes()
    assert load_snapshot(workbook).file_type == "xlsx"


def test_load_snapshot_uses_one_stable_private_source_after_copy(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "source-race.xlsx")
    original_sha256 = hashlib.sha256(workbook.read_bytes()).hexdigest()
    replacement = tmp_path / "replacement.xlsx"
    shutil.copyfile(workbook, replacement)
    with ZipFile(replacement) as archive:
        model_xml = archive.read("xl/worksheets/sheet2.xml")
    assert b"Inputs!B2*2" in model_xml
    _replace_member(replacement, "xl/worksheets/sheet2.xml", model_xml.replace(
        b"Inputs!B2*2", b"Inputs!B3*2", 1
    ))
    replacement_sha256 = hashlib.sha256(replacement.read_bytes()).hexdigest()
    stable_sources: list[Path] = []
    materialize_source = workbook_module._materialize_stable_workbook_source

    def materialize_then_replace(path: Path) -> Path:
        stable_source = materialize_source(path)
        stable_sources.append(stable_source)
        shutil.copyfile(replacement, path)
        return stable_source

    monkeypatch.setattr(
        workbook_module,
        "_materialize_stable_workbook_source",
        materialize_then_replace,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.path == workbook
    assert snapshot.sha256 == original_sha256
    assert snapshot.cells[("Model", "B2")].formula == "=Inputs!B2*2"
    assert hashlib.sha256(workbook.read_bytes()).hexdigest() == replacement_sha256
    assert stable_sources and not stable_sources[0].exists()


def test_load_snapshot_removes_stable_private_source_after_preflight_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "private-source-cleanup.xlsx")
    stable_sources: list[Path] = []
    materialize_source = workbook_module._materialize_stable_workbook_source

    def materialize_then_record(path: Path) -> Path:
        stable_source = materialize_source(path)
        stable_sources.append(stable_source)
        return stable_source

    def reject_private_source(path: Path):
        assert stable_sources == [path]
        raise WorkbookLoadError("controlled private-source preflight failure")

    monkeypatch.setattr(
        workbook_module,
        "_materialize_stable_workbook_source",
        materialize_then_record,
    )
    monkeypatch.setattr(workbook_module, "_validate_ooxml_archive", reject_private_source)

    with pytest.raises(WorkbookLoadError, match="controlled private-source"):
        load_snapshot(workbook)

    assert stable_sources and not stable_sources[0].exists()


def test_archive_preflight_rejects_source_size_before_any_zip_reader(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_ARCHIVE_MAX_SOURCE_BYTES", 1)

    def unexpected_zip_reader(*args, **kwargs):
        raise AssertionError("ZipFile ran before the source-size safety gate")

    monkeypatch.setattr(workbook_module, "ZipFile", unexpected_zip_reader)

    _reject_before_workbook_readers(monkeypatch, workbook)


def test_archive_preflight_rejects_excessive_entry_count_before_zip_reader(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-members.xlsx")
    with ZipFile(workbook) as archive:
        member_count = len(archive.infolist())
    monkeypatch.setattr(workbook_module, "_OOXML_ARCHIVE_MAX_ENTRY_COUNT", member_count - 1)

    def unexpected_zip_reader(*args, **kwargs):
        raise AssertionError("ZipFile ran before the central-directory safety gate")

    monkeypatch.setattr(workbook_module, "ZipFile", unexpected_zip_reader)

    _reject_before_workbook_readers(monkeypatch, workbook)


def test_archive_preflight_rejects_an_oversized_member_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-member.xlsx")
    with ZipFile(workbook) as archive:
        largest_member = max(member.file_size for member in archive.infolist())
    _append_member(workbook, "xl/media/filler.bin", b"x" * (largest_member + 1))
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_ARCHIVE_MAX_MEMBER_UNCOMPRESSED_BYTES",
        largest_member,
    )

    _reject_before_workbook_readers(monkeypatch, workbook)


def test_archive_preflight_rejects_excessive_aggregate_size_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-total.xlsx")
    with ZipFile(workbook) as archive:
        original_total = sum(member.file_size for member in archive.infolist())
    _append_member(workbook, "xl/media/filler.bin", b"x")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_ARCHIVE_MAX_TOTAL_UNCOMPRESSED_BYTES",
        original_total,
    )

    _reject_before_workbook_readers(monkeypatch, workbook)


def test_archive_preflight_rejects_a_compression_bomb_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "compression-bomb.xlsx")
    with ZipFile(workbook) as archive:
        ordinary_ratio = max(
            member.file_size // max(member.compress_size, 1)
            for member in archive.infolist()
        )
    _append_member(workbook, "xl/media/compressed.bin", b"x" * 200_000)
    with ZipFile(workbook) as archive:
        compressed_member = archive.getinfo("xl/media/compressed.bin")
    assert compressed_member.file_size > compressed_member.compress_size * ordinary_ratio
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_ARCHIVE_MAX_COMPRESSION_RATIO",
        ordinary_ratio,
    )

    _reject_before_workbook_readers(monkeypatch, workbook)


def test_semantic_reader_preflight_rejects_an_oversized_xml_part_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-xml.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_XML_PART_BYTES", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "semantic reader" in message
    assert "XML part" in message


def test_semantic_reader_preflight_rejects_aggregate_xml_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-xml-total.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_TOTAL_XML_BYTES", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "aggregate XML" in message


def test_semantic_reader_preflight_rejects_excessive_cells_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-cells.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_WORKSHEET_CELL_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "populated worksheet cells" in message


def test_dependency_edge_budget_bounds_named_formula_fanout(
    tmp_path: Path,
) -> None:
    workbook = _named_formula_fanout_workbook(tmp_path / "named-fanout.xlsx")

    snapshot = load_snapshot(workbook, max_dependency_edges=12)
    assert sum(len(values) for values in snapshot.reverse_dependencies.values()) == 12
    assert snapshot.range_dependencies == []

    with pytest.raises(
        WorkbookLoadError,
        match=r"dependency graph exceeds max_dependency_edges=11",
    ):
        load_snapshot(workbook, max_dependency_edges=11)
    with pytest.raises(
        WorkbookLoadError,
        match="max_dependency_edges must be at least 1",
    ):
        load_snapshot(workbook, max_dependency_edges=0)


def test_semantic_reader_preflight_rejects_excessive_row_dimensions_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-row-dimensions.xlsx")
    _append_row_dimension_declarations(workbook, 2)
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_ROW_DIMENSION_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "row-dimension declarations" in message


def test_semantic_reader_preflight_rejects_default_row_dimension_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-row-dimension-limit.xlsx")
    _append_row_dimension_declarations(
        workbook,
        workbook_module._OOXML_READER_MAX_ROW_DIMENSION_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "row-dimension declarations" in message


def test_semantic_reader_preflight_counts_row_dimensions_across_worksheets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-row-dimensions.xlsx")
    _append_row_dimension_declarations(workbook, 1)
    _append_row_dimension_declarations(workbook, 1, member_name="xl/worksheets/sheet2.xml")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_ROW_DIMENSION_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "row-dimension declarations" in message


def test_semantic_reader_preflight_accepts_row_dimensions_at_the_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "row-dimensions-at-limit.xlsx")
    _append_row_dimension_declarations(workbook, 2)
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_ROW_DIMENSION_COUNT", 2)

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


@pytest.mark.parametrize(
    "attributes",
    (
        {"spans": "1:1"},
        {"{urn:formulafence:archive-safety}audit": "1"},
    ),
)
def test_semantic_reader_preflight_ignores_row_attributes_that_do_not_create_dimensions(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    attributes: dict[str, str],
) -> None:
    workbook = make_model(tmp_path / "non-dimension-row-attributes.xlsx")
    _append_row_dimension_declarations(workbook, 1, attributes=attributes)
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_ROW_DIMENSION_COUNT", 0)

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_counts_unknown_unqualified_row_attributes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "unknown-row-dimension-attribute.xlsx")
    _append_row_dimension_declarations(
        workbook,
        1,
        attributes={"formulafenceAudit": "1"},
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_ROW_DIMENSION_COUNT", 0)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "row-dimension declarations" in message


def test_semantic_reader_preflight_ignores_foreign_namespace_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "foreign-row-dimensions.xlsx")
    _append_row_dimension_declarations(
        workbook,
        1,
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_ROW_DIMENSION_COUNT", 0)

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_rejects_excessive_column_dimensions_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-column-dimensions.xlsx")
    _append_column_dimension_declarations(workbook, 2)
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_COLUMN_DIMENSION_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "column-dimension declarations" in message


def test_semantic_reader_preflight_rejects_default_column_dimension_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-column-dimension-limit.xlsx")
    _append_column_dimension_declarations(
        workbook,
        workbook_module._OOXML_READER_MAX_COLUMN_DIMENSION_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "column-dimension declarations" in message


def test_semantic_reader_preflight_counts_column_dimensions_across_worksheets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-column-dimensions.xlsx")
    _append_column_dimension_declarations(workbook, 1)
    _append_column_dimension_declarations(
        workbook,
        1,
        member_name="xl/worksheets/sheet2.xml",
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_COLUMN_DIMENSION_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "column-dimension declarations" in message


def test_semantic_reader_preflight_accepts_column_dimensions_at_the_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "column-dimensions-at-limit.xlsx")
    _append_column_dimension_declarations(workbook, 2)
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_COLUMN_DIMENSION_COUNT", 2)

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_counts_unknown_unqualified_column_attributes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "unknown-column-dimension-attribute.xlsx")
    _append_column_dimension_declarations(
        workbook,
        1,
        attributes={"formulafenceAudit": "1"},
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_COLUMN_DIMENSION_COUNT", 0)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "column-dimension declarations" in message


def test_semantic_reader_preflight_ignores_foreign_namespace_columns(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "foreign-column-dimensions.xlsx")
    _append_column_dimension_declarations(
        workbook,
        1,
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_COLUMN_DIMENSION_COUNT", 0)

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_rejects_excessive_column_dimension_containers_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-column-dimension-containers.xlsx")
    _append_column_dimension_containers(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_COLUMN_DIMENSION_CONTAINER_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "column-dimension containers" in message


def test_semantic_reader_preflight_rejects_default_column_dimension_container_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-column-dimension-container-limit.xlsx")
    _append_column_dimension_containers(
        workbook,
        workbook_module._OOXML_READER_MAX_COLUMN_DIMENSION_CONTAINER_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "column-dimension containers" in message


def test_semantic_reader_preflight_counts_column_dimension_containers_across_worksheets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-column-dimension-containers.xlsx")
    _append_column_dimension_containers(workbook, 1)
    _append_column_dimension_containers(
        workbook,
        1,
        member_name="xl/worksheets/sheet2.xml",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_COLUMN_DIMENSION_CONTAINER_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "column-dimension containers" in message


def test_semantic_reader_preflight_accepts_column_dimension_containers_at_the_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "column-dimension-containers-at-limit.xlsx")
    _append_column_dimension_containers(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_COLUMN_DIMENSION_CONTAINER_COUNT",
        2,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_ignores_foreign_namespace_column_containers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "foreign-column-dimension-containers.xlsx")
    _append_column_dimension_containers(
        workbook,
        1,
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_COLUMN_DIMENSION_CONTAINER_COUNT",
        0,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_rejects_excessive_page_break_declarations_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-page-break-declarations.xlsx")
    _append_page_break_declarations(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


def test_semantic_reader_preflight_rejects_default_page_break_declaration_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-page-break-declaration-limit.xlsx")
    _append_page_break_declarations(
        workbook,
        workbook_module._OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


def test_semantic_reader_preflight_counts_page_break_declarations_across_worksheets_and_axes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-page-break-declarations.xlsx")
    _append_page_break_declarations(workbook, 1)
    _append_page_break_declarations(
        workbook,
        1,
        axis="column",
        member_name="xl/worksheets/sheet2.xml",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


def test_semantic_reader_preflight_accepts_one_complete_published_page_break_allowance(
    tmp_path: Path,
) -> None:
    workbook = make_model(tmp_path / "page-breaks-at-published-limit.xlsx")
    _append_page_break_declarations(workbook, 1_026)
    _append_page_break_declarations(workbook, 1_026, axis="column")

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


@pytest.mark.parametrize(
    ("child_name", "alternate_namespace"),
    (
        ("brk", "urn:formulafence:archive-safety"),
        ("formulafenceAudit", None),
    ),
)
def test_semantic_reader_preflight_counts_all_direct_page_break_children(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    child_name: str,
    alternate_namespace: str | None,
) -> None:
    workbook = make_model(tmp_path / "nonstandard-page-break-child.xlsx")
    _append_page_break_declarations(
        workbook,
        1,
        child_name=child_name,
        alternate_namespace=alternate_namespace,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


def test_semantic_reader_preflight_counts_strict_page_break_declarations_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_strict_worksheet_print_layout_model(
        tmp_path / "strict-page-break-declarations.xlsx"
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


def test_semantic_reader_preflight_rejects_custom_view_page_break_declarations_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_custom_workbook_view_model(
        tmp_path / "too-many-custom-view-page-break-declarations.xlsx"
    )
    _append_custom_sheet_view_page_break_declarations(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


def test_semantic_reader_preflight_accepts_custom_view_page_break_declarations_at_the_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_custom_workbook_view_model(
        tmp_path / "custom-view-page-break-declarations-at-limit.xlsx"
    )
    _append_custom_sheet_view_page_break_declarations(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT",
        1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_rejects_default_custom_view_page_break_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_custom_workbook_view_model(
        tmp_path / "default-custom-view-page-break-declaration-limit.xlsx"
    )
    _append_custom_sheet_view_page_break_declarations(
        workbook,
        workbook_module._OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


def test_semantic_reader_preflight_aggregates_custom_view_and_ordinary_page_break_declarations(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_custom_workbook_view_model(
        tmp_path / "aggregate-custom-view-page-break-declarations.xlsx"
    )
    _append_page_break_declarations(workbook, 1)
    _append_custom_sheet_view_page_break_declarations(workbook, 1, axis="column")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


def test_semantic_reader_preflight_counts_strict_custom_view_breaks_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_strict_custom_workbook_view_model(
        tmp_path / "strict-custom-view-page-break-declarations.xlsx"
    )
    _append_custom_sheet_view_page_break_declarations(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


@pytest.mark.parametrize(
    (
        "custom_sheet_view_namespace",
        "break_container_namespace",
        "child_name",
        "break_child_namespace",
    ),
    (
        (
            "urn:formulafence:opaque-custom-view",
            None,
            "brk",
            None,
        ),
        (
            None,
            "urn:formulafence:opaque-page-break-container",
            "brk",
            None,
        ),
        (None, None, "formulafenceAudit", None),
    ),
)
def test_semantic_reader_preflight_counts_opaque_custom_view_page_break_children(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    custom_sheet_view_namespace: str | None,
    break_container_namespace: str | None,
    child_name: str,
    break_child_namespace: str | None,
) -> None:
    workbook = make_custom_workbook_view_model(
        tmp_path / "opaque-custom-view-page-break-child.xlsx"
    )
    _append_custom_sheet_view_page_break_declarations(
        workbook,
        1,
        custom_sheet_view_namespace=custom_sheet_view_namespace,
        break_container_namespace=break_container_namespace,
        child_name=child_name,
        break_child_namespace=break_child_namespace,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_DECLARATION_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break declarations" in message


def test_semantic_reader_preflight_rejects_excessive_page_break_containers_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-page-break-containers.xlsx")
    _append_page_break_containers(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_CONTAINER_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break containers" in message


def test_semantic_reader_preflight_rejects_default_page_break_container_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-page-break-container-limit.xlsx")
    _append_page_break_containers(
        workbook,
        workbook_module._OOXML_READER_MAX_PAGE_BREAK_CONTAINER_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break containers" in message


def test_semantic_reader_preflight_counts_page_break_containers_across_worksheets_and_axes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-page-break-containers.xlsx")
    _append_page_break_containers(workbook, 1)
    _append_page_break_containers(
        workbook,
        1,
        axis="column",
        member_name="xl/worksheets/sheet2.xml",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_CONTAINER_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break containers" in message


def test_semantic_reader_preflight_accepts_page_break_containers_at_the_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "page-break-containers-at-limit.xlsx")
    _append_page_break_containers(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_CONTAINER_COUNT",
        2,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_ignores_foreign_namespace_page_break_containers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "foreign-page-break-containers.xlsx")
    _append_page_break_containers(
        workbook,
        1,
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_CONTAINER_COUNT",
        0,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_rejects_custom_view_page_break_containers_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_custom_workbook_view_model(
        tmp_path / "too-many-custom-view-page-break-containers.xlsx"
    )
    _append_custom_sheet_view_page_break_containers(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_CONTAINER_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "page-break containers" in message


def test_semantic_reader_preflight_accepts_custom_view_page_break_containers_at_the_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_custom_workbook_view_model(
        tmp_path / "custom-view-page-break-containers-at-limit.xlsx"
    )
    _append_custom_sheet_view_page_break_containers(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_PAGE_BREAK_CONTAINER_COUNT",
        2,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_rejects_excessive_content_type_declarations_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-content-types.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CONTENT_TYPE_DECLARATION_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "content-type declarations" in message


def test_semantic_reader_preflight_rejects_excessive_workbook_relationships_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-workbook-relationships.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_RELATIONSHIP_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook relationships" in message


def test_semantic_reader_preflight_rejects_excessive_relationship_part_xml_elements_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-relationship-part-elements.xlsx")
    relationship_member = "xl/_rels/workbook.xml.rels"
    _write_relationship_part_elements(workbook, relationship_member, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_RELATIONSHIP_PART_XML_ELEMENT_COUNT",
        _relationship_part_element_count(workbook, relationship_member) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "relationship-part XML elements" in message


def test_semantic_reader_preflight_counts_opaque_nested_relationship_part_xml_elements(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "opaque-relationship-part-elements.xlsx")
    relationship_member = "xl/private/_rels/opaque.xml.rels"
    _write_relationship_part_elements(
        workbook,
        relationship_member,
        1,
        nested=True,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_RELATIONSHIP_PART_XML_ELEMENT_COUNT",
        _relationship_part_element_count(workbook, relationship_member) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "relationship-part XML elements" in message


def test_semantic_reader_preflight_aggregates_relationship_part_xml_elements(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-relationship-part-elements.xlsx")
    relationship_members = (
        "xl/private/_rels/one.xml.rels",
        "xl/private/_rels/two.xml.rels",
    )
    for relationship_member in relationship_members:
        _write_relationship_part_elements(workbook, relationship_member, 1)
    with ZipFile(workbook) as archive:
        all_relationship_members = tuple(
            member.filename
            for member in archive.infolist()
            if member.filename.casefold().endswith(".rels")
        )
    element_counts = tuple(
        _relationship_part_element_count(workbook, relationship_member)
        for relationship_member in all_relationship_members
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_RELATIONSHIP_PART_XML_ELEMENT_COUNT",
        max(element_counts),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_RELATIONSHIP_XML_ELEMENT_COUNT",
        sum(element_counts) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "aggregate relationship-part XML elements" in message


def test_semantic_reader_preflight_accepts_relationship_part_xml_elements_at_exact_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "relationship-part-elements-at-limits.xlsx")
    with ZipFile(workbook) as archive:
        relationship_members = tuple(
            member.filename
            for member in archive.infolist()
            if member.filename.casefold().endswith(".rels")
        )
    element_counts = tuple(
        _relationship_part_element_count(workbook, relationship_member)
        for relationship_member in relationship_members
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_RELATIONSHIP_PART_XML_ELEMENT_COUNT",
        max(element_counts),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_RELATIONSHIP_XML_ELEMENT_COUNT",
        sum(element_counts),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_rejects_default_relationship_part_element_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-relationship-part-element-limit.xlsx")
    _write_relationship_part_elements(
        workbook,
        "xl/private/_rels/unused.xml.rels",
        workbook_module._OOXML_READER_MAX_RELATIONSHIP_PART_XML_ELEMENT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "relationship-part XML elements" in message


def test_semantic_reader_preflight_preserves_malformed_unused_relationship_part_coverage(
    tmp_path: Path,
) -> None:
    workbook = make_model(tmp_path / "malformed-unused-relationship-part.xlsx")
    _append_member(
        workbook,
        "xl/private/_rels/unused.xml.rels",
        b"<Relationships",
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"
    assert any(
        "could not parse a package relationship XML part" in warning
        for warning in snapshot.parser_warnings
    )


def test_semantic_reader_preflight_rejects_opaque_worksheet_drawing_xml_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_worksheet_drawing_shape_model(
        tmp_path / "opaque-worksheet-drawing.xml.xlsx"
    )
    drawing_member = "xl/drawings/drawing1.xml"
    _append_worksheet_drawing_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_DRAWING_PART_XML_ELEMENT_COUNT",
        _worksheet_drawing_xml_element_count(workbook, drawing_member) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet DrawingML XML structure" in message


def test_semantic_reader_preflight_counts_nested_opaque_worksheet_drawing_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_worksheet_drawing_shape_model(
        tmp_path / "nested-opaque-worksheet-drawing.xml.xlsx"
    )
    drawing_member = "xl/drawings/drawing1.xml"
    _append_worksheet_drawing_xml_elements(workbook, 1, nested=True)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_DRAWING_PART_XML_ELEMENT_COUNT",
        _worksheet_drawing_xml_element_count(workbook, drawing_member) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet DrawingML XML structure" in message


def test_semantic_reader_preflight_follows_strict_worksheet_drawing_relationships(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_worksheet_drawing_shape_model(
        tmp_path / "strict-worksheet-drawing-relationship.xlsx"
    )
    drawing_member = "xl/drawings/drawing1.xml"
    relationship_member = "xl/worksheets/_rels/sheet1.xml.rels"
    with ZipFile(workbook) as archive:
        relationships = archive.read(relationship_member)
    _replace_member(
        workbook,
        relationship_member,
        relationships.replace(
            b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
            b"http://purl.oclc.org/ooxml/officeDocument/relationships/drawing",
        ),
    )
    _append_worksheet_drawing_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_DRAWING_PART_XML_ELEMENT_COUNT",
        _worksheet_drawing_xml_element_count(workbook, drawing_member) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet DrawingML XML structure" in message


def test_semantic_reader_preflight_follows_strict_worksheet_drawing_parts(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_strict_worksheet_drawing_connector_model(
        tmp_path / "strict-worksheet-drawing-part.xlsx"
    )
    drawing_member = "xl/drawings/drawing1.xml"
    _append_worksheet_drawing_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_DRAWING_PART_XML_ELEMENT_COUNT",
        _worksheet_drawing_xml_element_count(workbook, drawing_member) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet DrawingML XML structure" in message


def test_semantic_reader_preflight_aggregates_worksheet_drawing_xml_targets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_worksheet_drawing_shape_model(
        tmp_path / "aggregate-worksheet-drawing.xml.xlsx"
    )
    first_member = "xl/drawings/drawing1.xml"
    second_member = "xl/drawings/drawing2.xml"
    _add_worksheet_drawing_relationship_target(workbook)
    element_counts = (
        _worksheet_drawing_xml_element_count(workbook, first_member),
        _worksheet_drawing_xml_element_count(workbook, second_member),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_DRAWING_PART_XML_ELEMENT_COUNT",
        max(element_counts),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_DRAWING_XML_ELEMENT_COUNT",
        sum(element_counts) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "aggregate worksheet DrawingML XML elements" in message


def test_semantic_reader_preflight_accepts_worksheet_drawing_xml_at_exact_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_worksheet_drawing_shape_model(
        tmp_path / "worksheet-drawing.xml-at-limits.xlsx"
    )
    element_count = _worksheet_drawing_xml_element_count(
        workbook,
        "xl/drawings/drawing1.xml",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_DRAWING_PART_XML_ELEMENT_COUNT",
        element_count,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_DRAWING_XML_ELEMENT_COUNT",
        element_count,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_drawing_shapes.present is True


def test_semantic_reader_preflight_rejects_default_worksheet_drawing_xml_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_worksheet_drawing_shape_model(
        tmp_path / "default-worksheet-drawing.xml-limit.xlsx"
    )
    drawing_member = "xl/drawings/drawing1.xml"
    existing_count = _worksheet_drawing_xml_element_count(workbook, drawing_member)
    _append_worksheet_drawing_xml_elements(
        workbook,
        (
            workbook_module._OOXML_READER_MAX_WORKSHEET_DRAWING_PART_XML_ELEMENT_COUNT
            - existing_count
            + 1
        ),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet DrawingML XML structure" in message


def test_semantic_reader_preflight_ignores_orphan_worksheet_drawing_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "orphan-worksheet-drawing.xml.xlsx")
    orphan_member = "xl/drawings/orphan.xml"
    _append_member(
        workbook,
        orphan_member,
        (
            b'<wsDr xmlns="http://schemas.openxmlformats.org/drawingml/2006/'
            b'spreadsheetDrawing"><opaque /><opaque /></wsDr>'
        ),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_DRAWING_PART_XML_ELEMENT_COUNT",
        1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_rejects_opaque_table_definition_xml_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_table_model(tmp_path / "opaque-table-definition.xml.xlsx")
    table_member = "xl/tables/table1.xml"
    _append_table_definition_xml_elements(workbook, 1)
    with ZipFile(workbook) as archive:
        table_xml = archive.read(table_member)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_TABLE_DEFINITION_PART_XML_ELEMENT_COUNT",
        _table_definition_xml_element_count(workbook, table_member) - 1,
    )
    original_tree_parse = workbook_module._xml_root_from_payload

    def unexpected_table_tree(payload: bytes) -> ElementTree.Element:
        if payload == table_xml:
            raise AssertionError("the over-budget table definition XML tree was materialized")
        return original_tree_parse(payload)

    monkeypatch.setattr(
        workbook_module,
        "_xml_root_from_payload",
        unexpected_table_tree,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "table-definition XML structure" in message


def test_semantic_reader_preflight_counts_nested_opaque_table_definition_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_table_model(tmp_path / "nested-table-definition.xml.xlsx")
    table_member = "xl/tables/table1.xml"
    _append_table_definition_xml_elements(workbook, 1, nested=True)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_TABLE_DEFINITION_PART_XML_ELEMENT_COUNT",
        _table_definition_xml_element_count(workbook, table_member) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "table-definition XML structure" in message


def test_semantic_reader_preflight_follows_strict_noncanonical_table_relationships(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_table_model(tmp_path / "strict-table-definition-target.xlsx")
    table_member = "xl/private/table-definition.xml"
    relationship_member = "xl/worksheets/_rels/sheet1.xml.rels"
    _add_worksheet_table_relationship_target(workbook, target_member=table_member)
    _append_table_definition_xml_elements(workbook, 1, member_name=table_member)
    with ZipFile(workbook) as archive:
        relationships = archive.read(relationship_member)
    _replace_member(
        workbook,
        relationship_member,
        relationships.replace(
            b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/table",
            b"http://purl.oclc.org/ooxml/officeDocument/relationships/table",
        ),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_TABLE_DEFINITION_PART_XML_ELEMENT_COUNT",
        _table_definition_xml_element_count(workbook, table_member) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "table-definition XML structure" in message


def test_semantic_reader_preflight_aggregates_table_definition_xml_targets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_table_model(tmp_path / "aggregate-table-definitions.xlsx")
    first_member = "xl/tables/table1.xml"
    second_member = "xl/tables/table2.xml"
    with ZipFile(workbook) as archive:
        table = archive.read(first_member)
    _append_member(workbook, second_member, table)
    element_counts = (
        _table_definition_xml_element_count(workbook, first_member),
        _table_definition_xml_element_count(workbook, second_member),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_TABLE_DEFINITION_PART_XML_ELEMENT_COUNT",
        max(element_counts),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_TABLE_DEFINITION_XML_ELEMENT_COUNT",
        sum(element_counts) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "aggregate table-definition XML elements" in message


def test_semantic_reader_preflight_accepts_table_definition_xml_at_exact_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_table_model(tmp_path / "table-definition-at-limits.xlsx")
    table_member = "xl/tables/table1.xml"
    element_count = _table_definition_xml_element_count(workbook, table_member)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_TABLE_DEFINITION_PART_XML_ELEMENT_COUNT",
        element_count,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_TABLE_DEFINITION_XML_ELEMENT_COUNT",
        element_count,
    )

    snapshot = load_snapshot(workbook)

    assert set(snapshot.tables) == {"Sales"}


def test_semantic_reader_preflight_rejects_default_table_definition_xml_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_table_model(tmp_path / "default-table-definition-limit.xlsx")
    table_member = "xl/tables/table1.xml"
    existing_count = _table_definition_xml_element_count(workbook, table_member)
    _append_table_definition_xml_elements(
        workbook,
        (
            workbook_module._OOXML_READER_MAX_TABLE_DEFINITION_PART_XML_ELEMENT_COUNT
            - existing_count
        ),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "table-definition XML structure" in message


def test_semantic_reader_preflight_preserves_malformed_orphan_table_coverage(
    tmp_path: Path,
) -> None:
    workbook = make_model(tmp_path / "malformed-orphan-table.xml.xlsx")
    _append_member(
        workbook,
        "xl/tables/orphan.xml",
        b'<table xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"
    assert any("Excel Table Style metadata" in warning for warning in snapshot.parser_warnings)


@pytest.mark.parametrize("nested", (False, True))
def test_rich_text_shared_string_scan_streams_opaque_root_children(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    nested: bool,
) -> None:
    workbook = make_rich_text_run_model(
        tmp_path / f"streamed-shared-string-root-{nested}.xlsx"
    )
    shared_member = "xl/sharedStrings.xml"
    _append_shared_string_root_xml_elements(workbook, 1, nested=nested)
    with ZipFile(workbook) as archive:
        shared_xml = archive.read(shared_member)
    original_tree_parse = workbook_module._xml_root_from_payload

    def unexpected_shared_string_tree(payload: bytes) -> ElementTree.Element:
        if payload == shared_xml:
            raise AssertionError("the shared-string XML table was materialized")
        return original_tree_parse(payload)

    monkeypatch.setattr(
        workbook_module,
        "_xml_root_from_payload",
        unexpected_shared_string_tree,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.rich_text_runs.shared_rich_text_run_count == 2


def test_semantic_reader_preflight_rejects_opaque_shared_string_root_xml_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_rich_text_run_model(tmp_path / "opaque-shared-string-root.xlsx")
    _append_shared_string_root_xml_elements(workbook, 1, nested=True)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_SHARED_STRING_OPAQUE_PART_XML_ELEMENT_COUNT",
        2,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "shared-string opaque XML structure" in message


@pytest.mark.parametrize("nested", (False, True))
def test_semantic_reader_preflight_rejects_opaque_worksheet_root_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    nested: bool,
) -> None:
    workbook = make_model(tmp_path / f"opaque-worksheet-root-{nested}.xlsx")
    element_count = _append_worksheet_opaque_root_xml_elements(
        workbook,
        1,
        nested=nested,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_OPAQUE_ROOT_PART_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet opaque root XML structure" in message


def test_semantic_reader_preflight_aggregates_opaque_worksheet_root_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-opaque-worksheet-root.xlsx")
    first_count = _append_worksheet_opaque_root_xml_elements(workbook, 1)
    second_count = _append_worksheet_opaque_root_xml_elements(
        workbook,
        1,
        member_name="xl/worksheets/sheet2.xml",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_OPAQUE_ROOT_PART_XML_ELEMENT_COUNT",
        max(first_count, second_count),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        first_count + second_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "aggregate worksheet opaque root XML elements" in message


def test_semantic_reader_preflight_accepts_opaque_worksheet_root_xml_at_exact_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "opaque-worksheet-root-at-limits.xlsx")
    element_count = _append_worksheet_opaque_root_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_OPAQUE_ROOT_PART_XML_ELEMENT_COUNT",
        element_count,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        element_count,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_keeps_standard_worksheet_root_content_unmetered(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "standard-worksheet-root-content.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_OPAQUE_ROOT_PART_XML_ELEMENT_COUNT",
        0,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        0,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_PART_XML_ELEMENT_COUNT",
        0,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_XML_ELEMENT_COUNT",
        0,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_covers_strict_opaque_worksheet_root_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_strict_worksheet_print_layout_model(
        tmp_path / "strict-opaque-worksheet-root.xlsx"
    )
    element_count = _append_worksheet_opaque_root_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_OPAQUE_ROOT_PART_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet opaque root XML structure" in message


def test_semantic_reader_preflight_rejects_default_opaque_worksheet_root_xml_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-opaque-worksheet-root-limit.xlsx")
    _append_worksheet_opaque_root_xml_elements(
        workbook,
        workbook_module._OOXML_READER_MAX_WORKSHEET_OPAQUE_ROOT_PART_XML_ELEMENT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet opaque root XML structure" in message


@pytest.mark.parametrize("nested", (False, True))
def test_semantic_reader_preflight_rejects_worksheet_extension_list_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    nested: bool,
) -> None:
    workbook = make_model(tmp_path / f"worksheet-extension-list-{nested}.xlsx")
    element_count = _append_worksheet_extension_list_xml_elements(
        workbook,
        1,
        nested=nested,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_PART_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet extension-list XML structure" in message


def test_semantic_reader_preflight_covers_nested_worksheet_extension_lists(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "nested-worksheet-extension-list.xlsx")
    element_count = _append_worksheet_extension_list_xml_elements(
        workbook,
        1,
        within_sheet_properties=True,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_PART_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet extension-list XML structure" in message


def test_semantic_reader_preflight_aggregates_worksheet_extension_list_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-worksheet-extension-list.xlsx")
    first_count = _append_worksheet_extension_list_xml_elements(workbook, 1)
    second_count = _append_worksheet_extension_list_xml_elements(
        workbook,
        1,
        member_name="xl/worksheets/sheet2.xml",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_PART_XML_ELEMENT_COUNT",
        max(first_count, second_count),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_XML_ELEMENT_COUNT",
        first_count + second_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "aggregate worksheet extension-list XML elements" in message


def test_semantic_reader_preflight_accepts_worksheet_extension_list_at_exact_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "worksheet-extension-list-at-limits.xlsx")
    element_count = _append_worksheet_extension_list_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_PART_XML_ELEMENT_COUNT",
        element_count,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_covers_strict_worksheet_extension_list_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_strict_worksheet_print_layout_model(
        tmp_path / "strict-worksheet-extension-list.xlsx"
    )
    element_count = _append_worksheet_extension_list_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_PART_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet extension-list XML structure" in message


def test_semantic_reader_preflight_rejects_default_worksheet_extension_list_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-worksheet-extension-list-limit.xlsx")
    _append_worksheet_extension_list_xml_elements(
        workbook,
        workbook_module._OOXML_READER_MAX_WORKSHEET_EXTENSION_LIST_PART_XML_ELEMENT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "worksheet extension-list XML structure" in message


@pytest.mark.parametrize("nested", (False, True))
def test_semantic_reader_preflight_rejects_opaque_workbook_root_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    nested: bool,
) -> None:
    workbook = make_model(tmp_path / f"opaque-workbook-root-{nested}.xlsx")
    element_count = _append_workbook_opaque_root_xml_elements(
        workbook,
        1,
        nested=nested,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook opaque root XML structure" in message


def test_semantic_reader_preflight_covers_alternate_namespace_workbook_root_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-workbook-root-opaque.xml.xlsx")
    _replace_workbook_root_namespace(
        workbook,
        "urn:formulafence:archive-safety-alternate-root",
    )
    element_count = _append_workbook_opaque_root_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook opaque root XML structure" in message


@pytest.mark.parametrize("nested", (False, True))
def test_semantic_reader_preflight_rejects_workbook_extension_list_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    nested: bool,
) -> None:
    workbook = make_model(tmp_path / f"workbook-extension-list-{nested}.xlsx")
    element_count = _append_workbook_extension_list_xml_elements(
        workbook,
        1,
        nested=nested,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook extension-list XML structure" in message


def test_semantic_reader_preflight_covers_nested_workbook_view_extension_lists(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "nested-workbook-view-extension-list.xlsx")
    element_count = _append_workbook_extension_list_xml_elements(
        workbook,
        1,
        within_book_view=True,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook extension-list XML structure" in message


def test_semantic_reader_preflight_covers_alternate_namespace_workbook_extension_lists(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-workbook-view-extension-list.xlsx")
    element_count = _append_workbook_extension_list_xml_elements(
        workbook,
        1,
        within_book_view=True,
        extension_namespace="urn:formulafence:archive-safety-alternate",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook extension-list XML structure" in message


def test_semantic_reader_preflight_accepts_workbook_opaque_root_xml_at_exact_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "opaque-workbook-root-at-limit.xlsx")
    element_count = _append_workbook_opaque_root_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        element_count,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_accepts_workbook_extension_list_at_exact_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "workbook-extension-list-at-limit.xlsx")
    element_count = _append_workbook_extension_list_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        0,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_keeps_standard_workbook_root_content_unmetered(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "standard-workbook-root-content.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        0,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_EXTENSION_LIST_XML_ELEMENT_COUNT",
        0,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_covers_strict_workbook_extension_list_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_strict_worksheet_print_layout_model(
        tmp_path / "strict-workbook-extension-list.xlsx"
    )
    element_count = _append_workbook_extension_list_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook extension-list XML structure" in message


def test_semantic_reader_preflight_rejects_default_opaque_workbook_root_xml_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-opaque-workbook-root-limit.xlsx")
    _append_workbook_opaque_root_xml_elements(
        workbook,
        workbook_module._OOXML_READER_MAX_WORKBOOK_OPAQUE_ROOT_XML_ELEMENT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook opaque root XML structure" in message


def test_semantic_reader_preflight_rejects_default_workbook_extension_list_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-workbook-extension-list-limit.xlsx")
    _append_workbook_extension_list_xml_elements(
        workbook,
        workbook_module._OOXML_READER_MAX_WORKBOOK_EXTENSION_LIST_XML_ELEMENT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook extension-list XML structure" in message


@pytest.mark.parametrize("sheet_kind", ("chartsheet", "dialogsheet"))
@pytest.mark.parametrize("nested", (False, True))
def test_semantic_reader_preflight_rejects_non_grid_sheet_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    sheet_kind: str,
    nested: bool,
) -> None:
    workbook, member_name = _make_non_grid_sheet_model(
        tmp_path / f"{sheet_kind}-xml-{nested}.xlsx",
        sheet_kind,
    )
    _append_non_grid_sheet_extension_list_xml_elements(
        workbook,
        1,
        nested=nested,
        member_name=member_name,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_NON_GRID_SHEET_PART_XML_ELEMENT_COUNT",
        _non_grid_sheet_xml_element_count(workbook, member_name) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "non-grid sheet XML structure" in message


@pytest.mark.parametrize("sheet_kind", ("chartsheet", "dialogsheet"))
def test_semantic_reader_preflight_rejects_non_grid_sheet_opaque_root_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    sheet_kind: str,
) -> None:
    workbook, member_name = _make_non_grid_sheet_model(
        tmp_path / f"{sheet_kind}-opaque-root.xml.xlsx",
        sheet_kind,
    )
    _append_non_grid_sheet_opaque_root_xml_elements(
        workbook,
        1,
        member_name=member_name,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_NON_GRID_SHEET_PART_XML_ELEMENT_COUNT",
        _non_grid_sheet_xml_element_count(workbook, member_name) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "non-grid sheet XML structure" in message


def test_semantic_reader_preflight_aggregates_non_grid_sheet_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_protection_model(
        tmp_path / "aggregate-chartsheet-xml.xlsx",
        include_chartsheet=True,
        chartsheet_count=2,
    )
    first_member = "xl/chartsheets/sheet1.xml"
    _convert_chartsheet_to_dialogsheet(workbook, sheet_number=2)
    second_member = "xl/dialogsheets/sheet2.xml"
    _append_non_grid_sheet_extension_list_xml_elements(
        workbook,
        1,
        member_name=first_member,
    )
    _append_non_grid_sheet_extension_list_xml_elements(
        workbook,
        1,
        member_name=second_member,
    )
    first_count = _non_grid_sheet_xml_element_count(workbook, first_member)
    second_count = _non_grid_sheet_xml_element_count(workbook, second_member)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_NON_GRID_SHEET_PART_XML_ELEMENT_COUNT",
        max(first_count, second_count),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_NON_GRID_SHEET_XML_ELEMENT_COUNT",
        first_count + second_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "aggregate non-grid sheet XML elements" in message


@pytest.mark.parametrize("sheet_kind", ("chartsheet", "dialogsheet"))
def test_semantic_reader_preflight_accepts_non_grid_sheet_xml_at_exact_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    sheet_kind: str,
) -> None:
    workbook, member_name = _make_non_grid_sheet_model(
        tmp_path / f"{sheet_kind}-xml-at-limits.xlsx",
        sheet_kind,
    )
    _append_non_grid_sheet_extension_list_xml_elements(
        workbook,
        1,
        member_name=member_name,
    )
    element_count = _non_grid_sheet_xml_element_count(workbook, member_name)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_NON_GRID_SHEET_PART_XML_ELEMENT_COUNT",
        element_count,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_NON_GRID_SHEET_XML_ELEMENT_COUNT",
        element_count,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


@pytest.mark.parametrize("sheet_kind", ("chartsheet", "dialogsheet"))
def test_semantic_reader_preflight_rejects_default_non_grid_sheet_xml_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    sheet_kind: str,
) -> None:
    workbook, member_name = _make_non_grid_sheet_model(
        tmp_path / f"default-{sheet_kind}-xml-limit.xlsx",
        sheet_kind,
    )
    _append_non_grid_sheet_extension_list_xml_elements(
        workbook,
        workbook_module._OOXML_READER_MAX_NON_GRID_SHEET_PART_XML_ELEMENT_COUNT,
        member_name=member_name,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "non-grid sheet XML structure" in message


def test_semantic_reader_preflight_covers_raw_relationship_selected_shared_strings(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_rich_text_run_model(
        tmp_path / "relationship-selected-shared-string.xml.xlsx"
    )
    alternate_member = "xl/raw-rich-text-sharedStrings.xml"
    with ZipFile(workbook) as archive:
        canonical_shared_strings = archive.read("xl/sharedStrings.xml")
    _append_member(workbook, alternate_member, canonical_shared_strings)
    _redirect_shared_string_relationship(workbook, "raw-rich-text-sharedStrings.xml")
    _append_shared_string_root_xml_elements(
        workbook,
        1,
        nested=True,
        member_name=alternate_member,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_SHARED_STRING_OPAQUE_PART_XML_ELEMENT_COUNT",
        2,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "shared-string opaque XML structure" in message


def test_semantic_reader_preflight_rejects_shared_string_item_xml_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_rich_text_run_model(tmp_path / "opaque-shared-string-item.xlsx")
    shared_member = "xl/sharedStrings.xml"
    _append_shared_string_item_xml_elements(workbook, 1)
    with ZipFile(workbook) as archive:
        shared_xml = archive.read(shared_member)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_SHARED_STRING_ITEM_XML_ELEMENT_COUNT",
        _shared_string_item_xml_element_count(workbook, shared_member) - 1,
    )
    original_tree_parse = workbook_module._xml_root_from_payload

    def unexpected_shared_string_tree(payload: bytes) -> ElementTree.Element:
        if payload == shared_xml:
            raise AssertionError("the over-budget shared-string item was materialized")
        return original_tree_parse(payload)

    monkeypatch.setattr(
        workbook_module,
        "_xml_root_from_payload",
        unexpected_shared_string_tree,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "shared-string item XML structure" in message


def test_semantic_reader_preflight_counts_nested_shared_string_item_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_rich_text_run_model(tmp_path / "nested-shared-string-item.xlsx")
    shared_member = "xl/sharedStrings.xml"
    _append_shared_string_item_xml_elements(workbook, 1, nested=True)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_SHARED_STRING_ITEM_XML_ELEMENT_COUNT",
        _shared_string_item_xml_element_count(workbook, shared_member) - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "shared-string item XML structure" in message


def test_semantic_reader_preflight_aggregates_complex_shared_string_xml_items(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_rich_text_run_model(tmp_path / "aggregate-shared-string-items.xlsx")
    shared_member = "xl/sharedStrings.xml"
    _duplicate_first_shared_string_item(workbook, shared_member)
    item_element_count = _shared_string_item_xml_element_count(workbook, shared_member)
    complex_element_count = _complex_shared_string_xml_element_count(
        workbook,
        shared_member,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_SHARED_STRING_ITEM_XML_ELEMENT_COUNT",
        item_element_count,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_COMPLEX_SHARED_STRING_XML_ELEMENT_COUNT",
        complex_element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "aggregate complex shared-string XML elements" in message


def test_semantic_reader_preflight_accepts_shared_string_xml_at_exact_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_rich_text_run_model(tmp_path / "shared-string-at-limits.xlsx")
    shared_member = "xl/sharedStrings.xml"
    item_element_count = _shared_string_item_xml_element_count(workbook, shared_member)
    complex_element_count = _complex_shared_string_xml_element_count(
        workbook,
        shared_member,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_SHARED_STRING_ITEM_XML_ELEMENT_COUNT",
        item_element_count,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_COMPLEX_SHARED_STRING_XML_ELEMENT_COUNT",
        complex_element_count,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.rich_text_runs.shared_rich_text_run_count == 2


def test_semantic_reader_preflight_rejects_default_shared_string_item_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_rich_text_run_model(tmp_path / "default-shared-string-item-limit.xlsx")
    shared_member = "xl/sharedStrings.xml"
    existing_count = _shared_string_item_xml_element_count(workbook, shared_member)
    _append_shared_string_item_xml_elements(
        workbook,
        (
            workbook_module._OOXML_READER_MAX_SHARED_STRING_ITEM_XML_ELEMENT_COUNT
            - existing_count
            + 1
        ),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "shared-string item XML structure" in message


def test_semantic_reader_preflight_rejects_excessive_workbook_sheet_declarations_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-workbook-sheets.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_SHEET_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook sheet declarations" in message


def test_semantic_reader_preflight_counts_alternate_namespace_workbook_sheet_declarations(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-namespace-workbook-sheets.xlsx")
    _append_workbook_sheet_declarations(
        workbook,
        1,
        namespace_prefix="audit",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_SHEET_COUNT",
        4,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook sheet declarations" in message


def test_semantic_reader_preflight_rejects_default_sheet_declaration_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-workbook-sheets.xlsx")
    _append_workbook_sheet_declarations(
        workbook,
        workbook_module._OOXML_READER_MAX_WORKBOOK_SHEET_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook sheet declarations" in message


def test_semantic_reader_preflight_rejects_excessive_workbook_defined_names_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-workbook-defined-names.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_DEFINED_NAME_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook defined-name declarations" in message


def test_semantic_reader_preflight_counts_alternate_namespace_workbook_defined_names(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-namespace-workbook-defined-names.xlsx")
    _append_workbook_defined_name_declarations(
        workbook,
        1,
        namespace_prefix="audit",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_DEFINED_NAME_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook defined-name declarations" in message


def test_semantic_reader_preflight_rejects_default_defined_name_declaration_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-workbook-defined-names.xlsx")
    _append_workbook_defined_name_declarations(
        workbook,
        workbook_module._OOXML_READER_MAX_WORKBOOK_DEFINED_NAME_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook defined-name declarations" in message


def test_semantic_reader_preflight_rejects_excessive_workbook_external_references_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_external_link_package_model(
        tmp_path / "too-many-workbook-external-references.xlsx"
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_EXTERNAL_REFERENCE_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook external-reference declarations" in message


def test_semantic_reader_preflight_counts_alternate_namespace_workbook_external_references(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_external_link_package_model(
        tmp_path / "alternate-namespace-workbook-external-references.xlsx"
    )
    _append_workbook_relationship_catalog_declarations(
        workbook,
        "externalReferences",
        1,
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_EXTERNAL_REFERENCE_COUNT",
        3,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook external-reference declarations" in message


def test_semantic_reader_preflight_rejects_default_external_reference_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_external_link_package_model(
        tmp_path / "repeated-workbook-external-references.xlsx"
    )
    _append_workbook_relationship_catalog_declarations(
        workbook,
        "externalReferences",
        workbook_module._OOXML_READER_MAX_WORKBOOK_EXTERNAL_REFERENCE_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook external-reference declarations" in message


def test_semantic_reader_preflight_rejects_excessive_workbook_pivot_caches_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_external_data_refresh_model(
        tmp_path / "too-many-workbook-pivot-caches.xlsx"
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_PIVOT_CACHE_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook pivot-cache declarations" in message


def test_semantic_reader_preflight_counts_alternate_namespace_workbook_pivot_caches(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_external_data_refresh_model(
        tmp_path / "alternate-namespace-workbook-pivot-caches.xlsx"
    )
    _append_workbook_relationship_catalog_declarations(
        workbook,
        "pivotCaches",
        1,
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_WORKBOOK_PIVOT_CACHE_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook pivot-cache declarations" in message


def test_semantic_reader_preflight_rejects_default_pivot_cache_declaration_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_external_data_refresh_model(
        tmp_path / "repeated-workbook-pivot-caches.xlsx"
    )
    _append_workbook_relationship_catalog_declarations(
        workbook,
        "pivotCaches",
        workbook_module._OOXML_READER_MAX_WORKBOOK_PIVOT_CACHE_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook pivot-cache declarations" in message


@pytest.mark.parametrize(
    "catalog_name, child_name, attributes, limit_name, expected_message",
    _WORKBOOK_AUXILIARY_CATALOG_CASES,
)
def test_semantic_reader_preflight_rejects_excessive_auxiliary_workbook_catalogs_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    catalog_name: str,
    child_name: str,
    attributes: dict[str, str],
    limit_name: str,
    expected_message: str,
) -> None:
    workbook = make_model(tmp_path / f"too-many-{catalog_name}.xlsx")
    _append_workbook_package_catalog_declarations(
        workbook,
        catalog_name,
        child_name,
        1,
        attributes=attributes,
    )
    monkeypatch.setattr(workbook_module, limit_name, 0)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert expected_message in message


@pytest.mark.parametrize(
    "catalog_name, child_name, attributes, limit_name, expected_message",
    _WORKBOOK_AUXILIARY_CATALOG_CASES,
)
def test_semantic_reader_preflight_counts_alternate_namespace_auxiliary_workbook_catalogs(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    catalog_name: str,
    child_name: str,
    attributes: dict[str, str],
    limit_name: str,
    expected_message: str,
) -> None:
    workbook = make_model(tmp_path / f"alternate-namespace-{catalog_name}.xlsx")
    if catalog_name != "bookViews":
        _append_workbook_package_catalog_declarations(
            workbook,
            catalog_name,
            child_name,
            1,
            attributes=attributes,
        )
    _append_workbook_package_catalog_declarations(
        workbook,
        catalog_name,
        child_name,
        1,
        attributes=attributes,
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(workbook_module, limit_name, 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert expected_message in message


@pytest.mark.parametrize(
    "catalog_name, child_name, attributes, limit_name, expected_message",
    _WORKBOOK_AUXILIARY_CATALOG_CASES,
)
def test_semantic_reader_preflight_rejects_default_auxiliary_workbook_catalogs_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    catalog_name: str,
    child_name: str,
    attributes: dict[str, str],
    limit_name: str,
    expected_message: str,
) -> None:
    workbook = make_model(tmp_path / f"repeated-{catalog_name}.xlsx")
    _append_workbook_package_catalog_declarations(
        workbook,
        catalog_name,
        child_name,
        getattr(workbook_module, limit_name) + 1,
        attributes=attributes,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert expected_message in message


def test_semantic_reader_preflight_rejects_excessive_custom_sheet_views_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-custom-sheet-views.xlsx")
    _append_custom_sheet_view_declarations(workbook, 1)
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_CUSTOM_SHEET_VIEW_COUNT", 0)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "custom sheet-view declarations" in message


def test_semantic_reader_preflight_counts_alternate_namespace_custom_sheet_views(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-namespace-custom-sheet-views.xlsx")
    _append_custom_sheet_view_declarations(workbook, 1)
    _append_custom_sheet_view_declarations(
        workbook,
        1,
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_CUSTOM_SHEET_VIEW_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "custom sheet-view declarations" in message


def test_semantic_reader_preflight_rejects_default_custom_sheet_view_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-custom-sheet-views.xlsx")
    _append_custom_sheet_view_declarations(
        workbook,
        workbook_module._OOXML_READER_MAX_CUSTOM_SHEET_VIEW_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "custom sheet-view declarations" in message


def test_semantic_reader_preflight_rejects_custom_sheet_view_descendants_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = _make_custom_sheet_view_subtree_model(
        tmp_path / "too-many-custom-sheet-view-descendants.xlsx"
    )
    _append_custom_sheet_view_opaque_descendants(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CUSTOM_SHEET_VIEW_DESCENDANT_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "custom sheet-view descendants" in message


def test_semantic_reader_preflight_rejects_default_custom_view_descendant_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = _make_custom_sheet_view_subtree_model(
        tmp_path / "default-custom-sheet-view-descendant-limit.xlsx"
    )
    _append_custom_sheet_view_opaque_descendants(
        workbook,
        workbook_module._OOXML_READER_MAX_CUSTOM_SHEET_VIEW_DESCENDANT_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "custom sheet-view descendants" in message


def test_semantic_reader_preflight_accepts_default_custom_view_descendant_capacity(
    tmp_path: Path,
) -> None:
    workbook = _make_custom_sheet_view_subtree_model(
        tmp_path / "custom-sheet-view-descendants-at-default-limit.xlsx"
    )
    _append_custom_sheet_view_opaque_descendants(
        workbook,
        workbook_module._OOXML_READER_MAX_CUSTOM_SHEET_VIEW_DESCENDANT_COUNT,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_aggregates_custom_sheet_view_descendants_across_sheets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = _make_custom_sheet_view_subtree_model(
        tmp_path / "aggregate-custom-sheet-view-descendants.xlsx"
    )
    _append_custom_sheet_view_declarations(
        workbook,
        1,
        member_name="xl/worksheets/sheet2.xml",
    )
    _append_custom_sheet_view_opaque_descendants(workbook, 1)
    _append_custom_sheet_view_opaque_descendants(
        workbook,
        1,
        member_name="xl/worksheets/sheet2.xml",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CUSTOM_SHEET_VIEW_DESCENDANT_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "custom sheet-view descendants" in message


def test_semantic_reader_preflight_accepts_custom_sheet_view_descendants_at_the_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = _make_custom_sheet_view_subtree_model(
        tmp_path / "custom-sheet-view-descendants-at-limit.xlsx"
    )
    _append_custom_sheet_view_opaque_descendants(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CUSTOM_SHEET_VIEW_DESCENDANT_COUNT",
        2,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_accepts_published_custom_view_page_break_allowance(
    tmp_path: Path,
) -> None:
    workbook = _make_custom_sheet_view_subtree_model(
        tmp_path / "custom-view-page-breaks-at-published-limit.xlsx"
    )
    _append_custom_sheet_view_page_break_declarations(workbook, 1_026)
    _append_custom_sheet_view_page_break_declarations(workbook, 1_026, axis="column")

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_counts_nested_custom_sheet_view_descendants(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = _make_custom_sheet_view_subtree_model(
        tmp_path / "nested-custom-sheet-view-descendants.xlsx"
    )
    _append_custom_sheet_view_opaque_descendants(workbook, 1, nested=True)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CUSTOM_SHEET_VIEW_DESCENDANT_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "custom sheet-view descendants" in message


def test_semantic_reader_preflight_counts_opaque_custom_sheet_view_descendants(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = _make_custom_sheet_view_subtree_model(
        tmp_path / "opaque-custom-sheet-view-descendants.xlsx"
    )
    _append_custom_sheet_view_opaque_descendants(
        workbook,
        1,
        custom_sheet_view_namespace="urn:formulafence:opaque-custom-sheet-view",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CUSTOM_SHEET_VIEW_DESCENDANT_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "custom sheet-view descendants" in message


def test_semantic_reader_preflight_ignores_foreign_custom_sheet_view_containers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "foreign-custom-sheet-view-container.xlsx")
    _append_custom_sheet_view_declarations(
        workbook,
        1,
        container_namespace="urn:formulafence:foreign-custom-sheet-views",
    )
    _append_custom_sheet_view_opaque_descendants(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CUSTOM_SHEET_VIEW_DESCENDANT_COUNT",
        0,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_counts_strict_custom_sheet_view_descendants(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_strict_custom_workbook_view_model(
        tmp_path / "strict-custom-sheet-view-descendants.xlsx"
    )
    _append_custom_sheet_view_opaque_descendants(
        workbook,
        workbook_module._OOXML_READER_MAX_CUSTOM_SHEET_VIEW_DESCENDANT_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "custom sheet-view descendants" in message


def test_semantic_reader_preflight_rejects_excessive_merged_cell_declarations_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-merged-cell-declarations.xlsx")
    _append_merged_cell_declarations(workbook, ("A1:B1",))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MERGED_CELL_RANGE_COUNT", 0)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "merged-cell declarations" in message


def test_semantic_reader_preflight_counts_alternate_namespace_merged_cell_declarations(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-namespace-merged-cell-declarations.xlsx")
    _append_merged_cell_declarations(workbook, ("A1:B1",))
    _append_merged_cell_declarations(
        workbook,
        ("C1:D1",),
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MERGED_CELL_RANGE_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "merged-cell declarations" in message


def test_semantic_reader_preflight_rejects_default_merged_cell_declaration_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-merged-cell-declarations.xlsx")
    _append_merged_cell_declarations(
        workbook,
        ("A1",) * (workbook_module._OOXML_READER_MAX_MERGED_CELL_RANGE_COUNT + 1),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "merged-cell declarations" in message


def test_semantic_reader_preflight_rejects_an_oversized_merged_cell_range_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-merged-cell-range.xlsx")
    _append_merged_cell_declarations(workbook, ("A1:K10",))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MERGED_CELL_COUNT", 100)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a merged-cell range" in message


def test_semantic_reader_preflight_rejects_excessive_merged_cell_area_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "excessive-merged-cell-area.xlsx")
    _append_merged_cell_declarations(workbook, ("A1:J10", "K1:T10"))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MERGED_CELL_COUNT", 150)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "merged-cell area" in message


def test_semantic_reader_preflight_rejects_a_full_worksheet_merge_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "full-worksheet-merge.xlsx")
    _append_merged_cell_declarations(workbook, ("A1:XFD1048576",))

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a merged-cell range" in message


def test_semantic_reader_preflight_rejects_an_oversized_merged_cell_reference_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-merged-cell-reference.xlsx")
    _append_merged_cell_declarations(workbook, ("A1:B1",))
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_MERGED_CELL_REFERENCE_CHARACTERS",
        4,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a merged-cell reference" in message


@pytest.mark.parametrize("reference", ("!", "A1:A0"))
def test_semantic_reader_preflight_rejects_an_unmeasurable_merged_cell_reference_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    reference: str,
) -> None:
    workbook = make_model(tmp_path / "unmeasurable-merged-cell-reference.xlsx")
    _append_merged_cell_declarations(workbook, (reference,))

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "could not be measured safely" in message


def test_semantic_reader_preflight_accepts_merged_cell_area_at_the_configured_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "merged-cell-area-at-limit.xlsx")
    _append_merged_cell_declarations(workbook, ("Inputs!A1:J10",))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MERGED_CELL_COUNT", 100)

    snapshot = load_snapshot(workbook)

    assert snapshot.sheets


def test_semantic_reader_preflight_rejects_excessive_data_validation_declarations_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-data-validations.xlsx")
    _append_data_validation_declarations(workbook, ("A1",))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_DATA_VALIDATION_COUNT", 0)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "data-validation declarations" in message


def test_semantic_reader_preflight_counts_alternate_namespace_data_validation_declarations(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-namespace-data-validations.xlsx")
    _append_data_validation_declarations(workbook, ("A1",))
    _append_data_validation_declarations(
        workbook,
        ("B1",),
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_DATA_VALIDATION_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "data-validation declarations" in message


def test_semantic_reader_preflight_rejects_default_data_validation_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-data-validations.xlsx")
    _append_data_validation_declarations(
        workbook,
        ("A1",) * (workbook_module._OOXML_READER_MAX_DATA_VALIDATION_COUNT + 1),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "data-validation declarations" in message


def test_semantic_reader_preflight_rejects_default_data_validation_reference_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-data-validation-reference-ranges.xlsx")
    _append_data_validation_declarations(
        workbook,
        (
            " ".join(
                ("A1",)
                * (workbook_module._OOXML_READER_MAX_MULTI_RANGE_REFERENCE_COUNT + 1)
            ),
        ),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a data-validation target reference" in message


def test_preflight_rejects_data_validation_reference_range_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "data-validation-reference-ranges.xlsx")
    _append_data_validation_declarations(workbook, ("A1 B1",))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MULTI_RANGE_REFERENCE_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a data-validation target reference" in message


def test_preflight_rejects_aggregate_data_validation_reference_ranges(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-data-validation-reference-ranges.xlsx")
    _append_data_validation_declarations(workbook, ("A1", "B1"))
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_DATA_VALIDATION_TARGET_RANGE_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "data-validation target ranges" in message


def test_semantic_reader_preflight_rejects_an_oversized_data_validation_reference_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-data-validation-reference.xlsx")
    _append_data_validation_declarations(workbook, ("A1:B1",))
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_MULTI_RANGE_REFERENCE_CHARACTERS",
        4,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a data-validation target reference" in message


def test_semantic_reader_preflight_rejects_an_oversized_data_validation_formula_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-data-validation-formula.xlsx")
    _append_data_validation_declarations(workbook, ("A1",), formula="TRUE")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_FORMULA_CHARACTERS", 3)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "formula text" in message


def test_semantic_reader_preflight_accepts_data_validation_reference_ranges_at_configured_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "data-validation-reference-ranges-at-limit.xlsx")
    _append_data_validation_declarations(workbook, ("A1 B1", "C1 D1"))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MULTI_RANGE_REFERENCE_COUNT", 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_DATA_VALIDATION_TARGET_RANGE_COUNT",
        4,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.summary()["data_validation_target_ranges"] == 4


def test_preflight_rejects_conditional_formatting_declaration_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-conditional-formatting-declarations.xlsx")
    _append_conditional_formatting_declarations(
        workbook,
        ("A1",),
        rule_namespaces=(),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CONDITIONAL_FORMATTING_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "conditional-formatting declarations" in message


def test_preflight_rejects_default_conditional_formatting_declaration_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-conditional-formatting-declarations.xlsx")
    _append_conditional_formatting_declarations(
        workbook,
        ("A1",) * (workbook_module._OOXML_READER_MAX_CONDITIONAL_FORMATTING_COUNT + 1),
        rule_namespaces=(),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "conditional-formatting declarations" in message


def test_semantic_reader_preflight_counts_alternate_namespace_conditional_formatting_rules(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-namespace-conditional-formatting-rules.xlsx")
    _append_conditional_formatting_declarations(
        workbook,
        ("A1",),
        rule_namespaces=(None, "urn:formulafence:archive-safety"),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CONDITIONAL_FORMATTING_RULE_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "conditional-formatting rules" in message


def test_preflight_rejects_default_conditional_formatting_rule_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-conditional-formatting-rules.xlsx")
    _append_conditional_formatting_declarations(
        workbook,
        ("A1",),
        rule_namespaces=(None,)
        * (workbook_module._OOXML_READER_MAX_CONDITIONAL_FORMATTING_RULE_COUNT + 1),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "conditional-formatting rules" in message


def test_preflight_rejects_conditional_formatting_reference_range_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "conditional-formatting-reference-ranges.xlsx")
    _append_conditional_formatting_declarations(workbook, ("A1 B1",))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MULTI_RANGE_REFERENCE_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a conditional-formatting target reference" in message


def test_preflight_rejects_aggregate_conditional_formatting_reference_ranges(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-conditional-formatting-reference-ranges.xlsx")
    _append_conditional_formatting_declarations(workbook, ("A1", "B1"))
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CONDITIONAL_FORMATTING_TARGET_RANGE_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "conditional-formatting target ranges" in message


def test_preflight_rejects_oversized_conditional_formatting_reference(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-conditional-formatting-reference.xlsx")
    _append_conditional_formatting_declarations(workbook, ("A1:B1",))
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_MULTI_RANGE_REFERENCE_CHARACTERS",
        4,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a conditional-formatting target reference" in message


def test_preflight_rejects_oversized_conditional_formatting_formula(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-conditional-formatting-formula.xlsx")
    _append_conditional_formatting_declarations(workbook, ("A1",), formula="TRUE")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_FORMULA_CHARACTERS", 3)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "formula text" in message


def test_preflight_accepts_conditional_formatting_reference_ranges_at_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "conditional-formatting-reference-ranges-at-limit.xlsx")
    _append_conditional_formatting_declarations(workbook, ("A1 B1", "C1 D1"))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MULTI_RANGE_REFERENCE_COUNT", 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_CONDITIONAL_FORMATTING_TARGET_RANGE_COUNT",
        4,
    )

    snapshot = load_snapshot(workbook)

    assert len(snapshot.conditional_formatting) == 2


def test_semantic_reader_preflight_rejects_excessive_scenario_manager_containers_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-scenario-manager-containers.xlsx")
    _append_scenario_manager_containers(workbook, ("A1",), scenario_count=0)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_SCENARIO_CONTAINER_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "Scenario Manager containers" in message


def test_semantic_reader_preflight_rejects_default_scenario_manager_container_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-scenario-manager-containers.xlsx")
    _append_scenario_manager_containers(
        workbook,
        ("A1",) * (workbook_module._OOXML_READER_MAX_SCENARIO_CONTAINER_COUNT + 1),
        scenario_count=0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "Scenario Manager containers" in message


def test_semantic_reader_preflight_counts_alternate_namespace_scenario_manager_declarations(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-namespace-scenario-manager-declarations.xlsx")
    _append_scenario_manager_containers(workbook, ("A1",))
    _append_scenario_manager_containers(
        workbook,
        ("B1",),
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_SCENARIO_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "Scenario Manager declarations" in message


def test_semantic_reader_preflight_counts_alternate_namespace_scenario_manager_input_cells(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-namespace-scenario-manager-input-cells.xlsx")
    _append_scenario_manager_containers(workbook, ("A1",))
    _append_scenario_manager_containers(
        workbook,
        ("B1",),
        alternate_namespace="urn:formulafence:archive-safety",
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_SCENARIO_INPUT_CELL_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "Scenario Manager input cells" in message


def test_semantic_reader_preflight_rejects_default_scenario_manager_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-scenario-manager-declarations.xlsx")
    _append_scenario_manager_containers(
        workbook,
        ("A1",),
        scenario_count=workbook_module._OOXML_READER_MAX_SCENARIO_COUNT + 1,
        input_cell_count=0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "Scenario Manager declarations" in message


def test_preflight_rejects_default_scenario_manager_input_cell_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-scenario-manager-input-cells.xlsx")
    _append_scenario_manager_containers(
        workbook,
        ("A1",),
        input_cell_count=workbook_module._OOXML_READER_MAX_SCENARIO_INPUT_CELL_COUNT + 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "Scenario Manager input cells" in message


def test_preflight_rejects_scenario_manager_reference_range_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "scenario-manager-reference-ranges.xlsx")
    _append_scenario_manager_containers(workbook, ("A1 B1",), scenario_count=0)
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MULTI_RANGE_REFERENCE_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a Scenario Manager target reference" in message


def test_preflight_rejects_aggregate_scenario_manager_reference_ranges(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-scenario-manager-reference-ranges.xlsx")
    _append_scenario_manager_containers(workbook, ("A1", "B1"), scenario_count=0)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_SCENARIO_TARGET_RANGE_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "Scenario Manager target ranges" in message


def test_semantic_reader_preflight_rejects_an_oversized_scenario_manager_reference_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-scenario-manager-reference.xlsx")
    _append_scenario_manager_containers(workbook, ("A1:B1",), scenario_count=0)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_MULTI_RANGE_REFERENCE_CHARACTERS",
        4,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "a Scenario Manager target reference" in message


def test_semantic_reader_preflight_accepts_scenario_manager_reference_ranges_at_configured_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "scenario-manager-reference-ranges-at-limit.xlsx")
    _append_scenario_manager_containers(workbook, ("A1 B1",))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_MULTI_RANGE_REFERENCE_COUNT", 2)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_SCENARIO_TARGET_RANGE_COUNT",
        2,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.scenario_manager.summary_reference_count == 2


def test_semantic_reader_preflight_rejects_excessive_xml_nesting_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "deep-reader-xml.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_XML_NESTING_DEPTH", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "XML nesting" in message


def test_semantic_reader_preflight_rejects_excessive_xml_elements_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "many-reader-elements.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_XML_ELEMENT_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "XML element count" in message


def test_xml_character_data_budget_accepts_its_exact_boundary_per_text_node(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_CHARACTER_DATA_CHARACTERS",
        5,
    )
    payload = b"<root><![CDATA[alpha]]><child/>bravo</root>"

    assert (
        workbook_module._xml_payload_structure_element_count_within_budget(
            payload,
            maximum_element_count=2,
        )
        == 2
    )
    assert workbook_module._xml_root_from_payload(payload).tag == "root"


@pytest.mark.parametrize("text", (b"abcdef", b"<![CDATA[abcdef]]>"))
def test_xml_character_data_budget_rejects_text_and_cdata_overages(
    text: bytes,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_CHARACTER_DATA_CHARACTERS",
        5,
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_XML_LEXICAL_CHUNK_BYTES", 1)
    payload = b"<root>" + text + b"</root>"

    assert (
        workbook_module._xml_payload_structure_element_count_within_budget(
            payload,
            maximum_element_count=1,
        )
        is None
    )
    with pytest.raises(ValueError, match="XML character data"):
        workbook_module._xml_root_from_payload(payload)


def test_semantic_reader_preflight_rejects_oversized_stylesheet_text_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "oversized-stylesheet-text.xlsx")
    _append_stylesheet_opaque_root_text(workbook, 129)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_CHARACTER_DATA_CHARACTERS",
        128,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "XML character data" in message


def test_stream_reader_reports_an_oversized_text_node_as_a_safety_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "stream-oversized-stylesheet-text.xlsx")
    _append_stylesheet_opaque_root_text(workbook, 129)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_CHARACTER_DATA_CHARACTERS",
        128,
    )

    with ZipFile(workbook) as archive:
        with pytest.raises(WorkbookLoadError, match="XML character data"):
            workbook_module._stream_ooxml_reader_xml(archive, "xl/styles.xml")


def test_semantic_reader_preflight_accepts_stylesheet_text_at_configured_boundary(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "stylesheet-text-at-limit.xlsx")
    _append_stylesheet_opaque_root_text(workbook, 128)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_CHARACTER_DATA_CHARACTERS",
        128,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.sheets


@pytest.mark.parametrize(
    ("prefix", "markup", "suffix", "maximum_markup_bytes"),
    (
        (b"<r>", b"<!--FormulaFence-->", b"</r>", len(b"<!--FormulaFence-->")),
        (b"<r>", b"<?ff FormulaFence?>", b"</r>", len(b"<?ff FormulaFence?>")),
        (b"<r>", b"<![CDATA[FormulaFence]]>", b"</r>", len(b"<![CDATA[")),
        (
            b"<FormulaFence>",
            b"</FormulaFence>",
            b"",
            len(b"</FormulaFence>"),
        ),
        (b"<r>", b"&FormulaFence;", b"</r>", len(b"&FormulaFence;")),
        (
            b"",
            b'<!DOCTYPE r SYSTEM "FormulaFence">',
            b"<r/>",
            len(b'<!DOCTYPE r SYSTEM "FormulaFence">'),
        ),
    ),
)
def test_xml_lexical_markup_budget_accepts_exact_boundaries(
    prefix: bytes,
    markup: bytes,
    suffix: bytes,
    maximum_markup_bytes: int,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(workbook_module, "_OOXML_READER_XML_LEXICAL_CHUNK_BYTES", 1)
    payload = prefix + markup + suffix

    assert workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_markup_bytes=maximum_markup_bytes,
    )
    assert not workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_markup_bytes=maximum_markup_bytes - 1,
    )


def test_xml_lexical_markup_budget_keeps_a_start_tag_violation_distinct() -> None:
    payload = b"<root/>"

    assert (
        workbook_module._xml_lexical_markup_budget_violation(
            io.BytesIO(payload),
            maximum_start_tag_bytes=len(payload) - 1,
            maximum_markup_bytes=len(payload) - 1,
        )
        == "start-tag"
    )


@pytest.mark.parametrize(
    ("encoding", "bom"),
    (
        ("utf-16-le", b"\xff\xfe"),
        ("utf-16-be", b"\xfe\xff"),
        ("utf-32-le", b"\xff\xfe\x00\x00"),
        ("utf-32-be", b"\x00\x00\xfe\xff"),
        ("utf-16-le", b""),
        ("utf-16-be", b""),
        ("utf-32-le", b""),
        ("utf-32-be", b""),
    ),
)
@pytest.mark.parametrize("markup", ("<!--FormulaFence-->", "<?ff FormulaFence?>"))
def test_xml_lexical_markup_budget_covers_fixed_width_xml_encodings(
    encoding: str,
    bom: bytes,
    markup: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = bom + f"<root>{markup}</root>".encode(encoding)
    maximum_markup_bytes = len(markup.encode(encoding))
    monkeypatch.setattr(workbook_module, "_OOXML_READER_XML_LEXICAL_CHUNK_BYTES", 1)

    assert workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_markup_bytes=maximum_markup_bytes,
    )
    assert not workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_markup_bytes=maximum_markup_bytes - 1,
    )


@pytest.mark.parametrize(
    ("label", "markup"),
    (
        ("comment", b"<!--" + (b"x" * 129) + b"-->"),
        ("processing-instruction", b"<?ff " + (b"x" * 129) + b"?>"),
        (
            "entity-reference",
            b'<ff:opaque xmlns:ff="urn:formulafence:archive-safety">&#'
            + (b"0" * 129)
            + b"65;</ff:opaque>",
        ),
    ),
)
def test_semantic_reader_preflight_rejects_oversized_ignored_markup_before_readers(
    label: str,
    markup: bytes,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / f"oversized-stylesheet-markup-{label}.xlsx")
    _append_stylesheet_opaque_root_markup(workbook, markup)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_NON_CHARACTER_DATA_MARKUP_BYTES",
        128,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "XML non-character-data markup" in message


@pytest.mark.parametrize(
    ("label", "markup"),
    (
        ("comment", b"<!--" + (b"x" * 129) + b"-->"),
        ("processing-instruction", b"<?ff " + (b"x" * 129) + b"?>"),
        (
            "entity-reference",
            b'<ff:opaque xmlns:ff="urn:formulafence:archive-safety">&#'
            + (b"0" * 129)
            + b"65;</ff:opaque>",
        ),
    ),
)
def test_stream_reader_rejects_ignored_markup_before_elementtree(
    label: str,
    markup: bytes,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / f"stream-stylesheet-markup-{label}.xlsx")
    _append_stylesheet_opaque_root_markup(workbook, markup)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_NON_CHARACTER_DATA_MARKUP_BYTES",
        128,
    )

    def unexpected_iterparse(*args, **kwargs):
        raise AssertionError("ElementTree started before the lexical markup budget")

    monkeypatch.setattr(workbook_module.ElementTree, "iterparse", unexpected_iterparse)
    with ZipFile(workbook) as archive:
        with pytest.raises(WorkbookLoadError, match="XML non-character-data markup"):
            workbook_module._stream_ooxml_reader_xml(archive, "xl/styles.xml")


def test_stream_reader_rejects_an_oversized_document_type_before_elementtree(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "stream-stylesheet-doctype-before-elementtree.xlsx")
    with ZipFile(workbook) as archive:
        stylesheet = archive.read("xl/styles.xml")
    document_type = b"<!DOCTYPE styleSheet [<!--" + (b"x" * 129) + b"-->]>"
    _replace_member(workbook, "xl/styles.xml", document_type + stylesheet)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_NON_CHARACTER_DATA_MARKUP_BYTES",
        128,
    )

    def unexpected_iterparse(*args, **kwargs):
        raise AssertionError("ElementTree started before the lexical markup budget")

    monkeypatch.setattr(workbook_module.ElementTree, "iterparse", unexpected_iterparse)
    with ZipFile(workbook) as archive:
        with pytest.raises(WorkbookLoadError, match="XML non-character-data markup"):
            workbook_module._stream_ooxml_reader_xml(archive, "xl/styles.xml")


@pytest.mark.parametrize(
    ("opening", "closing"),
    (
        (b"<!--", b"-->"),
        (b"<?ff ", b"?>"),
    ),
)
def test_semantic_reader_preflight_accepts_ignored_markup_at_configured_boundary(
    opening: bytes,
    closing: bytes,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / f"stylesheet-markup-at-limit-{opening!r}.xlsx")
    markup = opening + (b"x" * 4_096) + closing
    _append_stylesheet_opaque_root_markup(workbook, markup)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_NON_CHARACTER_DATA_MARKUP_BYTES",
        len(markup),
    )

    assert load_snapshot(workbook).sheets


def test_in_memory_xml_readers_reject_oversized_ignored_markup_before_elementtree(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = b"<root><!--" + (b"x" * 129) + b"--></root>"
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_XML_NON_CHARACTER_DATA_MARKUP_BYTES",
        128,
    )

    def unexpected_parser(*args, **kwargs):
        raise AssertionError("ElementTree started before the lexical markup budget")

    monkeypatch.setattr(
        workbook_module,
        "_ooxml_xml_parser_with_character_data_budget",
        unexpected_parser,
    )
    with pytest.raises(ValueError, match="XML non-character-data markup"):
        workbook_module._xml_root_from_payload(payload)


def test_bounded_xml_parser_forbids_document_type_declarations() -> None:
    with pytest.raises(ValueError, match="DTDForbidden"):
        list(
            workbook_module._iterparse_ooxml_xml_with_character_data_budget(
                io.BytesIO(b"<!DOCTYPE root><root/>")
            )
        )


@pytest.mark.parametrize(
    ("encoding", "bom"),
    (
        ("utf-8", b""),
        ("utf-16-le", b"\xff\xfe"),
        ("utf-16-be", b"\xfe\xff"),
    ),
)
def test_in_memory_xml_parser_forbids_document_type_declarations(
    encoding: str,
    bom: bytes,
) -> None:
    payload = bom + "<!DOCTYPE root><root/>".encode(encoding)

    with pytest.raises(ValueError, match="DTDForbidden"):
        workbook_module._xml_root_from_payload(payload)


@pytest.mark.parametrize(
    "payload",
    (
        b"<root><!-- <!DOCTYPE root> --></root>",
        b"<root><![CDATA[<!DOCTYPE root>]]></root>",
        b"<root>&lt;!DOCTYPE root&gt;</root>",
    ),
)
def test_in_memory_xml_parser_accepts_inert_document_type_spelling(payload: bytes) -> None:
    assert workbook_module._xml_root_from_payload(payload).tag == "root"


def test_xml_start_tag_budget_accepts_its_exact_physical_boundary() -> None:
    payload = b'<root alpha="FormulaFence"/>'

    assert workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_start_tag_bytes=len(payload),
    )
    assert not workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_start_tag_bytes=len(payload) - 1,
    )


def test_xml_start_tag_budget_skips_quoted_and_non_element_markup(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Comments, CDATA, and PIs cannot impersonate oversized elements."""
    fake_tag = b"<notAnElement " + (b'attribute="x" ' * 12) + b"/>"
    actual_start_tag = b'<root note="a quoted > delimiter"/>'
    payload = (
        b"<!-- "
        + fake_tag
        + b" -->"
        + b"<![CDATA[["
        + fake_tag
        + b"]]>"
        + b"<?FormulaFence "
        + fake_tag
        + b"?>"
        + b"<!DOCTYPE root [ <!ELEMENT root ANY> ]>"
        + actual_start_tag
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_XML_LEXICAL_CHUNK_BYTES", 1)

    assert workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_start_tag_bytes=len(actual_start_tag),
    )
    assert not workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_start_tag_bytes=len(actual_start_tag) - 1,
    )


@pytest.mark.parametrize(
    ("encoding", "bom"),
    (
        ("utf-16-le", b"\xff\xfe"),
        ("utf-16-be", b"\xfe\xff"),
        ("utf-32-le", b"\xff\xfe\x00\x00"),
        ("utf-32-be", b"\x00\x00\xfe\xff"),
        ("utf-16-le", b""),
        ("utf-16-be", b""),
        ("utf-32-le", b""),
        ("utf-32-be", b""),
    ),
)
def test_xml_start_tag_budget_covers_fixed_width_xml_encodings(
    encoding: str,
    bom: bytes,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    start_tag = "<root alpha='FormulaFence'/>".encode(encoding)
    payload = bom + start_tag
    monkeypatch.setattr(workbook_module, "_OOXML_READER_XML_LEXICAL_CHUNK_BYTES", 1)

    assert workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_start_tag_bytes=len(start_tag),
    )
    assert not workbook_module._xml_start_tags_within_budget(
        io.BytesIO(payload),
        maximum_start_tag_bytes=len(start_tag) - 1,
    )


@pytest.mark.parametrize(
    ("marker", "after"),
    (
        (b"<styleSheet", None),
        (b"<xf ", b"<cellXfs"),
    ),
)
def test_semantic_reader_preflight_rejects_oversized_style_attribute_maps_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    marker: bytes,
    after: bytes | None,
) -> None:
    workbook = make_model(tmp_path / f"oversized-style-attributes-{marker!r}.xlsx")
    start_tag_size = _append_stylesheet_start_tag_attributes(
        workbook,
        10_000,
        marker=marker,
        after=after,
    )
    assert start_tag_size > workbook_module._OOXML_READER_MAX_XML_START_TAG_BYTES

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "XML start tag" in message


def test_stream_reader_rejects_an_oversized_attribute_map_before_elementtree(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "stream-start-tag-before-elementtree.xlsx")
    _append_stylesheet_start_tag_attributes(
        workbook,
        10_000,
        marker=b"<styleSheet",
    )

    def unexpected_iterparse(*args, **kwargs):
        raise AssertionError("ElementTree started before the lexical start-tag budget")

    monkeypatch.setattr(workbook_module.ElementTree, "iterparse", unexpected_iterparse)
    with ZipFile(workbook) as archive:
        with pytest.raises(WorkbookLoadError, match="XML start tag"):
            workbook_module._stream_ooxml_reader_xml(archive, "xl/styles.xml")


def test_in_memory_xml_readers_reject_an_oversized_attribute_map_before_elementtree(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    attributes = b"".join(
        f' ff{index:08x}="x"'.encode() for index in range(10_000)
    )
    payload = b"<root" + attributes + b"/>"

    def unexpected_iterparse(*args, **kwargs):
        raise AssertionError("ElementTree started before the lexical start-tag budget")

    monkeypatch.setattr(workbook_module.ElementTree, "iterparse", unexpected_iterparse)
    assert (
        workbook_module._xml_payload_structure_element_count_within_budget(
            payload,
            maximum_element_count=1,
        )
        is None
    )

    def unexpected_fromstring(*args, **kwargs):
        raise AssertionError("ElementTree started before the lexical start-tag budget")

    monkeypatch.setattr(workbook_module.ElementTree, "fromstring", unexpected_fromstring)
    with pytest.raises(ValueError, match="XML start tag"):
        workbook_module._xml_root_from_payload(payload)


def test_semantic_reader_preflight_rejects_excessive_cell_styles_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-styles.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_CELL_STYLE_COUNT", 0)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "cell styles" in message


def test_semantic_reader_preflight_rejects_excessive_stylesheet_containers_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-stylesheet-containers.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_STYLESHEET_CONTAINER_COUNT", 0)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet containers" in message


@pytest.mark.parametrize(
    ("parent_names", "child_name", "limit_name", "message_fragment"),
    _STYLESHEET_CATALOG_CASES,
)
def test_semantic_reader_preflight_rejects_excessive_stylesheet_catalogs_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    parent_names: tuple[str, ...],
    child_name: str,
    limit_name: str,
    message_fragment: str,
) -> None:
    workbook = make_model(tmp_path / f"too-many-stylesheet-{child_name}.xlsx")
    record_count = _append_stylesheet_catalog_declarations(
        workbook,
        parent_names,
        child_name,
        1,
    )
    monkeypatch.setattr(workbook_module, limit_name, record_count - 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert message_fragment in message


def test_semantic_reader_preflight_rejects_default_font_catalog_over_limit_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-font-records.xlsx")
    _append_stylesheet_catalog_declarations(
        workbook,
        ("fonts",),
        "font",
        workbook_module._OOXML_READER_MAX_FONT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "font records" in message


def test_semantic_reader_preflight_accepts_font_catalog_at_configured_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "font-records-at-limit.xlsx")
    record_count = _append_stylesheet_catalog_declarations(
        workbook,
        ("fonts",),
        "font",
        1,
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_FONT_COUNT", record_count)

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_accepts_font_catalog_at_default_limit(
    tmp_path: Path,
) -> None:
    workbook = make_model(tmp_path / "font-records-at-default-limit.xlsx")
    existing_count = _append_stylesheet_catalog_declarations(
        workbook,
        ("fonts",),
        "font",
        0,
    )
    record_count = _append_stylesheet_catalog_declarations(
        workbook,
        ("fonts",),
        "font",
        workbook_module._OOXML_READER_MAX_FONT_COUNT - existing_count,
    )

    assert record_count == workbook_module._OOXML_READER_MAX_FONT_COUNT
    assert load_snapshot(workbook).file_type == "xlsx"


def test_semantic_reader_preflight_counts_alternate_namespace_cell_styles_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-cell-styles.xlsx")
    _append_stylesheet_catalog_declarations(
        workbook,
        ("cellXfs",),
        "xf",
        0,
        alternate_parent_namespace="urn:formulafence:alternate-style",
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_CELL_STYLE_COUNT", 0)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "cell styles" in message


def test_semantic_reader_preflight_counts_unknown_font_children_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "unexpected-font-record.xlsx")
    record_count = _append_stylesheet_catalog_declarations(
        workbook,
        ("fonts",),
        "unexpectedFontRecord",
        1,
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_FONT_COUNT", record_count - 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "font records" in message


@pytest.mark.parametrize("nested", (False, True))
def test_semantic_reader_preflight_rejects_opaque_stylesheet_root_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    nested: bool,
) -> None:
    workbook = make_model(tmp_path / f"opaque-stylesheet-root-{nested}.xlsx")
    element_count = _append_stylesheet_opaque_root_xml_elements(
        workbook,
        1,
        nested=nested,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet opaque root XML structure" in message


def test_semantic_reader_preflight_covers_alternate_namespace_stylesheet_root_xml(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-stylesheet-root-opaque.xml.xlsx")
    _replace_stylesheet_root_namespace(
        workbook,
        "urn:formulafence:archive-safety-alternate-root",
    )
    element_count = _append_stylesheet_opaque_root_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet opaque root XML structure" in message


def test_semantic_reader_preflight_rejects_foreign_stylesheet_root_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "foreign-stylesheet-root.xlsx")
    _replace_stylesheet_root_local_name(workbook, "foreignStyleSheet")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        0,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet opaque root XML structure" in message


@pytest.mark.parametrize("nested", (False, True))
def test_semantic_reader_preflight_rejects_opaque_stylesheet_catalog_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    nested: bool,
) -> None:
    workbook = make_model(tmp_path / f"opaque-stylesheet-catalog-{nested}.xlsx")
    element_count = _append_stylesheet_opaque_catalog_xml_elements(
        workbook,
        1,
        nested=nested,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_OPAQUE_CATALOG_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet opaque catalog XML structure" in message


@pytest.mark.parametrize("nested", (False, True))
def test_semantic_reader_preflight_rejects_opaque_cell_style_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    nested: bool,
) -> None:
    workbook = make_model(tmp_path / f"opaque-cell-style-{nested}.xlsx")
    _append_cell_style_xml_elements(workbook, 6 if not nested else 5, nested=nested)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_RECORD_XML_ELEMENT_COUNT",
        6,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet record XML structure" in message


def test_semantic_reader_preflight_rejects_repeated_known_cell_style_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "repeated-cell-style-alignment.xml.xlsx")
    _append_cell_style_xml_elements(workbook, 7, known_alignment=True)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_RECORD_XML_ELEMENT_COUNT",
        6,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet record XML structure" in message


def test_semantic_reader_preflight_rejects_aggregate_cell_style_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "aggregate-cell-style-xml.xlsx")
    _replace_member(
        workbook,
        "xl/styles.xml",
        (
            b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b"<cellXfs>"
            b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0"><alignment/></xf>'
            b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0"><alignment/></xf>'
            b"</cellXfs></styleSheet>"
        ),
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_RECORD_XML_ELEMENT_COUNT",
        1,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_AGGREGATE_RECORD_XML_ELEMENT_COUNT",
        1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "aggregate stylesheet record XML structure" in message


@pytest.mark.parametrize("nested", (False, True))
def test_semantic_reader_preflight_rejects_stylesheet_extension_list_xml_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    nested: bool,
) -> None:
    workbook = make_model(tmp_path / f"stylesheet-extension-list-{nested}.xlsx")
    element_count = _append_stylesheet_extension_list_xml_elements(
        workbook,
        1,
        nested=nested,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet extension-list XML structure" in message


def test_semantic_reader_preflight_covers_nested_cell_style_extension_lists(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "nested-cell-style-extension-list.xlsx")
    element_count = _append_stylesheet_extension_list_xml_elements(
        workbook,
        1,
        within_cell_style=True,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet extension-list XML structure" in message


def test_semantic_reader_preflight_covers_alternate_namespace_stylesheet_extension_lists(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "alternate-cell-style-extension-list.xlsx")
    element_count = _append_stylesheet_extension_list_xml_elements(
        workbook,
        1,
        within_cell_style=True,
        extension_namespace="urn:formulafence:archive-safety-alternate",
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count - 1,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet extension-list XML structure" in message


def test_semantic_reader_preflight_accepts_stylesheet_opaque_root_xml_at_exact_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "opaque-stylesheet-root-at-limit.xlsx")
    element_count = _append_stylesheet_opaque_root_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        element_count,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_accepts_stylesheet_extension_list_at_exact_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "stylesheet-extension-list-at-limit.xlsx")
    element_count = _append_stylesheet_extension_list_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        0,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_EXTENSION_LIST_XML_ELEMENT_COUNT",
        element_count,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_keeps_standard_stylesheet_root_content_unmetered(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "standard-stylesheet-root-content.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT",
        0,
    )
    monkeypatch.setattr(
        workbook_module,
        "_OOXML_READER_MAX_STYLESHEET_EXTENSION_LIST_XML_ELEMENT_COUNT",
        0,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.file_type == "xlsx"


def test_semantic_reader_preflight_rejects_default_opaque_stylesheet_root_xml_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-opaque-stylesheet-root-limit.xlsx")
    _append_stylesheet_opaque_root_xml_elements(
        workbook,
        workbook_module._OOXML_READER_MAX_STYLESHEET_OPAQUE_ROOT_XML_ELEMENT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet opaque root XML structure" in message


def test_semantic_reader_preflight_rejects_default_stylesheet_extension_list_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-stylesheet-extension-list-limit.xlsx")
    _append_stylesheet_extension_list_xml_elements(
        workbook,
        workbook_module._OOXML_READER_MAX_STYLESHEET_EXTENSION_LIST_XML_ELEMENT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet extension-list XML structure" in message


def test_semantic_reader_preflight_rejects_default_opaque_stylesheet_catalog_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-opaque-stylesheet-catalog-limit.xlsx")
    _append_stylesheet_opaque_catalog_xml_elements(
        workbook,
        workbook_module._OOXML_READER_MAX_STYLESHEET_OPAQUE_CATALOG_XML_ELEMENT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet opaque catalog XML structure" in message


def test_semantic_reader_preflight_rejects_default_stylesheet_record_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "default-stylesheet-record-limit.xlsx")
    _append_cell_style_xml_elements(
        workbook,
        workbook_module._OOXML_READER_MAX_STYLESHEET_RECORD_XML_ELEMENT_COUNT,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "stylesheet record XML structure" in message


def test_semantic_reader_preflight_rejects_excessive_cell_text_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-much-cell-text.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_CELL_TEXT_CHARACTERS", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "cell text" in message


def test_semantic_reader_preflight_rejects_excessive_formula_text_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-much-formula-text.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_FORMULA_CHARACTERS", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "formula text" in message


def test_semantic_reader_preflight_rejects_cell_text_beyond_excel_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "cell-text-limit.xlsx")
    member_name = "xl/worksheets/sheet1.xml"
    with ZipFile(workbook) as archive:
        worksheet = archive.read(member_name)
    _replace_member(
        workbook,
        member_name,
        worksheet.replace(
            b"<t>Revenue</t>",
            b"<t>" + b"x" * 32_768 + b"</t>",
        ),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "cell text" in message


def test_semantic_reader_preflight_rejects_formula_text_beyond_excel_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "formula-text-limit.xlsx")
    member_name = "xl/worksheets/sheet2.xml"
    with ZipFile(workbook) as archive:
        worksheet = archive.read(member_name)
    _replace_member(
        workbook,
        member_name,
        worksheet.replace(
            b"<f>Inputs!B2*2</f>",
            b"<f>=" + b"1" * 8_192 + b"</f>",
        ),
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "formula text" in message


def test_semantic_reader_preflight_rejects_excessive_shared_strings_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "too-many-shared-strings.xlsx")
    _append_member(
        workbook,
        "xl/sharedStrings.xml",
        (
            b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b"<si><t>one</t></si><si><t>two</t></si></sst>"
        ),
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_SHARED_STRING_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "shared-string table entries" in message


def test_semantic_reader_preflight_follows_relationship_selected_shared_strings(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "relationship-selected-shared-strings.xlsx")
    relationship_member = "xl/_rels/workbook.xml.rels"
    shared_member = "xl/private/strings.xml"
    with ZipFile(workbook) as archive:
        relationships = archive.read(relationship_member)
    _append_member(
        workbook,
        shared_member,
        (
            b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b"<si><t>one</t></si><si><t>two</t></si></sst>"
        ),
    )
    _replace_member(
        workbook,
        relationship_member,
        relationships.replace(
            b"</Relationships>",
            (
                b'<Relationship Id="rIdSharedStrings" '
                b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                b'relationships/sharedStrings" Target="private/strings.xml" />'
                b"</Relationships>"
            ),
        ),
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_SHARED_STRING_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "shared-string table entries" in message


def test_semantic_reader_preflight_follows_manifest_selected_shared_strings(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "manifest-selected-shared-strings.xlsx")
    content_types_member = "[Content_Types].xml"
    shared_member = "xl/private/strings.xml"
    with ZipFile(workbook) as archive:
        content_types = archive.read(content_types_member)
    _append_member(
        workbook,
        shared_member,
        (
            b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b"<si><t>one</t></si><si><t>two</t></si></sst>"
        ),
    )
    _replace_member(
        workbook,
        content_types_member,
        content_types.replace(
            b"</Types>",
            (
                b'<Override PartName="/xl/private/strings.xml" '
                b'ContentType="application/vnd.openxmlformats-officedocument.'
                b'spreadsheetml.sharedStrings+xml" />'
                b"</Types>"
            ),
        ),
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_SHARED_STRING_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "shared-string table entries" in message


def test_semantic_reader_preflight_counts_relationship_selected_worksheet_parts(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "relationship-selected-cells.xlsx")
    source_member = "xl/worksheets/sheet1.xml"
    selected_member = "xl/private/sheet-one.xml"
    relationship_member = "xl/_rels/workbook.xml.rels"
    with ZipFile(workbook) as archive:
        worksheet_payload = archive.read(source_member)
        relationships = archive.read(relationship_member)
    assert b"worksheets/sheet1.xml" in relationships
    _append_member(workbook, selected_member, worksheet_payload)
    _replace_member(
        workbook,
        relationship_member,
        relationships.replace(b"worksheets/sheet1.xml", b"private/sheet-one.xml"),
    )
    monkeypatch.setattr(workbook_module, "_OOXML_READER_MAX_WORKSHEET_CELL_COUNT", 1)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "populated worksheet cells" in message


def test_semantic_reader_preflight_rejects_xml_entity_declarations_before_scanners(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "xml-entity.xlsx")
    member_name = "xl/worksheets/sheet1.xml"
    with ZipFile(workbook) as archive:
        original = archive.read(member_name)
    _replace_member(
        workbook,
        member_name,
        b"<!DOCTYPE worksheet [<!ENTITY forbidden 'private'>]>\n" + original,
    )

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "workbook sheet metadata could not be scanned safely" in message


def test_macro_hash_streams_the_payload_without_zipfile_read(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "macro-stream.xlsx")
    payload = b"private macro payload" * 4096
    _append_member(workbook, "xl/vbaProject.bin", payload)

    original_read = ZipFile.read

    def reject_macro_read(self: ZipFile, name, *args, **kwargs):
        if name == "xl/vbaProject.bin":
            raise AssertionError("macro hashing loaded the complete payload")
        return original_read(self, name, *args, **kwargs)

    monkeypatch.setattr(ZipFile, "read", reject_macro_read)

    assert workbook_module._vba_hash(workbook) == hashlib.sha256(payload).hexdigest()


@pytest.mark.parametrize(
    ("member_name", "expected_detail"),
    [
        ("../not-a-workbook-part.xml", "unsafe or non-canonical"),
        ("XL/WORKBOOK.XML", "ambiguous"),
    ],
)
def test_archive_preflight_rejects_ambiguous_or_unsafe_member_paths(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    member_name: str,
    expected_detail: str,
) -> None:
    workbook = make_model(tmp_path / "unsafe-member-path.xlsx")
    _append_member(workbook, member_name, b"untrusted")

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert expected_detail in message
    assert member_name not in message


def test_archive_preflight_rejects_duplicate_zip_members_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "duplicate-member.xlsx")
    with ZipFile(workbook) as archive:
        workbook_payload = archive.read("xl/workbook.xml")
    with pytest.warns(UserWarning, match="Duplicate name"):
        _append_member(workbook, "xl/workbook.xml", workbook_payload)

    message = _reject_before_workbook_readers(monkeypatch, workbook)

    assert "ambiguous" in message
    assert "workbook.xml" not in message


def test_archive_preflight_rejects_encrypted_central_member_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "encrypted-member.xlsx")
    contents = bytearray(workbook.read_bytes())
    central_directory_offset = _last_central_directory_offset(contents)
    flag_offset = central_directory_offset + 8
    flag_bits = int.from_bytes(contents[flag_offset : flag_offset + 2], "little")
    contents[flag_offset : flag_offset + 2] = (flag_bits | 0x1).to_bytes(2, "little")
    workbook.write_bytes(contents)

    assert "encrypted" in _reject_before_workbook_readers(monkeypatch, workbook)


def test_archive_preflight_rejects_unicode_path_aliases_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "unicode-path-alias.xlsx")
    original_name = "xl/media/original.bin"
    unicode_alias = "xl/media/alias.bin"
    entry = ZipInfo(original_name)
    entry.extra = struct.pack(
        "<HHBI",
        0x7075,
        5 + len(unicode_alias.encode("utf-8")),
        1,
        binascii.crc32(original_name.encode("ascii")),
    ) + unicode_alias.encode("utf-8")
    with ZipFile(workbook, "a") as archive:
        archive.writestr(entry, b"untrusted")

    assert "Unicode-path aliases" in _reject_before_workbook_readers(
        monkeypatch,
        workbook,
    )


def test_archive_preflight_rejects_malformed_zip64_member_metadata_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "malformed-zip64.xlsx")
    contents = bytearray(workbook.read_bytes())
    central_directory_offset = _last_central_directory_offset(contents)
    uncompressed_size_offset = central_directory_offset + 24
    contents[uncompressed_size_offset : uncompressed_size_offset + 4] = b"\xff" * 4
    workbook.write_bytes(contents)

    assert "ZIP64" in _reject_before_workbook_readers(monkeypatch, workbook)


def test_archive_preflight_rejects_local_header_mismatch_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "local-header-mismatch.xlsx")
    contents = bytearray(workbook.read_bytes())
    local_flag_offset = 6
    flag_bits = int.from_bytes(contents[local_flag_offset : local_flag_offset + 2], "little")
    contents[local_flag_offset : local_flag_offset + 2] = (flag_bits | 0x8).to_bytes(
        2,
        "little",
    )
    workbook.write_bytes(contents)

    assert "local member metadata is inconsistent" in _reject_before_workbook_readers(
        monkeypatch,
        workbook,
    )


def test_archive_preflight_rejects_symbolic_link_members_before_readers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "symbolic-link.xlsx")
    link = ZipInfo("xl/media/link")
    link.create_system = 3
    link.external_attr = (stat.S_IFLNK | 0o777) << 16
    with ZipFile(workbook, "a") as archive:
        archive.writestr(link, b"not a real part")

    assert "symbolic-link" in _reject_before_workbook_readers(monkeypatch, workbook)


def test_cli_surfaces_archive_preflight_as_an_input_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    workbook = make_model(tmp_path / "cli-limit.xlsx")
    monkeypatch.setattr(workbook_module, "_OOXML_ARCHIVE_MAX_ENTRY_COUNT", 1)

    assert main(["profile", str(workbook)]) == 2

    assert "safety preflight" in capsys.readouterr().err


def test_cli_surfaces_malformed_openpyxl_cell_metadata_as_an_input_error(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    workbook = make_model(tmp_path / "empty-style.xlsx")
    member_name = "xl/worksheets/sheet1.xml"
    with ZipFile(workbook) as archive:
        worksheet = archive.read(member_name)
    _replace_member(
        workbook,
        member_name,
        worksheet.replace(b"<c ", b'<c s="" ', 1),
    )

    assert main(["profile", str(workbook)]) == 2

    assert "Could not read workbook" in capsys.readouterr().err


def test_load_snapshot_surfaces_openpyxl_index_errors_as_input_errors(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = make_model(tmp_path / "reader-index-error.xlsx")

    def malformed_reader(*args, **kwargs):
        raise IndexError("malformed workbook metadata")

    monkeypatch.setattr(workbook_module, "load_workbook", malformed_reader)

    with pytest.raises(WorkbookLoadError, match="Could not read workbook"):
        load_snapshot(workbook)
