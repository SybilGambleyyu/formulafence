from __future__ import annotations

import json

import pytest
from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

import formulafence.cli as cli_module
import formulafence.policy as policy_module
from formulafence.cli import main
from formulafence.lint import (
    DEFAULT_MAX_AGGREGATE_OMISSION_GAP_CELLS,
    DEFAULT_MAX_FORMULA_PATTERN_FINDINGS,
)
from formulafence.models import WorkbookSnapshot
from formulafence.output import DEFAULT_MAX_REPORT_BYTES
from formulafence.workbook import (
    DEFAULT_MAX_DEPENDENCY_EDGES,
    DEFAULT_MAX_FORMULA_DEFINED_NAME_STATES,
    DEFAULT_MAX_PROFILE_RECORDS,
)

from .helpers import (
    change_formula_dde_link_input,
    change_formula_defined_xlm_action_call,
    change_formula_defined_xlm_action_input,
    change_formula_defined_xlm_environment_information_call,
    change_formula_defined_xlm_environment_information_input,
    change_formula_defined_xlm_evaluation_call,
    change_formula_defined_xlm_evaluation_input,
    change_formula_defined_xlm_get_cell_call,
    change_formula_defined_xlm_get_cell_input,
    change_formula_defined_xlm_registration_call,
    change_formula_defined_xlm_registration_input,
    change_formula_environment_information_definition,
    change_formula_environment_information_input,
    change_formula_external_action_input,
    change_formula_external_action_target,
    change_office_custom_function_call,
    change_office_custom_function_input,
    change_unqualified_runtime_function_call,
    change_unqualified_runtime_function_input,
    change_worksheet_code_resource_registration_call,
    change_worksheet_code_resource_registration_input,
    make_calculated_column_model,
    make_formula_cached_result_model,
    make_formula_dde_link_model,
    make_formula_defined_xlm_action_model,
    make_formula_defined_xlm_environment_information_model,
    make_formula_defined_xlm_evaluation_model,
    make_formula_defined_xlm_get_cell_model,
    make_formula_defined_xlm_registration_model,
    make_formula_environment_information_model,
    make_formula_external_action_model,
    make_ignored_error_model,
    make_model,
    make_office_custom_function_model,
    make_python_in_excel_model,
    make_unqualified_runtime_function_model,
    make_worksheet_code_resource_registration_model,
    rewrite,
    set_python_in_excel_formula_source,
)


def _assert_html_review_artifact(rendered: str) -> None:
    assert rendered.startswith("<!doctype html>")
    assert 'id="review-filter"' in rendered
    assert 'id="severity-filter"' in rendered
    assert "<script src=" not in rendered
    assert "<link rel=" not in rendered


def test_check_emits_sarif_and_fails_for_a_policy_violation(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Model"]["B2"], "value", 200))
    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_formula_to_value: true\n",
        encoding="utf-8",
    )
    sarif = tmp_path / "result.sarif"

    result = main(
        [
            "check",
            str(baseline),
            str(candidate),
            "--policy",
            str(policy),
            "--format",
            "sarif",
            "--output",
            str(sarif),
        ]
    )

    assert result == 1
    payload = json.loads(sarif.read_text(encoding="utf-8"))
    assert payload["version"] == "2.1.0"
    results = payload["runs"][0]["results"]
    assert any(item["ruleId"] == "FFP001" for item in results)
    formula_override = next(item for item in results if item["ruleId"] == "FF001")
    assert formula_override["properties"]["impact_paths"] == [
        {"path": ["Model!B2", "Model!C2", "Dashboard!B12"], "target": "Dashboard!B12"},
        {"path": ["Model!B2", "Model!C2"], "target": "Model!C2"},
    ]


def test_check_rejects_an_oversized_policy_before_loading_workbooks(
    tmp_path,
    monkeypatch,
    capsys,
) -> None:
    policy = tmp_path / "oversized.yml"
    policy.write_bytes(b"x" * 33)
    monkeypatch.setattr(policy_module, "_POLICY_MAX_SOURCE_BYTES", 32)

    def unexpected_snapshot_load(*args, **kwargs):
        raise AssertionError("an oversized policy reached workbook inspection")

    monkeypatch.setattr(cli_module, "load_snapshot", unexpected_snapshot_load)

    result = cli_module.main(
        [
            "check",
            str(tmp_path / "before.xlsx"),
            str(tmp_path / "after.xlsx"),
            "--policy",
            str(policy),
        ]
    )

    assert result == 2
    assert "Policy source exceeds" in capsys.readouterr().err


def test_profile_does_not_expose_cell_values(tmp_path) -> None:
    workbook = make_model(tmp_path / "model.xlsx")
    output = tmp_path / "profile.json"

    assert main(["profile", str(workbook), "--format", "json", "--output", str(output)]) == 0

    profile = output.read_text(encoding="utf-8")
    assert "Calculated revenue" not in profile
    assert '"formula_cells"' in profile


def test_lint_emits_sarif_and_can_fail_a_ci_threshold(tmp_path) -> None:
    workbook_path = tmp_path / "model.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["A2"] = 1
    worksheet["B2"] = "=A2*2"
    worksheet["D2"] = "=C2*2"
    worksheet["E2"] = "=D2*2"
    workbook.save(workbook_path)
    output = tmp_path / "lint.sarif"

    assert (
        main(
            [
                "lint",
                str(workbook_path),
                "--format",
                "sarif",
                "--fail-on",
                "high",
                "--output",
                str(output),
            ]
        )
        == 1
    )

    payload = json.loads(output.read_text(encoding="utf-8"))
    result = payload["runs"][0]["results"][0]
    assert result["ruleId"] == "FF082"
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!C2"
    assert "=A2*2" not in json.dumps(payload)


def test_lint_high_threshold_leaves_manual_numeric_value_for_review(tmp_path) -> None:
    workbook_path = tmp_path / "model.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["A2"] = 1
    worksheet["B2"] = "=A2*2"
    worksheet["C2"] = 99
    worksheet["D2"] = "=C2*2"
    worksheet["E2"] = "=D2*2"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "medium"]) == 1


def test_lint_aggregate_omission_is_medium_and_uses_its_gap_bound(tmp_path) -> None:
    workbook_path = tmp_path / "model.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    for row, value in enumerate((10, 20, 30, 40, 50, 60), start=2):
        worksheet.cell(row=row, column=2, value=value)
    worksheet["B8"] = "=SUM(B2:B4)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "medium"]) == 1
    assert (
        main(
            [
                "lint",
                str(workbook_path),
                "--max-aggregate-omission-gap-cells",
                "2",
            ]
        )
        == 0
    )


def test_lint_direct_unlocked_formula_is_medium(tmp_path) -> None:
    workbook_path = tmp_path / "protected.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["A2"] = 10
    worksheet["B2"] = "=A2*2"
    worksheet["B2"].protection = Protection(locked=False)
    worksheet.protection.sheet = True
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "medium"]) == 1


def test_lint_incomplete_manual_formula_calculation_is_medium(tmp_path) -> None:
    workbook_path = tmp_path / "manual-incomplete.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["A2"] = 10
    worksheet["B2"] = "=A2*2"
    workbook.calculation.calcMode = "manual"
    workbook.calculation.calcCompleted = False
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "medium"]) == 1


def test_lint_error_checking_suppression_is_medium(tmp_path) -> None:
    workbook_path = make_ignored_error_model(tmp_path / "ignored-errors.xlsx")

    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "medium"]) == 1


def test_lint_table_calculated_column_exception_is_medium(tmp_path) -> None:
    workbook_path = make_calculated_column_model(
        tmp_path / "calculated-column.xlsx",
        exception=99,
    )

    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "medium"]) == 1


def test_lint_conditional_aggregate_range_shape_mismatch_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "conditional-aggregate-range-shape.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=MAXIFS(C2:C10,A2:A12,A14)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_sumproduct_range_shape_mismatch_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "sumproduct-range-shape.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=SUMPRODUCT(C2:C10,A2:A12)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_mmult_dimension_mismatch_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "mmult-dimension.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=MMULT(C2:D4,A2:B6)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_lookup_return_index_mismatch_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "lookup-return-index.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=VLOOKUP(A2,C2:D6,3,FALSE)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_choose_literal_index_mismatch_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "choose-literal-index.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=CHOOSE(3,C2,D2)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_randbetween_literal_bound_mismatch_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "randbetween-literal-bounds.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=RANDBETWEEN(2,1)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_subtotal_literal_function_num_mismatch_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "subtotal-literal-function-num.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=SUBTOTAL(12,C2:C10)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_index_literal_position_mismatch_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "index-literal-position.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=INDEX(C2:D4,4)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_approximate_lookup_unsorted_vector_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "approximate-lookup-unsorted.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    for row, value in enumerate((1, 3, 2), start=2):
        worksheet.cell(row=row, column=3, value=value)
        worksheet.cell(row=row, column=4, value=value * 10)
    worksheet["B2"] = "=VLOOKUP(A2,C2:D4,2)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_direct_self_reference_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "direct-self-reference.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=B2+1"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_multi_cell_static_cycle_is_high(tmp_path) -> None:
    workbook_path = tmp_path / "multi-cell-static-cycle.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=C2+1"
    worksheet["C2"] = "=B2+1"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_explicit_broken_reference_is_critical(tmp_path) -> None:
    workbook_path = tmp_path / "broken-reference.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["B2"] = "=IFERROR(#REF!,0)"
    workbook.save(workbook_path)

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 1
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_saved_broken_reference_result_is_high(tmp_path) -> None:
    workbook_path = make_formula_cached_result_model(
        tmp_path / "saved-broken-reference.xlsx",
        error_formula="=SUM(Inputs!A1:A1)",
        error_result="#REF!",
    )

    assert main(["lint", str(workbook_path), "--fail-on", "critical"]) == 0
    assert main(["lint", str(workbook_path), "--fail-on", "high"]) == 1


def test_lint_defaults_its_bounded_finding_limit() -> None:
    arguments = cli_module.build_parser().parse_args(["lint", "model.xlsx"])

    assert arguments.max_formula_pattern_findings == DEFAULT_MAX_FORMULA_PATTERN_FINDINGS
    assert (
        arguments.max_aggregate_omission_gap_cells
        == DEFAULT_MAX_AGGREGATE_OMISSION_GAP_CELLS
    )
    assert arguments.max_report_bytes == DEFAULT_MAX_REPORT_BYTES


def test_lint_rejects_an_impossible_one_cell_aggregate_omission_bound() -> None:
    with pytest.raises(SystemExit):
        cli_module.build_parser().parse_args(
            [
                "lint",
                "model.xlsx",
                "--max-aggregate-omission-gap-cells",
                "1",
            ]
        )


def test_profile_dependency_edge_limit_fails_before_writing_an_output(tmp_path, capsys) -> None:
    workbook_path = tmp_path / "named-fanout.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    for row in range(1, 4):
        worksheet.cell(row=row, column=1, value=row)
    workbook.defined_names.add(
        DefinedName("Fanout", attr_text="=SUM(Model!$A$1,Model!$A$2,Model!$A$3)")
    )
    for row in range(1, 5):
        worksheet.cell(row=row, column=2, value="=Fanout")
    workbook.save(workbook_path)
    output = tmp_path / "profile.json"

    assert (
        main(
            [
                "profile",
                str(workbook_path),
                "--max-dependency-edges",
                "11",
                "--format",
                "json",
                "--output",
                str(output),
            ]
        )
        == 2
    )
    assert not output.exists()
    assert "max_dependency_edges=11" in capsys.readouterr().err


def test_profile_formula_defined_name_state_limit_fails_before_writing_an_output(
    tmp_path,
    capsys,
) -> None:
    workbook_path = tmp_path / "action-chain.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    for index in range(4):
        previous = f"+ActionName{index - 1:05d}" if index else ""
        workbook.defined_names.add(
            DefinedName(
                f"ActionName{index:05d}",
                attr_text='=HYPERLINK("https://example.invalid","x")' + previous,
            )
        )
    workbook.save(workbook_path)
    output = tmp_path / "profile.json"

    assert (
        main(
            [
                "profile",
                str(workbook_path),
                "--max-formula-defined-name-states",
                "16",
                "--format",
                "json",
                "--output",
                str(output),
            ]
        )
        == 2
    )
    assert not output.exists()
    assert "max_formula_defined_name_states=16" in capsys.readouterr().err


def test_diff_passes_the_change_analysis_state_limit(tmp_path, capsys) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Model"]["B2"], "value", 200))

    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--max-change-analysis-states",
                "1",
            ]
        )
        == 2
    )
    assert "max_change_analysis_states=1" in capsys.readouterr().err


def test_diff_fails_before_writing_an_oversized_report(tmp_path, capsys) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Model"]["B2"], "value", 200))
    output = tmp_path / "report.json"

    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(output),
                "--max-report-bytes",
                "1",
            ]
        )
        == 2
    )
    assert not output.exists()
    assert "max_report_bytes=1" in capsys.readouterr().err


def test_diff_defaults_the_report_byte_limit() -> None:
    arguments = cli_module.build_parser().parse_args(["diff", "before.xlsx", "after.xlsx"])

    assert arguments.max_report_bytes == DEFAULT_MAX_REPORT_BYTES
    assert arguments.max_dependency_edges == DEFAULT_MAX_DEPENDENCY_EDGES
    assert (
        arguments.max_formula_defined_name_states
        == DEFAULT_MAX_FORMULA_DEFINED_NAME_STATES
    )


def test_profile_fails_before_writing_an_oversized_artifact(tmp_path, capsys) -> None:
    workbook = make_model(tmp_path / "model.xlsx")
    for output_format, suffix in (("json", "json"), ("markdown", "md")):
        output = tmp_path / f"profile.{suffix}"

        assert (
            main(
                [
                    "profile",
                    str(workbook),
                    "--format",
                    output_format,
                    "--output",
                    str(output),
                    "--max-report-bytes",
                    "1",
                ]
            )
            == 2
        )
        assert not output.exists()
    assert "max_report_bytes=1" in capsys.readouterr().err


def test_profile_defaults_the_report_byte_limit() -> None:
    arguments = cli_module.build_parser().parse_args(["profile", "model.xlsx"])

    assert arguments.max_report_bytes == DEFAULT_MAX_REPORT_BYTES
    assert arguments.max_profile_records == DEFAULT_MAX_PROFILE_RECORDS
    assert arguments.max_dependency_edges == DEFAULT_MAX_DEPENDENCY_EDGES
    assert (
        arguments.max_formula_defined_name_states
        == DEFAULT_MAX_FORMULA_DEFINED_NAME_STATES
    )


def test_profile_record_limit_fails_before_writing_an_output(tmp_path, monkeypatch, capsys) -> None:
    class OversizedDynamicInventory:
        def __len__(self) -> int:
            return DEFAULT_MAX_PROFILE_RECORDS + 1

        def values(self):
            raise AssertionError("an oversized profile inventory was traversed")

    snapshot = WorkbookSnapshot(
        path=tmp_path / "model.xlsx",
        sha256="0" * 64,
        file_type="xlsx",
        sheets={},
        cells={},
        reverse_dependencies={},
        range_dependencies=[],
        external_references=set(),
        broken_references=set(),
        defined_names={},
        macro_hash=None,
        calculation_settings={},
        parser_warnings=(),
        dynamic_reference_functions=OversizedDynamicInventory(),
    )
    output = tmp_path / "profile.json"
    monkeypatch.setattr(cli_module, "load_snapshot", lambda path, **kwargs: snapshot)

    assert (
        main(
            [
                "profile",
                str(snapshot.path),
                "--format",
                "json",
                "--output",
                str(output),
            ]
        )
        == 2
    )
    assert not output.exists()
    assert "max_profile_records=100000" in capsys.readouterr().err


def test_profile_output_swap_cannot_overwrite_the_workbook(tmp_path, monkeypatch) -> None:
    workbook = make_model(tmp_path / "model.xlsx")
    original_workbook = workbook.read_bytes()
    output = tmp_path / "profile.md"
    original_ensure_output_safe = cli_module._ensure_output_safe

    def ensure_then_swap(path, *inputs, **kwargs):
        original_ensure_output_safe(path, *inputs, **kwargs)
        assert path == output
        output.symlink_to(workbook)

    monkeypatch.setattr(cli_module, "_ensure_output_safe", ensure_then_swap)

    assert main(["profile", str(workbook), "--output", str(output)]) == 0

    assert workbook.read_bytes() == original_workbook
    assert not output.is_symlink()
    assert output.read_text(encoding="utf-8").startswith("# FormulaFence workbook profile")


def test_init_path_swap_refuses_to_replace_a_target_file(tmp_path, monkeypatch, capsys) -> None:
    target = tmp_path / "protected.txt"
    target.write_text("preserve this file", encoding="utf-8")
    policy = tmp_path / "formulafence.yml"
    original_create_text_atomically = cli_module._create_text_atomically

    def swap_then_create(path, content):
        assert path == policy
        policy.symlink_to(target)
        original_create_text_atomically(path, content)

    monkeypatch.setattr(cli_module, "_create_text_atomically", swap_then_create)

    assert main(["init", str(policy)]) == 2

    assert target.read_text(encoding="utf-8") == "preserve this file"
    assert policy.is_symlink()
    assert not list(tmp_path.glob("formulafence-output-*"))
    assert "Refusing to replace existing policy" in capsys.readouterr().err


def test_init_path_created_after_preflight_is_not_replaced(tmp_path, monkeypatch, capsys) -> None:
    policy = tmp_path / "formulafence.yml"
    original_create_text_atomically = cli_module._create_text_atomically

    def create_then_publish(path, content):
        assert path == policy
        policy.write_text("concurrent policy", encoding="utf-8")
        original_create_text_atomically(path, content)

    monkeypatch.setattr(cli_module, "_create_text_atomically", create_then_publish)

    assert main(["init", str(policy)]) == 2

    assert policy.read_text(encoding="utf-8") == "concurrent policy"
    assert not list(tmp_path.glob("formulafence-output-*"))
    assert "Refusing to replace existing policy" in capsys.readouterr().err


def test_init_hard_link_created_after_preflight_is_not_replaced(
    tmp_path, monkeypatch, capsys
) -> None:
    target = tmp_path / "protected.txt"
    target.write_text("preserve this file", encoding="utf-8")
    policy = tmp_path / "formulafence.yml"
    original_create_text_atomically = cli_module._create_text_atomically

    def hard_link_then_publish(path, content):
        assert path == policy
        policy.hardlink_to(target)
        original_create_text_atomically(path, content)

    monkeypatch.setattr(cli_module, "_create_text_atomically", hard_link_then_publish)

    assert main(["init", str(policy)]) == 2

    assert target.read_text(encoding="utf-8") == "preserve this file"
    assert policy.read_text(encoding="utf-8") == "preserve this file"
    assert not list(tmp_path.glob("formulafence-output-*"))
    assert "Refusing to replace existing policy" in capsys.readouterr().err


def test_init_refuses_an_existing_policy_without_force(tmp_path, capsys) -> None:
    policy = tmp_path / "formulafence.yml"
    policy.write_text("existing policy", encoding="utf-8")

    assert main(["init", str(policy)]) == 2

    assert policy.read_text(encoding="utf-8") == "existing policy"
    assert "Refusing to replace existing policy" in capsys.readouterr().err


def test_init_force_replaces_an_existing_policy(tmp_path) -> None:
    policy = tmp_path / "formulafence.yml"
    policy.write_text("old policy", encoding="utf-8")

    assert main(["init", str(policy), "--force"]) == 0

    assert policy.read_text(encoding="utf-8") == cli_module.DEFAULT_POLICY


def test_init_force_replaces_a_final_symlink_without_touching_its_target(tmp_path) -> None:
    target = tmp_path / "protected.txt"
    target.write_text("preserve this file", encoding="utf-8")
    policy = tmp_path / "formulafence.yml"
    policy.symlink_to(target)

    assert main(["init", str(policy), "--force"]) == 0

    assert target.read_text(encoding="utf-8") == "preserve this file"
    assert not policy.is_symlink()
    assert policy.read_text(encoding="utf-8") == cli_module.DEFAULT_POLICY


def test_cli_can_redact_external_workbook_link_material_from_shared_reports(tmp_path) -> None:
    baseline_marker = "PRIVATE-SHARED-BASELINE"
    candidate_marker = "PRIVATE-SHARED-CANDIDATE"

    def add_external_link_surfaces(workbook, marker: str) -> None:
        source = f"C:\\{marker}\\[Source.xlsx]Inputs"
        workbook["Model"]["D2"] = f"='{source}'!$B$2"
        workbook.defined_names.add(
            DefinedName("ExternalNamedLimit", attr_text=f"'{source}'!$C$2")
        )
        validation = DataValidation(
            type="whole",
            operator="greaterThan",
            formula1=f"='{source}'!$D$2",
        )
        validation.add("E2")
        workbook["Model"].add_data_validation(validation)

    def external_model(path, marker: str):
        workbook = make_model(path)
        rewrite(workbook, lambda model: add_external_link_surfaces(model, marker))
        return workbook

    baseline = external_model(tmp_path / "baseline.xlsx", baseline_marker)
    candidate = external_model(tmp_path / "candidate.xlsx", candidate_marker)
    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert baseline_marker in default_rendered
    assert candidate_marker in default_rendered

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-external-workbook-links",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert baseline_marker not in rendered
        assert candidate_marker not in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert "External-workbook link material:** redacted for sharing" in rendered
        else:
            assert "external-workbook link material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_external_workbook_link_surface_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "redacted-policy.sarif"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "sarif",
                "--redact-external-workbook-links",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FFP081" in policy_rendered
    assert baseline_marker not in policy_rendered
    assert candidate_marker not in policy_rendered

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    external_model(baseline_directory / "model.xlsx", baseline_marker)
    external_model(candidate_directory / "model.xlsx", candidate_marker)
    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"redacted-portfolio.{suffix}"
        assert (
            main(
                [
                    "portfolio",
                    str(baseline_directory),
                    str(candidate_directory),
                    "--format",
                    report_format,
                    "--redact-external-workbook-links",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert baseline_marker not in rendered
        assert candidate_marker not in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert "External-workbook link material:** redacted for sharing" in rendered
        else:
            assert "external-workbook link material redacted" in rendered


def test_cli_can_redact_formula_external_action_and_dde_material_from_shared_reports(
    tmp_path,
) -> None:
    action_baseline_marker = "PRIVATE-LINK-BASELINE"
    action_candidate_marker = "PRIVATE-LINK-CANDIDATE"
    baseline = make_formula_external_action_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_external_action_model(tmp_path / "candidate.xlsx")
    change_formula_external_action_target(candidate)

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert action_baseline_marker in default_rendered
    assert action_candidate_marker in default_rendered

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"action-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-formula-external-actions",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert action_baseline_marker not in rendered
        assert action_candidate_marker not in rendered
        assert "FF064" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert "Formula external-action / DDE material:** redacted for sharing" in rendered
        elif report_format == "json":
            assert "formula external-action material redacted" in rendered

    action_input_baseline = make_formula_external_action_model(
        tmp_path / "action-input-baseline.xlsx"
    )
    action_input_candidate = make_formula_external_action_model(
        tmp_path / "action-input-candidate.xlsx"
    )
    change_formula_external_action_input(action_input_candidate)
    action_input_output = tmp_path / "action-input-redacted.json"
    assert (
        main(
            [
                "diff",
                str(action_input_baseline),
                str(action_input_candidate),
                "--format",
                "json",
                "--redact-formula-external-actions",
                "--output",
                str(action_input_output),
            ]
        )
        == 0
    )
    action_input_rendered = action_input_output.read_text(encoding="utf-8")
    assert "PRIVATE-REFERENCED-LINK-BASELINE" not in action_input_rendered
    assert "PRIVATE-REFERENCED-LINK-CANDIDATE" not in action_input_rendered
    assert "formula external-action material redacted" in action_input_rendered

    dde_baseline = make_formula_dde_link_model(tmp_path / "dde-baseline.xlsx")
    dde_candidate = make_formula_dde_link_model(tmp_path / "dde-candidate.xlsx")
    change_formula_dde_link_input(dde_candidate)
    dde_output = tmp_path / "dde-redacted.json"
    assert (
        main(
            [
                "diff",
                str(dde_baseline),
                str(dde_candidate),
                "--format",
                "json",
                "--redact-formula-external-actions",
                "--output",
                str(dde_output),
            ]
        )
        == 0
    )
    dde_rendered = dde_output.read_text(encoding="utf-8")
    assert "PRIVATE-DDE-INPUT-BASELINE" not in dde_rendered
    assert "PRIVATE-DDE-INPUT-CANDIDATE" not in dde_rendered
    assert "FF074" in dde_rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_formula_external_action_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-formula-external-actions",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF064" in policy_rendered
    assert "FFP064" in policy_rendered
    assert action_baseline_marker not in policy_rendered
    assert action_candidate_marker not in policy_rendered

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_formula_external_action_model(baseline_directory / "model.xlsx")
    portfolio_candidate = make_formula_external_action_model(
        candidate_directory / "model.xlsx"
    )
    change_formula_external_action_target(portfolio_candidate)
    portfolio_output = tmp_path / "portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-formula-external-actions",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert action_baseline_marker not in portfolio_rendered
    assert action_candidate_marker not in portfolio_rendered
    assert "formula external-action material redacted" in portfolio_rendered


def test_cli_can_redact_python_in_excel_material_from_shared_reports(tmp_path) -> None:
    source_baseline = "PRIVATE-PY-SOURCE-BASELINE"
    source_candidate = "PRIVATE-PY-SOURCE-CANDIDATE"
    input_baseline = "PRIVATE-PY-INPUT-BASELINE"
    input_candidate = "PRIVATE-PY-INPUT-CANDIDATE"

    def python_source_model(path, source: str, input_value: str):
        workbook = make_python_in_excel_model(path, input_value=input_value)
        return set_python_in_excel_formula_source(workbook, source)

    baseline = python_source_model(
        tmp_path / "baseline.xlsx", source_baseline, input_baseline
    )
    candidate = python_source_model(
        tmp_path / "candidate.xlsx", source_candidate, input_candidate
    )
    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    for sensitive_value in (
        source_baseline,
        source_candidate,
        input_baseline,
        input_candidate,
    ):
        assert sensitive_value in default_rendered

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"python-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-python-in-excel",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        for sensitive_value in (
            source_baseline,
            source_candidate,
            input_baseline,
            input_candidate,
        ):
            assert sensitive_value not in rendered
        assert "FF065" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert "Python-in-Excel material:** redacted for sharing" in rendered
        elif report_format == "json":
            assert "Python-in-Excel material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_python_in_excel_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "python-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-python-in-excel",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF065" in policy_rendered
    assert "FFP065" in policy_rendered
    for sensitive_value in (
        source_baseline,
        source_candidate,
        input_baseline,
        input_candidate,
    ):
        assert sensitive_value not in policy_rendered

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    python_source_model(
        baseline_directory / "model.xlsx", source_baseline, input_baseline
    )
    python_source_model(
        candidate_directory / "model.xlsx", source_candidate, input_candidate
    )
    portfolio_output = tmp_path / "python-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-python-in-excel",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "Python-in-Excel material redacted" in portfolio_rendered
    for sensitive_value in (
        source_baseline,
        source_candidate,
        input_baseline,
        input_candidate,
    ):
        assert sensitive_value not in portfolio_rendered


def test_cli_can_redact_office_custom_function_material_from_shared_reports(
    tmp_path,
) -> None:
    baseline = make_office_custom_function_model(tmp_path / "baseline.xlsx")
    candidate = make_office_custom_function_model(tmp_path / "candidate.xlsx")
    change_office_custom_function_input(candidate)
    change_office_custom_function_call(candidate)
    sensitive_values = (
        "CONTOSO",
        "GETMARKETDATA",
        "GETRISKDATA",
        "PRIVATE-CUSTOM-FUNCTION-QUERY-BASELINE",
        "PRIVATE-CUSTOM-FUNCTION-INPUT-BASELINE",
        "PRIVATE-CUSTOM-FUNCTION-INPUT-CANDIDATE",
    )

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert all(value in default_rendered for value in sensitive_values)

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"custom-function-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-office-custom-functions",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert all(value not in rendered for value in sensitive_values)
        assert "FF066" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert "Office custom-function material:** redacted for sharing" in rendered
        elif report_format == "json":
            assert "Office custom-function material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_office_custom_function_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "custom-function-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-office-custom-functions",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF066" in policy_rendered
    assert "FFP066" in policy_rendered
    assert all(value not in policy_rendered for value in sensitive_values)

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_office_custom_function_model(baseline_directory / "model.xlsx")
    portfolio_candidate = make_office_custom_function_model(
        candidate_directory / "model.xlsx"
    )
    change_office_custom_function_input(portfolio_candidate)
    portfolio_output = tmp_path / "custom-function-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-office-custom-functions",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "Office custom-function material redacted" in portfolio_rendered
    assert all(value not in portfolio_rendered for value in sensitive_values)


def test_cli_can_redact_unqualified_runtime_function_material_from_shared_reports(
    tmp_path,
) -> None:
    baseline = make_unqualified_runtime_function_model(tmp_path / "baseline.xlsx")
    candidate = make_unqualified_runtime_function_model(tmp_path / "candidate.xlsx")
    change_unqualified_runtime_function_input(candidate)
    change_unqualified_runtime_function_call(candidate)
    sensitive_values = (
        "PRIVATEUDF",
        "UPDATEDUDF",
        "PRIVATE-RUNTIME-FUNCTION-QUERY-BASELINE",
        "PRIVATE-RUNTIME-FUNCTION-INPUT-BASELINE",
        "PRIVATE-RUNTIME-FUNCTION-INPUT-CANDIDATE",
    )

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert all(value in default_rendered for value in sensitive_values)

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"runtime-function-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-unqualified-runtime-functions",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert all(value not in rendered for value in sensitive_values)
        assert "FF075" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert "Unqualified runtime-function material:** redacted for sharing" in rendered
        elif report_format == "json":
            assert "unqualified runtime-function material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_unqualified_runtime_function_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "runtime-function-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-unqualified-runtime-functions",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF075" in policy_rendered
    assert "FFP075" in policy_rendered
    assert all(value not in policy_rendered for value in sensitive_values)

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_unqualified_runtime_function_model(baseline_directory / "model.xlsx")
    portfolio_candidate = make_unqualified_runtime_function_model(
        candidate_directory / "model.xlsx"
    )
    change_unqualified_runtime_function_input(portfolio_candidate)
    change_unqualified_runtime_function_call(portfolio_candidate)
    portfolio_output = tmp_path / "runtime-function-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-unqualified-runtime-functions",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "unqualified runtime-function material redacted" in portfolio_rendered
    assert all(value not in portfolio_rendered for value in sensitive_values)


def test_cli_can_redact_worksheet_code_resource_registration_material(
    tmp_path,
) -> None:
    baseline = make_worksheet_code_resource_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_worksheet_code_resource_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_worksheet_code_resource_registration_input(candidate)
    change_worksheet_code_resource_registration_call(candidate)
    sensitive_values = (
        "PRIVATE-REGISTRATION-MODULE-BASELINE",
        "PRIVATE-REGISTRATION-MODULE-CANDIDATE",
        "PRIVATE-REGISTRATION-MODULE-LITERAL-BASELINE",
        "PRIVATE-REGISTRATION-PROCEDURE-LITERAL-BASELINE",
        "PRIVATE-REGISTRATION-MODULE-LITERAL-CANDIDATE",
        "PRIVATE-REGISTRATION-PROCEDURE-LITERAL-CANDIDATE",
    )

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert all(value in default_rendered for value in sensitive_values)

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"registration-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-worksheet-code-resource-registrations",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert all(value not in rendered for value in sensitive_values)
        assert "FF067" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert (
                "Worksheet code-resource registration material:** redacted for sharing"
                in rendered
            )
        elif report_format == "json":
            assert "worksheet code-resource registration material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_worksheet_code_resource_registration_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "registration-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-worksheet-code-resource-registrations",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF067" in policy_rendered
    assert "FFP067" in policy_rendered
    assert all(value not in policy_rendered for value in sensitive_values)

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_worksheet_code_resource_registration_model(
        baseline_directory / "model.xlsx"
    )
    portfolio_candidate = make_worksheet_code_resource_registration_model(
        candidate_directory / "model.xlsx"
    )
    change_worksheet_code_resource_registration_input(portfolio_candidate)
    change_worksheet_code_resource_registration_call(portfolio_candidate)
    portfolio_output = tmp_path / "registration-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-worksheet-code-resource-registrations",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "worksheet code-resource registration material redacted" in portfolio_rendered
    assert all(value not in portfolio_rendered for value in sensitive_values)


def test_cli_can_redact_formula_defined_xlm_registration_material(tmp_path) -> None:
    baseline = make_formula_defined_xlm_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_registration_input(candidate)
    change_formula_defined_xlm_registration_call(candidate)
    sensitive_values = (
        "PRIVATE-XLM-REGISTRATION-MODULE-BASELINE",
        "PRIVATE-XLM-REGISTRATION-MODULE-CANDIDATE",
        "PRIVATE-XLM-REGISTRATION-LITERAL-PROCEDURE-BASELINE",
        "PRIVATE-XLM-REGISTRATION-LITERAL-PROCEDURE-CANDIDATE",
        "PRIVATE-XLM-REGISTRATION-LITERAL-TYPE-BASELINE",
        "PRIVATE-XLM-REGISTRATION-LITERAL-TYPE-CANDIDATE",
    )

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert all(value in default_rendered for value in sensitive_values)

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"xlm-registration-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-formula-defined-xlm-registrations",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert all(value not in rendered for value in sensitive_values)
        assert "FF068" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert (
                "Formula-defined XLM registration material:** redacted for sharing"
                in rendered
            )
        elif report_format == "json":
            assert "formula-defined XLM registration material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_formula_defined_xlm_registration_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "xlm-registration-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-formula-defined-xlm-registrations",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF068" in policy_rendered
    assert "FFP068" in policy_rendered
    assert all(value not in policy_rendered for value in sensitive_values)

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_formula_defined_xlm_registration_model(baseline_directory / "model.xlsx")
    portfolio_candidate = make_formula_defined_xlm_registration_model(
        candidate_directory / "model.xlsx"
    )
    change_formula_defined_xlm_registration_input(portfolio_candidate)
    change_formula_defined_xlm_registration_call(portfolio_candidate)
    portfolio_output = tmp_path / "xlm-registration-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-formula-defined-xlm-registrations",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "formula-defined XLM registration material redacted" in portfolio_rendered
    assert all(value not in portfolio_rendered for value in sensitive_values)


def test_cli_can_redact_formula_defined_xlm_evaluation_material(tmp_path) -> None:
    baseline = make_formula_defined_xlm_evaluation_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_evaluation_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_evaluation_input(candidate)
    change_formula_defined_xlm_evaluation_call(candidate)
    sensitive_values = (
        "PRIVATE-XLM-EVALUATE-EXPRESSION-BASELINE",
        "PRIVATE-XLM-EVALUATE-EXPRESSION-CANDIDATE",
        "PRIVATE-XLM-EVALUATE-LITERAL-BASELINE",
        "PRIVATE-XLM-EVALUATE-LITERAL-CANDIDATE",
    )

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert all(value in default_rendered for value in sensitive_values)

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"xlm-evaluation-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-formula-defined-xlm-evaluations",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert all(value not in rendered for value in sensitive_values)
        assert "FF069" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert (
                "Formula-defined XLM evaluation material:** redacted for sharing"
                in rendered
            )
        elif report_format == "json":
            assert "formula-defined XLM evaluation material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_formula_defined_xlm_evaluation_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "xlm-evaluation-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-formula-defined-xlm-evaluations",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF069" in policy_rendered
    assert "FFP069" in policy_rendered
    assert all(value not in policy_rendered for value in sensitive_values)

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_formula_defined_xlm_evaluation_model(baseline_directory / "model.xlsx")
    portfolio_candidate = make_formula_defined_xlm_evaluation_model(
        candidate_directory / "model.xlsx"
    )
    change_formula_defined_xlm_evaluation_input(portfolio_candidate)
    change_formula_defined_xlm_evaluation_call(portfolio_candidate)
    portfolio_output = tmp_path / "xlm-evaluation-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-formula-defined-xlm-evaluations",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "formula-defined XLM evaluation material redacted" in portfolio_rendered
    assert all(value not in portfolio_rendered for value in sensitive_values)


def test_cli_can_redact_formula_defined_xlm_action_material(tmp_path) -> None:
    baseline = make_formula_defined_xlm_action_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_defined_xlm_action_model(tmp_path / "candidate.xlsx")
    change_formula_defined_xlm_action_input(candidate)
    change_formula_defined_xlm_action_call(candidate)
    sensitive_values = (
        "PRIVATE-XLM-ACTION-INPUT-BASELINE",
        "PRIVATE-XLM-ACTION-INPUT-CANDIDATE",
        "PRIVATE-XLM-ACTION-EVENT",
        "PRIVATE-XLM-ACTION-EVENT-CANDIDATE",
    )

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert all(value in default_rendered for value in sensitive_values)

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"xlm-action-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-formula-defined-xlm-actions",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert all(value not in rendered for value in sensitive_values)
        assert "FF073" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert "Formula-defined XLM action material:** redacted for sharing" in rendered
        elif report_format == "json":
            assert "formula-defined XLM action material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_formula_defined_xlm_action_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "xlm-action-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-formula-defined-xlm-actions",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF073" in policy_rendered
    assert "FFP073" in policy_rendered
    assert all(value not in policy_rendered for value in sensitive_values)

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_formula_defined_xlm_action_model(baseline_directory / "model.xlsx")
    portfolio_candidate = make_formula_defined_xlm_action_model(
        candidate_directory / "model.xlsx"
    )
    change_formula_defined_xlm_action_input(portfolio_candidate)
    change_formula_defined_xlm_action_call(portfolio_candidate)
    portfolio_output = tmp_path / "xlm-action-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-formula-defined-xlm-actions",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "formula-defined XLM action material redacted" in portfolio_rendered
    assert all(value not in portfolio_rendered for value in sensitive_values)


def test_cli_can_redact_formula_defined_xlm_get_cell_material(tmp_path) -> None:
    baseline = make_formula_defined_xlm_get_cell_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_defined_xlm_get_cell_model(tmp_path / "candidate.xlsx")
    change_formula_defined_xlm_get_cell_input(candidate)
    change_formula_defined_xlm_get_cell_call(candidate)
    sensitive_values = (
        "PRIVATE-XLM-GET-CELL-INPUT-BASELINE",
        "PRIVATE-XLM-GET-CELL-INPUT-CANDIDATE",
        "GET.CELL(53,Inputs!$A$9)",
        "GET.CELL(54,Inputs!$A$9)",
    )

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert all(value in default_rendered for value in sensitive_values)

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"xlm-get-cell-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-formula-defined-xlm-get-cell-calls",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert all(value not in rendered for value in sensitive_values)
        assert "FF070" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert (
                "Formula-defined XLM GET.CELL material:** redacted for sharing"
                in rendered
            )
        elif report_format == "json":
            assert "formula-defined XLM GET.CELL material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_formula_defined_xlm_get_cell_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "xlm-get-cell-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-formula-defined-xlm-get-cell-calls",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF070" in policy_rendered
    assert "FFP070" in policy_rendered
    assert all(value not in policy_rendered for value in sensitive_values)

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_formula_defined_xlm_get_cell_model(baseline_directory / "model.xlsx")
    portfolio_candidate = make_formula_defined_xlm_get_cell_model(
        candidate_directory / "model.xlsx"
    )
    change_formula_defined_xlm_get_cell_input(portfolio_candidate)
    change_formula_defined_xlm_get_cell_call(portfolio_candidate)
    portfolio_output = tmp_path / "xlm-get-cell-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-formula-defined-xlm-get-cell-calls",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "formula-defined XLM GET.CELL material redacted" in portfolio_rendered
    assert all(value not in portfolio_rendered for value in sensitive_values)


def test_cli_can_redact_formula_defined_xlm_environment_information_material(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_environment_information_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_environment_information_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_environment_information_input(candidate)
    change_formula_defined_xlm_environment_information_call(candidate)
    sensitive_values = (
        "PRIVATE-XLM-ENVIRONMENT-INPUT-BASELINE",
        "PRIVATE-XLM-ENVIRONMENT-INPUT-CANDIDATE",
        "GET.WORKSPACE(2)",
        "GET.WORKSPACE(3)",
    )

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert all(value in default_rendered for value in sensitive_values)

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"xlm-environment-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-formula-defined-xlm-environment-information-calls",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert all(value not in rendered for value in sensitive_values)
        assert "FF071" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert (
                "Formula-defined XLM environment-information material:** redacted for "
                "sharing"
            ) in rendered
        elif report_format == "json":
            assert "formula-defined XLM environment-information material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_formula_defined_xlm_environment_information_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "xlm-environment-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-formula-defined-xlm-environment-information-calls",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF071" in policy_rendered
    assert "FFP071" in policy_rendered
    assert all(value not in policy_rendered for value in sensitive_values)

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_formula_defined_xlm_environment_information_model(
        baseline_directory / "model.xlsx"
    )
    portfolio_candidate = make_formula_defined_xlm_environment_information_model(
        candidate_directory / "model.xlsx"
    )
    change_formula_defined_xlm_environment_information_input(portfolio_candidate)
    change_formula_defined_xlm_environment_information_call(portfolio_candidate)
    portfolio_output = tmp_path / "xlm-environment-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-formula-defined-xlm-environment-information-calls",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "formula-defined XLM environment-information material redacted" in portfolio_rendered
    assert all(value not in portfolio_rendered for value in sensitive_values)


def test_cli_can_redact_formula_environment_information_material(tmp_path) -> None:
    baseline = make_formula_environment_information_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_environment_information_model(tmp_path / "candidate.xlsx")
    change_formula_environment_information_input(candidate)
    change_formula_environment_information_definition(candidate)
    sensitive_values = (
        "PRIVATE-NATIVE-ENVIRONMENT-INPUT-BASELINE",
        "PRIVATE-NATIVE-ENVIRONMENT-INPUT-CANDIDATE",
        "system",
        "osversion",
    )

    default_json = tmp_path / "default.json"
    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(default_json),
            ]
        )
        == 0
    )
    default_rendered = default_json.read_text(encoding="utf-8")
    assert all(value in default_rendered for value in sensitive_values)

    for report_format, suffix in (
        ("json", "json"),
        ("markdown", "md"),
        ("html", "html"),
        ("sarif", "sarif"),
    ):
        output = tmp_path / f"native-environment-redacted.{suffix}"
        assert (
            main(
                [
                    "diff",
                    str(baseline),
                    str(candidate),
                    "--format",
                    report_format,
                    "--redact-formula-environment-information",
                    "--output",
                    str(output),
                ]
            )
            == 0
        )
        rendered = output.read_text(encoding="utf-8")
        assert all(value not in rendered for value in sensitive_values)
        assert "FF072" in rendered
        if report_format == "html":
            _assert_html_review_artifact(rendered)
        elif report_format == "markdown":
            assert "Formula environment-information material:** redacted for sharing" in rendered
        elif report_format == "json":
            assert "formula environment-information material redacted" in rendered

    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_formula_environment_information_changes: true\n",
        encoding="utf-8",
    )
    policy_output = tmp_path / "native-environment-policy-redacted.json"
    assert (
        main(
            [
                "check",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy),
                "--format",
                "json",
                "--redact-formula-environment-information",
                "--output",
                str(policy_output),
            ]
        )
        == 1
    )
    policy_rendered = policy_output.read_text(encoding="utf-8")
    assert "FF072" in policy_rendered
    assert "FFP072" in policy_rendered
    assert all(value not in policy_rendered for value in sensitive_values)

    baseline_directory = tmp_path / "baseline-portfolio"
    candidate_directory = tmp_path / "candidate-portfolio"
    baseline_directory.mkdir()
    candidate_directory.mkdir()
    make_formula_environment_information_model(baseline_directory / "model.xlsx")
    portfolio_candidate = make_formula_environment_information_model(
        candidate_directory / "model.xlsx"
    )
    change_formula_environment_information_input(portfolio_candidate)
    change_formula_environment_information_definition(portfolio_candidate)
    portfolio_output = tmp_path / "native-environment-portfolio-redacted.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "json",
                "--redact-formula-environment-information",
                "--output",
                str(portfolio_output),
            ]
        )
        == 0
    )
    portfolio_rendered = portfolio_output.read_text(encoding="utf-8")
    assert "formula environment-information material redacted" in portfolio_rendered
    assert all(value not in portfolio_rendered for value in sensitive_values)

    portfolio_html_output = tmp_path / "native-environment-portfolio-redacted.html"
    assert (
        main(
            [
                "portfolio",
                str(baseline_directory),
                str(candidate_directory),
                "--format",
                "html",
                "--redact-formula-environment-information",
                "--output",
                str(portfolio_html_output),
            ]
        )
        == 0
    )
    portfolio_html = portfolio_html_output.read_text(encoding="utf-8")
    _assert_html_review_artifact(portfolio_html)
    assert "FF072" in portfolio_html
    assert "Formula environment-information material redacted for sharing" in portfolio_html
    assert all(value not in portfolio_html for value in sensitive_values)


def test_cli_refuses_to_overwrite_an_input_workbook(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    original = baseline.read_bytes()

    assert (
        main(
            [
                "diff",
                str(baseline),
                str(candidate),
                "--output",
                str(baseline),
            ]
        )
        == 2
    )
    assert baseline.read_bytes() == original


def test_init_policy_includes_modern_formula_coverage_controls(tmp_path) -> None:
    policy = tmp_path / "formulafence.yml"

    assert main(["init", str(policy)]) == 0

    content = policy.read_text(encoding="utf-8")
    assert "no_new_spill_references: true" in content
    assert "no_new_dynamic_array_output_references: true" in content
    assert "no_new_implicit_intersections: true" in content
    assert "no_array_formula_semantics_changes: true" in content
    assert "no_data_validation_changes: true" in content
    assert "no_conditional_formatting_changes: true" in content
    assert "no_protection_changes: true" in content
    assert "no_external_data_connection_changes: true" in content
    assert "no_external_link_package_changes: true" in content
    assert "no_external_workbook_link_surface_changes: true" in content
    assert "no_external_relationship_changes: true" in content
    assert "no_formula_external_action_changes: true" in content
    assert "no_formula_dde_link_changes: true" in content
    assert "no_python_in_excel_changes: true" in content
    assert "no_office_custom_function_changes: true" in content
    assert "no_unqualified_runtime_function_changes: true" in content
    assert "no_worksheet_code_resource_registration_changes: true" in content
    assert "no_formula_defined_xlm_registration_changes: true" in content
    assert "no_formula_defined_xlm_evaluation_changes: true" in content
    assert "no_formula_defined_xlm_action_changes: true" in content
    assert "no_formula_defined_xlm_get_cell_changes: true" in content
    assert "no_formula_defined_xlm_environment_information_changes: true" in content
    assert "no_formula_environment_information_changes: true" in content
    assert "no_xlm_macro_sheet_changes: true" in content
    assert "no_xlm_automatic_macro_binding_changes: true" in content
    assert "no_ribbon_customization_changes: true" in content
    assert "no_office_web_addin_changes: true" in content
    assert "no_pivot_table_definition_changes: true" in content
    assert "no_slicer_timeline_cache_changes: true" in content
    assert "no_power_pivot_data_model_changes: true" in content
    assert "no_what_if_data_table_changes: true" in content
    assert "no_scenario_manager_changes: true" in content
    assert "no_filter_visibility_changes: true" in content
    assert "no_ignored_error_changes: true" in content
    assert "no_named_sheet_view_changes: true" in content
    assert "no_number_format_changes: true" in content
    assert "no_cell_font_changes: true" in content
    assert "no_cell_fill_changes: true" in content
    assert "no_workbook_theme_changes: true" in content
    assert "no_cell_alignment_changes: true" in content
    assert "no_cell_border_changes: true" in content
    assert "no_worksheet_dimension_changes: true" in content
    assert "no_worksheet_display_control_changes: true" in content
    assert "no_worksheet_print_layout_changes: true" in content
    assert "no_formula_cached_result_changes: true" in content
    assert "no_rich_text_run_changes: true" in content
    assert "no_cell_hyperlink_changes: true" in content
    assert "no_worksheet_sparkline_changes: true" in content
    assert "no_xml_mapping_changes: true" in content
    assert "no_digital_signature_changes: true" in content
    assert "no_rich_data_changes: true" in content
    assert "no_custom_data_store_changes: true" in content
    assert "no_legacy_comment_changes: true" in content
    assert "no_threaded_comment_changes: true" in content
    assert "no_worksheet_drawing_shape_changes: true" in content
    assert "no_worksheet_image_changes: true" in content
    assert "no_worksheet_embedded_control_changes: true" in content
    assert "no_power_query_changes: true" in content
    assert "no_portfolio_membership_changes: true" in content
    assert "no_cross_workbook_impacts: true" in content
    assert "no_new_tokenization_failures: true" in content
