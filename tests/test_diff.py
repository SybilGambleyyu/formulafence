from __future__ import annotations

import json
import shutil
import warnings
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Protection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

import formulafence.workbook as workbook_module
from formulafence.diff import compare_snapshots
from formulafence.output import (
    EXTERNAL_WORKBOOK_LINK_REDACTION,
    FORMULA_DEFINED_XLM_ACTION_REDACTION,
    FORMULA_DEFINED_XLM_ENVIRONMENT_INFORMATION_REDACTION,
    FORMULA_DEFINED_XLM_EVALUATION_REDACTION,
    FORMULA_DEFINED_XLM_GET_CELL_REDACTION,
    FORMULA_DEFINED_XLM_REGISTRATION_REDACTION,
    FORMULA_ENVIRONMENT_INFORMATION_REDACTION,
    FORMULA_EXTERNAL_ACTION_REDACTION,
    OFFICE_CUSTOM_FUNCTION_REDACTION,
    PYTHON_IN_EXCEL_REDACTION,
    UNQUALIFIED_RUNTIME_FUNCTION_REDACTION,
    WORKSHEET_CODE_RESOURCE_REGISTRATION_REDACTION,
    profile_to_markdown,
    redact_external_workbook_link_material,
    redact_formula_defined_xlm_action_material,
    redact_formula_defined_xlm_action_report_payload,
    redact_formula_defined_xlm_environment_information_material,
    redact_formula_defined_xlm_environment_information_report_payload,
    redact_formula_defined_xlm_evaluation_material,
    redact_formula_defined_xlm_evaluation_report_payload,
    redact_formula_defined_xlm_get_cell_material,
    redact_formula_defined_xlm_get_cell_report_payload,
    redact_formula_defined_xlm_registration_material,
    redact_formula_defined_xlm_registration_report_payload,
    redact_formula_environment_information_material,
    redact_formula_environment_information_report_payload,
    redact_formula_external_action_material,
    redact_formula_external_action_report_payload,
    redact_office_custom_function_material,
    redact_office_custom_function_report_payload,
    redact_python_in_excel_material,
    redact_python_in_excel_report_payload,
    redact_unqualified_runtime_function_material,
    redact_unqualified_runtime_function_report_payload,
    redact_worksheet_code_resource_registration_material,
    redact_worksheet_code_resource_registration_report_payload,
    report_to_html,
    report_to_markdown,
    report_to_sarif,
)
from formulafence.workbook import load_snapshot, profile_snapshot

from .helpers import (
    add_conditional_formatting_databar_extension,
    add_ordinary_dimension_resize,
    add_power_pivot_data_model_direct_relationship,
    add_protected_range,
    add_python_in_excel_scripts_compatibility_part,
    add_unsupported_extended_chart_relationship,
    add_worksheet_dimension_baseline_adjustments,
    add_worksheet_smartart_component_relationship,
    add_xlm_automatic_macro_binding,
    break_slicer_timeline_pivot_cache_binding,
    change_alignment_definition,
    change_border_definition,
    change_border_logical_start_side,
    change_cell_hyperlink_display,
    change_cell_hyperlink_location,
    change_cell_hyperlink_target,
    change_cell_hyperlink_tooltip,
    change_chart_cached_data,
    change_chart_definition_material,
    change_chart_sheet_custom_workbook_view_scale,
    change_custom_data_payload,
    change_custom_data_store_storage_identifiers,
    change_custom_document_property_value,
    change_custom_workbook_view_filter,
    change_custom_xml_data_store_value,
    change_default_alignment_definition,
    change_default_border_definition,
    change_default_fill_definition,
    change_default_font_definition,
    change_default_zero_dimension_visibility_controls,
    change_extended_chart_definition_material,
    change_extended_chart_style_payload,
    change_external_data_refresh_controls,
    change_external_link_package_controls,
    change_external_relationship_target,
    change_fill_definition,
    change_filter_visibility_criterion,
    change_filter_visibility_hidden_column,
    change_filter_visibility_hidden_row,
    change_font_definition,
    change_formula_cached_result,
    change_formula_cached_result_with_visible_precedent,
    change_formula_dde_link_definition,
    change_formula_dde_link_input,
    change_formula_defined_xlm_action_call,
    change_formula_defined_xlm_action_definition,
    change_formula_defined_xlm_action_input,
    change_formula_defined_xlm_environment_information_call,
    change_formula_defined_xlm_environment_information_definition,
    change_formula_defined_xlm_environment_information_input,
    change_formula_defined_xlm_evaluation_call,
    change_formula_defined_xlm_evaluation_definition,
    change_formula_defined_xlm_evaluation_input,
    change_formula_defined_xlm_get_cell_call,
    change_formula_defined_xlm_get_cell_definition,
    change_formula_defined_xlm_get_cell_input,
    change_formula_defined_xlm_registration_call,
    change_formula_defined_xlm_registration_definition,
    change_formula_defined_xlm_registration_input,
    change_formula_environment_information_definition,
    change_formula_environment_information_input,
    change_formula_external_action_input,
    change_formula_external_action_target,
    change_formula_external_data_provider_definition,
    change_formula_external_data_provider_input,
    change_formula_external_data_provider_target,
    change_formula_workbook_structure_information_definition,
    change_formula_workbook_structure_information_input,
    change_gradient_fill_definition,
    change_ignored_error_extension_target,
    change_ignored_error_target,
    change_in_content_office_web_addin_anchor,
    change_in_content_office_web_addin_preview_payload,
    change_indirect_named_office_custom_function_definition,
    change_indirect_named_unqualified_runtime_function_definition,
    change_indirect_named_worksheet_code_resource_registration_definition,
    change_inline_rich_text_run_color,
    change_legacy_comment_text,
    change_legacy_note_visibility,
    change_legacy_placeholder_author_context,
    change_legacy_vml_control_controls,
    change_legacy_vml_note,
    change_named_formula_external_action_definition,
    change_named_formula_external_action_input,
    change_named_office_custom_function_definition,
    change_named_office_custom_function_input,
    change_named_sheet_view_criterion,
    change_named_unqualified_runtime_function_definition,
    change_named_unqualified_runtime_function_input,
    change_named_worksheet_code_resource_registration_definition,
    change_named_worksheet_code_resource_registration_input,
    change_number_format_code,
    change_number_format_default_style,
    change_office_custom_function_call,
    change_office_custom_function_input,
    change_office_web_addin_auto_show,
    change_office_web_addin_controls,
    change_office_web_addin_worksheet_binding,
    change_package_signature_certificate_payload,
    change_package_signature_reference,
    change_pivot_table_definition_material,
    change_pivot_table_refresh_control,
    change_power_pivot_data_model_declaration,
    change_power_pivot_data_model_payload,
    change_power_query_controls,
    change_power_query_refresh_noise,
    change_protected_range,
    change_python_in_excel_formula_binding,
    change_python_in_excel_input,
    change_python_in_excel_script,
    change_python_in_excel_scripts_script,
    change_ribbon_customization_callback,
    change_ribbon_customization_controls,
    change_rich_data_binding,
    change_rich_data_value,
    change_rich_data_web_image_target,
    change_rich_text_run_boundary,
    change_rich_text_run_color,
    change_rich_text_run_text_only,
    change_scenario_manager_input_value,
    change_shared_workbook_revision_controls,
    change_shared_workbook_revision_log,
    change_slicer_timeline_filter_material,
    change_strict_worksheet_display_controls,
    change_strict_worksheet_print_layout_controls,
    change_table_direct_dxf_assignment,
    change_table_filter_visibility_criterion,
    change_table_style_control,
    change_table_style_definition,
    change_threaded_comment_person_identity,
    change_threaded_comment_reply,
    change_unqualified_runtime_function_call,
    change_unqualified_runtime_function_input,
    change_vba_project_signature_payload,
    change_what_if_data_table_input,
    change_workbook_theme_colour,
    change_workbook_theme_image_payload,
    change_worksheet_code_resource_registration_call,
    change_worksheet_code_resource_registration_input,
    change_worksheet_dimension_controls,
    change_worksheet_display_controls,
    change_worksheet_drawing_connector_attachment,
    change_worksheet_drawing_connector_presentation,
    change_worksheet_drawing_shape_hyperlink,
    change_worksheet_drawing_shape_presentation,
    change_worksheet_embedded_control_controls,
    change_worksheet_embedded_control_payload,
    change_worksheet_image_payload,
    change_worksheet_image_presentation,
    change_worksheet_print_layout_controls,
    change_worksheet_smartart_data,
    change_worksheet_smartart_diagram_image_payload,
    change_worksheet_smartart_graphic_frame_uri,
    change_worksheet_sparkline_presentation,
    change_worksheet_sparkline_source,
    change_xlm_macro_sheet_controls,
    change_xlm_macro_sheet_related_part_payload,
    change_xml_mapping_refresh_behavior,
    change_xml_mapping_xpath,
    change_zero_dimension_visibility_controls,
    corrupt_alignment_column_control,
    corrupt_alignment_definition,
    corrupt_border_column_control,
    corrupt_border_definition,
    corrupt_cell_hyperlink_reference,
    corrupt_chart_definition_root,
    corrupt_chart_sheet_custom_workbook_view_page_setup,
    corrupt_custom_data_properties_root,
    corrupt_custom_workbook_view_control,
    corrupt_default_zero_dimension_visibility_controls,
    corrupt_extended_chart_definition_root,
    corrupt_external_relationship_metadata,
    corrupt_fill_column_control,
    corrupt_fill_definition,
    corrupt_filter_visibility_column_control,
    corrupt_filter_visibility_control,
    corrupt_font_column_control,
    corrupt_font_definition,
    corrupt_formula_cached_result,
    corrupt_ignored_error_control,
    corrupt_in_content_office_web_addin_reference,
    corrupt_legacy_comment_root,
    corrupt_legacy_vml_control_root,
    corrupt_named_sheet_view_control,
    corrupt_number_format_column_control,
    corrupt_number_format_definition,
    corrupt_office_web_addin_definition_root,
    corrupt_office_web_addin_worksheet_binding,
    corrupt_package_signature_root,
    corrupt_pivot_table_definition_root,
    corrupt_python_in_excel_package,
    corrupt_python_in_excel_scripts_package,
    corrupt_ribbon_customization_root,
    corrupt_rich_data_value_root,
    corrupt_rich_text_run,
    corrupt_scenario_manager_input,
    corrupt_shared_workbook_revision,
    corrupt_slicer_timeline_cache_root,
    corrupt_table_style_control,
    corrupt_threaded_comment_root,
    corrupt_what_if_data_table_input,
    corrupt_workbook_theme_root,
    corrupt_worksheet_dimension_control,
    corrupt_worksheet_display_control,
    corrupt_worksheet_drawing_connector_attachment,
    corrupt_worksheet_drawing_shape_root,
    corrupt_worksheet_embedded_control_activex_root,
    corrupt_worksheet_image_drawing_root,
    corrupt_worksheet_print_layout_control,
    corrupt_worksheet_smartart_component_root,
    corrupt_worksheet_smartart_relationships,
    corrupt_worksheet_sparkline_destination,
    corrupt_xlm_macro_sheet_root,
    corrupt_xml_mapping_single_cell_reference,
    corrupt_zero_dimension_visibility_controls,
    delete_what_if_data_table_input,
    duplicate_external_link_definition,
    duplicate_external_link_sheet_names,
    duplicate_ignored_error_container,
    duplicate_xlm_macro_sheet_workbook_relationship,
    externalize_chart_overlay_relationship,
    externalize_extended_chart_relationship,
    externalize_legacy_comment_relationship,
    externalize_legacy_note_vml_relationship,
    externalize_package_signature_relationship,
    externalize_pivot_table_cache_record_relationship,
    externalize_power_pivot_data_model,
    externalize_slicer_timeline_cache_relationship,
    externalize_threaded_comment_relationship,
    externalize_worksheet_image_relationship,
    externalize_worksheet_smartart_diagram_image_relationship,
    externalize_xml_mapping_relationship,
    lowercase_legacy_threaded_placeholder_identifiers,
    make_alignment_model,
    make_border_model,
    make_cell_hyperlink_model,
    make_cell_hyperlink_sparkline_model,
    make_chart_definition_model,
    make_chart_sheet_custom_workbook_view_model,
    make_conditional_formatting_model,
    make_current_row_table_model,
    make_custom_data_store_model,
    make_custom_workbook_view_model,
    make_data_validation_model,
    make_digital_signature_model,
    make_empty_chart_sheet_custom_view_container_model,
    make_extended_chart_definition_model,
    make_external_data_refresh_model,
    make_external_link_package_model,
    make_external_relationship_model,
    make_fill_model,
    make_filter_visibility_model,
    make_font_model,
    make_formula_cached_result_model,
    make_formula_dde_link_model,
    make_formula_defined_xlm_action_model,
    make_formula_defined_xlm_environment_information_model,
    make_formula_defined_xlm_evaluation_model,
    make_formula_defined_xlm_get_cell_model,
    make_formula_defined_xlm_registration_model,
    make_formula_environment_information_model,
    make_formula_external_action_model,
    make_formula_external_data_provider_model,
    make_formula_workbook_structure_information_model,
    make_ignored_error_model,
    make_implicit_intersection_model,
    make_in_content_office_web_addin_model,
    make_indirect_named_office_custom_function_model,
    make_indirect_named_unqualified_runtime_function_model,
    make_indirect_named_worksheet_code_resource_registration_model,
    make_legacy_array_model,
    make_legacy_comment_model,
    make_legacy_threaded_placeholder_model,
    make_legacy_vml_control_model,
    make_legacy_vml_note_model,
    make_let_model,
    make_model,
    make_named_formula_external_action_model,
    make_named_formula_model,
    make_named_lambda_model,
    make_named_office_custom_function_model,
    make_named_sheet_view_model,
    make_named_unqualified_runtime_function_model,
    make_named_worksheet_code_resource_registration_model,
    make_number_format_model,
    make_office_custom_function_model,
    make_office_web_addin_model,
    make_pivot_table_definition_model,
    make_power_pivot_data_model,
    make_power_query_model,
    make_protection_model,
    make_python_in_excel_model,
    make_python_in_excel_scripts_model,
    make_ribbon_customization_model,
    make_rich_data_model,
    make_rich_text_run_model,
    make_scenario_manager_model,
    make_scoped_named_lambda_model,
    make_shared_workbook_revision_model,
    make_slicer_timeline_cache_model,
    make_spill_model,
    make_strict_border_model,
    make_strict_custom_workbook_view_model,
    make_strict_shared_workbook_revision_model,
    make_strict_table_style_control_model,
    make_strict_workbook_theme_image_model,
    make_strict_worksheet_dimension_model,
    make_strict_worksheet_display_model,
    make_strict_worksheet_drawing_connector_model,
    make_strict_worksheet_image_model,
    make_strict_worksheet_print_layout_model,
    make_strict_worksheet_smartart_image_model,
    make_strict_worksheet_smartart_model,
    make_table_model,
    make_table_style_control_model,
    make_threaded_comment_model,
    make_three_d_model,
    make_unqualified_runtime_function_model,
    make_what_if_data_table_model,
    make_workbook_theme_image_model,
    make_worksheet_code_resource_registration_model,
    make_worksheet_dimension_model,
    make_worksheet_display_model,
    make_worksheet_drawing_connector_model,
    make_worksheet_drawing_shape_model,
    make_worksheet_embedded_control_model,
    make_worksheet_image_model,
    make_worksheet_print_layout_model,
    make_worksheet_smartart_image_model,
    make_worksheet_smartart_model,
    make_worksheet_sparkline_model,
    make_xlm_macro_sheet_model,
    make_xml_mapping_model,
    make_zero_dimension_visibility_model,
    mark_array_formula_dynamic,
    mark_array_formula_unclassified,
    mismatch_python_in_excel_scripts_relationship_contract,
    normalize_alignment_control_spelling,
    normalize_alignment_inheritance,
    normalize_border_control_spelling,
    normalize_border_inert_declarations,
    normalize_border_inheritance,
    normalize_custom_data_store_identifiers,
    normalize_custom_workbook_view_identifiers,
    normalize_digital_signature_control_spelling,
    normalize_fill_control_spelling,
    normalize_fill_inert_pattern_declarations,
    normalize_fill_inheritance,
    normalize_filter_visibility_control_spelling,
    normalize_font_control_spelling,
    normalize_font_inheritance,
    normalize_formula_cached_result_spelling,
    normalize_ignored_error_control_spelling,
    normalize_named_sheet_view_control_spelling,
    normalize_number_format_control_spelling,
    normalize_number_format_inheritance,
    normalize_rich_data_relationship_ids,
    normalize_rich_text_run_property_spelling,
    normalize_scenario_manager_reference_spelling,
    normalize_shared_workbook_revision_writer_noise,
    normalize_table_style_control_writer_noise,
    normalize_what_if_data_table_reference_spelling,
    normalize_workbook_theme_relationship_identifiers,
    normalize_worksheet_dimension_control_spelling,
    normalize_worksheet_dimension_inert_declarations,
    normalize_worksheet_display_control_spelling,
    normalize_worksheet_print_layout_control_spelling,
    normalize_worksheet_print_layout_inert_controls,
    normalize_worksheet_sparkline_control_spelling,
    normalize_xml_mapping_control_spelling,
    normalize_zero_dimension_visibility_control_spelling,
    overlap_what_if_data_table_outputs,
    rebind_external_link_declaration,
    rebind_legacy_comment_relationship,
    rebind_package_signature_relationship,
    rebind_pivot_table_cache_records,
    rebind_power_pivot_data_model,
    rebind_slicer_timeline_cache,
    rebind_worksheet_image_source_relationship,
    rebind_xml_mapping_relationship,
    remove_extended_chart_direct_relationship_id,
    remove_power_pivot_data_model_workbook_binding,
    remove_worksheet_smartart_diagram_image_payload,
    remove_xlm_macro_sheet_related_part_payload,
    renumber_cell_hyperlink_identifiers,
    renumber_chart_relationships,
    renumber_extended_chart_relationships,
    renumber_external_link_declaration_relationships,
    renumber_external_relationship_identifiers,
    renumber_in_content_office_web_addin_identifiers,
    renumber_legacy_comment_identifiers,
    renumber_legacy_threaded_placeholder_identifiers,
    renumber_legacy_vml_control_relationships,
    renumber_office_web_addin_relationships,
    renumber_pivot_table_cache_id,
    renumber_pivot_table_relationships,
    renumber_power_pivot_data_model_relationship,
    renumber_python_in_excel_relationship_identifier,
    renumber_python_in_excel_scripts_relationship_identifier,
    renumber_ribbon_customization_relationships,
    renumber_slicer_timeline_pivot_cache_id,
    renumber_slicer_timeline_relationships,
    renumber_threaded_comment_identifiers,
    renumber_worksheet_drawing_connector_identifiers,
    renumber_worksheet_drawing_shape_identifiers,
    renumber_worksheet_embedded_control_relationships,
    renumber_worksheet_image_identifiers,
    renumber_worksheet_smartart_diagram_image_relationship,
    renumber_worksheet_smartart_identifiers,
    renumber_xlm_macro_sheet_relationships,
    reorder_conditional_differential_styles,
    reorder_worksheet_sparklines,
    rewrite,
    rewrite_cell_hyperlink_as_revision_declaration,
    rewrite_chart_internal_target_spelling,
    rewrite_legacy_vml_control_internal_target_spelling,
    rewrite_office_web_addin_internal_target_spelling,
    rewrite_pivot_table_internal_target_spelling,
    rewrite_power_pivot_data_model_internal_target_spelling,
    rewrite_ribbon_customization_internal_target_spelling,
    rewrite_shared_rich_text_as_inline,
    rewrite_slicer_timeline_internal_target_spelling,
    rewrite_worksheet_embedded_control_internal_target_spelling,
    rewrite_xlm_macro_sheet_internal_target_spelling,
    set_chart_formula_external_workbook_target,
    set_external_data_connection_defaults,
    set_power_pivot_data_model_equivalent_guids,
    set_python_in_excel_formula_source,
    set_sheet_protection_defaults,
    set_sheet_protection_modern_verifier,
    set_slicer_timeline_equivalent_defaults,
    unbind_cell_hyperlink_relationship,
    unbind_custom_workbook_view,
    use_slicer_timeline_2011_relationship_type,
)


def test_formula_to_value_traces_cross_sheet_downstream_impact(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Model"]["B2"], "value", 200))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change = next(change for change in report.changes if change.location == ("Model", "B2"))

    assert change.kind == "formula_to_value"
    assert change.impact_count == 2
    assert ("Model", "C2") in change.impacted_cells
    assert ("Dashboard", "B12") in change.impacted_cells
    assert change.details["impact_paths"] == [
        {
            "target": "Dashboard!B12",
            "path": ["Model!B2", "Model!C2", "Dashboard!B12"],
        },
        {
            "target": "Model!C2",
            "path": ["Model!B2", "Model!C2"],
        },
    ]
    assert any(finding.rule_id == "FF001" for finding in report.findings)
    assert "`Model!B2` → `Model!C2` → `Dashboard!B12`" in report_to_markdown(report)


def test_html_report_is_self_contained_filterable_and_escapes_evidence(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    rewrite(
        candidate,
        lambda workbook: setattr(
            workbook["Model"]["B2"], "value", "<script>alert(1)</script>"
        ),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    rendered = report_to_html(report)

    assert rendered.startswith("<!doctype html>")
    assert 'id="review-filter"' in rendered
    assert 'id="severity-filter"' in rendered
    assert "FormulaFence change report" in rendered
    assert "FF001" in rendered
    assert "<script>alert(1)</script>" not in rendered
    assert "&lt;script&gt;alert(1)&lt;/script&gt;" in rendered
    assert "<script src=" not in rendered
    assert "<link rel=" not in rendered


def test_diff_detects_pattern_break_and_static_hazards(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def introduce_risk(workbook) -> None:
        workbook["Model"]["B3"] = "=Inputs!B3*3"
        workbook["Model"]["D2"] = "=[other.xlsx]Sheet1!A1"
        workbook["Model"]["D3"] = "=#REF!"
        workbook["Control"].sheet_state = "hidden"

    rewrite(candidate, introduce_risk)
    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    rule_ids = {finding.rule_id for finding in report.findings}
    assert {"FF003", "FF004", "FF006", "FF007"} <= rule_ids


def test_defined_name_change_is_semantic_control_change(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def move_name(workbook) -> None:
        workbook.defined_names["HeadlineOutput"].attr_text = "Model!$C$2"

    rewrite(candidate, move_name)
    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert any(change.kind == "defined_name_changed" for change in report.changes)
    assert any(finding.rule_id == "FF008" for finding in report.findings)


def test_named_ranges_resolve_into_the_dependency_index(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "named-ranges.xlsx")

    def add_named_range_formulas(workbook) -> None:
        model = workbook["Model"]
        model.defined_names.add(
            DefinedName(
                "LocalInput",
                attr_text="Inputs!$B$2",
                localSheetId=workbook.sheetnames.index("Model"),
            )
        )
        model["D2"] = "=HeadlineOutput*2"
        model["D3"] = "=LocalInput*3"
        workbook["Control"]["B2"] = "=Model!LocalInput*4"

    rewrite(workbook_path, add_named_range_formulas)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.unresolved_reference_tokens == {}
    assert ("Model", "D2") in snapshot.direct_dependents(("Dashboard", "B12"))
    assert ("Model", "D3") in snapshot.direct_dependents(("Inputs", "B2"))
    assert ("Control", "B2") in snapshot.direct_dependents(("Inputs", "B2"))
    assert "Model!LocalInput" in snapshot.defined_names


def test_formula_defined_names_expand_nested_and_local_static_dependencies(tmp_path) -> None:
    baseline = make_named_formula_model(tmp_path / "baseline.xlsx")
    candidate = make_named_formula_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Inputs"]["B2"], "value", 0.2))

    snapshot = load_snapshot(baseline)
    assert snapshot.unresolved_reference_tokens == {}
    assert snapshot.direct_dependents(("Inputs", "B2")) == {("Summary", "B2")}
    assert snapshot.direct_dependents(("Inputs", "B3")) == {("Summary", "B2")}
    assert snapshot.direct_dependents(("Inputs", "B4")) == {
        ("Report", "B2"),
        ("Summary", "B4"),
    }
    assert snapshot.direct_dependents(("Summary", "B3")) == set()
    assert "Summary!LocalMetric" in snapshot.defined_names

    report = compare_snapshots(snapshot, load_snapshot(candidate))
    change = next(change for change in report.changes if change.location == ("Inputs", "B2"))

    assert change.impacted_cells == (("Summary", "B2"),)


def test_formula_defined_name_change_remains_a_semantic_control_change(tmp_path) -> None:
    baseline = make_named_formula_model(tmp_path / "baseline.xlsx")
    candidate = make_named_formula_model(tmp_path / "candidate.xlsx")
    rewrite(
        candidate,
        lambda workbook: setattr(
            workbook.defined_names["TaxRate"], "attr_text", "=Inputs!$B$4"
        ),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert any(change.kind == "defined_name_changed" for change in report.changes)
    assert any(finding.rule_id == "FF008" for finding in report.findings)


def test_formula_defined_names_remain_coverage_gaps_when_not_fully_static(tmp_path) -> None:
    workbook_path = make_named_formula_model(tmp_path / "named-formulas.xlsx")

    def add_unsafe_formula_names(workbook) -> None:
        workbook.defined_names.add(DefinedName("RelativeMetric", attr_text="=B2"))
        workbook.defined_names.add(
            DefinedName("DynamicMetric", attr_text="=OFFSET(Inputs!$B$2,0,0)")
        )
        workbook.defined_names.add(DefinedName("CircularMetricA", attr_text="=CircularMetricB"))
        workbook.defined_names.add(DefinedName("CircularMetricB", attr_text="=CircularMetricA"))
        workbook.defined_names.add(DefinedName("PeriodMetric", attr_text="=SUM(Inputs:Report!B2)"))
        workbook.defined_names.add(DefinedName("SpillMetric", attr_text="=SUM(Inputs!$B$2#)"))
        workbook.defined_names.add(
            DefinedName(
                "SerializedSpillMetric",
                attr_text="=SUM(_xlfn.ANCHORARRAY(Inputs!$B$2))",
            )
        )
        workbook.defined_names.add(
            DefinedName(
                "ImplicitMetric",
                attr_text="=_xlfn.SINGLE(Inputs!$B$2:$B$4)",
            )
        )
        workbook.defined_names.add(
            DefinedName("LiteralImplicitMetric", attr_text="=@Inputs!$B$2:$B$4")
        )
        workbook["Summary"]["B5"] = "=RelativeMetric"
        workbook["Summary"]["B6"] = "=DynamicMetric"
        workbook["Summary"]["B7"] = "=CircularMetricA"
        workbook["Summary"]["B8"] = "=PeriodMetric"
        workbook["Summary"]["B9"] = "=SpillMetric"
        workbook["Summary"]["B10"] = "=SerializedSpillMetric"
        workbook["Summary"]["B11"] = "=ImplicitMetric"
        workbook["Summary"]["B12"] = "=LiteralImplicitMetric"

    rewrite(workbook_path, add_unsafe_formula_names)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.unresolved_reference_tokens == {
        ("Summary", "B5"): ("RelativeMetric",),
        ("Summary", "B6"): ("DynamicMetric",),
        ("Summary", "B7"): ("CircularMetricA",),
        ("Summary", "B8"): ("PeriodMetric",),
        ("Summary", "B9"): ("SpillMetric",),
        ("Summary", "B10"): ("SerializedSpillMetric",),
        ("Summary", "B11"): ("ImplicitMetric",),
        ("Summary", "B12"): ("LiteralImplicitMetric",),
    }
    assert ("Summary", "B5") not in snapshot.direct_dependents(("Inputs", "B2"))
    assert ("Summary", "B6") not in snapshot.direct_dependents(("Inputs", "B2"))
    assert ("Summary", "B7") not in snapshot.direct_dependents(("Inputs", "B2"))
    assert ("Summary", "B8") not in snapshot.direct_dependents(("Inputs", "B2"))
    assert ("Summary", "B9") not in snapshot.direct_dependents(("Inputs", "B2"))
    assert ("Summary", "B10") not in snapshot.direct_dependents(("Inputs", "B2"))
    assert ("Summary", "B11") not in snapshot.direct_dependents(("Inputs", "B2"))
    assert ("Summary", "B12") not in snapshot.direct_dependents(("Inputs", "B2"))


def test_formula_defined_names_expand_supported_static_table_references(tmp_path) -> None:
    workbook_path = make_table_model(tmp_path / "table.xlsx")

    def add_table_formula_name(workbook) -> None:
        workbook.defined_names.add(
            DefinedName("SalesAmount", attr_text="=SUM(Sales[Amount])")
        )
        workbook["Report"]["C2"] = "=SalesAmount"

    rewrite(workbook_path, add_table_formula_name)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.unresolved_reference_tokens == {}
    assert snapshot.direct_dependents(("Data", "A2")) >= {
        ("Report", "B2"),
        ("Report", "B3"),
        ("Report", "C2"),
    }


def test_let_variables_do_not_hide_static_dependency_paths(tmp_path) -> None:
    baseline = make_let_model(tmp_path / "baseline.xlsx")
    candidate = make_let_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Inputs"]["B2"], "value", 0.2))

    snapshot = load_snapshot(baseline)
    assert snapshot.unresolved_reference_tokens == {}
    assert snapshot.direct_dependents(("Inputs", "B2")) == {("Model", "B2")}
    assert snapshot.direct_dependents(("Inputs", "B3")) == {("Model", "B2")}

    report = compare_snapshots(snapshot, load_snapshot(candidate))
    change = next(change for change in report.changes if change.location == ("Inputs", "B2"))

    assert change.impacted_cells == (("Dashboard", "B2"), ("Model", "B2"))


def test_named_lambdas_expand_static_paths_through_nested_calls_and_named_formulas(
    tmp_path,
) -> None:
    baseline = make_named_lambda_model(tmp_path / "baseline.xlsx")
    candidate = make_named_lambda_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Inputs"]["B2"], "value", 0.2))

    snapshot = load_snapshot(baseline)
    assert snapshot.unresolved_reference_tokens == {}
    assert snapshot.direct_dependents(("Inputs", "B2")) == {
        ("Model", "B2"),
        ("Model", "B3"),
        ("Model", "B4"),
    }
    assert snapshot.direct_dependents(("Inputs", "B3")) == {
        ("Model", "B3"),
        ("Model", "B4"),
    }
    assert snapshot.direct_dependents(("Inputs", "B4")) == {
        ("Model", "B2"),
        ("Model", "B3"),
        ("Model", "B4"),
    }

    report = compare_snapshots(snapshot, load_snapshot(candidate))
    change = next(change for change in report.changes if change.location == ("Inputs", "B2"))

    assert change.impacted_cells == (
        ("Dashboard", "B2"),
        ("Model", "B2"),
        ("Model", "B3"),
        ("Model", "B4"),
    )


def test_unsafe_or_recursive_named_lambdas_remain_visible_coverage_gaps(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    workbook_path = make_model(tmp_path / "candidate.xlsx")

    def add_unsafe_named_lambdas(workbook) -> None:
        workbook.defined_names.add(
            DefinedName("UnsafeLookup", attr_text="=LAMBDA(address,INDIRECT(address))")
        )
        workbook.defined_names.add(
            DefinedName(
                "RecursiveCount",
                attr_text="=LAMBDA(value,IF(value=0,0,RecursiveCount(value-1)))",
            )
        )
        workbook["Model"]["D2"] = "=UnsafeLookup(Inputs!A1)"
        workbook["Model"]["D3"] = "=RecursiveCount(3)"

    rewrite(workbook_path, add_unsafe_named_lambdas)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.unresolved_reference_tokens == {
        ("Model", "D2"): ("UnsafeLookup",),
        ("Model", "D3"): ("RecursiveCount",),
    }
    assert snapshot.direct_dependents(("Inputs", "A1")) == {("Model", "D2")}
    report = compare_snapshots(load_snapshot(baseline), snapshot)
    assert {
        (finding.rule_id, finding.location)
        for finding in report.findings
    } >= {
        ("FF011", ("Model", "D2")),
        ("FF011", ("Model", "D3")),
    }


def test_named_lambda_calls_follow_worksheet_scope_and_qualified_local_names(tmp_path) -> None:
    snapshot = load_snapshot(make_scoped_named_lambda_model(tmp_path / "scoped.xlsx"))

    assert snapshot.unresolved_reference_tokens == {}
    assert snapshot.direct_dependents(("Inputs", "B2")) == {("Report", "B2")}
    assert snapshot.direct_dependents(("Inputs", "B3")) == {
        ("Model", "B2"),
        ("Report", "B3"),
    }
    assert snapshot.direct_dependents(("Model", "A2")) == {("Model", "B2")}
    assert snapshot.direct_dependents(("Report", "A2")) == {
        ("Report", "B2"),
        ("Report", "B3"),
    }


def test_diff_surfaces_new_static_coverage_gaps(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    rewrite(
        candidate,
        lambda workbook: setattr(
            workbook["Model"]["D2"],
            "value",
            '=UnknownMetric+INDIRECT("Inputs!B2")',
        ),
    )

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    assert profile["features"]["unresolved_reference_cells"] == [
        {"location": "Model!D2", "tokens": ["UnknownMetric"]}
    ]
    assert profile["features"]["dynamic_reference_cells"] == [
        {"location": "Model!D2", "functions": ["INDIRECT"]}
    ]

    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    rule_ids = {finding.rule_id for finding in report.findings}
    change_kinds = {change.kind for change in report.changes}
    assert {"FF011", "FF012"} <= rule_ids
    assert {
        "unresolved_formula_reference_added",
        "dynamic_formula_reference_added",
    } <= change_kinds


def test_spill_references_trace_anchors_but_remain_explicit_coverage_limits(tmp_path) -> None:
    baseline = make_spill_model(tmp_path / "baseline.xlsx")
    candidate = make_spill_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Inputs"]["B2"], "value", "=SEQUENCE(4)"))

    snapshot = load_snapshot(baseline)
    profile = profile_snapshot(snapshot)

    assert snapshot.unresolved_reference_tokens == {}
    assert snapshot.tokenization_failure_cells == set()
    assert snapshot.direct_dependents(("Inputs", "B2")) == {("Model", "B2")}
    assert snapshot.direct_dependents(("Inputs", "B3")) == {("Model", "B3")}
    assert snapshot.summary()["spill_reference_cells"] == 2
    assert profile["features"]["spill_reference_cells"] == [
        {"location": "Model!B2", "tokens": ["Inputs!B2#"]},
        {"location": "Model!B3", "tokens": ["_xlfn.ANCHORARRAY"]},
    ]
    assert "## Dynamic-array spill references" in profile_to_markdown(profile)

    report = compare_snapshots(snapshot, load_snapshot(candidate))
    change = next(change for change in report.changes if change.location == ("Inputs", "B2"))

    assert change.impacted_cells == (("Dashboard", "B2"), ("Model", "B2"))


def test_implicit_intersection_traces_the_selected_static_input_and_profiles_it(tmp_path) -> None:
    workbook = make_implicit_intersection_model(tmp_path / "implicit-intersection.xlsx")

    snapshot = load_snapshot(workbook)
    profile = profile_snapshot(snapshot)

    assert snapshot.unresolved_reference_tokens == {}
    assert snapshot.direct_dependents(("Inputs", "B2")) == {("Model", "B2")}
    assert snapshot.direct_dependents(("Inputs", "B3")) == {("Model", "B3")}
    assert snapshot.direct_dependents(("Inputs", "B4")) == set()
    assert snapshot.direct_dependents(("Model", "B2")) == {("Dashboard", "B2")}
    assert snapshot.summary()["implicit_intersection_cells"] == 2
    assert profile["features"]["implicit_intersection_cells"] == [
        {"location": "Model!B2", "tokens": ["_xlfn.SINGLE"]},
        {"location": "Model!B3", "tokens": ["@Inputs!B2:B4"]},
    ]
    assert "## Explicit implicit intersection" in profile_to_markdown(profile)


def test_legacy_cse_array_outputs_connect_input_changes_to_result_consumers(tmp_path) -> None:
    baseline = make_legacy_array_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_array_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Inputs"]["A2"], "value", "BBBB"))

    snapshot = load_snapshot(baseline)
    profile = profile_snapshot(snapshot)

    assert snapshot.direct_dependents(("Model", "B1")) == {
        ("Dashboard", "B2"),
        ("Model", "C2"),
    }
    assert snapshot.summary()["legacy_array_formula_cells"] == 1
    assert snapshot.summary()["legacy_array_formula_output_ranges"] == 1
    assert profile["features"]["legacy_array_formula_ranges"] == [
        {
            "anchor": "Model!B1",
            "ref": "Model!B1:B3",
            "output_cell_count": 3,
        }
    ]
    assert "## Legacy CSE array formulas" in profile_to_markdown(profile)

    report = compare_snapshots(snapshot, load_snapshot(candidate))
    change = next(change for change in report.changes if change.location == ("Inputs", "A2"))

    assert change.impacted_cells == (
        ("Dashboard", "B2"),
        ("Model", "B1"),
        ("Model", "C2"),
    )
    assert change.details["impact_paths"] == [
        {
            "target": "Dashboard!B2",
            "path": ["Inputs!A2", "Model!B1", "Dashboard!B2"],
        },
        {"target": "Model!B1", "path": ["Inputs!A2", "Model!B1"]},
        {
            "target": "Model!C2",
            "path": ["Inputs!A2", "Model!B1", "Model!C2"],
        },
    ]


def test_legacy_cse_output_aliases_do_not_expand_a_declared_huge_range(tmp_path) -> None:
    workbook = make_legacy_array_model(tmp_path / "large-cse.xlsx", "B1:XFD1048576")

    snapshot = load_snapshot(workbook)

    assert len(snapshot.cells) == 8
    assert snapshot.summary()["legacy_array_formula_output_cells"] > 1_000_000
    assert snapshot.direct_dependents(("Model", "B1")) == {
        ("Dashboard", "B2"),
        ("Model", "C2"),
    }


def test_dynamic_array_metadata_traces_observed_output_member_consumers(tmp_path) -> None:
    workbook = make_legacy_array_model(tmp_path / "dynamic.xlsx")
    mark_array_formula_dynamic(workbook)

    snapshot = load_snapshot(workbook)
    profile = profile_snapshot(snapshot)

    assert snapshot.legacy_array_formula_ranges == ()
    assert snapshot.dynamic_array_formula_cells == {("Model", "B1")}
    assert snapshot.unclassified_array_formula_cells == set()
    assert snapshot.direct_dependents(("Model", "B1")) == {
        ("Dashboard", "B2"),
        ("Model", "C2"),
    }
    assert snapshot.summary()["dynamic_array_observed_output_ranges"] == 1
    assert snapshot.summary()["dynamic_array_output_reference_cells"] == 2
    assert profile["features"]["dynamic_array_formula_cells"] == ["Model!B1"]
    assert profile["features"]["dynamic_array_observed_output_ranges"] == [
        {
            "anchor": "Model!B1",
            "ref": "Model!B1:B3",
            "output_cell_count": 3,
        }
    ]
    assert profile["features"]["dynamic_array_output_reference_cells"] == [
        {
            "location": "Dashboard!B2",
            "references": [
                {"anchor": "Model!B1", "observed_range": "Model!B1:B3"}
            ],
        },
        {
            "location": "Model!C2",
            "references": [
                {"anchor": "Model!B1", "observed_range": "Model!B1:B3"}
            ],
        },
    ]
    assert "## Dynamic-array formula anchors" in profile_to_markdown(profile)
    assert "observed from this workbook, not fixed" in profile_to_markdown(profile)


def test_dynamic_array_observed_output_aliases_connect_input_changes(tmp_path) -> None:
    baseline = make_legacy_array_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_array_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Inputs"]["A2"], "value", "BBBB"))
    mark_array_formula_dynamic(baseline)
    mark_array_formula_dynamic(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change = next(change for change in report.changes if change.location == ("Inputs", "A2"))

    assert change.impacted_cells == (
        ("Dashboard", "B2"),
        ("Model", "B1"),
        ("Model", "C2"),
    )
    assert change.details["impact_paths"] == [
        {
            "target": "Dashboard!B2",
            "path": ["Inputs!A2", "Model!B1", "Dashboard!B2"],
        },
        {"target": "Model!B1", "path": ["Inputs!A2", "Model!B1"]},
        {
            "target": "Model!C2",
            "path": ["Inputs!A2", "Model!B1", "Model!C2"],
        },
    ]


def test_dynamic_array_anchor_references_do_not_create_observed_member_aliases(tmp_path) -> None:
    workbook = make_legacy_array_model(tmp_path / "dynamic-anchor-only.xlsx")

    def use_anchor_only(workbook) -> None:
        workbook["Model"]["C2"] = "=B1*10"
        workbook["Dashboard"]["B2"] = "=Model!B1"

    rewrite(workbook, use_anchor_only)
    mark_array_formula_dynamic(workbook)

    snapshot = load_snapshot(workbook)

    assert snapshot.direct_dependents(("Model", "B1")) == {
        ("Dashboard", "B2"),
        ("Model", "C2"),
    }
    assert snapshot.dynamic_array_output_references == {}


def test_dynamic_array_observed_output_aliases_stay_compact_for_huge_ranges(tmp_path) -> None:
    workbook = make_legacy_array_model(
        tmp_path / "large-dynamic.xlsx", "B1:XFD1048576"
    )
    mark_array_formula_dynamic(workbook)

    snapshot = load_snapshot(workbook)

    assert len(snapshot.cells) == 8
    assert snapshot.summary()["dynamic_array_observed_output_ranges"] == 1
    assert snapshot.dynamic_array_formula_ranges[0].output_cell_count > 1_000_000
    assert snapshot.direct_dependents(("Model", "B1")) == {
        ("Dashboard", "B2"),
        ("Model", "C2"),
    }


def test_unclassified_array_metadata_is_a_visible_coverage_limit(tmp_path) -> None:
    baseline = make_legacy_array_model(tmp_path / "baseline.xlsx")
    workbook = make_legacy_array_model(tmp_path / "unclassified.xlsx")
    mark_array_formula_unclassified(workbook)

    snapshot = load_snapshot(workbook)
    profile = profile_snapshot(snapshot)

    assert snapshot.legacy_array_formula_ranges == ()
    assert snapshot.dynamic_array_formula_cells == set()
    assert snapshot.unclassified_array_formula_cells == {("Model", "B1")}
    assert snapshot.direct_dependents(("Model", "B1")) == set()
    assert snapshot.parser_warnings
    assert profile["features"]["unclassified_array_formula_cells"] == ["Model!B1"]
    assert "fixed-output aliases were not added" in profile_to_markdown(profile)
    report = compare_snapshots(load_snapshot(baseline), snapshot)
    assert any(finding.rule_id == "FF010" for finding in report.findings)


def test_array_formula_mode_and_legacy_output_range_changes_are_semantic(tmp_path) -> None:
    ordinary = make_legacy_array_model(tmp_path / "ordinary.xlsx")
    rewrite(
        ordinary,
        lambda workbook: setattr(
            workbook["Model"]["B1"], "value", "=LEN(Inputs!A1:A3)"
        ),
    )
    cse = make_legacy_array_model(tmp_path / "cse.xlsx")

    mode_report = compare_snapshots(load_snapshot(ordinary), load_snapshot(cse))
    mode_change = next(
        change for change in mode_report.changes if change.kind == "array_formula_mode_changed"
    )
    assert mode_change.location == ("Model", "B1")
    assert mode_change.details["before"] == {"mode": "ordinary", "output_range": None}
    assert mode_change.details["after"] == {
        "mode": "legacy_cse",
        "output_range": "B1:B3",
    }
    assert mode_change.impacted_cells == (("Dashboard", "B2"), ("Model", "C2"))
    assert any(finding.rule_id == "FF018" for finding in mode_report.findings)

    blank = make_legacy_array_model(tmp_path / "blank.xlsx")
    rewrite(blank, lambda workbook: setattr(workbook["Model"]["B1"], "value", None))
    dynamic = make_legacy_array_model(tmp_path / "dynamic.xlsx")
    mark_array_formula_dynamic(dynamic)
    new_dynamic_report = compare_snapshots(load_snapshot(blank), load_snapshot(dynamic))
    new_dynamic_change = next(
        change
        for change in new_dynamic_report.changes
        if change.kind == "array_formula_mode_changed"
    )
    assert new_dynamic_change.details["before"]["mode"] == "absent"
    assert new_dynamic_change.details["after"]["mode"] == "dynamic"
    assert any(finding.rule_id == "FF018" for finding in new_dynamic_report.findings)

    baseline = make_legacy_array_model(tmp_path / "baseline.xlsx", "B1:B3")
    candidate = make_legacy_array_model(tmp_path / "candidate.xlsx", "B1:B4")
    range_report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    range_change = next(
        change
        for change in range_report.changes
        if change.kind == "legacy_array_output_range_changed"
    )
    assert range_change.location == ("Model", "B1")
    assert range_change.details["before_output_range"] == "B1:B3"
    assert range_change.details["after_output_range"] == "B1:B4"
    assert range_change.impacted_cells == (("Dashboard", "B2"), ("Model", "C2"))
    assert any(finding.rule_id == "FF018" for finding in range_report.findings)


def test_dynamic_array_cached_extent_change_is_not_a_fixed_range_change(tmp_path) -> None:
    baseline = make_legacy_array_model(tmp_path / "baseline.xlsx", "B1:B3")
    candidate = make_legacy_array_model(tmp_path / "candidate.xlsx", "B1:B4")
    mark_array_formula_dynamic(baseline)
    mark_array_formula_dynamic(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert not {
        change.kind
        for change in report.changes
        if change.kind
        in {"array_formula_mode_changed", "legacy_array_output_range_changed"}
    }
    assert not {finding.rule_id for finding in report.findings if finding.rule_id == "FF018"}
    assert not {finding.rule_id for finding in report.findings if finding.rule_id == "FF019"}


def test_new_dynamic_array_output_member_consumers_emit_ff019(tmp_path) -> None:
    baseline = make_legacy_array_model(tmp_path / "baseline.xlsx")

    def remove_member_consumers(workbook) -> None:
        workbook["Model"]["C2"] = None
        workbook["Dashboard"]["B2"] = None

    rewrite(baseline, remove_member_consumers)
    candidate = make_legacy_array_model(tmp_path / "candidate.xlsx")
    mark_array_formula_dynamic(baseline)
    mark_array_formula_dynamic(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    changes = {
        change.location: change
        for change in report.changes
        if change.kind == "dynamic_array_output_reference_added"
    }
    findings = {
        finding.location: finding for finding in report.findings if finding.rule_id == "FF019"
    }

    assert set(changes) == {("Dashboard", "B2"), ("Model", "C2")}
    assert set(findings) == set(changes)
    assert changes[("Model", "C2")].details["references"] == [
        {"anchor": "Model!B1", "observed_range": "Model!B1:B3"}
    ]
    assert "observed dynamic-array spill" in findings[("Model", "C2")].message


def test_dynamic_array_extent_growth_only_flags_new_member_relationships(tmp_path) -> None:
    baseline = make_legacy_array_model(tmp_path / "baseline.xlsx", "B1:B3")
    candidate = make_legacy_array_model(tmp_path / "candidate.xlsx", "B1:B4")

    def add_future_member_consumer(workbook) -> None:
        workbook["Model"]["C4"] = "=B4*10"

    rewrite(baseline, add_future_member_consumer)
    rewrite(candidate, add_future_member_consumer)
    mark_array_formula_dynamic(baseline)
    mark_array_formula_dynamic(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    ff019 = [finding for finding in report.findings if finding.rule_id == "FF019"]

    assert [(finding.location, finding.details["references"]) for finding in ff019] == [
        (
            ("Model", "C4"),
            [{"anchor": "Model!B1", "observed_range": "Model!B1:B4"}],
        )
    ]
    assert not {finding.rule_id for finding in report.findings if finding.rule_id == "FF018"}


def test_diff_surfaces_new_spill_and_tokenization_coverage_limits(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def add_coverage_limits(workbook) -> None:
        workbook["Model"]["D2"] = "=SUM(Inputs!B2#)"
        workbook["Model"]["D3"] = "=SUM(Inputs!B2#1)"

    rewrite(candidate, add_coverage_limits)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert profile["features"]["spill_reference_cells"] == [
        {"location": "Model!D2", "tokens": ["Inputs!B2#"]}
    ]
    assert profile["features"]["tokenization_failure_cells"] == ["Model!D3"]
    assert "Formula tokenizer could not inspect `Model!D3`" in profile_to_markdown(profile)
    assert {finding.rule_id for finding in report.findings} >= {"FF015", "FF016"}
    assert {change.kind for change in report.changes} >= {
        "spill_reference_added",
        "formula_tokenization_failure_added",
    }


def test_diff_surfaces_new_implicit_intersection(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    rewrite(
        candidate,
        lambda workbook: setattr(
            workbook["Model"]["D2"], "value", "=@Inputs!B2:B4"
        ),
    )

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert profile_snapshot(candidate_snapshot)["features"]["implicit_intersection_cells"] == [
        {"location": "Model!D2", "tokens": ["@Inputs!B2:B4"]}
    ]
    assert {finding.rule_id for finding in report.findings} >= {"FF017"}
    assert {change.kind for change in report.changes} >= {"implicit_intersection_added"}


def test_external_defined_name_is_tracked_as_an_external_reference(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "external-name.xlsx")

    def add_external_name_formula(workbook) -> None:
        workbook.defined_names.add(
            DefinedName("ExternalInput", attr_text="'[other.xlsx]Inputs'!$B$2")
        )
        workbook["Model"]["D2"] = "=ExternalInput*2"

    rewrite(workbook_path, add_external_name_formula)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.external_references == {("Model", "D2")}
    assert snapshot.unresolved_reference_tokens == {}


def test_static_external_a1_references_are_retained_privately_for_portfolios(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "external-a1.xlsx")
    private_source = "PRIVATE-EXTERNAL-SOURCE-PATH"

    def add_external_a1_formula(workbook) -> None:
        workbook["Model"]["D2"] = (
            "='C:\\" + private_source + "\\[Inputs.xlsx]Data'!$B$2:$B$4"
        )

    rewrite(workbook_path, add_external_a1_formula)
    snapshot = load_snapshot(workbook_path)

    references = snapshot.external_workbook_references[("Model", "D2")]
    assert len(references) == 1
    assert references[0].source_path == f"C:\\{private_source}\\Inputs.xlsx"
    assert references[0].sheet == "Data"
    assert (
        references[0].min_column,
        references[0].min_row,
        references[0].max_column,
        references[0].max_row,
    ) == (2, 2, 2, 4)
    assert private_source not in json.dumps(profile_snapshot(snapshot))


def test_static_external_defined_names_are_retained_privately_for_portfolios(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "external-defined-name.xlsx")
    private_source = "PRIVATE-EXTERNAL-DEFINED-NAME-PATH"
    private_name = "PrivateExternalName"

    def add_external_name_formula(workbook) -> None:
        workbook["Model"]["D2"] = (
            "='C:\\" + private_source + "\\[Inputs.xlsx]" + private_name + "'"
        )

    rewrite(workbook_path, add_external_name_formula)
    snapshot = load_snapshot(workbook_path)

    references = snapshot.external_workbook_defined_name_references[("Model", "D2")]
    assert len(references) == 1
    assert references[0].source_path == f"C:\\{private_source}\\Inputs.xlsx"
    assert references[0].name_key == private_name.casefold()
    rendered = json.dumps(profile_snapshot(snapshot))
    assert private_source not in rendered
    assert private_name not in rendered


def test_external_workbook_link_surface_ledger_catches_same_surface_swaps(
    tmp_path,
) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    private_baseline = "PRIVATE-LINK-SURFACE-BASELINE"
    private_candidate = "PRIVATE-LINK-SURFACE-CANDIDATE"

    def add_link_surfaces(workbook, marker: str) -> None:
        source = f"C:\\{marker}\\[Source.xlsx]Inputs"
        workbook["Model"]["D2"] = f"='{source}'!$B$2"
        workbook.defined_names.add(
            DefinedName(
                "ExternalNamedLimit",
                attr_text=f"'{source}'!$C$2",
            )
        )
        validation = DataValidation(
            type="whole",
            operator="greaterThan",
            formula1=f"='{source}'!$D$2",
        )
        validation.add("E2")
        workbook["Model"].add_data_validation(validation)

    rewrite(baseline, lambda workbook: add_link_surfaces(workbook, private_baseline))
    rewrite(candidate, lambda workbook: add_link_surfaces(workbook, private_candidate))

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    profile = profile_snapshot(candidate_snapshot)
    surface_change = next(
        change
        for change in report.changes
        if change.kind == "external_workbook_link_surfaces_changed"
    )

    assert baseline_snapshot.external_references == {("Model", "D2")}
    assert candidate_snapshot.external_references == {("Model", "D2")}
    assert not any(finding.rule_id == "FF004" for finding in report.findings)
    assert any(finding.rule_id == "FF081" for finding in report.findings)
    assert surface_change.details == {
        "before": {
            "present": True,
            "surface_count": 3,
            "cell_formula_surface_count": 1,
            "defined_name_surface_count": 1,
            "data_validation_surface_count": 1,
            "chart_formula_surface_count": 0,
            "opaque_chart_part_count": 0,
            "external_reference_count": 3,
        },
        "after": {
            "present": True,
            "surface_count": 3,
            "cell_formula_surface_count": 1,
            "defined_name_surface_count": 1,
            "data_validation_surface_count": 1,
            "chart_formula_surface_count": 0,
            "opaque_chart_part_count": 0,
            "external_reference_count": 3,
        },
        "external_workbook_link_surface_material_changed": True,
    }
    assert profile["external_workbook_link_surfaces"] == surface_change.details["after"]

    rendered_profile = json.dumps(profile)
    rendered_markdown = profile_to_markdown(profile)
    external_surface_sarif = json.dumps(
        [
            result
            for result in report_to_sarif(report)["runs"][0]["results"]
            if result["ruleId"] == "FF081"
        ]
    )
    for private_value in (private_baseline, private_candidate):
        assert private_value not in rendered_profile
        assert private_value not in rendered_markdown
        assert private_value not in external_surface_sarif


def test_external_workbook_link_report_redaction_hides_static_and_dynamic_literals() -> None:
    payload = {
        "cell": "='C:\\PRIVATE-CELL\\[Source.xlsx]Inputs'!$B$2",
        "defined_name": "='C:\\PRIVATE-NAME\\[Source.xlsx]ExternalLimit'",
        "validation": "='C:\\PRIVATE-VALIDATION\\[Source.xlsx]Inputs'!$D$2",
        "structured": "='..\\PRIVATE-TABLE\\source.xlsx'!Sales[#Data]",
        "dynamic": '=INDIRECT("\'[PRIVATE-DYNAMIC]Inputs\'!A1")',
        "ordinary": "keep ordinary local review evidence",
        "local_table": "Table[Amount]",
        "nested": ["=[PRIVATE-NAMED]InputRange"],
    }

    redacted = redact_external_workbook_link_material(payload)

    assert payload["ordinary"] == "keep ordinary local review evidence"
    assert redacted["ordinary"] == payload["ordinary"]
    assert redacted["local_table"] == payload["local_table"]
    for private_value in (
        "PRIVATE-CELL",
        "PRIVATE-NAME",
        "PRIVATE-VALIDATION",
        "PRIVATE-TABLE",
        "PRIVATE-DYNAMIC",
        "PRIVATE-NAMED",
    ):
        assert private_value not in json.dumps(redacted)
    assert redacted["cell"] == EXTERNAL_WORKBOOK_LINK_REDACTION
    assert redacted["nested"] == [EXTERNAL_WORKBOOK_LINK_REDACTION]


def test_formula_external_action_redaction_hides_direct_action_and_dde_formulas() -> None:
    payload = {
        "hyperlink": '=HYPERLINK("https://private.example.test/PRIVATE-LINK", "Open")',
        "webservice": '=WEBSERVICE("https://private.example.test/PRIVATE-WEB")',
        "cube": '=CUBEVALUE("PRIVATE-CUBE-CONNECTION", "[Measures].[PRIVATE-REVENUE]")',
        "dde": "=DDE.TEST|'PRIVATE-DDE-TOPIC'!PRIVATE_ITEM",
        "ordinary": "=SUM(A1:A2)",
        "plain": "ordinary local review evidence",
    }

    redacted = redact_formula_external_action_material(payload)

    assert redacted["ordinary"] == payload["ordinary"]
    assert redacted["plain"] == payload["plain"]
    assert all(
        redacted[key] == FORMULA_EXTERNAL_ACTION_REDACTION
        for key in ("hyperlink", "webservice", "cube", "dde")
    )
    for private_value in (
        "PRIVATE-LINK",
        "PRIVATE-WEB",
        "PRIVATE-CUBE-CONNECTION",
        "PRIVATE-REVENUE",
        "PRIVATE-DDE-TOPIC",
        "PRIVATE_ITEM",
    ):
        assert private_value not in json.dumps(redacted)


def test_formula_external_action_report_redaction_hides_unsampled_static_inputs(
    tmp_path,
) -> None:
    """A shared report must not depend on its bounded impact sample for privacy."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-ACTION-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-ACTION-CANDIDATE"

    def add_action_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        # The report only keeps the first 20 sorted impact samples.  Z1 is a
        # known HYPERLINK consumer but intentionally falls beyond B1:U1.
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = '=HYPERLINK(A9, "Open")'

    rewrite(baseline, lambda workbook: add_action_fanout(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_action_fanout(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.formula_external_action_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )

    redacted = redact_formula_external_action_report_payload(report, raw_payload)
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert redacted_change["before"]["value"] == FORMULA_EXTERNAL_ACTION_REDACTION
    assert redacted_change["after"]["value"] == FORMULA_EXTERNAL_ACTION_REDACTION
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_python_in_excel_report_redaction_hides_unsampled_static_inputs(
    tmp_path,
) -> None:
    """PY input redaction must use full private impact, not displayed samples."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-PY-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-PY-CANDIDATE"

    def add_python_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        # Z1 is a known PY consumer but falls after B1:U1 in the report's
        # bounded sorted impact sample.
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = "=_xlfn._xlws.PY(0,0,A9)"

    rewrite(baseline, lambda workbook: add_python_fanout(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_python_fanout(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.python_in_excel_static_input_cells == frozenset({("Inputs", "A9")})

    redacted = redact_python_in_excel_report_payload(report, raw_payload)
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert redacted_change["before"]["value"] == PYTHON_IN_EXCEL_REDACTION
    assert redacted_change["after"]["value"] == PYTHON_IN_EXCEL_REDACTION
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_external_workbook_link_surface_ledger_covers_direct_external_names(
    tmp_path,
) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    private_baseline = "PRIVATE-LINK-NAME-BASELINE"
    private_candidate = "PRIVATE-LINK-NAME-CANDIDATE"

    def add_external_names(workbook, marker: str) -> None:
        source = f"C:\\{marker}\\[Source.xlsx]"
        workbook["Model"]["D2"] = f"='{source}ExternalInput'"
        workbook.defined_names.add(
            DefinedName(
                "ExternalNamedLimit",
                attr_text=f"='{source}ExternalLimit'",
            )
        )

    rewrite(baseline, lambda workbook: add_external_names(workbook, private_baseline))
    rewrite(candidate, lambda workbook: add_external_names(workbook, private_candidate))

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    details = next(
        finding.details for finding in report.findings if finding.rule_id == "FF081"
    )

    assert baseline_snapshot.external_references == {("Model", "D2")}
    assert candidate_snapshot.external_references == {("Model", "D2")}
    assert details["after"] == {
        "present": True,
        "surface_count": 2,
        "cell_formula_surface_count": 1,
        "defined_name_surface_count": 1,
        "data_validation_surface_count": 0,
        "chart_formula_surface_count": 0,
        "opaque_chart_part_count": 0,
        "external_reference_count": 2,
    }
    assert all(
        private_value not in json.dumps(profile_snapshot(candidate_snapshot))
        for private_value in (private_baseline, private_candidate)
    )


def test_external_workbook_link_surface_ledger_covers_chart_formula_targets(tmp_path) -> None:
    baseline = make_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_chart_definition_model(tmp_path / "candidate.xlsx")
    private_baseline = "PRIVATE-CHART-LINK-BASELINE"
    private_candidate = "PRIVATE-CHART-LINK-CANDIDATE"
    set_chart_formula_external_workbook_target(
        baseline,
        f"'C:\\{private_baseline}\\[Source.xlsx]Data'!$B$2:$B$4",
    )
    set_chart_formula_external_workbook_target(
        candidate,
        f"'C:\\{private_candidate}\\[Source.xlsx]Data'!$B$2:$B$4",
    )

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    profile = profile_snapshot(candidate_snapshot)

    assert (
        baseline_snapshot.chart_definitions.external_workbook_formula_surface_count
    ) == 1
    assert (
        candidate_snapshot.chart_definitions.external_workbook_formula_reference_count
    ) == 1
    assert profile["external_workbook_link_surfaces"] == {
        "present": True,
        "surface_count": 1,
        "cell_formula_surface_count": 0,
        "defined_name_surface_count": 0,
        "data_validation_surface_count": 0,
        "chart_formula_surface_count": 1,
        "opaque_chart_part_count": 0,
        "external_reference_count": 1,
    }
    assert any(finding.rule_id == "FF081" for finding in report.findings)
    assert "## Static external-workbook link surfaces" in profile_to_markdown(profile)

    rendered_profile = json.dumps(profile)
    rendered_markdown = profile_to_markdown(profile)
    rendered_sarif = json.dumps(report_to_sarif(report))
    for private_value in (private_baseline, private_candidate):
        assert private_value not in rendered_profile
        assert private_value not in rendered_markdown
        assert private_value not in rendered_sarif


def test_external_workbook_link_surface_ledger_fails_closed_for_unreadable_charts(
    tmp_path,
) -> None:
    baseline = make_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_chart_definition_model(tmp_path / "candidate.xlsx")
    corrupt_chart_definition_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    surface_change = next(
        change
        for change in report.changes
        if change.kind == "external_workbook_link_surfaces_changed"
    )

    assert candidate_snapshot.external_workbook_link_surfaces.opaque_chart_part_count == 1
    assert surface_change.details["opaque_chart_link_surface_coverage_changed"] is True
    assert any(finding.rule_id == "FF081" for finding in report.findings)


def test_external_workbook_link_surface_ledger_only_marks_unreadable_chart_formulas(
    tmp_path,
) -> None:
    baseline = make_extended_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_extended_chart_definition_model(tmp_path / "candidate.xlsx")
    add_unsupported_extended_chart_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.chart_definitions.unrecognized_part_count >= 1
    assert candidate_snapshot.external_workbook_link_surfaces.opaque_chart_part_count == 0
    assert not any(finding.rule_id == "FF081" for finding in report.findings)


def test_static_table_references_feed_dependency_paths_and_profiles(tmp_path) -> None:
    workbook_path = make_table_model(tmp_path / "table.xlsx")
    snapshot = load_snapshot(workbook_path)
    profile = profile_snapshot(snapshot)

    assert snapshot.summary()["table_count"] == 1
    assert snapshot.tables["Sales"].columns == ("Amount", "Rate", "Value")
    assert snapshot.unresolved_reference_tokens == {}
    assert ("Report", "B2") in snapshot.direct_dependents(("Data", "A2"))
    assert ("Report", "B3") in snapshot.direct_dependents(("Data", "B2"))
    assert ("Report", "B4") in snapshot.direct_dependents(("Data", "A1"))
    assert ("Report", "B5") in snapshot.direct_dependents(("Data", "B1"))
    assert profile["tables"] == [
        {
            "name": "Sales",
            "sheet": "Data",
            "ref": "A1:C4",
            "columns": ["Amount", "Rate", "Value"],
            "header_row_count": 1,
            "totals_row_count": 0,
        }
    ]
    assert "## Excel tables" in profile_to_markdown(profile)
    assert "| Sales | Data | A1:C4 | Amount, Rate, Value |" in profile_to_markdown(profile)


def test_table_definition_change_is_a_semantic_control_change(tmp_path) -> None:
    baseline = make_table_model(tmp_path / "baseline.xlsx")
    candidate = make_table_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Data"].tables["Sales"], "ref", "A1:C3"))

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert any(change.kind == "table_definition_changed" for change in report.changes)
    assert any(finding.rule_id == "FF013" for finding in report.findings)
    assert baseline_snapshot.xml_mapping_controls == candidate_snapshot.xml_mapping_controls
    assert not any(finding.rule_id == "FF049" for finding in report.findings)


def test_data_validation_controls_are_profiled_without_exposing_criteria(tmp_path) -> None:
    workbook = make_data_validation_model(tmp_path / "validation.xlsx")

    snapshot = load_snapshot(workbook)
    profile = profile_snapshot(snapshot)
    markdown = profile_to_markdown(profile)

    assert snapshot.summary()["data_validation_rules"] == 2
    assert snapshot.summary()["data_validation_target_ranges"] == 3
    assert profile["data_validations"] == [
        {
            "sheet": "Inputs",
            "ranges": ["Inputs!B2:B100", "Inputs!D2"],
            "type": "list",
            "operator": "between",
            "criteria_count": 1,
            "allow_blank": True,
            "dropdown_hidden": False,
            "prompts_disabled": False,
            "show_input_message": True,
            "show_error_message": True,
            "error_style": "stop",
            "has_error_alert_text": True,
            "has_input_prompt_text": True,
            "ime_mode": "noControl",
        },
        {
            "sheet": "Inputs",
            "ranges": ["Inputs!C2:C100"],
            "type": "decimal",
            "operator": "between",
            "criteria_count": 2,
            "allow_blank": False,
            "dropdown_hidden": False,
            "prompts_disabled": False,
            "show_input_message": False,
            "show_error_message": True,
            "error_style": "warning",
            "has_error_alert_text": True,
            "has_input_prompt_text": False,
            "ime_mode": "noControl",
        },
    ]
    assert "formula1" not in profile["data_validations"][0]
    assert "## Data-validation controls" in markdown
    assert "Limits!$A$2" not in markdown
    assert "Choose an approved status." not in markdown


def test_data_validation_writer_defaults_and_formula_spelling_are_canonical(tmp_path) -> None:
    baseline = make_data_validation_model(tmp_path / "baseline.xlsx")
    candidate = make_data_validation_model(tmp_path / "candidate.xlsx", reverse_status_targets=True)

    def use_omitted_ooxml_defaults(workbook) -> None:
        for validation in workbook["Inputs"].data_validations.dataValidation:
            if validation.formula1:
                validation.formula1 = validation.formula1.removeprefix("=")
            if validation.formula2:
                validation.formula2 = validation.formula2.removeprefix("=")
            validation.operator = None
            if validation.errorStyle == "stop":
                validation.errorStyle = None

    rewrite(candidate, use_omitted_ooxml_defaults)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert not {
        change.kind for change in report.changes if change.kind == "data_validation_changed"
    }
    assert not {finding.rule_id for finding in report.findings if finding.rule_id == "FF020"}


def test_data_validation_equivalent_target_grouping_is_canonical(tmp_path) -> None:
    baseline = make_data_validation_model(tmp_path / "baseline.xlsx")
    candidate = make_data_validation_model(tmp_path / "candidate.xlsx")

    def split_identical_status_control(workbook) -> None:
        inputs = workbook["Inputs"]
        _, amount = inputs.data_validations.dataValidation
        inputs.data_validations.dataValidation.clear()
        for target in ("B2:B100", "D2"):
            status = DataValidation(
                type="list",
                formula1="=Limits!$A$2:$A$4",
                allow_blank=True,
                showInputMessage=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Invalid status",
                error="Choose an approved status.",
                promptTitle="Approved status",
                prompt="Choose a documented status.",
            )
            status.add(target)
            inputs.add_data_validation(status)
        inputs.add_data_validation(amount)

    rewrite(candidate, split_identical_status_control)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert not {
        change.kind for change in report.changes if change.kind == "data_validation_changed"
    }
    assert not {finding.rule_id for finding in report.findings if finding.rule_id == "FF020"}


def test_data_validation_change_is_a_high_risk_semantic_control_change(tmp_path) -> None:
    baseline = make_data_validation_model(tmp_path / "baseline.xlsx")
    candidate = make_data_validation_model(tmp_path / "candidate.xlsx")

    def weaken_amount_control(workbook) -> None:
        amount = workbook["Inputs"].data_validations.dataValidation[1]
        amount.formula2 = "=Limits!$B$2"
        amount.showErrorMessage = False

    rewrite(candidate, weaken_amount_control)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change = next(change for change in report.changes if change.kind == "data_validation_changed")
    finding = next(finding for finding in report.findings if finding.rule_id == "FF020")

    assert change.severity == "high"
    assert change.details["sheet"] == "Inputs"
    assert change.details["before"][1]["formula2"] == "Limits!$B$3"
    assert change.details["after"][1]["formula2"] == "Limits!$B$2"
    assert change.details["after"][1]["show_error_message"] is False
    assert finding.location is None


def test_data_validation_global_prompt_disable_is_a_control_change(tmp_path) -> None:
    baseline = make_data_validation_model(tmp_path / "baseline.xlsx")
    candidate = make_data_validation_model(tmp_path / "candidate.xlsx")
    rewrite(
        candidate,
        lambda workbook: setattr(
            workbook["Inputs"].data_validations, "disablePrompts", True
        ),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change = next(change for change in report.changes if change.kind == "data_validation_changed")

    assert all(item["prompts_disabled"] is True for item in change.details["after"])
    assert "worksheet prompts disabled" in profile_to_markdown(
        profile_snapshot(load_snapshot(candidate))
    )


def test_data_validation_target_ranges_stay_compact_at_sheet_scale(tmp_path) -> None:
    workbook = make_data_validation_model(tmp_path / "large-validation.xlsx")

    def add_full_column_control(current_workbook) -> None:
        validation = DataValidation(
            type="whole",
            operator="greaterThan",
            formula1="0",
            showErrorMessage=True,
        )
        validation.add("E1:E1048576")
        current_workbook["Inputs"].add_data_validation(validation)

    rewrite(workbook, add_full_column_control)
    snapshot = load_snapshot(workbook)

    assert len(snapshot.cells) < 20
    assert snapshot.summary()["data_validation_rules"] == 3
    assert snapshot.summary()["data_validation_target_ranges"] == 4
    assert any(
        validation.ranges == ("E1:E1048576",)
        for validation in snapshot.data_validations
    )


def test_conditional_formatting_controls_are_profiled_without_exposing_criteria(tmp_path) -> None:
    workbook = make_conditional_formatting_model(tmp_path / "conditional-formatting.xlsx")

    snapshot = load_snapshot(workbook)
    profile = profile_snapshot(snapshot)
    markdown = profile_to_markdown(profile)

    assert snapshot.summary()["conditional_formatting_rules"] == 5
    assert snapshot.summary()["conditional_formatting_target_ranges"] == 5
    assert snapshot.summary()["conditional_formatting_extensions"] == 0
    assert profile["conditional_formatting"][0] == {
        "sheet": "Inputs",
        "ranges": ["Inputs!A2:A100"],
        "priority": 1,
        "type": "expression",
        "operator": None,
        "formula_count": 1,
        "has_text_criterion": False,
        "stop_if_true": True,
        "above_average": True,
        "percent": False,
        "bottom": False,
        "rank": None,
        "std_dev": None,
        "equal_average": False,
        "time_period": None,
        "formatting": ["differential style"],
        "extension_count": 0,
    }
    assert profile["conditional_formatting"][2]["formatting"] == ["color scale"]
    assert profile["conditional_formatting"][3]["formatting"] == ["data bar"]
    assert profile["conditional_formatting"][4]["formatting"] == ["icon set"]
    assert "formulas" not in profile["conditional_formatting"][0]
    assert "$A2<0" not in markdown
    assert "FFFFC7CE" not in markdown
    assert "## Conditional-formatting controls" in markdown


def test_conditional_formatting_defaults_formula_spelling_and_dxf_order_are_canonical(
    tmp_path,
) -> None:
    baseline = make_conditional_formatting_model(tmp_path / "baseline.xlsx")
    candidate = make_conditional_formatting_model(tmp_path / "candidate.xlsx")

    def use_equivalent_writer_spelling(workbook) -> None:
        rules = [
            rule
            for rule_group in workbook["Inputs"].conditional_formatting._cf_rules.values()
            for rule in rule_group
        ]
        for rule in rules:
            rule.priority *= 10
            rule.formula = [f"={formula}" for formula in rule.formula]
            if rule.stopIfTrue is None:
                rule.stopIfTrue = False
            if rule.aboveAverage is None:
                rule.aboveAverage = True
            if rule.percent is None:
                rule.percent = False
            if rule.bottom is None:
                rule.bottom = False
            if rule.equalAverage is None:
                rule.equalAverage = False

    rewrite(candidate, use_equivalent_writer_spelling)
    reorder_conditional_differential_styles(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert not {
        change.kind for change in report.changes if change.kind == "conditional_formatting_changed"
    }
    assert not {finding.rule_id for finding in report.findings if finding.rule_id == "FF021"}


def test_conditional_formatting_rule_change_and_precedence_change_are_high_risk(tmp_path) -> None:
    baseline = make_conditional_formatting_model(tmp_path / "baseline.xlsx")
    changed_rule = make_conditional_formatting_model(tmp_path / "changed-rule.xlsx")
    changed_precedence = make_conditional_formatting_model(
        tmp_path / "changed-precedence.xlsx"
    )

    def change_rule(workbook) -> None:
        rules = [
            rule
            for rule_group in workbook["Inputs"].conditional_formatting._cf_rules.values()
            for rule in rule_group
        ]
        rules[1].formula = ["90"]
        rules[1].stopIfTrue = True

    def swap_precedence(workbook) -> None:
        rules = [
            rule
            for rule_group in workbook["Inputs"].conditional_formatting._cf_rules.values()
            for rule in rule_group
        ]
        rules[0].priority, rules[1].priority = rules[1].priority, rules[0].priority

    rewrite(changed_rule, change_rule)
    rewrite(changed_precedence, swap_precedence)

    rule_report = compare_snapshots(load_snapshot(baseline), load_snapshot(changed_rule))
    precedence_report = compare_snapshots(
        load_snapshot(baseline), load_snapshot(changed_precedence)
    )
    rule_change = next(
        change for change in rule_report.changes if change.kind == "conditional_formatting_changed"
    )
    precedence_change = next(
        change
        for change in precedence_report.changes
        if change.kind == "conditional_formatting_changed"
    )

    assert rule_change.severity == "high"
    assert rule_change.details["before"]["rules"][1]["formulas"] == ["100"]
    assert rule_change.details["after"]["rules"][1]["formulas"] == ["90"]
    assert rule_change.details["after"]["rules"][1]["stop_if_true"] is True
    assert precedence_change.severity == "high"
    assert [item["type"] for item in precedence_change.details["after"]["rules"][:2]] == [
        "cellIs",
        "expression",
    ]
    assert {finding.rule_id for finding in rule_report.findings} >= {"FF021"}
    assert {finding.rule_id for finding in precedence_report.findings} >= {"FF021"}


def test_conditional_formatting_extensions_are_compared_without_guid_noise(tmp_path) -> None:
    baseline = make_conditional_formatting_model(tmp_path / "baseline.xlsx")
    equivalent = make_conditional_formatting_model(tmp_path / "equivalent.xlsx")
    changed = make_conditional_formatting_model(tmp_path / "changed.xlsx")
    changed_extension_type = make_conditional_formatting_model(
        tmp_path / "changed-extension-type.xlsx"
    )
    add_conditional_formatting_databar_extension(
        baseline,
        guid="{11111111-2222-3333-4444-555555555555}",
        axis_color="FF000000",
    )
    add_conditional_formatting_databar_extension(
        equivalent,
        guid="{AAAAAAAA-BBBB-CCCC-DDDD-EEEEEEEEEEEE}",
        axis_color="FF000000",
    )
    add_conditional_formatting_databar_extension(
        changed,
        guid="{99999999-8888-7777-6666-555555555555}",
        axis_color="FFFF0000",
    )
    add_conditional_formatting_databar_extension(
        changed_extension_type,
        guid="{99999999-8888-7777-6666-555555555555}",
        axis_color="FF000000",
        worksheet_extension_uri="{11111111-2222-3333-4444-555555555555}",
    )

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    equivalent_report = compare_snapshots(baseline_snapshot, load_snapshot(equivalent))
    changed_report = compare_snapshots(baseline_snapshot, load_snapshot(changed))
    extension_type_report = compare_snapshots(
        baseline_snapshot, load_snapshot(changed_extension_type)
    )

    assert baseline_snapshot.summary()["conditional_formatting_extensions"] == 1
    assert profile["conditional_formatting_extensions"] == [
        {"sheet": "Inputs", "element": "ext"}
    ]
    assert "FF000000" not in json.dumps(profile)
    markdown = profile_to_markdown(profile)
    assert "## Conditional-formatting extension coverage" in markdown
    assert "FF000000" not in markdown
    assert not {
        change.kind
        for change in equivalent_report.changes
        if change.kind == "conditional_formatting_changed"
    }
    changed_change = next(
        change
        for change in changed_report.changes
        if change.kind == "conditional_formatting_changed"
    )
    assert changed_change.details["before"]["extensions"] != changed_change.details["after"][
        "extensions"
    ]
    assert {finding.rule_id for finding in changed_report.findings} >= {"FF021"}
    extension_type_change = next(
        change
        for change in extension_type_report.changes
        if change.kind == "conditional_formatting_changed"
    )
    assert (
        extension_type_change.details["before"]["extensions"][0]["extension"]["attributes"][
            "uri"
        ]
        != extension_type_change.details["after"]["extensions"][0]["extension"]["attributes"][
            "uri"
        ]
    )


def test_protection_controls_are_profiled_without_verifier_or_identity_material(
    tmp_path,
) -> None:
    workbook = make_protection_model(
        tmp_path / "protected.xlsx", include_chartsheet=True
    )
    synthetic_hash = "c3ludGhldGljLWhhc2gtaGFzaA=="
    set_sheet_protection_modern_verifier(workbook, synthetic_hash)

    snapshot = load_snapshot(workbook)
    profile = profile_snapshot(snapshot)
    markdown = profile_to_markdown(profile)
    profile_text = json.dumps(profile)

    assert snapshot.summary()["workbook_protection_enabled"] is True
    assert snapshot.summary()["protected_sheet_count"] == 2
    assert snapshot.summary()["protected_range_count"] == 1
    assert snapshot.summary()["cell_protection_assignment_count"] == 4
    assert profile["sheet_protections"] == [
        {
            "sheet": "Dashboard",
            "sheet_type": "chartsheet",
            "enabled": True,
            "locked_actions": ["content", "objects"],
            "credential": {
                "configured": True,
                "has_legacy_verifier": True,
                "has_modern_verifier": False,
                "algorithm": None,
                "spin_count": None,
            },
            "opaque_metadata": {"present": False, "count": 0},
        },
        {
            "sheet": "Inputs",
            "sheet_type": "worksheet",
            "enabled": True,
            "locked_actions": [
                "format_columns",
                "format_rows",
                "insert_columns",
                "insert_rows",
                "insert_hyperlinks",
                "delete_columns",
                "delete_rows",
                "select_locked_cells",
                "pivot_tables",
            ],
            "credential": {
                "configured": True,
                "has_legacy_verifier": False,
                "has_modern_verifier": True,
                "algorithm": "SHA-512",
                "spin_count": 100000,
            },
            "opaque_metadata": {"present": False, "count": 0},
        },
    ]
    assert profile["protected_ranges"] == [
        {
            "sheet": "Inputs",
            "ranges": ["Inputs!B2:B5"],
            "has_name": True,
            "credential": {
                "configured": True,
                "has_legacy_verifier": True,
                "has_modern_verifier": False,
                "algorithm": None,
                "spin_count": None,
            },
            "has_security_descriptor": True,
            "opaque_metadata": {"present": False, "count": 0},
        }
    ]
    assert profile["cell_protection_default"] == {"locked": True, "hidden": False}
    assert profile["cell_protection_assignments"] == [
        {
            "sheet": "Inputs",
            "scope": "cell",
            "target": "B2",
            "locked": False,
            "hidden": False,
        },
        {
            "sheet": "Inputs",
            "scope": "cell",
            "target": "C2",
            "locked": True,
            "hidden": True,
        },
        {
            "sheet": "Inputs",
            "scope": "column",
            "target": "D:D",
            "locked": False,
            "hidden": False,
        },
        {
            "sheet": "Inputs",
            "scope": "row",
            "target": "5",
            "locked": True,
            "hidden": True,
        },
    ]
    for sensitive_value in (
        synthetic_hash,
        "c3ludGhldGljLXNhbHQ=",
        "Synthetic approved inputs",
        "synthetic-security-descriptor",
    ):
        assert sensitive_value not in profile_text
        assert sensitive_value not in markdown
    assert "## Workbook protection" in markdown
    assert "## Sheet protection controls" in markdown
    assert "## Protected ranges" in markdown
    assert "## Direct cell-protection assignments" in markdown


def test_sheet_protection_defaults_are_canonical_and_verifiers_diff_privately(tmp_path) -> None:
    baseline = make_protection_model(tmp_path / "baseline.xlsx")
    candidate = make_protection_model(tmp_path / "candidate.xlsx")
    set_sheet_protection_defaults(baseline, explicit=False)
    set_sheet_protection_defaults(candidate, explicit=True)

    equivalent_report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert not {
        change.kind
        for change in equivalent_report.changes
        if change.kind == "sheet_protection_changed"
    }
    assert "FF022" not in {finding.rule_id for finding in equivalent_report.findings}

    baseline_hash = "c3ludGhldGljLWhhc2gtYQ=="
    candidate_hash = "c3ludGhldGljLWhhc2gtYg=="
    set_sheet_protection_modern_verifier(baseline, baseline_hash)
    set_sheet_protection_modern_verifier(candidate, candidate_hash)
    verifier_report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change = next(
        change
        for change in verifier_report.changes
        if change.kind == "sheet_protection_changed"
    )

    assert change.details["credential_material_changed"] is True
    assert {finding.rule_id for finding in verifier_report.findings} >= {"FF022"}
    report_text = json.dumps(verifier_report.to_dict())
    assert baseline_hash not in report_text
    assert candidate_hash not in report_text
    assert "c3ludGhldGljLXNhbHQ=" not in report_text


def test_protection_control_changes_cover_workbook_ranges_and_cell_assignments(tmp_path) -> None:
    baseline = make_protection_model(tmp_path / "baseline.xlsx")
    candidate = make_protection_model(tmp_path / "candidate.xlsx")

    def weaken_protection(workbook) -> None:
        workbook.security.lockStructure = False
        inputs = workbook["Inputs"]
        inputs.protection.formatCells = True
        inputs["B2"].protection = Protection(locked=True)

    rewrite(candidate, weaken_protection)
    # openpyxl does not preserve protected ranges, so restore the test fixture's
    # raw range and then make every sensitive field change independently.
    add_protected_range(candidate)
    change_protected_range(
        candidate,
        sqref="B2:B6",
        name="Changed synthetic range name",
        password="C3D4",
        security_descriptor="changed-synthetic-security-descriptor",
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change_kinds = {change.kind for change in report.changes}
    protected_range_change = next(
        change
        for change in report.changes
        if change.kind == "protected_range_permissions_changed"
    )

    assert {
        "workbook_protection_changed",
        "sheet_protection_changed",
        "protected_range_permissions_changed",
        "cell_protection_assignments_changed",
    } <= change_kinds
    assert protected_range_change.details["range_name_material_changed"] is True
    assert protected_range_change.details["security_descriptor_material_changed"] is True
    assert protected_range_change.details["credential_material_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF022"}
    report_text = json.dumps(report.to_dict())
    assert "Synthetic approved inputs" not in report_text
    assert "Changed synthetic range name" not in report_text
    assert "synthetic-security-descriptor" not in report_text
    assert "changed-synthetic-security-descriptor" not in report_text


def test_external_data_refresh_controls_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_external_data_refresh_model(tmp_path / "baseline.xlsx")
    candidate = make_external_data_refresh_model(tmp_path / "candidate.xlsx")
    change_external_data_refresh_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    assert baseline_snapshot.summary()["external_data_connection_count"] == 2
    assert baseline_snapshot.summary()["external_data_connections_refresh_on_load"] == 1
    assert baseline_snapshot.summary()["query_table_refresh_control_count"] == 1
    assert baseline_snapshot.summary()["query_tables_refresh_on_load"] == 1
    assert baseline_snapshot.summary()["pivot_cache_refresh_control_count"] == 1
    assert baseline_snapshot.summary()["pivot_caches_refresh_on_load"] == 1
    assert profile["external_data_refresh_settings"] == {
        "update_links": "always",
        "allow_refresh_query": True,
        "refresh_all_connections": True,
        "save_external_link_values": False,
    }
    assert profile["external_data_connections"][0] == {
        "id": 1,
        "source_type": "ole_db",
        "deleted": False,
        "refresh_on_load": True,
        "refresh_interval_minutes": 60,
        "background": True,
        "keep_alive": True,
        "save_data": False,
        "save_password": True,
        "has_source_file": True,
        "has_connection_file": True,
        "only_use_connection_file": True,
        "reconnection_method": "always",
        "credential_method": "stored",
        "minimum_refreshable_version": 3,
        "has_single_sign_on_id": True,
        "awaiting_initial_refresh": True,
        "has_name": True,
        "has_description": True,
        "source_components": ["database", "parameters"],
        "parameter_count": 1,
        "parameters_refresh_on_change": 1,
        "opaque_metadata": {"present": True, "count": 1},
    }
    assert profile["query_table_refresh_controls"] == [
        {
            "sheet": "Inputs",
            "connection_id": 1,
            "refresh_on_load": True,
            "background_refresh": False,
            "refresh_disabled": False,
            "remove_data_on_save": True,
            "fill_formulas": True,
            "connection_edit_disabled": True,
            "growth_behavior": "overwrite_clear",
            "has_name": True,
            "has_refresh_metadata": False,
            "opaque_metadata": {"present": True, "count": 1},
        }
    ]
    assert profile["pivot_cache_refresh_controls"] == [
        {
            "cache_id": 7,
            "source_type": "external",
            "connection_id": 2,
            "refresh_on_load": True,
            "background_query": True,
            "refresh_enabled": False,
            "save_data": False,
            "upgrade_on_refresh": True,
            "opaque_metadata": {"present": True, "count": 1},
        }
    ]
    assert "## External-data connections" in markdown
    assert "## Query-table refresh controls" in markdown
    assert "## Pivot-cache refresh controls" in markdown

    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change_kinds = {change.kind for change in report.changes}
    connection_change = next(
        change
        for change in report.changes
        if change.kind == "external_data_connections_changed"
    )
    query_table_change = next(
        change
        for change in report.changes
        if change.kind == "query_table_refresh_controls_changed"
    )
    pivot_cache_change = next(
        change
        for change in report.changes
        if change.kind == "pivot_cache_refresh_controls_changed"
    )

    assert {
        "external_data_refresh_settings_changed",
        "external_data_connections_changed",
        "query_table_refresh_controls_changed",
        "pivot_cache_refresh_controls_changed",
    } <= change_kinds
    assert connection_change.details["identity_material_changed"] is True
    assert connection_change.details["source_configuration_material_changed"] is True
    assert query_table_change.details["identity_material_changed"] is True
    assert pivot_cache_change.details["source_configuration_material_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF023"}

    sensitive_values = (
        "synthetic confidential revenue connection",
        "changed synthetic confidential revenue connection",
        "private-baseline-password",
        "private-candidate-password",
        "C:/private/synthetic-revenue-source.accdb",
        "C:/private/changed-synthetic-source.accdb",
        "synthetic-private-sso-identifier",
        "changed-synthetic-private-sso-identifier",
        "synthetic confidential query table",
        "changed synthetic confidential query table",
        "synthetic private pivot extension payload",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_external_data_connection_defaults_are_canonical(tmp_path) -> None:
    baseline = make_external_data_refresh_model(tmp_path / "baseline.xlsx")
    candidate = make_external_data_refresh_model(tmp_path / "candidate.xlsx")
    set_external_data_connection_defaults(baseline, explicit=False)
    set_external_data_connection_defaults(candidate, explicit=True)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert not {
        change.kind
        for change in report.changes
        if change.kind == "external_data_connections_changed"
    }
    assert "FF023" not in {finding.rule_id for finding in report.findings}


def test_package_wide_external_relationships_are_profiled_and_diffed_privately(
    tmp_path,
) -> None:
    baseline = make_external_relationship_model(tmp_path / "baseline.xlsx")
    candidate = make_external_relationship_model(tmp_path / "candidate.xlsx")
    change_external_relationship_target(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    expected_relationships = {
        "present": True,
        "external_relationship_part_count": 2,
        "external_relationship_source_count": 2,
        "external_relationship_count": 3,
        "external_hyperlink_relationship_count": 1,
        "external_image_relationship_count": 1,
        "external_other_relationship_count": 1,
        "unrecognized_relationship_count": 0,
    }
    assert baseline_snapshot.external_relationships.to_dict() == expected_relationships
    assert baseline_snapshot.summary()["package_external_relationship_count"] == 3
    assert baseline_snapshot.summary()["package_external_relationship_source_count"] == 2
    assert baseline_snapshot.summary()["package_external_hyperlink_relationship_count"] == 1
    assert baseline_snapshot.summary()["package_external_image_relationship_count"] == 1
    assert baseline_snapshot.summary()["has_external_relationships"] is True
    assert profile["external_relationships"] == expected_relationships
    assert "## Package-wide external relationships" in markdown
    assert "**Relationship parts / sources / targets:** 2 / 2 / 3" in markdown
    assert "**Hyperlink / image / other targets:** 1 / 1 / 1" in markdown

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    relationship_change = next(
        change
        for change in report.changes
        if change.kind == "external_relationships_changed"
    )
    assert relationship_change.details["before"] == expected_relationships
    assert relationship_change.details["after"] == expected_relationships
    assert relationship_change.details["external_relationship_material_changed"] is True
    assert "FF063" in {finding.rule_id for finding in report.findings}

    sensitive_values = (
        "PRIVATE-PACKAGE-EXTERNAL-BASELINE",
        "PRIVATE-PACKAGE-EXTERNAL-CANDIDATE",
        "PRIVATE-PACKAGE-HYPERLINK-BASELINE",
        "PRIVATE-PACKAGE-IMAGE-BASELINE",
        "https://private.example.test/relationships/opaque-external",
        "rIdFenceOpaqueExternal",
        "rIdFenceOpaqueHyperlink",
        "rIdFenceOpaqueImage",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_formula_external_actions_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_formula_external_action_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_external_action_model(tmp_path / "candidate.xlsx")
    change_formula_external_action_target(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    expected_actions = {
        "present": True,
        "formula_external_action_cell_count": 8,
        "action_defined_name_count": 0,
        "hyperlink_function_count": 5,
        "webservice_function_count": 1,
        "image_function_count": 2,
        "rtd_function_count": 1,
        "stockhistory_function_count": 0,
        "cube_function_count": 0,
    }
    assert baseline_snapshot.formula_external_actions.to_dict() == expected_actions
    assert baseline_snapshot.summary()["formula_external_action_cell_count"] == 8
    assert baseline_snapshot.summary()["formula_external_action_defined_name_count"] == 0
    assert baseline_snapshot.summary()["formula_hyperlink_function_count"] == 5
    assert baseline_snapshot.summary()["formula_webservice_function_count"] == 1
    assert baseline_snapshot.summary()["formula_image_function_count"] == 2
    assert baseline_snapshot.summary()["formula_rtd_function_count"] == 1
    assert baseline_snapshot.summary()["has_formula_external_actions"] is True
    assert profile["formula_external_actions"] == expected_actions
    assert "## Formula external-action and data-provider surfaces" in markdown
    assert "**Formula cells / function calls / formula-defined names:** 8 / 9 / 0" in markdown
    assert "**HYPERLINK / WEBSERVICE / IMAGE / RTD:** 5 / 1 / 2 / 1" in markdown

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_external_actions_changed"
    )
    action_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF064"
    )
    assert action_change.details["before"] == expected_actions
    assert action_change.details["after"] == expected_actions
    assert action_change.details["formula_external_action_material_changed"] is True
    assert action_finding.details == action_change.details
    assert "FF004" not in {finding.rule_id for finding in report.findings}
    assert "FF047" not in {finding.rule_id for finding in report.findings}
    assert "FF063" not in {finding.rule_id for finding in report.findings}

    # The ordinary semantic-cell diff deliberately retains changed formulas for
    # reviewers. The action ledger, its profile, Markdown/SARIF findings, and
    # policy-facing details do not retain the formula's private arguments.
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(action_change.details),
        json.dumps(action_finding.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-LINK-BASELINE",
        "PRIVATE-LINK-CANDIDATE",
        "PRIVATE-WEBSERVICE",
        "PRIVATE-IMAGE.png",
        "PRIVATE-NAMESPACED-IMAGE.png",
        "PRIVATE.FormulaFence.Provider",
        "PRIVATE-SERVER",
        "PRIVATE-SECOND-LINK",
        "PRIVATE-THIRD-LINK",
        "PRIVATE-REFERENCED-LINK-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_direct_formula_dde_links_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_formula_dde_link_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_dde_link_model(tmp_path / "candidate.xlsx")
    change_formula_dde_link_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    expected_links = {
        "present": True,
        "dde_formula_cell_count": 4,
        "dde_link_count": 4,
        "dde_defined_name_count": 3,
    }

    assert baseline_snapshot.formula_dde_links.to_dict() == expected_links
    assert baseline_snapshot.formula_dde_links.dde_cells == frozenset(
        {("Inputs", f"B{row}") for row in range(2, 6)}
    )
    assert baseline_snapshot.summary()["formula_dde_link_formula_cell_count"] == 4
    assert baseline_snapshot.summary()["formula_dde_link_count"] == 4
    assert baseline_snapshot.summary()["formula_dde_link_defined_name_count"] == 3
    assert baseline_snapshot.summary()["has_formula_dde_links"] is True
    assert profile["formula_dde_links"] == expected_links
    assert profile["features"]["has_formula_dde_links"] is True
    assert "## Direct DDE-style formula links" in markdown
    assert "**Formula cells / DDE links / formula-defined names:** 4 / 4 / 3" in markdown

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    dde_change = next(
        change for change in report.changes if change.kind == "formula_dde_links_changed"
    )
    dde_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF074"
    )
    assert dde_change.details["before"] == expected_links
    assert dde_change.details["after"] == expected_links
    assert dde_change.details["formula_dde_link_definition_material_changed"] is True
    assert dde_finding.details == dde_change.details

    ff074_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF074"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["formula_dde_links"]),
        markdown,
        json.dumps(dde_change.details),
        json.dumps(dde_finding.to_dict()),
        json.dumps(ff074_sarif_result),
    )
    for sensitive_value in (
        "FENCE.DDE",
        "PRIVATE-DDE-DIRECT-TOPIC-BASELINE",
        "PRIVATE-DDE-NAMED-TOPIC-BASELINE",
        "PRIVATE-DDE-NAMED-TOPIC-CANDIDATE",
        "PRIVATE-DDE-LAMBDA-TOPIC-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_direct_formula_dde_link_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_formula_dde_link_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_dde_link_model(tmp_path / "candidate.xlsx")
    change_formula_dde_link_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.formula_dde_links == candidate_snapshot.formula_dde_links
    assert ("Inputs", "B5") in baseline_snapshot.reverse_dependencies[("Inputs", "A9")]

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    dde_change = next(
        change for change in report.changes if change.kind == "formula_dde_links_changed"
    )
    assert dde_change.details["formula_dde_link_static_input_changed"] is True
    assert dde_change.details["formula_dde_link_static_input_change_count"] == 1
    assert "formula_dde_link_invocation_material_changed" not in dde_change.details
    assert "formula_dde_link_definition_material_changed" not in dde_change.details
    assert "FF074" in {finding.rule_id for finding in report.findings}


def test_uninvoked_direct_formula_dde_name_is_profiled_without_an_endpoint(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "stored-dde-name.xlsx")

    def add_stored_dde_link(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.DDE.STORED",
                attr_text="=DDE.TEST|'PRIVATE-UNINVOKED-DDE-TOPIC'!PRIVATE_ITEM",
            )
        )

    rewrite(workbook_path, add_stored_dde_link)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_dde_links.to_dict() == {
        "present": True,
        "dde_formula_cell_count": 0,
        "dde_link_count": 0,
        "dde_defined_name_count": 1,
    }
    assert snapshot.formula_dde_links.dde_cells == frozenset()


def test_direct_formula_dde_names_follow_sheet_local_precedence(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "scoped-dde-name.xlsx")

    def add_scoped_dde_names(workbook) -> None:
        model = workbook["Model"]
        report = workbook["Dashboard"]
        model["D2"] = "=FENCE.DDE.SCOPED"
        report["D2"] = "=FENCE.DDE.SCOPED"
        workbook.defined_names.add(
            DefinedName(
                "FENCE.DDE.SCOPED",
                attr_text="=DDE.TEST|'PRIVATE-GLOBAL-DDE-TOPIC'!PRIVATE_ITEM",
            )
        )
        model.defined_names.add(
            DefinedName(
                "FENCE.DDE.SCOPED",
                attr_text="=DDE.TEST|'PRIVATE-LOCAL-DDE-TOPIC'!PRIVATE_ITEM",
                localSheetId=1,
            )
        )

    rewrite(workbook_path, add_scoped_dde_names)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_dde_links.to_dict() == {
        "present": True,
        "dde_formula_cell_count": 2,
        "dde_link_count": 2,
        "dde_defined_name_count": 2,
    }
    assert snapshot.formula_dde_links.dde_cells == frozenset(
        {("Model", "D2"), ("Dashboard", "D2")}
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_formula_external_action_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_formula_external_action_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_external_action_model(tmp_path / "candidate.xlsx")
    change_formula_external_action_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.formula_external_actions == candidate_snapshot.formula_external_actions

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_external_actions_changed"
    )
    assert action_change.details["formula_external_action_static_input_changed"] is True
    assert action_change.details["formula_external_action_static_input_change_count"] == 1
    assert "formula_external_action_material_changed" not in action_change.details
    assert "FF064" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(action_change.details),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-REFERENCED-LINK-BASELINE",
        "PRIVATE-REFERENCED-LINK-CANDIDATE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_formula_external_data_provider_calls_are_profiled_diffed_and_redacted(
    tmp_path,
) -> None:
    baseline = make_formula_external_data_provider_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_external_data_provider_model(tmp_path / "candidate.xlsx")
    change_formula_external_data_provider_target(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    expected_actions = {
        "present": True,
        "formula_external_action_cell_count": 11,
        "action_defined_name_count": 3,
        "hyperlink_function_count": 0,
        "webservice_function_count": 0,
        "image_function_count": 0,
        "rtd_function_count": 0,
        "stockhistory_function_count": 3,
        "cube_function_count": 8,
    }
    assert baseline_snapshot.formula_external_actions.to_dict() == expected_actions
    assert baseline_snapshot.summary()["formula_stockhistory_function_count"] == 3
    assert baseline_snapshot.summary()["formula_cube_function_count"] == 8
    assert profile["formula_external_actions"] == expected_actions
    assert "## Formula external-action and data-provider surfaces" in markdown
    assert "**Formula cells / function calls / formula-defined names:** 11 / 11 / 3" in markdown
    assert "**STOCKHISTORY / Cube functions:** 3 / 8" in markdown

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_external_actions_changed"
    )
    action_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF064"
    )
    assert action_change.details["before"] == expected_actions
    assert action_change.details["after"] == expected_actions
    assert action_change.details["formula_external_action_material_changed"] is True
    assert action_finding.details == action_change.details

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(action_change.details),
        json.dumps(action_finding.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-STOCK-BASELINE",
        "PRIVATE-CUBE-CONNECTION-BASELINE",
        "PRIVATE-CUBE-CONNECTION-CANDIDATE",
        "PRIVATE-REVENUE",
        "PRIVATE-MEMBER-PROPERTY",
        "PRIVATE-CUBE-CAPTION",
        "PRIVATE-KPI",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_named_formula_external_data_provider_calls_are_propagated_and_private(
    tmp_path,
) -> None:
    baseline = make_formula_external_data_provider_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_external_data_provider_model(tmp_path / "candidate.xlsx")
    change_formula_external_data_provider_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.formula_external_actions.action_cells == frozenset(
        {("Inputs", f"B{row}") for row in range(2, 13)}
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_external_actions_changed"
    )
    assert action_change.details[
        "formula_external_action_definition_material_changed"
    ] is True
    ff064_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF064"
    )

    rendered_ledger_artifacts = (
        json.dumps(profile_snapshot(baseline_snapshot)["formula_external_actions"]),
        json.dumps(action_change.details),
        json.dumps(ff064_sarif_result),
    )
    for sensitive_value in (
        "FENCE",
        "PRIVATE-NAMED-CUBE-CONNECTION-BASELINE",
        "PRIVATE-NAMED-CUBE-CONNECTION-CANDIDATE",
        "PRIVATE-NAMED-REVENUE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_formula_external_data_provider_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_formula_external_data_provider_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_external_data_provider_model(tmp_path / "candidate.xlsx")
    change_formula_external_data_provider_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.formula_external_actions == candidate_snapshot.formula_external_actions
    assert {("Inputs", "B2"), ("Inputs", "B10"), ("Inputs", "B11")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A2")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_external_actions_changed"
    )
    assert action_change.details["formula_external_action_static_input_changed"] is True
    assert action_change.details["formula_external_action_static_input_change_count"] == 1
    assert "formula_external_action_material_changed" not in action_change.details
    assert "formula_external_action_definition_material_changed" not in action_change.details
    assert "FF064" in {finding.rule_id for finding in report.findings}


def test_formula_external_data_provider_respects_named_lambda_shadowing(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "shadowed-data-provider.xlsx")

    def add_shadowing_lambdas(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "STOCKHISTORY",
                attr_text="=LAMBDA(stock,start_date,stock)",
            )
        )
        workbook.defined_names.add(
            DefinedName(
                "CUBEVALUE",
                attr_text="=LAMBDA(connection,member,connection)",
            )
        )
        workbook["Model"]["D2"] = "=STOCKHISTORY(Inputs!B2,Inputs!B3)"
        workbook["Model"]["D3"] = "=CUBEVALUE(Inputs!A1,Inputs!B2)"

    rewrite(workbook_path, add_shadowing_lambdas)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_external_actions.to_dict() == {
        "present": False,
        "formula_external_action_cell_count": 0,
        "action_defined_name_count": 0,
        "hyperlink_function_count": 0,
        "webservice_function_count": 0,
        "image_function_count": 0,
        "rtd_function_count": 0,
        "stockhistory_function_count": 0,
        "cube_function_count": 0,
    }
    assert snapshot.unresolved_reference_tokens == {}


def test_named_formula_external_actions_are_propagated_diffed_and_private(tmp_path) -> None:
    baseline = make_named_formula_external_action_model(tmp_path / "baseline.xlsx")
    candidate = make_named_formula_external_action_model(tmp_path / "candidate.xlsx")
    change_named_formula_external_action_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_actions = {
        "present": True,
        "formula_external_action_cell_count": 3,
        "action_defined_name_count": 3,
        "hyperlink_function_count": 2,
        "webservice_function_count": 1,
        "image_function_count": 0,
        "rtd_function_count": 0,
        "stockhistory_function_count": 0,
        "cube_function_count": 0,
    }
    assert baseline_snapshot.formula_external_actions.to_dict() == expected_actions
    assert baseline_snapshot.formula_external_actions.action_cells == frozenset(
        {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")}
    )
    assert profile["formula_external_actions"] == expected_actions

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_external_actions_changed"
    )
    action_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF064"
    )
    assert action_change.details["before"] == expected_actions
    assert action_change.details["after"] == expected_actions
    assert action_change.details["formula_external_action_definition_material_changed"] is True
    assert action_finding.details == action_change.details

    ff064_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF064"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["formula_external_actions"]),
        profile_to_markdown(profile),
        json.dumps(action_change.details),
        json.dumps(action_finding.to_dict()),
        json.dumps(ff064_sarif_result),
    )
    for sensitive_value in (
        "FENCE",
        "PRIVATE-NAMED-ACTION-LABEL-BASELINE",
        "PRIVATE-NAMED-ACTION-LABEL-CANDIDATE",
        "PRIVATE-NAMED-ACTION-INPUT-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_formula_external_action_report_redaction_hides_resolved_named_chain_inputs(
    tmp_path,
) -> None:
    """A name that calls an action-bearing name can still carry an endpoint."""
    baseline = make_named_formula_external_action_model(tmp_path / "baseline.xlsx")
    candidate = make_named_formula_external_action_model(tmp_path / "candidate.xlsx")
    private_marker = "PRIVATE-INDIRECT-NAMED-ENDPOINT-CANDIDATE"
    workbook = load_workbook(candidate)
    workbook.defined_names["FENCE.CHAIN"].attr_text = (
        '=LAMBDA(value,FENCE.WRAPPER("PRIVATE-INDIRECT-NAMED-ENDPOINT-CANDIDATE"))'
    )
    workbook.save(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    assert (
        report.before.formula_external_actions.definition_signature
        != report.after.formula_external_actions.definition_signature
    )
    assert private_marker in json.dumps(report.to_dict())

    redacted_payload = redact_formula_external_action_report_payload(report, report.to_dict())
    assert private_marker not in json.dumps(redacted_payload)
    changed_name = next(
        change
        for change in redacted_payload["changes"]
        if change["kind"] == "defined_name_changed"
    )
    assert changed_name["details"]["before"] == FORMULA_EXTERNAL_ACTION_REDACTION
    assert changed_name["details"]["after"] == FORMULA_EXTERNAL_ACTION_REDACTION

    redacted_sarif = report_to_sarif(report, redact_formula_external_actions=True)
    assert private_marker not in json.dumps(redacted_sarif)
    ff008 = next(
        result
        for result in redacted_sarif["runs"][0]["results"]
        if result["ruleId"] == "FF008"
    )
    assert ff008["properties"]["before"] == FORMULA_EXTERNAL_ACTION_REDACTION
    assert ff008["properties"]["after"] == FORMULA_EXTERNAL_ACTION_REDACTION


def test_named_formula_external_action_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_named_formula_external_action_model(tmp_path / "baseline.xlsx")
    candidate = make_named_formula_external_action_model(tmp_path / "candidate.xlsx")
    change_named_formula_external_action_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.formula_external_actions == candidate_snapshot.formula_external_actions
    assert {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_external_actions_changed"
    )
    assert action_change.details["formula_external_action_static_input_changed"] is True
    assert action_change.details["formula_external_action_static_input_change_count"] == 1
    assert "formula_external_action_material_changed" not in action_change.details
    assert "formula_external_action_definition_material_changed" not in action_change.details
    assert "FF064" in {finding.rule_id for finding in report.findings}


def test_uninvoked_formula_defined_external_action_is_profiled(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "stored-name.xlsx")

    def add_stored_action(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.STORED.ACTION",
                attr_text=(
                    '=HYPERLINK("https://private.example.test/PRIVATE-STORED-NAMED-ACTION",'
                    '"Open")'
                ),
            )
        )

    rewrite(workbook_path, add_stored_action)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_external_actions.to_dict() == {
        "present": True,
        "formula_external_action_cell_count": 0,
        "action_defined_name_count": 1,
        "hyperlink_function_count": 0,
        "webservice_function_count": 0,
        "image_function_count": 0,
        "rtd_function_count": 0,
        "stockhistory_function_count": 0,
        "cube_function_count": 0,
    }
    assert snapshot.formula_external_actions.action_cells == frozenset()


def test_recursive_named_formula_external_actions_are_cycle_safe(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "recursive.xlsx")

    def add_recursive_action(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.LOOP",
                attr_text=(
                    '=LAMBDA(value,HYPERLINK(value,"PRIVATE-RECURSIVE-ACTION")'
                    "+FENCE.LOOP(value))"
                ),
            )
        )
        workbook["Model"]["D2"] = "=FENCE.LOOP(Inputs!B2)"

    rewrite(workbook_path, add_recursive_action)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_external_actions.to_dict() == {
        "present": True,
        "formula_external_action_cell_count": 1,
        "action_defined_name_count": 1,
        "hyperlink_function_count": 1,
        "webservice_function_count": 0,
        "image_function_count": 0,
        "rtd_function_count": 0,
        "stockhistory_function_count": 0,
        "cube_function_count": 0,
    }
    assert snapshot.formula_external_actions.action_cells == frozenset(
        {("Model", "D2")}
    )
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == ("FENCE.LOOP",)


def test_python_in_excel_code_is_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_python_in_excel_model(tmp_path / "baseline.xlsx")
    candidate = make_python_in_excel_model(tmp_path / "candidate.xlsx")
    change_python_in_excel_script(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    expected_python = {
        "present": True,
        "python_part_count": 1,
        "python_formula_cell_count": 2,
        "python_function_count": 2,
        "python_script_count": 2,
        "python_environment_definition_count": 1,
        "python_initialization_count": 1,
        "unrecognized_python_in_excel_count": 0,
    }
    assert baseline_snapshot.python_in_excel.to_dict() == expected_python
    assert baseline_snapshot.summary()["python_in_excel_part_count"] == 1
    assert baseline_snapshot.summary()["python_in_excel_formula_cell_count"] == 2
    assert baseline_snapshot.summary()["python_in_excel_function_count"] == 2
    assert baseline_snapshot.summary()["python_in_excel_script_count"] == 2
    assert baseline_snapshot.summary()["python_in_excel_environment_definition_count"] == 1
    assert baseline_snapshot.summary()["python_in_excel_initialization_count"] == 1
    assert baseline_snapshot.summary()["has_python_in_excel"] is True
    assert profile["python_in_excel"] == expected_python
    assert profile["features"]["has_python_in_excel"] is True
    assert "## Python in Excel code" in markdown
    assert "**Package parts / PY formula cells / PY calls:** 1 / 2 / 2" in markdown
    assert "**Stored scripts / environment definitions / initializations:** 2 / 1 / 1" in markdown

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    python_change = next(
        change
        for change in report.changes
        if change.kind == "python_in_excel_changed"
    )
    python_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF065"
    )
    assert python_change.details["before"] == expected_python
    assert python_change.details["after"] == expected_python
    assert python_change.details["python_in_excel_definition_changed"] is True
    assert python_finding.details == python_change.details
    assert "FF064" not in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(python_change.details),
        json.dumps(python_finding.to_dict()),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-PYTHON-INIT-BASELINE",
        "PRIVATE-PYTHON-SCRIPT-BASELINE",
        "PRIVATE-PYTHON-SCRIPT-SECOND",
        "PRIVATE-PYTHON-SCRIPT-CANDIDATE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_python_in_excel_python_scripts_contract_is_profiled_and_private(
    tmp_path,
) -> None:
    baseline = make_python_in_excel_scripts_model(tmp_path / "baseline.xlsx")
    candidate = make_python_in_excel_scripts_model(tmp_path / "candidate.xlsx")
    change_python_in_excel_scripts_script(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    expected_python = {
        "present": True,
        "python_part_count": 1,
        "python_formula_cell_count": 1,
        "python_function_count": 1,
        "python_script_count": 2,
        "python_environment_definition_count": 0,
        "python_initialization_count": 0,
        "unrecognized_python_in_excel_count": 0,
    }
    assert baseline_snapshot.python_in_excel.to_dict() == expected_python
    assert profile["python_in_excel"] == expected_python
    assert "**Package parts / PY formula cells / PY calls:** 1 / 1 / 1" in markdown
    assert "**Stored scripts / environment definitions / initializations:** 2 / 0 / 0" in markdown

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    python_change = next(
        change
        for change in report.changes
        if change.kind == "python_in_excel_changed"
    )
    assert python_change.details["before"] == expected_python
    assert python_change.details["after"] == expected_python
    assert python_change.details["python_in_excel_definition_changed"] is True
    assert "FF065" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-PYTHON-SCRIPTS-BASELINE",
        "PRIVATE-PYTHON-SCRIPTS-SECOND",
        "PRIVATE-PYTHON-SCRIPTS-CANDIDATE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_python_in_excel_python_scripts_contract_is_guarded_alongside_python_part(
    tmp_path,
) -> None:
    baseline = make_python_in_excel_model(tmp_path / "baseline.xlsx")
    add_python_in_excel_scripts_compatibility_part(baseline)
    candidate = tmp_path / "candidate.xlsx"
    shutil.copyfile(baseline, candidate)
    change_python_in_excel_scripts_script(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.python_in_excel.to_dict() == {
        "present": True,
        "python_part_count": 2,
        "python_formula_cell_count": 2,
        "python_function_count": 2,
        "python_script_count": 4,
        "python_environment_definition_count": 1,
        "python_initialization_count": 1,
        "unrecognized_python_in_excel_count": 0,
    }

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    python_change = next(
        change
        for change in report.changes
        if change.kind == "python_in_excel_changed"
    )
    assert python_change.details["python_in_excel_definition_changed"] is True
    assert "FF065" in {finding.rule_id for finding in report.findings}


def test_python_in_excel_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_python_in_excel_model(tmp_path / "baseline.xlsx")
    candidate = make_python_in_excel_model(tmp_path / "candidate.xlsx")
    change_python_in_excel_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.python_in_excel == candidate_snapshot.python_in_excel

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    python_change = next(
        change
        for change in report.changes
        if change.kind == "python_in_excel_changed"
    )
    assert python_change.details["python_in_excel_static_input_changed"] is True
    assert python_change.details["python_in_excel_static_input_change_count"] == 1
    assert "python_in_excel_definition_changed" not in python_change.details
    assert "python_in_excel_formula_binding_changed" not in python_change.details
    assert "FF065" in {finding.rule_id for finding in report.findings}
    assert report.python_in_excel_static_input_cells == frozenset({("Inputs", "A9")})


def test_python_in_excel_shared_report_redaction_hides_source_and_static_inputs(
    tmp_path,
) -> None:
    source_baseline = "PRIVATE-PY-SOURCE-BASELINE"
    source_candidate = "PRIVATE-PY-SOURCE-CANDIDATE"
    input_baseline = "PRIVATE-PY-INPUT-BASELINE"
    input_candidate = "PRIVATE-PY-INPUT-CANDIDATE"
    baseline = make_python_in_excel_model(
        tmp_path / "baseline.xlsx", input_value=input_baseline
    )
    candidate = make_python_in_excel_model(
        tmp_path / "candidate.xlsx", input_value=input_candidate
    )
    set_python_in_excel_formula_source(baseline, source_baseline)
    set_python_in_excel_formula_source(candidate, source_candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    assert report.python_in_excel_static_input_cells == frozenset({("Inputs", "A9")})
    assert "FF065" in {finding.rule_id for finding in report.findings}

    default_payload = report.to_dict()
    default_rendered = json.dumps(default_payload)
    for sensitive_value in (
        source_baseline,
        source_candidate,
        input_baseline,
        input_candidate,
    ):
        assert sensitive_value in default_rendered

    action_only_rendered = json.dumps(
        redact_formula_external_action_report_payload(report, report.to_dict())
    )
    for sensitive_value in (
        source_baseline,
        source_candidate,
        input_baseline,
        input_candidate,
    ):
        assert sensitive_value in action_only_rendered

    redacted_payload = redact_python_in_excel_report_payload(report, default_payload)
    redacted_rendered = json.dumps(redacted_payload)
    for sensitive_value in (
        source_baseline,
        source_candidate,
        input_baseline,
        input_candidate,
    ):
        assert sensitive_value not in redacted_rendered
    assert PYTHON_IN_EXCEL_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == default_payload["summary"]

    redacted_markdown = report_to_markdown(report, redact_python_in_excel=True)
    redacted_sarif = report_to_sarif(report, redact_python_in_excel=True)
    assert "Python-in-Excel material:** redacted for sharing" in redacted_markdown
    assert "FF065" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, json.dumps(redacted_sarif)):
        for sensitive_value in (
            source_baseline,
            source_candidate,
            input_baseline,
            input_candidate,
        ):
            assert sensitive_value not in artifact

    local_lambda = '=LET(PY,LAMBDA(value,value),PY("ordinary local lambda"))'
    assert redact_python_in_excel_material(local_lambda) == local_lambda


def test_python_in_excel_formula_bindings_are_guarded_separately(tmp_path) -> None:
    baseline = make_python_in_excel_model(tmp_path / "baseline.xlsx")
    candidate = make_python_in_excel_model(tmp_path / "candidate.xlsx")
    change_python_in_excel_formula_binding(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.python_in_excel.definition_signature
        == candidate_snapshot.python_in_excel.definition_signature
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    python_change = next(
        change
        for change in report.changes
        if change.kind == "python_in_excel_changed"
    )
    assert python_change.details["python_in_excel_formula_binding_changed"] is True
    assert "python_in_excel_definition_changed" not in python_change.details
    assert "FF065" in {finding.rule_id for finding in report.findings}
    assert "FF064" not in {finding.rule_id for finding in report.findings}


def test_office_custom_function_calls_are_profiled_diffed_and_private(tmp_path) -> None:
    baseline = make_office_custom_function_model(tmp_path / "baseline.xlsx")
    candidate = make_office_custom_function_model(tmp_path / "candidate.xlsx")
    change_office_custom_function_call(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    expected_functions = {
        "present": True,
        "namespaced_custom_function_formula_cell_count": 4,
        "namespaced_custom_function_call_count": 5,
        "namespaced_custom_function_namespace_count": 2,
        "namespaced_custom_function_defined_name_count": 0,
    }
    assert baseline_snapshot.office_custom_functions.to_dict() == expected_functions
    assert (
        baseline_snapshot.summary()["namespaced_custom_function_formula_cell_count"]
        == 4
    )
    assert baseline_snapshot.summary()["namespaced_custom_function_call_count"] == 5
    assert (
        baseline_snapshot.summary()["namespaced_custom_function_namespace_count"]
        == 2
    )
    assert baseline_snapshot.summary()["has_namespaced_custom_function_calls"] is True
    assert profile["office_custom_functions"] == expected_functions
    assert profile["features"]["has_namespaced_custom_function_calls"] is True
    assert "## Namespaced custom-function calls" in markdown
    assert "**Formula cells / calls / namespaces / named definitions:** 4 / 5 / 2 / 0" in markdown

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    function_change = next(
        change
        for change in report.changes
        if change.kind == "office_custom_functions_changed"
    )
    function_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF066"
    )
    assert function_change.details["before"] == expected_functions
    assert function_change.details["after"] == expected_functions
    assert function_change.details["office_custom_function_material_changed"] is True
    assert function_finding.details == function_change.details
    assert "FF064" not in {finding.rule_id for finding in report.findings}
    assert "FF065" not in {finding.rule_id for finding in report.findings}

    # Formula edits remain in the ordinary semantic report by design. The
    # candidate ledger and its policy-facing finding keep function names,
    # formula arguments, and source values private.
    rendered_ledger_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(function_change.details),
        json.dumps(function_finding.to_dict()),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "CONTOSO",
        "FENCE",
        "GETMARKETDATA",
        "GETRISKDATA",
        "PRIVATE-CUSTOM-FUNCTION-QUERY-BASELINE",
        "PRIVATE-CUSTOM-FUNCTION-RISK-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_office_custom_function_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_office_custom_function_model(tmp_path / "baseline.xlsx")
    candidate = make_office_custom_function_model(tmp_path / "candidate.xlsx")
    change_office_custom_function_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.office_custom_functions == candidate_snapshot.office_custom_functions

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    function_change = next(
        change
        for change in report.changes
        if change.kind == "office_custom_functions_changed"
    )
    assert function_change.details["office_custom_function_static_input_changed"] is True
    assert function_change.details["office_custom_function_static_input_change_count"] == 1
    assert "office_custom_function_material_changed" not in function_change.details
    assert report.office_custom_function_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )
    assert "FF066" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(function_change.details),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-CUSTOM-FUNCTION-INPUT-BASELINE",
        "PRIVATE-CUSTOM-FUNCTION-INPUT-CANDIDATE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_named_custom_function_calls_are_propagated_diffed_and_private(tmp_path) -> None:
    baseline = make_named_office_custom_function_model(tmp_path / "baseline.xlsx")
    candidate = make_named_office_custom_function_model(tmp_path / "candidate.xlsx")
    change_named_office_custom_function_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_functions = {
        "present": True,
        "namespaced_custom_function_formula_cell_count": 3,
        "namespaced_custom_function_call_count": 5,
        "namespaced_custom_function_namespace_count": 1,
        "namespaced_custom_function_defined_name_count": 3,
    }
    assert baseline_snapshot.office_custom_functions.to_dict() == expected_functions
    assert baseline_snapshot.office_custom_functions.call_cells == frozenset(
        {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")}
    )
    assert profile["office_custom_functions"] == expected_functions

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    function_change = next(
        change
        for change in report.changes
        if change.kind == "office_custom_functions_changed"
    )
    function_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF066"
    )
    assert function_change.details["before"] == expected_functions
    assert function_change.details["after"] == expected_functions
    assert function_change.details["office_custom_function_material_changed"] is True
    assert function_change.details["office_custom_function_definition_changed"] is True
    assert function_finding.details == function_change.details

    ff066_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF066"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["office_custom_functions"]),
        profile_to_markdown(profile),
        json.dumps(function_change.details),
        json.dumps(function_finding.to_dict()),
        json.dumps(ff066_sarif_result),
    )
    for sensitive_value in (
        "CONTOSO",
        "FENCE",
        "GETDATA",
        "GETRISK",
        "PRIVATE-NAMED-CUSTOM-FUNCTION-INPUT-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)

    # The ordinary profile preserves defined-name labels and FF008 preserves a
    # changed definition for general workbook review. The dedicated FF066
    # ledger/result above is the intentionally redacted custom-call boundary.
    rendered_profile = json.dumps(profile)
    for sensitive_value in (
        "CONTOSO",
        "GETDATA",
        "GETRISK",
        "PRIVATE-NAMED-CUSTOM-FUNCTION-INPUT-BASELINE",
    ):
        assert sensitive_value not in rendered_profile


def test_named_custom_function_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_named_office_custom_function_model(tmp_path / "baseline.xlsx")
    candidate = make_named_office_custom_function_model(tmp_path / "candidate.xlsx")
    change_named_office_custom_function_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.office_custom_functions == candidate_snapshot.office_custom_functions
    assert {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    function_change = next(
        change
        for change in report.changes
        if change.kind == "office_custom_functions_changed"
    )
    assert function_change.details["office_custom_function_static_input_changed"] is True
    assert function_change.details["office_custom_function_static_input_change_count"] == 1
    assert "office_custom_function_material_changed" not in function_change.details
    assert "FF066" in {finding.rule_id for finding in report.findings}


def test_office_custom_function_report_redaction_hides_calls_and_static_inputs(
    tmp_path,
) -> None:
    baseline = make_office_custom_function_model(tmp_path / "baseline.xlsx")
    candidate = make_office_custom_function_model(tmp_path / "candidate.xlsx")
    change_office_custom_function_input(candidate)
    change_office_custom_function_call(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    sensitive_values = (
        "CONTOSO",
        "GETMARKETDATA",
        "GETRISKDATA",
        "PRIVATE-CUSTOM-FUNCTION-QUERY-BASELINE",
        "PRIVATE-CUSTOM-FUNCTION-INPUT-BASELINE",
        "PRIVATE-CUSTOM-FUNCTION-INPUT-CANDIDATE",
    )
    raw_rendered = json.dumps(raw_payload)
    assert all(value in raw_rendered for value in sensitive_values)

    redacted_payload = redact_office_custom_function_report_payload(report, raw_payload)
    redacted_rendered = json.dumps(redacted_payload)
    assert all(value not in redacted_rendered for value in sensitive_values)
    assert OFFICE_CUSTOM_FUNCTION_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == raw_payload["summary"]

    redacted_markdown = report_to_markdown(
        report, redact_office_custom_functions=True
    )
    redacted_sarif = report_to_sarif(report, redact_office_custom_functions=True)
    assert "Office custom-function material:** redacted for sharing" in redacted_markdown
    assert "FF066" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, json.dumps(redacted_sarif)):
        assert all(value not in artifact for value in sensitive_values)

    native_formula = "=ECMA.CEILING(4.1,1)"
    assert redact_office_custom_function_material(native_formula) == native_formula


def test_office_custom_function_report_redaction_hides_unsampled_static_inputs(
    tmp_path,
) -> None:
    """Custom-function input redaction must use full private impact evidence."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-CUSTOM-FUNCTION-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-CUSTOM-FUNCTION-CANDIDATE"

    def add_custom_function_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        # Z1 is a custom-function consumer after the report's B1:U1 sample.
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = '=CONTOSO.GETDATA(A9,"private-query")'

    rewrite(
        baseline,
        lambda workbook: add_custom_function_fanout(workbook, baseline_marker),
    )
    rewrite(
        candidate,
        lambda workbook: add_custom_function_fanout(workbook, candidate_marker),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.office_custom_function_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )

    redacted = redact_office_custom_function_report_payload(report, raw_payload)
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert redacted_change["before"]["value"] == OFFICE_CUSTOM_FUNCTION_REDACTION
    assert redacted_change["after"]["value"] == OFFICE_CUSTOM_FUNCTION_REDACTION
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_office_custom_function_report_redaction_hides_indirect_named_chain(
    tmp_path,
) -> None:
    baseline = make_indirect_named_office_custom_function_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_indirect_named_office_custom_function_model(
        tmp_path / "candidate.xlsx"
    )
    change_indirect_named_office_custom_function_definition(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    function_change = next(
        change
        for change in report.changes
        if change.kind == "office_custom_functions_changed"
    )
    assert function_change.details["office_custom_function_definition_changed"] is True
    assert "office_custom_function_material_changed" not in function_change.details

    baseline_marker = "PRIVATE-INDIRECT-NAMED-CUSTOM-FUNCTION-BASELINE"
    candidate_marker = "PRIVATE-INDIRECT-NAMED-CUSTOM-FUNCTION-CANDIDATE"
    ordinary_looking_definition = (
        f'=LAMBDA(value,CUSTOMWRAPPER("{baseline_marker}"))'
    )
    # Without private name-chain resolution this direct text is not a
    # namespaced call, which is why the report helper must use FF066's private
    # definition signature rather than rely only on lexical redaction.
    assert redact_office_custom_function_material(ordinary_looking_definition) == (
        ordinary_looking_definition
    )

    raw_payload = report.to_dict()
    raw_rendered = json.dumps(raw_payload)
    assert baseline_marker in raw_rendered
    assert candidate_marker in raw_rendered

    redacted_payload = redact_office_custom_function_report_payload(report, raw_payload)
    redacted_rendered = json.dumps(redacted_payload)
    assert baseline_marker not in redacted_rendered
    assert candidate_marker not in redacted_rendered
    assert OFFICE_CUSTOM_FUNCTION_REDACTION in redacted_rendered

    redacted_sarif = report_to_sarif(report, redact_office_custom_functions=True)
    redacted_sarif_rendered = json.dumps(redacted_sarif)
    assert baseline_marker not in redacted_sarif_rendered
    assert candidate_marker not in redacted_sarif_rendered
    assert "FF008" in redacted_sarif_rendered
    assert "FF066" in redacted_sarif_rendered


def test_unqualified_runtime_function_candidates_are_profiled_diffed_and_private(
    tmp_path,
) -> None:
    baseline = make_unqualified_runtime_function_model(tmp_path / "baseline.xlsx")
    candidate = make_unqualified_runtime_function_model(tmp_path / "candidate.xlsx")
    change_unqualified_runtime_function_call(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    expected_functions = {
        "present": True,
        "unqualified_runtime_function_formula_cell_count": 3,
        "unqualified_runtime_function_call_count": 4,
        "unqualified_runtime_function_defined_name_count": 0,
    }
    assert baseline_snapshot.unqualified_runtime_functions.to_dict() == expected_functions
    assert baseline_snapshot.unqualified_runtime_functions.call_cells == frozenset(
        {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")}
    )
    assert (
        baseline_snapshot.summary()["unqualified_runtime_function_formula_cell_count"]
        == 3
    )
    assert baseline_snapshot.summary()["unqualified_runtime_function_call_count"] == 4
    assert (
        baseline_snapshot.summary()["unqualified_runtime_function_defined_name_count"]
        == 0
    )
    assert baseline_snapshot.summary()["has_unqualified_runtime_function_calls"] is True
    assert profile["unqualified_runtime_functions"] == expected_functions
    assert profile["features"]["has_unqualified_runtime_function_calls"] is True
    assert "## Unqualified runtime-function candidates" in markdown
    assert "**Formula cells / calls / relevant named definitions:** 3 / 4 / 0" in markdown
    assert baseline_snapshot.office_custom_functions.present is False

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    function_change = next(
        change
        for change in report.changes
        if change.kind == "unqualified_runtime_functions_changed"
    )
    function_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF075"
    )
    assert function_change.details["before"] == expected_functions
    assert function_change.details["after"] == expected_functions
    assert function_change.details["unqualified_runtime_function_material_changed"] is True
    assert function_finding.details == function_change.details

    rendered_ledger_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(function_change.details),
        json.dumps(function_finding.to_dict()),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATEUDF",
        "UPDATEDUDF",
        "RISKUDF",
        "PRIVATE-RUNTIME-FUNCTION-QUERY-BASELINE",
        "PRIVATE-RUNTIME-FUNCTION-SECOND-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_unqualified_runtime_function_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_unqualified_runtime_function_model(tmp_path / "baseline.xlsx")
    candidate = make_unqualified_runtime_function_model(tmp_path / "candidate.xlsx")
    change_unqualified_runtime_function_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.unqualified_runtime_functions
        == candidate_snapshot.unqualified_runtime_functions
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    function_change = next(
        change
        for change in report.changes
        if change.kind == "unqualified_runtime_functions_changed"
    )
    assert function_change.details["unqualified_runtime_function_static_input_changed"] is True
    assert function_change.details["unqualified_runtime_function_static_input_change_count"] == 1
    assert "unqualified_runtime_function_material_changed" not in function_change.details
    assert "FF075" in {finding.rule_id for finding in report.findings}
    assert report.unqualified_runtime_function_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )

    rendered_artifacts = (
        json.dumps(function_change.details),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-RUNTIME-FUNCTION-INPUT-BASELINE",
        "PRIVATE-RUNTIME-FUNCTION-INPUT-CANDIDATE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_unqualified_runtime_function_report_redaction_hides_calls_and_static_inputs(
    tmp_path,
) -> None:
    baseline = make_unqualified_runtime_function_model(tmp_path / "baseline.xlsx")
    candidate = make_unqualified_runtime_function_model(tmp_path / "candidate.xlsx")
    change_unqualified_runtime_function_input(candidate)
    change_unqualified_runtime_function_call(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    sensitive_values = (
        "PRIVATEUDF",
        "UPDATEDUDF",
        "PRIVATE-RUNTIME-FUNCTION-QUERY-BASELINE",
        "PRIVATE-RUNTIME-FUNCTION-INPUT-BASELINE",
        "PRIVATE-RUNTIME-FUNCTION-INPUT-CANDIDATE",
    )
    raw_rendered = json.dumps(raw_payload)
    assert all(value in raw_rendered for value in sensitive_values)

    redacted_payload = redact_unqualified_runtime_function_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert all(value not in redacted_rendered for value in sensitive_values)
    assert UNQUALIFIED_RUNTIME_FUNCTION_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == raw_payload["summary"]

    redacted_markdown = report_to_markdown(
        report, redact_unqualified_runtime_functions=True
    )
    redacted_sarif = report_to_sarif(
        report, redact_unqualified_runtime_functions=True
    )
    assert (
        "Unqualified runtime-function material:** redacted for sharing"
        in redacted_markdown
    )
    assert "FF075" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, json.dumps(redacted_sarif)):
        assert all(value not in artifact for value in sensitive_values)

    native_formula = "=SUM(A1:A2)"
    assert redact_unqualified_runtime_function_material(native_formula) == native_formula


def test_unqualified_runtime_function_report_redaction_hides_unsampled_static_inputs(
    tmp_path,
) -> None:
    """Runtime-function input redaction must use full private impact evidence."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-RUNTIME-FUNCTION-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-RUNTIME-FUNCTION-CANDIDATE"

    def add_runtime_function_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        # Z1 is a bare-runtime-function consumer after the report's B1:U1 sample.
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = "=PRIVATEUDF(A9)"

    rewrite(
        baseline,
        lambda workbook: add_runtime_function_fanout(workbook, baseline_marker),
    )
    rewrite(
        candidate,
        lambda workbook: add_runtime_function_fanout(workbook, candidate_marker),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.unqualified_runtime_function_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )

    redacted = redact_unqualified_runtime_function_report_payload(report, raw_payload)
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert (
        redacted_change["before"]["value"]
        == UNQUALIFIED_RUNTIME_FUNCTION_REDACTION
    )
    assert (
        redacted_change["after"]["value"]
        == UNQUALIFIED_RUNTIME_FUNCTION_REDACTION
    )
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_unqualified_runtime_function_report_redaction_hides_indirect_named_chain(
    tmp_path,
) -> None:
    baseline = make_indirect_named_unqualified_runtime_function_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_indirect_named_unqualified_runtime_function_model(
        tmp_path / "candidate.xlsx"
    )
    change_indirect_named_unqualified_runtime_function_definition(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    function_change = next(
        change
        for change in report.changes
        if change.kind == "unqualified_runtime_functions_changed"
    )
    assert function_change.details["unqualified_runtime_function_definition_changed"] is True
    assert "unqualified_runtime_function_material_changed" not in function_change.details

    baseline_marker = "PRIVATE-INDIRECT-NAMED-RUNTIME-FUNCTION-BASELINE"
    candidate_marker = "PRIVATE-INDIRECT-NAMED-RUNTIME-FUNCTION-CANDIDATE"
    ordinary_looking_definition = (
        f'=LAMBDA(value,FENCE.RUNTIMEWRAPPER("{baseline_marker}"))'
    )
    # Without private name-chain resolution this direct text has only a dotted
    # workbook-defined call, not a bare runtime candidate.
    assert redact_unqualified_runtime_function_material(ordinary_looking_definition) == (
        ordinary_looking_definition
    )

    raw_payload = report.to_dict()
    raw_rendered = json.dumps(raw_payload)
    assert baseline_marker in raw_rendered
    assert candidate_marker in raw_rendered

    redacted_payload = redact_unqualified_runtime_function_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert baseline_marker not in redacted_rendered
    assert candidate_marker not in redacted_rendered
    assert UNQUALIFIED_RUNTIME_FUNCTION_REDACTION in redacted_rendered

    redacted_sarif = report_to_sarif(
        report, redact_unqualified_runtime_functions=True
    )
    redacted_sarif_rendered = json.dumps(redacted_sarif)
    assert baseline_marker not in redacted_sarif_rendered
    assert candidate_marker not in redacted_sarif_rendered
    assert "FF008" in redacted_sarif_rendered
    assert "FF075" in redacted_sarif_rendered


def test_named_unqualified_runtime_function_candidates_are_propagated_and_private(
    tmp_path,
) -> None:
    baseline = make_named_unqualified_runtime_function_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_named_unqualified_runtime_function_model(
        tmp_path / "candidate.xlsx"
    )
    change_named_unqualified_runtime_function_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_functions = {
        "present": True,
        "unqualified_runtime_function_formula_cell_count": 3,
        "unqualified_runtime_function_call_count": 5,
        "unqualified_runtime_function_defined_name_count": 3,
    }
    assert baseline_snapshot.unqualified_runtime_functions.to_dict() == expected_functions
    assert baseline_snapshot.unqualified_runtime_functions.call_cells == frozenset(
        {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")}
    )
    assert profile["unqualified_runtime_functions"] == expected_functions

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    function_change = next(
        change
        for change in report.changes
        if change.kind == "unqualified_runtime_functions_changed"
    )
    function_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF075"
    )
    assert function_change.details["before"] == expected_functions
    assert function_change.details["after"] == expected_functions
    assert function_change.details["unqualified_runtime_function_material_changed"] is True
    assert function_change.details["unqualified_runtime_function_definition_changed"] is True
    assert function_finding.details == function_change.details

    ff075_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF075"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["unqualified_runtime_functions"]),
        profile_to_markdown(profile),
        json.dumps(function_change.details),
        json.dumps(function_finding.to_dict()),
        json.dumps(ff075_sarif_result),
    )
    for sensitive_value in (
        "PRIVATEUDF",
        "UPDATEDUDF",
        "RISKUDF",
        "PRIVATE-NAMED-RUNTIME-FUNCTION-INPUT-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_named_unqualified_runtime_function_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_named_unqualified_runtime_function_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_named_unqualified_runtime_function_model(
        tmp_path / "candidate.xlsx"
    )
    change_named_unqualified_runtime_function_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.unqualified_runtime_functions
        == candidate_snapshot.unqualified_runtime_functions
    )
    assert {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    function_change = next(
        change
        for change in report.changes
        if change.kind == "unqualified_runtime_functions_changed"
    )
    assert function_change.details["unqualified_runtime_function_static_input_changed"] is True
    assert function_change.details["unqualified_runtime_function_static_input_change_count"] == 1
    assert "unqualified_runtime_function_material_changed" not in function_change.details
    assert "unqualified_runtime_function_definition_changed" not in function_change.details
    assert "FF075" in {finding.rule_id for finding in report.findings}


def test_uninvoked_runtime_function_definitions_are_inventory_and_private(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def add_uninvoked_runtime_function(workbook) -> None:
        workbook.defined_names.add(
            DefinedName("FENCE.DORMANT", attr_text="=PRIVATEUDF(Inputs!$A$1)")
        )

    rewrite(baseline, add_uninvoked_runtime_function)
    rewrite(candidate, add_uninvoked_runtime_function)
    candidate_workbook = load_workbook(candidate)
    candidate_workbook.defined_names["FENCE.DORMANT"].attr_text = (
        "=UPDATEDUDF(Inputs!$A$1)"
    )
    candidate_workbook.save(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    expected_functions = {
        "present": True,
        "unqualified_runtime_function_formula_cell_count": 0,
        "unqualified_runtime_function_call_count": 0,
        "unqualified_runtime_function_defined_name_count": 1,
    }
    assert baseline_snapshot.unqualified_runtime_functions.to_dict() == expected_functions
    assert baseline_snapshot.unqualified_runtime_functions.call_cells == frozenset()

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    function_change = next(
        change
        for change in report.changes
        if change.kind == "unqualified_runtime_functions_changed"
    )
    assert function_change.details["before"] == expected_functions
    assert function_change.details["after"] == expected_functions
    assert function_change.details["unqualified_runtime_function_definition_changed"] is True
    assert "unqualified_runtime_function_material_changed" not in function_change.details

    ff075_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF075"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile_snapshot(baseline_snapshot)),
        json.dumps(function_change.details),
        json.dumps(ff075_sarif_result),
    )
    for sensitive_value in ("PRIVATEUDF", "UPDATEDUDF"):
        assert all(
            sensitive_value not in artifact for artifact in rendered_ledger_artifacts
        )


def test_worksheet_code_resource_registrations_are_profiled_diffed_and_private(
    tmp_path,
) -> None:
    baseline = make_worksheet_code_resource_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_worksheet_code_resource_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_worksheet_code_resource_registration_call(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_registrations = {
        "present": True,
        "registration_formula_cell_count": 3,
        "register_id_function_count": 3,
        "registration_defined_name_count": 0,
    }
    assert (
        baseline_snapshot.worksheet_code_resource_registrations.to_dict()
        == expected_registrations
    )
    assert baseline_snapshot.worksheet_code_resource_registrations.registration_cells == (
        frozenset({("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")})
    )
    assert baseline_snapshot.office_custom_functions.present is False
    assert profile["worksheet_code_resource_registrations"] == expected_registrations

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    registration_change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_code_resource_registrations_changed"
    )
    registration_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF067"
    )
    assert registration_change.details["before"] == expected_registrations
    assert registration_change.details["after"] == expected_registrations
    assert (
        registration_change.details[
            "worksheet_code_resource_registration_formula_material_changed"
        ]
        is True
    )
    assert registration_finding.details == registration_change.details

    ff067_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF067"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["worksheet_code_resource_registrations"]),
        profile_to_markdown(profile),
        json.dumps(registration_change.details),
        json.dumps(registration_finding.to_dict()),
        json.dumps(ff067_sarif_result),
    )
    for sensitive_value in (
        "PRIVATE-REGISTRATION-MODULE-BASELINE",
        "PRIVATE-REGISTRATION-PROCEDURE-BASELINE",
        "PRIVATE-REGISTRATION-MODULE-LITERAL-BASELINE",
        "PRIVATE-REGISTRATION-PROCEDURE-LITERAL-BASELINE",
        "PRIVATE-REGISTRATION-MODULE-LITERAL-CANDIDATE",
        "PRIVATE-REGISTRATION-PROCEDURE-LITERAL-CANDIDATE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_worksheet_code_resource_registration_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_worksheet_code_resource_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_worksheet_code_resource_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_worksheet_code_resource_registration_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.worksheet_code_resource_registrations
        == candidate_snapshot.worksheet_code_resource_registrations
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    registration_change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_code_resource_registrations_changed"
    )
    assert (
        registration_change.details[
            "worksheet_code_resource_registration_static_input_changed"
        ]
        is True
    )
    assert (
        registration_change.details[
            "worksheet_code_resource_registration_static_input_change_count"
        ]
        == 1
    )
    assert (
        "worksheet_code_resource_registration_formula_material_changed"
        not in registration_change.details
    )
    assert "FF067" in {finding.rule_id for finding in report.findings}
    assert report.worksheet_code_resource_registration_static_input_cells == (
        frozenset({("Inputs", "A9")})
    )


def test_worksheet_code_resource_registration_report_redaction_hides_calls_and_inputs(
    tmp_path,
) -> None:
    baseline = make_worksheet_code_resource_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_worksheet_code_resource_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_worksheet_code_resource_registration_input(candidate)
    change_worksheet_code_resource_registration_call(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    sensitive_values = (
        "PRIVATE-REGISTRATION-MODULE-BASELINE",
        "PRIVATE-REGISTRATION-MODULE-CANDIDATE",
        "PRIVATE-REGISTRATION-MODULE-LITERAL-BASELINE",
        "PRIVATE-REGISTRATION-PROCEDURE-LITERAL-BASELINE",
        "PRIVATE-REGISTRATION-MODULE-LITERAL-CANDIDATE",
        "PRIVATE-REGISTRATION-PROCEDURE-LITERAL-CANDIDATE",
    )
    raw_rendered = json.dumps(raw_payload)
    assert all(value in raw_rendered for value in sensitive_values)

    redacted_payload = redact_worksheet_code_resource_registration_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert all(value not in redacted_rendered for value in sensitive_values)
    assert WORKSHEET_CODE_RESOURCE_REGISTRATION_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == raw_payload["summary"]

    redacted_markdown = report_to_markdown(
        report, redact_worksheet_code_resource_registrations=True
    )
    redacted_sarif = report_to_sarif(
        report, redact_worksheet_code_resource_registrations=True
    )
    assert (
        "Worksheet code-resource registration material:** redacted for sharing"
        in redacted_markdown
    )
    assert "FF067" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, json.dumps(redacted_sarif)):
        assert all(value not in artifact for value in sensitive_values)

    native_formula = "=SUM(A1:A2)"
    assert (
        redact_worksheet_code_resource_registration_material(native_formula)
        == native_formula
    )


def test_worksheet_code_resource_registration_report_redaction_hides_unsampled_inputs(
    tmp_path,
) -> None:
    """Registration input redaction must use full private impact evidence."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-REGISTRATION-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-REGISTRATION-CANDIDATE"

    def add_registration_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        # Z1 is a registration consumer after the report's B1:U1 sample.
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = '=REGISTER.ID(A9,"PRIVATE-UNSAMPLED-PROCEDURE","J!")'

    rewrite(
        baseline,
        lambda workbook: add_registration_fanout(workbook, baseline_marker),
    )
    rewrite(
        candidate,
        lambda workbook: add_registration_fanout(workbook, candidate_marker),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.worksheet_code_resource_registration_static_input_cells == (
        frozenset({("Inputs", "A9")})
    )

    redacted = redact_worksheet_code_resource_registration_report_payload(
        report, raw_payload
    )
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert (
        redacted_change["before"]["value"]
        == WORKSHEET_CODE_RESOURCE_REGISTRATION_REDACTION
    )
    assert (
        redacted_change["after"]["value"]
        == WORKSHEET_CODE_RESOURCE_REGISTRATION_REDACTION
    )
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_worksheet_code_resource_registration_report_redaction_hides_indirect_names(
    tmp_path,
) -> None:
    baseline = make_indirect_named_worksheet_code_resource_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_indirect_named_worksheet_code_resource_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_indirect_named_worksheet_code_resource_registration_definition(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    registration_change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_code_resource_registrations_changed"
    )
    assert (
        registration_change.details[
            "worksheet_code_resource_registration_definition_material_changed"
        ]
        is True
    )
    assert (
        "worksheet_code_resource_registration_formula_material_changed"
        not in registration_change.details
    )

    baseline_marker = "PRIVATE-INDIRECT-NAMED-REGISTRATION-MODULE-BASELINE"
    candidate_marker = "PRIVATE-INDIRECT-NAMED-REGISTRATION-MODULE-CANDIDATE"
    ordinary_looking_definition = (
        f'=LAMBDA(module,FENCE.REGISTRATIONWRAPPER("{baseline_marker}"))'
    )
    # Without private name-chain resolution this text does not itself call
    # REGISTER.ID, even though its dotted wrapper eventually does.
    assert (
        redact_worksheet_code_resource_registration_material(
            ordinary_looking_definition
        )
        == ordinary_looking_definition
    )

    raw_payload = report.to_dict()
    raw_rendered = json.dumps(raw_payload)
    assert baseline_marker in raw_rendered
    assert candidate_marker in raw_rendered

    redacted_payload = redact_worksheet_code_resource_registration_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert baseline_marker not in redacted_rendered
    assert candidate_marker not in redacted_rendered
    assert WORKSHEET_CODE_RESOURCE_REGISTRATION_REDACTION in redacted_rendered

    redacted_sarif = report_to_sarif(
        report, redact_worksheet_code_resource_registrations=True
    )
    redacted_sarif_rendered = json.dumps(redacted_sarif)
    assert baseline_marker not in redacted_sarif_rendered
    assert candidate_marker not in redacted_sarif_rendered
    assert "FF008" in redacted_sarif_rendered
    assert "FF067" in redacted_sarif_rendered


def test_uninvoked_formula_defined_code_resource_registration_is_profiled(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "stored-name.xlsx")

    def add_stored_registration(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.STORED.REGISTRATION",
                attr_text='=REGISTER.ID("PRIVATE-STORED-MODULE","PRIVATE-STORED-PROCEDURE","J!")',
            )
        )

    rewrite(workbook_path, add_stored_registration)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.worksheet_code_resource_registrations.to_dict() == {
        "present": True,
        "registration_formula_cell_count": 0,
        "register_id_function_count": 0,
        "registration_defined_name_count": 1,
    }
    assert snapshot.worksheet_code_resource_registrations.registration_cells == frozenset()


def test_named_worksheet_code_resource_registrations_are_propagated_and_private(
    tmp_path,
) -> None:
    baseline = make_named_worksheet_code_resource_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_named_worksheet_code_resource_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_named_worksheet_code_resource_registration_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_registrations = {
        "present": True,
        "registration_formula_cell_count": 3,
        "register_id_function_count": 3,
        "registration_defined_name_count": 3,
    }
    assert (
        baseline_snapshot.worksheet_code_resource_registrations.to_dict()
        == expected_registrations
    )
    assert baseline_snapshot.worksheet_code_resource_registrations.registration_cells == (
        frozenset({("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")})
    )
    assert profile["worksheet_code_resource_registrations"] == expected_registrations

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    registration_change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_code_resource_registrations_changed"
    )
    assert registration_change.details["before"] == expected_registrations
    assert registration_change.details["after"] == expected_registrations
    assert (
        registration_change.details[
            "worksheet_code_resource_registration_definition_material_changed"
        ]
        is True
    )

    ff067_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF067"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["worksheet_code_resource_registrations"]),
        profile_to_markdown(profile),
        json.dumps(registration_change.details),
        json.dumps(ff067_sarif_result),
    )
    for sensitive_value in (
        "FENCE",
        "PRIVATE-NAMED-REGISTRATION-MODULE-BASELINE",
        "PRIVATE-NAMED-REGISTRATION-PROCEDURE-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_named_worksheet_code_resource_registration_static_inputs_are_guarded(
    tmp_path,
) -> None:
    baseline = make_named_worksheet_code_resource_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_named_worksheet_code_resource_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_named_worksheet_code_resource_registration_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.worksheet_code_resource_registrations
        == candidate_snapshot.worksheet_code_resource_registrations
    )
    assert {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    registration_change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_code_resource_registrations_changed"
    )
    assert (
        registration_change.details[
            "worksheet_code_resource_registration_static_input_changed"
        ]
        is True
    )
    assert (
        registration_change.details[
            "worksheet_code_resource_registration_static_input_change_count"
        ]
        == 1
    )
    assert "FF067" in {finding.rule_id for finding in report.findings}


def test_recursive_named_worksheet_code_resource_registrations_are_cycle_safe(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "recursive.xlsx")

    def add_recursive_registration(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.LOOP",
                attr_text=(
                    '=LAMBDA(module,procedure,REGISTER.ID(module,procedure,"J!")'
                    "+FENCE.LOOP(module,procedure))"
                ),
            )
        )
        workbook["Model"]["D2"] = "=FENCE.LOOP(Inputs!B2,Inputs!B3)"

    rewrite(workbook_path, add_recursive_registration)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.worksheet_code_resource_registrations.to_dict() == {
        "present": True,
        "registration_formula_cell_count": 1,
        "register_id_function_count": 1,
        "registration_defined_name_count": 1,
    }
    assert snapshot.worksheet_code_resource_registrations.registration_cells == frozenset(
        {("Model", "D2")}
    )
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == ("FENCE.LOOP",)


def test_scoped_named_worksheet_code_resource_registrations_follow_local_precedence(
    tmp_path,
) -> None:
    workbook_path = make_scoped_named_lambda_model(tmp_path / "scoped.xlsx")
    workbook = load_workbook(workbook_path)
    model = workbook["Model"]
    report = workbook["Report"]
    model["D2"] = "=FENCE.REGISTER(A2,A3)"
    report["D2"] = "=Model!FENCE.REGISTER(A2,A3)"
    model.defined_names.add(
        DefinedName(
            "FENCE.REGISTER",
            attr_text='=LAMBDA(module,procedure,REGISTER.ID(module,procedure,"J!"))',
            localSheetId=1,
        )
    )
    workbook.save(workbook_path)

    snapshot = load_snapshot(workbook_path)

    assert snapshot.worksheet_code_resource_registrations.to_dict() == {
        "present": True,
        "registration_formula_cell_count": 2,
        "register_id_function_count": 2,
        "registration_defined_name_count": 1,
    }
    assert snapshot.worksheet_code_resource_registrations.registration_cells == frozenset(
        {("Model", "D2"), ("Report", "D2")}
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_formula_defined_xlm_registrations_are_propagated_diffed_and_private(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_registration_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_registrations = {
        "present": True,
        "registration_formula_cell_count": 3,
        "register_function_count": 3,
        "registration_defined_name_count": 3,
    }
    assert baseline_snapshot.formula_defined_xlm_registrations.to_dict() == (
        expected_registrations
    )
    assert baseline_snapshot.formula_defined_xlm_registrations.registration_cells == (
        frozenset({("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")})
    )
    assert baseline_snapshot.worksheet_code_resource_registrations.present is False
    assert baseline_snapshot.xlm_macro_sheets.present is False
    assert profile["formula_defined_xlm_registrations"] == expected_registrations
    assert "## Formula-defined XLM registrations" in profile_to_markdown(profile)

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    registration_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_registrations_changed"
    )
    registration_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF068"
    )
    assert registration_change.details["before"] == expected_registrations
    assert registration_change.details["after"] == expected_registrations
    assert (
        registration_change.details[
            "formula_defined_xlm_registration_definition_material_changed"
        ]
        is True
    )
    assert registration_finding.details == registration_change.details

    ff068_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF068"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["formula_defined_xlm_registrations"]),
        profile_to_markdown(profile),
        json.dumps(registration_change.details),
        json.dumps(registration_finding.to_dict()),
        json.dumps(ff068_sarif_result),
    )
    for sensitive_value in (
        "FENCE.XLM",
        "PRIVATE-XLM-REGISTRATION-MODULE-BASELINE",
        "PRIVATE-XLM-REGISTRATION-PROCEDURE-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_formula_defined_xlm_registration_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_formula_defined_xlm_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_registration_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.formula_defined_xlm_registrations
        == candidate_snapshot.formula_defined_xlm_registrations
    )
    assert {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    registration_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_registrations_changed"
    )
    assert (
        registration_change.details[
            "formula_defined_xlm_registration_static_input_changed"
        ]
        is True
    )
    assert (
        registration_change.details[
            "formula_defined_xlm_registration_static_input_change_count"
        ]
        == 1
    )
    assert "FF068" in {finding.rule_id for finding in report.findings}
    assert report.formula_defined_xlm_registration_static_input_cells == (
        frozenset({("Inputs", "A9")})
    )


def test_formula_defined_xlm_registration_report_redaction_hides_calls_and_inputs(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_registration_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_registration_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_registration_input(candidate)
    change_formula_defined_xlm_registration_call(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    sensitive_values = (
        "PRIVATE-XLM-REGISTRATION-MODULE-BASELINE",
        "PRIVATE-XLM-REGISTRATION-MODULE-CANDIDATE",
        "PRIVATE-XLM-REGISTRATION-LITERAL-PROCEDURE-BASELINE",
        "PRIVATE-XLM-REGISTRATION-LITERAL-PROCEDURE-CANDIDATE",
        "PRIVATE-XLM-REGISTRATION-LITERAL-TYPE-BASELINE",
        "PRIVATE-XLM-REGISTRATION-LITERAL-TYPE-CANDIDATE",
    )
    raw_rendered = json.dumps(raw_payload)
    assert all(value in raw_rendered for value in sensitive_values)

    redacted_payload = redact_formula_defined_xlm_registration_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert all(value not in redacted_rendered for value in sensitive_values)
    assert FORMULA_DEFINED_XLM_REGISTRATION_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == raw_payload["summary"]

    redacted_markdown = report_to_markdown(
        report, redact_formula_defined_xlm_registrations=True
    )
    redacted_sarif = report_to_sarif(
        report, redact_formula_defined_xlm_registrations=True
    )
    assert (
        "Formula-defined XLM registration material:** redacted for sharing"
        in redacted_markdown
    )
    assert "FF068" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, json.dumps(redacted_sarif)):
        assert all(value not in artifact for value in sensitive_values)

    native_formula = "=SUM(A1:A2)"
    assert redact_formula_defined_xlm_registration_material(native_formula) == native_formula


def test_formula_defined_xlm_registration_report_redaction_hides_unsampled_inputs(
    tmp_path,
) -> None:
    """XLM registration input redaction must use full private impact evidence."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-XLM-REGISTRATION-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-XLM-REGISTRATION-CANDIDATE"

    def add_registration_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = "=FENCE.XLM.UNSAMPLED(A9)"
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.UNSAMPLED",
                attr_text='=LAMBDA(module,REGISTER(module,"PRIVATE-UNSAMPLED-XLM-PROCEDURE","J!"))',
            )
        )

    rewrite(
        baseline,
        lambda workbook: add_registration_fanout(workbook, baseline_marker),
    )
    rewrite(
        candidate,
        lambda workbook: add_registration_fanout(workbook, candidate_marker),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.formula_defined_xlm_registration_static_input_cells == (
        frozenset({("Inputs", "A9")})
    )

    redacted = redact_formula_defined_xlm_registration_report_payload(
        report, raw_payload
    )
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert (
        redacted_change["before"]["value"]
        == FORMULA_DEFINED_XLM_REGISTRATION_REDACTION
    )
    assert (
        redacted_change["after"]["value"]
        == FORMULA_DEFINED_XLM_REGISTRATION_REDACTION
    )
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_formula_defined_xlm_registration_report_redaction_hides_indirect_names(
    tmp_path,
) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def add_registration_chain(workbook, marker: str) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.REGISTRATIONWRAPPER",
                attr_text='=LAMBDA(module,REGISTER(module,"J!"))',
            )
        )
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.REGISTRATIONCHAIN",
                attr_text=f'=LAMBDA(module,FENCE.XLM.REGISTRATIONWRAPPER("{marker}"))',
            )
        )

    baseline_marker = "PRIVATE-INDIRECT-XLM-REGISTRATION-BASELINE"
    candidate_marker = "PRIVATE-INDIRECT-XLM-REGISTRATION-CANDIDATE"
    rewrite(
        baseline,
        lambda workbook: add_registration_chain(workbook, baseline_marker),
    )
    rewrite(
        candidate,
        lambda workbook: add_registration_chain(workbook, candidate_marker),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    registration_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_registrations_changed"
    )
    assert (
        registration_change.details[
            "formula_defined_xlm_registration_definition_material_changed"
        ]
        is True
    )
    ordinary_looking_definition = (
        f'=LAMBDA(module,FENCE.XLM.REGISTRATIONWRAPPER("{baseline_marker}"))'
    )
    assert (
        redact_formula_defined_xlm_registration_material(ordinary_looking_definition)
        == ordinary_looking_definition
    )

    raw_payload = report.to_dict()
    raw_rendered = json.dumps(raw_payload)
    assert baseline_marker in raw_rendered
    assert candidate_marker in raw_rendered

    redacted_payload = redact_formula_defined_xlm_registration_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert baseline_marker not in redacted_rendered
    assert candidate_marker not in redacted_rendered
    assert FORMULA_DEFINED_XLM_REGISTRATION_REDACTION in redacted_rendered

    redacted_sarif = report_to_sarif(
        report, redact_formula_defined_xlm_registrations=True
    )
    redacted_sarif_rendered = json.dumps(redacted_sarif)
    assert baseline_marker not in redacted_sarif_rendered
    assert candidate_marker not in redacted_sarif_rendered
    assert "FF008" in redacted_sarif_rendered
    assert "FF068" in redacted_sarif_rendered


def test_uninvoked_formula_defined_xlm_registration_is_profiled(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "stored-name.xlsx")

    def add_stored_registration(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.STORED.REGISTRATION",
                attr_text=(
                    '=REGISTER("PRIVATE-STORED-XLM-MODULE",'
                    '"PRIVATE-STORED-XLM-PROCEDURE","J!")'
                ),
            )
        )

    rewrite(workbook_path, add_stored_registration)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_registrations.to_dict() == {
        "present": True,
        "registration_formula_cell_count": 0,
        "register_function_count": 0,
        "registration_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_registrations.registration_cells == frozenset()


def test_recursive_named_formula_defined_xlm_registrations_are_cycle_safe(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "recursive.xlsx")

    def add_recursive_registration(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.LOOP",
                attr_text=(
                    '=LAMBDA(module,procedure,REGISTER(module,procedure,"J!")'
                    "+FENCE.XLM.LOOP(module,procedure))"
                ),
            )
        )
        workbook["Model"]["D2"] = "=FENCE.XLM.LOOP(Inputs!B2,Inputs!B3)"

    rewrite(workbook_path, add_recursive_registration)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_registrations.to_dict() == {
        "present": True,
        "registration_formula_cell_count": 1,
        "register_function_count": 1,
        "registration_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_registrations.registration_cells == frozenset(
        {("Model", "D2")}
    )
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == ("FENCE.XLM.LOOP",)


def test_scoped_formula_defined_xlm_registrations_follow_local_precedence(
    tmp_path,
) -> None:
    workbook_path = make_scoped_named_lambda_model(tmp_path / "scoped.xlsx")
    workbook = load_workbook(workbook_path)
    model = workbook["Model"]
    report = workbook["Report"]
    model["D2"] = "=FENCE.XLM.REGISTER(A2,A3)"
    report["D2"] = "=Model!FENCE.XLM.REGISTER(A2,A3)"
    model.defined_names.add(
        DefinedName(
            "FENCE.XLM.REGISTER",
            attr_text='=LAMBDA(module,procedure,REGISTER(module,procedure,"J!"))',
            localSheetId=1,
        )
    )
    workbook.save(workbook_path)

    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_registrations.to_dict() == {
        "present": True,
        "registration_formula_cell_count": 2,
        "register_function_count": 2,
        "registration_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_registrations.registration_cells == frozenset(
        {("Model", "D2"), ("Report", "D2")}
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_direct_worksheet_register_is_outside_formula_defined_xlm_boundary(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "direct-register.xlsx")

    def add_direct_register(workbook) -> None:
        workbook["Model"]["D2"] = '=REGISTER("PRIVATE-DIRECT-MODULE",A2,"J!")'

    rewrite(workbook_path, add_direct_register)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_registrations.to_dict() == {
        "present": False,
        "registration_formula_cell_count": 0,
        "register_function_count": 0,
        "registration_defined_name_count": 0,
    }


def test_formula_defined_xlm_evaluations_are_propagated_diffed_and_private(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_evaluation_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_evaluation_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_evaluation_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_evaluations = {
        "present": True,
        "evaluation_formula_cell_count": 3,
        "evaluate_function_count": 3,
        "evaluation_defined_name_count": 3,
    }
    assert baseline_snapshot.formula_defined_xlm_evaluations.to_dict() == (
        expected_evaluations
    )
    assert baseline_snapshot.formula_defined_xlm_evaluations.evaluation_cells == (
        frozenset({("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")})
    )
    assert baseline_snapshot.formula_defined_xlm_registrations.present is False
    assert baseline_snapshot.xlm_macro_sheets.present is False
    assert profile["formula_defined_xlm_evaluations"] == expected_evaluations
    assert "## Formula-defined XLM expression evaluation" in profile_to_markdown(
        profile
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    evaluation_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_evaluations_changed"
    )
    evaluation_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF069"
    )
    assert evaluation_change.details["before"] == expected_evaluations
    assert evaluation_change.details["after"] == expected_evaluations
    assert (
        evaluation_change.details[
            "formula_defined_xlm_evaluation_definition_material_changed"
        ]
        is True
    )
    assert evaluation_finding.details == evaluation_change.details

    ff069_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF069"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["formula_defined_xlm_evaluations"]),
        profile_to_markdown(profile),
        json.dumps(evaluation_change.details),
        json.dumps(evaluation_finding.to_dict()),
        json.dumps(ff069_sarif_result),
    )
    for sensitive_value in (
        "FENCE.XLM.EVALUATE",
        "PRIVATE-XLM-EVALUATE-EXPRESSION-BASELINE",
        "PRIVATE-XLM-EVALUATE-DEFINITION-CANDIDATE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_formula_defined_xlm_evaluation_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_formula_defined_xlm_evaluation_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_evaluation_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_evaluation_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.formula_defined_xlm_evaluations
        == candidate_snapshot.formula_defined_xlm_evaluations
    )
    assert {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    evaluation_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_evaluations_changed"
    )
    assert (
        evaluation_change.details[
            "formula_defined_xlm_evaluation_static_input_changed"
        ]
        is True
    )
    assert (
        evaluation_change.details[
            "formula_defined_xlm_evaluation_static_input_change_count"
        ]
        == 1
    )
    assert "FF069" in {finding.rule_id for finding in report.findings}
    assert report.formula_defined_xlm_evaluation_static_input_cells == (
        frozenset({("Inputs", "A9")})
    )


def test_formula_defined_xlm_evaluation_report_redaction_hides_calls_and_inputs(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_evaluation_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_evaluation_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_evaluation_input(candidate)
    change_formula_defined_xlm_evaluation_call(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    sensitive_values = (
        "PRIVATE-XLM-EVALUATE-EXPRESSION-BASELINE",
        "PRIVATE-XLM-EVALUATE-EXPRESSION-CANDIDATE",
        "PRIVATE-XLM-EVALUATE-LITERAL-BASELINE",
        "PRIVATE-XLM-EVALUATE-LITERAL-CANDIDATE",
    )
    raw_rendered = json.dumps(raw_payload)
    assert all(value in raw_rendered for value in sensitive_values)

    redacted_payload = redact_formula_defined_xlm_evaluation_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert all(value not in redacted_rendered for value in sensitive_values)
    assert FORMULA_DEFINED_XLM_EVALUATION_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == raw_payload["summary"]

    redacted_markdown = report_to_markdown(
        report, redact_formula_defined_xlm_evaluations=True
    )
    redacted_sarif = report_to_sarif(
        report, redact_formula_defined_xlm_evaluations=True
    )
    assert (
        "Formula-defined XLM evaluation material:** redacted for sharing"
        in redacted_markdown
    )
    assert "FF069" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, json.dumps(redacted_sarif)):
        assert all(value not in artifact for value in sensitive_values)

    native_formula = "=SUM(A1:A2)"
    assert redact_formula_defined_xlm_evaluation_material(native_formula) == native_formula


def test_formula_defined_xlm_evaluation_report_redaction_hides_unsampled_inputs(
    tmp_path,
) -> None:
    """XLM evaluation input redaction must use full private impact evidence."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-XLM-EVALUATION-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-XLM-EVALUATION-CANDIDATE"

    def add_evaluation_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = "=FENCE.XLM.UNSAMPLED.EVALUATE(A9)"
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.UNSAMPLED.EVALUATE",
                attr_text="=LAMBDA(expression,EVALUATE(expression))",
            )
        )

    rewrite(
        baseline,
        lambda workbook: add_evaluation_fanout(workbook, baseline_marker),
    )
    rewrite(
        candidate,
        lambda workbook: add_evaluation_fanout(workbook, candidate_marker),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.formula_defined_xlm_evaluation_static_input_cells == (
        frozenset({("Inputs", "A9")})
    )

    redacted = redact_formula_defined_xlm_evaluation_report_payload(
        report, raw_payload
    )
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert (
        redacted_change["before"]["value"]
        == FORMULA_DEFINED_XLM_EVALUATION_REDACTION
    )
    assert (
        redacted_change["after"]["value"]
        == FORMULA_DEFINED_XLM_EVALUATION_REDACTION
    )
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_formula_defined_xlm_evaluation_report_redaction_hides_indirect_names(
    tmp_path,
) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def add_evaluation_chain(workbook, marker: str) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.EVALUATIONWRAPPER",
                attr_text="=LAMBDA(expression,EVALUATE(expression))",
            )
        )
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.EVALUATIONCHAIN",
                attr_text=(
                    f'=LAMBDA(expression,FENCE.XLM.EVALUATIONWRAPPER("{marker}"))'
                ),
            )
        )

    baseline_marker = "PRIVATE-INDIRECT-XLM-EVALUATION-BASELINE"
    candidate_marker = "PRIVATE-INDIRECT-XLM-EVALUATION-CANDIDATE"
    rewrite(
        baseline,
        lambda workbook: add_evaluation_chain(workbook, baseline_marker),
    )
    rewrite(
        candidate,
        lambda workbook: add_evaluation_chain(workbook, candidate_marker),
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    evaluation_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_evaluations_changed"
    )
    assert (
        evaluation_change.details[
            "formula_defined_xlm_evaluation_definition_material_changed"
        ]
        is True
    )
    ordinary_looking_definition = (
        f'=LAMBDA(expression,FENCE.XLM.EVALUATIONWRAPPER("{baseline_marker}"))'
    )
    assert (
        redact_formula_defined_xlm_evaluation_material(ordinary_looking_definition)
        == ordinary_looking_definition
    )

    raw_payload = report.to_dict()
    raw_rendered = json.dumps(raw_payload)
    assert baseline_marker in raw_rendered
    assert candidate_marker in raw_rendered

    redacted_payload = redact_formula_defined_xlm_evaluation_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert baseline_marker not in redacted_rendered
    assert candidate_marker not in redacted_rendered
    assert FORMULA_DEFINED_XLM_EVALUATION_REDACTION in redacted_rendered

    redacted_sarif = report_to_sarif(
        report, redact_formula_defined_xlm_evaluations=True
    )
    redacted_sarif_rendered = json.dumps(redacted_sarif)
    assert baseline_marker not in redacted_sarif_rendered
    assert candidate_marker not in redacted_sarif_rendered
    assert "FF008" in redacted_sarif_rendered
    assert "FF069" in redacted_sarif_rendered


def test_uninvoked_formula_defined_xlm_evaluation_is_profiled(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "stored-name.xlsx")

    def add_stored_evaluation(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.STORED.EVALUATION",
                attr_text='=EVALUATE("PRIVATE-STORED-XLM-EXPRESSION")',
            )
        )

    rewrite(workbook_path, add_stored_evaluation)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_evaluations.to_dict() == {
        "present": True,
        "evaluation_formula_cell_count": 0,
        "evaluate_function_count": 0,
        "evaluation_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_evaluations.evaluation_cells == frozenset()


def test_recursive_named_formula_defined_xlm_evaluations_are_cycle_safe(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "recursive.xlsx")

    def add_recursive_evaluation(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.EVALUATE.LOOP",
                attr_text=(
                    "=LAMBDA(expression,EVALUATE(expression)"
                    "+FENCE.XLM.EVALUATE.LOOP(expression))"
                ),
            )
        )
        workbook["Model"]["D2"] = "=FENCE.XLM.EVALUATE.LOOP(Inputs!B2)"

    rewrite(workbook_path, add_recursive_evaluation)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_evaluations.to_dict() == {
        "present": True,
        "evaluation_formula_cell_count": 1,
        "evaluate_function_count": 1,
        "evaluation_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_evaluations.evaluation_cells == frozenset(
        {("Model", "D2")}
    )
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == (
        "FENCE.XLM.EVALUATE.LOOP",
    )


def test_scoped_formula_defined_xlm_evaluations_follow_local_precedence(
    tmp_path,
) -> None:
    workbook_path = make_scoped_named_lambda_model(tmp_path / "scoped.xlsx")
    workbook = load_workbook(workbook_path)
    model = workbook["Model"]
    report = workbook["Report"]
    model["D2"] = "=FENCE.XLM.EVALUATE(A2)"
    report["D2"] = "=Model!FENCE.XLM.EVALUATE(A2)"
    model.defined_names.add(
        DefinedName(
            "FENCE.XLM.EVALUATE",
            attr_text="=LAMBDA(expression,EVALUATE(expression))",
            localSheetId=1,
        )
    )
    workbook.save(workbook_path)

    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_evaluations.to_dict() == {
        "present": True,
        "evaluation_formula_cell_count": 2,
        "evaluate_function_count": 2,
        "evaluation_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_evaluations.evaluation_cells == frozenset(
        {("Model", "D2"), ("Report", "D2")}
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_direct_worksheet_evaluate_is_outside_formula_defined_xlm_boundary(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "direct-evaluate.xlsx")

    def add_direct_evaluate(workbook) -> None:
        workbook["Model"]["D2"] = '=EVALUATE("PRIVATE-DIRECT-EXPRESSION")'

    rewrite(workbook_path, add_direct_evaluate)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_evaluations.to_dict() == {
        "present": False,
        "evaluation_formula_cell_count": 0,
        "evaluate_function_count": 0,
        "evaluation_defined_name_count": 0,
    }


def test_formula_defined_xlm_evaluation_respects_named_lambda_shadowing(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "shadowed-evaluate.xlsx")

    def add_shadowing_lambda(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "EVALUATE",
                attr_text="=LAMBDA(expression,expression)",
            )
        )
        workbook["Model"]["D2"] = "=EVALUATE(A2)"

    rewrite(workbook_path, add_shadowing_lambda)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_evaluations.to_dict() == {
        "present": False,
        "evaluation_formula_cell_count": 0,
        "evaluate_function_count": 0,
        "evaluation_defined_name_count": 0,
    }
    assert snapshot.unresolved_reference_tokens == {}


def test_formula_defined_xlm_evaluation_text_is_not_retokenized(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def add_runtime_text_evaluation(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.TEXT.EVALUATE",
                attr_text='=EVALUATE("=Inputs!$A$10*2")',
            )
        )
        workbook["Model"]["D2"] = "=FENCE.XLM.TEXT.EVALUATE"

    rewrite(baseline, add_runtime_text_evaluation)
    rewrite(candidate, add_runtime_text_evaluation)
    rewrite(candidate, lambda workbook: setattr(workbook["Inputs"]["A10"], "value", 5))

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.formula_defined_xlm_evaluations.evaluation_cells == (
        frozenset({("Model", "D2")})
    )
    assert ("Model", "D2") not in baseline_snapshot.reverse_dependencies.get(
        ("Inputs", "A10"), set()
    )
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    assert "FF069" not in {finding.rule_id for finding in report.findings}
    assert report.formula_defined_xlm_evaluation_static_input_cells == frozenset()
    raw_payload = report.to_dict()
    redacted_payload = redact_formula_defined_xlm_evaluation_report_payload(
        report, raw_payload
    )
    raw_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A10"
    )
    redacted_change = next(
        change
        for change in redacted_payload["changes"]
        if change["location"] == "Inputs!A10"
    )
    assert redacted_change["before"] == raw_change["before"]
    assert redacted_change["after"] == raw_change["after"]


def test_formula_defined_xlm_actions_are_propagated_diffed_and_private(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_action_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_defined_xlm_action_model(tmp_path / "candidate.xlsx")
    change_formula_defined_xlm_action_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_actions = {
        "present": True,
        "action_formula_cell_count": 3,
        "action_function_count": 5,
        "action_defined_name_count": 3,
    }
    assert baseline_snapshot.formula_defined_xlm_actions.to_dict() == expected_actions
    assert baseline_snapshot.formula_defined_xlm_actions.action_cells == frozenset(
        {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")}
    )
    assert baseline_snapshot.summary()["formula_defined_xlm_action_function_count"] == 5
    assert baseline_snapshot.summary()["has_formula_defined_xlm_actions"] is True
    assert baseline_snapshot.formula_defined_xlm_evaluations.present is False
    assert baseline_snapshot.xlm_macro_sheets.present is False
    assert profile["formula_defined_xlm_actions"] == expected_actions
    assert profile["features"]["has_formula_defined_xlm_actions"] is True
    assert "## Formula-defined XLM actions and event dispatch" in profile_to_markdown(
        profile
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_actions_changed"
    )
    action_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF073"
    )
    assert action_change.details["before"] == expected_actions
    assert action_change.details["after"] == expected_actions
    assert (
        action_change.details[
            "formula_defined_xlm_action_definition_material_changed"
        ]
        is True
    )
    assert action_finding.details == action_change.details

    ff073_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF073"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["formula_defined_xlm_actions"]),
        profile_to_markdown(profile),
        json.dumps(action_change.details),
        json.dumps(action_finding.to_dict()),
        json.dumps(ff073_sarif_result),
    )
    for sensitive_value in (
        "FENCE.XLM.ACTION",
        "PRIVATE-XLM-ACTION-INPUT-BASELINE",
        "PRIVATE-XLM-ACTION-MACRO",
        "PRIVATE-XLM-ACTION-MACRO-CANDIDATE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_formula_defined_xlm_action_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_formula_defined_xlm_action_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_defined_xlm_action_model(tmp_path / "candidate.xlsx")
    change_formula_defined_xlm_action_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.formula_defined_xlm_actions == (
        candidate_snapshot.formula_defined_xlm_actions
    )
    assert {("Inputs", "B2"), ("Inputs", "B4")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_actions_changed"
    )
    assert action_change.details["formula_defined_xlm_action_static_input_changed"] is True
    assert action_change.details["formula_defined_xlm_action_static_input_change_count"] == 1
    assert "FF073" in {finding.rule_id for finding in report.findings}
    assert report.formula_defined_xlm_action_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )


def test_formula_defined_xlm_action_report_redaction_hides_calls_and_inputs(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_action_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_defined_xlm_action_model(tmp_path / "candidate.xlsx")
    change_formula_defined_xlm_action_input(candidate)
    change_formula_defined_xlm_action_call(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    sensitive_values = (
        "PRIVATE-XLM-ACTION-INPUT-BASELINE",
        "PRIVATE-XLM-ACTION-INPUT-CANDIDATE",
        "PRIVATE-XLM-ACTION-EVENT",
        "PRIVATE-XLM-ACTION-EVENT-CANDIDATE",
    )
    raw_rendered = json.dumps(raw_payload)
    assert all(value in raw_rendered for value in sensitive_values)

    redacted_payload = redact_formula_defined_xlm_action_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert all(value not in redacted_rendered for value in sensitive_values)
    assert FORMULA_DEFINED_XLM_ACTION_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == raw_payload["summary"]

    redacted_markdown = report_to_markdown(
        report, redact_formula_defined_xlm_actions=True
    )
    redacted_sarif = report_to_sarif(
        report, redact_formula_defined_xlm_actions=True
    )
    assert "Formula-defined XLM action material:** redacted for sharing" in redacted_markdown
    assert "FF073" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, json.dumps(redacted_sarif)):
        assert all(value not in artifact for value in sensitive_values)

    native_formula = "=SUM(A1:A2)"
    assert redact_formula_defined_xlm_action_material(native_formula) == native_formula


def test_formula_defined_xlm_action_report_redaction_hides_unsampled_inputs(
    tmp_path,
) -> None:
    """XLM action input redaction must use full private impact evidence."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-XLM-ACTION-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-XLM-ACTION-CANDIDATE"

    def add_action_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = "=FENCE.XLM.UNSAMPLED.ACTION(A9)"
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.UNSAMPLED.ACTION",
                attr_text="=LAMBDA(payload,EXEC(payload))",
            )
        )

    rewrite(baseline, lambda workbook: add_action_fanout(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_action_fanout(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.formula_defined_xlm_action_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )

    redacted = redact_formula_defined_xlm_action_report_payload(report, raw_payload)
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert redacted_change["before"]["value"] == FORMULA_DEFINED_XLM_ACTION_REDACTION
    assert redacted_change["after"]["value"] == FORMULA_DEFINED_XLM_ACTION_REDACTION
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_formula_defined_xlm_action_report_redaction_hides_indirect_names(
    tmp_path,
) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def add_action_chain(workbook, marker: str) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.ACTIONWRAPPER",
                attr_text="=LAMBDA(payload,EXEC(payload))",
            )
        )
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.ACTIONCHAIN",
                attr_text=(
                    f'=LAMBDA(payload,FENCE.XLM.ACTIONWRAPPER("{marker}"))'
                ),
            )
        )

    baseline_marker = "PRIVATE-INDIRECT-XLM-ACTION-BASELINE"
    candidate_marker = "PRIVATE-INDIRECT-XLM-ACTION-CANDIDATE"
    rewrite(baseline, lambda workbook: add_action_chain(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_action_chain(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    action_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_actions_changed"
    )
    assert action_change.details["formula_defined_xlm_action_definition_material_changed"] is True
    ordinary_looking_definition = (
        f'=LAMBDA(payload,FENCE.XLM.ACTIONWRAPPER("{baseline_marker}"))'
    )
    assert (
        redact_formula_defined_xlm_action_material(ordinary_looking_definition)
        == ordinary_looking_definition
    )

    raw_payload = report.to_dict()
    raw_rendered = json.dumps(raw_payload)
    assert baseline_marker in raw_rendered
    assert candidate_marker in raw_rendered

    redacted_payload = redact_formula_defined_xlm_action_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert baseline_marker not in redacted_rendered
    assert candidate_marker not in redacted_rendered
    assert FORMULA_DEFINED_XLM_ACTION_REDACTION in redacted_rendered

    redacted_sarif = report_to_sarif(report, redact_formula_defined_xlm_actions=True)
    redacted_sarif_rendered = json.dumps(redacted_sarif)
    assert baseline_marker not in redacted_sarif_rendered
    assert candidate_marker not in redacted_sarif_rendered
    assert "FF008" in redacted_sarif_rendered
    assert "FF073" in redacted_sarif_rendered


def test_uninvoked_formula_defined_xlm_action_is_profiled(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "stored-name.xlsx")

    def add_stored_action(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.STORED.ACTION",
                attr_text='=EXEC("PRIVATE-STORED-XLM-ACTION")',
            )
        )

    rewrite(workbook_path, add_stored_action)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_actions.to_dict() == {
        "present": True,
        "action_formula_cell_count": 0,
        "action_function_count": 0,
        "action_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_actions.action_cells == frozenset()


def test_recursive_named_formula_defined_xlm_actions_are_cycle_safe(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "recursive.xlsx")

    def add_recursive_action(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.ACTION.LOOP",
                attr_text=(
                    "=LAMBDA(payload,EXEC(payload)"
                    "+FENCE.XLM.ACTION.LOOP(payload))"
                ),
            )
        )
        workbook["Model"]["D2"] = "=FENCE.XLM.ACTION.LOOP(Inputs!B2)"

    rewrite(workbook_path, add_recursive_action)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_actions.to_dict() == {
        "present": True,
        "action_formula_cell_count": 1,
        "action_function_count": 1,
        "action_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_actions.action_cells == frozenset(
        {("Model", "D2")}
    )
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == (
        "FENCE.XLM.ACTION.LOOP",
    )


def test_scoped_formula_defined_xlm_actions_follow_local_precedence(tmp_path) -> None:
    workbook_path = make_scoped_named_lambda_model(tmp_path / "scoped.xlsx")
    workbook = load_workbook(workbook_path)
    model = workbook["Model"]
    report = workbook["Report"]
    model["D2"] = "=FENCE.XLM.ACTION(A2)"
    report["D2"] = "=Model!FENCE.XLM.ACTION(A2)"
    model.defined_names.add(
        DefinedName(
            "FENCE.XLM.ACTION",
            attr_text="=LAMBDA(payload,EXEC(payload))",
            localSheetId=1,
        )
    )
    workbook.save(workbook_path)

    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_actions.to_dict() == {
        "present": True,
        "action_formula_cell_count": 2,
        "action_function_count": 2,
        "action_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_actions.action_cells == frozenset(
        {("Model", "D2"), ("Report", "D2")}
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_direct_worksheet_action_is_outside_formula_defined_xlm_boundary(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "direct-action.xlsx")

    def add_direct_action(workbook) -> None:
        workbook["Model"]["D2"] = '=RUN("PRIVATE-DIRECT-XLM-ACTION")'

    rewrite(workbook_path, add_direct_action)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_actions.to_dict() == {
        "present": False,
        "action_formula_cell_count": 0,
        "action_function_count": 0,
        "action_defined_name_count": 0,
    }


def test_formula_defined_xlm_actions_respect_named_lambda_shadowing(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "shadowed-action.xlsx")

    def add_shadowing_lambda(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "RUN",
                attr_text="=LAMBDA(payload,payload)",
            )
        )
        workbook["Model"]["D2"] = "=RUN(A2)"

    rewrite(workbook_path, add_shadowing_lambda)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_actions.to_dict() == {
        "present": False,
        "action_formula_cell_count": 0,
        "action_function_count": 0,
        "action_defined_name_count": 0,
    }
    assert snapshot.unresolved_reference_tokens == {}


def test_formula_defined_xlm_get_cell_calls_are_propagated_diffed_and_private(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_get_cell_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_defined_xlm_get_cell_model(tmp_path / "candidate.xlsx")
    change_formula_defined_xlm_get_cell_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_calls = {
        "present": True,
        "get_cell_formula_cell_count": 3,
        "get_cell_function_count": 3,
        "get_cell_defined_name_count": 3,
    }
    assert baseline_snapshot.formula_defined_xlm_get_cell_calls.to_dict() == (
        expected_calls
    )
    assert baseline_snapshot.formula_defined_xlm_get_cell_calls.get_cell_cells == (
        frozenset({("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")})
    )
    assert baseline_snapshot.office_custom_functions.present is False
    assert baseline_snapshot.xlm_macro_sheets.present is False
    assert profile["formula_defined_xlm_get_cell_calls"] == expected_calls
    assert "## Formula-defined XLM GET.CELL information" in profile_to_markdown(
        profile
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    get_cell_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_get_cell_calls_changed"
    )
    get_cell_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF070"
    )
    assert get_cell_change.details["before"] == expected_calls
    assert get_cell_change.details["after"] == expected_calls
    assert (
        get_cell_change.details[
            "formula_defined_xlm_get_cell_definition_material_changed"
        ]
        is True
    )
    assert get_cell_finding.details == get_cell_change.details

    ff070_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF070"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["formula_defined_xlm_get_cell_calls"]),
        profile_to_markdown(profile),
        json.dumps(get_cell_change.details),
        json.dumps(get_cell_finding.to_dict()),
        json.dumps(ff070_sarif_result),
    )
    for sensitive_value in (
        "FENCE.XLM.GET.CELL",
        "PRIVATE-XLM-GET-CELL-INPUT-BASELINE",
        "GET.CELL(7,reference)",
        "GET.CELL(8,reference)",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_formula_defined_xlm_get_cell_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_formula_defined_xlm_get_cell_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_defined_xlm_get_cell_model(tmp_path / "candidate.xlsx")
    change_formula_defined_xlm_get_cell_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.formula_defined_xlm_get_cell_calls
        == candidate_snapshot.formula_defined_xlm_get_cell_calls
    )
    assert {("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    get_cell_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_get_cell_calls_changed"
    )
    assert (
        get_cell_change.details[
            "formula_defined_xlm_get_cell_static_input_changed"
        ]
        is True
    )
    assert (
        get_cell_change.details[
            "formula_defined_xlm_get_cell_static_input_change_count"
        ]
        == 1
    )
    assert "FF070" in {finding.rule_id for finding in report.findings}
    assert report.formula_defined_xlm_get_cell_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )


def test_formula_defined_xlm_get_cell_report_redaction_hides_calls_and_inputs(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_get_cell_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_defined_xlm_get_cell_model(tmp_path / "candidate.xlsx")
    change_formula_defined_xlm_get_cell_input(candidate)
    change_formula_defined_xlm_get_cell_call(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    sensitive_values = (
        "PRIVATE-XLM-GET-CELL-INPUT-BASELINE",
        "PRIVATE-XLM-GET-CELL-INPUT-CANDIDATE",
        "GET.CELL(53,Inputs!$A$9)",
        "GET.CELL(54,Inputs!$A$9)",
    )
    raw_rendered = json.dumps(raw_payload)
    assert all(value in raw_rendered for value in sensitive_values)

    redacted_payload = redact_formula_defined_xlm_get_cell_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert all(value not in redacted_rendered for value in sensitive_values)
    assert FORMULA_DEFINED_XLM_GET_CELL_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == raw_payload["summary"]

    redacted_markdown = report_to_markdown(
        report, redact_formula_defined_xlm_get_cell_calls=True
    )
    redacted_sarif = report_to_sarif(
        report, redact_formula_defined_xlm_get_cell_calls=True
    )
    assert "Formula-defined XLM GET.CELL material:** redacted for sharing" in redacted_markdown
    assert "FF070" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, json.dumps(redacted_sarif)):
        assert all(value not in artifact for value in sensitive_values)

    native_formula = "=SUM(A1:A2)"
    assert redact_formula_defined_xlm_get_cell_material(native_formula) == native_formula


def test_formula_defined_xlm_get_cell_report_redaction_hides_unsampled_inputs(
    tmp_path,
) -> None:
    """GET.CELL input redaction must use full private impact evidence."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-XLM-GET-CELL-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-XLM-GET-CELL-CANDIDATE"

    def add_get_cell_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = "=FENCE.XLM.UNSAMPLED.GET.CELL(A9)"
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.UNSAMPLED.GET.CELL",
                attr_text="=LAMBDA(reference,GET.CELL(7,reference))",
            )
        )

    rewrite(baseline, lambda workbook: add_get_cell_fanout(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_get_cell_fanout(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.formula_defined_xlm_get_cell_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )

    redacted = redact_formula_defined_xlm_get_cell_report_payload(report, raw_payload)
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert (
        redacted_change["before"]["value"]
        == FORMULA_DEFINED_XLM_GET_CELL_REDACTION
    )
    assert (
        redacted_change["after"]["value"]
        == FORMULA_DEFINED_XLM_GET_CELL_REDACTION
    )
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_formula_defined_xlm_get_cell_report_redaction_hides_indirect_names(
    tmp_path,
) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def add_get_cell_chain(workbook, marker: str) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.GETCELLWRAPPER",
                attr_text="=LAMBDA(reference,GET.CELL(7,reference))",
            )
        )
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.GETCELLCHAIN",
                attr_text=(
                    f'=LAMBDA(reference,FENCE.XLM.GETCELLWRAPPER("{marker}"))'
                ),
            )
        )

    baseline_marker = "PRIVATE-INDIRECT-XLM-GET-CELL-BASELINE"
    candidate_marker = "PRIVATE-INDIRECT-XLM-GET-CELL-CANDIDATE"
    rewrite(baseline, lambda workbook: add_get_cell_chain(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_get_cell_chain(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    get_cell_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_get_cell_calls_changed"
    )
    assert (
        get_cell_change.details[
            "formula_defined_xlm_get_cell_definition_material_changed"
        ]
        is True
    )
    ordinary_looking_definition = (
        f'=LAMBDA(reference,FENCE.XLM.GETCELLWRAPPER("{baseline_marker}"))'
    )
    assert (
        redact_formula_defined_xlm_get_cell_material(ordinary_looking_definition)
        == ordinary_looking_definition
    )

    raw_payload = report.to_dict()
    raw_rendered = json.dumps(raw_payload)
    assert baseline_marker in raw_rendered
    assert candidate_marker in raw_rendered

    redacted_payload = redact_formula_defined_xlm_get_cell_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert baseline_marker not in redacted_rendered
    assert candidate_marker not in redacted_rendered
    assert FORMULA_DEFINED_XLM_GET_CELL_REDACTION in redacted_rendered

    redacted_sarif = report_to_sarif(
        report, redact_formula_defined_xlm_get_cell_calls=True
    )
    redacted_sarif_rendered = json.dumps(redacted_sarif)
    assert baseline_marker not in redacted_sarif_rendered
    assert candidate_marker not in redacted_sarif_rendered
    assert "FF008" in redacted_sarif_rendered
    assert "FF070" in redacted_sarif_rendered


def test_uninvoked_formula_defined_xlm_get_cell_call_is_profiled(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "stored-name.xlsx")

    def add_stored_get_cell_call(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.STORED.GET.CELL",
                attr_text="=GET.CELL(7,Inputs!$A$1)",
            )
        )

    rewrite(workbook_path, add_stored_get_cell_call)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_get_cell_calls.to_dict() == {
        "present": True,
        "get_cell_formula_cell_count": 0,
        "get_cell_function_count": 0,
        "get_cell_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_get_cell_calls.get_cell_cells == frozenset()


def test_recursive_named_formula_defined_xlm_get_cell_calls_are_cycle_safe(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "recursive.xlsx")

    def add_recursive_get_cell_call(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.GET.CELL.LOOP",
                attr_text=(
                    "=LAMBDA(reference,GET.CELL(7,reference)"
                    "+FENCE.XLM.GET.CELL.LOOP(reference))"
                ),
            )
        )
        workbook["Model"]["D2"] = "=FENCE.XLM.GET.CELL.LOOP(Inputs!B2)"

    rewrite(workbook_path, add_recursive_get_cell_call)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_get_cell_calls.to_dict() == {
        "present": True,
        "get_cell_formula_cell_count": 1,
        "get_cell_function_count": 1,
        "get_cell_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_get_cell_calls.get_cell_cells == frozenset(
        {("Model", "D2")}
    )
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == (
        "FENCE.XLM.GET.CELL.LOOP",
    )


def test_scoped_formula_defined_xlm_get_cell_calls_follow_local_precedence(
    tmp_path,
) -> None:
    workbook_path = make_scoped_named_lambda_model(tmp_path / "scoped.xlsx")
    workbook = load_workbook(workbook_path)
    model = workbook["Model"]
    report = workbook["Report"]
    model["D2"] = "=FENCE.XLM.GET.CELL(A2)"
    report["D2"] = "=Model!FENCE.XLM.GET.CELL(A2)"
    model.defined_names.add(
        DefinedName(
            "FENCE.XLM.GET.CELL",
            attr_text="=LAMBDA(reference,GET.CELL(7,reference))",
            localSheetId=1,
        )
    )
    workbook.save(workbook_path)

    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_get_cell_calls.to_dict() == {
        "present": True,
        "get_cell_formula_cell_count": 2,
        "get_cell_function_count": 2,
        "get_cell_defined_name_count": 1,
    }
    assert snapshot.formula_defined_xlm_get_cell_calls.get_cell_cells == frozenset(
        {("Model", "D2"), ("Report", "D2")}
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_direct_worksheet_get_cell_is_outside_formula_defined_xlm_boundary(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "direct-get-cell.xlsx")

    def add_direct_get_cell(workbook) -> None:
        workbook["Model"]["D2"] = "=GET.CELL(7,A2)"

    rewrite(workbook_path, add_direct_get_cell)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_get_cell_calls.to_dict() == {
        "present": False,
        "get_cell_formula_cell_count": 0,
        "get_cell_function_count": 0,
        "get_cell_defined_name_count": 0,
    }
    assert snapshot.office_custom_functions.present is False


def test_formula_defined_xlm_get_cell_respects_named_lambda_shadowing(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "shadowed-get-cell.xlsx")

    def add_shadowing_lambda(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "GET.CELL",
                attr_text="=LAMBDA(reference,reference)",
            )
        )
        workbook["Model"]["D2"] = "=GET.CELL(A2)"

    rewrite(workbook_path, add_shadowing_lambda)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_get_cell_calls.to_dict() == {
        "present": False,
        "get_cell_formula_cell_count": 0,
        "get_cell_function_count": 0,
        "get_cell_defined_name_count": 0,
    }
    assert snapshot.unresolved_reference_tokens == {}


def test_formula_defined_xlm_environment_information_calls_are_private_and_traced(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_environment_information_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_environment_information_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_environment_information_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_calls = {
        "present": True,
        "environment_information_formula_cell_count": 3,
        "environment_information_function_count": 3,
        "environment_information_defined_name_count": 3,
    }
    assert (
        baseline_snapshot.formula_defined_xlm_environment_information_calls.to_dict()
        == expected_calls
    )
    assert (
        baseline_snapshot.formula_defined_xlm_environment_information_calls.environment_information_cells
        == frozenset({("Inputs", "B2"), ("Inputs", "B3"), ("Inputs", "B4")})
    )
    assert baseline_snapshot.office_custom_functions.present is False
    assert baseline_snapshot.xlm_macro_sheets.present is False
    assert (
        profile["formula_defined_xlm_environment_information_calls"] == expected_calls
    )
    assert "## Formula-defined XLM environment information" in profile_to_markdown(
        profile
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_environment_information_calls_changed"
    )
    environment_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF071"
    )
    assert environment_change.details["before"] == expected_calls
    assert environment_change.details["after"] == expected_calls
    assert (
        environment_change.details[
            "formula_defined_xlm_environment_information_definition_material_changed"
        ]
        is True
    )
    assert environment_finding.details == environment_change.details

    ff071_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF071"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["formula_defined_xlm_environment_information_calls"]),
        profile_to_markdown(profile),
        json.dumps(environment_change.details),
        json.dumps(environment_finding.to_dict()),
        json.dumps(ff071_sarif_result),
    )
    for sensitive_value in (
        "FENCE.XLM.GET.WORKBOOK",
        "PRIVATE-XLM-ENVIRONMENT-INPUT-BASELINE",
        "GET.WORKBOOK(4,workbook_name)",
        "GET.WORKBOOK(1,workbook_name)",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_formula_defined_xlm_environment_information_static_inputs_are_guarded(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_environment_information_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_environment_information_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_environment_information_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.formula_defined_xlm_environment_information_calls
        == candidate_snapshot.formula_defined_xlm_environment_information_calls
    )
    assert {("Inputs", "B2"), ("Inputs", "B4")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_environment_information_calls_changed"
    )
    assert (
        environment_change.details[
            "formula_defined_xlm_environment_information_static_input_changed"
        ]
        is True
    )
    assert (
        environment_change.details[
            "formula_defined_xlm_environment_information_static_input_change_count"
        ]
        == 1
    )
    assert "FF071" in {finding.rule_id for finding in report.findings}
    assert report.formula_defined_xlm_environment_information_static_input_cells == (
        frozenset({("Inputs", "A9")})
    )


def test_formula_defined_xlm_environment_information_report_redaction_hides_calls_and_inputs(
    tmp_path,
) -> None:
    baseline = make_formula_defined_xlm_environment_information_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_environment_information_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_defined_xlm_environment_information_input(candidate)
    change_formula_defined_xlm_environment_information_call(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    sensitive_values = (
        "PRIVATE-XLM-ENVIRONMENT-INPUT-BASELINE",
        "PRIVATE-XLM-ENVIRONMENT-INPUT-CANDIDATE",
        "GET.WORKSPACE(2)",
        "GET.WORKSPACE(3)",
    )
    raw_rendered = json.dumps(raw_payload)
    assert all(value in raw_rendered for value in sensitive_values)

    redacted_payload = redact_formula_defined_xlm_environment_information_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert all(value not in redacted_rendered for value in sensitive_values)
    assert FORMULA_DEFINED_XLM_ENVIRONMENT_INFORMATION_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == raw_payload["summary"]

    redacted_markdown = report_to_markdown(
        report, redact_formula_defined_xlm_environment_information_calls=True
    )
    redacted_sarif = report_to_sarif(
        report, redact_formula_defined_xlm_environment_information_calls=True
    )
    assert (
        "Formula-defined XLM environment-information material:** redacted for sharing"
        in redacted_markdown
    )
    assert "FF071" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, json.dumps(redacted_sarif)):
        assert all(value not in artifact for value in sensitive_values)

    native_formula = "=SUM(A1:A2)"
    assert (
        redact_formula_defined_xlm_environment_information_material(native_formula)
        == native_formula
    )


def test_formula_defined_xlm_environment_information_report_redaction_hides_unsampled_inputs(
    tmp_path,
) -> None:
    """Environment-call input redaction must use full private impact evidence."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-XLM-ENVIRONMENT-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-XLM-ENVIRONMENT-CANDIDATE"

    def add_environment_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = "=FENCE.XLM.UNSAMPLED.ENVIRONMENT(A9)"
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.UNSAMPLED.ENVIRONMENT",
                attr_text="=LAMBDA(workbook_name,GET.WORKBOOK(4,workbook_name))",
            )
        )

    rewrite(baseline, lambda workbook: add_environment_fanout(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_environment_fanout(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.formula_defined_xlm_environment_information_static_input_cells == (
        frozenset({("Inputs", "A9")})
    )

    redacted = redact_formula_defined_xlm_environment_information_report_payload(
        report, raw_payload
    )
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert (
        redacted_change["before"]["value"]
        == FORMULA_DEFINED_XLM_ENVIRONMENT_INFORMATION_REDACTION
    )
    assert (
        redacted_change["after"]["value"]
        == FORMULA_DEFINED_XLM_ENVIRONMENT_INFORMATION_REDACTION
    )
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_formula_defined_xlm_environment_information_report_redaction_hides_indirect_names(
    tmp_path,
) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def add_environment_chain(workbook, marker: str) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.ENVIRONMENTWRAPPER",
                attr_text="=LAMBDA(workbook_name,GET.WORKBOOK(4,workbook_name))",
            )
        )
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.ENVIRONMENTCHAIN",
                attr_text=(
                    f'=LAMBDA(workbook_name,FENCE.XLM.ENVIRONMENTWRAPPER("{marker}"))'
                ),
            )
        )

    baseline_marker = "PRIVATE-INDIRECT-XLM-ENVIRONMENT-BASELINE"
    candidate_marker = "PRIVATE-INDIRECT-XLM-ENVIRONMENT-CANDIDATE"
    rewrite(baseline, lambda workbook: add_environment_chain(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_environment_chain(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_defined_xlm_environment_information_calls_changed"
    )
    assert (
        environment_change.details[
            "formula_defined_xlm_environment_information_definition_material_changed"
        ]
        is True
    )
    ordinary_looking_definition = (
        f'=LAMBDA(workbook_name,FENCE.XLM.ENVIRONMENTWRAPPER("{baseline_marker}"))'
    )
    assert (
        redact_formula_defined_xlm_environment_information_material(
            ordinary_looking_definition
        )
        == ordinary_looking_definition
    )

    raw_payload = report.to_dict()
    raw_rendered = json.dumps(raw_payload)
    assert baseline_marker in raw_rendered
    assert candidate_marker in raw_rendered

    redacted_payload = redact_formula_defined_xlm_environment_information_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert baseline_marker not in redacted_rendered
    assert candidate_marker not in redacted_rendered
    assert FORMULA_DEFINED_XLM_ENVIRONMENT_INFORMATION_REDACTION in redacted_rendered

    redacted_sarif = report_to_sarif(
        report, redact_formula_defined_xlm_environment_information_calls=True
    )
    redacted_sarif_rendered = json.dumps(redacted_sarif)
    assert baseline_marker not in redacted_sarif_rendered
    assert candidate_marker not in redacted_sarif_rendered
    assert "FF008" in redacted_sarif_rendered
    assert "FF071" in redacted_sarif_rendered


def test_xlm_environment_information_state_is_not_simulated(tmp_path) -> None:
    baseline = make_formula_defined_xlm_environment_information_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_defined_xlm_environment_information_model(
        tmp_path / "candidate.xlsx"
    )
    workbook = load_workbook(candidate)
    workbook.create_sheet("Candidate State")
    workbook.save(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.formula_defined_xlm_environment_information_calls
        == candidate_snapshot.formula_defined_xlm_environment_information_calls
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    assert "FF071" not in {finding.rule_id for finding in report.findings}


def test_uninvoked_formula_defined_xlm_environment_information_is_profiled(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "stored-name.xlsx")

    def add_stored_environment_call(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.STORED.ENVIRONMENT",
                attr_text="=GET.DOCUMENT(37)",
            )
        )

    rewrite(workbook_path, add_stored_environment_call)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_environment_information_calls.to_dict() == {
        "present": True,
        "environment_information_formula_cell_count": 0,
        "environment_information_function_count": 0,
        "environment_information_defined_name_count": 1,
    }
    assert (
        snapshot.formula_defined_xlm_environment_information_calls.environment_information_cells
        == frozenset()
    )


def test_recursive_named_xlm_environment_information_calls_are_cycle_safe(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "recursive.xlsx")

    def add_recursive_environment_call(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.ENVIRONMENT.LOOP",
                attr_text=(
                    "=LAMBDA(value,GET.WORKSPACE(2)"
                    "+FENCE.XLM.ENVIRONMENT.LOOP(value))"
                ),
            )
        )
        workbook["Model"]["D2"] = "=FENCE.XLM.ENVIRONMENT.LOOP(Inputs!B2)"

    rewrite(workbook_path, add_recursive_environment_call)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_environment_information_calls.to_dict() == {
        "present": True,
        "environment_information_formula_cell_count": 1,
        "environment_information_function_count": 1,
        "environment_information_defined_name_count": 1,
    }
    assert (
        snapshot.formula_defined_xlm_environment_information_calls.environment_information_cells
        == frozenset({("Model", "D2")})
    )
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == (
        "FENCE.XLM.ENVIRONMENT.LOOP",
    )


def test_scoped_xlm_environment_information_calls_follow_local_precedence(
    tmp_path,
) -> None:
    workbook_path = make_scoped_named_lambda_model(tmp_path / "scoped.xlsx")
    workbook = load_workbook(workbook_path)
    model = workbook["Model"]
    report = workbook["Report"]
    model["D2"] = "=FENCE.XLM.ENVIRONMENT(A2)"
    report["D2"] = "=Model!FENCE.XLM.ENVIRONMENT(A2)"
    model.defined_names.add(
        DefinedName(
            "FENCE.XLM.ENVIRONMENT",
            attr_text="=LAMBDA(value,GET.WORKBOOK(4,value))",
            localSheetId=1,
        )
    )
    workbook.save(workbook_path)

    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_environment_information_calls.to_dict() == {
        "present": True,
        "environment_information_formula_cell_count": 2,
        "environment_information_function_count": 2,
        "environment_information_defined_name_count": 1,
    }
    assert (
        snapshot.formula_defined_xlm_environment_information_calls.environment_information_cells
        == frozenset({("Model", "D2"), ("Report", "D2")})
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_direct_worksheet_environment_information_is_outside_stored_boundary(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "direct-environment.xlsx")

    def add_direct_environment_call(workbook) -> None:
        workbook["Model"]["D2"] = "=GET.WORKBOOK(4)"

    rewrite(workbook_path, add_direct_environment_call)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_environment_information_calls.to_dict() == {
        "present": False,
        "environment_information_formula_cell_count": 0,
        "environment_information_function_count": 0,
        "environment_information_defined_name_count": 0,
    }
    assert snapshot.office_custom_functions.present is False


def test_xlm_environment_information_respects_named_lambda_shadowing(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "shadowed-environment.xlsx")

    def add_shadowing_lambda(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "GET.WORKSPACE",
                attr_text="=LAMBDA(value,value)",
            )
        )
        workbook["Model"]["D2"] = "=GET.WORKSPACE(A2)"

    rewrite(workbook_path, add_shadowing_lambda)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_defined_xlm_environment_information_calls.to_dict() == {
        "present": False,
        "environment_information_formula_cell_count": 0,
        "environment_information_function_count": 0,
        "environment_information_defined_name_count": 0,
    }
    assert snapshot.unresolved_reference_tokens == {}


def test_native_environment_information_calls_are_private_and_traced(tmp_path) -> None:
    baseline = make_formula_environment_information_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_environment_information_model(tmp_path / "candidate.xlsx")
    change_formula_environment_information_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_calls = {
        "present": True,
        "environment_information_formula_cell_count": 5,
        "environment_information_function_count": 6,
        "environment_information_defined_name_count": 2,
        "implicit_cell_reference_function_count": 3,
        "implicit_sheets_reference_function_count": 0,
        "sheet_function_count": 0,
        "sheets_function_count": 0,
    }
    assert baseline_snapshot.formula_environment_information_calls.to_dict() == expected_calls
    assert (
        baseline_snapshot.summary()[
            "formula_environment_information_formula_cell_count"
        ]
        == 5
    )
    assert baseline_snapshot.summary()["formula_environment_information_function_count"] == 6
    assert (
        baseline_snapshot.summary()[
            "formula_environment_information_implicit_cell_reference_function_count"
        ]
        == 3
    )
    assert baseline_snapshot.summary()["has_formula_environment_information_calls"] is True
    assert (
        baseline_snapshot.formula_environment_information_calls.environment_information_cells
        == frozenset(
            {
                ("Inputs", "B2"),
                ("Inputs", "B3"),
                ("Inputs", "B4"),
                ("Inputs", "B5"),
                ("Inputs", "B6"),
            }
        )
    )
    assert profile["formula_environment_information_calls"] == expected_calls
    assert "## Native CELL, INFO, SHEET, and SHEETS information" in profile_to_markdown(
        profile
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_environment_information_calls_changed"
    )
    environment_finding = next(
        finding for finding in report.findings if finding.rule_id == "FF072"
    )
    assert environment_change.details["before"] == expected_calls
    assert environment_change.details["after"] == expected_calls
    assert (
        environment_change.details[
            "formula_environment_information_definition_material_changed"
        ]
        is True
    )
    assert environment_finding.details == environment_change.details

    ff072_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF072"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["formula_environment_information_calls"]),
        profile_to_markdown(profile),
        json.dumps(environment_change.details),
        json.dumps(environment_finding.to_dict()),
        json.dumps(ff072_sarif_result),
    )
    for sensitive_value in (
        "FENCE.NATIVE.ENVIRONMENT",
        "PRIVATE-NATIVE-ENVIRONMENT-INPUT-BASELINE",
        'CELL("filename")',
        'INFO("system")',
        'INFO("osversion")',
        "FORMULAFENCE_CELL_IMPLICIT_REFERENCE_MARKER",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_native_environment_information_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_formula_environment_information_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_environment_information_model(tmp_path / "candidate.xlsx")
    change_formula_environment_information_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.formula_environment_information_calls
        == candidate_snapshot.formula_environment_information_calls
    )
    assert {("Inputs", "B3"), ("Inputs", "B5")} <= set(
        baseline_snapshot.reverse_dependencies[("Inputs", "A9")]
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_environment_information_calls_changed"
    )
    assert (
        environment_change.details[
            "formula_environment_information_static_input_changed"
        ]
        is True
    )
    assert (
        environment_change.details[
            "formula_environment_information_static_input_change_count"
        ]
        == 1
    )
    assert "FF072" in {finding.rule_id for finding in report.findings}
    assert report.formula_environment_information_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )


def test_native_environment_information_report_redaction_hides_calls_and_inputs(
    tmp_path,
) -> None:
    baseline = make_formula_environment_information_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_environment_information_model(tmp_path / "candidate.xlsx")
    change_formula_environment_information_input(candidate)
    change_formula_environment_information_definition(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    sensitive_values = (
        "PRIVATE-NATIVE-ENVIRONMENT-INPUT-BASELINE",
        "PRIVATE-NATIVE-ENVIRONMENT-INPUT-CANDIDATE",
        "system",
        "osversion",
    )
    raw_rendered = json.dumps(raw_payload)
    assert all(value in raw_rendered for value in sensitive_values)

    redacted_payload = redact_formula_environment_information_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert all(value not in redacted_rendered for value in sensitive_values)
    assert FORMULA_ENVIRONMENT_INFORMATION_REDACTION in redacted_rendered
    assert redacted_payload["summary"] == raw_payload["summary"]

    redacted_markdown = report_to_markdown(
        report, redact_formula_environment_information=True
    )
    redacted_html = report_to_html(
        report, redact_formula_environment_information=True
    )
    redacted_sarif = report_to_sarif(
        report, redact_formula_environment_information=True
    )
    assert (
        "Formula environment-information material:** redacted for sharing"
        in redacted_markdown
    )
    assert "Sharing boundaries are active" in redacted_html
    assert FORMULA_ENVIRONMENT_INFORMATION_REDACTION in redacted_html
    assert "FF072" in json.dumps(redacted_sarif)
    for artifact in (redacted_markdown, redacted_html, json.dumps(redacted_sarif)):
        assert all(value not in artifact for value in sensitive_values)

    for environment_formula in (
        '=CELL("filename")',
        '=INFO("directory")',
        '=SHEET("Private tab")',
        "=SHEETS()",
    ):
        assert (
            redact_formula_environment_information_material(environment_formula)
            == FORMULA_ENVIRONMENT_INFORMATION_REDACTION
        )
    ordinary_formula = "=SUM(A1:A2)"
    assert redact_formula_environment_information_material(ordinary_formula) == ordinary_formula


def test_native_environment_information_report_redaction_hides_unsampled_inputs(
    tmp_path,
) -> None:
    """Native-call input redaction must use full private impact evidence."""
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    baseline_marker = "PRIVATE-UNSAMPLED-NATIVE-ENVIRONMENT-BASELINE"
    candidate_marker = "PRIVATE-UNSAMPLED-NATIVE-ENVIRONMENT-CANDIDATE"

    def add_environment_fanout(workbook, marker: str) -> None:
        inputs = workbook["Inputs"]
        inputs["A9"] = marker
        for column in range(2, 22):
            inputs.cell(row=1, column=column).value = "=A9"
        inputs["Z1"] = "=FENCE.NATIVE.UNSAMPLED.ENVIRONMENT(A9)"
        workbook.defined_names.add(
            DefinedName(
                "FENCE.NATIVE.UNSAMPLED.ENVIRONMENT",
                attr_text='=LAMBDA(value,CELL("filename",value))',
            )
        )

    rewrite(baseline, lambda workbook: add_environment_fanout(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_environment_fanout(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    raw_payload = report.to_dict()
    input_change = next(
        change for change in raw_payload["changes"] if change["location"] == "Inputs!A9"
    )
    assert "Inputs!Z1" not in input_change["impacted_cells"]
    assert report.formula_environment_information_static_input_cells == frozenset(
        {("Inputs", "A9")}
    )

    redacted = redact_formula_environment_information_report_payload(report, raw_payload)
    redacted_change = next(
        change for change in redacted["changes"] if change["location"] == "Inputs!A9"
    )
    assert (
        redacted_change["before"]["value"]
        == FORMULA_ENVIRONMENT_INFORMATION_REDACTION
    )
    assert (
        redacted_change["after"]["value"]
        == FORMULA_ENVIRONMENT_INFORMATION_REDACTION
    )
    rendered = json.dumps(redacted)
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered


def test_native_environment_information_report_redaction_hides_indirect_names(
    tmp_path,
) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")

    def add_environment_chain(workbook, marker: str) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.NATIVE.ENVIRONMENTWRAPPER",
                attr_text='=LAMBDA(value,CELL("filename",value))',
            )
        )
        workbook.defined_names.add(
            DefinedName(
                "FENCE.NATIVE.ENVIRONMENTCHAIN",
                attr_text=(
                    f'=LAMBDA(value,FENCE.NATIVE.ENVIRONMENTWRAPPER("{marker}"))'
                ),
            )
        )

    baseline_marker = "PRIVATE-INDIRECT-NATIVE-ENVIRONMENT-BASELINE"
    candidate_marker = "PRIVATE-INDIRECT-NATIVE-ENVIRONMENT-CANDIDATE"
    rewrite(baseline, lambda workbook: add_environment_chain(workbook, baseline_marker))
    rewrite(candidate, lambda workbook: add_environment_chain(workbook, candidate_marker))

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_environment_information_calls_changed"
    )
    assert (
        environment_change.details[
            "formula_environment_information_definition_material_changed"
        ]
        is True
    )
    ordinary_looking_definition = (
        f'=LAMBDA(value,FENCE.NATIVE.ENVIRONMENTWRAPPER("{baseline_marker}"))'
    )
    assert (
        redact_formula_environment_information_material(ordinary_looking_definition)
        == ordinary_looking_definition
    )

    raw_payload = report.to_dict()
    raw_rendered = json.dumps(raw_payload)
    assert baseline_marker in raw_rendered
    assert candidate_marker in raw_rendered

    redacted_payload = redact_formula_environment_information_report_payload(
        report, raw_payload
    )
    redacted_rendered = json.dumps(redacted_payload)
    assert baseline_marker not in redacted_rendered
    assert candidate_marker not in redacted_rendered
    assert FORMULA_ENVIRONMENT_INFORMATION_REDACTION in redacted_rendered

    redacted_sarif = report_to_sarif(
        report, redact_formula_environment_information=True
    )
    redacted_sarif_rendered = json.dumps(redacted_sarif)
    assert baseline_marker not in redacted_sarif_rendered
    assert candidate_marker not in redacted_sarif_rendered
    assert "FF008" in redacted_sarif_rendered
    assert "FF072" in redacted_sarif_rendered


def test_native_environment_information_state_is_not_simulated(tmp_path) -> None:
    baseline = make_formula_environment_information_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_environment_information_model(tmp_path / "candidate.xlsx")
    workbook = load_workbook(candidate)
    workbook.create_sheet("Candidate State")
    workbook.save(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.formula_environment_information_calls
        == candidate_snapshot.formula_environment_information_calls
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    assert "FF072" not in {finding.rule_id for finding in report.findings}


def test_native_environment_information_tracks_implicit_cell_reference_changes(
    tmp_path,
) -> None:
    baseline = make_formula_environment_information_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_environment_information_model(tmp_path / "candidate.xlsx")
    workbook = load_workbook(candidate)
    workbook["Inputs"]["B2"] = '=CELL("filename",A1)'
    workbook.save(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.formula_environment_information_calls.to_dict()[
            "implicit_cell_reference_function_count"
        ]
        == 3
    )
    assert (
        candidate_snapshot.formula_environment_information_calls.to_dict()[
            "implicit_cell_reference_function_count"
        ]
        == 2
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_environment_information_calls_changed"
    )
    assert (
        environment_change.details[
            "formula_environment_information_invocation_material_changed"
        ]
        is True
    )


def test_native_environment_information_respects_named_lambda_shadowing(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "shadowed-native-environment.xlsx")

    def add_shadowing_lambda(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "CELL",
                attr_text="=LAMBDA(value,value)",
            )
        )
        workbook["Model"]["D2"] = '=CELL("filename")'

    rewrite(workbook_path, add_shadowing_lambda)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_environment_information_calls.to_dict() == {
        "present": False,
        "environment_information_formula_cell_count": 0,
        "environment_information_function_count": 0,
        "environment_information_defined_name_count": 0,
        "implicit_cell_reference_function_count": 0,
        "implicit_sheets_reference_function_count": 0,
        "sheet_function_count": 0,
        "sheets_function_count": 0,
    }
    assert snapshot.unresolved_reference_tokens == {}


def test_uninvoked_native_environment_information_is_profiled(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "stored-native-environment.xlsx")

    def add_stored_environment_call(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.NATIVE.STORED.ENVIRONMENT",
                attr_text='=CELL("filename")',
            )
        )

    rewrite(workbook_path, add_stored_environment_call)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_environment_information_calls.to_dict() == {
        "present": True,
        "environment_information_formula_cell_count": 0,
        "environment_information_function_count": 0,
        "environment_information_defined_name_count": 1,
        "implicit_cell_reference_function_count": 0,
        "implicit_sheets_reference_function_count": 0,
        "sheet_function_count": 0,
        "sheets_function_count": 0,
    }
    assert (
        snapshot.formula_environment_information_calls.environment_information_cells
        == frozenset()
    )


def test_recursive_named_native_environment_information_calls_are_cycle_safe(
    tmp_path,
) -> None:
    workbook_path = make_model(tmp_path / "recursive-native-environment.xlsx")

    def add_recursive_environment_call(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.NATIVE.ENVIRONMENT.LOOP",
                attr_text=(
                    '=LAMBDA(value,CELL("filename")'
                    "+FENCE.NATIVE.ENVIRONMENT.LOOP(value))"
                ),
            )
        )
        workbook["Model"]["D2"] = "=FENCE.NATIVE.ENVIRONMENT.LOOP(Inputs!B2)"

    rewrite(workbook_path, add_recursive_environment_call)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_environment_information_calls.to_dict() == {
        "present": True,
        "environment_information_formula_cell_count": 1,
        "environment_information_function_count": 1,
        "environment_information_defined_name_count": 1,
        "implicit_cell_reference_function_count": 1,
        "implicit_sheets_reference_function_count": 0,
        "sheet_function_count": 0,
        "sheets_function_count": 0,
    }
    assert snapshot.formula_environment_information_calls.environment_information_cells == (
        frozenset({("Model", "D2")})
    )
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == (
        "FENCE.NATIVE.ENVIRONMENT.LOOP",
    )


def test_scoped_native_environment_information_calls_follow_local_precedence(
    tmp_path,
) -> None:
    workbook_path = make_scoped_named_lambda_model(tmp_path / "scoped-native.xlsx")
    workbook = load_workbook(workbook_path)
    model = workbook["Model"]
    report = workbook["Report"]
    model["D2"] = "=FENCE.NATIVE.ENVIRONMENT(A2)"
    report["D2"] = "=Model!FENCE.NATIVE.ENVIRONMENT(A2)"
    model.defined_names.add(
        DefinedName(
            "FENCE.NATIVE.ENVIRONMENT",
            attr_text='=LAMBDA(value,CELL("filename")+value)',
            localSheetId=1,
        )
    )
    workbook.save(workbook_path)

    snapshot = load_snapshot(workbook_path)

    assert snapshot.formula_environment_information_calls.to_dict() == {
        "present": True,
        "environment_information_formula_cell_count": 2,
        "environment_information_function_count": 2,
        "environment_information_defined_name_count": 1,
        "implicit_cell_reference_function_count": 2,
        "implicit_sheets_reference_function_count": 0,
        "sheet_function_count": 0,
        "sheets_function_count": 0,
    }
    assert snapshot.formula_environment_information_calls.environment_information_cells == (
        frozenset({("Model", "D2"), ("Report", "D2")})
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_native_workbook_tab_information_calls_are_private_and_traced(tmp_path) -> None:
    baseline = make_formula_workbook_structure_information_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_workbook_structure_information_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_workbook_structure_information_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    expected_calls = {
        "present": True,
        "environment_information_formula_cell_count": 6,
        "environment_information_function_count": 7,
        "environment_information_defined_name_count": 2,
        "implicit_cell_reference_function_count": 0,
        "implicit_sheets_reference_function_count": 3,
        "sheet_function_count": 3,
        "sheets_function_count": 4,
    }
    assert baseline_snapshot.formula_environment_information_calls.to_dict() == expected_calls
    assert baseline_snapshot.workbook_tab_order == ("Inputs", "Model", "Report")
    assert baseline_snapshot.workbook_tab_order_complete is True
    assert (
        baseline_snapshot.summary()["formula_sheet_function_count"]
        == expected_calls["sheet_function_count"]
    )
    assert (
        baseline_snapshot.summary()["formula_sheets_function_count"]
        == expected_calls["sheets_function_count"]
    )
    assert (
        baseline_snapshot.summary()[
            "formula_environment_information_implicit_sheets_reference_function_count"
        ]
        == expected_calls["implicit_sheets_reference_function_count"]
    )
    assert profile["formula_environment_information_calls"] == expected_calls
    assert "## Native CELL, INFO, SHEET, and SHEETS information" in profile_to_markdown(
        profile
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_environment_information_calls_changed"
    )
    assert environment_change.details["before"] == expected_calls
    assert environment_change.details["after"] == expected_calls
    assert (
        environment_change.details[
            "formula_environment_information_definition_material_changed"
        ]
        is True
    )

    ff072_sarif_result = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF072"
    )
    rendered_ledger_artifacts = (
        json.dumps(profile["formula_environment_information_calls"]),
        profile_to_markdown(profile),
        json.dumps(environment_change.details),
        json.dumps(ff072_sarif_result),
    )
    for sensitive_value in (
        "FENCE.TAB.INFORMATION",
        "PRIVATE-WORKBOOK-TAB-INPUT-BASELINE",
        "SHEET()",
        "SHEETS()",
        "FORMULAFENCE_SHEETS_IMPLICIT_REFERENCE_MARKER",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_ledger_artifacts)


def test_native_workbook_tab_information_static_inputs_are_guarded(tmp_path) -> None:
    baseline = make_formula_workbook_structure_information_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_workbook_structure_information_model(
        tmp_path / "candidate.xlsx"
    )
    change_formula_workbook_structure_information_input(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_environment_information_calls_changed"
    )
    assert (
        environment_change.details[
            "formula_environment_information_static_input_changed"
        ]
        is True
    )
    assert (
        environment_change.details[
            "formula_environment_information_static_input_change_count"
        ]
        == 1
    )
    assert "FF072" in {finding.rule_id for finding in report.findings}


def test_native_workbook_tab_information_detects_all_tab_catalog_changes(tmp_path) -> None:
    baseline = make_formula_workbook_structure_information_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_workbook_structure_information_model(
        tmp_path / "candidate.xlsx"
    )
    workbook = load_workbook(candidate)
    inputs = workbook["Inputs"]
    inputs["D1"] = 1
    inputs["D2"] = 2
    chart = BarChart()
    chart.add_data(Reference(inputs, min_col=4, min_row=1, max_row=2))
    chart_sheet = workbook.create_chartsheet("Chart", index=1)
    chart_sheet.add_chart(chart)
    workbook.save(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.sheet_order == candidate_snapshot.sheet_order
    assert candidate_snapshot.workbook_tab_order == (
        "Inputs",
        "Chart",
        "Model",
        "Report",
    )
    assert candidate_snapshot.worksheet_tab_order_complete is True
    assert candidate_snapshot.worksheet_tab_order == ("Inputs", "Model", "Report")
    assert (
        baseline_snapshot.formula_environment_information_calls
        == candidate_snapshot.formula_environment_information_calls
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    environment_change = next(
        change
        for change in report.changes
        if change.kind == "formula_environment_information_calls_changed"
    )
    assert (
        environment_change.details[
            "formula_environment_information_workbook_tab_catalog_changed"
        ]
        is True
    )
    assert "formula_environment_information_invocation_material_changed" not in (
        environment_change.details
    )
    assert "formula_environment_information_definition_material_changed" not in (
        environment_change.details
    )
    assert "formula_environment_information_static_input_changed" not in (
        environment_change.details
    )
    assert "FF072" in {finding.rule_id for finding in report.findings}
    assert "Chart" not in json.dumps(environment_change.details)


def test_native_workbook_tab_information_detects_tab_removal_and_reordering(
    tmp_path,
) -> None:
    baseline = make_formula_workbook_structure_information_model(
        tmp_path / "baseline.xlsx"
    )
    baseline_snapshot = load_snapshot(baseline)

    removed = make_formula_workbook_structure_information_model(
        tmp_path / "removed.xlsx"
    )
    workbook = load_workbook(removed)
    workbook.remove(workbook["Model"])
    workbook.save(removed)
    removed_snapshot = load_snapshot(removed)
    assert removed_snapshot.workbook_tab_order == ("Inputs", "Report")

    removed_report = compare_snapshots(baseline_snapshot, removed_snapshot)
    removed_change = next(
        change
        for change in removed_report.changes
        if change.kind == "formula_environment_information_calls_changed"
    )
    assert (
        removed_change.details[
            "formula_environment_information_workbook_tab_catalog_changed"
        ]
        is True
    )

    reordered = make_formula_workbook_structure_information_model(
        tmp_path / "reordered.xlsx"
    )
    workbook = load_workbook(reordered)
    report_sheet = workbook["Report"]
    workbook._sheets.remove(report_sheet)  # noqa: SLF001 - exercise tab-order persistence
    workbook._sheets.insert(0, report_sheet)  # noqa: SLF001 - exercise tab-order persistence
    workbook.save(reordered)
    reordered_snapshot = load_snapshot(reordered)
    assert reordered_snapshot.workbook_tab_order == ("Report", "Inputs", "Model")

    reordered_report = compare_snapshots(baseline_snapshot, reordered_snapshot)
    reordered_change = next(
        change
        for change in reordered_report.changes
        if change.kind == "formula_environment_information_calls_changed"
    )
    assert (
        reordered_change.details[
            "formula_environment_information_workbook_tab_catalog_changed"
        ]
        is True
    )


def test_native_workbook_tab_information_ignores_visibility_only(tmp_path) -> None:
    baseline = make_formula_workbook_structure_information_model(
        tmp_path / "baseline.xlsx"
    )
    candidate = make_formula_workbook_structure_information_model(
        tmp_path / "candidate.xlsx"
    )
    workbook = load_workbook(candidate)
    workbook["Model"].sheet_state = "hidden"
    workbook.save(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert baseline_snapshot.workbook_tab_order == candidate_snapshot.workbook_tab_order

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    assert "FF007" in {finding.rule_id for finding in report.findings}
    assert "FF072" not in {finding.rule_id for finding in report.findings}


def test_native_workbook_tab_information_does_not_guess_explicit_sheets_reference(
    tmp_path,
) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    for workbook_path in (baseline, candidate):
        workbook = load_workbook(workbook_path)
        workbook["Inputs"]["D2"] = "=SHEETS(Inputs!$A$1)"
        workbook.save(workbook_path)
    workbook = load_workbook(candidate)
    workbook.create_sheet("Inserted", 0)
    workbook.save(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    assert (
        baseline_snapshot.formula_environment_information_calls.sheets_function_count
        == 1
    )
    assert (
        baseline_snapshot.formula_environment_information_calls.implicit_sheets_reference_function_count
        == 0
    )

    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    assert "FF072" not in {finding.rule_id for finding in report.findings}


def test_recursive_named_custom_function_candidates_are_cycle_safe(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "recursive.xlsx")

    def add_recursive_custom_function(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCE.LOOP",
                attr_text=(
                    "=LAMBDA(value,CONTOSO.GETDATA(value)+FENCE.LOOP(value))"
                ),
            )
        )
        workbook["Model"]["D2"] = "=FENCE.LOOP(Inputs!B2)"

    rewrite(workbook_path, add_recursive_custom_function)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.office_custom_functions.to_dict() == {
        "present": True,
        "namespaced_custom_function_formula_cell_count": 1,
        "namespaced_custom_function_call_count": 1,
        "namespaced_custom_function_namespace_count": 1,
        "namespaced_custom_function_defined_name_count": 1,
    }
    assert snapshot.office_custom_functions.call_cells == frozenset({("Model", "D2")})
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == ("FENCE.LOOP",)


def test_recursive_named_unqualified_runtime_functions_are_cycle_safe(tmp_path) -> None:
    workbook_path = make_model(tmp_path / "recursive-runtime.xlsx")

    def add_recursive_runtime_function(workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "FENCELOOP",
                attr_text="=LAMBDA(value,PRIVATEUDF(value)+FENCELOOP(value))",
            )
        )
        workbook["Model"]["D2"] = "=FENCELOOP(Inputs!B2)"

    rewrite(workbook_path, add_recursive_runtime_function)
    snapshot = load_snapshot(workbook_path)

    assert snapshot.unqualified_runtime_functions.to_dict() == {
        "present": True,
        "unqualified_runtime_function_formula_cell_count": 1,
        "unqualified_runtime_function_call_count": 1,
        "unqualified_runtime_function_defined_name_count": 1,
    }
    assert snapshot.unqualified_runtime_functions.call_cells == frozenset(
        {("Model", "D2")}
    )
    assert snapshot.unresolved_reference_tokens[("Model", "D2")] == ("FENCELOOP",)


def test_scoped_named_custom_function_candidates_follow_local_precedence(tmp_path) -> None:
    workbook_path = make_scoped_named_lambda_model(tmp_path / "scoped.xlsx")
    workbook = load_workbook(workbook_path)
    model = workbook["Model"]
    report = workbook["Report"]
    model["C2"] = "=FENCE.LOCAL(A2)"
    report["C2"] = "=Model!FENCE.LOCAL(A2)"
    model.defined_names.add(
        DefinedName(
            "FENCE.LOCAL",
            attr_text="=LAMBDA(value,CONTOSO.GETDATA(value))",
            localSheetId=1,
        )
    )
    workbook.save(workbook_path)

    snapshot = load_snapshot(workbook_path)

    assert snapshot.office_custom_functions.to_dict() == {
        "present": True,
        "namespaced_custom_function_formula_cell_count": 2,
        "namespaced_custom_function_call_count": 2,
        "namespaced_custom_function_namespace_count": 1,
        "namespaced_custom_function_defined_name_count": 1,
    }
    assert snapshot.office_custom_functions.call_cells == frozenset(
        {("Model", "C2"), ("Report", "C2")}
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_scoped_unqualified_runtime_functions_follow_local_precedence(tmp_path) -> None:
    workbook_path = make_scoped_named_lambda_model(tmp_path / "scoped-runtime.xlsx")
    workbook = load_workbook(workbook_path)
    model = workbook["Model"]
    report = workbook["Report"]
    model["C2"] = "=LOCALUDF(A2)"
    report["C2"] = "=Model!LOCALUDF(A2)"
    model.defined_names.add(
        DefinedName(
            "LOCALUDF",
            attr_text="=LAMBDA(value,PRIVATEUDF(value))",
            localSheetId=1,
        )
    )
    workbook.save(workbook_path)

    snapshot = load_snapshot(workbook_path)

    assert snapshot.unqualified_runtime_functions.to_dict() == {
        "present": True,
        "unqualified_runtime_function_formula_cell_count": 2,
        "unqualified_runtime_function_call_count": 2,
        "unqualified_runtime_function_defined_name_count": 1,
    }
    assert snapshot.unqualified_runtime_functions.call_cells == frozenset(
        {("Model", "C2"), ("Report", "C2")}
    )
    assert snapshot.unresolved_reference_tokens == {}


def test_python_in_excel_relationship_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_python_in_excel_model(tmp_path / "baseline.xlsx")
    candidate = tmp_path / "candidate.xlsx"
    shutil.copyfile(baseline, candidate)
    renumber_python_in_excel_relationship_identifier(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert baseline_snapshot.python_in_excel == candidate_snapshot.python_in_excel
    assert "python_in_excel_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF065" not in {finding.rule_id for finding in report.findings}


def test_python_in_excel_python_scripts_relationship_identifier_rewrites_are_ignored(
    tmp_path,
) -> None:
    baseline = make_python_in_excel_scripts_model(tmp_path / "baseline.xlsx")
    candidate = tmp_path / "candidate.xlsx"
    shutil.copyfile(baseline, candidate)
    renumber_python_in_excel_scripts_relationship_identifier(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert baseline_snapshot.python_in_excel == candidate_snapshot.python_in_excel
    assert "python_in_excel_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF065" not in {finding.rule_id for finding in report.findings}


def test_malformed_python_in_excel_metadata_reports_a_coverage_gap(tmp_path) -> None:
    baseline = make_python_in_excel_model(tmp_path / "baseline.xlsx")
    candidate = make_python_in_excel_model(tmp_path / "candidate.xlsx")
    corrupt_python_in_excel_package(candidate)

    candidate_snapshot = load_snapshot(candidate)
    assert candidate_snapshot.python_in_excel.to_dict() == {
        "present": True,
        "python_part_count": 1,
        "python_formula_cell_count": 2,
        "python_function_count": 2,
        "python_script_count": 0,
        "python_environment_definition_count": 0,
        "python_initialization_count": 0,
        "unrecognized_python_in_excel_count": 2,
    }
    assert any(
        "Python-in-Excel" in warning for warning in candidate_snapshot.parser_warnings
    )

    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    assert "FF065" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(candidate_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-PYTHON-INIT-BASELINE",
        "PRIVATE-PYTHON-SCRIPT-BASELINE",
        "PRIVATE-PYTHON-SCRIPT-SECOND",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_malformed_python_in_excel_python_scripts_metadata_reports_a_coverage_gap(
    tmp_path,
) -> None:
    baseline = make_python_in_excel_scripts_model(tmp_path / "baseline.xlsx")
    candidate = make_python_in_excel_scripts_model(tmp_path / "candidate.xlsx")
    corrupt_python_in_excel_scripts_package(candidate)

    candidate_snapshot = load_snapshot(candidate)
    assert candidate_snapshot.python_in_excel.to_dict() == {
        "present": True,
        "python_part_count": 1,
        "python_formula_cell_count": 1,
        "python_function_count": 1,
        "python_script_count": 0,
        "python_environment_definition_count": 0,
        "python_initialization_count": 0,
        "unrecognized_python_in_excel_count": 2,
    }
    assert any(
        "Python-in-Excel" in warning for warning in candidate_snapshot.parser_warnings
    )

    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    assert "FF065" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(candidate_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-PYTHON-SCRIPTS-BASELINE",
        "PRIVATE-PYTHON-SCRIPTS-SECOND",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_mismatched_python_in_excel_python_scripts_contract_is_coverage_evidence(
    tmp_path,
) -> None:
    baseline = make_python_in_excel_scripts_model(tmp_path / "baseline.xlsx")
    candidate = make_python_in_excel_scripts_model(tmp_path / "candidate.xlsx")
    mismatch_python_in_excel_scripts_relationship_contract(candidate)

    candidate_snapshot = load_snapshot(candidate)
    assert candidate_snapshot.python_in_excel.python_part_count == 1
    assert candidate_snapshot.python_in_excel.python_script_count == 2
    assert candidate_snapshot.python_in_excel.unrecognized_python_in_excel_count > 0
    assert any(
        "Python-in-Excel" in warning for warning in candidate_snapshot.parser_warnings
    )

    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    assert "FF065" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(candidate_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-PYTHON-SCRIPTS-BASELINE",
        "PRIVATE-PYTHON-SCRIPTS-SECOND",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_package_wide_external_relationship_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_external_relationship_model(tmp_path / "baseline.xlsx")
    candidate = tmp_path / "candidate.xlsx"
    shutil.copyfile(baseline, candidate)
    renumber_external_relationship_identifiers(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert baseline_snapshot.external_relationships == candidate_snapshot.external_relationships
    assert "external_relationships_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF063" not in {finding.rule_id for finding in report.findings}


def test_unrecognized_package_relationship_metadata_fails_closed_and_stays_private(
    tmp_path,
) -> None:
    baseline = make_external_relationship_model(tmp_path / "baseline.xlsx")
    candidate = make_external_relationship_model(tmp_path / "candidate.xlsx")
    corrupt_external_relationship_metadata(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    relationship_change = next(
        change
        for change in report.changes
        if change.kind == "external_relationships_changed"
    )

    assert candidate_snapshot.external_relationships.unrecognized_relationship_count == 1
    assert any(
        "malformed or unsupported OPC relationship metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert relationship_change.details[
        "unrecognized_external_relationship_metadata_changed"
    ] is True
    assert "FF063" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-OPAQUE-ATTRIBUTE",
        "privateUnknownAttribute",
        "PRIVATE-PACKAGE-HYPERLINK-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_package_wide_external_relationship_scan_budget_fails_closed(tmp_path, monkeypatch) -> None:
    baseline = make_external_relationship_model(tmp_path / "baseline.xlsx")
    candidate = make_external_relationship_model(tmp_path / "candidate.xlsx")
    baseline_snapshot = load_snapshot(baseline)
    monkeypatch.setattr(workbook_module, "_EXTERNAL_RELATIONSHIP_MAX_XML_BYTES", 1)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert candidate_snapshot.external_relationships.present is True
    assert candidate_snapshot.external_relationships.unrecognized_relationship_count > 0
    assert any(
        "oversized package relationship XML part" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "FF063" in {finding.rule_id for finding in report.findings}


def test_external_link_packages_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_external_link_package_model(tmp_path / "baseline.xlsx")
    candidate = make_external_link_package_model(tmp_path / "candidate.xlsx")
    change_external_link_package_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    assert baseline_snapshot.summary()["external_link_package_count"] == 3
    assert baseline_snapshot.summary()["external_workbook_link_count"] == 1
    assert baseline_snapshot.summary()["dde_link_count"] == 1
    assert baseline_snapshot.summary()["ole_link_count"] == 1
    assert profile["external_link_packages"] == {
        "present": True,
        "external_link_count": 3,
        "external_workbook_count": 1,
        "dde_link_count": 1,
        "ole_link_count": 1,
        "unrecognized_link_count": 0,
        "external_workbook_sheet_count": 2,
        "external_defined_name_count": 1,
        "external_workbook_cached_sheet_count": 1,
        "external_workbook_cached_cell_count": 2,
        "external_workbook_cached_refresh_error_count": 1,
        "dde_item_count": 2,
        "dde_advise_item_count": 1,
        "dde_ole_item_count": 1,
        "dde_prefer_picture_item_count": 1,
        "dde_cached_value_count": 2,
        "ole_item_count": 2,
        "ole_advise_item_count": 1,
        "ole_icon_item_count": 1,
        "ole_prefer_picture_item_count": 1,
        "opaque_metadata": {"present": True, "count": 1},
    }
    assert "## External-link packages" in markdown
    assert "**Package parts:** 3 (1 workbook, 1 DDE, 1 OLE)" in markdown

    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    external_link_change = next(
        change
        for change in report.changes
        if change.kind == "external_link_packages_changed"
    )

    assert external_link_change.details["source_material_changed"] is True
    assert external_link_change.details["definition_material_changed"] is True
    assert external_link_change.details["cached_material_changed"] is True
    assert external_link_change.details["opaque_metadata_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF025"}

    sensitive_values = (
        "file:///private/baseline/external-workbook.xlsx",
        "file:///private/candidate/external-workbook.xlsx",
        "private external baseline sheet",
        "private external scenario",
        "private external baseline defined name",
        "private external baseline cached value",
        "private external candidate cached value",
        "private-baseline-dde-service",
        "private-baseline-dde-topic",
        "private-candidate-dde-topic",
        "private-baseline-dde-item",
        "private-candidate-dde-item",
        "private-baseline-dde-value",
        "private-candidate-dde-value",
        "private.baseline.ole.program",
        "private.candidate.ole.program",
        "private-baseline-ole-item",
        "private-candidate-ole-item",
        "private baseline external-link extension payload",
        "private candidate external-link extension payload",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_external_link_declaration_rebinding_is_a_source_change(tmp_path) -> None:
    baseline = make_external_link_package_model(tmp_path / "baseline.xlsx")
    candidate = make_external_link_package_model(tmp_path / "candidate.xlsx")
    rebind_external_link_declaration(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    external_link_change = next(
        change
        for change in report.changes
        if change.kind == "external_link_packages_changed"
    )

    assert external_link_change.details["source_material_changed"] is True
    assert "definition_material_changed" not in external_link_change.details
    assert "cached_material_changed" not in external_link_change.details
    assert {finding.rule_id for finding in report.findings} >= {"FF025"}


def test_external_link_relationship_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_external_link_package_model(tmp_path / "baseline.xlsx")
    candidate = make_external_link_package_model(tmp_path / "candidate.xlsx")
    renumber_external_link_declaration_relationships(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "external_link_packages_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF025" not in {finding.rule_id for finding in report.findings}


def test_ambiguous_external_link_definitions_fail_closed(tmp_path) -> None:
    baseline = make_external_link_package_model(tmp_path / "baseline.xlsx")
    candidate = make_external_link_package_model(tmp_path / "candidate.xlsx")
    duplicate_external_link_definition(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.external_link_packages.unrecognized_link_count == 1
    assert any(
        "without exactly one supported link definition" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "external_link_packages_changed" in {change.kind for change in report.changes}
    assert "FF025" in {finding.rule_id for finding in report.findings}


def test_repeated_external_link_children_are_retained_as_opaque_material(tmp_path) -> None:
    baseline = make_external_link_package_model(tmp_path / "baseline.xlsx")
    candidate = make_external_link_package_model(tmp_path / "candidate.xlsx")
    duplicate_external_link_sheet_names(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    external_link_change = next(
        change
        for change in report.changes
        if change.kind == "external_link_packages_changed"
    )

    assert any(
        "repeated sheetNames containers" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert external_link_change.details["opaque_metadata_changed"] is True
    assert "FF025" in {finding.rule_id for finding in report.findings}


def test_xlm_macro_sheets_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    change_xlm_macro_sheet_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    assert baseline_snapshot.sheets["Macro Automation"].formula_cells == 0
    assert baseline_snapshot.summary()["has_xlm_macro_sheets"] is True
    assert baseline_snapshot.summary()["xlm_macro_sheet_count"] == 1
    assert baseline_snapshot.summary()["xlm_macro_formula_cell_count"] == 2
    assert profile["xlm_macro_sheets"] == {
        "present": True,
        "declared_macro_sheet_count": 1,
        "macro_sheet_count": 1,
        "international_macro_sheet_count": 0,
        "unrecognized_macro_sheet_count": 0,
        "hidden_macro_sheet_count": 0,
        "very_hidden_macro_sheet_count": 1,
        "formula_cell_count": 2,
        "related_relationship_count": 3,
        "external_relationship_count": 1,
        "internal_related_part_count": 2,
        "fingerprinted_related_part_count": 2,
        "uninspected_related_part_count": 0,
        "embedded_object_relationship_count": 2,
        "embedded_package_relationship_count": 1,
    }
    assert "## Excel 4.0 / XLM macro sheets" in markdown
    assert "**Macro formula cells:** 2" in markdown

    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    macro_sheet_change = next(
        change for change in report.changes if change.kind == "xlm_macro_sheets_changed"
    )

    assert macro_sheet_change.details["workbook_binding_changed"] is True
    assert macro_sheet_change.details["macro_program_material_changed"] is True
    assert macro_sheet_change.details["related_part_relationships_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF026"}

    sensitive_values = (
        "private-baseline-xl-command-argument",
        "private-candidate-xl-command-argument",
        "private-baseline-xl-cell-value",
        "private-candidate-xl-cell-value",
        "private.baseline.xlm.embedded.object",
        "private.candidate.xlm.embedded.object",
        "file:///private/baseline-xl-linked-object.bin",
        "file:///private/candidate-xl-linked-object.bin",
        "private-baseline-xl-object.bin",
        "private-candidate-xl-object.bin",
        "private candidate XLM extension payload",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_xlm_automatic_macro_bindings_are_profiled_and_diffed_privately(
    tmp_path,
) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    add_xlm_automatic_macro_binding(
        baseline,
        name="_xlnm.Auto_Open",
        target="'Macro Automation'!$A$1",
    )
    add_xlm_automatic_macro_binding(
        candidate,
        name="_xlnm.Auto_Open",
        target="'Macro Automation'!$A$2",
    )

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    assert baseline_snapshot.summary()["has_xlm_automatic_macro_bindings"] is True
    assert baseline_snapshot.summary()["xlm_automatic_macro_binding_count"] == 1
    assert profile["features"]["has_xlm_automatic_macro_bindings"] is True
    assert profile["xlm_automatic_macro_bindings"] == {
        "present": True,
        "automatic_macro_binding_count": 1,
        "auto_open_binding_count": 1,
        "auto_close_binding_count": 0,
        "auto_activate_binding_count": 0,
        "auto_deactivate_binding_count": 0,
    }
    assert "## XLM automatic-macro bindings" in markdown
    assert "**Workbook bindings:** 1" in markdown

    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    automatic_macro_change = next(
        change
        for change in report.changes
        if change.kind == "xlm_automatic_macro_bindings_changed"
    )
    automatic_macro_sarif = next(
        result
        for result in report_to_sarif(report)["runs"][0]["results"]
        if result["ruleId"] == "FF076"
    )

    assert "xlm_macro_sheets_changed" not in {change.kind for change in report.changes}
    assert {finding.rule_id for finding in report.findings} >= {"FF008", "FF076"}
    assert automatic_macro_change.details == {
        "before": profile["xlm_automatic_macro_bindings"],
        "after": {
            "present": True,
            "automatic_macro_binding_count": 1,
            "auto_open_binding_count": 1,
            "auto_close_binding_count": 0,
            "auto_activate_binding_count": 0,
            "auto_deactivate_binding_count": 0,
        },
        "automatic_macro_binding_material_changed": True,
    }

    dedicated_artifacts = (
        json.dumps(profile["xlm_automatic_macro_bindings"]),
        markdown,
        json.dumps(automatic_macro_change.to_dict()),
        json.dumps(automatic_macro_sarif),
    )
    for sensitive_value in (
        "_xlnm.Auto_Open",
        "'Macro Automation'!$A$1",
        "'Macro Automation'!$A$2",
    ):
        assert all(sensitive_value not in artifact for artifact in dedicated_artifacts)


def test_xlm_automatic_macro_bindings_cover_all_documented_events(tmp_path) -> None:
    workbook = make_xlm_macro_sheet_model(tmp_path / "automatic-events.xlsm")
    for name in (
        "Auto_Open",
        "Auto_Close",
        "_xlnm.Auto_Activate",
        "AUTO_DEACTIVATE",
    ):
        add_xlm_automatic_macro_binding(workbook, name=name)

    bindings = load_snapshot(workbook).xlm_automatic_macro_bindings

    assert bindings.automatic_macro_binding_count == 4
    assert bindings.auto_open_binding_count == 1
    assert bindings.auto_close_binding_count == 1
    assert bindings.auto_activate_binding_count == 1
    assert bindings.auto_deactivate_binding_count == 1


def test_xlm_automatic_macro_bindings_require_workbook_scope_and_macro_target(
    tmp_path,
) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    add_xlm_automatic_macro_binding(
        candidate,
        name="Auto_Open",
        target="Inputs!$A$1",
    )
    add_xlm_automatic_macro_binding(
        candidate,
        name="Auto_Close",
        local_sheet_id=0,
    )
    add_xlm_automatic_macro_binding(
        candidate,
        name="Auto_Activate",
        target="'Macro Automation'!$A$1:$A$2",
    )

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.xlm_automatic_macro_bindings.present is False
    assert "FF076" not in {finding.rule_id for finding in report.findings}


def test_xlm_automatic_macro_bindings_fail_closed_for_ambiguous_sheet_binding(
    tmp_path,
) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    add_xlm_automatic_macro_binding(baseline)
    add_xlm_automatic_macro_binding(candidate)
    duplicate_xlm_macro_sheet_workbook_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.xlm_automatic_macro_bindings.present is False
    assert any(
        "duplicate XLM macro-sheet workbook relationship ids while inspecting "
        "automatic macro bindings" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "FF076" in {finding.rule_id for finding in report.findings}


def test_international_xlm_macro_sheet_parts_are_detected(tmp_path) -> None:
    workbook = make_xlm_macro_sheet_model(
        tmp_path / "international.xlsm", international=True
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.xlm_macro_sheets.international_macro_sheet_count == 1
    assert snapshot.xlm_macro_sheets.formula_cell_count == 2
    assert snapshot.parser_warnings == ()


def test_xlm_macro_sheet_relationship_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    renumber_xlm_macro_sheet_relationships(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "xlm_macro_sheets_changed" not in {change.kind for change in report.changes}
    assert "FF026" not in {finding.rule_id for finding in report.findings}


def test_xlm_macro_sheet_related_part_payloads_are_guarded_privately(tmp_path) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    change_xlm_macro_sheet_related_part_payload(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    macro_sheet_change = next(
        change for change in report.changes if change.kind == "xlm_macro_sheets_changed"
    )

    assert profile["workbook"]["xlm_related_part_payload_count"] == 2
    assert profile["xlm_macro_sheets"]["internal_related_part_count"] == 2
    assert profile["xlm_macro_sheets"]["fingerprinted_related_part_count"] == 2
    assert profile["xlm_macro_sheets"]["uninspected_related_part_count"] == 0
    assert macro_sheet_change.details["related_part_payload_material_changed"] is True
    assert "workbook_binding_changed" not in macro_sheet_change.details
    assert "macro_program_material_changed" not in macro_sheet_change.details
    assert "related_part_relationships_changed" not in macro_sheet_change.details
    assert {finding.rule_id for finding in report.findings} >= {"FF026"}

    sensitive_values = (
        "private baseline embedded XLM object payload",
        "private candidate XLM related-part payload only",
        "private-baseline-xl-object.bin",
    )
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_missing_xlm_related_part_payloads_fail_closed(tmp_path) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    remove_xlm_macro_sheet_related_part_payload(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.xlm_macro_sheets.internal_related_part_count == 2
    assert candidate_snapshot.xlm_macro_sheets.fingerprinted_related_part_count == 1
    assert candidate_snapshot.xlm_macro_sheets.uninspected_related_part_count == 1
    assert any(
        "could not locate an XLM macro-sheet internal related part" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "xlm_macro_sheets_changed" in {change.kind for change in report.changes}
    assert "FF026" in {finding.rule_id for finding in report.findings}


def test_oversized_xlm_related_part_payloads_remain_covered(tmp_path, monkeypatch) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    change_xlm_macro_sheet_related_part_payload(candidate)
    monkeypatch.setattr("formulafence.workbook._XLM_RELATED_PART_MAX_BYTES", 1)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    macro_sheet_change = next(
        change for change in report.changes if change.kind == "xlm_macro_sheets_changed"
    )

    assert baseline_snapshot.xlm_macro_sheets.fingerprinted_related_part_count == 0
    assert baseline_snapshot.xlm_macro_sheets.uninspected_related_part_count == 2
    assert any(
        "oversized XLM macro-sheet related part" in warning
        for warning in baseline_snapshot.parser_warnings
    )
    assert macro_sheet_change.details["related_part_payload_material_changed"] is True
    assert "FF026" in {finding.rule_id for finding in report.findings}


def test_xlm_related_part_count_budget_remains_covered(tmp_path, monkeypatch) -> None:
    workbook = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    budget_type = workbook_module._XlmRelatedPartBudget
    monkeypatch.setattr(
        workbook_module,
        "_XlmRelatedPartBudget",
        lambda: budget_type(remaining_parts=1),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.xlm_macro_sheets.internal_related_part_count == 2
    assert snapshot.xlm_macro_sheets.fingerprinted_related_part_count == 1
    assert snapshot.xlm_macro_sheets.uninspected_related_part_count == 1
    assert any(
        "XLM macro-sheet related-part count budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_xlm_related_part_byte_budget_remains_covered(tmp_path, monkeypatch) -> None:
    workbook = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    budget_type = workbook_module._XlmRelatedPartBudget
    monkeypatch.setattr(
        workbook_module,
        "_XlmRelatedPartBudget",
        lambda: budget_type(remaining_bytes=1),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.xlm_macro_sheets.internal_related_part_count == 2
    assert snapshot.xlm_macro_sheets.fingerprinted_related_part_count == 0
    assert snapshot.xlm_macro_sheets.uninspected_related_part_count == 2
    assert any(
        "XLM macro-sheet related-part read budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_xlm_macro_sheet_equivalent_internal_target_spellings_are_ignored(
    tmp_path,
) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    rewrite_xlm_macro_sheet_internal_target_spelling(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "xlm_macro_sheets_changed" not in {change.kind for change in report.changes}
    assert "FF026" not in {finding.rule_id for finding in report.findings}


def test_malformed_xlm_macro_sheet_parts_fail_closed(tmp_path) -> None:
    baseline = make_xlm_macro_sheet_model(tmp_path / "baseline.xlsm")
    candidate = make_xlm_macro_sheet_model(tmp_path / "candidate.xlsm")
    corrupt_xlm_macro_sheet_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.xlm_macro_sheets.unrecognized_macro_sheet_count == 1
    assert any(
        "XLM macro-sheet part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "xlm_macro_sheets_changed" in {change.kind for change in report.changes}
    assert "FF026" in {finding.rule_id for finding in report.findings}


def _append_ribbon_customization_xml_elements(
    path,
    count: int,
    *,
    nested: bool = False,
) -> None:
    """Add opaque RibbonX XML entries without relying on a workbook reader."""
    member = "customUI/customUI.xml"
    staging = path.with_suffix(".ribbon-xml-elements.xlsx")
    with ZipFile(path) as archive:
        contents = {
            entry.filename: archive.read(entry)
            for entry in archive.infolist()
        }
    root = ElementTree.fromstring(contents[member])
    namespace = root.tag.partition("}")[0].removeprefix("{")
    parent = root
    if nested:
        parent = ElementTree.SubElement(root, f"{{{namespace}}}opaque")
    for _ in range(count):
        ElementTree.SubElement(parent, f"{{{namespace}}}entry")
    contents[member] = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
    with ZipFile(staging, "w", compression=ZIP_DEFLATED) as archive:
        for name, payload in contents.items():
            archive.writestr(name, payload)
    staging.replace(path)


def _ribbon_customization_xml_element_count(path) -> int:
    """Count one fixture's RibbonX elements for exact budget tests."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("customUI/customUI.xml"))
    return sum(1 for _ in root.iter())


def test_ribbon_callback_changes_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_ribbon_customization_model(tmp_path / "baseline.xlsx")
    candidate = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    change_ribbon_customization_callback(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    ribbon_change = next(
        change for change in report.changes if change.kind == "ribbon_customization_changed"
    )

    assert baseline_snapshot.summary()["has_ribbon_customization"] is True
    assert baseline_snapshot.summary()["ribbon_customization_part_count"] == 1
    assert baseline_snapshot.summary()["ribbon_callback_attribute_count"] == 2
    assert profile["ribbon_customization"] == {
        "present": True,
        "declared_ribbon_part_count": 1,
        "ribbon_part_count": 1,
        "office_2010_ribbon_part_count": 0,
        "unrecognized_ribbon_part_count": 0,
        "control_count": 3,
        "callback_attribute_count": 2,
        "action_callback_count": 1,
        "image_relationship_count": 1,
        "external_relationship_count": 0,
    }
    assert "## Office RibbonX customization" in markdown
    assert "**Callback attributes:** 2 (1 onAction)" in markdown
    assert ribbon_change.details["ribbon_definition_material_changed"] is True
    assert "package_binding_changed" not in ribbon_change.details
    assert "image_relationships_changed" not in ribbon_change.details
    assert {finding.rule_id for finding in report.findings} >= {"FF027"}

    sensitive_values = (
        "PrivateBaselineRibbonAction",
        "PrivateCandidateRibbonAction",
        "PrivateBaselineRibbonLoad",
        "private baseline ribbon action",
        "private-baseline-ribbon.png",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_ribbon_customization_relationships_are_guarded_privately(tmp_path) -> None:
    baseline = make_ribbon_customization_model(tmp_path / "baseline.xlsx")
    candidate = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    change_ribbon_customization_controls(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    ribbon_change = next(
        change for change in report.changes if change.kind == "ribbon_customization_changed"
    )

    assert ribbon_change.details["ribbon_definition_material_changed"] is True
    assert ribbon_change.details["image_relationships_changed"] is True
    assert "package_binding_changed" not in ribbon_change.details
    assert {finding.rule_id for finding in report.findings} >= {"FF027"}


def test_office_2010_ribbon_customization_parts_are_detected(tmp_path) -> None:
    workbook = make_ribbon_customization_model(
        tmp_path / "office-2010-ribbon.xlsx", office_2010=True
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.ribbon_customization.office_2010_ribbon_part_count == 1
    assert snapshot.ribbon_customization.callback_attribute_count == 2
    assert snapshot.parser_warnings == ()


def test_office_2010_compatibility_ribbon_namespace_is_detected(tmp_path) -> None:
    workbook = make_ribbon_customization_model(
        tmp_path / "office-2010-compatibility-ribbon.xlsx",
        office_2010=True,
        compatibility_namespace=True,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.ribbon_customization.office_2010_ribbon_part_count == 1
    assert snapshot.ribbon_customization.unrecognized_ribbon_part_count == 0
    assert snapshot.parser_warnings == ()


def test_ribbon_customization_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_ribbon_customization_model(tmp_path / "baseline.xlsx")
    candidate = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    renumber_ribbon_customization_relationships(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "ribbon_customization_changed" not in {change.kind for change in report.changes}
    assert "FF027" not in {finding.rule_id for finding in report.findings}


def test_ribbon_customization_equivalent_target_spellings_are_ignored(tmp_path) -> None:
    baseline = make_ribbon_customization_model(tmp_path / "baseline.xlsx")
    candidate = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    rewrite_ribbon_customization_internal_target_spelling(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "ribbon_customization_changed" not in {change.kind for change in report.changes}
    assert "FF027" not in {finding.rule_id for finding in report.findings}


def test_oversized_ribbon_customization_parts_remain_covered(
    tmp_path, monkeypatch
) -> None:
    baseline = make_ribbon_customization_model(tmp_path / "baseline.xlsx")
    candidate = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    change_ribbon_customization_callback(candidate)
    monkeypatch.setattr("formulafence.workbook._RIBBON_CUSTOM_UI_MAX_PART_BYTES", 1)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.ribbon_customization.unrecognized_ribbon_part_count == 1
    assert any(
        "oversized RibbonX customization part" in warning
        for warning in baseline_snapshot.parser_warnings
    )
    assert "FF027" in {finding.rule_id for finding in report.findings}


def test_ribbon_customization_part_count_budget_remains_covered(
    tmp_path, monkeypatch
) -> None:
    workbook = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._RibbonCustomizationBudget
    monkeypatch.setattr(
        workbook_module,
        "_RibbonCustomizationBudget",
        lambda: budget_type(remaining_parts=0),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.ribbon_customization.unrecognized_ribbon_part_count == 1
    assert any(
        "RibbonX customization part count budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_ribbon_customization_byte_budget_remains_covered(tmp_path, monkeypatch) -> None:
    workbook = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._RibbonCustomizationBudget
    monkeypatch.setattr(
        workbook_module,
        "_RibbonCustomizationBudget",
        lambda: budget_type(remaining_bytes=1),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.ribbon_customization.unrecognized_ribbon_part_count == 1
    assert any(
        "RibbonX customization part read budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_ribbon_xml_element_budget_stops_before_materializing_the_tree(
    tmp_path, monkeypatch
) -> None:
    workbook = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    _append_ribbon_customization_xml_elements(workbook, 1)
    monkeypatch.setattr(
        workbook_module,
        "_RIBBON_CUSTOM_UI_MAX_XML_ELEMENT_COUNT",
        _ribbon_customization_xml_element_count(workbook) - 1,
    )

    def unexpected_tree_parse(*args, **kwargs):
        raise AssertionError("the RibbonX XML tree was materialized after its budget failed")

    monkeypatch.setattr(workbook_module, "_xml_root_from_payload", unexpected_tree_parse)
    warnings: set[str] = set()
    with ZipFile(workbook) as archive:
        inspection = workbook_module._ribbon_part_inspection(
            archive,
            "customUI/customUI.xml",
            warnings,
            workbook_module._RibbonCustomizationBudget(),
        )

    assert inspection.inspected is False
    assert any("RibbonX customization XML part whose structure" in warning for warning in warnings)


def test_ribbon_xml_element_budget_remains_covered(tmp_path, monkeypatch) -> None:
    workbook = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    _append_ribbon_customization_xml_elements(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_RIBBON_CUSTOM_UI_MAX_XML_ELEMENT_COUNT",
        _ribbon_customization_xml_element_count(workbook) - 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.ribbon_customization.unrecognized_ribbon_part_count == 1
    assert any(
        "RibbonX customization XML part whose structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_ribbon_xml_element_budget_accepts_the_configured_capacity(
    tmp_path, monkeypatch
) -> None:
    workbook = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    _append_ribbon_customization_xml_elements(workbook, 2)
    monkeypatch.setattr(
        workbook_module,
        "_RIBBON_CUSTOM_UI_MAX_XML_ELEMENT_COUNT",
        _ribbon_customization_xml_element_count(workbook),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.ribbon_customization.unrecognized_ribbon_part_count == 0


def test_ribbon_xml_element_budget_accepts_the_default_capacity(tmp_path) -> None:
    workbook = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    _append_ribbon_customization_xml_elements(
        workbook,
        workbook_module._RIBBON_CUSTOM_UI_MAX_XML_ELEMENT_COUNT
        - _ribbon_customization_xml_element_count(workbook),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.ribbon_customization.unrecognized_ribbon_part_count == 0


def test_ribbon_xml_element_budget_rejects_the_default_overage(tmp_path) -> None:
    workbook = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    _append_ribbon_customization_xml_elements(
        workbook,
        workbook_module._RIBBON_CUSTOM_UI_MAX_XML_ELEMENT_COUNT
        - _ribbon_customization_xml_element_count(workbook)
        + 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.ribbon_customization.unrecognized_ribbon_part_count == 1
    assert any(
        "RibbonX customization XML part whose structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_ribbon_xml_element_budget_counts_nested_opaque_subtrees(tmp_path, monkeypatch) -> None:
    workbook = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    _append_ribbon_customization_xml_elements(workbook, 1, nested=True)
    monkeypatch.setattr(
        workbook_module,
        "_RIBBON_CUSTOM_UI_MAX_XML_ELEMENT_COUNT",
        _ribbon_customization_xml_element_count(workbook) - 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.ribbon_customization.unrecognized_ribbon_part_count == 1
    assert any(
        "RibbonX customization XML part whose structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_malformed_ribbon_customization_parts_fail_closed(tmp_path) -> None:
    baseline = make_ribbon_customization_model(tmp_path / "baseline.xlsx")
    candidate = make_ribbon_customization_model(tmp_path / "candidate.xlsx")
    corrupt_ribbon_customization_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.ribbon_customization.unrecognized_ribbon_part_count == 1
    assert any(
        "RibbonX customization part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "ribbon_customization_changed" in {change.kind for change in report.changes}
    assert "FF027" in {finding.rule_id for finding in report.findings}


def _append_office_web_addin_xml_elements(
    path,
    member: str,
    count: int,
    *,
    nested: bool = False,
) -> None:
    """Add opaque Office Web Add-in XML entries without using a workbook reader."""
    staging = path.with_suffix(".office-web-addin-xml-elements.xlsx")
    with ZipFile(path) as archive:
        contents = {
            entry.filename: archive.read(entry)
            for entry in archive.infolist()
        }
    root = ElementTree.fromstring(contents[member])
    namespace = root.tag.partition("}")[0].removeprefix("{")
    parent = root
    if nested:
        parent = ElementTree.SubElement(root, f"{{{namespace}}}opaque")
    for _ in range(count):
        ElementTree.SubElement(parent, f"{{{namespace}}}entry")
    contents[member] = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
    with ZipFile(staging, "w", compression=ZIP_DEFLATED) as archive:
        for name, payload in contents.items():
            archive.writestr(name, payload)
    staging.replace(path)


def _office_web_addin_xml_element_count(path, member: str) -> int:
    """Count one fixture's Office Web Add-in XML elements for exact budget tests."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member))
    return sum(1 for _ in root.iter())


def test_office_web_addin_auto_show_changes_are_profiled_and_diffed_privately(
    tmp_path,
) -> None:
    baseline = make_office_web_addin_model(tmp_path / "baseline.xlsx")
    candidate = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    change_office_web_addin_auto_show(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    addin_change = next(
        change for change in report.changes if change.kind == "office_web_addins_changed"
    )

    assert baseline_snapshot.summary()["has_office_web_addins"] is True
    assert baseline_snapshot.summary()["office_web_addin_taskpane_part_count"] == 1
    assert baseline_snapshot.summary()["office_web_addin_auto_show_taskpane_count"] == 1
    assert profile["office_web_addins"] == {
        "present": True,
        "declared_taskpane_part_count": 1,
        "taskpane_part_count": 1,
        "web_extension_part_count": 1,
        "unrecognized_part_count": 0,
        "taskpane_count": 1,
        "visible_taskpane_count": 1,
        "locked_taskpane_count": 1,
        "web_extension_reference_count": 1,
        "auto_show_taskpane_count": 1,
        "store_reference_count": 1,
        "alternate_reference_count": 1,
        "binding_count": 1,
        "snapshot_reference_count": 1,
        "related_relationship_count": 3,
        "external_relationship_count": 1,
        "worksheet_binding_sheet_count": 0,
        "worksheet_binding_count": 0,
        "in_content_drawing_part_count": 0,
        "in_content_web_extension_reference_count": 0,
        "in_content_web_extension_part_count": 0,
    }
    assert "## Office Web Add-ins" in markdown
    assert "**Auto-show task-pane requests:** 1" in markdown
    assert addin_change.details["web_extension_definition_material_changed"] is True
    assert "workbook_binding_changed" not in addin_change.details
    assert "taskpane_configuration_material_changed" not in addin_change.details
    assert "related_part_relationships_changed" not in addin_change.details
    assert {finding.rule_id for finding in report.findings} >= {"FF028"}

    sensitive_values = (
        "PrivateBaselineAddin",
        "PrivateFallbackAddin",
        "private-baseline-manifest.xml",
        "private-fallback-manifest.xml",
        "PrivateBaselineBinding",
        "PrivateBaselineTable",
        "private baseline behavior",
        "private.example.invalid",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_office_web_addin_schema_reference_element_is_detected(tmp_path) -> None:
    workbook = make_office_web_addin_model(
        tmp_path / "schema-reference.xlsx",
        taskpane_reference_element="webextensionref",
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.office_web_addins.web_extension_reference_count == 1
    assert snapshot.office_web_addins.alternate_reference_count == 1
    assert snapshot.parser_warnings == ()


def test_office_web_addin_package_addition_is_critical(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_office_web_addin_model(tmp_path / "candidate.xlsx")

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    addin_change = next(
        change for change in report.changes if change.kind == "office_web_addins_changed"
    )

    assert addin_change.details["before"]["present"] is False
    assert addin_change.details["after"]["present"] is True
    assert addin_change.details["workbook_binding_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF028"}


def test_office_web_addin_taskpane_and_relationship_changes_are_guarded_privately(
    tmp_path,
) -> None:
    baseline = make_office_web_addin_model(tmp_path / "baseline.xlsx")
    candidate = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    change_office_web_addin_controls(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    addin_change = next(
        change for change in report.changes if change.kind == "office_web_addins_changed"
    )

    assert addin_change.details["taskpane_configuration_material_changed"] is True
    assert addin_change.details["related_part_relationships_changed"] is True
    assert addin_change.details["web_extension_definition_material_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF028"}


def test_office_web_addin_relationship_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_office_web_addin_model(tmp_path / "baseline.xlsx")
    candidate = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    renumber_office_web_addin_relationships(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "office_web_addins_changed" not in {change.kind for change in report.changes}
    assert "FF028" not in {finding.rule_id for finding in report.findings}


def test_office_web_addin_equivalent_target_spellings_are_ignored(tmp_path) -> None:
    baseline = make_office_web_addin_model(tmp_path / "baseline.xlsx")
    candidate = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    rewrite_office_web_addin_internal_target_spelling(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "office_web_addins_changed" not in {change.kind for change in report.changes}
    assert "FF028" not in {finding.rule_id for finding in report.findings}


def test_oversized_office_web_addin_parts_remain_covered(tmp_path, monkeypatch) -> None:
    baseline = make_office_web_addin_model(tmp_path / "baseline.xlsx")
    candidate = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    change_office_web_addin_auto_show(candidate)
    monkeypatch.setattr("formulafence.workbook._WEB_EXTENSION_MAX_PART_BYTES", 1)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "oversized Office Web Add-in package part" in warning
        for warning in baseline_snapshot.parser_warnings
    )
    assert "FF028" in {finding.rule_id for finding in report.findings}


def test_office_web_addin_part_count_budget_remains_covered(tmp_path, monkeypatch) -> None:
    workbook = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._OfficeWebAddinBudget
    monkeypatch.setattr(
        workbook_module,
        "_OfficeWebAddinBudget",
        lambda: budget_type(remaining_parts=0),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "Office Web Add-in part count budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_office_web_addin_byte_budget_remains_covered(tmp_path, monkeypatch) -> None:
    workbook = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._OfficeWebAddinBudget
    monkeypatch.setattr(
        workbook_module,
        "_OfficeWebAddinBudget",
        lambda: budget_type(remaining_bytes=1),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "Office Web Add-in part read budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_office_web_addin_xml_element_budget_stops_before_materializing_the_tree(
    tmp_path, monkeypatch
) -> None:
    parts = (
        (
            "xl/webextensions/taskpanes.xml",
            workbook_module._office_web_addin_taskpane_inspection,
        ),
        (
            "xl/webextensions/webextension1.xml",
            workbook_module._office_web_addin_extension_inspection,
        ),
    )

    def unexpected_tree_parse(*args, **kwargs):
        raise AssertionError(
            "the Office Web Add-in XML tree was materialized after its budget failed"
        )

    monkeypatch.setattr(workbook_module, "_xml_root_from_payload", unexpected_tree_parse)
    for member, inspection_function in parts:
        workbook = make_office_web_addin_model(
            tmp_path / f"{member.rsplit('/', 1)[-1]}.xlsx"
        )
        _append_office_web_addin_xml_elements(workbook, member, 1)
        monkeypatch.setattr(
            workbook_module,
            "_WEB_EXTENSION_MAX_XML_ELEMENT_COUNT",
            _office_web_addin_xml_element_count(workbook, member) - 1,
        )
        warnings: set[str] = set()
        with ZipFile(workbook) as archive:
            inspection = inspection_function(
                archive,
                member,
                warnings,
                workbook_module._OfficeWebAddinBudget(),
            )

        assert inspection.inspected is False
        assert any(
            "Office Web Add-in package part whose XML structure" in warning
            for warning in warnings
        )


def test_office_web_addin_xml_element_budget_remains_covered(tmp_path, monkeypatch) -> None:
    member = "xl/webextensions/taskpanes.xml"
    workbook = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    _append_office_web_addin_xml_elements(workbook, member, 2)
    monkeypatch.setattr(
        workbook_module,
        "_WEB_EXTENSION_MAX_XML_ELEMENT_COUNT",
        _office_web_addin_xml_element_count(workbook, member) - 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "Office Web Add-in package part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_office_web_addin_xml_element_budget_accepts_the_configured_capacity(
    tmp_path, monkeypatch
) -> None:
    member = "xl/webextensions/webextension1.xml"
    workbook = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    _append_office_web_addin_xml_elements(workbook, member, 2)
    monkeypatch.setattr(
        workbook_module,
        "_WEB_EXTENSION_MAX_XML_ELEMENT_COUNT",
        _office_web_addin_xml_element_count(workbook, member),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.office_web_addins.unrecognized_part_count == 0


def test_office_web_addin_xml_element_budget_aggregates_across_package_parts(
    tmp_path, monkeypatch
) -> None:
    taskpane_member = "xl/webextensions/taskpanes.xml"
    extension_member = "xl/webextensions/webextension1.xml"
    workbook = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    remaining_xml_elements = (
        _office_web_addin_xml_element_count(workbook, taskpane_member)
        + _office_web_addin_xml_element_count(workbook, extension_member)
        - 1
    )
    budget_type = workbook_module._OfficeWebAddinBudget
    monkeypatch.setattr(
        workbook_module,
        "_OfficeWebAddinBudget",
        lambda: budget_type(remaining_xml_elements=remaining_xml_elements),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "Office Web Add-in package part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_office_web_addin_xml_element_budget_accepts_the_default_capacity(tmp_path) -> None:
    member = "xl/webextensions/webextension1.xml"
    workbook = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    _append_office_web_addin_xml_elements(
        workbook,
        member,
        workbook_module._WEB_EXTENSION_MAX_XML_ELEMENT_COUNT
        - _office_web_addin_xml_element_count(workbook, member),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.office_web_addins.unrecognized_part_count == 0


def test_office_web_addin_xml_element_budget_rejects_the_default_overage(tmp_path) -> None:
    member = "xl/webextensions/taskpanes.xml"
    workbook = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    _append_office_web_addin_xml_elements(
        workbook,
        member,
        workbook_module._WEB_EXTENSION_MAX_XML_ELEMENT_COUNT
        - _office_web_addin_xml_element_count(workbook, member)
        + 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "Office Web Add-in package part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_office_web_addin_xml_element_budget_counts_nested_opaque_subtrees(
    tmp_path, monkeypatch
) -> None:
    member = "xl/webextensions/webextension1.xml"
    workbook = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    _append_office_web_addin_xml_elements(workbook, member, 1, nested=True)
    monkeypatch.setattr(
        workbook_module,
        "_WEB_EXTENSION_MAX_XML_ELEMENT_COUNT",
        _office_web_addin_xml_element_count(workbook, member) - 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "Office Web Add-in package part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_malformed_office_web_addin_parts_fail_closed(tmp_path) -> None:
    baseline = make_office_web_addin_model(tmp_path / "baseline.xlsx")
    candidate = make_office_web_addin_model(tmp_path / "candidate.xlsx")
    corrupt_office_web_addin_definition_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "Office Web Add-in definition part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "office_web_addins_changed" in {change.kind for change in report.changes}
    assert "FF028" in {finding.rule_id for finding in report.findings}


def test_worksheet_office_web_addin_bindings_are_diffed_privately(tmp_path) -> None:
    baseline = make_office_web_addin_model(
        tmp_path / "baseline.xlsx",
        worksheet_binding=True,
    )
    candidate = make_office_web_addin_model(
        tmp_path / "candidate.xlsx",
        worksheet_binding=True,
    )
    change_office_web_addin_worksheet_binding(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    addin_change = next(
        change for change in report.changes if change.kind == "office_web_addins_changed"
    )

    assert baseline_snapshot.cells == candidate_snapshot.cells
    assert baseline_snapshot.office_web_addins.worksheet_binding_sheet_count == 1
    assert baseline_snapshot.office_web_addins.worksheet_binding_count == 1
    assert baseline_snapshot.office_web_addins.unrecognized_part_count == 0
    assert baseline_snapshot.parser_warnings == ()
    assert profile["office_web_addins"]["worksheet_binding_sheet_count"] == 1
    assert profile["office_web_addins"]["worksheet_binding_count"] == 1
    assert "**Worksheet binding sheets / bindings:** 1 / 1" in markdown
    assert addin_change.details["worksheet_binding_material_changed"] is True
    assert "in_content_drawing_binding_changed" not in addin_change.details
    assert "FF028" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PrivateBaselineTable",
        "Inputs!$B$2:$B$4",
        "Inputs!$B$2:$B$3",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_malformed_or_oversized_worksheet_addin_bindings_fail_closed(
    tmp_path,
    monkeypatch,
) -> None:
    baseline = make_office_web_addin_model(
        tmp_path / "baseline.xlsx",
        worksheet_binding=True,
    )
    malformed = make_office_web_addin_model(
        tmp_path / "malformed.xlsx",
        worksheet_binding=True,
    )
    corrupt_office_web_addin_worksheet_binding(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "malformed or unsupported worksheet Office Web Add-in metadata" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert "FF028" in {finding.rule_id for finding in malformed_report.findings}

    oversized = make_office_web_addin_model(
        tmp_path / "oversized.xlsx",
        worksheet_binding=True,
    )
    monkeypatch.setattr(
        workbook_module,
        "_WORKSHEET_WEB_EXTENSION_MAX_XML_BYTES",
        1,
    )
    oversized_snapshot = load_snapshot(oversized)

    assert oversized_snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "oversized Office Web Add-in worksheet binding XML part" in warning
        for warning in oversized_snapshot.parser_warnings
    )


def test_in_content_office_web_addins_and_preview_fallbacks_are_covered(tmp_path) -> None:
    baseline = make_in_content_office_web_addin_model(tmp_path / "baseline.xlsx")
    anchor_candidate = make_in_content_office_web_addin_model(
        tmp_path / "anchor-candidate.xlsx"
    )
    preview_candidate = make_in_content_office_web_addin_model(
        tmp_path / "preview-candidate.xlsx"
    )
    change_in_content_office_web_addin_anchor(anchor_candidate)
    change_in_content_office_web_addin_preview_payload(preview_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    anchor_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(anchor_candidate),
    )
    preview_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(preview_candidate),
    )
    addin_change = next(
        change
        for change in anchor_report.changes
        if change.kind == "office_web_addins_changed"
    )
    preview_change = next(
        change
        for change in preview_report.changes
        if change.kind == "worksheet_image_controls_changed"
    )

    assert baseline_snapshot.parser_warnings == ()
    assert baseline_snapshot.office_web_addins.to_dict() == {
        "present": True,
        "declared_taskpane_part_count": 0,
        "taskpane_part_count": 0,
        "web_extension_part_count": 1,
        "unrecognized_part_count": 0,
        "taskpane_count": 0,
        "visible_taskpane_count": 0,
        "locked_taskpane_count": 0,
        "web_extension_reference_count": 0,
        "auto_show_taskpane_count": 0,
        "store_reference_count": 1,
        "alternate_reference_count": 0,
        "binding_count": 0,
        "snapshot_reference_count": 0,
        "related_relationship_count": 1,
        "external_relationship_count": 0,
        "worksheet_binding_sheet_count": 0,
        "worksheet_binding_count": 0,
        "in_content_drawing_part_count": 1,
        "in_content_web_extension_reference_count": 1,
        "in_content_web_extension_part_count": 1,
    }
    assert baseline_snapshot.worksheet_drawing_shapes.present is False
    assert baseline_snapshot.worksheet_images.anchored_picture_count == 1
    assert baseline_snapshot.worksheet_images.fingerprinted_image_part_count == 1
    assert "**In-content drawing parts / references / definitions:** 1 / 1 / 1" in markdown
    assert addin_change.details["in_content_drawing_binding_changed"] is True
    assert "FF028" in {finding.rule_id for finding in anchor_report.findings}
    assert preview_change.details["worksheet_image_payload_material_changed"] is True
    assert "FF059" in {finding.rule_id for finding in preview_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(anchor_report.to_dict()),
        report_to_markdown(anchor_report),
        json.dumps(report_to_sarif(anchor_report)),
        json.dumps(preview_report.to_dict()),
        report_to_markdown(preview_report),
        json.dumps(report_to_sarif(preview_report)),
    )
    for sensitive_value in (
        "PrivateInContentAddin",
        "private-in-content-manifest.xml",
        "private in-content behavior",
        "PRIVATE-IN-CONTENT-ADDIN-FRAME",
        "PRIVATE-IN-CONTENT-ADDIN-PREVIEW",
        "rIdFenceInContentExtension",
        "rIdFenceInContentPreview",
        "private-in-content-addin-preview.png",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_in_content_office_web_addin_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_in_content_office_web_addin_model(tmp_path / "baseline.xlsx")
    renumbered = make_in_content_office_web_addin_model(tmp_path / "renumbered.xlsx")
    renumber_in_content_office_web_addin_identifiers(renumbered)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered))

    assert report.changes == []
    assert report.findings == []


def test_malformed_in_content_office_web_addin_references_fail_closed(tmp_path) -> None:
    baseline = make_in_content_office_web_addin_model(tmp_path / "baseline.xlsx")
    candidate = make_in_content_office_web_addin_model(tmp_path / "candidate.xlsx")
    corrupt_in_content_office_web_addin_reference(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.office_web_addins.unrecognized_part_count >= 1
    assert any(
        "malformed or unsupported Office Web Add-in in-content DrawingML metadata"
        in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "FF028" in {finding.rule_id for finding in report.findings}


def test_worksheet_embedded_controls_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_worksheet_embedded_control_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    change_worksheet_embedded_control_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    control_change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_embedded_controls_changed"
    )

    assert baseline_snapshot.summary()["has_worksheet_embedded_controls"] is True
    assert baseline_snapshot.summary()["worksheet_active_x_part_count"] == 1
    assert baseline_snapshot.summary()["worksheet_ole_object_count"] == 2
    assert baseline_snapshot.parser_warnings == ()
    assert profile["worksheet_embedded_controls"] == {
        "present": True,
        "control_sheet_count": 1,
        "worksheet_control_count": 2,
        "active_x_part_count": 1,
        "active_x_binary_reference_count": 1,
        "form_control_property_part_count": 1,
        "legacy_vml_drawing_part_count": 0,
        "legacy_vml_control_count": 0,
        "legacy_vml_macro_assignment_count": 0,
        "legacy_vml_cell_link_count": 0,
        "legacy_vml_source_range_count": 0,
        "legacy_vml_camera_source_range_count": 0,
        "control_macro_assignment_count": 1,
        "control_cell_link_count": 4,
        "control_source_range_count": 2,
        "form_control_formula_binding_count": 4,
        "ole_object_count": 2,
        "linked_ole_object_count": 1,
        "auto_load_ole_object_count": 1,
        "auto_update_ole_object_count": 1,
        "related_relationship_count": 7,
        "external_relationship_count": 1,
        "internal_related_part_count": 2,
        "fingerprinted_related_part_count": 2,
        "uninspected_related_part_count": 0,
        "unrecognized_part_count": 0,
    }
    assert "## Worksheet embedded controls, legacy VML controls, and OLE objects" in markdown
    assert "**ActiveX persistence parts:** 1" in markdown
    assert control_change.details["worksheet_binding_changed"] is True
    assert control_change.details["worksheet_control_definition_material_changed"] is True
    assert control_change.details["active_x_definition_material_changed"] is True
    assert control_change.details["form_control_property_material_changed"] is True
    assert control_change.details["related_part_relationships_changed"] is True
    assert control_change.details["embedded_payload_material_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF029"}

    sensitive_values = (
        "PrivateBaselineCommandButton",
        "PrivateBaselineFormControl",
        "PrivateBaselineControlMacro",
        "Private.Baseline.Embedded.Object",
        "private-baseline-link-name",
        "private baseline ActiveX binary payload",
        "private baseline embedded OLE payload",
        "private control presentation",
        "private/baseline-linked-ole.bin",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_embedded_control_alternate_content_is_not_double_counted(tmp_path) -> None:
    workbook = make_worksheet_embedded_control_model(
        tmp_path / "alternate-content.xlsx",
        alternate_content=True,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_embedded_controls.worksheet_control_count == 2
    assert snapshot.worksheet_embedded_controls.control_macro_assignment_count == 1
    assert snapshot.parser_warnings == ()


def test_worksheet_embedded_control_package_addition_is_critical(tmp_path) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    control_change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_embedded_controls_changed"
    )

    assert control_change.details["before"]["present"] is False
    assert control_change.details["after"]["present"] is True
    assert control_change.details["worksheet_binding_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF029"}


def test_worksheet_embedded_control_payloads_are_guarded_privately(tmp_path) -> None:
    baseline = make_worksheet_embedded_control_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    change_worksheet_embedded_control_payload(candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    control_change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_embedded_controls_changed"
    )

    assert baseline_snapshot.worksheet_embedded_controls.internal_related_part_count == 2
    assert baseline_snapshot.worksheet_embedded_controls.fingerprinted_related_part_count == 2
    assert control_change.details["embedded_payload_material_changed"] is True
    assert "worksheet_control_definition_material_changed" not in control_change.details
    assert "related_part_relationships_changed" not in control_change.details


def test_worksheet_embedded_control_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_worksheet_embedded_control_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    renumber_worksheet_embedded_control_relationships(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "worksheet_embedded_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "legacy_comment_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF029" not in {finding.rule_id for finding in report.findings}
    assert "FF046" not in {finding.rule_id for finding in report.findings}


def test_worksheet_embedded_control_equivalent_target_spellings_are_ignored(tmp_path) -> None:
    baseline = make_worksheet_embedded_control_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    rewrite_worksheet_embedded_control_internal_target_spelling(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "worksheet_embedded_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF029" not in {finding.rule_id for finding in report.findings}


def test_legacy_vml_controls_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_legacy_vml_control_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_vml_control_model(tmp_path / "candidate.xlsx")
    change_legacy_vml_control_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    control_change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_embedded_controls_changed"
    )

    controls = profile["worksheet_embedded_controls"]
    assert baseline_snapshot.summary()["worksheet_legacy_vml_drawing_part_count"] == 1
    assert baseline_snapshot.summary()["worksheet_legacy_vml_control_count"] == 4
    assert controls["control_sheet_count"] == 1
    assert controls["worksheet_control_count"] == 0
    assert controls["legacy_vml_drawing_part_count"] == 1
    assert controls["legacy_vml_control_count"] == 4
    assert controls["legacy_vml_macro_assignment_count"] == 1
    assert controls["legacy_vml_cell_link_count"] == 3
    assert controls["legacy_vml_source_range_count"] == 1
    assert controls["legacy_vml_camera_source_range_count"] == 1
    assert controls["control_macro_assignment_count"] == 1
    assert controls["control_cell_link_count"] == 3
    assert controls["control_source_range_count"] == 1
    assert controls["related_relationship_count"] == 2
    assert baseline_snapshot.parser_warnings == ()
    assert "**Legacy VML drawing parts / controls:** 1 / 4" in markdown
    assert control_change.details["legacy_vml_control_definition_material_changed"] is True
    assert control_change.details["legacy_vml_related_part_relationships_changed"] is True
    assert control_change.details["related_part_relationships_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF029"}

    sensitive_values = (
        "PrivateLegacyVmlMacro",
        "PrivateCandidateVmlMacro",
        "Private legacy VML button caption",
        "Private legacy VML note text",
        "private-legacy-vml.png",
        "private-candidate-legacy-vml.png",
        "Inputs!$B$2:$B$4",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_legacy_vml_comment_notes_do_not_change_control_findings(tmp_path) -> None:
    baseline = make_legacy_vml_control_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_vml_control_model(tmp_path / "candidate.xlsx")
    change_legacy_vml_note(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "worksheet_embedded_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF029" not in {finding.rule_id for finding in report.findings}


def test_legacy_vml_comment_notes_are_not_profiled_as_controls(tmp_path) -> None:
    workbook = make_legacy_vml_note_model(tmp_path / "notes.xlsx")

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_embedded_controls.present is False
    assert snapshot.worksheet_embedded_controls.legacy_vml_drawing_part_count == 0
    assert snapshot.worksheet_embedded_controls.legacy_vml_control_count == 0
    assert snapshot.parser_warnings == ()


def test_legacy_vml_control_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_legacy_vml_control_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_vml_control_model(tmp_path / "candidate.xlsx")
    renumber_legacy_vml_control_relationships(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "worksheet_embedded_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF029" not in {finding.rule_id for finding in report.findings}


def test_legacy_vml_control_equivalent_target_spellings_are_ignored(tmp_path) -> None:
    baseline = make_legacy_vml_control_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_vml_control_model(tmp_path / "candidate.xlsx")
    rewrite_legacy_vml_control_internal_target_spelling(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "worksheet_embedded_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF029" not in {finding.rule_id for finding in report.findings}


def test_malformed_legacy_vml_control_parts_fail_closed(tmp_path) -> None:
    baseline = make_legacy_vml_control_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_vml_control_model(tmp_path / "candidate.xlsx")
    corrupt_legacy_vml_control_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.worksheet_embedded_controls.unrecognized_part_count >= 1
    assert any(
        "legacy VML control part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_embedded_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF029" in {finding.rule_id for finding in report.findings}


def test_legacy_vml_control_xml_budgets_fail_closed(tmp_path, monkeypatch) -> None:
    workbook = make_legacy_vml_control_model(tmp_path / "candidate.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_WORKSHEET_EMBEDDED_CONTROL_MAX_XML_PART_BYTES",
        1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_embedded_controls.unrecognized_part_count >= 1
    assert any(
        "oversized worksheet embedded-control XML part" in warning
        for warning in snapshot.parser_warnings
    )


def test_ordinary_worksheets_do_not_consume_control_xml_budget(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_model(tmp_path / "ordinary.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_WORKSHEET_EMBEDDED_CONTROL_MAX_XML_PART_BYTES",
        1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_embedded_controls.present is False
    assert not any(
        "worksheet embedded-control XML" in warning
        for warning in snapshot.parser_warnings
    )


def test_oversized_worksheet_embedded_control_parts_fail_closed(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    monkeypatch.setattr(
        "formulafence.workbook._WORKSHEET_EMBEDDED_CONTROL_MAX_XML_PART_BYTES",
        1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_embedded_controls.unrecognized_part_count >= 1
    assert any(
        "oversized worksheet embedded-control XML part" in warning
        for warning in snapshot.parser_warnings
    )


def test_worksheet_embedded_control_xml_byte_budget_remains_covered(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._WorksheetEmbeddedControlXmlBudget
    monkeypatch.setattr(
        workbook_module,
        "_WorksheetEmbeddedControlXmlBudget",
        lambda: budget_type(remaining_bytes=1),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_embedded_controls.unrecognized_part_count >= 1
    assert any(
        "worksheet embedded-control XML read budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_worksheet_embedded_control_xml_part_budget_remains_covered(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._WorksheetEmbeddedControlXmlBudget
    monkeypatch.setattr(
        workbook_module,
        "_WorksheetEmbeddedControlXmlBudget",
        lambda: budget_type(remaining_parts=1),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_embedded_controls.unrecognized_part_count >= 1
    assert any(
        "worksheet embedded-control XML part count budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_oversized_worksheet_embedded_control_payloads_remain_uninspected(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_WORKSHEET_EMBEDDED_CONTROL_RELATED_PART_MAX_BYTES",
        1,
    )

    snapshot = load_snapshot(workbook)
    controls = snapshot.worksheet_embedded_controls

    assert controls.internal_related_part_count == 2
    assert controls.fingerprinted_related_part_count == 0
    assert controls.uninspected_related_part_count == 2
    assert any(
        "oversized worksheet embedded-control payload part" in warning
        for warning in snapshot.parser_warnings
    )


def test_worksheet_embedded_control_payload_byte_budget_remains_covered(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._WorksheetEmbeddedControlRelatedPartBudget
    monkeypatch.setattr(
        workbook_module,
        "_WorksheetEmbeddedControlRelatedPartBudget",
        lambda: budget_type(remaining_bytes=1),
    )

    snapshot = load_snapshot(workbook)
    controls = snapshot.worksheet_embedded_controls

    assert controls.fingerprinted_related_part_count == 0
    assert controls.uninspected_related_part_count == 2
    assert any(
        "worksheet embedded-control payload read budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_worksheet_embedded_control_payload_part_budget_remains_covered(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._WorksheetEmbeddedControlRelatedPartBudget
    monkeypatch.setattr(
        workbook_module,
        "_WorksheetEmbeddedControlRelatedPartBudget",
        lambda: budget_type(remaining_parts=1),
    )

    snapshot = load_snapshot(workbook)
    controls = snapshot.worksheet_embedded_controls

    assert controls.fingerprinted_related_part_count == 1
    assert controls.uninspected_related_part_count == 1
    assert any(
        "worksheet embedded-control payload part count budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_malformed_worksheet_embedded_control_activex_parts_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_embedded_control_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_embedded_control_model(tmp_path / "candidate.xlsx")
    corrupt_worksheet_embedded_control_activex_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.worksheet_embedded_controls.unrecognized_part_count == 1
    assert any(
        "ActiveX control part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_embedded_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF029" in {finding.rule_id for finding in report.findings}


def _chart_xml_members(path) -> tuple[str, str, str]:
    """Return the chart fixture's DrawingML host, chart, and overlay XML parts."""
    with ZipFile(path) as archive:
        members = tuple(entry.filename for entry in archive.infolist())
    drawing_member = next(
        member
        for member in members
        if (
            member.startswith("xl/drawings/drawing")
            and member.endswith(".xml")
            and "/_rels/" not in member
        )
    )
    chart_member = next(
        member
        for member in members
        if member.startswith("xl/charts/chart") and member.endswith(".xml")
    )
    overlay_member = "xl/drawings/chartDrawing1.xml"
    if overlay_member not in members:
        raise ValueError("chart fixture does not contain an overlay XML part")
    return drawing_member, chart_member, overlay_member


def _append_chart_xml_elements(path, member: str, count: int, *, nested: bool = False) -> None:
    """Add opaque chart XML entries without invoking FormulaFence's reader."""
    staging = path.with_suffix(".chart-xml-elements.xlsx")
    with ZipFile(path) as archive:
        contents = {
            entry.filename: archive.read(entry)
            for entry in archive.infolist()
        }
    root = ElementTree.fromstring(contents[member])
    parent = root
    if nested:
        parent = ElementTree.SubElement(root, "{urn:formulafence:audit}opaque")
    for _ in range(count):
        ElementTree.SubElement(parent, "{urn:formulafence:audit}entry")
    contents[member] = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
    with ZipFile(staging, "w", compression=ZIP_DEFLATED) as archive:
        for name, payload in contents.items():
            archive.writestr(name, payload)
    staging.replace(path)


def _chart_xml_element_count(path, member: str) -> int:
    """Count one fixture's complete chart XML tree for exact budget tests."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member))
    return sum(1 for _ in root.iter())


def _pivot_xml_members(path) -> tuple[str, str]:
    """Return the PivotTable fixture's view and cache-definition XML members."""
    with ZipFile(path) as archive:
        members = tuple(entry.filename for entry in archive.infolist())
    pivot_table_member = next(
        member
        for member in members
        if member.startswith("xl/pivotTables/") and member.endswith(".xml")
    )
    cache_definition_member = next(
        member
        for member in members
        if (
            member.startswith("xl/pivotCache/pivotCacheDefinition")
            and member.endswith(".xml")
        )
    )
    return pivot_table_member, cache_definition_member


def _append_pivot_xml_elements(path, member: str, count: int, *, nested: bool = False) -> None:
    """Add opaque PivotTable XML entries without invoking FormulaFence's reader."""
    staging = path.with_suffix(".pivot-xml-elements.xlsx")
    with ZipFile(path) as archive:
        contents = {
            entry.filename: archive.read(entry)
            for entry in archive.infolist()
        }
    root = ElementTree.fromstring(contents[member])
    parent = root
    if nested:
        parent = ElementTree.SubElement(root, "{urn:formulafence:audit}opaque")
    for _ in range(count):
        ElementTree.SubElement(parent, "{urn:formulafence:audit}entry")
    contents[member] = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
    with ZipFile(staging, "w", compression=ZIP_DEFLATED) as archive:
        for name, payload in contents.items():
            archive.writestr(name, payload)
    staging.replace(path)


def _pivot_xml_element_count(path, member: str) -> int:
    """Count one fixture's complete PivotTable XML tree for exact budget tests."""
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(member))
    return sum(1 for _ in root.iter())


def test_chart_definitions_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_chart_definition_model(tmp_path / "candidate.xlsx")
    change_chart_definition_material(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    chart_change = next(
        change for change in report.changes if change.kind == "chart_definitions_changed"
    )

    charts = profile["chart_definitions"]
    assert baseline_snapshot.summary()["chart_host_sheet_count"] == 1
    assert baseline_snapshot.summary()["chart_part_count"] == 1
    assert baseline_snapshot.summary()["chart_series_count"] == 1
    assert baseline_snapshot.summary()["chart_cached_data_point_count"] == 6
    assert charts == {
        "present": True,
        "chart_host_sheet_count": 1,
        "chart_drawing_part_count": 1,
        "chart_reference_count": 1,
        "chart_part_count": 1,
        "chart_ex_reference_count": 0,
        "chart_ex_part_count": 0,
        "chart_ex_series_count": 0,
        "chart_ex_title_count": 0,
        "chart_ex_data_reference_count": 0,
        "chart_user_shape_part_count": 1,
        "chart_user_shape_count": 1,
        "chart_type_count": 1,
        "series_count": 1,
        "title_count": 1,
        "data_reference_count": 3,
        "numeric_data_reference_count": 2,
        "string_data_reference_count": 1,
        "literal_data_point_count": 0,
        "cached_data_point_count": 6,
        "pivot_source_count": 0,
        "external_data_reference_count": 0,
        "user_shape_reference_count": 1,
        "related_relationship_count": 3,
        "external_relationship_count": 0,
        "internal_related_part_count": 1,
        "fingerprinted_related_part_count": 1,
        "uninspected_related_part_count": 0,
        "unrecognized_part_count": 0,
    }
    assert baseline_snapshot.parser_warnings == ()
    assert "## Chart definitions and cached presentation data" in markdown
    assert "Office 2016+ ChartEx parts" in markdown
    assert chart_change.details["chart_definition_material_changed"] is True
    assert chart_change.details["overlay_shape_material_changed"] is True
    assert chart_change.details["related_part_relationships_changed"] is True
    assert chart_change.details["related_part_payload_material_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF030"}

    sensitive_values = (
        "Private baseline chart title",
        "Private candidate chart title",
        "Private baseline overlay text",
        "Private candidate overlay text",
        "Inputs!$B$3:$B$4",
        "private-chart-overlay-baseline.png",
        "private-chart-overlay-candidate.png",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_chart_cached_data_is_compared_separately_from_chart_definition(tmp_path) -> None:
    baseline = make_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_chart_definition_model(tmp_path / "candidate.xlsx")
    change_chart_cached_data(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    chart_change = next(
        change for change in report.changes if change.kind == "chart_definitions_changed"
    )

    assert chart_change.details["cached_series_material_changed"] is True
    assert "chart_definition_material_changed" not in chart_change.details
    assert "overlay_shape_material_changed" not in chart_change.details
    assert {finding.rule_id for finding in report.findings} >= {"FF030"}


def test_extended_chart_definitions_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_extended_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_extended_chart_definition_model(tmp_path / "candidate.xlsx")
    change_extended_chart_definition_material(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    chart_change = next(
        change for change in report.changes if change.kind == "chart_definitions_changed"
    )

    charts = profile["chart_definitions"]
    assert baseline_snapshot.summary()["chart_ex_part_count"] == 1
    assert baseline_snapshot.summary()["chart_ex_series_count"] == 1
    assert charts["chart_host_sheet_count"] == 1
    assert charts["chart_drawing_part_count"] == 1
    assert charts["chart_reference_count"] == 0
    assert charts["chart_part_count"] == 0
    assert charts["chart_ex_reference_count"] == 1
    assert charts["chart_ex_part_count"] == 1
    assert charts["chart_ex_series_count"] == 1
    assert charts["chart_ex_title_count"] == 1
    assert charts["chart_ex_data_reference_count"] == 2
    assert charts["related_relationship_count"] == 3
    assert charts["internal_related_part_count"] == 2
    assert charts["fingerprinted_related_part_count"] == 2
    assert charts["unrecognized_part_count"] == 0
    assert baseline_snapshot.worksheet_drawing_shapes.present is False
    assert baseline_snapshot.parser_warnings == ()
    assert "Office 2016+ ChartEx parts:** 1" in markdown
    assert chart_change.details["chart_definition_material_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF030"}

    sensitive_values = (
        "PRIVATE-CHARTEX-NAME",
        "PRIVATE-CHARTEX-TITLE",
        "PRIVATE-CANDIDATE-CHARTEX-TITLE",
        "PRIVATE-CHARTEX-VALUES",
        "PRIVATE-CANDIDATE-CHARTEX-VALUES",
        "PRIVATE-CHARTEX-FALLBACK-TEXT",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_extended_chart_style_payload_is_compared_separately(tmp_path) -> None:
    baseline = make_extended_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_extended_chart_definition_model(tmp_path / "candidate.xlsx")
    change_extended_chart_style_payload(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    chart_change = next(
        change for change in report.changes if change.kind == "chart_definitions_changed"
    )

    assert chart_change.details["related_part_payload_material_changed"] is True
    assert "chart_definition_material_changed" not in chart_change.details
    assert {finding.rule_id for finding in report.findings} >= {"FF030"}


def test_extended_chart_relationship_identifier_noise_is_ignored(tmp_path) -> None:
    baseline = make_extended_chart_definition_model(tmp_path / "baseline.xlsx")
    renumbered = make_extended_chart_definition_model(tmp_path / "renumbered.xlsx")
    renumber_extended_chart_relationships(renumbered)

    renumbered_snapshot = load_snapshot(renumbered)
    report = compare_snapshots(load_snapshot(baseline), renumbered_snapshot)

    assert renumbered_snapshot.parser_warnings == ()
    assert "chart_definitions_changed" not in {change.kind for change in report.changes}
    assert "FF030" not in {finding.rule_id for finding in report.findings}


def test_malformed_extended_chart_parts_fail_closed(tmp_path) -> None:
    baseline = make_extended_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_extended_chart_definition_model(tmp_path / "candidate.xlsx")
    corrupt_extended_chart_definition_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "extended-chart part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "chart_definitions_changed" in {change.kind for change in report.changes}
    assert "FF030" in {finding.rule_id for finding in report.findings}


def test_unsupported_extended_chart_relationships_fail_closed(tmp_path) -> None:
    workbook = make_extended_chart_definition_model(tmp_path / "candidate.xlsx")
    add_unsupported_extended_chart_relationship(workbook)

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "unsupported direct extended-chart relationship" in warning
        for warning in snapshot.parser_warnings
    )


def test_extended_chart_direct_relationships_without_ids_fail_closed(tmp_path) -> None:
    workbook = make_extended_chart_definition_model(tmp_path / "candidate.xlsx")
    remove_extended_chart_direct_relationship_id(workbook)

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "extended-chart direct relationship without an id" in warning
        for warning in snapshot.parser_warnings
    )


def test_external_extended_chart_bindings_fail_closed(tmp_path) -> None:
    baseline = make_extended_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_extended_chart_definition_model(tmp_path / "candidate.xlsx")
    externalize_extended_chart_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.chart_definitions.external_relationship_count >= 1
    assert candidate_snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "extended DrawingML chart reference without a safe internal target" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "chart_definitions_changed" in {change.kind for change in report.changes}


def test_chartsheet_chart_parts_are_discovered_from_drawing_relationships(tmp_path) -> None:
    workbook = make_protection_model(tmp_path / "chartsheet.xlsx", include_chartsheet=True)

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.present is True
    assert snapshot.chart_definitions.chart_host_sheet_count == 1
    assert snapshot.chart_definitions.chart_part_count == 1
    assert snapshot.chart_definitions.series_count == 1
    assert snapshot.parser_warnings == ()


def test_chartless_workbook_does_not_consume_chart_xml_budget(tmp_path, monkeypatch) -> None:
    workbook = make_model(tmp_path / "chartless.xlsx")
    monkeypatch.setattr(workbook_module, "_CHART_MAX_XML_PART_BYTES", 1)

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.present is False
    assert snapshot.chart_definitions.chart_part_count == 0
    assert not any("chart XML" in warning for warning in snapshot.parser_warnings)


def test_external_chart_targets_are_not_followed_or_exposed(tmp_path) -> None:
    baseline = make_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_chart_definition_model(tmp_path / "external-target.xlsx")
    externalize_chart_overlay_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.chart_definitions.external_relationship_count == 1
    assert candidate_snapshot.chart_definitions.internal_related_part_count == 0
    assert candidate_snapshot.parser_warnings == ()
    assert "FF030" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all(
        "example.invalid/private-chart-overlay.png" not in artifact
        for artifact in rendered_artifacts
    )


def test_chart_relationship_identifier_and_target_spelling_noise_is_ignored(tmp_path) -> None:
    baseline = make_chart_definition_model(tmp_path / "baseline.xlsx")
    renumbered = make_chart_definition_model(tmp_path / "renumbered.xlsx")
    target_rewritten = make_chart_definition_model(tmp_path / "target-rewritten.xlsx")
    renumber_chart_relationships(renumbered)
    rewrite_chart_internal_target_spelling(target_rewritten)

    renumbered_report = compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered))
    target_report = compare_snapshots(
        load_snapshot(baseline), load_snapshot(target_rewritten)
    )

    assert "chart_definitions_changed" not in {
        change.kind for change in renumbered_report.changes
    }
    assert "FF030" not in {finding.rule_id for finding in renumbered_report.findings}
    assert "chart_definitions_changed" not in {
        change.kind for change in target_report.changes
    }
    assert "FF030" not in {finding.rule_id for finding in target_report.findings}


def test_malformed_chart_parts_fail_closed(tmp_path) -> None:
    baseline = make_chart_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_chart_definition_model(tmp_path / "candidate.xlsx")
    corrupt_chart_definition_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "chart part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "chart_definitions_changed" in {change.kind for change in report.changes}
    assert "FF030" in {finding.rule_id for finding in report.findings}


def test_chart_xml_element_budget_stops_before_materializing_the_tree(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_chart_definition_model(tmp_path / "candidate.xlsx")
    _drawing_member, chart_member, _overlay_member = _chart_xml_members(workbook)
    _append_chart_xml_elements(workbook, chart_member, 1)
    monkeypatch.setattr(
        workbook_module,
        "_CHART_MAX_XML_ELEMENT_COUNT",
        _chart_xml_element_count(workbook, chart_member) - 1,
    )
    with ZipFile(workbook) as archive:
        chart_payload = archive.read(chart_member)
    original_tree_parse = workbook_module._xml_root_from_payload

    def unexpected_tree_parse(payload):
        if payload == chart_payload:
            raise AssertionError("the chart XML tree was materialized after its budget failed")
        return original_tree_parse(payload)

    monkeypatch.setattr(workbook_module, "_xml_root_from_payload", unexpected_tree_parse)
    warnings: set[str] = set()
    with ZipFile(workbook) as archive:
        inspection = workbook_module._chart_part_inspection(
            archive,
            chart_member,
            warnings,
            workbook_module._ChartXmlBudget(),
        )

    assert inspection.inspected is False
    assert any("chart XML part whose XML structure" in warning for warning in warnings)


def test_chart_xml_element_budget_remains_covered(tmp_path, monkeypatch) -> None:
    baseline = make_chart_definition_model(tmp_path / "baseline.xlsx")
    workbook = make_chart_definition_model(tmp_path / "candidate.xlsx")
    _drawing_member, chart_member, _overlay_member = _chart_xml_members(workbook)
    _append_chart_xml_elements(workbook, chart_member, 2)
    monkeypatch.setattr(
        workbook_module,
        "_CHART_MAX_XML_ELEMENT_COUNT",
        _chart_xml_element_count(workbook, chart_member) - 1,
    )

    snapshot = load_snapshot(workbook)
    report = compare_snapshots(load_snapshot(baseline), snapshot)

    assert snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "chart XML part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )
    assert "FF030" in {finding.rule_id for finding in report.findings}


def test_chart_xml_element_budget_accepts_the_configured_capacity(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_chart_definition_model(tmp_path / "candidate.xlsx")
    _drawing_member, chart_member, _overlay_member = _chart_xml_members(workbook)
    _append_chart_xml_elements(workbook, chart_member, 2)
    monkeypatch.setattr(
        workbook_module,
        "_CHART_MAX_XML_ELEMENT_COUNT",
        _chart_xml_element_count(workbook, chart_member),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.unrecognized_part_count == 0


def test_chart_xml_element_budget_aggregates_across_package_parts(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_chart_definition_model(tmp_path / "candidate.xlsx")
    xml_members = _chart_xml_members(workbook)
    remaining_xml_elements = (
        sum(_chart_xml_element_count(workbook, member) for member in xml_members) - 1
    )
    budget_type = workbook_module._ChartXmlBudget
    monkeypatch.setattr(
        workbook_module,
        "_ChartXmlBudget",
        lambda: budget_type(remaining_xml_elements=remaining_xml_elements),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "chart XML part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_chart_xml_element_budget_accepts_the_default_capacity(tmp_path) -> None:
    workbook = make_chart_definition_model(tmp_path / "candidate.xlsx")
    _drawing_member, chart_member, _overlay_member = _chart_xml_members(workbook)
    _append_chart_xml_elements(
        workbook,
        chart_member,
        workbook_module._CHART_MAX_XML_ELEMENT_COUNT
        - _chart_xml_element_count(workbook, chart_member),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.unrecognized_part_count == 0


def test_chart_xml_element_budget_rejects_the_default_overage(tmp_path) -> None:
    workbook = make_chart_definition_model(tmp_path / "candidate.xlsx")
    _drawing_member, chart_member, _overlay_member = _chart_xml_members(workbook)
    _append_chart_xml_elements(
        workbook,
        chart_member,
        workbook_module._CHART_MAX_XML_ELEMENT_COUNT
        - _chart_xml_element_count(workbook, chart_member)
        + 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "chart XML part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_chart_xml_element_budget_counts_nested_opaque_subtrees(tmp_path, monkeypatch) -> None:
    workbook = make_chart_definition_model(tmp_path / "candidate.xlsx")
    _drawing_member, chart_member, _overlay_member = _chart_xml_members(workbook)
    _append_chart_xml_elements(workbook, chart_member, 1, nested=True)
    monkeypatch.setattr(
        workbook_module,
        "_CHART_MAX_XML_ELEMENT_COUNT",
        _chart_xml_element_count(workbook, chart_member) - 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "chart XML part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_chart_xml_and_related_part_budgets_fail_closed(tmp_path, monkeypatch) -> None:
    workbook = make_chart_definition_model(tmp_path / "candidate.xlsx")
    monkeypatch.setattr(workbook_module, "_CHART_MAX_XML_PART_BYTES", 1)

    oversized_snapshot = load_snapshot(workbook)

    assert oversized_snapshot.chart_definitions.unrecognized_part_count >= 1
    assert any(
        "oversized chart XML part" in warning
        for warning in oversized_snapshot.parser_warnings
    )

    payload_workbook = make_chart_definition_model(tmp_path / "payload.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_CHART_MAX_XML_PART_BYTES",
        16 * 1024 * 1024,
    )
    monkeypatch.setattr(workbook_module, "_CHART_RELATED_PART_MAX_BYTES", 1)
    payload_snapshot = load_snapshot(payload_workbook)

    assert payload_snapshot.chart_definitions.internal_related_part_count == 1
    assert payload_snapshot.chart_definitions.fingerprinted_related_part_count == 0
    assert payload_snapshot.chart_definitions.uninspected_related_part_count == 1
    assert any(
        "oversized chart related part" in warning
        for warning in payload_snapshot.parser_warnings
    )


def test_pivot_xml_element_budget_stops_before_materializing_the_tree(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    pivot_table_member, _cache_definition_member = _pivot_xml_members(workbook)
    _append_pivot_xml_elements(workbook, pivot_table_member, 1)
    monkeypatch.setattr(
        workbook_module,
        "_PIVOT_MAX_XML_ELEMENT_COUNT",
        _pivot_xml_element_count(workbook, pivot_table_member) - 1,
    )
    with ZipFile(workbook) as archive:
        pivot_payload = archive.read(pivot_table_member)
    original_tree_parse = workbook_module._xml_root_from_payload

    def unexpected_tree_parse(payload):
        if payload == pivot_payload:
            raise AssertionError(
                "the PivotTable XML tree was materialized after its budget failed"
            )
        return original_tree_parse(payload)

    monkeypatch.setattr(workbook_module, "_xml_root_from_payload", unexpected_tree_parse)
    warnings: set[str] = set()
    with ZipFile(workbook) as archive:
        inspection = workbook_module._pivot_table_part_inspection(
            archive,
            pivot_table_member,
            warnings,
            workbook_module._PivotXmlBudget(),
            {},
        )

    assert inspection.inspected is False
    assert any("PivotTable XML part whose XML structure" in warning for warning in warnings)


def test_pivot_cache_xml_element_budget_stops_all_materializing_readers(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    _pivot_table_member, cache_definition_member = _pivot_xml_members(workbook)
    _append_pivot_xml_elements(workbook, cache_definition_member, 1)
    monkeypatch.setattr(
        workbook_module,
        "_PIVOT_MAX_XML_ELEMENT_COUNT",
        _pivot_xml_element_count(workbook, cache_definition_member) - 1,
    )
    with ZipFile(workbook) as archive:
        cache_definition_payload = archive.read(cache_definition_member)
    original_tree_parse = workbook_module._xml_root_from_payload

    def unexpected_tree_parse(payload):
        if payload == cache_definition_payload:
            raise AssertionError(
                "the PivotTable cache XML tree was materialized after its budget failed"
            )
        return original_tree_parse(payload)

    monkeypatch.setattr(workbook_module, "_xml_root_from_payload", unexpected_tree_parse)

    snapshot = load_snapshot(workbook)

    assert snapshot.pivot_table_definitions.unrecognized_part_count >= 1
    assert any(
        "PivotTable XML part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_pivot_xml_element_budget_remains_covered(tmp_path, monkeypatch) -> None:
    baseline = make_pivot_table_definition_model(tmp_path / "baseline.xlsx")
    workbook = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    pivot_table_member, _cache_definition_member = _pivot_xml_members(workbook)
    _append_pivot_xml_elements(workbook, pivot_table_member, 2)
    monkeypatch.setattr(
        workbook_module,
        "_PIVOT_MAX_XML_ELEMENT_COUNT",
        _pivot_xml_element_count(workbook, pivot_table_member) - 1,
    )

    snapshot = load_snapshot(workbook)
    report = compare_snapshots(load_snapshot(baseline), snapshot)

    assert snapshot.pivot_table_definitions.unrecognized_part_count >= 1
    assert any(
        "PivotTable XML part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )
    assert "FF031" in {finding.rule_id for finding in report.findings}


def test_pivot_xml_element_budget_accepts_the_configured_capacity(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    pivot_table_member, _cache_definition_member = _pivot_xml_members(workbook)
    _append_pivot_xml_elements(workbook, pivot_table_member, 2)
    monkeypatch.setattr(
        workbook_module,
        "_PIVOT_MAX_XML_ELEMENT_COUNT",
        _pivot_xml_element_count(workbook, pivot_table_member),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.pivot_table_definitions.unrecognized_part_count == 0


def test_pivot_xml_element_budget_aggregates_across_package_parts(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    xml_members = _pivot_xml_members(workbook)
    remaining_xml_elements = (
        sum(_pivot_xml_element_count(workbook, member) for member in xml_members) - 1
    )
    budget_type = workbook_module._PivotXmlBudget
    monkeypatch.setattr(
        workbook_module,
        "_PivotXmlBudget",
        lambda: budget_type(remaining_xml_elements=remaining_xml_elements),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.pivot_table_definitions.unrecognized_part_count >= 1
    assert any(
        "PivotTable XML part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_pivot_xml_element_budget_accepts_the_default_capacity(tmp_path) -> None:
    workbook = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    pivot_table_member, _cache_definition_member = _pivot_xml_members(workbook)
    _append_pivot_xml_elements(
        workbook,
        pivot_table_member,
        workbook_module._PIVOT_MAX_XML_ELEMENT_COUNT
        - _pivot_xml_element_count(workbook, pivot_table_member),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.pivot_table_definitions.unrecognized_part_count == 0


def test_pivot_xml_element_budget_rejects_the_default_overage(tmp_path) -> None:
    workbook = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    _pivot_table_member, cache_definition_member = _pivot_xml_members(workbook)
    _append_pivot_xml_elements(
        workbook,
        cache_definition_member,
        workbook_module._PIVOT_MAX_XML_ELEMENT_COUNT
        - _pivot_xml_element_count(workbook, cache_definition_member)
        + 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.pivot_table_definitions.unrecognized_part_count >= 1
    assert any(
        "PivotTable XML part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_pivot_xml_element_budget_counts_nested_opaque_subtrees(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    pivot_table_member, _cache_definition_member = _pivot_xml_members(workbook)
    _append_pivot_xml_elements(workbook, pivot_table_member, 1, nested=True)
    monkeypatch.setattr(
        workbook_module,
        "_PIVOT_MAX_XML_ELEMENT_COUNT",
        _pivot_xml_element_count(workbook, pivot_table_member) - 1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.pivot_table_definitions.unrecognized_part_count >= 1
    assert any(
        "PivotTable XML part whose XML structure" in warning
        for warning in snapshot.parser_warnings
    )


def test_pivot_table_definitions_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_pivot_table_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    change_pivot_table_definition_material(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    pivot_change = next(
        change
        for change in report.changes
        if change.kind == "pivot_table_definitions_changed"
    )

    pivots = profile["pivot_table_definitions"]
    assert baseline_snapshot.summary()["pivot_table_sheet_count"] == 1
    assert baseline_snapshot.summary()["pivot_table_part_count"] == 1
    assert baseline_snapshot.summary()["pivot_cache_definition_part_count"] == 1
    assert baseline_snapshot.summary()["pivot_cache_record_count"] == 2
    assert pivots == {
        "present": True,
        "pivot_table_sheet_count": 1,
        "pivot_table_part_count": 1,
        "pivot_cache_definition_part_count": 1,
        "pivot_cache_records_part_count": 1,
        "pivot_cache_binding_count": 1,
        "layout_location_count": 1,
        "pivot_field_count": 1,
        "row_field_count": 1,
        "column_field_count": 0,
        "page_field_count": 0,
        "data_field_count": 1,
        "filter_count": 0,
        "row_item_count": 1,
        "column_item_count": 0,
        "cache_field_count": 1,
        "shared_item_count": 2,
        "calculated_item_count": 0,
        "calculated_member_count": 0,
        "cache_record_count": 2,
        "related_relationship_count": 2,
        "external_relationship_count": 0,
        "fingerprinted_cache_record_part_count": 1,
        "uninspected_cache_record_part_count": 0,
        "unrecognized_part_count": 0,
    }
    assert baseline_snapshot.parser_warnings == ()
    assert "## PivotTable definitions and cached report material" in markdown
    assert pivot_change.details["pivot_table_layout_material_changed"] is True
    assert pivot_change.details["pivot_cache_definition_material_changed"] is True
    assert pivot_change.details["cached_shared_item_material_changed"] is True
    assert pivot_change.details["cache_record_payload_material_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF031"}
    assert "FF023" not in {finding.rule_id for finding in report.findings}

    sensitive_values = (
        "Private baseline pivot report",
        "Private baseline total",
        "Private candidate total",
        "Private baseline category",
        "Private candidate category",
        "Private baseline item one",
        "Private candidate item",
        "Private Pivot Source",
        "A2:B3",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_pivot_table_refresh_controls_remain_with_external_data_guard(tmp_path) -> None:
    baseline = make_pivot_table_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    change_pivot_table_refresh_control(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "pivot_cache_refresh_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "pivot_table_definitions_changed" not in {
        change.kind for change in report.changes
    }
    assert {finding.rule_id for finding in report.findings} >= {"FF023"}
    assert "FF031" not in {finding.rule_id for finding in report.findings}


def test_pivot_relationship_ids_targets_and_cache_ids_are_normalized(tmp_path) -> None:
    baseline = make_pivot_table_definition_model(tmp_path / "baseline.xlsx")
    renumbered = make_pivot_table_definition_model(tmp_path / "renumbered.xlsx")
    target_rewritten = make_pivot_table_definition_model(tmp_path / "target.xlsx")
    cache_id_renumbered = make_pivot_table_definition_model(tmp_path / "cache-id.xlsx")
    renumber_pivot_table_relationships(renumbered)
    rewrite_pivot_table_internal_target_spelling(target_rewritten)
    renumber_pivot_table_cache_id(cache_id_renumbered)

    reports = (
        compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered)),
        compare_snapshots(load_snapshot(baseline), load_snapshot(target_rewritten)),
        compare_snapshots(load_snapshot(baseline), load_snapshot(cache_id_renumbered)),
    )

    for report in reports:
        assert "pivot_table_definitions_changed" not in {
            change.kind for change in report.changes
        }
        assert "pivot_cache_refresh_controls_changed" not in {
            change.kind for change in report.changes
        }
        assert "FF031" not in {finding.rule_id for finding in report.findings}
        assert "FF023" not in {finding.rule_id for finding in report.findings}


def test_pivot_cache_record_rebinding_is_guarded(tmp_path) -> None:
    baseline = make_pivot_table_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    rebind_pivot_table_cache_records(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    pivot_change = next(
        change
        for change in report.changes
        if change.kind == "pivot_table_definitions_changed"
    )

    assert pivot_change.details["pivot_cache_definition_material_changed"] is True
    assert pivot_change.details["related_part_relationships_changed"] is True
    assert pivot_change.details["cache_record_payload_material_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF031"}


def test_pivot_packages_are_not_parsed_by_the_underlying_reader(
    tmp_path,
    monkeypatch,
) -> None:
    from openpyxl.pivot.cache import CacheDefinition
    from openpyxl.pivot.record import RecordList
    from openpyxl.pivot.table import TableDefinition

    workbook = make_pivot_table_definition_model(tmp_path / "candidate.xlsx")
    original_payload = workbook.read_bytes()

    def reject_pivot_parse(*args, **kwargs):
        raise AssertionError("the workbook reader must not parse PivotTable packages")

    monkeypatch.setattr(CacheDefinition, "from_tree", reject_pivot_parse)
    monkeypatch.setattr(RecordList, "from_tree", reject_pivot_parse)
    monkeypatch.setattr(TableDefinition, "from_tree", reject_pivot_parse)

    snapshot = load_snapshot(workbook)

    assert snapshot.pivot_table_definitions.fingerprinted_cache_record_part_count == 1
    assert workbook.read_bytes() == original_payload
    assert not any(
        "could not prepare a PivotTable-safe workbook reader copy" in warning
        for warning in snapshot.parser_warnings
    )


def test_external_pivot_cache_record_targets_are_not_followed_or_exposed(tmp_path) -> None:
    baseline = make_pivot_table_definition_model(tmp_path / "baseline.xlsx")
    candidate = make_pivot_table_definition_model(tmp_path / "external-target.xlsx")
    externalize_pivot_table_cache_record_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.pivot_table_definitions.external_relationship_count == 1
    assert candidate_snapshot.pivot_table_definitions.fingerprinted_cache_record_part_count == 1
    assert any(
        "cache-record reference without a safe internal target" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "FF031" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all(
        "example.invalid/private-pivot-cache-records" not in artifact
        for artifact in rendered_artifacts
    )


def test_pivot_table_parts_fail_closed_on_malformed_or_bounded_material(
    tmp_path,
    monkeypatch,
) -> None:
    baseline = make_pivot_table_definition_model(tmp_path / "baseline.xlsx")
    malformed = make_pivot_table_definition_model(tmp_path / "malformed.xlsx")
    corrupt_pivot_table_definition_root(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.pivot_table_definitions.unrecognized_part_count >= 1
    assert any(
        "PivotTable part with an unexpected root" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert "pivot_table_definitions_changed" in {
        change.kind for change in malformed_report.changes
    }
    assert "FF031" in {finding.rule_id for finding in malformed_report.findings}

    oversized_xml = make_pivot_table_definition_model(tmp_path / "oversized.xml.xlsx")
    monkeypatch.setattr(workbook_module, "_PIVOT_MAX_XML_PART_BYTES", 1)
    oversized_xml_snapshot = load_snapshot(oversized_xml)

    assert oversized_xml_snapshot.pivot_table_definitions.unrecognized_part_count >= 1
    assert any(
        "oversized PivotTable XML part" in warning
        for warning in oversized_xml_snapshot.parser_warnings
    )

    oversized_records = make_pivot_table_definition_model(tmp_path / "oversized.records.xlsx")
    monkeypatch.setattr(workbook_module, "_PIVOT_MAX_XML_PART_BYTES", 16 * 1024 * 1024)
    monkeypatch.setattr(workbook_module, "_PIVOT_CACHE_RECORD_MAX_BYTES", 1)
    oversized_record_snapshot = load_snapshot(oversized_records)

    assert (
        oversized_record_snapshot.pivot_table_definitions.fingerprinted_cache_record_part_count
        == 0
    )
    assert (
        oversized_record_snapshot.pivot_table_definitions.uninspected_cache_record_part_count
        == 1
    )
    assert any(
        "oversized PivotTable cache-record part" in warning
        for warning in oversized_record_snapshot.parser_warnings
    )


def test_pivotless_workbook_does_not_consume_pivottable_budget(tmp_path, monkeypatch) -> None:
    workbook = make_model(tmp_path / "pivotless.xlsx")
    monkeypatch.setattr(workbook_module, "_PIVOT_MAX_XML_PART_BYTES", 1)

    snapshot = load_snapshot(workbook)

    assert snapshot.pivot_table_definitions.present is False
    assert snapshot.pivot_table_definitions.pivot_table_part_count == 0
    assert not any("PivotTable XML" in warning for warning in snapshot.parser_warnings)


def test_slicer_timeline_cache_filters_are_profiled_and_diffed_privately(tmp_path) -> None:
    baseline = make_slicer_timeline_cache_model(tmp_path / "baseline.xlsx")
    candidate = make_slicer_timeline_cache_model(tmp_path / "candidate.xlsx")
    change_slicer_timeline_filter_material(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    filter_change = next(
        change
        for change in report.changes
        if change.kind == "slicer_timeline_cache_definitions_changed"
    )

    filters = profile["slicer_timeline_caches"]
    assert baseline_snapshot.summary()["slicer_cache_part_count"] == 2
    assert baseline_snapshot.summary()["timeline_cache_part_count"] == 1
    assert baseline_snapshot.summary()["selected_slicer_item_count"] == 1
    assert filters == {
        "present": True,
        "slicer_cache_part_count": 2,
        "timeline_cache_part_count": 1,
        "slicer_workbook_binding_count": 2,
        "timeline_workbook_binding_count": 1,
        "slicer_pivot_cache_binding_count": 1,
        "slicer_table_binding_count": 1,
        "timeline_pivot_cache_binding_count": 1,
        "slicer_pivot_table_binding_count": 1,
        "timeline_pivot_table_binding_count": 1,
        "slicer_item_count": 2,
        "selected_slicer_item_count": 1,
        "timeline_state_count": 1,
        "timeline_filter_count": 0,
        "related_relationship_count": 0,
        "external_relationship_count": 0,
        "unrecognized_part_count": 0,
    }
    assert "## Slicer and Timeline cache filter state" in markdown
    assert filter_change.details["slicer_filter_state_or_definition_material_changed"] is True
    assert filter_change.details["timeline_filter_state_or_definition_material_changed"] is True
    assert "FF032" in {finding.rule_id for finding in report.findings}
    assert "FF031" not in {finding.rule_id for finding in report.findings}

    sensitive_values = (
        "Private baseline revenue slicer",
        "Private baseline business unit",
        "Private baseline table slicer",
        "Private baseline transaction date",
        "Private baseline sales timeline",
        "Private baseline pivot report",
        "2024-01-01T00:00:00Z",
        "2024-05-14T00:00:00Z",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_slicer_timeline_relationships_targets_and_pivot_cache_ids_are_normalized(
    tmp_path,
) -> None:
    baseline = make_slicer_timeline_cache_model(tmp_path / "baseline.xlsx")
    renumbered = make_slicer_timeline_cache_model(tmp_path / "renumbered.xlsx")
    target_rewritten = make_slicer_timeline_cache_model(tmp_path / "target.xlsx")
    cache_id_renumbered = make_slicer_timeline_cache_model(tmp_path / "cache-id.xlsx")
    timeline_2011 = make_slicer_timeline_cache_model(tmp_path / "timeline-2011.xlsx")
    renumber_slicer_timeline_relationships(renumbered)
    rewrite_slicer_timeline_internal_target_spelling(target_rewritten)
    renumber_slicer_timeline_pivot_cache_id(cache_id_renumbered)
    use_slicer_timeline_2011_relationship_type(timeline_2011)

    reports = (
        compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered)),
        compare_snapshots(load_snapshot(baseline), load_snapshot(target_rewritten)),
        compare_snapshots(load_snapshot(baseline), load_snapshot(cache_id_renumbered)),
        compare_snapshots(load_snapshot(baseline), load_snapshot(timeline_2011)),
    )

    for report in reports:
        rule_ids = {finding.rule_id for finding in report.findings}
        change_kinds = {change.kind for change in report.changes}
        assert "slicer_timeline_cache_definitions_changed" not in change_kinds
        assert "pivot_table_definitions_changed" not in change_kinds
        assert "pivot_cache_refresh_controls_changed" not in change_kinds
        assert "FF032" not in rule_ids
        assert "FF031" not in rule_ids
        assert "FF023" not in rule_ids


def test_slicer_timeline_optional_defaults_and_uids_are_normalized(tmp_path) -> None:
    baseline = make_slicer_timeline_cache_model(tmp_path / "baseline.xlsx")
    equivalent = make_slicer_timeline_cache_model(tmp_path / "equivalent.xlsx")
    set_slicer_timeline_equivalent_defaults(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert "slicer_timeline_cache_definitions_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF032" not in {finding.rule_id for finding in report.findings}


def test_slicer_timeline_cache_binding_and_external_targets_are_guarded(tmp_path) -> None:
    baseline = make_slicer_timeline_cache_model(tmp_path / "baseline.xlsx")
    rebound = make_slicer_timeline_cache_model(tmp_path / "rebound.xlsx")
    external = make_slicer_timeline_cache_model(tmp_path / "external.xlsx")
    broken_pivot_binding = make_slicer_timeline_cache_model(tmp_path / "broken-pivot.xlsx")
    rebind_slicer_timeline_cache(rebound)
    externalize_slicer_timeline_cache_relationship(external)
    break_slicer_timeline_pivot_cache_binding(broken_pivot_binding)

    rebound_report = compare_snapshots(load_snapshot(baseline), load_snapshot(rebound))
    rebound_change = next(
        change
        for change in rebound_report.changes
        if change.kind == "slicer_timeline_cache_definitions_changed"
    )
    external_snapshot = load_snapshot(external)
    external_profile = profile_snapshot(external_snapshot)
    external_report = compare_snapshots(load_snapshot(baseline), external_snapshot)
    broken_pivot_snapshot = load_snapshot(broken_pivot_binding)
    broken_pivot_report = compare_snapshots(
        load_snapshot(baseline), broken_pivot_snapshot
    )

    assert rebound_change.details["workbook_cache_binding_changed"] is True
    assert "FF032" in {finding.rule_id for finding in rebound_report.findings}
    assert external_snapshot.slicer_timeline_caches.unrecognized_part_count >= 1
    assert any(
        "without a safe internal target" in warning
        for warning in external_snapshot.parser_warnings
    )
    assert "FF032" in {finding.rule_id for finding in external_report.findings}
    assert broken_pivot_snapshot.slicer_timeline_caches.unrecognized_part_count >= 2
    assert any(
        "could not bind a slicer cache to one PivotTable cache definition" in warning
        for warning in broken_pivot_snapshot.parser_warnings
    )
    assert "FF032" in {finding.rule_id for finding in broken_pivot_report.findings}
    assert "FF031" not in {finding.rule_id for finding in broken_pivot_report.findings}
    rendered_artifacts = (
        json.dumps(external_profile),
        profile_to_markdown(external_profile),
        json.dumps(external_report.to_dict()),
        report_to_markdown(external_report),
        json.dumps(report_to_sarif(external_report)),
    )
    assert all(
        "example.invalid/private-slicer-cache" not in artifact
        for artifact in rendered_artifacts
    )


def test_slicer_timeline_cache_parts_fail_closed_on_malformed_or_bounded_material(
    tmp_path,
    monkeypatch,
) -> None:
    baseline = make_slicer_timeline_cache_model(tmp_path / "baseline.xlsx")
    malformed = make_slicer_timeline_cache_model(tmp_path / "malformed.xlsx")
    corrupt_slicer_timeline_cache_root(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.slicer_timeline_caches.unrecognized_part_count >= 1
    assert any(
        "slicer/Timeline cache part with an unexpected root" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert "slicer_timeline_cache_definitions_changed" in {
        change.kind for change in malformed_report.changes
    }
    assert "FF032" in {finding.rule_id for finding in malformed_report.findings}

    oversized = make_slicer_timeline_cache_model(tmp_path / "oversized.xlsx")
    monkeypatch.setattr(workbook_module, "_SLICER_TIMELINE_MAX_XML_PART_BYTES", 1)
    oversized_snapshot = load_snapshot(oversized)

    assert oversized_snapshot.slicer_timeline_caches.unrecognized_part_count >= 1
    assert any(
        "oversized slicer/Timeline XML part" in warning
        for warning in oversized_snapshot.parser_warnings
    )


def test_slicerless_workbook_does_not_consume_filter_cache_budget(tmp_path, monkeypatch) -> None:
    workbook = make_model(tmp_path / "slicerless.xlsx")
    monkeypatch.setattr(workbook_module, "_SLICER_TIMELINE_MAX_XML_PART_BYTES", 1)

    snapshot = load_snapshot(workbook)

    assert snapshot.slicer_timeline_caches.present is False
    assert snapshot.slicer_timeline_caches.slicer_cache_part_count == 0
    assert snapshot.slicer_timeline_caches.timeline_cache_part_count == 0
    assert not any("slicer/Timeline XML" in warning for warning in snapshot.parser_warnings)


def test_power_pivot_data_model_is_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_power_pivot_data_model(tmp_path / "baseline.xlsx")
    payload_candidate = make_power_pivot_data_model(tmp_path / "payload.xlsx")
    declaration_candidate = make_power_pivot_data_model(tmp_path / "declaration.xlsx")
    change_power_pivot_data_model_payload(payload_candidate)
    change_power_pivot_data_model_declaration(declaration_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    payload_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(payload_candidate),
    )
    declaration_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(declaration_candidate),
    )
    payload_change = next(
        change
        for change in payload_report.changes
        if change.kind == "power_pivot_data_model_changed"
    )
    declaration_change = next(
        change
        for change in declaration_report.changes
        if change.kind == "power_pivot_data_model_changed"
    )

    data_model = profile["power_pivot_data_model"]
    assert baseline_snapshot.summary()["power_pivot_data_model_part_count"] == 1
    assert baseline_snapshot.summary()["power_pivot_data_model_table_count"] == 2
    assert data_model == {
        "present": True,
        "data_model_part_count": 1,
        "workbook_binding_count": 1,
        "data_model_declaration_count": 1,
        "model_table_count": 2,
        "model_relationship_count": 1,
        "related_relationship_count": 0,
        "external_relationship_count": 0,
        "fingerprinted_data_part_count": 1,
        "uninspected_data_part_count": 0,
        "unrecognized_part_count": 0,
    }
    assert "## Power Pivot Data Model" in markdown
    assert payload_change.details["embedded_data_model_payload_changed"] is True
    assert "workbook_data_model_declaration_changed" not in payload_change.details
    assert declaration_change.details["workbook_data_model_declaration_changed"] is True
    assert "embedded_data_model_payload_changed" not in declaration_change.details
    assert "FF033" in {finding.rule_id for finding in payload_report.findings}
    assert "FF033" in {finding.rule_id for finding in declaration_report.findings}

    sensitive_values = (
        "Private baseline revenue table",
        "Private baseline calendar table",
        "Private baseline data connection",
        "Private baseline calendar key",
        "private baseline Power Pivot data model payload",
        "private candidate Power Pivot data model payload",
        "Private candidate calendar key",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(payload_report.to_dict()),
        report_to_markdown(payload_report),
        json.dumps(report_to_sarif(payload_report)),
        json.dumps(declaration_report.to_dict()),
        report_to_markdown(declaration_report),
        json.dumps(report_to_sarif(declaration_report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_power_pivot_data_model_binding_noise_and_writer_guids_are_normalized(
    tmp_path,
) -> None:
    baseline = make_power_pivot_data_model(tmp_path / "baseline.xlsx")
    renumbered = make_power_pivot_data_model(tmp_path / "renumbered.xlsx")
    target_rewritten = make_power_pivot_data_model(tmp_path / "target.xlsx")
    regenerated_guids = make_power_pivot_data_model(tmp_path / "guids.xlsx")
    renumber_power_pivot_data_model_relationship(renumbered)
    rewrite_power_pivot_data_model_internal_target_spelling(target_rewritten)
    set_power_pivot_data_model_equivalent_guids(regenerated_guids)

    reports = (
        compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered)),
        compare_snapshots(load_snapshot(baseline), load_snapshot(target_rewritten)),
        compare_snapshots(load_snapshot(baseline), load_snapshot(regenerated_guids)),
    )

    for report in reports:
        assert "power_pivot_data_model_changed" not in {
            change.kind for change in report.changes
        }
        assert "FF033" not in {finding.rule_id for finding in report.findings}


def test_power_pivot_data_model_bindings_payload_bounds_and_direct_rels_fail_closed(
    tmp_path,
    monkeypatch,
) -> None:
    baseline = make_power_pivot_data_model(tmp_path / "baseline.xlsx")
    rebound = make_power_pivot_data_model(tmp_path / "rebound.xlsx")
    external = make_power_pivot_data_model(tmp_path / "external.xlsx")
    related = make_power_pivot_data_model(tmp_path / "related.xlsx")
    rebind_power_pivot_data_model(rebound)
    externalize_power_pivot_data_model(external)
    add_power_pivot_data_model_direct_relationship(related)

    rebound_report = compare_snapshots(load_snapshot(baseline), load_snapshot(rebound))
    external_snapshot = load_snapshot(external)
    external_profile = profile_snapshot(external_snapshot)
    external_report = compare_snapshots(load_snapshot(baseline), external_snapshot)
    related_snapshot = load_snapshot(related)
    related_report = compare_snapshots(load_snapshot(baseline), related_snapshot)

    rebound_change = next(
        change
        for change in rebound_report.changes
        if change.kind == "power_pivot_data_model_changed"
    )
    assert rebound_change.details["workbook_data_model_declaration_changed"] is True
    assert "FF033" in {finding.rule_id for finding in rebound_report.findings}
    assert external_snapshot.power_pivot_data_model.external_relationship_count == 1
    assert external_snapshot.power_pivot_data_model.unrecognized_part_count >= 1
    assert any(
        "without a safe internal target" in warning
        for warning in external_snapshot.parser_warnings
    )
    assert "FF033" in {finding.rule_id for finding in external_report.findings}
    assert related_snapshot.power_pivot_data_model.related_relationship_count == 1
    assert related_snapshot.power_pivot_data_model.unrecognized_part_count >= 1
    assert any(
        "direct relationships on a Power Pivot/Data Model" in warning
        for warning in related_snapshot.parser_warnings
    )
    assert "FF033" in {finding.rule_id for finding in related_report.findings}
    rendered_artifacts = (
        json.dumps(external_profile),
        profile_to_markdown(external_profile),
        json.dumps(external_report.to_dict()),
        report_to_markdown(external_report),
        json.dumps(report_to_sarif(external_report)),
    )
    assert all(
        "example.invalid/private-power-pivot-data-model" not in artifact
        for artifact in rendered_artifacts
    )

    oversized = make_power_pivot_data_model(tmp_path / "oversized.xlsx")
    monkeypatch.setattr(workbook_module, "_POWER_PIVOT_DATA_MAX_BYTES", 1)
    oversized_snapshot = load_snapshot(oversized)

    assert oversized_snapshot.power_pivot_data_model.uninspected_data_part_count == 1
    assert any(
        "oversized Power Pivot/Data Model part" in warning
        for warning in oversized_snapshot.parser_warnings
    )


def test_power_pivot_data_model_unbound_declaration_fails_closed(tmp_path) -> None:
    baseline = make_power_pivot_data_model(tmp_path / "baseline.xlsx")
    candidate = make_power_pivot_data_model(tmp_path / "candidate.xlsx")
    remove_power_pivot_data_model_workbook_binding(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    data_model = candidate_snapshot.power_pivot_data_model
    assert data_model.present is True
    assert data_model.workbook_binding_count == 0
    assert data_model.data_model_declaration_count == 1
    assert data_model.fingerprinted_data_part_count == 1
    assert data_model.unrecognized_part_count >= 2
    assert any(
        "workbook metadata without a Power Pivot data relationship" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert any(
        "package part not bound by an inspected workbook relationship" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "FF033" in {finding.rule_id for finding in report.findings}


def test_power_pivot_data_model_payload_budgets_remain_covered(tmp_path, monkeypatch) -> None:
    budget_type = workbook_module._PowerPivotDataBudget
    part_limited = make_power_pivot_data_model(tmp_path / "part-limited.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_PowerPivotDataBudget",
        lambda: budget_type(remaining_parts=0),
    )

    part_limited_snapshot = load_snapshot(part_limited)

    assert part_limited_snapshot.power_pivot_data_model.uninspected_data_part_count == 1
    assert any(
        "Power Pivot/Data Model part count budget" in warning
        for warning in part_limited_snapshot.parser_warnings
    )

    byte_limited = make_power_pivot_data_model(tmp_path / "byte-limited.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_PowerPivotDataBudget",
        lambda: budget_type(remaining_bytes=1),
    )

    byte_limited_snapshot = load_snapshot(byte_limited)

    assert byte_limited_snapshot.power_pivot_data_model.uninspected_data_part_count == 1
    assert any(
        "Power Pivot/Data Model read budget" in warning
        for warning in byte_limited_snapshot.parser_warnings
    )


def test_data_model_free_workbook_does_not_consume_payload_budget(tmp_path, monkeypatch) -> None:
    workbook = make_model(tmp_path / "data-model-free.xlsx")
    monkeypatch.setattr(workbook_module, "_POWER_PIVOT_DATA_MAX_BYTES", 1)

    snapshot = load_snapshot(workbook)

    assert snapshot.power_pivot_data_model.present is False
    assert snapshot.power_pivot_data_model.data_model_part_count == 0
    assert not any(
        "Power Pivot/Data Model part" in warning for warning in snapshot.parser_warnings
    )


def test_what_if_data_tables_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_what_if_data_table_model(tmp_path / "baseline.xlsx")
    candidate = make_what_if_data_table_model(tmp_path / "candidate.xlsx")
    change_what_if_data_table_input(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "what_if_data_tables_changed"
    )

    assert baseline_snapshot.summary()["what_if_data_table_count"] == 3
    assert baseline_snapshot.summary()["what_if_data_table_output_cell_count"] == 17
    assert baseline_snapshot.summary()["has_what_if_data_tables"] is True
    assert profile["what_if_data_tables"] == {
        "present": True,
        "data_table_count": 3,
        "one_variable_data_table_count": 2,
        "two_variable_data_table_count": 1,
        "one_variable_row_oriented_count": 1,
        "one_variable_column_oriented_count": 1,
        "declared_output_cell_count": 17,
        "recalculation_requested_count": 1,
        "deleted_input_reference_count": 0,
        "unrecognized_data_table_count": 0,
    }
    assert baseline_snapshot.cells[("Sensitivity", "D3")].formula == "=TABLE()"
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## What-If Data Tables" in markdown
    assert change.details["data_table_definition_material_changed"] is True
    assert "FF034" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in ("B2", "B4", "D3:D6", "Sensitivity!D3"):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_what_if_data_table_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_what_if_data_table_model(tmp_path / "baseline.xlsx")
    equivalent = make_what_if_data_table_model(tmp_path / "equivalent.xlsx")
    normalize_what_if_data_table_reference_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert "what_if_data_tables_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF034" not in {finding.rule_id for finding in report.findings}


def test_what_if_data_table_deleted_and_malformed_inputs_fail_closed(tmp_path) -> None:
    baseline = make_what_if_data_table_model(tmp_path / "baseline.xlsx")
    deleted = make_what_if_data_table_model(tmp_path / "deleted.xlsx")
    malformed = make_what_if_data_table_model(tmp_path / "malformed.xlsx")
    overlapping = make_what_if_data_table_model(tmp_path / "overlapping.xlsx")
    delete_what_if_data_table_input(deleted)
    corrupt_what_if_data_table_input(malformed)
    overlap_what_if_data_table_outputs(overlapping)

    deleted_snapshot = load_snapshot(deleted)
    deleted_report = compare_snapshots(load_snapshot(baseline), deleted_snapshot)
    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    malformed_report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)
    overlapping_snapshot = load_snapshot(overlapping)

    assert deleted_snapshot.what_if_data_tables.deleted_input_reference_count == 1
    assert deleted_snapshot.what_if_data_tables.unrecognized_data_table_count == 0
    assert "FF034" in {finding.rule_id for finding in deleted_report.findings}
    assert malformed_snapshot.what_if_data_tables.unrecognized_data_table_count == 1
    assert any(
        "malformed or unsupported What-If Data Table" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF034"} <= {
        finding.rule_id for finding in malformed_report.findings
    }
    assert overlapping_snapshot.what_if_data_tables.unrecognized_data_table_count == 2
    assert any(
        "overlapping What-If Data Table output ranges" in warning
        for warning in overlapping_snapshot.parser_warnings
    )

    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(malformed_report.to_dict()),
        report_to_markdown(malformed_report),
        json.dumps(report_to_sarif(malformed_report)),
    )
    assert all(
        "PrivateInputSheet!B2" not in artifact for artifact in rendered_artifacts
    )


def test_data_table_free_workbook_has_no_sensitivity_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.what_if_data_tables.present is False
    assert snapshot.what_if_data_tables.data_table_count == 0
    assert not any("What-If Data Table" in warning for warning in snapshot.parser_warnings)


def test_scenario_manager_is_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_scenario_manager_model(tmp_path / "baseline.xlsx")
    candidate = make_scenario_manager_model(tmp_path / "candidate.xlsx")
    change_scenario_manager_input_value(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(
        change for change in report.changes if change.kind == "scenario_manager_changed"
    )

    assert baseline_snapshot.summary()["scenario_manager_sheet_count"] == 1
    assert baseline_snapshot.summary()["scenario_manager_scenario_count"] == 2
    assert baseline_snapshot.summary()["scenario_manager_input_cell_count"] == 4
    assert baseline_snapshot.summary()["has_scenario_manager"] is True
    assert profile["scenario_manager"] == {
        "present": True,
        "scenario_sheet_count": 1,
        "scenario_count": 2,
        "input_cell_count": 4,
        "locked_scenario_count": 1,
        "hidden_scenario_count": 1,
        "scenario_with_comment_count": 2,
        "scenario_with_user_count": 1,
        "summary_reference_count": 2,
        "current_scenario_selection_count": 1,
        "shown_scenario_selection_count": 1,
        "deleted_input_cell_count": 0,
        "undone_input_cell_count": 0,
        "formatted_input_cell_count": 1,
        "unrecognized_scenario_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Scenario Manager" in markdown
    assert change.details["scenario_definition_material_changed"] is True
    assert "FF035" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "Private Upside",
        "Private Downside",
        "PRIVATE-UPSIDERATE",
        "CANDIDATE-PRIVATE-SCENARIO-VALUE",
        "PRIVATE-SCENARIO-COMMENT",
        "private-scenario-owner",
        "B2",
        "D2 D3",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_scenario_manager_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_scenario_manager_model(tmp_path / "baseline.xlsx")
    equivalent = make_scenario_manager_model(tmp_path / "equivalent.xlsx")
    normalize_scenario_manager_reference_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert "scenario_manager_changed" not in {change.kind for change in report.changes}
    assert "FF035" not in {finding.rule_id for finding in report.findings}


def test_scenario_manager_names_are_scoped_to_their_worksheet(tmp_path) -> None:
    snapshot = load_snapshot(
        make_scenario_manager_model(
            tmp_path / "worksheet-scoped-names.xlsx",
            duplicate_names_on_second_sheet=True,
        )
    )

    assert snapshot.scenario_manager.scenario_sheet_count == 2
    assert snapshot.scenario_manager.scenario_count == 4
    assert snapshot.scenario_manager.unrecognized_scenario_count == 0
    assert not any("Scenario Manager" in warning for warning in snapshot.parser_warnings)


def test_scenario_manager_malformed_input_fails_closed(tmp_path) -> None:
    baseline = make_scenario_manager_model(tmp_path / "baseline.xlsx")
    malformed = make_scenario_manager_model(tmp_path / "malformed.xlsx")
    corrupt_scenario_manager_input(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    malformed_report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.scenario_manager.unrecognized_scenario_count == 1
    assert any(
        "malformed or unsupported Scenario Manager" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF035"} <= {
        finding.rule_id for finding in malformed_report.findings
    }

    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(malformed_report.to_dict()),
        report_to_markdown(malformed_report),
        json.dumps(report_to_sarif(malformed_report)),
    )
    assert all("PrivateInputSheet!B2" not in artifact for artifact in rendered_artifacts)


def test_scenario_manager_free_workbook_has_no_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.scenario_manager.present is False
    assert snapshot.scenario_manager.scenario_count == 0
    assert not any("Scenario Manager" in warning for warning in snapshot.parser_warnings)


def test_filter_visibility_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_filter_visibility_model(tmp_path / "baseline.xlsx")
    candidate = make_filter_visibility_model(tmp_path / "candidate.xlsx")
    table_candidate = make_filter_visibility_model(tmp_path / "table-candidate.xlsx")
    row_candidate = make_filter_visibility_model(tmp_path / "row-candidate.xlsx")
    column_candidate = make_filter_visibility_model(tmp_path / "column-candidate.xlsx")
    change_filter_visibility_criterion(candidate)
    change_table_filter_visibility_criterion(table_candidate)
    change_filter_visibility_hidden_row(row_candidate)
    change_filter_visibility_hidden_column(column_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    table_report = compare_snapshots(baseline_snapshot, load_snapshot(table_candidate))
    row_report = compare_snapshots(baseline_snapshot, load_snapshot(row_candidate))
    column_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(column_candidate),
    )
    change = next(
        change
        for change in report.changes
        if change.kind == "filter_visibility_controls_changed"
    )

    assert baseline_snapshot.summary()["filter_visibility_auto_filter_count"] == 2
    assert baseline_snapshot.summary()["filter_visibility_hidden_row_count"] == 2
    assert baseline_snapshot.summary()["filter_visibility_hidden_column_count"] == 3
    assert baseline_snapshot.summary()["has_filter_visibility_controls"] is True
    assert profile["filter_visibility_controls"] == {
        "present": True,
        "worksheet_auto_filter_count": 1,
        "table_auto_filter_count": 1,
        "filter_column_count": 2,
        "filter_criterion_count": 2,
        "sort_state_count": 2,
        "sort_condition_count": 2,
        "default_hidden_sheet_count": 1,
        "default_zero_height_sheet_count": 0,
        "default_zero_width_sheet_count": 0,
        "hidden_row_count": 2,
        "zero_height_row_count": 0,
        "outlined_row_count": 2,
        "collapsed_row_count": 1,
        "visible_row_override_count": 1,
        "hidden_column_count": 3,
        "zero_width_column_count": 0,
        "outlined_column_count": 4,
        "collapsed_column_count": 1,
        "unrecognized_control_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Filter, sort, and visibility controls" in markdown
    assert change.details["filter_visibility_definition_material_changed"] is True
    assert "FF036" in {finding.rule_id for finding in report.findings}
    assert "FF036" in {finding.rule_id for finding in table_report.findings}
    assert "FF036" in {finding.rule_id for finding in row_report.findings}
    assert "FF036" in {finding.rule_id for finding in column_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-WORKSHEET-REGION",
        "CANDIDATE-PRIVATE-WORKSHEET-REGION",
        "PRIVATE-TABLE-SEGMENT",
        "PRIVATE-SORT-LIST",
        "C2:C5",
        "outlineLevel",
        "<col",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_filter_visibility_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_filter_visibility_model(tmp_path / "baseline.xlsx")
    equivalent = make_filter_visibility_model(tmp_path / "equivalent.xlsx")
    normalize_filter_visibility_control_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert "filter_visibility_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF036" not in {finding.rule_id for finding in report.findings}


def test_zero_dimension_visibility_controls_are_profiled_diffed_and_redacted(
    tmp_path,
) -> None:
    baseline = make_zero_dimension_visibility_model(tmp_path / "baseline.xlsx")
    candidate = make_zero_dimension_visibility_model(tmp_path / "candidate.xlsx")
    default_candidate = make_zero_dimension_visibility_model(
        tmp_path / "default-candidate.xlsx"
    )
    default_revealed = make_zero_dimension_visibility_model(
        tmp_path / "default-revealed.xlsx"
    )
    ordinary_resize = make_zero_dimension_visibility_model(
        tmp_path / "ordinary-resize.xlsx"
    )
    change_zero_dimension_visibility_controls(candidate)
    change_default_zero_dimension_visibility_controls(default_candidate)
    change_default_zero_dimension_visibility_controls(default_revealed)
    add_ordinary_dimension_resize(default_revealed)
    add_ordinary_dimension_resize(ordinary_resize)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    default_snapshot = load_snapshot(default_candidate)
    default_revealed_snapshot = load_snapshot(default_revealed)
    profile = profile_snapshot(candidate_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    default_report = compare_snapshots(baseline_snapshot, default_snapshot)
    default_override_report = compare_snapshots(
        default_snapshot,
        default_revealed_snapshot,
    )
    ordinary_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(ordinary_resize),
    )
    change = next(
        change
        for change in report.changes
        if change.kind == "filter_visibility_controls_changed"
    )

    assert candidate_snapshot.summary()["filter_visibility_zero_height_row_count"] == 1
    assert candidate_snapshot.summary()["filter_visibility_zero_width_column_count"] == 1
    assert profile["filter_visibility_controls"] == {
        "present": True,
        "worksheet_auto_filter_count": 0,
        "table_auto_filter_count": 0,
        "filter_column_count": 0,
        "filter_criterion_count": 0,
        "sort_state_count": 0,
        "sort_condition_count": 0,
        "default_hidden_sheet_count": 0,
        "default_zero_height_sheet_count": 0,
        "default_zero_width_sheet_count": 0,
        "hidden_row_count": 0,
        "zero_height_row_count": 1,
        "outlined_row_count": 0,
        "collapsed_row_count": 0,
        "visible_row_override_count": 0,
        "hidden_column_count": 0,
        "zero_width_column_count": 1,
        "outlined_column_count": 0,
        "collapsed_column_count": 0,
        "unrecognized_control_count": 0,
    }
    assert default_snapshot.filter_visibility_controls.default_zero_height_sheet_count == 1
    assert default_snapshot.filter_visibility_controls.default_zero_width_sheet_count == 1
    assert default_snapshot.filter_visibility_controls.zero_width_column_count == 16_384
    assert default_revealed_snapshot.filter_visibility_controls.zero_width_column_count == 16_383
    assert default_revealed_snapshot.filter_visibility_controls.visible_row_override_count == 1
    assert "FF036" in {finding.rule_id for finding in report.findings}
    assert "FF036" in {finding.rule_id for finding in default_report.findings}
    assert "FF036" in {finding.rule_id for finding in default_override_report.findings}
    assert "filter_visibility_controls_changed" not in {
        change.kind for change in ordinary_report.changes
    }
    assert "FF036" not in {finding.rule_id for finding in ordinary_report.findings}
    assert "zero-height rows" in markdown
    assert "zero-width columns" in markdown
    assert change.details["filter_visibility_definition_material_changed"] is True

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for private_value in ("customHeight", "customWidth", '"min": "2"', '"r": "3"'):
        assert all(private_value not in artifact for artifact in rendered_artifacts)


def test_zero_dimension_visibility_normalizes_equivalent_zero_spellings(tmp_path) -> None:
    baseline = make_zero_dimension_visibility_model(tmp_path / "baseline.xlsx")
    equivalent = make_zero_dimension_visibility_model(tmp_path / "equivalent.xlsx")
    change_zero_dimension_visibility_controls(baseline)
    change_zero_dimension_visibility_controls(equivalent)
    normalize_zero_dimension_visibility_control_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert "filter_visibility_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF036" not in {finding.rule_id for finding in report.findings}


def test_positive_dimension_resizes_are_distinct_from_visibility_controls(tmp_path) -> None:
    baseline = make_zero_dimension_visibility_model(tmp_path / "baseline.xlsx")
    candidate = make_zero_dimension_visibility_model(tmp_path / "candidate.xlsx")
    add_ordinary_dimension_resize(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "FF036" not in {finding.rule_id for finding in report.findings}
    assert "FF058" in {finding.rule_id for finding in report.findings}


def test_worksheet_dimension_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_worksheet_dimension_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_dimension_model(tmp_path / "candidate.xlsx")
    change_worksheet_dimension_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_dimension_controls_changed"
    )

    assert baseline_snapshot.cells == candidate_snapshot.cells
    assert baseline_snapshot.summary()["worksheet_dimension_control_count"] == 9
    assert baseline_snapshot.summary()["has_worksheet_dimension_controls"] is True
    assert profile["worksheet_dimension_controls"] == {
        "present": True,
        "default_row_height_count": 1,
        "default_column_width_count": 1,
        "default_baseline_adjustment_sheet_count": 0,
        "default_border_adjustment_sheet_count": 1,
        "row_height_assignment_count": 1,
        "row_baseline_adjustment_count": 0,
        "row_border_adjustment_count": 1,
        "column_width_assignment_count": 2,
        "best_fit_column_assignment_count": 2,
        "unrecognized_dimension_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Worksheet dimension controls" in markdown
    assert change.details["worksheet_dimension_definition_material_changed"] is True
    assert "FF058" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for private_value in (
        '"defaultRowHeight": "21"',
        '"defaultColWidth": "17.5"',
        '"width": "24"',
        '"min": "2"',
        '"r": "3"',
        "customWidth",
    ):
        assert all(private_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_dimension_writer_noise_and_inert_declarations_are_normalized(
    tmp_path,
) -> None:
    baseline = make_worksheet_dimension_model(tmp_path / "baseline.xlsx")
    spelling_equivalent = make_worksheet_dimension_model(
        tmp_path / "spelling-equivalent.xlsx"
    )
    inert_equivalent = make_worksheet_dimension_model(
        tmp_path / "inert-equivalent.xlsx"
    )
    normalize_worksheet_dimension_control_spelling(spelling_equivalent)
    normalize_worksheet_dimension_inert_declarations(inert_equivalent)

    spelling_report = compare_snapshots(
        load_snapshot(baseline),
        load_snapshot(spelling_equivalent),
    )
    inert_report = compare_snapshots(
        load_snapshot(baseline),
        load_snapshot(inert_equivalent),
    )

    assert spelling_report.changes == []
    assert spelling_report.findings == []
    assert inert_report.changes == []
    assert inert_report.findings == []


def test_worksheet_dimension_baseline_adjustments_are_material(tmp_path) -> None:
    baseline = make_worksheet_dimension_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_dimension_model(tmp_path / "candidate.xlsx")
    add_worksheet_dimension_baseline_adjustments(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert (
        candidate_snapshot.worksheet_dimension_controls.default_baseline_adjustment_sheet_count
        == 1
    )
    assert (
        candidate_snapshot.worksheet_dimension_controls.row_baseline_adjustment_count
        == 1
    )
    assert (
        candidate_snapshot.worksheet_dimension_controls.default_border_adjustment_sheet_count
        == 0
    )
    assert candidate_snapshot.worksheet_dimension_controls.unrecognized_dimension_count == 0
    assert not any(
        "worksheet-dimension" in warning for warning in candidate_snapshot.parser_warnings
    )
    assert "FF058" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("dyDescent" not in artifact for artifact in rendered_artifacts)


def test_worksheet_dimension_malformed_controls_fail_closed_and_redact_values(tmp_path) -> None:
    baseline = make_worksheet_dimension_model(tmp_path / "baseline.xlsx")
    malformed = make_worksheet_dimension_model(tmp_path / "malformed.xlsx")
    corrupt_worksheet_dimension_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, malformed_snapshot)

    assert malformed_snapshot.cells == baseline_snapshot.cells
    assert malformed_snapshot.worksheet_dimension_controls.unrecognized_dimension_count >= 1
    assert any(
        "malformed or unsupported worksheet-dimension" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF058"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for private_value in ("PRIVATE-INVALID-DIMENSION", '"ht": "NaN"'):
        assert all(private_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_dimension_preexisting_coverage_gap_stays_distinct(tmp_path) -> None:
    baseline = make_worksheet_dimension_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_dimension_model(tmp_path / "candidate.xlsx")
    change_worksheet_dimension_controls(candidate)
    corrupt_worksheet_dimension_control(baseline)
    corrupt_worksheet_dimension_control(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_dimension_controls_changed"
    )

    assert "FF058" in {finding.rule_id for finding in report.findings}
    assert "unrecognized_worksheet_dimension_metadata_changed" not in change.details


def test_ordinary_workbook_has_no_worksheet_dimension_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.worksheet_dimension_controls.present is False
    assert snapshot.worksheet_dimension_controls.column_width_assignment_count == 0
    assert not any("worksheet-dimension" in warning for warning in snapshot.parser_warnings)


def test_strict_worksheet_dimensions_are_supported(tmp_path) -> None:
    baseline = make_strict_worksheet_dimension_model(tmp_path / "baseline.xlsx")
    candidate = make_strict_worksheet_dimension_model(tmp_path / "candidate.xlsx")
    change_worksheet_dimension_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.worksheet_dimension_controls.row_height_assignment_count == 1
    assert baseline_snapshot.worksheet_dimension_controls.column_width_assignment_count == 2
    assert baseline_snapshot.worksheet_dimension_controls.best_fit_column_assignment_count == 2
    assert baseline_snapshot.worksheet_dimension_controls.unrecognized_dimension_count == 0
    assert not any(
        "worksheet-dimension" in warning for warning in baseline_snapshot.parser_warnings
    )
    assert "FF058" in {finding.rule_id for finding in report.findings}


def test_zero_dimension_visibility_malformed_controls_fail_closed(tmp_path) -> None:
    baseline = make_zero_dimension_visibility_model(tmp_path / "baseline.xlsx")
    malformed = make_zero_dimension_visibility_model(tmp_path / "malformed.xlsx")
    corrupt_zero_dimension_visibility_controls(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.filter_visibility_controls.unrecognized_control_count >= 1
    assert malformed_snapshot.filter_visibility_controls.zero_width_column_count == 0
    assert any(
        "malformed or unsupported filter, sort, or visibility" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF036"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for private_value in (
        '"width": "0_0"',
        '"ht": "NaN"',
        "customHeight",
        "customWidth",
    ):
        assert all(private_value not in artifact for artifact in rendered_artifacts)


def test_zero_dimension_default_controls_fail_closed(tmp_path) -> None:
    baseline = make_zero_dimension_visibility_model(tmp_path / "baseline.xlsx")
    malformed = make_zero_dimension_visibility_model(tmp_path / "malformed.xlsx")
    corrupt_default_zero_dimension_visibility_controls(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.filter_visibility_controls.unrecognized_control_count >= 1
    assert any(
        "malformed or unsupported filter, sort, or visibility" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF036"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for private_value in ('"defaultRowHeight": "-1"', '"defaultColWidth": "256"'):
        assert all(private_value not in artifact for artifact in rendered_artifacts)


def test_filter_visibility_malformed_control_fails_closed(tmp_path) -> None:
    baseline = make_filter_visibility_model(tmp_path / "baseline.xlsx")
    malformed = make_filter_visibility_model(tmp_path / "malformed.xlsx")
    corrupt_filter_visibility_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.filter_visibility_controls.unrecognized_control_count == 1
    assert any(
        "malformed or unsupported filter, sort, or visibility" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF036"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("4294967296" not in artifact for artifact in rendered_artifacts)


def test_filter_visibility_malformed_column_control_fails_closed(tmp_path) -> None:
    baseline = make_filter_visibility_model(tmp_path / "baseline.xlsx")
    malformed = make_filter_visibility_model(tmp_path / "malformed.xlsx")
    corrupt_filter_visibility_column_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.filter_visibility_controls.unrecognized_control_count == 1
    assert any(
        "malformed or unsupported filter, sort, or visibility" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF036"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("16385" not in artifact for artifact in rendered_artifacts)


def test_filter_visibility_free_workbook_has_no_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.filter_visibility_controls.present is False
    assert snapshot.filter_visibility_controls.worksheet_auto_filter_count == 0
    assert not any(
        "filter, sort, or visibility" in warning
        for warning in snapshot.parser_warnings
    )


def test_number_format_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_number_format_model(tmp_path / "baseline.xlsx")
    candidate = make_number_format_model(tmp_path / "candidate.xlsx")
    change_number_format_code(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "number_format_controls_changed"
    )

    assert baseline_snapshot.summary()["number_format_assignment_count"] == 6
    assert baseline_snapshot.summary()["number_format_custom_assignment_count"] == 5
    assert baseline_snapshot.summary()["has_number_format_controls"] is True
    assert profile["number_format_controls"] == {
        "present": True,
        "default_format_override_count": 0,
        "cell_format_assignment_count": 3,
        "row_format_assignment_count": 1,
        "column_format_assignment_count": 2,
        "built_in_format_assignment_count": 1,
        "custom_format_assignment_count": 5,
        "unrecognized_number_format_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Cell number-format controls" in markdown
    assert change.details["number_format_definition_material_changed"] is True
    assert "FF039" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-BASELINE-NUMBER-FORMAT",
        "CANDIDATE-PRIVATE-NUMBER-FORMAT",
        ";;;",
        "B2",
        "D:E",
        "numFmtId",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_number_format_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_number_format_model(tmp_path / "baseline.xlsx")
    equivalent = make_number_format_model(tmp_path / "equivalent.xlsx")
    normalize_number_format_control_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert "number_format_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF039" not in {finding.rule_id for finding in report.findings}


def test_number_format_xf_inheritance_and_apply_flag_are_normalized(tmp_path) -> None:
    baseline = make_number_format_model(tmp_path / "baseline.xlsx")
    equivalent = make_number_format_model(tmp_path / "equivalent.xlsx")
    normalize_number_format_inheritance(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert "number_format_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF039" not in {finding.rule_id for finding in report.findings}


def test_number_format_default_style_change_is_guarded(tmp_path) -> None:
    baseline = make_number_format_model(tmp_path / "baseline.xlsx")
    candidate = make_number_format_model(tmp_path / "candidate.xlsx")
    change_number_format_default_style(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.number_format_controls.default_format_override_count == 1
    assert candidate_snapshot.summary()["number_format_assignment_count"] == 6
    assert "FF039" in {finding.rule_id for finding in report.findings}


def test_number_format_malformed_column_control_fails_closed(tmp_path) -> None:
    baseline = make_number_format_model(tmp_path / "baseline.xlsx")
    malformed = make_number_format_model(tmp_path / "malformed.xlsx")
    corrupt_number_format_column_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.number_format_controls.unrecognized_number_format_count == 1
    assert any(
        "malformed or unsupported cell number-format" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF039"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("16385" not in artifact for artifact in rendered_artifacts)


def test_number_format_missing_definition_fails_closed(tmp_path) -> None:
    baseline = make_number_format_model(tmp_path / "baseline.xlsx")
    malformed = make_number_format_model(tmp_path / "malformed.xlsx")
    corrupt_number_format_definition(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.number_format_controls.unrecognized_number_format_count == 1
    assert any(
        "malformed or unsupported cell number-format" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF039"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("999987" not in artifact for artifact in rendered_artifacts)


def test_number_format_free_workbook_has_no_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.number_format_controls.present is False
    assert snapshot.number_format_controls.cell_format_assignment_count == 0
    assert not any("number-format" in warning for warning in snapshot.parser_warnings)


def test_font_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_font_model(tmp_path / "baseline.xlsx")
    candidate = make_font_model(tmp_path / "candidate.xlsx")
    change_font_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(change for change in report.changes if change.kind == "font_controls_changed")

    assert baseline_snapshot.summary()["font_assignment_count"] == 6
    assert baseline_snapshot.summary()["has_font_controls"] is True
    assert profile["font_controls"] == {
        "present": True,
        "default_font_definition_count": 1,
        "cell_font_assignment_count": 2,
        "row_font_assignment_count": 1,
        "column_font_assignment_count": 2,
        "unrecognized_font_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Cell font controls" in markdown
    assert change.details["font_definition_material_changed"] is True
    assert "FF040" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-BASELINE-FONT",
        "PRIVATE-WHITE-FONT",
        "PRIVATE-COLUMN-FONT",
        "FF112233",
        "FFFFFFFF",
        "B2",
        "D:E",
        "fontId",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_font_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_font_model(tmp_path / "baseline.xlsx")
    equivalent = make_font_model(tmp_path / "equivalent.xlsx")
    normalize_font_control_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []
    assert "font_controls_changed" not in {change.kind for change in report.changes}
    assert "FF040" not in {finding.rule_id for finding in report.findings}


def test_font_xf_inheritance_and_apply_flag_are_normalized(tmp_path) -> None:
    baseline = make_font_model(tmp_path / "baseline.xlsx")
    equivalent = make_font_model(tmp_path / "equivalent.xlsx")
    normalize_font_inheritance(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []
    assert "font_controls_changed" not in {change.kind for change in report.changes}
    assert "FF040" not in {finding.rule_id for finding in report.findings}


def test_default_font_definition_change_is_guarded(tmp_path) -> None:
    baseline = make_font_model(tmp_path / "baseline.xlsx")
    candidate = make_font_model(tmp_path / "candidate.xlsx")
    change_default_font_definition(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.font_controls.default_font_definition_count == 1
    assert candidate_snapshot.summary()["font_assignment_count"] == 6
    assert "FF040" in {finding.rule_id for finding in report.findings}


def test_font_malformed_column_control_fails_closed(tmp_path) -> None:
    baseline = make_font_model(tmp_path / "baseline.xlsx")
    malformed = make_font_model(tmp_path / "malformed.xlsx")
    corrupt_font_column_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.font_controls.unrecognized_font_count == 1
    assert any(
        "malformed or unsupported cell-font" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF040"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("16385" not in artifact for artifact in rendered_artifacts)


def test_font_missing_definition_fails_closed(tmp_path) -> None:
    baseline = make_font_model(tmp_path / "baseline.xlsx")
    malformed = make_font_model(tmp_path / "malformed.xlsx")
    corrupt_font_definition(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.font_controls.unrecognized_font_count == 1
    assert any(
        "malformed or unsupported cell-font" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF040"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("999999999" not in artifact for artifact in rendered_artifacts)


def test_ordinary_workbook_has_only_a_default_font_definition(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.font_controls.present is True
    assert snapshot.font_controls.default_font_definition_count == 1
    assert snapshot.font_controls.cell_font_assignment_count == 0
    assert not any("cell-font" in warning for warning in snapshot.parser_warnings)


def test_fill_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_fill_model(tmp_path / "baseline.xlsx")
    candidate = make_fill_model(tmp_path / "candidate.xlsx")
    change_fill_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(change for change in report.changes if change.kind == "fill_controls_changed")

    assert baseline_snapshot.summary()["fill_assignment_count"] == 6
    assert baseline_snapshot.summary()["has_fill_controls"] is True
    assert profile["fill_controls"] == {
        "present": True,
        "default_fill_definition_count": 0,
        "cell_fill_assignment_count": 3,
        "row_fill_assignment_count": 1,
        "column_fill_assignment_count": 2,
        "unrecognized_fill_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Cell fill controls" in markdown
    assert change.details["fill_definition_material_changed"] is True
    assert "FF041" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "FF112233",
        "FF445566",
        "FF99AABB",
        "FFCCDDEE",
        "FF102030",
        "B2",
        "D:E",
        "fillId",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_fill_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_fill_model(tmp_path / "baseline.xlsx")
    equivalent = make_fill_model(tmp_path / "equivalent.xlsx")
    normalize_fill_control_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []
    assert "fill_controls_changed" not in {change.kind for change in report.changes}
    assert "FF041" not in {finding.rule_id for finding in report.findings}


def test_inert_fill_pattern_declarations_are_normalized(tmp_path) -> None:
    baseline = make_fill_model(tmp_path / "baseline.xlsx")
    equivalent = make_fill_model(tmp_path / "equivalent.xlsx")
    normalize_fill_inert_pattern_declarations(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []
    assert "fill_controls_changed" not in {change.kind for change in report.changes}
    assert "FF041" not in {finding.rule_id for finding in report.findings}


def test_gradient_fill_change_is_guarded(tmp_path) -> None:
    baseline = make_fill_model(tmp_path / "baseline.xlsx")
    candidate = make_fill_model(tmp_path / "candidate.xlsx")
    change_gradient_fill_definition(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "FF041" in {finding.rule_id for finding in report.findings}


def test_fill_xf_inheritance_and_apply_flag_are_normalized(tmp_path) -> None:
    baseline = make_fill_model(tmp_path / "baseline.xlsx")
    equivalent = make_fill_model(tmp_path / "equivalent.xlsx")
    normalize_fill_inheritance(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []
    assert "fill_controls_changed" not in {change.kind for change in report.changes}
    assert "FF041" not in {finding.rule_id for finding in report.findings}


def test_default_fill_definition_change_is_guarded(tmp_path) -> None:
    baseline = make_fill_model(tmp_path / "baseline.xlsx")
    candidate = make_fill_model(tmp_path / "candidate.xlsx")
    change_default_fill_definition(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.fill_controls.default_fill_definition_count == 1
    assert "FF041" in {finding.rule_id for finding in report.findings}


def test_fill_malformed_column_control_fails_closed(tmp_path) -> None:
    baseline = make_fill_model(tmp_path / "baseline.xlsx")
    malformed = make_fill_model(tmp_path / "malformed.xlsx")
    corrupt_fill_column_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.fill_controls.unrecognized_fill_count == 1
    assert any(
        "malformed or unsupported cell-fill" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF041"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("16385" not in artifact for artifact in rendered_artifacts)


def test_fill_missing_definition_fails_closed(tmp_path) -> None:
    baseline = make_fill_model(tmp_path / "baseline.xlsx")
    malformed = make_fill_model(tmp_path / "malformed.xlsx")
    corrupt_fill_definition(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.fill_controls.unrecognized_fill_count == 1
    assert any(
        "malformed or unsupported cell-fill" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF041"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("999999999" not in artifact for artifact in rendered_artifacts)


def test_ordinary_workbook_has_no_fill_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.fill_controls.present is False
    assert snapshot.fill_controls.cell_fill_assignment_count == 0
    assert not any("cell-fill" in warning for warning in snapshot.parser_warnings)


def test_alignment_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_alignment_model(tmp_path / "baseline.xlsx")
    candidate = make_alignment_model(tmp_path / "candidate.xlsx")
    change_alignment_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "cell_alignment_controls_changed"
    )

    assert baseline_snapshot.cells == candidate_snapshot.cells
    assert baseline_snapshot.summary()["alignment_assignment_count"] == 5
    assert baseline_snapshot.summary()["has_alignment_controls"] is True
    assert profile["alignment_controls"] == {
        "present": True,
        "default_alignment_definition_count": 0,
        "cell_alignment_assignment_count": 2,
        "row_alignment_assignment_count": 1,
        "column_alignment_assignment_count": 2,
        "unrecognized_alignment_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Cell alignment controls" in markdown
    assert change.details["alignment_definition_material_changed"] is True
    assert "FF054" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-ROTATED-REVIEW-TEXT",
        "B2",
        "D:E",
        "textRotation",
        "relativeIndent",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_alignment_writer_noise_and_xf_inheritance_are_normalized(tmp_path) -> None:
    baseline = make_alignment_model(tmp_path / "baseline.xlsx")
    spelling_equivalent = make_alignment_model(tmp_path / "spelling-equivalent.xlsx")
    inheritance_equivalent = make_alignment_model(
        tmp_path / "inheritance-equivalent.xlsx"
    )
    normalize_alignment_control_spelling(spelling_equivalent)
    normalize_alignment_inheritance(inheritance_equivalent)

    spelling_report = compare_snapshots(
        load_snapshot(baseline),
        load_snapshot(spelling_equivalent),
    )
    inheritance_report = compare_snapshots(
        load_snapshot(baseline),
        load_snapshot(inheritance_equivalent),
    )

    assert spelling_report.changes == []
    assert spelling_report.findings == []
    assert inheritance_report.changes == []
    assert inheritance_report.findings == []


def test_default_alignment_definition_change_is_guarded(tmp_path) -> None:
    baseline = make_alignment_model(tmp_path / "baseline.xlsx")
    candidate = make_alignment_model(tmp_path / "candidate.xlsx")
    change_default_alignment_definition(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.alignment_controls.default_alignment_definition_count == 1
    assert "FF054" in {finding.rule_id for finding in report.findings}


def test_alignment_malformed_controls_fail_closed_and_redact_values(tmp_path) -> None:
    baseline = make_alignment_model(tmp_path / "baseline.xlsx")
    malformed_definition = make_alignment_model(tmp_path / "malformed-definition.xlsx")
    malformed_column = make_alignment_model(tmp_path / "malformed-column.xlsx")
    corrupt_alignment_definition(malformed_definition)
    corrupt_alignment_column_control(malformed_column)

    definition_snapshot = load_snapshot(malformed_definition)
    column_snapshot = load_snapshot(malformed_column)
    definition_report = compare_snapshots(
        load_snapshot(baseline),
        definition_snapshot,
    )
    column_report = compare_snapshots(load_snapshot(baseline), column_snapshot)

    assert definition_snapshot.alignment_controls.unrecognized_alignment_count == 1
    assert column_snapshot.alignment_controls.unrecognized_alignment_count == 1
    assert any(
        "malformed or unsupported cell-alignment" in warning
        for warning in definition_snapshot.parser_warnings
    )
    assert any(
        "malformed or unsupported cell-alignment" in warning
        for warning in column_snapshot.parser_warnings
    )
    assert {"FF010", "FF054"} <= {
        finding.rule_id for finding in definition_report.findings
    }
    assert {"FF010", "FF054"} <= {
        finding.rule_id for finding in column_report.findings
    }
    rendered_artifacts = (
        json.dumps(profile_snapshot(definition_snapshot)),
        profile_to_markdown(profile_snapshot(definition_snapshot)),
        json.dumps(definition_report.to_dict()),
        report_to_markdown(definition_report),
        json.dumps(report_to_sarif(definition_report)),
        json.dumps(profile_snapshot(column_snapshot)),
        profile_to_markdown(profile_snapshot(column_snapshot)),
        json.dumps(column_report.to_dict()),
        report_to_markdown(column_report),
        json.dumps(report_to_sarif(column_report)),
    )
    assert all("424242" not in artifact for artifact in rendered_artifacts)
    assert all("16385" not in artifact for artifact in rendered_artifacts)


def test_ordinary_workbook_has_no_alignment_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.alignment_controls.present is False
    assert snapshot.alignment_controls.cell_alignment_assignment_count == 0
    assert not any("cell-alignment" in warning for warning in snapshot.parser_warnings)


def test_border_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_border_model(tmp_path / "baseline.xlsx")
    candidate = make_border_model(tmp_path / "candidate.xlsx")
    change_border_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "cell_border_controls_changed"
    )

    assert baseline_snapshot.cells == candidate_snapshot.cells
    assert baseline_snapshot.summary()["border_assignment_count"] == 5
    assert baseline_snapshot.summary()["has_border_controls"] is True
    assert profile["border_controls"] == {
        "present": True,
        "default_border_definition_count": 0,
        "cell_border_assignment_count": 2,
        "row_border_assignment_count": 1,
        "column_border_assignment_count": 2,
        "unrecognized_border_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Cell border controls" in markdown
    assert change.details["border_definition_material_changed"] is True
    assert "FF057" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-BORDER-REVIEW-TEXT",
        "FF112233",
        "FF445566",
        "FF556677",
        "B2",
        "D:E",
        "borderId",
        "diagonalUp",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_border_writer_noise_and_xf_inheritance_are_normalized(tmp_path) -> None:
    baseline = make_border_model(tmp_path / "baseline.xlsx")
    spelling_equivalent = make_border_model(tmp_path / "spelling-equivalent.xlsx")
    inheritance_equivalent = make_border_model(
        tmp_path / "inheritance-equivalent.xlsx"
    )
    normalize_border_control_spelling(spelling_equivalent)
    normalize_border_inheritance(inheritance_equivalent)

    spelling_report = compare_snapshots(
        load_snapshot(baseline),
        load_snapshot(spelling_equivalent),
    )
    inheritance_report = compare_snapshots(
        load_snapshot(baseline),
        load_snapshot(inheritance_equivalent),
    )

    assert spelling_report.changes == []
    assert spelling_report.findings == []
    assert inheritance_report.changes == []
    assert inheritance_report.findings == []


def test_border_logical_sides_and_inert_declarations_are_handled(tmp_path) -> None:
    baseline = make_border_model(tmp_path / "baseline.xlsx")
    logical_side_candidate = make_border_model(tmp_path / "logical-side.xlsx")
    inert_equivalent = make_border_model(tmp_path / "inert-equivalent.xlsx")
    change_border_logical_start_side(logical_side_candidate)
    normalize_border_inert_declarations(inert_equivalent)

    logical_snapshot = load_snapshot(logical_side_candidate)
    logical_report = compare_snapshots(load_snapshot(baseline), logical_snapshot)
    inert_report = compare_snapshots(
        load_snapshot(baseline),
        load_snapshot(inert_equivalent),
    )

    assert logical_snapshot.border_controls.unrecognized_border_count == 0
    assert not any(
        "ordinary cell-border" in warning
        for warning in logical_snapshot.parser_warnings
    )
    assert {finding.rule_id for finding in logical_report.findings} == {"FF057"}
    assert inert_report.changes == []
    assert inert_report.findings == []


def test_default_border_definition_change_is_guarded(tmp_path) -> None:
    baseline = make_border_model(tmp_path / "baseline.xlsx")
    candidate = make_border_model(tmp_path / "candidate.xlsx")
    change_default_border_definition(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.border_controls.default_border_definition_count == 1
    assert "FF057" in {finding.rule_id for finding in report.findings}


def test_border_malformed_controls_fail_closed_and_redact_values(tmp_path) -> None:
    baseline = make_border_model(tmp_path / "baseline.xlsx")
    malformed_definition = make_border_model(tmp_path / "malformed-definition.xlsx")
    malformed_column = make_border_model(tmp_path / "malformed-column.xlsx")
    corrupt_border_definition(malformed_definition)
    corrupt_border_column_control(malformed_column)

    definition_snapshot = load_snapshot(malformed_definition)
    column_snapshot = load_snapshot(malformed_column)
    definition_report = compare_snapshots(
        load_snapshot(baseline),
        definition_snapshot,
    )
    column_report = compare_snapshots(load_snapshot(baseline), column_snapshot)

    assert definition_snapshot.border_controls.unrecognized_border_count == 1
    assert column_snapshot.border_controls.unrecognized_border_count == 1
    assert any(
        "malformed or unsupported ordinary cell-border" in warning
        for warning in definition_snapshot.parser_warnings
    )
    assert any(
        "malformed or unsupported ordinary cell-border" in warning
        for warning in column_snapshot.parser_warnings
    )
    assert {"FF010", "FF057"} <= {
        finding.rule_id for finding in definition_report.findings
    }
    assert {"FF010", "FF057"} <= {
        finding.rule_id for finding in column_report.findings
    }
    rendered_artifacts = (
        json.dumps(profile_snapshot(definition_snapshot)),
        profile_to_markdown(profile_snapshot(definition_snapshot)),
        json.dumps(definition_report.to_dict()),
        report_to_markdown(definition_report),
        json.dumps(report_to_sarif(definition_report)),
        json.dumps(profile_snapshot(column_snapshot)),
        profile_to_markdown(profile_snapshot(column_snapshot)),
        json.dumps(column_report.to_dict()),
        report_to_markdown(column_report),
        json.dumps(report_to_sarif(column_report)),
    )
    assert all(
        "PRIVATE-INVALID-BORDER-METADATA" not in artifact
        for artifact in rendered_artifacts
    )
    assert all("16385" not in artifact for artifact in rendered_artifacts)


def test_border_preexisting_coverage_gap_stays_distinct(tmp_path) -> None:
    baseline = make_border_model(tmp_path / "baseline.xlsx")
    candidate = make_border_model(tmp_path / "candidate.xlsx")
    change_border_definition(candidate)
    corrupt_border_definition(baseline)
    corrupt_border_definition(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "cell_border_controls_changed"
    )

    assert "FF057" in {finding.rule_id for finding in report.findings}
    assert "unrecognized_border_metadata_changed" not in change.details


def test_ordinary_workbook_has_no_border_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.border_controls.present is False
    assert snapshot.border_controls.cell_border_assignment_count == 0
    assert not any("cell-border" in warning for warning in snapshot.parser_warnings)


def test_strict_worksheet_border_assignments_are_supported(tmp_path) -> None:
    baseline = make_strict_border_model(tmp_path / "baseline.xlsx")
    candidate = make_strict_border_model(tmp_path / "candidate.xlsx")
    change_border_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.border_controls.cell_border_assignment_count == 2
    assert baseline_snapshot.border_controls.row_border_assignment_count == 1
    assert baseline_snapshot.border_controls.column_border_assignment_count == 2
    assert baseline_snapshot.border_controls.unrecognized_border_count == 0
    assert not any(
        "ordinary cell-border" in warning
        for warning in baseline_snapshot.parser_warnings
    )
    assert "FF057" in {finding.rule_id for finding in report.findings}


def test_worksheet_display_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_worksheet_display_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_display_model(tmp_path / "candidate.xlsx")
    change_worksheet_display_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_display_controls_changed"
    )

    assert baseline_snapshot.cells == candidate_snapshot.cells
    assert baseline_snapshot.summary()["worksheet_display_control_count"] == 11
    assert baseline_snapshot.summary()["has_worksheet_display_controls"] is True
    assert profile["worksheet_display_controls"] == {
        "present": True,
        "zero_hidden_view_count": 1,
        "formula_view_count": 1,
        "gridlines_hidden_view_count": 1,
        "custom_gridline_color_view_count": 1,
        "headers_hidden_view_count": 1,
        "outline_symbols_hidden_view_count": 1,
        "ruler_hidden_view_count": 1,
        "white_space_hidden_view_count": 1,
        "right_to_left_view_count": 1,
        "non_normal_view_count": 1,
        "split_or_frozen_pane_count": 1,
        "unrecognized_display_control_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Worksheet display controls" in markdown
    assert change.details["worksheet_display_definition_material_changed"] is True
    assert "FF055" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-DISPLAY-HEADER",
        "PRIVATE-DISPLAY-FOCUS",
        "C4",
        "showZeros",
        "defaultGridColor",
        "colorId",
        "showRuler",
        "showWhiteSpace",
        "topLeftCell",
        "xSplit",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_display_control_noise_is_normalized(tmp_path) -> None:
    baseline = make_worksheet_display_model(tmp_path / "baseline.xlsx")
    equivalent = make_worksheet_display_model(tmp_path / "equivalent.xlsx")
    normalize_worksheet_display_control_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []


def test_worksheet_display_malformed_controls_fail_closed_and_redact_values(
    tmp_path,
) -> None:
    baseline = make_worksheet_display_model(tmp_path / "baseline.xlsx")
    malformed = make_worksheet_display_model(tmp_path / "malformed.xlsx")
    corrupt_worksheet_display_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.worksheet_display_controls.unrecognized_display_control_count == 1
    assert any(
        "malformed or unsupported worksheet-display" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF055"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(malformed_snapshot)),
        profile_to_markdown(profile_snapshot(malformed_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("-987.5" not in artifact for artifact in rendered_artifacts)


def test_ordinary_workbook_has_no_worksheet_display_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.worksheet_display_controls.present is False
    assert snapshot.worksheet_display_controls.zero_hidden_view_count == 0
    assert not any(
        "worksheet-display" in warning for warning in snapshot.parser_warnings
    )


def test_strict_worksheet_display_controls_are_supported(tmp_path) -> None:
    baseline = make_strict_worksheet_display_model(tmp_path / "baseline.xlsx")
    candidate = make_strict_worksheet_display_model(tmp_path / "candidate.xlsx")
    change_strict_worksheet_display_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.worksheet_display_controls.zero_hidden_view_count == 1
    assert baseline_snapshot.worksheet_display_controls.unrecognized_display_control_count == 0
    assert not any(
        "worksheet-display" in warning
        for warning in baseline_snapshot.parser_warnings
    )
    assert "FF055" in {finding.rule_id for finding in report.findings}


def test_worksheet_print_layout_controls_are_profiled_diffed_and_redacted(
    tmp_path,
) -> None:
    baseline = make_worksheet_print_layout_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_print_layout_model(tmp_path / "candidate.xlsx")
    change_worksheet_print_layout_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_print_layout_controls_changed"
    )

    assert baseline_snapshot.cells == candidate_snapshot.cells
    assert baseline_snapshot.summary()["worksheet_print_layout_control_count"] == 11
    assert baseline_snapshot.summary()["has_worksheet_print_layout_controls"] is True
    assert profile["worksheet_print_layout_controls"] == {
        "present": True,
        "print_area_definition_count": 1,
        "print_title_definition_count": 1,
        "print_gridlines_sheet_count": 1,
        "print_headings_sheet_count": 1,
        "horizontally_centered_print_sheet_count": 1,
        "vertically_centered_print_sheet_count": 1,
        "page_margin_sheet_count": 1,
        "page_setup_sheet_count": 1,
        "header_footer_sheet_count": 1,
        "manual_row_page_break_count": 1,
        "manual_column_page_break_count": 1,
        "unrecognized_print_layout_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Worksheet print-layout controls" in markdown
    assert change.details["worksheet_print_layout_definition_material_changed"] is True
    assert [change.kind for change in report.changes] == [
        "worksheet_print_layout_controls_changed"
    ]
    assert {finding.rule_id for finding in report.findings} == {"FF056"}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-PRINT-HEADER",
        "PRIVATE-PRINT-FOOTER",
        "PRIVATE-PRINT-HEADER-TEXT",
        "PRIVATE-PRINT-HEADER-CANDIDATE",
        "_xlnm.Print_Area",
        "pageMargins",
        "pageSetup",
        "297mm",
        "16383",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_print_layout_control_noise_is_normalized(tmp_path) -> None:
    baseline = make_worksheet_print_layout_model(tmp_path / "baseline.xlsx")
    equivalent = make_worksheet_print_layout_model(tmp_path / "equivalent.xlsx")
    normalize_worksheet_print_layout_control_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []


def test_worksheet_print_layout_inert_controls_are_normalized(tmp_path) -> None:
    baseline = make_worksheet_print_layout_model(tmp_path / "baseline.xlsx")
    equivalent = make_worksheet_print_layout_model(tmp_path / "equivalent.xlsx")
    normalize_worksheet_print_layout_inert_controls(
        baseline,
        automatic_break_id=11,
        fit_to_height=3,
        fit_to_page=True,
        fit_to_width=2,
        first_page_number=7,
        scale=80,
        show_auto_page_breaks=True,
        inactive_header_suffix="BASELINE",
    )
    normalize_worksheet_print_layout_inert_controls(
        equivalent,
        automatic_break_id=101,
        fit_to_height=3,
        fit_to_page=True,
        fit_to_width=2,
        first_page_number=500,
        scale=160,
        show_auto_page_breaks=False,
        inactive_header_suffix="CANDIDATE",
    )

    baseline_snapshot = load_snapshot(baseline)
    equivalent_snapshot = load_snapshot(equivalent)
    report = compare_snapshots(baseline_snapshot, equivalent_snapshot)

    assert (
        baseline_snapshot.worksheet_print_layout_controls
        == equivalent_snapshot.worksheet_print_layout_controls
    )
    assert report.changes == []
    assert report.findings == []


def test_worksheet_print_layout_disabled_fit_dimensions_are_normalized(tmp_path) -> None:
    baseline = make_worksheet_print_layout_model(tmp_path / "baseline.xlsx")
    equivalent = make_worksheet_print_layout_model(tmp_path / "equivalent.xlsx")
    normalize_worksheet_print_layout_inert_controls(
        baseline,
        automatic_break_id=11,
        fit_to_height=3,
        fit_to_page=False,
        fit_to_width=2,
        first_page_number=7,
        scale=80,
        show_auto_page_breaks=True,
        inactive_header_suffix="BASELINE",
    )
    normalize_worksheet_print_layout_inert_controls(
        equivalent,
        automatic_break_id=101,
        fit_to_height=9,
        fit_to_page=False,
        fit_to_width=8,
        first_page_number=500,
        scale=80,
        show_auto_page_breaks=False,
        inactive_header_suffix="CANDIDATE",
    )

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []


def test_worksheet_print_layout_malformed_controls_fail_closed_and_redact_values(
    tmp_path,
) -> None:
    baseline = make_worksheet_print_layout_model(tmp_path / "baseline.xlsx")
    malformed = make_worksheet_print_layout_model(tmp_path / "malformed.xlsx")
    corrupt_worksheet_print_layout_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert (
        malformed_snapshot.worksheet_print_layout_controls.unrecognized_print_layout_count
        == 1
    )
    assert any(
        "malformed or unsupported worksheet print-layout" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF056"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(malformed_snapshot)),
        profile_to_markdown(profile_snapshot(malformed_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("987654321" not in artifact for artifact in rendered_artifacts)


def test_worksheet_print_layout_preexisting_coverage_gap_stays_distinct(tmp_path) -> None:
    baseline = make_worksheet_print_layout_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_print_layout_model(tmp_path / "candidate.xlsx")
    change_worksheet_print_layout_controls(candidate)
    corrupt_worksheet_print_layout_control(baseline)
    corrupt_worksheet_print_layout_control(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_print_layout_controls_changed"
    )

    assert "FF056" in {finding.rule_id for finding in report.findings}
    assert "unrecognized_worksheet_print_layout_metadata_changed" not in change.details


def test_strict_worksheet_print_layout_controls_are_supported(tmp_path) -> None:
    baseline = make_strict_worksheet_print_layout_model(tmp_path / "baseline.xlsx")
    candidate = make_strict_worksheet_print_layout_model(tmp_path / "candidate.xlsx")
    change_strict_worksheet_print_layout_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.worksheet_print_layout_controls.print_area_definition_count == 1
    assert baseline_snapshot.worksheet_print_layout_controls.page_setup_sheet_count == 1
    assert (
        baseline_snapshot.worksheet_print_layout_controls.unrecognized_print_layout_count
        == 0
    )
    assert not any(
        "worksheet print-layout" in warning
        for warning in baseline_snapshot.parser_warnings
    )
    assert "FF056" in {finding.rule_id for finding in report.findings}


def test_workbook_themes_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_workbook_theme_image_model(tmp_path / "baseline.xlsx")
    definition_candidate = make_workbook_theme_image_model(
        tmp_path / "definition-candidate.xlsx"
    )
    image_candidate = make_workbook_theme_image_model(
        tmp_path / "image-candidate.xlsx"
    )
    normalised = make_workbook_theme_image_model(tmp_path / "normalised.xlsx")
    change_workbook_theme_colour(definition_candidate)
    change_workbook_theme_image_payload(image_candidate)
    normalize_workbook_theme_relationship_identifiers(normalised)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    definition_snapshot = load_snapshot(definition_candidate)
    image_snapshot = load_snapshot(image_candidate)
    definition_report = compare_snapshots(baseline_snapshot, definition_snapshot)
    image_report = compare_snapshots(baseline_snapshot, image_snapshot)
    normalised_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(normalised),
    )
    definition_change = next(
        change
        for change in definition_report.changes
        if change.kind == "workbook_theme_changed"
    )
    image_change = next(
        change
        for change in image_report.changes
        if change.kind == "workbook_theme_changed"
    )

    assert baseline_snapshot.cells == definition_snapshot.cells
    assert baseline_snapshot.cells == image_snapshot.cells
    assert baseline_snapshot.summary()["workbook_theme_part_count"] == 1
    assert baseline_snapshot.summary()["workbook_theme_image_part_count"] == 1
    assert baseline_snapshot.summary()["has_workbook_theme"] is True
    assert profile["workbook_theme"] == {
        "present": True,
        "theme_part_count": 1,
        "colour_scheme_count": 1,
        "font_scheme_count": 1,
        "format_scheme_count": 1,
        "theme_relationship_count": 1,
        "external_theme_relationship_count": 0,
        "theme_image_part_count": 1,
        "theme_image_relationship_count": 1,
        "external_theme_image_relationship_count": 0,
        "unrecognized_theme_count": 0,
    }
    assert "## Workbook theme controls" in markdown
    assert definition_change.details["theme_definition_material_changed"] is True
    assert image_change.details["theme_definition_material_changed"] is True
    assert image_change.details["theme_image_payload_changed"] is True
    assert "FF053" in {finding.rule_id for finding in definition_report.findings}
    assert "FF053" in {finding.rule_id for finding in image_report.findings}
    assert "workbook_theme_changed" not in {
        change.kind for change in normalised_report.changes
    }
    assert "FF053" not in {
        finding.rule_id for finding in normalised_report.findings
    }

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(definition_report.to_dict()),
        report_to_markdown(definition_report),
        json.dumps(report_to_sarif(definition_report)),
        json.dumps(image_report.to_dict()),
        report_to_markdown(image_report),
        json.dumps(report_to_sarif(image_report)),
    )
    for sensitive_value in (
        "PRIVATE-THEME-SCHEME-BASELINE",
        "PRIVATE-THEME-IMAGE-BASELINE",
        "PRIVATE-THEME-IMAGE-CANDIDATE",
        "rIdFenceThemeImage",
        "C00000",
        "fence-theme-image.bin",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_strict_workbook_theme_image_controls_are_supported(tmp_path) -> None:
    baseline = make_strict_workbook_theme_image_model(tmp_path / "baseline.xlsx")
    image_candidate = make_strict_workbook_theme_image_model(
        tmp_path / "image-candidate.xlsx"
    )
    normalised = make_strict_workbook_theme_image_model(tmp_path / "normalised.xlsx")
    change_workbook_theme_image_payload(image_candidate)
    normalize_workbook_theme_relationship_identifiers(normalised)

    baseline_snapshot = load_snapshot(baseline)
    image_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(image_candidate),
    )
    normalised_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(normalised),
    )
    image_change = next(
        change
        for change in image_report.changes
        if change.kind == "workbook_theme_changed"
    )

    assert baseline_snapshot.workbook_theme.theme_image_part_count == 1
    assert baseline_snapshot.workbook_theme.theme_image_relationship_count == 1
    assert baseline_snapshot.workbook_theme.unrecognized_theme_count == 0
    assert image_change.details["theme_image_payload_changed"] is True
    assert "FF053" in {finding.rule_id for finding in image_report.findings}
    assert "workbook_theme_changed" not in {
        change.kind for change in normalised_report.changes
    }


def test_workbook_theme_metadata_fails_closed_and_is_redacted(tmp_path) -> None:
    baseline = make_workbook_theme_image_model(tmp_path / "baseline.xlsx")
    candidate = make_workbook_theme_image_model(tmp_path / "candidate.xlsx")
    corrupt_workbook_theme_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "workbook_theme_changed"
    )

    assert candidate_snapshot.workbook_theme.unrecognized_theme_count >= 1
    assert any(
        "malformed, unsupported, or incomplete workbook-theme metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert change.details["theme_definition_material_changed"] is True
    assert change.details["unrecognized_theme_metadata_changed"] is True
    assert {"FF010", "FF053"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "privateUnexpectedWorkbookTheme",
        "PRIVATE-THEME-SCHEME-BASELINE",
        "PRIVATE-THEME-IMAGE-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_workbook_theme_read_budget_fails_closed(tmp_path, monkeypatch) -> None:
    baseline = make_workbook_theme_image_model(tmp_path / "baseline.xlsx")
    candidate = make_workbook_theme_image_model(tmp_path / "candidate.xlsx")
    baseline_snapshot = load_snapshot(baseline)
    monkeypatch.setattr(workbook_module, "_WORKBOOK_THEME_MAX_PART_BYTES", 1)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert candidate_snapshot.workbook_theme.unrecognized_theme_count >= 1
    assert any(
        "oversized workbook-theme part" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert {"FF010", "FF053"} <= {finding.rule_id for finding in report.findings}


def test_formula_cached_results_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_formula_cached_result_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_cached_result_model(tmp_path / "candidate.xlsx")
    change_formula_cached_result(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "formula_cached_result_changed"
    )

    assert baseline_snapshot.summary()["formula_cached_result_cell_count"] == 4
    assert baseline_snapshot.summary()["formula_missing_cached_result_cell_count"] == 1
    assert baseline_snapshot.summary()["has_formula_cached_results"] is True
    assert profile["formula_cached_results"] == {
        "present": True,
        "formula_cell_count": 5,
        "cached_result_cell_count": 4,
        "missing_cached_result_cell_count": 1,
        "numeric_cached_result_count": 1,
        "string_cached_result_count": 1,
        "boolean_cached_result_count": 1,
        "error_cached_result_count": 1,
        "unrecognized_cached_result_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Stored formula results" in markdown
    assert change.details["unexplained_cached_result_change_count"] == 1
    assert change.details["cached_result_material_changed"] is True
    assert "FF042" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-CACHED-STRING",
        "999999",
        "#DIV/0!",
        "Report!B2",
        "B2",
        "<v>",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_formula_cached_results_recalculated_from_visible_precedent_stay_quiet(
    tmp_path,
) -> None:
    baseline = make_formula_cached_result_model(tmp_path / "baseline.xlsx")
    candidate = make_formula_cached_result_model(tmp_path / "candidate.xlsx")
    change_formula_cached_result_with_visible_precedent(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert {(change.kind, change.location) for change in report.changes} == {
        ("value_changed", ("Inputs", "A1"))
    }
    assert "formula_cached_result_changed" not in {change.kind for change in report.changes}
    assert "FF042" not in {finding.rule_id for finding in report.findings}


def test_formula_cached_result_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_formula_cached_result_model(tmp_path / "baseline.xlsx")
    equivalent = make_formula_cached_result_model(tmp_path / "equivalent.xlsx")
    normalize_formula_cached_result_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []


def test_formula_cached_result_malformed_metadata_fails_closed(tmp_path) -> None:
    baseline = make_formula_cached_result_model(tmp_path / "baseline.xlsx")
    malformed = make_formula_cached_result_model(tmp_path / "malformed.xlsx")
    corrupt_formula_cached_result(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.formula_cached_results.unrecognized_cached_result_count == 1
    assert any(
        "malformed or unsupported formula cached-result" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF042"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in ("PRIVATE-CACHED-STRING", "PRIVATE-NOT-A-NUMBER", "B2"):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_rich_text_run_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_rich_text_run_model(tmp_path / "baseline.xlsx")
    candidate = make_rich_text_run_model(tmp_path / "candidate.xlsx")
    inline_candidate = make_rich_text_run_model(tmp_path / "inline-candidate.xlsx")
    change_rich_text_run_color(candidate)
    change_inline_rich_text_run_color(inline_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    inline_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(inline_candidate),
    )
    change = next(
        change
        for change in report.changes
        if change.kind == "rich_text_run_controls_changed"
    )

    assert baseline_snapshot.summary()["rich_text_cell_count"] == 2
    assert baseline_snapshot.summary()["rich_text_run_count"] == 4
    assert baseline_snapshot.summary()["has_rich_text_runs"] is True
    assert profile["rich_text_runs"] == {
        "present": True,
        "shared_rich_text_item_count": 1,
        "shared_rich_text_cell_count": 1,
        "shared_rich_text_run_count": 2,
        "inline_rich_text_cell_count": 1,
        "inline_rich_text_run_count": 2,
        "phonetic_run_count": 0,
        "phonetic_property_count": 0,
        "unrecognized_rich_text_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Rich-text run controls" in markdown
    assert change.details["rich_text_run_control_change_count"] == 1
    assert change.details["rich_text_run_definition_material_changed"] is True
    assert "FF043" in {finding.rule_id for finding in report.findings}
    assert "FF043" in {finding.rule_id for finding in inline_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
        json.dumps(inline_report.to_dict()),
        report_to_markdown(inline_report),
        json.dumps(report_to_sarif(inline_report)),
    )
    for sensitive_value in (
        "PRIVATE-INLINE-HOLD",
        "PRIVATE-RICH-RUN-FONT",
        "FF000000",
        "FF334455",
        "FFFFFFFF",
        "Review!A3",
        "A3",
        "B3",
        "rPr",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_rich_text_run_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_rich_text_run_model(tmp_path / "baseline.xlsx")
    equivalent = make_rich_text_run_model(tmp_path / "equivalent.xlsx")
    normalize_rich_text_run_property_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []


def test_equivalent_shared_and_inline_rich_text_storage_stays_quiet(tmp_path) -> None:
    baseline = make_rich_text_run_model(tmp_path / "baseline.xlsx")
    equivalent = make_rich_text_run_model(tmp_path / "equivalent.xlsx")
    rewrite_shared_rich_text_as_inline(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []


def test_rich_text_run_style_boundary_changes_are_guarded(tmp_path) -> None:
    baseline = make_rich_text_run_model(tmp_path / "baseline.xlsx")
    candidate = make_rich_text_run_model(tmp_path / "candidate.xlsx")
    change_rich_text_run_boundary(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert {(change.kind, change.location) for change in report.changes} == {
        ("rich_text_run_controls_changed", None)
    }
    assert "FF043" in {finding.rule_id for finding in report.findings}


def test_rich_text_run_text_only_edits_stay_normal_cell_changes(tmp_path) -> None:
    baseline = make_rich_text_run_model(tmp_path / "baseline.xlsx")
    candidate = make_rich_text_run_model(tmp_path / "candidate.xlsx")
    change_rich_text_run_text_only(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert ("value_changed", ("Review", "A3")) in {
        (change.kind, change.location) for change in report.changes
    }
    assert "rich_text_run_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF043" not in {finding.rule_id for finding in report.findings}


def test_rich_text_run_malformed_metadata_fails_closed(tmp_path) -> None:
    baseline = make_rich_text_run_model(tmp_path / "baseline.xlsx")
    malformed = make_rich_text_run_model(tmp_path / "malformed.xlsx")
    corrupt_rich_text_run(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.rich_text_runs.unrecognized_rich_text_count == 1
    assert any(
        "malformed or unsupported rich-text run" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF043"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in ("PRIVATE-UNSUPPORTED-RUN-CONTROL", "A3", "FF000000"):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_cell_hyperlinks_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_cell_hyperlink_model(tmp_path / "baseline.xlsx")
    target_candidate = make_cell_hyperlink_model(tmp_path / "target-candidate.xlsx")
    tooltip_candidate = make_cell_hyperlink_model(tmp_path / "tooltip-candidate.xlsx")
    display_candidate = make_cell_hyperlink_model(tmp_path / "display-candidate.xlsx")
    location_candidate = make_cell_hyperlink_model(tmp_path / "location-candidate.xlsx")
    renumbered = make_cell_hyperlink_model(tmp_path / "renumbered.xlsx")
    change_cell_hyperlink_target(target_candidate)
    change_cell_hyperlink_tooltip(tooltip_candidate)
    change_cell_hyperlink_display(display_candidate)
    change_cell_hyperlink_location(location_candidate)
    renumber_cell_hyperlink_identifiers(renumbered)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    target_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(target_candidate),
    )
    tooltip_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(tooltip_candidate),
    )
    display_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(display_candidate),
    )
    location_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(location_candidate),
    )
    renumbered_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(renumbered),
    )
    target_change = next(
        change
        for change in target_report.changes
        if change.kind == "cell_hyperlink_controls_changed"
    )
    tooltip_change = next(
        change
        for change in tooltip_report.changes
        if change.kind == "cell_hyperlink_controls_changed"
    )
    display_change = next(
        change
        for change in display_report.changes
        if change.kind == "cell_hyperlink_controls_changed"
    )
    location_change = next(
        change
        for change in location_report.changes
        if change.kind == "cell_hyperlink_controls_changed"
    )

    assert baseline_snapshot.summary()["cell_hyperlink_count"] == 2
    assert baseline_snapshot.summary()["cell_hyperlink_external_relationship_count"] == 1
    assert baseline_snapshot.summary()["has_cell_hyperlinks"] is True
    assert profile["cell_hyperlinks"] == {
        "present": True,
        "worksheet_hyperlink_sheet_count": 1,
        "hyperlink_count": 2,
        "hyperlink_with_location_count": 1,
        "hyperlink_with_display_count": 1,
        "hyperlink_with_tooltip_count": 2,
        "binding_relationship_count": 1,
        "external_relationship_count": 1,
        "unrecognized_cell_hyperlink_count": 0,
    }
    assert "## Worksheet cell hyperlinks" in markdown
    assert target_change.details["cell_hyperlink_binding_changed"] is True
    assert target_change.details["cell_hyperlink_relationships_changed"] is True
    assert tooltip_change.details["cell_hyperlink_definition_material_changed"] is True
    assert "cell_hyperlink_relationships_changed" not in tooltip_change.details
    assert display_change.details["cell_hyperlink_definition_material_changed"] is True
    assert "cell_hyperlink_relationships_changed" not in display_change.details
    assert location_change.details["cell_hyperlink_binding_changed"] is True
    assert "cell_hyperlink_relationships_changed" not in location_change.details
    assert "FF047" in {finding.rule_id for finding in target_report.findings}
    assert "FF047" in {finding.rule_id for finding in tooltip_report.findings}
    assert "FF047" in {finding.rule_id for finding in display_report.findings}
    assert "FF047" in {finding.rule_id for finding in location_report.findings}
    assert "cell_hyperlink_controls_changed" not in {
        change.kind for change in renumbered_report.changes
    }
    assert "FF047" not in {finding.rule_id for finding in renumbered_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(target_report.to_dict()),
        report_to_markdown(target_report),
        json.dumps(report_to_sarif(target_report)),
        json.dumps(tooltip_report.to_dict()),
        report_to_markdown(tooltip_report),
        json.dumps(report_to_sarif(tooltip_report)),
        json.dumps(display_report.to_dict()),
        report_to_markdown(display_report),
        json.dumps(report_to_sarif(display_report)),
        json.dumps(location_report.to_dict()),
        report_to_markdown(location_report),
        json.dumps(report_to_sarif(location_report)),
    )
    for sensitive_value in (
        "https://approved.example.test/PRIVATE-LINK-BASELINE",
        "https://review.example.test/PRIVATE-LINK-CANDIDATE",
        "PRIVATE-EXTERNAL-LINK-TOOLTIP",
        "PRIVATE-INTERNAL-LINK-DISPLAY",
        "PRIVATE-INTERNAL-LINK-DISPLAY-CANDIDATE",
        "A1",
        "A2",
        "B2",
        "rIdFenceCellHyperlink",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_revision_cell_hyperlinks_are_guarded_without_reader_loss(tmp_path) -> None:
    baseline = make_cell_hyperlink_model(tmp_path / "baseline.xlsx")
    candidate = make_cell_hyperlink_model(tmp_path / "candidate.xlsx")
    rewrite_cell_hyperlink_as_revision_declaration(baseline)
    rewrite_cell_hyperlink_as_revision_declaration(candidate)
    change_cell_hyperlink_target(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert baseline_snapshot.cell_hyperlinks.hyperlink_count == 2
    assert baseline_snapshot.cell_hyperlinks.unrecognized_cell_hyperlink_count == 0
    assert candidate_snapshot.cell_hyperlinks.unrecognized_cell_hyperlink_count == 0
    assert "cell_hyperlink_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF047" in {finding.rule_id for finding in report.findings}


def test_unbound_cell_hyperlink_relationship_fails_closed_and_is_redacted(
    tmp_path,
) -> None:
    baseline = make_cell_hyperlink_model(tmp_path / "baseline.xlsx")
    candidate = make_cell_hyperlink_model(tmp_path / "candidate.xlsx")
    unbind_cell_hyperlink_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "cell_hyperlink_controls_changed"
    )

    assert candidate_snapshot.cell_hyperlinks.unrecognized_cell_hyperlink_count >= 1
    assert any(
        "malformed, unbound, or unsupported cell-hyperlink metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert change.details["unrecognized_cell_hyperlink_metadata_changed"] is True
    assert "FF047" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "https://approved.example.test/PRIVATE-LINK-BASELINE",
        "PRIVATE-EXTERNAL-LINK-TOOLTIP",
        "A1",
        "rId",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_malformed_cell_hyperlink_reference_is_quarantined_for_reader(tmp_path) -> None:
    baseline = make_cell_hyperlink_model(tmp_path / "baseline.xlsx")
    candidate = make_cell_hyperlink_model(tmp_path / "candidate.xlsx")
    corrupt_cell_hyperlink_reference(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.cell_hyperlinks.unrecognized_cell_hyperlink_count >= 1
    assert any(
        "malformed, unbound, or unsupported cell-hyperlink metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "cell_hyperlink_controls_changed" in {
        change.kind for change in report.changes
    }
    assert {"FF010", "FF047"} <= {finding.rule_id for finding in report.findings}


def test_cell_hyperlink_xml_budget_fails_closed(tmp_path, monkeypatch) -> None:
    baseline = make_cell_hyperlink_model(tmp_path / "baseline.xlsx")
    candidate = make_cell_hyperlink_model(tmp_path / "candidate.xlsx")
    baseline_snapshot = load_snapshot(baseline)
    monkeypatch.setattr(workbook_module, "_CELL_HYPERLINK_MAX_WORKSHEET_XML_BYTES", 1)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert candidate_snapshot.cell_hyperlinks.unrecognized_cell_hyperlink_count >= 1
    assert any(
        "oversized worksheet XML part while inspecting cell hyperlinks" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert {"FF010", "FF047"} <= {finding.rule_id for finding in report.findings}


def test_worksheet_sparklines_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_worksheet_sparkline_model(tmp_path / "baseline.xlsx")
    source_candidate = make_worksheet_sparkline_model(
        tmp_path / "source-candidate.xlsx"
    )
    presentation_candidate = make_worksheet_sparkline_model(
        tmp_path / "presentation-candidate.xlsx"
    )
    reordered = make_worksheet_sparkline_model(tmp_path / "reordered.xlsx")
    change_worksheet_sparkline_source(source_candidate)
    change_worksheet_sparkline_presentation(presentation_candidate)
    reorder_worksheet_sparklines(reordered)
    normalised = make_worksheet_sparkline_model(tmp_path / "normalised.xlsx")
    normalize_worksheet_sparkline_control_spelling(normalised)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    source_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(source_candidate),
    )
    presentation_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(presentation_candidate),
    )
    reordered_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(reordered),
    )
    normalised_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(normalised),
    )
    source_change = next(
        change
        for change in source_report.changes
        if change.kind == "worksheet_sparkline_controls_changed"
    )
    presentation_change = next(
        change
        for change in presentation_report.changes
        if change.kind == "worksheet_sparkline_controls_changed"
    )

    assert baseline_snapshot.summary()["worksheet_sparkline_group_count"] == 1
    assert baseline_snapshot.summary()["worksheet_sparkline_count"] == 2
    assert baseline_snapshot.summary()["has_worksheet_sparklines"] is True
    assert profile["worksheet_sparklines"] == {
        "present": True,
        "worksheet_sparkline_sheet_count": 1,
        "sparkline_group_count": 1,
        "sparkline_count": 2,
        "sparkline_with_source_count": 2,
        "group_date_axis_source_count": 1,
        "color_control_count": 2,
        "unrecognized_worksheet_sparkline_count": 0,
    }
    assert "## Worksheet sparklines" in markdown
    assert (
        "Sparkline Group extension is not supported and will be removed"
        not in baseline_snapshot.parser_warnings
    )
    assert source_change.details["worksheet_sparkline_bindings_changed"] is True
    assert (
        source_change.details["worksheet_sparkline_definition_material_changed"]
        is True
    )
    assert (
        presentation_change.details["worksheet_sparkline_definition_material_changed"]
        is True
    )
    assert "worksheet_sparkline_bindings_changed" not in presentation_change.details
    assert "FF048" in {finding.rule_id for finding in source_report.findings}
    assert "FF048" in {finding.rule_id for finding in presentation_report.findings}
    assert "worksheet_sparkline_controls_changed" not in {
        change.kind for change in reordered_report.changes
    }
    assert "FF048" not in {finding.rule_id for finding in reordered_report.findings}
    assert "worksheet_sparkline_controls_changed" not in {
        change.kind for change in normalised_report.changes
    }
    assert "FF048" not in {finding.rule_id for finding in normalised_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(source_report.to_dict()),
        report_to_markdown(source_report),
        json.dumps(report_to_sarif(source_report)),
        json.dumps(presentation_report.to_dict()),
        report_to_markdown(presentation_report),
        json.dumps(report_to_sarif(presentation_report)),
    )
    for sensitive_value in (
        "Inputs!$A$2:$A$4",
        "Inputs!$B$2:$B$4",
        "Inputs!$B$3:$B$5",
        "$F$1",
        "$F$2",
        "FF112233",
        "FF778899",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_malformed_worksheet_sparkline_metadata_fails_closed_and_is_redacted(
    tmp_path,
) -> None:
    baseline = make_worksheet_sparkline_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_sparkline_model(tmp_path / "candidate.xlsx")
    corrupt_worksheet_sparkline_destination(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_sparkline_controls_changed"
    )

    assert (
        candidate_snapshot.worksheet_sparklines.unrecognized_worksheet_sparkline_count
        >= 1
    )
    assert any(
        "malformed or unsupported worksheet-sparkline metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert (
        change.details["unrecognized_worksheet_sparkline_metadata_changed"] is True
    )
    assert {"FF010", "FF048"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-NOT-A-SPARKLINE-CELL",
        "Inputs!$B$2:$B$4",
        "$F$1",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_sparkline_xml_budget_fails_closed(tmp_path, monkeypatch) -> None:
    baseline = make_worksheet_sparkline_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_sparkline_model(tmp_path / "candidate.xlsx")
    baseline_snapshot = load_snapshot(baseline)
    monkeypatch.setattr(
        workbook_module,
        "_WORKSHEET_SPARKLINE_MAX_WORKSHEET_XML_BYTES",
        1,
    )

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert (
        candidate_snapshot.worksheet_sparklines.unrecognized_worksheet_sparkline_count
        >= 1
    )
    assert any(
        "oversized worksheet XML part while inspecting worksheet sparklines"
        in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert {"FF010", "FF048"} <= {finding.rule_id for finding in report.findings}


def test_sparkline_reader_overlay_preserves_hyperlink_isolation(tmp_path) -> None:
    workbook = make_cell_hyperlink_sparkline_model(tmp_path / "combined.xlsx")

    snapshot = load_snapshot(workbook)

    assert snapshot.cell_hyperlinks.hyperlink_count == 2
    assert snapshot.worksheet_sparklines.sparkline_count == 1
    assert snapshot.cells[("Inputs", "A1")].value == "Open approved source"
    assert (
        "Sparkline Group extension is not supported and will be removed"
        not in snapshot.parser_warnings
    )


def test_xml_mappings_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_xml_mapping_model(tmp_path / "baseline.xlsx")
    xpath_candidate = make_xml_mapping_model(tmp_path / "xpath-candidate.xlsx")
    refresh_candidate = make_xml_mapping_model(tmp_path / "refresh-candidate.xlsx")
    relationship_candidate = make_xml_mapping_model(
        tmp_path / "relationship-candidate.xlsx"
    )
    normalised = make_xml_mapping_model(tmp_path / "normalised.xlsx")
    change_xml_mapping_xpath(xpath_candidate)
    change_xml_mapping_refresh_behavior(refresh_candidate)
    rebind_xml_mapping_relationship(relationship_candidate)
    normalize_xml_mapping_control_spelling(normalised)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    xpath_snapshot = load_snapshot(xpath_candidate)
    refresh_snapshot = load_snapshot(refresh_candidate)
    relationship_snapshot = load_snapshot(relationship_candidate)
    xpath_report = compare_snapshots(baseline_snapshot, xpath_snapshot)
    refresh_report = compare_snapshots(baseline_snapshot, refresh_snapshot)
    relationship_report = compare_snapshots(
        baseline_snapshot,
        relationship_snapshot,
    )
    normalised_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(normalised),
    )
    xpath_change = next(
        change
        for change in xpath_report.changes
        if change.kind == "xml_mapping_controls_changed"
    )
    refresh_change = next(
        change
        for change in refresh_report.changes
        if change.kind == "xml_mapping_controls_changed"
    )
    relationship_change = next(
        change
        for change in relationship_report.changes
        if change.kind == "xml_mapping_controls_changed"
    )

    assert baseline_snapshot.cells == xpath_snapshot.cells
    assert baseline_snapshot.cells == refresh_snapshot.cells
    assert baseline_snapshot.cells == relationship_snapshot.cells
    assert baseline_snapshot.tables == xpath_snapshot.tables
    assert baseline_snapshot.tables == refresh_snapshot.tables
    assert baseline_snapshot.tables == relationship_snapshot.tables
    assert baseline_snapshot.summary()["xml_map_count"] == 1
    assert baseline_snapshot.summary()["xml_map_table_binding_count"] == 1
    assert baseline_snapshot.summary()["xml_map_single_cell_binding_count"] == 1
    assert baseline_snapshot.summary()["has_xml_mapping_controls"] is True
    assert profile["xml_mapping_controls"] == {
        "present": True,
        "xml_map_part_count": 1,
        "xml_schema_count": 1,
        "xml_map_count": 1,
        "xml_map_data_binding_count": 1,
        "xml_map_file_binding_count": 1,
        "xml_map_connection_binding_count": 1,
        "table_xml_binding_part_count": 1,
        "table_xml_binding_count": 1,
        "single_cell_xml_binding_sheet_count": 1,
        "single_cell_xml_binding_part_count": 1,
        "single_cell_xml_binding_count": 1,
        "single_cell_xml_connection_binding_count": 1,
        "unrecognized_xml_mapping_count": 0,
    }
    assert "## XML-mapped workbook controls" in markdown
    assert xpath_change.details["xml_mapping_bindings_changed"] is True
    assert refresh_change.details["xml_mapping_declarations_changed"] is True
    assert relationship_change.details["xml_mapping_relationships_changed"] is True
    assert "FF049" in {finding.rule_id for finding in xpath_report.findings}
    assert "FF049" in {finding.rule_id for finding in refresh_report.findings}
    assert "FF049" in {finding.rule_id for finding in relationship_report.findings}
    assert "xml_mapping_controls_changed" not in {
        change.kind for change in normalised_report.changes
    }
    assert "FF049" not in {finding.rule_id for finding in normalised_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(xpath_report.to_dict()),
        report_to_markdown(xpath_report),
        json.dumps(report_to_sarif(xpath_report)),
        json.dumps(refresh_report.to_dict()),
        report_to_markdown(refresh_report),
        json.dumps(report_to_sarif(refresh_report)),
        json.dumps(relationship_report.to_dict()),
        report_to_markdown(relationship_report),
        json.dumps(report_to_sarif(relationship_report)),
    )
    for sensitive_value in (
        "PRIVATE-XML-SCHEMA",
        "urn:formulafence:test:private",
        "PRIVATE-XML-MAP",
        "PRIVATE-XML-DATA-BINDING",
        "PRIVATE-XML-BINDING-FILE",
        "PRIVATE-XML-SINGLE-CELL",
        "/private:PrivateRoot/private:Record/private:Amount",
        "/private:PrivateRoot/private:Record/private:CandidateAmount",
        "/private:PrivateRoot/private:Header/private:AsOf",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_unsafe_xml_mapping_relationship_fails_closed_and_is_redacted(tmp_path) -> None:
    baseline = make_xml_mapping_model(tmp_path / "baseline.xlsx")
    candidate = make_xml_mapping_model(tmp_path / "candidate.xlsx")
    externalize_xml_mapping_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "xml_mapping_controls_changed"
    )

    assert candidate_snapshot.xml_mapping_controls.unrecognized_xml_mapping_count >= 1
    assert any(
        "malformed or unsupported XML-mapping metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert change.details["xml_mapping_relationships_changed"] is True
    assert change.details["unrecognized_xml_mapping_metadata_changed"] is True
    assert {"FF010", "FF049"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "https://private.example.test/PRIVATE-XML-MAP-RELATIONSHIP",
        "PRIVATE-XML-MAP",
        "/private:PrivateRoot/private:Record/private:Amount",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_malformed_xml_mapping_metadata_fails_closed_and_is_redacted(tmp_path) -> None:
    baseline = make_xml_mapping_model(tmp_path / "baseline.xlsx")
    candidate = make_xml_mapping_model(tmp_path / "candidate.xlsx")
    corrupt_xml_mapping_single_cell_reference(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "xml_mapping_controls_changed"
    )

    assert candidate_snapshot.xml_mapping_controls.unrecognized_xml_mapping_count >= 1
    assert any(
        "malformed or unsupported XML-mapping metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert change.details["unrecognized_xml_mapping_metadata_changed"] is True
    assert {"FF010", "FF049"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-NOT-AN-XML-MAP-CELL",
        "PRIVATE-XML-SINGLE-CELL",
        "/private:PrivateRoot/private:Header/private:AsOf",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_xml_mapping_xml_budget_fails_closed(tmp_path, monkeypatch) -> None:
    baseline = make_xml_mapping_model(tmp_path / "baseline.xlsx")
    candidate = make_xml_mapping_model(tmp_path / "candidate.xlsx")
    baseline_snapshot = load_snapshot(baseline)
    monkeypatch.setattr(workbook_module, "_XML_MAPPING_MAX_XML_PART_BYTES", 1)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert candidate_snapshot.xml_mapping_controls.unrecognized_xml_mapping_count >= 1
    assert any(
        "oversized XML-mapping XML part" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert {"FF010", "FF049"} <= {finding.rule_id for finding in report.findings}


def test_digital_signatures_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_digital_signature_model(tmp_path / "baseline.xlsx")
    package_candidate = make_digital_signature_model(
        tmp_path / "package-candidate.xlsx"
    )
    certificate_candidate = make_digital_signature_model(
        tmp_path / "certificate-candidate.xlsx"
    )
    vba_candidate = make_digital_signature_model(tmp_path / "vba-candidate.xlsx")
    relationship_candidate = make_digital_signature_model(
        tmp_path / "relationship-candidate.xlsx"
    )
    normalised = make_digital_signature_model(tmp_path / "normalised.xlsx")
    unsigned = make_model(tmp_path / "ordinary.xlsx")
    change_package_signature_reference(package_candidate)
    change_package_signature_certificate_payload(certificate_candidate)
    change_vba_project_signature_payload(vba_candidate)
    rebind_package_signature_relationship(relationship_candidate)
    normalize_digital_signature_control_spelling(normalised)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    package_snapshot = load_snapshot(package_candidate)
    certificate_snapshot = load_snapshot(certificate_candidate)
    vba_snapshot = load_snapshot(vba_candidate)
    relationship_snapshot = load_snapshot(relationship_candidate)
    package_report = compare_snapshots(baseline_snapshot, package_snapshot)
    certificate_report = compare_snapshots(
        baseline_snapshot,
        certificate_snapshot,
    )
    vba_report = compare_snapshots(baseline_snapshot, vba_snapshot)
    relationship_report = compare_snapshots(
        baseline_snapshot,
        relationship_snapshot,
    )
    normalised_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(normalised),
    )
    package_change = next(
        change
        for change in package_report.changes
        if change.kind == "digital_signature_controls_changed"
    )
    certificate_change = next(
        change
        for change in certificate_report.changes
        if change.kind == "digital_signature_controls_changed"
    )
    vba_change = next(
        change
        for change in vba_report.changes
        if change.kind == "digital_signature_controls_changed"
    )
    relationship_change = next(
        change
        for change in relationship_report.changes
        if change.kind == "digital_signature_controls_changed"
    )

    assert load_workbook(baseline).sheetnames == ["Inputs", "Model", "Dashboard", "Control"]
    assert baseline_snapshot.cells == package_snapshot.cells
    assert baseline_snapshot.cells == certificate_snapshot.cells
    assert baseline_snapshot.cells == vba_snapshot.cells
    assert baseline_snapshot.cells == relationship_snapshot.cells
    assert baseline_snapshot.summary()["package_xml_signature_count"] == 1
    assert baseline_snapshot.summary()["package_signature_certificate_part_count"] == 1
    assert baseline_snapshot.summary()["vba_project_signature_count"] == 3
    assert baseline_snapshot.summary()["has_digital_signatures"] is True
    assert profile["digital_signatures"] == {
        "present": True,
        "package_signature_origin_count": 1,
        "package_xml_signature_count": 1,
        "package_signature_reference_count": 1,
        "package_signature_certificate_count": 1,
        "package_signature_certificate_part_count": 1,
        "package_signature_certificate_relationship_count": 1,
        "vba_project_signature_count": 3,
        "vba_project_signature_relationship_count": 3,
        "unrecognized_digital_signature_count": 0,
    }
    assert "## Digital-signature controls" in markdown
    assert "does not validate cryptography" in markdown
    assert package_change.details["package_signature_material_changed"] is True
    assert certificate_change.details["package_signature_material_changed"] is True
    assert vba_change.details["vba_project_signature_payload_changed"] is True
    assert relationship_change.details["digital_signature_relationships_changed"] is True
    assert "FF005" not in {finding.rule_id for finding in vba_report.findings}
    assert "FF050" in {finding.rule_id for finding in package_report.findings}
    assert "FF050" in {finding.rule_id for finding in certificate_report.findings}
    assert "FF050" in {finding.rule_id for finding in vba_report.findings}
    assert "FF050" in {finding.rule_id for finding in relationship_report.findings}
    assert "digital_signature_controls_changed" not in {
        change.kind for change in normalised_report.changes
    }
    assert "FF050" not in {finding.rule_id for finding in normalised_report.findings}
    assert load_snapshot(unsigned).digital_signatures.to_dict() == {
        "present": False,
        "package_signature_origin_count": 0,
        "package_xml_signature_count": 0,
        "package_signature_reference_count": 0,
        "package_signature_certificate_count": 0,
        "package_signature_certificate_part_count": 0,
        "package_signature_certificate_relationship_count": 0,
        "vba_project_signature_count": 0,
        "vba_project_signature_relationship_count": 0,
        "unrecognized_digital_signature_count": 0,
    }

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(package_report.to_dict()),
        report_to_markdown(package_report),
        json.dumps(report_to_sarif(package_report)),
        json.dumps(certificate_report.to_dict()),
        report_to_markdown(certificate_report),
        json.dumps(report_to_sarif(certificate_report)),
        json.dumps(vba_report.to_dict()),
        report_to_markdown(vba_report),
        json.dumps(report_to_sarif(vba_report)),
        json.dumps(relationship_report.to_dict()),
        report_to_markdown(relationship_report),
        json.dumps(report_to_sarif(relationship_report)),
    )
    for sensitive_value in (
        "PRIVATE-PACKAGE-DIGEST-BASELINE",
        "PRIVATE-PACKAGE-SIGNATURE-BASELINE",
        "PRIVATE-SIGNER-CERTIFICATE-BASELINE",
        "PRIVATE-CERTIFICATE-PART-BASELINE",
        "PRIVATE-CERTIFICATE-PART-CANDIDATE",
        "PRIVATE-VBA-SIGNATURE-AGILE-CANDIDATE",
        "rIdFencePackageSignatureOrigin",
        "/xl/worksheets/sheet1.xml",
        "sig2.xml",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_unsafe_digital_signature_relationship_fails_closed_and_is_redacted(
    tmp_path,
) -> None:
    baseline = make_digital_signature_model(tmp_path / "baseline.xlsx")
    candidate = make_digital_signature_model(tmp_path / "candidate.xlsx")
    externalize_package_signature_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "digital_signature_controls_changed"
    )

    assert candidate_snapshot.digital_signatures.unrecognized_digital_signature_count >= 1
    assert any(
        "malformed or unsupported digital-signature metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert change.details["digital_signature_relationships_changed"] is True
    assert change.details["unrecognized_digital_signature_metadata_changed"] is True
    assert {"FF010", "FF050"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all(
        "https://private.example.test/PRIVATE-PACKAGE-SIGNATURE" not in artifact
        for artifact in rendered_artifacts
    )


def test_malformed_digital_signature_metadata_fails_closed_and_is_redacted(
    tmp_path,
) -> None:
    baseline = make_digital_signature_model(tmp_path / "baseline.xlsx")
    candidate = make_digital_signature_model(tmp_path / "candidate.xlsx")
    corrupt_package_signature_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "digital_signature_controls_changed"
    )

    assert candidate_snapshot.digital_signatures.unrecognized_digital_signature_count >= 1
    assert any(
        "malformed or unsupported digital-signature metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert change.details["unrecognized_digital_signature_metadata_changed"] is True
    assert {"FF010", "FF050"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("privateUnexpectedSignature" not in artifact for artifact in rendered_artifacts)


def test_digital_signature_read_budget_fails_closed(tmp_path, monkeypatch) -> None:
    baseline = make_digital_signature_model(tmp_path / "baseline.xlsx")
    candidate = make_digital_signature_model(tmp_path / "candidate.xlsx")
    baseline_snapshot = load_snapshot(baseline)
    monkeypatch.setattr(workbook_module, "_DIGITAL_SIGNATURE_MAX_PART_BYTES", 1)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert candidate_snapshot.digital_signatures.unrecognized_digital_signature_count >= 1
    assert any(
        "oversized digital-signature XML part" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert {"FF010", "FF050"} <= {finding.rule_id for finding in report.findings}


def test_rich_data_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_rich_data_model(tmp_path / "baseline.xlsx")
    value_candidate = make_rich_data_model(tmp_path / "value-candidate.xlsx")
    binding_candidate = make_rich_data_model(tmp_path / "binding-candidate.xlsx")
    web_image_candidate = make_rich_data_model(
        tmp_path / "web-image-candidate.xlsx"
    )
    normalised = make_rich_data_model(tmp_path / "normalised.xlsx")
    ordinary = make_model(tmp_path / "ordinary.xlsx")
    change_rich_data_value(value_candidate)
    change_rich_data_binding(binding_candidate)
    change_rich_data_web_image_target(web_image_candidate)
    normalize_rich_data_relationship_ids(normalised)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    value_snapshot = load_snapshot(value_candidate)
    binding_snapshot = load_snapshot(binding_candidate)
    web_image_snapshot = load_snapshot(web_image_candidate)
    value_report = compare_snapshots(baseline_snapshot, value_snapshot)
    binding_report = compare_snapshots(baseline_snapshot, binding_snapshot)
    web_image_report = compare_snapshots(baseline_snapshot, web_image_snapshot)
    normalised_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(normalised),
    )

    value_change = next(
        change
        for change in value_report.changes
        if change.kind == "rich_data_controls_changed"
    )
    binding_change = next(
        change
        for change in binding_report.changes
        if change.kind == "rich_data_controls_changed"
    )
    web_image_change = next(
        change
        for change in web_image_report.changes
        if change.kind == "rich_data_controls_changed"
    )

    assert baseline_snapshot.cells == value_snapshot.cells
    assert baseline_snapshot.cells == binding_snapshot.cells
    assert baseline_snapshot.cells == web_image_snapshot.cells
    assert baseline_snapshot.summary()["rich_value_count"] == 2
    assert baseline_snapshot.summary()["rich_value_bound_cell_count"] == 1
    assert baseline_snapshot.summary()["rich_data_external_relationship_count"] == 3
    assert baseline_snapshot.summary()["has_rich_data"] is True
    assert profile["rich_data"] == {
        "present": True,
        "rich_value_data_part_count": 1,
        "rich_value_structure_part_count": 1,
        "rich_value_type_part_count": 1,
        "rich_value_array_part_count": 1,
        "supporting_property_bag_part_count": 1,
        "supporting_property_bag_structure_part_count": 1,
        "rich_style_part_count": 1,
        "rich_value_web_image_part_count": 1,
        "rich_value_relationship_part_count": 1,
        "rich_value_count": 2,
        "rich_value_structure_count": 2,
        "linked_entity_structure_count": 1,
        "rich_value_array_count": 1,
        "supporting_property_bag_count": 1,
        "rich_value_metadata_binding_count": 2,
        "rich_value_bound_cell_count": 1,
        "web_image_count": 1,
        "web_image_relationship_count": 2,
        "external_web_image_relationship_count": 2,
        "rich_value_relationship_reference_count": 1,
        "external_rich_value_relationship_count": 1,
        "unrecognized_rich_data_count": 0,
    }
    assert "## Rich data controls" in markdown
    assert value_change.details["rich_data_values_changed"] is True
    assert binding_change.details["rich_data_metadata_bindings_changed"] is True
    assert web_image_change.details["rich_data_relationships_changed"] is True
    assert "FF051" in {finding.rule_id for finding in value_report.findings}
    assert "FF051" in {finding.rule_id for finding in binding_report.findings}
    assert "FF051" in {finding.rule_id for finding in web_image_report.findings}
    assert "rich_data_controls_changed" not in {
        change.kind for change in normalised_report.changes
    }
    assert "FF051" not in {finding.rule_id for finding in normalised_report.findings}
    assert load_snapshot(ordinary).rich_data.to_dict() == {
        "present": False,
        "rich_value_data_part_count": 0,
        "rich_value_structure_part_count": 0,
        "rich_value_type_part_count": 0,
        "rich_value_array_part_count": 0,
        "supporting_property_bag_part_count": 0,
        "supporting_property_bag_structure_part_count": 0,
        "rich_style_part_count": 0,
        "rich_value_web_image_part_count": 0,
        "rich_value_relationship_part_count": 0,
        "rich_value_count": 0,
        "rich_value_structure_count": 0,
        "linked_entity_structure_count": 0,
        "rich_value_array_count": 0,
        "supporting_property_bag_count": 0,
        "rich_value_metadata_binding_count": 0,
        "rich_value_bound_cell_count": 0,
        "web_image_count": 0,
        "web_image_relationship_count": 0,
        "external_web_image_relationship_count": 0,
        "rich_value_relationship_reference_count": 0,
        "external_rich_value_relationship_count": 0,
        "unrecognized_rich_data_count": 0,
    }

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(value_report.to_dict()),
        report_to_markdown(value_report),
        json.dumps(report_to_sarif(value_report)),
        json.dumps(binding_report.to_dict()),
        report_to_markdown(binding_report),
        json.dumps(report_to_sarif(binding_report)),
        json.dumps(web_image_report.to_dict()),
        report_to_markdown(web_image_report),
        json.dumps(report_to_sarif(web_image_report)),
    )
    for sensitive_value in (
        "PRIVATE-RICH-ENTITY-BASELINE",
        "PRIVATE-RICH-ENTITY-CANDIDATE",
        "PRIVATE-RICH-FIELD",
        "PRIVATE-RICH-PROPERTY",
        "PRIVATE-RICH-IMAGE-BASELINE",
        "PRIVATE-RICH-IMAGE-CANDIDATE",
        "PRIVATE-RICH-RELATIONSHIP-BASELINE",
        "rIdFenceRichImage",
        "B2",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_malformed_rich_data_metadata_fails_closed_and_is_redacted(tmp_path) -> None:
    baseline = make_rich_data_model(tmp_path / "baseline.xlsx")
    candidate = make_rich_data_model(tmp_path / "candidate.xlsx")
    corrupt_rich_data_value_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "rich_data_controls_changed"
    )

    assert candidate_snapshot.rich_data.unrecognized_rich_data_count >= 1
    assert any(
        "malformed, unsupported, or incomplete rich-data metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert change.details["rich_data_values_changed"] is True
    assert change.details["unrecognized_rich_data_metadata_changed"] is True
    assert {"FF010", "FF051"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all(
        sensitive_value not in artifact
        for sensitive_value in (
            "privateUnexpectedRichData",
            "PRIVATE-RICH-ENTITY-BASELINE",
            "PRIVATE-RICH-FIELD",
        )
        for artifact in rendered_artifacts
    )


def test_rich_data_read_budget_fails_closed(tmp_path, monkeypatch) -> None:
    baseline = make_rich_data_model(tmp_path / "baseline.xlsx")
    candidate = make_rich_data_model(tmp_path / "candidate.xlsx")
    baseline_snapshot = load_snapshot(baseline)
    monkeypatch.setattr(workbook_module, "_RICH_DATA_MAX_XML_PART_BYTES", 1)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert candidate_snapshot.rich_data.unrecognized_rich_data_count >= 1
    assert any(
        "oversized rich-data XML part" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert {"FF010", "FF051"} <= {finding.rule_id for finding in report.findings}


def test_custom_data_stores_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_custom_data_store_model(tmp_path / "baseline.xlsx")
    xml_candidate = make_custom_data_store_model(tmp_path / "xml-candidate.xlsx")
    binary_candidate = make_custom_data_store_model(
        tmp_path / "binary-candidate.xlsx"
    )
    storage_identifier_candidate = make_custom_data_store_model(
        tmp_path / "storage-identifier-candidate.xlsx"
    )
    property_candidate = make_custom_data_store_model(
        tmp_path / "property-candidate.xlsx"
    )
    normalised = make_custom_data_store_model(tmp_path / "normalised.xlsx")
    ordinary = make_model(tmp_path / "ordinary.xlsx")
    change_custom_xml_data_store_value(xml_candidate)
    change_custom_data_payload(binary_candidate)
    change_custom_data_store_storage_identifiers(storage_identifier_candidate)
    change_custom_document_property_value(property_candidate)
    normalize_custom_data_store_identifiers(normalised)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    xml_snapshot = load_snapshot(xml_candidate)
    binary_snapshot = load_snapshot(binary_candidate)
    storage_identifier_snapshot = load_snapshot(storage_identifier_candidate)
    property_snapshot = load_snapshot(property_candidate)
    xml_report = compare_snapshots(baseline_snapshot, xml_snapshot)
    binary_report = compare_snapshots(baseline_snapshot, binary_snapshot)
    storage_identifier_report = compare_snapshots(
        baseline_snapshot,
        storage_identifier_snapshot,
    )
    property_report = compare_snapshots(baseline_snapshot, property_snapshot)
    normalised_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(normalised),
    )
    xml_change = next(
        change
        for change in xml_report.changes
        if change.kind == "custom_data_store_changed"
    )
    binary_change = next(
        change
        for change in binary_report.changes
        if change.kind == "custom_data_store_changed"
    )
    storage_identifier_change = next(
        change
        for change in storage_identifier_report.changes
        if change.kind == "custom_data_store_changed"
    )
    property_change = next(
        change
        for change in property_report.changes
        if change.kind == "custom_data_store_changed"
    )

    assert baseline_snapshot.cells == xml_snapshot.cells
    assert baseline_snapshot.cells == binary_snapshot.cells
    assert baseline_snapshot.cells == storage_identifier_snapshot.cells
    assert baseline_snapshot.cells == property_snapshot.cells
    assert baseline_snapshot.summary()["custom_xml_part_count"] == 1
    assert baseline_snapshot.summary()["custom_data_part_count"] == 1
    assert baseline_snapshot.summary()["document_custom_property_count"] == 1
    assert baseline_snapshot.summary()["has_custom_data_stores"] is True
    assert profile["custom_data_stores"] == {
        "present": True,
        "custom_xml_part_count": 1,
        "custom_xml_property_part_count": 1,
        "custom_xml_schema_reference_count": 1,
        "custom_xml_relationship_count": 2,
        "external_custom_xml_relationship_count": 0,
        "custom_data_properties_part_count": 1,
        "custom_data_part_count": 1,
        "document_custom_property_part_count": 1,
        "document_custom_property_count": 1,
        "linked_document_custom_property_count": 0,
        "unrecognized_custom_data_store_count": 0,
    }
    assert "## Custom workbook data stores" in markdown
    assert xml_change.details["custom_xml_state_changed"] is True
    assert binary_change.details["custom_data_material_changed"] is True
    assert storage_identifier_change.details["custom_xml_state_changed"] is True
    assert storage_identifier_change.details["custom_data_material_changed"] is True
    assert property_change.details["document_custom_properties_changed"] is True
    assert "FF052" in {finding.rule_id for finding in xml_report.findings}
    assert "FF052" in {finding.rule_id for finding in binary_report.findings}
    assert "FF052" in {
        finding.rule_id for finding in storage_identifier_report.findings
    }
    assert "FF052" in {finding.rule_id for finding in property_report.findings}
    assert "custom_data_store_changed" not in {
        change.kind for change in normalised_report.changes
    }
    assert "FF052" not in {finding.rule_id for finding in normalised_report.findings}
    assert load_snapshot(ordinary).custom_data_stores.to_dict() == {
        "present": False,
        "custom_xml_part_count": 0,
        "custom_xml_property_part_count": 0,
        "custom_xml_schema_reference_count": 0,
        "custom_xml_relationship_count": 0,
        "external_custom_xml_relationship_count": 0,
        "custom_data_properties_part_count": 0,
        "custom_data_part_count": 0,
        "document_custom_property_part_count": 0,
        "document_custom_property_count": 0,
        "linked_document_custom_property_count": 0,
        "unrecognized_custom_data_store_count": 0,
    }

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(xml_report.to_dict()),
        report_to_markdown(xml_report),
        json.dumps(report_to_sarif(xml_report)),
        json.dumps(binary_report.to_dict()),
        report_to_markdown(binary_report),
        json.dumps(report_to_sarif(binary_report)),
        json.dumps(storage_identifier_report.to_dict()),
        report_to_markdown(storage_identifier_report),
        json.dumps(report_to_sarif(storage_identifier_report)),
        json.dumps(property_report.to_dict()),
        report_to_markdown(property_report),
        json.dumps(report_to_sarif(property_report)),
    )
    for sensitive_value in (
        "PRIVATE-CUSTOM-XML-BASELINE",
        "PRIVATE-CUSTOM-XML-CANDIDATE",
        "PRIVATE-CUSTOM-XML-ID",
        "PRIVATE-CUSTOM-XML-ID-CANDIDATE",
        "PRIVATE-CUSTOM-DATA-BASELINE",
        "PRIVATE-CUSTOM-DATA-CANDIDATE",
        "PRIVATE-CUSTOM-DATA-ID",
        "PRIVATE-CUSTOM-DATA-ID-CANDIDATE",
        "PRIVATE-CUSTOM-DOCUMENT-PROPERTY-NAME",
        "PRIVATE-CUSTOM-DOCUMENT-PROPERTY-BASELINE",
        "PRIVATE-CUSTOM-DOCUMENT-PROPERTY-CANDIDATE",
        "rIdFenceCustomDataPayload",
        "urn:formulafence:private-custom-xml-schema",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_custom_data_store_metadata_fails_closed_and_is_redacted(tmp_path) -> None:
    baseline = make_custom_data_store_model(tmp_path / "baseline.xlsx")
    candidate = make_custom_data_store_model(tmp_path / "candidate.xlsx")
    corrupt_custom_data_properties_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "custom_data_store_changed"
    )

    assert candidate_snapshot.custom_data_stores.unrecognized_custom_data_store_count >= 1
    assert any(
        "malformed, unsupported, or incomplete custom workbook data-store metadata"
        in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert change.details["custom_data_material_changed"] is True
    assert change.details["unrecognized_custom_data_store_metadata_changed"] is True
    assert {"FF010", "FF052"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "privateUnexpectedCustomDataProperties",
        "PRIVATE-CUSTOM-DATA-ID",
        "PRIVATE-CUSTOM-DATA-BASELINE",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_custom_data_store_read_budget_fails_closed(tmp_path, monkeypatch) -> None:
    baseline = make_custom_data_store_model(tmp_path / "baseline.xlsx")
    candidate = make_custom_data_store_model(tmp_path / "candidate.xlsx")
    baseline_snapshot = load_snapshot(baseline)
    monkeypatch.setattr(workbook_module, "_CUSTOM_DATA_STORE_MAX_PART_BYTES", 1)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert candidate_snapshot.custom_data_stores.unrecognized_custom_data_store_count >= 1
    assert any(
        "oversized custom data-store part" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert {"FF010", "FF052"} <= {finding.rule_id for finding in report.findings}


def test_power_query_custom_xml_is_not_double_counted_as_generic_state(tmp_path) -> None:
    baseline = make_power_query_model(tmp_path / "baseline.xlsx")
    candidate = make_power_query_model(tmp_path / "candidate.xlsx")
    change_power_query_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.custom_data_stores.present is False
    assert "custom_data_store_changed" not in {change.kind for change in report.changes}
    assert "FF052" not in {finding.rule_id for finding in report.findings}


def test_legacy_excel_notes_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_legacy_comment_model(tmp_path / "baseline.xlsx")
    text_candidate = make_legacy_comment_model(tmp_path / "text-candidate.xlsx")
    visibility_candidate = make_legacy_comment_model(
        tmp_path / "visibility-candidate.xlsx"
    )
    change_legacy_comment_text(text_candidate)
    change_legacy_note_visibility(visibility_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    text_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(text_candidate),
    )
    visibility_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(visibility_candidate),
    )
    text_change = next(
        change
        for change in text_report.changes
        if change.kind == "legacy_comment_controls_changed"
    )
    visibility_change = next(
        change
        for change in visibility_report.changes
        if change.kind == "legacy_comment_controls_changed"
    )

    assert baseline_snapshot.summary()["legacy_comment_count"] == 1
    assert baseline_snapshot.summary()["legacy_comment_author_count"] == 1
    assert baseline_snapshot.summary()["legacy_comment_note_shape_count"] == 1
    assert baseline_snapshot.summary()["has_legacy_comments"] is True
    assert profile["legacy_comments"] == {
        "present": True,
        "worksheet_comment_sheet_count": 1,
        "comment_part_count": 1,
        "comment_author_count": 1,
        "comment_count": 1,
        "comment_with_text_count": 1,
        "rich_text_comment_count": 0,
        "phonetic_comment_count": 0,
        "comment_property_count": 0,
        "threaded_placeholder_count": 0,
        "worksheet_note_drawing_sheet_count": 1,
        "note_vml_drawing_part_count": 1,
        "note_shape_count": 1,
        "visible_note_shape_count": 0,
        "anchored_note_shape_count": 0,
        "binding_relationship_count": 2,
        "external_relationship_count": 0,
        "unrecognized_legacy_comment_count": 0,
    }
    assert "## Legacy Excel Notes and threaded placeholders" in markdown
    assert text_change.details["legacy_comment_definition_material_changed"] is True
    assert visibility_change.details["legacy_note_vml_material_changed"] is True
    assert "FF046" in {finding.rule_id for finding in text_report.findings}
    assert "FF046" in {finding.rule_id for finding in visibility_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(text_report.to_dict()),
        report_to_markdown(text_report),
        json.dumps(report_to_sarif(text_report)),
        json.dumps(visibility_report.to_dict()),
        report_to_markdown(visibility_report),
        json.dumps(report_to_sarif(visibility_report)),
    )
    for sensitive_value in (
        "PRIVATE-LEGACY-NOTE-BASELINE",
        "PRIVATE-LEGACY-NOTE-CANDIDATE",
        "Private Legacy Note Author",
        "A1",
        "_x0000_s",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_unsafe_legacy_comment_relationship_fails_closed_and_is_redacted(
    tmp_path,
) -> None:
    baseline = make_legacy_comment_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_comment_model(tmp_path / "candidate.xlsx")
    externalize_legacy_comment_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "legacy_comment_controls_changed"
    )

    assert candidate_snapshot.legacy_comments.external_relationship_count == 1
    assert candidate_snapshot.legacy_comments.unrecognized_legacy_comment_count >= 1
    assert change.details["legacy_note_relationships_changed"] is True
    assert "FF046" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "https://example.invalid/private-legacy-note",
        "rId",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_legacy_comment_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_legacy_comment_model(tmp_path / "baseline.xlsx")
    renumbered = make_legacy_comment_model(tmp_path / "renumbered.xlsx")
    renumber_legacy_comment_identifiers(renumbered)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered))

    assert "legacy_comment_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF046" not in {finding.rule_id for finding in report.findings}


def test_legacy_comment_relationship_rebinding_is_guarded_and_redacted(
    tmp_path,
) -> None:
    baseline = make_legacy_comment_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_comment_model(tmp_path / "candidate.xlsx")
    rebind_legacy_comment_relationship(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "legacy_comment_controls_changed"
    )
    rendered_artifacts = (
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )

    assert change.details["legacy_note_relationships_changed"] is True
    assert "FF046" in {finding.rule_id for finding in report.findings}
    assert all("comment2.xml" not in artifact for artifact in rendered_artifacts)


def test_unsafe_legacy_note_vml_relationship_is_quarantined_for_reader(
    tmp_path,
) -> None:
    baseline = make_legacy_comment_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_comment_model(tmp_path / "candidate.xlsx")
    externalize_legacy_note_vml_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.legacy_comments.external_relationship_count == 1
    assert candidate_snapshot.legacy_comments.unrecognized_legacy_comment_count >= 1
    assert any(
        "isolated unsafe legacy Excel Note relationships" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "legacy_comment_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF046" in {finding.rule_id for finding in report.findings}


def test_hyperlink_reader_overlay_preserves_legacy_note_isolation(tmp_path) -> None:
    """Safe-reader overlays compose when one worksheet has both feature types."""
    candidate = make_legacy_comment_model(tmp_path / "candidate.xlsx")
    workbook = load_workbook(candidate)
    worksheet = workbook["Inputs"]
    worksheet["B2"].hyperlink = "https://approved.example.test/private-link"
    workbook.save(candidate)
    externalize_legacy_note_vml_relationship(candidate)

    reader_source, temporary_path, _warnings = workbook_module._openpyxl_safe_source(
        candidate
    )
    try:
        with ZipFile(reader_source) as archive:
            relationships = ElementTree.fromstring(
                archive.read("xl/worksheets/_rels/sheet1.xml.rels")
            )
            relationship_types = {
                relationship.get("Type", "")
                for relationship in relationships
            }
            worksheet_root = ElementTree.fromstring(
                archive.read("xl/worksheets/sheet1.xml")
            )

        assert not any(
            relationship_type.endswith("/vmlDrawing")
            for relationship_type in relationship_types
        )
        assert not any(
            relationship_type.endswith("/hyperlink")
            for relationship_type in relationship_types
        )
        assert not any(
            element.tag.endswith("legacyDrawing")
            for element in worksheet_root.iter()
        )
        assert not any(
            element.tag.endswith("hyperlink")
            for element in worksheet_root.iter()
        )
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def test_threaded_comment_placeholders_are_guarded_without_duplicating_threads(
    tmp_path,
) -> None:
    baseline = make_legacy_threaded_placeholder_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_threaded_placeholder_model(tmp_path / "candidate.xlsx")
    renumbered = make_legacy_threaded_placeholder_model(tmp_path / "renumbered.xlsx")
    lowercase = make_legacy_threaded_placeholder_model(tmp_path / "lowercase.xlsx")
    author_candidate = make_legacy_threaded_placeholder_model(
        tmp_path / "author-candidate.xlsx"
    )
    change_legacy_comment_text(candidate)
    renumber_legacy_threaded_placeholder_identifiers(renumbered)
    lowercase_legacy_threaded_placeholder_identifiers(lowercase)
    change_legacy_placeholder_author_context(author_candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    renumber_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(renumbered),
    )
    lowercase_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(lowercase),
    )
    author_snapshot = load_snapshot(author_candidate)
    author_report = compare_snapshots(baseline_snapshot, author_snapshot)

    assert baseline_snapshot.legacy_comments.threaded_placeholder_count == 1
    assert baseline_snapshot.threaded_comments.comment_count == 1
    assert author_snapshot.legacy_comments.threaded_placeholder_count == 0
    assert "legacy_comment_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "threaded_comment_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF046" in {finding.rule_id for finding in report.findings}
    assert "FF045" not in {finding.rule_id for finding in report.findings}
    assert "legacy_comment_controls_changed" not in {
        change.kind for change in renumber_report.changes
    }
    assert "threaded_comment_controls_changed" not in {
        change.kind for change in renumber_report.changes
    }
    assert "legacy_comment_controls_changed" not in {
        change.kind for change in lowercase_report.changes
    }
    assert "threaded_comment_controls_changed" not in {
        change.kind for change in lowercase_report.changes
    }
    assert "legacy_comment_controls_changed" in {
        change.kind for change in author_report.changes
    }
    assert "FF046" in {finding.rule_id for finding in author_report.findings}


def test_malformed_legacy_comments_fail_closed(tmp_path) -> None:
    baseline = make_legacy_comment_model(tmp_path / "baseline.xlsx")
    candidate = make_legacy_comment_model(tmp_path / "candidate.xlsx")
    corrupt_legacy_comment_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.legacy_comments.unrecognized_legacy_comment_count >= 1
    assert any(
        "legacy comments part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "legacy_comment_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF046" in {finding.rule_id for finding in report.findings}


def test_legacy_comment_xml_budget_fails_closed(tmp_path, monkeypatch) -> None:
    workbook = make_legacy_comment_model(tmp_path / "candidate.xlsx")
    monkeypatch.setattr(workbook_module, "_LEGACY_COMMENT_MAX_XML_PART_BYTES", 1)

    snapshot = load_snapshot(workbook)

    assert snapshot.legacy_comments.unrecognized_legacy_comment_count >= 1
    assert any(
        "oversized legacy-note" in warning
        for warning in snapshot.parser_warnings
    )


def test_threaded_comments_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_threaded_comment_model(tmp_path / "baseline.xlsx")
    candidate = make_threaded_comment_model(tmp_path / "candidate.xlsx")
    person_candidate = make_threaded_comment_model(tmp_path / "person-candidate.xlsx")
    change_threaded_comment_reply(candidate)
    change_threaded_comment_person_identity(person_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    person_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(person_candidate),
    )
    change = next(
        change
        for change in report.changes
        if change.kind == "threaded_comment_controls_changed"
    )

    assert baseline_snapshot.summary()["threaded_comment_count"] == 2
    assert baseline_snapshot.summary()["threaded_comment_thread_count"] == 1
    assert baseline_snapshot.summary()["threaded_comment_reply_count"] == 1
    assert baseline_snapshot.summary()["threaded_comment_person_count"] == 2
    assert baseline_snapshot.summary()["has_threaded_comments"] is True
    assert profile["threaded_comments"] == {
        "present": True,
        "worksheet_threaded_comment_sheet_count": 1,
        "threaded_comment_part_count": 1,
        "comment_thread_count": 1,
        "comment_count": 2,
        "reply_count": 1,
        "resolved_comment_count": 1,
        "comment_with_text_count": 2,
        "mention_count": 1,
        "mentioned_person_count": 1,
        "person_part_count": 1,
        "person_count": 2,
        "orphan_person_count": 0,
        "binding_relationship_count": 2,
        "external_relationship_count": 0,
        "unrecognized_threaded_comment_count": 0,
    }
    assert "## Modern threaded comments" in markdown
    assert change.details["threaded_comment_definition_material_changed"] is True
    assert "FF045" in {finding.rule_id for finding in report.findings}
    person_change = next(
        change
        for change in person_report.changes
        if change.kind == "threaded_comment_controls_changed"
    )
    assert person_change.details["threaded_comment_person_material_changed"] is True
    assert "FF045" in {finding.rule_id for finding in person_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
        json.dumps(person_report.to_dict()),
        report_to_markdown(person_report),
        json.dumps(report_to_sarif(person_report)),
    )
    for sensitive_value in (
        "PRIVATE-THREADED-COMMENT-BASELINE",
        "PRIVATE-THREADED-REPLY-BASELINE",
        "PRIVATE-THREADED-REPLY-CANDIDATE",
        "Private Reviewer",
        "private-reviewer@example.invalid",
        "2026-07-24T00:00:00Z",
        "{11111111-1111-1111-1111-111111111111}",
        "{55555555-5555-5555-5555-555555555555}",
        "rIdFenceThreadedComment",
        "A1",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_unsafe_threaded_comment_relationship_fails_closed_and_is_redacted(tmp_path) -> None:
    baseline = make_threaded_comment_model(tmp_path / "baseline.xlsx")
    candidate = make_threaded_comment_model(tmp_path / "candidate.xlsx")
    externalize_threaded_comment_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "threaded_comment_controls_changed"
    )

    assert candidate_snapshot.threaded_comments.external_relationship_count == 1
    assert candidate_snapshot.threaded_comments.unrecognized_threaded_comment_count >= 1
    assert change.details["threaded_comment_relationships_changed"] is True
    assert "FF045" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        profile_to_markdown(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "https://example.invalid/private-threaded-comment",
        "rIdFenceThreadedComment",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_threaded_comment_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_threaded_comment_model(tmp_path / "baseline.xlsx")
    renumbered = make_threaded_comment_model(tmp_path / "renumbered.xlsx")
    renumber_threaded_comment_identifiers(renumbered)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered))

    assert "threaded_comment_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF045" not in {finding.rule_id for finding in report.findings}


def test_malformed_threaded_comments_fail_closed(tmp_path) -> None:
    baseline = make_threaded_comment_model(tmp_path / "baseline.xlsx")
    candidate = make_threaded_comment_model(tmp_path / "candidate.xlsx")
    corrupt_threaded_comment_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.threaded_comments.unrecognized_threaded_comment_count >= 1
    assert any(
        "threaded-comment part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "threaded_comment_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF045" in {finding.rule_id for finding in report.findings}


def test_threaded_comment_xml_budget_fails_closed(tmp_path, monkeypatch) -> None:
    workbook = make_threaded_comment_model(tmp_path / "candidate.xlsx")
    monkeypatch.setattr(workbook_module, "_THREADED_COMMENT_MAX_XML_PART_BYTES", 1)

    snapshot = load_snapshot(workbook)

    assert snapshot.threaded_comments.unrecognized_threaded_comment_count >= 1
    assert any(
        "oversized threaded-comment" in warning
        for warning in snapshot.parser_warnings
    )


def test_worksheet_drawing_shapes_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_worksheet_drawing_shape_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_drawing_shape_model(tmp_path / "candidate.xlsx")
    hyperlink_candidate = make_worksheet_drawing_shape_model(
        tmp_path / "hyperlink-candidate.xlsx"
    )
    change_worksheet_drawing_shape_presentation(candidate)
    change_worksheet_drawing_shape_hyperlink(hyperlink_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    hyperlink_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(hyperlink_candidate),
    )
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_drawing_shape_controls_changed"
    )

    assert baseline_snapshot.summary()["worksheet_drawing_shape_count"] == 2
    assert baseline_snapshot.summary()["worksheet_drawing_text_shape_count"] == 2
    assert baseline_snapshot.summary()["has_worksheet_drawing_shapes"] is True
    assert baseline_snapshot.worksheet_images.present is False
    assert baseline_snapshot.worksheet_images.worksheet_image_sheet_count == 0
    assert profile["worksheet_drawing_shapes"] == {
        "present": True,
        "worksheet_drawing_sheet_count": 1,
        "worksheet_drawing_part_count": 1,
        "shape_anchor_count": 2,
        "shape_count": 2,
        "connector_shape_count": 0,
        "connector_attachment_count": 0,
        "group_shape_count": 1,
        "graphic_frame_count": 0,
        "diagram_graphic_frame_count": 0,
        "diagram_data_part_count": 0,
        "diagram_layout_part_count": 0,
        "diagram_quick_style_part_count": 0,
        "diagram_colour_part_count": 0,
        "diagram_drawing_part_count": 0,
        "diagram_image_part_count": 0,
        "fingerprinted_diagram_image_part_count": 0,
        "uninspected_diagram_image_part_count": 0,
        "text_shape_count": 2,
        "text_paragraph_count": 2,
        "text_run_count": 2,
        "macro_assignment_count": 1,
        "text_link_count": 1,
        "hyperlink_count": 1,
        "related_relationship_count": 1,
        "external_relationship_count": 1,
        "unrecognized_graphic_frame_count": 0,
        "unrecognized_shape_count": 0,
    }
    assert "## Worksheet DrawingML shape, connector, and graphic-frame controls" in markdown
    assert change.details["worksheet_drawing_shape_definition_material_changed"] is True
    assert "FF044" in {finding.rule_id for finding in report.findings}
    assert "FF044" in {finding.rule_id for finding in hyperlink_report.findings}
    hyperlink_change = next(
        change
        for change in hyperlink_report.changes
        if change.kind == "worksheet_drawing_shape_controls_changed"
    )
    assert hyperlink_change.details["worksheet_drawing_shape_relationships_changed"] is True

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
        json.dumps(hyperlink_report.to_dict()),
        report_to_markdown(hyperlink_report),
        json.dumps(report_to_sarif(hyperlink_report)),
    )
    for sensitive_value in (
        "PRIVATE-SHAPE-DO-NOT-APPROVE",
        "PRIVATE-GROUP-SHAPE-TEXT",
        "PrivateWorksheetShapeMacro",
        "=Inputs!$B$2",
        "private-worksheet-shape-target",
        "private-worksheet-shape-candidate",
        "Private worksheet warning shape",
        "000000",
        "FFFFFF",
        "rIdFenceShapeLink",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_drawing_shape_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_worksheet_drawing_shape_model(tmp_path / "baseline.xlsx")
    renumbered = make_worksheet_drawing_shape_model(tmp_path / "renumbered.xlsx")
    renumber_worksheet_drawing_shape_identifiers(renumbered)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered))

    assert "worksheet_drawing_shape_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF044" not in {finding.rule_id for finding in report.findings}


def test_worksheet_smartart_diagrams_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_worksheet_smartart_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_smartart_model(tmp_path / "candidate.xlsx")
    change_worksheet_smartart_data(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_drawing_shape_controls_changed"
    )

    assert baseline_snapshot.summary()["worksheet_drawing_graphic_frame_count"] == 1
    assert baseline_snapshot.summary()["worksheet_drawing_diagram_frame_count"] == 1
    assert baseline_snapshot.summary()["has_worksheet_drawing_shapes"] is True
    assert baseline_snapshot.chart_definitions.present is False
    assert profile["worksheet_drawing_shapes"] == {
        "present": True,
        "worksheet_drawing_sheet_count": 1,
        "worksheet_drawing_part_count": 1,
        "shape_anchor_count": 1,
        "shape_count": 0,
        "connector_shape_count": 0,
        "connector_attachment_count": 0,
        "group_shape_count": 0,
        "graphic_frame_count": 1,
        "diagram_graphic_frame_count": 1,
        "diagram_data_part_count": 1,
        "diagram_layout_part_count": 1,
        "diagram_quick_style_part_count": 1,
        "diagram_colour_part_count": 1,
        "diagram_drawing_part_count": 1,
        "diagram_image_part_count": 0,
        "fingerprinted_diagram_image_part_count": 0,
        "uninspected_diagram_image_part_count": 0,
        "text_shape_count": 0,
        "text_paragraph_count": 0,
        "text_run_count": 0,
        "macro_assignment_count": 1,
        "text_link_count": 0,
        "hyperlink_count": 0,
        "related_relationship_count": 5,
        "external_relationship_count": 0,
        "unrecognized_graphic_frame_count": 0,
        "unrecognized_shape_count": 0,
    }
    assert "SmartArt diagrams" in markdown
    assert change.details["worksheet_drawing_diagram_material_changed"] is True
    assert "FF044" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-SMARTART-DO-NOT-APPROVE",
        "PRIVATE-SMARTART-CANDIDATE-REVIEW-STATE",
        "PRIVATE-SMARTART-NAME",
        "PRIVATE-SMARTART-DESCRIPTION",
        "PrivateSmartArtMacro",
        "rIdFenceDiagramData",
        "rIdFenceDiagramLayout",
        "AAAAAAAA-AAAA-AAAA-AAAA-AAAAAAAAAAAA",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_smartart_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_worksheet_smartart_model(tmp_path / "baseline.xlsx")
    renumbered = make_worksheet_smartart_model(tmp_path / "renumbered.xlsx")
    renumber_worksheet_smartart_identifiers(renumbered)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered))

    assert "worksheet_drawing_shape_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF044" not in {finding.rule_id for finding in report.findings}


def test_worksheet_smartart_diagram_images_are_profiled_diffed_and_redacted(
    tmp_path,
) -> None:
    baseline = make_worksheet_smartart_image_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_smartart_image_model(tmp_path / "candidate.xlsx")
    change_worksheet_smartart_diagram_image_payload(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_drawing_shape_controls_changed"
    )

    shapes = profile["worksheet_drawing_shapes"]
    assert shapes["diagram_data_part_count"] == 1
    assert shapes["diagram_image_part_count"] == 1
    assert shapes["fingerprinted_diagram_image_part_count"] == 1
    assert shapes["uninspected_diagram_image_part_count"] == 0
    assert shapes["related_relationship_count"] == 6
    assert shapes["external_relationship_count"] == 0
    assert shapes["unrecognized_shape_count"] == 0
    assert baseline_snapshot.parser_warnings == ()
    assert "SmartArt Diagram Data images" in markdown
    assert change.details["worksheet_drawing_diagram_material_changed"] is True
    assert "FF044" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-SMARTART-DIAGRAM-IMAGE-CANDIDATE",
        "private-smartart-diagram.png",
        "rIdFenceDiagramImage",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_smartart_diagram_image_relationship_identifiers_are_ignored(
    tmp_path,
) -> None:
    baseline = make_worksheet_smartart_image_model(tmp_path / "baseline.xlsx")
    renumbered = make_worksheet_smartart_image_model(tmp_path / "renumbered.xlsx")
    renumber_worksheet_smartart_diagram_image_relationship(renumbered)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered))

    assert report.changes == []
    assert report.findings == []


def test_external_worksheet_smartart_diagram_images_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_smartart_image_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_smartart_image_model(tmp_path / "candidate.xlsx")
    externalize_worksheet_smartart_diagram_image_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    shapes = candidate_snapshot.worksheet_drawing_shapes
    assert shapes.diagram_image_part_count == 1
    assert shapes.uninspected_diagram_image_part_count == 1
    assert shapes.external_relationship_count >= 1
    assert shapes.unrecognized_shape_count >= 1
    assert any(
        "SmartArt Diagram Data image relationship without a safe internal target"
        in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_drawing_shape_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF044" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(candidate_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("private-smartart-image" not in artifact for artifact in rendered_artifacts)


def test_missing_worksheet_smartart_diagram_image_payloads_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_smartart_image_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_smartart_image_model(tmp_path / "candidate.xlsx")
    remove_worksheet_smartart_diagram_image_payload(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    shapes = candidate_snapshot.worksheet_drawing_shapes
    assert shapes.diagram_image_part_count == 1
    assert shapes.fingerprinted_diagram_image_part_count == 0
    assert shapes.uninspected_diagram_image_part_count == 1
    assert any(
        "could not locate a SmartArt Diagram Data image part" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_drawing_shape_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF044" in {finding.rule_id for finding in report.findings}


def test_oversized_worksheet_smartart_diagram_images_remain_covered(
    tmp_path,
    monkeypatch,
) -> None:
    baseline = make_worksheet_smartart_image_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_smartart_image_model(tmp_path / "candidate.xlsx")
    change_worksheet_smartart_diagram_image_payload(candidate)
    monkeypatch.setattr(
        workbook_module,
        "_WORKSHEET_DRAWING_SHAPE_RELATED_PART_MAX_BYTES",
        1,
    )

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    shapes = baseline_snapshot.worksheet_drawing_shapes
    assert shapes.diagram_image_part_count == 1
    assert shapes.fingerprinted_diagram_image_part_count == 0
    assert shapes.uninspected_diagram_image_part_count == 1
    assert any(
        "oversized SmartArt Diagram Data image" in warning
        for warning in baseline_snapshot.parser_warnings
    )
    assert "worksheet_drawing_shape_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF044" in {finding.rule_id for finding in report.findings}


def test_worksheet_smartart_diagram_image_count_budget_remains_covered(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_worksheet_smartart_image_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._WorksheetDrawingShapeRelatedPartBudget
    monkeypatch.setattr(
        workbook_module,
        "_WorksheetDrawingShapeRelatedPartBudget",
        lambda: budget_type(remaining_parts=0),
    )

    snapshot = load_snapshot(workbook)

    shapes = snapshot.worksheet_drawing_shapes
    assert shapes.diagram_image_part_count == 1
    assert shapes.fingerprinted_diagram_image_part_count == 0
    assert shapes.uninspected_diagram_image_part_count == 1
    assert any(
        "SmartArt Diagram Data image part count budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_worksheet_smartart_diagram_image_byte_budget_remains_covered(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_worksheet_smartart_image_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._WorksheetDrawingShapeRelatedPartBudget
    monkeypatch.setattr(
        workbook_module,
        "_WorksheetDrawingShapeRelatedPartBudget",
        lambda: budget_type(remaining_bytes=1),
    )

    snapshot = load_snapshot(workbook)

    shapes = snapshot.worksheet_drawing_shapes
    assert shapes.diagram_image_part_count == 1
    assert shapes.fingerprinted_diagram_image_part_count == 0
    assert shapes.uninspected_diagram_image_part_count == 1
    assert any(
        "SmartArt Diagram Data image read budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_malformed_worksheet_smartart_relationships_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_smartart_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_smartart_model(tmp_path / "candidate.xlsx")
    corrupt_worksheet_smartart_relationships(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.worksheet_drawing_shapes.unrecognized_shape_count >= 1
    assert any(
        "incomplete diagram relationship declaration" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_drawing_shape_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF044" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(candidate_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("rIdFenceDiagramColours" not in artifact for artifact in rendered_artifacts)


def test_malformed_worksheet_smartart_component_roots_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_smartart_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_smartart_model(tmp_path / "candidate.xlsx")
    corrupt_worksheet_smartart_component_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.worksheet_drawing_shapes.unrecognized_shape_count >= 1
    assert any(
        "SmartArt component with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_drawing_shape_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF044" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(candidate_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("unexpectedDataModel" not in artifact for artifact in rendered_artifacts)


def test_unknown_worksheet_graphic_frames_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_smartart_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_smartart_model(tmp_path / "candidate.xlsx")
    change_worksheet_smartart_graphic_frame_uri(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)
    profile = profile_snapshot(candidate_snapshot)

    assert candidate_snapshot.worksheet_drawing_shapes.graphic_frame_count == 1
    assert candidate_snapshot.worksheet_drawing_shapes.diagram_graphic_frame_count == 0
    assert (
        candidate_snapshot.worksheet_drawing_shapes.unrecognized_graphic_frame_count
        == 1
    )
    assert candidate_snapshot.worksheet_drawing_shapes.unrecognized_shape_count >= 1
    assert any(
        "unsupported non-chart Worksheet DrawingML graphic frames" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_drawing_shape_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF044" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("unknown-graphic" not in artifact for artifact in rendered_artifacts)


def test_worksheet_smartart_component_relationships_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_smartart_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_smartart_model(tmp_path / "candidate.xlsx")
    add_worksheet_smartart_component_relationship(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.worksheet_drawing_shapes.unrecognized_shape_count >= 1
    assert any(
        "relationships from a Worksheet DrawingML SmartArt component" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_drawing_shape_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF044" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(candidate_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("private-smartart-target" not in artifact for artifact in rendered_artifacts)


def test_worksheet_drawing_connectors_are_profiled_diffed_and_redacted(
    tmp_path,
) -> None:
    baseline = make_worksheet_drawing_connector_model(tmp_path / "baseline.xlsx")
    presentation_candidate = make_worksheet_drawing_connector_model(
        tmp_path / "presentation-candidate.xlsx"
    )
    attachment_candidate = make_worksheet_drawing_connector_model(
        tmp_path / "attachment-candidate.xlsx"
    )
    free_connector = make_worksheet_drawing_connector_model(
        tmp_path / "free-connector.xlsx",
        attached=False,
    )
    grouped_connector = make_worksheet_drawing_connector_model(
        tmp_path / "grouped-connector.xlsx",
        grouped=True,
    )
    change_worksheet_drawing_connector_presentation(presentation_candidate)
    change_worksheet_drawing_connector_attachment(attachment_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    presentation_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(presentation_candidate),
    )
    attachment_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(attachment_candidate),
    )
    free_snapshot = load_snapshot(free_connector)
    grouped_snapshot = load_snapshot(grouped_connector)

    assert baseline_snapshot.summary()["worksheet_drawing_connector_shape_count"] == 1
    assert (
        baseline_snapshot.summary()["worksheet_drawing_connector_attachment_count"]
        == 2
    )
    assert profile["worksheet_drawing_shapes"]["shape_anchor_count"] == 3
    assert profile["worksheet_drawing_shapes"]["shape_count"] == 2
    assert profile["worksheet_drawing_shapes"]["connector_shape_count"] == 1
    assert profile["worksheet_drawing_shapes"]["connector_attachment_count"] == 2
    assert profile["worksheet_drawing_shapes"]["group_shape_count"] == 1
    assert profile["worksheet_drawing_shapes"]["unrecognized_shape_count"] == 0
    assert free_snapshot.worksheet_drawing_shapes.connector_shape_count == 1
    assert free_snapshot.worksheet_drawing_shapes.connector_attachment_count == 0
    assert free_snapshot.worksheet_drawing_shapes.unrecognized_shape_count == 0
    assert grouped_snapshot.worksheet_drawing_shapes.connector_shape_count == 1
    assert grouped_snapshot.worksheet_drawing_shapes.connector_attachment_count == 2
    assert grouped_snapshot.worksheet_drawing_shapes.group_shape_count == 2
    assert grouped_snapshot.worksheet_drawing_shapes.unrecognized_shape_count == 0
    assert "Connector attachments" in markdown
    assert "FF044" in {finding.rule_id for finding in presentation_report.findings}
    assert "FF044" in {finding.rule_id for finding in attachment_report.findings}
    attachment_change = next(
        change
        for change in attachment_report.changes
        if change.kind == "worksheet_drawing_shape_controls_changed"
    )
    assert attachment_change.details[
        "worksheet_drawing_shape_definition_material_changed"
    ] is True

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(presentation_report.to_dict()),
        report_to_markdown(presentation_report),
        json.dumps(report_to_sarif(presentation_report)),
        json.dumps(attachment_report.to_dict()),
        report_to_markdown(attachment_report),
        json.dumps(report_to_sarif(attachment_report)),
    )
    for sensitive_value in (
        "PRIVATE-WORKFLOW-CONNECTOR-NAME",
        "PRIVATE-WORKFLOW-CONNECTOR-DESCRIPTION",
        "112233",
        "1025",
        "1026",
        "1027",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_drawing_connector_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_worksheet_drawing_connector_model(tmp_path / "baseline.xlsx")
    renumbered = make_worksheet_drawing_connector_model(tmp_path / "renumbered.xlsx")
    renumber_worksheet_drawing_connector_identifiers(renumbered)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered))

    assert "worksheet_drawing_shape_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF044" not in {finding.rule_id for finding in report.findings}


def test_malformed_worksheet_drawing_connector_attachments_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_drawing_connector_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_drawing_connector_model(tmp_path / "candidate.xlsx")
    corrupt_worksheet_drawing_connector_attachment(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.worksheet_drawing_shapes.unrecognized_shape_count >= 1
    assert any(
        "malformed or unsupported Worksheet DrawingML shape metadata" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_drawing_shape_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF044" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(candidate_snapshot)),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("999999" not in artifact for artifact in rendered_artifacts)


def test_strict_worksheet_drawing_connectors_are_supported(tmp_path) -> None:
    workbook = make_strict_worksheet_drawing_connector_model(tmp_path / "strict.xlsx")
    candidate = make_strict_worksheet_drawing_connector_model(
        tmp_path / "strict-candidate.xlsx"
    )
    change_worksheet_drawing_connector_attachment(candidate)

    snapshot = load_snapshot(workbook)
    report = compare_snapshots(snapshot, load_snapshot(candidate))

    assert snapshot.worksheet_drawing_shapes.shape_count == 2
    assert snapshot.worksheet_drawing_shapes.connector_shape_count == 1
    assert snapshot.worksheet_drawing_shapes.connector_attachment_count == 2
    assert snapshot.worksheet_drawing_shapes.unrecognized_shape_count == 0
    assert not any(
        "Worksheet DrawingML shape" in warning for warning in snapshot.parser_warnings
    )
    assert "FF044" in {finding.rule_id for finding in report.findings}


def test_strict_worksheet_smartart_diagrams_are_supported(tmp_path) -> None:
    baseline = make_strict_worksheet_smartart_model(tmp_path / "baseline.xlsx")
    candidate = make_strict_worksheet_smartart_model(tmp_path / "candidate.xlsx")
    change_worksheet_smartart_data(candidate)

    snapshot = load_snapshot(baseline)
    report = compare_snapshots(snapshot, load_snapshot(candidate))

    assert snapshot.worksheet_drawing_shapes.graphic_frame_count == 1
    assert snapshot.worksheet_drawing_shapes.diagram_graphic_frame_count == 1
    assert snapshot.worksheet_drawing_shapes.diagram_data_part_count == 1
    assert snapshot.worksheet_drawing_shapes.diagram_drawing_part_count == 1
    assert snapshot.worksheet_drawing_shapes.unrecognized_shape_count == 0
    assert not any(
        "SmartArt" in warning for warning in snapshot.parser_warnings
    )
    assert "FF044" in {finding.rule_id for finding in report.findings}


def test_strict_worksheet_smartart_diagram_images_are_supported(tmp_path) -> None:
    baseline = make_strict_worksheet_smartart_image_model(tmp_path / "baseline.xlsx")
    candidate = make_strict_worksheet_smartart_image_model(tmp_path / "candidate.xlsx")
    change_worksheet_smartart_diagram_image_payload(candidate)

    snapshot = load_snapshot(baseline)
    report = compare_snapshots(snapshot, load_snapshot(candidate))

    assert snapshot.worksheet_drawing_shapes.diagram_data_part_count == 1
    assert snapshot.worksheet_drawing_shapes.diagram_image_part_count == 1
    assert snapshot.worksheet_drawing_shapes.fingerprinted_diagram_image_part_count == 1
    assert snapshot.worksheet_drawing_shapes.uninspected_diagram_image_part_count == 0
    assert snapshot.worksheet_drawing_shapes.unrecognized_shape_count == 0
    assert not any("SmartArt" in warning for warning in snapshot.parser_warnings)
    assert "FF044" in {finding.rule_id for finding in report.findings}


def test_worksheet_images_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_worksheet_image_model(tmp_path / "baseline.xlsx")
    presentation_candidate = make_worksheet_image_model(
        tmp_path / "presentation-candidate.xlsx"
    )
    payload_candidate = make_worksheet_image_model(tmp_path / "payload-candidate.xlsx")
    change_worksheet_image_presentation(presentation_candidate)
    change_worksheet_image_payload(payload_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    presentation_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(presentation_candidate),
    )
    payload_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(payload_candidate),
    )
    presentation_change = next(
        change
        for change in presentation_report.changes
        if change.kind == "worksheet_image_controls_changed"
    )
    payload_change = next(
        change
        for change in payload_report.changes
        if change.kind == "worksheet_image_controls_changed"
    )

    assert self_report.findings == []
    assert baseline_snapshot.summary()["worksheet_anchored_picture_count"] == 1
    assert baseline_snapshot.summary()["worksheet_background_image_count"] == 1
    assert baseline_snapshot.summary()["worksheet_header_footer_image_count"] == 1
    assert baseline_snapshot.summary()["has_worksheet_images"] is True
    assert profile["worksheet_images"] == {
        "present": True,
        "worksheet_image_sheet_count": 1,
        "anchored_picture_count": 1,
        "anchored_picture_anchor_count": 1,
        "worksheet_background_image_count": 1,
        "header_footer_image_count": 1,
        "image_part_count": 3,
        "fingerprinted_image_part_count": 3,
        "uninspected_image_part_count": 0,
        "related_relationship_count": 5,
        "external_relationship_count": 0,
        "unrecognized_image_count": 0,
    }
    assert "## Native worksheet image controls" in markdown
    assert presentation_change.details["worksheet_image_definition_material_changed"] is True
    assert payload_change.details["worksheet_image_payload_material_changed"] is True
    assert "FF059" in {finding.rule_id for finding in presentation_report.findings}
    assert "FF059" in {finding.rule_id for finding in payload_report.findings}
    assert "FF044" not in {finding.rule_id for finding in presentation_report.findings}
    assert "FF044" not in {finding.rule_id for finding in payload_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(presentation_report.to_dict()),
        report_to_markdown(presentation_report),
        json.dumps(report_to_sarif(presentation_report)),
        json.dumps(payload_report.to_dict()),
        report_to_markdown(payload_report),
        json.dumps(report_to_sarif(payload_report)),
    )
    for sensitive_value in (
        "PRIVATE-ANCHORED-IMAGE-NAME",
        "PRIVATE-ANCHORED-IMAGE-DESCRIPTION",
        "PRIVATE-WATERMARK-IMAGE-TITLE",
        "PrivateHeaderFooterWatermark",
        "private-anchored.png",
        "private-background.png",
        "private-watermark.png",
        "rIdFenceAnchoredImage",
        "rIdFenceBackgroundImage",
        "rIdFenceHeaderFooterImage",
        "position:absolute",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_worksheet_image_identifier_rewrites_are_ignored(tmp_path) -> None:
    baseline = make_worksheet_image_model(tmp_path / "baseline.xlsx")
    renumbered = make_worksheet_image_model(tmp_path / "renumbered.xlsx")
    renumber_worksheet_image_identifiers(renumbered)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(renumbered))

    assert "worksheet_image_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF059" not in {finding.rule_id for finding in report.findings}


def test_worksheet_image_source_relationship_rebinding_is_detected(tmp_path) -> None:
    baseline = make_worksheet_image_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_image_model(tmp_path / "candidate.xlsx")
    rebind_worksheet_image_source_relationship(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "worksheet_image_controls_changed"
    )

    assert baseline_snapshot.worksheet_images.related_relationship_count == 5
    assert candidate_snapshot.worksheet_images.related_relationship_count == 5
    assert change.details["worksheet_image_binding_changed"] is True
    assert change.details["worksheet_image_relationships_changed"] is True
    assert "FF059" in {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(candidate_snapshot)),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("drawing2.xml" not in artifact for artifact in rendered_artifacts)


def test_external_worksheet_images_are_counted_without_exposing_targets(tmp_path) -> None:
    baseline = make_worksheet_image_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_image_model(tmp_path / "candidate.xlsx")
    externalize_worksheet_image_relationship(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    candidate_profile = profile_snapshot(candidate_snapshot)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert candidate_snapshot.worksheet_images.anchored_picture_count == 1
    assert candidate_snapshot.worksheet_images.image_part_count == 2
    assert candidate_snapshot.worksheet_images.external_relationship_count == 1
    assert candidate_snapshot.worksheet_images.unrecognized_image_count == 0
    assert {finding.rule_id for finding in report.findings} >= {"FF059", "FF063"}
    rendered_artifacts = (
        json.dumps(candidate_profile),
        profile_to_markdown(candidate_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all(
        "PRIVATE-EXTERNAL-WORKSHEET-IMAGE" not in artifact
        for artifact in rendered_artifacts
    )


def test_strict_worksheet_images_are_supported(tmp_path) -> None:
    baseline = make_strict_worksheet_image_model(tmp_path / "baseline.xlsx")
    candidate = make_strict_worksheet_image_model(tmp_path / "candidate.xlsx")
    change_worksheet_image_payload(candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.worksheet_images.anchored_picture_count == 1
    assert baseline_snapshot.worksheet_images.worksheet_background_image_count == 1
    assert baseline_snapshot.worksheet_images.header_footer_image_count == 1
    assert baseline_snapshot.worksheet_images.image_part_count == 3
    assert baseline_snapshot.worksheet_images.unrecognized_image_count == 0
    assert not any(
        "native worksheet-image" in warning
        for warning in baseline_snapshot.parser_warnings
    )
    assert "FF059" in {finding.rule_id for finding in report.findings}


def test_malformed_worksheet_images_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_image_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_image_model(tmp_path / "candidate.xlsx")
    corrupt_worksheet_image_drawing_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.worksheet_images.unrecognized_image_count == 1
    assert any(
        "Worksheet DrawingML picture part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_image_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF059" in {finding.rule_id for finding in report.findings}


def test_worksheet_image_xml_budget_fails_closed(tmp_path, monkeypatch) -> None:
    workbook = make_worksheet_image_model(tmp_path / "candidate.xlsx")
    monkeypatch.setattr(workbook_module, "_WORKSHEET_IMAGE_MAX_XML_PART_BYTES", 1)

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_images.unrecognized_image_count >= 1
    assert any(
        "oversized native worksheet-image XML" in warning
        for warning in snapshot.parser_warnings
    )


def test_worksheet_image_relationship_xml_budget_fails_closed(
    tmp_path,
    monkeypatch,
) -> None:
    workbook = make_worksheet_image_model(tmp_path / "candidate.xlsx")
    budget_type = workbook_module._WorksheetImageXmlBudget
    monkeypatch.setattr(
        workbook_module,
        "_WorksheetImageXmlBudget",
        lambda: budget_type(remaining_parts=1),
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_images.unrecognized_image_count >= 1
    assert any(
        "bounded native worksheet-image XML part count budget" in warning
        for warning in snapshot.parser_warnings
    )


def test_chart_drawing_is_not_misclassified_as_worksheet_shapes(tmp_path) -> None:
    workbook = make_chart_definition_model(tmp_path / "chart.xlsx")

    snapshot = load_snapshot(workbook)

    assert snapshot.chart_definitions.present is True
    assert snapshot.worksheet_drawing_shapes.present is False
    assert snapshot.worksheet_drawing_shapes.shape_count == 0
    assert snapshot.worksheet_images.present is False
    assert snapshot.worksheet_images.worksheet_image_sheet_count == 0


def test_malformed_worksheet_drawing_shapes_fail_closed(tmp_path) -> None:
    baseline = make_worksheet_drawing_shape_model(tmp_path / "baseline.xlsx")
    candidate = make_worksheet_drawing_shape_model(tmp_path / "candidate.xlsx")
    corrupt_worksheet_drawing_shape_root(candidate)

    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(load_snapshot(baseline), candidate_snapshot)

    assert candidate_snapshot.worksheet_drawing_shapes.unrecognized_shape_count == 1
    assert any(
        "Worksheet DrawingML shape part with an unexpected root" in warning
        for warning in candidate_snapshot.parser_warnings
    )
    assert "worksheet_drawing_shape_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF044" in {finding.rule_id for finding in report.findings}


def test_worksheet_drawing_shape_xml_budget_fails_closed(tmp_path, monkeypatch) -> None:
    workbook = make_worksheet_drawing_shape_model(tmp_path / "candidate.xlsx")
    monkeypatch.setattr(
        workbook_module,
        "_WORKSHEET_DRAWING_SHAPE_MAX_XML_PART_BYTES",
        1,
    )

    snapshot = load_snapshot(workbook)

    assert snapshot.worksheet_drawing_shapes.unrecognized_shape_count >= 1
    assert any(
        "oversized Worksheet DrawingML shape XML part" in warning
        for warning in snapshot.parser_warnings
    )


def test_ignored_error_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_ignored_error_model(tmp_path / "baseline.xlsx")
    candidate = make_ignored_error_model(tmp_path / "candidate.xlsx")
    extension_candidate = make_ignored_error_model(tmp_path / "extension-candidate.xlsx")
    change_ignored_error_target(candidate)
    change_ignored_error_extension_target(extension_candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    extension_report = compare_snapshots(
        baseline_snapshot,
        load_snapshot(extension_candidate),
    )
    change = next(
        change
        for change in report.changes
        if change.kind == "ignored_error_controls_changed"
    )

    assert baseline_snapshot.summary()["ignored_error_rule_count"] == 11
    assert baseline_snapshot.summary()["ignored_error_target_range_count"] == 5
    assert baseline_snapshot.summary()["has_ignored_error_controls"] is True
    assert profile["ignored_error_controls"] == {
        "present": True,
        "worksheet_count": 2,
        "standard_container_count": 1,
        "extension_container_count": 1,
        "ignored_error_rule_count": 11,
        "target_range_count": 5,
        "evaluation_error_count": 2,
        "inconsistent_formula_count": 2,
        "formula_range_omission_count": 1,
        "unlocked_formula_count": 1,
        "empty_cell_reference_count": 1,
        "list_data_validation_count": 1,
        "calculated_column_count": 1,
        "number_stored_as_text_count": 1,
        "two_digit_text_year_count": 1,
        "unrecognized_ignored_error_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Ignored Excel error-checking controls" in markdown
    assert change.details["ignored_error_definition_material_changed"] is True
    assert "FF037" in {finding.rule_id for finding in report.findings}
    assert "FF037" in {finding.rule_id for finding in extension_report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in ("C2:C3", "C4:C5"):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_ignored_error_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_ignored_error_model(tmp_path / "baseline.xlsx")
    equivalent = make_ignored_error_model(tmp_path / "equivalent.xlsx")
    normalize_ignored_error_control_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert "ignored_error_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF037" not in {finding.rule_id for finding in report.findings}


def test_ignored_error_malformed_control_fails_closed(tmp_path) -> None:
    baseline = make_ignored_error_model(tmp_path / "baseline.xlsx")
    malformed = make_ignored_error_model(tmp_path / "malformed.xlsx")
    corrupt_ignored_error_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.ignored_error_controls.unrecognized_ignored_error_count == 1
    assert any(
        "malformed or unsupported ignored-error" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF037"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all(
        "PrivateIgnoredErrorSheet!B2" not in artifact
        for artifact in rendered_artifacts
    )


def test_ignored_error_duplicate_standard_container_fails_closed(tmp_path) -> None:
    baseline = make_ignored_error_model(tmp_path / "baseline.xlsx")
    malformed = make_ignored_error_model(tmp_path / "malformed.xlsx")
    duplicate_ignored_error_container(malformed)

    malformed_snapshot = load_snapshot(malformed)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.ignored_error_controls.unrecognized_ignored_error_count == 1
    assert any(
        "malformed or unsupported ignored-error" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF037"} <= {finding.rule_id for finding in report.findings}


def test_ignored_error_free_workbook_has_no_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.ignored_error_controls.present is False
    assert snapshot.ignored_error_controls.ignored_error_rule_count == 0
    assert not any("ignored-error" in warning for warning in snapshot.parser_warnings)


def test_named_sheet_views_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_named_sheet_view_model(tmp_path / "baseline.xlsx")
    candidate = make_named_sheet_view_model(tmp_path / "candidate.xlsx")
    change_named_sheet_view_criterion(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "named_sheet_views_changed"
    )

    assert baseline_snapshot.summary()["named_sheet_view_count"] == 2
    assert baseline_snapshot.summary()["named_sheet_view_filter_count"] == 2
    assert baseline_snapshot.summary()["has_named_sheet_views"] is True
    assert profile["named_sheet_views"] == {
        "present": True,
        "worksheet_count": 1,
        "part_count": 1,
        "named_sheet_view_count": 2,
        "named_filter_count": 2,
        "column_filter_count": 2,
        "filter_criterion_count": 2,
        "sort_rule_count": 2,
        "sort_condition_count": 2,
        "unrecognized_named_sheet_view_count": 0,
    }
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Excel Named Sheet Views" in markdown
    assert change.details["named_sheet_view_definition_material_changed"] is True
    assert "FF038" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "Private baseline review",
        "Private alternate review",
        "PRIVATE-NAMED-VIEW-REGION",
        "PRIVATE-ALTERNATE-NAMED-VIEW-REGION",
        "CANDIDATE-PRIVATE-NAMED-VIEW-REGION",
        "PRIVATE-NAMED-VIEW-SORT-LIST",
        "A1:C5",
        "C2:C5",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_named_sheet_views_reconcile_table_owned_filters(tmp_path) -> None:
    baseline = make_named_sheet_view_model(
        tmp_path / "baseline.xlsx",
        table_owned=True,
    )
    candidate = make_named_sheet_view_model(
        tmp_path / "candidate.xlsx",
        table_owned=True,
    )
    change_named_sheet_view_criterion(candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.named_sheet_views.unrecognized_named_sheet_view_count == 0
    assert not any(
        "Named Sheet View" in warning for warning in baseline_snapshot.parser_warnings
    )
    assert "FF038" in {finding.rule_id for finding in report.findings}


def test_named_sheet_view_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_named_sheet_view_model(tmp_path / "baseline.xlsx")
    equivalent = make_named_sheet_view_model(tmp_path / "equivalent.xlsx")
    normalize_named_sheet_view_control_spelling(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert "named_sheet_views_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF038" not in {finding.rule_id for finding in report.findings}


def test_named_sheet_view_malformed_control_fails_closed(tmp_path) -> None:
    baseline = make_named_sheet_view_model(tmp_path / "baseline.xlsx")
    malformed = make_named_sheet_view_model(tmp_path / "malformed.xlsx")
    corrupt_named_sheet_view_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.named_sheet_views.unrecognized_named_sheet_view_count == 1
    assert any(
        "malformed or unsupported Named Sheet View" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF038"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("4294967296" not in artifact for artifact in rendered_artifacts)


def test_named_sheet_view_free_workbook_has_no_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.named_sheet_views.present is False
    assert snapshot.named_sheet_views.named_sheet_view_count == 0
    assert not any("Named Sheet View" in warning for warning in snapshot.parser_warnings)


def test_custom_workbook_views_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_custom_workbook_view_model(tmp_path / "baseline.xlsx")
    candidate = make_custom_workbook_view_model(tmp_path / "candidate.xlsx")
    change_custom_workbook_view_filter(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(
        change
        for change in report.changes
        if change.kind == "custom_workbook_views_changed"
    )

    assert baseline_snapshot.summary()["custom_workbook_view_count"] == 2
    assert baseline_snapshot.summary()["custom_sheet_view_count"] == 4
    assert baseline_snapshot.summary()["custom_view_sheet_count"] == 2
    assert baseline_snapshot.summary()["has_custom_workbook_views"] is True
    assert profile["custom_workbook_views"] == {
        "present": True,
        "custom_workbook_view_count": 2,
        "custom_sheet_view_count": 4,
        "custom_view_sheet_count": 2,
        "hidden_row_or_column_view_count": 1,
        "filtered_view_count": 1,
        "print_setting_view_count": 2,
        "display_setting_view_count": 1,
        "unrecognized_custom_view_count": 0,
    }
    assert baseline_snapshot.parser_warnings == ()
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Legacy Excel Custom Views" in markdown
    assert change.details["custom_workbook_view_definition_material_changed"] is True
    assert "FF060" in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE Executive View",
        "PRIVATE Print View",
        "PRIVATE-CUSTOM-VIEW-FILTER",
        "CANDIDATE-PRIVATE-CUSTOM-VIEW-FILTER",
        "{01234567-89AB-CDEF-0123-456789ABCDEF}",
        "A1:C3",
        "B2",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_custom_workbook_view_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_custom_workbook_view_model(tmp_path / "baseline.xlsx")
    equivalent = make_custom_workbook_view_model(tmp_path / "equivalent.xlsx")
    normalize_custom_workbook_view_identifiers(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []
    assert "custom_workbook_views_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF060" not in {finding.rule_id for finding in report.findings}


def test_custom_workbook_views_support_strict_spreadsheetml(tmp_path) -> None:
    baseline = make_custom_workbook_view_model(tmp_path / "baseline.xlsx")
    strict = make_strict_custom_workbook_view_model(tmp_path / "strict.xlsx")

    baseline_snapshot = load_snapshot(baseline)
    strict_snapshot = load_snapshot(strict)
    report = compare_snapshots(baseline_snapshot, strict_snapshot)

    assert strict_snapshot.custom_workbook_views.to_dict() == (
        baseline_snapshot.custom_workbook_views.to_dict()
    )
    assert strict_snapshot.custom_workbook_views.unrecognized_custom_view_count == 0
    assert not any("Custom View" in warning for warning in strict_snapshot.parser_warnings)
    assert "custom_workbook_views_changed" not in {
        change.kind for change in report.changes
    }


def test_custom_workbook_views_cover_chart_sheet_alternate_print_state(tmp_path) -> None:
    baseline = make_chart_sheet_custom_workbook_view_model(tmp_path / "baseline.xlsx")
    candidate = make_chart_sheet_custom_workbook_view_model(tmp_path / "candidate.xlsx")
    change_chart_sheet_custom_workbook_view_scale(candidate)

    baseline_snapshot = load_snapshot(baseline)
    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))

    assert baseline_snapshot.custom_workbook_views.to_dict() == {
        "present": True,
        "custom_workbook_view_count": 1,
        "custom_sheet_view_count": 2,
        "custom_view_sheet_count": 2,
        "hidden_row_or_column_view_count": 0,
        "filtered_view_count": 0,
        "print_setting_view_count": 1,
        "display_setting_view_count": 1,
        "unrecognized_custom_view_count": 0,
    }
    assert baseline_snapshot.summary()["sheet_count"] == 1
    assert not any("Custom View" in warning for warning in baseline_snapshot.parser_warnings)
    assert "custom_workbook_views_changed" in {
        change.kind for change in report.changes
    }
    assert "FF060" in {finding.rule_id for finding in report.findings}


def test_empty_chart_sheet_custom_view_container_is_not_malformed(tmp_path) -> None:
    snapshot = load_snapshot(
        make_empty_chart_sheet_custom_view_container_model(tmp_path / "chart.xlsx")
    )

    assert snapshot.custom_workbook_views.to_dict() == {
        "present": False,
        "custom_workbook_view_count": 0,
        "custom_sheet_view_count": 0,
        "custom_view_sheet_count": 0,
        "hidden_row_or_column_view_count": 0,
        "filtered_view_count": 0,
        "print_setting_view_count": 0,
        "display_setting_view_count": 0,
        "unrecognized_custom_view_count": 0,
    }
    assert not any("Custom View" in warning for warning in snapshot.parser_warnings)


def test_chart_sheet_custom_workbook_view_malformed_print_metadata_fails_closed(
    tmp_path,
) -> None:
    baseline = make_chart_sheet_custom_workbook_view_model(tmp_path / "baseline.xlsx")
    malformed = make_chart_sheet_custom_workbook_view_model(tmp_path / "malformed.xlsx")
    corrupt_chart_sheet_custom_workbook_view_page_setup(malformed)

    malformed_snapshot = load_snapshot(malformed)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.custom_workbook_views.unrecognized_custom_view_count >= 1
    assert any("Custom View" in warning for warning in malformed_snapshot.parser_warnings)
    assert {"FF010", "FF060"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(malformed_snapshot)),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all(
        "PRIVATE-INVALID-CHART-ORIENTATION" not in artifact
        for artifact in rendered_artifacts
    )


def test_custom_workbook_view_malformed_control_fails_closed(tmp_path) -> None:
    baseline = make_custom_workbook_view_model(tmp_path / "baseline.xlsx")
    malformed = make_custom_workbook_view_model(tmp_path / "malformed.xlsx")
    corrupt_custom_workbook_view_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.custom_workbook_views.unrecognized_custom_view_count >= 1
    assert any("Custom View" in warning for warning in malformed_snapshot.parser_warnings)
    assert {"FF010", "FF060"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all("4294967296" not in artifact for artifact in rendered_artifacts)


def test_custom_workbook_view_unbound_sheet_view_fails_closed(tmp_path) -> None:
    baseline = make_custom_workbook_view_model(tmp_path / "baseline.xlsx")
    malformed = make_custom_workbook_view_model(tmp_path / "malformed.xlsx")
    unbind_custom_workbook_view(malformed)

    malformed_snapshot = load_snapshot(malformed)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.custom_workbook_views.unrecognized_custom_view_count >= 1
    assert any("Custom View" in warning for warning in malformed_snapshot.parser_warnings)
    assert {"FF010", "FF060"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(profile_snapshot(malformed_snapshot)),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    assert all(
        "{BBBBBBBB-CCCC-DDDD-EEEE-FFFFFFFFFFFF}" not in artifact
        for artifact in rendered_artifacts
    )


def test_custom_workbook_view_free_workbook_has_no_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.custom_workbook_views.present is False
    assert snapshot.custom_workbook_views.custom_workbook_view_count == 0
    assert not any("Custom View" in warning for warning in snapshot.parser_warnings)


def test_table_style_controls_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_table_style_control_model(tmp_path / "baseline.xlsx")
    candidate = make_table_style_control_model(tmp_path / "candidate.xlsx")
    change_table_style_definition(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "table_style_controls_changed"
    )

    assert baseline_snapshot.summary()["table_style_info_count"] == 1
    assert baseline_snapshot.summary()["custom_table_style_count"] == 1
    assert baseline_snapshot.summary()["has_table_style_controls"] is True
    assert profile["table_style_controls"] == {
        "present": True,
        "table_style_info_count": 1,
        "styled_table_count": 1,
        "custom_table_style_count": 1,
        "custom_table_style_element_count": 3,
        "custom_style_applied_table_count": 1,
        "table_direct_dxf_assignment_count": 1,
        "table_direct_dxf_table_count": 1,
        "table_named_cell_style_assignment_count": 1,
        "row_striped_table_count": 1,
        "column_striped_table_count": 0,
        "emphasized_column_table_count": 1,
        "unrecognized_table_style_count": 0,
    }
    assert baseline_snapshot.parser_warnings == ()
    assert baseline_snapshot.tables == candidate_snapshot.tables
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Excel Table Style Controls" in markdown
    assert change.details["table_style_definition_material_changed"] is True
    assert "FF061" in {finding.rule_id for finding in report.findings}
    assert "FF013" not in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-FINANCE-TABLE-STYLE",
        "PRIVATE-FINANCE-DATA-STYLE",
        "FF113355",
        "FF557799",
        "FF99BBDD",
        "FFCC8844",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_table_style_control_toggles_are_detected_without_table_ref_noise(
    tmp_path,
) -> None:
    baseline = make_table_style_control_model(tmp_path / "baseline.xlsx")
    candidate = make_table_style_control_model(tmp_path / "candidate.xlsx")
    change_table_style_control(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "table_style_controls_changed"
    )

    assert baseline_snapshot.tables == candidate_snapshot.tables
    assert change.details["before"]["row_striped_table_count"] == 1
    assert change.details["after"]["row_striped_table_count"] == 0
    assert change.details["after"]["column_striped_table_count"] == 1
    assert "FF061" in {finding.rule_id for finding in report.findings}
    assert "FF013" not in {finding.rule_id for finding in report.findings}


def test_table_direct_dxf_format_changes_are_detected_without_cell_edits(
    tmp_path,
) -> None:
    baseline = make_table_style_control_model(tmp_path / "baseline.xlsx")
    candidate = make_table_style_control_model(tmp_path / "candidate.xlsx")
    change_table_direct_dxf_assignment(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)

    assert baseline_snapshot.tables == candidate_snapshot.tables
    assert (
        baseline_snapshot.table_style_controls.table_direct_dxf_assignment_count
        == 1
    )
    assert (
        candidate_snapshot.table_style_controls.table_direct_dxf_assignment_count
        == 1
    )
    assert "table_style_controls_changed" in {
        change.kind for change in report.changes
    }
    assert "FF061" in {finding.rule_id for finding in report.findings}
    assert "FF013" not in {finding.rule_id for finding in report.findings}


def test_table_style_control_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_table_style_control_model(tmp_path / "baseline.xlsx")
    equivalent = make_table_style_control_model(tmp_path / "equivalent.xlsx")
    normalize_table_style_control_writer_noise(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []
    assert "table_style_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF061" not in {finding.rule_id for finding in report.findings}


def test_table_style_controls_support_strict_spreadsheetml(tmp_path) -> None:
    baseline = make_table_style_control_model(tmp_path / "baseline.xlsx")
    strict = make_strict_table_style_control_model(tmp_path / "strict.xlsx")

    baseline_snapshot = load_snapshot(baseline)
    strict_snapshot = load_snapshot(strict)
    report = compare_snapshots(baseline_snapshot, strict_snapshot)

    assert strict_snapshot.table_style_controls.to_dict() == (
        baseline_snapshot.table_style_controls.to_dict()
    )
    assert strict_snapshot.table_style_controls.unrecognized_table_style_count == 0
    assert not any("Table Style" in warning for warning in strict_snapshot.parser_warnings)
    assert "table_style_controls_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF061" not in {finding.rule_id for finding in report.findings}


def test_malformed_table_style_control_fails_closed_and_is_redacted(tmp_path) -> None:
    baseline = make_table_style_control_model(tmp_path / "baseline.xlsx")
    malformed = make_table_style_control_model(tmp_path / "malformed.xlsx")
    corrupt_table_style_control(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert malformed_snapshot.table_style_controls.unrecognized_table_style_count >= 1
    assert any("Table Style" in warning for warning in malformed_snapshot.parser_warnings)
    assert {"FF010", "FF061"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-FINANCE-TABLE-STYLE",
        "PRIVATE-TABLE-STYLE-CONTROL",
        "4294967296",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_table_style_free_table_has_no_control_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_table_model(tmp_path / "ordinary-table.xlsx"))

    assert snapshot.table_style_controls.present is False
    assert snapshot.table_style_controls.table_style_info_count == 0
    assert snapshot.table_style_controls.custom_table_style_count == 0
    assert not any("Table Style" in warning for warning in snapshot.parser_warnings)


def test_shared_workbook_revisions_are_profiled_diffed_and_redacted(tmp_path) -> None:
    baseline = make_shared_workbook_revision_model(tmp_path / "baseline.xlsx")
    candidate = make_shared_workbook_revision_model(tmp_path / "candidate.xlsx")
    change_shared_workbook_revision_log(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)
    self_report = compare_snapshots(baseline_snapshot, load_snapshot(baseline))
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "shared_workbook_revisions_changed"
    )

    assert baseline_snapshot.summary()["revision_header_count"] == 1
    assert baseline_snapshot.summary()["revision_log_entry_count"] == 3
    assert baseline_snapshot.summary()["has_shared_workbook_revisions"] is True
    assert profile["shared_workbook_revisions"] == {
        "present": True,
        "revision_header_part_count": 1,
        "revision_header_count": 1,
        "revision_log_part_count": 1,
        "revision_log_entry_count": 3,
        "shared_workbook_enabled_count": 1,
        "track_revisions_enabled_count": 1,
        "revision_history_enabled_count": 1,
        "keep_change_history_enabled_count": 1,
        "revision_history_protected_count": 1,
        "unrecognized_shared_workbook_revision_count": 0,
    }
    assert baseline_snapshot.parser_warnings == ()
    assert baseline_snapshot.cells == candidate_snapshot.cells
    assert baseline_snapshot.sheets == candidate_snapshot.sheets
    assert self_report.changes == []
    assert self_report.findings == []
    assert "## Legacy Shared-Workbook Revision History" in markdown
    assert change.details["revision_log_material_changed"] is True
    assert "revision_header_material_changed" not in change.details
    assert "FF062" in {finding.rule_id for finding in report.findings}
    assert "FF001" not in {finding.rule_id for finding in report.findings}

    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-REVISION-AUTHOR",
        "PRIVATE-OLD-REVISION-VALUE",
        "CANDIDATE-PRIVATE-OLD-REVISION-VALUE",
        "PRIVATE-NEW-REVISION-VALUE",
        "{11111111-2222-3333-4444-555555555555}",
        "2026-07-26T01:02:03Z",
        "rIdPrivateRevisionLog",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_shared_workbook_revision_controls_are_detected_without_cell_noise(
    tmp_path,
) -> None:
    baseline = make_shared_workbook_revision_model(tmp_path / "baseline.xlsx")
    candidate = make_shared_workbook_revision_model(tmp_path / "candidate.xlsx")
    change_shared_workbook_revision_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    candidate_snapshot = load_snapshot(candidate)
    report = compare_snapshots(baseline_snapshot, candidate_snapshot)
    change = next(
        change
        for change in report.changes
        if change.kind == "shared_workbook_revisions_changed"
    )

    assert baseline_snapshot.cells == candidate_snapshot.cells
    assert baseline_snapshot.sheets == candidate_snapshot.sheets
    assert change.details["revision_header_material_changed"] is True
    assert change.details["before"]["keep_change_history_enabled_count"] == 1
    assert change.details["after"]["keep_change_history_enabled_count"] == 0
    assert change.details["after"]["revision_history_protected_count"] == 0
    assert "FF062" in {finding.rule_id for finding in report.findings}
    assert "FF001" not in {finding.rule_id for finding in report.findings}


def test_shared_workbook_revision_writer_noise_is_normalized(tmp_path) -> None:
    baseline = make_shared_workbook_revision_model(tmp_path / "baseline.xlsx")
    equivalent = make_shared_workbook_revision_model(tmp_path / "equivalent.xlsx")
    normalize_shared_workbook_revision_writer_noise(equivalent)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(equivalent))

    assert report.changes == []
    assert report.findings == []
    assert "shared_workbook_revisions_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF062" not in {finding.rule_id for finding in report.findings}


def test_shared_workbook_revisions_support_strict_spreadsheetml(tmp_path) -> None:
    baseline = make_shared_workbook_revision_model(tmp_path / "baseline.xlsx")
    strict = make_strict_shared_workbook_revision_model(tmp_path / "strict.xlsx")

    baseline_snapshot = load_snapshot(baseline)
    strict_snapshot = load_snapshot(strict)
    report = compare_snapshots(baseline_snapshot, strict_snapshot)

    assert strict_snapshot.shared_workbook_revisions.to_dict() == (
        baseline_snapshot.shared_workbook_revisions.to_dict()
    )
    assert (
        strict_snapshot.shared_workbook_revisions.unrecognized_shared_workbook_revision_count
        == 0
    )
    assert not any(
        "shared-workbook revision" in warning
        for warning in strict_snapshot.parser_warnings
    )
    assert "shared_workbook_revisions_changed" not in {
        change.kind for change in report.changes
    }
    assert "FF062" not in {finding.rule_id for finding in report.findings}


def test_malformed_shared_workbook_revision_fails_closed_and_is_redacted(
    tmp_path,
) -> None:
    baseline = make_shared_workbook_revision_model(tmp_path / "baseline.xlsx")
    malformed = make_shared_workbook_revision_model(tmp_path / "malformed.xlsx")
    corrupt_shared_workbook_revision(malformed)

    malformed_snapshot = load_snapshot(malformed)
    malformed_profile = profile_snapshot(malformed_snapshot)
    report = compare_snapshots(load_snapshot(baseline), malformed_snapshot)

    assert (
        malformed_snapshot.shared_workbook_revisions.unrecognized_shared_workbook_revision_count
        >= 1
    )
    assert any(
        "shared-workbook revision" in warning
        for warning in malformed_snapshot.parser_warnings
    )
    assert {"FF010", "FF062"} <= {finding.rule_id for finding in report.findings}
    rendered_artifacts = (
        json.dumps(malformed_profile),
        profile_to_markdown(malformed_profile),
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in (
        "PRIVATE-REVISION-AUTHOR",
        "PRIVATE-OLD-REVISION-VALUE",
        "PRIVATE-NEW-REVISION-VALUE",
        "PRIVATE-REVISION-CONTROL",
    ):
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_shared_workbook_revision_free_workbook_has_no_inventory(tmp_path) -> None:
    snapshot = load_snapshot(make_model(tmp_path / "ordinary.xlsx"))

    assert snapshot.shared_workbook_revisions.present is False
    assert snapshot.shared_workbook_revisions.revision_header_part_count == 0
    assert snapshot.shared_workbook_revisions.revision_log_entry_count == 0
    assert not any(
        "shared-workbook revision" in warning for warning in snapshot.parser_warnings
    )


def test_power_query_material_is_guarded_without_leaking_query_contents(tmp_path) -> None:
    baseline = make_power_query_model(tmp_path / "baseline.xlsx")
    candidate = make_power_query_model(tmp_path / "candidate.xlsx")
    change_power_query_controls(candidate)

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    markdown = profile_to_markdown(profile)

    assert baseline_snapshot.summary()["power_query_mashup_count"] == 1
    assert baseline_snapshot.summary()["power_query_formula_document_count"] == 1
    assert baseline_snapshot.summary()["power_query_metadata_item_count"] == 1
    assert profile["query_table_refresh_controls"] == [
        {
            "sheet": "Inputs",
            "connection_id": 1,
            "refresh_on_load": True,
            "background_refresh": True,
            "refresh_disabled": False,
            "remove_data_on_save": False,
            "fill_formulas": False,
            "connection_edit_disabled": True,
            "growth_behavior": "insert_clear",
            "has_name": True,
            "has_refresh_metadata": False,
            "opaque_metadata": {"present": False, "count": 0},
        }
    ]
    assert profile["power_query"] == {
        "present": True,
        "mashup_count": 1,
        "parsed_mashup_count": 1,
        "formula_document_count": 1,
        "package_part_count": 4,
        "embedded_content_part_count": 1,
        "metadata_document_count": 1,
        "metadata_item_count": 1,
        "permission_controls": {
            "payload_count": 1,
            "parsed_count": 1,
            "firewall_enabled_count": 1,
            "future_packages_allowed_count": 0,
            "workbook_group_type_count": 0,
            "opaque_metadata": {"present": False, "count": 0},
        },
        "permission_binding_count": 1,
        "opaque_metadata": {"present": False, "count": 0},
    }
    assert "## Power Query controls" in markdown
    assert "## Query-table refresh controls" in markdown

    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    power_query_change = next(
        change for change in report.changes if change.kind == "power_query_changed"
    )
    assert power_query_change.details["formula_material_changed"] is True
    assert power_query_change.details["metadata_control_material_changed"] is True
    assert power_query_change.details["permission_controls_changed"] is True
    assert {finding.rule_id for finding in report.findings} >= {"FF024"}

    sensitive_values = (
        "private-baseline-power-query-token",
        "private-candidate-power-query-token",
        "private-baseline-package-config",
        "private-baseline-embedded-content",
        "Private revenue query",
        "private-baseline-target",
        "private-refresh-message",
        "11111111-2222-3333-4444-555555555555",
        "aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee",
    )
    rendered_artifacts = (
        json.dumps(profile),
        markdown,
        json.dumps(report.to_dict()),
        report_to_markdown(report),
        json.dumps(report_to_sarif(report)),
    )
    for sensitive_value in sensitive_values:
        assert all(sensitive_value not in artifact for artifact in rendered_artifacts)


def test_power_query_ignores_volatile_refresh_and_user_binding_noise(tmp_path) -> None:
    baseline = make_power_query_model(tmp_path / "baseline.xlsx")
    candidate = make_power_query_model(tmp_path / "candidate.xlsx")
    change_power_query_refresh_noise(candidate)

    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert "power_query_changed" not in {change.kind for change in report.changes}
    assert "FF024" not in {finding.rule_id for finding in report.findings}


def test_current_row_table_references_trace_only_the_matching_row(tmp_path) -> None:
    baseline = make_current_row_table_model(tmp_path / "baseline.xlsx")
    candidate = make_current_row_table_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Data"]["A2"], "value", 100))

    baseline_snapshot = load_snapshot(baseline)
    assert baseline_snapshot.unresolved_reference_tokens == {}
    assert baseline_snapshot.direct_dependents(("Data", "A2")) == {
        ("Data", "C2"),
        ("Data", "E2"),
    }
    assert baseline_snapshot.direct_dependents(("Data", "A3")) == {("Data", "C3")}
    assert baseline_snapshot.direct_dependents(("Data", "A4")) == {("Data", "C4")}

    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(change for change in report.changes if change.location == ("Data", "A2"))

    assert change.impacted_cells == (("Data", "C2"), ("Data", "E2"), ("Report", "B2"))
    assert ("Data", "C3") not in change.impacted_cells
    assert ("Data", "C4") not in change.impacted_cells


def test_three_d_references_trace_every_sheet_in_the_tab_span(tmp_path) -> None:
    baseline = make_three_d_model(tmp_path / "baseline.xlsx")
    candidate = make_three_d_model(tmp_path / "candidate.xlsx")
    rewrite(candidate, lambda workbook: setattr(workbook["Feb"]["B2"], "value", 200))

    baseline_snapshot = load_snapshot(baseline)
    profile = profile_snapshot(baseline_snapshot)
    assert baseline_snapshot.unresolved_reference_tokens == {}
    assert baseline_snapshot.summary()["three_d_reference_cells"] == 1
    for sheet in ("Jan", "Feb", "Mar"):
        assert baseline_snapshot.direct_dependents((sheet, "B2")) == {("Summary", "B2")}
    assert baseline_snapshot.direct_dependents(("Jan:Mar", "B2")) == set()
    assert profile["features"]["three_d_reference_cells"] == [
        {"location": "Summary!B2", "tokens": ["Jan:Mar!B2"]}
    ]
    assert "## 3-D worksheet references" in profile_to_markdown(profile)

    report = compare_snapshots(baseline_snapshot, load_snapshot(candidate))
    change = next(change for change in report.changes if change.location == ("Feb", "B2"))

    assert change.impacted_cells == (("Summary", "B2"),)


def test_three_d_reference_scope_change_is_reported_when_tabs_move(tmp_path) -> None:
    baseline = make_three_d_model(tmp_path / "baseline.xlsx")
    candidate = make_three_d_model(tmp_path / "candidate.xlsx")

    def move_february_after_march(workbook) -> None:
        workbook._sheets = [  # noqa: SLF001 - sheet tab order is the scenario under test
            workbook["Jan"],
            workbook["Mar"],
            workbook["Feb"],
            workbook["Summary"],
        ]

    rewrite(candidate, move_february_after_march)
    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))
    scope_change = next(
        change
        for change in report.changes
        if change.kind == "three_d_reference_scope_changed"
    )

    assert scope_change.location == ("Summary", "B2")
    assert scope_change.details == {
        "references": [
            {
                "token": "Jan:Mar!B2",
                "before_sheets": ["Jan", "Feb", "Mar"],
                "after_sheets": ["Jan", "Mar"],
            }
        ]
    }
    assert any(finding.rule_id == "FF014" for finding in report.findings)


def test_three_d_references_include_a_sheet_inserted_between_tab_endpoints(tmp_path) -> None:
    baseline = make_three_d_model(tmp_path / "baseline.xlsx")
    candidate = make_three_d_model(tmp_path / "candidate.xlsx")

    def insert_period(workbook) -> None:
        inserted = workbook.create_sheet("Feb Extra", 1)
        inserted["A1"] = "Period input"
        inserted["B2"] = 25

    rewrite(candidate, insert_period)
    after = load_snapshot(candidate)
    assert after.direct_dependents(("Feb Extra", "B2")) == {("Summary", "B2")}

    report = compare_snapshots(load_snapshot(baseline), after)
    change = next(change for change in report.changes if change.location == ("Feb Extra", "B2"))

    assert change.kind == "value_added"
    assert change.impacted_cells == (("Summary", "B2"),)
    assert any(finding.rule_id == "FF014" for finding in report.findings)


def test_snapshot_captures_parser_coverage_warnings(tmp_path, monkeypatch) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    import formulafence.workbook as workbook_module

    original_load_workbook = workbook_module.load_workbook

    def noisy_load_workbook(*args, **kwargs):
        warnings.warn("fixture-only unsupported extension", UserWarning, stacklevel=2)
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(workbook_module, "load_workbook", noisy_load_workbook)
    snapshot = load_snapshot(baseline)

    assert snapshot.parser_warnings == ("fixture-only unsupported extension",)


def test_diff_surfaces_new_parser_coverage_warning(tmp_path, monkeypatch) -> None:
    baseline = make_model(tmp_path / "baseline.xlsx")
    candidate = make_model(tmp_path / "candidate.xlsx")
    import formulafence.workbook as workbook_module

    original_load_workbook = workbook_module.load_workbook

    def conditionally_noisy_load_workbook(path, *args, **kwargs):
        if str(path) == str(candidate):
            warnings.warn("candidate-only unsupported extension", UserWarning, stacklevel=2)
        return original_load_workbook(path, *args, **kwargs)

    monkeypatch.setattr(workbook_module, "load_workbook", conditionally_noisy_load_workbook)
    report = compare_snapshots(load_snapshot(baseline), load_snapshot(candidate))

    assert any(finding.rule_id == "FF010" for finding in report.findings)
