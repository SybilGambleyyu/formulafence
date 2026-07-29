from openpyxl.utils import FORMULAE

from formulafence.formulas import (
    _EXCEL_UNQUALIFIED_NATIVE_FUNCTIONS,
    ParsedReference,
    StructuredTable,
    conditional_aggregate_range_shape_mismatches,
    extract_references,
    formula_fingerprint,
    has_broken_reference,
    inspect_formula,
    lambda_parameter_count,
    parse_external_link_indexed_defined_name_reference,
    parse_external_link_indexed_sheet_defined_name_reference,
    parse_external_link_indexed_workbook_reference,
    parse_external_link_indexed_workbook_structured_reference,
    parse_external_link_indexed_workbook_three_d_reference,
    parse_external_workbook_defined_name_reference,
    parse_external_workbook_reference,
    parse_external_workbook_sheet_defined_name_reference,
    parse_external_workbook_structured_reference,
    parse_external_workbook_three_d_reference,
    parse_workbook_defined_name_alias,
)
from formulafence.models import (
    ExternalWorkbookReference,
    ExternalWorkbookStructuredReference,
    ExternalWorkbookThreeDReference,
)


def test_fingerprint_normalises_relative_copy_patterns() -> None:
    assert formula_fingerprint("=A2*2", "B2") == formula_fingerprint("=A3*2", "B3")
    assert formula_fingerprint("=$A2*2", "B2") == formula_fingerprint("=$A3*2", "B3")
    assert formula_fingerprint("=A2*2", "B2") != formula_fingerprint("=A2*3", "B2")
    assert formula_fingerprint("=A1#", "B1") == formula_fingerprint("=A2#", "B2")
    assert formula_fingerprint("=A1#", "B1") == formula_fingerprint(
        "=_xlfn.ANCHORARRAY(A1)", "B1"
    )
    assert formula_fingerprint("=A1#", "B1") != formula_fingerprint("=A1", "B1")
    assert formula_fingerprint("=@A1:A3", "B2") == formula_fingerprint(
        "=@A2:A4", "B3"
    )
    assert formula_fingerprint("=@A1:A3", "B2") != formula_fingerprint("=A1:A3", "B2")
    assert formula_fingerprint("=@A1:A3", "B2") == formula_fingerprint(
        "=_xlfn.SINGLE(A1:A3)", "B2"
    )


def test_broken_reference_detection_accepts_only_error_operands() -> None:
    assert has_broken_reference("=#REF!")
    assert has_broken_reference("=SUM(A1,#REF!,C1)")
    assert has_broken_reference("=IFERROR(#REF!,0)")

    assert not has_broken_reference('="#REF!"')
    assert not has_broken_reference('=IF(A1="#REF!",1,0)')
    assert not has_broken_reference("='#REF!'!A1")
    assert not has_broken_reference('=INDIRECT("#REF!")')


def test_conditional_aggregate_range_shape_mismatches_find_direct_conditional_aggregates() -> None:
    assert conditional_aggregate_range_shape_mismatches(
        "=SUMIFS(C2:C10,A2:A12,A14,B2:B12,B14)"
    ) == (("SUMIFS", 2),)
    assert conditional_aggregate_range_shape_mismatches(
        "=COUNTIFS(A2:A10,\">0\",B2:B9,\"open\")"
    ) == (("COUNTIFS", 1),)
    for function_name in ("AVERAGEIFS", "MAXIFS", "MINIFS"):
        assert conditional_aggregate_range_shape_mismatches(
            f"={function_name}(C2:C10,A2:A12,A14)"
        ) == ((function_name, 1),)
    assert conditional_aggregate_range_shape_mismatches(
        "=_xlfn.MAXIFS(C2:C10,A2:A12,A14)"
    ) == (("MAXIFS", 1),)
    assert conditional_aggregate_range_shape_mismatches(
        "=@_xlfn.MINIFS(C2:C10,A2:A12,A14)"
    ) == (("MINIFS", 1),)
    assert conditional_aggregate_range_shape_mismatches(
        "=IFERROR(SUMIFS('Input Sheet'!$C$2:$C$10,'Input Sheet'!$A$2:$A$11,$D$2)"
        "+COUNTIFS(B2:B10,1,C2:C8,2),0)"
    ) == (("SUMIFS", 1), ("COUNTIFS", 1))


def test_conditional_aggregate_range_shape_mismatches_keep_ambiguous_forms_quiet() -> None:
    quiet_formulas = (
        "=SUMIFS(C2:C10,A2:A10,A14,B2:B10,B14)",
        "=COUNTIFS($A:$A,\">0\",$B:$B,\"open\")",
        "=AVERAGEIFS(C2:C10,A2:A10,A14)",
        "=MAXIFS(C2:C10,A2:A10,A14)",
        "=MINIFS(C2:C10,A2:A10,A14)",
        "=_xlfn.MAXIFS(A2:A5,B3:B6,\"a\")",
        "=SUMIF(A2:A10,A14,C2:C12)",
        "=SUMIFS(Table1[Total],Table1[State],A1)",
        "=SUMIFS(C2:C10,NamedCriteriaRange,A1)",
        "=SUMIFS(C2:C10,OFFSET(A2,0,0,9,1),A1)",
        "=SUMIFS(C2:C10,A2#,A1)",
        "=SUMIFS(C2:C10,@A2:A10,A1)",
        "=SUMIFS(C2:C10,[Inputs.xlsx]Data!A2:A12,A1)",
        "=SUMIFS(C2:C10,2:12,A1)",
        "=SUMIFS(XFE2:XFE10,A2:A12,A1)",
        "=SUMIFS(C10:C2,A2:A12,A1)",
        "=SUMIFS(C1048577:C1048578,A2:A3,A1)",
        "=SUMIFS(C2:C10,A2:A10)",
        "=AVERAGEIFS(C2:C10,A2:A10)",
        "=MAXIFS(C2:C10,A2:A10)",
        "=MINIFS(C2:C10,A2:A10)",
        "=COUNTIFS(A2:A10,\">0\",)",
        "=Vendor.SUMIFS(C2:C10,A2:A12,A1)",
        "=Vendor.MAXIFS(C2:C10,A2:A12,A1)",
        "=Vendor._xlfn.MINIFS(C2:C10,A2:A12,A1)",
        "=_xlfn.AVERAGEIFS(C2:C10,A2:A12,A1)",
        "=_xlfn.SUMIFS(C2:C10,A2:A12,A1)",
        "=SUMIFS(C2:C10,A2:A10,#REF!)",
    )

    for formula in quiet_formulas:
        assert conditional_aggregate_range_shape_mismatches(formula) == ()


def test_extract_references_keeps_external_workbooks_separate() -> None:
    references = extract_references("='Input Sheet'!$B$2+[book.xlsx]Sheet1!A1+SUM(C1:C3)")

    assert len(references) == 3
    assert references[0].sheet == "Input Sheet"
    assert references[0].min_column == 2
    assert references[0].min_row == 2
    assert references[1].is_external
    assert references[2].sheet is None
    assert references[2].is_range


def test_external_workbook_a1_references_keep_private_source_spelling_for_portfolios() -> None:
    direct = parse_external_workbook_reference("[Inputs.xlsx]Data!$B$2:$B$4")
    relative = parse_external_workbook_reference("'..\\shared\\[Inputs.xlsx]Data Sheet'!A1")
    absolute = parse_external_workbook_reference("'C:\\Reports\\[Inputs.xlsx]Data'!A1")
    leading_equals = parse_external_workbook_reference(
        "='..\\shared\\[Inputs.xlsx]Data Sheet'!A1"
    )

    assert direct is not None
    assert direct.source_path == "Inputs.xlsx"
    assert direct.sheet == "Data"
    assert (direct.min_column, direct.min_row, direct.max_column, direct.max_row) == (
        2,
        2,
        2,
        4,
    )
    assert relative is not None
    assert relative.source_path == "..\\shared\\Inputs.xlsx"
    assert relative.sheet == "Data Sheet"
    assert absolute is not None
    assert absolute.source_path == "C:\\Reports\\Inputs.xlsx"
    assert leading_equals is not None
    assert leading_equals.source_path == "..\\shared\\Inputs.xlsx"
    assert parse_external_workbook_reference("[Inputs.xlsx]ExternalName") is None
    assert parse_external_workbook_reference("[Inputs.xlsx]Jan:Mar!A1") is None
    assert parse_external_workbook_reference("=+[Inputs.xlsx]Data!A1") is None
    assert parse_external_workbook_reference("=SUM([Inputs.xlsx]Data!A1)") is None
    assert (
        parse_external_workbook_reference("SUM(../inputs/[Inputs.xlsx]Data!A1")
        is None
    )
    assert (
        parse_external_workbook_reference("'..\\O'Brien\\[Inputs.xlsx]Data'!A1")
        is None
    )
    whole_column = parse_external_workbook_reference("[Inputs.xlsx]Data!A:A")
    whole_row = parse_external_workbook_reference("[Inputs.xlsx]Data!2:2")
    assert whole_column is not None
    assert (
        whole_column.min_column,
        whole_column.min_row,
        whole_column.max_column,
        whole_column.max_row,
    ) == (1, 1, 1, 1_048_576)
    assert whole_row is not None
    assert (
        whole_row.min_column,
        whole_row.min_row,
        whole_row.max_column,
        whole_row.max_row,
    ) == (1, 2, 16_384, 2)

    inspection = inspect_formula(
        "=SUM([Inputs.xlsx]Data!A1:A3)+'..\\shared\\[Other.xlsx]Data'!B2"
    )
    assert [
        (reference.source_path, reference.sheet)
        for reference in inspection.external_workbook_references
    ] == [
        ("Inputs.xlsx", "Data"),
        ("..\\shared\\Other.xlsx", "Data"),
    ]


def test_external_workbook_three_d_references_keep_private_endpoints_for_portfolios() -> None:
    direct = parse_external_workbook_three_d_reference(
        "[Inputs.xlsx]Jan:Mar!$B$2:$B$4"
    )
    relative = parse_external_workbook_three_d_reference(
        "'..\\shared\\[Inputs.xlsx]Jan 2026:Mar 2026'!A:A"
    )
    leading_equals = parse_external_workbook_three_d_reference(
        "='..\\shared\\[Inputs.xlsx]Jan 2026:Mar 2026'!A1"
    )

    assert direct is not None
    assert (
        direct.source_path,
        direct.first_sheet,
        direct.last_sheet,
        direct.min_column,
        direct.min_row,
        direct.max_column,
        direct.max_row,
    ) == ("Inputs.xlsx", "Jan", "Mar", 2, 2, 2, 4)
    assert relative is not None
    assert (
        relative.source_path,
        relative.first_sheet,
        relative.last_sheet,
        relative.min_column,
        relative.min_row,
        relative.max_column,
        relative.max_row,
    ) == (
        "..\\shared\\Inputs.xlsx",
        "Jan 2026",
        "Mar 2026",
        1,
        1,
        1,
        1_048_576,
    )
    assert leading_equals is not None
    assert leading_equals.source_path == "..\\shared\\Inputs.xlsx"
    assert parse_external_workbook_three_d_reference("[Inputs.xlsx]Data!A1") is None
    assert parse_external_workbook_three_d_reference("[1]Jan:Mar!A1") is None
    assert parse_external_workbook_three_d_reference("[Inputs.xlsx]Jan:Mar:Baz!A1") is None
    assert parse_external_workbook_three_d_reference("[Inputs.xlsx]Jan::Mar!A1") is None
    assert parse_external_workbook_three_d_reference("[Inputs.xlsx]Jan:Mar!InputRange") is None
    assert (
        parse_external_workbook_three_d_reference("=SUM([Inputs.xlsx]Jan:Mar!A1)")
        is None
    )

    inspection = inspect_formula(
        "=SUM([Inputs.xlsx]Jan:Mar!$B$2:$B$4)"
        "+SUM('..\\shared\\[Other.xlsx]Jan 2026:Mar 2026'!A1)"
    )

    assert inspection.unresolved_range_tokens == ()
    assert len(inspection.references) == 2
    assert all(reference.is_external for reference in inspection.references)
    assert [
        (
            reference.source_path,
            reference.first_sheet,
            reference.last_sheet,
        )
        for reference in inspection.external_workbook_three_d_references
    ] == [
        ("Inputs.xlsx", "Jan", "Mar"),
        ("..\\shared\\Other.xlsx", "Jan 2026", "Mar 2026"),
    ]


def test_external_workbook_structured_references_keep_private_table_lookup_data() -> None:
    direct = parse_external_workbook_structured_reference(
        "'..\\shared\\Inputs.xlsx'!Sales[#Data]"
    )
    bracketed = parse_external_workbook_structured_reference(
        "'..\\shared\\[Inputs.xlsx]'!Sales[[#Headers],[Amount]:[Rate]]"
    )
    leading_equals = parse_external_workbook_structured_reference(
        "='..\\shared\\Inputs.xlsx'!Sales[Amount]"
    )

    assert direct is not None
    assert (direct.source_path, direct.table_name, direct.table_reference) == (
        "..\\shared\\Inputs.xlsx",
        "Sales",
        "Sales[#Data]",
    )
    assert repr(direct) == "ExternalWorkbookStructuredReference()"
    assert bracketed is not None
    assert (
        bracketed.source_path,
        bracketed.table_name,
        bracketed.table_reference,
    ) == (
        "..\\shared\\Inputs.xlsx",
        "Sales",
        "Sales[[#Headers],[Amount]:[Rate]]",
    )
    assert leading_equals is not None
    assert leading_equals.source_path == "..\\shared\\Inputs.xlsx"
    assert parse_external_workbook_structured_reference(
        "'[Inputs.xlsx]Data'!Sales[Amount]"
    ) is None
    assert parse_external_workbook_structured_reference(
        "[Inputs.xlsx]Data!Sales[Amount]"
    ) is None
    assert parse_external_workbook_structured_reference(
        "[Inputs.xlsx]!Sales[@Amount]"
    ) is None
    assert parse_external_workbook_structured_reference(
        "'..\\shared\\Inputs.xlsx'!Sales"
    ) is None
    assert parse_external_workbook_structured_reference("[1]!Sales[Amount]") is None
    assert parse_external_workbook_structured_reference(
        "=SUM('[Inputs.xlsx]'!Sales[Amount])"
    ) is None
    assert parse_external_workbook_structured_reference(
        "'..\\O'Brien\\Inputs.xlsx'!Sales[Amount]"
    ) is None

    inspection = inspect_formula(
        "=SUM('..\\shared\\Inputs.xlsx'!Sales[Amount])"
        "+SUM('..\\shared\\[Other.xlsx]'!Margin[#All])"
    )
    assert inspection.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in inspection.references)
    assert [
        (reference.source_path, reference.table_name, reference.table_reference)
        for reference in inspection.external_workbook_structured_references
    ] == [
        ("..\\shared\\Inputs.xlsx", "Sales", "Sales[Amount]"),
        ("..\\shared\\Other.xlsx", "Margin", "Margin[#All]"),
    ]

    row_relative = inspect_formula(
        "=SUM('..\\shared\\Inputs.xlsx'!Sales[@Amount])"
    )
    assert len(row_relative.references) == 1
    assert row_relative.references[0].is_external
    assert row_relative.external_workbook_structured_references == ()
    assert row_relative.unresolved_range_tokens == ()

    bare_table = inspect_formula("=SUM('..\\shared\\Inputs.xlsx'!Sales)")
    assert len(bare_table.references) == 1
    assert bare_table.references[0].is_external
    assert bare_table.external_workbook_structured_references == ()
    assert bare_table.unresolved_range_tokens == ()


def test_external_workbook_defined_name_references_keep_private_lookup_data() -> None:
    direct = parse_external_workbook_defined_name_reference("[Inputs.xlsx]InputRange")
    relative = parse_external_workbook_defined_name_reference(
        "'..\\shared\\[Inputs.xlsx]InputRange'"
    )
    absolute = parse_external_workbook_defined_name_reference(
        "'C:\\Reports\\[Inputs.xlsx]InputRange'"
    )
    leading_equals = parse_external_workbook_defined_name_reference(
        "='..\\shared\\[Inputs.xlsx]InputRange'"
    )

    assert direct is not None
    assert direct.source_path == "Inputs.xlsx"
    assert direct.name_key == "inputrange"
    assert relative is not None
    assert relative.source_path == "..\\shared\\Inputs.xlsx"
    assert relative.name_key == "inputrange"
    assert absolute is not None
    assert absolute.source_path == "C:\\Reports\\Inputs.xlsx"
    assert absolute.name_key == "inputrange"
    assert leading_equals is not None
    assert leading_equals.source_path == "..\\shared\\Inputs.xlsx"
    assert leading_equals.name_key == "inputrange"
    assert parse_external_workbook_defined_name_reference(
        "[Inputs.xlsx]Data!InputRange"
    ) is None
    assert parse_external_workbook_defined_name_reference(
        "[Inputs.xlsx]InputRange[Column]"
    ) is None
    assert parse_external_workbook_defined_name_reference("[Inputs.xlsx]Jan:Mar!A1") is None
    assert parse_external_workbook_defined_name_reference("=[Inputs.xlsx]A1") is None
    assert parse_external_workbook_defined_name_reference("=+[Inputs.xlsx]InputRange") is None
    assert (
        parse_external_workbook_defined_name_reference("=SUM([Inputs.xlsx]InputRange)")
        is None
    )
    assert (
        parse_external_workbook_defined_name_reference(
            "SUM(../inputs/[Inputs.xlsx]InputRange"
        )
        is None
    )
    assert (
        parse_external_workbook_defined_name_reference(
            "'..\\O'Brien\\[Inputs.xlsx]InputRange'"
        )
        is None
    )

    inspection = inspect_formula(
        "=SUM([Inputs.xlsx]InputRange)+'..\\shared\\[Other.xlsx]Margin'"
    )
    assert inspection.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in inspection.references)
    assert [
        (reference.source_path, reference.name_key)
        for reference in inspection.external_workbook_defined_name_references
    ] == [
        ("Inputs.xlsx", "inputrange"),
        ("..\\shared\\Other.xlsx", "margin"),
    ]


def test_external_workbook_sheet_defined_names_keep_scope_private_and_static() -> None:
    direct = parse_external_workbook_sheet_defined_name_reference(
        "[Inputs.xlsx]Data!LocalInput"
    )
    relative = parse_external_workbook_sheet_defined_name_reference(
        "'..\\shared\\[Inputs.xlsx]Input Sheet'!Private.Input"
    )
    escaped_quote = parse_external_workbook_sheet_defined_name_reference(
        "'[Inputs.xlsx]O''Brien'!LocalInput"
    )

    assert direct is not None
    assert (direct.source_path, direct.scope_sheet, direct.name_key) == (
        "Inputs.xlsx",
        "Data",
        "localinput",
    )
    assert relative is not None
    assert (relative.source_path, relative.scope_sheet, relative.name_key) == (
        "..\\shared\\Inputs.xlsx",
        "Input Sheet",
        "private.input",
    )
    assert escaped_quote is not None
    assert (escaped_quote.scope_sheet, escaped_quote.name_key) == ("O'Brien", "localinput")
    assert parse_external_workbook_sheet_defined_name_reference(
        "[Inputs.xlsx]Data!A1"
    ) is None
    assert parse_external_workbook_sheet_defined_name_reference(
        "[Inputs.xlsx]Data!A1:B2"
    ) is None
    assert parse_external_workbook_sheet_defined_name_reference(
        "[Inputs.xlsx]Jan:Mar!LocalInput"
    ) is None
    assert parse_external_workbook_sheet_defined_name_reference(
        "[1]Data!LocalInput"
    ) is None
    assert parse_external_workbook_sheet_defined_name_reference(
        "'[Inputs.xlsx]Input Sheet!LocalInput"
    ) is None

    inspection = inspect_formula(
        "=SUM([Inputs.xlsx]Data!LocalInput)+'..\\shared\\[Other.xlsx]Input Sheet'!Margin"
    )
    assert inspection.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in inspection.references)
    assert [
        (reference.source_path, reference.scope_sheet, reference.name_key)
        for reference in inspection.external_workbook_defined_name_references
    ] == [
        ("Inputs.xlsx", "Data", "localinput"),
        ("..\\shared\\Other.xlsx", "Input Sheet", "margin"),
    ]


def test_workbook_defined_name_aliases_require_one_static_unqualified_name() -> None:
    assert parse_workbook_defined_name_alias("InputRange") == "inputrange"
    assert parse_workbook_defined_name_alias("= Private.Input ") == "private.input"
    assert parse_workbook_defined_name_alias("=ÜnicodeName") == "ünicodename"

    for value in (
        "",
        "A1",
        "Sheet1!InputRange",
        "='Sheet 1'!InputRange",
        "[Inputs.xlsx]InputRange",
        "=+InputRange",
        "=SUM(InputRange)",
        "=InputRange+1",
        "=InputRange[Column]",
        "=Input Range",
        "=\\InputRange",
        "==InputRange",
    ):
        assert parse_workbook_defined_name_alias(value) is None


def test_package_indexed_external_name_references_require_a_declared_one_based_index() -> None:
    assert parse_external_link_indexed_defined_name_reference("[1]!InputRange") == (
        1,
        "inputrange",
    )
    assert parse_external_link_indexed_defined_name_reference(
        "=[42]!Private.Input"
    ) == (42, "private.input")
    assert parse_external_link_indexed_defined_name_reference(
        "[2147483647]!InputRange"
    ) == (2_147_483_647, "inputrange")
    assert parse_external_link_indexed_defined_name_reference("[0]!InputRange") is None
    assert parse_external_link_indexed_defined_name_reference("[01]!InputRange") is None
    assert parse_external_link_indexed_defined_name_reference(
        "[2147483648]!InputRange"
    ) is None
    assert parse_external_link_indexed_defined_name_reference("[1]Data!A1") is None
    assert parse_external_link_indexed_defined_name_reference("[1]!A1") is None
    assert parse_external_link_indexed_defined_name_reference("[1]!A1:B2") is None
    assert parse_external_link_indexed_defined_name_reference("[1]!Input[Column]") is None
    assert parse_external_link_indexed_defined_name_reference("[1]!1InputRange") is None
    assert parse_external_link_indexed_defined_name_reference("[1]!\\InputRange") is None
    assert parse_external_link_indexed_defined_name_reference("[1]!$InputRange") is None
    assert parse_external_link_indexed_defined_name_reference("[1]!Input?Range") is None
    assert parse_external_link_indexed_defined_name_reference("[1]!Input+Range") is None

    inspection = inspect_formula(
        "=SUM([1]!InputRange)",
        indexed_external_workbook_paths={1: "../inputs/source.xlsx"},
    )
    assert inspection.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in inspection.references)
    assert [
        (reference.source_path, reference.name_key)
        for reference in inspection.external_workbook_defined_name_references
    ] == [("../inputs/source.xlsx", "inputrange")]

    unresolved = inspect_formula("=SUM([1]!InputRange)")
    assert unresolved.external_workbook_defined_name_references == ()
    assert unresolved.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in unresolved.references)


def test_package_indexed_external_sheet_defined_names_require_a_declared_index() -> None:
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "[1]Data!LocalInput"
    ) == (1, "Data", "localinput")
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "='[42]Input Sheet'!Private.Input"
    ) == (42, "Input Sheet", "private.input")
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "'[9]O''Brien'!LocalInput"
    ) == (9, "O'Brien", "localinput")
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "[0]Data!LocalInput"
    ) is None
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "[01]Data!LocalInput"
    ) is None
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "[1]!LocalInput"
    ) is None
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "[1]Data!A1"
    ) is None
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "[1]Data!A1:B2"
    ) is None
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "[1]Jan:Mar!LocalInput"
    ) is None
    assert parse_external_link_indexed_sheet_defined_name_reference(
        "'[1]Input Sheet!LocalInput"
    ) is None

    inspection = inspect_formula(
        "=SUM([1]Data!LocalInput)+'[2]Input Sheet'!Margin",
        indexed_external_workbook_paths={
            1: "../inputs/source.xlsx",
            2: "../inputs/other.xlsx",
        },
    )
    assert inspection.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in inspection.references)
    assert [
        (reference.source_path, reference.scope_sheet, reference.name_key)
        for reference in inspection.external_workbook_defined_name_references
    ] == [
        ("../inputs/source.xlsx", "Data", "localinput"),
        ("../inputs/other.xlsx", "Input Sheet", "margin"),
    ]

    unresolved = inspect_formula("=SUM([1]Data!LocalInput)")
    assert unresolved.external_workbook_defined_name_references == ()
    assert unresolved.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in unresolved.references)


def test_package_indexed_external_a1_references_require_a_declared_one_based_index() -> None:
    direct = parse_external_link_indexed_workbook_reference("[1]Data!$B$2:$B$4")
    quoted = parse_external_link_indexed_workbook_reference(
        "'[42]Input Sheet'!A:A"
    )
    escaped_quote = parse_external_link_indexed_workbook_reference(
        "'[9]O''Brien'!A1"
    )

    assert direct is not None
    assert (
        direct.index,
        direct.sheet,
        direct.min_column,
        direct.min_row,
        direct.max_column,
        direct.max_row,
    ) == (1, "Data", 2, 2, 2, 4)
    assert quoted is not None
    assert (
        quoted.index,
        quoted.sheet,
        quoted.min_column,
        quoted.min_row,
        quoted.max_column,
        quoted.max_row,
    ) == (42, "Input Sheet", 1, 1, 1, 1_048_576)
    assert escaped_quote is not None
    assert (escaped_quote.index, escaped_quote.sheet) == (9, "O'Brien")
    assert parse_external_workbook_reference("[1]Data!A1") is None
    assert parse_external_link_indexed_workbook_reference("[0]Data!A1") is None
    assert parse_external_link_indexed_workbook_reference("[01]Data!A1") is None
    assert (
        parse_external_link_indexed_workbook_reference("[2147483648]Data!A1")
        is None
    )
    assert parse_external_link_indexed_workbook_reference("[1]!A1") is None
    assert parse_external_link_indexed_workbook_reference("[1]Data!InputRange") is None
    assert parse_external_link_indexed_workbook_reference("[1]Jan:Mar!A1") is None
    assert parse_external_link_indexed_workbook_reference("'[1]Input Sheet!A1") is None
    assert parse_external_link_indexed_workbook_reference("[1]Data[Other]!A1") is None

    inspection = inspect_formula(
        "=SUM([1]Data!$B$2:$B$4)+'[2]Input Sheet'!A:A",
        indexed_external_workbook_paths={
            1: "../inputs/source.xlsx",
            2: "../inputs/other.xlsx",
        },
    )
    assert inspection.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in inspection.references)
    assert [
        (
            reference.source_path,
            reference.sheet,
            reference.min_column,
            reference.min_row,
            reference.max_column,
            reference.max_row,
        )
        for reference in inspection.external_workbook_references
    ] == [
        ("../inputs/source.xlsx", "Data", 2, 2, 2, 4),
        ("../inputs/other.xlsx", "Input Sheet", 1, 1, 1, 1_048_576),
    ]

    unresolved = inspect_formula("=SUM([1]Data!A1)")
    assert unresolved.unresolved_range_tokens == ()
    assert unresolved.external_workbook_references == ()
    assert all(reference.is_external for reference in unresolved.references)

    alias = inspect_formula(
        "=PackageExternalCell",
        named_external_workbook_references={
            "packageexternalcell": (
                ExternalWorkbookReference(
                    source_path="../inputs/source.xlsx",
                    sheet="Data",
                    min_column=2,
                    min_row=2,
                    max_column=2,
                    max_row=4,
                ),
            )
        },
    )
    assert alias.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in alias.references)
    assert [
        (reference.source_path, reference.sheet)
        for reference in alias.external_workbook_references
    ] == [("../inputs/source.xlsx", "Data")]


def test_package_indexed_external_three_d_references_require_a_declared_index() -> None:
    direct = parse_external_link_indexed_workbook_three_d_reference(
        "[1]Jan:Mar!$B$2:$B$4"
    )
    quoted = parse_external_link_indexed_workbook_three_d_reference(
        "'[42]Jan 2026:Mar 2026'!A:A"
    )
    leading_equals = parse_external_link_indexed_workbook_three_d_reference(
        "=[7]Jan:Mar!A1"
    )
    escaped_quote = parse_external_link_indexed_workbook_three_d_reference(
        "'[9]O''Brien:Year End'!A1"
    )

    assert direct is not None
    assert (
        direct.index,
        direct.first_sheet,
        direct.last_sheet,
        direct.min_column,
        direct.min_row,
        direct.max_column,
        direct.max_row,
    ) == (1, "Jan", "Mar", 2, 2, 2, 4)
    assert quoted is not None
    assert (
        quoted.index,
        quoted.first_sheet,
        quoted.last_sheet,
        quoted.min_column,
        quoted.min_row,
        quoted.max_column,
        quoted.max_row,
    ) == (42, "Jan 2026", "Mar 2026", 1, 1, 1, 1_048_576)
    assert escaped_quote is not None
    assert (escaped_quote.first_sheet, escaped_quote.last_sheet) == (
        "O'Brien",
        "Year End",
    )
    assert leading_equals is not None
    assert leading_equals.index == 7
    assert parse_external_link_indexed_workbook_three_d_reference("[1]Data!A1") is None
    assert parse_external_link_indexed_workbook_three_d_reference("[0]Jan:Mar!A1") is None
    assert parse_external_link_indexed_workbook_three_d_reference("[01]Jan:Mar!A1") is None
    assert (
        parse_external_link_indexed_workbook_three_d_reference("[1]Jan:Mar:Baz!A1")
        is None
    )
    assert parse_external_link_indexed_workbook_three_d_reference("[1]Jan::Mar!A1") is None
    assert parse_external_link_indexed_workbook_three_d_reference("[1]Jan:Mar!Input") is None

    inspection = inspect_formula(
        "=SUM([1]Jan:Mar!$B$2:$B$4)+SUM('[2]Jan 2026:Mar 2026'!A:A)",
        indexed_external_workbook_paths={
            1: "../inputs/source.xlsx",
            2: "../inputs/other.xlsx",
        },
    )
    assert inspection.unresolved_range_tokens == ()
    assert len(inspection.references) == 2
    assert all(reference.is_external for reference in inspection.references)
    assert [
        (
            reference.source_path,
            reference.first_sheet,
            reference.last_sheet,
            reference.min_column,
            reference.min_row,
            reference.max_column,
            reference.max_row,
        )
        for reference in inspection.external_workbook_three_d_references
    ] == [
        ("../inputs/source.xlsx", "Jan", "Mar", 2, 2, 2, 4),
        ("../inputs/other.xlsx", "Jan 2026", "Mar 2026", 1, 1, 1, 1_048_576),
    ]

    alias = inspect_formula(
        "=PackageExternalThreeD",
        named_external_workbook_three_d_references={
            "packageexternalthreed": (
                ExternalWorkbookThreeDReference(
                    source_path="../inputs/source.xlsx",
                    first_sheet="Jan",
                    last_sheet="Mar",
                    min_column=2,
                    min_row=2,
                    max_column=2,
                    max_row=4,
                ),
            )
        },
    )
    assert alias.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in alias.references)
    assert [
        (reference.source_path, reference.first_sheet, reference.last_sheet)
        for reference in alias.external_workbook_three_d_references
    ] == [("../inputs/source.xlsx", "Jan", "Mar")]


def test_package_indexed_external_structured_references_require_a_declared_index() -> None:
    direct = parse_external_link_indexed_workbook_structured_reference(
        "[1]!Sales[Amount]"
    )
    selector = parse_external_link_indexed_workbook_structured_reference(
        "[42]!Sales[[#Data],[Amount]:[Rate]]"
    )
    leading_equals = parse_external_link_indexed_workbook_structured_reference(
        "=[7]!Sales[#All]"
    )

    assert direct is not None
    assert (direct.index, direct.table_name, direct.table_reference) == (
        1,
        "Sales",
        "Sales[Amount]",
    )
    assert selector is not None
    assert (selector.index, selector.table_reference) == (
        42,
        "Sales[[#Data],[Amount]:[Rate]]",
    )
    assert leading_equals is not None
    assert leading_equals.index == 7
    assert parse_external_link_indexed_workbook_structured_reference(
        "[0]!Sales[Amount]"
    ) is None
    assert parse_external_link_indexed_workbook_structured_reference(
        "[01]!Sales[Amount]"
    ) is None
    assert parse_external_link_indexed_workbook_structured_reference(
        "'[1]'!Sales[Amount]"
    ) is None
    assert parse_external_link_indexed_workbook_structured_reference(
        "[1]Data!Sales[Amount]"
    ) is None
    assert parse_external_link_indexed_workbook_structured_reference(
        "[1]!Sales[@Amount]"
    ) is None
    assert parse_external_link_indexed_workbook_structured_reference(
        "[1]!Sales[#This Row]"
    ) is None
    assert parse_external_link_indexed_workbook_structured_reference("[1]!A1") is None

    inspection = inspect_formula(
        "=SUM([1]!Sales[Amount])+SUM([2]!Margin[[#Data],[Net]:[Tax]])",
        indexed_external_workbook_paths={
            1: "../inputs/source.xlsx",
            2: "../inputs/other.xlsx",
        },
    )
    assert inspection.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in inspection.references)
    assert [
        (reference.source_path, reference.table_name, reference.table_reference)
        for reference in inspection.external_workbook_structured_references
    ] == [
        ("../inputs/source.xlsx", "Sales", "Sales[Amount]"),
        ("../inputs/other.xlsx", "Margin", "Margin[[#Data],[Net]:[Tax]]"),
    ]

    unresolved = inspect_formula("=SUM([1]!Sales[Amount])")
    assert unresolved.unresolved_range_tokens == ()
    assert unresolved.external_workbook_structured_references == ()
    assert all(reference.is_external for reference in unresolved.references)

    alias = inspect_formula(
        "=PackageExternalTable",
        named_external_workbook_structured_references={
            "packageexternaltable": (
                ExternalWorkbookStructuredReference(
                    source_path="../inputs/source.xlsx",
                    table_name="Sales",
                    table_reference="Sales[Amount]",
                ),
            )
        },
    )
    assert alias.unresolved_range_tokens == ()
    assert all(reference.is_external for reference in alias.references)
    assert [
        (reference.source_path, reference.table_name, reference.table_reference)
        for reference in alias.external_workbook_structured_references
    ] == [("../inputs/source.xlsx", "Sales", "Sales[Amount]")]


def test_formula_inspection_resolves_names_and_marks_static_coverage_gaps() -> None:
    inspection = inspect_formula(
        '=HeadlineOutput+UnknownMetric+INDIRECT("Inputs!B2")',
        {
            "headlineoutput": (
                ParsedReference("Dashboard", 2, 12, 2, 12, raw="HeadlineOutput"),
            )
        },
    )

    assert inspection.references == (
        ParsedReference("Dashboard", 2, 12, 2, 12, raw="HeadlineOutput"),
    )
    assert inspection.unresolved_range_tokens == ("UnknownMetric",)
    assert inspection.dynamic_reference_functions == ("INDIRECT",)


def test_formula_inspection_inventories_external_action_and_provider_functions() -> None:
    inspection = inspect_formula(
        '=_xlfn.IMAGE("https://private.example.test/image.png")'
        '&HYPERLINK("#Inputs!A1","Internal")'
        '&WEBSERVICE("https://private.example.test/service")'
        '&RTD("Private.Provider","PrivateServer","Topic")'
        '&HYPERLINK("https://private.example.test/second","External")'
        '+_xlfn.STOCKHISTORY("XNAS:MSFT",DATE(2024,1,1))'
        '+CUBEVALUE("Finance","[Measures].[Revenue]")'
        '+CUBEMEMBER("Finance","[Date].[All]")'
        '+CUBEMEMBERPROPERTY("Finance","[Date].[All]","MEMBER_CAPTION")'
        '+CUBERANKEDMEMBER("Finance",A1,1)'
        '+CUBESET("Finance","[Date].[All].Children")'
        '+CUBESETCOUNT(A1)'
        '+CUBEKPIMEMBER("Finance","KPI",1)'
    )
    shadowed = inspect_formula(
        '=STOCKHISTORY(A1,DATE(2024,1,1))',
        named_function_references={"stockhistory": ()},
    )

    assert inspection.external_action_functions == (
        "IMAGE",
        "HYPERLINK",
        "WEBSERVICE",
        "RTD",
        "HYPERLINK",
        "STOCKHISTORY",
        "CUBEVALUE",
        "CUBEMEMBER",
        "CUBEMEMBERPROPERTY",
        "CUBERANKEDMEMBER",
        "CUBESET",
        "CUBESETCOUNT",
        "CUBEKPIMEMBER",
    )
    assert shadowed.external_action_functions == ()


def test_formula_inspection_propagates_external_actions_from_named_definitions() -> None:
    named_lambda = inspect_formula(
        "=FENCE.WRAPPER(A1)",
        named_function_references={"fence.wrapper": ()},
        named_function_formula_external_action_functions={
            "fence.wrapper": ("HYPERLINK",),
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_formula_external_action_functions={
            "fence.direct": ("WEBSERVICE",),
        },
    )
    named_provider = inspect_formula(
        "=FENCE.MARKET(A1)",
        named_function_references={"fence.market": ()},
        named_function_formula_external_action_functions={
            "fence.market": ("STOCKHISTORY", "CUBEVALUE"),
        },
    )

    assert named_lambda.external_action_functions == ("HYPERLINK",)
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
    )
    assert named_formula.external_action_functions == ("WEBSERVICE",)
    assert named_formula.unresolved_range_tokens == ()
    assert named_provider.external_action_functions == ("STOCKHISTORY", "CUBEVALUE")


def test_formula_inspection_inventories_python_in_excel_function_spellings() -> None:
    inspection = inspect_formula(
        "=_xlfn._xlws.PY(0,0,A1)+_xlws.PY(1,1,A2)+PY(2,0,A3)",
        origin=("Inputs", "B2"),
    )

    assert inspection.python_functions == ("PY", "PY", "PY")
    assert inspection.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
        ParsedReference(None, 1, 2, 1, 2, raw="A2"),
        ParsedReference(None, 1, 3, 1, 3, raw="A3"),
    )


def test_formula_inspection_inventories_namespaced_custom_function_candidates() -> None:
    inspection = inspect_formula(
        "=CONTOSO.ADD(A1)+CONTOSO.GETPRICE(A2)+MYFUNCTION.SPHEREVOLUME(A3)"
        "+ECMA.CEILING(A4,1)+ERF.PRECISE(A5)+WORKDAY.INTL(A6,1)"
        "+_xlfn._xlws.PY(0,0,A7)"
    )
    named_lambda = inspect_formula(
        "=LOCAL.RATE(A1)",
        named_function_references={"local.rate": ()},
    )
    named_formula = inspect_formula(
        "=MODEL.RATE(A1)",
        named_references={"model.rate": ()},
    )

    assert inspection.office_custom_function_candidates == (
        "CONTOSO.ADD",
        "CONTOSO.GETPRICE",
        "MYFUNCTION.SPHEREVOLUME",
    )
    assert named_lambda.office_custom_function_candidates == ()
    assert named_formula.office_custom_function_candidates == ()


def test_pinned_native_function_catalog_covers_openpyxl_bare_catalog() -> None:
    """Keep FormulaFence's stable allowlist ahead of its parser dependency.

    FormulaFence deliberately does not import this mutable dependency catalogue
    at runtime. This test instead makes an openpyxl catalogue update an explicit
    review event so a newly recognized bare native function cannot silently
    become a generic runtime-function candidate.
    """
    parser_native_functions = {
        name.upper() for name in FORMULAE if "." not in name
    }

    assert parser_native_functions <= _EXCEL_UNQUALIFIED_NATIVE_FUNCTIONS


def test_formula_inspection_inventories_unqualified_runtime_function_candidates() -> None:
    inspection = inspect_formula(
        "=LOCALUDF(A1)+ANOTHER_UDF(A2)+@MYUDF(A3)+SUM(A4)+XLOOKUP(A5,A5,A5)"
        "+VSTACK(A6,A7)+FIELDVALUE(A8,\"name\")+PY(\"1+1\",0)"
        "+CONTOSO.ADD(A9)+_xlfn._xlws.PY(0,0,A10)"
    )
    named_lambda = inspect_formula(
        "=WORKBOOKUDF(A1)",
        named_function_references={"workbookudf": ()},
    )
    named_formula = inspect_formula(
        "=FORMULANAME(A1)",
        named_references={"formulaname": ()},
    )
    local_lambda = inspect_formula(
        "=LET(LOCALUDF,LAMBDA(value,value),LOCALUDF(A1))"
    )

    assert inspection.unqualified_runtime_function_candidates == (
        "LOCALUDF",
        "ANOTHER_UDF",
        "MYUDF",
    )
    assert named_lambda.unqualified_runtime_function_candidates == ()
    assert named_formula.unqualified_runtime_function_candidates == ()
    assert local_lambda.unqualified_runtime_function_candidates == ()


def test_formula_inspection_propagates_unqualified_runtime_function_candidates() -> None:
    named_lambda = inspect_formula(
        "=FENCE.WRAPPER(A1)",
        named_function_references={"fence.wrapper": ()},
        named_function_unqualified_runtime_function_candidates={
            "fence.wrapper": ("PRIVATEUDF",),
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_unqualified_runtime_function_candidates={
            "fence.direct": ("PRIVATEUDF",),
        },
    )

    assert named_lambda.unqualified_runtime_function_candidates == ("PRIVATEUDF",)
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
    )
    assert named_formula.unqualified_runtime_function_candidates == ("PRIVATEUDF",)
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_propagates_candidates_from_named_definitions() -> None:
    named_lambda = inspect_formula(
        "=FENCE.WRAPPER(A1)",
        named_function_references={"fence.wrapper": ()},
        named_function_custom_function_candidates={
            "fence.wrapper": ("CONTOSO.GETDATA",),
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_custom_function_candidates={
            "fence.direct": ("CONTOSO.GETDATA",),
        },
    )

    assert named_lambda.office_custom_function_candidates == ("CONTOSO.GETDATA",)
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
    )
    assert named_formula.office_custom_function_candidates == ("CONTOSO.GETDATA",)
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_inventories_worksheet_code_resource_registrations() -> None:
    inspection = inspect_formula(
        '=REGISTER.ID(A1,A2,"J!")+@REGISTER.ID(A3,A4,"J!")+ECMA.CEILING(A5,1)'
    )
    shadowed = inspect_formula(
        "=REGISTER.ID(A1,A2)",
        named_function_references={"register.id": ()},
    )

    assert inspection.worksheet_code_resource_registration_functions == (
        "REGISTER.ID",
        "REGISTER.ID",
    )
    assert inspection.office_custom_function_candidates == ()
    assert shadowed.worksheet_code_resource_registration_functions == ()


def test_formula_inspection_propagates_worksheet_code_resource_registrations() -> None:
    named_lambda = inspect_formula(
        "=FENCE.REGISTER(A1,A2)",
        named_function_references={"fence.register": ()},
        named_function_worksheet_code_resource_registration_functions={
            "fence.register": ("REGISTER.ID",)
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_worksheet_code_resource_registration_functions={
            "fence.direct": ("REGISTER.ID",)
        },
    )

    assert named_lambda.worksheet_code_resource_registration_functions == (
        "REGISTER.ID",
    )
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
        ParsedReference(None, 1, 2, 1, 2, raw="A2"),
    )
    assert named_formula.worksheet_code_resource_registration_functions == (
        "REGISTER.ID",
    )
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_inventories_direct_dde_syntax_conservatively() -> None:
    command_style = inspect_formula("=cmd|' /C harmless'!A0")
    documented_style = inspect_formula("='Quote'|'NYSE'!ZAXX")
    unquoted_style = inspect_formula("=kepdde|_ddedata!Channel1.M340.Int_1")
    omitted_item = inspect_formula("=cmd|'/C harmless'!")
    embedded = inspect_formula("=@SUM(A1:A2)*cmd|'/C harmless'!A0")
    two_links = inspect_formula("=cmd|'topic'!A0+cmd|'topic'!A1")
    quoted_sheet = inspect_formula("='cmd|/C harmless'!A0")
    quoted_string = inspect_formula('=HYPERLINK("cmd|\'/C harmless\'!A0","Open")')
    ordinary_sheet = inspect_formula("=SUM('Sheet|One'!A1)")

    assert command_style.formula_dde_link_count == 1
    assert documented_style.formula_dde_link_count == 1
    assert unquoted_style.formula_dde_link_count == 1
    assert omitted_item.formula_dde_link_count == 1
    assert embedded.formula_dde_link_count == 1
    assert two_links.formula_dde_link_count == 2
    assert quoted_sheet.formula_dde_link_count == 0
    assert quoted_string.formula_dde_link_count == 0
    assert ordinary_sheet.formula_dde_link_count == 0
    assert command_style.tokenization_failed is True
    assert unquoted_style.tokenization_failed is False


def test_formula_inspection_propagates_direct_dde_markers_from_named_definitions() -> None:
    named_lambda = inspect_formula(
        "=FENCE.DDE(A1)",
        named_function_references={"fence.dde": None},
        named_function_formula_dde_link_markers={"fence.dde": ("DDE",)},
    )
    named_formula = inspect_formula(
        "=FENCE.DDE.DIRECT",
        named_references={"fence.dde.direct": ()},
        named_formula_dde_link_markers={"fence.dde.direct": ("DDE",)},
    )

    assert named_lambda.formula_dde_link_count == 1
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
    )
    assert named_formula.formula_dde_link_count == 1
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_inventories_formula_defined_xlm_registrations() -> None:
    ordinary = inspect_formula('=REGISTER(A1,A2,"J!")+@REGISTER(A3,A4,"J!")')
    definition = inspect_formula(
        '=REGISTER(A1,A2,"J!")+@REGISTER(A3,A4,"J!")',
        inspect_formula_defined_xlm_registrations=True,
    )
    shadowed = inspect_formula(
        '=REGISTER(A1,A2,"J!")',
        named_function_references={"register": ()},
        inspect_formula_defined_xlm_registrations=True,
    )

    assert ordinary.formula_defined_xlm_registration_functions == ()
    assert definition.formula_defined_xlm_registration_functions == (
        "REGISTER",
        "REGISTER",
    )
    assert shadowed.formula_defined_xlm_registration_functions == ()


def test_formula_inspection_propagates_formula_defined_xlm_registrations() -> None:
    named_lambda = inspect_formula(
        "=FENCE.REGISTER(A1,A2)",
        named_function_references={"fence.register": ()},
        named_function_formula_defined_xlm_registration_functions={
            "fence.register": ("REGISTER",)
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_formula_defined_xlm_registration_functions={
            "fence.direct": ("REGISTER",)
        },
    )

    assert named_lambda.formula_defined_xlm_registration_functions == ("REGISTER",)
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
        ParsedReference(None, 1, 2, 1, 2, raw="A2"),
    )
    assert named_formula.formula_defined_xlm_registration_functions == ("REGISTER",)
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_inventories_formula_defined_xlm_evaluations() -> None:
    ordinary = inspect_formula('=EVALUATE(A1)+@EVALUATE(A2)')
    definition = inspect_formula(
        '=EVALUATE(A1)+@EVALUATE(A2)',
        inspect_formula_defined_xlm_evaluations=True,
    )
    shadowed = inspect_formula(
        "=EVALUATE(A1)",
        named_function_references={"evaluate": ()},
        inspect_formula_defined_xlm_evaluations=True,
    )

    assert ordinary.formula_defined_xlm_evaluation_functions == ()
    assert definition.formula_defined_xlm_evaluation_functions == (
        "EVALUATE",
        "EVALUATE",
    )
    assert shadowed.formula_defined_xlm_evaluation_functions == ()


def test_formula_inspection_propagates_formula_defined_xlm_evaluations() -> None:
    named_lambda = inspect_formula(
        "=FENCE.EVALUATE(A1)",
        named_function_references={"fence.evaluate": ()},
        named_function_formula_defined_xlm_evaluation_functions={
            "fence.evaluate": ("EVALUATE",)
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_formula_defined_xlm_evaluation_functions={
            "fence.direct": ("EVALUATE",)
        },
    )

    assert named_lambda.formula_defined_xlm_evaluation_functions == ("EVALUATE",)
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
    )
    assert named_formula.formula_defined_xlm_evaluation_functions == ("EVALUATE",)
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_inventories_formula_defined_xlm_actions() -> None:
    formula = (
        '=CALL("library","procedure","J")+EXEC("command")+EXECUTE("dde")+'
        'ON.DATA("data","macro")+ON.DOUBLECLICK("macro")+ON.ENTRY("macro")+'
        'ON.KEY("key","macro")+ON.RECALC("macro")+ON.SHEET("macro")+'
        'ON.TIME(NOW(),"macro")+ON.WINDOW("macro")+RUN("macro")+SEND.KEYS("keys")'
    )
    ordinary = inspect_formula(formula)
    definition = inspect_formula(
        formula,
        inspect_formula_defined_xlm_actions=True,
    )
    shadowed = inspect_formula(
        '=RUN("macro")',
        named_function_references={"run": ()},
        inspect_formula_defined_xlm_actions=True,
    )

    assert ordinary.formula_defined_xlm_action_functions == ()
    assert definition.formula_defined_xlm_action_functions == (
        "CALL",
        "EXEC",
        "EXECUTE",
        "ON.DATA",
        "ON.DOUBLECLICK",
        "ON.ENTRY",
        "ON.KEY",
        "ON.RECALC",
        "ON.SHEET",
        "ON.TIME",
        "ON.WINDOW",
        "RUN",
        "SEND.KEYS",
    )
    assert shadowed.formula_defined_xlm_action_functions == ()


def test_formula_inspection_propagates_formula_defined_xlm_actions() -> None:
    named_lambda = inspect_formula(
        "=FENCE.ACTION(A1)",
        named_function_references={"fence.action": ()},
        named_function_formula_defined_xlm_action_functions={
            "fence.action": ("EXEC", "RUN")
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_formula_defined_xlm_action_functions={
            "fence.direct": ("ON.TIME",)
        },
    )

    assert named_lambda.formula_defined_xlm_action_functions == ("EXEC", "RUN")
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
    )
    assert named_formula.formula_defined_xlm_action_functions == ("ON.TIME",)
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_inventories_formula_defined_xlm_get_cell_calls() -> None:
    ordinary = inspect_formula("=GET.CELL(7,A1)+@GET.CELL(53,A2)")
    definition = inspect_formula(
        "=GET.CELL(7,A1)+@GET.CELL(53,A2)",
        inspect_formula_defined_xlm_get_cell_calls=True,
    )
    shadowed = inspect_formula(
        "=GET.CELL(7,A1)",
        named_function_references={"get.cell": ()},
        inspect_formula_defined_xlm_get_cell_calls=True,
    )

    assert ordinary.formula_defined_xlm_get_cell_functions == ()
    assert ordinary.office_custom_function_candidates == ()
    assert definition.formula_defined_xlm_get_cell_functions == (
        "GET.CELL",
        "GET.CELL",
    )
    assert shadowed.formula_defined_xlm_get_cell_functions == ()


def test_formula_inspection_propagates_formula_defined_xlm_get_cell_calls() -> None:
    named_lambda = inspect_formula(
        "=FENCE.GET.CELL(A1)",
        named_function_references={"fence.get.cell": ()},
        named_function_formula_defined_xlm_get_cell_functions={
            "fence.get.cell": ("GET.CELL",)
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_formula_defined_xlm_get_cell_functions={
            "fence.direct": ("GET.CELL",)
        },
    )

    assert named_lambda.formula_defined_xlm_get_cell_functions == ("GET.CELL",)
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
    )
    assert named_formula.formula_defined_xlm_get_cell_functions == ("GET.CELL",)
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_inventories_xlm_environment_information_calls() -> None:
    ordinary = inspect_formula(
        "=GET.WORKBOOK(4)+@GET.WORKSPACE(2)+GET.DOCUMENT(37)"
    )
    definition = inspect_formula(
        "=GET.WORKBOOK(4)+@GET.WORKSPACE(2)+GET.DOCUMENT(37)",
        inspect_formula_defined_xlm_environment_information_calls=True,
    )
    shadowed = inspect_formula(
        "=GET.WORKBOOK(4)",
        named_function_references={"get.workbook": ()},
        inspect_formula_defined_xlm_environment_information_calls=True,
    )

    assert ordinary.formula_defined_xlm_environment_information_functions == ()
    assert ordinary.office_custom_function_candidates == ()
    assert definition.formula_defined_xlm_environment_information_functions == (
        "GET.WORKBOOK",
        "GET.WORKSPACE",
        "GET.DOCUMENT",
    )
    assert shadowed.formula_defined_xlm_environment_information_functions == ()


def test_formula_inspection_propagates_xlm_environment_information_calls() -> None:
    named_lambda = inspect_formula(
        "=FENCE.ENVIRONMENT(A1)",
        named_function_references={"fence.environment": ()},
        named_function_formula_defined_xlm_environment_information_functions={
            "fence.environment": ("GET.WORKBOOK", "GET.WORKSPACE")
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_formula_defined_xlm_environment_information_functions={
            "fence.direct": ("GET.DOCUMENT",)
        },
    )

    assert named_lambda.formula_defined_xlm_environment_information_functions == (
        "GET.WORKBOOK",
        "GET.WORKSPACE",
    )
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
    )
    assert named_formula.formula_defined_xlm_environment_information_functions == (
        "GET.DOCUMENT",
    )
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_inventories_native_environment_information_calls() -> None:
    ordinary = inspect_formula(
        '=CELL("filename")+@CELL("type",A1)+INFO("directory")'
    )
    shadowed = inspect_formula(
        '=CELL("filename")',
        named_function_references={"cell": ()},
    )

    assert ordinary.formula_environment_information_functions == (
        "CELL",
        "CELL",
        "INFO",
    )
    assert ordinary.formula_environment_information_implicit_cell_reference_count == 1
    assert ordinary.formula_environment_information_function_count == 3
    assert shadowed.formula_environment_information_functions == ()


def test_formula_inspection_propagates_native_environment_information_calls() -> None:
    named_lambda = inspect_formula(
        "=FENCE.ENVIRONMENT(A1)",
        named_function_references={"fence.environment": ()},
        named_function_formula_environment_information_functions={
            "fence.environment": ("CELL", "INFO")
        },
    )
    named_formula = inspect_formula(
        "=FENCE.DIRECT",
        named_references={"fence.direct": ()},
        named_formula_environment_information_functions={
            "fence.direct": ("CELL",)
        },
    )

    assert named_lambda.formula_environment_information_functions == ("CELL", "INFO")
    assert named_lambda.references == (
        ParsedReference(None, 1, 1, 1, 1, raw="A1"),
    )
    assert named_formula.formula_environment_information_functions == ("CELL",)
    assert named_formula.unresolved_range_tokens == ()


def test_formula_inspection_inventories_workbook_tab_information_calls() -> None:
    ordinary = inspect_formula("=SHEET()+SHEETS()+SHEET(A1)+SHEETS(A1)")
    shadowed = inspect_formula(
        "=SHEET()+SHEETS()",
        named_function_references={"sheet": (), "sheets": ()},
    )
    propagated = inspect_formula(
        "=FENCE.TAB.INFORMATION(A1)",
        named_function_references={"fence.tab.information": ()},
        named_function_formula_environment_information_functions={
            "fence.tab.information": (
                "SHEET",
                "SHEETS",
                "FORMULAFENCE_SHEETS_IMPLICIT_REFERENCE_MARKER",
            )
        },
    )
    malformed = inspect_formula("=SHEETS(,)")

    assert ordinary.formula_environment_information_functions == (
        "SHEET",
        "SHEETS",
        "SHEET",
        "SHEETS",
    )
    assert ordinary.formula_environment_information_implicit_sheets_reference_count == 1
    assert shadowed.formula_environment_information_functions == ()
    assert propagated.formula_environment_information_functions == ("SHEET", "SHEETS")
    assert propagated.formula_environment_information_implicit_sheets_reference_count == 1
    assert malformed.formula_environment_information_functions == ("SHEETS",)
    assert malformed.formula_environment_information_implicit_sheets_reference_count == 0


def test_formula_inspection_recognises_a_known_named_constant() -> None:
    inspection = inspect_formula("=StaticRate", {"staticrate": ()})

    assert inspection.references == ()
    assert inspection.unresolved_range_tokens == ()


def test_formula_inspection_traces_static_spill_anchors_and_marks_the_boundary() -> None:
    literal = inspect_formula("=SUM(A1#)")
    qualified = inspect_formula("=SUM('My Sheet'!$B$2#)")
    serialized = inspect_formula("=SUM(_xlfn.ANCHORARRAY(A1))")
    string_literal = inspect_formula('=CONCAT("literal A1#",A1#)')

    assert literal.references == (ParsedReference(None, 1, 1, 1, 1, raw="A1"),)
    assert literal.spill_reference_tokens == ("A1#",)
    assert literal.tokenization_failed is False
    assert qualified.references == (
        ParsedReference("My Sheet", 2, 2, 2, 2, raw="'My Sheet'!$B$2"),
    )
    assert qualified.spill_reference_tokens == ("'My Sheet'!$B$2#",)
    assert serialized.references == (ParsedReference(None, 1, 1, 1, 1, raw="A1"),)
    assert serialized.spill_reference_tokens == ("_xlfn.ANCHORARRAY",)
    assert string_literal.references == (ParsedReference(None, 1, 1, 1, 1, raw="A1"),)
    assert string_literal.spill_reference_tokens == ("A1#",)


def test_formula_inspection_marks_unsupported_tokenization_failures() -> None:
    malformed = inspect_formula("=SUM(A1#1)")
    range_anchor = inspect_formula("=SUM(A1:B2#)")
    external_anchor = inspect_formula("=SUM([book.xlsx]Sheet1!A1#)")
    implicit_spill = inspect_formula("=SUM(@A1#)")

    assert malformed.tokenization_failed is True
    assert range_anchor.tokenization_failed is True
    assert external_anchor.tokenization_failed is True
    assert implicit_spill.tokenization_failed is True


def test_formula_inspection_traces_direct_implicit_intersection_and_marks_all_forms() -> None:
    literal = inspect_formula("=SUM(@A1:A3)", origin=("Model", "B2"))
    serialized = inspect_formula(
        "=SUM(_xlfn.SINGLE(Inputs!$B$2:$B$4))", origin=("Model", "C3")
    )
    function = inspect_formula("=@INDEX(A1:A3,B1)", origin=("Model", "B2"))
    dynamic_function = inspect_formula("=@OFFSET(A1,1,0)", origin=("Model", "B2"))
    horizontal = inspect_formula("=@A1:C1", origin=("Model", "B2"))
    rectangular = inspect_formula("=@A1:C3", origin=("Model", "B2"))
    outside_range = inspect_formula("=@A1:A3", origin=("Model", "B5"))
    named_range = inspect_formula("=@NamedRange", origin=("Model", "B2"))
    table_current_row = inspect_formula("=[@Amount]", origin=("Model", "B2"))
    string_literal = inspect_formula('=CONCAT("@A1:A3",@A1:A3)', origin=("Model", "B2"))

    assert literal.references == (ParsedReference(None, 1, 2, 1, 2, raw="A1:A3"),)
    assert literal.unresolved_range_tokens == ()
    assert literal.implicit_intersection_tokens == ("@A1:A3",)
    assert serialized.references == (
        ParsedReference("Inputs", 2, 3, 2, 3, raw="Inputs!$B$2:$B$4"),
    )
    assert serialized.implicit_intersection_tokens == ("_xlfn.SINGLE",)
    assert function.references == (
        ParsedReference(None, 1, 1, 1, 3, raw="A1:A3"),
        ParsedReference(None, 2, 1, 2, 1, raw="B1"),
    )
    assert function.implicit_intersection_tokens == ("@INDEX",)
    assert dynamic_function.dynamic_reference_functions == ("OFFSET",)
    assert dynamic_function.implicit_intersection_tokens == ("@OFFSET",)
    assert horizontal.references == (ParsedReference(None, 2, 1, 2, 1, raw="A1:C1"),)
    assert rectangular.references == (ParsedReference(None, 2, 2, 2, 2, raw="A1:C3"),)
    assert outside_range.references == (ParsedReference(None, 1, 1, 1, 3, raw="A1:A3"),)
    assert named_range.unresolved_range_tokens == ("@NamedRange",)
    assert named_range.implicit_intersection_tokens == ("@NamedRange",)
    assert table_current_row.implicit_intersection_tokens == ()
    assert string_literal.references == (ParsedReference(None, 1, 2, 1, 2, raw="A1:A3"),)
    assert string_literal.implicit_intersection_tokens == ("@A1:A3",)


def test_formula_inspection_respects_let_lexical_scope() -> None:
    inspection = inspect_formula(
        "=LET(rate,ExistingRate,LET(rate,Inputs!B3,rate)+rate+UnknownMetric)",
        {
            "existingrate": (
                ParsedReference("Inputs", 2, 2, 2, 2, raw="ExistingRate"),
            )
        },
    )

    assert inspection.references == (
        ParsedReference("Inputs", 2, 2, 2, 2, raw="ExistingRate"),
        ParsedReference("Inputs", 2, 3, 2, 3, raw="Inputs!B3"),
    )
    assert inspection.unresolved_range_tokens == ("UnknownMetric",)


def test_formula_inspection_handles_microsoft_let_example_without_false_gaps() -> None:
    inspection = inspect_formula(
        '=LET(filterCriteria,"Fred",filteredRange,FILTER(A2:D8,A2:A8=filterCriteria),'
        'IF(ISBLANK(filteredRange),"-",filteredRange))'
    )

    assert inspection.references == (
        ParsedReference(None, 1, 2, 4, 8, raw="A2:D8"),
        ParsedReference(None, 1, 2, 1, 8, raw="A2:A8"),
    )
    assert inspection.unresolved_range_tokens == ()


def test_formula_inspection_does_not_shadow_a_let_value_before_its_binding() -> None:
    inspection = inspect_formula(
        "=LET(rate,rate+Inputs!B2,rate*2)",
        {"rate": (ParsedReference("Inputs", 2, 3, 2, 3, raw="rate"),)},
    )

    assert inspection.references == (
        ParsedReference("Inputs", 2, 3, 2, 3, raw="rate"),
        ParsedReference("Inputs", 2, 2, 2, 2, raw="Inputs!B2"),
    )
    assert inspection.unresolved_range_tokens == ()


def test_formula_inspection_respects_inline_lambda_parameters() -> None:
    lambda_call = inspect_formula("=LAMBDA(rate,rate*Inputs!B2)(Inputs!B3)")
    reduce_call = inspect_formula(
        "=REDUCE(0,Inputs!B2:B3,LAMBDA(acc,value,acc+value))"
    )
    namespaced_let = inspect_formula("=_xlfn.LET(rate,Inputs!B2,rate*2)")

    assert lambda_call.references == (
        ParsedReference("Inputs", 2, 2, 2, 2, raw="Inputs!B2"),
        ParsedReference("Inputs", 2, 3, 2, 3, raw="Inputs!B3"),
    )
    assert lambda_call.unresolved_range_tokens == ()
    assert reduce_call.references == (
        ParsedReference("Inputs", 2, 2, 2, 3, raw="Inputs!B2:B3"),
    )
    assert reduce_call.unresolved_range_tokens == ()
    assert namespaced_let.unresolved_range_tokens == ()


def test_named_lambda_definition_detection_requires_one_valid_top_level_lambda() -> None:
    assert lambda_parameter_count("=LAMBDA(value,value+1)") == 1
    assert lambda_parameter_count("=_xlfn.LAMBDA(42)") == 0
    assert lambda_parameter_count("=LAMBDA(value,cmd|'topic'!value)") == 1
    assert (
        lambda_parameter_count(
            "=_xlfn.LAMBDA(_xlpm.temp,_xlpm.temp+Inputs!B2)"
        )
        == 1
    )
    assert lambda_parameter_count("=_xlfn.LAMBDA(_xlop.value,_xlop.value)") == 1
    assert lambda_parameter_count("=LAMBDA(value,value)(1)") is None
    assert lambda_parameter_count("=LAMBDA(A1,A1)") is None
    assert lambda_parameter_count("=SUM(LAMBDA(value,value)(1))") is None


def test_formula_inspection_expands_named_lambda_calls_without_shadowing_locals() -> None:
    named_functions = {
        "tocelsius": (
            ParsedReference("Inputs", 2, 2, 2, 2, raw="ToCelsius"),
        ),
        "unsafelookup": None,
        "f": (
            ParsedReference("ShouldNotAppear", 1, 1, 1, 1, raw="F"),
        ),
    }
    static_call = inspect_formula(
        "=ToCelsius(A2)", named_function_references=named_functions
    )
    unsafe_call = inspect_formula(
        "=UnsafeLookup(A2)", named_function_references=named_functions
    )
    locally_bound_call = inspect_formula(
        "=LET(f,LAMBDA(x,x+Inputs!B2),f(A2))",
        named_function_references=named_functions,
    )
    serialized_lambda = inspect_formula(
        "=_xlfn.LAMBDA(_xlpm.temp,_xlpm.temp+Inputs!B2)(A2)"
    )
    optional_serialized_lambda = inspect_formula(
        "=_xlfn.LAMBDA(_xlop.temp,_xlpm.temp+Inputs!B2)(A2)"
    )
    spaced_serialized_lambda = inspect_formula(
        "=_xlfn.LAMBDA(_xlop.temp, _xlfn.LET(_xlpm.rate, Inputs!B2, "
        "_xlpm.temp*_xlpm.rate))(A2)"
    )
    serialized_let = inspect_formula("=_xlfn.LET(_xlpm.rate,Inputs!B2,_xlpm.rate*2)")

    assert static_call.references == (
        ParsedReference("Inputs", 2, 2, 2, 2, raw="ToCelsius"),
        ParsedReference(None, 1, 2, 1, 2, raw="A2"),
    )
    assert static_call.unresolved_range_tokens == ()
    assert unsafe_call.references == (
        ParsedReference(None, 1, 2, 1, 2, raw="A2"),
    )
    assert unsafe_call.unresolved_range_tokens == ("UnsafeLookup",)
    assert locally_bound_call.references == (
        ParsedReference("Inputs", 2, 2, 2, 2, raw="Inputs!B2"),
        ParsedReference(None, 1, 2, 1, 2, raw="A2"),
    )
    assert locally_bound_call.unresolved_range_tokens == ()
    assert serialized_lambda.references == (
        ParsedReference("Inputs", 2, 2, 2, 2, raw="Inputs!B2"),
        ParsedReference(None, 1, 2, 1, 2, raw="A2"),
    )
    assert serialized_lambda.unresolved_range_tokens == ()
    assert optional_serialized_lambda.references == serialized_lambda.references
    assert optional_serialized_lambda.unresolved_range_tokens == ()
    assert spaced_serialized_lambda.references == serialized_lambda.references
    assert spaced_serialized_lambda.unresolved_range_tokens == ()
    assert serialized_let.references == (
        ParsedReference("Inputs", 2, 2, 2, 2, raw="Inputs!B2"),
    )
    assert serialized_let.unresolved_range_tokens == ()


def test_formula_inspection_propagates_all_external_endpoints_from_names_and_lambdas() -> None:
    workbook_reference = ExternalWorkbookReference(
        source_path="..\\inputs\\source.xlsx",
        sheet="Data",
        min_column=2,
        min_row=3,
        max_column=2,
        max_row=3,
    )
    structured_reference = ExternalWorkbookStructuredReference(
        source_path="..\\inputs\\source.xlsx",
        table_name="Sales",
        table_reference="Sales[Amount]",
    )
    named_formula = inspect_formula(
        "=SUM(ExternalFormula)",
        named_external_workbook_references={
            "externalformula": (workbook_reference,)
        },
        named_external_workbook_structured_references={
            "externalformula": (structured_reference,)
        },
    )
    named_lambda = inspect_formula(
        "=SUM(ExternalLambda(A1))",
        named_function_references={"externallambda": None},
        named_function_external_workbook_references={
            "externallambda": (workbook_reference,)
        },
        named_function_external_workbook_structured_references={
            "externallambda": (structured_reference,)
        },
    )

    for inspection in (named_formula, named_lambda):
        assert inspection.unresolved_range_tokens == ()
        assert inspection.external_workbook_references == (workbook_reference,)
        assert inspection.external_workbook_structured_references == (
            structured_reference,
        )
        assert sum(reference.is_external for reference in inspection.references) == 1


def test_formula_inspection_never_treats_a_cell_reference_as_a_local_variable() -> None:
    inspection = inspect_formula("=LET(A1,Inputs!B2,A1)")

    assert [reference.raw for reference in inspection.references] == [
        "A1",
        "Inputs!B2",
        "A1",
    ]


def test_formula_inspection_resolves_static_structured_table_references() -> None:
    sales = StructuredTable(
        name="Sales",
        sheet="Data",
        ref="A1:C4",
        columns=("Amount", "Rate", "Value"),
        header_row_count=1,
        totals_row_count=0,
    )
    inspection = inspect_formula(
        "=SUM(Sales[Amount])+SUM(Sales[[#Data],[Amount]:[Rate]])+ROWS(Sales[#All])",
        structured_tables={"sales": sales},
    )

    assert inspection.references == (
        ParsedReference("Data", 1, 2, 1, 4, raw="Sales[Amount]"),
        ParsedReference("Data", 1, 2, 2, 4, raw="Sales[[#Data],[Amount]:[Rate]]"),
        ParsedReference("Data", 1, 1, 3, 4, raw="Sales[#All]"),
    )
    assert inspection.unresolved_range_tokens == ()


def test_formula_inspection_resolves_three_d_references_in_tab_order() -> None:
    inspection = inspect_formula(
        "=SUM('Jan 2026:Mar 2026'!$B$2:$B$3)",
        sheet_order=("Jan 2026", "Feb 2026", "Mar 2026", "Summary"),
    )

    assert inspection.references == (
        ParsedReference("Jan 2026", 2, 2, 2, 3, raw="'Jan 2026:Mar 2026'!$B$2:$B$3"),
        ParsedReference("Feb 2026", 2, 2, 2, 3, raw="'Jan 2026:Mar 2026'!$B$2:$B$3"),
        ParsedReference("Mar 2026", 2, 2, 2, 3, raw="'Jan 2026:Mar 2026'!$B$2:$B$3"),
    )
    assert inspection.three_d_reference_tokens == ("'Jan 2026:Mar 2026'!$B$2:$B$3",)
    assert inspection.unresolved_range_tokens == ()


def test_formula_inspection_keeps_three_d_references_visible_without_sheet_order() -> None:
    inspection = inspect_formula("=SUM(Jan:Mar!B2)")
    missing_endpoint = inspect_formula(
        "=SUM(Jan:Mar!B2)", sheet_order=("Jan", "Summary")
    )
    external = inspect_formula(
        "=SUM('[book.xlsx]Jan:Mar'!B2)", sheet_order=("Jan", "Feb", "Mar")
    )

    assert inspection.references == ()
    assert inspection.three_d_reference_tokens == ()
    assert inspection.unresolved_range_tokens == ("Jan:Mar!B2",)
    assert missing_endpoint.unresolved_range_tokens == ("Jan:Mar!B2",)
    assert external.references[0].is_external
    assert external.unresolved_range_tokens == ()


def test_formula_inspection_requires_table_data_origin_for_this_row_references() -> None:
    sales = StructuredTable(
        name="Sales",
        sheet="Data",
        ref="A1:C4",
        columns=("Amount", "Rate", "Value"),
        header_row_count=1,
        totals_row_count=0,
    )

    inspection = inspect_formula("=Sales[@Amount]", structured_tables={"sales": sales})

    assert inspection.references == ()
    assert inspection.unresolved_range_tokens == ("Sales[@Amount]",)


def test_formula_inspection_resolves_current_row_table_references_in_context() -> None:
    sales = StructuredTable(
        name="Sales",
        sheet="Data",
        ref="A1:C4",
        columns=("Sales Amount", "Rate", "Value"),
        header_row_count=1,
        totals_row_count=0,
    )

    inspection = inspect_formula(
        "=[@[Sales Amount]]+[@Rate]+[Sales Amount]+[[Sales Amount]:[Rate]]"
        "+Sales[@Rate]+Sales[[#This Row],[Sales Amount]]"
        "+Sales[[#This Row],[Sales Amount]:[Rate]]",
        structured_tables={"sales": sales},
        origin=("Data", "C3"),
    )

    assert inspection.references == (
        ParsedReference("Data", 1, 3, 1, 3, raw="[@[Sales Amount]]"),
        ParsedReference("Data", 2, 3, 2, 3, raw="[@Rate]"),
        ParsedReference("Data", 1, 3, 1, 3, raw="[Sales Amount]"),
        ParsedReference("Data", 1, 3, 2, 3, raw="[[Sales Amount]:[Rate]]"),
        ParsedReference("Data", 2, 3, 2, 3, raw="Sales[@Rate]"),
        ParsedReference(
            "Data", 1, 3, 1, 3, raw="Sales[[#This Row],[Sales Amount]]"
        ),
        ParsedReference(
            "Data", 1, 3, 2, 3, raw="Sales[[#This Row],[Sales Amount]:[Rate]]"
        ),
    )
    assert inspection.unresolved_range_tokens == ()


def test_formula_inspection_keeps_invalid_current_row_references_visible() -> None:
    sales = StructuredTable(
        name="Sales",
        sheet="Data",
        ref="A1:C5",
        columns=("Amount", "Rate", "Value"),
        header_row_count=1,
        totals_row_count=1,
    )

    outside = inspect_formula(
        "=Sales[@Amount]", structured_tables={"sales": sales}, origin=("Report", "B2")
    )
    header = inspect_formula(
        "=[@Amount]", structured_tables={"sales": sales}, origin=("Data", "A1")
    )
    adjacent_unqualified = inspect_formula(
        "=[Amount]", structured_tables={"sales": sales}, origin=("Data", "D3")
    )
    total = inspect_formula(
        "=Sales[[#This Row],[Amount]]",
        structured_tables={"sales": sales},
        origin=("Data", "C5"),
    )

    assert outside.unresolved_range_tokens == ("Sales[@Amount]",)
    assert header.unresolved_range_tokens == ("[@Amount]",)
    assert adjacent_unqualified.unresolved_range_tokens == ("[Amount]",)
    assert total.unresolved_range_tokens == ("Sales[[#This Row],[Amount]]",)


def test_formula_inspection_resolves_qualified_this_row_from_an_adjacent_cell() -> None:
    sales = StructuredTable(
        name="Sales",
        sheet="Data",
        ref="A1:C5",
        columns=("Amount", "Rate", "Value"),
        header_row_count=1,
        totals_row_count=1,
    )

    inspection = inspect_formula(
        "=Sales[@Amount]+Sales[[#This Row],[Amount]:[Rate]]",
        structured_tables={"sales": sales},
        origin=("Data", "D3"),
    )

    assert inspection.references == (
        ParsedReference("Data", 1, 3, 1, 3, raw="Sales[@Amount]"),
        ParsedReference(
            "Data", 1, 3, 2, 3, raw="Sales[[#This Row],[Amount]:[Rate]]"
        ),
    )
    assert inspection.unresolved_range_tokens == ()


def test_formula_inspection_resolves_table_header_data_and_total_regions() -> None:
    sales = StructuredTable(
        name="Sales",
        sheet="Data",
        ref="A1:C5",
        columns=("Amount", "Rate", "Value"),
        header_row_count=1,
        totals_row_count=1,
    )
    inspection = inspect_formula(
        "=SUM(Sales[[#Headers],[#Data],[Rate]])+SUM(Sales[[#Totals],[Value]])",
        structured_tables={"sales": sales},
    )

    assert inspection.references == (
        ParsedReference("Data", 2, 1, 2, 1, raw="Sales[[#Headers],[#Data],[Rate]]"),
        ParsedReference("Data", 2, 2, 2, 4, raw="Sales[[#Headers],[#Data],[Rate]]"),
        ParsedReference("Data", 3, 5, 3, 5, raw="Sales[[#Totals],[Value]]"),
    )
