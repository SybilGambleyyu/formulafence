"""Executable contract checks for the public composite GitHub Action."""

from __future__ import annotations

import os
import shutil
import subprocess
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
ACTION_PATH = REPOSITORY_ROOT / "action.yml"
ACTION_SCRIPT = REPOSITORY_ROOT / "scripts" / "run-github-action.sh"


def _workbook(path: Path, value: object) -> None:
    workbook = Workbook()
    workbook.active.title = "Model"
    workbook.active["A1"] = value
    workbook.save(path)


def _run_action_script(
    tmp_path: Path,
    *,
    baseline: Path,
    candidate: Path,
    policy: Path | None = None,
    output: str = "reports/formulafence.md",
    report_format: str = "markdown",
    redact_external_workbook_links: str = "false",
    redact_formula_external_actions: str = "false",
    redact_python_in_excel: str = "false",
    redact_office_custom_functions: str = "false",
    redact_unqualified_runtime_functions: str = "false",
    redact_worksheet_code_resource_registrations: str = "false",
    redact_formula_defined_xlm_registrations: str = "false",
    redact_formula_defined_xlm_evaluations: str = "false",
    max_workbooks: str = "512",
    max_link_impact: str = "100000",
    upload_artifact: str = "true",
) -> tuple[subprocess.CompletedProcess[str], Path, Path]:
    outputs = tmp_path / "outputs.txt"
    summary = tmp_path / "summary.md"

    def workspace_input(path: Path) -> str:
        try:
            return str(path.relative_to(tmp_path))
        except ValueError:
            return str(path)

    environment = os.environ.copy()
    environment.update(
        {
            "GITHUB_ACTION_PATH": str(REPOSITORY_ROOT),
            "GITHUB_WORKSPACE": str(tmp_path),
            "GITHUB_OUTPUT": str(outputs),
            "GITHUB_STEP_SUMMARY": str(summary),
            "INPUT_BASELINE": workspace_input(baseline),
            "INPUT_CANDIDATE": workspace_input(candidate),
            "INPUT_POLICY": workspace_input(policy) if policy is not None else "",
            "INPUT_FORMAT": report_format,
            "INPUT_OUTPUT": output,
            "INPUT_REDACT_EXTERNAL_WORKBOOK_LINKS": redact_external_workbook_links,
            "INPUT_REDACT_FORMULA_EXTERNAL_ACTIONS": redact_formula_external_actions,
            "INPUT_REDACT_PYTHON_IN_EXCEL": redact_python_in_excel,
            "INPUT_REDACT_OFFICE_CUSTOM_FUNCTIONS": redact_office_custom_functions,
            "INPUT_REDACT_UNQUALIFIED_RUNTIME_FUNCTIONS": (
                redact_unqualified_runtime_functions
            ),
            "INPUT_REDACT_WORKSHEET_CODE_RESOURCE_REGISTRATIONS": (
                redact_worksheet_code_resource_registrations
            ),
            "INPUT_REDACT_FORMULA_DEFINED_XLM_REGISTRATIONS": (
                redact_formula_defined_xlm_registrations
            ),
            "INPUT_REDACT_FORMULA_DEFINED_XLM_EVALUATIONS": (
                redact_formula_defined_xlm_evaluations
            ),
            "INPUT_FAIL_ON": "none",
            "INPUT_MAX_WORKBOOKS": max_workbooks,
            "INPUT_MAX_LINK_IMPACT": max_link_impact,
            "INPUT_INSTALL": "false",
            "INPUT_UPLOAD_ARTIFACT": upload_artifact,
            "PYTHONPATH": str(REPOSITORY_ROOT / "src"),
            "PATH": f"{Path(sys.executable).parent}{os.pathsep}{environment['PATH']}",
        }
    )
    result = subprocess.run(
        ["bash", str(ACTION_SCRIPT)],
        cwd=tmp_path,
        env=environment,
        text=True,
        capture_output=True,
        check=False,
    )
    return result, outputs, summary


def test_action_metadata_exposes_policy_report_contract() -> None:
    action = yaml.safe_load(ACTION_PATH.read_text(encoding="utf-8"))

    assert action["runs"]["using"] == "composite"
    assert {
        "baseline",
        "candidate",
        "policy",
        "format",
        "output",
        "redact-external-workbook-links",
        "redact-formula-external-actions",
        "redact-python-in-excel",
        "redact-office-custom-functions",
        "redact-unqualified-runtime-functions",
        "redact-worksheet-code-resource-registrations",
        "redact-formula-defined-xlm-registrations",
        "redact-formula-defined-xlm-evaluations",
        "max-workbooks",
        "max-link-impact",
    } <= set(action["inputs"])
    assert {"report-path", "exit-code"} <= set(action["outputs"])
    steps = action["runs"]["steps"]
    upload_index = next(
        index
        for index, step in enumerate(steps)
        if step.get("uses")
        == "actions/upload-artifact@043fb46d1a93c77aae656e7c1c64a875d1fc6a0a"
    )
    enforce_index = next(
        index
        for index, step in enumerate(steps)
        if step.get("name") == "Enforce FormulaFence result"
    )
    assert upload_index < enforce_index


def test_action_preserves_a_policy_failure_until_after_report_generation(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    policy = tmp_path / "formulafence.yml"
    _workbook(baseline, "=1+1")
    _workbook(candidate, 2)
    policy.write_text(
        "version: 1\nrules:\n  no_formula_to_value: true\n", encoding="utf-8"
    )

    result, outputs, summary = _run_action_script(
        tmp_path, baseline=baseline, candidate=candidate, policy=policy
    )

    assert result.returncode == 0
    assert "exit-code=1" in outputs.read_text(encoding="utf-8")
    assert "report-written=true" in outputs.read_text(encoding="utf-8")
    report = tmp_path / "reports" / "formulafence.md"
    assert report.is_file()
    assert "FFP001" in report.read_text(encoding="utf-8")
    summary_text = summary.read_text(encoding="utf-8")
    assert "FFP001" in summary_text
    assert "\\n**Report:" not in summary_text
    assert "\n**Report:" in summary_text


def test_action_refuses_an_output_outside_the_workspace(tmp_path: Path) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        output="../outside.md",
    )

    assert result.returncode == 2
    assert "report output must stay" in result.stderr
    assert not (tmp_path.parent / "outside.md").exists()


def test_action_refuses_an_input_outside_the_workspace(tmp_path: Path) -> None:
    outside_baseline = tmp_path.parent / f"{tmp_path.name}-approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(outside_baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=outside_baseline,
        candidate=candidate,
    )

    assert result.returncode == 2
    assert "INPUT_BASELINE must stay inside GITHUB_WORKSPACE" in result.stderr


def test_action_rejects_an_ambiguous_artifact_switch(tmp_path: Path) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        upload_artifact="sometimes",
    )

    assert result.returncode == 2
    assert "Unsupported upload-artifact value" in result.stderr


def test_action_can_redact_external_workbook_link_material(tmp_path: Path) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    baseline_marker = "PRIVATE-ACTION-BASELINE"
    candidate_marker = "PRIVATE-ACTION-CANDIDATE"
    _workbook(baseline, f"='C:\\{baseline_marker}\\[Source.xlsx]Inputs'!$B$2")
    _workbook(candidate, f"='C:\\{candidate_marker}\\[Source.xlsx]Inputs'!$B$2")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        report_format="json",
        output="reports/formulafence.json",
        redact_external_workbook_links="true",
    )

    assert result.returncode == 0
    rendered = (tmp_path / "reports" / "formulafence.json").read_text(encoding="utf-8")
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered
    assert "external-workbook link material redacted" in rendered


def test_action_rejects_an_invalid_external_workbook_redaction_switch(tmp_path: Path) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        redact_external_workbook_links="sometimes",
    )

    assert result.returncode == 2
    assert "Unsupported redact-external-workbook-links value" in result.stderr


def test_action_can_redact_formula_external_action_material(tmp_path: Path) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    baseline_marker = "PRIVATE-ACTION-LINK-BASELINE"
    candidate_marker = "PRIVATE-ACTION-LINK-CANDIDATE"
    _workbook(
        baseline,
        f'=HYPERLINK("https://private.example.test/{baseline_marker}", "Open")',
    )
    _workbook(
        candidate,
        f'=HYPERLINK("https://private.example.test/{candidate_marker}", "Open")',
    )

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        report_format="json",
        output="reports/formulafence.json",
        redact_formula_external_actions="true",
    )

    assert result.returncode == 0
    rendered = (tmp_path / "reports" / "formulafence.json").read_text(encoding="utf-8")
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered
    assert "formula external-action material redacted" in rendered


def test_action_rejects_an_invalid_formula_external_action_redaction_switch(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        redact_formula_external_actions="sometimes",
    )

    assert result.returncode == 2
    assert "Unsupported redact-formula-external-actions value" in result.stderr


def test_action_can_redact_python_in_excel_material(tmp_path: Path) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    baseline_marker = "PRIVATE-PY-SOURCE-BASELINE"
    candidate_marker = "PRIVATE-PY-SOURCE-CANDIDATE"
    _workbook(baseline, f'=PY("{baseline_marker}",0)')
    _workbook(candidate, f'=PY("{candidate_marker}",0)')

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        report_format="json",
        output="reports/formulafence.json",
        redact_python_in_excel="true",
    )

    assert result.returncode == 0
    rendered = (tmp_path / "reports" / "formulafence.json").read_text(encoding="utf-8")
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered
    assert "Python-in-Excel material redacted" in rendered


def test_action_rejects_an_invalid_python_in_excel_redaction_switch(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        redact_python_in_excel="sometimes",
    )

    assert result.returncode == 2
    assert "Unsupported redact-python-in-excel value" in result.stderr


def test_action_can_redact_office_custom_function_material(tmp_path: Path) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    baseline_marker = "PRIVATE-CUSTOM-FUNCTION-SOURCE-BASELINE"
    candidate_marker = "PRIVATE-CUSTOM-FUNCTION-SOURCE-CANDIDATE"
    _workbook(baseline, f'=CONTOSO.GETDATA("{baseline_marker}")')
    _workbook(candidate, f'=CONTOSO.GETDATA("{candidate_marker}")')

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        report_format="json",
        output="reports/formulafence.json",
        redact_office_custom_functions="true",
    )

    assert result.returncode == 0
    rendered = (tmp_path / "reports" / "formulafence.json").read_text(encoding="utf-8")
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered
    assert "Office custom-function material redacted" in rendered


def test_action_rejects_an_invalid_office_custom_function_redaction_switch(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        redact_office_custom_functions="sometimes",
    )

    assert result.returncode == 2
    assert "Unsupported redact-office-custom-functions value" in result.stderr


def test_action_can_redact_unqualified_runtime_function_material(tmp_path: Path) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    baseline_marker = "PRIVATE-RUNTIME-FUNCTION-SOURCE-BASELINE"
    candidate_marker = "PRIVATE-RUNTIME-FUNCTION-SOURCE-CANDIDATE"
    _workbook(baseline, f'=PRIVATEUDF("{baseline_marker}")')
    _workbook(candidate, f'=PRIVATEUDF("{candidate_marker}")')

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        report_format="json",
        output="reports/formulafence.json",
        redact_unqualified_runtime_functions="true",
    )

    assert result.returncode == 0
    rendered = (tmp_path / "reports" / "formulafence.json").read_text(encoding="utf-8")
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered
    assert "unqualified runtime-function material redacted" in rendered


def test_action_rejects_an_invalid_unqualified_runtime_function_redaction_switch(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        redact_unqualified_runtime_functions="sometimes",
    )

    assert result.returncode == 2
    assert "Unsupported redact-unqualified-runtime-functions value" in result.stderr


def test_action_can_redact_worksheet_code_resource_registration_material(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    baseline_marker = "PRIVATE-REGISTRATION-SOURCE-BASELINE"
    candidate_marker = "PRIVATE-REGISTRATION-SOURCE-CANDIDATE"
    procedure_marker = "PRIVATE-REGISTRATION-PROCEDURE"
    _workbook(
        baseline,
        f'=REGISTER.ID("{baseline_marker}","{procedure_marker}","J!")',
    )
    _workbook(
        candidate,
        f'=REGISTER.ID("{candidate_marker}","{procedure_marker}","J!")',
    )

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        report_format="json",
        output="reports/formulafence.json",
        redact_worksheet_code_resource_registrations="true",
    )

    assert result.returncode == 0
    rendered = (tmp_path / "reports" / "formulafence.json").read_text(encoding="utf-8")
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered
    assert procedure_marker not in rendered
    assert "worksheet code-resource registration material redacted" in rendered


def test_action_rejects_an_invalid_worksheet_code_resource_registration_switch(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        redact_worksheet_code_resource_registrations="sometimes",
    )

    assert result.returncode == 2
    assert (
        "Unsupported redact-worksheet-code-resource-registrations value"
        in result.stderr
    )


def test_action_can_redact_formula_defined_xlm_registration_material(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    baseline_marker = "PRIVATE-ACTION-XLM-MODULE-BASELINE"
    candidate_marker = "PRIVATE-ACTION-XLM-MODULE-CANDIDATE"
    baseline_procedure = "PRIVATE-ACTION-XLM-PROCEDURE-BASELINE"
    candidate_procedure = "PRIVATE-ACTION-XLM-PROCEDURE-CANDIDATE"

    def xlm_registration_workbook(path: Path, module: str, procedure: str) -> None:
        workbook = Workbook()
        workbook.active.title = "Inputs"
        workbook.active["A9"] = module
        workbook.active["B2"] = "=FENCE.XLM.REGISTER(A9)"
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.REGISTER",
                attr_text=f'=LAMBDA(module,REGISTER(module,"{procedure}","J!"))',
            )
        )
        workbook.save(path)

    xlm_registration_workbook(baseline, baseline_marker, baseline_procedure)
    xlm_registration_workbook(candidate, candidate_marker, candidate_procedure)

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        report_format="json",
        output="reports/formulafence.json",
        redact_formula_defined_xlm_registrations="true",
    )

    assert result.returncode == 0
    rendered = (tmp_path / "reports" / "formulafence.json").read_text(encoding="utf-8")
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered
    assert baseline_procedure not in rendered
    assert candidate_procedure not in rendered
    assert "formula-defined XLM registration material redacted" in rendered
    assert "FF068" in rendered


def test_action_rejects_an_invalid_formula_defined_xlm_registration_switch(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        redact_formula_defined_xlm_registrations="sometimes",
    )

    assert result.returncode == 2
    assert "Unsupported redact-formula-defined-xlm-registrations value" in result.stderr


def test_action_can_redact_formula_defined_xlm_evaluation_material(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    baseline_marker = "PRIVATE-ACTION-XLM-EVALUATION-BASELINE"
    candidate_marker = "PRIVATE-ACTION-XLM-EVALUATION-CANDIDATE"
    baseline_expression = "PRIVATE-ACTION-XLM-EXPRESSION-BASELINE"
    candidate_expression = "PRIVATE-ACTION-XLM-EXPRESSION-CANDIDATE"

    def xlm_evaluation_workbook(path: Path, expression: str, input_value: str) -> None:
        workbook = Workbook()
        workbook.active.title = "Inputs"
        workbook.active["A9"] = input_value
        workbook.active["B2"] = "=FENCE.XLM.EVALUATE(A9)"
        workbook.defined_names.add(
            DefinedName(
                "FENCE.XLM.EVALUATE",
                attr_text=f'=LAMBDA(expression,EVALUATE("{expression}"))',
            )
        )
        workbook.save(path)

    xlm_evaluation_workbook(baseline, baseline_expression, baseline_marker)
    xlm_evaluation_workbook(candidate, candidate_expression, candidate_marker)

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        report_format="json",
        output="reports/formulafence.json",
        redact_formula_defined_xlm_evaluations="true",
    )

    assert result.returncode == 0
    rendered = (tmp_path / "reports" / "formulafence.json").read_text(encoding="utf-8")
    assert baseline_marker not in rendered
    assert candidate_marker not in rendered
    assert baseline_expression not in rendered
    assert candidate_expression not in rendered
    assert "formula-defined XLM evaluation material redacted" in rendered
    assert "FF069" in rendered


def test_action_rejects_an_invalid_formula_defined_xlm_evaluation_switch(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "approved.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _workbook(baseline, "=1+1")
    _workbook(candidate, "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        redact_formula_defined_xlm_evaluations="sometimes",
    )

    assert result.returncode == 2
    assert "Unsupported redact-formula-defined-xlm-evaluations value" in result.stderr


def test_action_runs_a_directory_portfolio_and_preserves_membership_evidence(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    baseline.mkdir()
    candidate.mkdir()
    _workbook(baseline / "model.xlsx", "=1+1")
    _workbook(candidate / "model.xlsx", "=1+1")
    _workbook(candidate / "added.xlsx", "=1+1")
    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_portfolio_membership_changes: true\n",
        encoding="utf-8",
    )

    result, outputs, summary = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        policy=policy,
    )

    assert result.returncode == 0
    assert "exit-code=1" in outputs.read_text(encoding="utf-8")
    report = tmp_path / "reports" / "formulafence.md"
    assert "FF077" in report.read_text(encoding="utf-8")
    assert "FFP077" in summary.read_text(encoding="utf-8")


def test_action_enforces_static_cross_workbook_portfolio_impacts(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    baseline.mkdir()
    source = baseline / "source.xlsx"
    summary_workbook = baseline / "summary.xlsx"
    _workbook(source, 1)
    _workbook(summary_workbook, "=[source.xlsx]Model!A1")
    shutil.copytree(baseline, candidate)
    workbook = Workbook()
    workbook.active.title = "Model"
    workbook.active["A1"] = 2
    workbook.save(candidate / "source.xlsx")
    policy = tmp_path / "formulafence.yml"
    policy.write_text(
        "version: 1\nrules:\n  no_cross_workbook_impacts: true\n",
        encoding="utf-8",
    )

    result, outputs, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        policy=policy,
    )

    assert result.returncode == 0
    assert "exit-code=1" in outputs.read_text(encoding="utf-8")
    assert "FF079" in (tmp_path / "reports" / "formulafence.md").read_text(
        encoding="utf-8"
    )
    assert "FFP079" in (tmp_path / "reports" / "formulafence.md").read_text(
        encoding="utf-8"
    )


def test_action_rejects_an_invalid_portfolio_limit(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    baseline.mkdir()
    candidate.mkdir()
    _workbook(baseline / "model.xlsx", "=1+1")
    _workbook(candidate / "model.xlsx", "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        max_workbooks="0",
    )

    assert result.returncode == 2
    assert "max-workbooks must be a positive integer" in result.stderr


def test_action_rejects_an_invalid_cross_workbook_impact_limit(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    baseline.mkdir()
    candidate.mkdir()
    _workbook(baseline / "model.xlsx", "=1+1")
    _workbook(candidate / "model.xlsx", "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        max_link_impact="0",
    )

    assert result.returncode == 2
    assert "max-link-impact must be a positive integer" in result.stderr


def test_action_refuses_a_report_inside_a_portfolio_input(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    baseline.mkdir()
    candidate.mkdir()
    _workbook(baseline / "model.xlsx", "=1+1")
    _workbook(candidate / "model.xlsx", "=1+1")

    result, _, _ = _run_action_script(
        tmp_path,
        baseline=baseline,
        candidate=candidate,
        output="baseline/report.md",
    )

    assert result.returncode == 2
    assert "report output must not be written inside an input directory" in result.stderr
    assert not (baseline / "report.md").exists()
