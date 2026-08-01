"""Tests for conservative single-workbook formula linting."""

from __future__ import annotations

from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Protection

from formulafence.lint import lint_snapshot
from formulafence.models import (
    ArrayFormulaRange,
    CellProtectionAssignmentSnapshot,
    FormulaFenceError,
)
from formulafence.output import as_json, lint_to_markdown, lint_to_sarif
from formulafence.workbook import load_snapshot

from .helpers import (
    make_calculated_column_model,
    make_formula_cached_result_model,
    make_ignored_error_model,
)


def _snapshot(
    tmp_path: Path,
    cells: dict[str, object],
    *,
    protected: bool = False,
    unlocked_cells: tuple[str, ...] = (),
    calculation_mode: str | None = None,
    calculation_completed: bool | None = None,
    calculation_iteration_enabled: bool | None = None,
    additional_sheets: dict[str, dict[str, object]] | None = None,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    for coordinate, value in cells.items():
        worksheet[coordinate] = value
    for title, sheet_cells in (additional_sheets or {}).items():
        additional_worksheet = workbook.create_sheet(title)
        for coordinate, value in sheet_cells.items():
            additional_worksheet[coordinate] = value
    for coordinate in unlocked_cells:
        worksheet[coordinate].protection = Protection(locked=False)
    worksheet.protection.sheet = protected
    if calculation_mode is not None:
        workbook.calculation.calcMode = calculation_mode
    if calculation_completed is not None:
        workbook.calculation.calcCompleted = calculation_completed
    if calculation_iteration_enabled is not None:
        workbook.calculation.iterate = calculation_iteration_enabled
    path = tmp_path / "model.xlsx"
    workbook.save(path)
    return load_snapshot(path)


def test_lint_reports_blank_gap_inside_supported_column_copy_block(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "A3": 2,
            "A4": 3,
            "A5": 4,
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
        },
    )

    report = lint_snapshot(snapshot)

    assert len(report.findings) == 1
    finding = report.findings[0]
    assert (finding.rule_id, finding.severity, finding.location) == (
        "FF082",
        "high",
        ("Model", "B3"),
    )
    assert finding.details == {
        "pattern_kind": "blank_gap",
        "pattern_evidence": [
            {
                "orientation": "column",
                "preceding_formula": "Model!B2",
                "following_formula": "Model!B4",
                "supporting_formula": "Model!B5",
            }
        ],
    }


def test_lint_reports_non_formula_gap_inside_supported_row_copy_block(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "C2": 99,
            "D2": "=C2*2",
            "E2": "=D2*2",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [
        ("FF082", "medium", ("Model", "C2"))
    ]
    assert report.findings[0].details["pattern_kind"] == "non_formula_gap"


def test_lint_treats_textual_formula_exceptions_as_low_severity(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "C2": "n.a.",
            "D2": "=C2*2",
            "E2": "=D2*2",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [
        ("FF082", "low", ("Model", "C2"))
    ]
    assert report.findings[0].details["pattern_kind"] == "text_gap"


def test_lint_treats_stored_error_interruption_as_high_severity(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "C2": "#N/A",
            "D2": "=C2*2",
            "E2": "=D2*2",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [
        ("FF082", "high", ("Model", "C2"))
    ]
    assert report.findings[0].details["pattern_kind"] == "error_gap"


def test_lint_reports_formula_outlier_inside_supported_row_copy_block(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "C2": "=B2*3",
            "D2": "=C2*2",
            "E2": "=D2*2",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [
        ("FF083", "medium", ("Model", "C2"))
    ]
    assert report.findings[0].details["pattern_kind"] == "formula_outlier"


def test_lint_requires_a_third_matching_peer_before_reporting(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "A4": 3,
            "B2": "=A2*2",
            "D2": "=C2*2",
        },
    )

    assert lint_snapshot(snapshot).findings == []


def test_lint_accepts_supporting_peer_before_the_interruption(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "C2": "=B2*2",
            "E2": "=D2*2",
        },
    )

    report = lint_snapshot(snapshot)

    assert [(finding.rule_id, finding.location) for finding in report.findings] == [
        ("FF082", ("Model", "D2"))
    ]
    assert report.findings[0].details["pattern_evidence"][0]["supporting_formula"] == "Model!B2"


def test_lint_skips_complete_copy_block(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "C2": "=B2*2",
            "D2": "=C2*2",
            "E2": "=D2*2",
        },
    )

    assert lint_snapshot(snapshot).findings == []


def test_lint_reports_omitted_numeric_column_run_after_simple_aggregate(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": 10,
            "B3": 20,
            "B4": 30,
            "B5": 40,
            "B6": 50,
            "B7": 60,
            "B8": "=SUM($B$2:$B$4)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF084", "medium", ("Model", "B8"))]
    assert report.findings[0].details == {
        "aggregate_function": "SUM",
        "orientation": "column",
        "referenced_range": "Model!B2:B4",
        "omitted_range": "Model!B5:B7",
        "omitted_cell_count": 3,
    }


def test_lint_handles_a_local_aggregate_in_excel_last_column(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "XFD2": 10,
            "XFD3": 20,
            "XFD4": 30,
            "XFD5": 40,
            "XFD6": 50,
            "XFD7": 60,
            "XFD8": "=SUM(XFD2:XFD4)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [(finding.rule_id, finding.location) for finding in report.findings] == [
        ("FF084", ("Model", "XFD8"))
    ]


def test_lint_reports_omitted_numeric_row_run_after_simple_aggregate(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": 10,
            "C2": 20,
            "D2": 30,
            "E2": 40,
            "F2": 50,
            "G2": "=AVERAGE(Model!B2:D2)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.location) for finding in report.findings
    ] == [("FF084", ("Model", "G2"))]
    assert report.findings[0].details == {
        "aggregate_function": "AVERAGE",
        "orientation": "row",
        "referenced_range": "Model!B2:D2",
        "omitted_range": "Model!E2:F2",
        "omitted_cell_count": 2,
    }


def test_lint_reports_a_directly_unlocked_formula_on_a_protected_sheet(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=A2*2"},
        protected=True,
        unlocked_cells=("B2",),
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF085", "medium", ("Model", "B2"))]
    assert report.findings[0].details == {"protection_scope": "direct_cell"}


def test_lint_keeps_non_direct_or_inactive_unlocked_formula_controls_quiet(
    tmp_path: Path,
) -> None:
    inactive = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=A2*2"},
        unlocked_cells=("B2",),
    )
    assert lint_snapshot(inactive).findings == []

    protected = _snapshot(tmp_path, {"A2": 10, "B2": "=A2*2"}, protected=True)
    protected.cell_protection_assignments = (
        CellProtectionAssignmentSnapshot(
            sheet="Model",
            scope="column",
            target="B:B",
            locked=False,
            hidden=False,
        ),
    )
    assert lint_snapshot(protected).findings == []


def test_lint_reports_explicit_incomplete_manual_formula_calculation(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=A2*2"},
        calculation_mode="manual",
        calculation_completed=False,
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF086", "medium", None)]
    assert report.findings[0].details == {
        "calculation_mode": "manual",
        "calculation_completed_before_save": False,
    }


def test_lint_keeps_completed_or_non_manual_calculation_states_quiet(tmp_path: Path) -> None:
    completed_manual = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=A2*2"},
        calculation_mode="manual",
        calculation_completed=True,
    )
    incomplete_automatic = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=A2*2"},
        calculation_mode="auto",
        calculation_completed=False,
    )
    formula_free_manual = _snapshot(
        tmp_path,
        {"A2": 10},
        calculation_mode="manual",
        calculation_completed=False,
    )

    assert lint_snapshot(completed_manual).findings == []
    assert lint_snapshot(incomplete_automatic).findings == []
    assert lint_snapshot(formula_free_manual).findings == []


def test_lint_reports_direct_static_self_reference_when_iteration_is_disabled(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(tmp_path, {"B2": "=B2+1"})

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF087", "high", ("Model", "B2"))]
    assert report.findings[0].details == {
        "calculation_iteration_enabled": False,
        "reference_scope": "direct_static",
    }


def test_lint_reports_excel_error_checking_suppressions(tmp_path: Path) -> None:
    snapshot = load_snapshot(make_ignored_error_model(tmp_path / "ignored-errors.xlsx"))

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF091", "medium", None)]
    assert report.findings[0].details == {
        "suppressed_warning_counts": {
            "evaluation_error": 2,
            "inconsistent_formula": 2,
            "formula_range_omission": 1,
            "unlocked_formula": 1,
            "empty_cell_reference": 1,
            "list_data_validation": 1,
            "calculated_column": 1,
            "number_stored_as_text": 1,
            "two_digit_text_year": 1,
        },
        "suppressed_warning_rule_count": 11,
        "suppressed_warning_target_range_count": 5,
    }


@pytest.mark.parametrize(
    ("exception", "expected_kind"),
    [
        (None, "blank"),
        (99, "non_formula_value"),
        ("manual exception", "text_value"),
        ("#DIV/0!", "stored_error_value"),
        ("=A3+B3", "formula_mismatch"),
    ],
)
def test_lint_reports_isolated_table_calculated_column_exceptions(
    tmp_path: Path,
    exception: object,
    expected_kind: str,
) -> None:
    snapshot = load_snapshot(
        make_calculated_column_model(
            tmp_path / "calculated-column.xlsx",
            exception=exception,
        )
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF092", "medium", ("Private Table Sheet", "C3"))]
    assert report.findings[0].details == {
        "exception_kind": expected_kind,
        "matching_adjacent_formula_peers": 2,
        "evidence_scope": "table_calculated_column",
    }
    declaration = snapshot.tables["PRIVATE_RESULT_TABLE"].calculated_column_formulas
    assert [(entry.column_index, entry.formula_fingerprint) for entry in declaration] == [
        (3, "R[0]C[-2]*R[0]C[-1]")
    ]


def test_lint_keeps_table_calculated_column_boundary_and_array_exceptions_quiet(
    tmp_path: Path,
) -> None:
    boundary = load_snapshot(
        make_calculated_column_model(
            tmp_path / "boundary.xlsx",
            exception=None,
            exception_coordinate="C2",
        )
    )
    array_master = load_snapshot(
        make_calculated_column_model(
            tmp_path / "array.xlsx",
            exception=99,
            array_formula=True,
        )
    )

    assert lint_snapshot(boundary).findings == []
    assert array_master.tables["PRIVATE_RESULT_TABLE"].calculated_column_formulas == ()
    assert lint_snapshot(array_master).findings == []


def test_lint_matches_a_structured_reference_table_calculated_column_master(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_calculated_column_model(
            tmp_path / "structured-master.xlsx",
            exception=99,
            structured_formula=True,
        )
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.location)
        for finding in report.findings
    ] == [("FF092", ("Private Table Sheet", "C3"))]


def test_lint_table_calculated_column_leaves_explicit_broken_formula_to_ff088(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_calculated_column_model(
            tmp_path / "broken-formula.xlsx",
            exception="=IFERROR(#REF!,0)",
        )
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF088", "critical", ("Private Table Sheet", "C3"))]


def test_lint_table_calculated_column_signal_does_not_repeat_stronger_copy_pattern(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_calculated_column_model(
            tmp_path / "copied-pattern.xlsx",
            exception=99,
            data_row_count=4,
        )
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF082", "medium", ("Private Table Sheet", "C3"))]


def test_lint_table_calculated_column_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_calculated_column_model(
            tmp_path / "too-many-exceptions.xlsx",
            data_row_count=5,
            exceptions={"C3": 99, "C5": 99},
        )
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_table_calculated_column_does_not_walk_an_oversized_sparse_ref(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(make_calculated_column_model(tmp_path / "sparse.xlsx"))
    cells = dict(snapshot.cells)
    cells.pop(("Private Table Sheet", "C3"))
    table = snapshot.tables["PRIVATE_RESULT_TABLE"]
    sparse_snapshot = replace(
        snapshot,
        cells=cells,
        tables={
            "PRIVATE_RESULT_TABLE": replace(table, ref="A1:C1048576"),
        },
    )

    report = lint_snapshot(sparse_snapshot)

    assert [
        (finding.rule_id, finding.location)
        for finding in report.findings
    ] == [("FF092", ("Private Table Sheet", "C3"))]


def test_lint_reports_direct_conditional_aggregate_range_shape_mismatches(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": (
                "=IFERROR(SUMIFS(C2:C10,A2:A12,A14)"
                "+COUNTIFS(B2:B10,1,C2:C8,2)"
                "+AVERAGEIFS(D2:D10,A2:A12,A14)"
                "+MAXIFS(E2:E10,A2:A12,A14)"
                "+MINIFS(F2:F10,A2:A12,A14),0)"
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF093", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "conditional_aggregate_call_count": 5,
        "mismatched_direct_range_argument_count": 5,
        "evidence_scope": "conditional_aggregate_direct_a1_ranges",
    }


def test_lint_conditional_aggregate_range_shape_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=SUMIFS(C2:C10,A2:A10,A1)",
            "A3": "=SUMIFS(C2:C10,NamedCriteriaRange,A1)",
            "A4": "=SUMIFS(Table1[Total],Table1[State],A1)",
            "A5": "=SUMIFS(C2:C10,OFFSET(A2,0,0,11,1),A1)",
            "A6": "=SUMIFS(C2:C10,A2#,A1)",
            "A7": "=SUMIFS(C2:C10,[Inputs.xlsx]Data!A2:A12,A1)",
            "A8": "=SUMIFS(C2:C10,A2:A10,#REF!)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [finding.rule_id for finding in report.findings] == ["FF114", "FF088"]
    assert report.findings[0].details == {
        "closed_external_criteria_function_call_count": 1,
        "affected_formula_cell_count": 1,
        "direct_external_a1_reference_argument_count": 1,
        "countblank_call_count": 0,
        "countif_call_count": 0,
        "countifs_call_count": 0,
        "sumif_call_count": 0,
        "sumifs_call_count": 1,
        "evidence_scope": "direct_external_a1_criteria_function_arguments",
    }


def test_lint_conditional_aggregate_range_shape_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=SUMIFS(C2:C10,A2:A10,A1)",
            "B3": "=SUMIFS(C3:C11,A3:A13,A2)",
            "B4": "=SUMIFS(C4:C12,A4:A12,A3)",
            "B5": "=SUMIFS(C5:C13,A5:A13,A4)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF093", "high", ("Model", "B3"))]


def test_lint_conditional_aggregate_range_shape_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=SUMIFS(C2:C10,A2:A12,A14)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_sumproduct_range_shape_mismatches(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": (
                "=IFERROR(SUMPRODUCT(C2:C10,A2:A12)"
                "+SUMPRODUCT(B2:C4,D2:E3),0)"
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF094", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "sumproduct_call_count": 2,
        "mismatched_direct_array_argument_count": 2,
        "evidence_scope": "sumproduct_direct_a1_ranges",
    }


def test_lint_sumproduct_range_shape_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=SUMPRODUCT(C2:C10,A2:A10)",
            "A3": "=SUMPRODUCT(C2:C10,NamedRange)",
            "A4": "=SUMPRODUCT(Table1[Price],Table1[Quantity])",
            "A5": "=SUMPRODUCT(C2:C10,OFFSET(A2,0,0,11,1))",
            "A6": "=SUMPRODUCT(C2:C10,A2#)",
            "A7": "=SUMPRODUCT(C2:C10,[Inputs.xlsx]Data!A2:A12)",
            "A8": "=SUMPRODUCT(C2:C10,A2:A10,#REF!)",
            "A9": "=SUMPRODUCT((C2:C10=A2:A12)*B2:B10)",
        },
    )

    report = lint_snapshot(snapshot)

    assert "FF094" not in {finding.rule_id for finding in report.findings}


def test_lint_sumproduct_range_shape_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=SUMPRODUCT(C2:C10,A2:A10)",
            "B3": "=SUMPRODUCT(C3:C11,A3:A13)",
            "B4": "=SUMPRODUCT(C4:C12,A4:A12)",
            "B5": "=SUMPRODUCT(C5:C13,A5:A13)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF094", "high", ("Model", "B3"))]


def test_lint_sumproduct_range_shape_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=SUMPRODUCT(C2:C10,A2:A12)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_structural_formula_rules_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=SUMIFS(C2:C10,A2:A12,A14)",
            "A3": "=SUMPRODUCT(C2:C10,A2:A12)",
            "A4": "=MMULT(A2:B4,C2:D5)",
            "A5": "=VLOOKUP(A2,C2:D4,3,FALSE)",
            "A6": "=CHOOSE(0,A2)",
            "A7": "=RANDBETWEEN(2,1)",
            "A8": "=SUBTOTAL(12,A2)",
            "A9": "=INDEX(A2:B4,4)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=7",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=7)


def test_lint_reports_direct_mmult_dimension_mismatches(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": (
                "=IFERROR(MMULT(C2:D4,A2:B6)"
                "+MMULT(B2:C4,D2:E5),0)"
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF095", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "mmult_call_count": 2,
        "incompatible_direct_matrix_pair_count": 2,
        "evidence_scope": "mmult_direct_a1_arrays",
    }


def test_lint_mmult_dimension_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=MMULT(C2:D4,A2:B3)",
            "A3": "=MMULT(C2:D4,NamedMatrix)",
            "A4": "=MMULT(Table1[One],Table1[Two])",
            "A5": "=MMULT(C2:D4,OFFSET(A2,0,0,5,2))",
            "A6": "=MMULT(C2:D4,A2#)",
            "A7": "=MMULT(C2:D4,[Inputs.xlsx]Data!A2:B6)",
            "A8": "=MMULT(C2:D4,A2:B3,#REF!)",
            "A9": "=MMULT(C2:D4,A2:B6*2)",
        },
    )

    report = lint_snapshot(snapshot)

    assert "FF095" not in {finding.rule_id for finding in report.findings}


def test_lint_mmult_dimension_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=MMULT(C2:D4,E2:F3)",
            "B3": "=MMULT(C3:D5,E3:F6)",
            "B4": "=MMULT(C4:D6,E4:F5)",
            "B5": "=MMULT(C5:D7,E5:F6)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF095", "high", ("Model", "B3"))]


def test_lint_mmult_dimension_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=MMULT(C2:D4,A2:B6)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_lookup_return_index_mismatches(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": (
                "=IFERROR(VLOOKUP(A2,'Input Sheet'!$C$2:$D$6,3,FALSE)"
                "+@HLOOKUP(B2,G2:H4,4,0),0)"
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF096", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "lookup_call_count": 2,
        "out_of_range_literal_index_count": 2,
        "evidence_scope": "lookup_direct_a1_table_literal_index",
    }


def test_lint_lookup_return_index_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=VLOOKUP(A2,C2:E6,3,FALSE)",
            "A3": "=HLOOKUP(A2,C2:E4,3,FALSE)",
            "A4": "=VLOOKUP(A2,NamedTable,4,FALSE)",
            "A5": "=HLOOKUP(A2,Table1[Key],4,FALSE)",
            "A6": "=VLOOKUP(A2,OFFSET(C2,0,0,5,3),4,FALSE)",
            "A7": "=HLOOKUP(A2,C2#,4,FALSE)",
            "A8": "=VLOOKUP(A2,[Inputs.xlsx]Data!C2:E6,4,FALSE)",
            "A9": "=HLOOKUP(A2,C2:E4,D2,FALSE)",
            "A10": "=VLOOKUP(A2,C2:E6,1+3,FALSE)",
            "A11": "=HLOOKUP(A2,C2:E4,4.0,FALSE)",
            "A12": "=VLOOKUP(A2,C2:E6,4,#REF!)",
        },
    )

    report = lint_snapshot(snapshot)

    assert "FF096" not in {finding.rule_id for finding in report.findings}


def test_lint_lookup_return_index_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=VLOOKUP(A2,C2:D6,2,FALSE)",
            "B3": "=VLOOKUP(A3,C3:D7,3,FALSE)",
            "B4": "=VLOOKUP(A4,C4:D8,2,FALSE)",
            "B5": "=VLOOKUP(A5,C5:D9,2,FALSE)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF096", "high", ("Model", "B3"))]


def test_lint_lookup_return_index_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=VLOOKUP(A2,C2:D6,3,FALSE)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_choose_literal_index_mismatches(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": '=IFERROR(CHOOSE(0,A2)+@CHOOSE(4,"one","two","three"),0)',
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF097", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "choose_call_count": 2,
        "out_of_range_literal_index_count": 2,
        "evidence_scope": "choose_literal_index_value_arity",
    }


def test_lint_choose_literal_index_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=CHOOSE(1,A2)",
            "A3": "=CHOOSE(2,A2,A3)",
            "A4": "=CHOOSE(B4,A2,A3)",
            "A5": "=CHOOSE(1+2,A2,A3)",
            "A6": "=CHOOSE(+3,A2,A3)",
            "A7": "=CHOOSE(-1,A2,A3)",
            "A8": "=CHOOSE(3.0,A2,A3)",
            "A9": "=CHOOSE({1,2},A2,A3)",
            "A10": "=CHOOSE(0,)",
            "A11": "=CHOOSE(3,A2,)",
            "A12": "=CHOOSE(0,A2,#REF!)",
            "A13": "=Vendor.CHOOSE(3,A2,A3)",
        },
    )

    report = lint_snapshot(snapshot)

    assert "FF097" not in {finding.rule_id for finding in report.findings}

    array_territory = _snapshot(tmp_path, {"B2": "=CHOOSE(3,C2,D2)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF097" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_choose_literal_index_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=CHOOSE(1,A2)",
            "B3": "=CHOOSE(3,A3)",
            "B4": "=CHOOSE(1,A4)",
            "B5": "=CHOOSE(1,A5)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF097", "high", ("Model", "B3"))]


def test_lint_choose_literal_index_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=CHOOSE(0,A2)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_randbetween_literal_bound_mismatches(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": "=IFERROR(RANDBETWEEN(2,1)+@RANDBETWEEN(-10,-11),0)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF098", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "randbetween_call_count": 2,
        "inverted_literal_bound_count": 2,
        "evidence_scope": "randbetween_direct_signed_integer_bounds",
    }


def test_lint_randbetween_literal_bound_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=RANDBETWEEN(1,2)",
            "A3": "=RANDBETWEEN(2,2)",
            "A4": "=RANDBETWEEN(-2,-1)",
            "A5": "=RANDBETWEEN(B5,1)",
            "A6": "=RANDBETWEEN(1+2,1)",
            "A7": "=RANDBETWEEN(2,1.0)",
            "A8": "=RANDBETWEEN({2},1)",
            "A9": "=RANDBETWEEN(2)",
            "A10": "=RANDBETWEEN(2,1,0)",
            "A11": "=RANDBETWEEN(2,#REF!)",
            "A12": "=Vendor.RANDBETWEEN(2,1)",
        },
    )

    report = lint_snapshot(snapshot)

    assert "FF098" not in {finding.rule_id for finding in report.findings}

    array_territory = _snapshot(tmp_path, {"B2": "=RANDBETWEEN(2,1)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF098" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_randbetween_literal_bound_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=RANDBETWEEN(1,2)",
            "B3": "=RANDBETWEEN(2,1)",
            "B4": "=RANDBETWEEN(1,2)",
            "B5": "=RANDBETWEEN(1,2)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF098", "high", ("Model", "B3"))]


def test_lint_randbetween_literal_bound_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=RANDBETWEEN(2,1)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_subtotal_literal_function_num_mismatches(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": "=IFERROR(SUBTOTAL(12,A2)+@SUBTOTAL(00112,A3),0)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF099", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "subtotal_call_count": 2,
        "unsupported_literal_function_num_count": 2,
        "evidence_scope": "subtotal_literal_function_num",
    }


def test_lint_subtotal_literal_function_num_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=SUBTOTAL(1,A2)",
            "A3": "=SUBTOTAL(11,A2,A3)",
            "A4": "=SUBTOTAL(101,A2)",
            "A5": "=SUBTOTAL(111,A2)",
            "A6": "=SUBTOTAL(B6,A2)",
            "A7": "=SUBTOTAL(1+11,A2)",
            "A8": "=SUBTOTAL(+12,A2)",
            "A9": "=SUBTOTAL(-1,A2)",
            "A10": "=SUBTOTAL(12.0,A2)",
            "A11": "=SUBTOTAL({12},A2)",
            "A12": "=SUBTOTAL(12)",
            "A13": "=SUBTOTAL(12,A2,)",
            "A14": "=SUBTOTAL(12,#REF!)",
            "A15": "=Vendor.SUBTOTAL(12,A2)",
        },
    )

    report = lint_snapshot(snapshot)

    assert "FF099" not in {finding.rule_id for finding in report.findings}

    array_territory = _snapshot(tmp_path, {"B2": "=SUBTOTAL(12,C2)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF099" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_subtotal_literal_function_num_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=SUBTOTAL(9,A2)",
            "B3": "=SUBTOTAL(12,A3)",
            "B4": "=SUBTOTAL(9,A4)",
            "B5": "=SUBTOTAL(9,A5)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF099", "high", ("Model", "B3"))]


def test_lint_subtotal_literal_function_num_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=SUBTOTAL(12,A2)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_index_literal_position_mismatches(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": (
                "=IFERROR(INDEX('Input Sheet'!$C$2:$D$4,4)"
                "+@INDEX(G2:H4,1,3),0)"
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF100", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "index_call_count": 2,
        "out_of_range_literal_index_count": 2,
        "evidence_scope": "index_direct_a1_array_literal_indices",
    }


def test_lint_index_literal_position_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=INDEX(C2:D4,3,2)",
            "A3": "=INDEX(C2:D4,0)",
            "A4": "=INDEX(C2:D4,0,2)",
            "A5": "=INDEX(C2:D4,3,0)",
            "A6": "=INDEX(C2:D4,E6)",
            "A7": "=INDEX(C2:D4,1+3)",
            "A8": "=INDEX(C2:D4,+4)",
            "A9": "=INDEX(C2:D4,-1)",
            "A10": "=INDEX(C2:D4,4.0)",
            "A11": "=INDEX({1,2;3,4},3)",
            "A12": "=INDEX(C2#,4)",
            "A13": "=INDEX(C2:D4,4,)",
            "A14": "=INDEX(C2:D4,,3)",
            "A15": "=INDEX(C2:D4,4,3,1)",
            "A16": "=INDEX(C2:D4,4,#REF!)",
            "A17": "=Vendor.INDEX(C2:D4,4)",
        },
    )

    report = lint_snapshot(snapshot)

    assert "FF100" not in {finding.rule_id for finding in report.findings}

    array_territory = _snapshot(tmp_path, {"B2": "=INDEX(C2:D4,4)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF100" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_index_literal_position_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=INDEX(C2:D4,1,2)",
            "B3": "=INDEX(C3:D5,4,1)",
            "B4": "=INDEX(C4:D6,1,2)",
            "B5": "=INDEX(C5:D7,1,2)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF100", "high", ("Model", "B3"))]


def test_lint_index_literal_position_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=INDEX(C2:D4,4)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_approximate_lookup_unsorted_direct_numeric_vectors(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "C2": 1,
            "D2": 10,
            "C3": 3,
            "D3": 30,
            "C4": 2,
            "D4": 20,
            "E2": (
                "=IFERROR(VLOOKUP(A2,C2:D4,2)"
                "+@HLOOKUP(A2,F2:H3,2,TRUE),0)"
            ),
            "F2": 2,
            "G2": 1,
            "H2": 3,
            "F3": 20,
            "G3": 10,
            "H3": 30,
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF101", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "approximate_lookup_call_count": 2,
        "unsorted_direct_numeric_lookup_vector_count": 2,
        "evidence_scope": "approximate_lookup_direct_numeric_a1_vectors",
    }


def test_lint_approximate_lookup_sort_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=VLOOKUP(B2,C2:D4,2)",
            "A3": "=HLOOKUP(B3,F2:H3,2,TRUE)",
            "A4": "=VLOOKUP(B4,J2:K4,2,FALSE)",
            "A5": "=VLOOKUP(B5,J2:K4,2,1)",
            "A6": "=VLOOKUP(B6,N2:O4,2)",
            "A7": "=VLOOKUP(B7,P:P,2)",
            "A8": "=HLOOKUP(B8,2:4,2,TRUE)",
            "A9": "=VLOOKUP(B9,Table1[#All],2)",
            "A10": "=HLOOKUP(B10,NamedTable,2,TRUE)",
            "A11": "=VLOOKUP(B11,[Inputs.xlsx]Data!C2:D4,2)",
            "A12": "=Vendor.VLOOKUP(B12,C2:D4,2)",
            "A13": "=VLOOKUP(B13,C2:D4,2,#REF!)",
            "C2": 1,
            "D2": 10,
            "C3": 2,
            "D3": 20,
            "C4": 3,
            "D4": 30,
            "F2": 1,
            "G2": 2,
            "H2": 3,
            "F3": 10,
            "G3": 20,
            "H3": 30,
            "J2": 3,
            "K2": 30,
            "J3": 1,
            "K3": 10,
            "J4": 2,
            "K4": 20,
            "N2": "text",
            "O2": 10,
            "N3": "another",
            "O3": 20,
            "N4": "value",
            "O4": 30,
        },
    )

    assert "FF101" not in {finding.rule_id for finding in lint_snapshot(snapshot).findings}

    array_territory = _snapshot(
        tmp_path,
        {
            "B2": "=VLOOKUP(A2,C2:D4,2)",
            "C2": 1,
            "D2": 10,
            "C3": 3,
            "D3": 30,
            "C4": 2,
            "D4": 20,
        },
    )
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF101" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_approximate_lookup_sort_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=VLOOKUP(A2,C2:D4,2,FALSE)",
            "B3": "=VLOOKUP(A3,C3:D5,2)",
            "B4": "=VLOOKUP(A4,C4:D6,2,FALSE)",
            "B5": "=VLOOKUP(A5,C5:D7,2,FALSE)",
            "C2": 1,
            "D2": 10,
            "C3": 3,
            "D3": 30,
            "C4": 2,
            "D4": 20,
            "C5": 4,
            "D5": 40,
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF101", "high", ("Model", "B3"))]


def test_lint_approximate_lookup_sort_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=VLOOKUP(A2,C2:D4,2)",
            "C2": 1,
            "D2": 10,
            "C3": 3,
            "D3": 30,
            "C4": 2,
            "D4": 20,
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_modern_lookup_unsupported_literal_modes(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": (
                "=IFERROR(XLOOKUP(A2,B2:B4,C2:C4,,3,-3)"
                "+@XMATCH(D2,F2:F4,3,0),0)"
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF102", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "unsupported_literal_mode_count": 4,
        "evidence_scope": "xlookup_xmatch_literal_mode_codes",
    }


def test_lint_modern_lookup_mode_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=XLOOKUP(B2,C2:C4,D2:D4)",
            "A3": "=XLOOKUP(B3,C3:C5,D3:D5,,-1,2)",
            "A4": "=XMATCH(B4,C2:C4,,-2)",
            "A5": "=_xlfn.XMATCH(B5,C2:C4,0,1)",
            "A6": "=XLOOKUP(B6,C2:C4,D2:D4,,E6)",
            "A7": "=XMATCH(B7,C2:C4,1+3)",
            "A8": "=Vendor.XLOOKUP(B8,C2:C4,D2:D4,,3,-3)",
            "A9": "=XLOOKUP(B9,C2:C4,D2:D4,,3,#REF!)",
        },
    )

    assert "FF102" not in {finding.rule_id for finding in lint_snapshot(snapshot).findings}

    array_territory = _snapshot(
        tmp_path,
        {"B2": "=XLOOKUP(A2,C2:C4,D2:D4,,3)"},
    )
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF102" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_modern_lookup_mode_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=XLOOKUP(A2,C2:C4,D2:D4)",
            "B3": "=XLOOKUP(A3,C3:C5,D3:D5,,3)",
            "B4": "=XLOOKUP(A4,C4:C6,D4:D6)",
            "B5": "=XLOOKUP(A5,C5:C7,D5:D7)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF102", "high", ("Model", "B3"))]


def test_lint_modern_lookup_mode_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=XLOOKUP(A2,C2:C4,D2:D4,,3)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_large_small_invalid_literal_ranks(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": "=IFERROR(LARGE(A2:B4,0)+@SMALL(C2:D4,+7),0)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF103", "high", ("Model", "E2"))]
    assert report.findings[0].details == {
        "large_small_call_count": 2,
        "invalid_literal_rank_count": 2,
        "evidence_scope": "large_small_direct_a1_array_literal_rank",
    }


def test_lint_large_small_rank_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=LARGE(B2:C4,1)",
            "A3": "=SMALL(B2:C4,6)",
            "A4": "=LARGE(B2:C4,D4)",
            "A5": "=SMALL(B2:C4,1+6)",
            "A6": "=LARGE(B2:C4,7.0)",
            "A7": "=SMALL({1,2},3)",
            "A8": "=LARGE(NamedArray,7)",
            "A9": "=SMALL(Table1[One],7)",
            "A10": "=LARGE(B2#,7)",
            "A11": "=SMALL([Inputs.xlsx]Data!B2:C4,7)",
            "A12": "=LARGE(B2:C4,7,#REF!)",
            "A13": "=Vendor.SMALL(B2:C4,7)",
        },
    )

    assert "FF103" not in {finding.rule_id for finding in lint_snapshot(snapshot).findings}

    array_territory = _snapshot(tmp_path, {"B2": "=LARGE(C2:D4,7)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF103" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_large_small_rank_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=LARGE(C2:D4,1)",
            "B3": "=LARGE(C3:D5,7)",
            "B4": "=LARGE(C4:D6,1)",
            "B5": "=LARGE(C5:D7,1)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF103", "high", ("Model", "B3"))]


def test_lint_large_small_rank_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=LARGE(C2:D4,7)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_invalid_text_literal_arguments(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "F2": (
                '=IFERROR(LEFT(A2,-1)+@RIGHT(B2,-2)+MID(C2,0,-3)'
                '+FIND("x",D2,0)+SEARCH("x",E2,-1),0)'
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF104", "high", ("Model", "F2"))]
    assert report.findings[0].details == {
        "invalid_literal_argument_count": 6,
        "evidence_scope": "text_direct_signed_integer_position_count",
    }


def test_lint_text_literal_argument_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=LEFT(B2,0)",
            "A3": "=RIGHT(B2,-0)",
            "A4": "=MID(B2,1,0)",
            "A5": '=FIND("x",B2,1)',
            "A6": '=SEARCH("x",B2,+1)',
            "A7": "=LEFT(B2,C2)",
            "A8": "=RIGHT(B2,1-2)",
            "A9": "=MID(B2,1.0,2)",
            "A10": "=MID(B2,1,2.0)",
            "A11": '=FIND("x",B2,0.0)',
            "A12": '=SEARCH("x",B2,1+1)',
            "A13": "=LEFT(B2)",
            "A14": "=RIGHT(B2,-1,0)",
            "A15": "=MID(B2,0)",
            "A16": '=FIND("x",B2,)',
            "A17": '=SEARCH("x",B2,0,#REF!)',
            "A18": "=Vendor.LEFT(B2,-1)",
            "A19": "=_xlfn.RIGHT(B2,-1)",
        },
    )

    assert "FF104" not in {
        finding.rule_id for finding in lint_snapshot(snapshot).findings
    }

    array_territory = _snapshot(tmp_path, {"B2": "=LEFT(C2,-1)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF104" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_text_literal_argument_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=LEFT(C2,1)",
            "B3": "=LEFT(C3,-1)",
            "B4": "=LEFT(C4,1)",
            "B5": "=LEFT(C5,1)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF104", "high", ("Model", "B3"))]


def test_lint_text_literal_argument_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=LEFT(C2,-1)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_zero_divisors(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "F2": "=IFERROR(A2/0+B2/-0+C2/0*1,0)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF105", "high", ("Model", "F2"))]
    assert report.findings[0].details == {
        "direct_zero_divisor_count": 3,
        "evidence_scope": "infix_division_direct_signed_integer_zero",
    }


def test_lint_direct_zero_divisor_rule_skips_ambiguous_forms(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=B2/1",
            "A3": "=B2/0.0",
            "A4": "=B2/0E0",
            "A5": "=B2/(0)",
            "A6": "=B2/0^1",
            "A7": "=B2/0%",
            "A8": "=B2/--0",
            "A9": "=B2/+(-0)",
            "A10": "=IFERROR(B2/#REF!,0)",
        },
    )

    assert "FF105" not in {
        finding.rule_id for finding in lint_snapshot(snapshot).findings
    }

    array_territory = _snapshot(tmp_path, {"B2": "=C2/0"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF105" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_direct_zero_divisor_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=C2/1",
            "B3": "=C3/0",
            "B4": "=C4/1",
            "B5": "=C5/1",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF105", "high", ("Model", "B3"))]


def test_lint_direct_zero_divisor_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=C2/0",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_sum_argument_range_overlaps(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "F2": (
                "=SUM(C2:D4,D4:E6)+SUM(F2:F4,F3:F5)+SUM(G2:G4,H2:H4)"
                "+SUM(I2:I4,I3:I5,I4:I6)"
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF110", "high", ("Model", "F2"))]
    assert report.findings[0].details == {
        "direct_sum_call_count": 3,
        "overlapping_direct_range_pair_count": 5,
        "evidence_scope": "sum_direct_a1_range_overlap",
    }
    single_finding_report = lint_snapshot(
        snapshot,
        max_formula_pattern_findings=1,
    )
    assert [finding.rule_id for finding in single_finding_report.findings] == ["FF110"]


def test_lint_direct_sum_overlap_rule_skips_ambiguous_or_nonoverlapping_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=SUM(B2:B4,C2:C4)",
            "A3": "=SUM('Inputs'!B2:B4,B2:B4)",
            "A4": "=SUM(B2:B4,1)",
            "A5": "=SUM(B:B,B:B)",
            "A6": "=SUM(NamedRange,B2:B4)",
            "A7": "=SUM(Table1[Amount],B2:B4)",
            "A8": "=SUM([Inputs.xlsx]Data!B2:B4,B2:B4)",
            "A9": "=Vendor.SUM(B2:B4,B2:B4)",
            "A10": "=SUM(B2:B4,#REF!)",
            "A11": "=SUM(B2:B4,B3:B5,)",
            "A12": "=SUM(B2:B4+0,B3:B5)",
            "A13": "=SUM(Ghost!B2:B4,Ghost!B3:B5)",
            "A14": "=SUM(B2:B4,B3:B5,Ghost!B2:B4)",
        },
        additional_sheets={"Inputs": {"B2": 1, "B3": 2, "B4": 3}},
    )

    assert "FF110" not in {finding.rule_id for finding in lint_snapshot(snapshot).findings}

    array_territory = _snapshot(tmp_path, {"B2": "=SUM(C2:C4,C4:C6)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF110" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_direct_sum_overlap_resolves_sheet_names_case_insensitively(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"B2": "=SUM(inputs!C2:C4,inputs!C4:C6)"},
        additional_sheets={"Inputs": {"C2": 1, "C3": 2, "C4": 3, "C5": 4, "C6": 5}},
    )

    report = lint_snapshot(snapshot)

    assert [finding.rule_id for finding in report.findings] == ["FF110"]
    assert report.findings[0].details == {
        "direct_sum_call_count": 1,
        "overlapping_direct_range_pair_count": 1,
        "evidence_scope": "sum_direct_a1_range_overlap",
    }


def test_lint_direct_sum_overlap_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=SUM(C2:C4,D2:D4)",
            "B3": "=SUM(C3:C5,C5:C7)",
            "B4": "=SUM(C4:C6,D4:D6)",
            "B5": "=SUM(C5:C7,D5:D7)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF110", "high", ("Model", "B3"))]


def test_lint_direct_sum_overlap_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=SUM(C2:C4,C4:C6)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_aggregate_literal_argument_errors(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "F2": (
                "=IFERROR(AGGREGATE(0,8,A2)+_xlfn.AGGREGATE(14,6,A3)"
                "+AGGREGATE(20,6,A4),0)"
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF111", "high", ("Model", "F2"))]
    assert report.findings[0].details == {
        "aggregate_call_count": 3,
        "unsupported_literal_function_num_count": 2,
        "unsupported_literal_option_count": 1,
        "missing_required_ref2_count": 1,
        "evidence_scope": "aggregate_direct_literal_codes_and_ref2_arity",
    }
    single_finding_report = lint_snapshot(
        snapshot,
        max_formula_pattern_findings=1,
    )
    assert [finding.rule_id for finding in single_finding_report.findings] == ["FF111"]


def test_lint_aggregate_literal_argument_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=AGGREGATE(1,0,B2)",
            "A3": "=_xlfn.AGGREGATE(19,7,B3,1)",
            "A4": "=AGGREGATE(B4,6,C4)",
            "A5": "=AGGREGATE(1,B5,C5)",
            "A6": "=AGGREGATE(1+19,6,B6)",
            "A7": "=AGGREGATE(1,6+2,B7)",
            "A8": "=AGGREGATE(1,,B8)",
            "A9": "=AGGREGATE(14,6,B9,)",
            "A10": "=AGGREGATE(1,6,B10,#REF!)",
            "A11": "=Vendor.AGGREGATE(0,8,B11)",
        },
    )

    assert "FF111" not in {finding.rule_id for finding in lint_snapshot(snapshot).findings}

    array_territory = _snapshot(tmp_path, {"B2": "=AGGREGATE(0,8,C2)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF111" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_aggregate_literal_argument_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=AGGREGATE(9,6,C2)",
            "B3": "=AGGREGATE(0,6,C3)",
            "B4": "=AGGREGATE(9,6,C4)",
            "B5": "=AGGREGATE(9,6,C5)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF111", "high", ("Model", "B3"))]


def test_lint_aggregate_literal_argument_candidates_share_the_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=A2*2",
            "B4": "=A4*2",
            "B5": "=A5*2",
            "E2": "=AGGREGATE(0,6,C2)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_aggregate_literal_argument_candidates_share_the_structural_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": "=SUM(C2:C4,C4:C6)",
            "F2": "=AGGREGATE(0,6,C2)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_mod_direct_literal_zero_divisors(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "F2": "=IFERROR(MOD(A2,0)+@MOD(B2,-0)+MOD(C2,2),0)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF112", "high", ("Model", "F2"))]
    assert report.findings[0].details == {
        "mod_literal_zero_divisor_count": 2,
        "evidence_scope": "mod_direct_signed_integer_zero_divisor",
    }


def test_lint_mod_literal_zero_divisor_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=MOD(B2,1)",
            "A3": "=MOD(B3,C3)",
            "A4": "=MOD(B4,0.0)",
            "A5": "=MOD(B5,0E0)",
            "A6": "=MOD(B6,(0))",
            "A7": "=MOD(B7,0^1)",
            "A8": "=MOD(B8,0%)",
            "A9": "=MOD(B9,--0)",
            "A10": "=MOD(B10,0,1)",
            "A11": "=MOD(B11,#REF!)",
            "A12": "=Vendor.MOD(B12,0)",
            "A13": "=QUOTIENT(B13,0)",
        },
    )

    assert "FF112" not in {finding.rule_id for finding in lint_snapshot(snapshot).findings}

    array_territory = _snapshot(tmp_path, {"B2": "=MOD(C2,0)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF112" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_mod_literal_zero_divisor_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=MOD(C2,1)",
            "B3": "=MOD(C3,0)",
            "B4": "=MOD(C4,1)",
            "B5": "=MOD(C5,1)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF112", "high", ("Model", "B3"))]


def test_lint_mod_literal_zero_divisor_candidates_share_the_structural_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": "=SUM(C2:C4,C4:C6)",
            "F2": "=MOD(C2,0)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_direct_date_function_literal_code_errors(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "F2": (
                "=IFERROR(YEARFRAC(A2,B2,5)+@WEEKDAY(C2,0)+WEEKNUM(D2,99),0)"
            ),
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF113", "high", ("Model", "F2"))]
    assert report.findings[0].details == {
        "date_function_call_count": 3,
        "unsupported_yearfrac_basis_count": 1,
        "unsupported_weekday_return_type_count": 1,
        "unsupported_weeknum_return_type_count": 1,
        "evidence_scope": "date_function_direct_literal_codes",
    }


def test_lint_date_function_literal_code_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": "=YEARFRAC(B2,C2,4)",
            "A3": "=WEEKDAY(B3,17)",
            "A4": "=WEEKNUM(B4,21)",
            "A5": "=YEARFRAC(B5,C5,D5)",
            "A6": "=WEEKDAY(B6,C6)",
            "A7": "=WEEKNUM(B7,0.0)",
            "A8": "=YEARFRAC(B8,C8,2^3)",
            "A9": "=WEEKDAY(B9,0%)",
            "A10": "=WEEKNUM(B10,--1)",
            "A11": "=YEARFRAC(B11,C11,)",
            "A12": "=WEEKDAY(B12,#REF!)",
            "A13": "=Vendor.WEEKNUM(B13,0)",
        },
    )

    assert "FF113" not in {finding.rule_id for finding in lint_snapshot(snapshot).findings}

    array_territory = _snapshot(tmp_path, {"B2": "=WEEKDAY(C2,0)"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF113" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_date_function_literal_code_rule_replaces_generic_outlier(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=YEARFRAC(C2,D2,1)",
            "B3": "=YEARFRAC(C3,D3,5)",
            "B4": "=YEARFRAC(C4,D4,1)",
            "B5": "=YEARFRAC(C5,D5,1)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF113", "high", ("Model", "B3"))]


def test_lint_date_function_literal_code_candidates_share_the_structural_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": "=SUM(C2:C4,C4:C6)",
            "F2": "=WEEKNUM(C2,0)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_reports_closed_external_criteria_function_risk_once(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=COUNTBLANK('[Source.xlsx]Data'!$A$2:$A$9)",
            "B3": "=COUNTIF('[Source.xlsx]Data'!$B$2:$B$9,1)",
            "B4": "=COUNTIFS([1]Data!A1:A9,1,[1]Data!B1:B9,2)",
            "B5": "=SUMIF('[Source.xlsx]Data'!$C$2:$C$9,1,'[Source.xlsx]Data'!$D$2:$D$9)",
            "B6": "=SUMIFS('[Source.xlsx]Data'!$E$2:$E$9,'[Source.xlsx]Data'!$F$2:$F$9,1)",
        },
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF114", "medium", None)]
    assert report.findings[0].details == {
        "closed_external_criteria_function_call_count": 5,
        "affected_formula_cell_count": 5,
        "direct_external_a1_reference_argument_count": 8,
        "countblank_call_count": 1,
        "countif_call_count": 1,
        "countifs_call_count": 1,
        "sumif_call_count": 1,
        "sumifs_call_count": 1,
        "evidence_scope": "direct_external_a1_criteria_function_arguments",
    }


def test_lint_closed_external_criteria_function_rule_skips_ambiguous_forms(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=COUNTIF(C2:C9,1)",
            "B3": "=COUNTIF('[Source.xlsx]Data'!A2:A9)",
            "B4": "=COUNTIF(\"'[Source.xlsx]Data'!A2:A9\",1)",
            "B5": "=Vendor.COUNTIF('[Source.xlsx]Data'!A2:A9,1)",
            "B6": "=COUNTIF('[Source.xlsx]Data'!A2:A9,#REF!)",
        },
    )

    assert "FF114" not in {finding.rule_id for finding in lint_snapshot(snapshot).findings}

    array_territory = _snapshot(
        tmp_path,
        {"B2": "=COUNTIF('[Source.xlsx]Data'!A2:A9,1)"},
    )
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert "FF114" not in {
        finding.rule_id for finding in lint_snapshot(array_territory).findings
    }


def test_lint_closed_external_criteria_function_risk_honors_finding_cap(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "E2": "=SUM(C2:C4,C4:C6)",
            "F2": "=COUNTIF('[Source.xlsx]Data'!A2:A9,1)",
        },
    )

    with pytest.raises(
        FormulaFenceError,
        match="max_formula_pattern_findings=1",
    ):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_closed_external_criteria_function_risk_keeps_saved_value_error(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "closed-external-criteria-value-error.xlsx",
            error_formula="=COUNTIF('[Source.xlsx]Data'!A2:A9,1)",
            error_result="#VALUE!",
        )
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [
        ("FF114", "medium", None),
        ("FF109", "high", ("Report", "B5")),
    ]


def test_lint_reports_static_multi_cell_cycle_when_iteration_is_disabled(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"B2": "=C2+1", "C2": "=B2+1", "D2": "=B2+1"},
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [
        ("FF090", "high", ("Model", "B2")),
        ("FF090", "high", ("Model", "C2")),
    ]
    assert [finding.details for finding in report.findings] == [
        {
            "calculation_iteration_enabled": False,
            "reference_scope": "multi_cell_static",
            "cycle_member_count": 2,
        },
        {
            "calculation_iteration_enabled": False,
            "reference_scope": "multi_cell_static",
            "cycle_member_count": 2,
        },
    ]


def test_lint_reports_cross_sheet_static_multi_cell_cycle(tmp_path: Path) -> None:
    workbook = Workbook()
    model = workbook.active
    model.title = "Model"
    inputs = workbook.create_sheet("Inputs")
    model["B2"] = "=Inputs!B2+1"
    inputs["B2"] = "=Model!B2+1"
    path = tmp_path / "cross-sheet-cycle.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))

    assert [
        (finding.rule_id, finding.location) for finding in report.findings
    ] == [
        ("FF090", ("Inputs", "B2")),
        ("FF090", ("Model", "B2")),
    ]
    assert {finding.details["cycle_member_count"] for finding in report.findings} == {2}


def test_lint_handles_a_deep_static_multi_cell_cycle_iteratively(tmp_path: Path) -> None:
    cycle_size = 1_025
    cells = {
        f"A{row}": f"=A{row + 1}+1" if row < cycle_size else "=A1+1"
        for row in range(1, cycle_size + 1)
    }

    report = lint_snapshot(_snapshot(tmp_path, cells))

    assert len(report.findings) == cycle_size
    assert {finding.rule_id for finding in report.findings} == {"FF090"}
    assert {finding.details["cycle_member_count"] for finding in report.findings} == {
        cycle_size
    }


def test_lint_keeps_iterative_or_ambiguous_multi_cell_cycles_quiet(
    tmp_path: Path,
) -> None:
    iterative = _snapshot(
        tmp_path,
        {"B2": "=C2+1", "C2": "=B2+1"},
        calculation_iteration_enabled=True,
    )
    dynamic = _snapshot(
        tmp_path,
        {"A1": 1, "B2": "=C2+OFFSET(A1,0,0)", "C2": "=B2+1"},
    )
    range_expression = _snapshot(
        tmp_path,
        {"B2": "=SUM(C2:C3)", "C2": "=B2+1", "C3": 1},
    )
    array_territory = _snapshot(tmp_path, {"B2": "=C2+1", "C2": "=B2+1"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B2",
            ref="B2:C2",
            min_column=2,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )
    three_d = _snapshot(tmp_path, {"B2": "=C2+1", "C2": "=B2+1"})
    three_d.three_d_reference_tokens[("Model", "B2")] = ("Model:Other!C2",)
    spill = _snapshot(tmp_path, {"B2": "=C2+1", "C2": "=B2+1"})
    spill.spill_reference_tokens[("Model", "B2")] = ("C2#",)
    implicit_intersection = _snapshot(tmp_path, {"B2": "=C2+1", "C2": "=B2+1"})
    implicit_intersection.implicit_intersection_tokens[("Model", "B2")] = ("@C2",)
    tokenization_failure = _snapshot(tmp_path, {"B2": "=C2+1", "C2": "=B2+1"})
    tokenization_failure.tokenization_failure_cells.add(("Model", "B2"))

    assert lint_snapshot(iterative).findings == []
    assert lint_snapshot(dynamic).findings == []
    assert lint_snapshot(range_expression).findings == []
    assert lint_snapshot(array_territory).findings == []
    assert lint_snapshot(three_d).findings == []
    assert lint_snapshot(spill).findings == []
    assert lint_snapshot(implicit_intersection).findings == []
    assert lint_snapshot(tokenization_failure).findings == []


def test_lint_deduplicates_direct_self_reference_inside_multi_cell_cycle(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(tmp_path, {"B2": "=B2+C2", "C2": "=B2+1"})

    report = lint_snapshot(snapshot)
    assert [
        (finding.rule_id, finding.location) for finding in report.findings
    ] == [
        ("FF087", ("Model", "B2")),
        ("FF090", ("Model", "C2")),
    ]
    assert report.findings[1].details == {
        "calculation_iteration_enabled": False,
        "reference_scope": "multi_cell_static",
        "cycle_member_count": 2,
    }


def test_lint_keeps_iteration_and_dynamic_or_range_self_forms_quiet(
    tmp_path: Path,
) -> None:
    iterative = _snapshot(
        tmp_path,
        {"B2": "=B2+1"},
        calculation_iteration_enabled=True,
    )
    dynamic = _snapshot(tmp_path, {"A1": "=OFFSET(A1,1,0)", "A2": 10})
    spill = _snapshot(tmp_path, {"A1": "=A1#"})
    implicit_intersection = _snapshot(tmp_path, {"A1": "=@A1"})
    range_expression = _snapshot(tmp_path, {"A1": "=SUM(A1:A2)", "A2": 10})
    array_territory = _snapshot(tmp_path, {"A1": "=A1+1"})
    array_territory.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="A1",
            ref="A1",
            min_column=1,
            min_row=1,
            max_column=1,
            max_row=1,
        ),
    )

    assert lint_snapshot(iterative).findings == []
    assert lint_snapshot(dynamic).findings == []
    assert lint_snapshot(spill).findings == []
    assert lint_snapshot(implicit_intersection).findings == []
    assert lint_snapshot(range_expression).findings == []
    assert lint_snapshot(array_territory).findings == []


def test_lint_reports_an_explicit_broken_reference_operand(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=SUM(A2,#REF!,C2)", "C2": 20},
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF088", "critical", ("Model", "B2"))]
    assert report.findings[0].details == {}


def test_lint_keeps_ref_text_but_not_an_error_operand_quiet(tmp_path: Path) -> None:
    literal = _snapshot(tmp_path, {"B2": '="#REF!"'})

    assert lint_snapshot(literal).findings == []


def test_lint_reports_a_saved_broken_reference_formula_result(tmp_path: Path) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-reference-error.xlsx",
            error_formula="=SUM(Inputs!A1:A1)",
            error_result="#REF!",
        )
    )

    report = lint_snapshot(snapshot)

    assert snapshot.broken_references == set()
    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF089", "high", ("Report", "B5"))]
    assert report.findings[0].details == {}


def test_lint_reports_a_saved_divide_by_zero_formula_result(tmp_path: Path) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-divide-by-zero.xlsx",
            error_formula="=Inputs!A1/Inputs!A2",
            error_result="#DIV/0!",
        )
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF106", "high", ("Report", "B5"))]
    assert report.findings[0].details == {"evidence_scope": "saved_formula_result"}


def test_lint_reports_a_saved_numeric_error_formula_result(tmp_path: Path) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-numeric-error.xlsx",
            error_formula="=IRR(Inputs!A1:A2)",
            error_result="#NUM!",
        )
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF107", "high", ("Report", "B5"))]
    assert report.findings[0].details == {"evidence_scope": "saved_formula_result"}


def test_lint_reports_a_saved_name_error_formula_result(tmp_path: Path) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-name-error.xlsx",
            error_formula="=UnknownCustomName(Inputs!A1)",
            error_result="#NAME?",
        )
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF108", "high", ("Report", "B5"))]
    assert report.findings[0].details == {"evidence_scope": "saved_formula_result"}


def test_lint_reports_a_saved_value_error_formula_result(tmp_path: Path) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-value-error.xlsx",
            error_formula='=VALUE("not-a-number")',
            error_result="#VALUE!",
        )
    )

    report = lint_snapshot(snapshot)

    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in report.findings
    ] == [("FF109", "high", ("Report", "B5"))]
    assert report.findings[0].details == {"evidence_scope": "saved_formula_result"}


def test_lint_keeps_other_saved_errors_quiet_and_avoids_static_duplicates(
    tmp_path: Path,
) -> None:
    other_error = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "other.xlsx",
            error_formula="=NA()",
            error_result="#N/A",
        )
    )
    static_reference = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "static-reference.xlsx",
            error_formula="=#REF!",
            error_result="#REF!",
        )
    )
    direct_zero_divisor = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "direct-zero-divisor.xlsx",
            error_formula="=Inputs!A1/0",
            error_result="#DIV/0!",
        )
    )
    mod_literal_zero_divisor = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "mod-literal-zero-divisor.xlsx",
            error_formula="=MOD(Inputs!A1,-0)",
            error_result="#DIV/0!",
        )
    )
    inverted_randbetween = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "inverted-randbetween.xlsx",
            error_formula="=RANDBETWEEN(2,1)",
            error_result="#NUM!",
        )
    )
    invalid_large_rank = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "invalid-large-rank.xlsx",
            error_formula="=LARGE(Inputs!A1:A1,2)",
            error_result="#NUM!",
        )
    )
    invalid_date_function_code = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "invalid-date-function-code.xlsx",
            error_formula="=WEEKDAY(Inputs!A1,0)",
            error_result="#NUM!",
        )
    )
    conditional_aggregate_value_error = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "conditional-aggregate-value-error.xlsx",
            error_formula="=SUMIFS(Inputs!A1:A2,Inputs!A1:A3,1)",
            error_result="#VALUE!",
        )
    )

    assert lint_snapshot(other_error).findings == []
    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in lint_snapshot(static_reference).findings
    ] == [("FF088", "critical", ("Report", "B5"))]
    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in lint_snapshot(direct_zero_divisor).findings
    ] == [("FF105", "high", ("Report", "B5"))]
    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in lint_snapshot(mod_literal_zero_divisor).findings
    ] == [("FF112", "high", ("Report", "B5"))]
    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in lint_snapshot(inverted_randbetween).findings
    ] == [("FF098", "high", ("Report", "B5"))]
    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in lint_snapshot(invalid_large_rank).findings
    ] == [("FF103", "high", ("Report", "B5"))]
    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in lint_snapshot(invalid_date_function_code).findings
    ] == [("FF113", "high", ("Report", "B5"))]
    assert [
        (finding.rule_id, finding.severity, finding.location)
        for finding in lint_snapshot(conditional_aggregate_value_error).findings
    ] == [("FF093", "high", ("Report", "B5"))]


def test_lint_keeps_short_or_non_numeric_aggregate_gaps_quiet(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": 10,
            "B3": 20,
            "B4": 30,
            "B5": "=SUM(B2:B3)",
            "D2": 10,
            "D3": 20,
            "D4": "intentional marker",
            "D5": 30,
            "D6": "=SUM(D2:D3)",
            "F2": 10,
            "F3": 20,
            "F4": 30,
            "F5": 40,
            "F6": "=SUM(F2:F3)+0",
        },
    )

    assert lint_snapshot(snapshot).findings == []


def test_lint_skips_array_territory_and_bounds_aggregate_gap_inspection(
    tmp_path: Path,
) -> None:
    cells = {f"B{row}": row for row in range(2, 132)}
    cells["B132"] = "=MAX(B2:B4)"
    snapshot = _snapshot(tmp_path, cells)
    snapshot.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="B5",
            ref="B5:B6",
            min_column=2,
            min_row=5,
            max_column=2,
            max_row=6,
        ),
    )

    assert lint_snapshot(snapshot).findings == []
    snapshot.dynamic_array_formula_ranges = ()
    assert lint_snapshot(snapshot, max_aggregate_omission_gap_cells=64).findings == []
    with pytest.raises(FormulaFenceError, match="max_aggregate_omission_gap_cells"):
        lint_snapshot(snapshot, max_aggregate_omission_gap_cells=1)


def test_lint_deduplicates_crossing_copy_patterns_with_all_evidence(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": 1,
            "C2": "=B2",
            "D2": 3,
            "E2": 4,
            "B3": "=B2",
            "D3": "=D2",
            "E3": "=E2",
            "B4": 5,
            "C4": "=B4",
            "D4": 7,
            "E4": 8,
            "B5": 9,
            "C5": "=B5",
        },
    )

    report = lint_snapshot(snapshot)

    assert [(finding.rule_id, finding.location) for finding in report.findings] == [
        ("FF082", ("Model", "C3"))
    ]
    assert [
        evidence["orientation"]
        for evidence in report.findings[0].details["pattern_evidence"]
    ] == ["column", "row"]


def test_lint_never_uses_a_tokenization_failure_as_copy_evidence(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "D2": "=C2*2",
            "E2": "=D2*2",
        },
    )
    snapshot.tokenization_failure_cells.add(("Model", "D2"))

    assert lint_snapshot(snapshot).findings == []


def test_lint_never_treats_declared_array_output_as_an_ordinary_gap(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "D2": "=C2*2",
            "E2": "=D2*2",
        },
    )
    snapshot.dynamic_array_formula_ranges = (
        ArrayFormulaRange(
            sheet="Model",
            anchor="C2",
            ref="C2",
            min_column=3,
            min_row=2,
            max_column=3,
            max_row=2,
        ),
    )

    assert lint_snapshot(snapshot).findings == []


def test_lint_fails_closed_when_array_metadata_is_incomplete(tmp_path: Path) -> None:
    snapshot = _snapshot(tmp_path, {"A1": 1, "B1": "=A1*2"})
    snapshot.array_formula_metadata_complete = False

    with pytest.raises(FormulaFenceError, match="complete array-formula metadata"):
        lint_snapshot(snapshot)


def test_lint_bounds_distinct_pattern_targets(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "D2": "=C2*2",
            "E2": "=D2*2",
            "A4": 1,
            "B4": "=A4*2",
            "D4": "=C4*2",
            "E4": "=D4*2",
        },
    )

    with pytest.raises(FormulaFenceError, match="max_formula_pattern_findings=1"):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_shares_its_finding_cap_with_aggregate_omissions(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "D2": "=C2*2",
            "E2": "=D2*2",
            "G2": 10,
            "G3": 20,
            "G4": 30,
            "G5": 40,
            "G6": 50,
            "G7": "=SUM(G2:G4)",
        },
    )

    with pytest.raises(FormulaFenceError, match="max_formula_pattern_findings=1"):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_shares_its_finding_cap_with_direct_unlocked_formulas(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=A2*2", "C2": "=A2*3"},
        protected=True,
        unlocked_cells=("B2", "C2"),
    )

    with pytest.raises(FormulaFenceError, match="max_formula_pattern_findings=1"):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_shares_its_finding_cap_with_incomplete_manual_calculation(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=A2*2"},
        protected=True,
        unlocked_cells=("B2",),
        calculation_mode="manual",
        calculation_completed=False,
    )

    with pytest.raises(FormulaFenceError, match="max_formula_pattern_findings=1"):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_shares_its_finding_cap_with_direct_self_references(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"B2": "=B2+1", "C2": "=C2+1"},
    )

    with pytest.raises(FormulaFenceError, match="max_formula_pattern_findings=1"):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_shares_its_finding_cap_with_multi_cell_static_cycles(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=C2+1",
            "C2": "=B2+1",
            "D2": "=E2+1",
            "E2": "=D2+1",
        },
    )

    with pytest.raises(FormulaFenceError, match="max_formula_pattern_findings=1"):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_shares_its_finding_cap_with_error_checking_suppressions(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(make_ignored_error_model(tmp_path / "ignored-errors.xlsx"))
    snapshot.calculation_settings["calcMode"] = "manual"
    snapshot.calculation_settings["calcCompleted"] = False

    with pytest.raises(FormulaFenceError, match="max_formula_pattern_findings=1"):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_shares_its_finding_cap_with_broken_references(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"B2": "=#REF!", "C2": "=IFERROR(#REF!,0)"},
    )

    with pytest.raises(FormulaFenceError, match="max_formula_pattern_findings=1"):
        lint_snapshot(snapshot, max_formula_pattern_findings=1)


def test_lint_renderers_keep_formula_text_out_of_review_artifacts(tmp_path: Path) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "A2": 1,
            "B2": "=A2*2",
            "D2": "=C2*2",
            "E2": "=D2*2",
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=A2*2" not in rendered
        assert "FF082" in rendered
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!C2"
    assert result["properties"]["pattern_kind"] == "blank_gap"
    with pytest.raises(FormulaFenceError, match="max_report_bytes=1"):
        lint_to_markdown(report, max_bytes=1)


def test_lint_aggregate_renderers_keep_formula_text_out_of_review_artifacts(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": 10,
            "B3": 20,
            "B4": 30,
            "B5": 40,
            "B6": 50,
            "B7": "=SUM(B2:B4)",
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=SUM(B2:B4)" not in rendered
        assert "FF084" in rendered
    assert "## Aggregate range evidence" in rendered_markdown
    assert "`Model!B2:B4`" in rendered_markdown
    assert rendered_sarif["runs"][0]["results"][0]["ruleId"] == "FF084"
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF084",
            "name": "FF084",
            "shortDescription": {
                "text": "A simple numeric aggregate stops before adjacent numeric cells."
            },
        }
    ]


def test_lint_unlocked_formula_renderers_keep_formula_text_out_of_review_artifacts(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=A2*2"},
        protected=True,
        unlocked_cells=("B2",),
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=A2*2" not in rendered
        assert "FF085" in rendered
    assert "## Formula protection evidence" in rendered_markdown
    assert rendered_sarif["runs"][0]["results"][0]["ruleId"] == "FF085"
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF085",
            "name": "FF085",
            "shortDescription": {
                "text": "A formula cell is explicitly unlocked on a protected worksheet."
            },
        }
    ]


def test_lint_incomplete_manual_calculation_renderers_keep_formula_text_out(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {"A2": 10, "B2": "=A2*2"},
        calculation_mode="manual",
        calculation_completed=False,
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=A2*2" not in rendered
        assert "FF086" in rendered
    assert "## Calculation freshness evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert "locations" not in result
    assert result["properties"] == {
        "severity": "medium",
        "calculation_mode": "manual",
        "calculation_completed_before_save": False,
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF086",
            "name": "FF086",
            "shortDescription": {
                "text": "A formula workbook was saved with incomplete manual calculation."
            },
        }
    ]


def test_lint_direct_self_reference_renderers_keep_formula_text_out(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(tmp_path, {"B2": "=B2+1"})
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=B2+1" not in rendered
        assert "FF087" in rendered
    assert "## Static circular-reference evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "calculation_iteration_enabled": False,
        "reference_scope": "direct_static",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF087",
            "name": "FF087",
            "shortDescription": {
                "text": (
                    "A formula directly references its own cell while calculation "
                    "iteration is disabled."
                )
            },
        }
    ]


def test_lint_multi_cell_static_cycle_renderers_keep_formula_text_out(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(tmp_path, {"B2": "=C2+1", "C2": "=B2+1"})
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=C2+1" not in rendered
        assert "=B2+1" not in rendered
        assert "FF090" in rendered
    assert "## Static circular-reference evidence" in rendered_markdown
    assert "a static multi-cell dependency component has 2 formula cells" in rendered_markdown
    assert [
        result["locations"][0]["logicalLocations"][0]["name"]
        for result in rendered_sarif["runs"][0]["results"]
    ] == ["Model!B2", "Model!C2"]
    assert {
        result["properties"]["cycle_member_count"]
        for result in rendered_sarif["runs"][0]["results"]
    } == {2}
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF090",
            "name": "FF090",
            "shortDescription": {
                "text": (
                    "A formula participates in a static multi-cell circular reference "
                    "while calculation iteration is disabled."
                )
            },
        }
    ]


def test_lint_broken_reference_renderers_keep_formula_text_out(tmp_path: Path) -> None:
    snapshot = _snapshot(tmp_path, {"B2": "=IFERROR(#REF!,0)"})
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=IFERROR(#REF!,0)" not in rendered
        assert "FF088" in rendered
    assert "## Explicit broken-reference evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {"severity": "critical"}
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF088",
            "name": "FF088",
            "shortDescription": {
                "text": "Formula contains an explicit broken #REF! reference."
            },
        }
    ]


def test_lint_saved_broken_reference_renderers_keep_cache_data_out(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-reference-error.xlsx",
            error_formula="=SUM(Inputs!A1:A1)",
            error_result="#REF!",
        )
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=SUM(Inputs!A1:A1)" not in rendered
        assert "#REF!" not in rendered
        assert "FF089" in rendered
    assert "## Saved broken-reference evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Report!B5"
    assert result["properties"] == {"severity": "high"}
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF089",
            "name": "FF089",
            "shortDescription": {
                "text": "A formula's saved result is a broken-reference error."
            },
        }
    ]


def test_lint_saved_divide_by_zero_renderers_keep_cache_data_out(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-divide-by-zero.xlsx",
            error_formula="=Inputs!A1/Inputs!A2",
            error_result="#DIV/0!",
        )
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=Inputs!A1/Inputs!A2" not in rendered
        assert "#DIV/0!" not in rendered
        assert "FF106" in rendered
    assert "## Saved division-by-zero evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Report!B5"
    assert result["properties"] == {
        "severity": "high",
        "evidence_scope": "saved_formula_result",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF106",
            "name": "FF106",
            "shortDescription": {
                "text": "A formula's saved result is a division-by-zero error."
            },
        }
    ]


def test_lint_saved_numeric_error_renderers_keep_cache_data_out(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-numeric-error.xlsx",
            error_formula="=IRR(Inputs!A1:A2)",
            error_result="#NUM!",
        )
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=IRR(Inputs!A1:A2)" not in rendered
        assert "#NUM!" not in rendered
        assert "FF107" in rendered
    assert "## Saved numeric-error evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Report!B5"
    assert result["properties"] == {
        "severity": "high",
        "evidence_scope": "saved_formula_result",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF107",
            "name": "FF107",
            "shortDescription": {
                "text": "A formula's saved result is a numeric error."
            },
        }
    ]


def test_lint_saved_name_error_renderers_keep_cache_data_out(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-name-error.xlsx",
            error_formula="=UnknownCustomName(Inputs!A1)",
            error_result="#NAME?",
        )
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "=UnknownCustomName(Inputs!A1)" not in rendered
        assert "#NAME?" not in rendered
        assert "FF108" in rendered
    assert "## Saved name-error evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Report!B5"
    assert result["properties"] == {
        "severity": "high",
        "evidence_scope": "saved_formula_result",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF108",
            "name": "FF108",
            "shortDescription": {
                "text": "A formula's saved result is a name error."
            },
        }
    ]


def test_lint_saved_value_error_renderers_keep_cache_data_out(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_formula_cached_result_model(
            tmp_path / "saved-value-error.xlsx",
            error_formula='=VALUE("PRIVATE-VALUE-ERROR")',
            error_result="#VALUE!",
        )
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert '=VALUE("PRIVATE-VALUE-ERROR")' not in rendered
        assert "#VALUE!" not in rendered
        assert "FF109" in rendered
    assert "## Saved value-error evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Report!B5"
    assert result["properties"] == {
        "severity": "high",
        "evidence_scope": "saved_formula_result",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF109",
            "name": "FF109",
            "shortDescription": {
                "text": "A formula's saved result is a value error."
            },
        }
    ]


def test_lint_error_checking_suppression_renderers_keep_targets_private(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(make_ignored_error_model(tmp_path / "ignored-errors.xlsx"))
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Error Review!B2" not in rendered
        assert "B2 B3" not in rendered
        assert "C2:C3" not in rendered
        assert "=1/0" not in rendered
        assert "FF091" in rendered
    assert "## Excel error-checking suppression evidence" in rendered_markdown
    assert "11 suppressed warning rules across 5 target ranges" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert "locations" not in result
    assert result["properties"] == {
        "severity": "medium",
        "suppressed_warning_counts": {
            "evaluation_error": 2,
            "inconsistent_formula": 2,
            "formula_range_omission": 1,
            "unlocked_formula": 1,
            "empty_cell_reference": 1,
            "list_data_validation": 1,
            "calculated_column": 1,
            "number_stored_as_text": 1,
            "two_digit_text_year": 1,
        },
        "suppressed_warning_rule_count": 11,
        "suppressed_warning_target_range_count": 5,
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF091",
            "name": "FF091",
            "shortDescription": {
                "text": (
                    "Workbook suppresses Excel error-checking prompts; review warnings may "
                    "be hidden."
                )
            },
        }
    ]


def test_lint_table_calculated_column_renderers_keep_master_formula_private(
    tmp_path: Path,
) -> None:
    snapshot = load_snapshot(
        make_calculated_column_model(
            tmp_path / "private-table-exception.xlsx",
            exception="=A3+B3",
        )
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "PRIVATE_RESULT_TABLE" not in rendered
        assert "Private Result" not in rendered
        assert "=A2*B2" not in rendered
        assert "=A3+B3" not in rendered
        assert "FF092" in rendered
    assert "## Excel Table calculated-column evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == (
        "'Private Table Sheet'!C3"
    )
    assert result["properties"] == {
        "severity": "medium",
        "exception_kind": "formula_mismatch",
        "matching_adjacent_formula_peers": 2,
        "evidence_scope": "table_calculated_column",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF092",
            "name": "FF092",
            "shortDescription": {
                "text": (
                    "An interior Excel Table cell differs from its declared "
                    "calculated-column formula."
                )
            },
        }
    ]


def test_lint_conditional_aggregate_range_shape_renderers_keep_ranges_private(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": (
                "=SUMIFS('Private Inputs'!$C$2:$C$10,"
                "'Private Inputs'!$A$2:$A$12,$D$2)"
            ),
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$C$10" not in rendered
        assert "$A$2:$A$12" not in rendered
        assert "FF093" in rendered
    assert "## Conditional-aggregate range-shape evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "conditional_aggregate_call_count": 1,
        "mismatched_direct_range_argument_count": 1,
        "evidence_scope": "conditional_aggregate_direct_a1_ranges",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF093",
            "name": "FF093",
            "shortDescription": {
                "text": (
                    "A conditional aggregate uses direct static ranges with different "
                    "shapes."
                )
            },
        }
    ]


def test_lint_sumproduct_range_shape_renderers_keep_ranges_private(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": (
                "=SUMPRODUCT('Private Inputs'!$C$2:$C$10,"
                "'Private Inputs'!$A$2:$A$12)"
            ),
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$C$10" not in rendered
        assert "$A$2:$A$12" not in rendered
        assert "FF094" in rendered
    assert "## SUMPRODUCT range-shape evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "sumproduct_call_count": 1,
        "mismatched_direct_array_argument_count": 1,
        "evidence_scope": "sumproduct_direct_a1_ranges",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF094",
            "name": "FF094",
            "shortDescription": {
                "text": "A SUMPRODUCT call uses direct static ranges with different shapes."
            },
        }
    ]


def test_lint_mmult_dimension_renderers_keep_ranges_private(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": (
                "=MMULT('Private Inputs'!$C$2:$D$4,"
                "'Private Inputs'!$A$2:$B$6)"
            ),
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$D$4" not in rendered
        assert "$A$2:$B$6" not in rendered
        assert "FF095" in rendered
    assert "## MMULT matrix-dimension evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "mmult_call_count": 1,
        "incompatible_direct_matrix_pair_count": 1,
        "evidence_scope": "mmult_direct_a1_arrays",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF095",
            "name": "FF095",
            "shortDescription": {
                "text": (
                    "An MMULT call uses direct static arrays with incompatible "
                    "matrix dimensions."
                )
            },
        }
    ]


def test_lint_lookup_return_index_renderers_keep_ranges_private(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=VLOOKUP(A2,'Private Inputs'!$C$2:$D$6,3,FALSE)",
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$D$6" not in rendered
        assert "FF096" in rendered
    assert "## Lookup return-index evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "lookup_call_count": 1,
        "out_of_range_literal_index_count": 1,
        "evidence_scope": "lookup_direct_a1_table_literal_index",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF096",
            "name": "FF096",
            "shortDescription": {
                "text": (
                    "A VLOOKUP or HLOOKUP call uses a literal return index outside "
                    "its direct static table range."
                )
            },
        }
    ]


def test_lint_choose_literal_index_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=CHOOSE(3,'Private Inputs'!$C$2:$D$6,\"confidential\")",
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$D$6" not in rendered
        assert "confidential" not in rendered
        assert "FF097" in rendered
    assert "## CHOOSE literal-index evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "choose_call_count": 1,
        "out_of_range_literal_index_count": 1,
        "evidence_scope": "choose_literal_index_value_arity",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF097",
            "name": "FF097",
            "shortDescription": {
                "text": (
                    "A CHOOSE call uses a literal index outside its available "
                    "value arguments."
                )
            },
        }
    ]


def test_lint_randbetween_literal_bound_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=RANDBETWEEN(99,1)+N('Private Inputs confidential')",
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "confidential" not in rendered
        assert "RANDBETWEEN(99,1)" not in rendered
        assert "FF098" in rendered
    assert "## RANDBETWEEN literal-bound evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "randbetween_call_count": 1,
        "inverted_literal_bound_count": 1,
        "evidence_scope": "randbetween_direct_signed_integer_bounds",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF098",
            "name": "FF098",
            "shortDescription": {
                "text": (
                    "A RANDBETWEEN call uses direct literal bounds with the bottom "
                    "above the top."
                )
            },
        }
    ]


def test_lint_subtotal_literal_function_num_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=SUBTOTAL(12,'Private Inputs'!$C$2:$D$6)+N('confidential')",
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$D$6" not in rendered
        assert "confidential" not in rendered
        assert "SUBTOTAL(12" not in rendered
        assert "FF099" in rendered
    assert "## SUBTOTAL function-code evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "subtotal_call_count": 1,
        "unsupported_literal_function_num_count": 1,
        "evidence_scope": "subtotal_literal_function_num",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF099",
            "name": "FF099",
            "shortDescription": {
                "text": (
                    "A SUBTOTAL call uses a literal function number outside Excel's "
                    "supported codes."
                )
            },
        }
    ]


def test_lint_index_literal_position_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    snapshot = _snapshot(
        tmp_path,
        {
            "B2": "=INDEX('Private Inputs'!$C$2:$D$4,4)+N('confidential')",
        },
    )
    report = lint_snapshot(snapshot)

    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$D$4" not in rendered
        assert "confidential" not in rendered
        assert "INDEX('Private Inputs'" not in rendered
        assert "FF100" in rendered
    assert "## INDEX literal-position evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "index_call_count": 1,
        "out_of_range_literal_index_count": 1,
        "evidence_scope": "index_direct_a1_array_literal_indices",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF100",
            "name": "FF100",
            "shortDescription": {
                "text": (
                    "An INDEX call uses a literal row or column number outside "
                    "its direct static array."
                )
            },
        }
    ]


def test_lint_approximate_lookup_sort_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    for row, value in enumerate((1, 3, 2), start=2):
        inputs.cell(row=row, column=3, value=value)
        inputs.cell(row=row, column=4, value=value * 10)
    model = workbook.create_sheet("Model")
    model["B2"] = (
        "=VLOOKUP(A2,'Private Inputs'!$C$2:$D$4,2)+N(\"confidential\")"
    )
    path = tmp_path / "private-approximate-lookup.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$D$4" not in rendered
        assert "confidential" not in rendered
        assert "VLOOKUP(A2" not in rendered
        assert "FF101" in rendered
    assert "## Approximate lookup sort evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "approximate_lookup_call_count": 1,
        "unsorted_direct_numeric_lookup_vector_count": 1,
        "evidence_scope": "approximate_lookup_direct_numeric_a1_vectors",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF101",
            "name": "FF101",
            "shortDescription": {
                "text": (
                    "An approximate VLOOKUP or HLOOKUP call uses a direct static "
                    "numeric lookup vector that is not sorted ascending."
                )
            },
        }
    ]


def test_lint_modern_lookup_mode_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    model = workbook.create_sheet("Model")
    model["B2"] = (
        "=XLOOKUP(A2,'Private Inputs'!$C$2:$C$4,"
        "'Private Inputs'!$D$2:$D$4,,3)+N(\"confidential\")"
    )
    path = tmp_path / "private-modern-lookup-mode.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$C$4" not in rendered
        assert "confidential" not in rendered
        assert "XLOOKUP(A2" not in rendered
        assert "FF102" in rendered
    assert "## XLOOKUP/XMATCH mode-code evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "unsupported_literal_mode_count": 1,
        "evidence_scope": "xlookup_xmatch_literal_mode_codes",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF102",
            "name": "FF102",
            "shortDescription": {
                "text": (
                    "An XLOOKUP or XMATCH call uses a literal mode outside Excel's "
                    "supported codes."
                )
            },
        }
    ]


def test_lint_large_small_rank_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    model = workbook.create_sheet("Model")
    model["B2"] = "=LARGE('Private Inputs'!$C$2:$D$4,7)+N(\"confidential\")"
    path = tmp_path / "private-large-small-rank.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$D$4" not in rendered
        assert "confidential" not in rendered
        assert "LARGE('Private Inputs'" not in rendered
        assert "FF103" in rendered
    assert "## LARGE/SMALL literal-rank evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "large_small_call_count": 1,
        "invalid_literal_rank_count": 1,
        "evidence_scope": "large_small_direct_a1_array_literal_rank",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF103",
            "name": "FF103",
            "shortDescription": {
                "text": (
                    "A LARGE or SMALL call uses a literal rank that is nonpositive or "
                    "exceeds its direct static array capacity."
                )
            },
        }
    ]


def test_lint_text_literal_argument_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    model = workbook.create_sheet("Model")
    model["B2"] = "=LEFT('Private Inputs'!$C$2,-1)+N(\"confidential\")"
    path = tmp_path / "private-text-literal-argument.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2" not in rendered
        assert "confidential" not in rendered
        assert "LEFT('Private Inputs'" not in rendered
        assert "FF104" in rendered
    assert "## Text literal-argument evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "invalid_literal_argument_count": 1,
        "evidence_scope": "text_direct_signed_integer_position_count",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF104",
            "name": "FF104",
            "shortDescription": {
                "text": (
                    "A LEFT, RIGHT, MID, FIND, or SEARCH call uses an invalid direct "
                    "literal character position or count."
                )
            },
        }
    ]


def test_lint_direct_zero_divisor_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    model = workbook.create_sheet("Model")
    model["B2"] = "='Private Inputs'!$C$2/-0+N(\"confidential\")"
    path = tmp_path / "private-direct-zero-divisor.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2" not in rendered
        assert "confidential" not in rendered
        assert "'Private Inputs'!$C$2/-0" not in rendered
        assert "FF105" in rendered
    assert "## Direct zero-divisor evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "direct_zero_divisor_count": 1,
        "evidence_scope": "infix_division_direct_signed_integer_zero",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF105",
            "name": "FF105",
            "shortDescription": {
                "text": "A division expression uses a direct literal zero divisor."
            },
        }
    ]


def test_lint_direct_sum_overlap_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    model = workbook.create_sheet("Model")
    model["B2"] = (
        "=SUM('Private Inputs'!$C$2:$C$4,'Private Inputs'!$C$4:$C$5)"
        '+N("confidential")'
    )
    path = tmp_path / "private-direct-sum-overlap.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2:$C$4" not in rendered
        assert "confidential" not in rendered
        assert "SUM('Private Inputs'" not in rendered
        assert "FF110" in rendered
    assert "## Direct SUM overlap evidence" in rendered_markdown
    assert "1 overlapping direct static range pair across 1 SUM call" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "direct_sum_call_count": 1,
        "overlapping_direct_range_pair_count": 1,
        "evidence_scope": "sum_direct_a1_range_overlap",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF110",
            "name": "FF110",
            "shortDescription": {
                "text": (
                    "A SUM call uses direct static ranges that overlap, so at least "
                    "one cell is included more than once."
                )
            },
        }
    ]


def test_lint_aggregate_literal_argument_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    model = workbook.create_sheet("Model")
    model["B2"] = '=AGGREGATE(0,8,\'Private Inputs\'!$C$2)+N("confidential")'
    path = tmp_path / "private-aggregate-literal-argument.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2" not in rendered
        assert "confidential" not in rendered
        assert "AGGREGATE(0,8" not in rendered
        assert "FF111" in rendered
    assert "## AGGREGATE literal-argument evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "aggregate_call_count": 1,
        "unsupported_literal_function_num_count": 1,
        "unsupported_literal_option_count": 1,
        "missing_required_ref2_count": 0,
        "evidence_scope": "aggregate_direct_literal_codes_and_ref2_arity",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF111",
            "name": "FF111",
            "shortDescription": {
                "text": (
                    "An AGGREGATE call uses an unsupported direct literal function "
                    "number or option, or omits a required second reference."
                )
            },
        }
    ]


def test_lint_mod_literal_zero_divisor_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    model = workbook.create_sheet("Model")
    model["B2"] = '=MOD(\'Private Inputs\'!$C$2,-0)+N("confidential")'
    path = tmp_path / "private-mod-literal-zero-divisor.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2" not in rendered
        assert "confidential" not in rendered
        assert "MOD('Private Inputs'" not in rendered
        assert "FF112" in rendered
    assert "## MOD direct zero-divisor evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "mod_literal_zero_divisor_count": 1,
        "evidence_scope": "mod_direct_signed_integer_zero_divisor",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF112",
            "name": "FF112",
            "shortDescription": {
                "text": "A MOD call uses a direct literal zero divisor."
            },
        }
    ]


def test_lint_date_function_literal_code_renderers_keep_values_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Private Inputs"
    model = workbook.create_sheet("Model")
    model["B2"] = '=WEEKDAY(\'Private Inputs\'!$C$2,0)+N("confidential")'
    path = tmp_path / "private-date-function-literal-code.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "Private Inputs" not in rendered
        assert "$C$2" not in rendered
        assert "confidential" not in rendered
        assert "WEEKDAY('Private Inputs'" not in rendered
        assert "FF113" in rendered
    assert "## Date-function literal-code evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert result["locations"][0]["logicalLocations"][0]["name"] == "Model!B2"
    assert result["properties"] == {
        "severity": "high",
        "date_function_call_count": 1,
        "unsupported_yearfrac_basis_count": 0,
        "unsupported_weekday_return_type_count": 1,
        "unsupported_weeknum_return_type_count": 0,
        "evidence_scope": "date_function_direct_literal_codes",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF113",
            "name": "FF113",
            "shortDescription": {
                "text": (
                    "A YEARFRAC, WEEKDAY, or WEEKNUM call uses an unsupported "
                    "direct literal code."
                )
            },
        }
    ]


def test_lint_closed_external_criteria_function_renderers_keep_links_private(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    model = workbook.active
    model.title = "Model"
    model["B2"] = (
        "=COUNTIF('../confidential/[PrivateSource.xlsx]Private Inputs'!$C$2:$C$9,1)"
    )
    path = tmp_path / "private-closed-external-criteria-function.xlsx"
    workbook.save(path)

    report = lint_snapshot(load_snapshot(path))
    rendered_json = as_json(report.to_dict())
    rendered_markdown = lint_to_markdown(report)
    rendered_sarif = lint_to_sarif(report)

    for rendered in (rendered_json, rendered_markdown, str(rendered_sarif)):
        assert "PrivateSource" not in rendered
        assert "confidential" not in rendered
        assert "Private Inputs" not in rendered
        assert "$C$2" not in rendered
        assert "COUNTIF('../" not in rendered
        assert "FF114" in rendered
    assert "## Closed-external-workbook criteria-function evidence" in rendered_markdown
    result = rendered_sarif["runs"][0]["results"][0]
    assert "locations" not in result
    assert result["properties"] == {
        "severity": "medium",
        "closed_external_criteria_function_call_count": 1,
        "affected_formula_cell_count": 1,
        "direct_external_a1_reference_argument_count": 1,
        "countblank_call_count": 0,
        "countif_call_count": 1,
        "countifs_call_count": 0,
        "sumif_call_count": 0,
        "sumifs_call_count": 0,
        "evidence_scope": "direct_external_a1_criteria_function_arguments",
    }
    assert rendered_sarif["runs"][0]["tool"]["driver"]["rules"] == [
        {
            "id": "FF114",
            "name": "FF114",
            "shortDescription": {
                "text": (
                    "Criteria functions directly reference an external workbook; Excel "
                    "returns #VALUE! if its source workbook is closed."
                )
            },
        }
    ]
