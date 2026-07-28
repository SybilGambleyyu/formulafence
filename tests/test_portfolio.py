"""Portfolio comparison contract tests."""

from __future__ import annotations

import json
import shutil
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

import formulafence.portfolio as portfolio_module
from formulafence.cli import build_parser, main
from formulafence.diff import DEFAULT_MAX_CHANGE_ANALYSIS_STATES
from formulafence.models import ExternalWorkbookStructuredReference, FormulaFenceError
from formulafence.output import (
    DEFAULT_MAX_REPORT_BYTES,
    as_json,
    portfolio_to_html,
    portfolio_to_markdown,
    portfolio_to_sarif,
)
from formulafence.policy import parse_policy
from formulafence.portfolio import (
    DEFAULT_MAX_INVENTORY_ENTRIES,
    DEFAULT_MAX_PORTFOLIO_SNAPSHOT_CELLS,
    DEFAULT_MAX_PORTFOLIO_SOURCE_BYTES,
    PortfolioError,
    _canonical_external_table_references,
    _canonical_three_d_sheet_span,
    compare_portfolios,
    discover_workbooks,
)
from formulafence.workbook import (
    DEFAULT_MAX_DEPENDENCY_EDGES,
    DEFAULT_MAX_FORMULA_DEFINED_NAME_STATES,
    load_snapshot,
    profile_snapshot,
)

from .helpers import (
    duplicate_external_link_definition,
    duplicate_indexed_external_link_part_binding,
    make_indexed_external_workbook_a1_link_model,
    make_indexed_external_workbook_name_link_model,
    make_indexed_external_workbook_sheet_defined_name_link_model,
    make_model,
    rewrite,
)


def _copy_workbook(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, destination)


def _write_workbook(path: Path, sheet_name: str, cells: dict[str, object]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for coordinate, value in cells.items():
        worksheet[coordinate] = value
    workbook.save(path)
    return path


def _write_named_formula_fanout_workbook(path: Path) -> Path:
    """Create four retained local graph edges from a compact named formula."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    worksheet["A1"] = 1
    worksheet["A2"] = 2
    workbook.defined_names.add(
        DefinedName("Fanout", attr_text="=SUM(Model!$A$1,Model!$A$2)")
    )
    worksheet["B1"] = "=Fanout"
    worksheet["B2"] = "=Fanout"
    workbook.save(path)
    return path


def _write_formula_defined_name_action_chain_workbook(path: Path) -> Path:
    """Create four names whose action ledgers grow as 1 + 2 + 3 + 4."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model"
    for index in range(4):
        previous = f"+ActionName{index - 1:05d}" if index else ""
        workbook.defined_names.add(
            DefinedName(
                f"ActionName{index:05d}",
                attr_text='=HYPERLINK("https://example.invalid","x")' + previous,
            )
        )
    workbook.save(path)
    return path


def _write_table_source_workbook(path: Path) -> Path:
    """Create a small source table whose data cells can change independently."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for coordinate, value in {
        "B2": "Amount",
        "C2": "Rate",
        "D2": "Label",
        "B3": 10,
        "C3": 0.1,
        "D3": "North",
        "B4": 20,
        "C4": 0.2,
        "D4": "South",
        "B5": 30,
        "C5": 0.3,
        "D5": "West",
    }.items():
        worksheet[coordinate] = value
    worksheet.add_table(Table(displayName="Sales", ref="B2:D5"))
    workbook.save(path)
    return path


def _portfolio_pair(tmp_path: Path) -> tuple[Path, Path]:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    baseline.mkdir()
    candidate.mkdir()
    (baseline / "models").mkdir()
    source = make_model(baseline / "models" / "shared.xlsx")
    _copy_workbook(source, candidate / "models" / "shared.xlsx")
    return baseline, candidate


def _cross_workbook_portfolio_pair(tmp_path: Path) -> tuple[Path, Path]:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    _write_workbook(baseline / "inputs.xlsx", "Data", {"B2": 100})
    _write_workbook(
        baseline / "calculation.xlsx",
        "Model",
        {
            "D2": "=[inputs.xlsx]Data!B2",
            "E2": "=D2*2",
        },
    )
    _write_workbook(
        baseline / "reports" / "summary.xlsx",
        "Summary",
        {
            "D2": "='..\\[calculation.xlsx]Model'!E2",
            "E2": "=D2*2",
        },
    )
    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs.xlsx",
        lambda workbook: setattr(workbook["Data"]["B2"], "value", 200),
    )
    return baseline, candidate


def test_portfolio_matches_relative_paths_and_reports_membership_changes(tmp_path: Path) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)

    def replace_formula(workbook) -> None:
        workbook["Model"]["B2"] = 200

    rewrite(candidate / "models" / "shared.xlsx", replace_formula)
    make_model(baseline / "removed.xlsx")
    make_model(candidate / "added.xlsx")

    report = compare_portfolios(baseline, candidate)
    payload = report.to_dict()
    entries = {entry["path"]: entry for entry in payload["workbooks"]}

    assert payload["before"]["workbook_count"] == 2
    assert payload["after"]["workbook_count"] == 2
    assert entries["models/shared.xlsx"]["status"] == "changed"
    assert entries["added.xlsx"]["status"] == "added"
    assert entries["removed.xlsx"]["status"] == "removed"
    assert entries["models/shared.xlsx"]["before"]["path"] == "models/shared.xlsx"
    assert any(
        change["kind"] == "formula_to_value"
        for change in entries["models/shared.xlsx"]["changes"]
    )
    assert {finding["rule_id"] for finding in entries["added.xlsx"]["findings"]} == {
        "FF077"
    }
    assert {finding["rule_id"] for finding in entries["removed.xlsx"]["findings"]} == {
        "FF077"
    }
    assert str(tmp_path) not in json.dumps(payload)


def test_portfolio_policy_applies_per_workbook_and_blocks_membership(tmp_path: Path) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)

    def replace_formula(workbook) -> None:
        workbook["Model"]["B2"] = 200

    rewrite(candidate / "models" / "shared.xlsx", replace_formula)
    make_model(candidate / "added.xlsx")
    policy = parse_policy(
        {
            "version": 1,
            "rules": {
                "no_formula_to_value": True,
                "no_portfolio_membership_changes": True,
            },
        }
    )

    report = compare_portfolios(baseline, candidate, policy=policy)
    by_path = {entry.path: entry for entry in report.workbooks}
    assert {finding.rule_id for finding in by_path["models/shared.xlsx"].findings} >= {
        "FFP001"
    }
    assert {finding.rule_id for finding in by_path["added.xlsx"].findings} == {
        "FF077",
        "FFP077",
    }

    policy_path = tmp_path / "formulafence.yml"
    policy_path.write_text(
        "version: 1\nrules:\n  no_formula_to_value: true\n"
        "  no_portfolio_membership_changes: true\n",
        encoding="utf-8",
    )
    output = tmp_path / "portfolio.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--policy",
                str(policy_path),
                "--format",
                "json",
                "--output",
                str(output),
            ]
        )
        == 1
    )
    assert "FFP077" in output.read_text(encoding="utf-8")


def test_portfolio_reports_transitive_static_cross_workbook_impacts(tmp_path: Path) -> None:
    baseline, candidate = _cross_workbook_portfolio_pair(tmp_path)
    policy = parse_policy(
        {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
    )

    report = compare_portfolios(baseline, candidate, policy=policy)
    source = next(entry for entry in report.workbooks if entry.path == "inputs.xlsx")
    finding = next(finding for finding in source.findings if finding.rule_id == "FF079")

    assert finding.location == ("Data", "B2")
    assert finding.details["impacted_workbook_count"] == 2
    assert finding.details["impacted_formula_count"] == 4
    assert finding.details["sample_impacts"] == [
        {
            "workbook": "calculation.xlsx",
            "location": "Model!D2",
            "path": [
                {"workbook": "inputs.xlsx", "location": "Data!B2"},
                {"workbook": "calculation.xlsx", "location": "Model!D2"},
            ],
        },
        {
            "workbook": "calculation.xlsx",
            "location": "Model!E2",
            "path": [
                {"workbook": "inputs.xlsx", "location": "Data!B2"},
                {"workbook": "calculation.xlsx", "location": "Model!D2"},
                {"workbook": "calculation.xlsx", "location": "Model!E2"},
            ],
        },
        {
            "workbook": "reports/summary.xlsx",
            "location": "Summary!D2",
            "path": [
                {"workbook": "inputs.xlsx", "location": "Data!B2"},
                {"workbook": "calculation.xlsx", "location": "Model!D2"},
                {"workbook": "calculation.xlsx", "location": "Model!E2"},
                {"workbook": "reports/summary.xlsx", "location": "Summary!D2"},
            ],
        },
        {
            "workbook": "reports/summary.xlsx",
            "location": "Summary!E2",
            "path": [
                {"workbook": "inputs.xlsx", "location": "Data!B2"},
                {"workbook": "calculation.xlsx", "location": "Model!D2"},
                {"workbook": "calculation.xlsx", "location": "Model!E2"},
                {"workbook": "reports/summary.xlsx", "location": "Summary!D2"},
                {"workbook": "reports/summary.xlsx", "location": "Summary!E2"},
            ],
        },
    ]
    assert {finding.rule_id for finding in source.findings} >= {"FF079", "FFP079"}
    assert not report.incomplete

    payload = report.to_dict()
    source_payload = next(
        entry for entry in payload["workbooks"] if entry["path"] == "inputs.xlsx"
    )
    assert source_payload["summary"]["raw_finding_count"] == 1
    assert source_payload["summary"]["policy_finding_count"] == 1
    assert payload["summary"]["cross_workbook_impact_source_count"] == 1
    assert payload["summary"]["cross_workbook_impacted_formula_count"] == 4
    assert not payload["summary"]["cross_workbook_impact_incomplete"]
    markdown = portfolio_to_markdown(report)
    assert "FF079" in markdown
    assert "`inputs.xlsx` `Data!B2` → `calculation.xlsx` `Model!D2`" in markdown
    sarif_results = portfolio_to_sarif(report)["runs"][0]["results"]
    assert any(result["ruleId"] == "FFP079" for result in sarif_results)
    sarif_finding = next(result for result in sarif_results if result["ruleId"] == "FF079")
    assert sarif_finding["locations"][0]["physicalLocation"]["artifactLocation"]["uri"] == (
        "inputs.xlsx"
    )
    assert sarif_finding["locations"][0]["logicalLocations"][0]["name"] == "Data!B2"
    assert sarif_finding["properties"]["sample_impacts"][0]["workbook"] == (
        "calculation.xlsx"
    )


def test_portfolio_resolves_relative_ranges_case_insensitively_without_basename_guessing(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    _write_workbook(
        baseline / "inputs" / "shared.xlsx",
        "Data",
        {"B2": 10, "B3": 20, "B4": 30},
    )
    _write_workbook(
        baseline / "reports" / "summary.xlsx",
        "Summary",
        {"D2": "=SUM('..\\INPUTS\\[SHARED.XLSX]data'!B2:B4)"},
    )
    _write_workbook(baseline / "other" / "shared.xlsx", "Data", {"B3": 999})
    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "shared.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 21),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source = next(entry for entry in report.workbooks if entry.path == "inputs/shared.xlsx")
    finding = next(finding for finding in source.findings if finding.rule_id == "FF079")

    assert finding.details["impacted_workbook_count"] == 1
    assert finding.details["impacted_formula_count"] == 1
    assert finding.details["sample_impacts"][0]["workbook"] == "reports/summary.xlsx"
    assert finding.details["sample_impacts"][0]["location"] == "Summary!D2"


def test_portfolio_resolves_relative_external_workbook_defined_names_privately(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    source_path = _write_workbook(
        baseline / "inputs" / "source.xlsx",
        "Data",
        {"B2": 10, "B3": 20, "B4": 30},
    )
    source = load_workbook(source_path)
    source.defined_names.add(
        DefinedName("PrivateInputRange", attr_text="Data!$B$2:$B$4")
    )
    source.defined_names.add(
        DefinedName("PrivateInputAlias", attr_text="=PrivateInputRange")
    )
    source.defined_names.add(
        DefinedName(
            "DynamicInputRange",
            attr_text="=OFFSET(Data!$B$2,0,0,3,1)",
        )
    )
    source.save(source_path)
    consumer_path = _write_workbook(
        baseline / "reports" / "summary.xlsx",
        "Summary",
        {
            "D2": "=SUM('..\\inputs\\[SOURCE.XLSX]privateinputalias')",
            "D3": "=SUM('..\\inputs\\[SOURCE.XLSX]dynamicinputrange')",
            "D4": "=SUM(ExternalNameFormula)",
            "D5": "=SUM(ExternalNameDynamicFormula)",
            "D6": "=SUM(ExternalNameLambda(7))",
        },
    )
    consumer = load_workbook(consumer_path)
    consumer.defined_names.add(
        DefinedName(
            "ExternalNameFormula",
            attr_text="=SUM('..\\inputs\\[SOURCE.XLSX]privateinputalias')",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "ExternalNameDynamicFormula",
            attr_text="=OFFSET('..\\inputs\\[SOURCE.XLSX]privateinputalias',0,0,1,1)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "ExternalNameLambda",
            attr_text="=LAMBDA(value,SUM(value,ExternalNameFormula))",
        )
    )
    consumer.save(consumer_path)
    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 21),
    )

    policy = parse_policy(
        {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
    )
    report = compare_portfolios(baseline, candidate, policy=policy)
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding for finding in source_entry.findings if finding.rule_id == "FF079"
    )
    consumer_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/summary.xlsx"
    )
    assert consumer_entry.after is not None
    report_rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )
    profile_rendered = as_json(profile_snapshot(consumer_entry.after))
    rendered = (*report_rendered, profile_rendered)

    assert finding.details["impacted_workbook_count"] == 1
    assert finding.details["impacted_formula_count"] == 3
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    assert [impact["location"] for impact in finding.details["sample_impacts"]] == [
        "Summary!D2",
        "Summary!D4",
        "Summary!D6",
    ]
    assert all("PrivateInputRange" not in value for value in rendered)
    assert all("PrivateInputAlias" not in value for value in rendered)
    assert all("DynamicInputRange" not in value for value in rendered)
    assert all("ExternalNameFormula" not in value for value in report_rendered)
    assert all("ExternalNameDynamicFormula" not in value for value in report_rendered)
    assert all("ExternalNameLambda" not in value for value in report_rendered)


def test_portfolio_resolves_static_external_named_lambdas_privately(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    _write_workbook(
        baseline / "inputs" / "source.xlsx",
        "Data",
        {"B3": 20},
    )
    consumer_path = _write_workbook(
        baseline / "reports" / "summary.xlsx",
        "Summary",
        {
            "D2": "=SUM(ExternalLambda(7))",
            "E2": "=SUM(ExternalLambdaSecond(7))",
            "F2": "=SUM(ExternalLambdaDynamic(7))",
            "G2": "=SUM(ExternalLambdaRelative(7))",
            "H2": "=SUM(ExternalLambdaLocal(7))",
            "I2": "=SUM(ExternalLambdaShadowed(7))",
            "J2": "=SUM(ExternalFormulaUsesLambda)",
            "K2": "=SUM(ExternalLambda)",
            "L2": "=SUM(ExternalFormulaWithInternalInput)",
        },
    )
    consumer = load_workbook(consumer_path)
    consumer.create_sheet("Inputs")["B2"] = 5
    consumer.defined_names.add(
        DefinedName(
            "ExternalLambda",
            attr_text=(
                "=LAMBDA(value,SUM(value,Inputs!$B$2,"
                "'..\\inputs\\[SOURCE.XLSX]Data'!$B$3))"
            ),
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "ExternalLambdaSecond",
            attr_text="=LAMBDA(value,ExternalLambda(value))",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "ExternalFormulaUsesLambda",
            attr_text="=SUM(ExternalLambda(7))",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "ExternalFormulaWithInternalInput",
            attr_text="=SUM(ExternalLambda(7),Inputs!$B$2)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "ExternalLambdaDynamic",
            attr_text=(
                "=LAMBDA(value,SUM(value,Inputs!$B$2,"
                "OFFSET('..\\inputs\\[SOURCE.XLSX]Data'!$B$3,0,0,1,1)))"
            ),
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "ExternalLambdaRelative",
            attr_text="=LAMBDA(value,SUM(value,Inputs!$B$2,A1))",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "ExternalLambdaShadowed",
            attr_text=(
                "=LAMBDA(value,SUM(value,'..\\inputs\\[SOURCE.XLSX]Data'!$B$3))"
            ),
        )
    )
    consumer["Summary"].defined_names.add(
        DefinedName(
            "ExternalLambdaLocal",
            attr_text=(
                "=LAMBDA(value,SUM(value,'..\\inputs\\[SOURCE.XLSX]Data'!$B$3))"
            ),
            localSheetId=0,
        )
    )
    consumer["Summary"].defined_names.add(
        DefinedName(
            "ExternalLambdaShadowed",
            attr_text="=LAMBDA(value,value)",
            localSheetId=0,
        )
    )
    consumer.save(consumer_path)
    consumer_snapshot = load_snapshot(consumer_path)
    assert consumer_snapshot.direct_dependents(("Inputs", "B2")) == {
        ("Summary", "D2"),
        ("Summary", "E2"),
        ("Summary", "J2"),
        ("Summary", "L2"),
    }
    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 21),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding for finding in source_entry.findings if finding.rule_id == "FF079"
    )
    consumer_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/summary.xlsx"
    )
    assert consumer_entry.after is not None
    report_rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )
    profile_rendered = as_json(profile_snapshot(consumer_entry.after))

    assert finding.details["impacted_workbook_count"] == 1
    assert finding.details["impacted_formula_count"] == 4
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    assert [impact["location"] for impact in finding.details["sample_impacts"]] == [
        "Summary!D2",
        "Summary!E2",
        "Summary!J2",
        "Summary!L2",
    ]
    for private_value in (
        "ExternalLambda",
        "ExternalLambdaSecond",
        "ExternalFormulaUsesLambda",
        "ExternalFormulaWithInternalInput",
        "ExternalLambdaDynamic",
        "ExternalLambdaRelative",
        "ExternalLambdaLocal",
        "ExternalLambdaShadowed",
        "..\\inputs\\[SOURCE.XLSX]Data",
    ):
        assert all(private_value not in value for value in report_rendered)
    assert "ExternalLambda" in profile_rendered


def test_portfolio_resolves_static_external_tables_privately(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    source_path = _write_table_source_workbook(baseline / "inputs" / "source.xlsx")
    consumer_path = _write_workbook(
        baseline / "reports" / "summary.xlsx",
        "Summary",
        {
            "D2": "=SUM('..\\inputs\\source.xlsx'!Sales[Amount])",
            "E2": "=SUM('..\\inputs\\[source.xlsx]'!Sales[[#Data],[Amount]:[Rate]])",
            "F2": "=SUM(DirectExternalTable)",
            "G2": "=SUM(DirectExternalTableBridge)",
            "H2": "=SUM('..\\inputs\\source.xlsx'!Sales[#Headers])",
            "I2": "=SUM('..\\inputs\\source.xlsx'!Sales[@Amount])",
            "J2": "=SUM('..\\inputs\\source.xlsx'!MissingTable[Amount])",
            "K2": "=SUM('[source.xlsx]Data'!Sales[Amount])",
            "L2": "=SUM('C:\\Private\\source.xlsx'!Sales[Amount])",
            "M2": "=SUM(DirectExternalTableShadowed)",
            "N2": "=SUM(DirectExternalTableFormula)",
            "O2": "=SUM(DirectExternalTableDynamicFormula)",
            "P2": "=SUM(DirectExternalTableLocalFormula)",
            "Q2": "=SUM(DirectExternalTableRelativeFormula)",
            "R2": "=SUM(DirectExternalTableLambda(7))",
            "S2": "=SUM(DirectExternalTableMixedFormula)",
            "T2": "=SUM(DirectExternalTableMixedLambda(7))",
        },
    )
    consumer = load_workbook(consumer_path)
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalTable",
            attr_text="'..\\inputs\\source.xlsx'!Sales[Amount]",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalTableBridge",
            attr_text="=DirectExternalTable",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalTableFormula",
            attr_text="=SUM(DirectExternalTableBridge)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalTableDynamicFormula",
            attr_text="=OFFSET(DirectExternalTable,0,0,1,1)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalTableRelativeFormula",
            attr_text="=SUM(DirectExternalTable,A1)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalTableLambda",
            attr_text="=LAMBDA(value,SUM(value,DirectExternalTableBridge))",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalTableMixedFormula",
            attr_text=(
                "=SUM(DirectExternalTable,'..\\inputs\\[source.xlsx]Data'!$C$3)"
            ),
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalTableMixedLambda",
            attr_text=(
                "=LAMBDA(value,SUM(value,DirectExternalTable,"
                "'..\\inputs\\[source.xlsx]Data'!$C$3))"
            ),
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalTableShadowed",
            attr_text="'..\\inputs\\source.xlsx'!Sales[Amount]",
        )
    )
    consumer["Summary"].defined_names.add(
        DefinedName(
            "DirectExternalTableShadowed",
            attr_text="Summary!$A$1",
            localSheetId=0,
        )
    )
    consumer["Summary"].defined_names.add(
        DefinedName(
            "DirectExternalTableLocalFormula",
            attr_text="=SUM(DirectExternalTable)",
            localSheetId=0,
        )
    )
    consumer.save(consumer_path)

    source_snapshot = load_snapshot(source_path)
    table_reference = ExternalWorkbookStructuredReference(
        source_path="../inputs/source.xlsx",
        table_name="Sales",
        table_reference="Sales[Amount]",
    )
    assert [
        (
            reference.sheet,
            reference.min_column,
            reference.min_row,
            reference.max_column,
            reference.max_row,
        )
        for reference in _canonical_external_table_references(
            source_snapshot,
            table_reference,
        )
    ] == [("Data", 2, 3, 2, 5)]
    for selector, expected_bounds in {
        "Sales[#Headers]": [(2, 2, 4, 2)],
        "Sales[#Data]": [(2, 3, 4, 5)],
        "Sales[#All]": [(2, 2, 4, 5)],
        "Sales[[#Data],[Amount]:[Rate]]": [(2, 3, 3, 5)],
        "Sales[#Totals]": [],
    }.items():
        assert [
            (
                reference.min_column,
                reference.min_row,
                reference.max_column,
                reference.max_row,
            )
            for reference in _canonical_external_table_references(
                source_snapshot,
                replace(table_reference, table_reference=selector),
            )
        ] == expected_bounds
    sales = next(table for table in source_snapshot.tables.values() if table.name == "Sales")
    ambiguous_snapshot = replace(
        source_snapshot,
        tables={
            **source_snapshot.tables,
            "case-collision": replace(sales, name="sales"),
        },
    )
    assert _canonical_external_table_references(ambiguous_snapshot, table_reference) == ()

    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 11),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding for finding in source_entry.findings if finding.rule_id == "FF079"
    )
    consumer_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/summary.xlsx"
    )
    assert consumer_entry.after is not None
    report_rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )
    profile_rendered = as_json(profile_snapshot(consumer_entry.after))

    assert not report.incomplete
    assert finding.location == ("Data", "B3")
    assert finding.details["impacted_workbook_count"] == 1
    assert finding.details["impacted_formula_count"] == 8
    assert [impact["location"] for impact in finding.details["sample_impacts"]] == [
        "Summary!D2",
        "Summary!E2",
        "Summary!F2",
        "Summary!G2",
        "Summary!N2",
        "Summary!R2",
        "Summary!S2",
        "Summary!T2",
    ]
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    snapshot_repr = repr(consumer_entry.after.external_workbook_structured_references)
    for private_value in (
        "..\\inputs\\source.xlsx",
        "Sales[Amount]",
        "Sales[[#Data],[Amount]:[Rate]]",
        "Sales[@Amount]",
        "MissingTable[Amount]",
    ):
        assert private_value not in snapshot_repr
        assert all(private_value not in value for value in report_rendered)
        assert private_value not in profile_rendered
    for private_alias in (
        "DirectExternalTable",
        "DirectExternalTableBridge",
        "DirectExternalTableFormula",
        "DirectExternalTableDynamicFormula",
        "DirectExternalTableLocalFormula",
        "DirectExternalTableRelativeFormula",
        "DirectExternalTableLambda",
        "DirectExternalTableMixedFormula",
        "DirectExternalTableMixedLambda",
        "DirectExternalTableShadowed",
    ):
        assert all(private_alias not in value for value in report_rendered)


def test_portfolio_resolves_declared_package_indexed_external_tables_privately(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    _write_table_source_workbook(baseline / "inputs" / "source.xlsx")
    _write_workbook(baseline / "decoy.xlsx", "Data", {"B3": 999})
    make_indexed_external_workbook_name_link_model(
        baseline / "reports" / "summary.xlsx",
        target_paths=("../decoy.xlsx", "../inputs/source.xlsx"),
        link_index=2,
        external_reference="[2]!Sales[Amount]",
        consumer_alias_name="PackageExternalTable",
        consumer_formula_alias=True,
        consumer_static_formula_name="PackageExternalTableFormula",
        consumer_static_lambda_name="PackageExternalTableLambda",
    )
    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 11),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding for finding in source_entry.findings if finding.rule_id == "FF079"
    )
    consumer_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/summary.xlsx"
    )
    assert consumer_entry.after is not None
    report_rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )
    profile_rendered = as_json(profile_snapshot(consumer_entry.after))

    assert not report.incomplete
    assert finding.details["impacted_workbook_count"] == 1
    assert finding.details["impacted_formula_count"] == 4
    assert [impact["location"] for impact in finding.details["sample_impacts"]] == [
        "Model!D2",
        "Model!E2",
        "Model!F2",
        "Model!G2",
    ]
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    for private_value in ("Sales[Amount]", "../inputs/source.xlsx"):
        assert all(private_value not in value for value in report_rendered)
        assert private_value not in profile_rendered
    assert all("PackageExternalTable" not in value for value in report_rendered)
    assert "PackageExternalTable" in profile_rendered
    assert all("PackageExternalTableFormula" not in value for value in report_rendered)
    assert "PackageExternalTableFormula" in profile_rendered
    assert all("PackageExternalTableLambda" not in value for value in report_rendered)
    assert "PackageExternalTableLambda" in profile_rendered


def test_portfolio_resolves_declared_package_indexed_external_names_privately(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    source_path = _write_workbook(
        baseline / "inputs" / "source.xlsx",
        "Data",
        {"B2": 10, "B3": 20, "B4": 30},
    )
    source = load_workbook(source_path)
    source.defined_names.add(
        DefinedName("PrivatePackageSourceRange", attr_text="Data!$B$2:$B$4")
    )
    source.defined_names.add(
        DefinedName(
            "PrivatePackageSourceAlias",
            attr_text="=PrivatePackageSourceRange",
        )
    )
    source.save(source_path)
    _write_workbook(baseline / "decoy.xlsx", "Data", {"B3": 999})
    make_indexed_external_workbook_name_link_model(
        baseline / "reports" / "summary.xlsx",
        target_paths=("../decoy.xlsx", "../inputs/source.xlsx"),
        source_name="PrivatePackageSourceAlias",
        link_index=2,
        consumer_static_lambda_name="PackageExternalInputLambda",
    )
    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 21),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding for finding in source_entry.findings if finding.rule_id == "FF079"
    )
    consumer_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/summary.xlsx"
    )
    assert consumer_entry.after is not None
    report_rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )
    profile_rendered = as_json(profile_snapshot(consumer_entry.after))
    rendered = (*report_rendered, profile_rendered)

    assert not report.incomplete
    assert finding.details["impacted_workbook_count"] == 1
    assert finding.details["impacted_formula_count"] == 3
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    assert [impact["location"] for impact in finding.details["sample_impacts"]] == [
        "Model!D2",
        "Model!E2",
        "Model!G2",
    ]
    assert all("PrivatePackageSourceRange" not in value for value in rendered)
    assert all("PrivatePackageSourceAlias" not in value for value in rendered)
    assert all("PackageExternalInput" not in value for value in report_rendered)
    assert "PackageExternalInput" in profile_rendered
    assert all(
        "PackageExternalInputLambda" not in value for value in report_rendered
    )
    assert "PackageExternalInputLambda" in profile_rendered


def test_portfolio_resolves_declared_package_indexed_external_a1_links_privately(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    _write_workbook(
        baseline / "inputs" / "source.xlsx",
        "Data",
        {"B2": 10, "B3": 20, "B4": 30},
    )
    _write_workbook(baseline / "decoy.xlsx", "Data", {"B3": 999})
    make_indexed_external_workbook_a1_link_model(
        baseline / "reports" / "summary.xlsx",
        target_paths=("../decoy.xlsx", "../inputs/source.xlsx"),
        link_index=2,
        source_sheet="Data",
        source_range="$B$2:$B$4",
    )
    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 21),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding for finding in source_entry.findings if finding.rule_id == "FF079"
    )
    consumer_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/summary.xlsx"
    )
    assert consumer_entry.after is not None
    report_rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )
    profile_rendered = as_json(profile_snapshot(consumer_entry.after))

    assert not report.incomplete
    assert finding.details["impacted_workbook_count"] == 1
    assert finding.details["impacted_formula_count"] == 2
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    assert [impact["location"] for impact in finding.details["sample_impacts"]] == [
        "Model!D2",
        "Model!E2",
    ]
    assert all("PackageExternalCell" not in value for value in report_rendered)
    assert "PackageExternalCell" in profile_rendered


def test_portfolio_resolves_package_indexed_external_alias_chains_privately(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    source_path = _write_workbook(
        baseline / "inputs" / "source.xlsx",
        "Data",
        {"B2": 10, "B3": 20, "B4": 30},
    )
    source = load_workbook(source_path)
    source.defined_names.add(
        DefinedName("PrivatePackageChainGlobal", attr_text="Data!$B$2:$B$4")
    )
    source["Data"].defined_names.add(
        DefinedName(
            "PrivatePackageChainLocal",
            attr_text="Data!$B$2:$B$4",
            localSheetId=0,
        )
    )
    source.save(source_path)
    make_indexed_external_workbook_name_link_model(
        baseline / "reports" / "name-chain.xlsx",
        source_name="PrivatePackageChainGlobal",
        include_direct_indexed_formula=False,
        consumer_formula_alias=True,
    )
    make_indexed_external_workbook_a1_link_model(
        baseline / "reports" / "a1-chain.xlsx",
        include_direct_indexed_formula=False,
        consumer_formula_alias=True,
    )
    make_indexed_external_workbook_sheet_defined_name_link_model(
        baseline / "reports" / "local-name-chain.xlsx",
        source_name="PrivatePackageChainLocal",
        include_direct_indexed_formula=False,
        consumer_formula_alias=True,
    )

    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 21),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding for finding in source_entry.findings if finding.rule_id == "FF079"
    )
    report_rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )

    assert not report.incomplete
    assert finding.details["impacted_workbook_count"] == 3
    assert finding.details["impacted_formula_count"] == 3
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    assert {
        (impact["workbook"], impact["location"])
        for impact in finding.details["sample_impacts"]
    } == {
        ("reports/a1-chain.xlsx", "Model!D2"),
        ("reports/local-name-chain.xlsx", "Model!D2"),
        ("reports/name-chain.xlsx", "Model!D2"),
    }
    for private_value in (
        "PrivatePackageChainGlobal",
        "PrivatePackageChainLocal",
        "PackageExternalFormulaAlias",
        "PackageExternalInput",
        "PackageExternalCell",
        "PackageExternalSheetName",
    ):
        assert all(private_value not in value for value in report_rendered)


def test_portfolio_resolves_static_external_sheet_local_names_privately(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    source_path = _write_workbook(
        baseline / "inputs" / "source.xlsx",
        "Data",
        {"B2": 10, "B3": 20, "B4": 30},
    )
    source = load_workbook(source_path)
    other = source.create_sheet("Other")
    other["B2"] = 999
    source["Data"].defined_names.add(
        DefinedName(
            "PrivateLocalSource",
            attr_text="Data!$B$2:$B$4",
            localSheetId=0,
        )
    )
    source["Data"].defined_names.add(
        DefinedName(
            "PrivateLocalAlias",
            attr_text="=PrivateLocalSource",
            localSheetId=0,
        )
    )
    source["Data"].defined_names.add(
        DefinedName(
            "PrivateDynamicLocal",
            attr_text="=OFFSET(Data!$B$2,0,0,3,1)",
            localSheetId=0,
        )
    )
    source["Other"].defined_names.add(
        DefinedName(
            "PrivateWrongScope",
            attr_text="Other!$B$2",
            localSheetId=1,
        )
    )
    source.defined_names.add(
        DefinedName("PrivateGlobalOnly", attr_text="Data!$B$2:$B$4")
    )
    source.save(source_path)

    direct_path = _write_workbook(
        baseline / "reports" / "direct.xlsx",
        "Summary",
        {
            "D2": "=SUM('..\\inputs\\[SOURCE.XLSX]data'!privatelocalalias)",
            "D3": "=SUM(DirectLocalAlias)",
            "D4": "=SUM('..\\inputs\\[SOURCE.XLSX]Data'!PrivateDynamicLocal)",
            "D5": "=SUM('..\\inputs\\[SOURCE.XLSX]Data'!PrivateGlobalOnly)",
            "D6": "=SUM('..\\inputs\\[SOURCE.XLSX]Data'!PrivateWrongScope)",
            "D7": "=SUM(DirectLocalFormulaAlias)",
        },
    )
    direct = load_workbook(direct_path)
    direct.defined_names.add(
        DefinedName(
            "DirectLocalAlias",
            attr_text="'..\\inputs\\[SOURCE.XLSX]Data'!PrivateLocalAlias",
        )
    )
    direct.defined_names.add(
        DefinedName(
            "DirectLocalFormulaAlias",
            attr_text="=DirectLocalAlias",
        )
    )
    direct.save(direct_path)
    make_indexed_external_workbook_sheet_defined_name_link_model(
        baseline / "reports" / "package.xlsx",
        source_name="PrivateLocalAlias",
    )

    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 21),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding for finding in source_entry.findings if finding.rule_id == "FF079"
    )
    report_rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )

    assert not report.incomplete
    assert finding.details["impacted_workbook_count"] == 2
    assert finding.details["impacted_formula_count"] == 5
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    assert {
        (impact["workbook"], impact["location"])
        for impact in finding.details["sample_impacts"]
    } == {
        ("reports/direct.xlsx", "Summary!D2"),
        ("reports/direct.xlsx", "Summary!D3"),
        ("reports/direct.xlsx", "Summary!D7"),
        ("reports/package.xlsx", "Model!D2"),
        ("reports/package.xlsx", "Model!E2"),
    }
    assert all("PrivateLocalSource" not in value for value in report_rendered)
    assert all("PrivateLocalAlias" not in value for value in report_rendered)
    assert all("PrivateDynamicLocal" not in value for value in report_rendered)
    assert all("PrivateGlobalOnly" not in value for value in report_rendered)
    assert all("PrivateWrongScope" not in value for value in report_rendered)
    assert all("DirectLocalFormulaAlias" not in value for value in report_rendered)


def test_portfolio_resolves_static_direct_external_alias_chains_privately(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    source_path = _write_workbook(
        baseline / "inputs" / "source.xlsx",
        "Data",
        {"B2": 10, "B3": 20, "B4": 30},
    )
    source = load_workbook(source_path)
    source.defined_names.add(
        DefinedName("PrivateDirectSourceRange", attr_text="Data!$B$2:$B$4")
    )
    source.defined_names.add(
        DefinedName(
            "PrivateDirectSourceAlias",
            attr_text="=PrivateDirectSourceRange",
        )
    )
    source.save(source_path)

    consumer_path = _write_workbook(
        baseline / "reports" / "summary.xlsx",
        "Summary",
        {
            "D2": "=SUM(DirectExternalCell)",
            "E2": "=SUM(DirectExternalName)",
            "F2": "=SUM(DirectExternalLeadingEqualsCell)",
            "G2": "=SUM(DirectExternalFormulaAlias)",
            "H2": "=SUM(DirectExternalLocalAlias)",
            "I2": "=SUM(DirectExternalAbsoluteAlias)",
            "J2": "=SUM(DirectExternalMalformedAlias)",
            "K2": "=SUM(DirectExternalShadowed)",
            "L2": "=SUM(DirectExternalChainedCellAlias)",
            "M2": "=SUM(DirectExternalChainedNameAlias)",
            "N2": "=SUM(DirectExternalFormulaWrapper)",
            "O2": "=SUM(DirectExternalAliasCycleA)",
            "P2": "=SUM(DirectExternalUnresolvedAlias)",
            "Q2": "=SUM(DirectExternalChainedShadowed)",
            "R2": "=SUM(DirectExternalLocalFormulaAlias)",
            "S2": "=SUM(DirectExternalFormulaWrapperSecond)",
            "T2": "=SUM(DirectExternalUnresolvedFormula)",
            "U2": "=SUM(DirectExternalFormulaDuplicate)",
        },
    )
    consumer = load_workbook(consumer_path)
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalCell",
            attr_text="'..\\inputs\\[SOURCE.XLSX]Data'!$B$2:$B$4",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalName",
            attr_text="'..\\inputs\\[SOURCE.XLSX]PrivateDirectSourceAlias'",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalLeadingEqualsCell",
            attr_text="='..\\inputs\\[SOURCE.XLSX]Data'!$B$2:$B$4",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalFormulaAlias",
            attr_text="=DirectExternalCell",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalChainedCellAlias",
            attr_text="=DirectExternalFormulaAlias",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalChainedNameAlias",
            attr_text="DirectExternalName",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalFormulaWrapper",
            attr_text="=SUM(DirectExternalCell)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalFormulaWrapperSecond",
            attr_text="=SUM(DirectExternalFormulaWrapper)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalUnresolvedFormula",
            attr_text="=SUM(DirectExternalCell,NoSuchExternalFormulaName)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalFormulaDuplicate",
            attr_text="=SUM(DirectExternalCell,DirectExternalCell)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalAliasCycleA",
            attr_text="=DirectExternalAliasCycleB",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalAliasCycleB",
            attr_text="=DirectExternalAliasCycleA",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalUnresolvedAlias",
            attr_text="=NoSuchAlias",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalChainedShadowed",
            attr_text="=DirectExternalCell",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalLocalAlias",
            attr_text="'..\\inputs\\[SOURCE.XLSX]Data'!$B$2:$B$4",
            localSheetId=0,
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalAbsoluteAlias",
            attr_text="'C:\\Private\\[SOURCE.XLSX]Data'!$B$2:$B$4",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalMalformedAlias",
            attr_text="=SUM('..\\inputs\\[SOURCE.XLSX]Data'!$B$2:$B$4)",
        )
    )
    consumer.defined_names.add(
        DefinedName(
            "DirectExternalShadowed",
            attr_text="'..\\inputs\\[SOURCE.XLSX]Data'!$B$2:$B$4",
        )
    )
    consumer["Summary"].defined_names.add(
        DefinedName(
            "DirectExternalShadowed",
            attr_text="Summary!$A$1",
            localSheetId=0,
        )
    )
    consumer["Summary"].defined_names.add(
        DefinedName(
            "DirectExternalChainedShadowed",
            attr_text="Summary!$A$1",
            localSheetId=0,
        )
    )
    consumer["Summary"].defined_names.add(
        DefinedName(
            "DirectExternalLocalFormulaAlias",
            attr_text="=DirectExternalCell",
            localSheetId=0,
        )
    )
    consumer.save(consumer_path)

    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 21),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding for finding in source_entry.findings if finding.rule_id == "FF079"
    )
    rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )

    assert not report.incomplete
    assert finding.details["impacted_workbook_count"] == 1
    assert finding.details["impacted_formula_count"] == 10
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    assert [impact["location"] for impact in finding.details["sample_impacts"]] == [
        "Summary!D2",
        "Summary!E2",
        "Summary!F2",
        "Summary!G2",
        "Summary!J2",
        "Summary!L2",
        "Summary!M2",
        "Summary!N2",
        "Summary!S2",
        "Summary!U2",
    ]
    for private_value in (
        "PrivateDirectSourceRange",
        "PrivateDirectSourceAlias",
        "DirectExternalCell",
        "DirectExternalName",
        "DirectExternalFormulaAlias",
        "DirectExternalChainedCellAlias",
        "DirectExternalChainedNameAlias",
        "DirectExternalFormulaWrapper",
        "DirectExternalFormulaWrapperSecond",
        "DirectExternalUnresolvedFormula",
        "DirectExternalFormulaDuplicate",
        "DirectExternalAliasCycleA",
        "DirectExternalUnresolvedAlias",
        "DirectExternalChainedShadowed",
        "DirectExternalLocalFormulaAlias",
        "SOURCE.XLSX",
    ):
        assert all(private_value not in value for value in rendered)


def test_external_three_d_spans_require_complete_candidate_tab_metadata(
    tmp_path: Path,
) -> None:
    source_path = _write_workbook(
        tmp_path / "source.xlsx",
        "Jan",
        {"B2": 10},
    )
    source = load_workbook(source_path)
    source.create_sheet("Feb")["B2"] = 20
    source.create_sheet("Mar")["B2"] = 30
    source.save(source_path)

    snapshot = load_snapshot(source_path)

    assert snapshot.workbook_tab_order_complete
    assert snapshot.worksheet_tab_order_complete
    assert snapshot.worksheet_tab_order == ("Jan", "Feb", "Mar")
    assert _canonical_three_d_sheet_span(snapshot, "Jan", "Mar") == (
        "Jan",
        "Feb",
        "Mar",
    )
    assert (
        _canonical_three_d_sheet_span(
            replace(snapshot, workbook_tab_order_complete=False),
            "Jan",
            "Mar",
        )
        == ()
    )
    assert (
        _canonical_three_d_sheet_span(
            replace(snapshot, worksheet_tab_order_complete=False),
            "Jan",
            "Mar",
        )
        == ()
    )
    assert (
        _canonical_three_d_sheet_span(
            replace(snapshot, worksheet_tab_order=("Jan", "Mar")),
            "Jan",
            "Mar",
        )
        == ()
    )


def test_portfolio_resolves_static_external_three_d_a1_spans_privately(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    source_path = _write_workbook(
        baseline / "inputs" / "source.xlsx",
        "Jan",
        {"B2": 10, "B3": 11, "B4": 12},
    )
    source = load_workbook(source_path)
    for title, values in (
        ("Feb", {"B2": 20, "B3": 21, "B4": 22}),
        ("Mar", {"B2": 30, "B3": 31, "B4": 32}),
        ("Outside", {"B2": 40, "B3": 41, "B4": 42}),
    ):
        worksheet = source.create_sheet(title)
        for coordinate, value in values.items():
            worksheet[coordinate] = value
    source.save(source_path)

    direct_path = _write_workbook(
        baseline / "reports" / "direct.xlsx",
        "Summary",
        {
            "D2": "=SUM('..\\inputs\\[SOURCE.XLSX]Jan:Mar'!$B$2:$B$4)",
            "E2": "=SUM(DirectExternalThreeDBridge)",
            "F2": "=SUM(DirectExternalThreeDLeadingEquals)",
            "G2": "=SUM(DirectExternalThreeDReverse)",
            "H2": "=SUM(DirectExternalThreeDShadowed)",
            "I2": "=SUM(DirectExternalThreeDLocal)",
            "J2": "=SUM('..\\inputs\\[SOURCE.XLSX]Jan:Missing'!$B$2:$B$4)",
            "K2": "=SUM(DirectExternalThreeDFormula)",
            "L2": "=SUM(DirectExternalThreeDLambda(7))",
        },
    )
    direct = load_workbook(direct_path)
    direct.defined_names.add(
        DefinedName(
            "DirectExternalThreeD",
            attr_text="'..\\inputs\\[SOURCE.XLSX]Jan:Mar'!$B$2:$B$4",
        )
    )
    direct.defined_names.add(
        DefinedName(
            "DirectExternalThreeDBridge",
            attr_text="=DirectExternalThreeD",
        )
    )
    direct.defined_names.add(
        DefinedName(
            "DirectExternalThreeDFormula",
            attr_text="=SUM(DirectExternalThreeDBridge)",
        )
    )
    direct.defined_names.add(
        DefinedName(
            "DirectExternalThreeDLambda",
            attr_text="=LAMBDA(value,SUM(value,DirectExternalThreeDBridge))",
        )
    )
    direct.defined_names.add(
        DefinedName(
            "DirectExternalThreeDLeadingEquals",
            attr_text="='..\\inputs\\[SOURCE.XLSX]Jan:Mar'!$B$2:$B$4",
        )
    )
    direct.defined_names.add(
        DefinedName(
            "DirectExternalThreeDReverse",
            attr_text="'..\\inputs\\[SOURCE.XLSX]Mar:Jan'!$B$2:$B$4",
        )
    )
    direct.defined_names.add(
        DefinedName(
            "DirectExternalThreeDShadowed",
            attr_text="'..\\inputs\\[SOURCE.XLSX]Jan:Mar'!$B$2:$B$4",
        )
    )
    direct["Summary"].defined_names.add(
        DefinedName(
            "DirectExternalThreeDShadowed",
            attr_text="Summary!$A$1",
            localSheetId=0,
        )
    )
    direct["Summary"].defined_names.add(
        DefinedName(
            "DirectExternalThreeDLocal",
            attr_text="'..\\inputs\\[SOURCE.XLSX]Jan:Mar'!$B$2:$B$4",
            localSheetId=0,
        )
    )
    direct.save(direct_path)
    make_indexed_external_workbook_a1_link_model(
        baseline / "reports" / "package.xlsx",
        source_sheet="Jan:Mar",
        source_range="$B$2:$B$4",
        consumer_formula_alias=True,
    )

    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Feb"]["B3"], "value", 23),
    )

    report = compare_portfolios(
        baseline,
        candidate,
        policy=parse_policy(
            {"version": 1, "rules": {"no_cross_workbook_impacts": True}}
        ),
    )
    source_entry = next(
        entry for entry in report.workbooks if entry.path == "inputs/source.xlsx"
    )
    finding = next(
        finding
        for finding in source_entry.findings
        if finding.rule_id == "FF079" and finding.location == ("Feb", "B3")
    )
    rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )

    assert not report.incomplete
    assert finding.details["impacted_workbook_count"] == 2
    assert finding.details["impacted_formula_count"] == 7
    assert {finding.rule_id for finding in source_entry.findings} >= {"FF079", "FFP079"}
    assert {
        (impact["workbook"], impact["location"])
        for impact in finding.details["sample_impacts"]
    } == {
        ("reports/direct.xlsx", "Summary!D2"),
        ("reports/direct.xlsx", "Summary!E2"),
        ("reports/direct.xlsx", "Summary!F2"),
        ("reports/direct.xlsx", "Summary!K2"),
        ("reports/direct.xlsx", "Summary!L2"),
        ("reports/package.xlsx", "Model!D2"),
        ("reports/package.xlsx", "Model!E2"),
    }
    outside_candidate = tmp_path / "candidate-outside"
    shutil.copytree(baseline, outside_candidate)
    rewrite(
        outside_candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Outside"]["B3"], "value", 43),
    )
    outside_report = compare_portfolios(baseline, outside_candidate)
    outside_source_entry = next(
        entry
        for entry in outside_report.workbooks
        if entry.path == "inputs/source.xlsx"
    )
    assert "FF079" not in {finding.rule_id for finding in outside_source_entry.findings}
    for private_value in (
        "DirectExternalThreeD",
        "DirectExternalThreeDBridge",
        "DirectExternalThreeDFormula",
        "DirectExternalThreeDLambda",
        "DirectExternalThreeDLeadingEquals",
        "DirectExternalThreeDReverse",
        "DirectExternalThreeDShadowed",
        "DirectExternalThreeDLocal",
        "PackageExternalCell",
        "PackageExternalFormulaAlias",
        "SOURCE.XLSX",
    ):
        assert all(private_value not in value for value in rendered)


def test_portfolio_fails_closed_for_dynamic_or_absolute_package_indexed_names(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    private_path = "PRIVATE-PACKAGE-INDEXED-EXTERNAL-PATH"
    source_path = _write_workbook(
        baseline / "inputs" / "source.xlsx",
        "Data",
        {"B2": 10, "B3": 20, "B4": 30},
    )
    source = load_workbook(source_path)
    source.defined_names.add(
        DefinedName(
            "PrivatePackageDynamicName",
            attr_text="=OFFSET(Data!$B$2,0,0,3,1)",
        )
    )
    source.defined_names.add(
        DefinedName("PrivatePackageStaticName", attr_text="Data!$B$2:$B$4")
    )
    source.save(source_path)
    make_indexed_external_workbook_name_link_model(
        baseline / "reports" / "dynamic.xlsx",
        target_paths=("../inputs/source.xlsx",),
        source_name="PrivatePackageDynamicName",
    )
    make_indexed_external_workbook_name_link_model(
        baseline / "reports" / "absolute.xlsx",
        target_paths=(f"C:\\{private_path}\\source.xlsx",),
        source_name="PrivatePackageStaticName",
    )
    make_indexed_external_workbook_name_link_model(
        baseline / "reports" / "ambiguous.xlsx",
        target_paths=("../inputs/source.xlsx",),
        source_name="PrivatePackageStaticName",
    )
    make_indexed_external_workbook_name_link_model(
        baseline / "reports" / "rebound.xlsx",
        target_paths=("../inputs/source.xlsx", "../decoy.xlsx"),
        source_name="PrivatePackageStaticName",
    )
    make_indexed_external_workbook_name_link_model(
        baseline / "reports" / "sheet-local.xlsx",
        target_paths=("../inputs/source.xlsx",),
        source_name="PrivatePackageStaticName",
        consumer_alias_local_sheet_id=1,
        include_direct_indexed_formula=False,
    )
    make_indexed_external_workbook_a1_link_model(
        baseline / "reports" / "a1-absolute.xlsx",
        target_paths=(f"C:\\{private_path}\\source.xlsx",),
    )
    make_indexed_external_workbook_a1_link_model(
        baseline / "reports" / "a1-sheet-local.xlsx",
        target_paths=("../inputs/source.xlsx",),
        consumer_alias_local_sheet_id=1,
        include_direct_indexed_formula=False,
    )
    shutil.copytree(baseline, candidate)
    duplicate_external_link_definition(candidate / "reports" / "ambiguous.xlsx")
    duplicate_indexed_external_link_part_binding(candidate / "reports" / "rebound.xlsx")
    rewrite(
        candidate / "inputs" / "source.xlsx",
        lambda workbook: setattr(workbook["Data"]["B3"], "value", 21),
    )

    report = compare_portfolios(baseline, candidate)
    rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )

    assert not report.incomplete
    assert not any(
        finding.rule_id == "FF079"
        for entry in report.workbooks
        for finding in entry.findings
    )
    assert all(private_path not in value for value in rendered)
    assert all("PrivatePackageDynamicName" not in value for value in rendered)
    assert all("PrivatePackageStaticName" not in value for value in rendered)
    assert all("PackageExternalCell" not in value for value in rendered)
    absolute_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/absolute.xlsx"
    )
    assert absolute_entry.after is not None
    assert private_path not in as_json(profile_snapshot(absolute_entry.after))
    a1_absolute_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/a1-absolute.xlsx"
    )
    assert a1_absolute_entry.after is not None
    assert private_path not in as_json(profile_snapshot(a1_absolute_entry.after))
    ambiguous_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/ambiguous.xlsx"
    )
    assert ambiguous_entry.after is not None
    assert any(
        "without exactly one external workbook" in warning
        for warning in ambiguous_entry.after.parser_warnings
    )
    rebound_entry = next(
        entry for entry in report.workbooks if entry.path == "reports/rebound.xlsx"
    )
    assert rebound_entry.after is not None
    assert any(
        "package-indexed external-reference declaration or package part" in warning
        for warning in rebound_entry.after.parser_warnings
    )


def test_portfolio_never_guesses_or_discloses_unresolved_external_link_paths(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    private_source = "PRIVATE-EXTERNAL-LINK-PATH"
    source_path = _write_workbook(baseline / "inputs.xlsx", "Data", {"B2": 100})
    source = load_workbook(source_path)
    source.defined_names.add(DefinedName("ExternalName", attr_text="Data!$B$2"))
    source.save(source_path)
    _write_workbook(
        baseline / "reports" / "summary.xlsx",
        "Summary",
        {
            "D2": "=[inputs.xlsx]Data!B2",
            "E2": f"='C:\\{private_source}\\[inputs.xlsx]Data'!B2",
            "F2": "='..\\..\\[inputs.xlsx]Data'!B2",
            "G2": "='..\\[inputs.xlsx]Data'!B2",
            "H2": "=[inputs.xlsx]ExternalName",
            "I2": f"='C:\\{private_source}\\[inputs.xlsx]ExternalName'",
            "J2": "='..\\..\\[inputs.xlsx]ExternalName'",
        },
    )
    shutil.copytree(baseline, candidate)
    rewrite(
        candidate / "inputs.xlsx",
        lambda workbook: setattr(workbook["Data"]["B2"], "value", 200),
    )

    report = compare_portfolios(baseline, candidate)
    source = next(entry for entry in report.workbooks if entry.path == "inputs.xlsx")
    rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )

    finding = next(finding for finding in source.findings if finding.rule_id == "FF079")
    assert finding.details["impacted_formula_count"] == 1
    assert finding.details["sample_impacts"][0]["location"] == "Summary!G2"
    assert all(private_source not in value for value in rendered)
    assert all("ExternalName" not in value for value in rendered)


def test_portfolio_fails_closed_when_cross_workbook_impact_bound_is_reached(
    tmp_path: Path,
) -> None:
    baseline, candidate = _cross_workbook_portfolio_pair(tmp_path)

    report = compare_portfolios(baseline, candidate, max_link_impact=2)
    source = next(entry for entry in report.workbooks if entry.path == "inputs.xlsx")

    assert report.incomplete
    assert {finding.rule_id for finding in source.findings} >= {"FF079", "FF080"}
    assert next(
        finding for finding in source.findings if finding.rule_id == "FF080"
    ).details == {"max_link_impact": 2}

    output = tmp_path / "portfolio.json"
    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--max-link-impact",
                "2",
                "--format",
                "json",
                "--output",
                str(output),
            ]
        )
        == 2
    )
    assert "FF080" in output.read_text(encoding="utf-8")


def test_portfolio_added_workbook_outputs_keep_contents_private(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    baseline.mkdir()
    candidate.mkdir()
    workbook_path = make_model(candidate / "added.xlsx")
    secret = "PORTFOLIO-PRIVATE-CONTENT-DO-NOT-REPORT"
    workbook = load_workbook(workbook_path)
    workbook["Inputs"]["A1"] = secret
    workbook.save(workbook_path)

    report = compare_portfolios(baseline, candidate)
    rendered = (
        as_json(report.to_dict()),
        portfolio_to_markdown(report),
        as_json(portfolio_to_sarif(report)),
    )
    assert all(secret not in value for value in rendered)
    sarif_result = portfolio_to_sarif(report)["runs"][0]["results"][0]
    assert sarif_result["ruleId"] == "FF077"
    assert sarif_result["locations"][0]["physicalLocation"]["artifactLocation"]["uri"] == (
        "added.xlsx"
    )


def test_portfolio_preserves_a_report_for_an_unreadable_workbook(tmp_path: Path) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    (candidate / "models" / "shared.xlsx").write_bytes(b"not a workbook")

    report = compare_portfolios(baseline, candidate)
    entry = report.to_dict()["workbooks"][0]
    assert report.incomplete
    assert entry["status"] == "unreadable"
    assert entry["findings"] == [
        {
            "rule_id": "FF078",
            "severity": "critical",
            "message": "Workbook could not be inspected; portfolio comparison is incomplete.",
            "location": None,
            "details": {"unreadable_sides": ["candidate"]},
        }
    ]

    output = tmp_path / "portfolio.md"
    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--output",
                str(output),
            ]
        )
        == 2
    )
    assert output.is_file()
    assert "FF078" in output.read_text(encoding="utf-8")


def test_portfolio_preserves_membership_evidence_for_an_unreadable_new_workbook(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    baseline.mkdir()
    candidate.mkdir()
    (candidate / "new.xlsx").write_bytes(b"not a workbook")
    policy = parse_policy(
        {
            "version": 1,
            "rules": {"no_portfolio_membership_changes": True},
        }
    )

    report = compare_portfolios(baseline, candidate, policy=policy)
    entry = report.to_dict()["workbooks"][0]

    assert report.incomplete
    assert entry["status"] == "unreadable"
    assert len(entry["changes"]) == 1
    assert entry["changes"][0]["kind"] == "workbook_added"
    assert entry["changes"][0]["details"] == {"portfolio_change": "added"}
    assert {finding["rule_id"] for finding in entry["findings"]} == {
        "FF077",
        "FF078",
        "FFP077",
    }


def test_portfolio_cli_refuses_a_report_inside_an_input_directory(tmp_path: Path) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    output = baseline / "report.md"

    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--output",
                str(output),
            ]
        )
        == 2
    )
    assert not output.exists()


def test_portfolio_inventory_fails_closed_for_limits_and_unsupported_formats(
    tmp_path: Path,
) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    make_model(baseline / "second.xlsx")
    make_model(candidate / "second.xlsx")

    with pytest.raises(PortfolioError, match="max_workbooks=1"):
        compare_portfolios(baseline, candidate, max_workbooks=1)

    (baseline / "legacy.xlsb").write_bytes(b"unsupported")
    with pytest.raises(PortfolioError, match="unsupported spreadsheet files"):
        discover_workbooks(baseline, label="baseline")


def test_portfolio_inventory_bounds_all_filesystem_entries_before_filtering(
    tmp_path: Path,
) -> None:
    root = tmp_path / "portfolio"
    root.mkdir()
    for index in range(3):
        (root / f"note-{index}.txt").write_text("ignored", encoding="utf-8")

    assert discover_workbooks(root, label="baseline", max_inventory_entries=3) == {}

    (root / "note-3.txt").write_text("over budget", encoding="utf-8")
    with pytest.raises(PortfolioError, match="max_inventory_entries=3"):
        discover_workbooks(root, label="baseline", max_inventory_entries=3)
    with pytest.raises(PortfolioError, match="max_inventory_entries must be at least 1"):
        discover_workbooks(root, label="baseline", max_inventory_entries=0)


def test_portfolio_source_byte_budget_is_exact_and_precedes_snapshot_reads(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    first = baseline / "models" / "shared.xlsx"
    _copy_workbook(first, baseline / "second.xlsx")
    _copy_workbook(first, candidate / "second.xlsx")
    total_source_bytes = sum(
        workbook.stat().st_size
        for workbook in (baseline / "models" / "shared.xlsx", baseline / "second.xlsx")
    )

    exact_report = compare_portfolios(
        baseline,
        candidate,
        max_portfolio_source_bytes=total_source_bytes,
    )
    assert exact_report.incomplete is False

    snapshot_reads: list[Path] = []

    def fail_if_called(path: Path, **kwargs: object):
        snapshot_reads.append(Path(path))
        raise AssertionError("source-byte preflight must run before snapshot reads")

    monkeypatch.setattr(portfolio_module, "load_snapshot", fail_if_called)
    with pytest.raises(
        PortfolioError,
        match=f"max_portfolio_source_bytes={total_source_bytes - 1}",
    ):
        compare_portfolios(
            baseline,
            candidate,
            max_portfolio_source_bytes=total_source_bytes - 1,
        )
    assert snapshot_reads == []


def test_portfolio_source_byte_budget_rejects_a_nonpositive_limit(tmp_path: Path) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)

    with pytest.raises(
        PortfolioError, match="max_portfolio_source_bytes must be at least 1"
    ):
        compare_portfolios(baseline, candidate, max_portfolio_source_bytes=0)
    with pytest.raises(
        PortfolioError, match="max_portfolio_snapshot_cells must be at least 1"
    ):
        compare_portfolios(baseline, candidate, max_portfolio_snapshot_cells=0)
    with pytest.raises(PortfolioError, match="max_dependency_edges must be at least 1"):
        compare_portfolios(baseline, candidate, max_dependency_edges=0)
    with pytest.raises(
        PortfolioError,
        match="max_formula_defined_name_states must be at least 1",
    ):
        compare_portfolios(baseline, candidate, max_formula_defined_name_states=0)


def test_portfolio_source_byte_budget_checks_the_candidate_before_reads(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    candidate_source = candidate / "models" / "shared.xlsx"

    def grow_candidate_source(workbook) -> None:
        workbook["Inputs"]["A1"] = "".join(f"{index:04x}" for index in range(4_096))

    rewrite(candidate_source, grow_candidate_source)
    baseline_source = baseline / "models" / "shared.xlsx"
    assert candidate_source.stat().st_size > baseline_source.stat().st_size
    snapshot_reads: list[Path] = []

    def fail_if_called(path: Path, **kwargs: object):
        snapshot_reads.append(Path(path))
        raise AssertionError("source-byte preflight must run before snapshot reads")

    monkeypatch.setattr(portfolio_module, "load_snapshot", fail_if_called)
    with pytest.raises(PortfolioError, match="Candidate portfolio contains"):
        compare_portfolios(
            baseline,
            candidate,
            max_portfolio_source_bytes=baseline_source.stat().st_size,
        )
    assert snapshot_reads == []


def test_portfolio_snapshot_cell_budget_is_exact_and_stops_later_reads(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    baseline_first = baseline / "models" / "shared.xlsx"
    candidate_first = candidate / "models" / "shared.xlsx"
    baseline_second = baseline / "second.xlsx"
    candidate_second = candidate / "second.xlsx"
    _copy_workbook(baseline_first, baseline_second)
    _copy_workbook(candidate_first, candidate_second)
    baseline_snapshot_cells = sum(
        len(load_snapshot(workbook).cells) for workbook in (baseline_first, baseline_second)
    )
    candidate_snapshot_cells = sum(
        len(load_snapshot(workbook).cells) for workbook in (candidate_first, candidate_second)
    )
    exact_limit = max(baseline_snapshot_cells, candidate_snapshot_cells)

    exact_report = compare_portfolios(
        baseline,
        candidate,
        max_portfolio_snapshot_cells=exact_limit,
    )
    assert exact_report.incomplete is False

    original_load_snapshot = portfolio_module.load_snapshot
    snapshot_reads: list[Path] = []

    def record_load(path: Path, **kwargs: object):
        snapshot_reads.append(Path(path))
        return original_load_snapshot(path, **kwargs)

    monkeypatch.setattr(portfolio_module, "load_snapshot", record_load)
    with pytest.raises(
        PortfolioError,
        match=f"max_portfolio_snapshot_cells={baseline_snapshot_cells - 1}",
    ):
        compare_portfolios(
            baseline,
            candidate,
            max_portfolio_snapshot_cells=baseline_snapshot_cells - 1,
        )
    assert snapshot_reads == [baseline_first, candidate_first, baseline_second]


def test_portfolio_snapshot_cell_budget_checks_the_candidate_before_later_reads(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    baseline_source = baseline / "models" / "shared.xlsx"
    candidate_source = candidate / "models" / "shared.xlsx"

    def grow_candidate_snapshot(workbook) -> None:
        for row in range(1, 33):
            workbook["Model"].cell(row=row, column=8, value=f"=Inputs!A{row}")

    rewrite(candidate_source, grow_candidate_snapshot)
    _copy_workbook(baseline_source, baseline / "second.xlsx")
    _copy_workbook(candidate_source, candidate / "second.xlsx")
    baseline_snapshot_cells = len(load_snapshot(baseline_source).cells)
    assert len(load_snapshot(candidate_source).cells) > baseline_snapshot_cells
    original_load_snapshot = portfolio_module.load_snapshot
    snapshot_reads: list[Path] = []

    def record_load(path: Path, **kwargs: object):
        snapshot_reads.append(Path(path))
        return original_load_snapshot(path, **kwargs)

    monkeypatch.setattr(portfolio_module, "load_snapshot", record_load)
    with pytest.raises(PortfolioError, match="Candidate portfolio retains"):
        compare_portfolios(
            baseline,
            candidate,
            max_portfolio_snapshot_cells=baseline_snapshot_cells,
        )
    assert snapshot_reads == [baseline_source, candidate_source]


def test_portfolio_dependency_edge_budget_is_exact_and_stops_later_reads(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    for root in (baseline, candidate):
        _write_named_formula_fanout_workbook(root / "first.xlsx")
        _write_named_formula_fanout_workbook(root / "second.xlsx")

    exact_report = compare_portfolios(
        baseline,
        candidate,
        max_dependency_edges=8,
    )
    assert exact_report.incomplete is False

    original_load_snapshot = portfolio_module.load_snapshot
    snapshot_reads: list[Path] = []

    def record_load(path: Path, **kwargs: object):
        snapshot_reads.append(Path(path))
        return original_load_snapshot(path, **kwargs)

    monkeypatch.setattr(portfolio_module, "load_snapshot", record_load)
    with pytest.raises(
        PortfolioError,
        match=r"Baseline portfolio dependency graph exceeds max_dependency_edges=7",
    ):
        compare_portfolios(
            baseline,
            candidate,
            max_dependency_edges=7,
        )
    assert snapshot_reads == [
        baseline / "first.xlsx",
        candidate / "first.xlsx",
        baseline / "second.xlsx",
    ]


def test_portfolio_formula_defined_name_state_budget_is_shared_and_stops_later_reads(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    for root in (baseline, candidate):
        _write_formula_defined_name_action_chain_workbook(root / "first.xlsx")
        _write_formula_defined_name_action_chain_workbook(root / "second.xlsx")

    exact_report = compare_portfolios(
        baseline,
        candidate,
        max_formula_defined_name_states=34,
    )
    assert exact_report.incomplete is False

    original_load_snapshot = portfolio_module.load_snapshot
    snapshot_reads: list[Path] = []

    def record_load(path: Path, **kwargs: object):
        snapshot_reads.append(Path(path))
        return original_load_snapshot(path, **kwargs)

    monkeypatch.setattr(portfolio_module, "load_snapshot", record_load)
    with pytest.raises(
        PortfolioError,
        match=(
            r"Baseline portfolio formula-defined-name propagation exceeds "
            r"max_formula_defined_name_states=33"
        ),
    ):
        compare_portfolios(
            baseline,
            candidate,
            max_formula_defined_name_states=33,
        )
    assert snapshot_reads == [
        baseline / "first.xlsx",
        candidate / "first.xlsx",
        baseline / "second.xlsx",
    ]


def test_portfolio_change_analysis_budget_is_shared_across_matched_workbooks(
    tmp_path: Path,
) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    for root, value in ((baseline, 1), (candidate, 2)):
        for name in ("first.xlsx", "second.xlsx"):
            _write_workbook(
                root / name,
                "Model",
                {"A1": value, "B1": "=A1"},
            )

    exact = compare_portfolios(
        baseline,
        candidate,
        max_change_analysis_states=4,
    )
    assert exact.incomplete is False
    with pytest.raises(
        PortfolioError,
        match="Portfolio change analysis exceeds max_change_analysis_states=3",
    ):
        compare_portfolios(
            baseline,
            candidate,
            max_change_analysis_states=3,
        )
    with pytest.raises(
        PortfolioError,
        match="max_change_analysis_states must be at least 1",
    ):
        compare_portfolios(
            baseline,
            candidate,
            max_change_analysis_states=0,
        )


def test_portfolio_cli_passes_the_inventory_entry_limit(tmp_path: Path) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    (baseline / "note.txt").write_text("ignored", encoding="utf-8")
    (candidate / "note.txt").write_text("ignored", encoding="utf-8")

    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--max-inventory-entries",
                "2",
            ]
        )
        == 2
    )


def test_portfolio_cli_passes_the_source_byte_limit(tmp_path: Path) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)

    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--max-portfolio-source-bytes",
                "1",
            ]
        )
        == 2
    )


def test_portfolio_cli_passes_the_snapshot_cell_limit(tmp_path: Path) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)

    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--max-portfolio-snapshot-cells",
                "1",
            ]
        )
        == 2
    )


def test_portfolio_cli_passes_the_change_analysis_state_limit(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    _write_workbook(baseline / "model.xlsx", "Model", {"A1": 1, "B1": "=A1"})
    _write_workbook(candidate / "model.xlsx", "Model", {"A1": 2, "B1": "=A1"})

    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--max-change-analysis-states",
                "1",
            ]
        )
        == 2
    )


def test_portfolio_cli_fails_before_writing_an_oversized_report(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    _write_workbook(baseline / "model.xlsx", "Model", {"A1": 1})
    _write_workbook(candidate / "model.xlsx", "Model", {"A1": 2})
    output = tmp_path / "portfolio.json"

    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--format",
                "json",
                "--output",
                str(output),
                "--max-report-bytes",
                "1",
            ]
        )
        == 2
    )
    assert not output.exists()


def test_portfolio_renderers_respect_the_report_byte_limit(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline"
    candidate = tmp_path / "candidate"
    _write_workbook(baseline / "model.xlsx", "Model", {"A1": 1})
    _write_workbook(candidate / "model.xlsx", "Model", {"A1": 2})
    report = compare_portfolios(baseline, candidate)

    assert portfolio_to_markdown(report, max_bytes=1_000_000).startswith(
        "# FormulaFence portfolio report"
    )
    assert portfolio_to_html(report, max_bytes=1_000_000).startswith("<!doctype html>")
    assert as_json(report.to_dict(), max_bytes=1_000_000).startswith("{")
    assert as_json(portfolio_to_sarif(report), max_bytes=1_000_000).startswith("{")

    with pytest.raises(FormulaFenceError, match="max_report_bytes=1"):
        portfolio_to_markdown(report, max_bytes=1)
    with pytest.raises(FormulaFenceError, match="max_report_bytes=1"):
        portfolio_to_html(report, max_bytes=1)
    with pytest.raises(FormulaFenceError, match="max_report_bytes=1"):
        as_json(report.to_dict(), max_bytes=1)
    with pytest.raises(FormulaFenceError, match="max_report_bytes=1"):
        as_json(portfolio_to_sarif(report), max_bytes=1)


def test_portfolio_cli_defaults_the_inventory_entry_limit() -> None:
    arguments = build_parser().parse_args(["portfolio", "before", "after"])

    assert arguments.max_inventory_entries == DEFAULT_MAX_INVENTORY_ENTRIES


def test_portfolio_cli_defaults_the_source_byte_limit() -> None:
    arguments = build_parser().parse_args(["portfolio", "before", "after"])

    assert arguments.max_portfolio_source_bytes == DEFAULT_MAX_PORTFOLIO_SOURCE_BYTES


def test_portfolio_cli_defaults_the_snapshot_cell_limit() -> None:
    arguments = build_parser().parse_args(["portfolio", "before", "after"])

    assert arguments.max_portfolio_snapshot_cells == DEFAULT_MAX_PORTFOLIO_SNAPSHOT_CELLS


def test_portfolio_cli_defaults_the_dependency_edge_limit() -> None:
    arguments = build_parser().parse_args(["portfolio", "before", "after"])

    assert arguments.max_dependency_edges == DEFAULT_MAX_DEPENDENCY_EDGES


def test_portfolio_cli_defaults_the_formula_defined_name_state_limit() -> None:
    arguments = build_parser().parse_args(["portfolio", "before", "after"])

    assert (
        arguments.max_formula_defined_name_states
        == DEFAULT_MAX_FORMULA_DEFINED_NAME_STATES
    )


def test_portfolio_cli_defaults_the_change_analysis_state_limit() -> None:
    arguments = build_parser().parse_args(["portfolio", "before", "after"])

    assert arguments.max_change_analysis_states == DEFAULT_MAX_CHANGE_ANALYSIS_STATES


def test_portfolio_cli_defaults_the_report_byte_limit() -> None:
    arguments = build_parser().parse_args(["portfolio", "before", "after"])

    assert arguments.max_report_bytes == DEFAULT_MAX_REPORT_BYTES


def test_portfolio_inventory_ignores_office_lock_files(tmp_path: Path) -> None:
    baseline, _ = _portfolio_pair(tmp_path)
    (baseline / "~$shared.xlsx").write_bytes(b"transient office lock")

    assert set(discover_workbooks(baseline, label="baseline")) == {"models/shared.xlsx"}


def test_portfolio_inventory_refuses_symlinked_paths(tmp_path: Path) -> None:
    baseline, _ = _portfolio_pair(tmp_path)
    (baseline / "linked-models").symlink_to(baseline / "models", target_is_directory=True)

    with pytest.raises(PortfolioError, match="symlinked path"):
        discover_workbooks(baseline, label="baseline")


def test_portfolio_inventory_fails_closed_when_a_subdirectory_cannot_be_read(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    blocked = baseline / "blocked"
    blocked.mkdir()
    original_scandir = portfolio_module.os.scandir

    def refuse_blocked_directory(path: str | Path):
        if Path(path) == blocked:
            raise PermissionError("controlled unreadable directory")
        return original_scandir(path)

    monkeypatch.setattr(portfolio_module.os, "scandir", refuse_blocked_directory)

    with pytest.raises(PortfolioError, match="Could not inventory baseline portfolio"):
        discover_workbooks(baseline, label="baseline")
    assert main(["portfolio", str(baseline), str(candidate)]) == 2


@pytest.mark.parametrize(
    "replacement_mode",
    ("in-place", "regular", "symlink"),
)
def test_portfolio_marks_a_workbook_unreadable_when_its_source_changes_after_inventory(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    replacement_mode: str,
) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    inventoried_source = baseline / "models" / "shared.xlsx"
    replacement = make_model(tmp_path / "outside.xlsx")
    replacement_workbook = load_workbook(replacement)
    replacement_workbook["Model"]["B2"] = "=Inputs!B3*2"
    replacement_workbook.save(replacement)
    original_load_snapshot = portfolio_module.load_snapshot
    swapped = False

    def swap_then_load(path: Path, **kwargs: object):
        nonlocal swapped
        if Path(path) == inventoried_source:
            if replacement_mode == "in-place":
                shutil.copyfile(replacement, inventoried_source)
            else:
                inventoried_source.unlink()
            if replacement_mode == "symlink":
                inventoried_source.symlink_to(replacement)
            elif replacement_mode == "regular":
                shutil.copyfile(replacement, inventoried_source)
            swapped = True
        return original_load_snapshot(path, **kwargs)

    monkeypatch.setattr(portfolio_module, "load_snapshot", swap_then_load)

    report = compare_portfolios(baseline, candidate)
    entry = report.workbooks[0]

    assert swapped
    assert inventoried_source.is_symlink() is (replacement_mode == "symlink")
    assert report.incomplete
    assert entry.status == "unreadable"
    assert entry.before is None
    assert entry.after is not None
    assert {finding.rule_id for finding in entry.findings} == {"FF078"}


def test_portfolio_cli_preserves_incomplete_evidence_for_a_late_source_change(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    baseline, candidate = _portfolio_pair(tmp_path)
    inventoried_source = baseline / "models" / "shared.xlsx"
    replacement = make_model(tmp_path / "replacement.xlsx")
    original_load_snapshot = portfolio_module.load_snapshot
    replaced = False

    def replace_then_load(path: Path, **kwargs: object):
        nonlocal replaced
        if Path(path) == inventoried_source:
            shutil.copyfile(replacement, inventoried_source)
            replaced = True
        return original_load_snapshot(path, **kwargs)

    monkeypatch.setattr(portfolio_module, "load_snapshot", replace_then_load)
    output = tmp_path / "portfolio.md"

    assert (
        main(
            [
                "portfolio",
                str(baseline),
                str(candidate),
                "--output",
                str(output),
            ]
        )
        == 2
    )
    assert replaced
    assert "FF078" in output.read_text(encoding="utf-8")


def test_portfolio_inventory_refuses_case_colliding_paths(tmp_path: Path) -> None:
    baseline, _ = _portfolio_pair(tmp_path)
    _copy_workbook(
        baseline / "models" / "shared.xlsx",
        baseline / "models" / "SHARED.xlsx",
    )

    with pytest.raises(PortfolioError, match="differ only by case"):
        discover_workbooks(baseline, label="baseline")
